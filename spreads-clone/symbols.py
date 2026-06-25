"""Symbol discovery and OANDA/MT5 name mapping for the spread monitor."""

from __future__ import annotations

import os
from pathlib import Path
import re
import json
from typing import Iterable, List, Optional, Sequence, Set


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_JOURNAL_PATH = ROOT_DIR / "journal" / "Trading Journal.xlsx"
DEFAULT_WATCHLIST_PATH = ROOT_DIR / "watchlist.json"

FX_CODES = {
    "AUD",
    "CAD",
    "CHF",
    "EUR",
    "GBP",
    "HKD",
    "JPY",
    "NOK",
    "NZD",
    "SEK",
    "SGD",
    "TRY",
    "USD",
    "ZAR",
}
CFD_CODES = {"XAU", "XAG", "XPT", "XPD", "WTI", "BCO", "NAS", "SPX", "US30", "GER30"}
CRYPTO_TOKENS = {"BTC", "ETH", "SOL", "USDT", "USDC", "BNB", "XRP", "DOGE", "ADA"}

DEFAULT_SYMBOLS = [
    "EUR_USD",
    "GBP_USD",
    "USD_JPY",
    "AUD_USD",
    "NZD_USD",
    "USD_CAD",
    "USD_CHF",
    "XAU_USD",
]


def _clean_symbol_token(value: str) -> str:
    return re.sub(r"[^A-Z0-9_]", "", str(value or "").upper())


def is_crypto_symbol(value: str) -> bool:
    token = _clean_symbol_token(value).replace("_", "")
    if token.endswith(("USDT", "USDC")):
        return True
    if len(token) >= 6 and (token[:3] in CRYPTO_TOKENS or token[-3:] in CRYPTO_TOKENS):
        return True
    return False


def normalize_oanda_symbol(value: object) -> Optional[str]:
    token = _clean_symbol_token(str(value or ""))
    if not token:
        return None
    if is_crypto_symbol(token):
        return None

    if re.fullmatch(r"[A-Z]{3}_[A-Z]{3}", token):
        base, quote = token.split("_")
        if (base in FX_CODES or base in CFD_CODES) and quote in FX_CODES:
            return token

    compact = token.replace("_", "")
    if re.fullmatch(r"[A-Z]{6}", compact):
        base, quote = compact[:3], compact[3:]
        if (base in FX_CODES or base in CFD_CODES) and quote in FX_CODES:
            return f"{base}_{quote}"

    if compact in {"XAUUSD", "XAGUSD", "XPTUSD", "XPDUSD"}:
        return f"{compact[:3]}_{compact[3:]}"

    return None


def oanda_to_mt5_symbol(oanda_symbol: str) -> str:
    return str(oanda_symbol or "").upper().replace("_", "")


def display_symbol(oanda_symbol: str) -> str:
    return str(oanda_symbol or "").replace("_", "/")


def symbols_from_env() -> List[str]:
    raw = os.getenv("SPREAD_MONITOR_SYMBOLS", "").strip()
    if not raw:
        return []
    result = []
    for item in re.split(r"[\s,;]+", raw):
        symbol = normalize_oanda_symbol(item)
        if symbol:
            result.append(symbol)
    return sorted(dict.fromkeys(result))


def symbols_from_journal(path: Path = DEFAULT_JOURNAL_PATH) -> List[str]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    found: Set[str] = set()
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                for value in row:
                    if value in (None, ""):
                        continue
                    text = str(value).upper()
                    for match in re.findall(r"\b[A-Z]{3}[_/]?[A-Z]{3}\b", text):
                        symbol = normalize_oanda_symbol(match.replace("/", "_"))
                        if symbol:
                            found.add(symbol)
    finally:
        workbook.close()
    return sorted(found)


def symbols_from_watchlist(path: Path = DEFAULT_WATCHLIST_PATH) -> List[str]:
    if not path.exists():
        return []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []
    raw_values: List[object] = []
    if isinstance(payload, list):
        raw_values.extend(payload)
    elif isinstance(payload, dict):
        for key in ("symbols", "watchlist", "items"):
            values = payload.get(key)
            if isinstance(values, list):
                raw_values.extend(values)
    found = []
    for value in raw_values:
        symbol = normalize_oanda_symbol(value)
        if symbol:
            found.append(symbol)
    return sorted(dict.fromkeys(found))


def normalize_available_oanda_symbols(values: Iterable[object]) -> List[str]:
    found = []
    for value in values:
        symbol = normalize_oanda_symbol(value)
        if symbol:
            found.append(symbol)
    return sorted(dict.fromkeys(found))


def normalize_available_mt5_symbols(values: Iterable[object]) -> List[str]:
    found = []
    for item in values:
        if isinstance(item, str):
            name = item
        else:
            name = str(getattr(item, "name", "") or "")
        symbol = normalize_oanda_symbol(name)
        if symbol is None:
            compact = re.sub(r"[^A-Z0-9]", "", name.upper())
            candidates = []
            if len(compact) >= 6:
                candidates.extend([compact[:6], compact[-6:]])
            for candidate in candidates:
                symbol = normalize_oanda_symbol(candidate)
                if symbol:
                    break
        if symbol:
            found.append(symbol)
    return sorted(dict.fromkeys(found))


def resolve_mt5_symbol(oanda_symbol: str, available_symbols: Sequence[object]) -> Optional[str]:
    base = oanda_to_mt5_symbol(oanda_symbol)
    if not base:
        return None
    names = [
        item if isinstance(item, str) else str(getattr(item, "name", "") or "")
        for item in available_symbols
    ]
    names = [name for name in names if name]

    for candidate in (oanda_symbol, base):
        for name in names:
            if name.upper() == candidate.upper():
                return name

    base_upper = base.upper()
    for name in names:
        normalized = re.sub(r"[^A-Z0-9]", "", name.upper())
        if normalized == base_upper:
            return name

    for name in names:
        normalized = re.sub(r"[^A-Z0-9]", "", name.upper())
        if normalized.startswith(base_upper) or normalized.endswith(base_upper):
            return name
    return None


def build_symbol_universe(
    *,
    journal_path: Path = DEFAULT_JOURNAL_PATH,
    watchlist_path: Path = DEFAULT_WATCHLIST_PATH,
    oanda_symbols: Optional[Iterable[object]] = None,
    mt5_symbols: Optional[Iterable[object]] = None,
    include_all_available: Optional[bool] = None,
) -> List[str]:
    env_symbols = symbols_from_env()
    if env_symbols:
        return env_symbols
    if include_all_available is None:
        include_all_available = str(os.getenv("SPREAD_MONITOR_INCLUDE_ALL_OANDA_INSTRUMENTS", "")).strip().lower() in {"1", "true", "yes", "y"}

    journal_symbols = symbols_from_journal(journal_path)
    watchlist_symbols = symbols_from_watchlist(watchlist_path)
    available_symbols: List[str] = []
    if oanda_symbols is not None:
        available_symbols.extend(normalize_available_oanda_symbols(oanda_symbols))
    if mt5_symbols is not None:
        available_symbols.extend(normalize_available_mt5_symbols(mt5_symbols))

    available_set = set(available_symbols)
    requested = list(dict.fromkeys([*journal_symbols, *watchlist_symbols]))
    found = [symbol for symbol in requested if not available_set or symbol in available_set]
    if include_all_available:
        found.extend(symbol for symbol in sorted(available_set) if symbol not in set(found))
    if not found:
        found.extend(symbol for symbol in DEFAULT_SYMBOLS if not available_set or symbol in available_set)
    if not found:
        found.extend(DEFAULT_SYMBOLS)
    return list(dict.fromkeys(found))
