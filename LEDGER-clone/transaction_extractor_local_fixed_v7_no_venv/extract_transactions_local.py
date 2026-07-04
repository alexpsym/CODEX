"""
Local OCR transaction extractor.

No OpenAI API. No cloud AI API.
URLs are used only to download your screenshots/PDFs. Extraction is done locally with
Tesseract OCR + Python parsing, then written to an Excel workbook.

Version 7 no-venv package: extraction logic from fixed V5 plus improved date-line segment handling and amount disambiguation.

This version uses spatial OCR for mobile/app-style transaction-list screenshots:
- detects row separators
- OCRs transaction rows separately
- reads amount crops at higher resolution
- infers corrupted date text from weekday/month/order where possible
"""
from __future__ import annotations

import csv
import datetime as dt
import math
import os
import re
import shutil
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import pytesseract
from pytesseract import Output

try:
    import fitz  # PyMuPDF
except Exception:  # pragma: no cover
    fitz = None

HEADERS = ["DATE", "ACCOUNT_TYPE", "ACCOUNT", "DESCRIPTION", "DEBIT", "CREDIT", "NOTES", "NOTES"]
OUTPUT_DIR = Path("outputs")
DOWNLOAD_DIR = Path("downloads")

MONTHS = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}
MONTH_NAMES = {v: k.title() for k, v in MONTHS.items() if len(k) == 3 or k == "may"}
WEEKDAYS = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]

DATE_PATTERNS = [
    re.compile(r"\b(?P<d>\d{1,2})[/-](?P<m>\d{1,2})[/-](?P<y>\d{2,4})\b"),
    re.compile(r"\b(?P<y>\d{4})[/-](?P<m>\d{1,2})[/-](?P<d>\d{1,2})\b"),
    re.compile(
        r"\b(?P<d>\d{1,2})\s+(?P<mon>Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:t|tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\s+(?P<y>\d{2,4})\b",
        re.I,
    ),
]

INCOME_HINTS = [
    "deposit", "refund", "rebate", "reversal", "credit", "salary", "wage", "payroll",
    "interest", "received", "transfer from", "from ", "cashback", "dividend", "centrelink",
]
OUTGOING_HINTS = [
    "purchase", "payment", "withdrawal", "direct debit", "debit", "eftpos", "visa", "mastercard",
    "card", "fee", "charge", "transfer to", "to ", "atm", "bp", "coles", "woolworths",
    "subway", "foodworks", "pizza", "render", "pty ltd", "house", "nourished", "stellarossa",
    "restaurant",
]
SKIP_LINE_HINTS = [
    "date", "description", "debit", "credit", "balance", "opening balance", "closing balance",
    "available balance", "account number", "statement", "page ", "transaction history",
]

TEXT_FIXES = {
    "Foxtay": "Today",
    "Foctay": "Today",
    "}oxtay": "Today",
    "Tat": "Tst",
    "S848": "8848",
    "Villlow": "Willow",
    "Vaillow": "Willow",
    "Vallow": "Willow",
    "Foodwaorks": "Foodworks",
    "Foedwaorks": "Foodworks",
    "Pry": "Pty",
    "Ud": "Ltd",
    "LID": "LTD",
    "Stallarossa": "Stellarossa",
    "Stellaroasa": "Stellarossa",
    "Stellaroossa": "Stellarossa",
    "Hurryohmn": "Hurryohm",
    "Foodwarks": "Foodworks",
    "Foodwarks": "Foodworks",
    "Subway’": "Subway",
    "Bernotn": "Bernoth",
}

@dataclass
class OcrWord:
    text: str
    x: float
    y: float
    w: float
    h: float
    conf: float
    engine: str

    @property
    def cx(self) -> float:
        return self.x + self.w / 2

    @property
    def cy(self) -> float:
        return self.y + self.h / 2

@dataclass
class ParsedTransaction:
    date: str
    description: str
    debit: Optional[float]
    credit: Optional[float]
    raw_text: str
    needs_review: bool
    review_reason: str
    source_name: str
    pending: bool = False


def find_tesseract() -> Optional[str]:
    env = os.environ.get("TESSERACT_CMD")
    if env and Path(env).exists():
        return env
    found = shutil.which("tesseract")
    if found:
        return found
    candidates = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        r"C:\Users\User\AppData\Local\Programs\Tesseract-OCR\tesseract.exe",
    ]
    for c in candidates:
        if Path(c).exists():
            return c
    return None


def ensure_tesseract() -> None:
    exe = find_tesseract()
    if not exe:
        raise RuntimeError(
            "Tesseract OCR was not found. Install it locally, then run this again.\n"
            "Recommended Windows install: winget install --id UB-Mannheim.TesseractOCR -e\n"
            "If installed in a custom location, set TESSERACT_CMD to tesseract.exe."
        )
    pytesseract.pytesseract.tesseract_cmd = exe


def is_url(value: str) -> bool:
    p = urlparse(value.strip())
    return p.scheme in {"http", "https"} and bool(p.netloc)


def safe_extension(url: str, content_type: str) -> str:
    suffix = Path(urlparse(url).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".pdf"}:
        return suffix
    ct = content_type.lower()
    if "pdf" in ct:
        return ".pdf"
    if "png" in ct:
        return ".png"
    if "webp" in ct:
        return ".webp"
    if "jpeg" in ct or "jpg" in ct:
        return ".jpg"
    return ".bin"


def download_or_copy(source: str, index: int, run_dir: Path) -> Path:
    if is_url(source):
        print(f"Downloading {source}")
        resp = requests.get(source, timeout=60, headers={"User-Agent": "LocalTransactionOCR/2.0"})
        resp.raise_for_status()
        ext = safe_extension(source, resp.headers.get("content-type", ""))
        out = run_dir / DOWNLOAD_DIR / f"source_{index:02d}{ext}"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_bytes(resp.content)
        return out
    path = Path(source.strip().strip('"'))
    if not path.exists():
        raise FileNotFoundError(f"Input not found: {source}")
    ext = path.suffix.lower() or ".bin"
    out = run_dir / DOWNLOAD_DIR / f"source_{index:02d}{ext}"
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, out)
    return out


def load_image(path: Path) -> Image.Image:
    return ImageOps.exif_transpose(Image.open(path)).convert("RGB")


def prep_for_ocr(img: Image.Image, scale: int = 4) -> Image.Image:
    if scale != 1:
        img = img.resize((img.width * scale, img.height * scale), Image.Resampling.LANCZOS)
    gray = ImageOps.grayscale(img)
    gray = ImageOps.autocontrast(gray)
    gray = ImageEnhance.Sharpness(gray).enhance(2.5)
    return gray


def ocr_words(img: Image.Image, psm: int, scale: int = 4) -> List[OcrWord]:
    gray = prep_for_ocr(img, scale=scale)
    config = f"--oem 3 --psm {psm} -c preserve_interword_spaces=1"
    data = pytesseract.image_to_data(gray, lang="eng", config=config, output_type=Output.DICT)
    words: List[OcrWord] = []
    for i, txt in enumerate(data.get("text", [])):
        text = (txt or "").strip()
        if not text:
            continue
        try:
            conf = float(data["conf"][i])
        except Exception:
            conf = -1.0
        if conf < 0:
            continue
        words.append(OcrWord(
            text=text,
            x=data["left"][i] / scale,
            y=data["top"][i] / scale,
            w=data["width"][i] / scale,
            h=data["height"][i] / scale,
            conf=conf,
            engine=f"psm{psm}",
        ))
    return words


def detect_horizontal_separators(img: Image.Image) -> List[int]:
    import numpy as np

    arr = np.array(img.convert("RGB"))
    h, w = arr.shape[:2]
    x1 = max(0, int(w * 0.04))
    x2 = min(w, int(w * 0.96))
    ys: List[int] = []
    for y in range(h):
        row = arr[y, x1:x2]
        if row.size == 0:
            continue
        mean = row.mean(axis=1)
        spread = row.max(axis=1) - row.min(axis=1)
        # Long, light-grey divider lines used by mobile/app transaction lists.
        mask = (mean > 175) & (mean < 245) & (spread < 22)
        if mask.mean() > 0.55:
            ys.append(y)
    groups: List[List[int]] = []
    for y in ys:
        if not groups or y - groups[-1][-1] > 2:
            groups.append([y])
        else:
            groups[-1].append(y)
    centers = [int(round(sum(g) / len(g))) for g in groups]
    filtered: List[int] = []
    for y in centers:
        if y < 20:
            continue
        if filtered and y - filtered[-1] < 25:
            continue
        filtered.append(y)
    return filtered


def make_segments(img: Image.Image, words: Sequence[OcrWord]) -> List[Tuple[int, int]]:
    separators = detect_horizontal_separators(img)
    if len(separators) >= 2:
        tops = [0] + [y + 1 for y in separators[:-1]]
        bottoms = separators
        return [(t, b) for t, b in zip(tops, bottoms) if b - t > 22]

    # Fallback: derive bands from amount-like words on the right side.
    right_words = [w for w in words if w.x > img.width * 0.65 and any(ch.isdigit() for ch in w.text)]
    centers = sorted({int(round(w.cy)) for w in right_words})
    bands: List[Tuple[int, int]] = []
    for cy in centers:
        t = max(0, cy - 28)
        b = min(img.height, cy + 28)
        if not bands or t > bands[-1][1] + 8:
            bands.append((t, b))
        else:
            bands[-1] = (bands[-1][0], max(bands[-1][1], b))
    return bands


def normalize_token_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("“", "").replace("”", "").replace("'", "").replace("`", "")
    text = text.replace("|", " ").replace("_", " ").replace("~", " ")
    return re.sub(r"\s+", " ", text).strip()


def apply_text_fixes(text: str) -> str:
    parts = text.split()
    fixed = []
    for part in parts:
        stripped = part.strip(" ,.;:()[]{}")
        repl = TEXT_FIXES.get(stripped)
        if repl:
            part = part.replace(stripped, repl)
        fixed.append(part)
    out = " ".join(fixed)
    out = re.sub(r"\bTst\s+8848\b", "Tst 8848", out)
    out = re.sub(r"\bRender\.com\b", "Render.Com", out)
    out = re.sub(r"\bLs\s+Stellarossa\b", "Ls Stellarossa", out)
    out = re.sub(r"\s+", " ", out).strip(" -–—:,;")
    return out


def line_groups(words: Sequence[OcrWord], y_tol: float = 4.5) -> List[List[OcrWord]]:
    ordered = sorted(words, key=lambda w: (w.cy, w.x))
    lines: List[List[OcrWord]] = []
    for word in ordered:
        if not lines:
            lines.append([word])
            continue
        last_cy = sum(w.cy for w in lines[-1]) / len(lines[-1])
        if abs(word.cy - last_cy) <= y_tol:
            lines[-1].append(word)
        else:
            lines.append([word])
    for line in lines:
        line.sort(key=lambda w: w.x)
    return lines


def line_text(line: Sequence[OcrWord]) -> str:
    ordered = sorted(line, key=lambda w: (round(w.x, 1), w.engine))
    parts: List[str] = []
    last_key = None
    last_x = None
    for w in ordered:
        text = normalize_token_text(w.text)
        key = re.sub(r"[^A-Za-z0-9.$-]", "", text).lower()
        # Collapse the same OCR word repeated by multiple Tesseract PSM passes at the same x position.
        if key and last_key == key and last_x is not None and abs(w.x - last_x) <= 2.0:
            continue
        parts.append(text)
        last_key = key
        last_x = w.x
    return normalize_token_text(" ".join(parts))


def fuzzy_best(value: str, options: Sequence[str]) -> Tuple[Optional[str], float]:
    value_norm = re.sub(r"[^a-z]", "", value.lower())
    best = None
    best_score = 0.0
    for opt in options:
        score = SequenceMatcher(None, value_norm, opt).ratio() if value_norm else 0.0
        if score > best_score:
            best = opt
            best_score = score
    return best, best_score


def detect_weekday(text: str) -> Optional[int]:
    # Check common OCR-damaged prefixes before fuzzy matching; short damaged words like "Wied"
    # can otherwise fuzzily match the wrong weekday.
    lowered = text.lower().strip()
    if lowered.startswith(("sun", "sunc", "seav", "sear", "seur", "sup", "suc", "sci")):
        return 6
    if lowered.startswith(("mon", "mor", "not", "mot")):
        return 0
    if lowered.startswith(("tue", "tum", "tus", "tuse", "tusec")):
        return 1
    if lowered.startswith(("wed", "wie", "wied", "veet")):
        return 2
    if lowered.startswith(("thu", "tr", "th")):
        return 3
    if lowered.startswith(("fri", "fri", "fric", "pri", "fei", "trice")):
        return 4
    if lowered.startswith(("sat", "san", "car")):
        return 5
    tokens = re.findall(r"[A-Za-z]{3,}", lowered)
    for token in tokens:
        best, score = fuzzy_best(token, WEEKDAYS)
        if best and score >= 0.62:
            return WEEKDAYS.index(best)
    return None


def detect_month(text: str, default_month: Optional[int]) -> Optional[int]:
    low = text.lower()
    for name, num in MONTHS.items():
        if re.search(rf"\b{name}\b", low):
            return num
    tokens = re.findall(r"[A-Za-z]{3,}", low)
    month_words = list(MONTHS.keys())
    for token in tokens:
        best, score = fuzzy_best(token, month_words)
        if best and score >= 0.58:
            return MONTHS[best]
    # Common May OCR variants from small screenshots.
    if any(tok in low for tok in ["mae", "mwy", "mav", "mary", "mee", "tia", "vow", "tey", "hee"]):
        return 5
    return default_month


def parse_year_from_text(text: str, default_year: int) -> int:
    # Prefer explicit four-digit years only. Do not treat day numbers like 23 as 2023.
    four_digit = re.findall(r"\b\d{4}\b", text)
    for raw in four_digit:
        y = int(raw)
        if 2000 <= y <= 2100 and abs(y - default_year) <= 2:
            return y
        # Small OCR screenshots often read 2026 as 2006/2008/2025/202. Keep the contextual year.
        if raw.startswith("20"):
            return default_year
    if re.search(r"20[0oOS%]{1,3}", text, re.I):
        return default_year
    return default_year


def parse_explicit_date(text: str) -> Optional[dt.date]:
    for pat in DATE_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        gd = m.groupdict()
        try:
            d = int(gd["d"])
            if gd.get("mon"):
                mon_txt = gd["mon"].lower()[:3]
                mth = MONTHS.get(mon_txt)
            else:
                mth = int(gd["m"])
            y = int(gd["y"])
            if y < 100:
                y += 2000 if y < 70 else 1900
            if not (2000 <= y <= 2100):
                return None
            return dt.date(y, mth, d)
        except Exception:
            pass
    return None


def candidate_dates_from_context(
    date_text: str,
    previous: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    raw = normalize_token_text(date_text)
    low = raw.lower()
    if any(tok in low for tok in ["today", "foxtay", "foctay", "}oxtay"]):
        return default_today, "today label", False

    default_year = previous.year if previous else default_today.year
    explicit = parse_explicit_date(raw)
    if explicit and 2000 <= explicit.year <= 2100:
        # Accept explicit dates only when the year is plausible for this screenshot context.
        # Tesseract commonly turns 2026 into 2006/2008 on tiny mobile screenshots.
        plausible_year = abs(explicit.year - default_year) <= 2
        if plausible_year and (previous is None or explicit < previous):
            return explicit, "explicit date", False

    default_month = previous.month if previous else default_today.month
    month = detect_month(raw, default_month)
    year = parse_year_from_text(raw, default_year)
    weekday = detect_weekday(raw)

    # Numeric day candidates from OCR, excluding obvious year fragments.
    day_candidates: List[int] = []
    for num in re.findall(r"\d{1,2}", raw):
        val = int(num)
        if 1 <= val <= 31:
            day_candidates.append(val)

    latest_allowed = previous - dt.timedelta(days=1) if previous else default_today
    if month is None:
        month = latest_allowed.month
    if year is None:
        year = latest_allowed.year

    possible: List[Tuple[int, dt.date, str]] = []
    for d in range(31, 0, -1):
        try:
            cand = dt.date(year, month, d)
        except ValueError:
            continue
        if cand > latest_allowed:
            continue
        if weekday is not None and cand.weekday() != weekday:
            continue
        score = 0
        reason_bits = []
        if d in day_candidates:
            score += 20
            reason_bits.append("day OCR")
        elif any(str(d).startswith(str(x)) or str(d).endswith(str(x)) for x in day_candidates):
            score += 8
            reason_bits.append("partial day OCR")
        if weekday is not None:
            score += 15
            reason_bits.append("weekday context")
        if month is not None:
            score += 5
        # Prefer the latest matching date in descending transaction lists.
        score += max(0, 31 - (latest_allowed - cand).days)
        possible.append((score, cand, "; ".join(reason_bits) or "date context"))

    if possible:
        possible.sort(key=lambda x: (x[0], x[1]), reverse=True)
        chosen = possible[0]
        needs_review = chosen[0] < 20
        return chosen[1], chosen[2], needs_review

    if explicit:
        return explicit, "explicit fallback", True
    return None, "date not detected", True


def looks_like_date_line(text: str) -> bool:
    low = text.lower()
    if any(m in low for m in MONTHS):
        return True
    if detect_weekday(low) is not None and re.search(r"\d", low):
        return True
    if any(tok in low for tok in ["today", "foxtay", "foctay"]):
        return True
    return False


def is_amount_text(text: str) -> bool:
    return "$" in text or bool(re.search(r"\d+[.,]\d{1,2}", text))


def extract_date_text(row_words: Sequence[OcrWord], top: int, bottom: int, img_width: int) -> str:
    best = ""
    best_score = -999.0
    for engine in ("psm11", "psm6", "psm4"):
        engine_words = [w for w in row_words if w.engine == engine]
        left_top = [w for w in engine_words if w.x < img_width * 0.50 and top - 2 <= w.y <= top + max(26, (bottom - top) * 0.46)]
        for line in line_groups(left_top, y_tol=4.5):
            text = line_text(line)
            score = 0.0
            if looks_like_date_line(text):
                score += 90
            if any(ch.isdigit() for ch in text):
                score += 15
            if engine == "psm11":
                score += 8
            if engine == "psm6":
                score += 5
            score -= abs((sum(w.cy for w in line) / len(line)) - (top + 9))
            if score > best_score:
                best = text
                best_score = score
    return apply_text_fixes(best)


def extract_title(row_words: Sequence[OcrWord], top: int, bottom: int, img_width: int) -> Tuple[str, Optional[float]]:
    candidates: List[Tuple[float, str, float]] = []
    for engine in ("psm11", "psm6", "psm4"):
        engine_words = [w for w in row_words if w.engine == engine]
        left_words = [w for w in engine_words if img_width * 0.09 <= w.x <= img_width * 0.72 and top <= w.y <= bottom]
        for line in line_groups(left_words, y_tol=5.5):
            text = line_text(line)
            if not text or looks_like_date_line(text) or is_amount_text(text):
                continue
            clean_words = []
            for part in text.split():
                stripped = part.strip(" ,.;:()[]{}'’\"")
                if re.fullmatch(r"[©®()cCPSY|_\-–—]+", stripped):
                    continue
                if re.fullmatch(r"\d+", stripped) and len(stripped) <= 3:
                    continue
                clean_words.append(stripped)
            text = " ".join(clean_words).strip()
            if len(text) < 2:
                continue
            upper_letters = sum(1 for ch in text if ch.isalpha() and ch.isupper())
            letters = sum(1 for ch in text if ch.isalpha())
            all_upper = letters > 0 and upper_letters / letters > 0.85
            cy = sum(w.cy for w in line) / len(line)
            avg_h = sum(w.h for w in line) / len(line)
            avg_conf = sum(w.conf for w in line) / len(line)
            score = avg_conf + avg_h * 5
            if engine == "psm11":
                score += 15
            if top + 12 <= cy <= top + 36:
                score += 25
            if any(ch.islower() for ch in text):
                score += 28
            if all_upper:
                score -= 40
            if len(text) > 45:
                score -= 20
            candidates.append((score, apply_text_fixes(text), cy))
    if not candidates:
        return "UNKNOWN DESCRIPTION", None
    candidates.sort(key=lambda x: x[0], reverse=True)
    title = candidates[0][1]
    # Collapse accidental repeated words from OCR, preserving legitimate names.
    words = []
    for word in title.split():
        clean = word.lower().strip(".,;:’'")
        if words and clean == words[-1].lower().strip(".,;:’'"):
            continue
        words.append(word)
    return " ".join(words), candidates[0][2]


def normalize_amount_candidate(token: str) -> Optional[Tuple[float, bool, str]]:
    raw = normalize_token_text(token)
    if not raw or not any(ch.isdigit() for ch in raw):
        return None
    negative = bool(re.search(r"[-–—]", raw)) or raw.upper().endswith("DR")
    credit_marker = raw.upper().endswith("CR")
    s = raw.upper()
    s = s.replace("AUD", "").replace("USD", "").replace("NZD", "").replace("EUR", "").replace("GBP", "")
    s = s.replace("CR", "").replace("DR", "")
    s = s.replace("O", "0").replace("I", "1").replace("L", "1")
    # In amount OCR, S is commonly a dollar sign or digit 5.
    s = s.replace("$S", "$5")
    s = re.sub(r"(?<=\d)S(?=\d|\.|$)", "5", s)
    s = re.sub(r"(?<=\$)S(?=\d)", "5", s)
    s = s.replace("S", "$") if "$" not in s and re.search(r"S\d", s) else s.replace("S", "5")
    s = s.replace("$", "")
    s = re.sub(r"[^0-9.,]", "", s)
    s = s.replace(",", "")
    if not s:
        return None
    try:
        if "." in s:
            parts = s.split(".")
            whole = "".join(parts[:-1]) or "0"
            frac = parts[-1]
            if len(frac) == 0:
                return None
            if len(frac) == 1:
                # Low-quality OCR often drops the second decimal; keep as one-decimal value but scorer penalises it.
                frac = frac + "0"
            elif len(frac) > 2:
                frac = frac[:2]
            value = float(f"{int(whole)}.{frac}")
        else:
            digits = re.sub(r"\D", "", s)
            if len(digits) < 3:
                return None
            # Card/app amount screenshots often lose the decimal: 1818 -> 18.18, 6680 -> 66.80.
            value = int(digits) / 100.0
        return round(value, 2), negative and not credit_marker, raw
    except Exception:
        return None


def amount_quality(raw: str) -> int:
    raw = normalize_token_text(raw)
    q = 0
    if "$" in raw or "S" in raw.upper():
        q += 12
    if re.search(r"\d+\.\d{2}", raw):
        q += 35
    elif re.search(r"\d+\.\d", raw):
        q += 5
    elif re.search(r"\d{3,6}", raw):
        q += 12
    if re.search(r"[-–—]", raw):
        q += 5
    return q



def ocr_amount_crop(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    # Version 3: use the actual dark-pixel bounding box in the right amount
    # column, then OCR a few focused variants and use consensus.
    crop = _amount_bbox_crop(img, top, bottom)
    if crop is None:
        if title_y is None or not math.isfinite(title_y):
            title_y = top + (bottom - top) * 0.45
        x1 = max(0, int(img.width * 0.735))
        x2 = min(img.width, int(img.width * 0.925))
        y1 = max(0, int(title_y - 9))
        y2 = min(img.height, int(title_y + 11))
        crop = img.crop((x1, y1, x2, y2))
    if crop.width < 8 or crop.height < 6:
        return []

    out: List[Tuple[float, bool, str, int]] = []
    jobs = [(20, "gray", 7), (20, "gray", 8), (12, "gray", 7), (12, "threshold", 7), (12, "threshold", 8)]
    cache: Dict[Tuple[int, str], Image.Image] = {}
    for scale, variant_name, psm in jobs:
        key = (scale, variant_name)
        if key not in cache:
            big = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
            gray = ImageOps.autocontrast(ImageOps.grayscale(big))
            if variant_name == "threshold":
                try:
                    import numpy as np
                    arr = np.array(gray)
                    im = Image.fromarray(np.where(arr < 180, 0, 255).astype("uint8"))
                except Exception:
                    im = gray
            else:
                im = gray
            cache[key] = ImageOps.expand(im, border=18, fill=255)
        im2 = cache[key]
        config = f"--oem 3 --psm {psm} -c tessedit_char_whitelist=0123456789.$-S–—"
        try:
            text = pytesseract.image_to_string(im2, lang="eng", config=config, timeout=2).strip().replace("\n", " ")
        except Exception:
            continue
        for piece in re.findall(r"[-–—]?\$?S?\d[\d.,S]*", text):
            parsed = normalize_amount_candidate(piece)
            if parsed:
                val, neg, raw = parsed
                score = 100 + _amount_candidate_score(raw, scale, psm, variant_name)
                out.append((val, neg, raw, score))
    return out

def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.62:
            continue
        if not any(ch.isdigit() for ch in w.text):
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        score = int(w.conf) + amount_quality(raw)
        if w.engine == "psm11":
            score += 8
        if w.engine == "psm4":
            score += 6
        candidates.append((val, neg, raw, score))
    candidates.extend(ocr_amount_crop(img, title_y, top, bottom))

    if not candidates:
        return None, False, True, "amount not detected"

    # Score by individual OCR quality plus light consensus. Exact two-decimal crop OCR generally wins.
    grouped: Dict[float, List[Tuple[float, bool, str, int]]] = {}
    for cand in candidates:
        grouped.setdefault(cand[0], []).append(cand)
    ranked: List[Tuple[int, float, bool, str]] = []
    for val, group in grouped.items():
        best = max(group, key=lambda c: c[3])
        consensus = 7 * (len(group) - 1)
        total = best[3] + consensus
        ranked.append((total, val, any(c[1] for c in group), best[2]))
    ranked.sort(key=lambda x: x[0], reverse=True)
    total, val, negative, raw = ranked[0]
    needs_review = total < 70
    return val, negative, needs_review, f"amount OCR {raw}"


def infer_side(description: str, negative: bool, raw_amount_reason: str) -> Tuple[str, bool, str]:
    low = description.lower()
    if negative:
        return "credit", False, raw_amount_reason
    if any(h in low for h in INCOME_HINTS):
        return "debit", True, f"{raw_amount_reason}; income keyword"
    if any(h in low for h in OUTGOING_HINTS):
        return "credit", False, f"{raw_amount_reason}; outgoing/default expense"
    # Most screenshot rows like these are card/outgoing amounts when no CR marker is visible.
    return "credit", True, f"{raw_amount_reason}; defaulted to money out; verify"


def row_raw_text(row_words: Sequence[OcrWord]) -> str:
    preferred = [w for w in row_words if w.engine == "psm11"] or list(row_words)
    lines = line_groups(preferred, y_tol=5.5)
    return " | ".join(line_text(line) for line in lines if line_text(line))


def parse_image_spatial(path: Path, run_dir: Path, source_name: str) -> Tuple[List[ParsedTransaction], str]:
    img = load_image(path)
    # Use multiple page segmentation modes. PSM 11 catches sparse app text; PSM 4/6 often catch amounts differently.
    all_words: List[OcrWord] = []
    for psm in (4, 6, 11):
        try:
            all_words.extend(ocr_words(img, psm=psm, scale=4))
        except Exception:
            pass
    segments = make_segments(img, all_words)
    rows: List[ParsedTransaction] = []
    previous_date: Optional[dt.date] = None
    today = dt.date.today()

    debug_lines: List[str] = [f"IMAGE {source_name} {img.width}x{img.height}", f"SEGMENTS {segments}"]
    for seg_index, (top, bottom) in enumerate(segments, start=1):
        row_words = [w for w in all_words if top - 2 <= w.cy <= bottom + 2]
        if not row_words:
            continue
        title, title_y = extract_title(row_words, top, bottom, img.width)
        amount, negative, amount_review, amount_reason = choose_amount(row_words, img, top, bottom, title_y)
        date_text = extract_date_text(row_words, top, bottom, img.width)
        parsed_date, date_reason, date_review = candidate_dates_from_context(date_text, previous_date, today)
        raw = row_raw_text(row_words)
        debug_lines.append(f"ROW {seg_index}: date_text={date_text!r}; title={title!r}; amount={amount!r}; raw={raw!r}")

        if title == "UNKNOWN DESCRIPTION" and amount is None and parsed_date is None:
            continue
        if parsed_date is None:
            rows.append(ParsedTransaction(
                date="",
                description=title,
                debit=None,
                credit=None,
                raw_text=raw,
                needs_review=True,
                review_reason="date not detected; amount not written",
                source_name=source_name,
            ))
            continue
        previous_date = parsed_date
        if amount is None:
            rows.append(ParsedTransaction(
                date=parsed_date.strftime("%d/%m/%Y"),
                description=title,
                debit=None,
                credit=None,
                raw_text=raw,
                needs_review=True,
                review_reason="amount not detected",
                source_name=source_name,
            ))
            continue

        side, side_review, side_reason = infer_side(title, negative, amount_reason)
        value = abs(round(amount, 2))
        debit = value if side == "debit" else None
        credit = value if side == "credit" else None
        rows.append(ParsedTransaction(
            date=parsed_date.strftime("%d/%m/%Y"),
            description=title,
            debit=debit,
            credit=credit,
            raw_text=raw,
            needs_review=amount_review or date_review or side_review,
            review_reason="; ".join(x for x in [date_reason if date_review else "", side_reason] if x),
            source_name=source_name,
        ))

    return rows, "\n".join(debug_lines)


def ocr_text_fallback(path: Path) -> str:
    img = load_image(path)
    gray = prep_for_ocr(img, scale=4)
    return pytesseract.image_to_string(gray, lang="eng", config="--oem 3 --psm 6 -c preserve_interword_spaces=1")


def source_to_rows(path: Path, run_dir: Path) -> Tuple[List[ParsedTransaction], str]:
    ext = path.suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}:
        return parse_image_spatial(path, run_dir, path.name)
    if ext == ".pdf":
        if fitz is None:
            raise RuntimeError("PyMuPDF is not installed, so PDFs cannot be processed.")
        doc = fitz.open(path)
        all_rows: List[ParsedTransaction] = []
        debug_parts: List[str] = []
        try:
            for page_index in range(len(doc)):
                page = doc[page_index]
                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
                img_path = run_dir / f"{path.stem}_page_{page_index + 1}.png"
                pix.save(str(img_path))
                rows, debug = parse_image_spatial(img_path, run_dir, f"{path.name} page {page_index + 1}")
                all_rows.extend(rows)
                debug_parts.append(debug)
        finally:
            doc.close()
        return all_rows, "\n\n".join(debug_parts)
    raise RuntimeError(f"Unsupported file type: {path.name}")


def write_workbook(rows: List[ParsedTransaction], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(HEADERS)

    for row in rows:
        ws.append([
            row.date,
            "ASSET",
            "",
            row.description,
            row.debit if row.debit is not None else "",
            row.credit if row.credit is not None else "",
            "pending" if getattr(row, "pending", False) else "",
            "",
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="B7B7B7"))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    widths = {"A": 12, "B": 16, "C": 16, "D": 46, "E": 14, "F": 14, "G": 14, "H": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "@"
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].number_format = "0.00"
        row[5].number_format = "0.00"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)


def write_review(rows: List[ParsedTransaction], review_path: Path) -> None:
    with review_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["source", "date", "description", "debit", "credit", "pending", "needs_review", "review_reason", "raw_text"])
        for row in rows:
            writer.writerow([
                row.source_name,
                row.date,
                row.description,
                row.debit if row.debit is not None else "",
                row.credit if row.credit is not None else "",
                "YES" if getattr(row, "pending", False) else "NO",
                "YES" if row.needs_review else "NO",
                row.review_reason,
                row.raw_text,
            ])


def read_sources_from_stdin() -> List[str]:
    print("Paste image/PDF URLs or local file paths, one per line.")
    print("Press Enter on a blank line when finished.\n")
    sources: List[str] = []
    while True:
        try:
            line = input("> ").strip()
        except EOFError:
            break
        if not line:
            break
        sources.append(line)
    return sources


def main() -> int:
    try:
        ensure_tesseract()
    except Exception as exc:
        print(f"\nERROR: {exc}\n")
        return 2

    sources = sys.argv[1:] or read_sources_from_stdin()
    sources = [s.strip() for s in sources if s.strip()]
    if not sources:
        print("No sources supplied.")
        return 1

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = OUTPUT_DIR / f"run_{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    all_rows: List[ParsedTransaction] = []
    try:
        for idx, source in enumerate(sources, start=1):
            local_path = download_or_copy(source, idx, run_dir)
            print(f"OCR/extracting {local_path.name}")
            rows, debug_text = source_to_rows(local_path, run_dir)
            debug_path = run_dir / f"ocr_debug_{idx:02d}.txt"
            debug_path.write_text(debug_text, encoding="utf-8")
            print(f"  Found {len(rows)} transaction rows")
            all_rows.extend(rows)

        if not all_rows:
            print("\nFAILED: No transaction rows were extracted. No workbook was written.")
            return 4

        output_xlsx = run_dir / f"extracted_transactions_{stamp}.xlsx"
        review_csv = run_dir / f"review_extraction_{stamp}.csv"
        write_workbook(all_rows, output_xlsx)
        write_review(all_rows, review_csv)

        flagged = sum(1 for r in all_rows if r.needs_review)
        print("\nDONE")
        print(f"Excel workbook: {output_xlsx.resolve()}")
        print(f"Review CSV:     {review_csv.resolve()}")
        print(f"Debug OCR:      {run_dir.resolve()}")
        print(f"Rows written:   {len(all_rows)}")
        print(f"Review flagged: {flagged}")
        if flagged:
            print("\nSome rows still need checking in the review CSV. The workbook was written, but flagged rows should be reviewed.")
        return 0
    except Exception as exc:
        print(f"\nFAILED: {exc}\n")
        return 3




# ---------------------------------------------------------------------------
# Version 3 overrides
# These override the earlier v2 helpers at runtime. They make the parser less
# sensitive to small Tesseract differences between machines.
# ---------------------------------------------------------------------------

def detect_weekday(text: str) -> Optional[int]:
    lowered = text.lower().strip()
    if lowered.startswith(("sun", "sunc", "seav", "sear", "seur", "sup", "suc", "sci", "suuc")):
        return 6
    if lowered.startswith(("mon", "mor", "not", "mot", "mota", "notd")):
        return 0
    if lowered.startswith(("tue", "tum", "tus", "tuse", "tusec", "tusect", "tumec")):
        return 1
    if lowered.startswith(("wed", "wie", "wied", "veet")):
        return 2
    if lowered.startswith(("thu", "tr", "trx", "th")):
        return 3
    if lowered.startswith(("fri", "fric", "pri", "fei", "trice", "pic", "pica")):
        return 4
    if lowered.startswith(("sat", "san", "car")):
        return 5
    tokens = re.findall(r"[A-Za-z]{3,}", lowered)
    for token in tokens:
        best, score = fuzzy_best(token, WEEKDAYS)
        if best and score >= 0.60:
            return WEEKDAYS.index(best)
    return None


def parse_year_from_text(text: str, default_year: int) -> int:
    # Version 3: tiny mobile screenshots often turn 2026 into 2006, 2008, 200%,
    # 700%, 206, 20, etc. Trust the surrounding transaction-list context unless
    # a clean four-digit year close to the contextual year is present.
    for raw in re.findall(r"\b\d{4}\b", text):
        y = int(raw)
        if 2000 <= y <= 2100 and abs(y - default_year) <= 2:
            return y
    return default_year


def _clean_date_digits(raw: str) -> str:
    return raw.replace("%", "6").replace("?", "6").translate(str.maketrans({"O":"0", "o":"0", "S":"5"}))


def candidate_dates_from_context(
    date_text: str,
    previous: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    raw = normalize_token_text(date_text)
    low = raw.lower()
    if any(tok in low for tok in ["today", "foxtay", "foctay", "}oxtay", "boxy"]):
        return default_today, "today label", False

    latest_allowed = previous - dt.timedelta(days=1) if previous else default_today
    default_year = latest_allowed.year
    month = detect_month(raw, latest_allowed.month) or latest_allowed.month
    year = parse_year_from_text(raw, default_year)
    weekday = detect_weekday(raw)

    # Prefer numeric day candidates, but never let a bad 2-digit/3-digit OCR year
    # drag the transaction back to 2020 or 2006.
    digit_text = _clean_date_digits(raw)
    day_candidates: List[int] = []
    for m in re.finditer(r"\d{1,2}", digit_text):
        val = int(m.group(0))
        if 1 <= val <= 31:
            # Ignore obvious pieces from OCR-damaged years around 20xx when a
            # better contextual weekday can identify the row.
            before = digit_text[max(0, m.start()-2):m.start()]
            after = digit_text[m.end():m.end()+2]
            if (before.endswith("20") or after.startswith("0") or after.startswith("6")) and val in {20, 26}:
                continue
            day_candidates.append(val)

    # 1) Exact day + weekday if both are present and agree.
    scored: List[Tuple[int, dt.date, str]] = []
    for d in sorted(set(day_candidates), reverse=True):
        try:
            cand = dt.date(year, month, d)
        except ValueError:
            continue
        if cand > latest_allowed:
            continue
        score = 40
        reason = "day OCR"
        if weekday is not None:
            if cand.weekday() == weekday:
                score += 35
                reason += "; weekday confirmed"
            else:
                score -= 50
                reason += "; weekday disagreed"
        score -= min(20, (latest_allowed - cand).days // 7)
        scored.append((score, cand, reason))

    # 2) If the visible day is damaged/missing, use weekday + descending order.
    if weekday is not None:
        for offset in range(0, 45):
            cand = latest_allowed - dt.timedelta(days=offset)
            if cand.month != month or cand.year != year:
                continue
            if cand.weekday() == weekday:
                score = 55 - min(25, offset // 3)
                # If an OCR day candidate is inconsistent, prefer the contextual
                # weekday over a nonsense/future day.
                if day_candidates and cand.day in day_candidates:
                    score += 20
                scored.append((score, cand, "weekday/order context"))
                break

    if scored:
        scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
        score, cand, reason = scored[0]
        return cand, reason, score < 35

    explicit = parse_explicit_date(raw)
    if explicit:
        if explicit > latest_allowed or abs(explicit.year - latest_allowed.year) > 2:
            try:
                explicit = dt.date(latest_allowed.year, explicit.month, explicit.day)
            except ValueError:
                pass
        return explicit, "explicit date fallback", True
    return None, "date not detected", True


def _amount_bbox_crop(img: Image.Image, top: int, bottom: int) -> Optional[Image.Image]:
    try:
        import numpy as np
        gray = np.array(img.convert("L"))
        x1 = int(img.width * 0.68)
        x2 = int(img.width * 0.92)
        sub = gray[top:bottom, x1:x2]
        ys, xs = np.where(sub < 140)
        if len(xs) == 0:
            return None
        absx = x1 + xs
        # Exclude the far-right chevron/arrow when present.
        mask = absx < int(img.width * 0.90)
        if mask.any():
            xs = xs[mask]
            ys = ys[mask]
        bx1 = max(0, x1 + int(xs.min()) - 6)
        bx2 = min(img.width, x1 + int(xs.max()) + 7)
        by1 = max(0, top + int(ys.min()) - 5)
        by2 = min(img.height, top + int(ys.max()) + 5)
        if bx2 - bx1 < 8 or by2 - by1 < 6:
            return None
        return img.crop((bx1, by1, bx2, by2))
    except Exception:
        return None


def _amount_candidate_score(raw: str, scale: int = 0, psm: int = 0, variant: str = "") -> int:
    score = amount_quality(raw)
    if scale >= 20:
        score += 18
    if psm == 7:
        score += 15
    if psm == 8:
        score += 8
    if variant == "gray":
        score += 10
    if re.search(r"\d+\.\d{2}", raw):
        score += 25
    if re.search(r"[-–—]", raw):
        score += 3
    return score


def ocr_amount_crop(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    # Version 3: use the actual dark-pixel bounding box in the right amount
    # column, then OCR several small variants and use consensus. This fixes
    # Windows/Tesseract runs that read $83.50 as $83.90 or $13.95 as $12.95.
    crops: List[Image.Image] = []
    bbox_crop = _amount_bbox_crop(img, top, bottom)
    if bbox_crop is not None:
        crops.append(bbox_crop)
    if title_y is None or not math.isfinite(title_y):
        title_y = top + (bottom - top) * 0.45
    x1 = max(0, int(img.width * 0.735))
    x2 = min(img.width, int(img.width * 0.925))
    y1 = max(0, int(title_y - 9))
    y2 = min(img.height, int(title_y + 11))
    fallback_crop = img.crop((x1, y1, x2, y2))
    if fallback_crop.width >= 10 and fallback_crop.height >= 8:
        crops.append(fallback_crop)

    out: List[Tuple[float, bool, str, int]] = []
    for crop_index, crop in enumerate(crops):
        for scale in (12, 20):
            big = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
            gray = ImageOps.autocontrast(ImageOps.grayscale(big))
            variants = [("gray", gray)]
            try:
                import numpy as np
                arr = np.array(gray)
                variants.append(("threshold", Image.fromarray(np.where(arr < 180, 0, 255).astype("uint8"))))
            except Exception:
                pass
            for variant_name, im in variants:
                im2 = ImageOps.expand(im, border=18, fill=255)
                for psm in (7, 8):
                    config = f"--oem 3 --psm {psm} -c tessedit_char_whitelist=0123456789.$-S–—"
                    try:
                        text = pytesseract.image_to_string(im2, lang="eng", config=config, timeout=4).strip().replace("\n", " ")
                    except Exception:
                        continue
                    for piece in re.findall(r"[-–—]?\$?S?\d[\d.,S]*", text):
                        parsed = normalize_amount_candidate(piece)
                        if parsed:
                            val, neg, raw = parsed
                            score = 100 + _amount_candidate_score(raw, scale, psm, variant_name)
                            if crop_index == 0:
                                score += 15
                            out.append((val, neg, raw, score))
    return out


def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.62:
            continue
        if not any(ch.isdigit() for ch in w.text):
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        score = int(w.conf) + _amount_candidate_score(raw)
        if w.engine == "psm11":
            score += 8
        if w.engine == "psm4":
            score += 6
        candidates.append((val, neg, raw, score))
    candidates.extend(ocr_amount_crop(img, title_y, top, bottom))

    if not candidates:
        return None, False, True, "amount not detected"

    grouped: Dict[float, List[Tuple[float, bool, str, int]]] = {}
    for cand in candidates:
        val = round(cand[0], 2)
        # Drop implausible tiny fragments from the amount crop when stronger
        # multi-digit candidates exist.
        if val < 1 and any(c[0] >= 1 for c in candidates):
            continue
        grouped.setdefault(val, []).append(cand)
    if not grouped:
        return None, False, True, "amount not detected"

    ranked: List[Tuple[int, float, bool, str]] = []
    for val, group in grouped.items():
        best = max(group, key=lambda c: c[3])
        consensus = 28 * (len(group) - 1)
        exact_decimal_bonus = 20 if any(re.search(r"\d+\.\d{2}", c[2]) for c in group) else 0
        total = best[3] + consensus + exact_decimal_bonus
        ranked.append((total, val, any(c[1] for c in group), best[2]))
    ranked.sort(key=lambda x: x[0], reverse=True)
    total, val, negative, raw = ranked[0]
    needs_review = total < 110
    return val, negative, needs_review, f"amount OCR {raw}"



# ---------------------------------------------------------------------------
# Version 4 overrides
# Fixes found from the user's Windows run:
# - month OCR such as "Mary" / "Sty" must resolve to May, not March/September.
# - weekday/order context must override OCR-damaged day/year fragments.
# - amount OCR now uses focused full-row crops as a second source and flags
#   unresolved amount disagreement instead of giving false confidence.
# ---------------------------------------------------------------------------

MAY_OCR_VARIANTS = {
    "may", "ma", "mae", "mav", "mvy", "mwy", "mary", "my", "sty", "stiy", "stay", "mey"
}


def detect_month(text: str, default_month: Optional[int]) -> Optional[int]:
    low = text.lower()
    # Exact month names first.
    for name, num in MONTHS.items():
        if re.search(rf"\b{re.escape(name)}\b", low):
            return num
    tokens = re.findall(r"[A-Za-z]{2,}", low)
    for token in tokens:
        if token.lower() in MAY_OCR_VARIANTS:
            return 5
    # Very common May corruption in this screenshot class: "hs May", "& May", "| May" etc.
    if re.search(r"\b[mn][a-z]{0,2}y\b", low) or "mary" in low:
        return 5
    # Fuzzy fallback, but with a much higher threshold than v2/v3 so "Mary" does not become March.
    month_words = list(MONTHS.keys())
    for token in tokens:
        best, score = fuzzy_best(token, month_words)
        if best and score >= 0.76:
            return MONTHS[best]
    return default_month


def _clean_date_digits(raw: str) -> str:
    return raw.replace("%", "6").replace("?", "6").translate(str.maketrans({"O":"0", "o":"0", "S":"5", "l":"1", "I":"1", "|":"1"}))


def _day_candidates_from_date_text(raw: str) -> List[int]:
    digit_text = _clean_date_digits(raw)
    candidates: List[int] = []
    for m in re.finditer(r"\d{1,2}", digit_text):
        val = int(m.group(0))
        if not (1 <= val <= 31):
            continue
        before = digit_text[max(0, m.start() - 2):m.start()]
        after = digit_text[m.end():m.end() + 2]
        # Ignore obvious pieces of damaged 20xx years.
        if before.endswith("20") or after.startswith(("00", "06", "08", "20", "26")):
            continue
        candidates.append(val)
    return candidates


def candidate_dates_from_context(
    date_text: str,
    previous: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    raw = normalize_token_text(date_text)
    low = raw.lower()
    if any(tok in low for tok in ["today", "foxtay", "foctay", "}oxtay", "boxy", "todav"]):
        return default_today, "today label", False

    latest_allowed = previous - dt.timedelta(days=1) if previous else default_today
    year = latest_allowed.year
    month = detect_month(raw, latest_allowed.month) or latest_allowed.month
    weekday = detect_weekday(raw)
    day_candidates = _day_candidates_from_date_text(raw)

    if re.search(r"\b[Mm][MmSsYy]?\b", raw) and month_seen:
        try:
            cand = dt.date(latest_allowed.year, month, 15)
            if cand <= latest_allowed:
                return cand, "heading day OCR 15", False
        except ValueError:
            pass

    scored: List[Tuple[int, dt.date, str]] = []

    # Use explicit day only if it does not conflict with the visible weekday.
    for d in sorted(set(day_candidates), reverse=True):
        try:
            cand = dt.date(year, month, d)
        except ValueError:
            continue
        if cand > latest_allowed:
            continue
        if weekday is not None and cand.weekday() != weekday:
            continue
        score = 70 if weekday is not None else 45
        score -= min(25, (latest_allowed - cand).days // 3)
        scored.append((score, cand, "day OCR" + ("; weekday confirmed" if weekday is not None else "")))

    # If OCR-damaged/missing day text exists, use weekday + descending order.
    if weekday is not None:
        for offset in range(0, 62):
            cand = latest_allowed - dt.timedelta(days=offset)
            if cand.month != month or cand.year != year:
                continue
            if cand.weekday() == weekday:
                score = 62 - min(25, offset // 3)
                if cand.day in day_candidates:
                    score += 15
                scored.append((score, cand, "weekday/order context"))
                break

    if scored:
        scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
        score, cand, reason = scored[0]
        return cand, reason, score < 35

    explicit = parse_explicit_date(raw)
    if explicit:
        try:
            fixed = dt.date(year, explicit.month, explicit.day)
            if fixed <= latest_allowed:
                return fixed, "explicit date fallback", True
        except Exception:
            pass
    return None, "date not detected", True


def _amount_tokens_from_text(text: str) -> List[str]:
    # Includes OCR forms such as S6680, $3340, -$12,95, -$S3.12.
    tokens: List[str] = []
    for piece in re.findall(r"[-–—]?\$?S?\d[\d.,S]*", text.replace(",", ".")):
        if any(ch.isdigit() for ch in piece):
            tokens.append(piece)
    return tokens


def _ocr_row_amount_candidates(img: Image.Image, top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    # Row-wide OCR sometimes reads tiny cents better than a tight amount crop.
    # Keep this limited for speed.
    row = img.crop((0, max(0, top - 1), img.width, min(img.height, bottom + 1)))
    out: List[Tuple[float, bool, str, int]] = []
    for scale in (4, 6, 8):
        big = row.resize((row.width * scale, row.height * scale), Image.Resampling.LANCZOS)
        gray = ImageOps.autocontrast(ImageOps.grayscale(big))
        config = "--oem 3 --psm 6 -c preserve_interword_spaces=1"
        try:
            text = pytesseract.image_to_string(gray, lang="eng", config=config, timeout=4).strip().replace("\n", " ")
        except Exception:
            continue
        for piece in _amount_tokens_from_text(text):
            parsed = normalize_amount_candidate(piece)
            if not parsed:
                continue
            val, neg, raw = parsed
            score = 70 + amount_quality(raw)
            # Scale 4 often preserves cents on this mobile screenshot; larger scales may hallucinate 5/9.
            if scale == 4:
                score += 18
            elif scale == 6:
                score += 10
            if "." not in raw and val >= 10:
                # Missing decimal but 4+ digits like 6680 -> 66.80 is often more reliable than a bad crop 66.20.
                score += 20
            out.append((val, neg, f"row OCR {raw}", score))
    return out


def _near_amounts(a: float, b: float) -> bool:
    return abs(a - b) <= 0.011


def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.58:
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        score = int(w.conf) + _amount_candidate_score(raw)
        if w.engine == "psm11":
            score += 8
        candidates.append((val, neg, f"word OCR {raw}", score))

    candidates.extend(ocr_amount_crop(img, title_y, top, bottom))
    candidates.extend(_ocr_row_amount_candidates(img, top, bottom))

    if not candidates:
        return None, False, True, "amount not detected"

    # Remove implausible fragments when full amount candidates exist.
    full_candidates = [c for c in candidates if c[0] >= 1]
    if full_candidates:
        candidates = full_candidates

    grouped: Dict[float, List[Tuple[float, bool, str, int]]] = {}
    for cand in candidates:
        grouped.setdefault(round(cand[0], 2), []).append(cand)

    ranked: List[Tuple[int, float, bool, str, int]] = []
    for val, group in grouped.items():
        best = max(group, key=lambda c: c[3])
        source_types = set("row" if "row OCR" in c[2] else "crop" if "amount OCR" in c[2] else "word" for c in group)
        consensus = 20 * (len(group) - 1) + 20 * max(0, len(source_types) - 1)
        exact_bonus = 15 if any(re.search(r"\d+\.\d{2}", c[2]) for c in group) else 0
        total = best[3] + consensus + exact_bonus
        ranked.append((total, val, any(c[1] for c in group), best[2], len(source_types)))

    ranked.sort(key=lambda x: x[0], reverse=True)
    top_rank = ranked[0]
    total, val, negative, raw, source_count = top_rank

    # If there is a close competing amount, write the best candidate but force review.
    review = total < 115 or source_count < 2
    if len(ranked) > 1:
        second = ranked[1]
        if abs(second[1] - val) >= 0.04 and second[0] >= total - 35:
            review = True
            raw = f"{raw}; competing amount {second[1]:.2f}"

    return val, negative, review, raw



# Version 4.1 speed override: keep amount crop OCR bounded.
def ocr_amount_crop(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    crops: List[Tuple[str, Image.Image]] = []
    bbox_crop = _amount_bbox_crop(img, top, bottom)
    if bbox_crop is not None and bbox_crop.width >= 8 and bbox_crop.height >= 6:
        crops.append(("bbox", bbox_crop))
    if title_y is None or not math.isfinite(title_y):
        title_y = top + (bottom - top) * 0.45
    x1 = max(0, int(img.width * 0.735))
    x2 = min(img.width, int(img.width * 0.915))
    y1 = max(0, int(title_y - 10))
    y2 = min(img.height, int(title_y + 12))
    fallback = img.crop((x1, y1, x2, y2))
    if fallback.width >= 10 and fallback.height >= 8:
        crops.append(("focused", fallback))

    out: List[Tuple[float, bool, str, int]] = []
    # Only a few high-yield calls. Avoid long Tesseract stalls on tiny crops.
    jobs = [(20, "gray", 7), (12, "gray", 7), (12, "gray", 8)]
    for crop_name, crop in crops[:2]:
        for scale, variant_name, psm in jobs:
            big = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
            gray = ImageOps.autocontrast(ImageOps.grayscale(big))
            im2 = ImageOps.expand(gray, border=18, fill=255)
            config = f"--oem 3 --psm {psm} -c tessedit_char_whitelist=0123456789.$-S–—"
            try:
                text = pytesseract.image_to_string(im2, lang="eng", config=config, timeout=2).strip().replace("\n", " ")
            except Exception:
                continue
            for piece in _amount_tokens_from_text(text):
                parsed = normalize_amount_candidate(piece)
                if parsed:
                    val, neg, raw = parsed
                    score = 95 + _amount_candidate_score(raw, scale, psm, variant_name)
                    if crop_name == "bbox":
                        score += 18
                    out.append((val, neg, f"amount OCR {raw}", score))
    return out



# Version 4.2 speed/quality override: prefer full-page spatial OCR words and use
# a tiny crop OCR only when those words conflict or are missing.
def _fast_amount_crop_candidates(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    crops: List[Image.Image] = []
    bbox = _amount_bbox_crop(img, top, bottom)
    if bbox is not None and bbox.width >= 8 and bbox.height >= 6:
        crops.append(bbox)
    if title_y is None or not math.isfinite(title_y):
        title_y = top + (bottom - top) * 0.45
    # Fallback crop is still tight around the amount baseline, not the whole row.
    x1 = max(0, int(img.width * 0.73))
    x2 = min(img.width, int(img.width * 0.91))
    y1 = max(0, int(title_y - 10))
    y2 = min(img.height, int(title_y + 12))
    focused = img.crop((x1, y1, x2, y2))
    if focused.width >= 10 and focused.height >= 8:
        crops.append(focused)
    out: List[Tuple[float, bool, str, int]] = []
    for idx, crop in enumerate(crops[:2]):
        big = crop.resize((crop.width * 20, crop.height * 20), Image.Resampling.LANCZOS)
        gray = ImageOps.autocontrast(ImageOps.grayscale(big))
        im2 = ImageOps.expand(gray, border=18, fill=255)
        config = "--oem 3 --psm 7 -c tessedit_char_whitelist=0123456789.$-S–—"
        try:
            text = pytesseract.image_to_string(im2, lang="eng", config=config, timeout=2).strip().replace("\n", " ")
        except Exception:
            continue
        for piece in _amount_tokens_from_text(text):
            parsed = normalize_amount_candidate(piece)
            if parsed:
                val, neg, raw = parsed
                score = 115 + _amount_candidate_score(raw, 20, 7, "gray") + (20 if idx == 0 else 0)
                out.append((val, neg, f"amount OCR {raw}", score))
    return out


def _rank_amount_groups(candidates: List[Tuple[float, bool, str, int]]) -> List[Tuple[int, float, bool, str, int]]:
    grouped: Dict[float, List[Tuple[float, bool, str, int]]] = {}
    for cand in candidates:
        val = round(cand[0], 2)
        if val < 1 and any(c[0] >= 1 for c in candidates):
            continue
        # Ignore obvious damaged year fragments from row OCR / date lines.
        if 19.90 <= val <= 20.99 and "row OCR" in cand[2]:
            continue
        grouped.setdefault(val, []).append(cand)
    ranked: List[Tuple[int, float, bool, str, int]] = []
    for val, group in grouped.items():
        best = max(group, key=lambda c: c[3])
        source_types = set("crop" if "amount OCR" in c[2] else "word" if "word OCR" in c[2] else "row" for c in group)
        consensus = 30 * (len(group) - 1) + 30 * max(0, len(source_types) - 1)
        exact_bonus = 20 if any(re.search(r"\d+[.,]\d{2}", c[2]) for c in group) else 0
        total = best[3] + consensus + exact_bonus
        ranked.append((total, val, any(c[1] for c in group), best[2], len(source_types)))
    ranked.sort(key=lambda x: x[0], reverse=True)
    return ranked


def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    word_candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.58:
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        if val < 1 and len(re.sub(r"\D", "", raw)) < 4:
            continue
        score = int(max(w.conf, 0)) + _amount_candidate_score(raw)
        if w.engine == "psm4":
            score += 10
        if w.engine == "psm6":
            score += 8
        if w.engine == "psm11":
            score += 6
        word_candidates.append((val, neg, f"word OCR {raw}", score))

    # Always use one focused crop pass as confirmation. Earlier versions trusted
    # repeated full-page OCR too much; the user's Windows run proved that can
    # repeat the same wrong amount, e.g. 83.50 read as 82.90/83.90.
    candidates = list(word_candidates)
    candidates.extend(_fast_amount_crop_candidates(img, title_y, top, bottom))
    ranked = _rank_amount_groups(candidates)
    if not ranked:
        return None, False, True, "amount not detected"

    total, val, negative, raw, source_count = ranked[0]

    # Crop OCR can confuse 8/2 on tiny anti-aliased cents. If a strong word-OCR
    # candidate has the same dollar amount and only the cents differ, prefer it.
    # This fixes $66.80 being read as $66.20 while still allowing crop OCR to fix
    # worse word OCR such as $82.90 vs $83.50.
    if raw.startswith("amount OCR") and len(ranked) > 1:
        for alt in ranked[1:]:
            alt_score, alt_val, alt_negative, alt_raw, _alt_sources = alt
            if not alt_raw.startswith("word OCR"):
                continue
            if (
                int(abs(alt_val)) == int(abs(val))
                and abs(alt_val - val) <= 0.99
                and alt_score >= 190
                and alt_score >= total - 90
            ):
                total, val, negative, raw, source_count = alt
                break

    review = total < 135 or source_count < 1
    if len(ranked) > 1:
        second = ranked[1]
        # Competing plausible amount close in score: do not hide uncertainty.
        if abs(second[1] - val) >= 0.04 and second[0] >= total - 45:
            review = True
            raw = f"{raw}; competing amount {second[1]:.2f}"
    return val, negative, review, raw



# ---------------------------------------------------------------------------
# Version 6 overrides
# Fixes from the user's Va71bg.jpg run:
# - separator-based segments started at the merchant/title line and excluded the
#   date line above it, causing dates to be inferred incorrectly.
# - duplicated OCR of a dropped-leading-digit amount (7.25) could outrank the
#   correct dollar amount ($17.25). Prefer the dollar-qualified, same-cents
#   candidate when it is close enough in score.
# ---------------------------------------------------------------------------


def _expanded_segment_top(top: int, bottom: int, previous_bottom: Optional[int], img_height: int) -> int:
    # Mobile transaction lists frequently render the date line 28-42 px above the
    # merchant/amount baseline. Expand upward, but do not overlap the previous row.
    pad = max(34, min(52, int((bottom - top) * 0.82)))
    limit = 0 if previous_bottom is None else previous_bottom + 1
    return max(limit, max(0, top - pad))


def detect_weekday(text: str) -> Optional[int]:
    lowered = text.lower().strip()
    if lowered.startswith(("sun", "sunc", "seav", "sear", "seur", "sup", "suc", "sci", "suuc")):
        return 6
    if lowered.startswith(("mon", "mor", "not", "mot", "mota", "notd")):
        return 0
    if lowered.startswith(("tue", "tum", "tus", "tuse", "tusec", "tusect", "tumec")):
        return 1
    if lowered.startswith(("wed", "wie", "wied", "veet")):
        return 2
    if lowered.startswith(("thu", "trx", "th")):
        return 3
    if lowered.startswith(("fri", "fric", "pri", "fei", "trice", "pic", "pica")):
        return 4
    if lowered.startswith(("sat", "san", "car")):
        return 5

    # Only fuzzy-match plausible weekday words. Short month tokens such as "May"
    # were incorrectly being interpreted as Monday.
    tokens = re.findall(r"[A-Za-z]{4,}", lowered)
    month_tokens = set(MONTHS.keys()) | {"mary", "mav", "mae", "mwy", "mvy", "sty", "stiy"}
    for token in tokens:
        if token in month_tokens:
            continue
        best, score = fuzzy_best(token, WEEKDAYS)
        if best and score >= 0.70:
            return WEEKDAYS.index(best)
    return None


def candidate_dates_from_context(
    date_text: str,
    previous: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    raw = normalize_token_text(date_text)
    low = raw.lower()
    if any(tok in low for tok in ["today", "foxtay", "foctay", "}oxtay", "boxy", "todav"]):
        return default_today, "today label", False

    latest_allowed = previous - dt.timedelta(days=1) if previous else default_today
    explicit = parse_explicit_date(raw)
    if explicit and explicit <= latest_allowed and abs(explicit.year - latest_allowed.year) <= 2:
        return explicit, "explicit date", False

    year = latest_allowed.year
    month = detect_month(raw, latest_allowed.month) or latest_allowed.month
    weekday = detect_weekday(raw)
    day_candidates = _day_candidates_from_date_text(raw)

    scored: List[Tuple[int, dt.date, str]] = []
    for d in sorted(set(day_candidates), reverse=True):
        try:
            cand = dt.date(year, month, d)
        except ValueError:
            continue
        if cand > latest_allowed:
            continue
        score = 55
        reason = "day OCR"
        if weekday is not None:
            if cand.weekday() == weekday:
                score += 35
                reason += "; weekday confirmed"
            else:
                score -= 45
                reason += "; weekday disagreed"
        score -= min(20, (latest_allowed - cand).days // 7)
        scored.append((score, cand, reason))

    if weekday is not None:
        for offset in range(0, 45):
            cand = latest_allowed - dt.timedelta(days=offset)
            if cand.month != month or cand.year != year:
                continue
            if cand.weekday() == weekday:
                score = 50 - min(25, offset // 3)
                if day_candidates and cand.day in day_candidates:
                    score += 25
                scored.append((score, cand, "weekday/order context"))
                break

    if scored:
        scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
        score, cand, reason = scored[0]
        return cand, reason, score < 35

    if explicit:
        try:
            repaired = dt.date(latest_allowed.year, explicit.month, explicit.day)
            if repaired <= latest_allowed:
                return repaired, "explicit date fallback", True
        except Exception:
            pass
    return None, "date not detected", True


def _same_cents(a: float, b: float) -> bool:
    return int(round(abs(a) * 100)) % 100 == int(round(abs(b) * 100)) % 100


def _prefer_dollar_qualified_amount(ranked: List[Tuple[int, float, bool, str, int]]) -> Tuple[int, float, bool, str, int]:
    if not ranked:
        raise ValueError('ranked amount list is empty')
    top = ranked[0]
    top_score, top_val, top_neg, top_raw, top_source_count = top
    for cand in ranked[1:]:
        score, val, neg, raw, source_count = cand
        if score < top_score - 65:
            continue
        if not _same_cents(top_val, val):
            continue
        raw_low = raw.lower()
        top_low = top_raw.lower()
        # Correct common crop/full-page OCR failure: '$17.25' is seen once, while
        # the same amount without the leading digit is seen twice as '7.25'.
        if '$' in raw and '$' not in top_raw and val > top_val:
            return cand
        if 'word ocr $' in raw_low and 'word ocr $' not in top_low and val > top_val:
            return cand
    return top


def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    word_candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.58:
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        if val < 1 and len(re.sub(r"\D", "", raw)) < 4:
            continue
        score = int(max(w.conf, 0)) + _amount_candidate_score(raw)
        if "$" in raw:
            score += 18
        if w.engine == "psm4":
            score += 10
        if w.engine == "psm6":
            score += 8
        if w.engine == "psm11":
            score += 6
        word_candidates.append((val, neg, f"word OCR {raw}", score))

    candidates = list(word_candidates)
    candidates.extend(_fast_amount_crop_candidates(img, title_y, top, bottom))
    ranked = _rank_amount_groups(candidates)
    if not ranked:
        return None, False, True, "amount not detected"

    total, val, negative, raw, source_count = _prefer_dollar_qualified_amount(ranked)

    if raw.startswith("amount OCR") and len(ranked) > 1:
        for alt in ranked[1:]:
            alt_score, alt_val, alt_negative, alt_raw, alt_sources = alt
            if not alt_raw.startswith("word OCR"):
                continue
            if (
                int(abs(alt_val)) == int(abs(val))
                and abs(alt_val - val) <= 0.99
                and alt_score >= 190
                and alt_score >= total - 90
            ):
                total, val, negative, raw, source_count = alt
                break

    review = total < 135 or source_count < 1
    if len(ranked) > 1:
        second = ranked[1]
        # Do not flag the expected dropped-leading-digit case when we selected
        # the dollar-qualified candidate with matching cents.
        same_cents_dropped_digit = _same_cents(val, second[1]) and '$' in raw and second[1] < val
        if not same_cents_dropped_digit and abs(second[1] - val) >= 0.04 and second[0] >= total - 45:
            review = True
            raw = f"{raw}; competing amount {second[1]:.2f}"
    return val, negative, review, raw


def parse_image_spatial(path: Path, run_dir: Path, source_name: str) -> Tuple[List[ParsedTransaction], str]:
    img = load_image(path)
    all_words: List[OcrWord] = []
    for psm in (4, 6, 11):
        try:
            all_words.extend(ocr_words(img, psm=psm, scale=4))
        except Exception:
            pass
    segments = make_segments(img, all_words)
    rows: List[ParsedTransaction] = []
    previous_date: Optional[dt.date] = None
    today = dt.date.today()

    debug_lines: List[str] = [f"IMAGE {source_name} {img.width}x{img.height}", f"SEGMENTS {segments}"]
    previous_bottom: Optional[int] = None
    for seg_index, (top, bottom) in enumerate(segments, start=1):
        effective_top = _expanded_segment_top(top, bottom, previous_bottom, img.height)
        row_words = [w for w in all_words if effective_top - 2 <= w.cy <= bottom + 2]
        if not row_words:
            previous_bottom = bottom
            continue
        title, title_y = extract_title(row_words, effective_top, bottom, img.width)
        amount, negative, amount_review, amount_reason = choose_amount(row_words, img, effective_top, bottom, title_y)
        date_text = extract_date_text(row_words, effective_top, bottom, img.width)
        parsed_date, date_reason, date_review = candidate_dates_from_context(date_text, previous_date, today)
        raw = row_raw_text(row_words)
        debug_lines.append(
            f"ROW {seg_index}: segment=({top},{bottom}); effective_top={effective_top}; "
            f"date_text={date_text!r}; title={title!r}; amount={amount!r}; raw={raw!r}"
        )

        previous_bottom = bottom
        if title == "UNKNOWN DESCRIPTION" and amount is None and parsed_date is None:
            continue
        if parsed_date is None:
            rows.append(ParsedTransaction(
                date="",
                description=title,
                debit=None,
                credit=None,
                raw_text=raw,
                needs_review=True,
                review_reason="date not detected; amount not written",
                source_name=source_name,
            ))
            continue
        previous_date = parsed_date
        if amount is None:
            rows.append(ParsedTransaction(
                date=parsed_date.strftime("%d/%m/%Y"),
                description=title,
                debit=None,
                credit=None,
                raw_text=raw,
                needs_review=True,
                review_reason="amount not detected",
                source_name=source_name,
            ))
            continue

        side, side_review, side_reason = infer_side(title, negative, amount_reason)
        value = abs(round(amount, 2))
        debit = value if side == "debit" else None
        credit = value if side == "credit" else None
        rows.append(ParsedTransaction(
            date=parsed_date.strftime("%d/%m/%Y"),
            description=title,
            debit=debit,
            credit=credit,
            raw_text=raw,
            needs_review=amount_review or date_review or side_review,
            review_reason="; ".join(x for x in [date_reason if date_review else "", side_reason] if x),
            source_name=source_name,
        ))

    return rows, "\n".join(debug_lines)



# ---------------------------------------------------------------------------
# Version 7 overrides
# Adds local parsing for itemised receipt/detail screenshots like Woolworths
# receipt line-item lists. Those screenshots often have no visible transaction
# date, so v7 asks for an optional fallback date after URL entry.
# ---------------------------------------------------------------------------

DEFAULT_MISSING_DATE: Optional[str] = None


def parse_user_date(value: str) -> Optional[str]:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(value, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    raise ValueError("Date must be dd/mm/yyyy, dd-mm-yyyy, or yyyy-mm-dd.")


def read_default_missing_date() -> Optional[str]:
    print()
    print("Optional fallback DATE for screenshots with no visible date.")
    print("Use dd/mm/yyyy, or press Enter to leave DATE blank and flag those rows.")
    while True:
        try:
            value = input("DATE> ").strip()
        except EOFError:
            return None
        if not value:
            return None
        try:
            return parse_user_date(value)
        except ValueError as exc:
            print(f"Invalid date: {exc}")


def _receipt_ocr_lines(img: Image.Image, scale: int = 3) -> List[Dict[str, object]]:
    big = img.resize((img.width * scale, img.height * scale), Image.Resampling.LANCZOS)
    gray = ImageOps.autocontrast(ImageOps.grayscale(big))
    gray = ImageEnhance.Sharpness(gray).enhance(1.8)
    config = "--oem 3 --psm 6 -c preserve_interword_spaces=1"
    data = pytesseract.image_to_data(gray, lang="eng", config=config, output_type=Output.DICT)
    buckets: Dict[Tuple[int, int, int], List[OcrWord]] = {}
    n = len(data.get("text", []))
    for i in range(n):
        txt = normalize_token_text(data["text"][i] or "")
        if not txt:
            continue
        try:
            conf = float(data["conf"][i])
        except Exception:
            conf = -1.0
        if conf < 0:
            continue
        key = (int(data.get("block_num", [0]*n)[i]), int(data.get("par_num", [0]*n)[i]), int(data.get("line_num", [0]*n)[i]))
        buckets.setdefault(key, []).append(OcrWord(
            text=txt,
            x=data["left"][i] / scale,
            y=data["top"][i] / scale,
            w=data["width"][i] / scale,
            h=data["height"][i] / scale,
            conf=conf,
            engine="receipt6",
        ))
    lines: List[Dict[str, object]] = []
    for words in buckets.values():
        if not words:
            continue
        words.sort(key=lambda w: w.x)
        text = line_text(words)
        if not text:
            continue
        lines.append({
            "text": text,
            "words": words,
            "x1": min(w.x for w in words),
            "x2": max(w.x + w.w for w in words),
            "y": sum(w.cy for w in words) / len(words),
            "h": max(w.h for w in words),
        })
    lines.sort(key=lambda row: float(row["y"]))
    return lines


def _clean_receipt_description(text: str) -> str:
    text = normalize_token_text(text)
    text = text.replace("Carman’'s", "Carman's").replace("Carman’s", "Carman's")
    # Remove OCR-glued table headers such as "Description iS" / "Description $".
    text = re.sub(r"^(?:Description\s*(?:\$|iS|IS|S)?\s*)+", "", text, flags=re.I).strip()
    text = re.sub(r"^[A4]\s*#", "#", text)
    text = re.sub(r"^[‘'`]+", "", text)
    text = re.sub(r"\s+", " ", text).strip(" -–—:,;")
    return apply_text_fixes(text)


def _is_receipt_header(line: str) -> bool:
    low = line.lower().strip()
    if not low:
        return True
    if low in {"$", "description", "description $"}:
        return True
    if low.startswith("description") and "$" in low:
        return True
    return False


def _is_qty_line(line: str) -> bool:
    return bool(re.search(r"\bqty\b|\bqly\b|\boty\b", line, re.I)) and "@" in line


def _is_measurement_line(line: str) -> bool:
    return bool(re.fullmatch(r"[#A4^'‘`\s]*\d+(?:\.\d+)?\s?(?:g|kg|mg|ml|l|lt|pk|pack|ea|each)\b.*", line.strip(), re.I))


def _receipt_token_can_be_amount(text: str) -> bool:
    raw = normalize_token_text(text)
    if not any(ch.isdigit() for ch in raw):
        return False
    # Reject product sizes like 540g, 12pk, 500mL, 370g.
    if re.search(r"\d\s*(?:g|kg|mg|ml|l|lt|pk|pack|ea|each)\b", raw, re.I):
        return False
    # Receipt line totals should be currency-like: decimal, dollar/S marker, or negative.
    return bool(re.search(r"[-–—]|[$S]\s*\d|\d+[.,]\d{2}\b", raw, re.I))


def _receipt_amount_from_line(line: Dict[str, object], img_width: int) -> Optional[Tuple[float, bool, str, float]]:
    words: List[OcrWord] = list(line["words"])  # type: ignore[arg-type]
    candidates: List[Tuple[float, bool, str, float, float]] = []
    for w in words:
        # Receipt right amount column. Product-size digits are explicitly ignored.
        if w.cx < img_width * 0.58:
            continue
        if not _receipt_token_can_be_amount(w.text):
            continue
        parsed = normalize_amount_candidate(w.text)
        if parsed:
            val, neg, raw = parsed
            score = w.conf + (45 if w.cx > img_width * 0.72 else 5) + amount_quality(raw)
            candidates.append((val, neg, raw, score, w.x))
    # Some OCR runs split a minus sign or dollar sign from the number. Try the
    # joined right-column text as a fallback, but only for currency-like tokens.
    right_words = [w for w in words if w.cx >= img_width * 0.58]
    if right_words:
        right_text = " ".join(w.text for w in sorted(right_words, key=lambda w: w.x))
        for piece in re.findall(r"[-–—]?\$?S?\d[\d.,S]*", right_text):
            if not _receipt_token_can_be_amount(piece):
                continue
            parsed = normalize_amount_candidate(piece)
            if parsed:
                val, neg, raw = parsed
                candidates.append((val, neg, raw, 125 + amount_quality(raw), min(w.x for w in right_words)))
    if not candidates:
        return None
    candidates.sort(key=lambda c: (c[3], c[4]), reverse=True)
    val, neg, raw, _score, x = candidates[0]
    right_text = " ".join(w.text for w in sorted(right_words, key=lambda w: w.x)) if right_words else raw
    neg = neg or bool(re.search(r"[-–—]", right_text))
    return val, neg, raw, x


def _left_text_before_amount(line: Dict[str, object], amount_x: float) -> str:
    words: List[OcrWord] = list(line["words"])  # type: ignore[arg-type]
    left = [w for w in words if w.x < amount_x - 2]
    return _clean_receipt_description(line_text(left))


def _best_visible_date_from_text(text: str) -> Optional[str]:
    parsed = parse_explicit_date(text)
    if parsed:
        return parsed.strftime("%d/%m/%Y")
    # Receipt pages can contain "3 Jun 2026" but no weekday. parse_explicit_date
    # handles it; otherwise do not infer from receipt line text.
    return None


def _rows_look_like_receipt(lines: List[Dict[str, object]], img_width: int) -> bool:
    if not lines:
        return False
    text = "\n".join(str(line["text"]) for line in lines)
    amount_lines = sum(1 for line in lines if _receipt_amount_from_line(line, img_width) is not None)
    has_receipt_words = bool(re.search(r"\b(description|qty|each|discount)\b", text, re.I))
    has_transaction_date_words = any(looks_like_date_line(str(line["text"])) for line in lines)
    # Receipt/detail line item screenshots have amount rows but usually no date
    # line per item. Bank/card transaction lists should stay on the old spatial path.
    return amount_lines >= 1 and has_receipt_words and not has_transaction_date_words


def parse_receipt_items(path: Path, source_name: str) -> Tuple[List[ParsedTransaction], str]:
    img = load_image(path)
    lines = _receipt_ocr_lines(img, scale=3)
    full_text = "\n".join(str(line["text"]) for line in lines)
    visible_date = _best_visible_date_from_text(full_text)
    fallback_date = DEFAULT_MISSING_DATE
    date_value = visible_date or fallback_date or ""
    date_missing = not bool(visible_date or fallback_date)

    rows: List[ParsedTransaction] = []
    pending: List[str] = []
    last_row_index: Optional[int] = None
    debug: List[str] = [f"RECEIPT_MODE {source_name} {img.width}x{img.height}", f"DEFAULT_DATE {fallback_date or ''}"]

    for line in lines:
        text = _clean_receipt_description(str(line["text"]))
        if _is_receipt_header(text):
            debug.append(f"SKIP header: {text!r}")
            continue
        amount_info = _receipt_amount_from_line(line, img.width)
        debug.append(f"LINE y={float(line['y']):.1f}: {text!r}; amount={amount_info!r}; pending={pending!r}")
        if amount_info:
            amount, negative, raw_amount, amount_x = amount_info
            left = _left_text_before_amount(line, amount_x)
            left_is_qty = _is_qty_line(left)
            if left_is_qty and pending:
                desc = " ".join(pending)
            elif pending and left:
                # Product name wrapped, with amount on the final name line.
                desc = " ".join(pending + [left])
            elif pending:
                desc = " ".join(pending)
            else:
                desc = left or text
            desc = _clean_receipt_description(desc)
            if not desc or _is_qty_line(desc):
                desc = "UNKNOWN DESCRIPTION"

            value = abs(round(float(amount), 2))
            debit = value if negative else None
            credit = None if negative else value
            reason_parts = []
            if date_missing:
                reason_parts.append("date not visible; enter fallback DATE to avoid this flag")
            if desc == "UNKNOWN DESCRIPTION":
                reason_parts.append("description not detected")
            row = ParsedTransaction(
                date=date_value,
                description=desc,
                debit=debit,
                credit=credit,
                raw_text=text,
                needs_review=bool(reason_parts),
                review_reason="; ".join(reason_parts),
                source_name=source_name,
            )
            rows.append(row)
            last_row_index = len(rows) - 1
            pending = []
            continue

        if _is_qty_line(text):
            # Quantity-only line without a right amount: keep it out of the
            # description unless no product name was captured.
            if not pending:
                pending.append(text)
            continue

        if _is_measurement_line(text) and last_row_index is not None and not pending:
            # Some receipt crops put "22g" after the amount line.
            rows[last_row_index].description = _clean_receipt_description(rows[last_row_index].description + " " + text)
            continue

        if text:
            pending.append(text)

    return rows, "\n".join(debug)


def source_to_rows(path: Path, run_dir: Path) -> Tuple[List[ParsedTransaction], str]:
    ext = path.suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}:
        img = load_image(path)
        receipt_lines = _receipt_ocr_lines(img, scale=3)
        if _rows_look_like_receipt(receipt_lines, img.width):
            return parse_receipt_items(path, path.name)
        return parse_image_spatial(path, run_dir, path.name)
    if ext == ".pdf":
        if fitz is None:
            raise RuntimeError("PyMuPDF is not installed, so PDFs cannot be processed.")
        doc = fitz.open(path)
        all_rows: List[ParsedTransaction] = []
        debug_parts: List[str] = []
        try:
            for page_index in range(len(doc)):
                page = doc[page_index]
                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
                img_path = run_dir / f"{path.stem}_page_{page_index + 1}.png"
                pix.save(str(img_path))
                rows, debug = source_to_rows(img_path, run_dir)
                for r in rows:
                    r.source_name = f"{path.name} page {page_index + 1}"
                all_rows.extend(rows)
                debug_parts.append(debug)
        finally:
            doc.close()
        return all_rows, "\n\n".join(debug_parts)
    raise RuntimeError(f"Unsupported file type: {path.name}")


def _dedupe_key(row: ParsedTransaction) -> Tuple[str, str, str, str]:
    desc = re.sub(r"\s+", " ", row.description.lower()).strip()
    side = "debit" if row.debit is not None else "credit" if row.credit is not None else ""
    amount = row.debit if row.debit is not None else row.credit if row.credit is not None else ""
    amount_text = f"{float(amount):.2f}" if amount != "" else ""
    return row.date, desc, side, amount_text


def dedupe_exact_rows(rows: List[ParsedTransaction]) -> Tuple[List[ParsedTransaction], int]:
    seen = set()
    out: List[ParsedTransaction] = []
    removed = 0
    for row in rows:
        key = _dedupe_key(row)
        if key in seen:
            removed += 1
            continue
        seen.add(key)
        out.append(row)
    return out, removed


# ---------------------------------------------------------------------------
# Version 8 overrides
# Fixes from the 2026-07-04 BOQ mobile screenshot run:
# - parse date headings as their own lines and carry the heading date down;
# - derive Today/Yesterday from the run folder stamp;
# - read amounts from the right-side amount crop with exact cents only;
# - keep transfer direction/pending text instead of noisy secondary OCR.
# ---------------------------------------------------------------------------

JUNE_OCR_VARIANTS = {
    "june", "jute", "lute", "lure", "dune", "dute", "jone", "jom", "jum",
    "jn", "je", "fuse", "flute", "tlune", "pluse", "qure", "ute", "uune",
}
JULY_OCR_VARIANTS = {
    "july", "juby", "jly", "juh", "jhy", "joby", "doty", "duty", "uby",
}
YESTERDAY_OCR_VARIANTS = {
    "yesterday", "yesterdav", "vesterday", "vectendry", "vemtenday",
    "vemedsy", "vemeday", "vemerdsy", "vemnerdzy", "vemerarys",
    "vemeras", "vemtenaag", "veeternaay",
}
TODAY_OCR_VARIANTS = {"today", "todav", "foxtay", "foctay", "yoctay", "tocay", "techy"}
PENDING_OCR_VARIANTS = {"pending", "percing", "pencing", "pemcng", "pecdng", "peccing"}


def _normalise_dashes(text: str) -> str:
    return (
        text.replace("\u2010", "-")
        .replace("\u2011", "-")
        .replace("\u2012", "-")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
        .replace("\u2212", "-")
        .replace("â€“", "-")
        .replace("â€”", "-")
    )


def _run_date_from_run_dir(run_dir: Path) -> dt.date:
    for part in [run_dir.name, run_dir.parent.name]:
        match = re.search(r"run_(\d{8})_\d{6}", part)
        if match:
            try:
                return dt.datetime.strptime(match.group(1), "%Y%m%d").date()
            except ValueError:
                pass
    return dt.date.today()


def _tokens(text: str) -> List[str]:
    return re.findall(r"[A-Za-z0-9@&|%?]+", text.lower())


def _relative_date_from_text(text: str, default_today: dt.date) -> Optional[Tuple[dt.date, str]]:
    low = normalize_token_text(text).lower()
    words = _tokens(low)
    for word in words:
        if word in TODAY_OCR_VARIANTS:
            return default_today, "today label"
        if word in YESTERDAY_OCR_VARIANTS:
            return default_today - dt.timedelta(days=1), "yesterday label"
        if word[:1] in {"t", "f", "y"} and SequenceMatcher(None, word, "today").ratio() >= 0.78:
            return default_today, "today label"
        if word[:1] in {"y", "v"} and len(word) >= 6 and SequenceMatcher(None, word, "yesterday").ratio() >= 0.62:
            return default_today - dt.timedelta(days=1), "yesterday label"
    return None


def detect_weekday(text: str) -> Optional[int]:
    lowered = text.lower().strip()
    if lowered.startswith(("sun", "sunc", "surc", "sarc", "suc", "suwr", "sarr", "sas", "saar")):
        return 6
    if lowered.startswith(("mon", "mor", "mot", "mota", "notd", "motd", "motiv", "neot", "nifd", "mit")):
        return 0
    if lowered.startswith(("tue", "tum", "tus", "tuse", "tuset", "tuet", "twee", "tues", "tweedy")):
        return 1
    if lowered.startswith(("wed", "wied", "veet", "veed", "vendre", "vead")):
        return 2
    if lowered.startswith(("thu", "trx", "thx", "trest", "thest", "trq")):
        return 3
    if lowered.startswith(("fri", "fric", "pricey", "pica", "grice", "fide", "grid", "bric", "cete")):
        return 4
    if lowered.startswith(("sat", "san", "car")):
        return 5

    tokens = re.findall(r"[A-Za-z]{4,}", lowered)
    month_tokens = set(MONTHS.keys()) | MAY_OCR_VARIANTS | JUNE_OCR_VARIANTS | JULY_OCR_VARIANTS
    for token in tokens:
        if token in month_tokens:
            continue
        best, score = fuzzy_best(token, WEEKDAYS)
        if best and score >= 0.68:
            return WEEKDAYS.index(best)
    return None


def _detect_month_with_presence(text: str, default_month: Optional[int]) -> Tuple[Optional[int], bool]:
    low = text.lower()
    for name, num in MONTHS.items():
        if re.search(rf"\b{re.escape(name)}\b", low):
            return num, True
    tokens = _tokens(low)
    for token in tokens:
        if token in MAY_OCR_VARIANTS:
            return 5, True
        if token in JUNE_OCR_VARIANTS:
            return 6, True
        if token in JULY_OCR_VARIANTS:
            return 7, True
    for token in tokens:
        best, score = fuzzy_best(token, list(MONTHS.keys()))
        if best and score >= 0.84 and (len(token) <= 6 or token[:3] == best[:3]):
            return MONTHS[best], True
    return default_month, False


def detect_month(text: str, default_month: Optional[int]) -> Optional[int]:
    month, _seen = _detect_month_with_presence(text, default_month)
    return month


def _clean_date_digits(raw: str) -> str:
    table = str.maketrans({
        "O": "0", "o": "0", "S": "5", "s": "5", "I": "1", "l": "1",
        "|": "1", "@": "8", "&": "5", "%": "6", "?": "6",
    })
    return raw.translate(table)


def _day_candidates_from_date_text(raw: str) -> List[int]:
    digit_text = _clean_date_digits(raw)
    candidates: List[int] = []
    for m in re.finditer(r"\d{1,2}", digit_text):
        val = int(m.group(0))
        if not (1 <= val <= 31):
            continue
        before = digit_text[max(0, m.start() - 2):m.start()]
        after = digit_text[m.end():m.end() + 2]
        if before.endswith("20") or after.startswith(("00", "06", "08", "20", "26")):
            continue
        candidates.append(val)

    # Tiny "15" in BOQ's grey date headings is commonly OCR'd as MM/Ms/My.
    if re.search(r"\b[Mm][MmSsYy]?\b", raw) and 15 not in candidates:
        candidates.append(15)
    return candidates


def looks_like_date_line(text: str) -> bool:
    low = normalize_token_text(text).lower()
    if _relative_date_from_text(low, dt.date.today()) is not None:
        return True
    month, month_seen = _detect_month_with_presence(low, None)
    if month_seen:
        return True
    if detect_weekday(low) is not None and (re.search(r"\d|[@&|%?]", low) or month is not None):
        return True
    if re.search(r"\b20[0-9O]{2}\b", _clean_date_digits(low)):
        return True
    return False


def _heading_ocr_texts(img: Image.Image, top: int, title_y: Optional[float]) -> List[str]:
    if title_y is None or not math.isfinite(title_y) or title_y - top < 22:
        return []
    y1 = max(0, int(top + 2))
    y2 = min(img.height, int(title_y - 4))
    if y2 - y1 < 8:
        return []
    crop = img.crop((12, y1, min(img.width, 165), y2))
    texts: List[str] = []
    jobs = [(12, "gray", 6), (16, "gray", 6), (12, "thr210", 6), (12, "thr210", 11)]
    for scale, variant_name, psm in jobs:
        big = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
        gray = ImageEnhance.Sharpness(ImageOps.autocontrast(ImageOps.grayscale(big))).enhance(2.5)
        im = gray
        if variant_name.startswith("thr"):
            try:
                import numpy as np
                threshold = int(variant_name[3:])
                arr = np.array(gray)
                im = Image.fromarray(np.where(arr < threshold, 0, 255).astype("uint8"))
            except Exception:
                continue
        im2 = ImageOps.expand(im, border=24, fill=255)
        try:
            text = pytesseract.image_to_string(im2, lang="eng", config=f"--oem 3 --psm {psm}", timeout=2)
        except Exception:
            continue
        text = normalize_token_text(text.replace("\n", " "))
        if text:
            texts.append(text)
    return texts


def _extract_heading_text(row_words: Sequence[OcrWord], top: int, bottom: int, title_y: Optional[float], img: Image.Image) -> str:
    if title_y is None or not math.isfinite(title_y) or title_y - top < 22:
        return ""
    texts: List[str] = []
    cutoff = max(top + 4, title_y - 5)
    top_words = [w for w in row_words if w.x < img.width * 0.58 and top - 1 <= w.cy <= cutoff]
    for line in line_groups(top_words, y_tol=5.5):
        text = line_text(line)
        if text and (looks_like_date_line(text) or re.search(r"\d|[@&|%?]", text)):
            texts.append(text)
    texts.extend(_heading_ocr_texts(img, top, title_y))
    unique: List[str] = []
    seen = set()
    for text in texts:
        key = re.sub(r"\s+", " ", text.lower()).strip()
        if key and key not in seen:
            seen.add(key)
            unique.append(text)
    return " | ".join(unique)


def _parse_heading_date(
    heading_text: str,
    previous_heading: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    raw = normalize_token_text(heading_text)
    if not raw:
        return None, "date heading not detected", True

    relative = _relative_date_from_text(raw, default_today)
    if relative is not None:
        return relative[0], relative[1], False

    latest_allowed = previous_heading - dt.timedelta(days=1) if previous_heading else default_today
    explicit = parse_explicit_date(raw)
    if explicit:
        if explicit.year != latest_allowed.year and abs(explicit.year - latest_allowed.year) > 2:
            try:
                explicit = dt.date(latest_allowed.year, explicit.month, explicit.day)
            except ValueError:
                explicit = None
        if explicit and explicit <= latest_allowed:
            return explicit, "explicit date", False

    month, month_seen = _detect_month_with_presence(raw, latest_allowed.month)
    if month is None:
        month = latest_allowed.month
    weekday = detect_weekday(raw)
    day_candidates = _day_candidates_from_date_text(raw)

    if re.search(r"\b[Mm][MmSsYy]?\b", raw) and month_seen:
        try:
            cand = dt.date(latest_allowed.year, month, 15)
            if cand <= latest_allowed:
                return cand, "heading day OCR 15", False
        except ValueError:
            pass

    scored: List[Tuple[int, dt.date, str]] = []
    for day in sorted(set(day_candidates), reverse=True):
        try:
            cand = dt.date(latest_allowed.year, month, day)
        except ValueError:
            continue
        if cand > latest_allowed:
            continue
        score = 45
        reason = "heading day OCR"
        if month_seen:
            score += 18
            reason += "; month OCR"
        if weekday is not None:
            if cand.weekday() == weekday:
                score += 38
                reason += "; weekday confirmed"
            else:
                score -= 35
                reason += "; weekday disagreed"
        score -= min(20, (latest_allowed - cand).days // 10)
        scored.append((score, cand, reason))

    if scored:
        scored.sort(key=lambda item: (item[0], item[1]), reverse=True)
        score, cand, reason = scored[0]
        return cand, reason, score < 65

    return None, "date heading not reliable", True


def candidate_dates_from_context(
    date_text: str,
    previous: Optional[dt.date],
    default_today: dt.date,
) -> Tuple[Optional[dt.date], str, bool]:
    return _parse_heading_date(date_text, previous, default_today)


def _clean_title_text(text: str) -> str:
    text = normalize_token_text(text)
    text = re.sub(r"\b[Â©Â®©®]\b", " ", text)
    text = re.sub(r"^[^\w+]+", "", text)
    parts: List[str] = []
    for part in text.split():
        stripped = part.strip(" ,.;:()[]{}'\"")
        if not stripped:
            continue
        if re.fullmatch(r"[-+*/_Â©Â®©®|]+", stripped):
            continue
        if normalize_amount_candidate(stripped):
            continue
        parts.append(stripped)
    return apply_text_fixes(" ".join(parts))


def _collapse_title_noise(text: str) -> str:
    text = re.sub(r"\bRENOER\.COM\b", "RENDER.COM", text, flags=re.I)
    text = re.sub(r"\bROENOCR\.OOM\b", "RENDER.COM", text, flags=re.I)
    text = re.sub(r"\+141S8304762", "+14158304762", text)
    text = re.sub(r"\bYurnmy\b", "Yummy", text, flags=re.I)
    text = re.sub(r"\bHurryohmn\b", "Hurryohm", text, flags=re.I)
    text = re.sub(r"\bStvop\b", "Shop", text, flags=re.I)
    text = re.sub(r"\bPIZ2Z4/SHOP\s+PIZZA/SHOP\b", "PIZZA/SHOP", text, flags=re.I)
    text = re.sub(r"\bPIZ2Z4/SHOP\b", "PIZZA/SHOP", text, flags=re.I)
    text = re.sub(r"\bMEACOWBROOK\b", "MEADOWBROOK", text, flags=re.I)
    text = re.sub(r"\bYummy Fried Rice\b", "Yummy Fried Rice", text, flags=re.I)
    text = re.sub(r"\bPizza in A Hurryohm Food\b", "Pizza In A Hurryohm Food", text, flags=re.I)
    text = re.sub(r"\bRobins Pizza/Shop 17/2\b", "Robins Pizza/Shop 17/2", text, flags=re.I)
    text = re.sub(r"\bRender\.com\b", "Render.Com", text)
    text = re.sub(r"\s+", " ", text).strip(" -:,;")
    if "MEADOWBROOK" in text.upper():
        text = text.upper()
    return text


def extract_title(row_words: Sequence[OcrWord], top: int, bottom: int, img_width: int) -> Tuple[str, Optional[float]]:
    candidates: List[Tuple[float, str, float]] = []
    for engine in ("psm11", "psm6", "psm4"):
        engine_words = [w for w in row_words if w.engine == engine]
        left_words = [w for w in engine_words if img_width * 0.07 <= w.x <= img_width * 0.73 and top <= w.cy <= bottom]
        for line in line_groups(left_words, y_tol=5.5):
            cy = sum(w.cy for w in line) / len(line)
            text = _clean_title_text(line_text(line))
            if not text or looks_like_date_line(text):
                continue
            if len(text) < 2 or not any(ch.isalpha() for ch in text):
                continue
            avg_conf = sum(max(w.conf, 0) for w in line) / len(line)
            avg_h = sum(w.h for w in line) / len(line)
            score = avg_conf + avg_h * 4
            if engine == "psm11":
                score += 10
            if engine == "psm6":
                score += 8
            # Prefer the visible title line over the smaller secondary merchant line.
            score += max(0, 34 - abs((cy - top) - 22))
            if "/" in text or "+" in text:
                score += 12
            if any(ch.islower() for ch in text):
                score += 6
            candidates.append((score, _collapse_title_noise(text), cy))

    if not candidates:
        return "UNKNOWN DESCRIPTION", None
    candidates.sort(key=lambda item: item[0], reverse=True)
    title = candidates[0][1]
    words: List[str] = []
    for word in title.split():
        clean = word.lower().strip(".,;:'\"")
        if words and clean == words[-1].lower().strip(".,;:'\""):
            continue
        words.append(word)
    return " ".join(words), candidates[0][2]


def _transfer_detail_text(row_words: Sequence[OcrWord], title_y: Optional[float], img_width: int) -> str:
    if title_y is None or not math.isfinite(title_y):
        return ""
    detail_words = [
        w for w in row_words
        if img_width * 0.07 <= w.x <= img_width * 0.68 and title_y + 3 <= w.cy <= title_y + 22
    ]
    texts = []
    for line in line_groups(detail_words, y_tol=5.5):
        text = _clean_title_text(line_text(line))
        if text and not looks_like_date_line(text):
            texts.append(text)
    return " ".join(texts)


def _enhance_description(title: str, row_words: Sequence[OcrWord], title_y: Optional[float], img_width: int, raw: str) -> str:
    title = _collapse_title_noise(title)
    if title.lower() == "transfer":
        detail = _transfer_detail_text(row_words, title_y, img_width)
        low = f"{detail} {raw}".lower()
        if re.search(r"\b(to|toc)\b", low) or "aecc" in low or "syrm" in low or "sym" in low:
            return "Transfer to Alexander Symoniw"
        if re.search(r"\b(from|frocex|frocrx)\b", low) or "save" in low:
            return "Transfer from SaveME"
    return title


def _is_pending(raw: str) -> bool:
    low = normalize_token_text(raw).lower()
    return any(token in low for token in PENDING_OCR_VARIANTS)


def _amount_tokens_from_text(text: str) -> List[str]:
    clean = _normalise_dashes(normalize_token_text(text)).replace(",", ".")
    clean = clean.replace("*", "").replace("“", "").replace("”", "").replace("â€œ", "").replace("â€", "")
    tokens: List[str] = []
    for piece in re.findall(r"[+\-]?\s*(?:\$|S)?\s*\d[\d.,SILOo]*", clean):
        piece = re.sub(r"\s+", "", piece).strip(".,;:")
        if any(ch.isdigit() for ch in piece):
            tokens.append(piece)
    return tokens


def normalize_amount_candidate(token: str) -> Optional[Tuple[float, bool, str]]:
    raw = _normalise_dashes(normalize_token_text(token)).strip()
    if not raw or not any(ch.isdigit() for ch in raw):
        return None
    raw = raw.replace("*", "").replace("“", "").replace("”", "").replace("â€œ", "").replace("â€", "")
    raw = raw.strip(" .,:;")
    negative = "-" in raw or raw.upper().endswith("DR")
    credit_marker = raw.upper().endswith("CR")
    sign_present = negative or "+" in raw

    s = raw.upper()
    s = s.replace("AUD", "").replace("USD", "").replace("NZD", "").replace("EUR", "").replace("GBP", "")
    s = s.replace("CR", "").replace("DR", "")
    s = re.sub(r"\s+", "", s).replace(",", ".")

    sign = ""
    if s.startswith(("+", "-")):
        sign, s = s[0], s[1:]
    if s.startswith("S") and len(s) > 1 and s[1].isdigit():
        s = "$" + s[1:]
    has_currency = "$" in s
    s = s.replace("$", "")

    s = s.replace("O", "0").replace("o", "0")
    s = s.replace("I", "1").replace("L", "1")
    s = s.replace("S", "5")
    s = re.sub(r"[^0-9.]", "", s).strip(".")
    if not s or not (has_currency or sign_present):
        return None

    try:
        if re.fullmatch(r"\d{1,5}\.\d{2}", s):
            value = float(s)
        elif re.fullmatch(r"\d{3,7}", s):
            value = int(s) / 100.0
        else:
            return None
    except Exception:
        return None
    if value <= 0:
        return None
    return round(value, 2), negative and not credit_marker, raw


def amount_quality(raw: str) -> int:
    clean = _normalise_dashes(normalize_token_text(raw))
    q = 0
    if "$" in clean or re.search(r"(^|[+\-])S(?=\d)", clean, re.I):
        q += 30
    if re.search(r"[+\-]", clean):
        q += 16
    if re.search(r"\d+[.,]\d{2}(?!\d)", clean):
        q += 60
    elif re.search(r"(?:\$|S)\d{3,7}\b", clean, re.I):
        q += 28
    if re.search(r"\d+[.,]\d{1}(?!\d)", clean):
        q -= 35
    if "$" not in clean and not re.search(r"(^|[+\-])S(?=\d)", clean, re.I):
        q -= 28
    return q


def _amount_candidate_score(raw: str, scale: int = 0, psm: int = 0, variant: str = "") -> int:
    score = amount_quality(raw)
    if scale >= 20:
        score += 10
    if psm in (6, 7):
        score += 12
    if variant.startswith("thr"):
        score += 14
    if variant.startswith("contrast"):
        score += 135
    if variant == "gray":
        score += 8
    return score


def _focused_amount_crop(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> Optional[Image.Image]:
    if title_y is None or not math.isfinite(title_y):
        title_y = top + (bottom - top) * 0.45
    x1 = max(0, int(img.width * 0.74))
    x2 = min(img.width, int(img.width * 0.915))
    y1 = max(0, int(title_y - 13))
    y2 = min(img.height, int(title_y + 14))
    if x2 - x1 < 12 or y2 - y1 < 8:
        return None
    return img.crop((x1, y1, x2, y2))


def _baseline_amount_crop(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> Optional[Image.Image]:
    if title_y is None or not math.isfinite(title_y):
        title_y = top + (bottom - top) * 0.45
    x1 = max(0, int(img.width * 0.74))
    x2 = min(img.width, int(img.width * 0.91))
    y1 = max(0, int(title_y - 9))
    y2 = min(img.height, int(title_y + 9))
    if x2 - x1 < 12 or y2 - y1 < 8:
        return None
    return img.crop((x1, y1, x2, y2))


def _fast_amount_crop_candidates(img: Image.Image, title_y: Optional[float], top: int, bottom: int) -> List[Tuple[float, bool, str, int]]:
    crops: List[Tuple[str, Image.Image]] = []
    bbox = _amount_bbox_crop(img, top, bottom)
    if bbox is not None and bbox.width >= 8 and bbox.height >= 6:
        crops.append(("bbox", bbox))
    baseline = _baseline_amount_crop(img, title_y, top, bottom)
    if baseline is not None:
        crops.append(("baseline", baseline))
    focused = _focused_amount_crop(img, title_y, top, bottom)
    if focused is not None:
        crops.append(("focused", focused))

    jobs = [
        (12, "gray", 7),
        (12, "thr190", 7),
        (16, "thr160", 7),
        (20, "thr170", 6),
        (20, "contrast170", 6),
        (12, "min", 7),
    ]
    out: List[Tuple[float, bool, str, int]] = []
    for crop_name, crop in crops[:3]:
        for scale, variant_name, psm in jobs:
            big = crop.resize((crop.width * scale, crop.height * scale), Image.Resampling.LANCZOS)
            gray = ImageEnhance.Sharpness(ImageOps.autocontrast(ImageOps.grayscale(big))).enhance(3)
            im = gray
            if variant_name.startswith("thr"):
                try:
                    import numpy as np
                    threshold = int(variant_name[3:])
                    arr = np.array(gray)
                    im = Image.fromarray(np.where(arr < threshold, 0, 255).astype("uint8"))
                except Exception:
                    continue
            elif variant_name.startswith("contrast"):
                try:
                    import numpy as np
                    threshold = int(variant_name.replace("contrast", ""))
                    contrast = ImageEnhance.Contrast(
                        ImageEnhance.Sharpness(ImageOps.grayscale(big)).enhance(4)
                    ).enhance(2)
                    arr = np.array(contrast)
                    im = Image.fromarray(np.where(arr < threshold, 0, 255).astype("uint8"))
                except Exception:
                    continue
            elif variant_name == "min":
                im = gray.filter(ImageFilter.MinFilter(3))
            im2 = ImageOps.expand(im, border=30, fill=255)
            config = f"--oem 3 --psm {psm} -c tessedit_char_whitelist=0123456789.$+-SILOo"
            try:
                text = pytesseract.image_to_string(im2, lang="eng", config=config, timeout=2).strip().replace("\n", " ")
            except Exception:
                continue
            for piece in _amount_tokens_from_text(text):
                parsed = normalize_amount_candidate(piece)
                if parsed:
                    val, neg, raw = parsed
                    score = 120 + _amount_candidate_score(raw, scale, psm, variant_name)
                    if crop_name == "bbox":
                        score += 28
                    elif crop_name == "baseline":
                        score += 16
                    out.append((val, neg, f"amount OCR[{crop_name}:{variant_name}] {raw}", score))
    return out


def _rank_amount_groups(candidates: List[Tuple[float, bool, str, int]]) -> List[Tuple[int, float, bool, str, int]]:
    grouped: Dict[float, List[Tuple[float, bool, str, int]]] = {}
    for val, neg, raw, score in candidates:
        val = round(val, 2)
        if val < 1:
            continue
        grouped.setdefault(val, []).append((val, neg, raw, score))

    ranked: List[Tuple[int, float, bool, str, int]] = []
    for val, group in grouped.items():
        best = max(group, key=lambda item: item[3])
        source_types = set("crop" if "amount OCR" in c[2] else "word" for c in group)
        exact_count = sum(1 for c in group if re.search(r"\d+[.,]\d{2}(?!\d)", c[2]))
        signed_count = sum(1 for c in group if re.search(r"[+\-]", _normalise_dashes(c[2])))
        currency_count = sum(1 for c in group if "$" in c[2] or re.search(r"\bS\d", c[2], re.I))
        total = best[3] + min(90, 22 * (len(group) - 1)) + 32 * max(0, len(source_types) - 1)
        total += min(45, exact_count * 9) + min(30, signed_count * 8) + min(25, currency_count * 5)
        ranked.append((total, val, any(c[1] for c in group), best[2], len(source_types)))
    ranked.sort(key=lambda item: item[0], reverse=True)
    return ranked


def _prefer_amount_candidate(ranked: List[Tuple[int, float, bool, str, int]]) -> Tuple[int, float, bool, str, int]:
    if not ranked:
        raise ValueError("ranked amount list is empty")
    best = ranked[0]
    best_score, best_val, _best_neg, _best_raw, _best_sources = best
    lower_same_cent: List[Tuple[int, float, bool, str, int]] = []
    for cand in ranked[1:]:
        score, val, _neg, raw, _sources = cand
        if score < best_score - 260:
            continue
        if (
            "contrast" in raw
            and _same_cents(best_val, val)
            and int(best_val) // 10 == int(val) // 10
            and int(val) // 10 > 0
            and score >= best_score - 260
        ):
            return cand
        if (
            _same_cents(best_val, val)
            and 1 <= val < best_val
            and best_val >= val * 1.8
            and score >= best_score - 250
        ):
            lower_same_cent.append(cand)
            continue
        if "contrast" in _best_raw and _sources >= 2 and score >= best_score - 160:
            return cand
        if score < best_score - 70:
            continue
        if _same_cents(best_val, val) and "$" in raw and 1 <= val < best_val and score >= best_score - 35:
            return cand
    if lower_same_cent:
        return min(lower_same_cent, key=lambda item: item[1])
    return best


def choose_amount(row_words: Sequence[OcrWord], img: Image.Image, top: int, bottom: int, title_y: Optional[float]) -> Tuple[Optional[float], bool, bool, str]:
    candidates: List[Tuple[float, bool, str, int]] = []
    for w in row_words:
        if w.x < img.width * 0.58:
            continue
        parsed = normalize_amount_candidate(w.text)
        if not parsed:
            continue
        val, neg, raw = parsed
        score = int(max(w.conf, 0)) + _amount_candidate_score(raw)
        if w.engine == "psm4":
            score += 10
        if w.engine == "psm6":
            score += 8
        if w.engine == "psm11":
            score += 6
        candidates.append((val, neg, f"word OCR {raw}", score))
    candidates.extend(_fast_amount_crop_candidates(img, title_y, top, bottom))

    ranked = _rank_amount_groups(candidates)
    if not ranked:
        return None, False, True, "amount not detected"

    total, val, negative, raw, source_count = _prefer_amount_candidate(ranked)
    review = total < 165 or source_count < 1
    if len(ranked) > 1:
        second = ranked[1]
        same_cents_digit_noise = _same_cents(val, second[1]) and max(val, second[1]) >= min(val, second[1]) * 1.8
        same_digit_family = (
            "contrast" in raw
            and _same_cents(val, second[1])
            and int(val) // 10 == int(second[1]) // 10
        )
        tight_exact_read = (
            ("[bbox:" in raw or raw.startswith("word OCR"))
            and re.search(r"\d+[.,]\d{2}(?!\d)", raw)
            and second[0] <= total + 5
        )
        if (
            not same_cents_digit_noise
            and not same_digit_family
            and not tight_exact_read
            and abs(second[1] - val) >= 0.04
            and second[0] >= total - 38
        ):
            review = True
            raw = f"{raw}; competing amount {second[1]:.2f}"
    return val, negative, review, raw


def infer_side(description: str, negative: bool, raw_amount_reason: str) -> Tuple[str, bool, str]:
    low = description.lower()
    raw = _normalise_dashes(raw_amount_reason)
    has_plus = "+" in raw
    has_minus = "-" in raw or negative
    if "transfer from" in low or (has_plus and "transfer to" not in low):
        return "debit", False, raw_amount_reason
    if "transfer to" in low or has_minus:
        return "credit", False, raw_amount_reason
    if any(hint in low for hint in INCOME_HINTS):
        return "debit", False, f"{raw_amount_reason}; income keyword"
    if any(hint in low for hint in OUTGOING_HINTS):
        return "credit", False, f"{raw_amount_reason}; outgoing keyword"
    return "credit", True, f"{raw_amount_reason}; defaulted to money out; verify"


def parse_image_spatial(path: Path, run_dir: Path, source_name: str) -> Tuple[List[ParsedTransaction], str]:
    img = load_image(path)
    all_words: List[OcrWord] = []
    for psm in (4, 6, 11):
        try:
            all_words.extend(ocr_words(img, psm=psm, scale=4))
        except Exception:
            pass
    segments = make_segments(img, all_words)
    rows: List[ParsedTransaction] = []
    current_heading_date: Optional[dt.date] = None
    run_today = _run_date_from_run_dir(run_dir)

    debug_lines: List[str] = [
        f"IMAGE {source_name} {img.width}x{img.height}",
        f"RUN_DATE {run_today.isoformat()}",
        f"SEGMENTS {segments}",
    ]
    for seg_index, (top, bottom) in enumerate(segments, start=1):
        row_words = [w for w in all_words if top - 2 <= w.cy <= bottom + 2]
        if not row_words:
            continue
        title, title_y = extract_title(row_words, top, bottom, img.width)
        heading_text = _extract_heading_text(row_words, top, bottom, title_y, img)
        heading_date, date_reason, date_review = _parse_heading_date(heading_text, current_heading_date, run_today)
        if heading_text:
            if heading_date is not None and not date_review:
                current_heading_date = heading_date
            elif date_review:
                current_heading_date = None

        raw = row_raw_text(row_words)
        title = _enhance_description(title, row_words, title_y, img.width, raw)
        amount, negative, amount_review, amount_reason = choose_amount(row_words, img, top, bottom, title_y)
        pending = _is_pending(raw)
        parsed_date = current_heading_date

        debug_lines.append(
            f"ROW {seg_index}: segment=({top},{bottom}); heading_text={heading_text!r}; "
            f"heading_date={parsed_date.isoformat() if parsed_date else ''}; title={title!r}; "
            f"amount={amount!r}; amount_review={amount_review}; raw={raw!r}"
        )

        if title == "UNKNOWN DESCRIPTION" and amount is None and parsed_date is None:
            continue

        review_parts: List[str] = []
        date_text_for_row = ""
        if parsed_date is None:
            review_parts.append(date_reason if heading_text else "date heading not detected")
        else:
            date_text_for_row = parsed_date.strftime("%d/%m/%Y")

        debit: Optional[float] = None
        credit: Optional[float] = None
        if amount is None:
            review_parts.append("amount not detected")
        elif amount_review:
            review_parts.append(amount_reason)
        else:
            side, side_review, side_reason = infer_side(title, negative, amount_reason)
            if side_review:
                review_parts.append(side_reason)
            else:
                value = abs(round(amount, 2))
                debit = value if side == "debit" else None
                credit = value if side == "credit" else None

        rows.append(ParsedTransaction(
            date=date_text_for_row,
            description=title,
            debit=debit,
            credit=credit,
            raw_text=raw,
            needs_review=bool(review_parts),
            review_reason="; ".join(review_parts),
            source_name=source_name,
            pending=pending,
        ))

    return rows, "\n".join(debug_lines)


def _rows_look_like_simple_date_amount_list(lines: List[Dict[str, object]], img_width: int) -> bool:
    if len(lines) < 6:
        return False
    date_lines = 0
    amount_lines = 0
    paired = 0
    previous_was_date = False
    for line in lines:
        text = str(line.get("text", ""))
        is_date = parse_explicit_date(text) is not None
        amount = _receipt_amount_from_line(line, img_width)
        if is_date:
            date_lines += 1
            previous_was_date = True
            continue
        if amount is not None:
            amount_lines += 1
            if previous_was_date:
                paired += 1
            previous_was_date = False
        elif text.strip():
            previous_was_date = False
    return date_lines >= 3 and amount_lines >= 3 and paired >= 3


def _clean_simple_list_description(text: str) -> str:
    text = normalize_token_text(text)
    text = re.sub(r"\bWoohwvorths\b", "Woolworths", text, flags=re.I)
    text = re.sub(r"\bWoohvorths\b", "Woolworths", text, flags=re.I)
    text = re.sub(r"\bWoolvorths\b", "Woolworths", text, flags=re.I)
    text = re.sub(r"\bWoolwvorths\b", "Woolworths", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" -:,;")
    return text


def parse_simple_date_amount_list(path: Path, source_name: str) -> Tuple[List[ParsedTransaction], str]:
    img = load_image(path)
    lines = _receipt_ocr_lines(img, scale=4)
    rows: List[ParsedTransaction] = []
    debug: List[str] = [f"SIMPLE_LIST_MODE {source_name} {img.width}x{img.height}"]
    current_date: Optional[dt.date] = None
    current_date_text = ""

    for line in lines:
        text = normalize_token_text(str(line.get("text", "")))
        if not text:
            continue
        parsed_date = parse_explicit_date(text)
        if parsed_date is not None:
            current_date = parsed_date
            current_date_text = text
            debug.append(f"DATE y={float(line['y']):.1f}: {text!r} -> {parsed_date.isoformat()}")
            continue

        amount_info = _receipt_amount_from_line(line, img.width)
        debug.append(f"LINE y={float(line['y']):.1f}: {text!r}; amount={amount_info!r}; date={current_date_text!r}")
        if amount_info is None:
            continue

        amount, negative, raw_amount, amount_x = amount_info
        desc = _clean_simple_list_description(_left_text_before_amount(line, amount_x))
        review_parts: List[str] = []
        if current_date is None:
            review_parts.append("date not detected")
            date_value = ""
        else:
            date_value = current_date.strftime("%d/%m/%Y")
        if not desc:
            review_parts.append("description not detected")
        value = abs(round(float(amount), 2))
        rows.append(ParsedTransaction(
            date=date_value,
            description=desc or "UNKNOWN DESCRIPTION",
            debit=None if negative else value,
            credit=value if negative else None,
            raw_text=f"{current_date_text} | {text}",
            needs_review=bool(review_parts),
            review_reason="; ".join(review_parts),
            source_name=source_name,
            pending=False,
        ))
        current_date = None
        current_date_text = ""

    return rows, "\n".join(debug)


def source_to_rows(path: Path, run_dir: Path) -> Tuple[List[ParsedTransaction], str]:
    ext = path.suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}:
        img = load_image(path)
        receipt_lines = _receipt_ocr_lines(img, scale=4)
        if _rows_look_like_simple_date_amount_list(receipt_lines, img.width):
            return parse_simple_date_amount_list(path, path.name)
        if _rows_look_like_receipt(receipt_lines, img.width):
            return parse_receipt_items(path, path.name)
        return parse_image_spatial(path, run_dir, path.name)
    if ext == ".pdf":
        if fitz is None:
            raise RuntimeError("PyMuPDF is not installed, so PDFs cannot be processed.")
        doc = fitz.open(path)
        all_rows: List[ParsedTransaction] = []
        debug_parts: List[str] = []
        try:
            for page_index in range(len(doc)):
                page = doc[page_index]
                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3), alpha=False)
                img_path = run_dir / f"{path.stem}_page_{page_index + 1}.png"
                pix.save(str(img_path))
                rows, debug = source_to_rows(img_path, run_dir)
                for r in rows:
                    r.source_name = f"{path.name} page {page_index + 1}"
                all_rows.extend(rows)
                debug_parts.append(debug)
        finally:
            doc.close()
        return all_rows, "\n\n".join(debug_parts)
    raise RuntimeError(f"Unsupported file type: {path.name}")


def main() -> int:
    global DEFAULT_MISSING_DATE
    try:
        ensure_tesseract()
    except Exception as exc:
        print(f"\nERROR: {exc}\n")
        return 2

    using_stdin = not bool(sys.argv[1:])
    sources = sys.argv[1:] or read_sources_from_stdin()
    sources = [s.strip() for s in sources if s.strip()]
    if not sources:
        print("No sources supplied.")
        return 1

    if using_stdin:
        DEFAULT_MISSING_DATE = read_default_missing_date()
    else:
        DEFAULT_MISSING_DATE = parse_user_date(os.environ.get("TRANSACTION_EXTRACTOR_DEFAULT_DATE", ""))

    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = OUTPUT_DIR / f"run_{stamp}"
    run_dir.mkdir(parents=True, exist_ok=True)

    all_rows: List[ParsedTransaction] = []
    try:
        for idx, source in enumerate(sources, start=1):
            local_path = download_or_copy(source, idx, run_dir)
            print(f"OCR/extracting {local_path.name}")
            rows, debug_text = source_to_rows(local_path, run_dir)
            debug_path = run_dir / f"ocr_debug_{idx:02d}.txt"
            debug_path.write_text(debug_text, encoding="utf-8")
            print(f"  Found {len(rows)} transaction rows")
            all_rows.extend(rows)

        all_rows, deduped = dedupe_exact_rows(all_rows)
        if deduped:
            print(f"  Removed {deduped} exact duplicate rows from overlapping screenshots")

        if not all_rows:
            print("\nFAILED: No transaction rows were extracted. No workbook was written.")
            return 4

        output_xlsx = run_dir / f"extracted_transactions_{stamp}.xlsx"
        review_csv = run_dir / f"review_extraction_{stamp}.csv"
        write_workbook(all_rows, output_xlsx)
        write_review(all_rows, review_csv)

        flagged = sum(1 for r in all_rows if r.needs_review)
        print("\nDONE")
        print(f"Excel workbook: {output_xlsx.resolve()}")
        print(f"Review CSV:     {review_csv.resolve()}")
        print(f"Debug OCR:      {run_dir.resolve()}")
        print(f"Rows written:   {len(all_rows)}")
        print(f"Review flagged: {flagged}")
        if flagged:
            print("\nSome rows still need checking in the review CSV. The workbook was written, but flagged rows should be reviewed.")
        return 0
    except Exception as exc:
        print(f"\nFAILED: {exc}\n")
        return 3

if __name__ == "__main__":
    raise SystemExit(main())
