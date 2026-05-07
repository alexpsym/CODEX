#!/usr/bin/env python3
"""
Build a 32-bit-Android-safe Excel replica of the Trading Journal from workbooks in CODEX-master/journal.
No pandas, no FastAPI, no local web server.
"""
from __future__ import annotations

import argparse
import math
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, date, time, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    print(f"ERROR: missing openpyxl: {exc}", file=sys.stderr)
    print("Install with: python -m pip install openpyxl xlrd==2.0.1 python-dateutil", file=sys.stderr)
    raise SystemExit(2)

try:
    from dateutil import parser as date_parser
except Exception:  # pragma: no cover
    date_parser = None

try:
    import xlrd
except Exception:  # pragma: no cover
    xlrd = None

OUTPUT_NAME = "TradingJournal_Android_Replica.xlsx"
EXCLUDED_SOURCE_NAMES = {
    OUTPUT_NAME.lower(),
    "account_cashflows.xlsx",
}

ALIASES = {
    "open_time": ["opening_time", "open_time", "entry_time", "time_open", "opened", "opened_at", "date_open"],
    "close_time": ["closing_time", "close_time", "exit_time", "time_close", "closed_at", "closed", "date_close"],
    "side": ["type_buy_sell", "side", "direction", "buy_sell", "type", "long_short"],
    "symbol": ["symbol", "instrument", "pair", "market", "ticker"],
    "account": ["account", "account_label", "portfolio", "book"],
    "currency": ["account_currency", "currency", "ccy", "deposit_currency"],
    "setup": ["setup", "strategy", "entry_setup"],
    "qty": ["size_quantity", "qty", "quantity", "size", "units", "volume", "lots"],
    "entry": ["entry_price", "entry", "open_price", "price_open"],
    "exit": ["closing_price", "exit_price", "exit", "close_price", "price_close"],
    "swap": ["swap"],
    "commission": ["commission", "fee", "fees", "cost"],
    "net_profit": ["net_profit", "realized_pnl", "pnl", "profit", "pl", "net_pnl", "result", "p_l"],
    "balance_after": ["balance_after_trade", "bal_after_trade", "balance_after", "bal_after"],
    "stop_loss": ["stop_loss_optional", "stop_loss", "sl"],
    "take_profit": ["take_profit_optional", "take_profit", "tp", "target"],
    "high": ["highest_price_optional", "highest_price", "high"],
    "low": ["lowest_price_optional", "lowest_price", "low"],
    "notes": ["notes", "pre_trade_comments", "entry_comments", "trade_management", "exit_comments"],
    "breakeven": ["breakeven", "break_even", "be"],
    "timeframe": ["timeframe", "tf"],
    "error": ["error"],
    "is_test_trade": ["is_test_trade", "test_trade", "test"],
    "order_id": ["order_id", "orderid", "order_no", "order_no_"],
    "fill_count": ["fill_count", "fillcount"],
    "source": ["source", "source_file", "import_source"],
    "result_pct": ["result_pct", "profit_pct", "result_percent", "profit_percent"],
    "r_multiple": ["r_multiple", "r", "r_mult"],
    "held_through_news": ["held_through_news"],
    "spiked_out": ["spiked_out"],
    "early_close": ["early_close"],
}

DARK = "111827"
DARK2 = "0F172A"
PANEL = "1F2937"
HEADER = "2563EB"
HEADER2 = "1D4ED8"
TEXT = "E5E7EB"
MUTED = "9CA3AF"
GREEN = "22C55E"
RED = "F87171"
AMBER = "F59E0B"
BORDER = "374151"
WHITE = "FFFFFF"


def norm_col(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "—", "-"}:
        return None
    text = text.replace(",", "")
    text = re.sub(r"[^0-9+\-.]", "", text)
    if not text or text in {"+", "-", "."}:
        return None
    try:
        return float(text)
    except Exception:
        return None


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def parse_dt(value: Any, datemode: Optional[int] = None) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)) and datemode is not None and xlrd is not None:
        try:
            return xlrd.xldate.xldate_as_datetime(value, datemode).replace(tzinfo=None)
        except Exception:
            pass
    if isinstance(value, (int, float)):
        # Excel serial fallback for xlsx numeric dates.
        try:
            if 20000 < float(value) < 90000:
                return datetime(1899, 12, 30) + timedelta(days=float(value))
        except Exception:
            pass
    text = clean_text(value)
    if not text:
        return None
    text = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except Exception:
        pass
    if date_parser is not None:
        try:
            return date_parser.parse(text, dayfirst=False, fuzzy=True).replace(tzinfo=None)
        except Exception:
            return None
    return None


def first_present(header_map: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    for alias in aliases:
        idx = header_map.get(norm_col(alias))
        if idx is not None:
            return idx
    return None


def find_header_row(rows: List[List[Any]]) -> Tuple[int, List[str]]:
    best_i, best_score = 0, -1
    for i, row in enumerate(rows[:20]):
        names = [norm_col(x) for x in row]
        if not any(names):
            continue
        score = 0
        for aliases in ALIASES.values():
            if any(norm_col(a) in names for a in aliases):
                score += 1
        if score > best_score:
            best_i, best_score = i, score
    header = [clean_text(x) or f"Column {i+1}" for i, x in enumerate(rows[best_i])] if rows else []
    return best_i, header


def load_sheet_rows_xlsx(path: Path) -> Iterable[Tuple[str, List[List[Any]], Optional[int]]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for ws in wb.worksheets:
            rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
            yield ws.title, rows, None
    finally:
        try:
            wb.close()
        except Exception:
            pass


def load_sheet_rows_xls(path: Path) -> Iterable[Tuple[str, List[List[Any]], Optional[int]]]:
    if xlrd is None:
        raise RuntimeError("xlrd is required for .xls files. Install with: python -m pip install xlrd==2.0.1")
    book = xlrd.open_workbook(str(path))
    for sheet in book.sheets():
        rows = [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
        yield sheet.name, rows, book.datemode


def canonical_symbol(value: Any) -> str:
    text = clean_text(value).upper().replace("/", "").replace(" ", "")
    if text.endswith(".A") or text.endswith(".B"):
        text = text[:-2]
    return text


def boolish(value: Any) -> str:
    text = clean_text(value).lower()
    if text in {"yes", "y", "true", "1"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return clean_text(value)


def infer_asset_class(symbol: str, account: str) -> str:
    s = symbol.upper()
    a = account.lower()
    if "oanda" in a or "forex" in a or re.fullmatch(r"[A-Z]{6}", s):
        return "FX"
    return "Crypto"


def is_trade_row(row: Dict[str, Any]) -> bool:
    return bool(clean_text(row.get("symbol"))) and any(
        row.get(k) is not None for k in ("entry", "exit", "net_profit", "open_time", "close_time")
    )


def is_test_trade_row(row: Dict[str, Any]) -> bool:
    return clean_text(row.get("is_test_trade")).lower() in {"1", "true", "yes", "y"}


def row_pnl(row: Dict[str, Any]) -> Optional[float]:
    return safe_float(row.get("net_profit"))


def row_pnl_currency(row: Dict[str, Any]) -> str:
    return clean_text(row.get("realized_pnl_currency") or row.get("currency")) or "UNKNOWN"


def is_win(row: Dict[str, Any]) -> bool:
    pnl = row_pnl(row)
    return pnl is not None and pnl > 0


def is_loss(row: Dict[str, Any]) -> bool:
    pnl = row_pnl(row)
    return pnl is not None and pnl < 0


def is_be(row: Dict[str, Any]) -> bool:
    be = clean_text(row.get("breakeven")).lower() in {"yes","y","true","1"}
    pnl = row_pnl(row)
    return be or (pnl is not None and abs(pnl) < 1e-12)


def trade_duration_seconds(row: Dict[str, Any]) -> Optional[float]:
    o, c = row.get("open_time"), row.get("close_time")
    if isinstance(o, datetime) and isinstance(c, datetime):
        delta=(c-o).total_seconds()
        if delta < 0:
            return None
        return max(1.0, float(math.ceil(delta)))
    return None


def pip_size_for_symbol(symbol: str) -> float:
    s = clean_text(symbol).upper()
    return 0.01 if s.endswith("JPY") else 0.0001


def signed_price_move(row: Dict[str, Any]) -> Optional[float]:
    entry, exit_p = safe_float(row.get("entry")), safe_float(row.get("exit"))
    if entry is None or exit_p is None:
        return None
    move = exit_p - entry
    side = clean_text(row.get("side")).upper()
    if side.startswith("SELL") or side == "SHORT":
        move = -move
    return move


def stop_pct(row: Dict[str, Any]) -> Optional[float]:
    entry, sl = safe_float(row.get("entry")), safe_float(row.get("stop_loss"))
    if entry and sl and entry > 0:
        return abs(entry - sl) / entry * 100.0
    return None


def target_pct(row: Dict[str, Any]) -> Optional[float]:
    entry, tp = safe_float(row.get("entry")), safe_float(row.get("take_profit"))
    if entry and tp and entry > 0:
        return abs(tp - entry) / entry * 100.0
    return None


def metric_values(rows: List[Dict[str, Any]], key: str) -> List[float]:
    vals = [safe_float(r.get(key)) for r in rows]
    return [v for v in vals if v is not None]


def avg_metric(rows: List[Dict[str, Any]], key: str) -> Optional[float]:
    return avg(metric_values(rows, key))


def safe_min(vals: List[float]) -> Optional[float]:
    return min(vals) if vals else None


def safe_max(vals: List[float]) -> Optional[float]:
    return max(vals) if vals else None


def parse_workbook(path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    warnings: List[str] = []
    trades: List[Dict[str, Any]] = []
    loader = load_sheet_rows_xls if path.suffix.lower() == ".xls" else load_sheet_rows_xlsx
    name_lower = path.name.lower()
    is_bybit_demo = name_lower in {"bybit demo.xlsx", "bybit demo.xlsm", "bybit demo.xls"}
    account_default = "Bybit Demo" if is_bybit_demo else (path.stem.strip() or path.name)

    for sheet_name, rows, datemode in loader(path):
        non_empty = [r for r in rows if any(clean_text(x) for x in r)]
        if not non_empty:
            continue
        header_i, header = find_header_row(non_empty)
        header_map: Dict[str, int] = {}
        for idx, name in enumerate(header):
            n = norm_col(name)
            if n and n not in header_map:
                header_map[n] = idx
        col = {key: first_present(header_map, aliases) for key, aliases in ALIASES.items()}
        signal_cols = [col.get("symbol"), col.get("side"), col.get("entry"), col.get("exit"), col.get("net_profit"), col.get("open_time"), col.get("close_time")]
        if not col.get("symbol") or sum(x is not None for x in signal_cols) < 2:
            continue
        data_rows = non_empty[header_i + 1:]
        for row_index, row in enumerate(data_rows, start=header_i + 2):
            def cell(key: str) -> Any:
                idx = col.get(key)
                if idx is None or idx >= len(row):
                    return None
                return row[idx]
            symbol = canonical_symbol(cell("symbol"))
            if not symbol:
                continue
            open_dt = parse_dt(cell("open_time"), datemode)
            close_dt = parse_dt(cell("close_time"), datemode)
            pnl = safe_float(cell("net_profit"))
            commission = safe_float(cell("commission"))
            swap = safe_float(cell("swap"))
            if pnl is None and commission is None and cell("exit") is None and cell("entry") is None:
                # Avoid treating account/balance rows as trades.
                continue
            account = clean_text(cell("account")) or account_default
            side = clean_text(cell("side")).upper()
            entry = safe_float(cell("entry"))
            exit_price = safe_float(cell("exit"))
            qty = safe_float(cell("qty"))
            balance_after = safe_float(cell("balance_after"))
            stop_loss = safe_float(cell("stop_loss"))
            take_profit = safe_float(cell("take_profit"))
            breakeven = boolish(cell("breakeven"))
            notes_parts = []
            for key in ["notes", "error", "held_through_news", "spiked_out", "early_close"]:
                val = boolish(cell(key)) if key != "notes" else clean_text(cell(key))
                if val:
                    notes_parts.append(f"{key}: {val}" if key != "notes" else val)
            close_for_id = close_dt.isoformat(sep=" ") if close_dt else ""
            trade = {
                "id": f"{path.name}:{sheet_name}:{row_index}:{symbol}:{close_for_id}",
                "source_file": path.name,
                "sheet": sheet_name,
                "row": row_index,
                "account": account,
                "currency": clean_text(cell("currency")) or ("USDT" if is_bybit_demo else ""),
                "realized_pnl_currency": clean_text(cell("currency")) or ("USDT" if is_bybit_demo else ""),
                "asset_class": "Crypto" if is_bybit_demo else infer_asset_class(symbol, account),
                "symbol": symbol,
                "side": side,
                "setup": clean_text(cell("setup")),
                "timeframe": clean_text(cell("timeframe")),
                "open_time": open_dt,
                "close_time": close_dt,
                "qty": qty,
                "entry": entry,
                "exit": exit_price,
                "stop_loss": stop_loss,
                "take_profit": take_profit,
                "commission": commission,
                "swap": swap,
                "net_profit": pnl,
                "balance_after": balance_after,
                "breakeven": breakeven,
                "notes": " | ".join(notes_parts),
                "is_test_trade": clean_text(cell("is_test_trade")),
                "order_id": clean_text(cell("order_id")),
                "fill_count": clean_text(cell("fill_count")),
                "import_source": clean_text(cell("source")),
                "result_pct": safe_float(cell("result_pct")),
                "r_multiple": safe_float(cell("r_multiple")),
            }
            if not trade["side"] and is_bybit_demo:
                trade["side"] = clean_text(cell("side")).upper()
            if trade["result_pct"] is None and trade.get("balance_after") is not None and pnl is not None:
                bal_before = trade["balance_after"] - pnl
                if bal_before > 0:
                    trade["result_pct"] = (pnl / bal_before) * 100.0
            if trade["r_multiple"] is None:
                move = signed_price_move(trade)
                if move is not None and entry is not None and stop_loss is not None and abs(entry - stop_loss) > 0:
                    trade["r_multiple"] = move / abs(entry - stop_loss)
            trade["trade_duration_seconds"] = trade_duration_seconds(trade)
            trades.append(trade)
    if not trades:
        warnings.append(f"No trade rows parsed from {path.name}")
    return trades, warnings


def find_journal_dir(repo_dir: Optional[str]) -> Path:
    candidates: List[Path] = []
    if repo_dir:
        candidates.append(Path(repo_dir) / "journal")
    env_repo = os.environ.get("CODEX_REPO_DIR")
    if env_repo:
        candidates.append(Path(env_repo) / "journal")
    candidates.extend([
        Path("/storage/emulated/0/Download/CODEX-master (4)/CODEX-master/journal"),
        Path.home() / "storage" / "downloads" / "CODEX-master (4)" / "CODEX-master" / "journal",
    ])
    for c in candidates:
        if c.exists() and c.is_dir():
            return c
    download_roots = [Path.home() / "storage" / "downloads", Path("/storage/emulated/0/Download"), Path("/sdcard/Download")]
    for root in download_roots:
        if not root.exists():
            continue
        for found in root.glob("**/CODEX-master/journal"):
            if found.is_dir():
                return found
    raise FileNotFoundError("Could not find CODEX-master/journal under Downloads")


def list_source_workbooks(journal_dir: Path) -> List[Path]:
    out = []
    for path in sorted(journal_dir.iterdir(), key=lambda p: p.name.lower()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
            continue
        if path.name.lower() in EXCLUDED_SOURCE_NAMES:
            continue
        if path.name.startswith("~$"):
            continue
        out.append(path)
    return out


def result_label(pnl: Optional[float], breakeven: str = "") -> str:
    if breakeven.lower() == "yes":
        return "Breakeven"
    if pnl is None:
        return "Unknown"
    if pnl > 0:
        return "Win"
    if pnl < 0:
        return "Loss"
    return "Breakeven"


def pct(num: float, den: float) -> Optional[float]:
    return (num / den) if den else None


def summarize(trades: List[Dict[str, Any]]) -> Dict[str, Any]:
    wins = sum(1 for t in trades if result_label(t.get("net_profit"), t.get("breakeven", "")) == "Win")
    losses = sum(1 for t in trades if result_label(t.get("net_profit"), t.get("breakeven", "")) == "Loss")
    bes = sum(1 for t in trades if result_label(t.get("net_profit"), t.get("breakeven", "")) == "Breakeven")
    known = wins + losses
    pnls = [t.get("net_profit") for t in trades if isinstance(t.get("net_profit"), (int, float))]
    gross_gain = sum(x for x in pnls if x > 0)
    gross_loss = sum(x for x in pnls if x < 0)
    return {
        "trades": len(trades),
        "wins": wins,
        "losses": losses,
        "breakeven": bes,
        "win_rate": pct(wins, known),
        "net_pl": sum(pnls),
        "gross_gain": gross_gain,
        "gross_loss": gross_loss,
        "avg_pl": pct(sum(pnls), len(pnls)) if pnls else None,
        "max_win": max(pnls) if pnls else None,
        "max_loss": min(pnls) if pnls else None,
    }




def _is_long_side(row: Dict[str, Any]) -> bool:
    side = clean_text(row.get("side")).upper()
    return side.startswith("BUY") or side == "LONG"


def _is_short_side(row: Dict[str, Any]) -> bool:
    side = clean_text(row.get("side")).upper()
    return side.startswith("SELL") or side == "SHORT"


def _asset_class(row: Dict[str, Any]) -> str:
    return clean_text(row.get("asset_class") or "").lower()


def _money_by_currency(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    by = defaultdict(list)
    for r in rows:
        pnl = row_pnl(r)
        if pnl is None:
            continue
        by[row_pnl_currency(r)].append(float(pnl))
    currencies = sorted(by.keys())
    net = {c: sum(v) for c, v in by.items()}
    gain = {c: sum(x for x in v if x > 0) for c, v in by.items()}
    loss_abs = {c: [abs(x) for x in v if x < 0] for c, v in by.items()}
    loss = {c: sum(vals) for c, vals in loss_abs.items()}
    return {
        "net_profit_total": net,
        "gross_gain": {c:(gain[c] if gain[c] else 0.0) for c in by.keys()},
        "gross_loss": {c:(loss[c] if loss[c] else 0.0) for c in by.keys()},
        "avg_gain": {c: (avg([x for x in by[c] if x > 0]) or 0.0) for c in by.keys()},
        "avg_loss": {c: (avg(loss_abs[c]) or 0.0) for c in by.keys()},
        "max_gain": {c: (safe_max([x for x in by[c] if x > 0]) or 0.0) for c in by.keys()},
        "max_loss": {c: (safe_max(loss_abs[c]) or 0.0) for c in by.keys()},
        "currencies": currencies,
        "mixed_currency": len(currencies) > 1,
    }


def _drawdowns(rows: List[Dict[str, Any]]) -> List[float]:
    by_acct = defaultdict(list)
    for r in rows:
        bal = safe_float(r.get("balance_after"))
        dt = r.get("close_time") or r.get("open_time")
        if bal is None or not isinstance(dt, datetime):
            continue
        by_acct[clean_text(r.get("account")) or "UNKNOWN"].append((dt, bal))
    dds=[]
    for vals in by_acct.values():
        vals.sort(key=lambda x:x[0])
        if len(vals) < 2:
            continue
        peak = vals[0][1]
        for _, bal in vals:
            peak = max(peak, bal)
            if peak > 0:
                dd=(peak - bal) / peak * 100.0
                if dd > 0 and math.isfinite(dd):
                    dds.append(dd)
    return dds


def _streak(rows: List[Dict[str, Any]], want_win: bool) -> Dict[str, Any]:
    best=None;cur=[]
    def close(st):
        nonlocal best
        if not st: return
        if best is None or len(st) > len(best): best=list(st)
    for r in sorted(rows, key=lambda t: (t.get("close_time") or t.get("open_time") or datetime.min, str(t.get("id") or ""))):
        ok = is_win(r) if want_win else is_loss(r)
        if ok: cur.append(r)
        else: close(cur);cur=[]
    close(cur)
    if not best:
        return None
    syms=defaultdict(int)
    for r in best: syms[clean_text(r.get("symbol"))]+=1
    st=best[0].get("close_time") or best[0].get("open_time")
    en=best[-1].get("close_time") or best[-1].get("open_time")
    return {
        "type": "winning" if want_win else "losing",
        "trade_count": len(best), "start_time": st, "end_time": en,
        "elapsed_seconds": (en-st).total_seconds() if isinstance(st,datetime) and isinstance(en,datetime) else None,
        "dominant_symbol": max(syms.items(), key=lambda kv: kv[1])[0] if syms else None,
        "symbol_counts": dict(syms),
        "net_r_multiple": sum(v for v in [safe_float(r.get("r_multiple")) for r in best] if v is not None),
        "net_result_pct": sum(v for v in [safe_float(r.get("result_pct")) for r in best] if v is not None),
        "trade_ids": [r.get("id") for r in best if r.get("id")],
    }

def _market_bucket(rows: List[Dict[str, Any]], label: str) -> Dict[str, Any]:
    pnls=[row_pnl(r) for r in rows if row_pnl(r) is not None]
    wins=[r for r in rows if is_win(r)]; losses=[r for r in rows if is_loss(r)]; bes=[r for r in rows if is_be(r)]
    known=len(wins)+len(losses)
    totals=_money_by_currency(rows)
    losses_abs=[abs(v) for v in pnls if v < 0]
    gains=[v for v in pnls if v > 0]
    metric_sources={}
    def src(metric_key, key, fn):
        vals=[(r,safe_float(r.get(key))) for r in rows if safe_float(r.get(key)) is not None]
        if not vals: return None
        row,val=fn(vals,key=lambda t:t[1])
        return {"id":row.get("id"),"symbol":row.get("symbol"),"asset_class":clean_text(row.get("asset_class")).lower(),"side":row.get("side"),"open_time":row.get("open_time"),"close_time":row.get("close_time"),"date":row.get("close_time") or row.get("open_time"),"currency":row_pnl_currency(row),"account":row.get("account"),"source":row.get("import_source"),"timeframe":row.get("timeframe"),"net_profit":row.get("net_profit"),"result_pct":row.get("result_pct"),"r_multiple":row.get("r_multiple"),"trade_duration_seconds":row.get("trade_duration_seconds"),"metric_key":metric_key,"metric_value":val}
    metric_sources["min_result_pct"]=src("min_result_pct","result_pct",min)
    metric_sources["max_result_pct"]=src("max_result_pct","result_pct",max)
    metric_sources["min_r_multiple"]=src("min_r_multiple","r_multiple",min)
    metric_sources["max_r_multiple"]=src("max_r_multiple","r_multiple",max)
    metric_sources["max_gain"]=src("max_gain","net_profit",max)
    _ml=src("max_loss","net_profit",min)
    if _ml and _ml.get("metric_value") is not None: _ml["metric_value"]=abs(_ml["metric_value"])
    metric_sources["max_loss"]=_ml
    return {
        "label": label, "trades": len(rows), "wins": len(wins), "losses": len(losses), "break_even": len(bes),
        "win_rate_pct": (len(wins)/known*100.0) if known else None,
        "net_profit_total": sum(pnls), "gross_gain": sum(gains), "gross_loss": (sum(losses_abs) if losses_abs else None),
        "avg_gain": avg(gains), "avg_loss": avg(losses_abs),
        "max_gain": safe_max(gains), "max_loss": safe_max(losses_abs),
        "avg_result_pct": avg_metric(rows, "result_pct"), "min_result_pct": safe_min(metric_values(rows, "result_pct")), "max_result_pct": safe_max(metric_values(rows, "result_pct")),
        "avg_r_multiple": avg_metric(rows, "r_multiple"), "min_r_multiple": safe_min(metric_values(rows, "r_multiple")), "max_r_multiple": safe_max(metric_values(rows, "r_multiple")),
        "avg_stop_pct": avg([stop_pct(t) for t in rows if stop_pct(t) is not None]),
        "avg_target_pct": avg([target_pct(t) for t in rows if target_pct(t) is not None]),
        "avg_duration_seconds": avg_metric(rows, "trade_duration_seconds"),
        "money_by_currency": totals,
        "metric_sources": metric_sources,
    }

def compute_journal_stats_replica(trades: List[Dict[str, Any]]) -> Dict[str, Any]:
    live=[t for t in trades if is_trade_row(t) and not is_test_trade_row(t)]
    fx=[t for t in live if _asset_class(t)=="fx"]; crypto=[t for t in live if _asset_class(t)=="crypto"]
    wins=[t for t in live if is_win(t)]; losses=[t for t in live if is_loss(t)]; be=[t for t in live if is_be(t)]
    longs=[t for t in live if _is_long_side(t)]; shorts=[t for t in live if _is_short_side(t)]
    by_inst=instrument_stats(live)
    dds=_drawdowns(live)
    totals=_market_bucket(live, "Overall")
    totals.update({
        "long_trades":len(longs),"short_trades":len(shorts),
        "long_wins":sum(1 for t in longs if is_win(t)),"long_losses":sum(1 for t in longs if is_loss(t)),"long_break_even":sum(1 for t in longs if is_be(t)),
        "short_wins":sum(1 for t in shorts if is_win(t)),"short_losses":sum(1 for t in shorts if is_loss(t)),"short_break_even":sum(1 for t in shorts if is_be(t)),
        "fx_win_rate_pct": _market_bucket(fx,"Forex")["win_rate_pct"], "crypto_win_rate_pct": _market_bucket(crypto,"Crypto")["win_rate_pct"],
        "avg_stop_pct_winners": avg([stop_pct(t) for t in wins if stop_pct(t) is not None]), "avg_stop_pct_losers": avg([stop_pct(t) for t in losses if stop_pct(t) is not None]),
        "avg_target_pct_winners": avg([target_pct(t) for t in wins if target_pct(t) is not None]), "avg_target_pct_losers": avg([target_pct(t) for t in losses if target_pct(t) is not None]),
        "avg_result_pct_winners": avg_metric(wins,"result_pct"), "avg_result_pct_losers": avg_metric(losses,"result_pct"),
        "avg_r_multiple_winners": avg_metric(wins,"r_multiple"), "avg_r_multiple_losers": avg_metric(losses,"r_multiple"),
        "min_trade_duration_seconds": safe_min(metric_values(live,"trade_duration_seconds")), "max_trade_duration_seconds": safe_max(metric_values(live,"trade_duration_seconds")),
        "avg_winner_duration_seconds": avg_metric(wins,"trade_duration_seconds"), "avg_loser_duration_seconds": avg_metric(losses,"trade_duration_seconds"),
        "avg_fx_duration_seconds": avg_metric(fx,"trade_duration_seconds"),"avg_crypto_duration_seconds": avg_metric(crypto,"trade_duration_seconds"),
        "min_fx_trade_duration_seconds": safe_min(metric_values(fx,"trade_duration_seconds")),"max_fx_trade_duration_seconds": safe_max(metric_values(fx,"trade_duration_seconds")),
        "min_crypto_trade_duration_seconds": safe_min(metric_values(crypto,"trade_duration_seconds")),"max_crypto_trade_duration_seconds": safe_max(metric_values(crypto,"trade_duration_seconds")),
        "max_drawdown_pct": safe_max(dds) if len(dds)>=1 else None, "avg_drawdown_pct": avg(dds) if len(dds)>=1 else None, "min_drawdown_pct": safe_min(dds) if len(dds)>=1 else None,
    })
    def leader(rows, key):
        cand=[r for r in rows if r.get(key,0)>0]
        return max(cand,key=lambda x:(x.get(key,0),x.get("symbol",""))) if cand else None
    groups={
      "overview": totals,
      "risk_expectancy": {k: totals.get(k) for k in ["avg_stop_pct","avg_target_pct","avg_result_pct","avg_r_multiple","avg_stop_pct_winners","avg_stop_pct_losers","avg_target_pct_winners","avg_target_pct_losers","avg_result_pct_winners","avg_result_pct_losers","avg_r_multiple_winners","avg_r_multiple_losers","max_drawdown_pct","avg_drawdown_pct","min_drawdown_pct"]},
      "duration": {"overall_avg_seconds":totals.get("avg_duration_seconds"),"overall_shortest_seconds":totals.get("min_trade_duration_seconds"),"overall_longest_seconds":totals.get("max_trade_duration_seconds"),"overall_avg_winner_seconds":totals.get("avg_winner_duration_seconds"),"overall_avg_loser_seconds":totals.get("avg_loser_duration_seconds"),"overall_longest_winner_seconds":safe_max(metric_values(wins,"trade_duration_seconds")),"overall_longest_loser_seconds":safe_max(metric_values(losses,"trade_duration_seconds")),"fx_avg_seconds":totals.get("avg_fx_duration_seconds"),"fx_shortest_seconds":totals.get("min_fx_trade_duration_seconds"),"fx_longest_seconds":totals.get("max_fx_trade_duration_seconds"),"fx_avg_winner_seconds":avg_metric([t for t in fx if is_win(t)],"trade_duration_seconds"),"fx_avg_loser_seconds":avg_metric([t for t in fx if is_loss(t)],"trade_duration_seconds"),"fx_shortest_winner_seconds":safe_min(metric_values([t for t in fx if is_win(t)],"trade_duration_seconds")),"fx_shortest_loser_seconds":safe_min(metric_values([t for t in fx if is_loss(t)],"trade_duration_seconds")),"fx_longest_winner_seconds":safe_max(metric_values([t for t in fx if is_win(t)],"trade_duration_seconds")),"fx_longest_loser_seconds":safe_max(metric_values([t for t in fx if is_loss(t)],"trade_duration_seconds")),"crypto_avg_seconds":totals.get("avg_crypto_duration_seconds"),"crypto_shortest_seconds":totals.get("min_crypto_trade_duration_seconds"),"crypto_longest_seconds":totals.get("max_crypto_trade_duration_seconds"),"crypto_avg_winner_seconds":avg_metric([t for t in crypto if is_win(t)],"trade_duration_seconds"),"crypto_avg_loser_seconds":avg_metric([t for t in crypto if is_loss(t)],"trade_duration_seconds"),"crypto_shortest_winner_seconds":safe_min(metric_values([t for t in crypto if is_win(t)],"trade_duration_seconds")),"crypto_shortest_loser_seconds":safe_min(metric_values([t for t in crypto if is_loss(t)],"trade_duration_seconds")),"crypto_longest_winner_seconds":safe_max(metric_values([t for t in crypto if is_win(t)],"trade_duration_seconds")),"crypto_longest_loser_seconds":safe_max(metric_values([t for t in crypto if is_loss(t)],"trade_duration_seconds")),"metric_sources":{}},
      "by_market": {"overall": _market_bucket(live,"Overall"),"fx": _market_bucket(fx,"Forex"),"crypto": _market_bucket(crypto,"Crypto")},
      "market_breakdown": [_market_bucket(live,"Overall"),_market_bucket(fx,"Forex"),_market_bucket(crypto,"Crypto")],
      "leaders": {"most_wins_instrument": leader(by_inst,"wins"),"most_losses_instrument": leader(by_inst,"losses"),"fx_most_wins_instrument": leader([r for r in by_inst if clean_text(r.get("asset_class")).lower()=="fx"],"wins"),"fx_most_losses_instrument": leader([r for r in by_inst if clean_text(r.get("asset_class")).lower()=="fx"],"losses"),"crypto_most_wins_instrument": leader([r for r in by_inst if clean_text(r.get("asset_class")).lower()=="crypto"],"wins"),"crypto_most_losses_instrument": leader([r for r in by_inst if clean_text(r.get("asset_class")).lower()=="crypto"],"losses")},
      "direction": {"long_trades":len(longs),"short_trades":len(shorts),"long_win_rate_pct":(sum(1 for t in longs if is_win(t))/ (sum(1 for t in longs if is_win(t) or is_loss(t)) or 1) *100.0 if longs else None),"short_win_rate_pct":(sum(1 for t in shorts if is_win(t))/ (sum(1 for t in shorts if is_win(t) or is_loss(t)) or 1) *100.0 if shorts else None)},
      "streaks": {"longest_winning": _streak(live,True), "longest_losing": _streak(live,False)}
    }
    return {"totals": totals, "groups": groups, "by_instrument": by_inst}

def instrument_stats(trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    buckets=defaultdict(list)
    for t in trades: buckets[t.get("symbol") or "UNKNOWN"].append(t)
    out=[]
    for sym,items in buckets.items():
        wins=[x for x in items if is_win(x)]; losses=[x for x in items if is_loss(x)]
        def dist(kind, row):
            e,sl,tp=safe_float(row.get("entry")),safe_float(row.get("stop_loss")),safe_float(row.get("take_profit"))
            if e is None: return None
            if kind=="sl" and sl is not None: return abs(e-sl)
            if kind=="tp" and tp is not None: return abs(tp-e)
            return None
        durs=metric_values(items,"trade_duration_seconds")
        out.append({"symbol":sym,"asset_class":clean_text(items[0].get("asset_class","")).lower(),"total_trades":len(items),"long_trades":sum(1 for x in items if _is_long_side(x)),"short_trades":sum(1 for x in items if _is_short_side(x)),"wins":len(wins),"losses":len(losses),"break_even":sum(1 for x in items if is_be(x)),"long_wins":sum(1 for x in items if _is_long_side(x) and is_win(x)),"long_losses":sum(1 for x in items if _is_long_side(x) and is_loss(x)),"long_break_even":sum(1 for x in items if _is_long_side(x) and is_be(x)),"short_wins":sum(1 for x in items if _is_short_side(x) and is_win(x)),"short_losses":sum(1 for x in items if _is_short_side(x) and is_loss(x)),"short_break_even":sum(1 for x in items if _is_short_side(x) and is_be(x)),"avg_trade_duration_seconds":avg(durs),"min_trade_duration_seconds":safe_min(durs),"max_trade_duration_seconds":safe_max(durs),"avg_sl_distance_pips":avg([abs(safe_float(x.get("entry"))-safe_float(x.get("stop_loss")))/pip_size_for_symbol(sym) for x in items if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("stop_loss")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_tp_distance_pips":avg([abs(safe_float(x.get("take_profit"))-safe_float(x.get("entry")))/pip_size_for_symbol(sym) for x in items if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("take_profit")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_sl_distance_quote":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("sl",x) for x in items if dist("sl",x) is not None])),"avg_tp_distance_quote":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("tp",x) for x in items if dist("tp",x) is not None])),"avg_sl_distance_pips_wins":avg([abs(safe_float(x.get("entry"))-safe_float(x.get("stop_loss")))/pip_size_for_symbol(sym) for x in wins if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("stop_loss")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_sl_distance_pips_losses":avg([abs(safe_float(x.get("entry"))-safe_float(x.get("stop_loss")))/pip_size_for_symbol(sym) for x in losses if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("stop_loss")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_tp_distance_pips_wins":avg([abs(safe_float(x.get("take_profit"))-safe_float(x.get("entry")))/pip_size_for_symbol(sym) for x in wins if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("take_profit")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_tp_distance_pips_losses":avg([abs(safe_float(x.get("take_profit"))-safe_float(x.get("entry")))/pip_size_for_symbol(sym) for x in losses if clean_text(items[0].get("asset_class","")).lower()=="fx" and safe_float(x.get("entry")) is not None and safe_float(x.get("take_profit")) is not None]) if clean_text(items[0].get("asset_class","")).lower()=="fx" else None,"avg_sl_distance_quote_wins":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("sl",x) for x in wins if dist("sl",x) is not None])),"avg_sl_distance_quote_losses":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("sl",x) for x in losses if dist("sl",x) is not None])),"avg_tp_distance_quote_wins":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("tp",x) for x in wins if dist("tp",x) is not None])),"avg_tp_distance_quote_losses":(None if clean_text(items[0].get("asset_class","")).lower()=="fx" else avg([dist("tp",x) for x in losses if dist("tp",x) is not None])),"quote_currency":items[0].get("currency") or "USDT"})
    out.sort(key=lambda r:(-int(r["total_trades"] or 0),r["symbol"]))
    return out

def instrument_display_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out=[]
    for r in rows:
        out.append({"Symbol":r.get("symbol"),"Asset":r.get("asset_class"),"Total Trades":r.get("total_trades"),"Long Trades":r.get("long_trades"),"Short Trades":r.get("short_trades"),"Wins":r.get("wins"),"Losses":r.get("losses"),"Breakeven":r.get("break_even"),"Long Wins":r.get("long_wins"),"Long Losses":r.get("long_losses"),"Short Wins":r.get("short_wins"),"Short Losses":r.get("short_losses"),"Avg SL W":(r.get("avg_sl_distance_pips_wins") if clean_text(r.get("asset_class")).lower()=="fx" else r.get("avg_sl_distance_quote_wins")),"Avg SL L":(r.get("avg_sl_distance_pips_losses") if clean_text(r.get("asset_class")).lower()=="fx" else r.get("avg_sl_distance_quote_losses")),"Avg TP W":(r.get("avg_tp_distance_pips_wins") if clean_text(r.get("asset_class")).lower()=="fx" else r.get("avg_tp_distance_quote_wins")),"Avg TP L":(r.get("avg_tp_distance_pips_losses") if clean_text(r.get("asset_class")).lower()=="fx" else r.get("avg_tp_distance_quote_losses")),"Avg Duration":r.get("avg_trade_duration_seconds"),"Shortest Duration":r.get("min_trade_duration_seconds"),"Longest Duration":r.get("max_trade_duration_seconds")})
    return out

def avg(values: Iterable[Any]) -> Optional[float]:
    nums = [float(v) for v in values if isinstance(v, (int, float)) and not math.isnan(float(v))]
    return sum(nums) / len(nums) if nums else None


def sorted_trades(trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def key(t: Dict[str, Any]) -> Tuple[datetime, str]:
        dt = t.get("close_time") or t.get("open_time") or datetime.min
        return (dt if isinstance(dt, datetime) else datetime.min, str(t.get("id") or ""))
    return sorted(trades, key=key, reverse=True)


def equity_rows(trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    chronological = list(reversed(sorted_trades(trades)))
    equity = 0.0
    rows = []
    for i, t in enumerate(chronological, start=1):
        pnl = t.get("net_profit") if isinstance(t.get("net_profit"), (int, float)) else 0.0
        if isinstance(t.get("balance_after"), (int, float)):
            equity = float(t["balance_after"])
        else:
            equity += float(pnl)
        rows.append({
            "#": i,
            "Close Time": t.get("close_time"),
            "Symbol": t.get("symbol"),
            "Net P/L": pnl,
            "Equity": equity,
        })
    return rows


def calendar_rows(trades: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_month_day: Dict[str, Dict[int, float]] = defaultdict(lambda: defaultdict(float))
    for t in trades:
        dt = t.get("close_time") or t.get("open_time")
        pnl = t.get("net_profit")
        if not isinstance(dt, datetime) or not isinstance(pnl, (int, float)):
            continue
        by_month_day[dt.strftime("%Y-%m")][dt.day] += float(pnl)
    out = []
    for month in sorted(by_month_day.keys(), reverse=True):
        row = {"Month": month}
        total = 0.0
        for day in range(1, 32):
            v = by_month_day[month].get(day)
            row[str(day)] = v if v not in (None, 0) else None
            total += float(v or 0)
        row["Monthly Total"] = total
        out.append(row)
    return out


def style_sheet(ws, max_col: int, freeze: Optional[str] = None) -> None:
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = freeze
    thin = Side(style="thin", color=BORDER)
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=DARK)
            cell.font = Font(color=TEXT, name="Calibri", size=11)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 16
    ws.auto_filter.ref = ws.dimensions


def write_table(ws, headers: List[str], rows: List[Dict[str, Any]], start_row: int = 1, start_col: int = 1) -> None:
    for j, h in enumerate(headers, start=start_col):
        ws.cell(start_row, j, h)
    for i, row in enumerate(rows, start=start_row + 1):
        for j, h in enumerate(headers, start=start_col):
            value = row.get(h)
            c = ws.cell(i, j, value)
            if isinstance(value, datetime):
                c.number_format = "yyyy-mm-dd hh:mm"
            elif h.lower().endswith("rate"):
                c.number_format = "0.00%"
            elif isinstance(value, (int, float)):
                c.number_format = "#,##0.00"


def set_pl_format(ws, ranges: Iterable[str]) -> None:
    for rng in ranges:
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], font=Font(color=GREEN)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=Font(color=RED)))


def fmt_money_breakdown(bucket: Dict[str, Any], key: str) -> str:
    m=(bucket.get("money_by_currency") or {}).get(key) or {}
    if not isinstance(m, dict) or not m: return "—"
    return " / ".join(f"{c} {m.get(c,0):.2f}" for c in sorted(m.keys()))

def write_dashboard(wb: Workbook, trades: List[Dict[str, Any]], sources: List[Path], warnings: List[str]) -> None:
    ws = wb.active
    ws.title = "Dashboard"
    stats = compute_journal_stats_replica(trades)
    g = stats["groups"]
    ws.append(["Trading Journal Android Replica", "Generated", datetime.now()])
    ws.append(["Source workbooks", len(sources), "Source folder", str(sources[0].parent if sources else "")])
    def section(title, rows):
        ws.append([title, "Value"])
        for k,v in rows: ws.append([k,v])
        ws.append([None,None])
    def market_rows(bucket):
        return [("Trades",bucket.get("trades")),("Wins",bucket.get("wins")),("Losses",bucket.get("losses")),("Break-even",bucket.get("break_even")),("Win rate",bucket.get("win_rate_pct")),("Net P/L",fmt_money_breakdown(bucket,"net_profit_total")),("Gross gain",fmt_money_breakdown(bucket,"gross_gain")),("Gross loss",fmt_money_breakdown(bucket,"gross_loss")),("Avg result %",bucket.get("avg_result_pct")),("Max loss %",bucket.get("min_result_pct")),("Max win %",bucket.get("max_result_pct")),("Avg R",bucket.get("avg_r_multiple")),("Max R loss",bucket.get("min_r_multiple")),("Max R win",bucket.get("max_r_multiple")),("Max gain",fmt_money_breakdown(bucket,"max_gain")),("Max loss",fmt_money_breakdown(bucket,"max_loss")),("Avg stop %",bucket.get("avg_stop_pct")),("Avg target %",bucket.get("avg_target_pct")),("Avg duration",bucket.get("avg_duration_seconds"))]
    bym=g["by_market"]
    section("Overall", market_rows(bym["overall"]))
    section("Winners", [("Avg stop %",stats["totals"].get("avg_stop_pct_winners")),("Avg target %",stats["totals"].get("avg_target_pct_winners")),("Avg result %",stats["totals"].get("avg_result_pct_winners")),("Avg R",stats["totals"].get("avg_r_multiple_winners"))])
    section("Losers", [("Avg stop %",stats["totals"].get("avg_stop_pct_losers")),("Avg target %",stats["totals"].get("avg_target_pct_losers")),("Avg result %",stats["totals"].get("avg_result_pct_losers")),("Avg R",stats["totals"].get("avg_r_multiple_losers"))])
    section("Drawdown", [("Max drawdown",g["risk_expectancy"].get("max_drawdown_pct")),("Avg drawdown",g["risk_expectancy"].get("avg_drawdown_pct")),("Min drawdown",g["risk_expectancy"].get("min_drawdown_pct"))])
    section("Duration", [("Overall avg",stats["totals"].get("avg_duration_seconds")),("Overall shortest",stats["totals"].get("min_trade_duration_seconds")),("Overall longest",stats["totals"].get("max_trade_duration_seconds")),("FX shortest",stats["totals"].get("min_fx_trade_duration_seconds")),("FX longest",stats["totals"].get("max_fx_trade_duration_seconds")),("Crypto shortest",stats["totals"].get("min_crypto_trade_duration_seconds")),("Crypto longest",stats["totals"].get("max_crypto_trade_duration_seconds"))])
    section("FX", market_rows(bym["fx"]))
    section("Crypto", market_rows(bym["crypto"]))
    L=g["leaders"]
    section("Instrument leaders", [("Overall most wins", (L.get("most_wins_instrument") or {}).get("symbol")),("Overall most losses", (L.get("most_losses_instrument") or {}).get("symbol")),("FX most wins", (L.get("fx_most_wins_instrument") or {}).get("symbol")),("FX most losses", (L.get("fx_most_losses_instrument") or {}).get("symbol")),("Crypto most wins", (L.get("crypto_most_wins_instrument") or {}).get("symbol")),("Crypto most losses", (L.get("crypto_most_losses_instrument") or {}).get("symbol"))])
    ws.append(["Money by currency","Net","Gross gain","Gross loss"])
    mb = bym["overall"].get("money_by_currency",{})
    for c in mb.get("currencies",[]): ws.append([c, mb["net_profit_total"].get(c), mb["gross_gain"].get(c), mb["gross_loss"].get(c)])
    ws.append(["Diagnostics / warnings", None])
    for w in warnings: ws.append([w,None])

def build_output(journal_dir: Path, output_path: Path) -> Tuple[int, int, List[str]]:
    sources = list_source_workbooks(journal_dir)
    all_trades: List[Dict[str, Any]] = []
    warnings: List[str] = []
    if not sources:
        warnings.append(f"No journal workbooks found in {journal_dir}")
    rows_by_source = defaultdict(int)
    for path in sources:
        try:
            trades, w = parse_workbook(path)
            all_trades.extend(trades)
            rows_by_source[path.name] += len(trades)
            warnings.extend(w)
        except Exception as exc:
            warnings.append(f"{path.name}: {exc}")

    # Dedupe by stable trade signature.
    seen = set()
    deduped = []
    for t in all_trades:
        key = (t.get("symbol"), t.get("side"), t.get("open_time"), t.get("close_time"), t.get("qty"), t.get("entry"), t.get("exit"), t.get("net_profit"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(t)
    all_trades = deduped

    stats = compute_journal_stats_replica(all_trades)
    wb = Workbook()
    write_dashboard(wb, all_trades, sources, warnings)

    trades_ws = wb.create_sheet("All Trades")
    trade_headers = ["Open Time", "Close Time", "Account", "Symbol", "Side", "Timeframe", "Test", "Setup", "Qty", "Entry", "Exit", "Stop Loss", "Target", "Commission", "Net Profit", "Profit %", "R-Multiple", "Balance After", "Trade Duration", "Breakeven", "Chart", "Actions", "Source", "Notes", "Order ID", "Fill Count"]
    rows = []
    for t in sorted_trades(all_trades):
        rows.append({
            "Open Time": t.get("open_time"),
            "Close Time": t.get("close_time"),
            "Account": t.get("account"),
            "Symbol": t.get("symbol"),
            "Side": t.get("side"),
            "Timeframe": t.get("timeframe"),
            "Test": t.get("is_test_trade"),
            "Setup": t.get("setup"),
            "Qty": t.get("qty"),
            "Entry": t.get("entry"),
            "Exit": t.get("exit"),
            "Stop Loss": t.get("stop_loss"),
            "Target": t.get("take_profit"),
            "Commission": t.get("commission"),
            "Net Profit": t.get("net_profit"),
            "Profit %": t.get("result_pct"),
            "R-Multiple": t.get("r_multiple"),
            "Balance After": t.get("balance_after"),
            "Trade Duration": t.get("trade_duration_seconds"),
            "Breakeven": t.get("breakeven"),
            "Chart": "",
            "Actions": "",
            "Source": t.get("import_source") or f"{t.get('source_file')} / {t.get('sheet')} / row {t.get('row')}",
            "Notes": t.get("notes"),
            "Order ID": t.get("order_id"),
            "Fill Count": t.get("fill_count"),
        })
    write_table(trades_ws, trade_headers, rows)
    style_sheet(trades_ws, len(trade_headers), freeze="A2")
    trades_ws.column_dimensions["A"].width = 20
    trades_ws.column_dimensions["B"].width = 20
    trades_ws.column_dimensions["T"].width = 36
    trades_ws.column_dimensions["U"].width = 44
    set_pl_format(trades_ws, [f"O2:Q{max(2, len(rows)+1)}"])

    inst_ws = wb.create_sheet("Instrument Averages")
    inst_headers = ["Symbol", "Asset", "Total Trades", "Long Trades", "Short Trades", "Wins", "Losses", "Breakeven", "Long Wins", "Long Losses", "Short Wins", "Short Losses", "Avg SL W", "Avg SL L", "Avg TP W", "Avg TP L", "Avg Duration", "Shortest Duration", "Longest Duration"]
    inst_rows = instrument_display_rows(stats["by_instrument"])
    write_table(inst_ws, inst_headers, inst_rows)
    style_sheet(inst_ws, len(inst_headers), freeze="A2")
    set_pl_format(inst_ws, [f"H2:I{max(2, len(inst_rows)+1)}"])

    cal_ws = wb.create_sheet("PL Calendar")
    cal_headers = ["Month"] + [str(i) for i in range(1, 32)] + ["Monthly Total"]
    cal_data = calendar_rows(all_trades)
    write_table(cal_ws, cal_headers, cal_data)
    style_sheet(cal_ws, len(cal_headers), freeze="B2")
    for c in range(2, 34):
        cal_ws.column_dimensions[get_column_letter(c)].width = 9
    set_pl_format(cal_ws, [f"B2:AF{max(2, len(cal_data)+1)}", f"AG2:AG{max(2, len(cal_data)+1)}"])

    eq_ws = wb.create_sheet("Equity Curve")
    eq_headers = ["#", "Close Time", "Symbol", "Net P/L", "Equity"]
    eq_data = equity_rows(all_trades)
    write_table(eq_ws, eq_headers, eq_data)
    style_sheet(eq_ws, len(eq_headers), freeze="A2")
    eq_ws.column_dimensions["B"].width = 20
    set_pl_format(eq_ws, [f"D2:E{max(2, len(eq_data)+1)}"])
    if len(eq_data) >= 2:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.y_axis.title = "Equity"
        chart.x_axis.title = "Trade #"
        data = Reference(eq_ws, min_col=5, min_row=1, max_row=len(eq_data)+1)
        cats = Reference(eq_ws, min_col=1, min_row=2, max_row=len(eq_data)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 28
        eq_ws.add_chart(chart, "G2")

    diag_ws = wb.create_sheet("Diagnostics")
    diag_rows = []
    diag_rows.append({"Item": "Journal folder", "Value": str(journal_dir)})
    diag_rows.append({"Item": "Output workbook", "Value": str(output_path)})
    diag_rows.append({"Item": "Source workbooks", "Value": len(sources)})
    diag_rows.append({"Item": "Parsed trades", "Value": len(all_trades)})
    diag_rows.append({"Item": "Test rows excluded from stats", "Value": sum(1 for t in all_trades if is_test_trade_row(t))})
    bybit_count = sum(c for n, c in rows_by_source.items() if Path(n).stem.lower() == "bybit demo")
    diag_rows.append({"Item": "Bybit Demo parsed row count", "Value": bybit_count})
    diag_rows.append({"Item": "Money by currency", "Value": str(stats.get("totals", {}).get("money_by_currency", {}))})
    for src in sources:
        diag_rows.append({"Item": "Source file", "Value": src.name})
        diag_rows.append({"Item": f"Parsed rows ({src.name})", "Value": rows_by_source.get(src.name, 0)})
    diag_rows.append({"Item": "Missing balance_after_trade", "Value": sum(1 for t in all_trades if t.get("balance_after") is None)})
    diag_rows.append({"Item": "Missing stop_loss", "Value": sum(1 for t in all_trades if t.get("stop_loss") is None)})
    diag_rows.append({"Item": "Missing take_profit", "Value": sum(1 for t in all_trades if t.get("take_profit") is None)})
    diag_rows.append({"Item": "Missing result_pct", "Value": sum(1 for t in all_trades if t.get("result_pct") is None)})
    diag_rows.append({"Item": "Missing r_multiple", "Value": sum(1 for t in all_trades if t.get("r_multiple") is None)})
    diag_rows.append({"Item": "Missing open_time", "Value": sum(1 for t in all_trades if t.get("open_time") is None)})
    diag_rows.append({"Item": "Missing close_time", "Value": sum(1 for t in all_trades if t.get("close_time") is None)})
    for warning in warnings:
        diag_rows.append({"Item": "Warning", "Value": warning})
    write_table(diag_ws, ["Item", "Value"], diag_rows)
    style_sheet(diag_ws, 2, freeze="A2")
    diag_ws.column_dimensions["A"].width = 24
    diag_ws.column_dimensions["B"].width = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(sources), len(all_trades), warnings


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--repo", default=None, help="Path to CODEX-master repo. Defaults to Downloads/CODEX-master (4)/CODEX-master.")
    ap.add_argument("--output", default=None, help="Output .xlsx path. Defaults to repo/journal/TradingJournal_Android_Replica.xlsx")
    args = ap.parse_args(argv)
    try:
        journal_dir = find_journal_dir(args.repo)
        output_path = Path(args.output) if args.output else journal_dir / OUTPUT_NAME
        source_count, trade_count, warnings = build_output(journal_dir, output_path)
        print(f"OK: wrote {output_path}")
        print(f"Source workbooks: {source_count}")
        print(f"Parsed trades: {trade_count}")
        if warnings:
            print("Warnings:")
            for warning in warnings[:20]:
                print(f"- {warning}")
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
