from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from tools.backfill_legacy_crypto_trade_metadata import run_backfill
from tools.master_journal_workbook import (
    TRADE_LOG_DATA_START_ROW,
    TRADE_NUMBER_HEADER,
    build_master_journal_workbook,
    _trade_log_header_map,
)


def _journal(path: Path, rows):
    snapshot = {
        "items": rows,
        "stats": {"totals": {}, "groups": {"by_market": {}, "risk_expectancy": {}, "leaders": {}, "duration": {}}},
        "balances": [],
    }
    build_master_journal_workbook(snapshot, path)


def _legacy(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "TRADE LOG"
    for row_idx, payload in enumerate(rows, start=2):
        for col, value in payload.items():
            ws[f"{col}{row_idx}"] = value
    wb.save(path)
    wb.close()


def _legacy_row(number=1, *, local=True, utc=False, error_cell=False):
    row = {
        "A": number,
        "B": "BTCUSDT",
        "AJ": "LONG",
        "Z": 100.0,
        "AA": 110.0,
        "AC": 95.0,
        "AE": 120.0,
        "N": 1.5,
        "M": 2.0,
        "BL": datetime(2026, 1, 1, 10, 30),
        "BM": 300,
        "BN": 101.0,
        "BO": 1.0,
        "BP": 19.0,
        "BQ": datetime(2026, 1, 1, 10, 45),
        "BR": 600,
        "BV": 103.0,
        "BW": "3%",
        "BX": "4%",
    }
    if local:
        row.update({"W": datetime(2026, 1, 1, 10, 0), "X": datetime(2026, 1, 1, 11, 0)})
    if utc:
        row.update({"U": datetime(2026, 1, 1, 0, 0), "V": datetime(2026, 1, 1, 1, 0)})
    if error_cell:
        row["BV"] = "#REF!"
    return row


def _repo_row(row_id="r1"):
    return {
        "id": row_id,
        "row_type": "trade",
        "account": "BINANCE",
        "asset_class": "crypto",
        "symbol": "BTCUSDT",
        "side": "BUY",
        "open_time": "2026-01-01T10:00:00",
        "close_time": "2026-01-01T11:00:00",
        "entry_price": 100.0,
        "exit_price": 110.0,
        "stop_loss": 95.0,
        "take_profit": 120.0,
        "r_multiple": 1.5,
        "result_pct": 2.0,
        "net_profit": 10.0,
    }


def test_backfill_exact_match_dry_run_and_apply(tmp_path: Path):
    journal = tmp_path / "Trading Journal.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    _journal(journal, [_repo_row()])
    _legacy(legacy, [_legacy_row(error_cell=True)])

    dry = run_backfill(journal, legacy, apply_changes=False)
    assert dry["legacy_trades_parsed"] == 1
    assert dry["matches"] == 1
    assert dry["trade_numbers_to_write"] == 1
    assert dry["move_be_rows_to_write"] == 1
    assert dry["move_profit_rows_to_write"] == 1
    assert dry["ignored_invalid_cells"] == 1
    assert dry["applied"] is False

    applied = run_backfill(journal, legacy, apply_changes=True)
    assert applied["applied"] is True
    wb = load_workbook(journal)
    ws = wb["Trade Log"]
    headers = _trade_log_header_map(ws)
    row = TRADE_LOG_DATA_START_ROW
    assert ws.cell(row, headers[TRADE_NUMBER_HEADER]).value == "C1"
    assert ws.cell(row, headers["Move to Break Even Duration"]).value == 300
    assert ws.cell(row, headers["Move to Break Even Duration"]).number_format == r"00\:00\:00\:00"
    assert ws.cell(row, headers["Move to Break Even Distance From Entry %"]).value == 0.01
    assert ws.cell(row, headers["Move to Break Even Distance From Exit %"]).value == pytest.approx(19.0 / 120.0)
    assert ws.cell(row, headers["Move to Profit Distance From Entry %"]).value == 0.03
    wb.close()


def test_backfill_matches_by_utc_plus_ten(tmp_path: Path):
    journal = tmp_path / "Trading Journal.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    _journal(journal, [_repo_row()])
    _legacy(legacy, [_legacy_row(local=False, utc=True)])
    summary = run_backfill(journal, legacy, apply_changes=False)
    assert summary["matches"] == 1
    assert summary["unmatched"] == 0


def test_backfill_aborts_apply_on_ambiguous_and_duplicate_targets(tmp_path: Path):
    legacy = tmp_path / "legacy.xlsx"
    journal_ambiguous = tmp_path / "ambiguous.xlsx"
    _journal(journal_ambiguous, [_repo_row("r1"), _repo_row("r2")])
    _legacy(legacy, [_legacy_row()])
    ambiguous = run_backfill(journal_ambiguous, legacy, apply_changes=True)
    assert ambiguous["applied"] is False
    assert ambiguous["error"] == "unsafe_match_result"
    assert ambiguous["ambiguous"] == 1

    journal_duplicate = tmp_path / "duplicate.xlsx"
    _journal(journal_duplicate, [_repo_row("r1")])
    _legacy(legacy, [_legacy_row(1), _legacy_row(2)])
    duplicate = run_backfill(journal_duplicate, legacy, apply_changes=True)
    assert duplicate["applied"] is False
    assert duplicate["error"] == "unsafe_match_result"
    assert duplicate["duplicated_repo_targets"] == 1
