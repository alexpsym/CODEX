from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.ensure_local_journal_templates import ensure_local_journal_templates


EXPECTED_HEADERS = ["account", "date", "amount", "new_balance", "currency", "reason"]


def test_helper_creates_missing_cashflow_template(tmp_path: Path) -> None:
    journal_dir = tmp_path / "journal"

    rc = ensure_local_journal_templates(journal_dir)

    assert rc == 0
    target = journal_dir / "account_cashflows.xlsx"
    assert target.exists()

    wb = load_workbook(target)
    ws = wb["Cashflows"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, 7)]
    assert headers == EXPECTED_HEADERS


def test_helper_does_not_overwrite_existing_workbook(tmp_path: Path) -> None:
    journal_dir = tmp_path / "journal"
    journal_dir.mkdir(parents=True)
    target = journal_dir / "account_cashflows.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Sentinel"
    ws["A1"] = "keep-me"
    wb.save(target)

    rc = ensure_local_journal_templates(journal_dir)

    assert rc == 0
    persisted = load_workbook(target)
    assert persisted.sheetnames == ["Sentinel"]
    assert persisted["Sentinel"]["A1"].value == "keep-me"


def test_helper_fails_when_journal_dir_is_an_existing_file(tmp_path: Path) -> None:
    bad_path = tmp_path / "journal"
    bad_path.write_text("not a dir", encoding="utf-8")

    rc = ensure_local_journal_templates(bad_path)

    assert rc != 0
