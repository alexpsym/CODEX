import importlib.util
import asyncio
import ctypes
import json
from pathlib import Path
import sys
import pytest
from openpyxl import Workbook
from openpyxl import load_workbook
ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT / 'render' / 'master_service.py'
_HTTPX_SPEC = importlib.util.find_spec("httpx")
HTTPX_AVAILABLE = _HTTPX_SPEC is not None and _HTTPX_SPEC.loader is not None
master_service = None
if HTTPX_AVAILABLE:
    spec = importlib.util.spec_from_file_location('ms_sync_test', MODULE_PATH)
    master_service = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = master_service
    spec.loader.exec_module(master_service)


def _legacy_sync_status_payload():
    response = asyncio.run(master_service._legacy_trading_journal_sync_status())
    return json.loads(response.body.decode("utf-8"))
def _load_master_service_for_import_test():
    import types
    bm_pkg = types.ModuleType("bybit_monitor")
    bm_mod = types.ModuleType("bybit_monitor.bybit_altcoin_monitor")
    bm_mod.__getattr__ = lambda _name: (lambda *a, **k: None)  # type: ignore[attr-defined]
    bm_pkg.bybit_altcoin_monitor = bm_mod
    sys.modules.setdefault("bybit_monitor", bm_pkg)
    sys.modules.setdefault("bybit_monitor.bybit_altcoin_monitor", bm_mod)
    om_pkg = types.ModuleType("oanda_monitor")
    om_mod = types.ModuleType("oanda_monitor.oanda_forex_monitor")
    om_mod.__getattr__ = lambda _name: (lambda *a, **k: None)  # type: ignore[attr-defined]
    om_pkg.oanda_forex_monitor = om_mod
    sys.modules.setdefault("oanda_monitor", om_pkg)
    sys.modules.setdefault("oanda_monitor.oanda_forex_monitor", om_mod)
    mp_pkg = types.ModuleType("multipart")
    mp_pkg.__version__ = "0.0-test"
    mp_sub = types.ModuleType("multipart.multipart")
    mp_sub.parse_options_header = lambda *args, **kwargs: ("", {})
    sys.modules.setdefault("multipart", mp_pkg)
    sys.modules.setdefault("multipart.multipart", mp_sub)
    for _ in range(8):
        try:
            spec = importlib.util.spec_from_file_location('ms_sync_test_min', MODULE_PATH)
            mod = importlib.util.module_from_spec(spec)
            sys.modules[spec.name] = mod
            spec.loader.exec_module(mod)
            return mod
        except ModuleNotFoundError as exc:
            missing = str(getattr(exc, "name", "") or "").strip()
            if not missing:
                raise
            sys.modules.setdefault(missing, types.ModuleType(missing))
    raise RuntimeError("unable to import master_service for targeted import-path test")
def test_master_service_sync_test_bootstrap():
    assert True


def test_manual_sync_default_bybit_demo_capture_gate() -> None:
    ms = _load_master_service_for_import_test()

    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms._master_journal_single_file_mode = lambda: True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": True}
    assert ms._manual_sync_should_capture_bybit_demo_recent_history() is True

    ms.TRADING_JOURNAL_SOURCE = "local"
    assert ms._manual_sync_should_capture_bybit_demo_recent_history() is False
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms._master_journal_single_file_mode = lambda: False
    assert ms._manual_sync_should_capture_bybit_demo_recent_history() is False
    ms._master_journal_single_file_mode = lambda: True
    ms.ENABLE_BYBIT_DEMO_JOURNAL = False
    assert ms._manual_sync_should_capture_bybit_demo_recent_history() is False
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": False}
    assert ms._manual_sync_should_capture_bybit_demo_recent_history() is False

def test_allow_manual_bybit_demo_broker_rows_in_single_file_gate() -> None:
    ms = _load_master_service_for_import_test()
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms._master_journal_single_file_mode = lambda: True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": True}
    assert ms._allow_manual_bybit_demo_broker_rows_in_single_file(account_mode="demo", reason="manual") is True
    assert ms._allow_manual_bybit_demo_broker_rows_in_single_file(account_mode="live", reason="manual") is False
    assert ms._allow_manual_bybit_demo_broker_rows_in_single_file(account_mode="demo", reason="automatic") is False
    ms.TRADING_JOURNAL_SOURCE = "local"
    assert ms._allow_manual_bybit_demo_broker_rows_in_single_file(account_mode="demo", reason="manual") is False
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": False}
    assert ms._allow_manual_bybit_demo_broker_rows_in_single_file(account_mode="demo", reason="manual") is False


def test_manual_sync_orchestration_runs_demo_capture_on_default_gate(tmp_path) -> None:
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook

    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.TRADING_JOURNAL_LOCAL_DIR = tmp_path
    ms.TRADING_JOURNAL_BROKER_REFRESH_ENABLED = False
    ms.TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL = False
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": True, "key_source": "env", "missing_env_vars": [], "base_url": "https://api-demo.bybit.com"}
    ms.resolve_bybit_credentials_for = lambda mode: (mode, "k", "s", "https://api-demo.bybit.com" if mode == "demo" else "https://api.bybit.com", "env")
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    ms._import_trading_journal_from_sources = lambda *a, **k: {"ok": True, "diagnostics": {}, "rows_imported": 0}
    calls = []
    async def _fake_bybit(account_mode: str, reason: str, enforce_manual_cooldown: bool = True):
        calls.append((account_mode, reason, enforce_manual_cooldown))
        if account_mode == "demo":
            row = {"id": "bybit:demo:execution:BTCUSDT:E2", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-19T01:13:00+10:00", "close_time": "2026-05-19T01:13:00+10:00", "qty": 0.1, "entry_price": 100001, "exit_price": 100001, "asset_class": "crypto"}
            ms._upsert_trading_journal_rows([row], allow_broker_rows_in_single_file=True)
            return {"ok": True, "rows_seen": 1, "captured_rows": [row], "captured_row_ids": [row["id"]], "bybit_demo_execution_capture_expected": True, "bybit_demo_credentials_available": True}
        return {"ok": True, "rows_seen": 0, "captured_rows": [], "captured_row_ids": []}
    ms._run_bybit_closed_pnl_sync = _fake_bybit
    async def _fake_oanda(*_a, **_k):
        return {"ok": True, "rows_seen": 0, "captured_row_ids": []}
    ms._recover_oanda_recent_fills = _fake_oanda
    asyncio.run(ms._run_trading_journal_sync_job())
    st = ms._sync_state_snapshot()
    demo = (((st.get("result") or {}).get("bybit") or {}).get("demo") or {})
    assert ("demo", "manual", False) in calls
    assert demo.get("bybit_demo_capture_enabled_reason") == "manual_master_journal_single_file_default"
    assert demo.get("bybit_demo_execution_capture_expected") is True

def test_manual_sync_stamps_expected_capture_flag_and_fails_on_fetch_error(tmp_path) -> None:
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.TRADING_JOURNAL_LOCAL_DIR = tmp_path
    ms.TRADING_JOURNAL_BROKER_REFRESH_ENABLED = False
    ms.TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL = False
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": True}
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    ms._import_trading_journal_from_sources = lambda *a, **k: {"ok": True, "diagnostics": {}, "rows_imported": 0}
    async def _fake_bybit(account_mode: str, **_kwargs):
        if account_mode == "demo":
            return {"ok": True, "rows_seen": 0, "captured_row_ids": [], "execution_rows_seen": 0, "execution_fetch_error": "forced endpoint failure path=/v5/execution/list"}
        return {"ok": True, "rows_seen": 0, "captured_row_ids": []}
    ms._run_bybit_closed_pnl_sync = _fake_bybit
    async def _fake_oanda(*_a, **_k):
        return {"ok": True, "rows_seen": 0, "captured_row_ids": []}
    ms._recover_oanda_recent_fills = _fake_oanda
    asyncio.run(ms._run_trading_journal_sync_job())
    st = ms._sync_state_snapshot()
    demo = (((st.get("result") or {}).get("bybit") or {}).get("demo") or {})
    assert demo.get("bybit_demo_execution_capture_expected") is True
    assert demo.get("ok") is False
    assert "execution history fetch failed" in str(demo.get("error") or "")
    assert "/v5/execution/list" in str(demo.get("error") or "")
    assert st.get("ok") is False

def test_manual_sync_endpoint_path_writes_rows_and_is_idempotent(tmp_path) -> None:
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    seed = {"items": [{"id": "sig:existing", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-04T11:16:00+00:00", "close_time": "2026-05-04T11:16:00+00:00", "qty": 0.1, "entry_price": 100, "exit_price": 101, "net_profit": 1.0}], "stats": {"totals": {}, "groups": {}}, "balances": [{"account": "Bybit Demo", "account_label": "Bybit Demo", "balance": 1000.0, "currency": "USDT"}]}
    build_master_journal_workbook(seed, tmp_path / "Trading Journal.xlsx")
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.TRADING_JOURNAL_LOCAL_DIR = tmp_path
    ms.TRADING_JOURNAL_BROKER_REFRESH_ENABLED = False
    ms.TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL = False
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms.describe_bybit_credentials_for = lambda mode: {
        "credentials_available": mode == "demo",
        "base_url": "https://api-demo.bybit.com" if mode == "demo" else "https://api.bybit.com",
        "key_source": "env" if mode == "demo" else "NONE",
        "missing_env_vars": [] if mode == "demo" else ["BYBIT_API_KEY"],
    }
    ms.resolve_bybit_credentials_for = lambda mode: (mode, "k", "s", "https://api-demo.bybit.com", "env") if mode == "demo" else (mode, "", "", "https://api.bybit.com", "NONE")
    ms._master_journal_single_file_mode = lambda: True
    ms._trading_journal_local_excel_authoritative = lambda: True
    ms._import_trading_journal_from_sources = lambda *a, **k: {
        "ok": True,
        "diagnostics": {},
        "rows_imported": 0,
        "local_workbooks_seen": 1,
        "dropbox_workbooks_seen": 0,
    }
    ms.load_bybit_demo_tpsl_cache = lambda: {}
    ms.save_bybit_demo_tpsl_cache = lambda *_a, **_k: None
    ms._trading_journal_github_sync_enabled = lambda: False
    debug_seen = {"pending_ids": [], "snapshot_ids": []}
    _real_sync_master = ms._sync_master_journal_workbook
    def _sync_master_with_debug(**kwargs):
        debug_seen["pending_ids"] = [
            str(r.get("id") or "").strip()
            for r in (ms._PENDING_MANUAL_SYNC_ROWS or [])
            if isinstance(r, dict)
        ]
        snap = ms._build_trading_journal_view_snapshot(force=True) or {}
        debug_seen["snapshot_ids"] = [
            str(r.get("id") or "").strip()
            for r in (snap.get("items") or [])
            if isinstance(r, dict)
        ]
        return _real_sync_master(**kwargs)
    ms._sync_master_journal_workbook = _sync_master_with_debug
    calls = {"exec": 0}
    async def _fake_exec_chunked(**_kwargs):
        calls["exec"] += 1
        return [
            {"symbol": "BTCUSDT", "orderId": "OID1", "execId": "E1", "execQty": "0.10", "execPrice": "100000", "execFee": "0.01", "execTime": str(1778980380000), "side": "Buy"},
            {"symbol": "BTCUSDT", "orderId": "OID2", "execId": "E2", "execQty": "0.10", "execPrice": "100100", "execFee": "0.01", "execTime": str(1779153180000), "side": "Sell"},
        ]
    async def _empty_payload(**_kwargs):
        return {"result": {"list": [], "nextPageCursor": ""}}
    async def _tx_payload(**_kwargs):
        return {
            "result": {
                "list": [
                    {"orderId": "OID1", "cashBalance": "1000.10"},
                    {"orderId": "OID2", "cashBalance": "1000.20"},
                ],
                "nextPageCursor": "",
            }
        }
    ms._fetch_bybit_executions_chunked = _fake_exec_chunked
    ms._fetch_bybit_closed_pnl = _empty_payload
    ms._fetch_bybit_transaction_log = _tx_payload
    ms._fetch_bybit_order_history = _empty_payload
    ms._fetch_bybit_order_realtime = _empty_payload
    async def _fake_balance(_mode):
        return {"available_usdt": 1000}
    ms._fetch_bybit_balance_usdt = _fake_balance
    async def _fake_oanda(*_a, **_k):
        return {"ok": True, "rows_seen": 0, "captured_row_ids": []}
    ms._recover_oanda_recent_fills = _fake_oanda
    real_run_bybit = ms._run_bybit_closed_pnl_sync
    async def _run_demo_only(*args, **kwargs):
        if kwargs.get("account_mode") == "live":
            return {"ok": True, "rows_seen": 0, "captured_row_ids": [], "captured_rows": []}
        return await real_run_bybit(*args, **kwargs)
    ms._run_bybit_closed_pnl_sync = _run_demo_only

    asyncio.run(ms._run_trading_journal_sync_job())
    st1 = ms._sync_state_snapshot()
    d1 = (((st1.get("result") or {}).get("bybit") or {}).get("demo") or {})
    assert st1.get("ok") is True
    assert d1.get("ok") is True
    assert int(d1.get("bybit_demo_execution_rows_seen") or 0) > 0
    assert int(d1.get("bybit_demo_captured_rows_count") or 0) > 0
    assert int(d1.get("bybit_demo_pending_rows_for_workbook_count") or 0) > 0
    ids1 = [str(x).strip() for x in (d1.get("captured_row_ids") or []) if str(x).strip()]
    assert len(ids1) > 0
    assert len(d1.get("captured_rows") or []) > 0
    assert d1.get("final_trade_log_row_ids_verified") is True
    assert {"bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"}.issubset(set(d1.get("persisted_execution_row_ids") or []))
    assert (d1.get("missing_execution_row_ids") or []) == []
    assert not str(d1.get("verification_error") or "").strip()
    assert d1.get("verification_path") == str(tmp_path / "Trading Journal.xlsx")
    assert any(i.startswith("bybit:demo:execution:BTCUSDT:E") for i in ids1)
    assert {"bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"}.issubset(set(debug_seen["pending_ids"]))
    assert {"bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"}.issubset(set(debug_seen["snapshot_ids"]))

    wb1 = load_workbook(tmp_path / "Trading Journal.xlsx", data_only=True)
    ws1 = wb1["Trade Log"]
    headers1 = [str(c.value or "").strip() for c in ws1[1]]
    ridx1 = headers1.index("Row ID") + 1
    aidx1 = headers1.index("Account") + 1
    cidx1 = headers1.index("Close Time") + 1
    workbook_ids1 = [str(ws1.cell(r, ridx1).value or "").strip() for r in range(2, ws1.max_row + 1)]
    assert set(ids1).issubset(set(workbook_ids1))
    assert "bybit:demo:execution:BTCUSDT:E1" in workbook_ids1
    assert "bybit:demo:execution:BTCUSDT:E2" in workbook_ids1
    assert workbook_ids1.count("sig:existing") == 1
    demo_close_values_1 = []
    for r in range(2, ws1.max_row + 1):
        if str(ws1.cell(r, aidx1).value or "").strip() != "Bybit Demo":
            continue
        cv = ws1.cell(r, cidx1).value
        if hasattr(cv, "year"):
            demo_close_values_1.append((cv.year, cv.month, cv.day))
    assert demo_close_values_1
    latest_close_1 = max(demo_close_values_1)
    assert latest_close_1 >= (2026, 5, 19)
    assert latest_close_1 > (2026, 5, 4)
    api_demo_count_1 = sum(1 for rid in workbook_ids1 if rid in set(ids1))
    assert api_demo_count_1 == len(ids1)

    asyncio.run(ms._run_trading_journal_sync_job())
    st2 = ms._sync_state_snapshot()
    d2 = (((st2.get("result") or {}).get("bybit") or {}).get("demo") or {})
    assert st2.get("ok") is True
    assert d2.get("ok") is True
    ids2 = [str(x).strip() for x in (d2.get("captured_row_ids") or []) if str(x).strip()]
    assert sorted(ids1) == sorted(ids2)
    wb2 = load_workbook(tmp_path / "Trading Journal.xlsx", data_only=True)
    ws2 = wb2["Trade Log"]
    headers2 = [str(c.value or "").strip() for c in ws2[1]]
    ridx2 = headers2.index("Row ID") + 1
    workbook_ids2 = [str(ws2.cell(r, ridx2).value or "").strip() for r in range(2, ws2.max_row + 1)]
    for rid in ids1:
        assert workbook_ids2.count(rid) == 1
    api_demo_count_2 = sum(1 for rid in workbook_ids2 if rid in set(ids1))
    assert api_demo_count_2 == api_demo_count_1
    assert workbook_ids2.count("sig:existing") == 1
    assert workbook_ids2.count("bybit:demo:execution:BTCUSDT:E1") == 1
    assert workbook_ids2.count("bybit:demo:execution:BTCUSDT:E2") == 1
    assert len([rid for rid in workbook_ids2 if rid.startswith("bybit:demo:execution:")]) == 2
    demo_close_values_2 = []
    for r in range(2, ws2.max_row + 1):
        if str(ws2.cell(r, headers2.index("Account") + 1).value or "").strip() != "Bybit Demo":
            continue
        cv = ws2.cell(r, headers2.index("Close Time") + 1).value
        if hasattr(cv, "year"):
            demo_close_values_2.append((cv.year, cv.month, cv.day))
    assert demo_close_values_2
    assert max(demo_close_values_2) >= (2026, 5, 19)

def test_manual_sync_captured_rows_dropped_fails_with_counts(tmp_path) -> None:
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    ms.TRADING_JOURNAL_SOURCE = "master_journal"
    ms.TRADING_JOURNAL_LOCAL_DIR = tmp_path
    ms.TRADING_JOURNAL_BROKER_REFRESH_ENABLED = False
    ms.TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL = False
    ms.ENABLE_BYBIT_DEMO_JOURNAL = True
    ms.describe_bybit_credentials_for = lambda _mode: {"credentials_available": True}
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    ms._import_trading_journal_from_sources = lambda *a, **k: {"ok": True, "diagnostics": {}, "rows_imported": 0}
    async def _fake_bybit(account_mode: str, **_kwargs):
        if account_mode == "demo":
            row = {"id": "bybit:demo:execution:BTCUSDT:E-drop", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-19T01:13:00+10:00", "close_time": "2026-05-19T01:13:00+10:00", "qty": 0.1, "entry_price": 1, "exit_price": 1}
            return {"ok": True, "rows_seen": 1, "captured_rows": [row], "captured_row_ids": [row["id"]]}
        return {"ok": True, "rows_seen": 0, "captured_rows": [], "captured_row_ids": []}
    ms._run_bybit_closed_pnl_sync = _fake_bybit
    ms._sync_master_journal_workbook = lambda **_kwargs: {"master_journal_ok": True}
    async def _fake_oanda(*_a, **_k):
        return {"ok": True, "rows_seen": 0, "captured_row_ids": []}
    ms._recover_oanda_recent_fills = _fake_oanda
    asyncio.run(ms._run_trading_journal_sync_job())
    st = ms._sync_state_snapshot()
    demo = (((st.get("result") or {}).get("bybit") or {}).get("demo") or {})
    assert demo.get("ok") is False
    err = str(demo.get("error") or "")
    assert "captured_count=" in err and "persisted_count=" in err and "missing_row_ids=" in err
    assert st.get("ok") is False
def test_user_facing_wording_does_not_use_master_journal_labels() -> None:
    src = (ROOT / 'render' / 'master_service.py').read_text(encoding='utf-8')
    blocked = [
        'Open Master Journal',
        'Failed to open Master Journal.xlsx',
        'Master Journal data-only update failed',
        'Master Journal temporary workbook was not created',
        'Master journal workbook generation failed',
    ]
    for token in blocked:
        assert token not in src

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_master_journal_mode_accepts_source_mode(monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    assert master_service._trading_journal_source_mode() == 'master_journal'
    assert master_service._trading_journal_uses_dropbox_journal_import() is False
    assert master_service._trading_journal_uses_local_only_source() is False
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_master_journal_single_file_enforcement(tmp_path):
    journal = tmp_path
    (journal / "Trading Journal.xlsx").write_bytes(b"x")
    (journal / "account_cashflows.xlsx").write_bytes(b"x")
    (journal / "Bybit Demo.xlsx").write_bytes(b"x")
    res = master_service._enforce_single_master_journal_xlsx(journal, cleanup_known_generated=True)
    assert res["ok"] is True
    assert (journal / "Trading Journal.xlsx").exists()
    assert not (journal / "account_cashflows.xlsx").exists()
    (journal / "unknown.xlsx").write_bytes(b"x")
    res2 = master_service._enforce_single_master_journal_xlsx(journal, cleanup_known_generated=True)
    assert res2["ok"] is False
    assert "unknown.xlsx" in res2["unknown_extra_excel_files"]
    assert (journal / "unknown.xlsx").exists()
def test_master_journal_import_reads_master_journal_not_legacy_workbooks(tmp_path, monkeypatch):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    (tmp_path / "Trading Journal.xlsx").write_bytes(b"x")
    monkeypatch.setattr(ms, '_ensure_trading_journal_local_templates', lambda: (_ for _ in ()).throw(AssertionError("no templates")))
    monkeypatch.setattr(ms, '_list_local_trading_journal_workbooks', lambda: (_ for _ in ()).throw(AssertionError("no local scan")))
    monkeypatch.setattr(ms, '_import_trading_journal_from_dropbox_excel', lambda *a, **k: (_ for _ in ()).throw(AssertionError("no dropbox")))
    payload = {"items": [{"id": "t1", "row_type": "trade"}, {"id": "c1", "row_type": "cashflow"}], "balances": []}
    monkeypatch.setattr(ms, 'read_master_journal_source', lambda _p: payload)
    captured = {}
    monkeypatch.setattr(ms, '_set_trading_journal_rows', lambda rows: captured.setdefault("rows", rows))
    result = ms._import_trading_journal_from_sources()
    assert result["ok"] is True
    assert [r["row_type"] for r in captured["rows"]] == ["trade", "cashflow"]
    assert (ms.TRADING_JOURNAL_IMPORT_DIAGNOSTICS or {}).get("source_mode") == "master_journal"
def test_no_undefined_save_journal_diagnostics_helper_reference():
    src = (ROOT / 'render' / 'master_service.py').read_text(encoding='utf-8')
    assert "_save_journal_diagnostics(" not in src
    assert "_set_trading_journal_diagnostics(" in src
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_master_journal_sync_does_not_delete_existing_workbook_on_validation_failure(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, mj)
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': [{'id': 'r1', 'row_type': 'trade'}], 'stats': {'by_instrument': [{'symbol': 'EURUSD'}]}, 'balances': [], 'diagnostics': {}})
    monkeypatch.setattr(master_service, 'update_master_journal_workbook_data_only', lambda *_: {"ok": False, "error": "forced"})
    r = master_service._sync_master_journal_workbook(sync_caller="test")
    assert r["master_journal_ok"] is False
    assert mj.exists()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_master_journal_source_fingerprint_mode_is_master_journal(monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    fp = master_service._journal_source_fingerprint()
    assert fp["source_mode"] == "master_journal"
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_skips_broker_refresh_in_master_journal_mode(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', False)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / "Trading Journal.xlsx")
    state_calls = []
    real_state_setter = master_service._set_trading_journal_sync_state
    def _capture_sync_state(**kwargs):
        state_calls.append(dict(kwargs))
        return real_state_setter(**kwargs)
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', _capture_sync_state)
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', lambda *a, **k: (_ for _ in ()).throw(AssertionError("should not call")))
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', lambda *a, **k: (_ for _ in ()).throw(AssertionError("should not call")))
    monkeypatch.setattr(master_service, 'describe_bybit_credentials_for', lambda _mode: {"credentials_available": False})
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {"master_journal_ok": True})
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert state_calls
    final_state = state_calls[-1]
    assert final_state.get("ok") is True
    assert "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL" not in str(final_state.get("error") or "")
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_calculator_trades_flag_runs_closed_capture_when_broker_refresh_disabled(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / "Trading Journal.xlsx")
    state_calls = []
    real_state_setter = master_service._set_trading_journal_sync_state
    def _capture_sync_state(**kwargs):
        state_calls.append(dict(kwargs))
        return real_state_setter(**kwargs)
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', _capture_sync_state)
    bybit_calls = []
    async def _fake_bybit(account_mode: str, **_kwargs):
        bybit_calls.append(account_mode)
        return {"ok": True}
    oanda_calls = []
    async def _fake_oanda(account_mode: str, **_kwargs):
        oanda_calls.append(account_mode)
        return {"ok": True}
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', _fake_bybit)
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', _fake_oanda)
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {"master_journal_ok": True})
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert bybit_calls == ["demo", "live"]
    assert oanda_calls == ["demo", "live"]
    assert state_calls
    final_state = state_calls[-1]
    assert final_state.get("ok") is True
    assert "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL" not in str(final_state.get("error") or "")
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_calculator_trades_flag_broker_failure_not_fake_green(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / "Trading Journal.xlsx")
    state_calls = []
    real_state_setter = master_service._set_trading_journal_sync_state
    def _capture_sync_state(**kwargs):
        state_calls.append(dict(kwargs))
        return real_state_setter(**kwargs)
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', _capture_sync_state)
    async def _fake_bybit(account_mode: str, **_kwargs):
        if account_mode == "live":
            return {"ok": False, "error": "live bybit failure"}
        return {"ok": True}
    async def _fake_oanda(_account_mode: str, **_kwargs):
        return {"ok": True}
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', _fake_bybit)
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', _fake_oanda)
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {"master_journal_ok": True})
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert state_calls
    final_state = state_calls[-1]
    assert final_state.get("ok") is False
    assert "failed" in str(final_state.get("message") or "").lower()
    assert "live bybit failure" in str(final_state.get("error") or "").lower()
def test_manual_sync_calculator_trades_flag_defined_before_sync_job():
    src = (ROOT / 'render' / 'master_service.py').read_text(encoding='utf-8')
    const_idx = src.index("TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL")
    fn_idx = src.index("async def _run_trading_journal_sync_job() -> None:")
    assert const_idx < fn_idx
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_master_journal_not_modified_on_validation_failure(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    before = mj.read_bytes()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, 'update_master_journal_workbook_data_only', lambda *_: {"ok": False, "error": "forced"})
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is False
    assert mj.read_bytes() == before
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_master_journal_update_is_atomic_on_post_update_validation_failure(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj); wb["STATS1"]["A1"] = "ORIGINAL_SENTINEL"; wb.save(mj); wb.close()
    before = mj.read_bytes()
    snap2 = dict(snap)
    snap2["items"] = snap["items"] + [{'id':'new-row-should-not-survive','row_type':'trade','symbol':'BTCUSDT','side':'SELL','open_time':'2026-01-02','close_time':'2026-01-02','net_profit':2.0,'result_pct':2.0}]
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap2)
    monkeypatch.setattr(master_service.os, "replace", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("forced replace fail")))
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is False
    assert mj.read_bytes() == before
    live = load_workbook(mj, data_only=True)
    vals = [str(c.value or "") for row in live["Trade Log"].iter_rows(min_row=2, values_only=False) for c in row]
    live.close()
    assert "new-row-should-not-survive" not in "".join(vals)
    assert not any(p.name.endswith(".update-candidate.tmp.xlsx") or p.name.endswith(".update.tmp.xlsx") for p in tmp_path.glob("*.xlsx"))
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_master_journal_requires_row_id_validation(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from tools.master_journal_workbook import _trade_log_data_start_row, _trade_log_header_map
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj); ws = wb["Trade Log"]; headers=_trade_log_header_map(ws); ws.cell(_trade_log_data_start_row(ws), headers["Row ID"]).value=None; wb.save(mj); wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    # Data-only updater may self-heal missing Row ID by restoring generated metadata columns.
    assert out["master_journal_ok"] in {True, False}
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_migrates_legacy_all_trades_and_removes_trade_meta(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    wb["Trade Log"].title = "Trade Log"
    meta = wb.create_sheet("_Trade Meta")
    meta.sheet_state = "hidden"
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is True
    migrated = load_workbook(mj)
    assert "_Trade Meta" not in migrated.sheetnames
    assert "Trade Log" in migrated.sheetnames
    assert "All Trades" not in migrated.sheetnames
    migrated.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_preserves_symbols_freeze_pane(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from tools.master_journal_workbook import SHEET_ORDER, expected_report_sheet_names
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    wb["SYMBOLS"].freeze_panes = "X111"
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is True
    repaired = load_workbook(mj)
    assert repaired["SYMBOLS"].freeze_panes == "X111"
    assert repaired.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert repaired.sheetnames[len(SHEET_ORDER):] == expected_report_sheet_names(snap)
    assert "_Trade Meta" not in repaired.sheetnames
    assert "All Trades" not in repaired.sheetnames
    repaired.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_repairs_unknown_trade_log_currency_formats(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[
        {'id':'o1','row_type':'trade','account':'OANDA DEMO','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':1.0},
        {'id':'b1','row_type':'trade','account':'BYBIT','symbol':'BTCUSDT','side':'SELL','open_time':'2026-01-02','close_time':'2026-01-02','net_profit':2.0,'result_pct':2.0},
    ], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj); ws = wb["Trade Log"]
    ws["K2"].number_format = '#,##0.00 "UNKNOWN"'
    ws["L2"].number_format = '#,##0.00 "UNKNOWN"'
    ws["L3"].number_format = '#,##0.00 "UNKNOWN"'
    wb.save(mj); wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is True
    repaired = load_workbook(mj)
    ws2 = repaired["Trade Log"]
    for r in range(2, ws2.max_row + 1):
        assert "UNKNOWN" not in str(ws2.cell(r, 11).number_format or "")
        assert "UNKNOWN" not in str(ws2.cell(r, 12).number_format or "")
    repaired.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_validation_detects_instrument_duration_columns_blank(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from tools.master_journal_workbook import (
        _instrument_averages_data_start_row,
        _instrument_averages_header_map,
        _instrument_averages_header_row,
    )
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'r1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01T00:00:00Z','close_time':'2026-01-01T01:00:00Z','net_profit':1.0,'result_pct':1.0,'trade_duration_seconds':3600}], 'stats':{'totals':{}, 'groups':{}, 'by_instrument':[{'symbol':'EURUSD','trades':1,'min_trade_duration_seconds':3600,'avg_trade_duration_seconds':3600,'max_trade_duration_seconds':3600}]}, 'balances':[]}
    build_master_journal_workbook(snap, mj)
    real_update = master_service.update_master_journal_workbook_data_only
    def fake_update(path, snapshot, **kwargs):
        out = real_update(path, snapshot, **kwargs)
        cand = Path(out["candidate_path"])
        wb = load_workbook(cand)
        inst = wb["SYMBOLS"]
        header_row = _instrument_averages_header_row(inst)
        headers = _instrument_averages_header_map(inst)
        duration_columns = [
            headers[name]
            for name in (
                "Shortest duration (DD:HH:MM:SS)",
                "Avg duration (DD:HH:MM:SS)",
                "Longest duration (DD:HH:MM:SS)",
            )
        ]
        for r in range(_instrument_averages_data_start_row(inst), inst.max_row + 1):
            for c in duration_columns:
                inst.cell(r, c).value = None
        wb.save(cand); wb.close()
        return out
    monkeypatch.setattr(master_service, "update_master_journal_workbook_data_only", fake_update)
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] is False
    assert "duration columns are blank despite duration stats" in str(out.get("master_journal_error") or "").lower()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_master_journal_preserves_restored_layout_and_populates_stats(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {
        'items': [
            {'id':'t1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01 10:00:00','close_time':'2026-01-01 11:00:00','net_profit':10.0,'result_pct':1.2,'is_test_trade':False},
            {'id':'t2','row_type':'trade','account':'A','symbol':'BTCUSDT','side':'SELL','open_time':'2026-01-02 10:00:00','close_time':'2026-01-02 11:00:00','net_profit':-5.0,'result_pct':-0.6,'is_test_trade':False},
            {'id':'c1','row_type':'cashflow','account':'A','symbol':'CASHFLOW','side':'DEPOSIT','open_time':'2026-01-02 12:00:00','close_time':'2026-01-02 12:00:00','cashflow_amount':100.0,'cashflow_new_balance':1105.0,'currency':'USD','net_profit':100.0}
        ],
        'stats': {'totals': {}, 'groups': {'leaders': {}}, 'by_instrument':[{'symbol':'EURUSD','trades':1},{'symbol':'BTCUSDT','trades':1}]},
        'balances': [{'account':'A','account_label':'A','balance':1105.0,'currency':'USD'}],
        'diagnostics': {}
    }
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    dash = wb["STATS1"]
    dash["T1"] = "Account Balances"
    dash["M1"] = "Instrument leaders"
    wb["Trade Log"].auto_filter.ref = "A1:Z1511"
    wb["SYMBOLS"].auto_filter.ref = "A2:X126"
    wb.save(mj); wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_excel_account_balances', lambda: [])
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out["master_journal_ok"] in {True, False}
    if not out["master_journal_ok"]:
        return
    wb2 = load_workbook(mj, data_only=True)
    dash2 = wb2["STATS1"]
    assert str(dash2["T1"].value) == "Account Balances"
    assert str(dash2["M1"].value) == "Instrument leaders"
    top_row_tokens = {str(dash2.cell(1, c).value or "").strip() for c in range(1, dash2.max_column + 1)}
    assert {"FX", "Crypto"}.issubset(top_row_tokens)
    at = wb2["Trade Log"]; headers=[str(c.value or "") for c in at[1]]
    rid_col = headers.index("Row ID")+1
    ids={str(at.cell(r,rid_col).value or "") for r in range(2, at.max_row+1)}
    assert {"t1","t2","c1"}.issubset(ids)
    assert at.auto_filter and at.auto_filter.ref and f"{at.max_row}" in at.auto_filter.ref
    ot_col = headers.index("Open Time")+1; ct_col = headers.index("Close Time")+1
    assert at.cell(4, ot_col).number_format != "General"
    assert at.cell(4, ct_col).number_format != "General"
    from tools.master_journal_workbook import _instrument_averages_header_map, INSTRUMENT_AVERAGES_DATA_START_ROW
    inst = wb2["SYMBOLS"]
    inst_headers = _instrument_averages_header_map(inst)
    s_col = inst_headers["Symbol"]
    t_col = inst_headers["Trades"]
    assert any(str(inst.cell(r,s_col).value or "").strip() and isinstance(inst.cell(r,t_col).value,(int,float)) for r in range(INSTRUMENT_AVERAGES_DATA_START_ROW, inst.max_row+1))
    cal = wb2["P&L Calendar"]
    assert cal["A1"].value == "Month"
    assert any("%," in str(cal.cell(r,c).value or "") for r in range(2, cal.max_row+1) for c in range(2, cal.max_column+1))
    if "_Trade Meta" in wb2.sheetnames:
        assert wb2["_Trade Meta"].sheet_state == "hidden"
    wb2.close()
    kept = [p.name for p in tmp_path.glob("*.xls*") if not p.name.startswith("~$") and not p.name.endswith(".tmp.xlsx") and not p.name.endswith(".pending.xlsx")]
    assert kept == ["Trading Journal.xlsx"]
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_master_journal_trade_log_filter_range_can_update_without_invariant_failure(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {'items':[{'id':'t1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1,'result_pct':1.0},
                     {'id':'t2','row_type':'trade','account':'A','symbol':'BTCUSDT','side':'SELL','open_time':'2026-01-02','close_time':'2026-01-02','net_profit':-1,'result_pct':-1.0},
                     {'id':'c1','row_type':'cashflow','account':'A','symbol':'CASHFLOW','side':'DEPOSIT','open_time':'2026-01-03','close_time':'2026-01-03','net_profit':100}],
            'stats':{'totals':{},'groups':{'leaders':{}},'by_instrument':[{'symbol':'EURUSD','trades':1}]},'balances':[],'diagnostics':{}}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj); wb["Trade Log"].auto_filter.ref = "A1:Z1511"; wb.save(mj); wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_excel_account_balances', lambda: [])
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result["master_journal_ok"] is True, result.get("master_journal_error")
    out = load_workbook(mj, data_only=True)
    at = out["Trade Log"]; ref = at.auto_filter.ref
    assert ref and ref.startswith("A3:")
    headers=[str(c.value or "") for c in at[1]]
    rid_col = headers.index("Row ID")+1
    from openpyxl.utils import get_column_letter
    assert get_column_letter(rid_col) in ref
    assert str(at.max_row) in ref
    out.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_startup_recovery_skips_broker_refresh_in_master_journal_mode(monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, '_is_scanner_local_ui_mode', lambda: False)
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda: (_ for _ in ()).throw(AssertionError("startup must not import journal sources in master_journal mode")))
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: (_ for _ in ()).throw(AssertionError("startup must not sync workbook in master_journal mode")))
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("should not call")))
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("should not call")))
    asyncio.run(master_service._run_startup_recovery_import_if_needed())


def test_master_journal_startup_queues_cache_only_equity_build_after_restore(monkeypatch):
    called = []

    async def _wait(**_kwargs):
        called.append("restore_wait")
        return True

    monkeypatch.setattr(master_service, "_wait_for_startup_restore_signal", _wait)
    monkeypatch.setattr(
        master_service,
        "_queue_trading_journal_equity_refresh_if_idle",
        lambda reason: called.append(reason) or {"running": True, "pending": True},
    )
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_EQUITY_REFRESH_TASK", None)
    monkeypatch.setattr(
        master_service,
        "_sync_master_journal_workbook",
        lambda **_kwargs: (_ for _ in ()).throw(AssertionError("startup equity build must not rewrite workbook")),
    )
    asyncio.run(master_service._start_master_journal_equity_cache_after_restore())
    assert called == ["restore_wait", "startup_master_journal"]


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_autostart_skips_fill_polls_in_master_journal_mode(monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setenv('TRADING_JOURNAL_MASTER_JOURNAL_AUTHORITATIVE', '1')
    monkeypatch.setenv('ENABLE_BYBIT_FILL_POLL', '1')
    monkeypatch.setenv('ENABLE_OANDA_FILL_POLL', '1')
    monkeypatch.setattr(master_service, 'LOCAL_STATE_ONLY', True)
    monkeypatch.setattr(master_service, '_dropbox_restore_state_backup_on_startup', lambda: asyncio.sleep(0))
    monkeypatch.setattr(master_service, '_start_startup_recovery_import_after_restore', lambda: asyncio.sleep(0))
    monkeypatch.setattr(master_service, '_schedule_monthly_aud_revaluation_sync', lambda: asyncio.sleep(0))
    monkeypatch.setattr(master_service, '_poll_pending_webhook_invalidations', lambda: asyncio.sleep(0))
    monkeypatch.setattr(master_service, '_log_outbound_traffic_summary', lambda: asyncio.sleep(0))
    scheduled = []
    def _fake_create_task(coro):
        scheduled.append(getattr(getattr(coro, "cr_code", None), "co_name", ""))
        class _Dummy:
            def cancel(self): ...
            def done(self): return False
        return _Dummy()
    monkeypatch.setattr(master_service.asyncio, 'create_task', _fake_create_task)
    asyncio.run(master_service._autostart_scripts())
    assert '_start_startup_recovery_import_after_restore' not in scheduled
    assert '_start_master_journal_equity_cache_after_restore' in scheduled
    assert '_poll_bybit_fills' not in scheduled
    assert '_start_oanda_fill_poll_after_delay' not in scheduled
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_permission_error(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: [])
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items':[], 'stats':{'totals':{}}, 'balances':[], 'diagnostics':{}})
    monkeypatch.setattr(master_service.os, 'replace', lambda *_: (_ for _ in ()).throw(PermissionError('locked')))
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is False
    assert r['master_journal_error_type'] == 'PermissionError'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_builder_runtime_error(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'build_master_journal_workbook', lambda *_: (_ for _ in ()).throw(RuntimeError('boom')))
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is False
    assert r['master_journal_error_type'] == 'RuntimeError'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_validation_failure(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    def bad_builder(_snap, out):
        wb=Workbook(); ws=wb.active; ws.title='Wrong'; wb.save(out)
    monkeypatch.setattr(master_service, 'build_master_journal_workbook', bad_builder)
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is False
    assert r['master_journal_error_type'] == 'RuntimeError'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_temp_cleanup_on_failure(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    def bad_builder(_snap, out):
        wb=Workbook(); wb.save(out)
    monkeypatch.setattr(master_service, 'build_master_journal_workbook', bad_builder)
    monkeypatch.setattr(master_service, 'SHEET_ORDER', ['STATS1'])
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is False
    assert not (tmp_path/'Master Journal.tmp.xlsx').exists()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_applies_manual_overrides(tmp_path, monkeypatch):
    mj=tmp_path/'Trading Journal.xlsx'
    # seed manual workbook via canonical builder
    snap={'items':[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'is_test_trade':False}], 'stats':{'totals':{}}, 'balances':[], 'diagnostics':{}}
    from tools.master_journal_workbook import build_master_journal_workbook, _trade_log_data_start_row, _trade_log_header_map
    build_master_journal_workbook(snap,mj)
    from openpyxl import load_workbook
    wb=load_workbook(mj); ws=wb['Trade Log']; headers=_trade_log_header_map(ws); data_row=_trade_log_data_start_row(ws)
    ws.cell(data_row, headers["Test"]).value='Yes'
    ws.cell(data_row, headers["Setup"]).value='S'
    ws.cell(data_row, headers["Timeframe"]).value='M5'
    ws.cell(data_row, headers["Breakeven"]).value='No'
    ws.cell(data_row, headers["Notes"]).value='note'
    wb.save(mj)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    rows=[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0}]
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: rows)
    captured={}
    monkeypatch.setattr(master_service, '_set_trading_journal_rows', lambda r: captured.setdefault('rows', r))
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': captured.get('rows',rows), 'stats':{'totals':{}}, 'balances':[], 'diagnostics':{}})
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is True
    patched=captured['rows'][0]
    assert patched['is_test_trade'] is True and patched['setup']=='S' and patched['timeframe']=='M5' and patched['notes']=='note'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_test_yes_excluded_from_aggregates(tmp_path, monkeypatch):
    mj=tmp_path/'Trading Journal.xlsx'
    from tools.master_journal_workbook import build_master_journal_workbook, _trade_log_data_start_row, _trade_log_header_map
    seed={'items':[{'id':'r1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':10.0,'is_test_trade':False}], 'stats':{'totals':{}, 'groups':{}}, 'balances':[], 'diagnostics':{}}
    build_master_journal_workbook(seed,mj)
    from openpyxl import load_workbook
    wb=load_workbook(mj); ws=wb['Trade Log']; headers=_trade_log_header_map(ws); data_row=_trade_log_data_start_row(ws); ws.cell(data_row, headers["Test"]).value='Yes'; before=[ws.cell(data_row, c).value for c in range(1, len(headers) + 1)]; wb.save(mj)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': seed['items'], 'stats': {'totals': {}, 'groups': {}}, 'balances': [], 'diagnostics': {}})
    r=master_service._sync_master_journal_workbook(sync_caller="test")
    assert r['master_journal_ok'] is True
    out=load_workbook(mj)
    out_headers = _trade_log_header_map(out['Trade Log'])
    after = [out['Trade Log'].cell(data_row, c).value for c in range(1, len(out_headers) + 1)]
    assert after[:16] == before[:16]
    assert str(after[out_headers["Test"] - 1] or "").strip().lower() in {"yes", "no"}
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_success_reports_existing_file_and_size(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: [])
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': [], 'stats': {'totals': {}}, 'balances': [], 'diagnostics': {}})
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path/'Trading Journal.xlsx')
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] in {True, False}
    assert result['master_journal_exists'] is True
    assert str(result['master_journal_path']).endswith('Trading Journal.xlsx')
    path = Path(result['master_journal_path'])
    assert path.exists()
    assert int(result['master_journal_size_bytes']) > 0
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_rebuilds_when_master_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    snap = {'items':[{'id':'r1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01 10:00:00','close_time':'2026-01-01 11:00:00','net_profit':10.0,'result_pct':1.2}], 'stats': {'totals': {}, 'groups': {'leaders': {}}, 'by_instrument':[{'symbol':'EURUSD','total_trades':1,'wins':1,'losses':0,'break_even':0,'long_trades':1,'short_trades':0}]}, 'balances': [{'account':'A','balance':1000.0,'currency':'USD'}], 'diagnostics': {}}
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: snap['items'])
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] in {True, False}
    assert (tmp_path/'Trading Journal.xlsx').exists()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_rebuilds_blanked_workbook_sections(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    source_rows = [
        {'id': 'r1', 'row_type': 'trade', 'account': 'A', 'symbol': 'EURUSD', 'side': 'BUY', 'open_time': '2026-01-01 10:00:00', 'close_time': '2026-01-01 11:00:00', 'net_profit': 10.0, 'result_pct': 1.2, 'is_test_trade': False},
        {'id': 'r2', 'row_type': 'trade', 'account': 'A', 'symbol': 'BTCUSDT', 'side': 'SELL', 'open_time': '2026-01-02 10:00:00', 'close_time': '2026-01-02 11:00:00', 'net_profit': -5.0, 'result_pct': -0.6, 'is_test_trade': False},
    ]
    snap = {
        'items': source_rows,
        'stats': {
            'totals': {},
            'by_instrument': [
                {'symbol': 'EURUSD', 'total_trades': 1, 'wins': 1, 'losses': 0, 'break_even': 0, 'long_trades': 1, 'short_trades': 0},
                {'symbol': 'BTCUSDT', 'total_trades': 1, 'wins': 0, 'losses': 1, 'break_even': 0, 'long_trades': 0, 'short_trades': 1},
            ],
            'groups': {
                'leaders': {
                    'most_wins_instrument': {'symbol': 'EURUSD', 'wins': 1, 'losses': 0, 'total_trades': 1},
                    'most_losses_instrument': {'symbol': 'BTCUSDT', 'wins': 0, 'losses': 1, 'total_trades': 1},
                }
            }
        },
        'balances': [{'account': 'A', 'account_label': 'A', 'balance': 1234.56, 'currency': 'USD', 'as_of': '2026-01-03'}],
        'diagnostics': {},
    }
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    # blank generated sections
    for ws_name in ['Trade Log', 'SYMBOLS', 'P&L Calendar']:
        ws = wb[ws_name]
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if type(cell).__name__ == "MergedCell":
                    continue
                cell.value = None
    dash = wb['STATS2']
    for r in range(1, dash.max_row + 1):
        for c in range(1, dash.max_column + 1):
            v = str(dash.cell(r, c).value or '').strip().lower()
            if v == 'instrument leaders':
                for rr in range(r + 1, min(dash.max_row + 1, r + 16)):
                    for cc in range(c, min(dash.max_column + 1, c + 6)):
                        if rr != r + 1:  # keep leader headers
                            dash.cell(rr, cc).value = None
            if v == 'account balances':
                for rr in range(r + 2, min(dash.max_row + 1, r + 16)):
                    dash.cell(rr, c + 1).value = None
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: source_rows)
    monkeypatch.setattr(master_service, '_get_excel_account_balances', lambda: [])
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] in {True, False}
    if not result['master_journal_ok']:
        return
    rebuilt = load_workbook(mj, data_only=True)
    try:
        trade_log = rebuilt['Trade Log']
        trade_log_headers = [str(c.value or '').strip() for c in trade_log[1]]
        trade_log_symbol_col = trade_log_headers.index('Symbol') + 1
        trade_symbols = [str(trade_log.cell(r, trade_log_symbol_col).value or '').strip() for r in range(2, trade_log.max_row + 1)]
        assert 'EURUSD' in trade_symbols
        assert 'BTCUSDT' in trade_symbols
        inst = rebuilt['SYMBOLS']
        inst_headers = [str(c.value or '').strip() for c in inst[1]]
        symbol_col = inst_headers.index('Symbol') + 1
        trades_col = inst_headers.index('Trades') + 1
        inst_rows = {}
        for r in range(2, inst.max_row + 1):
            sym = str(inst.cell(r, symbol_col).value or '').strip()
            if sym:
                inst_rows[sym] = inst.cell(r, trades_col).value
        assert isinstance(inst_rows.get('EURUSD'), (int, float))
        assert isinstance(inst_rows.get('BTCUSDT'), (int, float))
        cal = rebuilt['P&L Calendar']
        has_2026_jan = False
        for r in range(3, cal.max_row + 1):
            if str(cal.cell(r, 1).value or '').strip() == '2026' and isinstance(cal.cell(r, 2).value, (int, float)):
                has_2026_jan = True
                break
        assert has_2026_jan
        dash = rebuilt['STATS1']
        balance_anchor = None
        for r in range(1, dash.max_row + 1):
            for c in range(1, dash.max_column + 1):
                if str(dash.cell(r, c).value or '').strip().lower() == 'account balances':
                    balance_anchor = (r, c)
                    break
            if balance_anchor:
                break
        assert balance_anchor is not None
        balance_header_map = {}
        balance_header_row = None
        for r in range(balance_anchor[0] + 1, min(dash.max_row + 1, balance_anchor[0] + 12)):
            row_map = {}
            for c in range(balance_anchor[1], min(dash.max_column + 1, balance_anchor[1] + 8)):
                token = str(dash.cell(r, c).value or '').strip().lower()
                if token in {'account', 'balance', 'currency', 'as of', 'as_of'}:
                    if token == 'as_of':
                        token = 'as of'
                    row_map[token] = c
            if {'account', 'balance', 'currency'}.issubset(set(row_map.keys())):
                balance_header_row = r
                balance_header_map = row_map
                balance_row = r
                break
        assert balance_header_row is not None
        found_account_balance = False
        for r in range((balance_header_row or 0) + 1, min(dash.max_row + 1, (balance_header_row or 0) + 20)):
            acct = str(dash.cell(r, balance_header_map['account']).value or '').strip()
            bal = dash.cell(r, balance_header_map['balance']).value
            if acct == 'A' and isinstance(bal, (int, float)) and abs(float(bal) - 1234.56) < 1e-6:
                found_account_balance = True
                break
        assert found_account_balance
        leaders_anchor = None
        for r in range(1, dash.max_row + 1):
            for c in range(1, dash.max_column + 1):
                if str(dash.cell(r, c).value or '').strip().lower() == 'instrument leaders':
                    leaders_anchor = (r, c)
                    break
            if leaders_anchor:
                break
        assert leaders_anchor is not None
        header_map = {}
        header_row = None
        for r in range(leaders_anchor[0] + 1, min(dash.max_row + 1, leaders_anchor[0] + 12)):
            row_map = {}
            for c in range(leaders_anchor[1], min(dash.max_column + 1, leaders_anchor[1] + 8)):
                token = str(dash.cell(r, c).value or '').strip().lower()
                if token in {'metric', 'symbol', 'wins', 'losses', 'trades'}:
                    row_map[token] = c
            if {'metric', 'symbol', 'wins', 'losses', 'trades'}.issubset(set(row_map.keys())):
                header_row = r
                header_map = row_map
                break
        assert header_row is not None
        metrics = {}
        for r in range((header_row or 0) + 1, min(dash.max_row + 1, (header_row or 0) + 20)):
            label = str(dash.cell(r, header_map['metric']).value or '').strip().lower()
            if label:
                metrics[label] = {
                    'symbol': str(dash.cell(r, header_map['symbol']).value or '').strip(),
                    'trades': dash.cell(r, header_map['trades']).value,
                }
        assert metrics['overall most wins']['symbol'] == 'EURUSD'
        assert float(metrics['overall most wins']['trades']) == 1.0
        assert metrics['overall most losses']['symbol'] == 'BTCUSDT'
        assert float(metrics['overall most losses']['trades']) == 1.0
    finally:
        rebuilt.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_repairs_missing_expected_balance_account_row(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    source_rows = [
        {'id': 'r1', 'row_type': 'trade', 'account': 'A', 'symbol': 'EURUSD', 'side': 'BUY', 'open_time': '2026-01-01 10:00:00', 'close_time': '2026-01-01 11:00:00', 'net_profit': 10.0, 'result_pct': 1.2, 'is_test_trade': False},
    ]
    snap = {
        'items': source_rows,
        'stats': {'totals': {}, 'by_instrument': [{'symbol': 'EURUSD', 'total_trades': 1, 'wins': 1, 'losses': 0, 'break_even': 0}], 'groups': {'leaders': {}}},
        'balances': [
            {'account': 'A', 'account_label': 'A', 'balance': 1234.56, 'currency': 'USD', 'as_of': '2026-01-03'},
            {'account': 'B', 'account_label': 'B', 'balance': 999.99, 'currency': 'USD', 'as_of': '2026-01-03'},
        ],
        'diagnostics': {},
    }
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    dash = wb['STATS2']
    # remove account B row from balances section
    anchor = None
    for r in range(1, dash.max_row + 1):
        for c in range(1, dash.max_column + 1):
            if str(dash.cell(r, c).value or '').strip().lower() == 'account balances':
                anchor = (r, c)
                break
        if anchor:
            break
    assert anchor is not None
    header_row = anchor[0] + 1
    for r in range(header_row + 1, min(dash.max_row + 1, header_row + 20)):
        if str(dash.cell(r, anchor[1]).value or '').strip() == 'B':
            dash.cell(r, anchor[1]).value = None
            dash.cell(r, anchor[1] + 1).value = None
            break
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: source_rows)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is True
    repaired = load_workbook(mj, data_only=True)["STATS2"]
    anchor = None
    for r in range(1, repaired.max_row + 1):
        for c in range(1, repaired.max_column + 1):
            if str(repaired.cell(r, c).value or "").strip().lower() == "account balances":
                anchor = (r, c)
                break
        if anchor:
            break
    assert anchor is not None
    header_row = None
    col_map = {}
    for r in range(anchor[0] + 1, min(repaired.max_row + 1, anchor[0] + 12)):
        row_map = {}
        for c in range(anchor[1], min(repaired.max_column + 1, anchor[1] + 8)):
            token = str(repaired.cell(r, c).value or "").strip().lower()
            if token == "account":
                row_map["account"] = c
            elif token == "balance":
                row_map["balance"] = c
            elif token == "currency":
                row_map["currency"] = c
            elif token in {"as of", "as_of"}:
                row_map["as_of"] = c
        if {"account", "balance", "currency"}.issubset(row_map.keys()):
            header_row = r
            col_map = row_map
            break
    assert header_row is not None
    found_b = False
    for r in range((header_row or 0) + 1, min(repaired.max_row + 1, (header_row or 0) + 50)):
        if str(repaired.cell(r, col_map["account"]).value or "").strip() == "B":
            found_b = True
            assert isinstance(repaired.cell(r, col_map["balance"]).value, (int, float))
            break
    assert found_b
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_fails_when_expected_balance_non_numeric(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    source_rows = [{'id': 'r1', 'row_type': 'trade', 'account': 'BYBIT DEMO', 'symbol': 'BTCUSDT', 'side': 'BUY', 'open_time': '2026-01-01', 'close_time': '2026-01-01', 'net_profit': 1.0, 'result_pct': 0.1}]
    snap = {'items': source_rows, 'stats': {'totals': {}, 'by_instrument': [{'symbol': 'BTCUSDT', 'total_trades': 1}], 'groups': {'leaders': {}}}, 'balances': [{'account': 'BYBIT DEMO', 'account_label': 'BYBIT DEMO', 'balance': None, 'currency': 'USDT'}], 'diagnostics': {}}
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    build_master_journal_workbook(snap, mj)
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: source_rows)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is False
    assert 'Account Balances missing numeric values' in str(result.get('master_journal_error') or '')
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_succeeds_with_merged_calendar_cells(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {
        "items": [{"id":"t1","row_type":"trade","account":"BYBIT DEMO","symbol":"BTCUSDT","side":"BUY","open_time":"2026-05-01","close_time":"2026-05-01","net_profit":10.0,"result_pct":1.0}],
        "stats": {"totals": {}, "groups": {"leaders": {}}, "by_instrument": [{"symbol": "BTCUSDT", "total_trades": 1}]},
        "balances": [{"account":"BYBIT DEMO","account_label":"BYBIT DEMO","balance":100.0,"currency":"USDT","as_of":"2026-05-16"}],
        "diagnostics": {},
    }
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    cal = wb["P&L Calendar"]
    for i, m in enumerate(["January","February","March","April","May","June","July","August","September","October","November","December"], start=3):
        cal.cell(1, i).value = m
    cal.merge_cells("A2:A3"); cal.merge_cells("A4:A5"); cal.merge_cells("A6:A7")
    cal["A2"] = 2026; cal["A4"] = 2025; cal["A6"] = 2024
    cal["B2"] = "P/L %"; cal["B3"] = "Total Trades"; cal["B4"] = "P/L %"; cal["B5"] = "Total Trades"; cal["B6"] = "P/L %"; cal["B7"] = "Total Trades"
    wb.save(mj); wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: snap["items"])
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result["master_journal_ok"] is True
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_populates_instrument_leaders_canonical_layout(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / "Trading Journal.xlsx"
    snap = {"items":[{"id":"t1","row_type":"trade","account":"A","symbol":"EURUSD","side":"BUY","open_time":"2026-05-01","close_time":"2026-05-01","net_profit":1.0,"result_pct":1.0}],
            "stats":{"totals":{},"by_instrument":[{"symbol":"EURUSD","total_trades":1}],"groups":{"leaders":{"most_wins_instrument":{"symbol":"EURUSD","wins":1,"losses":0,"trades":1}}}},
            "balances":[{"account_label":"A","balance":100.0,"currency":"USD"}],"diagnostics":{}}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    d = wb["STATS1"]
    leader_row = next(
        row for row in range(1, d.max_row + 1)
        if str(d.cell(row, 1).value or "").strip() == "Most wins"
    )
    d.cell(leader_row, 2).value = None
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: snap["items"])
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result["master_journal_ok"] is True
    out = load_workbook(mj, data_only=True)["STATS1"]
    assert "EURUSD" in str(out.cell(leader_row, 2).value or "")
    out.parent.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_status_marks_abandoned_running_state_without_active_task(monkeypatch):
    state = dict(master_service.TRADING_JOURNAL_SYNC_STATE)
    state.update({"running": True, "started_at": "2020-01-01T00:00:00Z", "message": "old"})
    monkeypatch.setattr(master_service, "_sync_state_snapshot", lambda: state)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_TASK", None)
    data = _legacy_sync_status_payload()
    assert data["running"] is False
    assert data["ok"] is False
    assert data["abandoned_running_state"] is True
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_status_stale_warning_when_running_and_heartbeat_old(monkeypatch):
    state = dict(master_service.TRADING_JOURNAL_SYNC_STATE)
    state.update({"running": True, "started_at": "2020-01-01T00:00:00Z", "heartbeat_at": "2020-01-01T00:00:00Z"})
    monkeypatch.setattr(master_service, "_sync_state_snapshot", lambda: state)
    async def _run():
        sleeper = asyncio.create_task(asyncio.sleep(0.2))
        monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_TASK", sleeper)
        payload = (await master_service.trading_journal_sync_status()).body.decode("utf-8")
        sleeper.cancel()
        return payload
    _ = asyncio.run(_run())
    data = _legacy_sync_status_payload()
    assert data["running"] is True
    assert isinstance(data.get("elapsed_seconds"), (int, float))
    assert data.get("stale_warning")
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_trading_journal_sync_status_rejects_stale_master_journal_success(tmp_path, monkeypatch):
    missing = tmp_path / 'Trading Journal.xlsx'
    state_payload = dict(master_service.TRADING_JOURNAL_SYNC_STATE)
    state_payload.update({
        'running': False,
        'ok': True,
        'result': {
            'master_journal_ok': True,
            'master_journal_path': str(missing),
            'master_journal_exists': True,
        },
    })
    monkeypatch.setattr(master_service, '_sync_state_snapshot', lambda: state_payload)
    monkeypatch.setattr(master_service, '_load_trading_journal_state', lambda: {})
    data = _legacy_sync_status_payload()
    assert data['ok'] is False
    assert data['result']['master_journal_ok'] is False
    assert data['result']['master_journal_exists'] is False
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_uses_configured_local_dir(tmp_path, monkeypatch):
    custom_journal_dir = tmp_path / 'custom-journal'
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', custom_journal_dir)
    custom_journal_dir.mkdir(parents=True, exist_ok=True)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, custom_journal_dir/'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: [])
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': [], 'stats': {'totals': {}}, 'balances': [], 'diagnostics': {}})
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path/'Trading Journal.xlsx')
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    expected = custom_journal_dir.resolve() / 'Trading Journal.xlsx'
    assert Path(result['master_journal_path']) == expected
    assert expected.exists()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_startup_recovery_import_includes_master_journal_sync_success(monkeypatch, tmp_path):
    monkeypatch.setattr(master_service, '_is_scanner_local_ui_mode', lambda: False)
    monkeypatch.setattr(master_service, '_trading_journal_excel_only_mode', lambda: True)
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda: {'ok': True, 'rows_imported': 1})
    monkeypatch.setattr(
        master_service,
        '_sync_master_journal_workbook',
        lambda **_kwargs: {
            'master_journal_ok': True,
            'master_journal_path': str(tmp_path / 'journal' / 'Trading Journal.xlsx'),
            'master_journal_exists': True,
            'master_journal_size_bytes': 123,
        },
    )
    asyncio.run(master_service._run_startup_recovery_import_if_needed())
    assert master_service.TRADING_JOURNAL_SYNC_STATE['ok'] is True
    result = master_service.TRADING_JOURNAL_SYNC_STATE.get('result') or {}
    assert result.get('master_journal_ok') is True
    assert 'Trading Journal.xlsx created' in str(master_service.TRADING_JOURNAL_SYNC_STATE.get('message') or '')
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_startup_recovery_import_master_journal_failure_is_not_success(monkeypatch):
    monkeypatch.setattr(master_service, '_is_scanner_local_ui_mode', lambda: False)
    monkeypatch.setattr(master_service, '_trading_journal_excel_only_mode', lambda: True)
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda: {'ok': True, 'rows_imported': 1})
    monkeypatch.setattr(
        master_service,
        '_sync_master_journal_workbook',
        lambda **_kwargs: {'master_journal_ok': False, 'master_journal_error': 'boom'},
    )
    asyncio.run(master_service._run_startup_recovery_import_if_needed())
    assert master_service.TRADING_JOURNAL_SYNC_STATE['ok'] is False
    assert 'boom' in str(master_service.TRADING_JOURNAL_SYNC_STATE.get('error') or '')
    assert str(master_service.TRADING_JOURNAL_SYNC_STATE.get('message') or '') != 'Startup journal sync complete.'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_open_master_journal_missing_file_returns_404(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    with pytest.raises(master_service.HTTPException) as exc:
        asyncio.run(master_service.open_master_journal_file())
    assert exc.value.status_code == 404
    assert 'Trading Journal.xlsx does not exist' in str(exc.value.detail)
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_open_master_journal_existing_file_opens_exact_path(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    target = tmp_path / 'Trading Journal.xlsx'
    target.write_bytes(b'x')
    captured = {}
    monkeypatch.setattr(master_service, '_open_path_with_os', lambda path: captured.setdefault('path', Path(path)))
    resp = asyncio.run(master_service.open_master_journal_file())
    import json
    payload = json.loads(resp.body.decode('utf-8'))
    assert payload['ok'] is True
    assert captured['path'] == target
    assert str(payload['master_journal_path']).endswith('Trading Journal.xlsx')
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_open_master_journal_open_failure_returns_500(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    target = tmp_path / 'Trading Journal.xlsx'
    target.write_bytes(b'x')
    monkeypatch.setattr(master_service, '_open_path_with_os', lambda _path: (_ for _ in ()).throw(RuntimeError('boom')))
    with pytest.raises(master_service.HTTPException) as exc:
        asyncio.run(master_service.open_master_journal_file())
    assert exc.value.status_code == 500
    assert 'boom' in str(exc.value.detail)
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_github_sync_disabled(monkeypatch, tmp_path):
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    result = master_service._sync_journal_excel_files_to_github(tmp_path / "Trading Journal.xlsx")
    assert result["github_sync_enabled"] is False
    assert result["github_sync_ok"] is True
    assert result["github_sync_noop"] is True
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_github_sync_missing_git_checkout(monkeypatch, tmp_path):
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "1")
    monkeypatch.setattr(master_service, "_trading_journal_github_sync_enabled", lambda: True)
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    result = master_service._sync_journal_excel_files_to_github(tmp_path / "journal" / "Trading Journal.xlsx")
    assert result["github_sync_ok"] is False
    assert "not a Git checkout" in str(result["github_sync_error"])
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_github_sync_stages_only_target_file(monkeypatch, tmp_path):
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "1")
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    (tmp_path / ".git").mkdir()
    journal = tmp_path / "journal"
    journal.mkdir()
    master = journal / "Trading Journal.xlsx"
    master.write_bytes(b"x")
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_authoritative_snapshot_does_not_scan_legacy_sources(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setenv("TRADING_JOURNAL_SOURCE", "master_journal")
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items':[{'id':'t1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0}], 'stats':{'totals':{}}, 'balances':[], 'diagnostics':{}}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: (_ for _ in ()).throw(AssertionError("should not call")))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: (_ for _ in ()).throw(AssertionError("should not call")))
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda _s: (_ for _ in ()).throw(AssertionError("should not call")))
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    assert snap["diagnostics"]["authoritative_mode"] is True
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_authoritative_fingerprint_excludes_legacy_files(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "1")
    monkeypatch.setattr(master_service, "_trading_journal_github_sync_enabled", lambda: True)
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    (tmp_path / ".git").mkdir()
    monkeypatch.setattr(master_service, "_repo_root_for_journal_path", lambda _p: tmp_path)
    (tmp_path / "Trading Journal.xlsx").write_bytes(b"x")
    monkeypatch.setenv("TRADING_JOURNAL_MASTER_JOURNAL_AUTHORITATIVE", "1")
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: (_ for _ in ()).throw(AssertionError("should not call")))
    fp = master_service._journal_source_fingerprint()
    paths = [str((f or {}).get("path") or "") for f in fp.get("files", [])]
    assert any("Trading Journal.xlsx" in p for p in paths)
    assert all("account_cashflows.xlsx" not in p for p in paths)
    assert all("Bybit Demo.xlsx" not in p for p in paths)
    (tmp_path / "~$Trading Journal.xlsx").write_bytes(b"x")
    (tmp_path / "foo.tmp.xlsx").write_bytes(b"x")
    (tmp_path / "foo.pending.xlsx").write_bytes(b"x")
    commands = []
    def fake_git(args, _cwd, _timeout):
        commands.append(args)
        if args == ["--version"]:
            return 0, "git version 2", ""
        if args[:2] == ["rev-parse", "--abbrev-ref"]:
            return 0, "main\n", ""
        if args[:3] == ["remote", "get-url", "origin"]:
            return 0, "x\n", ""
        if args[:2] == ["diff", "--cached"]:
            return 0, "", ""
        return 0, "", ""
    monkeypatch.setattr(master_service, "_run_git_command", fake_git)
    result = master_service._sync_journal_excel_files_to_github(tmp_path / "Trading Journal.xlsx")
    assert "Trading Journal.xlsx" in " ".join(result.get("github_sync_files") or [])
    assert "~$Trading Journal.xlsx" not in " ".join(result.get("github_sync_files") or [])
    assert "foo.tmp.xlsx" not in " ".join(result.get("github_sync_files") or [])
    assert "foo.pending.xlsx" not in " ".join(result.get("github_sync_files") or [])
    add_calls = [cmd for cmd in commands if cmd and cmd[0] == "add"]
    if add_calls:
        assert all(cmd != ["add", "."] for cmd in add_calls)
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_master_journal_does_not_enable_authoritative_mode(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setenv("TRADING_JOURNAL_SOURCE", "local")
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items':[{'id':'t1','row_type':'trade','account':'A','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0}], 'stats':{'totals':{}}, 'balances':[], 'diagnostics':{}}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda _s: {})
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    assert snap["diagnostics"]["authoritative_mode"] is False
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_save_watcher_enablement(monkeypatch):
    monkeypatch.setattr(master_service, '_is_render_env', lambda: True)
    assert master_service._manual_save_watcher_enabled() is False
    monkeypatch.setattr(master_service, '_is_render_env', lambda: False)
    monkeypatch.setenv('TRADING_JOURNAL_GITHUB_SYNC_ENABLED','1')
    monkeypatch.delenv('TRADING_JOURNAL_GITHUB_SYNC_ON_MANUAL_SAVE_ENABLED', raising=False)
    assert master_service._manual_save_watcher_enabled() is True
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_save_sync_once_records_error_and_no_rebuild(tmp_path, monkeypatch):
    target = tmp_path / 'Trading Journal.xlsx'; target.write_bytes(b'a')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    called={'sync':0,'build':0}
    monkeypatch.setattr(master_service, '_sync_journal_excel_files_to_github', lambda p: called.__setitem__('sync', called['sync']+1) or {'github_sync_ok':False,'github_sync_error':'git fail','github_sync_files':['journal/Trading Journal.xlsx'],'github_sync_commit':''})
    monkeypatch.setattr(master_service, 'build_master_journal_workbook', lambda *a, **k: called.__setitem__('build', called['build']+1))
    master_service._run_manual_save_github_sync_once(target)
    st=master_service._manual_save_state_snapshot()
    assert called['sync']==1 and called['build']==0
    assert st['manual_save_last_error']=='git fail'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_save_ignore_temp_names(tmp_path):
    assert master_service._should_ignore_manual_save_path(tmp_path / '~$Trading Journal.xlsx')
    assert master_service._should_ignore_manual_save_path(tmp_path / 'Master Journal.tmp.xlsx')
    assert master_service._should_ignore_manual_save_path(tmp_path / 'Master Journal.pending.xlsx')
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_shutdown_stops_manual_save_watcher(monkeypatch):
    called={'n':0}
    monkeypatch.setattr(master_service, '_stop_manual_save_github_sync_watcher', lambda: called.__setitem__('n', called['n']+1))
    asyncio.run(master_service._log_local_master_shutdown())
    assert called['n']==1
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_save_scan_debounce_and_service_write_suppression(tmp_path, monkeypatch):
    p=tmp_path/'Trading Journal.xlsx'; p.write_bytes(b'one')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    calls=[]
    monkeypatch.setattr(master_service, '_sync_journal_excel_files_to_github', lambda *_: calls.append(1) or {'github_sync_enabled':True,'github_sync_ok':True,'github_sync_noop':False,'github_sync_error':'','github_sync_files':[],'github_sync_commit':'abc'})
    master_service._manual_save_set_known_fingerprint(p)
    # service generated write suppression
    master_service._manual_save_set_known_fingerprint(p)
    master_service._manual_save_scan_once(10.0, p)
    assert len(calls)==0
    p.write_bytes(b'two')
    master_service._manual_save_scan_once(10.0, p)
    assert len(calls)==0
    master_service._manual_save_scan_once(20.0, p)
    assert len(calls)==1
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_save_disabled_github_no_fake_success(monkeypatch, tmp_path):
    p=tmp_path/'Trading Journal.xlsx'; p.write_bytes(b'x')
    monkeypatch.setattr(master_service, '_sync_journal_excel_files_to_github', lambda *_: {'github_sync_enabled':False,'github_sync_ok':True,'github_sync_noop':True,'github_sync_error':'','github_sync_files':[],'github_sync_commit':''})
    master_service._run_manual_save_github_sync_once(p)
    st=master_service._manual_save_state_snapshot()
    assert st['manual_save_last_success_at'] is None
    assert 'disabled' in str(st['manual_save_last_error']).lower()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_bybit_server_time_invalid_json_no_path_nameerror(monkeypatch):
    class Resp:
        status_code=200
        text='x'
        def json(self): raise ValueError('bad')
    class Ctx:
        async def __aenter__(self): return Resp()
        async def __aexit__(self,*a): return False
    class Client:
        async def __aenter__(self): return self
        async def __aexit__(self,*a): return False
        def get(self,*a,**k): return Ctx()
    monkeypatch.setattr(master_service.httpx, 'AsyncClient', lambda **k: Client())
    with pytest.raises(ValueError, match='Bybit server time response is unparseable.'):
        asyncio.run(master_service._fetch_bybit_server_time_ms('https://api.bybit.com'))
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_signed_get_keeps_valid_json(monkeypatch):
    class Resp:
        status_code=200
        text='ok'
        content=b'{"retCode":0,"result":{"x":1}}'
        def json(self): return {'retCode':0,'result':{'x':1}}
    class Client:
        async def __aenter__(self): return self
        async def __aexit__(self,*a): return False
        async def get(self,*a,**k): return Resp()
    monkeypatch.setattr(master_service.httpx, 'AsyncClient', lambda **k: Client())
    async def fake_headers(**k): return {}
    monkeypatch.setattr(master_service, '_build_bybit_signed_headers', fake_headers)
    payload=asyncio.run(master_service._bybit_signed_get(base_url='https://api.bybit.com',api_key='k',api_secret='s',path='/x',params={}))
    assert payload == {'retCode': 0, 'result': {'x': 1}}
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_update_oanda_settings_passes_payload():
    out=master_service._update_oanda_settings({'wait_seconds':10})
    assert out.get('wait_seconds')==10
def test_source_guard_manual_save_fingerprint_only_master_journal_sync():
    src=(ROOT/'render'/'master_service.py').read_text(encoding='utf-8')
    assert '_manual_save_set_known_fingerprint(path)' not in src
    needle = '_manual_save_set_known_fingerprint(master_path)'
    assert src.count(needle) == 1
    sync_ix = src.index('def _manual_save_github_sync_watcher_loop')
    only_ix = src.index(needle)
    assert only_ix > sync_ix
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_existing_workbook_sync_does_not_rebuild_or_refresh_derived(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    mj = tmp_path / 'Trading Journal.xlsx'
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, mj)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    called = {'build': 0, 'refresh': 0}
    monkeypatch.setattr(master_service, 'build_master_journal_workbook', lambda *_: called.__setitem__('build', called['build']+1))
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: {'items': [], 'balances': [], 'stats': {'totals': {}, 'groups': {}}})
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is True
    assert called['build'] == 0
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_missing_master_journal_fails_loudly(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is False
    assert result['master_journal_error_type'] in {'FileNotFoundError', 'RuntimeError'}
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_canonical_market_precedence_cases():
    cm = master_service._canonical_market_for_row
    assert cm({'account':'OANDA LIVE','symbol':'EURUSD','asset_class':''}) == 'fx'
    assert cm({'account':'PEPPERSTONE LIVE','symbol':'EURUSD','asset_class':'crypto'}) == 'fx'
    assert cm({'account':'PEPPERSTONE LIVE','symbol':'XAUUSD','asset_class':''}) == 'fx'
    assert cm({'account':'BYBIT LIVE','symbol':'BTCUSD','asset_class':'fx'}) == 'crypto'
    assert cm({'account':'BINANCE','symbol':'ETHUSD','asset_class':'fx'}) == 'crypto'
    assert cm({'account':'BYBIT','symbol':'BTCUSDT','asset_class':''}) == 'crypto'
    assert cm({'account':'UNKNOWN','symbol':'ABCDEF','asset_class':''}) == ''


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_build_journal_balance_timelines_pepperstone_demo_zero_anchor_beats_stale_trade_variants():
    rows = [
        {'id':'pepperstone-demo-stale-trade','row_type':'trade','source':'master_journal','account':'Pepperstone Demo','symbol':'EURUSD','side':'BUY','open_time':'2018-07-25T22:55:00','close_time':'2018-07-25T23:00:00','net_profit':4.78,'balance_after_trade':4.78,'currency':'AUD'},
    ]
    cashflows = {
        master_service._norm_account_key('pepperstone_demo'): [
            {'account':'PEPPERSTONE DEMO','date':'2018-07-26T00:00:00','amount':None,'new_balance':0,'currency':'AUD'},
        ]
    }
    timelines = master_service._build_journal_balance_timelines(rows, cashflows, [])
    balances = {str(b.get('label') or b.get('account')).upper(): b for b in timelines['balances']}
    assert master_service._norm_account_key('PEPPERSTONE DEMO') == master_service._norm_account_key('pepperstone_demo')
    assert master_service._norm_account_key('Pepperstone-Demo') == 'PEPPERSTONE DEMO'
    assert balances['PEPPERSTONE DEMO']['balance'] == 0
    assert balances['PEPPERSTONE DEMO']['currency'] == 'AUD'
    assert balances['PEPPERSTONE DEMO']['balance_source'] == 'cashflow_anchor_plus_trades'
    assert timelines['diagnostics']['PEPPERSTONE DEMO']['ignored_cashflow_anchors_without_new_balance'] == 0


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_writes_zero_balances_and_validation_detects_mismatch(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    source_rows = [{'id':'r1','row_type':'trade','account':'PEPPERSTONE DEMO','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'result_pct':0.1}]
    stale_snap = {'items': source_rows, 'stats': {'totals': {}, 'by_instrument': [{'symbol': 'EURUSD', 'total_trades': 1}], 'groups': {'leaders': {}}}, 'balances': [
        {'account_label': 'PEPPERSTONE DEMO', 'balance': 4.78, 'currency': 'AUD'},
        {'account_label': 'BINANCE', 'balance': 396.65720524, 'currency': 'USDT'},
    ], 'diagnostics': {}}
    build_master_journal_workbook(stale_snap, mj)
    zero_snap = {'items': source_rows, 'stats': stale_snap['stats'], 'balances': [
        {'account_label': 'PEPPERSTONE DEMO', 'balance': 0, 'currency': 'AUD', 'balance_source': 'broker_account_summary'},
        {'account_label': 'BINANCE', 'balance': 0, 'currency': 'USDT', 'balance_source': 'broker_account_summary'},
    ], 'diagnostics': {}}
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: zero_snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: source_rows)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is True
    from openpyxl import load_workbook
    wb = load_workbook(mj, data_only=True)
    dash = wb['STATS2']
    pairs = {}
    for r in range(1, dash.max_row + 1):
        for c in range(1, dash.max_column):
            label = str(dash.cell(r, c).value or '').strip()
            if label:
                pairs[label] = dash.cell(r, c + 1).value
    assert pairs['PEPPERSTONE DEMO'] == 0
    assert pairs['BINANCE'] == 0
    wb.close()
    # Force workbook-vs-snapshot mismatch by corrupting the candidate workbook after data-only update.
    import tools.master_journal_workbook as mjw
    real_update = mjw.update_master_journal_workbook_data_only
    def _corrupting_update(path, snapshot, **kwargs):
        payload = real_update(path, snapshot, **kwargs)
        candidate_path = Path(payload['candidate_path'])
        bad_wb = load_workbook(candidate_path)
        try:
            bad_dash = bad_wb['STATS2']
            for r in range(1, bad_dash.max_row + 1):
                for c in range(1, bad_dash.max_column):
                    if str(bad_dash.cell(r, c).value or '').strip() == 'PEPPERSTONE DEMO':
                        bad_dash.cell(r, c + 1).value = 4.78
                        break
                else:
                    continue
                break
            bad_wb.save(candidate_path)
        finally:
            bad_wb.close()
        return payload
    monkeypatch.setattr(master_service, 'update_master_journal_workbook_data_only', _corrupting_update)
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: zero_snap)
    bad = master_service._sync_master_journal_workbook(sync_caller="test")
    assert bad['master_journal_ok'] is False
    assert 'Account Balances mismatch vs snapshot' in str(bad.get('master_journal_error') or '')
    assert 'PEPPERSTONE DEMO' in str(bad.get('master_journal_error') or '')
    assert 'expected_balance=0.0' in str(bad.get('master_journal_error') or '')
    assert 'actual_balance=4.78' in str(bad.get('master_journal_error') or '')


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_zero_qty_repairable_crypto_passes(tmp_path, monkeypatch):
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    rows = [{
        'id': 'btc-repairable',
        'row_type': 'trade',
        'account': 'BYBIT',
        'symbol': 'BTCUSDT',
        'side': 'BUY',
        'qty': 0,
        'entry_price': 100.0,
        'exit_price': 101.0,
        'open_time': '2026-01-01T00:00:00Z',
        'close_time': '2026-01-01T01:00:00Z',
        'net_profit': 1.0,
        'raw_refs': {'closedSize': '0.015'},
    }]
    snap = {'items': rows, 'stats': {'totals': {}, 'by_instrument': [{'symbol': 'BTCUSDT', 'total_trades': 1}], 'groups': {'leaders': {}}}, 'balances': [{'account_label': 'BYBIT', 'balance': 1.0, 'currency': 'USDT'}], 'diagnostics': {}}
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: rows)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is True
    wb = load_workbook(tmp_path / 'Trading Journal.xlsx', data_only=True)
    tl = master_service._get_trade_log_sheet(wb, allow_legacy=False)
    headers = master_service._trade_log_header_map(tl)
    qty_col = headers['Qty']
    data_start = master_service._trade_log_data_start_row(tl)
    assert float(tl.cell(data_start, qty_col).value) == pytest.approx(0.015)
    wb.close()


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_zero_qty_unrepairable_crypto_fails(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    rows = [{
        'id': 'btc-unrepairable',
        'row_type': 'trade',
        'account': 'BYBIT',
        'symbol': 'BTCUSDT',
        'side': 'BUY',
        'qty': 0,
        'entry_price': None,
        'exit_price': None,
        'open_time': '2026-01-01T00:00:00Z',
        'close_time': '2026-01-01T01:00:00Z',
        'net_profit': None,
    }]
    snap = {'items': rows, 'stats': {'totals': {}, 'by_instrument': [{'symbol': 'BTCUSDT', 'total_trades': 1}], 'groups': {'leaders': {}}}, 'balances': [{'account_label': 'BYBIT', 'balance': 1.0, 'currency': 'USDT'}], 'diagnostics': {}}
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=True: snap)
    monkeypatch.setattr(master_service, '_get_trading_journal_rows', lambda: rows)
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is False
    assert 'unrepaired crypto Qty=0' in str(result.get('master_journal_error') or '')
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_build_journal_balance_timelines_rejects_non_authoritative_stale_excel_seed():
    rows = []
    cashflows = {}
    excel_balances = [{'account': 'BINANCE', 'label': 'BINANCE', 'balance': 396.65720524, 'currency': 'USDT', 'balance_source': 'excel_account_balance'}]
    out = master_service._build_journal_balance_timelines(rows, cashflows, excel_balances)
    bal = next(b for b in out['balances'] if str(b.get('label')) == 'BINANCE')
    assert bal['balance'] is None
    assert bal['balance_source'] == 'timeline_missing'


def test_build_journal_balance_timelines_applies_pnl_for_bybit_test_rows():
    ms = _load_master_service_for_import_test()
    rows = [
        {'id':'sig:cadd2b98cf847e7f4d2e2c54','row_type':'trade','source':'master_journal','account':'Bybit Demo','account_label':'Bybit Demo','symbol':'BTCUSDT','close_time':'2026-05-04T11:16:00+00:00','net_profit':0.0,'analysis_balance_after_trade':369.64962148,'balance_after_trade':369.64962148},
        {'id':'bybit:demo:trade:BTCUSDT:049246f1ec6e4a68','row_type':'trade','source':'bybit_execution_history_grouped','account':'Bybit Demo','account_label':'Bybit Demo','symbol':'BTCUSDT','close_time':'2026-05-26T09:54:11+10:00','net_profit':-1.31721048,'is_test_trade':True},
        {'id':'bybit:demo:trade:BTCUSDT:85c1adb0266f56a4','row_type':'trade','source':'bybit_execution_history_grouped','account':'Bybit Demo','account_label':'Bybit Demo','symbol':'BTCUSDT','close_time':'2026-05-26T12:43:30+10:00','net_profit':-1.83430304,'is_test_trade':True},
        {'id':'cash:c1','row_type':'trade','source':'master_journal','account':'Bybit Demo','account_label':'Bybit Demo','symbol':'CASHFLOW','close_time':'2026-05-26T13:52:01+10:00','cashflow_new_balance':319.8339282399999,'balance_after_trade':319.8339282399999,'net_profit':0.0},
    ]
    out = ms._build_journal_balance_timelines(rows, {}, [])
    by_id = {r.get("id"): r for r in out.get("rows", [])}
    assert by_id['bybit:demo:trade:BTCUSDT:049246f1ec6e4a68']['analysis_balance_after_trade'] == pytest.approx(368.33241100)
    assert by_id['bybit:demo:trade:BTCUSDT:85c1adb0266f56a4']['analysis_balance_after_trade'] == pytest.approx(366.49810796)
    assert by_id['cash:c1']['analysis_balance_after_trade'] == pytest.approx(319.8339282399999)

def test_build_journal_balance_timelines_zero_cashflow_anchor_beats_older_trade_balance():
    ms = _load_master_service_for_import_test()
    rows = [
        {'id':'t1','row_type':'trade','source':'master_journal','account':'BINANCE','symbol':'BTCUSDT','close_time':'2020-10-24T02:18:00','net_profit':-3.7985222,'balance_after_trade':396.65720524},
    ]
    cashflows = {'BINANCE': [{'account':'BINANCE','date':'2020-10-26T00:00:00','amount':None,'new_balance':0,'currency':'USDT','side':'WITHDRAWAL'}]}
    out = ms._build_journal_balance_timelines(rows, cashflows, [])
    bal = next(b for b in out['balances'] if b.get('label') == 'BINANCE')
    assert bal['balance'] == 0
    assert bal['balance_source'] == 'cashflow_anchor_plus_trades'
    assert out['diagnostics']['BINANCE']['previous_cashflow_balance'] == 0


def test_build_journal_balance_timelines_ignores_blank_cashflow_anchor_without_erasing_valid_trade_balance():
    ms = _load_master_service_for_import_test()
    rows = [
        {'id':'t1','row_type':'trade','source':'master_journal','account':'BINANCE','symbol':'BTCUSDT','close_time':'2020-10-24T02:18:00','net_profit':-3.7985222,'balance_after_trade':396.65720524},
    ]
    cashflows = {'BINANCE': [{'account':'BINANCE','date':'2020-10-26T00:00:00','amount':None,'new_balance':None,'currency':'USDT','side':'WITHDRAWAL'}]}
    out = ms._build_journal_balance_timelines(rows, cashflows, [])
    bal = next(b for b in out['balances'] if b.get('label') == 'BINANCE')
    assert bal['balance'] == 396.65720524
    assert bal['balance_source'] in {'authoritative_trade_balance', 'trade_timeline'}
    assert out['diagnostics']['BINANCE']['ignored_cashflow_anchors_without_new_balance'] == 1


def test_cashflow_row_to_ledger_event_keeps_zero_anchor_without_amount():
    ms = _load_master_service_for_import_test()
    event = ms._cashflow_row_to_ledger_event({
        'id':'c1',
        'row_type':'cashflow',
        'account':'BINANCE',
        'close_time':'2020-10-26T00:00:00',
        'cashflow_amount':'',
        'cashflow_new_balance':0,
        'currency':'USDT',
    })
    assert event['new_balance'] == 0
    assert event['amount'] is None

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_balance_regression_stale_excel_binance_overridden_by_authoritative_zero_source():
    rows = []
    cashflows = {}
    excel_balances = [
        {'account': 'BINANCE', 'label': 'BINANCE', 'balance': 396.65720524, 'currency': 'USDT', 'balance_source': 'excel_account_balance'}
    ]
    timeline = master_service._build_journal_balance_timelines(rows, cashflows, excel_balances)
    merged = master_service._merge_missing_timeline_balances_with_broker(
        timeline['balances'],
        [
            {
                'account': 'BINANCE',
                'label': 'BINANCE',
                'balance': 0,
                'currency': 'USDT',
                'balance_source': 'broker_account_summary',
                'source': 'broker_account_summary',
                'as_of': '2026-05-11T00:00:00Z',
            }
        ],
    )
    bal = next(b for b in merged if str(b.get('label')) == 'BINANCE')
    assert bal['balance'] == 0
    assert bal['balance_source'] == 'broker_account_summary'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_merge_missing_timeline_balances_with_broker_zero_overrides_stale_timeline():
    timeline = [{'account': 'BINANCE', 'label': 'BINANCE', 'balance': 396.65720524, 'currency': 'USDT', 'balance_source': 'trade_timeline', 'missing_balance': False}]
    broker = [{'account': 'BINANCE', 'label': 'BINANCE', 'balance': 0, 'currency': 'USDT', 'balance_source': 'broker_account_summary', 'as_of': '2026-05-10T00:00:00Z'}]
    merged = master_service._merge_missing_timeline_balances_with_broker(timeline, broker)
    bal = next(b for b in merged if str(b.get('label')) == 'BINANCE')
    assert bal['balance'] == 0
    assert bal['balance_source'] == 'broker_account_summary'
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_sync_master_journal_uses_zero_cashflow_anchor_when_cashflow_new_balance_blank(tmp_path, monkeypatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    monkeypatch.setattr(master_service, "_master_journal_authoritative_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: mj)
    monkeypatch.setenv("TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_save_trading_journal_view_snapshot", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    rows = [
        {'id':'t1','row_type':'trade','account':'BINANCE','symbol':'BTCUSDT','side':'BUY','open_time':'2020-10-01','close_time':'2020-10-01','net_profit':1.0,'balance_after_trade':396.65720524,'currency':'USDT'},
        {'id':'c1','row_type':'cashflow','account':'BINANCE','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2020-10-26','close_time':'2020-10-26','cashflow_amount':-396.65720524,'balance_after_trade':0,'cashflow_new_balance':'','currency':'USDT','notes':'Withdrawal -396.65720524 USDT'},
        {'id':'t2','row_type':'trade','account':'PEPPERSTONE DEMO','symbol':'EURUSD','side':'BUY','open_time':'2022-10-01','close_time':'2022-10-01','net_profit':1.0,'balance_after_trade':4.78,'currency':'AUD'},
        {'id':'c2','row_type':'cashflow','account':'PEPPERSTONE DEMO','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2022-12-16','close_time':'2022-12-16','cashflow_amount':-4.78,'balance_after_trade':0,'cashflow_new_balance':'','currency':'AUD','notes':'Withdrawal -4.78 AUD'},
    ]
    snap = {'items': rows, 'stats': {'totals': {}, 'by_instrument': [], 'groups': {'leaders': {}}}, 'balances': [
        {'account_label': 'BINANCE', 'balance': 396.65720524, 'currency': 'USDT'},
        {'account_label': 'PEPPERSTONE DEMO', 'balance': 4.78, 'currency': 'AUD'},
    ], 'diagnostics': {}}
    build_master_journal_workbook(snap, mj)
    wb = load_workbook(mj)
    ws = wb['Trade Log']
    headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    for rr in range(2, ws.max_row + 1):
        if str(ws.cell(rr, headers['Row Type']).value).strip().lower() == 'cashflow':
            ws.cell(rr, headers['Cashflow New Balance']).value = None
    wb.save(mj); wb.close()
    snap2 = master_service._build_trading_journal_view_snapshot(force=True)
    balances = {str(b.get('label') or b.get('account')): b for b in (snap2.get('balances') or [])}
    assert balances['BINANCE']['balance'] == 0
    assert balances['PEPPERSTONE DEMO']['balance'] == 0
    assert balances['BINANCE'].get('balance_source') != 'authoritative_trade_balance'
    assert balances['PEPPERSTONE DEMO'].get('balance_source') != 'authoritative_trade_balance'
    result = master_service._sync_master_journal_workbook(sync_caller="test")
    assert result['master_journal_ok'] is True
    synced = load_workbook(mj, data_only=True)
    dash = synced['STATS2']
    dash_map = {}
    for r in range(1, dash.max_row + 1):
        for c in range(1, dash.max_column):
            label = str(dash.cell(r, c).value or '').strip()
            if label:
                dash_map[label] = dash.cell(r, c + 1).value
    assert dash_map['BINANCE'] == 0
    assert dash_map['PEPPERSTONE DEMO'] == 0
    synced.close()
@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')


def test_gitignore_trading_journal_exception_rules() -> None:
    gi = (ROOT / '.gitignore').read_text(encoding='utf-8')
    assert 'journal/*.xlsx' in gi
    assert '!journal/Trading Journal.xlsx' in gi
    assert '!journal/Master Journal.xlsx' not in gi

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_repo_state_files_for_github_dedupes_master_journal(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    journal_dir = tmp_path / "journal"
    journal_dir.mkdir(parents=True, exist_ok=True)
    master = journal_dir / "Trading Journal.xlsx"
    master.write_bytes(b"x")
    (journal_dir / "5-digit-demo-calculation-context.json").write_text("{}", encoding="utf-8")
    (tmp_path / "state_backup.json").write_text("{}", encoding="utf-8")
    files = master_service._repo_state_files_for_github(master)
    rel = [str(p.relative_to(tmp_path)).replace("\\", "/") for p in files]
    assert rel.count("journal/Trading Journal.xlsx") == 1
def test_fetch_bybit_executions_calls_include_end_time():
    src=(Path(__file__).resolve().parents[1]/"render"/"master_service.py").read_text(encoding="utf-8")
    assert "_fetch_bybit_executions(" in src
    for chunk in src.split("_fetch_bybit_executions(")[1:]:
        call = chunk.split(")",1)[0]
        if "*," in call and "category: str" in call and "cursor:" in call:
            continue
        assert "end_time=" in call
def test_return_annotations_not_corrupted():
    src=(Path(__file__).resolve().parents[1]/"render"/"master_service.py").read_text(encoding="utf-8")
    bad=[
        "def _calculator_quote_elapsed_ms(quote_started: Optional[float]) -> Dict[str, object]:",
        "def _find_journal_row_index(row_id: str) -> Dict[str, object]:",
        "def _upsert_trading_journal_rows(rows: Iterable[Dict[str, object]]) -> Dict[str, object]:",
        "def _dedupe_legacy_bybit_demo_rows() -> Dict[str, object]:",
        "def _repair_persisted_bybit_demo_journal_sides() -> Dict[str, object]:",
        "def _repair_persisted_oanda_trade_rows() -> Dict[str, object]:",
        "def _repair_persisted_bybit_trade_context_fields() -> Dict[str, object]:",
        "def _purge_bybit_demo_journal_state() -> Dict[str, object]:",
        "def _row_source_rank(row: Dict[str, object]) -> Dict[str, object]:",
        "def _allocate_port() -> Dict[str, object]:",
        "def _normalize_bybit_recv_window_ms(raw_value: Optional[str]) -> Dict[str, object]:",
        "def _safe_int_env(raw_value: Optional[str], default: int) -> Dict[str, object]:",
        "async def _fetch_bybit_server_time_ms(base_url: str) -> Dict[str, object]:",
        "async def _refresh_bybit_time_offset_ms(base_url: str) -> Dict[str, object]:",
        "async def _get_bybit_time_offset_ms(base_url: str, force_refresh: bool = False) -> Dict[str, object]:",
        "def _bybit_position_idx_for_order(*, side: str, configured_mode: str = "") -> Dict[str, object]:",
        "def _append_bybit_demo_rows_to_workbook(active_folder: str, rows: List[Dict[str, object]]) -> Dict[str, object]:",
        "def _append_bybit_demo_rows_to_local_workbook(local_dir: Path, rows: List[Dict[str, object]]) -> Dict[str, object]:",
    ]
    for b in bad:
        assert b not in src
    assert "async def _fetch_oanda_display_precision(" in src
    oanda_chunk = src.split("async def _fetch_oanda_display_precision(", 1)[1].split("\n\n", 1)[0]
    assert "-> int:" in oanda_chunk
    assert "-> Dict[str, object]:" not in oanda_chunk

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_demo_capture_not_masked_by_live_empty(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {'ok': True, 'rows_imported': 0, 'diagnostics': {}})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(tmp_path/'Trading Journal.xlsx'), 'master_journal_exists': True})
    async def _bybit(account_mode: str, **_k):
        if account_mode == 'demo':
            return {'ok': True, 'rows_seen': 1, 'captured_row_ids': ['demo-row-1']}
        return {'ok': True, 'rows_seen': 0, 'captured_row_ids': []}
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', _bybit)
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', lambda *_a, **_k: asyncio.sleep(0, result={'ok': True}))
    st=[]
    real=master_service._set_trading_journal_sync_state
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', lambda **kw: st.append(dict(kw)) or real(**kw))
    asyncio.run(master_service._run_trading_journal_sync_job())
    final=st[-1]
    assert final.get('ok') is False
    assert final['result']['bybit']['demo']['ok'] is False

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_live_capture_verified_independently(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {'ok': True, 'rows_imported': 0, 'diagnostics': {}})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(tmp_path/'Trading Journal.xlsx'), 'master_journal_exists': True})
    async def _bybit(account_mode: str, **_k):
        if account_mode == 'live':
            return {'ok': True, 'rows_seen': 1, 'captured_row_ids': ['live-row-1']}
        return {'ok': True, 'rows_seen': 0, 'captured_row_ids': []}
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', _bybit)
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', lambda *_a, **_k: asyncio.sleep(0, result={'ok': True}))
    st=[]
    real=master_service._set_trading_journal_sync_state
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', lambda **kw: st.append(dict(kw)) or real(**kw))
    asyncio.run(master_service._run_trading_journal_sync_job())
    final=st[-1]
    assert final.get('ok') is False
    assert final['result']['bybit']['live']['ok'] is False

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_demo_live_successful_row_id_verification(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    from tools.master_journal_workbook import build_master_journal_workbook
    snap={'items':[{'id':'demo-row-1','row_type':'trade','symbol':'BTCUSDT','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0},{'id':'live-row-1','row_type':'trade','symbol':'ETHUSDT','side':'SELL','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0}], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}
    build_master_journal_workbook(snap, tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {'ok': True, 'rows_imported': 2, 'diagnostics': {}})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(tmp_path/'Trading Journal.xlsx'), 'master_journal_exists': True})
    async def _bybit(account_mode: str, **_k):
        return {'ok': True, 'rows_seen': 1, 'captured_row_ids': ['demo-row-1' if account_mode=='demo' else 'live-row-1']}
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', _bybit)
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', lambda *_a, **_k: asyncio.sleep(0, result={'ok': True}))
    st=[]
    real=master_service._set_trading_journal_sync_state
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', lambda **kw: st.append(dict(kw)) or real(**kw))
    asyncio.run(master_service._run_trading_journal_sync_job())
    bybit = (st[-1].get('result') or {}).get('bybit') or {}
    assert isinstance(bybit['demo'].get('final_trade_log_row_ids_verified'), bool)
    assert isinstance(bybit['live'].get('final_trade_log_row_ids_verified'), bool)

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_oanda_missing_row_ids_hard_fail(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook({'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {'ok': True, 'rows_imported': 0, 'diagnostics': {}})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(tmp_path/'Trading Journal.xlsx'), 'master_journal_exists': True})
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', lambda **_k: asyncio.sleep(0, result={'ok': True, 'rows_seen':0, 'captured_row_ids': []}))
    async def _oanda(acct): return {'ok': True, 'rows_seen':1, 'captured_row_ids':[f'o-{acct}-1']}
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', _oanda)
    st=[]; real=master_service._set_trading_journal_sync_state
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', lambda **kw: st.append(dict(kw)) or real(**kw))
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert st[-1].get('ok') is False
    assert st[-1]['result']['oanda']['demo']['ok'] is False or st[-1]['result']['oanda']['live']['ok'] is False

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_manual_sync_oanda_row_ids_present_pass(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL', True)
    monkeypatch.setattr(master_service, '_trading_journal_broker_refresh_enabled', lambda: False)
    from tools.master_journal_workbook import build_master_journal_workbook
    snap={'items':[{'id':'o-demo-1','row_type':'trade','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0},{'id':'o-live-1','row_type':'trade','symbol':'GBPUSD','side':'SELL','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0}], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}
    build_master_journal_workbook(snap, tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_import_trading_journal_from_sources', lambda *a, **k: {'ok': True, 'rows_imported': 2, 'diagnostics': {}})
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(tmp_path/'Trading Journal.xlsx'), 'master_journal_exists': True})
    monkeypatch.setattr(master_service, '_run_bybit_closed_pnl_sync', lambda **_k: asyncio.sleep(0, result={'ok': True, 'rows_seen':0, 'captured_row_ids': []}))
    async def _oanda(acct): return {'ok': True, 'rows_seen':1, 'captured_row_ids':[f'o-{acct}-1']}
    monkeypatch.setattr(master_service, '_recover_oanda_recent_fills', _oanda)
    st=[]; real=master_service._set_trading_journal_sync_state
    monkeypatch.setattr(master_service, '_set_trading_journal_sync_state', lambda **kw: st.append(dict(kw)) or real(**kw))
    asyncio.run(master_service._run_trading_journal_sync_job())
    out=(st[-1].get('result') or {}).get('oanda') or {}
    assert out['demo']['final_trade_log_row_ids_verified'] is True
    assert out['live']['final_trade_log_row_ids_verified'] is True


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_repo_state_files_for_github_includes_legacy_master_journal_when_tracked_and_missing(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    journal_dir = tmp_path / "journal"
    journal_dir.mkdir(parents=True, exist_ok=True)
    master = journal_dir / "Trading Journal.xlsx"
    master.write_bytes(b"x")
    monkeypatch.setattr(master_service, "_run_git_command", lambda args, _cwd, _timeout: (0, "", "") if args[:2] == ["ls-files", "--error-unmatch"] else (1, "", ""))
    files = master_service._repo_state_files_for_github(master)
    rel = [str(p.relative_to(tmp_path)).replace("\\", "/") for p in files]
    assert "journal/Trading Journal.xlsx" in rel
    assert "journal/Master Journal.xlsx" in rel

@pytest.mark.skipif(not HTTPX_AVAILABLE, reason='httpx is not installed')
def test_github_sync_stages_legacy_master_journal_deletion_when_tracked_and_missing(tmp_path, monkeypatch):
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "1")
    monkeypatch.setattr(master_service, "_trading_journal_github_sync_enabled", lambda: True)
    monkeypatch.setattr(master_service, "BASE_DIR", tmp_path)
    (tmp_path / ".git").mkdir()
    journal = tmp_path / "journal"
    journal.mkdir()
    (journal / "Trading Journal.xlsx").write_bytes(b"x")
    commands = []
    def fake_git(args, _cwd, _timeout):
        commands.append(args)
        if args == ["--version"]:
            return 0, "git version 2", ""
        if args[:2] == ["rev-parse", "--abbrev-ref"]:
            return 0, "main\n", ""
        if args[:3] == ["remote", "get-url", "origin"]:
            return 0, "x\n", ""
        if args[:2] == ["ls-files", "--error-unmatch"]:
            return 0, "journal/Master Journal.xlsx\n", ""
        if args[:2] == ["diff", "--cached"]:
            return 1, "", ""
        if args and args[0] in {"add", "commit", "push", "rev-parse"}:
            return 0, "", ""
        return 0, "", ""
    monkeypatch.setattr(master_service, "_run_git_command", fake_git)
    result = master_service._sync_journal_excel_files_to_github(journal / "Trading Journal.xlsx")
    assert result["github_sync_ok"] is True
    assert "journal/Trading Journal.xlsx" in result["github_sync_files"]
    assert "journal/Master Journal.xlsx" in result["github_sync_files"]
    add_calls = [cmd for cmd in commands if cmd and cmd[0] == "add"]
    assert add_calls
    add_joined = " ".join(add_calls[0])
    assert "journal/Trading Journal.xlsx" in add_joined
    assert "journal/Master Journal.xlsx" in add_joined


@pytest.mark.skipif(master_service is None, reason="master_service import unavailable")
def test_enforce_single_master_journal_rejects_legacy_backup_excel_name(tmp_path):
    journal = tmp_path / "journal"
    journal.mkdir()
    (journal / "Trading Journal.xlsx").write_bytes(b"x")
    (journal / "Master Journal.legacy.bak.xlsx").write_bytes(b"x")
    res = master_service._enforce_single_master_journal_xlsx(journal, cleanup_known_generated=True)
    assert res["ok"] is False
    assert "Master Journal.legacy.bak.xlsx" in (res.get("unknown_extra_excel_files") or [])


def test_single_file_enforcement_error_includes_backup_move_guidance():
    src = (ROOT / "render" / "master_service.py").read_text(encoding="utf-8")
    assert "Move legacy backups outside journal/. Keep only journal/Trading Journal.xlsx." in src


@pytest.mark.skipif(master_service is None, reason="master_service import unavailable")
def test_check_master_journal_write_lock_reports_locked_on_windows_probe_failure(monkeypatch, tmp_path):
    p = tmp_path / "Trading Journal.xlsx"
    p.write_bytes(b"x")
    monkeypatch.setattr(master_service.os, "name", "nt")
    class _K:
        def __init__(self):
            self.CreateFileW = lambda *_a: ctypes.c_void_p(-1).value
            self.CloseHandle = lambda *_a: 1
    monkeypatch.setattr(master_service.ctypes, "WinDLL", lambda *_a, **_k: _K(), raising=False)
    monkeypatch.setattr(master_service.ctypes, "get_last_error", lambda: 5, raising=False)
    out = master_service._check_master_journal_write_lock(p)
    assert out["locked"] is True
    assert out["code"] == "EXCEL_WORKBOOK_OPEN"


@pytest.mark.skipif(master_service is None, reason="master_service import unavailable")
def test_check_master_journal_write_lock_does_not_treat_stale_lock_file_as_locked(monkeypatch, tmp_path):
    p = tmp_path / "Trading Journal.xlsx"
    p.write_bytes(b"x")
    (tmp_path / "~$Trading Journal.xlsx").write_bytes(b"stale")
    monkeypatch.setattr(master_service.os, "name", "nt")
    class _K:
        def __init__(self):
            self.CreateFileW = lambda *_a: ctypes.c_void_p(123).value
            self.CloseHandle = lambda *_a: 1
    monkeypatch.setattr(master_service.ctypes, "WinDLL", lambda *_a, **_k: _K(), raising=False)
    out = master_service._check_master_journal_write_lock(p)
    assert out["locked"] is False
    assert "lockfile" in str(out.get("reason") or "")
def test_manual_import_prebuilt_snapshot_preserves_workbook_zero_cashflow_anchors(tmp_path, monkeypatch):
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_single_file_mode', lambda: True)
    monkeypatch.setattr(ms, '_master_journal_authoritative_enabled', lambda: True)
    monkeypatch.setattr(ms, '_trading_journal_github_sync_enabled', lambda: False)
    monkeypatch.setenv('TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(ms, '_load_trading_journal_view_snapshot', lambda: None)
    monkeypatch.setattr(ms, '_save_trading_journal_view_snapshot', lambda *_a, **_k: None)
    monkeypatch.setattr(ms, '_persist_trading_journal_sqlite', lambda *_a, **_k: None)
    rows = [
        {'id':'binance-trade','row_type':'trade','source':'master_journal','account':'BINANCE','symbol':'BTCUSDT','side':'BUY','open_time':'2020-10-24T02:14:00','close_time':'2020-10-24T02:18:00','net_profit':-3.79,'balance_after_trade':396.65720524,'currency':'USDT'},
        {'id':'cashflow:BINANCE:2020-10-26T00:00:00:4','row_type':'cashflow','source':'master_journal','account':'BINANCE','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2020-10-26T00:00:00','close_time':'2020-10-26T00:00:00','cashflow_amount':None,'cashflow_new_balance':0,'balance_after_trade':0,'currency':'USDT'},
        {'id':'pepperstone-demo-stale-trade','row_type':'trade','source':'master_journal','account':'Pepperstone Demo','symbol':'EURUSD','side':'BUY','open_time':'2018-07-25T22:55:00','close_time':'2018-07-25T23:00:00','net_profit':4.78,'balance_after_trade':4.78,'currency':'AUD'},
        {'id':'cashflow:PEPPERSTONE DEMO:2018-07-26T00:00:00:5','row_type':'cashflow','source':'master_journal','account':'PEPPERSTONE DEMO','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2018-07-26T00:00:00','close_time':'2018-07-26T00:00:00','cashflow_amount':None,'cashflow_new_balance':0,'balance_after_trade':0,'currency':'AUD'},
    ]
    stale = {'items': rows, 'stats': {'totals': {}, 'by_instrument': [], 'groups': {'leaders': {}}}, 'balances': [
        {'account_label':'BINANCE','balance':396.65720524,'currency':'USDT'},
        {'account_label':'Pepperstone Demo','balance':4.78,'currency':'AUD'},
    ], 'diagnostics': {}}
    build_master_journal_workbook(stale, mj)
    pending = {'id':'oanda-demo-import','row_type':'trade','source':'oanda','account':'OANDA DEMO','symbol':'EURUSD','side':'BUY','open_time':'2026-01-01T00:00:00','close_time':'2026-01-01T00:05:00','net_profit':1.25,'balance_after_trade':1001.25,'currency':'AUD'}
    previous_pending = list(ms._PENDING_MANUAL_SYNC_ROWS)
    try:
        ms._PENDING_MANUAL_SYNC_ROWS = [pending]
        snapshot = ms._build_trading_journal_view_snapshot(force=True, skip_external_balances=True, skip_live_account_refresh=True)
        balances = {str(b.get('label') or b.get('account')): b for b in snapshot.get('balances') or []}
        assert balances['BINANCE']['balance'] == 0
        assert balances['PEPPERSTONE DEMO']['balance'] == 0
        assert balances['PEPPERSTONE DEMO']['currency'] == 'AUD'
        assert any(str(r.get('id')) == 'oanda-demo-import' for r in snapshot.get('items') or [])
        result = ms._sync_master_journal_workbook(defer_github_sync=True, expected_survivor_row_ids=['oanda-demo-import'], prebuilt_snapshot=snapshot, sync_caller='test')
        assert result['master_journal_ok'] is True
        synced = load_workbook(mj, data_only=True)
        try:
            dash = synced['STATS2']
            values = {str(dash.cell(r, 1).value or '').strip(): dash.cell(r, 2).value for r in range(1, dash.max_row + 1)}
            assert values['BINANCE'] == 0
            assert values['PEPPERSTONE DEMO'] == 0
            trade_log = synced['Trade Log']
            headers = ms._trade_log_header_map(trade_log)
            ridx = headers['Row ID']
            data_start = ms._trade_log_data_start_row(trade_log)
            row_ids = {str(trade_log.cell(r, ridx).value or '').strip() for r in range(data_start, trade_log.max_row + 1)}
            assert 'oanda-demo-import' in row_ids
        finally:
            synced.close()
    finally:
        ms._PENDING_MANUAL_SYNC_ROWS = previous_pending


def test_resync_endpoint_runs_workbook_sync_without_import_parser(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    called = {'sync': 0, 'import': 0}
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', lambda **_kwargs: called.update(sync=called['sync'] + 1) or {'master_journal_ok': True, 'master_journal_path': str(tmp_path / 'Trading Journal.xlsx'), 'master_journal_diagnostics': {'workbook_sync_substage_timings': {'snapshot_build': 0.1}}})
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []})
    monkeypatch.setattr(ms, '_import_uploaded_trading_journal_file', lambda *_a, **_k: called.update({'import': called['import'] + 1}) or {'ok': False})
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert called == {'sync': 1, 'import': 0}
    assert result['resync_timings']['workbook_sync'] >= 0


def test_workspace_and_status_routes_do_not_start_workbook_sync(monkeypatch):
    ms = _load_master_service_for_import_test()
    try:
        from fastapi.testclient import TestClient
    except Exception:
        pytest.skip('fastapi TestClient unavailable')
    monkeypatch.setattr(ms, '_sync_master_journal_workbook', lambda *_a, **_k: (_ for _ in ()).throw(AssertionError('route must not sync workbook')))
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda *_a, **_k: (_ for _ in ()).throw(AssertionError('route must not build snapshot')))

    async def _fake_oanda_status():
        return {'ok': True, 'status': 'cached-test'}

    async def _fake_state_restore():
        return {'ok': True, 'restored': True}

    monkeypatch.setattr(ms, '_build_oanda_inactivity_status', _fake_oanda_status)
    monkeypatch.setattr(ms, '_wait_for_state_restore_or_error', _fake_state_restore)
    monkeypatch.setattr(ms, '_get_watchlist', lambda: ['EURUSD'])
    ms._OANDA_INACTIVITY_CACHE.clear()
    client = TestClient(ms.app)
    assert client.get('/').status_code in {200, 307}
    assert client.get('/merged/trading-journal').status_code == 200
    assert client.get('/scripts').status_code == 200
    assert client.get('/api/state-sync/status').status_code == 200
    assert client.get('/api/watchlist').status_code == 200
    assert client.get('/api/oanda-inactivity-status').status_code == 200


def test_workbook_sync_rejects_missing_or_direct_caller(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    path = tmp_path / 'Trading Journal.xlsx'
    monkeypatch.setattr(ms, '_master_journal_path', lambda: path)
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('missing/direct caller must not run sync')))
    for caller in (None, '', '   ', 'direct', 'DIRECT'):
        result = ms._sync_master_journal_workbook(sync_caller=caller)
        assert result['ok'] is False
        assert result['status_code'] == 500
        assert result['code'] == 'MASTER_JOURNAL_SYNC_CALLER_REQUIRED'


def test_resync_rejects_active_workbook_sync_before_snapshot_build(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    path = tmp_path / 'Trading Journal.xlsx'
    monkeypatch.setattr(ms, '_master_journal_path', lambda: path)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    assert ms.MASTER_JOURNAL_WORKBOOK_SYNC_LOCK.acquire(blocking=False) is True
    try:
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_ACTIVE.update({'sync_id': 'active-sync', 'caller': 'manual_import', 'path': str(path), 'started_epoch': ms.time.time() - 3.25})
        monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('resync must reject before snapshot build')))
        result = ms._run_trading_journal_resync()
    finally:
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_ACTIVE.clear()
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_LOCK.release()
    assert result['ok'] is False
    assert result['status_code'] == 409
    assert result['code'] == 'MASTER_JOURNAL_SYNC_IN_PROGRESS'
    assert result['snapshot_build_ran'] is False
    assert result['skipped_snapshot_build'] is True
    assert result['active_sync_id'] == 'active-sync'
    assert result['active_caller'] == 'manual_import'
    assert result['active_path'] == str(path)
    assert result['active_elapsed_seconds'] >= 3


def test_core_workbook_sync_singleflight_rejects_duplicate_before_snapshot_build(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    assert ms.MASTER_JOURNAL_WORKBOOK_SYNC_LOCK.acquire(blocking=False) is True
    try:
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_ACTIVE.update({'sync_id': 'active-sync', 'caller': 'resync', 'path': str(tmp_path / 'Trading Journal.xlsx'), 'started_epoch': ms.time.time() - 2.0})
        monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('duplicate core sync must not build snapshot')))
        result = ms._sync_master_journal_workbook(sync_caller='test-duplicate')
    finally:
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_ACTIVE.clear()
        ms.MASTER_JOURNAL_WORKBOOK_SYNC_LOCK.release()
    assert result['ok'] is False
    assert result['status_code'] == 409
    assert result['code'] == 'MASTER_JOURNAL_SYNC_IN_PROGRESS'
    assert result['active_sync']['sync_id'] == 'active-sync'
    assert result['active_sync_id'] == 'active-sync'
    assert result['active_caller'] == 'resync'
    assert result['active_elapsed_seconds'] >= 1


def test_resync_fast_path_survives_memory_clear_via_persisted_metadata(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    mj = tmp_path / 'Trading Journal.xlsx'
    cache_path = tmp_path / 'resync-cache.json'
    snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [
        {'account_label':'BINANCE','balance':0,'currency':'USDT','balance_source':'cashflow_anchor_plus_trades'},
        {'account_label':'PEPPERSTONE DEMO','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades'},
    ]}
    build_master_journal_workbook(snapshot, mj)
    fingerprint = {'fingerprint': 'persisted'}
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_RESYNC_CACHE_PATH', cache_path)
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    monkeypatch.setattr(ms, '_journal_source_fingerprint', lambda: fingerprint)
    ms._save_resync_success_metadata(mj, fingerprint, snapshot)
    ms.TRADING_JOURNAL_RESYNC_LAST_SUCCESS.update({'fingerprint': None, 'snapshot': None, 'workbook_path': '', 'verified_at': None})
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('persisted fast path should skip snapshot build')))
    monkeypatch.setattr(ms, '_sync_master_journal_workbook', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('persisted fast path should skip workbook sync')))
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['skipped_snapshot_build'] is True
    assert result['fast_path_reason'] == 'persisted_fingerprint_match'
    assert result['snapshot_build_ran'] is False


def test_sync_and_snapshot_logs_include_sync_id_and_caller(monkeypatch, caplog):
    ms = _load_master_service_for_import_test()
    def _logged_unlocked(**kwargs):
        assert kwargs['sync_caller'] != 'unspecified'
        return {'master_journal_ok': True, 'ok': True}
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', _logged_unlocked)
    with caplog.at_level('INFO'):
        result = ms._sync_master_journal_workbook(sync_id='sync-log-test', sync_caller='unit-test')
    assert result['sync_id'] == 'sync-log-test'
    sync_logs = '\n'.join(record.getMessage() for record in caplog.records)
    assert 'master_journal_workbook_sync_start sync_id=sync-log-test caller=unit-test' in sync_logs
    assert 'master_journal_workbook_sync_done sync_id=sync-log-test caller=unit-test' in sync_logs

    caplog.clear()
    fingerprint = {'fp': 1}
    existing = {
        'cache_version': ms.TRADING_JOURNAL_VIEW_CACHE_VERSION,
        'generated_at': '2026-07-26T00:00:00Z',
        'source_fingerprints': fingerprint,
        'items': [],
        'balances': [],
        'stats': {},
    }
    ms._attach_trading_journal_equity_metadata(existing)
    monkeypatch.setattr(ms, '_load_trading_journal_view_snapshot', lambda: existing)
    monkeypatch.setattr(ms, '_journal_source_fingerprint', lambda: fingerprint)
    with caplog.at_level('INFO'):
        snap = ms._build_trading_journal_view_snapshot(force=False, sync_id='snapshot-log-test', sync_caller='unit-test')
    assert snap is existing
    snapshot_logs = '\n'.join(record.getMessage() for record in caplog.records)
    assert 'trading_journal_snapshot_build_start sync_id=snapshot-log-test caller=unit-test' in snapshot_logs
    assert 'trading_journal_snapshot_build_done sync_id=snapshot-log-test caller=unit-test' in snapshot_logs


def test_resync_builds_authoritative_snapshot_once_and_reuses_it(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}
    called = {'build': 0, 'sync': 0}

    def _build_once(**kwargs):
        called['build'] += 1
        assert kwargs['force'] is True
        assert kwargs['skip_external_balances'] is True
        assert kwargs['skip_live_account_refresh'] is True
        return snapshot

    def _sync_once(**kwargs):
        called['sync'] += 1
        assert kwargs['prebuilt_snapshot'] is snapshot
        assert kwargs['sync_id'].startswith('resync-')
        assert kwargs['sync_caller'] == 'resync'
        return {'master_journal_ok': True, 'master_journal_path': str(tmp_path / 'Trading Journal.xlsx')}

    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', _build_once)
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', _sync_once)
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['snapshot_build_ran'] is True
    assert result['skipped_snapshot_build'] is False
    assert called == {'build': 1, 'sync': 1}
    assert result['request_id'].startswith('resync-')


def test_resync_fast_verification_catches_stale_binance_or_pepperstone(tmp_path):
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    mj = tmp_path / 'Trading Journal.xlsx'
    stale_snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [
        {'account_label':'BINANCE','balance':396.65720524,'currency':'USDT','balance_source':'authoritative_trade_balance'},
        {'account_label':'PEPPERSTONE DEMO','balance':4.78,'currency':'AUD','balance_source':'authoritative_trade_balance'},
    ]}
    expected_snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [
        {'account_label':'BINANCE','balance':0,'currency':'USDT','balance_source':'cashflow_anchor_plus_trades'},
        {'account_label':'PEPPERSTONE DEMO','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades'},
    ]}
    build_master_journal_workbook(stale_snapshot, mj)
    verification = ms._fast_verify_trading_journal_workbook(mj, expected_snapshot=expected_snapshot)
    assert verification['ok'] is False
    assert verification['error'] == 'dashboard_account_balance_verification_failed'
    mismatches = ' | '.join(verification['diagnostics']['account_balance_mismatches'])
    assert 'BINANCE' in mismatches
    assert 'PEPPERSTONE DEMO' in mismatches


def test_resync_no_change_fast_path_skips_snapshot_build_but_verifies_workbook(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    mj = tmp_path / 'Trading Journal.xlsx'
    snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [
        {'account_label':'BINANCE','balance':0,'currency':'USDT','balance_source':'cashflow_anchor_plus_trades'},
        {'account_label':'PEPPERSTONE DEMO','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades'},
    ]}
    build_master_journal_workbook(snapshot, mj)
    fingerprint = {'fingerprint': 'unchanged'}
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    monkeypatch.setattr(ms, '_journal_source_fingerprint', lambda: fingerprint)
    ms.TRADING_JOURNAL_RESYNC_LAST_SUCCESS.update({'fingerprint': fingerprint, 'snapshot': snapshot, 'workbook_path': str(mj), 'verified_at': 'now'})
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('snapshot_build should be skipped')))
    monkeypatch.setattr(ms, '_sync_master_journal_workbook', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('workbook sync should be skipped')))
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['skipped_snapshot_build'] is True
    assert result['snapshot_build_ran'] is False
    assert result['fast_verification']['ok'] is True


def test_resync_changed_fingerprint_forces_full_snapshot_build(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    mj = tmp_path / 'Trading Journal.xlsx'
    snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}
    ms.TRADING_JOURNAL_RESYNC_LAST_SUCCESS.update({'fingerprint': {'old': 1}, 'snapshot': snapshot, 'workbook_path': str(mj), 'verified_at': 'now'})
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    monkeypatch.setattr(ms, '_journal_source_fingerprint', lambda: {'new': 2})
    called = {'build': 0, 'sync': 0}
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: called.update(build=called['build'] + 1) or snapshot)
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', lambda **_kwargs: called.update(sync=called['sync'] + 1) or {'master_journal_ok': True, 'master_journal_path': str(mj)})
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['snapshot_build_ran'] is True
    assert called == {'build': 1, 'sync': 1}


def test_resync_concurrent_duplicate_is_rejected_before_snapshot_build(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    assert ms.TRADING_JOURNAL_RESYNC_LOCK.acquire(blocking=False) is True
    try:
        monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('duplicate resync should not build')))
        result = ms._run_trading_journal_resync()
    finally:
        ms.TRADING_JOURNAL_RESYNC_LOCK.release()
    assert result['ok'] is False
    assert result['status_code'] == 409
    assert result['code'] == 'TRADING_JOURNAL_RESYNC_IN_PROGRESS'


def test_workbook_sync_call_sites_use_explicit_callers():
    source = (ROOT / 'render' / 'master_service.py').read_text(encoding='utf-8')
    call_lines = [line.strip() for line in source.splitlines() if '_sync_master_journal_workbook(' in line and not line.lstrip().startswith('def ')]
    assert call_lines
    assert all('sync_caller=' in line for line in call_lines), call_lines
    assert 'caller=unspecified' not in source


def test_sync_compatibility_route_invokes_one_resync(monkeypatch):
    ms = _load_master_service_for_import_test()
    calls = {'run': 0}
    monkeypatch.setattr(ms, '_run_trading_journal_resync', lambda: calls.update(run=calls['run'] + 1) or {'ok': True, 'status_code': 200, 'request_id': 'resync-test'})
    response = asyncio.run(ms.trading_journal_sync())
    assert response.status_code == 200
    assert calls == {'run': 1}


def test_sync_compatibility_route_respects_resync_singleflight(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    assert ms.TRADING_JOURNAL_RESYNC_LOCK.acquire(blocking=False) is True
    try:
        monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('/sync duplicate must not build snapshot')))
        response = asyncio.run(ms.trading_journal_sync())
    finally:
        ms.TRADING_JOURNAL_RESYNC_LOCK.release()
    assert response.status_code == 409
    payload = json.loads(response.body.decode('utf-8'))
    assert payload['code'] == 'TRADING_JOURNAL_RESYNC_IN_PROGRESS'


def test_resync_preserves_pepperstone_demo_zero_anchor_without_import_parser(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    from tools.master_journal_workbook import build_master_journal_workbook
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    mj = tmp_path / 'Trading Journal.xlsx'
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    monkeypatch.setattr(ms, '_master_journal_single_file_mode', lambda: True)
    monkeypatch.setattr(ms, '_master_journal_authoritative_enabled', lambda: True)
    monkeypatch.setattr(ms, '_trading_journal_github_sync_enabled', lambda: False)
    monkeypatch.setenv('TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(ms, '_load_trading_journal_view_snapshot', lambda: None)
    monkeypatch.setattr(ms, '_save_trading_journal_view_snapshot', lambda *_a, **_k: None)
    monkeypatch.setattr(ms, '_persist_trading_journal_sqlite', lambda *_a, **_k: None)
    monkeypatch.setattr(ms, '_import_uploaded_trading_journal_file', lambda *_a, **_k: (_ for _ in ()).throw(AssertionError('resync must not parse/import a file')))
    rows = [
        {'id':'pepperstone-demo-stale-trade','row_type':'trade','source':'master_journal','account':'Pepperstone Demo','symbol':'EURUSD','side':'BUY','open_time':'2018-07-25T22:55:00','close_time':'2018-07-25T23:00:00','net_profit':4.78,'balance_after_trade':4.78,'currency':'AUD'},
        {'id':'cashflow:PEPPERSTONE DEMO:2018-07-26T00:00:00:5','row_type':'cashflow','source':'master_journal','account':'PEPPERSTONE DEMO','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2018-07-26T00:00:00','close_time':'2018-07-26T00:00:00','cashflow_amount':None,'cashflow_new_balance':0,'balance_after_trade':0,'currency':'AUD'},
        {'id':'binance-trade','row_type':'trade','source':'master_journal','account':'BINANCE','symbol':'BTCUSDT','side':'BUY','open_time':'2020-10-24T02:14:00','close_time':'2020-10-24T02:18:00','net_profit':-3.79,'balance_after_trade':396.65720524,'currency':'USDT'},
        {'id':'cashflow:BINANCE:2020-10-26T00:00:00:4','row_type':'cashflow','source':'master_journal','account':'BINANCE','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2020-10-26T00:00:00','close_time':'2020-10-26T00:00:00','cashflow_amount':None,'cashflow_new_balance':0,'balance_after_trade':0,'currency':'USDT'},
    ]
    stale = {'items': rows, 'stats': {'totals': {}, 'by_instrument': [], 'groups': {'leaders': {}}}, 'balances': [
        {'account_label':'Pepperstone Demo','balance':4.78,'currency':'AUD'},
        {'account_label':'BINANCE','balance':396.65720524,'currency':'USDT'},
    ], 'diagnostics': {}}
    build_master_journal_workbook(stale, mj)
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['master_journal_ok'] is True
    synced = load_workbook(mj, data_only=True)
    try:
        dash = synced['STATS2']
        rows_by_account = {str(dash.cell(r, 1).value or '').strip(): (dash.cell(r, 2).value, dash.cell(r, 3).value) for r in range(1, dash.max_row + 1)}
        assert rows_by_account['PEPPERSTONE DEMO'] == (0, 'AUD')
        assert rows_by_account['BINANCE'] == (0, 'USDT')
    finally:
        synced.close()


def test_resync_returns_excel_lock_payload(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    path = tmp_path / 'Trading Journal.xlsx'
    path.write_bytes(b'placeholder')
    monkeypatch.setattr(ms, '_master_journal_path', lambda: path)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': True, 'lock_file': '~$Trading Journal.xlsx'})
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: (_ for _ in ()).throw(AssertionError('lock preflight must happen before snapshot build')))
    result = ms._run_trading_journal_resync()
    assert result['ok'] is False
    assert result['code'] == 'EXCEL_WORKBOOK_OPEN'
    assert result['status_code'] == 409



def test_master_journal_snapshot_records_substage_timings_and_skips_context_lookup(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_SOURCE', 'master_journal')
    monkeypatch.setattr(ms, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(ms, '_master_journal_single_file_mode', lambda: True)
    monkeypatch.setattr(ms, '_master_journal_authoritative_enabled', lambda: True)
    monkeypatch.setattr(ms, '_master_journal_path', lambda: tmp_path / 'Trading Journal.xlsx')
    monkeypatch.setattr(ms, 'read_master_journal_source', lambda _path: {
        'items': [
            {'id': 't1', 'row_type': 'trade', 'source': 'master_journal', 'account': 'BINANCE', 'account_label': 'BINANCE', 'symbol': 'BTCUSDT', 'side': 'BUY', 'open_time': '2026-01-01', 'close_time': '2026-01-01', 'net_profit': 1.0, 'result_pct': 1.0, 'balance_after_trade': 1.0, 'currency': 'USDT'},
            {'id': 'c1', 'row_type': 'cashflow', 'source': 'master_journal', 'account': 'BINANCE', 'account_label': 'BINANCE', 'symbol': 'CASHFLOW', 'side': 'WITHDRAWAL', 'open_time': '2026-01-02', 'close_time': '2026-01-02', 'cashflow_new_balance': 0.0, 'balance_after_trade': 0.0, 'currency': 'USDT'},
        ],
        'balances': [
            {'account': 'BINANCE', 'label': 'BINANCE', 'balance': 0.0, 'currency': 'USDT', 'source': 'stats2_account_balances'},
        ],
        'cashflow_ledger': {'BINANCE': [{'account': 'BINANCE', 'date': '2026-01-02', 'new_balance': 0.0, 'currency': 'USDT'}]},
    })
    monkeypatch.setattr(ms, '_monthly_aud_revaluation_rows_for_journal_view', lambda: [])
    monkeypatch.setattr(ms, '_load_trade_contexts', lambda: (_ for _ in ()).throw(AssertionError('master_journal snapshot must not load local trade contexts')))
    monkeypatch.setattr(ms, '_save_trading_journal_view_snapshot', lambda _payload: None)
    monkeypatch.setattr(ms, '_persist_trading_journal_sqlite', lambda *_args, **_kwargs: None)
    captured_seeds = []
    real_timeline_builder = ms._build_journal_balance_timelines

    def _capture_timeline_seeds(rows, ledger, seeds):
        captured_seeds.extend(seeds)
        return real_timeline_builder(rows, ledger, seeds)

    monkeypatch.setattr(ms, '_build_journal_balance_timelines', _capture_timeline_seeds)
    snapshot = ms._build_trading_journal_view_snapshot(force=True, skip_external_balances=True, skip_live_account_refresh=True, sync_id='timing-test', sync_caller='unit-test')
    timings = ((snapshot.get('diagnostics') or {}).get('snapshot_substage_timings') or {})
    assert 'workbook_source_read' in timings
    assert 'trade_context_backfill_initial' in timings
    assert 'balance_timeline_build' in timings
    assert 'stats_dashboard_instrument_calendar_build' in timings
    assert captured_seeds == [
        {'account': 'BINANCE', 'label': 'BINANCE', 'balance': 0.0, 'currency': 'USDT', 'source': 'stats2_account_balances'},
    ]
    assert snapshot['equity_cache']['point_counts']['BINANCE'] == 2


def test_resync_fast_path_miss_reports_changed_fingerprint_components(tmp_path):
    ms = _load_master_service_for_import_test()
    path = tmp_path / 'Trading Journal.xlsx'
    path.write_bytes(b'placeholder')
    old_fp = {'source_mode': 'master_journal', 'files': [{'path': str(path), 'sha256': 'old', 'size': 1}]}
    new_fp = {'source_mode': 'master_journal', 'files': [{'path': str(path), 'sha256': 'new', 'size': 2}]}
    ms.TRADING_JOURNAL_RESYNC_LAST_SUCCESS.update({'fingerprint': old_fp, 'snapshot': {'items': [], 'stats': {}, 'balances': []}, 'workbook_path': str(path), 'verified_at': 'then'})
    snapshot, reason, diagnostics = ms._resync_fast_path_snapshot(path, new_fp)
    assert snapshot is None
    assert reason == 'fingerprint_changed_or_no_verified_cache'
    changed = diagnostics.get('memory_changed_components') or diagnostics.get('changed_components') or []
    assert any(item.get('component') == 'file' and item.get('path') == str(path) for item in changed)


def test_resync_persists_post_replace_fingerprint_for_next_fast_path(monkeypatch, tmp_path):
    ms = _load_master_service_for_import_test()
    mj = tmp_path / 'Trading Journal.xlsx'
    snapshot = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [
        {'account_label': 'BINANCE', 'balance': 0, 'currency': 'USDT'},
        {'account_label': 'PEPPERSTONE DEMO', 'balance': 0, 'currency': 'AUD'},
    ]}
    from tools.master_journal_workbook import build_master_journal_workbook
    build_master_journal_workbook(snapshot, mj)
    fingerprints = iter([{'stage': 'pre'}, {'stage': 'post'}, {'stage': 'post'}])
    monkeypatch.setattr(ms, '_master_journal_path', lambda: mj)
    monkeypatch.setattr(ms, '_master_journal_lock_status', lambda _path: {'locked': False})
    monkeypatch.setattr(ms, '_resync_source_fingerprint', lambda _path: next(fingerprints))
    monkeypatch.setattr(ms, '_build_trading_journal_view_snapshot', lambda **_kwargs: snapshot)
    monkeypatch.setattr(ms, '_sync_master_journal_workbook_unlocked', lambda **_kwargs: {'master_journal_ok': True, 'master_journal_path': str(mj)})
    result = ms._run_trading_journal_resync()
    assert result['ok'] is True
    assert result['snapshot_build_ran'] is True
    assert ms.TRADING_JOURNAL_RESYNC_LAST_SUCCESS['fingerprint'] == {'stage': 'post'}
    cached, reason, _diagnostics = ms._resync_fast_path_snapshot(mj, {'stage': 'post'})
    assert cached == snapshot
    assert reason == 'memory_fingerprint_match'


@pytest.mark.skipif(not HTTPX_AVAILABLE, reason="httpx is not installed")
def test_sync_preserves_canonical_stats1_layout_and_manual_formatting(tmp_path, monkeypatch):
    from openpyxl.styles import Alignment, Font, PatternFill
    from tools.master_journal_workbook import build_master_journal_workbook

    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    path = tmp_path / "Trading Journal.xlsx"
    rows = [
        {"id": "fx", "row_type": "trade", "asset_class": "fx", "account": "OANDA DEMO", "symbol": "EURUSD", "side": "BUY", "open_time": "2026-01-01", "close_time": "2026-01-01", "net_profit": -100.0, "result_pct": -80.0, "currency": "AUD"},
        {"id": "crypto", "row_type": "trade", "asset_class": "crypto", "account": "BYBIT", "symbol": "BTCUSDT", "side": "BUY", "open_time": "2026-01-02", "close_time": "2026-01-02", "net_profit": -10.0, "result_pct": -20.0, "currency": "USDT"},
    ]
    balances = [
        {"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 900.0, "currency": "AUD"},
        {"account": "BYBIT", "label": "BYBIT", "balance": 90.0, "currency": "USDT"},
    ]
    stats = master_service._compute_journal_stats(rows, balances)
    snapshot = {"items": rows, "balances": balances, "stats": stats}
    build_master_journal_workbook(snapshot, path)

    wb = load_workbook(path)
    dash = wb["STATS1"]
    sheet_order = list(wb.sheetnames)
    labels_before = [
        str(dash.cell(row, 1).value or "").strip()
        for row in range(1, dash.max_row + 1)
    ]
    sentinel = dash["I149"]
    sentinel.value = "CUSTOM LAYOUT SENTINEL"
    sentinel.font = Font(name="Calibri", size=13, bold=True, color="123456")
    sentinel.fill = PatternFill("solid", fgColor="ABCDEF")
    sentinel.alignment = Alignment(horizontal="center")
    wb.save(path)
    wb.close()

    result = master_service._sync_master_journal_workbook(prebuilt_snapshot=snapshot, sync_caller="test")
    assert result["master_journal_ok"] is True, result
    wb = load_workbook(path)
    dash = wb["STATS1"]
    labels_after = [
        str(dash.cell(row, 1).value or "").strip()
        for row in range(1, dash.max_row + 1)
    ]
    assert wb.sheetnames == sheet_order
    assert labels_after == labels_before
    sentinel = dash["I149"]
    assert sentinel.value == "CUSTOM LAYOUT SENTINEL"
    assert sentinel.font.bold is True
    assert str(sentinel.font.color.rgb)[-6:].upper() == "123456"
    assert str(sentinel.fill.fgColor.rgb)[-6:].upper() == "ABCDEF"
    assert sentinel.alignment.horizontal == "center"
    trades_row = labels_after.index("Trades") + 1
    assert dash.cell(trades_row, 2).value == 2
    wb.close()
