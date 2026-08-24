import asyncio
import importlib.util
import json
import shutil
import subprocess
import sys

import re
from pathlib import Path
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from tools.master_journal_workbook import TRADE_LOG_DATA_ROW_HEIGHT, TRADE_LOG_DATA_START_ROW, TRADE_LOG_HEADERS, TRADE_LOG_FILTER_HEADER_ROW

ROOT = Path(__file__).resolve().parents[1]
MASTER_SERVICE_PATH = ROOT / 'render' / 'master_service.py'


def _repo_tmp_dir(name: str) -> Path:
    path = ROOT / f".pytest_tmp_{name}"
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True)
    return path


def _extract_html_template(source: str) -> str:
    match = re.search(r'HTML_TEMPLATE\s*=\s*"""(.*?)"""\n\nINSTRUMENT_SPECS_TEMPLATE', source, re.S)
    assert match, 'HTML_TEMPLATE block not found'
    return match.group(1)


def test_dashboard_home_removes_instrument_specs_recent_trades_open_orders() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    html = _extract_html_template(source)

    assert 'id="instrument-specs-widget"' not in html
    assert 'id="recent-trades-panel"' not in html
    assert 'id="open-orders-panel"' not in html

    assert 'script-toolbar-title' not in html
    assert '>Scripts<' not in html
    assert 'Watchlist' in html
    assert 'OANDA Inactivity' in html
    assert 'Orders / Positions' in html
    assert 'id="dashboard-workspace"' in html
    assert 'id="dashboard-workspace-title"' in html
    assert 'id="dashboard-workspace-status"' in html
    assert 'id="dashboard-workspace-empty"' in html
    assert 'id="dashboard-workspace-frame"' in html
    assert 'src="/merged/open-orders?_dashboard=1"' in html
    assert 'class="dashboard-main-content"' in html
    assert 'id="pine-scripts-panel"' in html
    assert 'id="pine-files"' in html
    assert 'id="pine-fallback"' in html
    assert 'id="dashboard-instrument-lookup-panel"' not in html
    assert 'id="dashboard-instrument-lookup-form"' not in html
    assert 'id="dashboard-instrument-lookup-input"' not in html
    assert 'id="dashboard-history-panel"' not in html
    assert '{{HISTORY_EXPORT_TOOL}}' not in html
    assert '{{HISTORY_PAGE_SCRIPT_TAG}}' not in html
    for equity_id in (
        "journal-equity-panel",
        "journal-equity-account",
        "journal-equity-refresh-btn",
        "journal-equity-summary",
        "journal-equity-canvas",
        "journal-equity-overlay-canvas",
        "journal-equity-state",
        "journal-equity-hover-live",
    ):
        assert f'id="{equity_id}"' in html
    assert '{{TRADING_JOURNAL_EQUITY_SCRIPT_TAG}}' in html
    assert 'Select a script from the toolbar above to load it here.' not in html
    assert 'Select a script from the left to load it here.' not in html
    assert '.local-exit-btn' in html
    assert 'class="panel dashboard-script-panel" id="dashboard-scripts-panel"' in html
    assert 'class="script-toolbar-row"' not in html
    assert 'id="scripts-grid" class="script-stack"' in html
    assert 'id="exit-button-slot" class="exit-button-slot"' in html
    assert 'id="exit-panel"' not in html

    assert 'id="journal-sync-widget"' not in html
    assert 'id="sync-journal-btn"' not in html
    assert 'id="sync-journal-status"' not in html

    watchlist_idx = html.find('Watchlist')
    oanda_idx = html.find('OANDA Inactivity')
    assert watchlist_idx != -1 and oanda_idx != -1
    assert watchlist_idx < oanda_idx

    scripts_grid_idx = html.find('id="scripts-grid"')
    exit_slot_idx = html.find('id="exit-button-slot"')
    watchlist_widget_idx = html.find('id="watchlist-widget"')
    oanda_widget_idx = html.find('id="oanda-inactivity-widget"')
    workspace_idx = html.find('id="dashboard-workspace"')
    equity_idx = html.find('id="journal-equity-panel"')
    pine_idx = html.find('id="pine-scripts-panel"')
    assert scripts_grid_idx != -1 and exit_slot_idx != -1
    assert watchlist_widget_idx != -1 and oanda_widget_idx != -1 and workspace_idx != -1 and equity_idx != -1 and pine_idx != -1
    assert scripts_grid_idx < watchlist_widget_idx
    assert scripts_grid_idx < workspace_idx
    assert watchlist_widget_idx < oanda_widget_idx
    assert workspace_idx < equity_idx < pine_idx

    rail_start = html.find('<div class="dashboard-rail">')
    workspace_start = html.find('<section class="panel" id="dashboard-workspace">')
    rail_html = html[rail_start:workspace_start]
    assert '<aside class="panel sidebar">' not in rail_html
    assert 'id="scripts-grid"' in rail_html


def test_dashboard_equity_toolbar_is_responsive_and_non_overlapping() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    html = _extract_html_template(source)

    assert '.journal-equity-toolbar{' in html
    assert 'align-items:flex-end;' in html
    assert 'flex-wrap:wrap;' in html
    assert '#journal-equity-refresh-btn{' in html
    assert '.equity-chart-wrap{' in html
    assert 'min-height:420px;' in html
    assert 'box-sizing:border-box;' in html
    assert 'min-width:0;' in html


def test_dashboard_script_css_keeps_scripts_vertical_above_watchlist() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    html = _extract_html_template(source)

    assert '.dashboard-script-panel' in html
    assert '.dashboard-script-toolbar' not in html
    assert '.script-toolbar-row' not in html
    assert '.script-toolbar-grid' not in html
    assert 'script-toolbar-title' not in html
    assert '>Scripts<' not in html
    assert 'flex-direction:column;' in html
    assert 'text-overflow: ellipsis;' in html
    assert 'white-space: nowrap;' in html
    assert '.script-toolbar-grid .script-btn[data-script-name="history"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="trading-journal"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="open-orders"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="instrument-lookup"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="spreads-clone"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="mt5"]' not in html
    assert '.script-toolbar-grid .script-btn[data-script-name="pine"]' not in html
    assert '.local-exit-btn{\n            margin-top: 0;' in html


def test_dashboard_orders_workspace_uses_content_height_sync() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    html = _extract_html_template(source)
    dashboard_js = (ROOT / "render" / "static" / "dashboard.js").read_text(encoding="utf-8")

    assert 'min-height: 780px' not in html
    assert 'height: calc(100vh - 7rem)' not in html
    assert 'min-height: 680px' not in html
    assert 'height: 70vh' not in html
    assert 'installWorkspaceHeightSync' in dashboard_js
    assert 'const frameWindow = doc.defaultView || workspaceFrame.contentWindow || window;' in dashboard_js
    assert 'const FrameResizeObserver = frameWindow.ResizeObserver || window.ResizeObserver;' in dashboard_js
    assert 'const FrameMutationObserver = frameWindow.MutationObserver || window.MutationObserver;' in dashboard_js
    assert 'observeNode(workspaceMutationObserver, doc.body' in dashboard_js
    assert "workspaceFrame?.addEventListener('load', installWorkspaceHeightSync);" in dashboard_js


def test_dashboard_home_loads_equity_script_without_inline_history_tool() -> None:
    module = _load_master_service_module()
    module.APP_PROFILE = "local"
    response = asyncio.run(module.home_page())
    body = response.body.decode("utf-8")

    assert 'id="dashboard-history-panel"' not in body
    assert 'id="history-export"' not in body
    assert '/static/history_page.js' not in body
    assert body.count('id="journal-equity-canvas"') == 1
    assert '/static/trading_journal_equity_curve.js?v=' in body


def test_dashboard_profile_layout_removes_local_columns_server_side() -> None:
    module = _load_master_service_module()

    module.APP_PROFILE = "local"
    local_body = asyncio.run(module.home_page()).body.decode("utf-8")
    assert 'class="layout"' in local_body
    assert 'class="dashboard-main-content"' in local_body
    assert 'id="watchlist-widget"' in local_body

    module.APP_PROFILE = "render"
    render_body = asyncio.run(module.home_page()).body.decode("utf-8")
    assert 'class="layout render-dashboard-layout"' in render_body
    assert 'class="dashboard-main-content"' not in render_body
    assert 'id="watchlist-widget"' not in render_body
    assert "LOCAL_DASHBOARD_" not in render_body

    source = MASTER_SERVICE_PATH.read_text(encoding="utf-8")
    html = _extract_html_template(source)
    assert ".layout.render-dashboard-layout{" in html
    assert "grid-template-columns: minmax(0, 720px);" in html
    assert "{{DASHBOARD_LAYOUT_CLASS}}" in html


def test_instrument_lookup_owns_specs_and_journal_markup() -> None:
    master_service_py = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    calculator_js = (ROOT / 'render' / 'static' / 'calculator.js').read_text(encoding='utf-8')
    lookup_js = (ROOT / 'render' / 'static' / 'instrument_lookup.js').read_text(encoding='utf-8')

    assert 'id="calc-instrument-specs"' not in master_service_py
    assert 'id="calc-journal-summary"' not in master_service_py
    assert 'Instrument Lookup' in master_service_py
    assert 'Journal Stats' in master_service_py
    instrument_block = master_service_py[master_service_py.index('INSTRUMENT_SPECS_TEMPLATE'):master_service_py.index('CALCULATOR_TEMPLATE')]
    assert 'id="asset-toggle"' not in instrument_block
    assert 'Download JPG' not in instrument_block
    assert 'Broker rules' not in instrument_block
    assert 'min-width: 1800px' not in master_service_py
    assert '"/instrument-lookup"' in master_service_py
    assert 'id="recent-trades-section"' in instrument_block
    assert 'id="trade-scroll-top"' in instrument_block
    assert 'id="trade-scroll-spacer"' in instrument_block
    assert 'id="trade-table-wrap"' in instrument_block
    assert instrument_block.index('id="trade-scroll-top"') < instrument_block.index('Recent Trades <span') < instrument_block.index('id="trade-table-wrap"')
    journal_panel_block = instrument_block[instrument_block.index('<h2>Journal Stats</h2>'):instrument_block.index('id="recent-trades-section"')]
    assert 'Recent Trades' not in journal_panel_block
    assert '.recent-trades-panel { margin-top:1rem; }' in instrument_block
    assert '/api/instrument-specs?query=' in lookup_js
    assert '/api/calculator/journal-summary?asset=' in lookup_js
    assert 'history.replaceState(null, \'\', `/instrument-lookup?q=' in lookup_js
    assert 'recentTradesMarkup' in lookup_js
    assert 'syncTradeScroll' in lookup_js
    assert 'tradeScrollSpacer.style.width' in lookup_js
    assert '&asset=' not in lookup_js
    assert "'openInterest'" in lookup_js
    assert "openInterestValue: 'Open interest value (USD)'" in lookup_js
    assert "keys: ['lastPrice', 'fundingRate', 'nextFundingTime', 'launchTime', 'openInterestValue'" in lookup_js
    assert 'Trading Rules' not in lookup_js
    assert 'Broker rules' not in lookup_js
    assert "'tickSize'" in lookup_js
    assert "'category', 'status', 'baseCoin', 'quoteCoin'" in lookup_js
    assert 'upgradeLegacyMarkup' in lookup_js
    assert 'instrument-lookup-runtime-css' in lookup_js
    assert 'flattenMetrics' not in lookup_js
    assert 'JSON.stringify(item)' not in lookup_js


def test_instrument_lookup_hides_rules_and_renders_journal_counts_neutrally() -> None:
    node = shutil.which("node")
    assert node, "node is required for instrument lookup JS behavior test"
    harness = r'''
const fs = require('fs');
const source = fs.readFileSync(process.argv[1], 'utf8');

class Element {
  constructor(id) {
    this.id = id;
    this.value = '';
    this.textContent = '';
    this.innerHTML = '';
    this.tagName = 'DIV';
    this.listeners = {};
    this.classList = { toggle: () => {}, add: () => {}, remove: () => {} };
    this.dataset = {};
    this.style = {};
    this.scrollLeft = 0;
    this.scrollWidth = id === 'trade-table' || id === 'trade-table-wrap' ? 1440 : 0;
    this.clientWidth = id === 'trade-table-wrap' || id === 'trade-scroll-top' ? 640 : 0;
  }
  addEventListener(event, callback) { this.listeners[event] = callback; }
  closest() { return null; }
  querySelectorAll() { return []; }
}

const elements = Object.fromEntries([
  'q', 'load', 'rows', 'err', 'journal-status', 'journal-metrics',
  'recent-trades-section', 'trade-scroll-top', 'trade-scroll-spacer', 'trade-table-wrap',
  'trade-table', 'trade-head', 'trade-body'
].map((id) => [id, new Element(id)]));
global.window = { location: { search: '?q=BTCUSDT' } };
global.history = { replaceState: () => {} };
global.document = {
  head: { appendChild: () => {} },
  createElement: (tag) => new Element(tag),
  getElementById: (id) => elements[id] || null,
};
global.fetch = async (url) => {
  const textUrl = String(url);
  if (textUrl.includes('/api/instrument-specs')) {
    return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({
      resolved_symbol: 'BTCUSDT',
      displayName: 'Bitcoin',
      category: 'linear',
      status: 'Trading',
      baseCoin: 'BTC',
      quoteCoin: 'USDT',
      type: 'perpetual',
      tickSize: '0.10',
      minOrderQty: '0.001',
      lastPrice: '100',
      'range.1m': 0.01
    }) };
  }
  if (textUrl.includes('/api/calculator/journal-summary')) {
    return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({
      status: 'ok',
      canonical_symbol: 'BTCUSDT',
      stats: { total_trades: 2, wins: 1, losses: 1, win_rate: '50.00%' },
      metrics: {
        trades: 2,
        wins: 1,
        losses: 1,
        win_rate_pct: 50,
        net_profit_total: 10,
        avg_stop_pct: -1.2,
        min_stop_pct: -2.5,
        max_stop_pct: 3.5,
        avg_target_pct: 4,
        min_target_pct: 1,
        max_target_pct: 8,
        net_r_multiple: 1.5,
        avg_r_multiple: 0.75,
        shortest_duration_seconds: 60,
        longest_duration_seconds: 120,
        avg_drawdown_pct: 12.3,
        max_drawdown_pct: 45.6,
        longest_winning_streak: { trade_count: 3 },
        longest_losing_streak: { trade_count: 2 }
      },
      trades: [{ symbol: 'BTCUSDT', close_time: '2026-01-01T00:00:00Z', side: 'Buy', net_profit: 10, result_pct: 1, currency: 'USDT' }]
    }) };
  }
  return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ resolved_symbol: 'BTCUSDT' }) };
};
eval(source);
(async () => {
  await Promise.resolve();
  await Promise.resolve();
  await new Promise((resolve) => setImmediate(resolve));
  elements['trade-scroll-top'].scrollLeft = 321;
  elements['trade-scroll-top'].listeners.scroll();
  const tableAfterTop = elements['trade-table-wrap'].scrollLeft;
  elements['trade-table-wrap'].scrollLeft = 123;
  elements['trade-table-wrap'].listeners.scroll();
  const topAfterTable = elements['trade-scroll-top'].scrollLeft;
  console.log(JSON.stringify({
    rows: elements.rows.innerHTML,
    journal: elements['journal-metrics'].innerHTML,
    tradeHead: elements['trade-head'].innerHTML,
    tradeBody: elements['trade-body'].innerHTML,
    spacerWidth: elements['trade-scroll-spacer'].style.width,
    tableAfterTop,
    topAfterTable
  }));
})();
'''
    completed = subprocess.run(
        [node, "-e", harness, str(ROOT / "render" / "static" / "instrument_lookup.js")],
        check=True,
        capture_output=True,
        text=True,
    )
    result = json.loads(completed.stdout.strip().splitlines()[-1])
    specs_html = result["rows"]
    journal_html = result["journal"]
    assert "Trading Rules" not in specs_html
    for hidden in ("Category", "Status", "Base", "Quote", "Tick size", "Min order qty"):
        assert hidden not in specs_html
    distances = journal_html[journal_html.index("Distances"):journal_html.index("Winners / Losers")]
    assert "positive" not in distances
    assert "negative" not in distances
    assert "[object Object]" not in journal_html
    assert "Avg drawdown" not in journal_html
    assert "Max drawdown" not in journal_html
    assert "Best win streak" in journal_html
    assert ">3<" in journal_html
    assert "Worst losing streak" in journal_html
    assert ">2<" in journal_html
    assert "Net P/L" in result["tradeHead"]
    assert "BTCUSDT" in result["tradeBody"] or "Buy" in result["tradeBody"]
    assert result["spacerWidth"] == "1440px"
    assert result["tableAfterTop"] == 321
    assert result["topAfterTable"] == 123


def test_instrument_lookup_top_scrollbar_updates_on_window_resize_without_resize_observer() -> None:
    node = shutil.which("node")
    assert node, "node is required for instrument lookup JS behavior test"
    harness = r'''
const fs = require('fs');
const source = fs.readFileSync(process.argv[1], 'utf8');
let resizeHandler = null;
let resizeRegistrations = 0;

class Element {
  constructor(id) {
    this.id = id;
    this.value = '';
    this.textContent = '';
    this.innerHTML = '';
    this.tagName = 'DIV';
    this.listeners = {};
    this.classList = { toggle: () => {}, add: () => {}, remove: () => {} };
    this.dataset = {};
    this.style = {};
    this.scrollLeft = 0;
    this.scrollWidth = id === 'trade-table' || id === 'trade-table-wrap' ? 960 : 0;
    this.clientWidth = id === 'trade-table-wrap' || id === 'trade-scroll-top' ? 420 : 0;
  }
  addEventListener(event, callback) { this.listeners[event] = callback; }
  closest() { return null; }
  querySelectorAll() { return []; }
}

const elements = Object.fromEntries([
  'q', 'load', 'rows', 'err', 'journal-status', 'journal-metrics',
  'recent-trades-section', 'trade-scroll-top', 'trade-scroll-spacer', 'trade-table-wrap',
  'trade-table', 'trade-head', 'trade-body'
].map((id) => [id, new Element(id)]));
global.window = {
  location: { search: '?q=BTCUSDT' },
  addEventListener: (event, callback) => {
    if (event === 'resize') {
      resizeHandler = callback;
      resizeRegistrations += 1;
    }
  },
  requestAnimationFrame: (callback) => setTimeout(callback, 0),
};
global.history = { replaceState: () => {} };
global.document = {
  head: { appendChild: () => {} },
  createElement: (tag) => new Element(tag),
  getElementById: (id) => elements[id] || null,
};
global.fetch = async (url) => {
  const textUrl = String(url);
  if (textUrl.includes('/api/instrument-specs')) {
    return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ resolved_symbol: 'BTCUSDT', lastPrice: '100' }) };
  }
  if (textUrl.includes('/api/calculator/journal-summary')) {
    return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({
      status: 'ok',
      canonical_symbol: 'BTCUSDT',
      stats: { total_trades: 1, wins: 1, losses: 0, win_rate: '100.00%' },
      metrics: { trades: 1, wins: 1, losses: 0, win_rate_pct: 100, net_profit_total: 10, net_r_multiple: 1, avg_r_multiple: 1 },
      trades: [{ symbol: 'BTCUSDT', close_time: '2026-01-01T00:00:00Z', side: 'Buy', net_profit: 10, result_pct: 1, currency: 'USDT' }]
    }) };
  }
  return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ resolved_symbol: 'BTCUSDT' }) };
};
eval(source);
(async () => {
  await Promise.resolve();
  await Promise.resolve();
  await new Promise((resolve) => setImmediate(resolve));
  const initialSpacerWidth = elements['trade-scroll-spacer'].style.width;
  elements['trade-table'].scrollWidth = 1777;
  elements['trade-table-wrap'].scrollWidth = 1777;
  if (!resizeHandler) throw new Error('resize handler was not registered');
  resizeHandler();
  await new Promise((resolve) => setTimeout(resolve, 5));
  elements['trade-scroll-top'].scrollLeft = 333;
  elements['trade-scroll-top'].listeners.scroll();
  const tableAfterTop = elements['trade-table-wrap'].scrollLeft;
  elements['trade-table-wrap'].scrollLeft = 222;
  elements['trade-table-wrap'].listeners.scroll();
  const topAfterTable = elements['trade-scroll-top'].scrollLeft;
  console.log(JSON.stringify({
    resizeRegistrations,
    initialSpacerWidth,
    resizedSpacerWidth: elements['trade-scroll-spacer'].style.width,
    tableAfterTop,
    topAfterTable,
  }));
})();
'''
    completed = subprocess.run(
        [node, "-e", harness, str(ROOT / "render" / "static" / "instrument_lookup.js")],
        check=True,
        capture_output=True,
        text=True,
    )
    result = json.loads(completed.stdout.strip().splitlines()[-1])
    assert result["resizeRegistrations"] == 1
    assert result["initialSpacerWidth"] == "960px"
    assert result["resizedSpacerWidth"] == "1777px"
    assert result["tableAfterTop"] == 333
    assert result["topAfterTable"] == 222


def test_instrument_lookup_detects_separator_normalized_fx_and_metals() -> None:
    node = shutil.which("node")
    assert node, "node is required for instrument lookup JS behavior test"
    harness = r'''
const fs = require('fs');
const source = fs.readFileSync(process.argv[1], 'utf8');
const variants = JSON.parse(process.argv[2]);

class Element {
  constructor(id) {
    this.id = id;
    this.value = '';
    this.textContent = '';
    this.innerHTML = '';
    this.tagName = 'DIV';
    this.listeners = {};
    this.classList = { toggle: () => {}, add: () => {}, remove: () => {} };
    this.dataset = {};
  }
  addEventListener(event, callback) { this.listeners[event] = callback; }
  closest() { return null; }
  querySelectorAll() { return []; }
}

async function runVariant(query) {
  const elements = Object.fromEntries(['q', 'load', 'rows', 'err', 'journal-status', 'journal-metrics', 'trade-head', 'trade-body'].map((id) => [id, new Element(id)]));
  const urls = [];
  global.window = { location: { search: `?q=${encodeURIComponent(query)}` } };
  global.history = { replaceState: () => {} };
  global.document = {
    head: { appendChild: () => {} },
    createElement: (tag) => new Element(tag),
    getElementById: (id) => elements[id] || null,
  };
  global.fetch = async (url) => {
    urls.push(String(url));
    let body = {};
    if (String(url).includes('/api/calculator/journal-summary')) body = { status: 'no_data', trades: [] };
    return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify(body) };
  };
  eval(source);
  await Promise.resolve();
  await Promise.resolve();
  await new Promise((resolve) => setImmediate(resolve));
  return urls;
}

(async () => {
  const result = {};
  for (const variant of variants) result[variant] = await runVariant(variant);
  console.log(JSON.stringify(result));
})();
'''
    variants = ["XAG/USD", "XAG-USD", "XAG USD", "XAG_USD", "XAGUSD", "EUR/USD", "EUR-USD", "EUR USD", "EUR_USD", "EURUSD"]
    completed = subprocess.run(
        [node, "-e", harness, str(ROOT / "render" / "static" / "instrument_lookup.js"), json.dumps(variants)],
        check=True,
        capture_output=True,
        text=True,
    )
    requested = json.loads(completed.stdout)
    for variant in variants:
        urls = requested[variant]
        assert any("/api/instrument-specs?query=" in url and "&prefer=oanda" in url for url in urls), variant
        assert any("/api/calculator/journal-summary?asset=fx" in url for url in urls), variant
        assert not any("/api/resolve-symbol" in url for url in urls), variant


def test_instrument_lookup_keeps_successful_panel_when_sibling_request_fails() -> None:
    node = shutil.which("node")
    assert node, "node is required for instrument lookup JS behavior test"
    harness = r'''
const fs = require('fs');
const source = fs.readFileSync(process.argv[1], 'utf8');

class Element {
  constructor(id) {
    this.id = id;
    this.value = '';
    this.textContent = '';
    this.innerHTML = '';
    this.tagName = 'DIV';
    this.listeners = {};
    this.classList = { toggle: () => {}, add: () => {}, remove: () => {} };
    this.dataset = {};
  }
  addEventListener(event, callback) { this.listeners[event] = callback; }
  closest() { return null; }
  querySelectorAll() { return []; }
}

async function runScenario(mode) {
  const elements = Object.fromEntries(['q', 'load', 'rows', 'err', 'journal-status', 'journal-metrics', 'trade-head', 'trade-body'].map((id) => [id, new Element(id)]));
  global.window = { location: { search: '?q=BTCUSDT' } };
  global.history = { replaceState: () => {} };
  global.document = {
    head: { appendChild: () => {} },
    createElement: (tag) => new Element(tag),
    getElementById: (id) => elements[id] || null,
  };
  global.fetch = async (url) => {
    const textUrl = String(url);
    if (mode === 'journal-fails' && textUrl.includes('/api/calculator/journal-summary')) {
      return { ok: false, status: 500, statusText: 'Server Error', text: async () => JSON.stringify({ detail: 'journal exploded' }) };
    }
    if (mode === 'specs-fails' && textUrl.includes('/api/instrument-specs')) {
      return { ok: false, status: 502, statusText: 'Bad Gateway', text: async () => JSON.stringify({ detail: 'specs exploded' }) };
    }
    if (textUrl.includes('/api/instrument-specs')) {
      return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ resolved_symbol: 'BTCUSDT', lastPrice: '100' }) };
    }
    if (textUrl.includes('/api/calculator/journal-summary')) {
      return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ status: 'ok', canonical_symbol: 'BTCUSDT', stats: { total_trades: 1, wins: 1, losses: 0, win_rate: '100.00%' }, metrics: { trades: 1, wins: 1, losses: 0, win_rate_pct: 100, net_profit_total: 10 }, trades: [{ symbol: 'BTCUSDT', close_time: '2026-01-01T00:00:00Z', side: 'Buy', net_profit: 10, result_pct: 1, currency: 'USDT' }] }) };
    }
    if (textUrl.includes('/api/resolve-symbol')) {
      return { ok: true, status: 200, statusText: 'OK', text: async () => JSON.stringify({ resolved_symbol: 'BTCUSDT' }) };
    }
    return { ok: true, status: 200, statusText: 'OK', text: async () => '{}' };
  };
  eval(source);
  await Promise.resolve();
  await Promise.resolve();
  await new Promise((resolve) => setImmediate(resolve));
  return {
    rows: elements.rows.innerHTML,
    journal: elements['journal-metrics'].innerHTML,
    tradeBody: elements['trade-body'].innerHTML,
    err: elements.err.textContent,
  };
}

(async () => {
  console.log(JSON.stringify({
    journalFails: await runScenario('journal-fails'),
    specsFails: await runScenario('specs-fails'),
  }));
})();
'''
    completed = subprocess.run(
        [node, "-e", harness, str(ROOT / "render" / "static" / "instrument_lookup.js")],
        check=True,
        capture_output=True,
        text=True,
    )
    result = json.loads(completed.stdout)
    assert "Last price" in result["journalFails"]["rows"]
    assert "Journal stats failed" in result["journalFails"]["err"]
    assert "Journal stats are unavailable" in result["journalFails"]["journal"]
    assert "No instrument specs loaded yet" in result["specsFails"]["rows"]
    assert "Instrument specs failed" in result["specsFails"]["err"]
    assert "Trades" in result["specsFails"]["journal"]
    assert "Buy" in result["specsFails"]["tradeBody"]


def test_dashboard_home_removed_legacy_open_trading_journal_panel() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    html = _extract_html_template(source)
    assert 'id="open-master-journal-btn"' not in html


def _load_master_service_module():
    spec = importlib.util.spec_from_file_location("render_master_service_dashboard", ROOT / "render" / "master_service.py")
    module = importlib.util.module_from_spec(spec)
    assert spec and spec.loader
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def test_dashboard_home_uses_versioned_dashboard_js_url() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    assert '/static/dashboard.js?v=' in source


def test_local_profile_sets_no_cache_for_home_and_static_assets(monkeypatch) -> None:
    pytest = __import__('pytest')
    try:
        from fastapi.testclient import TestClient
    except Exception:
        pytest.skip('fastapi TestClient is unavailable in this environment')

    module = _load_master_service_module()
    monkeypatch.setattr(module, 'APP_PROFILE', 'local')
    client = TestClient(module.app)

    home = client.get('/')
    cache = home.headers.get('cache-control', '')
    assert 'no-store' in cache or 'no-cache' in cache

    static = client.get('/static/dashboard.js')
    static_cache = static.headers.get('cache-control', '')
    assert 'no-store' in static_cache or 'no-cache' in static_cache


def test_local_profile_buttons_use_configured_render_routes(monkeypatch) -> None:
    module = _load_master_service_module()
    monkeypatch.setattr(module, "APP_PROFILE", "local")
    monkeypatch.setenv("RENDER_CALCULATOR_BASE_URL", "https://tools.example/")
    buttons = module._profile_main_buttons()
    by_name = {str(item.get("name")): item for item in buttons}
    assert by_name["calculator"]["open_url"] == "/merged/calculator"
    assert by_name["bounce-trader"]["open_url"] == "https://tools.example/merged/bounce-trader"
    assert by_name["bounce-trader"]["remote_owned"] is True
    assert by_name["fxweekend"]["open_url"] == "https://tools.example/apps/fxweekend-clone"
    assert by_name["fxweekend"]["remote_owned"] is True
    assert by_name["trading-journal"]["open_url"] == "/dashboard/trading-journal"
    assert by_name["monitor"]["label"] == "Alerts"
    assert by_name["monitor"]["open_url"] == "/merged/monitor"
    assert by_name["atr-scanner"]["label"] == "Scanner"
    assert by_name["atr-scanner"]["open_url"] == "/merged/atr-scanner"
    assert by_name["monitor"]["id"] != by_name["atr-scanner"]["id"]
    assert by_name["instrument-lookup"]["open_url"] == "/instrument-lookup"
    assert by_name["instrument-lookup"]["dashboard_main_view"] is True
    assert by_name["history"]["open_url"] == "/merged/history"
    assert by_name["history"]["dashboard_main_view"] is True
    assert by_name["spreads-clone"]["label"] == "Oanda Spreads"
    assert [item["name"] for item in buttons] == [
        "calculator",
        "bounce-trader",
        "fxweekend",
        "trading-journal",
        "instrument-lookup",
        "history",
        "monitor",
        "atr-scanner",
        "ivindicator-clone",
        "spreads-clone",
    ]
    assert "open-orders" not in by_name
    assert "pine" not in by_name
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    assert '@app.get("/dashboard/trading-journal")' in source
    assert '@app.get("/trading-journal", response_class=HTMLResponse)' in source
    assert '@app.get("/dashboard/pine", response_class=HTMLResponse)' in source


def test_remote_tool_buttons_refuse_localhost_and_render_has_no_tool_buttons(monkeypatch) -> None:
    module = _load_master_service_module()
    monkeypatch.setattr(module, "APP_PROFILE", "local")
    monkeypatch.setenv("RENDER_CALCULATOR_BASE_URL", "http://127.0.0.1:9000")
    monkeypatch.setenv("RENDER_EXTERNAL_URL", "https://render-fallback.example")
    fallback_by_name = {str(item.get("name")): item for item in module._profile_main_buttons()}
    assert fallback_by_name["bounce-trader"]["open_url"] == "https://render-fallback.example/merged/bounce-trader"
    assert fallback_by_name["fxweekend"]["open_url"] == "https://render-fallback.example/apps/fxweekend-clone"

    monkeypatch.delenv("RENDER_EXTERNAL_URL", raising=False)
    by_name = {str(item.get("name")): item for item in module._profile_main_buttons()}
    assert by_name["bounce-trader"]["open_url"] == "/render-tools-configuration-error"
    assert by_name["fxweekend"]["open_url"] == "/render-tools-configuration-error"

    error_page = asyncio.run(module.render_tools_configuration_error())
    assert error_page.status_code == 503
    assert "RENDER_CALCULATOR_BASE_URL" in error_page.body.decode("utf-8")

    monkeypatch.setattr(module, "APP_PROFILE", "render")
    assert module._profile_main_buttons() == []


def test_pine_dashboard_api_lists_reads_and_blocks_traversal(monkeypatch) -> None:
    module = _load_master_service_module()
    tmp_root = _repo_tmp_dir("dashboard_pine")
    try:
        pine_root = tmp_root / "pinescripts"
        pine_root.mkdir()
        (pine_root / "custom_indicator.pine").write_text("//@version=6\nindicator('x')\n", encoding="utf-8")
        (pine_root / "atr_percentage.pine").write_text("//@version=6\nindicator('atr')\n", encoding="utf-8")
        (pine_root / "notes.md").write_text("not a pine script", encoding="utf-8")
        monkeypatch.setattr(module, "APP_PROFILE", "local")
        monkeypatch.setattr(module, "PINE_SCRIPTS_DIR", pine_root)

        files_response = asyncio.run(module.pine_files())
        files_payload = json.loads(files_response.body.decode("utf-8"))
        assert files_payload["files"] == ["atr_percentage.pine", "custom_indicator.pine"]

        file_response = asyncio.run(module.pine_file("custom_indicator.pine"))
        file_payload = json.loads(file_response.body.decode("utf-8"))
        assert file_payload["display_path"] == "pinescripts/custom_indicator.pine"
        assert file_payload["code"].startswith("//@version=6")

        atr_response = asyncio.run(module.pine_file("atr_percentage.pine"))
        atr_payload = json.loads(atr_response.body.decode("utf-8"))
        assert atr_payload["display_path"] == "pinescripts/atr_percentage.pine"
        assert "indicator('atr')" in atr_payload["code"]

        with pytest.raises(Exception) as excinfo:
            asyncio.run(module.pine_file("../secret.pine"))
        assert getattr(excinfo.value, "status_code", None) == 400
    finally:
        shutil.rmtree(tmp_root, ignore_errors=True)


def test_pine_dashboard_page_has_clipboard_and_textarea_fallback() -> None:
    source = MASTER_SERVICE_PATH.read_text(encoding="utf-8")
    dashboard_js = (ROOT / "render" / "static" / "dashboard.js").read_text(encoding="utf-8")
    assert "navigator.clipboard.writeText(code)" in source
    assert "navigator.clipboard.writeText(code)" in dashboard_js
    assert '<textarea id="fallback" hidden>' in source
    assert 'id="pine-fallback"' in source
    assert "Clipboard unavailable. Use manual copy below." in source
    assert "Clipboard unavailable. Use manual copy below." in dashboard_js


def test_trading_journal_workspace_contains_action_buttons_in_order():
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    assert "open-journal-btn" in source
    assert "import-journal-btn" in source
    assert "journal-resync-btn" in source
    assert ">Resync<" in source
    assert "crypto-monthly-pnl-btn" in source
    assert "bybit-demo-balance-adjustment-btn" in source
    assert "Bybit Demo Balance Adjustment" in source
    assert source.index("open-journal-btn") < source.index("import-journal-btn") < source.index("journal-resync-btn") < source.index("crypto-monthly-pnl-btn") < source.index("bybit-demo-balance-adjustment-btn")


def test_open_master_journal_polish_autofits_and_clears_stale_recommendation_fill(tmp_path: Path):
    module = _load_master_service_module()
    path = tmp_path / "Trading Journal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Stats 1"
    ws["A1"] = "Recommendation"
    ws["B1"] = "Reduce target"
    ws["B1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws["C1"] = "A very long journal note that should widen the column before Excel opens"
    ws.column_dimensions["C"].width = 8
    wb.save(path)
    wb.close()

    result = module._polish_master_journal_for_excel_open(path)

    assert result["ok"] is True
    checked = load_workbook(path)
    try:
        ws2 = checked["Stats 1"]
        assert str(ws2["B1"].fill.fgColor.rgb or "")[-6:].upper() != "FFF2CC"
        assert float(ws2.column_dimensions["C"].width or 0) > 8
    finally:
        checked.close()


def test_open_master_journal_polish_keeps_trade_log_data_row_heights(tmp_path: Path):
    module = _load_master_service_module()
    path = tmp_path / "Trading Journal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        ws.cell(TRADE_LOG_FILTER_HEADER_ROW, col).value = header
    notes_col = TRADE_LOG_HEADERS.index("Notes") + 1
    row_id_col = TRADE_LOG_HEADERS.index("Row ID") + 1
    data_row = TRADE_LOG_DATA_START_ROW
    ws.cell(data_row, 1).value = "2026-01-01T00:00:00Z"
    ws.cell(data_row, 2).value = "2026-01-01T00:05:00Z"
    ws.cell(data_row, 3).value = "Bybit Demo"
    ws.cell(data_row, 4).value = "BTCUSDT"
    ws.cell(data_row, row_id_col).value = "row-long-note"
    ws.cell(data_row, notes_col).value = "Long wrapped note " * 30
    ws.row_dimensions[data_row].height = TRADE_LOG_DATA_ROW_HEIGHT
    wb.save(path)
    wb.close()

    result = module._polish_master_journal_for_excel_open(path)

    assert result["ok"] is True
    checked = load_workbook(path)
    try:
        ws2 = checked["Trade Log"]
        assert float(ws2.row_dimensions[data_row].height) == float(TRADE_LOG_DATA_ROW_HEIGHT)
    finally:
        checked.close()


def test_open_master_journal_polish_inserts_stats1_recommendation_rows(tmp_path: Path):
    module = _load_master_service_module()
    path = tmp_path / "Trading Journal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "STATS1"
    ws["B1"] = "Overall"
    ws["C1"] = "FX"
    ws["D1"] = "Crypto"
    ws["A2"] = "Avg stop %"
    ws["A3"] = "Min stop %"
    ws["A4"] = "Max stop %"
    ws["A5"] = "Avg target %"
    ws["A6"] = "Min target %"
    ws["A7"] = "Max target %"
    ws["A8"] = "Winners"
    ws["A9"] = "Avg stop %"
    ws["B9"] = 0.01
    ws["C9"] = 0.02
    ws["D9"] = 0.03
    ws["A10"] = "Avg target %"
    ws["B10"] = 0.04
    ws["C10"] = 0.01
    ws["D10"] = 0.05
    ws["A11"] = "Losers"
    ws["A12"] = "Avg stop %"
    ws["B12"] = 0.02
    ws["C12"] = 0.01
    ws["D12"] = 0.03
    ws["A13"] = "Avg target %"
    ws["B13"] = 0.03
    ws["C13"] = 0.02
    ws["D13"] = 0.05
    wb.save(path)
    wb.close()

    result = module._polish_master_journal_for_excel_open(path)

    assert result["ok"] is True
    assert result["stats1_recommendation_cells_repaired"] == 3
    checked = load_workbook(path)
    try:
        ws2 = checked["STATS1"]
        labels = {row: str(ws2.cell(row, 1).value or "") for row in range(1, ws2.max_row + 1)}
        max_stop_row = next(row for row, label in labels.items() if label == "Max stop %")
        stop_recommendation_row = max_stop_row + 1
        max_target_row = next(row for row, label in labels.items() if label == "Max target %")
        target_recommendation_row = max_target_row + 1
        assert "Source" not in labels.values()
        assert ws2.cell(stop_recommendation_row, 1).value == "Recommendation"
        assert [ws2.cell(stop_recommendation_row, col).value for col in range(2, 5)] == [
            "Decrease stop \u2014 Recommended: 1.00% (1.00 pp below loss average)",
            "Increase stop \u2014 Recommended: 2.00% (1.00 pp above loss average)",
            "Decrease stop \u2014 Recommended: 2.99% (0.01 pp below loss average; exact tie, so a small decrease is preferred)",
        ]
        assert ws2.cell(target_recommendation_row, 1).value == "Recommendation"
        assert [ws2.cell(target_recommendation_row, col).value for col in range(2, 5)] == [None, None, None]
    finally:
        checked.close()


@pytest.mark.parametrize(
    ("winner_pct", "loser_pct", "prefix"),
    [
        ("1.0000000000005", "1.0", "Increase stop"),
        ("0.9999999999995", "1.0", "Decrease stop"),
        ("1.0", "1.0", "Decrease stop"),
    ],
)
def test_open_master_journal_polish_uses_decimal_stop_recommendation_payload(
    tmp_path: Path,
    winner_pct: str,
    loser_pct: str,
    prefix: str,
):
    module = _load_master_service_module()
    path = tmp_path / "Trading Journal.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "STATS1"
    ws["B1"] = "Overall"
    ws["C1"] = "FX"
    ws["D1"] = "Crypto"
    ws["A2"] = "Avg stop %"
    ws["A3"] = "Min stop %"
    ws["A4"] = "Max stop %"
    ws["A5"] = "Winners"
    ws["A6"] = "Avg stop %"
    ws["B6"] = float(Decimal(winner_pct) / Decimal("100"))
    ws["C6"] = 0.02
    ws["D6"] = 0.03
    ws["A7"] = "Losers"
    ws["A8"] = "Avg stop %"
    ws["B8"] = float(Decimal(loser_pct) / Decimal("100"))
    ws["C8"] = 0.01
    ws["D8"] = 0.04
    wb.save(path)
    wb.close()

    result = module._polish_master_journal_for_excel_open(path)

    assert result["ok"] is True
    checked = load_workbook(path)
    try:
        ws2 = checked["STATS1"]
        max_stop_row = next(row for row in range(1, ws2.max_row + 1) if ws2.cell(row, 1).value == "Max stop %")
        recommendation = ws2.cell(max_stop_row + 1, 2).value
    finally:
        checked.close()
    expected = module._stop_recommendation_payload([Decimal(winner_pct)], [Decimal(loser_pct)])[module.STOP_RECOMMENDATION_HEADER]
    assert recommendation == expected
    assert str(recommendation).startswith(prefix)
    if winner_pct == loser_pct:
        assert "exact tie, so a small decrease is preferred" in str(recommendation)
        assert "exact_tie_goal_preference_decrease" not in str(recommendation)



def test_trading_journal_actions_crypto_monthly_diagnostics_are_rendered() -> None:
    js = (ROOT / 'render' / 'static' / 'trading_journal_actions.js').read_text(encoding='utf-8')
    assert 'renderCryptoMonthlyDiagnostics' in js
    assert 'crypto-monthly-pnl-diagnostics' in js
    assert 'Crypto Monthly P&amp;L diagnostics / raw JSON' in js
    for token in ['now_brisbane', 'current_month', 'last_completed_month', 'state_months', 'workbook_months', 'ignored_invalid_workbook_anchors', 'missing_workbook_months', 'verified_row_ids', 'code_version', 'app_commit']:
        assert token in js
    assert 'payload.rows_inserted || 0' not in js

def test_trading_journal_workspace_bybit_account_dropdown_has_no_auto_option():
    source = MASTER_SERVICE_PATH.read_text(encoding='utf-8')
    assert '>Auto<' not in source
    assert '<option value="">Auto</option>' not in source
    assert '<option value="" selected disabled>Select Demo or Live</option>' in source
    assert '<option value="demo">Demo</option>' in source
    assert '<option value="live">Live</option>' in source
