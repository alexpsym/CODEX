from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from tools.backfill_legacy_forex_trade_metadata import run_backfill
from tools.master_journal_workbook import build_master_journal_workbook


def test_forex_backfill_reports_duplicate_execution_as_ambiguous(tmp_path: Path):
    journal = tmp_path / "Trading Journal.xlsx"
    legacy = tmp_path / "FOREX JOURNAL.xlsx"
    row = {
        "row_type": "trade", "account": "OANDA LIVE", "asset_class": "fx",
        "symbol": "NZDUSD", "side": "BUY",
        "open_time": "2024-03-05 16:19:35", "close_time": "2024-03-05 16:23:15",
        "entry_price": 0.60814, "exit_price": 0.60737, "stop_loss": 0.60739,
        "take_profit": 0.61009, "result_pct": -1.0, "r_multiple": -1.02,
    }
    build_master_journal_workbook({
        "items": [{"id": "a", **row}, {"id": "b", **row}],
        "stats": {"totals": {}, "groups": {}},
        "balances": [],
    }, journal)

    wb = Workbook()
    ws = wb.active
    ws.title = "TRADE LOG"
    ws["A4"] = 1008
    ws["B4"] = "NZDUSD"
    ws["C4"] = datetime(2024, 3, 5, 16, 19, 35)
    ws["D4"] = datetime(2024, 3, 5, 16, 23, 15)
    ws["R4"] = 0.60814
    ws["S4"] = 0.60737
    ws["T4"] = "LONG"
    ws["X4"] = 0.60739
    ws["AB4"] = 0.61009
    wb.save(legacy)
    wb.close()

    dry = run_backfill(journal, legacy)
    assert dry["legacy_trades_parsed"] == 1
    assert dry["matches"] == 0
    assert dry["ambiguous"] == 1
    applied = run_backfill(journal, legacy, apply_changes=True)
    assert applied["applied"] is False
    assert applied["error"] == "unsafe_match_result"
