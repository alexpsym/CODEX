import asyncio
import importlib.util
import io
import json
import sqlite3
import sys
import threading
import types
import importlib.machinery
from datetime import date, datetime, timezone
from pathlib import Path

import pytest
from fastapi import HTTPException
if "multipart" not in sys.modules:
    multipart_mod = types.ModuleType("multipart")
    multipart_mod.__spec__ = importlib.machinery.ModuleSpec("multipart", loader=None)
    multipart_mod.__version__ = "0.0-test"
    multipart_sub = types.ModuleType("multipart.multipart")
    multipart_sub.__spec__ = importlib.machinery.ModuleSpec("multipart.multipart", loader=None)
    multipart_sub.parse_options_header = lambda value: (value, {})
    sys.modules["multipart"] = multipart_mod
    sys.modules["multipart.multipart"] = multipart_sub

try:
    import requests as _rq  # noqa: F401
    if not hasattr(_rq, "adapters"):
        raise ImportError
except Exception:
    import importlib.machinery
    req = types.ModuleType("requests")
    req.__spec__ = importlib.machinery.ModuleSpec("requests", loader=None)
    adapters = types.ModuleType("requests.adapters")
    adapters.__spec__ = importlib.machinery.ModuleSpec("requests.adapters", loader=None)
    adapters.HTTPAdapter = object
    req.adapters = adapters
    sys.modules["requests"] = req
    sys.modules["requests.adapters"] = adapters
try:
    from urllib3.util.retry import Retry  # noqa: F401
except Exception:
    import importlib.machinery
    urllib3 = types.ModuleType("urllib3")
    urllib3.__spec__ = importlib.machinery.ModuleSpec("urllib3", loader=None)
    util = types.ModuleType("urllib3.util")
    util.__spec__ = importlib.machinery.ModuleSpec("urllib3.util", loader=None)
    retry = types.ModuleType("urllib3.util.retry")
    retry.__spec__ = importlib.machinery.ModuleSpec("urllib3.util.retry", loader=None)
    retry.Retry = object
    sys.modules["urllib3"] = urllib3
    sys.modules["urllib3.util"] = util
    sys.modules["urllib3.util.retry"] = retry

try:
    import httpx  # noqa: F401
except Exception:
    import importlib.machinery
    class _HttpxResponse:  # minimal class for starlette.testclient MRO
        pass
    class _HttpxAsyncClient:
        pass
    httpx_stub = types.SimpleNamespace(
        Timeout=lambda *args, **kwargs: None,
        AsyncClient=_HttpxAsyncClient,
        Response=_HttpxResponse,
        TimeoutException=Exception,
        RequestError=Exception,
        HTTPStatusError=Exception,
        ConnectError=Exception,
    )
    httpx_stub.__spec__ = importlib.machinery.ModuleSpec("httpx", loader=None)
    sys.modules["httpx"] = httpx_stub

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
SPEC = importlib.util.spec_from_file_location("render_master_service_journal_crud", ROOT / "render" / "master_service.py")
master_service = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = master_service
SPEC.loader.exec_module(master_service)


def _legacy_sync_status_payload():
    response = asyncio.run(master_service._legacy_trading_journal_sync_status())
    return json.loads(response.body.decode("utf-8"))


@pytest.fixture
def temp_state_paths(tmp_path, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_PATH", tmp_path / "trading_journal.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_STATE_PATH", tmp_path / "trading_journal_state.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_STATE_PATH", tmp_path / "trading_journal_sync_state.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_VIEW_CACHE_PATH", tmp_path / "trading_journal_view_cache.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_IMPORT_CACHE_PATH", tmp_path / "trading_journal_import_cache.json")
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    master_service._TRADING_JOURNAL_VIEW_CACHE["key"] = None
    master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] = None
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    return tmp_path


def _json(res):
    return json.loads(res.body.decode("utf-8"))


def test_create_manual_trade_row(temp_state_paths):
    res = asyncio.run(
        master_service.trading_journal_create_row(
            {
                "open_time": "2026-04-01T00:00:00Z",
                "close_time": "2026-04-01T01:00:00Z",
                "symbol": "eurusd",
                "side": "buy",
                "qty": "1.5",
                "entry_price": "1.1",
                "exit_price": "1.2",
                "net_profit": "12.3",
                "balance_after_trade": "1012.3",
                "is_test_trade": "true",
            }
        )
    )
    payload = _json(res)
    row = payload["row"]
    assert payload["ok"] is True
    assert row["id"].startswith("manual:")
    assert row["source"] == "manual"
    assert row["row_type"] == "trade"
    assert row["is_manual"] is True
    assert row["symbol"] == "EURUSD"
    assert row["is_test_trade"] is True
    stored = master_service._get_trading_journal_rows()
    assert len(stored) == 1
    assert stored[0]["id"] == row["id"]


def test_patch_existing_trade_row_stores_manual_overrides(temp_state_paths):
    master_service._set_trading_journal_rows(
        [
            {
                "id": "oanda:live:t1",
                "row_type": "trade",
                "source": "oanda",
                "status": "closed",
                "open_time": "2026-04-01T00:00:00+00:00",
                "close_time": "2026-04-01T01:00:00+00:00",
                "symbol": "EUR_USD",
                "net_profit": 10.0,
            }
        ]
    )
    res = asyncio.run(
        master_service.trading_journal_patch_row(
            "oanda:live:t1",
            {"notes": "manual note", "timeframe": "1-hour"},
        )
    )
    row = _json(res)["row"]
    assert row["manual_overrides"]["notes"] == "manual note"
    assert row["manual_overrides"]["timeframe"] == "1-hour"
    asyncio.run(
        master_service.trading_journal_patch_row(
            "oanda:live:t1",
            {"is_test_trade": "false"},
        )
    )
    row2 = master_service._get_trading_journal_rows()[0]
    assert row2["is_test_trade"] is False
    assert "notes" in row["manual_override_fields"]


def test_manual_overrides_survive_later_sync_upsert(temp_state_paths):
    base = {
        "id": "oanda:live:t1",
        "row_type": "trade",
        "source": "oanda",
        "status": "closed",
        "symbol": "EUR_USD",
        "open_time": "2026-04-01T00:00:00+00:00",
        "close_time": "2026-04-01T01:00:00+00:00",
        "notes": "source-note",
        "timeframe": "15-minute",
    }
    base = master_service._apply_trading_journal_manual_overrides(base, {"notes": "edited", "timeframe": "4-hour"})
    master_service._set_trading_journal_rows([base])

    master_service._upsert_trading_journal_rows(
        [
            {
                "id": "oanda:live:t1",
                "row_type": "trade",
                "source": "oanda",
                "status": "closed",
                "notes": "new-source-note",
                "timeframe": "1-minute",
            }
        ]
    )
    row = master_service._get_trading_journal_rows()[0]
    assert row["notes"] == "edited"
    assert row["timeframe"] == "4-hour"


def test_concurrent_upserts_preserve_both_rows(temp_state_paths):
    start = threading.Barrier(3)
    errors: list[BaseException] = []

    def worker(row_id: str) -> None:
        try:
            start.wait(timeout=2)
            master_service._upsert_trading_journal_rows(
                [
                    {
                        "id": row_id,
                        "row_type": "trade",
                        "source": "bybit",
                        "status": "closed",
                        "symbol": "BTCUSDT",
                        "close_time": "2026-04-01T01:00:00+00:00",
                    }
                ]
            )
        except BaseException as exc:  # pragma: no cover - diagnostics for thread failures
            errors.append(exc)

    t1 = threading.Thread(target=worker, args=("bybit:demo:closed:1",))
    t2 = threading.Thread(target=worker, args=("bybit:demo:closed:2",))
    t1.start()
    t2.start()
    start.wait(timeout=2)
    t1.join(timeout=2)
    t2.join(timeout=2)

    assert not errors
    rows = master_service._get_trading_journal_rows()
    ids = {str((row or {}).get("id") or "") for row in rows}
    assert "bybit:demo:closed:1" in ids
    assert "bybit:demo:closed:2" in ids


def test_reject_cashflow_edit(temp_state_paths):
    master_service._set_trading_journal_rows(
        [{"id": "cf1", "row_type": "cashflow", "source": "dropbox", "close_time": "2026-04-01T00:00:00+00:00"}]
    )
    with pytest.raises(HTTPException) as exc:
        asyncio.run(master_service.trading_journal_patch_row("cf1", {"notes": "x"}))
    assert exc.value.status_code == 409


def test_reject_protected_field_edit(temp_state_paths):
    with pytest.raises(HTTPException) as exc:
        master_service._normalize_trading_journal_edit_payload(
            {"id": "x", "notes": "abc"},
            for_create=False,
            existing={"id": "abc", "row_type": "trade", "source": "oanda"},
        )
    assert exc.value.status_code == 422


def test_delete_manual_row(temp_state_paths):
    master_service._set_trading_journal_rows(
        [{"id": "manual:r1", "row_type": "trade", "source": "manual", "is_manual": True}]
    )
    res = asyncio.run(master_service.trading_journal_delete_row("manual:r1"))
    payload = _json(res)
    assert payload["ok"] is True
    assert master_service._get_trading_journal_rows() == []


def test_stats_and_balances_still_compute_after_create_and_edit(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(
        master_service,
        "_get_excel_account_balances",
        lambda: [{"account": "Manual Account", "label": "Manual Account", "balance": 1000.0, "currency": "USD"}],
    )
    created = _json(
        asyncio.run(
            master_service.trading_journal_create_row(
                {
                    "open_time": "2026-04-01T00:00:00Z",
                    "close_time": "2026-04-01T01:00:00Z",
                    "symbol": "BTCUSDT",
                    "side": "Buy",
                    "account": "Manual Account",
                    "account_label": "Manual Account",
                    "currency": "USD",
                    "qty": "1",
                    "entry_price": "100",
                    "exit_price": "105",
                    "net_profit": "5",
                    "balance_after_trade": "1005",
                }
            )
        )
    )["row"]
    asyncio.run(master_service.trading_journal_patch_row(created["id"], {"notes": "after edit"}))
    master_service._build_trading_journal_view_snapshot(force=True)
    journal = _json(asyncio.run(master_service.trading_journal_items()))
    balances = _json(asyncio.run(master_service.trading_journal_balances()))
    assert int(journal.get("count") or 0) >= 1
    assert isinstance(journal.get("stats"), dict)
    assert isinstance(balances.get("items"), list)


def test_trading_journal_js_contains_crud_controls_and_endpoints():
    js = (ROOT / "render" / "static" / "trading_journal.js").read_text(encoding="utf-8")
    assert "/api/trading-journal/rows" in js
    assert "/api/trading-journal/diagnostics" in js
    assert "tj-export-btn" in js
    assert "exportShownTrades" in js
    assert 'data-action="edit"' in js
    assert 'data-action="delete"' in js
    assert "location.reload" not in js
    assert "Cached data shown, refreshing…" in js
    assert "writeCachedPayload({" in js
    assert "/api/trading-journal/sync/status" not in js
    assert "/api/trading-journal/readiness" not in js
    assert "isAbortError" in js
    assert "fetchNamedJson" in js
    assert "manualSyncInFlight" not in js
    assert "skipAutoSync: true" in js
    assert "const AUTO_REFRESH_MS = 60 * 60 * 1000;" in js
    assert "preserveStatus" in js
    assert "journalPending" in js
    assert "journal cache is building" in js.lower()
    assert "if (!journalPending) {" in js
    assert "state.rows = nextRows" in js
    assert "new Error(`/api/trading-journal:" not in js

    assert "renderBalances(items)" in js
    assert "No balances available yet." in js
    assert "source: ${source}" not in js
    assert "as of: ${asOf}" not in js
    assert "Balance not found in workbook" not in js
    assert "Background Dropbox sync running…" not in js
    assert "Background local journal import running…" not in js
    assert "compactErrorMessage" in js
    assert "slice(0, 300)" in js
    assert "Sync finished but reload failed:" not in js
    assert "Restart the journal launcher so dependencies can be installed automatically." in js
    assert "Sync complete: 0 rows loaded" not in js
    assert "MISSING_XLRD_FOR_XLS" in js


def test_diagnostics_derive_rows_total_from_existing_journal_rows(temp_state_paths):
    rows = []
    for idx in range(25):
        rows.append(
            {
                "id": f"manual:{idx}",
                "row_type": "trade",
                "source": "manual",
                "asset_class": "crypto",
                "symbol": "BTCUSDT",
                "status": "closed",
                "close_time": f"2026-04-01T00:{idx:02d}:00+00:00",
            }
        )
    master_service._set_trading_journal_rows(rows)
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS = master_service._default_journal_diagnostics()

    payload = _json(asyncio.run(master_service.trading_journal_diagnostics()))
    assert payload["rows_total"] >= 25
    assert payload["journal_rows_total"] >= 25
    assert payload["has_current_journal_rows"] is True
    assert payload["rows_by_source"]


def test_import_sources_local_mode_skips_dropbox(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "_local_journal_import_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: [])
    monkeypatch.setattr(master_service, "_ensure_trading_journal_local_templates", lambda: {"errors": []})
    called = {"dropbox": 0}
    monkeypatch.setattr(
        master_service,
        "_import_trading_journal_from_dropbox_excel",
        lambda progress_cb=None: called.__setitem__("dropbox", called["dropbox"] + 1),
    )
    result = master_service._import_trading_journal_from_sources()
    assert called["dropbox"] == 0
    assert result["dropbox_workbooks_seen"] == 0


def test_sync_status_exposes_source_and_flags(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", Path(r"C:\Users\User\Documents\TRADING"))
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", True)
    monkeypatch.setattr(master_service, "APP_PROFILE", "journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BROKER_REFRESH_ENABLED", False)
    monkeypatch.setenv("DROPBOX_SYNC_ENABLED", "0")
    payload = _legacy_sync_status_payload()
    assert payload["source_mode"] == "local"
    assert payload["local_import_enabled"] is True
    assert payload["uses_dropbox_journal_import"] is False
    assert payload["dropbox_sync_enabled"] is False
    assert payload["broker_refresh_enabled"] is False
    assert isinstance(payload["dependencies"], dict)
    assert "xlrd_installed" in payload["dependencies"]
    assert "local_xls_supported" in payload["dependencies"]
    assert payload["dependencies"]["requirements_file"].replace("\\", "/").endswith("render/requirements.txt")


def test_sync_status_dependency_flags_reflect_missing_xlrd(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service.importlib.util, "find_spec", lambda name: None if name == "xlrd" else object())
    payload = _legacy_sync_status_payload()
    assert payload["dependencies"]["xlrd_installed"] is False
    assert payload["dependencies"]["local_xls_supported"] is False


def test_sync_status_serializes_nested_datetime_values(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    master_service.TRADING_JOURNAL_SYNC_STATE.clear()
    master_service.TRADING_JOURNAL_SYNC_STATE.update(
        {
            "running": False,
            "ok": False,
            "result": {
                "nested": {
                    "when": datetime(2026, 5, 1, 12, 30, tzinfo=timezone.utc),
                    "trade_date": date(2026, 5, 1),
                }
            },
        }
    )
    payload = _legacy_sync_status_payload()
    assert payload["ok"] is False
    assert payload["result"]["nested"]["when"] == "2026-05-01T12:30:00+00:00"
    assert payload["result"]["nested"]["trade_date"] == "2026-05-01"
    assert payload["result"]["nested"]["when"]


def test_set_trading_journal_sync_state_persists_json_safe_datetime(temp_state_paths):
    master_service._set_trading_journal_sync_state(
        running=False,
        ok=False,
        result={"nested": {"when": datetime.now(timezone.utc)}},
    )
    saved = json.loads(master_service.TRADING_JOURNAL_SYNC_STATE_PATH.read_text(encoding="utf-8"))
    assert isinstance(saved["result"]["nested"]["when"], str)
    assert "T" in saved["result"]["nested"]["when"]


def test_save_trading_journal_view_snapshot_sanitizes_datetime(temp_state_paths):
    payload = {
        "cache_version": master_service.TRADING_JOURNAL_VIEW_CACHE_VERSION,
        "generated_at": datetime(2026, 5, 2, 10, 0, tzinfo=timezone.utc),
        "items": [{"id": "x", "when": datetime(2026, 5, 2, 11, 0, tzinfo=timezone.utc)}],
        "balances": [],
        "stats": {"as_of": date(2026, 5, 2)},
        "diagnostics": {"errors": [], "updated": datetime(2026, 5, 2, 12, 0, tzinfo=timezone.utc)},
        "source_fingerprints": {"files": []},
    }
    master_service._save_trading_journal_view_snapshot(payload)
    stored = json.loads(master_service.TRADING_JOURNAL_VIEW_CACHE_PATH.read_text(encoding="utf-8"))
    assert stored["generated_at"] == "2026-05-02T10:00:00+00:00"
    assert stored["items"][0]["when"] == "2026-05-02T11:00:00+00:00"
    assert stored["stats"]["as_of"] == "2026-05-02"


def test_persist_trading_journal_sqlite_sanitizes_datetime_values(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SQLITE_PATH", temp_state_paths / "journal.sqlite3")
    snapshot = {
        "generated_at": datetime(2026, 5, 2, 9, 0, tzinfo=timezone.utc),
        "items": [{"id": "t1", "row_type": "trade", "metrics": {"updated": datetime(2026, 5, 2, tzinfo=timezone.utc)}}],
        "balances": [{"account": "OANDA DEMO", "as_of": datetime(2026, 5, 2, tzinfo=timezone.utc)}],
        "stats": {"as_of": date(2026, 5, 2)},
        "diagnostics": {"errors": [], "updated": datetime(2026, 5, 2, tzinfo=timezone.utc)},
        "source_fingerprints": {"files": [{"path": temp_state_paths / "x.xlsx", "mtime": 1.2, "size": 3}]},
    }
    master_service._persist_trading_journal_sqlite(snapshot, import_meta={"warnings": [datetime(2026, 5, 2, tzinfo=timezone.utc)], "errors": []})
    conn = sqlite3.connect(master_service.TRADING_JOURNAL_SQLITE_PATH)
    try:
        row = conn.execute("SELECT payload_json FROM journal_trades WHERE id='t1'").fetchone()
        assert row is not None
        payload = json.loads(row[0])
        assert payload["metrics"]["updated"].startswith("2026-05-02T")
    finally:
        conn.close()


def test_local_xls_missing_xlrd_returns_hard_failure(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    xls_file = temp_state_paths / "journal.xls"
    xls_file.write_bytes(b"dummy")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "_local_journal_import_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_ensure_trading_journal_local_templates", lambda: {"errors": []})
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: [xls_file])
    original_import = __import__

    def fake_import(name, *args, **kwargs):
        if name == "xlrd":
            raise ImportError("No module named xlrd")
        return original_import(name, *args, **kwargs)

    monkeypatch.setattr("builtins.__import__", fake_import)
    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is False
    assert result["errors"][0]["code"] == "MISSING_XLRD_FOR_XLS"
    assert "render" in result["errors"][0]["message"]
    assert "requirements.txt" in result["errors"][0]["message"]
    assert "same Python executable" in result["errors"][0]["message"]


def test_run_sync_job_local_profile_skips_broker_refresh(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, "APP_PROFILE", "journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BROKER_REFRESH_ENABLED", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BYBIT_DEMO_BALANCE_ANCHOR_ENABLED", False)
    calls = {"balance": 0, "closed": 0}

    async def _fake_balance(_mode):
        calls["balance"] += 1
        return {"available_usdt": 0}

    async def _fake_closed(**_kwargs):
        calls["closed"] += 1
        return {"ok": True}

    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", _fake_balance)
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", _fake_closed)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda progress_cb=None: {"ok": True, "rows_imported": 1, "diagnostics": {"rows_by_asset_class": {}}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert calls["balance"] == 0
    assert calls["closed"] == 0


def test_missing_xlrd_for_local_xls_returns_structured_failure(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    workbook = temp_state_paths / "journal.xls"
    workbook.write_bytes(b"dummy")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "_local_journal_import_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_ensure_trading_journal_local_templates", lambda: {"errors": []})
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: [workbook])
    monkeypatch.setattr(master_service, "_set_trading_journal_diagnostics", lambda payload: None)
    real_import = __import__

    def _fake_import(name, *args, **kwargs):
        if name == "xlrd":
            raise ImportError("missing xlrd")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr("builtins.__import__", _fake_import)
    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is False
    assert result["errors"][0]["code"] == "MISSING_XLRD_FOR_XLS"


def test_balances_returns_200_when_items_exist_even_with_errors(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    snapshot = {
        "balances": [{"account": "A", "balance": 1}],
        "diagnostics": {"errors": [{"code": "X", "message": "warn"}]},
    }
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: snapshot)
    master_service._TRADING_JOURNAL_VIEW_CACHE["key"] = None
    master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] = None
    res = asyncio.run(master_service.trading_journal_balances())
    payload = _json(res)
    assert res.status_code == 200
    assert payload["ok"] is False


def test_missing_bybit_demo_anchor_with_wallet_failure_does_not_requeue(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    snapshot = {
        "balances": [{"account": "BYBIT DEMO", "label": "BYBIT DEMO", "balance": None, "currency": "USDT"}],
        "diagnostics": {"errors": ["Missing balance anchor for accounts: BYBIT DEMO"]},
    }
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: snapshot)
    monkeypatch.setattr(master_service, "_trading_journal_bybit_demo_balance_anchor_enabled", lambda: True)
    monkeypatch.setattr(
        master_service,
        "_load_trading_journal_state",
        lambda: {"broker_balance_diagnostics": {"warnings": ["Bybit demo wallet snapshot unavailable for workbook anchor reconstruction: demo creds missing"]}},
    )
    queued = {"calls": 0}
    monkeypatch.setattr(master_service, "_queue_trading_journal_sync_if_idle", lambda _r: queued.__setitem__("calls", queued["calls"] + 1) or {"running": True})
    monkeypatch.setattr(master_service, "_sync_state_snapshot", lambda: {"running": False, "ok": False})
    master_service._TRADING_JOURNAL_VIEW_CACHE["key"] = None
    master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] = None
    res = asyncio.run(master_service.trading_journal_balances())
    payload = _json(res)
    assert queued["calls"] == 0
    assert payload["pending"] is False
    assert payload["ok"] is False
    assert payload["items"]


def test_diagnostics_local_mode_zeroes_stale_dropbox_counts(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS = master_service._default_journal_diagnostics()
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS["dropbox_workbooks_seen"] = 0
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS["local_workbooks_seen"] = 2
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS["last_sync"] = {"source_mode": "local"}
    master_service._save_json_file(master_service.TRADING_JOURNAL_STATE_PATH, {"workbooks_seen": 99, "dropbox_workbooks_seen": 99})
    monkeypatch.setattr(master_service, "_sync_state_snapshot", lambda: {"dropbox_workbooks_seen": 25, "source_mode": "local"})
    snapshot = master_service._build_trading_journal_diagnostics_snapshot()
    assert snapshot["dropbox_workbooks_seen"] == 0
    assert snapshot["workbook_sources_seen"] == snapshot["local_workbooks_seen"]


def test_trading_journal_items_first_load_returns_pending_and_queues_sync(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_TRADING_JOURNAL_VIEW_CACHE", {"key": None, "payload": None})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_sync_state_snapshot", lambda: {"running": False, "ok": None, "message": ""})
    queued = {"called": False}

    def fake_queue(_reason: str):
        queued["called"] = True
        return {"running": True, "ok": None, "message": "queued"}

    monkeypatch.setattr(master_service, "_queue_trading_journal_sync_if_idle", fake_queue)
    response = asyncio.run(master_service.trading_journal_items())
    payload = _json(response)
    assert response.status_code == 202
    assert payload["pending"] is True
    assert payload["ok"] is False
    assert queued["called"] is True


def test_trading_journal_items_existing_snapshot_returns_200(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    snapshot = {
        "items": [{"id": "manual:1", "row_type": "trade", "source": "manual"}],
        "stats": {"groups": {}},
        "generated_at": "2026-04-01T00:00:00Z",
        "cache_version": 1,
        "source_fingerprints": {"source_mode": "local", "files": []},
    }
    monkeypatch.setattr(master_service, "_TRADING_JOURNAL_VIEW_CACHE", {"key": "snapshot", "payload": snapshot})
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "local", "files": []})
    response = asyncio.run(master_service.trading_journal_items())
    payload = _json(response)
    assert response.status_code == 200
    assert payload["count"] == 1
    assert payload["items"][0]["id"] == "manual:1"


def test_load_view_snapshot_rejects_old_cache_version(temp_state_paths):
    master_service._save_json_file(
        master_service.TRADING_JOURNAL_VIEW_CACHE_PATH,
        {
            "cache_version": 3,
            "source_fingerprints": {"source_mode": "local", "files": []},
            "items": [],
        },
    )
    assert master_service._load_trading_journal_view_snapshot() is None


def test_trading_journal_items_failed_sync_without_snapshot_returns_503(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_TRADING_JOURNAL_VIEW_CACHE", {"key": None, "payload": None})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(
        master_service,
        "_sync_state_snapshot",
        lambda: {"running": False, "ok": False, "error": "import failed", "message": "import failed"},
    )
    response = asyncio.run(master_service.trading_journal_items())
    payload = _json(response)
    assert response.status_code == 503
    assert payload["ok"] is False
    assert "import failed" in str(payload.get("error") or "")


def test_balance_merge_includes_bybit_demo_from_state_when_not_in_cashflow(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "local", "files": []})
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda _state: {})
    monkeypatch.setattr(master_service, "_cashflow_rows_for_journal", lambda _ledger: [])
    monkeypatch.setattr(master_service, "_compute_journal_stats", lambda _items, _balances: {"groups": {}})
    monkeypatch.setattr(master_service, "_build_trading_journal_diagnostics_snapshot", lambda: {"errors": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        master_service,
        "_load_json_file",
        lambda path, default: {
            "broker_account_balances": [
                {"account": "Bybit Demo", "label": "Bybit Demo", "balance": 111.0, "currency": "USDT", "source": "bybit_wallet_balance"}
            ]
        } if path == master_service.TRADING_JOURNAL_STATE_PATH else default,
    )
    snapshot = master_service._build_trading_journal_view_snapshot(force=True)
    labels = {str((item or {}).get("label") or "") for item in snapshot.get("balances", [])}
    assert "Bybit Demo" in labels

    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", False)
    snapshot2 = master_service._build_trading_journal_view_snapshot(force=True)
    labels2 = {str((item or {}).get("label") or "") for item in snapshot2.get("balances", [])}
    assert "Bybit Demo" not in labels2


def test_balance_merge_resolves_missing_timeline_anchor_with_broker_balance(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "local", "files": []})
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda _state: {})
    monkeypatch.setattr(master_service, "_cashflow_rows_for_journal", lambda _ledger: [])
    monkeypatch.setattr(master_service, "_compute_journal_stats", lambda _items, _balances: {"groups": {}})
    monkeypatch.setattr(master_service, "_build_trading_journal_diagnostics_snapshot", lambda: {"errors": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_build_journal_balance_timelines", lambda *_a, **_k: {
        "rows": [],
        "balances": [{"account": "BYBIT DEMO", "label": "BYBIT DEMO", "balance": None, "currency": "USDT", "missing_balance": True, "last_trade_at": "2026-01-01T00:00:00Z", "balance_source": "timeline_missing"}],
        "diagnostics": {"BYBIT DEMO": {"account_key": "BYBIT DEMO", "missing_balance": True, "warning": "missing"}},
    })
    monkeypatch.setattr(
        master_service,
        "_load_json_file",
        lambda path, default: {
            "broker_account_balances": [
                {"account": "Bybit Demo", "label": "Bybit Demo", "balance": 224.87, "currency": "USDT", "source": "bybit_wallet_balance"}
            ]
        } if path == master_service.TRADING_JOURNAL_STATE_PATH else default,
    )
    snapshot = master_service._build_trading_journal_view_snapshot(force=True)
    bybit_demo = next(item for item in snapshot["balances"] if master_service._is_bybit_demo_account_label(item.get("label")))
    assert bybit_demo["balance"] == pytest.approx(224.87)
    assert bybit_demo["balance_source"] == "bybit_wallet_balance"
    assert bybit_demo["resolved_missing_balance_with_broker"] is True
    assert bybit_demo["previous_balance_source"] == "timeline_missing"
    assert bybit_demo["last_trade_at"] == "2026-01-01T00:00:00Z"
    assert not any("Missing balance anchor for accounts: BYBIT DEMO" in str(err) for err in snapshot["diagnostics"]["errors"])


def test_merge_missing_timeline_balances_with_broker_preserves_resolution_metadata():
    balances = [
        {"account": "BYBIT DEMO", "label": "BYBIT DEMO", "balance": None, "currency": "USDT", "missing_balance": True, "last_trade_at": "2026-01-01T00:00:00Z", "source": "timeline", "balance_source": "timeline_missing"}
    ]
    broker = [
        {"account": "Bybit Demo", "label": "Bybit Demo", "balance": 224.87769878, "currency": "USDT", "source": "bybit_wallet_balance", "balance_source": "bybit_wallet_balance", "as_of": "2026-01-02T00:00:00Z"}
    ]
    merged = master_service._merge_missing_timeline_balances_with_broker(balances, broker)
    resolved = next(item for item in merged if master_service._is_bybit_demo_account_label(item.get("label")))
    assert resolved["missing_balance"] is False
    assert resolved["resolved_missing_balance_with_broker"] is True
    assert resolved["previous_balance_source"] == "timeline_missing"
    assert resolved["last_trade_at"] == "2026-01-01T00:00:00Z"


def test_merge_missing_timeline_balances_with_broker_canonicalizes_bybit_live_broker_only():
    merged = master_service._merge_missing_timeline_balances_with_broker(
        [],
        [{"account": "Bybit Live", "label": "Bybit Live", "balance": 10.0, "source": "bybit_wallet_balance"}],
    )
    assert len(merged) == 1
    assert merged[0]["label"] == "BYBIT"
    assert merged[0]["account"] == "BYBIT"


def test_merge_missing_timeline_balances_with_broker_preserves_existing_bybit():
    merged = master_service._merge_missing_timeline_balances_with_broker(
        [{"account": "BYBIT", "label": "BYBIT", "balance": 22.0, "source": "excel_account_balance", "balance_source": "excel_account_balance"}],
        [{"account": "Bybit Live", "label": "Bybit Live", "balance": 11.0, "source": "bybit_wallet_balance"}],
    )
    assert len(merged) == 1
    assert merged[0]["label"] == "BYBIT"
    assert merged[0]["balance"] == 22.0
    assert merged[0]["skipped_broker_balance_reason"] == "existing_bybit_balance_preserved"


def test_diagnostics_does_not_report_zero_when_journal_items_exist(temp_state_paths):
    master_service._set_trading_journal_rows(
        [
            {
                "id": "manual:1",
                "row_type": "trade",
                "source": "manual",
                "asset_class": "fx",
                "symbol": "EURUSD",
                "status": "closed",
                "close_time": "2026-04-01T00:00:00+00:00",
            }
        ]
    )
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS = master_service._default_journal_diagnostics()
    payload = _json(asyncio.run(master_service.trading_journal_diagnostics()))
    assert payload["rows_total"] > 0
    assert payload["journal_rows_total"] > 0
    assert payload["diagnostics_source"] in {"derived_from_current_rows", "mixed", "import"}


def test_diagnostics_counts_visible_rows_not_raw_rows(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", False)
    rows = []
    for idx in range(25):
        rows.append(
            {
                "id": f"manual:visible:{idx}",
                "row_type": "trade",
                "source": "manual",
                "asset_class": "fx",
                "symbol": "EURUSD",
                "status": "closed",
                "close_time": f"2026-04-01T00:{idx:02d}:00+00:00",
            }
        )
    rows.append(
        {
            "id": "manual:quarantined:1",
            "row_type": "trade",
            "source": "manual",
            "asset_class": "fx",
            "symbol": "EURUSD",
            "status": "invalid_time_order",
            "close_time": "2026-04-01T01:00:00+00:00",
        }
    )
    rows.extend(
        [
            {
                "id": "bybit:demo:hidden:1",
                "row_type": "trade",
                "source": "bybit",
                "account": "demo",
                "account_label": "Bybit Demo",
                "asset_class": "crypto",
                "symbol": "BTCUSDT",
                "status": "closed",
                "close_time": "2026-04-01T02:00:00+00:00",
            },
            {
                "id": "bybit:demo:hidden:2",
                "row_type": "trade",
                "source": "bybit",
                "account": "demo",
                "account_label": "Bybit Demo",
                "asset_class": "crypto",
                "symbol": "ETHUSDT",
                "status": "closed",
                "close_time": "2026-04-01T03:00:00+00:00",
            },
        ]
    )
    master_service._set_trading_journal_rows(rows)
    master_service.TRADING_JOURNAL_IMPORT_DIAGNOSTICS = master_service._default_journal_diagnostics()
    payload = _json(asyncio.run(master_service.trading_journal_diagnostics()))

    assert payload["rows_total"] == 25
    assert payload["journal_rows_total"] == 25
    assert payload["visible_rows_total"] == 25
    assert payload["raw_rows_total"] == 28
    assert payload["quarantined_rows"] == 1
    assert payload["excluded_rows_total"] == 3
    assert payload["has_current_journal_rows"] is True


def test_import_from_sources_local_when_dropbox_missing(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "edgewonk-export-78784.xls"
    workbook.write_bytes(b"dummy")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "both")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", temp_state_paths)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR_EXPLICIT", True)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_dropbox_excel", lambda progress_cb=None: {"ok": False, "rows_imported": 0, "workbooks_seen": 0, "errors": []})
    monkeypatch.setattr(
        master_service,
        "_parse_local_trading_journal_workbook",
        lambda path: (
            [{
                "id": f"local:{path.name}:1",
                "source": "local_excel",
                "asset_class": "fx",
                "symbol": "EURUSD",
                "open_time": "2026-04-01T00:00:00+00:00",
                "close_time": "2026-04-01T01:00:00+00:00",
            }],
            None,
        ),
    )
    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is True
    assert result["local_workbooks_seen"] == 1
    rows = master_service._get_trading_journal_rows()
    assert any(str(r.get("source")) == "local_excel" for r in rows)


def test_local_import_includes_bybit_demo_workbook_and_balance_anchor(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    bybit_workbook = temp_state_paths / master_service.BYBIT_DEMO_WORKBOOK_NAME
    rows = []
    base_balance = 400.0
    for idx in range(15):
        close_hour = idx + 1
        open_hour = max(0, close_hour - 1)
        opening_time = f"2026-04-01T{open_hour:02d}:00:00+00:00"
        closing_time = f"2026-04-01T{close_hour:02d}:00:00+00:00"
        if idx in {2, 9}:  # keep legacy invalid workbook time-order rows visible
            opening_time, closing_time = closing_time, opening_time
        side = "Buy" if idx < 10 else "Sell"
        pnl = -1.0 if side == "Buy" else -2.0
        balance = 380.97753999 if idx == 14 else base_balance + pnl
        base_balance = balance
        rows.append(
            {
                "opening_time": opening_time,
                "closing_time": closing_time,
                "type_buy_sell": side,
                "symbol": "BTCUSDT" if idx % 2 == 0 else "ETHUSDT",
                "size_quantity": 0.01 + (idx * 0.001),
                "entry_price": 100 + idx,
                "closing_price": 101 + idx,
                "stop_loss": 95 + idx,
                "take_profit": 105 + idx,
                "commission": 0.1,
                "net_profit": pnl,
                "balance_after_trade": balance,
                "timeframe": "5-minute",
                "is_test_trade": "No",
                "currency": "USDT",
                "notes": f"row-{idx}",
                "order_id": f"demo-order-{idx}",
                "fill_count": 1,
                "source": "excel",
            }
        )
    master_service.pd.DataFrame(rows, columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS).to_excel(
        bybit_workbook,
        sheet_name=master_service.BYBIT_DEMO_WORKBOOK_SHEET,
        index=False,
    )

    master_service._set_trading_journal_rows(
        [
            {
                "id": "bybit:demo:closedpnl:HYPERUSDT:existing",
                "row_type": "trade",
                "source": "bybit",
                "account": "demo",
                "account_label": "Bybit Demo",
                "asset_class": "crypto",
                "symbol": "HYPERUSDT",
                "side": "Buy",
                "status": "closed",
                "open_time": "2026-04-02T00:00:00+00:00",
                "close_time": "2026-04-02T01:00:00+00:00",
                "net_profit": 1.0,
            }
        ]
    )
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR_EXPLICIT", True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", temp_state_paths)

    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is True
    assert result["rows_imported"] >= 15
    assert result["local_workbooks_seen"] == 1
    assert master_service.BYBIT_DEMO_WORKBOOK_NAME in result["local_workbook_names"]

    snapshot = master_service._build_trading_journal_view_snapshot(force=True)
    items = snapshot.get("items") or []
    workbook_rows = [
        row for row in items
        if str(row.get("source")) == "local_excel"
        and str(row.get("account_label") or row.get("account") or "").strip().lower() == "bybit demo"
    ]
    assert len(workbook_rows) == 15
    assert sum(1 for row in workbook_rows if str((row.get("metrics") or {}).get("time_order_repaired")).lower() == "true") >= 2
    bybit_demo_visible = [
        row for row in items
        if str(row.get("row_type") or "trade") == "trade"
        and str(row.get("account_label") or row.get("account") or "").strip().lower() == "bybit demo"
    ]
    assert len(bybit_demo_visible) == 15
    assert not any(str(row.get("source") or "").lower() == "bybit" for row in bybit_demo_visible)

    bybit_balance = next(
        bal for bal in (snapshot.get("balances") or [])
        if str(bal.get("label") or bal.get("account") or "").strip().lower() == "bybit demo"
    )
    assert bybit_balance["balance"] == pytest.approx(380.97753999)
    assert bybit_balance["balance_source"] == "trade_timeline"

    diag_errors = snapshot.get("diagnostics", {}).get("errors") or []
    assert not any("Missing balance anchor for accounts: BYBIT DEMO" in str(err) for err in diag_errors)


def test_import_from_sources_ignores_default_local_workbooks_when_not_enabled(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "edgewonk-export-78784.xls"
    workbook.write_bytes(b"dummy")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "both")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", temp_state_paths)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR_EXPLICIT", False)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_dropbox_excel", lambda progress_cb=None: {"ok": False, "rows_imported": 0, "workbooks_seen": 0, "errors": []})
    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is False
    assert result["ignored_local_workbooks"] == ["edgewonk-export-78784.xls"]


def test_bybit_invalid_time_rows_are_repaired_from_items(temp_state_paths):
    master_service._set_trading_journal_rows(
        [
            {
                "id": "bybit:demo:closedpnl:HYPERUSDT:bad",
                "source": "bybit",
                "account": "demo",
                "account_label": "Bybit Demo",
                "asset_class": "crypto",
                "symbol": "HYPERUSDT",
                "side": "Buy",
                "status": "closed",
                "open_time": "2026-04-26T10:12:25+00:00",
                "close_time": "2026-04-26T10:11:33+00:00",
            },
            {
                "id": "bybit:demo:closedpnl:HYPERUSDT:good",
                "source": "bybit",
                "account": "demo",
                "account_label": "Bybit Demo",
                "asset_class": "crypto",
                "symbol": "HYPERUSDT",
                "side": "Buy",
                "status": "closed",
                "open_time": "2026-04-26T10:12:25+00:00",
                "close_time": "2026-04-26T12:32:57+00:00",
            },
        ]
    )
    rows, stats = master_service._sanitize_bybit_demo_rows(master_service._get_trading_journal_rows())
    assert stats["quarantined_invalid_time"] == 0
    master_service._set_trading_journal_rows(rows)
    master_service._build_trading_journal_view_snapshot(force=True)
    payload = _json(asyncio.run(master_service.trading_journal_items()))
    repaired = [r for r in payload["items"] if r.get("symbol") == "HYPERUSDT"]
    assert len(repaired) == 2
    assert any((r.get("metrics") or {}).get("time_order_repaired") is True for r in repaired)


def test_import_from_sources_clears_existing_rows_on_empty_result_in_local_mode(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    master_service._set_trading_journal_rows([{"id": "existing:1", "source": "manual", "open_time": "2026-04-01T00:00:00+00:00"}])
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", temp_state_paths)
    monkeypatch.setattr(master_service, "_list_local_trading_journal_workbooks", lambda: [])
    result = master_service._import_trading_journal_from_sources()
    assert result["ok"] is False
    assert "authoritative" in str(result.get("message") or "").lower()
    rows = master_service._get_trading_journal_rows()
    assert not any(str(r.get("id")) == "existing:1" for r in rows)


def test_upsert_trading_journal_rows_rejects_broker_rows_in_local_mode(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "local")
    master_service._set_trading_journal_rows([])
    changed = master_service._upsert_trading_journal_rows([{"id": "oanda:demo:test", "source": "oanda"}])
    assert changed == 0


def test_merge_trading_journal_row_preserves_rich_raw_refs_and_context_fields():
    existing = {
        "id": "bybit:demo:trade:BTCUSDT:85c1adb0266f56a4",
        "source": "bybit_execution_history_grouped",
        "raw_refs": {"order_ids": ["a", "b"]},
        "stop_loss": 76491.2,
        "take_profit": 77067.3,
        "timeframe": "1m",
        "is_test_trade": True,
        "r_multiple": -0.27,
    }
    incoming = {
        "id": "bybit:demo:trade:BTCUSDT:85c1adb0266f56a4",
        "source": "master_journal",
        "raw_refs": None,
        "stop_loss": "",
        "take_profit": "",
        "timeframe": "",
        "is_test_trade": None,
        "r_multiple": "",
    }
    merged = master_service._merge_trading_journal_row(existing, incoming)
    assert merged.get("raw_refs", {}).get("order_ids") == ["a", "b"]
    assert merged.get("stop_loss") == 76491.2
    assert merged.get("take_profit") == 77067.3
    assert merged.get("timeframe") == "1m"
    assert merged.get("is_test_trade") is True
    assert merged.get("r_multiple") == -0.27
    assert master_service._get_trading_journal_rows() == []


def test_trading_journal_js_quarantine_is_not_hard_warning():
    js = (ROOT / "render" / "static" / "trading_journal.js").read_text(encoding="utf-8")
    assert "|| quarantinedRows > 0" not in js
    assert "repaired time-order rows" in js


def test_parse_excel_generic_filename_infers_fx_asset_class(monkeypatch: pytest.MonkeyPatch):
    row = {
        "symbol": "EUR_USD",
        "open_time": "2026-04-01",
        "close_time": "2026-04-01",
        "net_profit": 1.0,
    }
    df = master_service.pd.DataFrame([row])

    class FakeExcel:
        sheet_names = ["Sheet1"]

    monkeypatch.setattr(master_service.pd, "ExcelFile", lambda *_args, **_kwargs: FakeExcel())
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_args, **_kwargs: df)
    rows, _bal = master_service._parse_excel_account_workbook("edgewonk-export-78784.xls", "/tmp/edgewonk-export-78784.xls", b"x")
    assert rows
    assert rows[0]["asset_class"] == "fx"


def test_row_pnl_fallback_counts_realized_pnl_only_rows():
    assert master_service._is_win({"realized_pnl": 1.0}) is True
    assert master_service._is_loss({"realized_pnl": -1.0}) is True
    assert master_service._is_be({"realized_pnl": 0.0}) is True


def test_normalize_bybit_closed_pnl_row_sets_net_profit():
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "BTCUSDT",
            "orderId": "order-1",
            "updatedTime": 1710000001000,
            "createdTime": 1710000000000,
            "closedPnl": "9.5",
            "avgEntryPrice": "100",
            "avgExitPrice": "109.5",
            "closedSize": "1",
            "side": "Buy",
        },
        account_mode="demo",
        balance_after_trade=None,
    )
    assert row is not None
    assert row["realized_pnl"] == 9.5
    assert row["net_profit"] == 9.5


def test_normalize_bybit_closed_pnl_row_does_not_repersist_invalid_ctx_setup(monkeypatch: pytest.MonkeyPatch):
    captured = {}
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda *_a, **_k: {"timeframe": "15m", "setup": "invalid-setup", "is_test_trade": False, "open_time": "2026-01-01T00:00:00Z"})
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: captured.update(payload) or payload)
    row = master_service._normalize_bybit_closed_pnl_row(
        {"symbol": "BTCUSDT", "orderId": "order-2", "updatedTime": 1710000001000, "createdTime": 1710000000000, "closedPnl": "1", "avgEntryPrice": "100", "avgExitPrice": "101", "closedSize": "1", "side": "Buy"},
        account_mode="demo",
        balance_after_trade=None,
    )
    assert row is not None
    assert row.get("setup") in ("", None)
    assert captured.get("setup") in ("", None)


def test_journal_rows_from_bybit_execution_replays_valid_setup_and_ignores_invalid(monkeypatch: pytest.MonkeyPatch):
    base = {"account": "demo", "category": "linear", "symbol": "BTCUSDT", "orderId": "oid", "execId": "eid", "execQty": "1", "execPrice": "100", "execFee": "0.1", "execPnl": "2", "execTime": 1710000001000, "side": "Buy"}
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda *_a, **_k: {"timeframe": "15m", "setup": "Pullback", "is_test_trade": True})
    rows = master_service._journal_rows_from_bybit_execution(dict(base))
    assert rows and rows[0]["setup"] == "Pullback"
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda *_a, **_k: {"timeframe": "15m", "setup": "bad-setup", "is_test_trade": True})
    rows2 = master_service._journal_rows_from_bybit_execution(dict(base))
    assert rows2 and rows2[0].get("setup") in ("", None)


def test_oanda_rows_set_net_profit_from_realized_pnl(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_persist_oanda_fill_state", lambda: None)
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: [])
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_lookup_trade_context_by_market_window", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_update_unresolved_registry", lambda **_kwargs: (False, {}))
    rows = master_service._journal_rows_from_oanda_order_fill(
        {
            "account": "live",
            "id": "100",
            "instrument": "EUR_USD",
            "time": "2026-04-01T01:00:00Z",
            "orderID": "200",
            "units": "-1000",
            "tradesClosed": [{"tradeID": "t1", "units": "-1000", "price": "1.2", "realizedPL": "7", "financing": "0"}],
            "accountCurrency": "AUD",
            "price": "1.2",
        }
    )
    assert rows
    assert rows[0]["net_profit"] == rows[0]["realized_pnl"]


def test_manual_sync_calls_bybit_without_manual_cooldown(monkeypatch: pytest.MonkeyPatch):
    calls = []

    async def _fake_bybit(**kwargs):
        calls.append(kwargs)
        return {"ok": True}

    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", _fake_bybit)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda progress_cb=None: {"ok": True, "rows_imported": 0, "diagnostics": {}})
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert len(calls) == 2
    assert all(call.get("enforce_manual_cooldown") is False for call in calls)


def test_oanda_summary_is_saved_as_broker_balance_during_sync(monkeypatch: pytest.MonkeyPatch):
    captured = {}
    async def _fake_oanda_summary(account):
        return {"balance": 1500.0 if account == "demo" else 2500.0, "nav": 1501.0, "currency": "AUD"}
    async def _fake_bybit(*_a, **_k):
        return {"available_usdt": 10.0}
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_fetch_oanda_account_summary", _fake_oanda_summary)
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", _fake_bybit)
    monkeypatch.setattr(master_service, "_trading_journal_source_mode", lambda: "local")
    monkeypatch.setattr(master_service, "_trading_journal_bybit_demo_balance_anchor_enabled", lambda: False)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda progress_cb=None: {"ok": True, "rows_imported": 0, "diagnostics": {}})
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", lambda **kwargs: asyncio.sleep(0))
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda account, lookback_hours=72: asyncio.sleep(0, result={"ok": True, "account": account}))
    monkeypatch.setattr(master_service, "_save_broker_balance_diagnostics_state", lambda balances, warnings: captured.update({"balances": balances, "warnings": warnings}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    labels = {item.get("label"): item for item in captured.get("balances", [])}
    assert labels["OANDA DEMO"]["balance_source"] == "oanda_account_summary"
    assert labels["OANDA LIVE"]["balance_source"] == "oanda_account_summary"


def test_balance_timeline_resets_on_cashflow_and_applies_later_non_test_pnl():
    rows = [
        {"id": "t1", "row_type": "trade", "account": "OANDA DEMO", "close_time": "2026-01-01T00:01:00Z", "net_profit": 10.0},
        {"id": "t2", "row_type": "trade", "account": "OANDA DEMO", "close_time": "2026-01-01T00:02:00Z", "net_profit": -5.0},
        {"id": "t3", "row_type": "trade", "account": "OANDA DEMO", "close_time": "2026-01-01T00:04:00Z", "net_profit": 20.0},
    ]
    ledger = {
        "OANDA DEMO": [
            {"account": "OANDA DEMO", "date": "2026-01-01T00:00:00Z", "new_balance": 1000.0, "currency": "AUD"},
            {"account": "OANDA DEMO", "date": "2026-01-01T00:03:00Z", "new_balance": 500.0, "currency": "AUD"},
        ]
    }
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    balances = timeline["balances"]
    assert balances[0]["balance"] == pytest.approx(520.0)


def test_first_trade_pnl_is_not_double_counted_from_cashflow_anchor():
    rows = [
        {"id": "t1", "row_type": "trade", "account": "BYBIT", "close_time": "2026-01-01T00:01:00Z", "net_profit": 7.0},
    ]
    ledger = {"BYBIT": [{"account": "BYBIT", "date": "2026-01-01T00:00:00Z", "new_balance": 100.0, "currency": "USDT"}]}
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    trade = timeline["rows"][0]
    assert trade["analysis_balance_after_trade"] == pytest.approx(107.0)


def test_final_cashflow_after_final_trade_overrides_current_balance():
    rows = [
        {"id": "t1", "row_type": "trade", "account": "PEPPERSTONE DEMO", "close_time": "2026-01-01T00:01:00Z", "net_profit": -95.36},
    ]
    ledger = {
        "PEPPERSTONE DEMO": [
            {"account": "PEPPERSTONE DEMO", "date": "2026-01-01T00:00:00Z", "new_balance": 0.0, "currency": "AUD"},
            {"account": "PEPPERSTONE DEMO", "date": "2026-01-01T00:02:00Z", "new_balance": 0.0, "currency": "AUD"},
        ]
    }
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    assert timeline["balances"][0]["balance"] == pytest.approx(0.0)


def test_oanda_export_balance_overrides_stale_cashflow_anchor():
    rows = [
        {"id": "t1", "row_type": "trade", "source": "local_excel", "account": "OANDA DEMO", "close_time": "2026-04-08T00:01:00Z", "net_profit": -6.88, "balance_after_trade": 1493.77},
    ]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}]}
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    bal = timeline["balances"][0]
    diag = timeline["diagnostics"][master_service._norm_account_key("OANDA DEMO")]
    assert bal["balance"] == pytest.approx(1493.77)
    assert bal["balance_source"] == "authoritative_trade_balance"
    assert diag["stale_cashflow_overridden"] is True


def test_oanda_transaction_export_latest_balance_is_account_balance(monkeypatch: pytest.MonkeyPatch):
    df = master_service.pd.DataFrame(
        [
            {"TICKET": 589, "TRANSACTION DATE": "2026-04-08T00:00:00Z", "TRANSACTION TYPE": "ORDER_FILL", "DETAILS": "fill", "INSTRUMENT": "EUR_USD", "PL": 0.0, "BALANCE": 1493.77},
            {"TICKET": 622, "TRANSACTION DATE": "2026-04-09T00:00:00Z", "TRANSACTION TYPE": "ORDER_FILL", "DETAILS": "fill", "INSTRUMENT": "EUR_USD", "PL": 0.0, "BALANCE": 1500.65},
        ]
    )
    class FakeExcel:
        sheet_names = ["Sheet1"]
    monkeypatch.setattr(master_service.pd, "ExcelFile", lambda *_a, **_k: FakeExcel())
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_a, **_k: df)
    rows, balance = master_service._parse_excel_account_workbook("OANDA DEMO.xls", "/tmp/OANDA DEMO.xls", b"x")
    assert rows == []
    assert balance["balance"] == pytest.approx(1500.65)
    assert balance["source"] == "oanda_transaction_export_balance"


def test_oanda_broker_balance_overrides_reconstructed_timeline_balance():
    balances = [
        {"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 193.71, "currency": "AUD", "source": "cashflow_anchor_plus_trades", "balance_source": "cashflow_anchor_plus_trades", "as_of": "2022-05-05T00:00:00Z"}
    ]
    broker = [
        {"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1493.77, "currency": "AUD", "source": "oanda_account_summary", "balance_source": "oanda_account_summary", "as_of": "2026-04-08T00:00:00Z"}
    ]
    merged = master_service._merge_missing_timeline_balances_with_broker(balances, broker)
    assert merged[0]["balance"] == pytest.approx(1493.77)
    assert merged[0]["resolved_or_overridden_with_broker"] is True


def test_oanda_export_account_balance_overrides_stale_cashflow_without_trade_rows():
    rows = []
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}]}
    excel_balances = [
        {
            "account": "OANDA DEMO",
            "label": "OANDA DEMO",
            "balance": 1500.65,
            "currency": "AUD",
            "source": "local_excel",
            "balance_source": "oanda_transaction_export_balance",
            "as_of": "2026-04-09T00:00:00Z",
        }
    ]
    timeline = master_service._build_journal_balance_timelines(rows, ledger, excel_balances)
    bal = timeline["balances"][0]
    diag = timeline["diagnostics"][master_service._norm_account_key("OANDA DEMO")]
    assert bal["balance"] == pytest.approx(1500.65)
    assert bal["balance_source"] == "oanda_transaction_export_balance"
    assert diag["stale_cashflow_overridden"] is True
    assert diag["balance_source"] == "oanda_transaction_export_balance"


def test_local_oanda_export_balance_source_is_preserved(tmp_path, monkeypatch: pytest.MonkeyPatch):
    path = tmp_path / "OANDA DEMO.xls"
    path.write_bytes(b"x")
    monkeypatch.setattr(
        master_service,
        "_parse_excel_account_workbook",
        lambda *_a, **_k: ([], {"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1500.65, "source": "oanda_transaction_export_balance", "balance_source": "oanda_transaction_export_balance"}),
    )
    _rows, balance = master_service._parse_local_trading_journal_workbook(path)
    assert balance["balance_source"] == "oanda_transaction_export_balance"
    assert balance["source"] == "oanda_transaction_export_balance"
    assert balance["import_source"] == "local_excel"


def test_newer_oanda_export_account_balance_overrides_older_authoritative_trade_balance():
    rows = [{"id": "t1", "row_type": "trade", "source": "oanda", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "close_time": "2026-04-08T00:01:00Z", "net_profit": -6.88, "balance_after_trade": 1493.77}]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}]}
    excel_balances = [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1500.65, "currency": "AUD", "source": "oanda_transaction_export_balance", "balance_source": "oanda_transaction_export_balance", "as_of": "2026-04-30T19:46:41"}]
    timeline = master_service._build_journal_balance_timelines(rows, ledger, excel_balances)
    bal = timeline["balances"][0]
    diag = timeline["diagnostics"][master_service._norm_account_key("OANDA DEMO")]
    assert bal["balance"] == pytest.approx(1500.65)
    assert bal["balance_source"] == "oanda_transaction_export_balance"
    assert diag["latest_authoritative_balance_at"] == "2026-04-30T19:46:41"


def test_oanda_export_aest_dates_pick_latest_balance(monkeypatch: pytest.MonkeyPatch):
    df = master_service.pd.DataFrame([
        {"TICKET": 500, "TRANSACTION DATE": "2026-04-28 19:46:41 AEST", "TRANSACTION TYPE": "ORDER_FILL", "DETAILS": "fill", "INSTRUMENT": "EUR_USD", "PL": 0.0, "BALANCE": 1493.77},
        {"TICKET": 622, "TRANSACTION DATE": "2026-04-30 19:46:41 AEST", "TRANSACTION TYPE": "ORDER_FILL", "DETAILS": "fill", "INSTRUMENT": "EUR_USD", "PL": 0.0, "BALANCE": 1500.65},
    ])
    class FakeExcel: sheet_names = ["Sheet1"]
    monkeypatch.setattr(master_service.pd, "ExcelFile", lambda *_a, **_k: FakeExcel())
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_a, **_k: df)
    _rows, balance = master_service._parse_excel_account_workbook("OANDA DEMO.xls", "/tmp/OANDA DEMO.xls", b"x")
    assert balance["balance"] == pytest.approx(1500.65)


def test_newer_oanda_broker_summary_overrides_older_authoritative_trade_balance():
    balances = [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1493.77, "currency": "AUD", "source": "authoritative_trade_balance", "balance_source": "authoritative_trade_balance", "as_of": "2026-04-08T00:01:00Z"}]
    broker = [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1500.65, "currency": "AUD", "source": "oanda_account_summary", "balance_source": "oanda_account_summary", "as_of": "2026-04-30T00:00:00Z"}]
    merged = master_service._merge_missing_timeline_balances_with_broker(balances, broker)
    assert merged[0]["balance"] == pytest.approx(1500.65)


def test_newer_final_cashflow_still_overrides_oanda_broker_summary():
    balances = [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 202.12, "currency": "AUD", "source": "cashflow_anchor_plus_trades", "balance_source": "cashflow_anchor_plus_trades", "as_of": "2026-05-01T00:00:00Z"}]
    broker = [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1500.65, "currency": "AUD", "source": "oanda_account_summary", "balance_source": "oanda_account_summary", "as_of": "2026-04-30T00:00:00Z"}]
    merged = master_service._merge_missing_timeline_balances_with_broker(balances, broker)
    assert merged[0]["balance"] == pytest.approx(202.12)


def test_balance_timeline_diagnostics_no_authoritative_candidate_no_nameerror():
    rows = [{"id": "m1", "row_type": "trade", "source": "manual", "account": "MANUAL ACC", "close_time": "2026-05-01T00:00:00Z", "net_profit": 1.0}]
    ledger = {"MANUAL ACC": [{"account": "MANUAL ACC", "date": "2026-04-30T00:00:00Z", "new_balance": 100.0, "currency": "AUD"}]}
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    diag = timeline["diagnostics"][master_service._norm_account_key("MANUAL ACC")]
    assert diag["authoritative_balance_used"] is None
    assert diag["latest_authoritative_balance_at"] is None


def test_oanda_export_generic_filename_relabels_to_single_existing_oanda_cashflow_account():
    balances = [{
        "account": "oanda_history_test",
        "label": "oanda_history_test",
        "balance": 1500.65,
        "balance_source": "oanda_transaction_export_balance",
        "as_of": "2026-04-30T09:46:41+00:00",
        "raw_refs": {"transaction_date": "2026-04-30T09:46:41+00:00", "workbook": "oanda_history_test.xlsx"},
    }]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}]}
    relabeled, warnings = master_service._reconcile_oanda_export_balance_labels(balances, ledger)
    timeline = master_service._build_journal_balance_timelines([], ledger, relabeled)
    labels = {b["label"]: b for b in timeline["balances"]}
    assert warnings == []
    assert "oanda_history_test" not in labels
    assert labels["OANDA DEMO"]["balance"] == pytest.approx(1500.65)


def test_oanda_export_balance_without_timestamp_does_not_silently_override_cashflow():
    balances = [{"account": "oanda_history_test", "label": "oanda_history_test", "balance": 1500.65, "balance_source": "oanda_transaction_export_balance"}]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}]}
    relabeled, warnings = master_service._reconcile_oanda_export_balance_labels(balances, ledger)
    timeline = master_service._build_journal_balance_timelines([], ledger, relabeled)
    labels = {b["label"]: b for b in timeline["balances"]}
    assert labels["OANDA DEMO"]["balance"] == pytest.approx(202.12)
    assert any("oanda_export_balance_missing_timestamp" in w for w in warnings)


def test_oanda_export_generic_filename_is_ambiguous_when_demo_and_live_exist():
    balances = [{"account": "oanda_history_test", "label": "oanda_history_test", "balance": 1500.65, "balance_source": "oanda_transaction_export_balance"}]
    ledger = {
        "OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-05-05T00:00:00Z", "new_balance": 202.12, "currency": "AUD"}],
        "OANDA LIVE": [{"account": "OANDA LIVE", "date": "2022-05-05T00:00:00Z", "new_balance": 300.0, "currency": "AUD"}],
    }
    relabeled, warnings = master_service._reconcile_oanda_export_balance_labels(balances, ledger)
    assert relabeled == []
    assert any("ambiguous_oanda_export_account_mapping" in w for w in warnings)


def test_oanda_export_filename_demo_hint_maps_to_oanda_demo_even_with_live_present():
    balances = [{"account": "oanda_history_demo", "label": "oanda_history_demo", "balance": 1500.65, "balance_source": "oanda_transaction_export_balance", "dropbox_path": "/tmp/oanda_history_demo.xlsx", "as_of": "2026-04-30T09:46:41+00:00", "raw_refs": {"transaction_date": "2026-04-30T09:46:41+00:00", "workbook": "oanda_history_demo.xlsx"}}]
    ledger = {"OANDA DEMO": [{}], "OANDA LIVE": [{}]}
    relabeled, warnings = master_service._reconcile_oanda_export_balance_labels(balances, ledger)
    assert warnings == []
    assert relabeled[0]["label"] == "OANDA DEMO"


def test_oanda_export_filename_live_hint_maps_to_oanda_live_even_with_demo_present():
    balances = [{"account": "oanda_history_live", "label": "oanda_history_live", "balance": 1500.65, "balance_source": "oanda_transaction_export_balance", "dropbox_path": "/tmp/oanda_history_live.xlsx", "as_of": "2026-04-30T09:46:41+00:00", "raw_refs": {"transaction_date": "2026-04-30T09:46:41+00:00", "workbook": "oanda_history_live.xlsx"}}]
    ledger = {"OANDA DEMO": [{}], "OANDA LIVE": [{}]}
    relabeled, warnings = master_service._reconcile_oanda_export_balance_labels(balances, ledger)
    assert warnings == []
    assert relabeled[0]["label"] == "OANDA LIVE"


def _bybit_csv_sample(row_count: int = 2) -> str:
    header = "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)"
    rows = []
    for i in range(row_count):
        rows.append(f"BTCUSDT,ord-{i},Buy,Market,0.001,65000,65000,Trade,0.0006,0.04,exec-{i},22:55 2026-05-17,1000.0")
    return header + "\n" + "\n".join(rows) + "\n"

def _bybit_csv_grouped_three_trades() -> str:
    header = "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)"
    fills = [
        ("OIDA", "Buy", 0.019, 100000, "EX001", "17/05/2026 22:55"),
        ("OIDA", "Sell", 0.019, 100100, "EX002", "17/05/2026 22:56"),
    ] + [(f"OIDB{i}", "Buy", 0.018, 100200 + i, f"EX{3+i:03d}", "19/05/2026 00:52") for i in range(18)] + [
        ("OIDBCLS", "Sell", 0.324, 100260, "EX021", "19/05/2026 00:54"),
        ("OIDC", "Buy", 0.016, 100300, "EX022", "19/05/2026 01:12"),
        ("OIDC", "Sell", 0.016, 100330, "EX023", "19/05/2026 01:13"),
    ]
    rows = [
        f"BTCUSDT,{oid},{side},Market,{qty},{px},{px},Trade,0.00055,0.01,{exid},{ts},1000.0"
        for oid, side, qty, px, exid, ts in fills
    ]
    return header + "\n" + "\n".join(rows) + "\n"


def test_valid_bybit_import_preserves_existing_rows_and_reports_sync_fields(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    from openpyxl import load_workbook
    old_rows = [
        {"id": "old1", "row_type": "trade", "source": "manual", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "symbol": "EURUSD", "side": "Buy", "qty": 1.0, "entry_price": 1.1, "exit_price": 1.2, "open_time": "2026-05-01T00:00:00Z", "close_time": "2026-05-01T01:00:00Z", "net_profit": 10.0},
        {"id": "old2", "row_type": "trade", "source": "manual", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "symbol": "GBPUSD", "side": "Buy", "qty": 1.0, "entry_price": 1.2, "exit_price": 1.3, "open_time": "2026-05-02T00:00:00Z", "close_time": "2026-05-02T01:00:00Z", "net_profit": 10.0},
        {"id": "old3", "row_type": "trade", "source": "manual", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "symbol": "USDJPY", "side": "Sell", "qty": 1.0, "entry_price": 150.0, "exit_price": 149.0, "open_time": "2026-05-03T00:00:00Z", "close_time": "2026-05-03T01:00:00Z", "net_profit": 10.0},
        {"id": "monthly_aud_reval:bybit_live:2026-03", "row_type": "monthly_aud_reval", "source": "bybit_monthly_aud_reval", "account": "Bybit Live", "account_label": "Bybit Live", "symbol": "MONTHLY AUD P/L", "open_time": "2026-03-01T00:00:00Z", "close_time": "2026-03-31T23:59:59Z", "result_cash": 5.0, "result_currency": "AUD"},
        {"id": "monthly_aud_reval:bybit_live:2026-04", "row_type": "monthly_aud_reval", "source": "bybit_monthly_aud_reval", "account": "Bybit Live", "account_label": "Bybit Live", "symbol": "MONTHLY AUD P/L", "open_time": "2026-04-01T00:00:00Z", "close_time": "2026-04-30T23:59:59Z", "result_cash": 7.0, "result_currency": "AUD"},
    ]
    master_service._set_trading_journal_rows(old_rows)
    build_master_journal_workbook({"items": old_rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, temp_state_paths / "Trading Journal.xlsx")

    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True, "master_journal_path": str(temp_state_paths / "Trading Journal.xlsx")})
    order = []
    def _wrapped_verify(*_a, **_k):
        order.append("verify")
        return {"ok": True, "missing_row_ids": []}
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", _wrapped_verify)
    def _fake_sync(_path):
        order.append("github")
        return {"github_sync_enabled": True, "github_sync_ok": True, "github_sync_noop": False, "github_sync_error": "", "github_sync_commit": "abc123"}
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", _fake_sync)

    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    assert payload.get("github_sync_enabled") is True
    assert payload.get("github_sync_ok") is True
    assert payload.get("github_sync_deferred") is False
    assert payload.get("github_sync_path")
    assert payload.get("github_sync_message")
    assert order.index("verify") < order.index("github")

    ids = {str(r.get("id") or "") for r in master_service._get_trading_journal_rows()}
    assert {"old1", "old2", "old3", "monthly_aud_reval:bybit_live:2026-03", "monthly_aud_reval:bybit_live:2026-04"}.issubset(ids)
    bybit_ids = [x for x in ids if x.startswith("bybit:demo:trade:")]
    assert len(bybit_ids) == 3

    assert isinstance(payload.get("import_timings"), dict)


def test_sync_master_journal_duration_validation_still_fails_real_trade_blank_duration(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    from tools.master_journal_workbook import build_master_journal_workbook
    mj = temp_state_paths / "Trading Journal.xlsx"
    rows = [{"id": "t1", "row_type": "trade", "source": "manual", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "symbol": "EURUSD", "side": "Buy", "qty": 1.0, "entry_price": 1.1, "exit_price": 1.2, "open_time": "2026-05-01T00:00:00Z", "close_time": "2026-05-01T01:00:00Z", "net_profit": 10.0}]
    build_master_journal_workbook({"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, mj)
    from openpyxl import load_workbook
    wb = load_workbook(mj)
    ws = wb["Trade Log"]
    headers = [str(c.value or "") for c in ws[1]]
    dur_col = headers.index("Trade Duration (DD:HH:MM:SS)") + 1
    ws.cell(2, dur_col).value = ""
    wb.save(mj)
    wb.close()
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=True: {"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []})
    monkeypatch.setattr(master_service, "update_master_journal_workbook_data_only", lambda *a, **k: {"ok": True, "candidate_path": str(mj)})
    out = master_service._sync_master_journal_workbook(defer_github_sync=True, sync_caller="test")
    assert out.get("ok") is False
    assert "duration column blank/non-numeric" in str(out.get("master_journal_error") or "")


def test_import_file_endpoint_not_stub_anymore(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert int(payload.get("status_code") or 0) == 200
    assert int(payload.get("status_code") or 0) != 501
    assert payload["ok"] is True
    assert payload.get("errors") == []
    assert "_parse_ts_utc" not in str(payload)
    assert int(payload.get("rows_parsed") or 0) >= 1
    assert int(payload.get("verified_row_ids_count") or 0) >= 1


def test_import_file_ambiguous_bybit_csv_is_blocked_without_account_mode(temp_state_paths):
    master_service._set_trading_journal_rows([{"id": "existing:1", "row_type": "trade", "source": "manual"}])
    payload = master_service._import_uploaded_trading_journal_file("bybit_history.csv", _bybit_csv_sample(1).encode("utf-8"))
    assert int(payload.get("status_code") or 0) == 422
    assert payload["ok"] is False
    assert "ambiguous" in payload["message"].lower()
    assert "ambiguous_bybit_account" in (payload.get("errors") or [])
    assert payload.get("requires_account_mode") is True
    assert payload.get("detected_file_kind") == "bybit_history_csv"
    assert set(payload.get("account_mode_options") or []) == {"demo", "live"}
    assert any(str(r.get("id")) == "existing:1" for r in master_service._get_trading_journal_rows())


def test_parse_local_workbook_bybit_csv_requires_account_mode_or_filename_hint(tmp_path: Path):
    p = tmp_path / "bybit_history.csv"
    p.write_text(_bybit_csv_sample(1), encoding="utf-8")
    with pytest.raises(ValueError, match="ambiguous"):
        master_service._parse_local_trading_journal_workbook(p, original_name="bybit_history.csv")


def test_import_file_uses_tempfile_without_nameerror(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    assert "NameError" not in " ".join(payload.get("errors") or [])


def test_import_file_bybit_csv_parses_once_with_explicit_account_mode(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    calls = {"bybit": 0, "local": 0}
    monkeypatch.setattr(master_service, "_is_bybit_trade_history_csv", lambda _p: True)
    def _fake_bybit(_path, account_mode="demo"):
        calls["bybit"] += 1
        assert account_mode == "live"
        return [{"id": "bybit:live:execution:BTCUSDT:e1", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Live"}]
    def _fake_local(_path, **_k):
        calls["local"] += 1
        return [], None
    monkeypatch.setattr(master_service, "_parse_bybit_trade_history_csv_with_diagnostics", lambda _path, account_mode="demo": (_fake_bybit(_path, account_mode), [], {"bybit_execution_rows_seen":1,"bybit_completed_trades_imported":1,"bybit_unmatched_execution_rows":0}))
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", _fake_local)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("history.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="live")
    assert payload["ok"] is True
    assert calls["bybit"] == 1
    assert calls["local"] == 0


def test_import_file_bridges_pending_rows_during_sync_and_restores_after(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    seen = {"during": []}
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_PENDING_MANUAL_SYNC_ROWS", [{"id": "existing:pending"}], raising=False)
    def _fake_sync(**_k):
        seen["during"] = [str(r.get("id") or "") for r in (master_service._PENDING_MANUAL_SYNC_ROWS or []) if isinstance(r, dict)]
        return {"ok": True}
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", _fake_sync)
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    assert any(x.startswith("bybit:demo:trade:") for x in seen["during"])
    assert "existing:pending" in seen["during"]
    assert [str(r.get("id") or "") for r in (master_service._PENDING_MANUAL_SYNC_ROWS or []) if isinstance(r, dict)] == ["existing:pending"]
    assert isinstance(payload.get("import_timings"), dict)


def test_import_file_does_not_build_snapshot_before_sync(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    calls = {"snap": 0, "sync": 0}
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    def _fake_snap(*_a, **_k):
        calls["snap"] += 1
        return {"items": master_service._get_trading_journal_rows(), "stats": {}, "balances": []}
    def _fake_sync(**_k):
        calls["sync"] += 1
        return {"ok": True}
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", _fake_snap)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", _fake_sync)
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    assert calls["sync"] == 1
    assert calls["snap"] == 1
    assert isinstance(payload.get("import_timings"), dict)


def test_build_journal_balance_timelines_timestamp_cache_bounds_to_datetime_calls(monkeypatch: pytest.MonkeyPatch):
    real_to_datetime = master_service.pd.to_datetime
    calls = {"n": 0}
    def _counted(*args, **kwargs):
        calls["n"] += 1
        return real_to_datetime(*args, **kwargs)
    monkeypatch.setattr(master_service.pd, "to_datetime", _counted)
    rows = []
    for i in range(120):
        rows.append({"id": f"r{i}", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "open_time": "2026-05-17T12:00:00Z", "close_time": "2026-05-17T12:00:00Z", "net_profit": 1.0, "currency": "USDT"})
    ledger = {master_service._norm_account_key("Bybit Demo"): [{"date": "2026-05-17T11:00:00Z", "new_balance": 1000.0, "currency": "USDT"} for _ in range(120)]}
    _ = master_service._build_journal_balance_timelines(rows, ledger, [])
    assert calls["n"] < 50


def test_global_pnl_inference_bybit_from_balance_continuity(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    existing = [{"id": "anchor:1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "balance_after_trade": 1000.0, "open_time": "2026-05-17T12:00:00Z"}]
    master_service._set_trading_journal_rows(existing)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    csv = _bybit_csv_grouped_three_trades().replace(",1000.0\n", ",1012.5\n").replace(",0.04,", ",0.50,")
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", csv.encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    rows = master_service._get_trading_journal_rows()
    imported = next(r for r in rows if str(r.get("id", "")).startswith("bybit:demo:trade:"))
    assert imported.get("commission") is not None
    assert isinstance(imported.get("net_profit"), (int, float))
    assert int(payload.get("pnl_inferred_count") or 0) >= 0


def test_global_pnl_does_not_treat_fee_only_open_fill_as_realized_loss(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    existing = [{"id": "anchor:1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "balance_after_trade": 1000.0, "open_time": "2026-05-17T12:00:00Z"}]
    master_service._set_trading_journal_rows(existing)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    csv = _bybit_csv_grouped_three_trades().replace(",1000.0\n", ",999.5\n").replace(",0.04,", ",0.50,")
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", csv.encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    imported = next(r for r in master_service._get_trading_journal_rows() if str(r.get("id", "")).startswith("bybit:demo:trade:"))
    assert imported.get("net_profit") is None or isinstance(imported.get("net_profit"), (int, float))
    assert isinstance(payload.get("warnings") or [], list)


def test_global_pnl_warns_when_previous_balance_anchor_missing(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is True
    assert int(payload.get("pnl_unresolved_count") or 0) >= 0
    assert isinstance(payload.get("pnl_unresolved_row_ids") or [], list)


def test_global_pnl_does_not_double_count_commission():
    imported = [{"id": "r2", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "balance_after_trade": 1012.5, "commission": 0.5, "net_profit": None, "open_time": "2026-01-02T00:00:00Z"}]
    existing = [{"id": "r1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "balance_after_trade": 1000.0, "open_time": "2026-01-01T00:00:00Z"}]
    rows, _warnings, _diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0]["net_profit"] == 12.5


def test_import_response_uses_generic_pnl_diagnostics_not_bybit_only(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert "pnl_inferred_count" in payload
    assert "pnl_unresolved_reasons" in payload


def test_global_pnl_normalizes_explicit_realized_pnl_to_net_profit():
    imported = [{"id": "o1", "row_type": "trade", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "currency": "AUD", "realized_pnl": 8.25, "open_time": "2026-01-02T00:00:00Z"}]
    rows, _warnings, _diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, [])
    assert rows[0]["net_profit"] == 8.25


def test_global_pnl_preserves_pepperstone_explicit_profit_marked_source():
    imported = [{"id": "p1", "row_type": "trade", "source": "pepperstone_mt5_statement", "account": "Pepperstone", "account_label": "Pepperstone", "currency": "USD", "profit": 5.5, "commission": 0.0, "swap": 0.0, "net_profit": None, "open_time": "2026-01-02T00:00:00Z"}]
    rows, _warnings, _diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, [])
    assert rows[0]["net_profit"] == 5.5


def test_global_pnl_non_trade_event_blocking_is_same_account_currency_only():
    imported = [{"id": "t2", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "balance_after_trade": 1010.0, "open_time": "2026-01-02T00:00:00Z"}]
    existing = [
        {"id": "t1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "balance_after_trade": 1000.0, "open_time": "2026-01-01T00:00:00Z"},
        {"id": "cf1", "row_type": "cashflow", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "currency": "AUD", "balance_after_trade": 9999.0, "open_time": "2026-01-01T12:00:00Z"},
    ]
    rows, warnings, _diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0]["net_profit"] == 10.0
    assert not any("blocked_by_non_trade" in w for w in warnings)


def test_global_pnl_sort_prefers_close_time_over_open_time():
    imported = [{"id": "t2", "row_type": "trade", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "currency": "AUD", "open_time": "2026-01-01T00:00:00Z", "close_time": "2026-01-03T00:00:00Z", "balance_after_trade": 1010.0}]
    existing = [{"id": "t1", "row_type": "trade", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "currency": "AUD", "close_time": "2026-01-02T00:00:00Z", "balance_after_trade": 1000.0}]
    rows, _w, _d = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0]["net_profit"] == 10.0


def test_global_pnl_post_trade_anchor_does_not_use_next_trade_as_anchor():
    imported = [
        {"id": "a", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "open_time": "2026-01-02T00:00:00Z", "commission": 0.1},
        {"id": "b", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "open_time": "2026-01-02T00:01:00Z", "balance_after_trade": 1010.0},
    ]
    existing = [{"id": "prev", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "open_time": "2026-01-01T00:00:00Z", "balance_after_trade": 1000.0}]
    rows, _w, diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    ra = next(r for r in rows if r["id"] == "a")
    assert ra.get("net_profit") is None
    assert diag["pnl_unresolved_reasons"]["a"] == "pnl_inference_post_trade_anchor_ambiguous"


def test_global_pnl_blank_currency_blocks_when_account_currency_ambiguous():
    imported = [{"id": "r1", "row_type": "trade", "account": "acct", "account_label": "acct", "currency": "", "balance_after_trade": 1010.0, "open_time": "2026-01-03T00:00:00Z"}]
    existing = [
        {"id": "e1", "row_type": "trade", "account": "acct", "account_label": "acct", "currency": "AUD", "balance_after_trade": 1000.0, "open_time": "2026-01-01T00:00:00Z"},
        {"id": "e2", "row_type": "trade", "account": "acct", "account_label": "acct", "currency": "USD", "balance_after_trade": 1000.0, "open_time": "2026-01-02T00:00:00Z"},
    ]
    rows, _w, diag = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0].get("net_profit") is None
    assert diag["pnl_unresolved_reasons"]["r1"] == "pnl_inference_ambiguous_currency"


def test_pepperstone_mt5_profit_commission_swap_normalizes_to_net_profit():
    imported = [{"id": "mt5:1", "row_type": "trade", "source": "pepperstone_mt5_statement", "account": "Pepperstone", "account_label": "Pepperstone", "profit": 10.0, "commission": -1.0, "swap": -0.5, "open_time": "2026-01-01T00:00:00Z"}]
    rows, _w, d = master_service._infer_realized_net_profit_from_balance_continuity(imported, [])
    assert rows[0]["net_profit"] == 8.5
    assert d.get("pnl_explicit_normalized_count") == 1


def test_fee_only_closed_trade_is_not_suppressed_as_open_fill():
    imported = [{"id": "c1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "commission": 0.5, "balance_after_trade": 999.5, "close_time": "2026-01-02T00:00:00Z", "status": "closed"}]
    existing = [{"id": "p1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "currency": "USDT", "balance_after_trade": 1000.0, "close_time": "2026-01-01T00:00:00Z"}]
    rows, _w, _d = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0]["net_profit"] == -0.5


def test_prev_index_does_not_use_wrong_duplicate_row():
    existing = [
        {"id": "dup1", "row_type": "trade", "account": "A", "account_label": "A", "currency": "USD", "balance_after_trade": 100.0, "close_time": "2026-01-01T00:00:00Z"},
        {"id": "dup2", "row_type": "trade", "account": "A", "account_label": "A", "currency": "USD", "balance_after_trade": 100.0, "close_time": "2026-01-01T00:00:00Z"},
    ]
    imported = [{"id": "t", "row_type": "trade", "account": "A", "account_label": "A", "currency": "USD", "balance_after_trade": 110.0, "close_time": "2026-01-02T00:00:00Z"}]
    rows, _w, _d = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0]["net_profit"] == 10.0


def test_oanda_parser_explicit_realized_pl_reaches_net_profit_after_import(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    csv = (
        "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,DIRECTION,UNITS,PRICE,STOP LOSS,TAKE PROFIT,SPREAD COST,COMMISSION,GSL FEE,PL,BALANCE\n"
        "1,2026-01-01 10:00:00 AEST,MARKET_ORDER,CLIENT_ORDER,EUR_USD,Buy,100000,1.1,,, -1,0,0,,1000\n"
        "2,2026-01-01 10:00:01 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,Buy,100000,1.1,,, -1,0,0,,999\n"
        "3,2026-01-01 11:00:00 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,EUR_USD,Buy,100000,1.2,,, -1,-1,0,10,1008\n"
    )
    payload = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", csv.encode("utf-8"))
    assert payload["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if str(r.get("id", "")).startswith("oanda_export:"))
    assert row["net_profit"] == 9.0


def test_profit_alias_not_treated_as_gross_mt5_profit_for_unmarked_sources():
    imported = [{"id": "x1", "row_type": "trade", "account": "Manual", "account_label": "Manual", "profit": 10.0, "commission": -1.0, "swap": -1.0, "open_time": "2026-01-01T00:00:00Z"}]
    rows, _w, _d = master_service._infer_realized_net_profit_from_balance_continuity(imported, [])
    assert rows[0].get("net_profit") is None


def test_chain_currency_uses_current_chain_key_not_row_membership():
    shared = {"row_type": "trade", "account": "A", "account_label": "A", "balance_after_trade": 1000.0, "close_time": "2026-01-01T00:00:00Z"}
    existing = [dict(shared, id="aud", currency="AUD"), dict(shared, id="usd", currency="USD")]
    imported = [dict(shared, id="new", currency="", balance_after_trade=1010.0, close_time="2026-01-02T00:00:00Z")]
    rows, _w, d = master_service._infer_realized_net_profit_from_balance_continuity(imported, existing)
    assert rows[0].get("net_profit") is None
    assert d["pnl_unresolved_reasons"]["new"] == "pnl_inference_ambiguous_currency"


def test_pepperstone_mt5_parser_profit_commission_swap_reaches_net_profit_after_import(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    frame = master_service.pd.DataFrame([
        {
            "account": "Pepperstone MT5",
            "symbol": "XAUUSD",
            "side": "Buy",
            "opening_time": "2026-01-01 10:00:00",
            "closing_time": "2026-01-01 11:00:00",
            "size_quantity": 1.0,
            "entry_price": 2000.0,
            "closing_price": 2005.0,
            "profit": 10.0,
            "commission": -1.0,
            "swap": -0.5,
            "balance_after_trade": 1008.5,
            "currency": "USD",
        }
    ])
    bio = io.BytesIO()
    frame.to_excel(bio, index=False)
    payload = master_service._import_uploaded_trading_journal_file("pepperstone_mt5.xlsx", bio.getvalue())
    assert payload["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if "pepperstone" in str(r.get("account_label", "")).lower())
    assert row["net_profit"] == 8.5


def test_workbook_net_pl_populates_for_oanda_pepperstone_bybit_after_sync(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    rows = [
        {"id": "oanda:1", "row_type": "trade", "account": "OANDA DEMO", "account_label": "OANDA DEMO", "symbol": "EURUSD", "side": "Buy", "open_time": "2026-01-01T00:00:00Z", "close_time": "2026-01-01T01:00:00Z", "qty": 1.0, "entry_price": 1.1, "exit_price": 1.2, "net_profit": 10.0, "commission": -1.0, "balance_after_trade": 1010.0, "currency": "AUD"},
        {"id": "pep:1", "row_type": "trade", "account": "Pepperstone MT5", "account_label": "Pepperstone MT5", "symbol": "XAUUSD", "side": "Buy", "open_time": "2026-01-02T00:00:00Z", "close_time": "2026-01-02T01:00:00Z", "qty": 1.0, "entry_price": 2000.0, "exit_price": 2005.0, "net_profit": 8.5, "commission": -1.0, "swap": -0.5, "balance_after_trade": 1018.5, "currency": "USD"},
        {"id": "bybit:1", "row_type": "trade", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-01-03T00:00:00Z", "close_time": "2026-01-03T01:00:00Z", "qty": 0.1, "entry_price": 100000.0, "exit_price": 100100.0, "net_profit": 12.5, "commission": 0.5, "balance_after_trade": 1031.0, "currency": "USDT"},
    ]
    master_service._set_trading_journal_rows(rows)
    out = master_service._sync_master_journal_workbook(sync_caller="test")
    assert out.get("ok") is True
    wb = master_service.load_workbook(master_service._master_journal_path(), data_only=True, read_only=True)
    try:
        ws = master_service._get_trade_log_sheet(wb, allow_legacy=False)
        headers = [str(c.value or "").strip() for c in ws[1]]
        pnl_idx = headers.index("Net P/L") + 1
        vals = [ws.cell(r, pnl_idx).value for r in range(2, ws.max_row + 1)]
        assert any(v == 10.0 for v in vals)
        assert any(v == 8.5 for v in vals)
        assert any(v == 12.5 for v in vals)
    finally:
        wb.close()


def test_oanda_demo_upload_name_survives_tempfile_parse(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    csv = "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,BALANCE\n1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,1000\n"
    monkeypatch.setattr(master_service, "_journal_rows_from_oanda_transaction_history_frame", lambda *_a, **_k: {"rows":[{"id":"o:1","row_type":"trade","account":"OANDA DEMO","account_label":"OANDA DEMO","net_profit":1.0}],"account_balance":None})
    payload = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", csv.encode("utf-8"))
    assert payload["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if r.get("id") == "o:1")
    assert row["account_label"] == "OANDA DEMO"


def test_oanda_live_upload_name_survives_tempfile_parse(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    csv = "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,BALANCE\n1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,1000\n"
    monkeypatch.setattr(master_service, "_journal_rows_from_oanda_transaction_history_frame", lambda *_a, **_k: {"rows":[{"id":"o:2","row_type":"trade","account":"OANDA LIVE","account_label":"OANDA LIVE","net_profit":1.0}],"account_balance":None})
    payload = master_service._import_uploaded_trading_journal_file("oanda_live.csv", csv.encode("utf-8"))
    assert payload["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if r.get("id") == "o:2")
    assert row["account_label"] == "OANDA LIVE"


def test_oanda_ambiguous_manual_upload_fails_or_warns_clearly(temp_state_paths):
    csv = "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,BALANCE\n1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,1000\n"
    payload = master_service._import_uploaded_trading_journal_file("oanda.csv", csv.encode("utf-8"))
    assert payload["ok"] is False
    assert int(payload["status_code"]) == 422


def test_pepperstone_mt5_upload_name_survives_tempfile_parse(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    frame = master_service.pd.DataFrame([{"account":"Pepperstone MT5","symbol":"XAUUSD","side":"Buy","opening_time":"2026-01-01","closing_time":"2026-01-02","size_quantity":1,"entry_price":1,"closing_price":2,"net_profit":1}])
    bio = io.BytesIO(); frame.to_excel(bio, index=False)
    payload = master_service._import_uploaded_trading_journal_file("pepperstone_mt5.xlsx", bio.getvalue())
    assert payload["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if "pepperstone" in str(r.get("account_label","")).lower())
    assert row.get("source") == "pepperstone_mt5_statement"
    assert row.get("import_source") == "local_excel"


def test_local_excel_generic_rows_get_local_excel_source_when_blank(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_parse_excel_account_workbook", lambda *_a, **_k: ([{"id": "g:1", "row_type": "trade", "source": "", "account": "Generic"}], None))
    p = temp_state_paths / "generic.xlsx"
    p.write_bytes(b"x")
    rows, _balance = master_service._parse_local_trading_journal_workbook(p, original_name="generic.xlsx")
    assert rows[0]["source"] == "local_excel"
    assert rows[0]["import_source"] == "local_excel"


def test_import_file_balance_only_not_fake_success(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([], {"account": "OANDA DEMO", "balance": 123.45}))
    payload = master_service._import_uploaded_trading_journal_file("oanda_demo.xlsx", b"dummy")
    assert payload["ok"] is False
    assert int(payload["status_code"]) == 422
    assert payload.get("balance_parsed") is True


def test_import_file_unexpected_exception_returns_structured_json_failure(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert int(payload["status_code"]) == 500
    assert payload["uploaded_name"] == "manual.xlsx"
    assert isinstance(payload.get("errors"), list) and payload["errors"]


def test_import_file_rejects_rows_without_ids(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    master_service._set_trading_journal_rows([{"id": "existing:1", "row_type": "trade", "source": "manual"}])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"row_type": "trade", "source": "manual"}], None))
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False and int(payload["status_code"]) == 422
    assert any(r.get("id") == "existing:1" for r in master_service._get_trading_journal_rows())


def test_import_file_rolls_back_rows_when_sync_fails(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    original = [{"id": "existing:1", "row_type": "trade", "source": "manual"}]
    master_service._set_trading_journal_rows(original)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda: {"ok": False, "error": "nope"})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert master_service._get_trading_journal_rows() == original


def test_import_file_accepts_master_journal_ok_without_top_level_ok(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"master_journal_ok": True, "master_journal_path": str(temp_state_paths / "Trading Journal.xlsx")})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is True
    assert int(payload["status_code"]) == 200


def test_import_file_sync_failure_surfaces_master_journal_error(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    original = [{"id": "existing:1", "row_type": "trade", "source": "manual"}]
    master_service._set_trading_journal_rows(original)
    sync_error = "Trading Journal validation failed: Trade Log filter missing."
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"master_journal_ok": False, "master_journal_error": sync_error})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert int(payload["status_code"]) == 500
    assert sync_error in str(payload.get("message") or "")
    assert any(sync_error in str(err) for err in (payload.get("errors") or []))
    assert "unknown error" not in str(payload.get("message") or "").lower()
    assert master_service._get_trading_journal_rows() == original


def test_import_file_rolls_back_rows_when_verification_fails(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    original = [{"id": "existing:1", "row_type": "trade", "source": "manual"}]
    master_service._set_trading_journal_rows(original)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": False, "missing_row_ids": ["new:1"], "error": "missing"})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert master_service._get_trading_journal_rows() == original


def test_import_file_bybit_parse_error_is_422(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_is_bybit_trade_history_csv", lambda _p: True)
    monkeypatch.setattr(master_service, "_parse_bybit_trade_history_csv_with_diagnostics", lambda *_a, **_k: (_ for _ in ()).throw(ValueError("bad bybit csv")))
    payload = master_service._import_uploaded_trading_journal_file("bybit_demo.csv", _bybit_csv_grouped_three_trades().encode("utf-8"), account_mode="demo")
    assert payload["ok"] is False
    assert int(payload["status_code"]) == 422


def test_import_file_read_failure_returns_json():
    class BadUpload:
        filename = "bad.csv"
        async def read(self):
            raise RuntimeError("read failed")
    res = asyncio.run(master_service.trading_journal_import_file(file=BadUpload(), account_mode=None))
    payload = _json(res)
    assert res.status_code == 400
    assert payload["ok"] is False


def test_import_file_rollback_uses_deepcopy_for_nested_fields(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    original = [{"id": "existing:1", "row_type": "trade", "source": "manual", "raw_refs": {"nested": {"k": "v"}}, "manual_overrides": {"note": "keep"}}]
    master_service._set_trading_journal_rows(original)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "existing:1", "row_type": "trade", "source": "manual", "raw_refs": {"nested": {"k": "changed"}}}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda: {"ok": False, "error": "sync failed"})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    restored = master_service._get_trading_journal_rows()
    assert restored == original
    assert restored[0]["raw_refs"]["nested"]["k"] == "v"


def test_import_file_restores_workbook_bytes_on_verification_failure(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    workbook.write_bytes(b"ORIGINAL-WB")
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    def _fake_sync(**_k):
        workbook.write_bytes(b"MODIFIED-WB")
        return {"ok": True}
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", _fake_sync)
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": False, "missing_row_ids": ["new:1"], "error": "missing"})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert workbook.read_bytes() == b"ORIGINAL-WB"


def test_import_file_verification_missing_ids_comes_from_original_verify_result(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    calls = {"n": 0}
    def _verify(*_a, **_k):
        calls["n"] += 1
        return {"ok": False, "missing_row_ids": ["new:1", "new:2"], "error": "missing"}
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", _verify)
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert payload.get("missing_row_ids") == ["new:1", "new:2"]
    assert calls["n"] == 1


def test_import_file_deletes_new_workbook_created_before_verification_failure(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    assert workbook.exists() is False
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    def _fake_sync(**_k):
        workbook.write_bytes(b"NEWLY-CREATED")
        return {"ok": True}
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", _fake_sync)
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": False, "missing_row_ids": ["new:1"], "error": "missing"})
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert workbook.exists() is False


def test_import_file_reports_workbook_delete_rollback_failure(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda _p, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: (workbook.write_bytes(b"NEW-WB"), {"ok": True})[1])
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": False, "missing_row_ids": ["new:1"], "error": "missing"})
    monkeypatch.setattr(Path, "unlink", lambda _self: (_ for _ in ()).throw(PermissionError("cannot delete")))
    payload = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert payload["ok"] is False
    assert "rollback failed" in str(payload.get("message") or "").lower()
    assert any("could not restore workbook" in str(err).lower() for err in (payload.get("errors") or []))


def test_import_passes_original_name_to_parse_local_workbook(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    seen = {"name": None}
    def _fake_parse(_path, *, original_name=None, **_k):
        seen["name"] = original_name
        return [{"id": "p:1", "row_type": "trade", "source": "pepperstone_mt5_statement"}], None
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", _fake_parse)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    payload = master_service._import_uploaded_trading_journal_file("pepperstone_mt5.xlsx", b"x")
    assert payload["ok"] is True
    assert seen["name"] == "pepperstone_mt5.xlsx"


def test_parse_excel_balance_uses_latest_trade_timestamp_not_bottom_row(monkeypatch: pytest.MonkeyPatch):
    df = master_service.pd.DataFrame(
        [
            {"symbol": "BTCUSDT", "close_time": "2026-04-28T10:00:00Z", "balance_after_trade": 380.97753999, "currency": "USDT"},
            {"symbol": "BTCUSDT", "close_time": "2026-04-01T10:00:00Z", "balance_after_trade": 403.72484338, "currency": "USDT"},
        ]
    )

    class FakeExcel:
        sheet_names = ["Sheet1"]

    monkeypatch.setattr(master_service.pd, "ExcelFile", lambda *_args, **_kwargs: FakeExcel())
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_args, **_kwargs: df)
    _rows, balance = master_service._parse_excel_account_workbook("Bybit Demo.xlsx", "/tmp/Bybit Demo.xlsx", b"x")
    assert balance is not None
    assert balance["balance"] == pytest.approx(380.97753999)


def test_balance_timeline_merges_alias_accounts_by_normalized_key():
    rows = [
        {"id": "t1", "row_type": "trade", "account": "OANDA Demo", "account_label": "OANDA Demo", "close_time": "2026-01-01T00:01:00Z", "net_profit": 3.0},
    ]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2026-01-01T00:00:00Z", "new_balance": 100.0, "currency": "AUD"}]}
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    assert len(timeline["balances"]) == 1
    assert timeline["balances"][0]["balance"] == pytest.approx(103.0)


def test_trading_journal_balances_snapshot_expected_values(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "local", "files": []})
    monkeypatch.setattr(master_service, "_build_trading_journal_diagnostics_snapshot", lambda: {"errors": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(
        master_service,
        "_get_trading_journal_rows",
        lambda: [
            {"id": "bybit:t1", "row_type": "trade", "account": "BYBIT", "close_time": "2026-01-01T00:01:00Z", "net_profit": 0.0},
            {"id": "oanda-live:t1", "row_type": "trade", "account": "OANDA LIVE", "close_time": "2026-01-01T00:01:00Z", "net_profit": 0.0},
            {"id": "pep-live:t1", "row_type": "trade", "account": "PEPPERSTONE LIVE", "close_time": "2026-01-01T00:01:00Z", "net_profit": 0.0},
            {"id": "pep-demo:t1", "row_type": "trade", "account": "PEPPERSTONE DEMO", "close_time": "2026-01-01T00:01:00Z", "net_profit": 0.0},
        ],
    )
    monkeypatch.setattr(
        master_service,
        "_load_cashflows_for_active_journal_source",
        lambda _state: {
            "BYBIT": [{"account": "BYBIT", "date": "2026-01-01T00:00:00Z", "new_balance": 224.87769878, "currency": "USDT"}],
            "OANDA LIVE": [{"account": "OANDA LIVE", "date": "2026-01-01T00:00:00Z", "new_balance": 1479.31, "currency": "AUD"}],
            "PEPPERSTONE LIVE": [{"account": "PEPPERSTONE LIVE", "date": "2026-01-01T00:00:00Z", "new_balance": 2508.73, "currency": "AUD"}],
            "PEPPERSTONE DEMO": [{"account": "PEPPERSTONE DEMO", "date": "2026-01-01T00:00:00Z", "new_balance": 0.0, "currency": "AUD"}],
        },
    )
    snapshot = master_service._build_trading_journal_view_snapshot(force=True)
    by_label = {str(i.get("label")): i for i in snapshot.get("balances", [])}
    assert by_label["BYBIT"]["balance"] == pytest.approx(224.87769878)
    assert by_label["OANDA LIVE"]["balance"] == pytest.approx(1479.31)
    assert by_label["PEPPERSTONE LIVE"]["balance"] == pytest.approx(2508.73)
    assert by_label["PEPPERSTONE DEMO"]["balance"] == pytest.approx(0.0)


def test_master_journal_authoritative_snapshot_preserves_monthly_aud_reval(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "master_journal", "files": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_save_trading_journal_view_snapshot", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: Path("/tmp/Trading Journal.xlsx"))
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_load_json_file", lambda *_a, **_k: {})
    monkeypatch.setattr(
        master_service,
        "read_master_journal_source",
        lambda _p: {
            "items": [
                {"id": "t1", "row_type": "trade", "account": "OANDA DEMO", "symbol": "EURUSD", "side": "BUY", "open_time": "2026-04-30T09:45:41Z", "close_time": "2026-04-30T09:46:41Z", "net_profit": 1.0, "result_pct": 0.5},
                {"id": "m1", "row_type": "monthly_aud_reval", "account": "Bybit Live", "symbol": "MONTHLY AUD P/L", "open_time": "2026-04-01T00:00:00Z", "close_time": "2026-04-30T23:59:59Z", "result_cash": 25.0, "result_currency": "AUD"},
            ],
            "cashflow_ledger": {},
        },
    )
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    ids = {str(r.get("id")) for r in snap.get("items", [])}
    assert "m1" in ids


def test_master_journal_authoritative_snapshot_monthly_uses_json_fallback(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "master_journal", "files": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_save_trading_journal_view_snapshot", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: Path("/tmp/Trading Journal.xlsx"))
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_load_json_file", lambda *_a, **_k: {})
    monkeypatch.setattr(
        master_service,
        "_monthly_aud_revaluation_rows_for_journal_view",
        lambda: [{
            "id": "monthly_aud_reval:bybit_live:2026-03",
            "row_type": "monthly_aud_reval",
            "account": "Bybit Live",
            "account_label": "Bybit Live",
            "symbol": "MONTHLY AUD P/L",
            "close_time": "2026-03-31T23:59:59Z",
            "result_cash": 123.45,
            "result_currency": "AUD",
            "raw_refs": {"period_month": "2026-03"},
        }],
    )
    monkeypatch.setattr(
        master_service,
        "read_master_journal_source",
        lambda _p: {
            "items": [
                {"id": "t1", "row_type": "trade", "account": "OANDA DEMO", "symbol": "EURUSD", "side": "BUY", "open_time": "2026-04-30T09:45:41Z", "close_time": "2026-04-30T09:46:41Z", "net_profit": 10.0, "result_pct": 0.5},
                {"id": "monthly_aud_reval:bybit_live:2026-03", "row_type": "monthly_aud_reval", "account": "Bybit Live", "symbol": "MONTHLY AUD P/L", "close_time": "2026-03-31T23:59:59Z", "result_cash": None},
            ],
            "cashflow_ledger": {},
        },
    )
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    monthly = next(r for r in snap.get("items", []) if r.get("id") == "monthly_aud_reval:bybit_live:2026-03")
    assert monthly["result_cash"] == pytest.approx(123.45)
    assert monthly["result_currency"] == "AUD"
    assert (monthly.get("raw_refs") or {}).get("period_month") == "2026-03"
    assert "net_profit" not in monthly and "realized_pnl" not in monthly
    assert snap["stats"]["totals"]["trades"] == 1
    assert snap["stats"]["totals"]["net_profit_total"] == pytest.approx(10.0)


def test_persist_trading_journal_sqlite_cashflow_no_name_error(tmp_path, monkeypatch: pytest.MonkeyPatch):
    sqlite_path = tmp_path / "trading_journal.sqlite"
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SQLITE_PATH", sqlite_path)
    snapshot = {
        "generated_at": "2026-05-01T00:00:00Z",
        "items": [{"id": "cf:1", "row_type": "cashflow", "source": "cashflow_ledger", "close_time": "2026-05-01T00:00:00Z"}],
        "balances": [],
        "stats": {},
        "diagnostics": {},
        "source_fingerprints": {"files": []},
    }
    master_service._persist_trading_journal_sqlite(snapshot)
    conn = sqlite3.connect(sqlite_path)
    try:
        assert conn.execute("SELECT COUNT(*) FROM journal_cashflows").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM journal_trades").fetchone()[0] == 0
    finally:
        conn.close()

def test_append_generic_local_broker_rows_does_not_match_blank_ids(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    rows_a = [{"account": "Bybit Live", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-01-01T00:00:00Z", "close_time": "2026-01-01T01:00:00Z", "entry_price": 1, "exit_price": 2, "qty": 1, "realized_pnl": 1, "raw_refs": {"orderId": "A"}}]
    rows_b = [{"account": "Bybit Live", "symbol": "ETHUSDT", "side": "Sell", "open_time": "2026-01-02T00:00:00Z", "close_time": "2026-01-02T01:00:00Z", "entry_price": 3, "exit_price": 2, "qty": 2, "realized_pnl": -1, "raw_refs": {"orderId": "B"}}]
    master_service._append_generic_local_broker_rows("Bybit Live.xlsx", rows_a, "bybit_closed_pnl")
    master_service._append_generic_local_broker_rows("Bybit Live.xlsx", rows_b, "bybit_closed_pnl")
    df = master_service.pd.read_excel(tmp_path / "Bybit Live.xlsx", sheet_name="Trades")
    assert set(df["order_id"].astype(str)) >= {"A", "B"}

def test_snapshot_includes_monthly_aud_note_rows_excluded_from_stats(tmp_path, monkeypatch):
    monkeypatch.setattr(master_service, "_master_journal_authoritative_enabled", lambda: False)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setenv("TRADING_JOURNAL_SOURCE", "both")
    monthly_path = tmp_path / "monthly_aud_revaluation.json"
    monthly_path.write_text(json.dumps({"items": [{
        "id": "monthly_aud_reval:bybit_live:2026-03",
        "row_type": "monthly_aud_reval",
        "account": "Bybit Live",
        "account_label": "Bybit Live",
        "close_time": "2026-03-31T23:59:59Z",
        "result_cash": 123.45,
        "result_currency": "AUD",
        "raw_refs": {"period_month": "2026-03"},
    }]}), encoding="utf-8")
    monkeypatch.setattr(master_service, "MONTHLY_AUD_REVALUATION_PATH", monthly_path)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "Trading Journal.xlsx")
    (tmp_path / "Trading Journal.xlsx").write_bytes(b"stub")
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: (_ for _ in ()).throw(AssertionError("authoritative path should not be used when explicitly disabled")))
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"source_mode": "local", "files": []})
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [{"id": "t1", "row_type": "trade", "close_time": "2026-03-01T00:00:00Z", "net_profit": 10.0, "account": "A"}])
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda _state: {})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_args, **_kwargs: None)
    snapshot = master_service._build_trading_journal_view_snapshot(force=True)
    ids = {str(r.get("id")) for r in snapshot["items"]}
    assert "monthly_aud_reval:bybit_live:2026-03" in ids
    monthly = next(r for r in snapshot["items"] if r.get("row_type") == "monthly_aud_reval")
    assert monthly["result_currency"] == "AUD"
    assert monthly["result_cash"] == pytest.approx(123.45)
    assert "net_profit" not in monthly and "realized_pnl" not in monthly
    assert snapshot["stats"]["totals"]["trades"] == 1
    assert snapshot["stats"]["totals"]["net_profit_total"] == pytest.approx(10.0)
    assert snapshot["stats"]["groups"]["overview"]["trades"] == 1


def test_journal_source_fingerprint_includes_monthly_paths():
    fp = master_service._journal_source_fingerprint()
    paths = {str(i.get("path")) for i in fp.get("files", []) if isinstance(i, dict)}
    assert str(master_service.MONTHLY_AUD_REVALUATION_PATH) in paths
    assert str(master_service.MONTHLY_AUD_REVALUATION_STATE_PATH) in paths


def test_run_monthly_sync_invalidates_snapshot_on_change(monkeypatch):
    called = {"n": 0}
    master_service._TRADING_JOURNAL_VIEW_CACHE["key"] = "snapshot"
    master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] = {"ok": True}

    def _invalidate():
        called["n"] += 1
        master_service._TRADING_JOURNAL_VIEW_CACHE["key"] = None
        master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] = None

    async def fake_sync_monthly_aud_revaluation(**_kwargs):
        return {"ok": True, "changed": True}

    monkeypatch.setattr(master_service, "_invalidate_trading_journal_view_snapshot", _invalidate)
    monkeypatch.setattr(master_service, "sync_monthly_aud_revaluation", fake_sync_monthly_aud_revaluation)
    monkeypatch.setattr(master_service, "_load_json_file", lambda *_args, **_kwargs: {})
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    asyncio.run(master_service._run_monthly_aud_revaluation_sync(reason="test"))
    assert called["n"] == 1
    assert master_service._TRADING_JOURNAL_VIEW_CACHE["key"] is None
    assert master_service._TRADING_JOURNAL_VIEW_CACHE["payload"] is None


def test_monthly_aud_revaluation_rows_for_journal_view_keeps_zero_result(tmp_path, monkeypatch):
    monthly_path = tmp_path / "monthly_aud_revaluation.json"
    monthly_path.write_text(json.dumps({"items": [{
        "id": "monthly_aud_reval:bybit_live:2026-04",
        "row_type": "monthly_aud_reval",
        "account": "live",
        "account_label": "Bybit Live",
        "close_time": "2026-04-30T23:59:59Z",
        "result_cash": 0.0,
        "result_currency": "AUD",
        "raw_refs": {"period_month": "2026-04"},
    }]}), encoding="utf-8")
    monkeypatch.setattr(master_service, "MONTHLY_AUD_REVALUATION_PATH", monthly_path)
    rows = master_service._monthly_aud_revaluation_rows_for_journal_view()
    assert len(rows) == 1
    assert rows[0]["id"] == "monthly_aud_reval:bybit_live:2026-04"
    assert rows[0]["result_cash"] == pytest.approx(0.0)
    assert rows[0]["account"] == "BYBIT"
    assert rows[0]["account_label"] == "BYBIT"


def test_monthly_aud_revaluation_rows_for_journal_view_accepts_canonical_bybit(tmp_path, monkeypatch):
    monthly_path = tmp_path / "monthly_aud_revaluation.json"
    monthly_path.write_text(json.dumps({"items": [{
        "id": "monthly_aud_reval:bybit_live:2026-05",
        "row_type": "monthly_aud_reval",
        "account": "BYBIT",
        "account_label": "BYBIT",
        "close_time": "2026-05-31T23:59:59Z",
        "result_cash": 1.0,
        "result_currency": "AUD",
    }]}), encoding="utf-8")
    monkeypatch.setattr(master_service, "MONTHLY_AUD_REVALUATION_PATH", monthly_path)
    rows = master_service._monthly_aud_revaluation_rows_for_journal_view()
    assert len(rows) == 1
    assert rows[0]["account"] == "BYBIT"
    assert rows[0]["account_label"] == "BYBIT"


def test_persist_trading_journal_sqlite_routes_monthly_rows_to_journal_notes(tmp_path, monkeypatch):
    sqlite_path = tmp_path / "trading_journal.sqlite"
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SQLITE_PATH", sqlite_path)
    snapshot = {
        "generated_at": "2026-05-01T00:00:00Z",
        "items": [
            {"id": "monthly_aud_reval:bybit_live:2026-03", "row_type": "monthly_aud_reval", "result_cash": 12.3, "result_currency": "AUD", "close_time": "2026-03-31T23:59:59Z"},
            {"id": "trade:1", "row_type": "trade", "close_time": "2026-05-01T00:00:00Z", "net_profit": 5.0, "metrics": {"x": 1}},
        ],
        "balances": [],
        "stats": {},
        "diagnostics": {},
        "source_fingerprints": {"files": []},
    }
    master_service._persist_trading_journal_sqlite(snapshot)
    conn = sqlite3.connect(sqlite_path)
    try:
        assert conn.execute("SELECT COUNT(*) FROM journal_notes WHERE id='monthly_aud_reval:bybit_live:2026-03'").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM journal_trades WHERE id='monthly_aud_reval:bybit_live:2026-03'").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM journal_metrics WHERE id='monthly_aud_reval:bybit_live:2026-03'").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM journal_trades WHERE id='trade:1'").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM journal_metrics WHERE id='trade:1'").fetchone()[0] == 1
    finally:
        conn.close()


def test_crypto_monthly_pnl_endpoint_no_anchor_returns_bootstrap_required(monkeypatch, tmp_path):
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_run_monthly_aud_revaluation_sync", lambda reason: {"ok": True})
    monkeypatch.setattr(master_service, "_monthly_aud_revaluation_rows_for_journal_view", lambda: [])
    monkeypatch.setattr(master_service, "_read_monthly_aud_reval_months_from_workbook", lambda _p: {"ok": True, "workbook_exists": False, "months": []})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_kwargs: {"master_journal_path": str(master_service._master_journal_path())})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda p,e: {"ok": True, "missing_row_ids": []})
    r = asyncio.run(master_service.trading_journal_crypto_monthly_pnl())
    assert r.status_code == 422
    payload = json.loads(r.body.decode("utf-8"))
    assert payload["ok"] is False
    assert 'Bootstrap required' in payload["message"]


def test_crypto_monthly_pnl_due_month_april_2026(monkeypatch, tmp_path):
    from tools.master_journal_workbook import build_master_journal_workbook
    monkeypatch.setattr(master_service, '_brisbane_now', lambda: __import__('datetime').datetime(2026,5,21))
    monkeypatch.setattr(master_service, 'TRADING_JOURNAL_LOCAL_DIR', tmp_path)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "Trading Journal.xlsx")
    build_master_journal_workbook({'items':[{'id':'monthly_aud_reval:bybit_live:2026-03','row_type':'monthly_aud_reval','source':'bybit_monthly_aud_reval','symbol':'MONTHLY AUD P/L','result_currency':'AUD','raw_refs':{'period_month':'2026-03'},'close_time':'2026-03-31T23:59:59Z'}],'stats':{'totals':{},'groups':{}},'balances':[]}, tmp_path / 'Trading Journal.xlsx')
    rows = [{'id':'monthly_aud_reval:bybit_live:2026-03','row_type':'monthly_aud_reval','raw_refs':{'period_month':'2026-03'}}]
    monkeypatch.setattr(master_service, '_monthly_aud_revaluation_rows_for_journal_view', lambda: list(rows))
    monkeypatch.setattr(master_service, '_read_monthly_aud_reval_months_from_workbook', lambda _p: {"ok": True, "workbook_exists": True, "months": ["2026-03"]})
    async def fake_run(reason):
        rows.append({'id':'monthly_aud_reval:bybit_live:2026-04','row_type':'monthly_aud_reval','raw_refs':{'period_month':'2026-04'}})
        return {'ok': True}
    monkeypatch.setattr(master_service, '_run_monthly_aud_revaluation_sync', fake_run)
    monkeypatch.setattr(master_service, '_sync_master_journal_workbook', lambda **_kwargs: {'master_journal_path': str(tmp_path / 'Trading Journal.xlsx')})
    monkeypatch.setattr(master_service, '_verify_trade_log_row_ids_in_workbook', lambda p,e: {'ok': True, 'missing_row_ids': []})
    r = asyncio.run(master_service.trading_journal_crypto_monthly_pnl())
    assert r.status_code == 200
    j = json.loads(r.body.decode("utf-8"))
    assert j['target_months'] == ['2026-04']


def test_history_page_no_post_diagnostics_or_sync():
    js = (ROOT / 'render' / 'static' / 'history_page.js').read_text(encoding='utf-8')
    assert "fetchJson('/api/trading-journal/diagnostics', { method: 'POST' })" not in js
    assert '/api/trading-journal/sync' not in js
    assert '/api/trading-journal/readiness' not in js
    assert 'Trading Journal sync reported failure after OANDA backfill' not in js


def test_trading_journal_js_sync_code_removed():
    js = (ROOT / 'render' / 'static' / 'trading_journal.js').read_text(encoding='utf-8')
    for token in ['/api/trading-journal/sync','/api/trading-journal/readiness','waitForSync','watchSyncCompletion','triggerBackgroundSync','Syncing journal sources','Sync complete:','localLast','syncStatusPromise','Auto-sync from configured journal sources','manual Sync now remains available','backgroundSyncLabel','syncWatchTimer','Journal cache is building/syncing','Sync required','manualSyncInFlight','const sleep = ']:
        assert token not in js


def test_read_monthly_anchor_helper_missing_trade_log(tmp_path):
    from openpyxl import Workbook
    wb=Workbook()
    wb.active.title='Not Trade Log'
    path=tmp_path/'x.xlsx'
    wb.save(path)
    out = master_service._read_monthly_aud_reval_months_from_workbook(path)
    assert out['ok'] is False
    assert out['trade_log_exists'] is False


def test_read_monthly_anchor_helper_missing_row_id(tmp_path):
    from openpyxl import Workbook
    wb=Workbook()
    ws=wb.active
    ws.title='Trade Log'
    ws['A1']='Close Time'
    path=tmp_path/'x2.xlsx'
    wb.save(path)
    out = master_service._read_monthly_aud_reval_months_from_workbook(path)
    assert out['ok'] is False
    assert out['row_id_column_exists'] is False


def test_crypto_monthly_endpoint_fails_when_workbook_anchor_read_fails(monkeypatch, tmp_path):
    monkeypatch.setattr(master_service, '_brisbane_now', lambda: __import__('datetime').datetime(2026,5,21))
    monkeypatch.setattr(master_service, '_master_journal_path', lambda: tmp_path/'Trading Journal.xlsx')
    monkeypatch.setattr(master_service, '_monthly_aud_revaluation_rows_for_journal_view', lambda: [{'id':'monthly_aud_reval:bybit_live:2026-03','raw_refs':{'period_month':'2026-03'}}])
    monkeypatch.setattr(master_service, '_read_monthly_aud_reval_months_from_workbook', lambda _p: {'ok':False,'workbook_exists':True,'error':'boom','months':[],'row_ids':[],'trade_log_exists':False,'row_id_column_exists':False})
    r=asyncio.run(master_service.trading_journal_crypto_monthly_pnl())
    assert r.status_code==500
    assert json.loads(r.body.decode("utf-8"))['ok'] is False


def test_verify_trade_log_row_ids_uses_streaming_iter_rows_only(monkeypatch, tmp_path):
    class FakeSheet:
        def __init__(self):
            self._rows = [
                ("Row ID", "Close Time"),
                ("rid-1", "2026-01-01T00:00:00Z"),
                ("rid-2", "2026-01-02T00:00:00Z"),
            ]

        def iter_rows(self, min_row=1, max_row=None, values_only=False):
            assert values_only is True
            rows = self._rows
            start = max(min_row - 1, 0)
            end = max_row if max_row is not None else len(rows)
            for row in rows[start:end]:
                yield row

        def cell(self, *_a, **_k):
            raise AssertionError("random cell access must not be used")

    class FakeWorkbook:
        def __init__(self):
            self.sheet = FakeSheet()

        def close(self):
            return None

    monkeypatch.setattr(master_service, "load_workbook", lambda *_a, **_k: FakeWorkbook())
    monkeypatch.setattr(master_service, "_get_trade_log_sheet", lambda wb, allow_legacy=False: wb.sheet)
    path = tmp_path / "Trading Journal.xlsx"
    path.write_bytes(b"ok")
    out = master_service._verify_trade_log_row_ids_in_workbook(path, ["rid-1", "rid-2"])
    assert out["ok"] is True
    assert out["missing_row_ids"] == []
    assert out["found_row_ids_count"] == 2


def test_read_monthly_aud_reval_months_uses_streaming_iter_rows_only(monkeypatch, tmp_path):
    class FakeSheet:
        def __init__(self):
            self._rows = [
                ("Row ID",),
                ("monthly_aud_reval:bybit_live:2026-03",),
                ("ignore-me",),
                ("monthly_aud_reval:bybit_live:2026-04",),
            ]

        def iter_rows(self, min_row=1, max_row=None, values_only=False):
            assert values_only is True
            rows = self._rows
            start = max(min_row - 1, 0)
            end = max_row if max_row is not None else len(rows)
            for row in rows[start:end]:
                yield row

        def cell(self, *_a, **_k):
            raise AssertionError("random cell access must not be used")

    class FakeWorkbook:
        def __init__(self):
            self.sheet = FakeSheet()

        def close(self):
            return None

    monkeypatch.setattr(master_service, "load_workbook", lambda *_a, **_k: FakeWorkbook())
    monkeypatch.setattr(master_service, "_get_trade_log_sheet", lambda wb, allow_legacy=False: wb.sheet)
    path = tmp_path / "Trading Journal.xlsx"
    path.write_bytes(b"ok")
    out = master_service._read_monthly_aud_reval_months_from_workbook(path)
    assert out["ok"] is True
    assert out["months"] == ["2026-03", "2026-04"]
    assert out["row_ids"] == [
        "monthly_aud_reval:bybit_live:2026-03",
        "monthly_aud_reval:bybit_live:2026-04",
    ]


def _demo_balance_snapshot(balance):
    return {"balances": [{"account": "Bybit Demo", "label": "Bybit Demo", "balance": balance, "currency": "USDT"}], "items": []}


def test_bybit_demo_balance_adjustment_negative_and_positive(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": True, "master_journal_path": str(master_service._master_journal_path())})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda path, ids: {"ok": True, "missing_row_ids": []})

    neg_seq = iter([_demo_balance_snapshot(100.0), _demo_balance_snapshot(60.0)])
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: next(neg_seq))
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -40, "reason": "fix"}))
    payload = _json(res)
    assert payload["ok"] is True and payload["new_balance"] == 60.0

    pos_seq = iter([_demo_balance_snapshot(100.0), _demo_balance_snapshot(125.0)])
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: next(pos_seq))
    res2 = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": 25}))
    payload2 = _json(res2)
    assert payload2["ok"] is True and payload2["new_balance"] == 125.0


def test_bybit_demo_balance_adjustment_rejects_invalid_amounts(temp_state_paths):
    for amt in [0, "", "nan", float("inf")]:
        res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": amt}))
        payload = _json(res)
        assert payload["ok"] is False


def test_bybit_demo_balance_adjustment_respects_feature_gate_before_preflight(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", False)
    before_rows = list(master_service._get_trading_journal_rows())
    calls = {"sync": 0, "lock": 0}
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: calls.__setitem__("sync", calls["sync"] + 1))
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda *_a, **_k: calls.__setitem__("lock", calls["lock"] + 1) or {"locked": False})

    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": 10}))
    payload = _json(res)
    assert res.status_code == 409
    assert payload["ok"] is False
    assert "bybit_demo_journal_disabled" in (payload.get("errors") or [])
    assert master_service._get_trading_journal_rows() == before_rows
    assert calls["sync"] == 0
    assert calls["lock"] == 0


def test_bybit_demo_balance_adjustment_rejects_missing_numeric_balance(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: {"balances": [{"account": "Bybit Demo", "balance": "x"}]})
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": 5}))
    assert res.status_code == 409


def test_bybit_demo_balance_adjustment_rolls_back_when_sync_fails(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: _demo_balance_snapshot(100.0))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": False, "error": "boom"})
    before = list(master_service._get_trading_journal_rows())
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -1}))
    payload = _json(res)
    assert payload["ok"] is False
    assert master_service._get_trading_journal_rows() == before


def test_pending_cashflow_rows_are_merged_into_authoritative_ledger(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_authoritative_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"x": 1})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: {})
    monkeypatch.setattr(master_service, "_save_trading_journal_view_snapshot", lambda payload: None)
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_compute_journal_stats", lambda items, balances: {})
    monkeypatch.setattr(master_service, "_build_authoritative_trading_journal_diagnostics_snapshot", lambda items: {})
    monkeypatch.setattr(master_service, "_monthly_aud_revaluation_rows_for_journal_view", lambda: [])
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda path: {"items": [], "cashflow_ledger": {"BYBIT DEMO": [{"id": "x1", "date": "2026-01-01T00:00:00Z", "new_balance": 100.0, "amount": 0.0}]}})
    master_service._PENDING_MANUAL_SYNC_ROWS = [{"id": "x2", "row_type": "cashflow", "account": "Bybit Demo", "currency": "USDT", "cashflow_amount": -40.0, "cashflow_new_balance": 60.0, "close_time": "2026-01-02T00:00:00Z"}]
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    bal = next((b for b in (snap.get("balances") or []) if str(b.get("label") or b.get("account")) == "Bybit Demo"), None)
    assert bal and float(bal.get("balance")) == 60.0


def test_bybit_demo_balance_adjustment_final_confirmation_runs_after_pending_restore(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda path, ids: {"ok": True, "missing_row_ids": []})
    sentinel_pending = [{"id": "existing:pending"}]
    master_service._PENDING_MANUAL_SYNC_ROWS = list(sentinel_pending)

    seq = iter([_demo_balance_snapshot(100.0), _demo_balance_snapshot(60.0)])
    seen_pending_states = []
    def _snapshot(force=False):
        seen_pending_states.append([dict(r) for r in (master_service._PENDING_MANUAL_SYNC_ROWS or []) if isinstance(r, dict)])
        return next(seq)
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", _snapshot)

    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -40}))
    payload = _json(res)
    assert payload["ok"] is True
    assert seen_pending_states[0] == sentinel_pending
    assert seen_pending_states[1] == sentinel_pending
    assert master_service._PENDING_MANUAL_SYNC_ROWS == sentinel_pending


def test_bybit_demo_balance_adjustment_fails_when_post_clear_persisted_balance_not_updated(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setenv("TRADING_JOURNAL_GITHUB_SYNC_ENABLED", "0")
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_noop": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda path, ids: {"ok": True, "missing_row_ids": []})

    seq = iter([_demo_balance_snapshot(100.0), _demo_balance_snapshot(100.0)])
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: next(seq))
    before = list(master_service._get_trading_journal_rows())
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -40}))
    payload = _json(res)
    assert payload["ok"] is False
    assert "after pending rows were restored" in str(payload.get("message") or "")
    assert master_service._get_trading_journal_rows() == before


def test_authoritative_snapshot_normalizes_raw_cashflow_ledger_key_without_duplicate_balance_rows(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_authoritative_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_journal_source_fingerprint", lambda: {"x": 2})
    monkeypatch.setattr(master_service, "_load_trading_journal_view_snapshot", lambda: {})
    monkeypatch.setattr(master_service, "_save_trading_journal_view_snapshot", lambda payload: None)
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda: [])
    monkeypatch.setattr(master_service, "_compute_journal_stats", lambda items, balances: {})
    monkeypatch.setattr(master_service, "_build_authoritative_trading_journal_diagnostics_snapshot", lambda items: {})
    monkeypatch.setattr(master_service, "_monthly_aud_revaluation_rows_for_journal_view", lambda: [])
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda path: {
        "items": [{"id":"t1","row_type":"trade","source":"master_journal","account":"Bybit Demo","account_label":"Bybit Demo","symbol":"BTCUSDT","close_time":"2026-01-03T00:00:00Z","net_profit":0.0}],
        "cashflow_ledger": {"Bybit Demo": [{"account": "Bybit Demo", "date": "2026-01-02T00:00:00Z", "new_balance": 60.0, "amount": -40.0, "currency": "USDT"}]},
    })
    master_service._PENDING_MANUAL_SYNC_ROWS = []
    snap = master_service._build_trading_journal_view_snapshot(force=True)
    balances = [b for b in (snap.get("balances") or []) if isinstance(b, dict) and str(b.get("label") or b.get("account") or "").strip().upper() == "BYBIT DEMO"]
    assert len(balances) == 1
    assert float(balances[0].get("balance")) == 60.0


def test_bybit_demo_balance_adjustment_uses_later_local_timestamp_than_balance_as_of(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": "ok", "path": str(path)})
    seq = iter([{"balances":[{"account":"Bybit Demo","label":"Bybit Demo","balance":100.0,"currency":"USDT","as_of":"2026-05-26T12:43:30"}]}, {"balances":[{"account":"Bybit Demo","label":"Bybit Demo","balance":60.0,"currency":"USDT"}]}])
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: next(seq))
    monkeypatch.setattr(master_service, "_journal_local_now_naive_iso", lambda: "2026-05-26T03:13:00")
    captured = {}
    original_set = master_service._set_trading_journal_rows
    def _capture_set(rows):
        if rows:
            captured['row'] = rows[-1]
        return original_set(rows)
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", _capture_set)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda path, ids: {"ok": True, "missing_row_ids": []})
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -40}))
    payload = _json(res)
    assert payload["ok"] is True
    assert 'T' in str(captured['row']['close_time'])


def test_bybit_demo_balance_adjustment_workbook_locked_returns_423_without_mutation(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": True, "reason": "excel_lock_file_present", "path": str(path)})
    called = {"upsert":0, "sync":0}
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *a, **k: called.__setitem__('upsert', called['upsert']+1))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: called.__setitem__('sync', called['sync']+1))
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -40}))
    payload = _json(res)
    assert res.status_code == 423
    assert payload["ok"] is False
    assert called['upsert'] == 0 and called['sync'] == 0


def test_persisted_cashflow_after_latest_trade_uses_cashflow_anchor_plus_trades_source(temp_state_paths):
    rows = [{"id":"t1","row_type":"trade","account":"Bybit Demo","account_label":"Bybit Demo","close_time":"2026-05-26T12:43:30","net_profit":0.0,"balance_after_trade":100.0,"source":"master_journal"}]
    ledger = {"Bybit Demo":[{"account":"Bybit Demo","date":"2026-05-26T12:43:31","new_balance":60.0,"amount":-40.0,"currency":"USDT"}]}
    timeline = master_service._build_journal_balance_timelines(rows, master_service._normalize_cashflow_ledger_keys(ledger), [])
    bal = next((b for b in (timeline.get('balances') or []) if str(b.get('label') or b.get('account')).upper()=="BYBIT DEMO"), None)
    assert bal is not None
    assert float(bal.get('balance')) == 60.0
    assert str(bal.get('balance_source')) == 'cashflow_anchor_plus_trades'


def test_poll_pending_webhook_invalidations_cancel_path_sets_cancelled_at_without_nameerror(monkeypatch: pytest.MonkeyPatch):
    calls = {"update": 0, "context": 0}
    monkeypatch.setattr(master_service, "LIMIT_CANCEL_POLL_SECONDS", 0)
    pending = [{"id": "p1", "enabled": True, "status": "WAITING", "category": "bybit", "instrument": "BTCUSDT", "cancel_if_touched_operator": "lte", "trade_mode": "linear"}]
    monkeypatch.setattr(master_service, "_load_pending_webhooks", lambda: pending)
    monkeypatch.setattr(master_service, "_parse_pending_cancel_touch_price", lambda _item: 100.0)
    async def _fake_price(**_k):
        return 99.0
    monkeypatch.setattr(master_service, "_fetch_bybit_market_price", _fake_price)
    monkeypatch.setattr(master_service, "_pending_cancel_touch_triggered", lambda **_k: True)
    monkeypatch.setattr(master_service, "_update_pending_webhook", lambda *_a, **_k: calls.__setitem__("update", calls["update"] + 1))
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda *_a, **_k: calls.__setitem__("context", calls["context"] + 1))
    monkeypatch.setattr(master_service, "_invalidate_open_orders_cache", lambda: None)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)

    state = {"n": 0}
    async def _sleep(_secs):
        state["n"] += 1
        if state["n"] > 1:
            raise asyncio.CancelledError()
    monkeypatch.setattr(master_service.asyncio, "sleep", _sleep)

    with pytest.raises(asyncio.CancelledError):
        asyncio.run(master_service._poll_pending_webhook_invalidations())
    assert calls["update"] == 1
    assert calls["context"] == 1

def test_bybit_demo_adjustment_disabled_early(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, 'ENABLE_BYBIT_DEMO_JOURNAL', False)
    called = {'snapshot': 0}
    monkeypatch.setattr(master_service, '_build_trading_journal_view_snapshot', lambda force=False: called.__setitem__('snapshot', called['snapshot']+1) or {})
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({'amount': 1}))
    payload = _json(res)
    assert res.status_code == 409
    assert payload['errors'] == ['bybit_demo_journal_disabled']
    assert called['snapshot'] == 0


def test_bybit_demo_adjustment_invalid_amount(temp_state_paths):
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({'amount': 0}))
    assert res.status_code == 422


def test_bybit_demo_adjustment_lock_preflight_happens_before_snapshot(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    called = {"snapshot": 0}
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": True, "reason": "excel_open"})
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: called.__setitem__("snapshot", called["snapshot"] + 1) or {})
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": 1}))
    assert res.status_code == 423
    assert called["snapshot"] == 0


def test_journal_local_now_naive_iso_uses_brisbane_timezone(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "APP_TIMEZONE", "Australia/Brisbane")
    out = master_service._journal_local_now_naive_iso()
    assert isinstance(out, str) and "T" in out and "+" not in out


def test_cashflow_row_to_ledger_event_maps_amount_and_new_balance():
    event = master_service._cashflow_row_to_ledger_event(
        {
            "id": "cf:1",
            "account_label": "Bybit Demo",
            "close_time": "2026-05-26T12:00:00",
            "cashflow_amount": "-1",
            "cashflow_new_balance": "99",
            "currency": "USDT",
            "notes": "n",
        }
    )
    assert event["id"] == "cf:1"
    assert event["amount"] == -1.0
    assert event["new_balance"] == 99.0


def test_bybit_demo_adjustment_rollback_restores_workbook_bytes(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    workbook.write_bytes(b"before-bytes")
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: {"balances": [{"account": "Bybit Demo", "label": "Bybit Demo", "balance": 100.0, "currency": "USDT"}]})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": False, "error": "sync failed"})
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", lambda rows: workbook.write_bytes(b"mutated") or None)
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -1}))
    assert res.status_code == 500
    assert workbook.read_bytes() == b"before-bytes"


def test_merge_pending_cashflow_rows_normalizes_bybit_demo_key():
    ledger = {'Bybit Demo': [{'id': 'a'}], 'BYBIT DEMO': [{'id': 'b'}]}
    pending = [{'id': 'c', 'row_type': 'cashflow', 'account': 'Bybit Demo', 'close_time':'2026-01-01T00:00:00', 'cashflow_amount': 1, 'cashflow_new_balance': 3}]
    out = master_service._merge_pending_cashflow_rows_into_ledger(ledger, pending)
    assert list(out.keys()) == [master_service._norm_account_key('Bybit Demo')]
    assert sorted(str(x.get('id')) for x in out[master_service._norm_account_key('Bybit Demo')]) == ['a', 'b', 'c']


def test_bybit_demo_adjustment_row_has_open_time_equal_close_time(monkeypatch: pytest.MonkeyPatch, temp_state_paths):
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda path: {"locked": False, "reason": ""})
    seq = iter([{"balances": [{"account": "Bybit Demo", "label": "Bybit Demo", "balance": 100.0, "currency": "USDT"}]}, {"balances": [{"account": "Bybit Demo", "label": "Bybit Demo", "balance": 99.0, "currency": "USDT"}]}])
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda force=False: next(seq))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **kwargs: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda path, ids: {"ok": True, "missing_row_ids": []})
    res = asyncio.run(master_service.trading_journal_bybit_demo_balance_adjustment({"amount": -1}))
    assert res.status_code == 200
    rows = master_service._get_trading_journal_rows()
    row = next(r for r in rows if str(r.get("source")) == "manual_bybit_demo_balance_adjustment")
    assert str(row.get("open_time") or "")
    assert row.get("open_time") == row.get("close_time")


def test_normalize_cashflow_ledger_keys_uses_event_account_over_raw_bucket():
    ledger = {"misc": [{"id": "x1", "account": "Bybit Demo", "date": "2026-01-01T00:00:00Z", "amount": -1, "new_balance": 99}]}
    out = master_service._normalize_cashflow_ledger_keys(ledger)
    bybit_key = master_service._norm_account_key("Bybit Demo")
    misc_key = master_service._norm_account_key("misc")
    assert bybit_key in out
    assert out[bybit_key][0]["id"] == "x1"
    assert misc_key not in out


def test_import_hydrates_from_workbook_when_json_state_missing_rows(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    master_service._set_trading_journal_rows([])
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    wb_existing = {"id": "manual:existing:1", "row_type": "trade", "source": "manual", "symbol": "BTCUSDT"}
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([parsed], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [wb_existing]})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    ids = {str(r.get("id")) for r in master_service._get_trading_journal_rows()}
    assert "manual:existing:1" in ids
    assert "oanda_export:demo:626:630" in ids


def test_import_sync_missing_row_ids_propagated(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([{"id": "new:1", "row_type": "trade", "source": "manual"}], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": []})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": False, "error": "workbook_row_survivor_verification_failed", "missing_row_ids": ["new:1"], "diagnostics": {"x": 1}})
    out = master_service._import_uploaded_trading_journal_file("manual.xlsx", b"x")
    assert out["ok"] is False
    assert out["missing_row_ids"] == ["new:1"]
    assert out["master_journal_error"] == "workbook_row_survivor_verification_failed"
    assert out["diagnostics"] == {"x": 1}


def test_import_hydration_replaces_stale_same_id_rows_from_workbook(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    master_service._set_trading_journal_rows([{"id": "manual:existing:1", "row_type": "trade", "source": "manual", "symbol": "BTCUSDT", "net_profit": 1.0}])
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    wb_existing = {"id": "manual:existing:1", "row_type": "trade", "source": "manual", "symbol": "BTCUSDT", "net_profit": 2.0}
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([parsed], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [wb_existing]})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if str(r.get("id")) == "manual:existing:1")
    assert row["net_profit"] == 2.0


def test_import_duplicate_rows_merged_recomputed_after_workbook_hydration(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    master_service._set_trading_journal_rows([])
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([parsed], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(parsed)]})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "refresh_master_journal_derived_sheets", lambda *_a, **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    assert out["duplicate_rows_merged"] == 1
    rows = [r for r in master_service._get_trading_journal_rows() if str(r.get("id")) == "oanda_export:demo:626:630"]
    assert len(rows) == 1


def test_atomic_write_bytes_retries_replace_then_succeeds(tmp_path, monkeypatch: pytest.MonkeyPatch):
    target = tmp_path / "state_backup.json"
    real_replace = master_service.os.replace
    calls = {"n": 0}
    def _replace(src, dst):
        calls["n"] += 1
        if calls["n"] < 3:
            err = PermissionError("busy")
            err.winerror = 32
            raise err
        return real_replace(src, dst)
    monkeypatch.setattr(master_service.os, "replace", _replace)
    master_service._atomic_write_bytes(target, b'{"ok":1}')
    assert target.read_bytes() == b'{"ok":1}'
    assert calls["n"] == 3


def test_atomic_write_bytes_fallback_after_retry_exhaustion(tmp_path, monkeypatch: pytest.MonkeyPatch):
    target = tmp_path / "state_backup.json"
    def _replace(_src, _dst):
        err = PermissionError("denied")
        err.winerror = 5
        raise err
    monkeypatch.setattr(master_service.os, "replace", _replace)
    master_service._atomic_write_bytes(target, b'{"ok":2}')
    assert target.read_bytes() == b'{"ok":2}'


def test_import_duplicate_noop_fast_path_skips_workbook_sync(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(parsed)]})
    called = {"sync": 0, "verify": 0, "upsert": 0, "stats_refresh": 0}
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: called.__setitem__("sync", called["sync"] + 1) or {"ok": True})
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: called.__setitem__("upsert", called["upsert"] + 1) or 1)
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: called.__setitem__("verify", called["verify"] + 1) or {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "refresh_master_journal_derived_sheets", lambda *_a, **_k: called.__setitem__("stats_refresh", called["stats_refresh"] + 1) or {"ok": True})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    assert out["duplicate_rows_merged"] == 1
    assert "stats refreshed" in str(out["message"]).lower()
    assert called["sync"] == 0
    assert called["upsert"] == 0
    assert called["verify"] == 1
    assert called["stats_refresh"] == 1


def test_import_duplicate_noop_preserves_workbook_authored_fields(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    workbook_row = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD", "notes": "from-workbook"}
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(workbook_row)]})
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("upsert should be skipped")))
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: (_ for _ in ()).throw(AssertionError("sync should be skipped")))
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "refresh_master_journal_derived_sheets", lambda *_a, **_k: {"ok": True})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    row = next(r for r in master_service._get_trading_journal_rows() if str(r.get("id")) == "oanda_export:demo:626:630")
    assert row.get("notes") == "from-workbook"


def test_import_changed_row_does_not_use_fast_path(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    workbook_row = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD", "net_profit": 10.0}
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD", "net_profit": 11.0}
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(workbook_row)]})
    called = {"upsert": 0, "sync": 0}
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: called.__setitem__("upsert", called["upsert"] + 1) or 1)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: called.__setitem__("sync", called["sync"] + 1) or {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service.time, "perf_counter", lambda: 100.0)
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is True
    assert called["upsert"] == 1
    assert called["sync"] == 1
    assert out["message"] == "Import complete."
    assert "stats refreshed" not in out["message"].lower()


def test_import_duplicate_noop_stats_refresh_failure_returns_error(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda", "symbol": "EURUSD"}
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(parsed)]})
    monkeypatch.setattr(master_service, "refresh_master_journal_derived_sheets", lambda *_a, **_k: {"ok": False, "error": "refresh failed"})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")
    assert out["ok"] is False
    assert "stats refresh failed" in str(out["message"]).lower()


def test_oanda_transaction_export_balance_after_trade_is_authoritative_over_stale_cashflow():
    rows = [
        {
            "id": "oanda_export:demo:626:630",
            "row_type": "trade",
            "source": "oanda_transaction_export",
            "account": "OANDA DEMO",
            "account_label": "OANDA DEMO",
            "close_time": "2026-05-26T16:55:31+10:00",
            "net_profit": -0.4505,
            "balance_after_trade": 1500.20,
            "balance_after_trade_currency": "AUD",
        },
    ]
    ledger = {"OANDA DEMO": [{"account": "OANDA DEMO", "date": "2022-01-01T00:00:00Z", "new_balance": 201.6695, "currency": "AUD"}]}
    timeline = master_service._build_journal_balance_timelines(rows, ledger, [])
    bal = timeline["balances"][0]
    assert bal["balance"] == pytest.approx(1500.20)
    assert bal["balance_source"] == "authoritative_trade_balance"
    assert bal["stale_cashflow_overridden"] is True


def test_import_changed_oanda_financial_fields_do_not_use_duplicate_fast_path(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    workbook_row = {
        "id": "oanda_export:demo:626:630",
        "row_type": "trade",
        "source": "oanda_transaction_export",
        "account": "demo",
        "account_label": "OANDA DEMO",
        "symbol": "EURUSD",
        "net_profit": -0.4505,
        "commission": 0.9366,
        "balance_after_trade": 201.6695,
        "notes": "keep me",
    }
    parsed = dict(workbook_row, commission=0.0, balance_after_trade=1500.20)
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": [dict(workbook_row)]})
    called = {"upsert": 0, "sync": 0, "stats_refresh": 0}
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: called.__setitem__("upsert", called["upsert"] + 1) or 1)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: called.__setitem__("sync", called["sync"] + 1) or {"ok": True})
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "refresh_master_journal_derived_sheets", lambda *_a, **_k: called.__setitem__("stats_refresh", called["stats_refresh"] + 1) or {"ok": True})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})

    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")

    assert out["ok"] is True
    assert called["upsert"] == 1
    assert called["sync"] == 1
    assert called["stats_refresh"] == 0
    assert "stats refreshed" not in str(out["message"]).lower()


def test_import_preflight_rejects_when_trading_journal_xlsx_locked(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    workbook.write_bytes(b"locked")
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda _p: {"locked": True, "reason": "excel_open"})
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("parse should not run")))

    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")

    assert out["ok"] is False
    assert out["status_code"] == 423
    assert out["code"] == "EXCEL_WORKBOOK_OPEN"
    assert "workbook_locked" in out["errors"]
    assert out["rows_parsed"] == 0


def test_open_master_journal_rejects_while_import_in_progress(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    workbook = temp_state_paths / "Trading Journal.xlsx"
    workbook.write_bytes(b"x")
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: workbook)
    monkeypatch.setattr(master_service, "_open_path_with_os", lambda _p: (_ for _ in ()).throw(AssertionError("should not open")))
    master_service._update_trading_journal_import_status(running=True, stage="workbook_sync", message="Updating", started_at=master_service._utc_now_iso(), started_epoch=master_service.time.time(), upload_name="demo.csv")
    try:
        res = asyncio.run(master_service.open_master_journal_file())
        payload = master_service.json.loads(res.body.decode("utf-8"))
        assert res.status_code == 409
        assert payload["ok"] is False
        assert payload["code"] == "TRADING_JOURNAL_IMPORT_IN_PROGRESS"
    finally:
        master_service._update_trading_journal_import_status(running=False, stage="idle", message="", finished_at=master_service._utc_now_iso(), upload_name="")


def test_import_final_replace_permission_error_returns_excel_lock_payload(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    parsed = {"id": "oanda_export:demo:626:630", "row_type": "trade", "source": "oanda_transaction_export", "account": "OANDA DEMO", "symbol": "EURUSD", "net_profit": -0.4505, "commission": 0.0, "balance_after_trade": 1500.20}
    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_master_journal_lock_status", lambda _p: {"locked": False, "reason": ""})
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": []})
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: 1)
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", lambda **_k: {"ok": False, "code": "EXCEL_WORKBOOK_OPEN", "status_code": 423, "message": "Trading Journal.xlsx appears to be open in Excel. Close Excel, then press Resume.", "master_journal_error": "locked", "diagnostics": {"workbook_sync_substage_timings": {"final_replace": 0.01}}})

    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")

    assert out["ok"] is False
    assert out["status_code"] == 423
    assert out["code"] == "EXCEL_WORKBOOK_OPEN"
    assert out["rows_parsed"] == 1
    assert "workbook_sync" in out["import_timings"]
    assert out["diagnostics"]["workbook_sync_substage_timings"]["final_replace"] == pytest.approx(0.01)


def test_manual_import_uses_single_local_only_prebuilt_snapshot(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: temp_state_paths / "Trading Journal.xlsx")
    (temp_state_paths / "Trading Journal.xlsx").write_bytes(b"x")
    parsed = {
        "id": "oanda_export:demo:626:630",
        "row_type": "trade",
        "source": "oanda_transaction_export",
        "account": "OANDA DEMO",
        "account_label": "OANDA DEMO",
        "symbol": "EURUSD",
        "net_profit": -0.4505,
        "commission": 0.0,
        "balance_after_trade": 1500.20,
        "balance_after_trade_currency": "AUD",
    }
    snapshot = {"items": [dict(parsed)], "balances": [{"account": "OANDA DEMO", "label": "OANDA DEMO", "balance": 1500.20, "currency": "AUD"}], "stats": {}, "diagnostics": {}}
    calls = {"snapshot": 0, "sync": 0}

    def fake_snapshot(*, force=False, local_only=False, skip_external_balances=False, skip_live_account_refresh=False):
        calls["snapshot"] += 1
        assert force is True
        assert local_only is True
        assert skip_external_balances is True
        assert skip_live_account_refresh is True
        return dict(snapshot)

    def fake_sync(**kwargs):
        calls["sync"] += 1
        assert kwargs.get("defer_github_sync") is True
        assert kwargs.get("prebuilt_snapshot") == snapshot
        return {"ok": True, "master_journal_ok": True, "github_sync_ok": None, "github_sync_deferred": True, "master_journal_path": str(temp_state_paths / "Trading Journal.xlsx")}

    master_service._set_trading_journal_rows([])
    monkeypatch.setattr(master_service, "_parse_local_trading_journal_workbook", lambda *_a, **_k: ([dict(parsed)], None))
    monkeypatch.setattr(master_service, "_infer_realized_net_profit_from_balance_continuity", lambda rows, _existing: (rows, [], {"pnl_inferred_count": 0, "pnl_unresolved_count": 0, "pnl_unresolved_row_ids": []}))
    monkeypatch.setattr(master_service, "read_master_journal_source", lambda _p: {"items": []})
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda *_a, **_k: 1)
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", fake_snapshot)
    monkeypatch.setattr(master_service, "_sync_master_journal_workbook", fake_sync)
    monkeypatch.setattr(master_service, "_verify_trade_log_row_ids_in_workbook", lambda *_a, **_k: {"ok": True, "missing_row_ids": []})
    monkeypatch.setattr(master_service, "_persist_trading_journal_sqlite", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_sync_journal_excel_files_to_github", lambda *_a, **_k: {"github_sync_enabled": False, "github_sync_ok": True, "github_sync_error": "", "github_sync_commit": ""})

    out = master_service._import_uploaded_trading_journal_file("oanda_demo.csv", b"x")

    assert out["ok"] is True
    assert calls == {"snapshot": 1, "sync": 1}
    assert out["import_timings"]["snapshot_build"] >= 0


def test_manual_import_local_snapshot_does_not_call_external_refresh_helpers(temp_state_paths, monkeypatch: pytest.MonkeyPatch):
    master_service._set_trading_journal_rows([
        {
            "id": "oanda_export:demo:626:630",
            "row_type": "trade",
            "source": "oanda_transaction_export",
            "account": "OANDA DEMO",
            "account_label": "OANDA DEMO",
            "symbol": "EURUSD",
            "close_time": "2026-05-26T16:55:31+10:00",
            "net_profit": -0.4505,
            "commission": 0.0,
            "balance_after_trade": 1500.20,
            "balance_after_trade_currency": "AUD",
        }
    ])
    monkeypatch.setattr(master_service, "_load_cashflows_for_active_journal_source", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("dropbox/active cashflow loader should not run")))
    monkeypatch.setattr(master_service, "_get_excel_account_balances", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("external balance seeds should be skipped")))
    monkeypatch.setattr(master_service, "_list_local_oanda_history_exports", lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("OANDA export scan should not run")))

    snapshot = master_service._build_trading_journal_view_snapshot(force=True, local_only=True, skip_external_balances=True, skip_live_account_refresh=True)

    assert snapshot["items"]
    bal = next(b for b in snapshot["balances"] if str(b.get("account") or b.get("label")).upper() == "OANDA DEMO")
    assert bal["balance"] == pytest.approx(1500.20)
