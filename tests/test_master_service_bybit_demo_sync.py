import asyncio
import importlib.util
import json
import warnings
import sys
import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
HTTPX_AVAILABLE = importlib.util.find_spec("httpx") is not None
pytestmark = pytest.mark.skipif(not HTTPX_AVAILABLE, reason="httpx is not installed")

if HTTPX_AVAILABLE:
    SPEC = importlib.util.spec_from_file_location("render_master_service_bybit_sync", ROOT / "render" / "master_service.py")
    master_service = importlib.util.module_from_spec(SPEC)
    assert SPEC and SPEC.loader
    sys.modules[SPEC.name] = master_service
    SPEC.loader.exec_module(master_service)


def _isolate_trading_journal_runtime(monkeypatch, tmp_path: Path) -> None:
    """Keep sync tests from reading or writing repository runtime-state files."""

    monkeypatch.setattr(master_service, "TRADING_JOURNAL_PATH", tmp_path / "trading_journal.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_STATE_PATH", tmp_path / "trading_journal_state.json")
    monkeypatch.setattr(
        master_service,
        "TRADING_JOURNAL_SYNC_STATE_PATH",
        tmp_path / "trading_journal_sync_state.json",
    )


def test_save_json_file_retries_transient_permission_error(tmp_path, monkeypatch) -> None:
    from shared import atomic_json

    target = tmp_path / "trading_journal.json"
    attempts = {"count": 0}
    seen_src: list[str] = []
    real_replace = atomic_json.os.replace

    def flaky_replace(src: str | os.PathLike[str], dst: str | os.PathLike[str]) -> None:
        attempts["count"] += 1
        seen_src.append(Path(src).name)
        if attempts["count"] < 3:
            raise PermissionError("[WinError 5] Access is denied")
        real_replace(src, dst)

    monkeypatch.setattr(atomic_json.time, "sleep", lambda _n: None)
    monkeypatch.setattr(atomic_json.os, "replace", flaky_replace)

    master_service._save_json_file(target, {"items": [{"id": "r1"}]})
    payload = json.loads(target.read_text(encoding="utf-8"))
    assert payload["items"][0]["id"] == "r1"
    assert attempts["count"] == 3
    assert seen_src
    assert all(name != "trading_journal.json.tmp" for name in seen_src)


def test_save_json_file_surfaces_permanent_permission_error(tmp_path, monkeypatch) -> None:
    from shared import atomic_json

    target = tmp_path / "trading_journal.json"
    monkeypatch.setattr(atomic_json.time, "sleep", lambda _n: None)
    monkeypatch.setattr(
        atomic_json.os,
        "replace",
        lambda _src, _dst: (_ for _ in ()).throw(PermissionError("[WinError 5] Access is denied")),
    )

    with pytest.raises(PermissionError):
        master_service._save_json_file(target, {"items": []})


def test_manual_demo_sync_uses_7_day_recovery_window(monkeypatch) -> None:
    now_s = 1_700_000_000.0
    now_ms = int(now_s * 1000)
    captured = {}
    monkeypatch.setattr(master_service.time, "time", lambda: now_s)
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _mode: ("demo", "k", "s", "https://api.bybit.com", "env"))
    monkeypatch.setattr(master_service, "_persist_bybit_closed_pnl_last_seen", lambda: None)
    master_service._BYBIT_CLOSED_PNL_LAST_SEEN["demo"] = None

    async def fake_sync(**kwargs):
        captured.update(kwargs)
        return kwargs["start_time"] + 1000

    monkeypatch.setattr(master_service, "_sync_bybit_closed_pnl_window", fake_sync)
    result = asyncio.run(master_service._run_bybit_closed_pnl_sync(account_mode="demo", reason="manual"))
    expected = now_ms - master_service._BYBIT_CLOSED_PNL_RECOVERY_WINDOW_MS + master_service._BYBIT_CLOSED_PNL_RECOVERY_SAFETY_MARGIN_MS
    assert result["ok"] is True
    assert captured["start_time"] == expected


def test_startup_recovery_forces_7_day_window_even_with_last_seen(monkeypatch) -> None:
    now_s = 1_700_100_000.0
    now_ms = int(now_s * 1000)
    captured = {}
    monkeypatch.setattr(master_service.time, "time", lambda: now_s)
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _mode: ("demo", "k", "s", "https://api.bybit.com", "env"))
    monkeypatch.setattr(master_service, "_persist_bybit_closed_pnl_last_seen", lambda: None)
    master_service._BYBIT_CLOSED_PNL_LAST_SEEN["demo"] = now_ms - (60 * 1000)

    async def fake_sync(**kwargs):
        captured.update(kwargs)
        return kwargs["start_time"] + 2000

    monkeypatch.setattr(master_service, "_sync_bybit_closed_pnl_window", fake_sync)
    asyncio.run(master_service._run_bybit_closed_pnl_sync(account_mode="demo", reason="startup_recovery"))
    expected = now_ms - master_service._BYBIT_CLOSED_PNL_RECOVERY_WINDOW_MS + master_service._BYBIT_CLOSED_PNL_RECOVERY_SAFETY_MARGIN_MS
    assert captured["start_time"] == expected


def test_persisted_closed_pnl_last_seen_restored_after_restart(tmp_path, monkeypatch) -> None:
    state_path = tmp_path / "trading_journal_state.json"
    state_path.write_text(
        json.dumps({"bybit_closed_pnl_last_seen": {"demo": 111, "live": 222}}),
        encoding="utf-8",
    )
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_STATE_PATH", state_path)
    master_service._BYBIT_CLOSED_PNL_LAST_SEEN["demo"] = None
    master_service._BYBIT_CLOSED_PNL_LAST_SEEN["live"] = None

    master_service._restore_bybit_closed_pnl_last_seen_from_state()
    assert master_service._BYBIT_CLOSED_PNL_LAST_SEEN["demo"] == 111
    assert master_service._BYBIT_CLOSED_PNL_LAST_SEEN["live"] == 222


def test_recovered_rows_upsert_without_duplicates(monkeypatch) -> None:
    saved = {"rows": []}
    row = {"id": "bybit:demo:closedpnl:BTCUSDT:123", "symbol": "BTCUSDT", "raw_refs": {"orderId": "123"}}
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: list(saved["rows"]))
    monkeypatch.setattr(
        master_service,
        "_save_trading_journal",
        lambda rows, **_kwargs: saved.update({"rows": list(rows)}),
    )
    changed1 = master_service._upsert_trading_journal_rows([row])
    changed2 = master_service._upsert_trading_journal_rows([row])
    assert changed1 == 1
    assert changed2 == 1
    assert len(saved["rows"]) == 1


def test_bybit_demo_poll_waits_for_restore_signal(monkeypatch) -> None:
    poll_called = {"value": False}
    wait_called = {"value": False}

    async def fake_wait_for(awaitable, timeout):
        wait_called["value"] = True
        await awaitable
        return None

    async def fake_poll():
        poll_called["value"] = True

    event = asyncio.Event()
    monkeypatch.setattr(master_service, "_STARTUP_STATE_RESTORE_DONE", event)
    monkeypatch.setattr(master_service.asyncio, "wait_for", fake_wait_for)
    monkeypatch.setattr(master_service, "_poll_bybit_demo_closed_pnl", fake_poll)

    async def runner():
        task = asyncio.create_task(master_service._start_bybit_demo_closed_pnl_poll_after_restore())
        await asyncio.sleep(0)
        assert wait_called["value"] is True
        assert poll_called["value"] is False
        event.set()
        await task

    asyncio.run(runner())


def test_startup_recovery_waits_for_restore_signal(monkeypatch) -> None:
    recovery_called = {"value": False}
    wait_called = {"value": False}

    async def fake_wait_for(awaitable, timeout):
        wait_called["value"] = True
        await awaitable
        return None

    async def fake_recovery():
        recovery_called["value"] = True

    event = asyncio.Event()
    monkeypatch.setattr(master_service, "APP_PROFILE", "local")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "dropbox")
    monkeypatch.setattr(master_service, "DROPBOX_SYNC_ENABLED", True)
    monkeypatch.setattr(master_service, "_STARTUP_STATE_RESTORE_DONE", event)
    monkeypatch.setattr(master_service.asyncio, "wait_for", fake_wait_for)
    monkeypatch.setattr(master_service, "_run_startup_recovery_import_if_needed", fake_recovery)

    async def runner():
        task = asyncio.create_task(master_service._start_startup_recovery_import_after_restore())
        await asyncio.sleep(0)
        assert wait_called["value"] is True
        assert recovery_called["value"] is False
        event.set()
        await task
        assert recovery_called["value"] is True

    asyncio.run(runner())


def test_workbook_upsert_normalizes_blank_numeric_values(monkeypatch) -> None:
    existing = pd.DataFrame(
        [{"order_id": "oid-1", "entry_price": 100.0, "stop_loss": 95.0, "take_profit": 110.0}],
        columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS,
    )
    captured = {"uploaded": None}

    def fake_read_excel(*_args, **_kwargs):
        return existing.copy()

    class DummyWriter:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda _path: b"dummy")
    monkeypatch.setattr(master_service.pd, "read_excel", fake_read_excel)
    monkeypatch.setattr(master_service.pd, "ExcelWriter", DummyWriter)
    monkeypatch.setattr(master_service.pd.DataFrame, "to_excel", lambda self, *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda _path, payload: captured.update({"uploaded": payload}))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda _folder: {"changed": 0})
    monkeypatch.setattr(
        master_service,
        "_bybit_demo_workbook_row",
        lambda _row: {
            "opening_time": "2026-01-01",
            "closing_time": "2026-01-01",
            "type_buy_sell": "Buy",
            "symbol": "BTCUSDT",
            "size_quantity": "",
            "entry_price": "",
            "closing_price": "",
            "stop_loss": "",
            "take_profit": "",
            "commission": "",
            "net_profit": "",
            "balance_after_trade": "",
            "currency": "USDT",
            "notes": "",
            "fill_count": "",
            "order_id": "oid-1",
            "source": "bybit",
        },
    )

    changed = master_service._append_bybit_demo_rows_to_workbook("/tmp", [{"id": "x"}])
    assert changed == 1
    assert captured["uploaded"] is not None


def test_place_bybit_order_upserts_final_absolute_tpsl_context(monkeypatch) -> None:
    captured_contexts = []
    invalidations = {"count": 0}

    class DummyResponse:
        def raise_for_status(self):
            return None

        def json(self):
            return {"retCode": 0, "retMsg": "OK", "result": {"orderId": "oid-123", "orderLinkId": "link-123"}}

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def post(self, *_args, **_kwargs):
            return DummyResponse()

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _mode: ("demo", "k", "s", "https://api.bybit.com", "env"))
    async def fake_wait_for_position_entry(**_kwargs):
        return {"avgPrice": "100.0", "positionIdx": 0}

    async def fake_set_trading_stop(**_kwargs):
        return {"retCode": 0}

    monkeypatch.setattr(master_service, "_wait_for_position_entry", fake_wait_for_position_entry)
    monkeypatch.setattr(master_service, "_set_bybit_trading_stop", fake_set_trading_stop)
    monkeypatch.setattr(master_service, "_delete_pending_webhook", lambda _id: False)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: captured_contexts.append(payload) or payload)
    monkeypatch.setattr(master_service, "_invalidate_open_orders_cache", lambda: invalidations.__setitem__("count", invalidations["count"] + 1))

    payload = {
        "symbol": "DASHUSDT",
        "action": "buy",
        "quantity": "1",
        "account": "demo",
        "tp_offset": "10",
        "sl_offset": "-5",
        "timeframe": "15-minute",
        "order_type": "market",
    }
    result = asyncio.run(master_service._place_bybit_order(payload, request_id="req-1"))
    assert result["order"]["orderId"] == "oid-123"
    assert len(captured_contexts) >= 2
    assert captured_contexts[0]["timeframe"] == "15MIN"
    assert captured_contexts[-1]["entry_price"] == 100.0
    assert captured_contexts[-1]["stop_loss"] == 95.0
    assert captured_contexts[-1]["take_profit"] == 110.0
    assert captured_contexts[-1]["status"] == "ACTIVE"
    assert invalidations["count"] == 1


def test_bybit_repair_backfills_blank_fields_and_persists(monkeypatch) -> None:
    rows = [
        {
            "id": "bybit:demo:closedpnl:DASHUSDT:oid-1",
            "source": "bybit",
            "account": "demo",
            "symbol": "DASHUSDT",
            "side": "Buy",
            "status": "closed",
            "timeframe": "",
            "stop_loss": "",
            "take_profit": "",
            "raw_refs": {"orderId": "oid-1"},
        }
    ]
    saved = {"rows": None}
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: list(rows))
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", lambda updated: saved.update({"rows": list(updated)}))
    monkeypatch.setattr(
        master_service,
        "_backfill_trade_row_context_fields",
        lambda row: {**row, "timeframe": "15-minute", "stop_loss": 90.0, "take_profit": 120.0},
    )
    changed = master_service._repair_persisted_bybit_trade_context_fields()
    assert changed == 1
    assert saved["rows"] is not None
    assert saved["rows"][0]["timeframe"] == "15-minute"
    assert saved["rows"][0]["stop_loss"] == 90.0
    assert saved["rows"][0]["take_profit"] == 120.0


def test_closed_pnl_row_backfills_tpsl_from_context(monkeypatch) -> None:
    monkeypatch.setattr(
        master_service,
        "_lookup_trade_context_for_journal_row",
        lambda _row: {"timeframe": "1-hour", "stop_loss": "99.5", "take_profit": "110.2"},
    )
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "BTCUSDT",
            "orderId": "abc123",
            "orderLinkId": "link-1",
            "openFee": "0.1",
            "closeFee": "0.2",
            "fillCount": "1",
            "side": "Buy",
            "createdTime": 1_000,
            "updatedTime": 5_000,
            "avgEntryPrice": "100",
            "avgExitPrice": "105",
            "closedSize": "0.1",
            "closedPnl": "1.0",
        },
        account_mode="demo",
        balance_after_trade=1000.0,
        stop_loss=None,
        take_profit=None,
    )
    assert row is not None
    assert row["stop_loss"] == 99.5
    assert row["take_profit"] == 110.2
    assert row["timeframe"] == "1H"


def test_closed_pnl_row_prefers_context_open_time(monkeypatch) -> None:
    resolved_ctx = {
        "open_time": "2026-04-11T01:20:00+00:00",
        "created_at": "2026-04-11T01:19:00+00:00",
        "timeframe": "4-hour",
    }
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda _row: None)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda _payload: _payload)
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "DASHUSDT",
            "orderId": "dash-order-1",
            "createdTime": 1_775_884_246_000,
            "updatedTime": 1_775_884_246_000,
            "avgEntryPrice": "22",
            "avgExitPrice": "23",
            "closedSize": "10",
            "closedPnl": "4.2",
        },
        account_mode="demo",
        balance_after_trade=1000.0,
        resolved_trade_context=resolved_ctx,
    )
    assert row is not None
    assert row["open_time"] == "2026-04-11T01:20:00+00:00"
    assert row["close_time"] == "2026-04-11T05:10:46+00:00"


def test_closed_pnl_row_stale_context_falls_back_to_created_time(monkeypatch) -> None:
    stale_ctx = {
        "open_time": "2026-04-11T06:10:46+00:00",
        "created_at": "2026-04-11T06:10:40+00:00",
        "timeframe": "4-hour",
        "stop_loss": "20",
        "take_profit": "30",
    }
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda _row: None)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda _payload: _payload)
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "HYPERUSDT",
            "orderId": "hyper-order-1",
            "createdTime": 1_775_870_000_000,
            "updatedTime": 1_775_884_246_000,
            "avgEntryPrice": "22",
            "avgExitPrice": "23",
            "closedSize": "10",
            "closedPnl": "4.2",
            "side": "Buy",
        },
        account_mode="demo",
        balance_after_trade=1000.0,
        resolved_trade_context=stale_ctx,
    )
    assert row is not None
    assert row["status"] == "closed"
    assert row["open_time"] == "2026-04-11T01:13:20+00:00"


def test_closed_pnl_row_stale_context_without_valid_created_time_is_quarantined(monkeypatch) -> None:
    stale_ctx = {
        "open_time": "2026-04-11T06:10:46+00:00",
        "timeframe": "4-hour",
        "stop_loss": "20",
        "take_profit": "30",
    }
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda _row: None)
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "HYPERUSDT",
            "orderId": "hyper-order-2",
            "createdTime": 1_775_884_246_000,
            "updatedTime": 1_775_884_246_000,
            "avgEntryPrice": "22",
            "avgExitPrice": "23",
            "closedSize": "10",
            "closedPnl": "4.2",
            "side": "Buy",
        },
        account_mode="demo",
        balance_after_trade=1000.0,
        resolved_trade_context=stale_ctx,
    )
    assert row is not None
    assert row["status"] == "invalid_time_order"
    assert row["row_type"] == "quarantine"


def test_resolve_bybit_closed_pnl_trade_context_prefers_valid_ref_match(monkeypatch) -> None:
    monkeypatch.setattr(
        master_service,
        "_load_trade_contexts",
        lambda: [
            {
                "broker": "bybit",
                "account": "demo",
                "instrument": "HYPERUSDT",
                "side": "Buy",
                "order_id": "oid-1",
                "open_time": "2026-04-11T06:10:46+00:00",
            },
            {
                "broker": "bybit",
                "account": "demo",
                "instrument": "HYPERUSDT",
                "side": "Buy",
                "order_id": "oid-1",
                "open_time": "2026-04-11T01:10:46+00:00",
            },
        ],
    )
    ctx = master_service._resolve_bybit_closed_pnl_trade_context(
        account_mode="demo",
        symbol="HYPERUSDT",
        side="Buy",
        order_id="oid-1",
        close_time="2026-04-11T05:10:46+00:00",
    )
    assert ctx is not None
    assert ctx["open_time"] == "2026-04-11T01:10:46+00:00"


def test_sync_bybit_closed_pnl_window_stale_context_does_not_raise(monkeypatch) -> None:
    statuses = []
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: ("/tmp", []))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda _folder: None)
    async def fake_empty_payload(**_kwargs):
        return {"result": {"list": []}}

    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", fake_empty_payload)
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda rows, **_k: len(rows))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_workbook", lambda *_args, **_kwargs: 0)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_args, **_kwargs: {"deduped_by_order_id": 0, "deduped_by_fingerprint": 0})
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **kwargs: statuses.append(kwargs))
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: payload)
    monkeypatch.setattr(master_service, "load_bybit_demo_tpsl_cache", lambda: {})
    monkeypatch.setattr(
        master_service,
        "_load_trade_contexts",
        lambda: [{"order_id": "oid-stale", "open_time": "2026-04-11T06:10:46+00:00"}],
    )
    async def fake_closed_pnl(**_kwargs):
        return {"result": {"list": [{
            "symbol": "HYPERUSDT", "orderId": "oid-stale", "orderLinkId": "", "side": "Buy",
            "createdTime": 1_775_884_246_000, "updatedTime": 1_775_884_246_000,
            "avgEntryPrice": "100", "avgExitPrice": "101", "closedSize": "1",
            "closedPnl": "1", "openFee": "0", "closeFee": "0",
        }]}}

    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", fake_closed_pnl)
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))
    assert statuses
    assert statuses[-1].get("last_error") is None


def test_sync_bybit_closed_pnl_window_quarantines_invalid_time_row(monkeypatch) -> None:
    statuses = []
    upsert_calls = []
    workbook_calls = []
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: ("/tmp", []))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda _folder: None)

    async def fake_empty_payload(**_kwargs):
        return {"result": {"list": []}}

    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", fake_empty_payload)
    async def fake_executions(**_kwargs):
        return []
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", fake_executions)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_args, **_kwargs: {"changed": 0})
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **kwargs: statuses.append(kwargs))
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: payload)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "load_bybit_demo_tpsl_cache", lambda: {})
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", lambda _account: asyncio.sleep(0, result={"available_usdt": 1000, "total_equity": 1000}))

    def fake_upsert(rows):
        upsert_calls.append(rows)
        return len(rows)

    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", fake_upsert)
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_workbook", lambda *_args, **_kwargs: workbook_calls.append(_args) or 0)

    async def fake_closed_pnl(**_kwargs):
        return {
            "result": {
                "list": [
                    {
                        "symbol": "HYPERUSDT",
                        "orderId": "oid-quarantine",
                        "orderLinkId": "",
                        "side": "Buy",
                        "createdTime": 1_775_884_246_000,
                        "updatedTime": 1_775_884_246_000,
                        "avgEntryPrice": "100",
                        "avgExitPrice": "101",
                        "closedSize": "1",
                        "closedPnl": "1",
                        "openFee": "0",
                        "closeFee": "0",
                    }
                ]
            }
        }

    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", fake_closed_pnl)
    monkeypatch.setattr(
        master_service,
        "_load_trade_contexts",
        lambda: [{"order_id": "oid-quarantine", "open_time": "2026-04-11T06:10:46+00:00"}],
    )

    asyncio.run(
        master_service._sync_bybit_closed_pnl_window(
            account_mode="demo",
            base_url="u",
            api_key="k",
            api_secret="s",
            start_time=0,
            end_time=3000,
        )
    )

    assert statuses
    status = statuses[-1]
    assert status.get("last_error") is None
    assert status.get("last_invalid_time_rows_dropped") == 1
    dropped = status.get("last_invalid_time_row_details") or []
    assert len(dropped) == 1
    assert dropped[0]["symbol"] == "HYPERUSDT"
    assert dropped[0]["side"] == "Buy"
    assert dropped[0]["orderId"] == "oid-quarantine"
    assert dropped[0]["open_time"] == "2026-04-11T06:10:46+00:00"
    assert dropped[0]["close_time"] == "2026-04-11T05:10:46+00:00"
    assert dropped[0]["reason"] == "invalid_time_order"
    assert upsert_calls == []
    assert workbook_calls == []


def test_sync_records_broker_balance_warning_without_failing_import(
    monkeypatch, tmp_path
) -> None:
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", lambda **_kwargs: asyncio.sleep(0, result={"ok": True}))
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda progress_cb=None: {"ok": True, "rows_imported": 1, "diagnostics": {"rows_by_asset_class": {}}})
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", lambda account: (_ for _ in ()).throw(RuntimeError(f"{account} missing creds")))
    async def fake_oanda_summary(account_mode):
        raise RuntimeError(f"{account_mode} missing creds")

    async def fake_oanda_recovery(_account_mode):
        return {"ok": True, "captured_rows": [], "captured_row_ids": []}

    monkeypatch.setattr(master_service, "_fetch_oanda_account_summary", fake_oanda_summary)
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", fake_oanda_recovery)
    monkeypatch.setattr(master_service, "_manual_sync_should_capture_bybit_demo_recent_history", lambda: False)
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: False)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "missing-journal.xlsx")
    monkeypatch.setattr(
        master_service,
        "_sync_master_journal_workbook",
        lambda **_kwargs: {
            "master_journal_ok": True,
            "master_journal_path": str(tmp_path / "missing-journal.xlsx"),
        },
    )
    saved_state = {}
    monkeypatch.setattr(master_service, "_load_trading_journal_state", lambda: dict(saved_state))
    monkeypatch.setattr(master_service, "_save_trading_journal_state", lambda payload: saved_state.update(payload))
    updates = []
    monkeypatch.setattr(master_service, "_set_trading_journal_sync_state", lambda **kwargs: updates.append(kwargs))
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert isinstance(saved_state.get("broker_balance_diagnostics"), dict)
    warnings = saved_state["broker_balance_diagnostics"].get("warnings") or []
    assert any("Bybit demo balance unavailable" in str(w) for w in warnings)
    assert any(update.get("ok") is True for update in updates)


def test_repair_existing_bybit_row_open_time_from_context(monkeypatch) -> None:
    bad_row = {
        "id": "bybit:demo:closedpnl:DASHUSDT:123",
        "source": "bybit",
        "account": "demo",
        "symbol": "DASHUSDT",
        "side": "Buy",
        "open_time": "2026-04-11T05:10:46+00:00",
        "close_time": "2026-04-11T05:10:46+00:00",
        "raw_refs": {"orderId": "123"},
    }
    monkeypatch.setattr(
        master_service,
        "_lookup_trade_context_for_journal_row",
        lambda _row: {"open_time": "2026-04-11T01:20:00+00:00", "created_at": "2026-04-11T01:19:30+00:00"},
    )
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: [])
    repaired, changed = master_service._repair_persisted_bybit_open_times([bad_row])
    assert changed == 1
    assert repaired[0]["open_time"] == "2026-04-11T01:20:00+00:00"
    assert repaired[0]["close_time"] == "2026-04-11T05:10:46+00:00"


def test_repair_existing_bybit_row_open_time_when_open_after_close(monkeypatch) -> None:
    bad_row = {
        "id": "bybit:demo:closedpnl:DASHUSDT:124",
        "source": "bybit",
        "account": "demo",
        "symbol": "DASHUSDT",
        "side": "Buy",
        "open_time": "2026-04-11T06:10:46+00:00",
        "close_time": "2026-04-11T05:10:46+00:00",
        "raw_refs": {"orderId": "124"},
    }
    monkeypatch.setattr(
        master_service,
        "_lookup_trade_context_for_journal_row",
        lambda _row: {"open_time": "2026-04-11T01:20:00+00:00"},
    )
    repaired, changed = master_service._repair_persisted_bybit_open_times([bad_row])
    assert changed == 1
    assert repaired[0]["open_time"] == "2026-04-11T01:20:00+00:00"


def test_normalize_bybit_closed_pnl_prefers_execution_times(monkeypatch) -> None:
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda _row: None)
    row = master_service._normalize_bybit_closed_pnl_row(
        {
            "symbol": "BTCUSDT",
            "orderId": "oid-exec",
            "createdTime": 1_700_000_000_000,
            "updatedTime": 1_700_000_900_000,
            "avgEntryPrice": "100",
            "avgExitPrice": "101",
            "closedSize": "1",
            "closedPnl": "1",
            "side": "Buy",
        },
        account_mode="demo",
        balance_after_trade=1000.0,
        execution_times={"open_time": "2026-04-17T00:24:42.814000+00:00", "close_time": "2026-04-17T00:29:23.301318+00:00"},
    )
    assert row is not None
    assert row["open_time"] == "2026-04-17T00:24:42.814000+00:00"
    assert row["close_time"] == "2026-04-17T00:29:23.301318+00:00"
    assert row["raw_refs"]["time_source"] == "execution"


def test_repair_existing_bybit_row_missing_open_time(monkeypatch) -> None:
    bad_row = {
        "id": "bybit:demo:closedpnl:DASHUSDT:125",
        "source": "bybit",
        "account": "demo",
        "symbol": "DASHUSDT",
        "side": "Buy",
        "open_time": "",
        "close_time": "2026-04-11T05:10:46+00:00",
        "raw_refs": {"orderId": "125"},
    }
    monkeypatch.setattr(
        master_service,
        "_lookup_trade_context_for_journal_row",
        lambda _row: {"created_at": "2026-04-11T01:19:30+00:00"},
    )
    repaired, changed = master_service._repair_persisted_bybit_open_times([bad_row])
    assert changed == 1
    assert repaired[0]["open_time"] == "2026-04-11T01:19:30+00:00"


def test_same_id_upsert_overwrites_bad_open_time_without_duplicate(monkeypatch) -> None:
    saved = {"rows": [{"id": "bybit:demo:closedpnl:DASHUSDT:123", "open_time": "2026-04-11T05:10:46+00:00"}]}
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: list(saved["rows"]))
    monkeypatch.setattr(
        master_service,
        "_save_trading_journal",
        lambda rows, **_kwargs: saved.update({"rows": list(rows)}),
    )
    master_service._upsert_trading_journal_rows(
        [{"id": "bybit:demo:closedpnl:DASHUSDT:123", "open_time": "2026-04-11T01:20:00+00:00"}]
    )
    assert len(saved["rows"]) == 1
    assert saved["rows"][0]["open_time"] == "2026-04-11T01:20:00+00:00"


def test_health_and_root_head_routes_return_200() -> None:
    health = asyncio.run(master_service.healthcheck())
    root_head = asyncio.run(master_service.root_head_health())
    route_paths = {getattr(route, "path", "") for route in master_service.app.routes}
    assert health.status_code == 200
    assert root_head.status_code == 200
    assert "/health" in route_paths
    assert "/" in route_paths


def test_workbook_upsert_handles_float_text_column_notes(monkeypatch) -> None:
    existing = pd.DataFrame(
        [{"order_id": "oid-1", "notes": float("nan")}],
        columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS,
    )
    existing["notes"] = pd.to_numeric(existing["notes"], errors="coerce")
    captured = {"uploaded": None}

    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda _path: b"dummy")
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_args, **_kwargs: existing.copy())

    class DummyWriter:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(master_service.pd, "ExcelWriter", DummyWriter)
    monkeypatch.setattr(master_service.pd.DataFrame, "to_excel", lambda self, *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda _path, payload: captured.update({"uploaded": payload}))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda _folder: {"changed": 0})
    monkeypatch.setattr(
        master_service,
        "_bybit_demo_workbook_row",
        lambda _row: {
            "opening_time": "2026-01-01",
            "closing_time": "2026-01-01",
            "type_buy_sell": "Buy",
            "symbol": "BTCUSDT",
            "size_quantity": 0.1,
            "entry_price": 100.0,
            "closing_price": 101.0,
            "stop_loss": "",
            "take_profit": "",
            "commission": 0.02,
            "net_profit": 1.0,
            "balance_after_trade": "",
            "currency": "USDT",
            "notes": "",
            "fill_count": "",
            "order_id": "oid-1",
            "source": "bybit",
        },
    )

    changed = master_service._append_bybit_demo_rows_to_workbook("/tmp", [{"id": "x"}])
    assert changed == 1
    assert captured["uploaded"] is not None


def test_workbook_row_roundtrip_preserves_timeframe(monkeypatch) -> None:
    row = {
        "open_time": "2026-01-01T00:00:00+00:00",
        "close_time": "2026-01-01T01:00:00+00:00",
        "side": "Buy",
        "symbol": "BTCUSDT",
        "qty": 1.0,
        "entry_price": 100.0,
        "exit_price": 110.0,
        "realized_pnl": 10.0,
        "timeframe": "15-minute",
        "is_test_trade": True,
        "raw_refs": {"orderId": "oid-tf", "fillCount": 1, "source": "closed_pnl"},
    }
    wb_row = master_service._bybit_demo_workbook_row(row)
    assert wb_row["timeframe"] == "15-minute"
    assert wb_row["is_test_trade"] == "Yes"

    frame = master_service._coerce_bybit_demo_workbook_frame(pd.DataFrame([wb_row]))
    reparsed = {
        "timeframe": master_service._normalize_timeframe(master_service._excel_cell_to_python(frame.iloc[0].get("timeframe"))),
        "is_test_trade": master_service._normalize_test_trade_flag(master_service._excel_cell_to_python(frame.iloc[0].get("is_test_trade"))),
    }
    assert reparsed["timeframe"] == "15-minute"
    assert reparsed["is_test_trade"] is True


def test_workbook_upsert_handles_multiple_float_text_columns(monkeypatch) -> None:
    existing = pd.DataFrame(
        [{"order_id": "oid-1", "notes": float("nan"), "source": float("nan"), "symbol": float("nan")}],
        columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS,
    )
    for col in ("notes", "source", "symbol"):
        existing[col] = pd.to_numeric(existing[col], errors="coerce")

    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda _path: b"dummy")
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_args, **_kwargs: existing.copy())

    class DummyWriter:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(master_service.pd, "ExcelWriter", DummyWriter)
    monkeypatch.setattr(master_service.pd.DataFrame, "to_excel", lambda self, *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda _folder: {"changed": 0})
    monkeypatch.setattr(
        master_service,
        "_bybit_demo_workbook_row",
        lambda _row: {
            "opening_time": "2026-01-01",
            "closing_time": "2026-01-01",
            "type_buy_sell": "Sell",
            "symbol": "ETHUSDT",
            "size_quantity": 0.2,
            "entry_price": 200.0,
            "closing_price": 198.0,
            "stop_loss": "",
            "take_profit": "",
            "commission": 0.02,
            "net_profit": -0.4,
            "balance_after_trade": "",
            "currency": "USDT",
            "notes": "",
            "fill_count": "",
            "order_id": "oid-1",
            "source": "closed_pnl",
        },
    )

    changed = master_service._append_bybit_demo_rows_to_workbook("/tmp", [{"id": "x"}])
    assert changed == 1


def test_workbook_upsert_raises_no_future_warning(monkeypatch) -> None:
    existing = pd.DataFrame(
        [{"order_id": "oid-1", "notes": float("nan")}],
        columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS,
    )
    existing["notes"] = pd.to_numeric(existing["notes"], errors="coerce")

    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda _path: b"dummy")
    monkeypatch.setattr(master_service.pd, "read_excel", lambda *_args, **_kwargs: existing.copy())

    class DummyWriter:
        def __init__(self, *_args, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

    monkeypatch.setattr(master_service.pd, "ExcelWriter", DummyWriter)
    monkeypatch.setattr(master_service.pd.DataFrame, "to_excel", lambda self, *args, **kwargs: None)
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda _folder: {"changed": 0})
    monkeypatch.setattr(master_service, "_bybit_demo_workbook_row", lambda _row: {
        "opening_time": "2026-01-01", "closing_time": "2026-01-01", "type_buy_sell": "Buy",
        "symbol": "BTCUSDT", "size_quantity": "", "entry_price": "", "closing_price": "",
        "stop_loss": "", "take_profit": "", "commission": "", "net_profit": "",
        "balance_after_trade": "", "currency": "USDT", "notes": "", "order_id": "oid-1",
        "fill_count": "", "source": "closed_pnl",
    })

    with warnings.catch_warnings():
        warnings.simplefilter("error", FutureWarning)
        changed = master_service._append_bybit_demo_rows_to_workbook("/tmp", [{"id": "x"}])
    assert changed == 1


def test_bybit_unresolved_tpsl_warns_once_and_persists_registry(tmp_path, monkeypatch) -> None:
    warnings = []
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_STATE_PATH", tmp_path / "state.json")
    master_service._STARTUP_STATE_RESTORE_DONE.set()

    class DummyLogger:
        def warning(self, message, *args):
            warnings.append(message % args if args else message)

    monkeypatch.setattr(master_service, "BYBIT_LOGGER", DummyLogger())
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: ("/tmp", []))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda _folder: None)
    async def fake_empty_payload(**_kwargs):
        return {"result": {"list": []}}

    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", fake_empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", fake_empty_payload)
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda rows, **_k: len(rows))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_workbook", lambda *_args, **_kwargs: 0)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_args, **_kwargs: {"deduped_by_order_id": 0, "deduped_by_fingerprint": 0})
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_kwargs: None)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: payload)
    monkeypatch.setattr(master_service, "load_bybit_demo_tpsl_cache", lambda: {})
    async def fake_closed_pnl(**_kwargs):
        return {"result": {"list": [{
            "symbol": "BTCUSDT", "orderId": "oid-1", "orderLinkId": "", "side": "Buy",
            "createdTime": 1000, "updatedTime": 2000, "avgEntryPrice": "100", "avgExitPrice": "101",
            "closedSize": "1", "closedPnl": "1", "openFee": "0", "closeFee": "0",
        }]}}

    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", fake_closed_pnl)

    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))

    assert len([line for line in warnings if "BYBIT_DEMO_TPSL unresolved" in line]) == 1
    state = master_service._load_trading_journal_state()
    key = "demo|oid-1|||BTCUSDT|"
    entry = state.get("unresolved_registry", {}).get("bybit_demo_tpsl", {}).get(key, {})
    assert entry.get("count") == 2
    assert entry.get("resolved") is False


def test_parent_order_link_id_lookup_supported(monkeypatch) -> None:
    monkeypatch.setattr(
        master_service,
        "_load_trade_contexts",
        lambda: [{"parent_order_link_id": "parent-1", "timeframe": "1-hour"}],
    )
    row = master_service._lookup_trade_context_for_journal_row({"raw_refs": {"parentOrderLinkId": "parent-1"}})
    assert row and row.get("timeframe") == "1-hour"


def test_local_authoritative_sync_bybit_demo_does_not_touch_dropbox(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: (_ for _ in ()).throw(AssertionError("dropbox resolve called")))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox ensure called")))
    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox download called")))
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox upload called")))
    monkeypatch.setattr(master_service, "_ensure_local_bybit_demo_files", lambda _d: {"demo_workbook_created": True, "trade_history_template_created": True})
    async def _empty_payload(**_): return {"result": {"list": []}}
    async def _empty_exec(**_): return []
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _empty_exec)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty_payload)
    monkeypatch.setattr(master_service, "load_bybit_demo_tpsl_cache", lambda: {})
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_: None)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: payload)
    async def _empty_balance(): return {}
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", _empty_balance)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0, "deduped_by_order_id": 0, "deduped_by_fingerprint": 0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", lambda _rows: None)
    calls = {"local_append": 0}
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_local_workbook", lambda *_: calls.__setitem__("local_append", calls["local_append"] + 1) or 1)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_: {"changed": 0, "deduped_by_order_id": 0, "deduped_by_fingerprint": 0})
    async def fake_closed(**_):
        return {"result": {"list": [{"symbol": "BTCUSDT", "orderId": "oid-b", "side": "Buy", "avgEntryPrice": "100", "avgExitPrice": "101", "closedSize": "1", "closedPnl": "1", "openFee": "0", "closeFee": "0", "createdTime": 1000, "updatedTime": 2000}]}}
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", fake_closed)
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))
    assert calls["local_append"] == 1


def test_append_bybit_demo_rows_to_local_workbook_writes_and_is_idempotent(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    row = {
        "status": "closed",
        "open_time": "2026-04-01T00:00:00Z",
        "close_time": "2026-04-01T01:00:00Z",
        "side": "Buy",
        "symbol": "BTCUSDT",
        "qty": 1.5,
        "entry_price": 100.0,
        "exit_price": 110.0,
        "stop_loss": 95.0,
        "take_profit": 120.0,
        "commission": 0.2,
        "realized_pnl": 14.8,
        "balance_after_trade": 1000.0,
        "timeframe": "1-hour",
        "is_test_trade": False,
        "notes": "n",
        "raw_refs": {"orderId": "oid-local-1", "fillCount": 1, "source": "closed_pnl"},
    }
    changed = master_service._append_bybit_demo_rows_to_local_workbook(tmp_path, [row])
    assert changed == 1
    workbook = tmp_path / master_service.BYBIT_DEMO_WORKBOOK_NAME
    assert workbook.exists()
    df = master_service.pd.read_excel(workbook, sheet_name=master_service.BYBIT_DEMO_WORKBOOK_SHEET)
    assert len(df) == 1
    assert str(df.iloc[0]["order_id"]) == "oid-local-1"
    changed2 = master_service._append_bybit_demo_rows_to_local_workbook(tmp_path, [row])
    assert changed2 == 1
    df2 = master_service.pd.read_excel(workbook, sheet_name=master_service.BYBIT_DEMO_WORKBOOK_SHEET)
    assert len(df2) == 1


def test_local_authoritative_sync_bybit_demo_writes_real_workbook(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: (_ for _ in ()).throw(AssertionError("dropbox resolve called")))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox ensure called")))
    monkeypatch.setattr(master_service, "_dropbox_download_bytes", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox download called")))
    monkeypatch.setattr(master_service, "_dropbox_upload_bytes", lambda *_: (_ for _ in ()).throw(AssertionError("dropbox upload called")))
    monkeypatch.setattr(master_service, "_ensure_local_bybit_demo_files", lambda _d: {"demo_workbook_created": True, "trade_history_template_created": True})
    async def _empty_payload(**_): return {"result": {"list": []}}
    async def _empty_exec(**_): return []
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty_payload)
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _empty_exec)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty_payload)
    monkeypatch.setattr(master_service, "load_bybit_demo_tpsl_cache", lambda: {})
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_: None)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda payload: payload)
    async def _empty_balance(): return {}
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", _empty_balance)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0, "deduped_by_order_id": 0, "deduped_by_fingerprint": 0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", lambda _rows: None)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_: {"changed": 0, "deduped_by_order_id": 0, "deduped_by_fingerprint": 0})
    async def fake_closed(**_):
        return {"result": {"list": [{"symbol": "BTCUSDT", "orderId": "oid-real-1", "side": "Buy", "avgEntryPrice": "100", "avgExitPrice": "101", "closedSize": "1", "closedPnl": "1", "openFee": "0", "closeFee": "0", "createdTime": 1000, "updatedTime": 2000}]}}
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", fake_closed)
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))
    workbook = tmp_path / master_service.BYBIT_DEMO_WORKBOOK_NAME
    assert workbook.exists()
    df = master_service.pd.read_excel(workbook, sheet_name=master_service.BYBIT_DEMO_WORKBOOK_SHEET)
    assert any(str(v) == "oid-real-1" for v in df["order_id"].tolist())


def test_run_sync_job_pre_sanitizes_local_demo_workbook_with_wallet_snapshot(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "ENABLE_BYBIT_DEMO_JOURNAL", True)
    monkeypatch.setattr(master_service, "_trading_journal_source_mode", lambda: "local")
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: False)
    monkeypatch.setattr(master_service, "_trading_journal_bybit_demo_balance_anchor_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_set_trading_journal_sync_state", lambda **_: None)
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", lambda **_: asyncio.sleep(0, result={"ok": True}))
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *_: asyncio.sleep(0, result={"ok": True}))
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda **_: {"ok": True, "rows_imported": 0, "diagnostics": {}, "warnings": []})
    monkeypatch.setattr(master_service, "_read_excel_sheet_or_empty", lambda *_: master_service.pd.DataFrame([{"net_profit": 1.0, "balance_after_trade": None}]))
    monkeypatch.setattr(master_service, "_coerce_bybit_demo_workbook_frame", lambda frame: frame)
    monkeypatch.setattr(master_service, "_bybit_demo_workbook_has_rows_needing_balance", lambda _f: True)
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", lambda _acct: asyncio.sleep(0, result={"available_usdt": 10}))
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", lambda: asyncio.sleep(0, result={"current_balance": 224.87769878}))
    calls = {"sanitize": 0}
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_a, **_k: calls.__setitem__("sanitize", calls["sanitize"] + 1) or {"changed": 1})
    saved = {"broker_account_balances": [{"account": "Bybit Live", "label": "Bybit Live", "balance": 11.0}]}
    monkeypatch.setattr(master_service, "_load_trading_journal_state", lambda: dict(saved))
    monkeypatch.setattr(master_service, "_save_trading_journal_state", lambda s: saved.update(s))
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert calls["sanitize"] >= 1
    assert any(str((b or {}).get("label")) == "Bybit Demo" for b in saved.get("broker_account_balances", []))
    assert any(str((b or {}).get("label")) == "Bybit Live" for b in saved.get("broker_account_balances", []))


def test_run_sync_job_demo_anchor_failure_keeps_missing_warning(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "_trading_journal_source_mode", lambda: "local")
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: False)
    monkeypatch.setattr(master_service, "_trading_journal_bybit_demo_balance_anchor_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_set_trading_journal_sync_state", lambda **_: None)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda **_: {"ok": True, "rows_imported": 0, "diagnostics": {}, "warnings": []})
    monkeypatch.setattr(master_service, "_read_excel_sheet_or_empty", lambda *_: master_service.pd.DataFrame([{"net_profit": 1.0, "balance_after_trade": None}]))
    monkeypatch.setattr(master_service, "_coerce_bybit_demo_workbook_frame", lambda frame: frame)
    monkeypatch.setattr(master_service, "_bybit_demo_workbook_has_rows_needing_balance", lambda _f: True)
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", lambda: (_ for _ in ()).throw(RuntimeError("demo creds missing")))
    calls = {"sanitize": 0}
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_a, **_k: calls.__setitem__("sanitize", calls["sanitize"] + 1) or {"changed": 1})
    state = {}
    monkeypatch.setattr(master_service, "_load_trading_journal_state", lambda: dict(state))
    monkeypatch.setattr(master_service, "_save_trading_journal_state", lambda s: state.update(s))
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert calls["sanitize"] == 0
    warnings = ((state.get("broker_balance_diagnostics") or {}).get("warnings") or [])
    assert any("wallet snapshot unavailable" in str(w).lower() for w in warnings)


def test_bybit_demo_workbook_has_rows_needing_balance_only_when_missing_balance():
    df_full = master_service.pd.DataFrame(
        [{"net_profit": 1.0, "balance_after_trade": 10.0}, {"net_profit": -1.0, "balance_after_trade": 9.0}]
    )
    assert master_service._bybit_demo_workbook_has_rows_needing_balance(df_full) is False
    df_missing = master_service.pd.DataFrame(
        [{"net_profit": 1.0, "balance_after_trade": None}, {"net_profit": -1.0, "balance_after_trade": 9.0}]
    )
    assert master_service._bybit_demo_workbook_has_rows_needing_balance(df_missing) is True


def test_run_sync_job_demo_anchor_non_numeric_balance_warns_and_skips_sanitize(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "_trading_journal_source_mode", lambda: "local")
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: False)
    monkeypatch.setattr(master_service, "_trading_journal_bybit_demo_balance_anchor_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_set_trading_journal_sync_state", lambda **_: None)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda **_: {"ok": True, "rows_imported": 0, "diagnostics": {}, "warnings": []})
    monkeypatch.setattr(master_service, "_read_excel_sheet_or_empty", lambda *_: master_service.pd.DataFrame([{"net_profit": 1.0, "balance_after_trade": None}]))
    monkeypatch.setattr(master_service, "_coerce_bybit_demo_workbook_frame", lambda frame: frame)
    monkeypatch.setattr(master_service, "_bybit_demo_workbook_has_rows_needing_balance", lambda _f: True)
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", lambda: asyncio.sleep(0, result={"current_balance": None}))
    calls = {"sanitize": 0}
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_a, **_k: calls.__setitem__("sanitize", calls["sanitize"] + 1) or {"changed": 1})
    state = {}
    monkeypatch.setattr(master_service, "_load_trading_journal_state", lambda: dict(state))
    monkeypatch.setattr(master_service, "_save_trading_journal_state", lambda s: state.update(s))
    asyncio.run(master_service._run_trading_journal_sync_job())
    assert calls["sanitize"] == 0
    assert not any(str((b or {}).get("label")) == "Bybit Demo" for b in state.get("broker_account_balances", []))
    warnings = ((state.get("broker_balance_diagnostics") or {}).get("warnings") or [])
    assert any("no numeric current_balance" in str(w).lower() for w in warnings)


def test_bybit_signed_get_retries_after_timestamp_window_error(monkeypatch) -> None:
    calls: list[tuple[str, dict[str, str]]] = []
    market_time_calls = {"count": 0}
    now = {"value": 1_700_000_000.0}
    master_service._BYBIT_TIME_OFFSET_CACHE.clear()
    master_service._BYBIT_TIME_OFFSET_CACHE["https://api.bybit.com"] = {
        "synced_at": int(now["value"] * 1000),
        "offset_ms": 0,
        "rtt_ms": 0,
    }

    class DummyResponse:
        def __init__(self, status_code: int, payload: dict[str, object]):
            self.status_code = status_code
            self._payload = payload
            self.content = json.dumps(payload).encode("utf-8")
            self.text = json.dumps(payload)

        def json(self):
            return self._payload

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def get(self, url: str, headers=None):
            headers = headers or {}
            calls.append((url, headers))
            if url.endswith("/v5/market/time"):
                market_time_calls["count"] += 1
                return DummyResponse(200, {"retCode": 0, "result": {"timeSecond": "1700000010"}})
            if len([u for u, _ in calls if "/v5/order/history" in u]) == 1:
                return DummyResponse(
                    200,
                    {
                        "retCode": 10002,
                        "retMsg": "invalid request, req_timestamp[1700000000000],server_timestamp[1700000006000],recv_window[5000]",
                    },
                )
            return DummyResponse(200, {"retCode": 0, "result": {"list": []}})

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    monkeypatch.setattr(master_service.time, "time", lambda: now["value"])
    payload = asyncio.run(
        master_service._bybit_signed_get(
            base_url="https://api.bybit.com",
            api_key="k",
            api_secret="s",
            path="/v5/order/history",
            params={"category": "linear"},
        )
    )
    order_headers = [headers for url, headers in calls if "/v5/order/history" in url]
    assert payload["retCode"] == 0
    assert market_time_calls["count"] >= 1
    assert len(order_headers) == 2
    assert int(order_headers[1]["X-BAPI-TIMESTAMP"]) >= int(order_headers[0]["X-BAPI-TIMESTAMP"])


def test_bybit_signed_get_persistent_timestamp_error_raises(monkeypatch) -> None:
    class DummyResponse:
        status_code = 200
        content = b"{}"
        text = "{}"

        def json(self):
            return {"retCode": 10002, "retMsg": "invalid request, recv_window[5000]"}

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def get(self, *_args, **_kwargs):
            return DummyResponse()

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    with pytest.raises(ValueError, match="path=/v5/order/history"):
        asyncio.run(
            master_service._bybit_signed_get(
                base_url="https://api.bybit.com",
                api_key="k",
                api_secret="s",
                path="/v5/order/history",
                params={"category": "linear"},
            )
        )


def test_bybit_signed_get_non_timestamp_error_does_not_retry(monkeypatch) -> None:
    requests = {"count": 0}
    master_service._BYBIT_TIME_OFFSET_CACHE.clear()
    master_service._BYBIT_TIME_OFFSET_CACHE["https://api.bybit.com"] = {"synced_at": int(master_service.time.time() * 1000), "offset_ms": 0, "rtt_ms": 0}

    class DummyResponse:
        status_code = 200
        content = b"{}"
        text = "{}"

        def json(self):
            return {"retCode": 10004, "retMsg": "signature error"}

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def get(self, url, *_args, **_kwargs):
            if "/v5/order/history" in str(url):
                requests["count"] += 1
            return DummyResponse()

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    with pytest.raises(ValueError, match="retCode=10004"):
        asyncio.run(
            master_service._bybit_signed_get(
                base_url="https://api.bybit.com",
                api_key="k",
                api_secret="s",
                path="/v5/order/history",
                params={"category": "linear"},
            )
        )
    assert requests["count"] == 1


def test_bybit_signed_post_uses_consistent_recv_window_and_body_on_retry(monkeypatch) -> None:
    sent_headers: list[dict[str, str]] = []
    sent_bodies: list[str] = []
    master_service._BYBIT_TIME_OFFSET_CACHE.clear()

    class DummyResponse:
        def __init__(self, payload: dict[str, object]):
            self.status_code = 200
            self.content = json.dumps(payload).encode("utf-8")
            self._payload = payload

        def raise_for_status(self):
            return None

        def json(self):
            return self._payload

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def post(self, _url, headers=None, content=None):
            sent_headers.append(dict(headers or {}))
            sent_bodies.append(str(content))
            if len(sent_bodies) == 1:
                return DummyResponse({"retCode": 10002, "retMsg": "timestamp expired recv_window[5000]"})
            return DummyResponse({"retCode": 0, "result": {"ok": True}})

        async def get(self, _url, headers=None):
            return DummyResponse({"retCode": 0, "result": {"timeSecond": "1700000010"}})

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    monkeypatch.setattr(master_service, "BYBIT_RECV_WINDOW_MS", 15000)
    body = {"category": "linear", "orderLinkId": "same-link-id"}
    payload = asyncio.run(
        master_service._bybit_signed_post(
            base_url="https://api.bybit.com",
            api_key="k",
            api_secret="s",
            path="/v5/order/create",
            body=body,
        )
    )
    assert payload["retCode"] == 0
    assert len(sent_bodies) == 2
    assert sent_bodies[0] == sent_bodies[1]
    assert sent_headers[0]["X-BAPI-RECV-WINDOW"] == "15000"
    expected_signature = master_service._bybit_sign_request(
        sent_headers[0]["X-BAPI-TIMESTAMP"],
        "k",
        "s",
        sent_bodies[0],
        recv_window="15000",
    )
    assert sent_headers[0]["X-BAPI-SIGN"] == expected_signature


def test_bybit_signed_post_timeout_not_retried(monkeypatch) -> None:
    calls = {"count": 0}

    class DummyClient:
        def __init__(self, *args, **kwargs):
            pass

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def post(self, *_args, **_kwargs):
            calls["count"] += 1
            raise master_service.httpx.TimeoutException("timeout")

    monkeypatch.setattr(master_service.httpx, "AsyncClient", DummyClient)
    with pytest.raises(master_service.httpx.TimeoutException):
        asyncio.run(
            master_service._bybit_signed_post(
                base_url="https://api.bybit.com",
                api_key="k",
                api_secret="s",
                path="/v5/order/create",
                body={"category": "linear", "orderLinkId": "timeout-link"},
            )
        )
    assert calls["count"] == 1


def test_run_closed_pnl_sync_records_error_details(monkeypatch) -> None:
    statuses = []
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _mode: ("demo", "k", "s", "https://api.bybit.com", "env"))
    monkeypatch.setattr(master_service.time, "time", lambda: 1_700_000_000.0)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **kwargs: statuses.append(kwargs))

    async def failing_sync(**_kwargs):
        raise ValueError("Bybit signed GET failed path=/v5/order/history retCode=10002 retMsg=invalid request recv_window[5000]")

    monkeypatch.setattr(master_service, "_sync_bybit_closed_pnl_window", failing_sync)
    with pytest.raises(ValueError):
        asyncio.run(master_service._run_bybit_closed_pnl_sync(account_mode="demo", reason="scheduled"))
    assert statuses
    assert "retCode=10002" in str(statuses[-1].get("last_error"))
    assert "path=/v5/order/history" in str(statuses[-1].get("last_error"))


def test_resolve_local_journal_file_case_insensitive_and_append_reuses_existing(tmp_path, monkeypatch) -> None:
    local_dir = tmp_path / "journal"
    local_dir.mkdir()
    existing = local_dir / "BYBIT DEMO.xlsx"
    pd.DataFrame(columns=master_service.BYBIT_DEMO_WORKBOOK_COLUMNS).to_excel(existing, index=False)

    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", local_dir)
    resolved = master_service._resolve_local_journal_file("Bybit Demo.xlsx")
    assert resolved == existing

    rows = [{"order_id": "oid-1", "symbol": "BTCUSDT"}]
    monkeypatch.setattr(master_service, "_bybit_demo_workbook_row", lambda _row: {col: "" for col in master_service.BYBIT_DEMO_WORKBOOK_COLUMNS} | {"order_id": "oid-1", "symbol": "BTCUSDT"})
    changed = master_service._append_bybit_demo_rows_to_local_workbook(local_dir, rows)
    assert changed == 1
    assert existing.exists()
    matching_entries = [
        path for path in local_dir.iterdir() if path.name.casefold() == "bybit demo.xlsx"
    ]
    assert matching_entries == [existing]

def test_backfill_bybit_demo_balances_reverse_pnl_ordering() -> None:
    rows = [
        {"account": "Bybit Demo", "status": "closed", "close_time": "2026-01-01T00:00:00Z", "net_profit": -1.386334, "raw_refs": {"orderId": "1"}, "currency": "USDT"},
        {"account": "Bybit Demo", "status": "closed", "close_time": "2026-01-04T00:00:00Z", "net_profit": -1.421031, "raw_refs": {"orderId": "4"}, "currency": "USDT"},
        {"account": "Bybit Demo", "status": "closed", "close_time": "2026-01-03T00:00:00Z", "net_profit": 2.153793, "raw_refs": {"orderId": "3"}, "currency": "USDT"},
        {"account": "Bybit Demo", "status": "closed", "close_time": "2026-01-02T00:00:00Z", "net_profit": -3.664636, "raw_refs": {"orderId": "2"}, "currency": "USDT"},
    ]
    out, stats = master_service._backfill_bybit_demo_balances_from_current_balance(rows, {"current_balance": 224.87769878, "snapshot_at": "2026-01-04T12:00:00Z"})
    by_oid = {str((r.get("raw_refs") or {}).get("orderId")): r for r in out}
    assert by_oid["4"]["balance_after_trade"] == pytest.approx(224.87769878)
    assert by_oid["3"]["balance_after_trade"] == pytest.approx(224.87769878 - (-1.421031))
    assert by_oid["2"]["balance_after_trade"] == pytest.approx(by_oid["3"]["balance_after_trade"] - 2.153793)
    assert by_oid["1"]["balance_after_trade"] == pytest.approx(by_oid["2"]["balance_after_trade"] - (-3.664636))
    assert all(r.get("balance_after_trade_currency") == "USDT" for r in out)
    assert all(r.get("balance_source") == "bybit_demo_wallet_reverse_pnl" for r in out)
    assert stats["changed"] is True


def test_sync_demo_persists_balance_only_changes(monkeypatch) -> None:
    seed = [{"account": "Bybit Demo", "status": "closed", "close_time": "2026-01-01T00:00:00Z", "net_profit": 1.0, "raw_refs": {"orderId": "1"}, "currency": "USDT", "balance_after_trade": None}]
    saved = {"rows": None}
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: False)
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: ("/x", []))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda _f: None)
    async def _empty_async(**_k):
        return {"result": {"list": [], "nextPageCursor": ""}}
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty_async)
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", _empty_async)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty_async)
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: list(seed))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0}))
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", lambda: asyncio.sleep(0, result={"current_balance": 10.0, "snapshot_at": "2026-01-01T00:00:00Z"}))
    monkeypatch.setattr(master_service, "_set_trading_journal_rows", lambda rows: saved.update({"rows": rows}))
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_workbook", lambda *_a, **_k: 0)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_a, **_k: {"changed": 0})
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_k: None)
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=1))
    assert saved["rows"] is not None


def test_wallet_snapshot_prefers_total_wallet_when_coin_wallet_missing(monkeypatch) -> None:
    async def fake_signed_get(**_kwargs):
        return {"result": {"list": [{"totalEquity": "100", "totalWalletBalance": "90", "totalAvailableBalance": "80", "coin": [{"coin": "USDT", "availableToTrade": "70"}]}]}}
    monkeypatch.setattr(master_service, "_bybit_signed_get", fake_signed_get)
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _a: ("demo", "k", "s", "https://api-demo.bybit.com", "env"))
    snap = asyncio.run(master_service._fetch_bybit_balance_usdt("demo"))
    assert "wallet_balance_usdt" not in snap


def test_single_file_broker_upsert_bypass_only_for_explicit_flag(monkeypatch):
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    saved={"rows":[]}
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: list(saved["rows"]))
    monkeypatch.setattr(
        master_service,
        "_save_trading_journal",
        lambda rows, **_kwargs: saved.update({"rows": list(rows)}),
    )
    row={"id":"x1","source":"bybit","account":"demo"}
    assert master_service._upsert_trading_journal_rows([row]) == 0
    assert master_service._upsert_trading_journal_rows([row], allow_broker_rows_in_single_file=True) == 1


def test_single_file_mode_no_side_workbook_touches(monkeypatch):
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    calls={"ensure_local":0,"ensure_dropbox":0,"resolve_dropbox":0,"sanitize_local":0,"sanitize_dropbox":0}
    monkeypatch.setattr(master_service, "_ensure_local_bybit_demo_files", lambda *_: calls.__setitem__("ensure_local", calls["ensure_local"]+1))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda *_: calls.__setitem__("ensure_dropbox", calls["ensure_dropbox"]+1))
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: calls.__setitem__("resolve_dropbox", calls["resolve_dropbox"]+1) or ("/tmp", []))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_local_workbook", lambda *_: calls.__setitem__("sanitize_local", calls["sanitize_local"]+1) or {"changed":0})
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_: calls.__setitem__("sanitize_dropbox", calls["sanitize_dropbox"]+1) or {"changed":0})
    async def _empty(**_): return {"result":{"list":[]}}
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", _empty)
    async def _executions(**_):
        return []
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _executions)
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed":0}))
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_: None)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    out=asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=1))
    assert out["rows_seen"] == 0
    assert calls == {"ensure_local":0,"ensure_dropbox":0,"resolve_dropbox":0,"sanitize_local":0,"sanitize_dropbox":0}


def test_single_file_live_does_not_append_side_workbook_and_uses_upsert(monkeypatch):
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    calls={"append_live":0, "upsert_kwargs":None}
    async def _empty(**_): return {"result":{"list":[], "nextPageCursor":""}}
    async def _closed(**_):
        return {"result":{"list":[{"symbol":"BTCUSDT","orderId":"1","side":"Buy","createdTime":1000,"updatedTime":2000,"avgEntryPrice":"1","avgExitPrice":"2","closedSize":"1","closedPnl":"1","openFee":"0","closeFee":"0"}], "nextPageCursor":""}}
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", _closed)
    async def _executions(**_):
        return []
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _executions)
    monkeypatch.setattr(master_service, "_append_generic_local_broker_rows", lambda *_a, **_k: calls.__setitem__("append_live", calls["append_live"]+1) or 0)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_: None)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed":0, "deduped_by_order_id":0, "deduped_by_fingerprint":0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda rows, **k: calls.__setitem__("upsert_kwargs", k) or len(rows))
    out=asyncio.run(
        master_service._sync_bybit_closed_pnl_window(
            account_mode="live",
            base_url="u",
            api_key="k",
            api_secret="s",
            start_time=0,
            end_time=3000,
            allow_broker_rows_in_single_file=True,
        )
    )
    assert calls["append_live"] == 0
    assert calls["upsert_kwargs"] == {"allow_broker_rows_in_single_file": True}
    assert out["rows_seen"] > 0 and out["rows_upserted"] > 0


def test_single_file_live_status_has_no_local_workbook_path(monkeypatch):
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    captured=[]
    async def _empty(**_): return {"result":{"list":[], "nextPageCursor":""}}
    async def _closed(**_):
        return {"result":{"list":[{"symbol":"BTCUSDT","orderId":"1","side":"Buy","createdTime":1000,"updatedTime":2000,"avgEntryPrice":"1","avgExitPrice":"2","closedSize":"1","closedPnl":"1","openFee":"0","closeFee":"0"}], "nextPageCursor":""}}
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", _closed)
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", lambda **_: [])
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **k: captured.append(k))
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed":0, "deduped_by_order_id":0, "deduped_by_fingerprint":0}))
    monkeypatch.setattr(master_service, "_get_trading_journal_rows", lambda: [])
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda rows, **k: len(rows))
    asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="live", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=3000))
    assert any(c.get("last_local_workbook_path") in {None, ""} for c in captured if "last_local_workbook_path" in c)


def test_oanda_single_file_demo_live_no_side_workbook_append(monkeypatch):
    monkeypatch.setattr(master_service, '_trading_journal_local_excel_authoritative', lambda: True)
    monkeypatch.setattr(master_service, '_master_journal_single_file_mode', lambda: True)
    monkeypatch.setattr(master_service, '_get_oanda_config', lambda a: {'mode':a,'base_url':'u','account_id':'i','token':'t'})
    async def _tx(**_): return ([{'id':'1','time':'2026-01-01T00:00:00Z'}], '1')
    monkeypatch.setattr(master_service, '_fetch_oanda_transactions', _tx)
    monkeypatch.setattr(master_service, '_journal_rows_from_oanda_order_fill', lambda e: [{'id':f"o-{e.get('account')}-1",'source':'oanda','account':e.get('account')}])
    calls={'append':0,'upsert':0,'kw':None}
    monkeypatch.setattr(master_service, '_append_generic_local_broker_rows', lambda *a, **k: calls.__setitem__('append', calls['append']+1) or 0)
    monkeypatch.setattr(master_service, '_upsert_trading_journal_rows', lambda rows, **k: calls.__setitem__('upsert', calls['upsert']+len(rows)) or calls.__setitem__('kw', k) or len(rows))
    monkeypatch.setattr(master_service, '_record_oanda_fill_diagnostic', lambda *a, **k: None)
    out_d = asyncio.run(master_service._recover_oanda_recent_fills('demo'))
    out_l = asyncio.run(master_service._recover_oanda_recent_fills('live'))
    assert calls['append'] == 0
    assert calls['upsert'] >= 2
    assert calls['kw'] == {'allow_broker_rows_in_single_file': True}
    assert out_d['rows_seen'] > 0 and out_d['rows_upserted'] > 0
    assert out_l['rows_seen'] > 0 and out_l['rows_upserted'] > 0


def test_bybit_trade_history_csv_parser_full_fixture(tmp_path: Path) -> None:
    headers = [
        "contracts","Order No.","Direction","Order Type","Filled Qty","Filled Price","Order Price","Filled Type",
        "Trading Fee Rate","Fees Paid","Trasaction ID","Transaction Time(UTC+10)","Final Balance (USDT)"
    ]
    lines = [",".join(headers)]
    fills = [
        ("OIDA","Buy",0.019,100000,"EX001","17/05/2026 22:55"),
        ("OIDA","Sell",0.019,100100,"EX002","17/05/2026 22:56"),
    ] + [(f"OIDB{i}","Buy",0.018,100200+i,f"EX{3+i:03d}","19/05/2026 00:52") for i in range(18)] + [
        ("OIDBCLS","Sell",0.324,100260,"EX021","19/05/2026 00:54"),
        ("OIDC","Buy",0.016,100300,"EX022","19/05/2026 01:12"),
        ("OIDC","Sell",0.016,100330,"EX023","19/05/2026 01:13"),
    ]
    for oid, side, qty, px, exid, ts in fills:
        lines.append(f"BTCUSDT,{oid},{side},Market,{qty},{px},{px},Trade,0.00055,0.01,{exid},{ts},1000.0")
    p = tmp_path / "bybit_history_fixture.csv"
    p.write_text("\n".join(lines), encoding="utf-8")
    assert master_service._is_bybit_trade_history_csv(headers) is True
    rows, unmatched, diag = master_service._parse_bybit_trade_history_csv_with_diagnostics(p, account_mode="demo")
    assert len(rows) == 3
    assert unmatched == []
    assert diag["bybit_execution_rows_seen"] == 23
    assert diag["bybit_completed_trades_imported"] == 3
    assert diag["bybit_unmatched_execution_rows"] == 0
    assert [round(float(r["qty"]), 3) for r in rows] == [0.019, 0.324, 0.016]
    assert all(str(r["id"]).startswith("bybit:demo:trade:") for r in rows)

def test_rows_only_bybit_parser_raises_on_unmatched(tmp_path: Path) -> None:
    p = tmp_path / "buy_only.csv"
    p.write_text(
        "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)\n"
        "BTCUSDT,OID1,Buy,Market,0.01,100000,100000,Trade,0.00055,0.01,EX001,17/05/2026 22:55,1000\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="Unmatched Bybit execution rows remain open"):
        master_service._parse_bybit_trade_history_csv(p, account_mode="demo")

def test_parse_iso_datetime_source_has_no_github_sync_terms() -> None:
    import inspect
    src = inspect.getsource(master_service._parse_iso_datetime)
    lowered = src.lower()
    assert "github_sync_result" not in lowered
    assert "sync_journal_excel_files_to_github" not in lowered
    assert "github_sync" not in lowered

def test_bybit_reversal_rows_block_with_explicit_diagnostics(tmp_path: Path) -> None:
    p = tmp_path / "reversal.csv"
    p.write_text(
        "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)\n"
        "BTCUSDT,OID1,Buy,Market,0.01,100000,100000,Trade,0.00055,0.01,EX001,17/05/2026 22:55,1000\n"
        "BTCUSDT,OID2,Sell,Market,0.02,100100,100100,Trade,0.00055,0.01,EX002,17/05/2026 22:56,1000\n",
        encoding="utf-8",
    )
    rows, unmatched, diag = master_service._parse_bybit_trade_history_csv_with_diagnostics(p, account_mode="demo")
    assert rows == []
    assert len(unmatched) >= 1
    assert diag.get("bybit_reversal_execution_rows_seen", 0) >= 1
    assert diag.get("bybit_reversal_import_blocked") is True

def test_manual_import_reversal_response_includes_reversal_diagnostics(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    payload = (
        "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)\n"
        "BTCUSDT,OID1,Buy,Market,0.01,100000,100000,Trade,0.00055,0.01,EX001,17/05/2026 22:55,1000\n"
        "BTCUSDT,OID2,Sell,Market,0.02,100100,100100,Trade,0.00055,0.01,EX002,17/05/2026 22:56,1000\n"
    ).encode("utf-8")
    out = master_service._import_uploaded_trading_journal_file("bybit_history.csv", payload, account_mode="demo")
    assert out["ok"] is False
    assert "unmatched_bybit_executions" in out.get("errors", [])
    assert "zero_rows" not in out.get("errors", [])
    assert out.get("bybit_execution_rows_seen") == 2
    assert out.get("bybit_completed_trades_imported") == 0
    assert out.get("bybit_unmatched_execution_rows", 0) >= 1
    assert out.get("bybit_reversal_import_blocked") is True
    assert out.get("bybit_reversal_execution_rows_seen", 0) >= 1

def test_manual_import_buy_only_response_includes_unmatched_diagnostics(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    payload = (
        "contracts,Order No.,Direction,Order Type,Filled Qty,Filled Price,Order Price,Filled Type,Trading Fee Rate,Fees Paid,Trasaction ID,Transaction Time(UTC+10),Final Balance (USDT)\n"
        "BTCUSDT,OID1,Buy,Market,0.01,100000,100000,Trade,0.00055,0.01,EX001,17/05/2026 22:55,1000\n"
    ).encode("utf-8")
    out = master_service._import_uploaded_trading_journal_file("bybit_history.csv", payload, account_mode="demo")
    assert out["ok"] is False
    assert "unmatched_bybit_executions" in out.get("errors", [])
    assert "zero_rows" not in out.get("errors", [])
    assert out.get("bybit_execution_rows_seen") == 1
    assert out.get("bybit_completed_trades_imported") == 0
    assert out.get("bybit_unmatched_execution_rows", 0) >= 1

def test_epoch_or_iso_to_iso_dayfirst_au_date_no_warning():
    import warnings
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        out = master_service._epoch_or_iso_to_iso("17/05/2026 22:55")
    assert out is not None
    assert str(out).startswith("2026-05-17")
    assert not any("dayfirst=False" in str(w.message) for w in caught)


def test_sync_bybit_closed_pnl_window_uses_execution_rows_when_closed_pnl_empty(monkeypatch) -> None:
    monkeypatch.setattr(master_service, "_trading_journal_local_excel_authoritative", lambda: False)
    monkeypatch.setattr(master_service, "_resolve_trading_journal_dropbox_folder", lambda: ("/tmp", []))
    monkeypatch.setattr(master_service, "_ensure_bybit_demo_dropbox_files", lambda *_: None)
    async def _empty(**_): return {"result": {"list": [], "nextPageCursor": ""}}
    async def _execs(**_):
        return [
            {"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E1", "execPrice": "100", "execQty": "0.1", "execFee": "0.01", "feeRate": "0.00055", "execTime": "1715900000000", "side": "Buy"},
            {"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E2", "execPrice": "101", "execQty": "0.1", "execFee": "0.01", "feeRate": "0.00055", "execTime": "1715900001000", "side": "Buy"},
        ]
    monkeypatch.setattr(master_service, "_fetch_bybit_order_history", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_order_realtime", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_transaction_log", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_closed_pnl", _empty)
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _execs)
    captured = {"rows": None}
    monkeypatch.setattr(master_service, "_upsert_trading_journal_rows", lambda rows, **_k: captured.update({"rows": rows}) or len(rows))
    monkeypatch.setattr(master_service, "_append_bybit_demo_rows_to_workbook", lambda *_a, **_k: 0)
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_workbook", lambda *_a, **_k: {"changed": 0})
    monkeypatch.setattr(master_service, "_sanitize_bybit_demo_rows", lambda rows: (rows, {"changed": 0, "deduped_by_order_id": 0, "deduped_by_fingerprint": 0}))
    monkeypatch.setattr(
        master_service,
        "_get_trading_journal_rows",
        lambda: list(captured.get("rows") or []),
    )
    monkeypatch.setattr(master_service, "_record_bybit_demo_sync_status", lambda **_k: None)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda: None)
    monkeypatch.setattr(master_service, "_fetch_bybit_demo_current_balance_snapshot", lambda: asyncio.sleep(0, result={}))
    out = asyncio.run(master_service._sync_bybit_closed_pnl_window(account_mode="demo", base_url="u", api_key="k", api_secret="s", start_time=0, end_time=1))
    assert out["execution_rows_seen"] == 2
    assert out["execution_rows_upserted"] == 2
    assert out["rows_seen"] == 2
    assert captured["rows"] is not None and len(captured["rows"]) == 2
    ids = [r["id"] for r in captured["rows"]]
    assert ids == ["bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"]
    assert all(r.get("net_profit") is None for r in captured["rows"])


def test_fetch_bybit_executions_chunked_splits_large_windows(monkeypatch) -> None:
    calls = []
    async def _fake_fetch(**kwargs):
        calls.append((kwargs["start_time"], kwargs["end_time"]))
        suffix = str(len(calls))
        return [{"execId": f"E{suffix}", "orderId": "O1", "symbol": "BTCUSDT", "execTime": str(kwargs["start_time"])}]
    monkeypatch.setattr(master_service, "_fetch_bybit_executions", _fake_fetch)
    start_time = 1777816800000  # 2026-05-04T00:00:00+10:00
    end_time = 1779199199000    # 2026-05-19T23:59:59+10:00
    rows = asyncio.run(master_service._fetch_bybit_executions_chunked(base_url="u", api_key="k", api_secret="s", category="linear", start_time=start_time, end_time=end_time))
    assert len(calls) > 1
    assert all((e - s) <= ((7 * 24 * 60 * 60 * 1000) - 1) for s, e in calls)
    assert len(rows) == len(calls)


def test_manual_sync_fails_when_captured_bybit_rows_missing_from_workbook(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(
        master_service,
        "_sync_master_journal_workbook",
        lambda *_args, **_kwargs: {"master_journal_ok": True},
    )
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", lambda *_a, **_k: asyncio.sleep(0, result={"available_usdt": 1000}))
    monkeypatch.setattr(master_service, "_fetch_oanda_account_summary", lambda *_a, **_k: asyncio.sleep(0, result={"balance": 1000, "nav": 1000, "currency": "AUD"}))
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 2, "captured_row_ids": ["bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"], "execution_rows_seen": 2, "execution_rows_normalized": 2, "latest_execution_time": "2026-05-19T01:13:00+10:00"}))
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    st = master_service._sync_state_snapshot()
    assert st.get("ok") is False
    err = str(st.get("error") or "") + str(st.get("message") or "")
    assert "not persisted to Trade Log" in err


def test_manual_sync_fails_when_bybit_execution_prefetch_fails(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(
        master_service,
        "_sync_master_journal_workbook",
        lambda *_args, **_kwargs: {"master_journal_ok": True},
    )
    monkeypatch.setattr(master_service, "_trading_journal_broker_refresh_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_fetch_bybit_balance_usdt", lambda *_a, **_k: asyncio.sleep(0, result={"available_usdt": 1000}))
    monkeypatch.setattr(master_service, "_fetch_oanda_account_summary", lambda *_a, **_k: asyncio.sleep(0, result={"balance": 1000, "nav": 1000, "currency": "AUD"}))
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", lambda *a, **k: asyncio.sleep(0, result={"ok": False, "rows_seen": 0, "execution_rows_seen": 0, "error": "Bybit execution prefetch failed: Bybit execution API failed"}))
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    st = master_service._sync_state_snapshot()
    assert st.get("ok") is False
    assert "Bybit execution prefetch failed" in (str(st.get("error") or "") + str(st.get("message") or ""))


def test_run_bybit_closed_pnl_sync_propagates_execution_diagnostics(monkeypatch) -> None:
    now_s = 1_800_000_000.0
    monkeypatch.setattr(master_service.time, "time", lambda: now_s)
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda _mode: ("demo", "k", "s", "https://api.bybit.com", "env"))
    monkeypatch.setattr(master_service, "_persist_bybit_closed_pnl_last_seen", lambda: None)
    master_service._BYBIT_CLOSED_PNL_LAST_SEEN["demo"] = None
    async def _fake_sync(**_kwargs):
        return {
            "max_seen": 123,
            "rows_seen": 2,
            "rows_upserted": 2,
            "captured_row_ids": ["bybit:demo:execution:BTCUSDT:E1"],
            "execution_rows_seen": 2,
            "execution_rows_normalized": 2,
            "execution_rows_upserted": 1,
            "latest_execution_time": "2026-05-19T01:13:00+10:00",
            "execution_fetch_error": None,
        }
    monkeypatch.setattr(master_service, "_sync_bybit_closed_pnl_window", _fake_sync)
    out = asyncio.run(master_service._run_bybit_closed_pnl_sync(reason="manual", account_mode="demo"))
    assert out["execution_rows_seen"] == 2
    assert out["execution_rows_normalized"] == 2
    assert out["execution_rows_upserted"] == 1
    assert out["latest_execution_time"] == "2026-05-19T01:13:00+10:00"
    assert out["execution_fetch_error"] is None


def test_manual_sync_writes_execution_ids_to_master_workbook(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL", True)
    monkeypatch.setattr(master_service, "_bybit_demo_credentials_available", lambda: True)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    async def _fake_bybit_sync(*_a, **_k):
        rows = [
            {"id": "bybit:demo:execution:BTCUSDT:E1", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-17T01:13:00+10:00", "close_time": "2026-05-17T01:13:00+10:00", "qty": 0.1, "entry_price": 100000, "exit_price": 100000, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
            {"id": "bybit:demo:execution:BTCUSDT:E2", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-19T01:13:00+10:00", "close_time": "2026-05-19T01:13:00+10:00", "qty": 0.1, "entry_price": 100001, "exit_price": 100001, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
        ]
        return {"ok": True, "rows_seen": 2, "rows_upserted": 2, "captured_row_ids": [r["id"] for r in rows], "captured_rows": rows, "execution_rows_seen": 2, "execution_rows_normalized": 2, "execution_rows_upserted": 2, "latest_execution_time": "2026-05-19T01:13:00+10:00"}
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", _fake_bybit_sync)
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    wb = load_workbook(tmp_path / "Trading Journal.xlsx", data_only=True)
    ws = wb["Trade Log"]
    headers = [str(c.value or "") for c in ws[1]]
    ridx = headers.index("Row ID") + 1
    cidx = headers.index("Close Time") + 1
    ids = [str(ws.cell(r, ridx).value or "").strip() for r in range(2, ws.max_row + 1)]
    assert "bybit:demo:execution:BTCUSDT:E1" in ids
    assert "bybit:demo:execution:BTCUSDT:E2" in ids
    close_vals = [
        value.isoformat() if hasattr(value, "isoformat") else str(value or "")
        for value in (ws.cell(r, cidx).value for r in range(2, ws.max_row + 1))
    ]
    assert any(str(v or "").startswith("2026-05-19") for v in close_vals)


def test_manual_sync_real_path_with_demo_credentials_writes_execution_ids_without_manual_override_flag(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BROKER_REFRESH_ENABLED", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL", False)
    monkeypatch.setattr(master_service, "_bybit_demo_credentials_available", lambda: True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda mode: (mode, "k", "s", "https://api-demo.bybit.com" if mode == "demo" else "https://api.bybit.com", "KEY2" if mode == "demo" else "KEY1"))
    monkeypatch.setattr(master_service, "_persist_bybit_closed_pnl_last_seen", lambda: None)
    async def _fake_window(**kwargs):
        if kwargs.get("account_mode") == "demo":
            rows = [
                {"id": "bybit:demo:execution:BTCUSDT:E1", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-17T01:13:00+10:00", "close_time": "2026-05-17T01:13:00+10:00", "qty": 0.1, "entry_price": 100000, "exit_price": 100000, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
                {"id": "bybit:demo:execution:BTCUSDT:E2", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-19T01:13:00+10:00", "close_time": "2026-05-19T01:13:00+10:00", "qty": 0.1, "entry_price": 100001, "exit_price": 100001, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
            ]
            return {"max_seen": kwargs["end_time"], "rows_seen": 2, "rows_upserted": 2, "captured_row_ids": [r["id"] for r in rows], "captured_rows": rows, "execution_rows_seen": 2, "execution_rows_normalized": 2, "execution_rows_upserted": 2, "latest_execution_time": "2026-05-19T01:13:00+10:00", "execution_fetch_error": None, "missing_execution_row_ids": []}
        return {"max_seen": kwargs["end_time"], "rows_seen": 0, "rows_upserted": 0, "captured_row_ids": []}
    monkeypatch.setattr(master_service, "_sync_bybit_closed_pnl_window", _fake_window)
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    asyncio.run(master_service._run_trading_journal_sync_job())
    wb = load_workbook(tmp_path / "Trading Journal.xlsx", data_only=True)
    ws = wb["Trade Log"]
    headers = [str(c.value or "") for c in ws[1]]
    ridx = headers.index("Row ID") + 1
    cidx = headers.index("Close Time") + 1
    ids = [str(ws.cell(r, ridx).value or "").strip() for r in range(2, ws.max_row + 1)]
    assert "bybit:demo:execution:BTCUSDT:E1" in ids and "bybit:demo:execution:BTCUSDT:E2" in ids
    close_vals = [
        value.isoformat() if hasattr(value, "isoformat") else str(value or "")
        for value in (ws.cell(r, cidx).value for r in range(2, ws.max_row + 1))
    ]
    assert any(str(v or "").startswith("2026-05-19") for v in close_vals)


def test_manual_sync_missing_demo_credentials_sets_verification_false(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_ENABLE_LOCAL_IMPORT", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BROKER_REFRESH_ENABLED", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL", True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "resolve_bybit_credentials_for", lambda mode: (mode, "", "", "https://api-demo.bybit.com" if mode == "demo" else "https://api.bybit.com", "NONE"))
    monkeypatch.setattr(
        master_service,
        "describe_bybit_credentials_for",
        lambda mode: {
            "mode": mode,
            "base_url": "https://api-demo.bybit.com" if mode == "demo" else "https://api.bybit.com",
            "key_source": "NONE",
            "credentials_available": False,
            "missing_env_vars": ["BYBIT_API_KEY", "BYBIT_API_SECRET"],
        },
    )
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    st = master_service._sync_state_snapshot()
    assert st.get("ok") is True
    assert st.get("message") == "Completed with warnings"
    demo = ((st.get("result") or {}).get("bybit") or {}).get("demo") or {}
    assert demo.get("ok") is False
    assert "Bybit Demo API credentials are not configured" in str(demo.get("error") or st.get("error") or "")
    assert demo.get("final_trade_log_row_ids_verified") is False
    assert int(demo.get("execution_rows_seen") or 0) == 0
    assert demo.get("latest_execution_time") in (None, "")


def test_manual_sync_bridges_captured_rows_into_master_workbook(tmp_path: Path, monkeypatch) -> None:
    from tools.master_journal_workbook import build_master_journal_workbook
    _isolate_trading_journal_runtime(monkeypatch, tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", tmp_path)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_BROKER_REFRESH_ENABLED", False)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_CALCULATOR_TRADES_ON_MANUAL", True)
    monkeypatch.setattr(master_service, "_bybit_demo_credentials_available", lambda: True)
    build_master_journal_workbook({"items": [], "stats": {"totals": {}, "groups": {}}, "balances": []}, tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *a, **k: {"ok": True, "rows_imported": 0, "rows_by_asset_class": {}, "local_workbooks_seen": 1, "dropbox_workbooks_seen": 0})
    rows = [
        {"id": "bybit:demo:execution:BTCUSDT:E1", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-17T01:13:00+10:00", "close_time": "2026-05-17T01:13:00+10:00", "qty": 0.1, "entry_price": 100000, "exit_price": 100000, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
        {"id": "bybit:demo:execution:BTCUSDT:E2", "row_type": "trade", "source": "bybit_execution_history", "account": "Bybit Demo", "account_label": "Bybit Demo", "symbol": "BTCUSDT", "side": "Buy", "open_time": "2026-05-19T01:13:00+10:00", "close_time": "2026-05-19T01:13:00+10:00", "qty": 0.1, "entry_price": 100001, "exit_price": 100001, "asset_class": "crypto", "balance_after_trade": 1000.0, "balance_after_trade_source": "master_journal", "currency": "USDT"},
    ]
    async def _fake_bybit(*_a, **_k):
        return {"ok": True, "rows_seen": 2, "rows_upserted": 2, "captured_row_ids": [r["id"] for r in rows], "captured_rows": rows}
    monkeypatch.setattr(master_service, "_run_bybit_closed_pnl_sync", _fake_bybit)
    monkeypatch.setattr(master_service, "_recover_oanda_recent_fills", lambda *a, **k: asyncio.sleep(0, result={"ok": True, "rows_seen": 0, "captured_row_ids": []}))
    asyncio.run(master_service._run_trading_journal_sync_job())
    wb = load_workbook(tmp_path / "Trading Journal.xlsx", data_only=True)
    ws = wb["Trade Log"]
    headers = [str(c.value or "") for c in ws[1]]
    ridx = headers.index("Row ID") + 1
    ids = [str(ws.cell(r, ridx).value or "").strip() for r in range(2, ws.max_row + 1)]
    assert "bybit:demo:execution:BTCUSDT:E1" in ids
    assert "bybit:demo:execution:BTCUSDT:E2" in ids


def test_sanitize_equal_time_execution_row_survives() -> None:
    raw = {"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E1", "execQty": "0.1", "execPrice": "100", "execTime": "1779199199000", "side": "Buy"}
    row = master_service._normalize_bybit_execution_history_row(raw, "demo")
    assert row["open_time"] == row["close_time"]
    sanitized, stats = master_service._sanitize_bybit_demo_rows([row])
    assert len(sanitized) == 1
    assert str(sanitized[0].get("row_type") or "").lower() != "quarantine"
    assert stats["quarantined_invalid_time"] == 0
    assert sanitized[0]["id"] == "bybit:demo:execution:BTCUSDT:E1"


def test_sanitize_same_order_multiple_exec_ids_preserved() -> None:
    r1 = master_service._normalize_bybit_execution_history_row({"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E1", "execQty": "0.1", "execPrice": "100", "execTime": "1779199199000", "side": "Buy"}, "demo")
    r2 = master_service._normalize_bybit_execution_history_row({"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E2", "execQty": "0.1", "execPrice": "101", "execTime": "1779199199000", "side": "Buy"}, "demo")
    sanitized, stats = master_service._sanitize_bybit_demo_rows([r1, r2])
    assert len(sanitized) == 2
    assert stats["deduped_by_order_id"] == 0
    assert {x["id"] for x in sanitized} == {"bybit:demo:execution:BTCUSDT:E1", "bybit:demo:execution:BTCUSDT:E2"}


def test_sanitize_duplicate_execution_id_is_deduped() -> None:
    row = master_service._normalize_bybit_execution_history_row(
        {"symbol": "BTCUSDT", "orderId": "OID-1", "execId": "E1", "execQty": "0.1", "execPrice": "100", "execTime": "1779199199000", "side": "Buy"},
        "demo",
    )

    sanitized, stats = master_service._sanitize_bybit_demo_rows([row, dict(row)])

    assert len(sanitized) == 1
    assert stats["deduped_by_fingerprint"] == 1


def test_sanitize_full_23_row_csv_fixture_survives(tmp_path: Path) -> None:
    headers = ["contracts","Order No.","Direction","Order Type","Filled Qty","Filled Price","Order Price","Filled Type","Trading Fee Rate","Fees Paid","Trasaction ID","Transaction Time(UTC+10)","Final Balance (USDT)"]
    lines = [",".join(headers)]
    for i in range(23):
        lines.append(f"BTCUSDT,OID-1,Buy,Market,0.001,100000,100000,Trade,0.00055,0.01,EX{i:03d},2026-05-19 01:13:00+10:00,1000.{i}")
    p = tmp_path / "hist.csv"
    p.write_text("\n".join(lines), encoding="utf-8")
    records = pd.read_csv(p, encoding="utf-8-sig").to_dict(orient="records")
    rows = [
        row
        for index, record in enumerate(records, start=2)
        if (
            row := master_service._normalize_bybit_execution_history_row(
                record,
                "demo",
                str(p),
                index,
            )
        )
    ]
    sanitized, stats = master_service._sanitize_bybit_demo_rows(rows)
    assert len(sanitized) == 23
    assert stats["quarantined_invalid_time"] == 0
    assert stats["deduped_by_order_id"] == 0
    assert stats["trade_group_merged"] == 0


def test_lookup_trade_context_supports_grouped_order_ids(monkeypatch):
    contexts = [{"order_id":"oid-2","stop_loss":1,"take_profit":2,"timeframe":"1m","is_test_trade":True}]
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: contexts)
    row = {"raw_refs": {"order_ids": ["oid-1", "oid-2"]}}
    ctx = master_service._lookup_trade_context_for_journal_row(row)
    assert ctx and ctx.get("order_id") == "oid-2"


def test_trade_duration_rounding_helper():
    assert master_service._round_trade_duration_seconds(60.1) == 60
    assert master_service._round_trade_duration_seconds(60.5) == 61
    assert master_service._round_trade_duration_seconds(0.2) == 1
    assert master_service._round_trade_duration_seconds(60) == 60


def test_lookup_trade_context_for_open_item_supports_order_ids(monkeypatch):
    contexts = [{"order_id": "oid-2", "timeframe": "1m"}]
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: contexts)
    item = {"raw_refs": {"order_ids": ["oid-1", "oid-2"]}}
    matched = master_service._lookup_trade_context_for_open_item(item)
    assert matched and matched.get("order_id") == "oid-2"


def test_backfill_grouped_row_populates_context_and_r_multiple(monkeypatch):
    row = {
        "id": "bybit:demo:trade:BTCUSDT:x",
        "source": "bybit_execution_history_grouped",
        "symbol": "BTCUSDT",
        "side": "Buy",
        "entry_price": 100.0,
        "exit_price": 101.0,
        "raw_refs": {"order_ids": ["oid-2"]},
    }
    ctx = {"order_id": "oid-2", "stop_loss": 99.0, "take_profit": 103.0, "timeframe": "1m", "is_test_trade": True}
    monkeypatch.setattr(master_service, "_lookup_trade_context_for_journal_row", lambda _r: ctx)
    monkeypatch.setattr(master_service, "_lookup_trade_context_by_market_window", lambda *_a, **_k: None)
    patched = master_service._backfill_trade_row_context_fields(row)
    assert patched["stop_loss"] == 99.0
    assert patched["take_profit"] == 103.0
    assert patched["timeframe"] == "1m"
    assert patched["is_test_trade"] is True
    assert patched.get("r_multiple") not in (None, "")


def test_grouped_order_ids_ambiguous_context_refuses_match(monkeypatch):
    contexts = [{"order_id": "oid-1"}, {"order_id": "oid-2"}]
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: contexts)
    row = {"raw_refs": {"order_ids": ["oid-1", "oid-2"]}}
    assert master_service._lookup_trade_context_for_journal_row(row) is None


def test_backfill_persisted_grouped_row_populates_context_and_is_idempotent(monkeypatch):
    base_row = {
        "id": "bybit:demo:trade:BTCUSDT:abc",
        "row_type": "trade",
        "source": "bybit_execution_history_grouped",
        "symbol": "BTCUSDT",
        "side": "Buy",
        "entry_price": 100.0,
        "exit_price": 101.0,
        "open_time": "2026-05-26T09:53:00+10:00",
        "close_time": "2026-05-26T09:54:00+10:00",
        "raw_refs": {"order_ids": ["oid-2"]},
    }
    contexts = [{"order_id": "oid-2", "stop_loss": 99.0, "take_profit": 103.0, "timeframe": "1m", "is_test_trade": True}]
    monkeypatch.setattr(master_service, "_load_trade_contexts", lambda: contexts)

    first_rows, first_changed = master_service._backfill_persisted_bybit_trade_fields([base_row])
    assert first_changed == 1
    first = first_rows[0]
    assert first["stop_loss"] == 99.0
    assert first["take_profit"] == 103.0
    assert first["timeframe"] == "1m"
    assert first["is_test_trade"] is True
    assert first.get("r_multiple") not in (None, "")

    second_rows, second_changed = master_service._backfill_persisted_bybit_trade_fields(first_rows)
    assert second_changed == 0
    assert second_rows == first_rows
