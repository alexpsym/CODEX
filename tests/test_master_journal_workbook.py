from pathlib import Path
from collections import defaultdict
from copy import copy
from datetime import datetime
import shutil
import re
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.views import Selection
import pytest
from tools.master_journal_workbook import build_master_journal_workbook, read_master_journal_manual_overrides, read_master_journal_source, update_master_journal_workbook_data_only, SHEET_ORDER, STATS1_SHEET, STATS2_SHEET, SYMBOLS_SHEET, TARGET_R_METADATA_SHEET, TRADE_LOG_HEADERS, TRADE_LOG_HEADERS_V1, OLD_TRADE_LOG_HEADERS, PRE_MOVE_TRADE_LOG_HEADERS, MOVE_TO_FIELD_MAP, TRADE_LOG_HEADER_ROWS, TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_ROW_HEIGHT, TRADE_LOG_FILTER_HEADER_ROW, TRADE_NUMBER_HEADER, STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_INSUFFICIENT, REPORT_YEARLY_SHEET, REPORT_METRIC_LABELS, INSTRUMENT_AVERAGES_HEADERS, INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, INSTRUMENT_AVERAGES_DATA_START_ROW, DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL, DASHBOARD_MOVE_TO_PROFIT_LABEL, PROFIT_FILL, LOSS_FILL, DURATION_NUMBER_FORMAT, adaptive_percent_number_format, adaptive_number_format, resolve_trade_folder_link, expected_report_sheet_names, _apply_trade_number_hyperlinks, _ensure_trade_log_schema, _ensure_instrument_averages_schema, _ensure_symbols_freeze_panes, _ensure_pnl_calendar_freeze_panes, _repair_trade_log_move_to_durations, _trade_log_header_map, _instrument_averages_header_map, _result_percentage_totals_by_market
from tools.master_journal_workbook import _format_duration_display, _parse_duration_text, _repair_legacy_duration_number_formats, _populate_symbols_metrics_preserving_layout, _repair_symbols_header_merges_preserving_layout
from tools.master_journal_workbook import RECOMMENDATION_TRADE_LOG_HEADERS, _trade_log_three_row_header_values_for
from tools.master_journal_workbook import _period_drawdown_metrics, _sanitize_recommendation_text
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter, range_boundaries

def _cf_ranges(ws):
    return [str(k.sqref) for k in ws.conditional_formatting._cf_rules.keys()]

def _cell_covered(ranges, cell):
    row, col = coordinate_to_tuple(cell)
    from openpyxl.utils.cell import range_boundaries
    for rg in ranges:
        for part in rg.split():
            min_col, min_row, max_col, max_row = range_boundaries(part)
            if min_row <= row <= max_row and min_col <= col <= max_col:
                return True
    return False



def _header_col(ws, name: str) -> int:
    if ws.title in {"Trade Log", "All Trades"}:
        return _trade_log_header_map(ws)[name]
    if ws.title in {SYMBOLS_SHEET, "Instrument Averages"}:
        return _instrument_averages_header_map(ws)[name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return headers.index(name) + 1


def _assert_target_recommendation_is_numeric_or_insufficient(value: object) -> None:
    text = str(value or "").strip()
    if text in {TARGET_RECOMMENDATION_INSUFFICIENT, "No eligible losing trades"}:
        return
    assert "Recommended:" in text
    assert text not in {"Reduce target", "Increase target", "Keep target"}
    number_text = text.split("Recommended:", 1)[1].split("R", 1)[0].strip()
    assert number_text
    if "." in number_text:
        assert len(number_text.split(".", 1)[1]) <= 2
    number = float(number_text)
    assert number > 0


def _move_group_ranges(ws):
    header_map = _trade_log_header_map(ws)
    be_cols = [header_map[h] for h in list(MOVE_TO_FIELD_MAP.keys())[:5]]
    profit_cols = [header_map[h] for h in list(MOVE_TO_FIELD_MAP.keys())[5:]]
    return (
        f"{get_column_letter(min(be_cols))}1:{get_column_letter(max(be_cols))}1",
        f"{get_column_letter(min(profit_cols))}1:{get_column_letter(max(profit_cols))}1",
    )


def _trade_data_row(index: int = 0) -> int:
    return TRADE_LOG_DATA_START_ROW + index

def _dashboard_account_balances(ws):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row, col).value or "").strip().lower() != "account balances":
                continue
            header_row = row + 1
            headers = {str(ws.cell(header_row, c).value or "").strip().lower(): c for c in range(col, ws.max_column + 1)}
            account_col = headers["account"]
            balance_col = headers["balance"]
            currency_col = headers["currency"]
            return {
                str(ws.cell(r, account_col).value or "").strip(): (ws.cell(r, balance_col).value, ws.cell(r, currency_col).value)
                for r in range(header_row + 1, ws.max_row + 1)
                if str(ws.cell(r, account_col).value or "").strip()
            }
    return {}


def _ensure_trade_log_headers(wb) -> None:
    ws = wb["Trade Log"] if "Trade Log" in wb.sheetnames else wb.create_sheet("Trade Log")
    existing = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    while existing and not existing[-1]:
        existing.pop()
    if existing and existing == TRADE_LOG_HEADERS_V1[:len(existing)]:
        ws.insert_cols(1)
        existing = []
    if not existing or existing == TRADE_LOG_HEADERS[:len(existing)]:
        for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
            ws.cell(1, col, header)
    _ensure_trade_log_schema(ws)
    instrument_name = SYMBOLS_SHEET if SYMBOLS_SHEET in wb.sheetnames else "Instrument Averages"
    if instrument_name in wb.sheetnames:
        inst = wb[instrument_name]
        _ensure_instrument_averages_schema(inst)


def _sheetnames_without_target_r_metadata(wb) -> list[str]:
    return [name for name in wb.sheetnames if name != TARGET_R_METADATA_SHEET]


def _all_rule_colors(ws):
    out = []
    for rules in ws.conditional_formatting._cf_rules.values():
        for rule in rules:
            dxf = getattr(rule, "dxf", None)
            fill = getattr(getattr(dxf, "fill", None), "fgColor", None)
            font = getattr(getattr(dxf, "font", None), "color", None)
            out.append(((fill.rgb or "") if fill else "", (font.rgb or "") if font else ""))
    return out


def _custom_number_formats(path: Path):
    with zipfile.ZipFile(path) as zf:
        root = ET.fromstring(zf.read("xl/styles.xml"))
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    return [
        str(node.attrib.get("formatCode") or "")
        for node in root.findall(".//x:numFmt", ns)
    ]


def sample_snapshot():
    return {
        'updated_at':'2026-05-10T00:00:00Z',
        'items':[
            {'id':'t1','row_type':'trade','symbol':'EURUSD','asset_class':'fx','side':'BUY','open_time':'2026-05-01T00:00:00Z','close_time':'2026-05-01T01:00:00Z','net_profit':120.5,'result_pct':2.3,'r_multiple':1.2,'stop_loss':1.09,'take_profit':1.12,'entry_price':1.1,'trade_duration_seconds':3700,'analysis_balance_after_trade':1000,'account':'OANDA DEMO','setup':'S1'},
            {'id':'t2','row_type':'trade','symbol':'BTCUSDT','asset_class':'crypto','side':'SELL','open_time':'2026-05-02T00:00:00Z','close_time':'2026-05-02T02:00:00Z','net_profit':-50.0,'result_pct':-1.1,'r_multiple':-0.8,'stop_loss':61000,'take_profit':59000,'entry_price':60000,'trade_duration_seconds':7215,'analysis_balance_after_trade':950,'account':'BYBIT','setup':'S2'},
        ],
        'stats':{'totals':{'trades':2,'wins':1,'losses':1,'break_even':0,'win_rate_pct':50.0,'net_profit_total':70.5,'gross_gain':120.5,'gross_loss':50.0,'money_by_currency':{'net_profit_total':{'AUD':70.5},'gross_gain':{'AUD':120.5},'gross_loss':{'AUD':50.0},'max_gain':{'AUD':120.5},'max_loss':{'AUD':50.0}}},'groups':{'by_market':{'overall':{'trades':2,'wins':1,'losses':1,'break_even':0,'win_rate_pct':50.0,'net_profit_total':70.5,'gross_gain':120.5,'gross_loss':50.0,'avg_result_pct':0.6,'min_result_pct':-1.1,'max_result_pct':2.3,'avg_r_multiple':0.2,'min_r_multiple':-0.8,'max_r_multiple':1.2,'max_gain':120.5,'max_loss':50.0,'avg_stop_pct':1.1,'avg_target_pct':2.2,'avg_duration_seconds':5457,'money_by_currency':{'net_profit_total':{'AUD':70.5},'gross_gain':{'AUD':120.5},'gross_loss':{'AUD':50.0},'max_gain':{'AUD':120.5},'max_loss':{'AUD':50.0}},'metric_sources':{'min_result_pct':{'symbol':'BTCUSDT','date':'2026-05-02'},'max_result_pct':{'symbol':'EURUSD','date':'2026-05-01'}}},'fx':{},'crypto':{}},'risk_expectancy':{'avg_stop_pct_winners':1,'avg_stop_pct_losers':2,'avg_target_pct_winners':3,'avg_target_pct_losers':4,'avg_result_pct_winners':2.3,'avg_result_pct_losers':-1.1,'avg_r_multiple_winners':1.2,'avg_r_multiple_losers':-0.8,'max_drawdown_pct':5,'avg_drawdown_pct':2},'duration':{'overall_avg_seconds':5457,'overall_shortest_seconds':3700,'overall_longest_seconds':7215,'fx_shortest_seconds':3700,'fx_longest_seconds':3700,'crypto_shortest_seconds':7215,'crypto_longest_seconds':7215},'leaders':{}},'by_instrument':[{'symbol':'EURUSD','asset_class':'fx','total_trades':1,'long_trades':1,'short_trades':0,'wins':1,'losses':0,'break_even':0,'net_profit_total':120.5,'avg_net_profit':120.5,'win_rate_pct':100,'avg_sl_pct_wins':1,'avg_sl_pct_losses':None,'avg_tp_pct_wins':2,'avg_tp_pct_losses':None,'avg_trade_duration_seconds':3700,'min_trade_duration_seconds':3700,'max_trade_duration_seconds':3700}]},
        'balances':[{'account_label':'BYBIT','balance':10.123456789,'currency':'USDT','as_of':'2026-05-10'}]
    }


def test_single_builder_definition():
    src = Path('tools/master_journal_workbook.py').read_text(encoding='utf-8')
    assert src.count('def build_master_journal_workbook') == 1


def test_recommendation_sanitizer_hides_internal_reason_keys():
    raw = (
        "Decrease stop \u2014 Recommended: 0.99% "
        "(0.01 pp below loss average; exact_tie_goal_preference_decrease; "
        "unexpected_debug_reason_code)"
    )
    cleaned = _sanitize_recommendation_text(STOP_RECOMMENDATION_HEADER, raw)
    assert "exact tie, so a small decrease is preferred" in cleaned
    assert "unexpected debug reason code" in cleaned
    assert "exact_tie_goal_preference_decrease" not in cleaned
    assert "unexpected_debug_reason_code" not in cleaned


def test_stats_and_symbols_fall_back_to_trade_rows_when_stats_are_empty(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["stats"] = {
        "totals": {},
        "groups": {
            "by_market": {"overall": {}, "fx": {}, "crypto": {}},
            "risk_expectancy": {},
            "duration": {},
            "leaders": {},
        },
        "by_instrument": [],
    }
    out = tmp_path / "trade-row-derived-stats.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    stats1 = wb[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or "").strip(): row for row in range(1, stats1.max_row + 1)}

    assert stats1.cell(labels["Trades"], 2).value == 2
    assert stats1.cell(labels["Trades"], 3).value == 1
    assert stats1.cell(labels["Trades"], 4).value == 1
    assert stats1.cell(labels["R expectancy"], 2).value == pytest.approx(0.2)
    assert stats1.cell(labels["R expectancy"], 3).value == pytest.approx(1.2)
    assert stats1.cell(labels["R expectancy"], 4).value == pytest.approx(-0.8)

    symbols = wb[SYMBOLS_SHEET]
    headers = _instrument_averages_header_map(symbols)
    assert headers["Longs"] != headers["Shorts"]
    symbol_rows = {
        str(symbols.cell(row, headers["Symbol"]).value or "").strip().upper(): row
        for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1)
        if str(symbols.cell(row, headers["Symbol"]).value or "").strip()
    }
    assert {"EURUSD", "BTCUSDT"}.issubset(symbol_rows)
    assert symbols.cell(symbol_rows["EURUSD"], headers["Trades"]).value == 1
    assert symbols.cell(symbol_rows["EURUSD"], headers["Longs"]).value == 1
    assert symbols.cell(symbol_rows["BTCUSDT"], headers["Shorts"]).value == 1
    assert headers["Avg stop %"] == headers["Win Rate %"] + 1
    assert headers["Avg stop % (W)"] == headers["Avg stop %"] + 1
    assert headers[STOP_RECOMMENDATION_HEADER] == headers["Avg stop % (L)"] + 1
    assert headers["Avg target %"] == headers[STOP_RECOMMENDATION_HEADER] + 1
    assert headers["Avg target % (W)"] == headers["Avg target %"] + 1
    assert headers[TARGET_RECOMMENDATION_HEADER] == headers["Avg target % (L)"] + 1
    eurusd = symbol_rows["EURUSD"]
    btcusdt = symbol_rows["BTCUSDT"]
    assert symbols.cell(eurusd, headers["Avg stop %"]).value == pytest.approx(abs(1.1 - 1.09) / 1.1)
    assert symbols.cell(eurusd, headers["Avg target %"]).value == pytest.approx(abs(1.12 - 1.1) / 1.1)
    assert symbols.cell(btcusdt, headers["Avg stop %"]).value == pytest.approx(abs(61000 - 60000) / 60000)
    assert symbols.cell(btcusdt, headers["Avg target %"]).value == pytest.approx(abs(59000 - 60000) / 60000)
    wb.close()


def test_stats1_streak_details_are_repaired_under_their_own_streak_rows(tmp_path: Path):
    snapshot = sample_snapshot()
    path = tmp_path / "streak-order.xlsx"
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or "").strip(): row for row in range(1, stats1.max_row + 1)}
    best = labels["Best Win Streak"]
    rows = [
        [stats1.cell(row, col).value for col in range(1, 5)]
        for row in range(best, best + 6)
    ]
    for offset, source in enumerate([0, 3, 1, 2, 4, 5]):
        for col, value in enumerate(rows[source], start=1):
            stats1.cell(best + offset, col).value = value
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or "").strip(): row for row in range(1, stats1.max_row + 1)}
    best = labels["Best Win Streak"]
    assert [stats1.cell(best + offset, 1).value for offset in range(6)] == [
        "Best Win Streak",
        "Start",
        "End",
        "Worst Losing Streak",
        "Start",
        "End",
    ]
    wb.close()


def test_stats_tables_get_clean_group_borders(tmp_path: Path):
    out = tmp_path / "clean-borders.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    stats1 = wb[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or "").strip(): row for row in range(1, stats1.max_row + 1)}
    assert stats1["A1"].border.left.style == "medium"
    assert stats1.cell(labels["R expectancy"], 4).border.right.style == "medium"
    assert stats1.cell(labels["Best Win Streak"], 1).border.top.style == "medium"
    assert not any(rng.min_col <= 4 and rng.min_row <= stats1.max_row for rng in stats1.merged_cells.ranges)

    stats2 = wb[STATS2_SHEET]
    assert stats2["A1"].border.left.style == "medium"
    assert stats2["F2"].border.right.style == "medium"
    assert _cell_fill_rgb(stats2["E3"]) == ""
    wb.close()


@pytest.mark.parametrize(
    ("seconds", "expected"),
    [
        (3661, "01 hours, 01 minutes, 01 seconds"),
        (180, "03 minutes, 00 seconds"),
        (59, "59 seconds"),
        (90061, "01 days, 01 hours, 01 minutes, 01 seconds"),
    ],
)
def test_format_duration_display_suppresses_leading_units(seconds, expected):
    assert _format_duration_display(seconds) == expected
    assert _parse_duration_text(expected) == pytest.approx(seconds)


def test_generated_workbook_duration_styles_are_excel_safe(tmp_path: Path):
    journal_dir = tmp_path / "journal"
    out = journal_dir / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)

    invalid_tokens = ("[>=1000000]", "[>=10000]", "[>=100]")
    custom_formats = _custom_number_formats(out)
    assert DURATION_NUMBER_FORMAT == r"00\:00\:00\:00"
    assert not any(all(token in fmt for token in invalid_tokens) for fmt in custom_formats)

    wb = load_workbook(out)
    try:
        assert wb.active.title == STATS1_SHEET
    finally:
        wb.close()


def test_legacy_duration_format_registry_entries_are_removed_from_styles(tmp_path: Path):
    invalid_format = (
        '[>=1000000]00 "days", 00 "hours", 00 "minutes", 00 "seconds";'
        '[>=10000]00 "hours", 00 "minutes", 00 "seconds";'
        '[>=100]00 "minutes", 00 "seconds";'
        '00 "seconds"'
    )
    path = tmp_path / "stale_duration_style.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = 10101
    ws["A1"].number_format = invalid_format
    wb.save(path)
    wb.close()

    wb = load_workbook(path)
    try:
        wb.active["A1"].number_format = "General"
        _repair_legacy_duration_number_formats(wb)
        wb.save(path)
    finally:
        wb.close()

    invalid_tokens = ("[>=1000000]", "[>=10000]", "[>=100]")
    custom_formats = _custom_number_formats(path)
    assert not any(all(token in fmt for token in invalid_tokens) for fmt in custom_formats)


def test_dashboard_parity_and_equity(tmp_path: Path):
    out=tmp_path/'Trading Journal.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    assert wb.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(wb)[len(SHEET_ORDER):] == expected_report_sheet_names(sample_snapshot())
    vals=[
        str(ws.cell(r,c).value or '')
        for ws in (wb[STATS1_SHEET], wb[STATS2_SHEET])
        for r in range(1,220)
        for c in range(1,15)
    ]
    assert 'Account Balances' in vals and 'Main Stats' not in vals and 'Label' not in vals
    for label in ['Overall','Winners','Losers','Drawdown','FX','Crypto','Win rate','R expectancy','Gross IR gain','Max R loss','Max R win','Min duration','Max Move to Profit','Side','Patterns','Timeframe','Commission']:
        assert label in vals
    assert 'Instrument leaders' not in vals
    assert any(isinstance(wb[STATS1_SHEET].cell(r,c).value, float) for r in range(1,220) for c in range(1,13))
    assert wb[STATS1_SHEET]["A8"].value == "Net P/L Percentage"
    assert wb[STATS1_SHEET]["A9"].value == "Net P/L R multiples"
    assert wb[STATS1_SHEET]["A14"].value == "Percentage expectancy"
    assert wb[STATS1_SHEET]["A15"].value == "R expectancy"
    assert wb[STATS1_SHEET]["B8"].value == pytest.approx(0.012)
    assert all(wb[STATS1_SHEET][coord].number_format == '0.00%' for coord in ('C8','D8','C10','D10'))
    assert all(wb[STATS1_SHEET][coord].number_format == '0.000"R"' for coord in ('C9','D9'))
    assert any(
        isinstance(wb[STATS1_SHEET].cell(r, c).value, str) and
        str(wb[STATS1_SHEET].cell(r, c).number_format or "") == "General" and
        _parse_duration_text(wb[STATS1_SHEET].cell(r, c).value) is not None
        for r in range(1,220) for c in range(1,13)
    )
    assert 'Equity Curve' not in wb.sheetnames
    ranges = _cf_ranges(wb[STATS1_SHEET])
    assert all(not r.startswith("B1:K") for r in ranges)
    assert not _cell_covered(ranges, "B3")  # Trades count should not be profit/loss colored


def test_dashboard_overall_return_average_labels_and_move_durations(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][0].update({
        "open_time": "2026-05-01 09:00:00",
        "move_to_break_even_time": "2026-05-01 10:00:00",
        "move_to_profit_time": "2026-05-01 11:00:00",
    })
    out = tmp_path / "dashboard_acceptance.xlsx"
    build_master_journal_workbook(snapshot, out)
    dash = load_workbook(out, data_only=True)[STATS1_SHEET]
    assert dash["B8"].value not in (None, "")
    labels = {str(dash.cell(row, 1).value or ""): row for row in range(1, dash.max_row + 1)}
    assert labels["Min Move to Break Even"] < labels[DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL] < labels["Max Move to Break Even"]
    assert labels["Min Move to Profit"] < labels[DASHBOARD_MOVE_TO_PROFIT_LABEL] < labels["Max Move to Profit"]
    assert dash.cell(labels[DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL], 2).value == "01 hours, 00 minutes, 00 seconds"
    assert dash.cell(labels[DASHBOARD_MOVE_TO_PROFIT_LABEL], 2).value == "02 hours, 00 minutes, 00 seconds"
    for label in ("Min Move to Break Even", "Max Move to Break Even", "Min Move to Profit", "Max Move to Profit"):
        assert dash.cell(labels[label] + 1, 1).value == "Source"
        assert "EURUSD · 2026-05-01" in str(dash.cell(labels[label] + 1, 2).value)


def test_mixed_currency_overall_return_uses_account_balances():
    totals = _result_percentage_totals_by_market(
        [
            {
                "row_type": "trade", "account": "OANDA LIVE", "asset_class": "fx",
                "symbol": "EURUSD", "net_profit": 100, "currency": "AUD", "result_pct": 99,
            },
            {
                "row_type": "trade", "account": "BYBIT LIVE", "asset_class": "crypto",
                "symbol": "BTCUSDT", "net_profit": 50, "currency": "USDT", "result_pct": 88,
            },
        ],
        [
            {"account_label": "OANDA LIVE", "balance": 1100, "currency": "AUD"},
            {"account_label": "BYBIT LIVE", "balance": 550, "currency": "USDT"},
        ],
    )
    assert totals["fx"]["market_return_pct"] == pytest.approx(10.0)
    assert totals["crypto"]["market_return_pct"] == pytest.approx(10.0)
    assert totals["overall"]["market_return_pct"] == pytest.approx(10.0)
    assert totals["fx"]["net_result_pct"] == pytest.approx(99.0)
    assert totals["crypto"]["net_result_pct"] == pytest.approx(88.0)
    assert totals["overall"]["net_result_pct"] == pytest.approx(187.0)


def test_net_pl_percentages_are_linear_trade_percent_sums():
    rows = [
        {"row_type": "trade", "account": "A", "asset_class": "fx", "net_profit": -30, "currency": "AUD", "result_pct": -3},
        {"row_type": "trade", "account": "B", "asset_class": "fx", "net_profit": -40, "currency": "AUD", "result_pct": -4},
        {"row_type": "trade", "account": "C", "asset_class": "fx", "net_profit": -50, "currency": "AUD", "result_pct": -5},
    ]
    balances = [
        {"account_label": "A", "balance": 970, "currency": "AUD"},
        {"account_label": "B", "balance": 960, "currency": "AUD"},
        {"account_label": "C", "balance": 950, "currency": "AUD"},
    ]
    totals = _result_percentage_totals_by_market(rows, balances)
    assert totals["overall"]["market_return_pct"] == pytest.approx(-4.0)
    assert totals["overall"]["net_result_pct"] == pytest.approx(-12.0)


def test_market_return_percentages_weight_unequal_starting_equity():
    rows = [
        {"row_type": "trade", "account": "SMALL", "asset_class": "fx", "net_profit": -10, "currency": "AUD", "result_pct": -10},
        {"row_type": "trade", "account": "LARGE", "asset_class": "fx", "net_profit": -100, "currency": "AUD", "result_pct": -100},
    ]
    balances = [
        {"account_label": "SMALL", "balance": 90, "currency": "AUD"},
        {"account_label": "LARGE", "balance": 9900, "currency": "AUD"},
    ]
    totals = _result_percentage_totals_by_market(rows, balances)
    assert totals["overall"]["market_return_pct"] == pytest.approx(-110 / 10100 * 100)
    assert totals["overall"]["net_result_pct"] == pytest.approx(-110)


def test_net_pl_percentages_convert_usdt_to_aud_with_fallback(monkeypatch):
    monkeypatch.setenv("JOURNAL_USDT_TO_AUD_FALLBACK", "1.45")
    rows = [
        {"row_type": "trade", "account": "AUD", "asset_class": "fx", "net_profit": 100, "currency": "AUD", "result_pct": 9},
        {"row_type": "trade", "account": "USDT", "asset_class": "crypto", "net_profit": 100, "currency": "USDT", "result_pct": 8},
    ]
    balances = [
        {"account_label": "AUD", "balance": 1100, "currency": "AUD"},
        {"account_label": "USDT", "balance": 1100, "currency": "USDT"},
    ]
    totals = _result_percentage_totals_by_market(rows, balances)
    assert totals["overall"]["net_result_pct"] == pytest.approx(17)
    assert totals["overall"]["market_return_pct"] == pytest.approx(((100 + 100 * 1.45) / (1000 + 1000 * 1.45)) * 100)
    assert totals["overall"]["return_unavailable_reason"] is None
    assert totals["overall"]["return_method"] == "capital_weighted_account_return_aud"


def test_stats1_net_pl_percentage_uses_trade_profit_percent_sum(tmp_path: Path):
    snapshot = {
        "items": [
            {
                "id": "fx-loss",
                "row_type": "trade",
                "account": "OANDA DEMO",
                "asset_class": "fx",
                "symbol": "EURUSD",
                "open_time": "2026-01-01",
                "close_time": "2026-01-01",
                "net_profit": -100.0,
                "result_pct": -80.0,
                "r_multiple": -0.8,
                "analysis_balance_after_trade": 900.0,
                "currency": "AUD",
            }
        ],
        "balances": [{"account_label": "OANDA DEMO", "balance": 900.0, "currency": "AUD"}],
        "stats": {"totals": {}, "groups": {"by_market": {"overall": {}, "fx": {}, "crypto": {}}, "risk_expectancy": {}, "duration": {}, "leaders": {}}, "by_instrument": []},
    }
    out = tmp_path / "linear-net-pl.xlsx"
    build_master_journal_workbook(snapshot, out)
    stats1 = load_workbook(out, data_only=True)[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or ""): row for row in range(1, stats1.max_row + 1)}
    row = labels["Net P/L Percentage"]
    assert stats1.cell(row, 2).value == pytest.approx(-0.8)
    assert stats1.cell(row, 3).value == pytest.approx(-0.8)
    assert "Unavailable" not in str(stats1.cell(row, 2).value)


def test_streak_rows_split_count_start_end_for_stats1_and_reports(tmp_path: Path):
    snapshot = sample_snapshot()
    detail = {"start_time": "2026-05-01T00:00:00Z", "end_time": "2026-05-03T00:00:00Z"}
    snapshot["stats"]["groups"]["by_market"]["overall"]["winning_streak"] = 2
    snapshot["stats"]["groups"]["by_market"]["overall"]["longest_winning_streak"] = detail
    snapshot["stats"]["groups"]["by_market"]["fx"]["winning_streak"] = 2
    snapshot["stats"]["groups"]["by_market"]["fx"]["longest_winning_streak"] = detail
    out = tmp_path / "streak-split.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out, data_only=True)
    try:
        stats1 = wb[STATS1_SHEET]
        labels = defaultdict(list)
        for row in range(1, stats1.max_row + 1):
            labels[str(stats1.cell(row, 1).value or "")].append(row)
        row = labels["Best Win Streak"][0]
        assert stats1.cell(row, 2).value == 2
        assert isinstance(stats1.cell(row, 2).value, int)
        assert stats1.cell(row + 1, 1).value == "Start"
        assert stats1.cell(row + 2, 1).value == "End"
        assert stats1.cell(row + 1, 2).value == "2026-05-01 00:00:00"
        assert stats1.cell(row + 2, 2).value == "2026-05-03 00:00:00"
        assert _cell_font_rgb(stats1.cell(row + 1, 2)) != "9C0006"
        assert _cell_font_rgb(stats1.cell(row + 2, 2)) != "9C0006"

        yearly = wb[REPORT_YEARLY_SHEET]
        report_rows = {str(yearly.cell(r, 1).value or ""): r for r in range(1, yearly.max_row + 1)}
        year_col = next(col for col in range(2, yearly.max_column + 1) if yearly.cell(1, col).value == 2026)
        assert yearly.cell(report_rows["Best Win Streak"], year_col).value == 1
        assert yearly.cell(report_rows["Best Win Streak"] + 1, 1).value == "Start"
        assert yearly.cell(report_rows["Worst Losing Streak"] + 1, 1).value == "Start"
        assert _cell_font_rgb(yearly.cell(report_rows["Best Win Streak"] + 1, year_col)) != "9C0006"
    finally:
        wb.close()


def test_stats2_as_of_column_uses_plain_black_and_white_style(tmp_path: Path):
    out = tmp_path / "stats2-asof-style.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    try:
        stats2 = wb[STATS2_SHEET]
        for row in range(3, 10):
            cell = stats2.cell(row, 5)
            if cell.value in (None, ""):
                continue
            assert "T" not in str(cell.value)
            assert cell.number_format in {"General", "yyyy-mm-dd hh:mm:ss"}
            assert _cell_fill_rgb(cell) == ""
            assert _cell_font_rgb(cell) == "000000"
    finally:
        wb.close()


def test_symbols_latest_blank_metric_columns_populate_and_header_merges_survive():
    wb = Workbook()
    ws = wb.active
    ws.title = SYMBOLS_SHEET
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = "A2:BE3"
    ws["B1"] = "Class"
    ws.merge_cells("B1:B2")
    ws["A2"] = "Symbol"
    ws["C2"] = "Trades"
    for col, label in {
        29: "1M", 30: "5M", 31: "15M", 32: "30M", 33: "1H", 34: "4H",
        35: "DAILY", 36: "WEEKLY", 37: "MONTHLY",
    }.items():
        ws.cell(2, col).value = label
    ws["AC1"] = "Timeframe"
    ws.merge_cells("AC1:AK1")
    ws["AS1"] = "Stops"
    ws["AS2"] = "Min stop %"
    ws["AT2"] = "Avg stop %"
    ws["AU2"] = "Max stop %"
    ws["AX1"] = "Targets"
    ws["AX2"] = "Min  %"
    ws["AY2"] = "Avg  %"
    ws["AZ2"] = "Max  %"
    ws["A3"] = "EURUSD"
    rows = [
        {
            "row_type": "trade", "symbol": "EURUSD", "timeframe": "1M",
            "entry_price": 1.0, "stop_loss": 0.99, "take_profit": 1.02,
        },
        {
            "row_type": "trade", "symbol": "EURUSD", "timeframe": "5M",
            "entry_price": 1.0, "stop_loss": 0.98, "take_profit": 1.03,
        },
    ]
    diagnostics = {}
    _populate_symbols_metrics_preserving_layout(ws, rows, diagnostics)
    _repair_symbols_header_merges_preserving_layout(ws, diagnostics)

    assert ws["AC3"].value == 1
    assert ws["AD3"].value == 1
    assert ws["AS3"].value == pytest.approx(0.01)
    assert ws["AT3"].value == pytest.approx(0.015)
    assert ws["AU3"].value == pytest.approx(0.02)
    assert ws["AX3"].value == pytest.approx(0.02)
    assert ws["AY3"].value == pytest.approx(0.025)
    assert ws["AZ3"].value == pytest.approx(0.03)
    merges = {str(rng) for rng in ws.merged_cells.ranges}
    assert "B1:B2" in merges
    assert "C1:C2" in merges
    assert "AC1:AK1" in merges
    assert "AC1:AC2" not in merges
    assert ws.freeze_panes == "B3"
    assert ws.auto_filter.ref == "A2:BE3"


def test_move_to_duration_repair_uses_open_and_move_times():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        ws.cell(1, col, header)
    _ensure_trade_log_schema(ws)
    headers = _trade_log_header_map(ws)
    row = TRADE_LOG_DATA_START_ROW
    ws.cell(row, headers["Open Time"]).value = datetime(2026, 1, 1, 9, 0, 0)
    ws.cell(row, headers["Move to Break Even Time"]).value = datetime(2026, 1, 1, 10, 1, 2)
    ws.cell(row, headers["Move to Profit Time"]).value = datetime(2026, 1, 1, 8, 0, 0)
    diagnostics = {}
    assert _repair_trade_log_move_to_durations(ws, diagnostics) == 2
    assert ws.cell(row, headers["Move to Break Even Duration"]).value == 10102
    assert ws.cell(row, headers["Move to Profit Duration"]).value == 0


def test_instrument_averages_new_columns_order_formats_and_alignment(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][0].update({
        "pattern": "range", "ema": "Yes", "aths_atls": "All-Time High",
        "order_type": "Limit", "round_number": "Yes", "spiked_out": "No",
        "close_stopout": "No", "near_perfect_entry": "Yes", "near_win": "No",
        "early_close": "No", "timeframe": "1H", "move_to_break_even_duration": 3600,
        "move_to_profit_duration": 7200,
    })
    snapshot["items"].append({
        "id": "t3", "row_type": "trade", "symbol": "EURUSD", "asset_class": "fx",
        "side": "SELL", "account": "OANDA DEMO", "open_time": "2026-05-03T00:00:00Z",
        "close_time": "2026-05-03T01:00:00Z", "r_multiple": -0.3,
        "order_type": "Market", "round_number": True, "spiked_out": 1,
        "close_stopout": "Yes", "near_perfect_entry": "No", "near_win": "true",
        "early_close": "Y",
    })
    out = tmp_path / "instrument_acceptance.xlsx"
    build_master_journal_workbook(snapshot, out)
    ws = load_workbook(out)[SYMBOLS_SHEET]
    headers = [str(cell.value or "") for cell in ws[INSTRUMENT_AVERAGES_FILTER_HEADER_ROW]]
    expected_headers = [
        "Recommendation" if header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER} else header
        for header in INSTRUMENT_AVERAGES_HEADERS
    ]
    assert headers == expected_headers
    by_header = {header: index + 1 for index, header in enumerate(headers)}
    row = INSTRUMENT_AVERAGES_DATA_START_ROW
    assert ws.cell(row, by_header["Move to break even"]).value == 1
    assert ws.cell(row, by_header["Move to profit"]).value == 1
    assert ws.cell(row, by_header["Net R Multiple"]).value == pytest.approx(0.9)
    assert ws.cell(row, by_header["Net R Multiple"]).number_format == '0.000"R"'
    assert ws.cell(row, by_header["All-time highs"]).value == 1
    assert ws.cell(row, by_header["All-time lows"]).value == 0
    assert ws.cell(row, by_header["Market"]).value == 1
    assert ws.cell(row, by_header["Limit"]).value == 1
    assert ws.cell(row, by_header["Round number"]).value == 2
    assert ws.cell(row, by_header["Spiked out"]).value == 1
    assert ws.cell(row, by_header["Close stop out"]).value == 1
    assert ws.cell(row, by_header["Near perfect entry"]).value == 1
    assert ws.cell(row, by_header["Near win"]).value == 1
    assert ws.cell(row, by_header["Early close"]).value == 1
    assert ws.cell(row, by_header["Move to break even"]).font.color.rgb == "FF000000"
    assert ws.cell(row, by_header["Move to profit"]).font.color.rgb == "FF000000"
    assert all(
        ws.cell(row, col).alignment.horizontal == "left"
        for col in range(2, ws.max_column + 1)
        if ws.cell(row, col).value is not None
    )
    order_range = (
        f"{get_column_letter(by_header['Market'])}1:"
        f"{get_column_letter(by_header['Limit'])}1"
    )
    assert order_range in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws.cell(1, by_header["Market"]).value == "Order"
    assert ws.freeze_panes == "B3"
    assert ws.auto_filter.ref.endswith(f"{get_column_letter(len(INSTRUMENT_AVERAGES_HEADERS))}{ws.max_row}")


def test_manual_override_roundtrip(tmp_path: Path):
    out=tmp_path/'Trading Journal.xlsx'; build_master_journal_workbook(sample_snapshot(), out)
    wb=load_workbook(out); ws=wb['Trade Log']; ws.cell(_trade_data_row(), _header_col(ws, 'Test')).value='Yes'; ws.cell(_trade_data_row(), _header_col(ws, 'Setup')).value='AAA'; wb.save(out)
    ov=read_master_journal_manual_overrides(out)
    assert ov['t1']['is_test_trade'] is True and ov['t1']['setup']=='AAA'


def test_trade_number_schema_reports_and_source_roundtrip(tmp_path: Path):
    snap = sample_snapshot()
    snap["items"][0]["trade_number"] = "C42"
    out = tmp_path / "trade_number.xlsx"
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    assert TRADE_LOG_HEADERS[0] == TRADE_NUMBER_HEADER
    assert ws["A1"].value == TRADE_NUMBER_HEADER
    assert "A1:A3" in {str(rng) for rng in ws.merged_cells.ranges}
    assert _header_col(ws, "Open Time") == 2
    assert ws.cell(_trade_data_row(), _header_col(ws, TRADE_NUMBER_HEADER)).value == "C42"
    assert ws.cell(_trade_data_row(), _header_col(ws, TRADE_NUMBER_HEADER)).number_format == "@"
    row_id_col = _header_col(ws, "Row ID")
    assert row_id_col == len(TRADE_LOG_HEADERS)
    assert ws.column_dimensions[get_column_letter(row_id_col)].hidden is True
    assert wb.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(wb)[len(SHEET_ORDER):] == expected_report_sheet_names(snap)
    yearly = wb[REPORT_YEARLY_SHEET]
    assert yearly["B1"].value == 2018
    assert yearly.cell(1, yearly.max_column).value >= 2026
    assert [yearly.cell(row, 1).value for row in range(2, 2 + len(REPORT_METRIC_LABELS))] == REPORT_METRIC_LABELS
    assert yearly["B7"].number_format == "0.00%"
    avg_r_row = REPORT_METRIC_LABELS.index("R expectancy") + 2
    assert yearly.cell(avg_r_row, 2).number_format == '0.000"R"'
    duration_row = REPORT_METRIC_LABELS.index("Shortest (DD:HH:MM:SS)") + 2
    assert yearly.cell(duration_row, 2).number_format == "General"
    year_sheet = wb["2026"]
    assert year_sheet["B1"].value == "January"
    assert year_sheet["M1"].value == "December"
    assert wb["2018"]["B1"].value == "May"
    assert wb["2018"]["I1"].value == "December"
    wb.close()
    parsed = read_master_journal_source(out)
    assert parsed["items"][0]["trade_number"] == "C42"


def test_old_current_two_row_trade_log_migrates_with_trade_number_leftmost():
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for col, header in enumerate(TRADE_LOG_HEADERS_V1, start=1):
        if header in MOVE_TO_FIELD_MAP:
            continue
        visible_header = "Recommendation" if header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER} else header
        ws.cell(1, col, visible_header)
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
    for label, headers in (("Move to Break-Even", list(MOVE_TO_FIELD_MAP)[:5]), ("Move to Profit", list(MOVE_TO_FIELD_MAP)[5:])):
        cols = [TRADE_LOG_HEADERS_V1.index(header) + 1 for header in headers]
        ws.merge_cells(start_row=1, start_column=min(cols), end_row=1, end_column=max(cols))
        ws.cell(1, min(cols), label)
        for idx, col in enumerate(cols):
            ws.cell(2, col, ["Time", "Duration", "Trigger Price", "Distance From Entry %", "Distance From Exit %"][idx])
    ws.cell(3, TRADE_LOG_HEADERS_V1.index("Open Time") + 1, "2026-01-01")
    ws.cell(3, TRADE_LOG_HEADERS_V1.index("Row ID") + 1, "old-row")
    _ensure_trade_log_schema(ws)
    assert ws["A1"].value == TRADE_NUMBER_HEADER
    assert ws.cell(TRADE_LOG_DATA_START_ROW, _header_col(ws, TRADE_NUMBER_HEADER)).value in (None, "")
    assert ws.cell(TRADE_LOG_DATA_START_ROW, _header_col(ws, "Open Time")).value == "2026-01-01"
    assert ws.cell(TRADE_LOG_DATA_START_ROW, _header_col(ws, "Row ID")).value == "old-row"
    assert _header_col(ws, "Row ID") == len(TRADE_LOG_HEADERS)
    assert ws.auto_filter.ref == f"A3:{get_column_letter(len(TRADE_LOG_HEADERS))}4"


def test_current_pnl_calendar_layout_freezes_months_and_left_axis():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Calendar"
    for col, month in enumerate(["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"], start=3):
        ws.cell(1, col, month)
    ws.merge_cells("A2:A3")
    ws["A2"] = 2026
    ws["B2"] = "P/L %"
    ws["B3"] = "Total Trades"
    _ensure_pnl_calendar_freeze_panes(ws)
    assert ws.freeze_panes == "C2"


def test_transposed_pnl_calendar_layout_freezes_months_and_years():
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Calendar"
    ws["A1"] = "Month"
    ws["B1"] = 2026
    ws["A2"] = "January"
    ws["B2"] = "1.00%, 1 trade"

    _ensure_pnl_calendar_freeze_panes(ws)

    assert ws.freeze_panes == "B2"


def test_pre_move_schema_migrates_stop_out_value_and_hides_row_id():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for col, header in enumerate(PRE_MOVE_TRADE_LOG_HEADERS, start=1):
        ws.cell(1, col, header)
    ws.cell(2, PRE_MOVE_TRADE_LOG_HEADERS.index("Stop Out") + 1, "Yes")
    ws.cell(2, PRE_MOVE_TRADE_LOG_HEADERS.index("Row ID") + 1, "row-1")
    _ensure_trade_log_schema(ws)
    assert ws.cell(_trade_data_row(), _header_col(ws, "Close Stopout")).value == "Yes"
    assert ws.cell(_trade_data_row(), _header_col(ws, "Row ID")).value == "row-1"
    last_letter = get_column_letter(len(TRADE_LOG_HEADERS))
    assert get_column_letter(_header_col(ws, "Row ID")) == last_letter
    assert ws.column_dimensions[last_letter].hidden is True
    assert ws.auto_filter.ref == f"A3:{last_letter}4"


def test_trade_log_three_row_umbrella_headers_and_filter(tmp_path: Path):
    out = tmp_path / "three_row_headers.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    header_map = _trade_log_header_map(ws)
    be_cols = [header_map[h] for h in list(MOVE_TO_FIELD_MAP.keys())[:5]]
    profit_cols = [header_map[h] for h in list(MOVE_TO_FIELD_MAP.keys())[5:]]
    be_range = f"{get_column_letter(min(be_cols))}1:{get_column_letter(max(be_cols))}1"
    profit_range = f"{get_column_letter(min(profit_cols))}1:{get_column_letter(max(profit_cols))}1"
    assert be_range in {str(rng) for rng in ws.merged_cells.ranges}
    assert profit_range in {str(rng) for rng in ws.merged_cells.ranges}
    assert ws.cell(1, min(be_cols)).value == "Move to Break-Even"
    assert ws.cell(1, min(profit_cols)).value == "Move to Profit"
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        if header not in MOVE_TO_FIELD_MAP:
            letter = get_column_letter(col)
            assert f"{letter}1:{letter}3" in merged
            expected_header = "Recommendation" if header in {"Stop Loss Recommendation", "Target Recommendation"} else header
            assert ws.cell(1, col).value == expected_header
            assert ws.cell(2, col).value is None
            assert ws.cell(3, col).value is None
            assert ws.cell(1, col).alignment.horizontal == "left"
            assert ws.cell(1, col).alignment.vertical == "center"
    expected = ["Time", "Duration", "Trigger Price", "Distance From Entry %", "Distance From Exit %"]
    assert [ws.cell(2, col).value for col in be_cols] == expected
    assert [ws.cell(2, col).value for col in profit_cols] == expected
    for col in be_cols + profit_cols:
        letter = get_column_letter(col)
        assert f"{letter}2:{letter}3" in merged
    assert ws.freeze_panes == "A4"
    last_letter = get_column_letter(len(TRADE_LOG_HEADERS))
    assert ws.auto_filter.ref == f"A3:{last_letter}{ws.max_row}"
    assert ws.cell(4, _header_col(ws, "Row ID")).value == "t1"
    assert ws.column_dimensions[last_letter].hidden is True
    for row in range(1, TRADE_LOG_HEADER_ROWS + 1):
        for col in range(1, len(TRADE_LOG_HEADERS) + 1):
            if ws.cell(row, col).__class__.__name__ == "MergedCell":
                continue
            assert _cell_fill_rgb(ws.cell(row, col)) == ""
    pattern_validations = _validation_for_col(ws, get_column_letter(_header_col(ws, "Pattern")))
    assert any(dv.formula1 == '"range,channel"' and dv.allow_blank for dv in pattern_validations)
    wb.close()



def test_duplicate_two_row_headers_migrate_to_vertical_merges_without_losing_rows(tmp_path: Path):
    out = tmp_path / "duplicate_headers.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    before_ids = [ws.cell(row, _header_col(ws, "Row ID")).value for row in range(4, ws.max_row + 1)]
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 2 and merged.max_row <= 3:
            ws.unmerge_cells(str(merged))
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        if header not in MOVE_TO_FIELD_MAP:
            ws.cell(2, col).value = "Recommendation" if header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER} else header
    ws.delete_rows(3, 1)
    _ensure_trade_log_schema(ws)
    after_ids = [ws.cell(row, _header_col(ws, "Row ID")).value for row in range(4, ws.max_row + 1)]
    assert after_ids == before_ids
    assert _trade_log_header_map(ws) == {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    assert all(ws.cell(2, col).value is None for col, header in enumerate(TRADE_LOG_HEADERS, start=1) if header not in MOVE_TO_FIELD_MAP)
    assert ws.freeze_panes == "A4"
    last_letter = get_column_letter(len(TRADE_LOG_HEADERS))
    assert ws.auto_filter.ref == f"A3:{last_letter}{ws.max_row}"
    assert ws.column_dimensions[last_letter].hidden is True
    wb.close()


def test_dashboard_build_includes_canonical_move_duration_rows(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][0].update({
        "move_to_break_even_duration": 3600,
        "move_to_profit_time": "2026-01-01 02:00:00",
        "open_time": "2026-01-01 00:00:00",
    })
    out = tmp_path / "canonical_dashboard_rows.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    dash = wb[STATS1_SHEET]
    labels = {str(dash.cell(row, 1).value or ""): row for row in range(1, dash.max_row + 1)}
    assert DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL in labels
    assert DASHBOARD_MOVE_TO_PROFIT_LABEL in labels
    break_even_row = labels[DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL]
    profit_row = labels[DASHBOARD_MOVE_TO_PROFIT_LABEL]
    assert labels["Min Move to Break Even"] + 2 == break_even_row
    assert labels["Max Move to Break Even"] == break_even_row + 1
    assert dash.cell(labels["Min Move to Break Even"] + 1, 1).value == "Source"
    assert dash.cell(labels["Max Move to Break Even"] + 1, 1).value == "Source"
    assert labels["Min Move to Profit"] + 2 == profit_row
    assert labels["Max Move to Profit"] == profit_row + 1
    assert dash.cell(labels["Min Move to Profit"] + 1, 1).value == "Source"
    assert dash.cell(labels["Max Move to Profit"] + 1, 1).value == "Source"
    for value_col in (2, 3, 4):
        assert dash.cell(break_even_row, value_col).number_format == "General"
        assert dash.cell(profit_row, value_col).number_format == "General"
    wb.close()


def test_data_only_update_inserts_missing_dashboard_move_rows_and_preserves_metrics(tmp_path: Path):
    from openpyxl import Workbook

    path = tmp_path / "missing_move_rows.xlsx"
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    wb.create_sheet("Trade Log")
    wb.create_sheet("Instrument Averages")
    wb.create_sheet("P&L Calendar")
    dash["B1"] = "Overall"
    dash["C1"] = "FX"
    dash["D1"] = "Crypto"
    dash["F1"] = "Account Balances"
    dash["F2"] = "Account"
    dash["G2"] = "Balance"
    dash["H2"] = "Currency"
    dash["F11"] = "Instrument leaders"
    dash["F12"] = "Metric"
    dash["G12"] = "Symbol"
    dash["H12"] = "Wins"
    dash["I12"] = "Losses"
    dash["J12"] = "Trades"
    dash["A2"] = "Trades"
    dash["A18"] = "Avg duration (DD:HH:MM:SS)"
    dash["A19"] = "Max loss %"
    dash["A35"] = "Winners"
    dash["A40"] = "Losers"
    dash["A45"] = "Drawdown"
    for cell in (dash["B18"], dash["C18"], dash["D18"]):
        cell.number_format = r'00\:00\:00\:00'
    dash.row_dimensions[18].height = 27
    _ensure_trade_log_headers(wb)
    trade = wb["Trade Log"]
    trade.cell(_trade_data_row(), _header_col(trade, "Row ID")).value = "move-1"
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Break Even Duration")).value = 10000
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Break Even Duration")).number_format = r'00\:00\:00\:00'
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Profit Time")).value = "2026-05-01 11:00:00"
    wb.save(path)
    wb.close()

    snapshot = {
        "items": [{
            "id": "move-1", "row_type": "trade", "account": "BYBIT", "asset_class": "crypto",
            "symbol": "BTCUSDT", "open_time": "2026-05-01 09:00:00", "close_time": "2026-05-01 12:00:00",
        }],
        "stats": {
            "totals": {"trades": 1},
            "groups": {
                "by_market": {"overall": {"trades": 1}, "fx": {"trades": 0}, "crypto": {"trades": 1}},
                "risk_expectancy": {}, "leaders": {}, "duration": {},
            },
        },
        "balances": [],
    }
    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    assert result["diagnostics"]["inserted_dashboard_metric_rows"] == [
        DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        DASHBOARD_MOVE_TO_PROFIT_LABEL,
        "Avg winning trade duration",
        "Avg losing trade duration",
    ]
    Path(result["candidate_path"]).replace(path)
    wb = load_workbook(path)
    dash = wb[STATS1_SHEET]
    labels = {str(dash.cell(row, 1).value or ""): row for row in range(1, dash.max_row + 1)}
    break_even_row = labels[DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL]
    profit_row = labels[DASHBOARD_MOVE_TO_PROFIT_LABEL]
    assert labels["Avg winning trade duration"] == 19
    assert labels["Avg losing trade duration"] == 20
    assert break_even_row == 21
    assert profit_row == 22
    assert labels["Max loss %"] == 23
    assert labels["Winners"] == 39
    assert dash.cell(break_even_row, 2).value == "01 hours, 00 minutes, 00 seconds"
    assert dash.cell(break_even_row, 4).value == "01 hours, 00 minutes, 00 seconds"
    assert dash.cell(profit_row, 2).value == "02 hours, 00 minutes, 00 seconds"
    assert dash.cell(profit_row, 4).value == "02 hours, 00 minutes, 00 seconds"
    assert dash.cell(2, 2).value == 1
    assert dash.row_dimensions[break_even_row].height == 27
    assert dash.cell(break_even_row, 2).number_format == "General"
    trade = wb["Trade Log"]
    merged = {str(rng) for rng in trade.merged_cells.ranges}
    be_range, profit_range = _move_group_ranges(trade)
    assert "A1:A3" in merged and be_range in merged and profit_range in merged
    assert trade.freeze_panes == "A4"
    wb.close()


def test_dashboard_manual_move_rows_survive_and_populate_by_label(tmp_path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side

    path = tmp_path / "manual_move_metrics.xlsx"
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    wb.create_sheet("Trade Log")
    wb.create_sheet("Instrument Averages")
    wb.create_sheet("P&L Calendar")
    dash["A1"] = "Metric"
    dash["B1"] = "Overall"
    dash["C1"] = "FX"
    dash["D1"] = "Crypto"
    dash["G1"] = "Winners"
    dash["G8"] = "Losers"
    dash["G14"] = "Drawdown"
    dash["M1"] = "Instrument leaders"
    dash["T1"] = "Account Balances"
    dash["T2"] = "Account"
    dash["U2"] = "Balance"
    dash["V2"] = "Currency"
    dash["W2"] = "As Of"
    dash["A2"] = "Move to Break Even (DD:HH:MM:SS)"
    dash["A3"] = "Move to Profit (DD:HH:MM:SS)"
    dash["A4"] = "Trades"
    dash["A2"].fill = PatternFill("solid", fgColor="ABCDEF")
    label_font = copy(dash["A2"].font)
    label_font.bold = True
    dash["A2"].font = label_font
    dash["A2"].border = Border(bottom=Side(style="thick", color="123456"))
    dash.row_dimensions[2].height = 31
    _ensure_trade_log_headers(wb)
    trade = wb["Trade Log"]
    trade.cell(_trade_data_row(), _header_col(trade, "Row ID")).value = "move-1"
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Break Even Duration")).value = 10000
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Break Even Duration")).number_format = r'00\:00\:00\:00'
    trade.cell(_trade_data_row(), _header_col(trade, "Move to Profit Time")).value = "2026-05-01 11:00:00"
    wb.save(path)
    wb.close()

    snapshot = {
        "items": [{
            "id": "move-1", "row_type": "trade", "account": "BYBIT", "asset_class": "crypto",
            "symbol": "BTCUSDT", "open_time": "2026-05-01 09:00:00", "close_time": "2026-05-01 12:00:00",
        }],
        "stats": {
            "totals": {"trades": 1},
            "groups": {
                "by_market": {"overall": {"trades": 1}, "fx": {"trades": 0}, "crypto": {"trades": 1}},
                "risk_expectancy": {}, "leaders": {}, "duration": {},
            },
        },
        "balances": [],
    }
    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    wb = load_workbook(path)
    dash = wb[STATS1_SHEET]
    assert dash["A2"].value == DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL
    assert dash["A3"].value == DASHBOARD_MOVE_TO_PROFIT_LABEL
    assert dash["B2"].value == "01 hours, 00 minutes, 00 seconds"
    assert dash["D2"].value == "01 hours, 00 minutes, 00 seconds"
    assert dash["B3"].value == "02 hours, 00 minutes, 00 seconds"
    assert dash["D3"].value == "02 hours, 00 minutes, 00 seconds"
    assert dash["B4"].value == 1 and dash["D4"].value == 1
    assert str(dash["A2"].fill.fgColor.rgb).endswith("ABCDEF")
    assert dash["A2"].font.bold is True
    assert dash["A2"].border.bottom.style == "thick"
    assert dash["A2"].alignment.horizontal == "left"
    assert dash.row_dimensions[2].height == 31
    wb.close()

def test_schema_migration_preserves_trade_and_derived_sheet_content(tmp_path: Path):
    out = tmp_path / "preserve_content.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    # Convert the generated sheet back to the pre-move flat schema to exercise migration.
    data = [[ws.cell(row, _header_col(ws, header)).value for header in PRE_MOVE_TRADE_LOG_HEADERS if header != "Close" and header != "Stop Out"] for row in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1)]
    wb.close()

    from openpyxl import Workbook
    legacy = Workbook()
    legacy.remove(legacy.active)
    for name in ["Dashboard", "Trade Log", "Instrument Averages", "P&L Calendar"]:
        legacy.create_sheet(name)
    trade = legacy["Trade Log"]
    for col, header in enumerate(PRE_MOVE_TRADE_LOG_HEADERS, start=1):
        trade.cell(1, col, header)
    for row_idx, source in enumerate(data, start=2):
        source_map = {header: value for header, value in zip([h for h in PRE_MOVE_TRADE_LOG_HEADERS if h not in {"Close", "Stop Out"}], source)}
        for col, header in enumerate(PRE_MOVE_TRADE_LOG_HEADERS, start=1):
            trade.cell(row_idx, col, source_map.get(header, "Yes" if header == "Stop Out" else None))
    legacy["Instrument Averages"].append(["Symbol", "Trades"])
    legacy["Instrument Averages"].append(["EURUSD", 2])
    legacy["P&L Calendar"]["C1"] = "January"
    legacy["P&L Calendar"]["A2"] = 2026
    legacy["P&L Calendar"]["B2"] = "P/L %"
    legacy["P&L Calendar"]["C2"] = 0.01
    before_trade_rows = trade.max_row - 1
    before_instrument = legacy["Instrument Averages"]["A2"].value
    before_calendar = legacy["P&L Calendar"]["C2"].value
    _ensure_trade_log_schema(trade)
    assert trade.max_row - TRADE_LOG_HEADER_ROWS == before_trade_rows
    assert legacy["Instrument Averages"]["A2"].value == before_instrument
    assert legacy["P&L Calendar"]["C2"].value == before_calendar
    assert all(trade.cell(row, _header_col(trade, "Row ID")).value for row in range(TRADE_LOG_DATA_START_ROW, trade.max_row + 1))
    legacy.close()


def test_data_only_update_aborts_before_wiping_populated_sheets(tmp_path: Path):
    path = tmp_path / "no_wipe.xlsx"
    build_master_journal_workbook(sample_snapshot(), path)
    empty_snapshot = {
        "items": [],
        "stats": {"totals": {}, "groups": {"by_market": {"overall": {}, "fx": {}, "crypto": {}}, "risk_expectancy": {}, "leaders": {}, "duration": {}}},
        "balances": [],
    }
    with pytest.raises(RuntimeError, match="would be wiped"):
        update_master_journal_workbook_data_only(path, empty_snapshot)

def test_calendar_month_conditional_formatting_rows(tmp_path: Path):
    snap=sample_snapshot()
    snap['items']=[
        {'id':'p','row_type':'trade','account':'A','open_time':'2026-05-01','close_time':'2026-05-01','net_profit':10,'result_pct':1.2,'is_test_trade':False},
        {'id':'n','row_type':'trade','account':'A','open_time':'2026-06-01','close_time':'2026-06-01','net_profit':-5,'result_pct':-0.4,'is_test_trade':False},
    ]
    out=tmp_path/'Trading Journal.xlsx'; build_master_journal_workbook(snap,out); wb=load_workbook(out)
    cal=wb['P&L Calendar']
    may=cal['F2']; jun=cal['G2']; mar=cal['D2']
    assert may.value == "1.20%, 1 trade"
    assert jun.value == "-0.40%, 1 trade"
    assert _cell_fill_rgb(may) == PROFIT_FILL
    assert _cell_fill_rgb(jun) == LOSS_FILL
    assert _cf_ranges(cal) == []
    assert cal.freeze_panes == "B2"
    for c in range(1, 14):
        assert cal.cell(1, c).font.bold is True
    assert mar.value in ('', None)
    assert _cell_fill_rgb(mar) == ""
    assert not any(str(cal.cell(row, 1).value or "").endswith(" Trades") for row in range(2, cal.max_row + 1))


def test_calendar_data_only_update_clears_blank_month_fill(tmp_path: Path):
    snap = sample_snapshot()
    snap["items"] = [
        {"id": "may", "row_type": "trade", "account": "A", "open_time": "2026-05-01", "close_time": "2026-05-01", "net_profit": 10, "result_pct": 1.2, "is_test_trade": False},
    ]
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    cal = wb["P&L Calendar"]
    cal["D2"].fill = PatternFill("solid", fgColor="E5E7EB")
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    Path(result["candidate_path"]).replace(out)

    checked = load_workbook(out)
    try:
        mar = checked["P&L Calendar"]["D2"]
        assert mar.value in ("", None)
        assert _cell_fill_rgb(mar) == ""
    finally:
        checked.close()



def test_no_equity_curve_sheet(tmp_path: Path):
    out=tmp_path/'m.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    assert 'Equity Curve' not in wb.sheetnames

def test_unanchored_account_does_not_fabricate_equity(tmp_path: Path):
    s=sample_snapshot()
    s['items']=[
        {'id':'u1','row_type':'trade','account':'U','open_time':'2026-05-01','close_time':'2026-05-01','net_profit':10},
        {'id':'u2','row_type':'trade','account':'U','open_time':'2026-05-02','close_time':'2026-05-02','net_profit':5},
    ]
    out=tmp_path/'u.xlsx'; build_master_journal_workbook(s,out); wb=load_workbook(out)
    ws=wb['Trade Log']
    assert ws.cell(_trade_data_row(), _header_col(ws, 'Balance After')).value in ('', None)
    assert ws.cell(4, _header_col(ws, 'Balance After')).value in ('', None)


def test_trade_log_hidden_row_id_and_unsorted_override(tmp_path: Path):
    out=tmp_path/'Trading Journal.xlsx'; build_master_journal_workbook(sample_snapshot(), out)
    wb=load_workbook(out); ws=wb['Trade Log']
    headers=[ws.cell(1,c).value for c in range(1,ws.max_column+1)]
    assert '__row_id' not in headers
    assert ws.max_column == len(TRADE_LOG_HEADERS)
    assert ws.cell(_trade_data_row(), 1).comment is None
    ws.cell(_trade_data_row(), _header_col(ws, 'Test')).value='Yes'; ws.cell(_trade_data_row(), _header_col(ws, 'Setup')).value='setup-x'; wb.save(out)
    ov=read_master_journal_manual_overrides(out)
    assert ov["t1"]["is_test_trade"] is True
    assert ov["t1"]["setup"] == "setup-x"
    assert len(ws.conditional_formatting) > 0
    assert ws.cell(_trade_data_row(), _header_col(ws, "Profit %")).number_format == "0.00%"
    assert ws.cell(_trade_data_row(), 1).comment is None
    assert ws.cell(_trade_data_row(), _header_col(ws, "Profit %")).value in (0.023, -0.011)


def test_legacy_comment_row_id_preferred_over_trade_meta_after_row_move(tmp_path: Path):
    out=tmp_path/'legacy.xlsx'; build_master_journal_workbook(sample_snapshot(), out)
    wb=load_workbook(out); ws=wb['Trade Log']
    from openpyxl.comments import Comment
    ws["A4"].comment = Comment("row_id:t1", "legacy")
    ws["A5"].comment = Comment("row_id:t2", "legacy")
    for c in range(1, ws.max_column+1):
        ws.cell(4,c).value, ws.cell(5,c).value = ws.cell(5,c).value, ws.cell(4,c).value
    # stale _Trade Meta row mapping now conflicts with moved comments
    ws.cell(_trade_data_row(), _header_col(ws, "Setup")).value = "moved-comment-target"
    wb.save(out)
    ov=read_master_journal_manual_overrides(out)
    assert ov["t1"]["setup"] == "moved-comment-target"

def test_balance_after_resolution_and_duration_display(tmp_path: Path):
    s=sample_snapshot()
    s['items'] = [
        {'id':'t1','row_type':'trade','account':'A','open_time':'2026-05-01','close_time':'2026-05-01','net_profit':10,'analysis_balance_after_trade':100,'trade_duration_seconds':41},
        {'id':'t2','row_type':'trade','account':'A','open_time':'2026-05-02','close_time':'2026-05-02','net_profit':5,'trade_duration_seconds':303},
        {'id':'t3','row_type':'trade','account':'B','open_time':'2026-05-01','close_time':'2026-05-01','net_profit':3,'trade_duration_seconds':3661},
    ]
    out=tmp_path/'m3.xlsx'; build_master_journal_workbook(s,out); wb=load_workbook(out)
    ws=wb['Trade Log']
    assert ws.cell(_trade_data_row(), _header_col(ws, 'Balance After')).value == 100
    assert ws.cell(5, _header_col(ws, 'Balance After')).value == 105
    assert ws.cell(6, _header_col(ws, 'Balance After')).value in ("", None)
    assert ws.cell(1, _header_col(ws, 'Trade Duration (DD:HH:MM:SS)')).value == 'Trade Duration (DD:HH:MM:SS)'
    assert ws.cell(_trade_data_row(), _header_col(ws, 'Trade Duration (DD:HH:MM:SS)')).value == "41 seconds"
    assert ws.cell(5, _header_col(ws, 'Trade Duration (DD:HH:MM:SS)')).value == "05 minutes, 03 seconds"
    assert ws.cell(_trade_data_row(), _header_col(ws, 'Trade Duration (DD:HH:MM:SS)')).number_format == "General"
    inst=wb[SYMBOLS_SHEET]
    inst_headers = _instrument_averages_header_map(inst)
    for header in ("Shortest duration (DD:HH:MM:SS)", "Avg duration (DD:HH:MM:SS)", "Longest duration (DD:HH:MM:SS)"):
        assert _parse_duration_text(inst.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, inst_headers[header]).value) is not None
        assert inst.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, inst_headers[header]).number_format == "General"

def test_sheet_order_and_hidden_meta(tmp_path: Path):
    out=tmp_path/'x.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    assert 'Diagnostics' not in SHEET_ORDER
    assert wb.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(wb)[len(SHEET_ORDER):] == expected_report_sheet_names(sample_snapshot())
    assert '_Trade Meta' not in wb.sheetnames
    assert len(wb[STATS1_SHEET].conditional_formatting) > 0
    assert len(wb[SYMBOLS_SHEET].conditional_formatting) > 0
    assert len(wb["P&L Calendar"].conditional_formatting) == 0
    assert _cell_fill_rgb(wb["P&L Calendar"]["F2"]) == PROFIT_FILL


def test_trade_log_preserves_explicit_bybit_execution_row_id(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    snap["items"] = [
        {
            "id": "bybit:demo:execution:BTCUSDT:E1",
            "row_type": "trade",
            "source": "bybit_execution_history",
            "account": "Bybit Demo",
            "account_label": "Bybit Demo",
            "symbol": "BTCUSDT",
            "side": "Buy",
            "open_time": "2026-05-19T01:13:00+10:00",
            "close_time": "2026-05-19T01:13:00+10:00",
            "qty": 0.1,
            "entry_price": 100000.0,
            "exit_price": 100000.0,
            "commission": 0.01,
            "net_profit": None,
            "currency": "USDT",
            "asset_class": "crypto",
        }
    ]
    build_master_journal_workbook(snap, out)
    ws = load_workbook(out)["Trade Log"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rid_col = headers.index("Row ID") + 1
    assert ws.cell(_trade_data_row(), rid_col).value == "bybit:demo:execution:BTCUSDT:E1"

def test_update_data_only_migrates_legacy_all_trades_and_removes_trade_meta(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    wb["Trade Log"].title = "All Trades"
    meta = wb.create_sheet("_Trade Meta")
    meta.sheet_state = "hidden"
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    assert result["diagnostics"].get("migrated_trade_log_sheet") is True
    assert result["diagnostics"].get("removed_legacy_trade_meta") is True
    candidate = Path(result["candidate_path"])
    candidate.replace(out)

    migrated = load_workbook(out)
    assert migrated.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(migrated)[len(SHEET_ORDER):] == expected_report_sheet_names(snap)
    assert "All Trades" not in migrated.sheetnames
    assert "_Trade Meta" not in migrated.sheetnames
    migrated.close()

def _symbols_sheet_view_snapshot(ws) -> dict:
    pane = ws.sheet_view.pane
    def selection_state(selection) -> dict:
        active_cell = selection.activeCell
        sqref = str(selection.sqref or "")
        if selection.pane in {"topRight", "bottomLeft"} and active_cell in (None, "A1") and sqref in ("", "A1"):
            active_cell = None
            sqref = ""
        return {"pane": selection.pane, "activeCell": active_cell, "sqref": sqref}

    return {
        "freeze_panes": ws.freeze_panes,
        "pane": {
            "xSplit": pane.xSplit,
            "ySplit": pane.ySplit,
            "topLeftCell": pane.topLeftCell,
            "activePane": pane.activePane,
            "state": pane.state,
        },
        "selection": [
            selection_state(selection)
            for selection in ws.sheet_view.selection
        ],
    }


def test_update_data_only_repairs_symbols_freeze_pane_and_canonical_selections(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    symbols = wb[SYMBOLS_SHEET]
    symbols.freeze_panes = "B40"
    symbols.sheet_view.selection = [
        Selection(pane="topRight", activeCell=None, sqref=None),
        Selection(pane="bottomLeft", activeCell=None, sqref=None),
        Selection(pane="bottomRight", activeCell="A1", sqref="A1"),
        Selection(pane="bottomLeft", activeCell="A1", sqref="A1"),
        Selection(pane="bottomRight", activeCell="A1", sqref="A1"),
        Selection(pane="bottomLeft", activeCell="A1", sqref="A1"),
        Selection(pane="bottomRight", activeCell="G62", sqref="G62"),
    ]
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    assert result["diagnostics"].get("repaired_symbols_freeze_pane") is True
    Path(result["candidate_path"]).replace(out)

    repaired = load_workbook(out)
    view = _symbols_sheet_view_snapshot(repaired[SYMBOLS_SHEET])
    assert view["freeze_panes"] == "B3"
    assert view["pane"] == {
        "xSplit": 1.0,
        "ySplit": 2.0,
        "topLeftCell": "B3",
        "activePane": "bottomRight",
        "state": "frozen",
    }
    assert view["selection"] == [
        {"pane": "topRight", "activeCell": None, "sqref": ""},
        {"pane": "bottomLeft", "activeCell": None, "sqref": ""},
        {"pane": "bottomRight", "activeCell": "A1", "sqref": "A1"},
    ]
    assert "G62" not in {item["activeCell"] for item in view["selection"]}
    _ensure_symbols_freeze_panes(repaired[SYMBOLS_SHEET], {})
    second_view = _symbols_sheet_view_snapshot(repaired[SYMBOLS_SHEET])
    assert second_view == view
    repaired.save(out)
    repaired.close()

    second = update_master_journal_workbook_data_only(out, snap)
    assert second["ok"] is True
    Path(second["candidate_path"]).replace(out)
    reloaded = load_workbook(out)
    assert _symbols_sheet_view_snapshot(reloaded[SYMBOLS_SHEET]) == view
    assert reloaded.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(reloaded)[len(SHEET_ORDER):] == expected_report_sheet_names(snap)
    assert "_Trade Meta" not in reloaded.sheetnames
    assert "All Trades" not in reloaded.sheetnames
    reloaded.close()


def test_update_data_only_expands_symbols_filter_for_preserved_manual_rows(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb[SYMBOLS_SHEET]
    manual_col = ws.max_column + 1
    manual_row = ws.max_row + 5
    ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, manual_col).value = "Manual Metric"
    ws.cell(manual_row, manual_col).value = "Keep"
    ws.auto_filter.ref = (
        f"A{INSTRUMENT_AVERAGES_FILTER_HEADER_ROW}:"
        f"{get_column_letter(manual_col)}{INSTRUMENT_AVERAGES_DATA_START_ROW}"
    )
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)

    repaired = load_workbook(out)
    try:
        ws = repaired[SYMBOLS_SHEET]
        assert ws.cell(manual_row, manual_col).value == "Keep"
        _min_col, _min_row, max_col, max_row = range_boundaries(ws.auto_filter.ref)
        assert max_col >= manual_col
        assert max_row >= manual_row
    finally:
        repaired.close()


def test_update_data_only_survivor_guard_fails_when_row_id_header_missing(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
    ridx = headers.index("Row ID") + 1
    ws.cell(1, ridx).value = "Row ID Removed"
    wb.save(out)
    wb.close()
    with pytest.raises(RuntimeError, match="Trade Log headers cannot be migrated safely"):
        update_master_journal_workbook_data_only(out, snap, expected_survivor_row_ids=["old1", "old2"])

def test_conditional_format_colors_and_dashboard_semantics(tmp_path: Path):
    out=tmp_path/'cf.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    dash = wb[STATS1_SHEET]
    trade_log = wb["Trade Log"]
    # Dashboard semantic metrics use direct full-cell fills rather than text-only formatting.
    assert _cell_fill_rgb(dash["B3"]) == "C6EFCE"
    assert _cell_fill_rgb(dash["B4"]) == "FFC7CE"
    labels = {str(dash.cell(row, 1).value or "").strip(): row for row in range(1, dash.max_row + 1)}
    assert "Expectancy %" not in labels
    assert _cell_fill_rgb(dash.cell(labels["Min win %"], 2)) == "C6EFCE"
    assert _cell_fill_rgb(dash.cell(labels["Max loss %"], 2)) == "FFC7CE"
    assert _cell_fill_rgb(dash["H13"]) == "C6EFCE"
    assert _cell_fill_rgb(dash["I13"]) == "FFC7CE"

    # Trade Log row-level rules cover the full visible row without overlapping value-cell fill rules.
    tr = _cf_ranges(trade_log)
    assert any(r.startswith(f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}") for r in tr)
    assert not any("M2:M" in r for r in tr)
    assert not any("N2:P" in r for r in tr)
    colors = _all_rule_colors(trade_log) + _all_rule_colors(dash) + _all_rule_colors(wb["P&L Calendar"]) + _all_rule_colors(wb[SYMBOLS_SHEET])
    assert any("C6EFCE" in f and "006100" in c for f, c in colors)
    assert any("FFC7CE" in f and "9C0006" in c for f, c in colors)

def test_read_source_cashflow_new_balance_falls_back_to_balance_after_zero(tmp_path: Path):
    snap = sample_snapshot()
    snap["items"] = [
        {'id':'t1','row_type':'trade','account':'BINANCE','symbol':'BTCUSDT','side':'BUY','open_time':'2026-01-01','close_time':'2026-01-01','net_profit':1.0,'balance_after_trade':396.65720524,'currency':'USDT'},
        {'id':'c1','row_type':'cashflow','account':'BINANCE','symbol':'CASHFLOW','side':'WITHDRAWAL','open_time':'2026-01-02','close_time':'2026-01-02','net_profit':-396.65720524,'balance_after_trade':0,'cashflow_amount':-396.65720524,'cashflow_new_balance':'','currency':'USDT'},
    ]
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    ws.cell(_trade_data_row(), headers["Cashflow New Balance"]).value = None
    wb.save(out)
    wb.close()

    parsed = read_master_journal_source(out)
    cashflow_row = next(i for i in parsed["items"] if i.get("row_type") == "cashflow")
    assert cashflow_row["cashflow_new_balance"] == 0
    assert cashflow_row["cashflow_new_balance"] is not None
    assert parsed["cashflow_ledger"]["BINANCE"][-1]["new_balance"] == 0


def test_build_workbook_forces_blank_setup_for_semantic_rows(tmp_path: Path):
    snap = {
        "items": [
            {"id":"m1","row_type":"monthly_aud_reval","account":"BYBIT","symbol":"MONTHLY AUD P/L","side":"","open_time":"2026-01-31","close_time":"2026-01-31","result_cash":1.0,"notes":"monthly","setup":"STALE"},
            {"id":"c1","row_type":"cashflow","account":"BYBIT","symbol":"CASHFLOW","side":"DEPOSIT","open_time":"2026-01-20","close_time":"2026-01-20","cashflow_amount":10.0,"cashflow_new_balance":100.0,"notes":"cash detail","setup":"STALE"},
        ],
        "stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},
        "balances":[],
    }
    out = tmp_path / "semantic_blank_setup.xlsx"
    build_master_journal_workbook(snap, out)
    ws = load_workbook(out)["Trade Log"]
    setup_col = _header_col(ws, 'Setup')
    notes_col = _header_col(ws, 'Notes')
    assert ws.cell(4, setup_col).value in ("", None) and ws.cell(4, notes_col).value == "monthly"
    assert ws.cell(5, setup_col).value in ("", None) and ws.cell(5, notes_col).value == "cash detail"

def test_instrument_currency_and_percent_formats(tmp_path: Path):
    out=tmp_path/'fmt.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    inst = wb[SYMBOLS_SHEET]
    headers = _instrument_averages_header_map(inst)
    row = INSTRUMENT_AVERAGES_DATA_START_ROW
    assert inst.cell(row, headers["Win Rate %"]).number_format == "0.00%"
    assert inst.cell(row, headers["Win Rate %"]).value == 1.0
    assert inst.cell(row, headers["Net P/L %"]).number_format == "0.00%"
    assert inst.cell(row, headers["Wins"]).number_format == "0;-0;;@"
    assert inst.cell(row, headers["Losses"]).number_format == "0;-0;;@"

def test_dashboard_layout_style_columns(tmp_path: Path):
    out=tmp_path/'db.xlsx'; build_master_journal_workbook(sample_snapshot(), out); wb=load_workbook(out)
    dash=wb[STATS1_SHEET]
    assert dash['A1'].fill.fgColor.type != 'rgb' or dash['A1'].fill.fgColor.rgb != '000B1220'
    assert dash['I2'].value != 'Instrument leaders'
    expected_duration_labels = [
        "Min duration", "Avg duration", "Max duration",
        "Min Move to Break Even", "Source", "Average Move to Break Even", "Max Move to Break Even", "Source",
        "Min Move to Profit", "Source", "Average Move to Profit", "Max Move to Profit", "Source",
    ]
    labels = defaultdict(list)
    for row in range(1, dash.max_row + 1):
        label = str(dash.cell(row, 1).value or "").strip()
        if label:
            labels[label].append(row)
    duration_start = next(row for row in labels["Min duration"] if row < labels["Winners"][0])
    duration_rows = range(duration_start, duration_start + len(expected_duration_labels))
    assert [dash.cell(row, 1).value for row in duration_rows] == expected_duration_labels
    for source_row in (duration_start + 4, duration_start + 7, duration_start + 9, duration_start + 12):
        label_cell = dash.cell(source_row, 1)
        assert label_cell.value == "Source"
        assert label_cell.font.bold is False
        assert label_cell.font.italic is True
        assert label_cell.alignment.horizontal == "right"

    for section, expected_seconds in (("Winners", 3700), ("Losers", 7215)):
        section_row = labels[section][0]
        next_section_row = labels["Losers"][0] if section == "Winners" else labels["Side"][0]
        section_labels = {
            str(dash.cell(row, 1).value or "").strip(): row
            for row in range(section_row + 1, next_section_row)
            if str(dash.cell(row, 1).value or "").strip()
        }
        assert [dash.cell(section_labels[label], 1).value for label in ("Min duration", "Avg duration", "Max duration")] == [
            "Min duration", "Avg duration", "Max duration",
        ]
        for label in ("Min duration", "Avg duration", "Max duration"):
            row = section_labels[label]
            assert _parse_duration_text(dash.cell(row, 2).value) == expected_seconds
            assert dash.cell(row, 1).font.bold is True
            assert dash.cell(row, 1).font.italic is False
            assert dash.cell(row, 1).alignment.horizontal == "left"
        section_cell = dash.cell(section_row, 1)
        assert section_cell.font.bold is True
        assert section_cell.alignment.horizontal == "left"

def test_read_master_journal_source_parses_core_fields(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    parsed = read_master_journal_source(out)
    trades = [r for r in parsed["items"] if r.get("row_type") == "trade"]
    assert trades
    row = trades[0]
    for key in ("id", "qty", "entry_price", "exit_price", "stop_loss", "take_profit", "net_profit", "result_pct", "r_multiple", "trade_duration_seconds", "is_test_trade"):
        assert key in row


def test_read_source_repairs_stale_excel_id_and_dedupes_execution(tmp_path: Path):
    snapshot = {
        "items": [{
            "id": "oanda_export:live:367:370",
            "trade_number": "F1010",
            "row_type": "trade",
            "account": "OANDA LIVE",
            "asset_class": "fx",
            "symbol": "AUDJPY",
            "side": "BUY",
            "open_time": "2024-03-22 12:05:11",
            "close_time": "2024-03-22 12:23:45",
            "qty": 0.06663,
            "entry_price": 99.334,
            "exit_price": 99.185,
            "stop_loss": 99.187,
            "take_profit": 99.628,
            "net_profit": -10.0586,
            "notes": "canonical",
        }],
        "stats": {"totals": {}, "groups": {}},
        "balances": [],
    }
    path = tmp_path / "stale_duplicate.xlsx"
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    ws = wb["Trade Log"]
    headers = _trade_log_header_map(ws)
    source_row = TRADE_LOG_DATA_START_ROW
    duplicate_row = source_row + 1
    for col in range(1, ws.max_column + 1):
        ws.cell(duplicate_row, col).value = ws.cell(source_row, col).value
    ws.cell(duplicate_row, headers["Row ID"]).value = (
        "excel:PEPPERSTONE DEMO:Sheet0:541:AUDCAD:2018-07-04T11:34:00"
    )
    ws.cell(duplicate_row, headers["Entry Price"]).value = 99.33
    ws.cell(duplicate_row, headers["Exit Price"]).value = 99.19
    ws.cell(duplicate_row, headers["Notes"]).value = "duplicate"
    wb.save(path)
    wb.close()

    parsed = read_master_journal_source(path)
    trades = [row for row in parsed["items"] if row.get("row_type") == "trade"]
    assert len(trades) == 1
    assert trades[0]["id"] == "oanda_export:live:367:370"
    assert trades[0]["notes"] == "canonical"
    assert parsed["diagnostics"]["duplicate_execution_rows_removed"] == 1
    repairs = parsed["diagnostics"]["repaired_corrupted_row_ids"]
    assert len(repairs) == 1
    assert repairs[0]["reason"] == "stale_excel_row_id"
    assert repairs[0]["mismatched_fields"] == ["account", "symbol", "date"]

    update = update_master_journal_workbook_data_only(
        path,
        {"items": parsed["items"], "stats": {"totals": {}, "groups": {}}, "balances": []},
    )
    candidate = Path(update["candidate_path"])
    updated = load_workbook(candidate)
    updated_trade_log = updated["Trade Log"]
    _, _, _, filter_max_row = range_boundaries(updated_trade_log.auto_filter.ref)
    assert filter_max_row == TRADE_LOG_DATA_START_ROW
    assert all(
        updated_trade_log.cell(TRADE_LOG_DATA_START_ROW + 1, col).value in (None, "")
        for col in range(1, updated_trade_log.max_column + 1)
    )
    updated.close()


def test_read_master_journal_source_parses_result_percent_alias(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / "result_pct_alias.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    ws.append(["Open Time", "Close Time", "Account", "Symbol", "Side", "Result %"])
    ws.append(["2026-01-01 00:00:00", "2026-01-01 01:00:00", "OANDA DEMO", "EURUSD", "BUY", 0.0125])
    wb.save(p)
    out = read_master_journal_source(p)
    trade = next(r for r in out["items"] if r.get("row_type") == "trade")
    assert trade["result_pct"] == pytest.approx(1.25)


def test_read_master_journal_source_parses_numeric_ddhhmmss_duration_as_seconds(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / "duration_numeric.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    ws.append(["Open Time", "Close Time", "Account", "Symbol", "Side", "Trade Duration (DD:HH:MM:SS)", "Row ID"])
    ws.append(["2026-01-01 00:00:00", "2026-01-01 00:00:41", "A", "EURUSD", "BUY", 41, "r1"])
    ws.append(["2026-01-01 00:00:00", "2026-01-01 00:05:03", "A", "EURUSD", "BUY", 503, "r2"])
    ws.append(["2026-01-01 00:00:00", "2026-01-01 01:01:01", "A", "EURUSD", "BUY", 10101, "r3"])
    wb.save(p)
    out = read_master_journal_source(p)
    rows = {r["id"]: r for r in out["items"]}
    assert rows["r1"]["trade_duration_seconds"] == 41
    assert rows["r2"]["trade_duration_seconds"] == 303
    assert rows["r3"]["trade_duration_seconds"] == 3661


def test_read_master_journal_source_derives_duration_when_cell_blank(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / "duration_derived.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    ws.append(["Open Time", "Close Time", "Account", "Symbol", "Side", "Trade Duration (DD:HH:MM:SS)", "Row ID"])
    ws.append(["2026-01-01 00:00:00", "2026-01-01 00:01:07", "A", "EURUSD", "BUY", "", "r1"])
    wb.save(p)
    out = read_master_journal_source(p)
    row = next(r for r in out["items"] if r["id"] == "r1")
    assert row["trade_duration_seconds"] == 67


def test_instrument_averages_op_and_duration_not_blank_after_data_only_update(tmp_path: Path):
    p = tmp_path / "master.xlsx"
    snap = sample_snapshot()
    if snap.get("stats") and snap["stats"].get("by_instrument"):
        for rec in snap["stats"]["by_instrument"]:
            if isinstance(rec, dict):
                rec["net_result_pct"] = rec.get("net_result_pct", 2.5)
                rec["avg_result_pct"] = rec.get("avg_result_pct", 1.25)
                rec["min_trade_duration_seconds"] = rec.get("min_trade_duration_seconds", 41)
                rec["avg_trade_duration_seconds"] = rec.get("avg_trade_duration_seconds", 303)
                rec["max_trade_duration_seconds"] = rec.get("max_trade_duration_seconds", 3661)
    build_master_journal_workbook(snap, p)
    wb = load_workbook(p)
    inst = wb[SYMBOLS_SHEET]
    legacy_headers = _instrument_averages_header_map(inst)
    for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, inst.max_row + 1):
        if inst.cell(row, legacy_headers["Symbol"]).value in (None, ""):
            continue
        for header, value in (
            ("Shortest duration (DD:HH:MM:SS)", 41),
            ("Avg duration (DD:HH:MM:SS)", 303),
            ("Longest duration (DD:HH:MM:SS)", 3661),
        ):
            cell = inst.cell(row, legacy_headers[header])
            cell.value = value
            cell.number_format = r"00\:00\:00\:00"
    headers = [str(c.value or "") for c in inst[INSTRUMENT_AVERAGES_FILTER_HEADER_ROW]]
    # mimic live alias header style
    inst.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, headers.index("Shortest duration (DD:HH:MM:SS)") + 1).value = "Shortest (DD:HH:MM:SS)"
    inst.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, headers.index("Longest duration (DD:HH:MM:SS)") + 1).value = "Longest (DD:HH:MM:SS)"
    _ensure_trade_log_headers(wb)
    for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + 2):
        wb["Trade Log"].row_dimensions[row].height = 15
    wb.save(p); wb.close()
    res = update_master_journal_workbook_data_only(p, snap)
    Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)
    inst2 = out[SYMBOLS_SHEET]
    h2 = _instrument_averages_header_map(inst2)
    net_col = h2["Net P/L %"]
    avg_col = h2["Avg P/L %"]
    s_col = h2["Shortest duration (DD:HH:MM:SS)"]
    a_col = h2["Avg duration (DD:HH:MM:SS)"]
    l_col = h2["Longest duration (DD:HH:MM:SS)"]
    vals = [
        (
            inst2.cell(r, net_col).value,
            inst2.cell(r, avg_col).value,
            inst2.cell(r, s_col).value,
            inst2.cell(r, a_col).value,
            inst2.cell(r, l_col).value,
        )
        for r in range(INSTRUMENT_AVERAGES_DATA_START_ROW, inst2.max_row + 1)
    ]
    assert any(isinstance(v[0], (int, float)) or isinstance(v[1], (int, float)) for v in vals)
    assert any(all(isinstance(x, str) and _parse_duration_text(x) is not None for x in v[2:]) for v in vals)
    for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, inst2.max_row + 1):
        if inst2.cell(row, h2["Symbol"]).value in (None, ""):
            continue
        for col in (s_col, a_col, l_col):
            assert inst2.cell(row, col).number_format == "General"
        assert _parse_duration_text(inst2.cell(row, col).value) is not None
    for col in (s_col, a_col, l_col):
        header_cell = inst2.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, col)
        if header_cell.value in (None, "") and inst2.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col).value not in (None, ""):
            header_cell = inst2.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col)
        assert header_cell.alignment.wrap_text is True
        assert header_cell.alignment.vertical == "center"
    assert inst2.auto_filter.ref
    out.close()


def test_monthly_aud_row_uses_result_currency_and_excluded_from_metrics(tmp_path: Path):
    s = sample_snapshot()
    s["items"] = [
        {
            "id": "t1",
            "row_type": "trade",
            "symbol": "EURUSD",
            "account": "OANDA",
            "open_time": "2026-04-02T00:00:00+10:00",
            "close_time": "2026-04-02T01:00:00+10:00",
            "net_profit": 10,
            "result_pct": 1.0,
        },
        {
            "id": "monthly_aud_reval:bybit_live:2026-04",
            "row_type": "monthly_aud_reval",
            "account": "live",
            "account_label": "Bybit Live",
            "symbol": "MONTHLY AUD P/L",
            "open_time": "2026-04-01T00:00:00+10:00",
            "close_time": "2026-04-30T23:59:59+10:00",
            "result_cash": 123.45,
            "result_currency": "AUD",
        },
    ]
    out = tmp_path / "monthly.xlsx"
    build_master_journal_workbook(s, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    symbol_col = _header_col(ws, "Symbol")
    monthly_rows = [r for r in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1) if ws.cell(r, symbol_col).value == "MONTHLY AUD P/L"]
    assert len(monthly_rows) == 1
    mr = monthly_rows[0]
    assert ws.cell(mr, _header_col(ws, "Net P/L")).value == 123.45
    fmt = str(ws.cell(mr, _header_col(ws, "Net P/L")).number_format or "")
    assert "AUD" in fmt
    assert "UNKNOWN" not in fmt
    # metrics remain from trade rows only
    cal = wb["P&L Calendar"]
    assert cal["E2"].value == "1.00%, 1 trade"  # April trade row only


def test_read_master_journal_source_monthly_aud_roundtrip_fields(tmp_path: Path):
    snapshot = {
        "items": [{
            "id": "monthly_aud_reval:bybit_live:2026-04",
            "row_type": "monthly_aud_reval",
            "account": "Bybit Live",
            "account_label": "Bybit Live",
            "symbol": "MONTHLY AUD P/L",
            "open_time": "2026-04-01T00:00:00+10:00",
            "close_time": "2026-04-30T23:59:59+10:00",
            "result_cash": 123.45,
            "result_currency": "AUD",
        }],
        "stats": {"totals": {}, "groups": {}},
        "balances": [],
    }
    out = tmp_path / "monthly_roundtrip.xlsx"
    build_master_journal_workbook(snapshot, out)
    parsed = read_master_journal_source(out)
    monthly = next(r for r in parsed["items"] if r.get("row_type") == "monthly_aud_reval")
    assert monthly["result_cash"] == pytest.approx(123.45)
    assert monthly["result_currency"] == "AUD"
    assert (monthly.get("raw_refs") or {}).get("period_month") == "2026-04"
    assert monthly.get("net_profit") in (None, "")
    assert monthly.get("realized_pnl") in (None, "")

def test_trade_log_commission_zero_none_blank_and_nonzero(tmp_path: Path):
    s = sample_snapshot()
    s["items"] = [
        {"id":"c0","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-01","close_time":"2026-01-01","commission":0.0,"net_profit":1.0,"result_pct":1.0,"commission_currency":"AUD"},
        {"id":"c1","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-02","close_time":"2026-01-02","commission":None,"net_profit":1.0,"result_pct":1.0,"commission_currency":"AUD"},
        {"id":"c2","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-03","close_time":"2026-01-03","commission":"","net_profit":1.0,"result_pct":1.0,"commission_currency":"AUD"},
        {"id":"c3","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-04","close_time":"2026-01-04","commission":1.25,"net_profit":1.0,"result_pct":1.0,"commission_currency":"AUD"},
    ]
    out = tmp_path / "commission.xlsx"
    build_master_journal_workbook(s, out)
    ws = load_workbook(out)["Trade Log"]
    comm_col = _header_col(ws, 'Commission')
    assert ws.cell(4, comm_col).value in ("", None)
    assert ws.cell(5, comm_col).value in ("", None)
    assert ws.cell(6, comm_col).value in ("", None)
    assert ws.cell(7, comm_col).value == 1.25
    assert "AUD" in str(ws.cell(7, comm_col).number_format or "")


def test_trade_log_never_uses_oanda_spread_metrics_as_commission(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"] = [{
        "id": "oanda_export:demo:618:622",
        "row_type": "trade",
        "source": "oanda_transaction_export",
        "account": "OANDA DEMO",
        "symbol": "EURUSD",
        "side": "BUY",
        "open_time": "2026-04-30T19:45:59+10:00",
        "close_time": "2026-04-30T19:46:41+10:00",
        "qty": 0.06546,
        "entry_price": 1.16929,
        "exit_price": 1.16910,
        "commission": None,
        "net_profit": -1.7591,
        "balance_after_trade": 1500.65,
        "metrics": {
            "oanda_open_spread_cost": 0.3667,
            "oanda_close_spread_cost": 0.4125,
            "oanda_total_spread_cost": 0.7792,
        },
    }]
    out = tmp_path / "oanda_repaired.xlsx"
    build_master_journal_workbook(snapshot, out)
    ws = load_workbook(out)["Trade Log"]
    row_id_col = _header_col(ws, "Row ID")
    row_number = next(
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, row_id_col).value == "oanda_export:demo:618:622"
    )
    assert ws.cell(row_number, _header_col(ws, "Commission")).value in ("", None)


def test_trade_log_oanda_live_zero_commission_blank_and_spread_metrics_ignored(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"] = [{
        "id": "oanda_export:live:460:464",
        "row_type": "trade",
        "source": "oanda_transaction_export",
        "account": "OANDA LIVE",
        "symbol": "EURUSD",
        "side": "BUY",
        "open_time": "2025-11-06T07:16:35+10:00",
        "close_time": "2025-11-06T11:28:13+10:00",
        "qty": 0.06801,
        "entry_price": 1.14910,
        "exit_price": 1.15084,
        "commission": 0.0,
        "swap": 0.9922,
        "net_profit": -17.3026,
        "balance_after_trade": 1479.31,
        "metrics": {
            "oanda_open_spread_cost": 0.5407,
            "oanda_close_spread_cost": 0.5574,
            "oanda_total_spread_cost": 1.0981,
        },
    }]
    out = tmp_path / "oanda_live_repaired.xlsx"
    build_master_journal_workbook(snapshot, out)
    ws = load_workbook(out)["Trade Log"]
    row_id_col = _header_col(ws, "Row ID")
    row_number = next(
        row
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, row_id_col).value == "oanda_export:live:460:464"
    )
    assert ws.cell(row_number, _header_col(ws, "Commission")).value in ("", None)


def test_trade_log_currency_inference_avoids_unknown_and_respects_fx_vs_crypto(tmp_path: Path):
    s = sample_snapshot()
    s["items"] = [
        {"id":"o1","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-01","close_time":"2026-01-01","commission":None,"net_profit":1.0,"result_pct":1.0},
        {"id":"p1","row_type":"trade","account":"PEPPERSTONE LIVE","symbol":"GBPUSD","side":"BUY","open_time":"2026-01-02","close_time":"2026-01-02","net_profit":2.0,"result_pct":2.0},
        {"id":"m1","row_type":"monthly_aud_reval","account":"Bybit Live","symbol":"MONTHLY AUD P/L","side":"","open_time":"2026-01-31","close_time":"2026-01-31","result_cash":3.0,"net_profit":3.0},
        {"id":"b1","row_type":"trade","account":"BYBIT","symbol":"BTCUSDT","side":"SELL","open_time":"2026-01-03","close_time":"2026-01-03","net_profit":4.0,"result_pct":4.0},
        {"id":"f1","row_type":"trade","account":"OANDA DEMO","symbol":"USDCAD","side":"BUY","open_time":"2026-01-04","close_time":"2026-01-04","net_profit":5.0,"result_pct":5.0},
        {"id":"f2","row_type":"trade","account":"OANDA DEMO","symbol":"USDCHF","side":"BUY","open_time":"2026-01-05","close_time":"2026-01-05","net_profit":6.0,"result_pct":6.0},
    ]
    out = tmp_path / "currency_infer.xlsx"
    build_master_journal_workbook(s, out)
    ws = load_workbook(out)["Trade Log"]
    headers = [str(c.value or "") for c in ws[1]]
    row_id_col = headers.index("Row ID") + 1
    row_map = {str(ws.cell(r, row_id_col).value): r for r in range(2, ws.max_row + 1)}
    for rid in ("o1","p1","m1","b1","f1","f2"):
        assert rid in row_map
        assert "UNKNOWN" not in str(ws.cell(row_map[rid], _header_col(ws, 'Commission')).number_format or "")
        assert "UNKNOWN" not in str(ws.cell(row_map[rid], _header_col(ws, 'Net P/L')).number_format or "")
    assert "AUD" in str(ws.cell(row_map["o1"], _header_col(ws, 'Net P/L')).number_format or "")
    assert "AUD" in str(ws.cell(row_map["p1"], _header_col(ws, 'Net P/L')).number_format or "")
    assert "AUD" in str(ws.cell(row_map["m1"], _header_col(ws, 'Net P/L')).number_format or "")
    assert "USDT" in str(ws.cell(row_map["b1"], _header_col(ws, 'Net P/L')).number_format or "")
    assert "AUD" in str(ws.cell(row_map["f1"], _header_col(ws, 'Net P/L')).number_format or "")
    assert "AUD" in str(ws.cell(row_map["f2"], _header_col(ws, 'Net P/L')).number_format or "")
    assert ws.cell(row_map["o1"], _header_col(ws, "Commission")).value in ("", None)

def test_update_data_only_repairs_unknown_trade_log_currency_formats(tmp_path: Path):
    s = sample_snapshot()
    s["items"] = [
        {"id":"o1","row_type":"trade","account":"OANDA DEMO","symbol":"EURUSD","side":"BUY","open_time":"2026-01-01","close_time":"2026-01-01","net_profit":1.0,"result_pct":1.0},
        {"id":"b1","row_type":"trade","account":"BYBIT","symbol":"BTCUSDT","side":"SELL","open_time":"2026-01-02","close_time":"2026-01-02","net_profit":2.0,"result_pct":2.0},
        {"id":"m1","row_type":"monthly_aud_reval","account":"Bybit Live","symbol":"MONTHLY AUD P/L","open_time":"2026-01-31","close_time":"2026-01-31","result_cash":3.0,"net_profit":3.0},
    ]
    out = tmp_path / "repair_unknown.xlsx"
    build_master_journal_workbook(s, out)
    wb = load_workbook(out); ws = wb["Trade Log"]
    ws.cell(_trade_data_row(), _header_col(ws, "Commission")).number_format = '#,##0.00 "UNKNOWN"'
    ws.cell(_trade_data_row(), _header_col(ws, "Net P/L")).number_format = '#,##0.00 "UNKNOWN"'
    ws.cell(5, _header_col(ws, "Net P/L")).number_format = '#,##0.00 "UNKNOWN"'
    ws.cell(4, _header_col(ws, "Net P/L")).number_format = '#,##0.00 "UNKNOWN"'
    wb.save(out); wb.close()
    res = update_master_journal_workbook_data_only(out, s)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(out)
    ws2 = load_workbook(out)["Trade Log"]
    assert ws2.cell(1, _header_col(ws2, "Target Price")).value == "Target Price"
    assert ws2.cell(1, _header_col(ws2, "Target Distance")).value == "Target Distance"
    assert ws2.cell(1, _header_col(ws2, "Commission")).value == "Commission"
    assert ws2.cell(1, _header_col(ws2, "Net P/L")).value == "Net P/L"
    commission_col = _header_col(ws2, "Commission")
    net_pl_col = _header_col(ws2, "Net P/L")
    for r in range(2, ws2.max_row + 1):
        assert "UNKNOWN" not in str(ws2.cell(r, commission_col).number_format or "")
        assert "UNKNOWN" not in str(ws2.cell(r, net_pl_col).number_format or "")
        assert "UNKNOWN" not in str(ws2.cell(r, _header_col(ws2, "Target Price")).number_format or "")
        assert "UNKNOWN" not in str(ws2.cell(r, _header_col(ws2, "Target Distance")).number_format or "")

def test_metric_refresh_same_row_sections_and_non_a_balance_block(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path/'mj.xlsx'
    wb=Workbook(); ws=wb.active; ws.title='Dashboard'; wb.create_sheet('Trade Log'); wb.create_sheet('Instrument Averages')
    # anchors on same row
    ws['A1']='Overall'; ws['D1']='FX'; ws['G1']='Crypto'; ws['J1']='Winners'; ws['J8']='Losers'; ws['J14']='Drawdown'; ws['M1']='Instrument leaders'
    ws['P1']='Account Balances'
    # labels
    for base in ['A','D','G']:
        ws[f'{base}2']='Trades'; ws[f'{base}3']='Avg R'; ws[f'{base}4']='Win rate'; ws[f'{base}5']='Max loss %'; ws[f'{base}6']='Source'
    ws['P2']='Account'; ws['Q2']='Balance'; ws['R2']='Currency'; ws['S2']='As Of'; ws['P3']='BYBIT'
    ws['B4'].number_format='0.00%'; ws['E4'].number_format='0.00%'; ws['H4'].number_format='0.00%'
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={'stats':{'totals':{},'groups':{'by_market':{'overall':{'trades':1,'avg_r_multiple':2.0,'win_rate_pct':50.0,'min_result_pct':-1.25,'metric_sources':{'min_result_pct':{'symbol':'EURUSD','date':'2026-01-01'}}},'fx':{'trades':2,'avg_r_multiple':3.0,'win_rate_pct':25.0,'min_result_pct':-2.0,'metric_sources':{'min_result_pct':{'symbol':'GBPUSD','date':'2026-01-02'}}},'crypto':{'trades':3,'avg_r_multiple':4.0,'win_rate_pct':75.0,'min_result_pct':-3.0,'metric_sources':{'min_result_pct':{'symbol':'BTCUSDT','date':'2026-01-03'}}}},'risk_expectancy':{},'leaders':{}}},'balances':[{'account_label':'BYBIT','balance':12.5,'currency':'USDT','as_of':'2026-01-04'}]}
    res = update_master_journal_workbook_data_only(p,snap); Path(res["candidate_path"]).replace(p)
    out=load_workbook(p)
    d=out[STATS1_SHEET]
    assert d['B2'].value==1 and d['E2'].value==2 and d['H2'].value==3
    assert d['B3'].value==2.0 and d['E3'].value==3.0 and d['H3'].value==4.0
    assert d['B4'].value==0.5 and d['E4'].value==0.25 and d['H4'].value==0.75
    assert d['Q3'].value == 12.5
    assert d['R3'].value == 'USDT'
    assert 'GBPUSD' in str(d['E6'].value)

def test_embedded_fx_crypto_duration_without_duration_section(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p=tmp_path/'d.xlsx'
    wb=Workbook(); ws=wb.active; ws.title='Dashboard'; wb.create_sheet('Trade Log'); wb.create_sheet('Instrument Averages')
    ws['A1']='Overall'; ws['D1']='FX'; ws['G1']='Crypto'; ws['J1']='Winners'; ws['J8']='Losers'; ws['J14']='Drawdown'; ws['M1']='Instrument leaders'; ws['T1']='Account Balances'
    ws['D2']='FX shortest'; ws['D3']='Source'; ws['D4']='FX longest'; ws['D5']='Source'
    ws['G2']='Crypto shortest'; ws['G3']='Source'; ws['G4']='Crypto longest'; ws['G5']='Source'
    ws['T2']='Account'; ws['U2']='Balance'; ws['V2']='Currency'; ws['T3']='Bybit Live'
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={'stats':{'totals':{},'groups':{'by_market':{'overall':{},'fx':{},'crypto':{}},'risk_expectancy':{},'leaders':{},'duration':{'fx_shortest_seconds':10,'fx_longest_seconds':20,'crypto_shortest_seconds':30,'crypto_longest_seconds':40,'metric_sources':{'fx_shortest_seconds':{'symbol':'EURUSD','date':'2026-01-01'},'fx_longest_seconds':{'symbol':'GBPUSD','date':'2026-01-02'},'crypto_shortest_seconds':{'symbol':'BTCUSDT','date':'2026-01-03'},'crypto_longest_seconds':{'symbol':'ETHUSDT','date':'2026-01-04'}}}}},'balances':[{'account_label':'BYBIT','balance':1,'currency':'USDT'}]}
    res = update_master_journal_workbook_data_only(p,snap); Path(res["candidate_path"]).replace(p)
    out=load_workbook(p)[STATS1_SHEET]
    assert out['E2'].value == "10 seconds" and out['E2'].number_format == "General"
    assert out['E4'].value == "20 seconds" and out['E4'].number_format == "General"
    assert out['H2'].value == "30 seconds" and out['H2'].number_format == "General"
    assert out['H4'].value == "40 seconds" and out['H4'].number_format == "General"
    assert 'EURUSD' in str(out['E3'].value) and 'ETHUSDT' in str(out['H5'].value)

def test_read_master_journal_source_asset_class_regressions(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / 'asset_class.xlsx'
    wb = Workbook(); ws = wb.active; ws.title = 'Trade Log'
    headers = ['Open Time','Close Time','Account','Symbol','Side']
    ws.append(headers)
    ws.append(['2026-01-01','2026-01-01','UNKNOWN','ABCDEF','BUY'])
    ws.append(['2026-01-01','2026-01-01','UNKNOWN','ABC/DEF','BUY'])
    ws.append(['2026-01-01','2026-01-01','PEPPERSTONE LIVE','EURUSD','BUY'])
    ws.append(['2026-01-01','2026-01-01','OANDA LIVE','EUR/USD','BUY'])
    ws.append(['2026-01-01','2026-01-01','BYBIT LIVE','BTCUSD','BUY'])
    ws.append(['2026-01-01','2026-01-01','BINANCE LIVE','ETHUSD','BUY'])
    _ensure_trade_log_headers(wb); wb.save(p)
    parsed = read_master_journal_source(p)
    rows = parsed['items']
    assert rows[0]['asset_class'] == ''
    assert rows[1]['asset_class'] == ''
    assert rows[2]['asset_class'] == 'fx'
    assert rows[3]['asset_class'] == 'fx'
    assert rows[4]['asset_class'] == 'crypto'
    assert rows[5]['asset_class'] == 'crypto'


def test_read_master_journal_source_recomputes_corrupted_monthly_trade_row_id(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "corrupt_read.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for idx, header in enumerate(TRADE_LOG_HEADERS, start=1):
        ws.cell(1, idx).value = header
    row = {
        "Open Time": "2026-05-02",
        "Close Time": "2026-05-02",
        "Account": "PEPPERSTONE DEMO",
        "Symbol": "EURUSD",
        "Side": "BUY",
        "Qty": 1,
        "Entry Price": 1.1,
        "Exit Price": 1.2,
        "Net P/L": 10,
        "Row Type": "trade",
        "Row ID": "monthly_aud_reval:bybit_live:2026-05",
    }
    for header, value in row.items():
        ws.cell(2, TRADE_LOG_HEADERS.index(header) + 1).value = value
    wb.save(p)

    parsed = read_master_journal_source(p)
    item = parsed["items"][0]
    assert item["row_type"] == "trade"
    assert item["id"].startswith("sig:")
    assert item["id"] != "monthly_aud_reval:bybit_live:2026-05"
    assert parsed["diagnostics"]["repaired_corrupted_row_ids"][0]["reason"] == "invalid_monthly_aud_reval_row_id"


def test_update_master_journal_workbook_data_only_repairs_corrupted_row_id_cells(tmp_path: Path):
    snapshot = sample_snapshot()
    path = tmp_path / "sync_repair.xlsx"
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    ws = wb["Trade Log"]
    row_id_col = _header_col(ws, "Row ID")
    ws.cell(_trade_data_row(), row_id_col).value = "monthly_aud_reval:bybit_live:2026-05"
    wb.save(path)

    result = update_master_journal_workbook_data_only(path, snapshot)
    Path(result["candidate_path"]).replace(path)
    repaired = load_workbook(path)["Trade Log"]
    assert repaired.cell(_trade_data_row(), row_id_col).value == snapshot["items"][0]["id"]
    assert repaired.cell(_trade_data_row(), row_id_col).value != "monthly_aud_reval:bybit_live:2026-05"
    assert result["diagnostics"].get("repaired_trade_log_row_ids", 0) >= 0

def test_instrument_leaders_skips_missing_optional_rows(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p=tmp_path/'leaders.xlsx'
    wb=Workbook(); ws=wb.active; ws.title='Dashboard'; wb.create_sheet('Trade Log'); wb.create_sheet('Instrument Averages')
    ws['A1']='Overall'; ws['D1']='FX'; ws['G1']='Crypto'; ws['J1']='Winners'; ws['J8']='Losers'; ws['J14']='Drawdown'; ws['M1']='Instrument leaders'; ws['T1']='Account Balances'
    ws['M2']='Metric'; ws['N2']='Symbol'; ws['O2']='Wins'; ws['P2']='Losses'; ws['Q2']='Trades'
    ws['M3']='FX most wins'; ws['M4']='FX most losses'; ws['M5']='Crypto most wins'  # missing crypto most losses row intentionally
    ws['T2']='Account'; ws['U2']='Balance'; ws['V2']='Currency'
    ws['T3']='Bybit Live'; ws['U3']='1'; ws['V3']='USDT'
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={'stats':{'totals':{},'groups':{'by_market':{'overall':{},'fx':{},'crypto':{}},'risk_expectancy':{},'duration':{},'leaders':{
        'most_wins_instrument':{'symbol':'EURUSD','wins':4,'losses':1,'trades':5},
        'most_losses_instrument':{'symbol':'GBPUSD','wins':1,'losses':4,'trades':5},
        'fx_most_wins_instrument':{'symbol':'EURUSD','wins':3,'losses':1,'trades':4},
        'fx_most_losses_instrument':{'symbol':'XAUUSD','wins':1,'losses':3,'trades':4},
        'crypto_most_wins_instrument':{'symbol':'BTCUSDT','wins':6,'losses':2,'trades':8},
        'crypto_most_losses_instrument':{'symbol':'ETHUSDT','wins':2,'losses':6,'trades':8},
    }}},'balances':[{'account_label':'BYBIT','balance':2,'currency':'USDT'}]}
    result=update_master_journal_workbook_data_only(p,snap); Path(result["candidate_path"]).replace(p)
    out=load_workbook(p)[STATS1_SHEET]
    assert all(out[cell].value in (None, "") for cell in ("M1", "M2", "M3", "M4", "M5", "N3", "O3", "P3", "Q3"))
    assert result['diagnostics'].get('cleared_stats2_instrument_leaders_table') is True

    result2=update_master_journal_workbook_data_only(p,snap); Path(result2["candidate_path"]).replace(p)
    out2=load_workbook(p)[STATS1_SHEET]
    labels = [str(out2.cell(r, 13).value or '').strip().lower() for r in range(1, out2.max_row + 1)]
    assert labels.count('crypto most losses') == 0
    assert result2['diagnostics'].get('cleared_stats2_instrument_leaders_table') in (None, True)

def test_account_balances_restores_missing_rows_without_layout_mutation(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    src = tmp_path / "m.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    _ensure_trade_log_headers(wb); wb.save(src)
    snap = {'stats':{'totals':{},'groups':{'by_market':{'overall':{},'fx':{},'crypto':{}},'risk_expectancy':{},'leaders':{},'duration':{}}},'balances':[
        {"account_label": "Bybit Demo", "balance": 123.456789, "currency": "USDT", "as_of": "2026-05-16"},
        {"account_label": "BYBIT", "balance": 10.123456789, "currency": "USDT", "as_of": "2026-05-16"},
    ]}
    res = update_master_journal_workbook_data_only(src, snap)
    Path(res["candidate_path"]).replace(src)
    out = load_workbook(src)
    d = out[STATS1_SHEET]
    found = {}
    for r in range(3, d.max_row + 1):
        label = str(d.cell(r, 20).value or "").strip()
        if label in {"BYBIT DEMO", "BYBIT"}:
            found[label] = r
    assert "BYBIT DEMO" in found and "BYBIT" in found
    assert isinstance(d.cell(found["BYBIT DEMO"], 21).value, (int, float))
    assert d.cell(found["BYBIT DEMO"], 22).value == "USDT"
    assert str(d.cell(found["BYBIT DEMO"], 23).value) == "2026-05-16"
    assert _sheetnames_without_target_r_metadata(out) == [*SHEET_ORDER, *expected_report_sheet_names(snap)]
    out.close()

def test_account_balances_reuses_blank_row_before_append(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    p = tmp_path / "reuse.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    ws["T3"]="BYBIT"; ws["U3"]=1.0; ws["V3"]="USDT"
    ws["T4"]=None; ws["U4"]=None; ws["V4"]=None
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={'stats':{'totals':{},'groups':{'by_market':{'overall':{},'fx':{},'crypto':{}},'risk_expectancy':{},'leaders':{},'duration':{}}},'balances':[{'account_label':'Bybit Demo','balance':2.5,'currency':'USDT','as_of':'2026-05-16'}]}
    res = update_master_journal_workbook_data_only(p, snap); Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)[STATS1_SHEET]
    assert out["T4"].value == "BYBIT DEMO"
    assert out["U4"].value == 2.5


def test_account_balances_renames_stale_bybit_live_row_to_bybit(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    p = tmp_path / "stale_only.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    ws["T3"]="Bybit Live"; ws["U3"]=1.0; ws["V3"]="USDT"
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[{"account_label":"BYBIT","balance":9.5,"currency":"USDT","as_of":"2026-05-16"}]}
    res = update_master_journal_workbook_data_only(p, snap); Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)[STATS1_SHEET]
    assert out["T3"].value == "BYBIT"
    assert out["U3"].value == 9.5
    assert out["V3"].value == "USDT"
    labels = [str(out.cell(r, 20).value or "").strip() for r in range(3, out.max_row + 1)]
    assert "Bybit Live" not in labels


def test_account_balances_clears_duplicate_stale_bybit_live_row(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    p = tmp_path / "stale_dup.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    ws["T3"]="BYBIT"; ws["U3"]=3.0; ws["V3"]="USDT"
    ws["T4"]="Bybit Live"; ws["U4"]=2.0; ws["V4"]="USDT"
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[{"account_label":"BYBIT","balance":7.0,"currency":"USDT"}]}
    res = update_master_journal_workbook_data_only(p, snap); Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)[STATS1_SHEET]
    assert out["T3"].value == "BYBIT"
    assert out["U3"].value == 7.0
    assert out["T4"].value in (None, "")
    assert out["U4"].value in (None, "")


def test_account_balances_clears_duplicate_stale_bybit_live_alias_row(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    p = tmp_path / "stale_dup_alias.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    ws["T3"]="BYBIT"; ws["U3"]=3.0; ws["V3"]="USDT"
    ws["T4"]="bybit_live"; ws["U4"]=2.0; ws["V4"]="USDT"
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[{"account_label":"BYBIT","balance":7.0,"currency":"USDT"}]}
    res = update_master_journal_workbook_data_only(p, snap); Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)[STATS1_SHEET]
    assert out["T3"].value == "BYBIT"
    assert out["T4"].value in (None, "")

def test_update_data_only_rewrites_calendar_to_one_line_layout(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path / "merged-calendar.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    for merged in list(cal.merged_cells.ranges):
        cal.unmerge_cells(str(merged))
    for row in range(1, cal.max_row + 1):
        for col in range(1, cal.max_column + 1):
            cal.cell(row, col).value = None
    months = ["January","February","March","April","May","June","July","August","September","October","November","December"]
    for i, m in enumerate(months, start=3):
        cal.cell(1, i).value = m
    cal.merge_cells("A2:A3"); cal.merge_cells("A4:A5"); cal.merge_cells("A6:A7")
    cal["A2"]=2026; cal["A4"]=2025; cal["A6"]=2024
    cal["B2"]="P/L %"; cal["B3"]="Total Trades"; cal["B4"]="P/L %"; cal["B5"]="Total Trades"; cal["B6"]="P/L %"; cal["B7"]="Total Trades"
    wb.save(p)
    wb.close()
    snap = {
        "items": [{"id":"t1","row_type":"trade","account":"BYBIT DEMO","symbol":"BTCUSDT","side":"BUY","open_time":"2026-05-01","close_time":"2026-05-01","net_profit":10.0,"result_pct":1.0}],
        "stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},
        "balances":[{"account_label":"BYBIT DEMO","balance":100.0,"currency":"USDT","as_of":"2026-05-16"}],
    }
    res = update_master_journal_workbook_data_only(p, snap)
    Path(res["candidate_path"]).replace(p)
    out = load_workbook(p, data_only=True)
    out_cal = out["P&L Calendar"]
    assert list(out_cal.merged_cells.ranges) == []
    assert out_cal["A1"].value == "Month"
    assert [out_cal.cell(1, col).value for col in range(2, 5)] == [2024, 2025, 2026]
    assert out_cal["A6"].value == "May"
    assert out_cal["D6"].value == "1.00%, 1 trade"
    assert out_cal["D2"].value in (None, "")
    assert _cell_fill_rgb(out_cal["D2"]) == ""
    assert out_cal.freeze_panes == "B2"
    assert out_cal.auto_filter.ref == "A1:D13"
    assert not any(str(out_cal.cell(row, 1).value or "").endswith(" Trades") for row in range(2, out_cal.max_row + 1))
    d = out[STATS2_SHEET]
    values = [d.cell(row, col).value for row in range(1, d.max_row + 1) for col in range(1, d.max_column + 1)]
    assert "BYBIT DEMO" in values
    assert any(isinstance(value, (int, float)) and value == 100.0 for value in values)
    out.close()


def test_update_data_only_fails_on_unrepaired_crypto_zero_qty(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path / "zero_qty_fail.xlsx"
    wb=Workbook(); ws=wb.active; ws.title='Dashboard'; wb.create_sheet('Trade Log'); wb.create_sheet('Instrument Averages'); wb.create_sheet('P&L Calendar')
    ws["A1"]="Overall"; ws["D1"]="FX"; ws["G1"]="Crypto"; ws["J1"]="Winners"; ws["J8"]="Losers"; ws["J14"]="Drawdown"; ws["M1"]="Instrument leaders"; ws["T1"]="Account Balances"
    ws["T2"]="Account"; ws["U2"]="Balance"; ws["V2"]="Currency"; ws["W2"]="As Of"
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"items":[{"id":"z1","row_type":"trade","account":"BYBIT","symbol":"BTCUSDT","side":"BUY","qty":0,"entry_price":100,"exit_price":100,"net_profit":0,"open_time":"2026-01-01","close_time":"2026-01-01"}],"stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[]}
    out = update_master_journal_workbook_data_only(p, snap)
    assert out["ok"] is False
    assert "Unrepaired crypto zero-quantity" in str(out.get("error"))


def test_zero_qty_repair_from_raw_refs_closed_size():
    from tools.master_journal_workbook import _repair_or_flag_zero_trade_qty
    row = {"id":"r1","row_type":"trade","account":"BYBIT","symbol":"BTCUSDT","qty":0,"raw_refs":{"closedSize":"0.015"}}
    fixed = _repair_or_flag_zero_trade_qty(dict(row))
    assert fixed["qty"] == 0.015


def test_zero_qty_repair_from_pnl_inference_crypto():
    from tools.master_journal_workbook import _repair_or_flag_zero_trade_qty
    row = {"id":"r2","row_type":"trade","account":"BYBIT","symbol":"BTCUSDT","side":"BUY","qty":0,"entry_price":100.0,"exit_price":110.0,"commission":1.0,"net_profit":9.0}
    fixed = _repair_or_flag_zero_trade_qty(dict(row))
    assert fixed["qty"] == 1.0


def test_zero_qty_fx_is_diagnosed_not_inferred():
    from tools.master_journal_workbook import _repair_or_flag_zero_trade_qty
    row = {"id":"r3","row_type":"trade","account":"OANDA LIVE","symbol":"EUR/USD","side":"BUY","qty":0,"entry_price":1.1,"exit_price":1.2,"commission":0.1,"net_profit":1.0}
    fixed = _repair_or_flag_zero_trade_qty(dict(row))
    assert fixed["qty"] == 0
    assert "zero_qty_unrepaired_fx" in (fixed.get("diagnostics") or [])

def test_update_data_only_adds_missing_calendar_year_row(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path / "append-year.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    for merged in list(cal.merged_cells.ranges):
        cal.unmerge_cells(str(merged))
    for row in range(1, cal.max_row + 1):
        for col in range(1, cal.max_column + 1):
            cal.cell(row, col).value = None
    for i, m in enumerate(["January","February","March","April","May","June","July","August","September","October","November","December"], start=3):
        cal.cell(1, i).value = m
    cal.merge_cells("A2:A3"); cal["A2"]=2026; cal["B2"]="P/L %"; cal["B3"]="Total Trades"
    wb.save(p)
    wb.close()
    snap = {"items":[{"id":"t1","row_type":"trade","account":"BYBIT DEMO","symbol":"BTCUSDT","side":"BUY","open_time":"2027-01-05","close_time":"2027-01-05","result_pct":2.0}],
            "stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},
            "balances":[{"account_label":"BYBIT DEMO","balance":100.0,"currency":"USDT","as_of":"2026-05-16"}]}
    res = update_master_journal_workbook_data_only(p, snap); Path(res["candidate_path"]).replace(p)
    out = load_workbook(p, data_only=True)["P&L Calendar"]
    assert list(out.merged_cells.ranges) == []
    assert int(out["A2"].value) == 2027
    assert out["B2"].value == "2.00%, 1 trade"
    assert not any(str(out.cell(row, 1).value or "").endswith(" Trades") for row in range(2, out.max_row + 1))

def test_instrument_leaders_custom_layout_populates_values(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path / "leaders-layout.xlsx"
    wb = Workbook(); d = wb.active; d.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); cal = wb.create_sheet("P&L Calendar")
    d["A1"]="Account Balances"; d["A2"]="Account"; d["B2"]="Balance"; d["C2"]="Currency"; d["A3"]="BYBIT DEMO"
    d["G1"]="Overall"; d["J1"]="FX"; d["M1"]="Crypto"; d["A11"]="Instrument leaders"; d["G12"]="Winners"; d["G17"]="Losers"; d["G22"]="Drawdown"
    d["A12"]="Metric"; d["B12"]="Symbol"; d["C12"]="Wins"; d["D12"]="Losses"; d["E12"]="Trades"
    labels=["Overall most wins","Overall most losses","FX most wins","FX most losses","Crypto most wins","Crypto most losses"]
    for i, lbl in enumerate(labels, start=13): d.cell(i,1).value=lbl
    for i,m in enumerate(["January","February","March","April","May","June","July","August","September","October","November","December"], start=3): cal.cell(1,i).value=m
    cal.merge_cells("A2:A3"); cal["A2"]=2026; cal["B2"]="P/L %"; cal["B3"]="Total Trades"
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"items":[{"id":"t1","row_type":"trade","close_time":"2026-05-01","result_pct":1.0}],
          "stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"duration":{},"leaders":{
              "most_wins_instrument":{"symbol":"EURUSD","wins":5,"losses":1,"trades":6},
              "most_losses_instrument":{"symbol":"BTCUSDT","wins":1,"losses":5,"trades":6},
              "fx_most_wins_instrument":{"symbol":"EURUSD","wins":5,"losses":1,"trades":6},
              "fx_most_losses_instrument":{"symbol":"GBPUSD","wins":1,"losses":5,"trades":6},
              "crypto_most_wins_instrument":{"symbol":"ETHUSDT","wins":4,"losses":2,"trades":6},
              "crypto_most_losses_instrument":{"symbol":"SOLUSDT","wins":1,"losses":5,"trades":6},
          }}},
          "balances":[{"account_label":"BYBIT DEMO","balance":100.0,"currency":"USDT"}]}
    res=update_master_journal_workbook_data_only(p,snap); Path(res["candidate_path"]).replace(p)
    out=load_workbook(p)[STATS1_SHEET]
    assert all(out.cell(row, col).value in (None, "") for row in range(11, 19) for col in range(1, 6))
    assert res["diagnostics"].get("cleared_stats2_instrument_leaders_table") is True

def test_legacy_all_trades_migrates_to_trade_log(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only, build_master_journal_workbook
    p = tmp_path / "legacy.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    wb["Trade Log"].title = "All Trades"
    wb.save(p)
    snap = sample_snapshot()
    res=update_master_journal_workbook_data_only(p,snap)
    assert res["ok"] is True
    assert res["diagnostics"].get("migrated_trade_log_sheet") is True
    Path(res["candidate_path"]).replace(p)
    out=load_workbook(p)
    assert "Trade Log" in out.sheetnames and "All Trades" not in out.sheetnames
    out.close()

def test_both_all_trades_and_trade_log_fails_ambiguous(tmp_path: Path):
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    from openpyxl import Workbook
    p = tmp_path / "ambiguous.xlsx"
    wb = Workbook(); wb.remove(wb.active)
    wb.create_sheet("Dashboard"); wb.create_sheet("All Trades"); wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    _ensure_trade_log_headers(wb); wb.save(p)
    snap={"items":[],"stats":{"totals":{},"groups":{"by_market":{"overall":{},"fx":{},"crypto":{}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[]}
    with pytest.raises(RuntimeError, match="ambiguous trade sheets"):
        update_master_journal_workbook_data_only(p,snap)


def test_update_data_only_overwrites_stale_dashboard_account_balances_with_zero(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    stale = sample_snapshot()
    stale["balances"] = [
        {"account_label": "PEPPERSTONE DEMO", "balance": 4.78, "currency": "AUD", "as_of": "2026-05-10"},
        {"account_label": "BINANCE", "balance": 396.65720524, "currency": "USDT", "as_of": "2026-05-10"},
    ]
    build_master_journal_workbook(stale, out)
    snap = sample_snapshot()
    snap["balances"] = [
        {"account_label": "PEPPERSTONE DEMO", "balance": 0, "currency": "AUD", "as_of": "2026-05-11", "balance_source": "broker_account_summary"},
        {"account_label": "BINANCE", "balance": 0, "currency": "USDT", "as_of": "2026-05-11", "balance_source": "broker_account_summary"},
    ]
    res = update_master_journal_workbook_data_only(out, snap)
    Path(res["candidate_path"]).replace(out)
    wb = load_workbook(out, data_only=True)
    assert wb.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert _sheetnames_without_target_r_metadata(wb)[len(SHEET_ORDER):] == expected_report_sheet_names(snap)
    vals = _dashboard_account_balances(wb[STATS2_SHEET])
    assert vals["PEPPERSTONE DEMO"][0] == 0
    assert vals["BINANCE"][0] == 0
    wb.close()


def test_infer_trade_duration_rounds_to_nearest_second():
    from tools.master_journal_workbook import _infer_trade_duration_seconds
    assert _infer_trade_duration_seconds({"row_type":"trade","trade_duration_seconds":60.1}) == 60
    assert _infer_trade_duration_seconds({"row_type":"trade","trade_duration_seconds":60.5}) == 61
    assert _infer_trade_duration_seconds({"row_type":"trade","trade_duration_seconds":0.2}) == 1
    assert _infer_trade_duration_seconds({"row_type":"trade","trade_duration_seconds":60}) == 60


def test_canonicalize_and_dedupe_balances_prefers_cashflow_anchor():
    from tools.master_journal_workbook import _canonicalize_and_dedupe_balances
    balances = [
        {"account_label": "Bybit Demo", "balance": 369.64962148, "currency": "USDT", "balance_source": "trade_timeline", "as_of": "2026-05-26T01:00:00Z"},
        {"account_label": "Bybit Demo", "balance": 319.8339282399999, "currency": "USDT", "balance_source": "cashflow_anchor_plus_trades", "as_of": "2026-05-26T02:00:00Z"},
    ]
    deduped = _canonicalize_and_dedupe_balances(balances)
    assert len(deduped) == 1
    assert deduped[0]["balance"] == 319.8339282399999


def test_canonicalize_balances_prefers_cashflow_zero_over_stale_authoritative_trade():
    from tools.master_journal_workbook import _canonicalize_and_dedupe_balances
    balances = _canonicalize_and_dedupe_balances([
        {'account_label':'BINANCE','balance':396.65720524,'currency':'USDT','balance_source':'authoritative_trade_balance','as_of':'2020-10-24T02:18:00'},
        {'account_label':'BINANCE','balance':0,'currency':'USDT','balance_source':'cashflow_anchor_plus_trades','as_of':'2020-10-26T00:00:00'},
    ])
    assert len(balances) == 1
    assert balances[0]['balance'] == 0
    assert balances[0]['balance_source'] == 'cashflow_anchor_plus_trades'


def test_canonicalize_balances_prefers_pepperstone_demo_cashflow_zero_over_stale_trade_variants():
    from tools.master_journal_workbook import _canonicalize_and_dedupe_balances
    balances = _canonicalize_and_dedupe_balances([
        {'account_label':'Pepperstone Demo','balance':4.78,'currency':'AUD','balance_source':'authoritative_trade_balance','as_of':'2018-07-25T23:00:00'},
        {'account_label':'pepperstone_demo','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades','as_of':'2018-07-26T00:00:00'},
        {'account_label':'PEPPERSTONE-DEMO','balance':None,'currency':'AUD','balance_source':'timeline_missing','as_of':'2018-07-27T00:00:00'},
    ])
    assert len(balances) == 1
    assert balances[0]['account_label'] == 'PEPPERSTONE DEMO'
    assert balances[0]['account'] == 'PEPPERSTONE DEMO'
    assert balances[0]['balance'] == 0
    assert balances[0]['currency'] == 'AUD'
    assert balances[0]['balance_source'] == 'cashflow_anchor_plus_trades'


def test_update_data_only_verifies_dashboard_binance_zero_balance(tmp_path: Path):
    from tools.master_journal_workbook import build_master_journal_workbook, update_master_journal_workbook_data_only
    p = tmp_path / 'Trading Journal.xlsx'
    stale = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'BINANCE','balance':396.65720524,'currency':'USDT','balance_source':'authoritative_trade_balance'}]}
    build_master_journal_workbook(stale, p)
    fresh = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'BINANCE','balance':0,'currency':'USDT','balance_source':'cashflow_anchor_plus_trades','as_of':'2020-10-26T00:00:00'}]}
    result = update_master_journal_workbook_data_only(p, fresh)
    assert result['ok'] is True
    assert 'BINANCE' in result['diagnostics']['account_balance_verified']
    candidate = Path(result['candidate_path'])
    wb = load_workbook(candidate, data_only=True)
    try:
        values = _dashboard_account_balances(wb[STATS2_SHEET])
        assert values['BINANCE'][0] == 0
    finally:
        wb.close()


def test_update_data_only_overwrites_stale_pepperstone_demo_balance_with_zero(tmp_path: Path):
    from tools.master_journal_workbook import build_master_journal_workbook, update_master_journal_workbook_data_only
    p = tmp_path / 'Trading Journal.xlsx'
    stale = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'Pepperstone Demo','balance':4.78,'currency':'AUD','balance_source':'authoritative_trade_balance'}]}
    build_master_journal_workbook(stale, p)
    fresh = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'pepperstone_demo','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades','as_of':'2018-07-26T00:00:00'}]}
    result = update_master_journal_workbook_data_only(p, fresh)
    assert result['ok'] is True
    assert 'PEPPERSTONE DEMO' in result['diagnostics']['account_balance_verified']
    candidate = Path(result['candidate_path'])
    wb = load_workbook(candidate, data_only=True)
    try:
        values = _dashboard_account_balances(wb[STATS2_SHEET])
        assert values['PEPPERSTONE DEMO'] == (0, 'AUD')
    finally:
        wb.close()


def test_update_data_only_fails_if_pepperstone_demo_balance_remains_stale(tmp_path: Path, monkeypatch):
    import tools.master_journal_workbook as mjw
    from tools.master_journal_workbook import build_master_journal_workbook, update_master_journal_workbook_data_only
    p = tmp_path / 'Trading Journal.xlsx'
    stale = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'Pepperstone Demo','balance':4.78,'currency':'AUD','balance_source':'authoritative_trade_balance'}]}
    build_master_journal_workbook(stale, p)
    real_write = mjw._write_value_preserving_cell

    def _block_pepperstone_demo_balance_write(ws, row, col, value):
        if value == 0 and str(ws.cell(row, col - 1).value or '').strip().upper() == 'PEPPERSTONE DEMO':
            return False
        return real_write(ws, row, col, value)

    monkeypatch.setattr(mjw, '_write_value_preserving_cell', _block_pepperstone_demo_balance_write)
    fresh = {'items': [], 'stats': {'totals': {}, 'groups': {}}, 'balances': [{'account_label':'PEPPERSTONE DEMO','balance':0,'currency':'AUD','balance_source':'cashflow_anchor_plus_trades','as_of':'2018-07-26T00:00:00'}]}
    result = update_master_journal_workbook_data_only(p, fresh)
    assert result['ok'] is False
    assert result['error'] == 'dashboard_account_balance_verification_failed'
    mismatches = result['diagnostics']['account_balance_mismatches']
    assert len(mismatches) == 1
    assert mismatches[0]['account'] == 'PEPPERSTONE DEMO'
    assert mismatches[0]['expected'] == 0.0
    assert mismatches[0]['actual'] == 4.78
    assert isinstance(mismatches[0]['row'], int)

def test_update_data_only_writes_overall_fx_crypto_streak_metrics(tmp_path: Path):
    from openpyxl import Workbook
    from tools.master_journal_workbook import update_master_journal_workbook_data_only
    p = tmp_path / "streaks.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"; wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["A1"] = "Overall"; ws["D1"] = "FX"; ws["G1"] = "Crypto"; ws["J1"] = "Winners"; ws["J8"] = "Losers"; ws["J14"] = "Drawdown"; ws["M1"] = "Instrument leaders"; ws["T1"] = "Account Balances"
    for col in ("A", "D", "G"):
        ws[f"{col}2"] = "Winning Streak"
        ws[f"{col}3"] = "Losing Streak"
    ws["T2"] = "Account"; ws["U2"] = "Balance"; ws["V2"] = "Currency"; ws["T3"] = "BINANCE"
    _ensure_trade_log_headers(wb)
    for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + 2):
        wb["Trade Log"].row_dimensions[row].height = 15
    wb.save(p); wb.close()
    snap = {"stats":{"totals":{},"groups":{"by_market":{"overall":{"winning_streak":4,"losing_streak":3},"fx":{"winning_streak":2,"losing_streak":1},"crypto":{"winning_streak":5,"losing_streak":6}},"risk_expectancy":{},"leaders":{},"duration":{}}},"balances":[{"account_label":"BINANCE","balance":0,"currency":"USDT"}]}
    res = update_master_journal_workbook_data_only(p, snap)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(p)
    out = load_workbook(p)[STATS1_SHEET]
    assert out["B2"].value == 4 and out["B3"].value == 3
    assert out["E2"].value == 2 and out["E3"].value == 1
    assert out["H2"].value == 5 and out["H3"].value == 6



def test_read_master_journal_source_uses_read_only_workbook(monkeypatch, tmp_path):
    from tools import master_journal_workbook as mjw
    from tools.master_journal_workbook import build_master_journal_workbook
    path = tmp_path / 'Trading Journal.xlsx'
    build_master_journal_workbook({'items': [{'id': 't1', 'row_type': 'trade', 'account': 'BINANCE', 'symbol': 'BTCUSDT', 'side': 'BUY', 'open_time': '2026-01-01', 'close_time': '2026-01-01'}], 'stats': {'totals': {}, 'groups': {}}, 'balances': []}, path)
    real_load = mjw.load_workbook
    calls = []
    def wrapped_load_workbook(*args, **kwargs):
        calls.append(kwargs)
        return real_load(*args, **kwargs)
    monkeypatch.setattr(mjw, 'load_workbook', wrapped_load_workbook)
    payload = mjw.read_master_journal_source(path)
    assert payload['items']
    assert calls and calls[0].get('read_only') is True
    assert calls[0].get('data_only') is True


def test_read_master_journal_source_distance_percent_format_is_source_aware(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / "distance_percent.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    ws.append(["Open Time", "Close Time", "Account", "Symbol", "Side", "Stop Loss Distance", "Target Distance", "Row ID"])
    ws.append(["2026-01-01", "2026-01-01", "OANDA DEMO", "EURUSD", "BUY", 0.01, 1.0, "pct-row"])
    ws.cell(2, 6).number_format = "0.00%"
    ws.cell(2, 7).number_format = "General"
    wb.save(p)
    out = read_master_journal_source(p)
    row = next(r for r in out["items"] if r["id"] == "pct-row")
    assert row["stop_loss_distance_pct"] == pytest.approx(1.0)
    assert row["target_distance_pct"] == pytest.approx(1.0)


def _target_r_roundtrip_fx_rows(count: int = 5) -> list[dict]:
    return [
        {
            "id": f"fx-target-{idx}",
            "row_type": "trade",
            "asset_class": "fx",
            "account": "OANDA DEMO",
            "symbol": "EURUSD" if idx % 2 else "GBPUSD",
            "side": "BUY",
            "open_time": f"2026-01-{idx:02d}T00:00:00Z",
            "close_time": f"2026-01-{idx:02d}T01:00:00Z",
            "entry_price": 1.1,
            "exit_price": 1.1 + (0.001 * (1.0 + (idx / 10.0))),
            "stop_loss": 1.099,
            "take_profit": 1.102,
            "planned_entry_price": 1.1,
            "planned_stop_price": 1.099,
            "planned_target_price": 1.102,
            "qty": 0.1,
            "qty_raw": 10000,
            "qty_unit": "lots",
            "net_profit": 15.0 * (1.0 + (idx / 10.0)),
            "currency": "AUD",
            "original_loss_conversion_factor": 1.5,
            "original_loss_conversion_factor_source": "homeConversionFactors.lossQuoteHome.factor",
            "original_loss_conversion_factor_time": f"2026-01-{idx:02d}T00:00:00Z",
            "original_open_transaction_id": f"open-{idx}",
        }
        for idx in range(1, count + 1)
    ]


def _target_r_recommendation_cells(path: Path) -> dict:
    wb = load_workbook(path, data_only=True)
    try:
        stats1 = wb[STATS1_SHEET]
        max_target_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Max target %")
        target_recommendation_row = max_target_row + 2
        stats1_recs = tuple(stats1.cell(target_recommendation_row, col).value for col in (2, 3, 4))
        symbols = wb[SYMBOLS_SHEET]
        headers = _instrument_averages_header_map(symbols)
        symbol_recs = {
            str(symbols.cell(row, headers["Symbol"]).value or "").strip().upper(): symbols.cell(row, headers[TARGET_RECOMMENDATION_HEADER]).value
            for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1)
            if str(symbols.cell(row, headers["Symbol"]).value or "").strip()
        }
        return {"stats1": stats1_recs, "symbols": symbol_recs}
    finally:
        wb.close()


def _visible_workbook_layout_signature(path: Path) -> dict:
    from tools import master_journal_workbook as mjw

    wb = load_workbook(path, data_only=False)
    try:
        visible_names = [name for name in wb.sheetnames if name != TARGET_R_METADATA_SHEET]
        signature = {"sheetnames": visible_names}
        for name in visible_names:
            ws = wb[name]
            max_style_row = min(ws.max_row or 1, 6)
            max_style_col = min(ws.max_column or 1, 8)
            formulas = {}
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formulas[cell.coordinate] = cell.value
            signature[name] = {
                "freeze": ws.freeze_panes,
                "filter": ws.auto_filter.ref if ws.auto_filter else None,
                "merged": sorted(str(rng) for rng in ws.merged_cells.ranges),
                "widths": {key: dim.width for key, dim in ws.column_dimensions.items() if dim.width is not None},
                "heights": {key: dim.height for key, dim in ws.row_dimensions.items() if dim.height is not None},
                "formulas": formulas,
                "styles": {
                    (row, col): mjw._style_signature(ws.cell(row, col))
                    for row in range(1, max_style_row + 1)
                    for col in range(1, max_style_col + 1)
                },
            }
        return signature
    finally:
        wb.close()


def test_target_r_metadata_sheet_is_not_created_and_legacy_sheet_is_removed(tmp_path: Path):
    rows = _target_r_roundtrip_fx_rows()
    loss_row = dict(rows[0])
    loss_row.update({"id": "fx-target-loss", "net_profit": -15.0, "exit_price": 1.099})
    rows.append(loss_row)
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook({"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, out)

    wb = load_workbook(out)
    try:
        assert TARGET_R_METADATA_SHEET not in wb.sheetnames
        wb.create_sheet(TARGET_R_METADATA_SHEET)
        wb[TARGET_R_METADATA_SHEET].sheet_state = "veryHidden"
        wb.save(out)
    finally:
        wb.close()

    parsed = read_master_journal_source(out)
    assert "original_loss_conversion_factor" not in parsed["items"][0]
    assert "original_open_transaction_id" not in parsed["items"][0]

    result = update_master_journal_workbook_data_only(
        out,
        {"items": parsed["items"], "stats": {"totals": {}, "groups": {}}, "balances": []},
    )
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)

    final_wb = load_workbook(out, data_only=True)
    try:
        assert TARGET_R_METADATA_SHEET not in final_wb.sheetnames
    finally:
        final_wb.close()


def test_visible_row_ids_survive_visible_rows_move_without_target_r_metadata(tmp_path: Path):
    rows = _target_r_roundtrip_fx_rows(2)
    out = tmp_path / "metadata-moved.xlsx"
    build_master_journal_workbook({"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, out)

    wb = load_workbook(out)
    ws = wb["Trade Log"]
    row_a = TRADE_LOG_DATA_START_ROW
    row_b = TRADE_LOG_DATA_START_ROW + 1
    for col in range(1, ws.max_column + 1):
        ws.cell(row_a, col).value, ws.cell(row_b, col).value = ws.cell(row_b, col).value, ws.cell(row_a, col).value
    wb.save(out)
    wb.close()

    parsed = read_master_journal_source(out)
    by_id = {row["id"]: row for row in parsed["items"]}
    assert set(by_id) == {"fx-target-1", "fx-target-2"}
    assert all("original_open_transaction_id" not in row for row in parsed["items"])
    assert all("original_loss_conversion_factor" not in row for row in parsed["items"])


def test_source_ignores_legacy_target_r_metadata_and_diagnoses_visible_duplicate_ids(tmp_path: Path):
    rows = _target_r_roundtrip_fx_rows(2)
    duplicate_visible = tmp_path / "metadata-duplicate-visible.xlsx"
    build_master_journal_workbook({"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, duplicate_visible)
    wb = load_workbook(duplicate_visible)
    ws = wb["Trade Log"]
    row_id_col = _trade_log_header_map(ws)["Row ID"]
    ws.cell(TRADE_LOG_DATA_START_ROW + 1, row_id_col).value = "fx-target-1"
    wb.save(duplicate_visible)
    wb.close()
    parsed_duplicate = read_master_journal_source(duplicate_visible)
    assert parsed_duplicate["diagnostics"]["target_r_duplicate_trade_row_ids"] == ["fx-target-1"]
    assert all("original_loss_conversion_factor" not in row for row in parsed_duplicate["items"])

    legacy_metadata = tmp_path / "metadata-legacy-hidden.xlsx"
    build_master_journal_workbook({"items": rows[:1], "stats": {"totals": {}, "groups": {}}, "balances": []}, legacy_metadata)
    wb = load_workbook(legacy_metadata)
    ws = wb.create_sheet(TARGET_R_METADATA_SHEET)
    ws.sheet_state = "veryHidden"
    ws.append(["Row ID", "Payload JSON"])
    ws.append(["fx-target-1", '{"original_loss_conversion_factor":9.9,"original_open_transaction_id":"wrong"}'])
    wb.save(legacy_metadata)
    wb.close()
    parsed_legacy_metadata = read_master_journal_source(legacy_metadata)
    assert "target_r_metadata_duplicate_row_ids" not in parsed_legacy_metadata["diagnostics"]
    assert "original_loss_conversion_factor" not in parsed_legacy_metadata["items"][0]
    assert "original_open_transaction_id" not in parsed_legacy_metadata["items"][0]


def test_missing_target_r_metadata_still_allows_recorded_fx_price_r(tmp_path: Path):
    from tools.master_journal_workbook import _target_r_recommendation

    rows = _target_r_roundtrip_fx_rows()
    loss_row = dict(rows[0])
    loss_row.update({"id": "fx-target-loss", "net_profit": -15.0, "exit_price": 1.099})
    rows.append(loss_row)
    for row in rows:
        row.pop("original_loss_conversion_factor", None)
        row.pop("original_loss_conversion_factor_source", None)
        row.pop("original_loss_conversion_factor_time", None)
        row.pop("original_open_transaction_id", None)
    out = tmp_path / "metadata-missing.xlsx"
    build_master_journal_workbook({"items": rows, "stats": {"totals": {}, "groups": {}}, "balances": []}, out)
    wb = load_workbook(out)
    if TARGET_R_METADATA_SHEET in wb.sheetnames:
        del wb[TARGET_R_METADATA_SHEET]
    wb.save(out)
    wb.close()

    risk = _target_r_recommendation(read_master_journal_source(out)["items"])

    assert risk["eligible_target_r_wins"] == 5
    assert risk["eligible_target_r_losses"] == 1
    assert risk["target_recommendation"].startswith(("Increase target", "Decrease target"))
    assert risk["target_r_calculation_method_counts"]["price_captured_r_from_original_plan"] == 5
    assert risk["target_r_winning_exclusion_reasons"] == {}
    assert risk["target_r_winning_exclusion_reasons_by_market"]["fx"] == {}


def test_generated_trade_log_distance_fraction_displays_one_percent(tmp_path: Path):
    out = tmp_path / "one_percent.xlsx"
    snap = sample_snapshot()
    snap["items"] = [{
        "id": "onepct", "row_type": "trade", "account": "OANDA DEMO", "symbol": "EURUSD",
        "side": "BUY", "open_time": "2026-01-01", "close_time": "2026-01-01",
        "entry_price": 100.0, "stop_loss": 99.0, "take_profit": 101.0,
        "net_profit": 1.0, "result_pct": 1.0,
    }]
    build_master_journal_workbook(snap, out)
    ws = load_workbook(out)["Trade Log"]
    stop_distance = ws.cell(_trade_data_row(), _header_col(ws, "Stop Loss Distance"))
    target_distance = ws.cell(_trade_data_row(), _header_col(ws, "Target Distance"))
    assert stop_distance.value == pytest.approx(0.01)
    assert stop_distance.number_format == "0.00%"
    assert target_distance.value == pytest.approx(0.01)
    assert target_distance.number_format == "0.00%"


def test_recommendations_are_removed_from_trade_log_and_kept_in_symbols_and_stats1(tmp_path: Path):
    out = tmp_path / "recommendations.xlsx"
    snap = sample_snapshot()
    snap["items"] = [
        {
            "id": f"eur-win-{idx}", "row_type": "trade", "account": "OANDA DEMO",
            "asset_class": "fx", "symbol": "EURUSD", "side": "BUY",
            "open_time": f"2026-01-{idx:02d}", "close_time": f"2026-01-{idx:02d}",
            "entry_price": 100.0, "exit_price": 100.0 + (10.0 * r_multiple),
            "stop_loss": 99.0, "take_profit": 140.0,
            "planned_entry_price": 100.0, "planned_stop_price": 90.0, "planned_target_price": 140.0,
            "net_profit": 10.0 * r_multiple, "result_pct": 1.0, "r_multiple": r_multiple,
            "original_risk_amount": 10.0,
            "original_risk_currency": "AUD",
        }
        for idx, r_multiple in enumerate([2.4, 2.5, 2.7, 3.0, 3.1, 3.4, 3.6, 4.8], start=1)
    ] + [
        {
            "id": "eur-loss", "row_type": "trade", "account": "OANDA DEMO",
            "asset_class": "fx", "symbol": "EURUSD", "side": "BUY",
            "open_time": "2026-01-20", "close_time": "2026-01-20",
            "entry_price": 100.0, "exit_price": 95.0, "stop_loss": 98.0, "take_profit": 104.0,
            "planned_entry_price": 100.0, "planned_stop_price": 80.0, "planned_target_price": 140.0,
            "net_profit": -5.0, "result_pct": -0.5, "r_multiple": -0.5,
        },
    ]
    snap["stats"]["by_instrument"] = [
        {
            "symbol": "EURUSD", "asset_class": "fx", "total_trades": 9,
            "wins": 8, "losses": 1, "break_even": 0,
            "long_trades": 9, "short_trades": 0,
            "avg_sl_pct_wins": 1.0, "avg_sl_pct_losses": 2.0,
            "avg_tp_pct_wins": 40.0, "avg_tp_pct_losses": 4.0,
        }
    ]
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    trade_log = wb["Trade Log"]
    trade_headers = _trade_log_header_map(trade_log)
    assert STOP_RECOMMENDATION_HEADER not in trade_headers
    assert TARGET_RECOMMENDATION_HEADER not in trade_headers

    symbols = wb[SYMBOLS_SHEET]
    symbol_headers = _instrument_averages_header_map(symbols)
    assert symbol_headers[STOP_RECOMMENDATION_HEADER] == symbol_headers["Avg stop % (L)"] + 1
    assert symbol_headers[TARGET_RECOMMENDATION_HEADER] == symbol_headers["Avg target % (L)"] + 1
    assert symbols.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, symbol_headers[STOP_RECOMMENDATION_HEADER]).value == "Recommendation"
    assert symbols.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, symbol_headers[TARGET_RECOMMENDATION_HEADER]).value == "Recommendation"
    expected_stop = "Decrease stop — Recommended: 10.00% (10.00 pp below loss average)"
    assert symbols.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, symbol_headers[STOP_RECOMMENDATION_HEADER]).value == expected_stop
    assert symbols.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, symbol_headers[TARGET_RECOMMENDATION_HEADER]).value == "Decrease target \u2014 Recommended: 3.25R (current median: 4.0R)"
    assert _cell_fill_rgb(symbols.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, symbol_headers[STOP_RECOMMENDATION_HEADER])) not in {"FFF2CC", PROFIT_FILL, LOSS_FILL}
    assert _cell_fill_rgb(symbols.cell(INSTRUMENT_AVERAGES_DATA_START_ROW, symbol_headers[TARGET_RECOMMENDATION_HEADER])) not in {"FFF2CC", PROFIT_FILL, LOSS_FILL}

    stats1 = wb[STATS1_SHEET]
    recommendation_rows = [
        row for row in range(1, stats1.max_row + 1)
        if str(stats1.cell(row, 1).value or "").strip() == "Recommendation"
    ]
    assert len(recommendation_rows) >= 2
    assert [stats1.cell(recommendation_rows[0], col).value for col in (2, 3, 4)] == [expected_stop, expected_stop, "Need wins & losses"]
    assert [stats1.cell(recommendation_rows[1], col).value for col in (2, 3, 4)] == [
        "Decrease target \u2014 Recommended: 3.25R (current median: 4.0R)",
        "Decrease target \u2014 Recommended: 3.25R (current median: 4.0R)",
        "No eligible winning trades",
    ]
    wb.close()


def test_real_trading_journal_source_data_produces_numeric_target_recommendations():
    source = Path("journal") / "Trading Journal.xlsx"
    if not source.exists():
        pytest.skip("checked-in Trading Journal.xlsx is not available")
    from tools.master_journal_workbook import _target_r_recommendation

    snap = read_master_journal_source(source)
    active_rows = [
        row for row in (snap.get("items") or [])
        if str(row.get("row_type") or "trade").strip().lower() == "trade" and not row.get("is_test_trade")
    ]
    groups = {
        "overall": active_rows,
        "fx": [row for row in active_rows if str(row.get("asset_class") or "").lower() == "fx"],
        "crypto": [row for row in active_rows if str(row.get("asset_class") or "").lower() == "crypto"],
        "EURUSD": [row for row in active_rows if str(row.get("symbol") or "").upper() == "EURUSD"],
        "USDJPY": [row for row in active_rows if str(row.get("symbol") or "").upper() == "USDJPY"],
        "BTCUSDT": [row for row in active_rows if str(row.get("symbol") or "").upper() == "BTCUSDT"],
        "ETHUSDT": [row for row in active_rows if str(row.get("symbol") or "").upper() == "ETHUSDT"],
    }
    for name, rows in groups.items():
        rec = _target_r_recommendation(rows)
        _assert_target_recommendation_is_numeric_or_insufficient(rec["target_recommendation"])
        if name in {"overall", "crypto", "BTCUSDT", "ETHUSDT"}:
            assert rec["target_recommendation"] != TARGET_RECOMMENDATION_INSUFFICIENT


def test_checked_in_trading_journal_has_no_direction_only_or_zero_target_recommendations():
    path = Path("journal") / "Trading Journal.xlsx"
    if not path.exists():
        pytest.skip("checked-in Trading Journal.xlsx is not available")
    wb = load_workbook(path, data_only=True)
    try:
        trade_headers = _trade_log_header_map(wb["Trade Log"])
        assert TARGET_RECOMMENDATION_HEADER not in trade_headers

        stats1 = wb[STATS1_SHEET]
        max_target_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Max target %")
        target_recommendation_row = max_target_row + 2
        for col in (2, 3, 4):
            _assert_target_recommendation_is_numeric_or_insufficient(stats1.cell(target_recommendation_row, col).value)

        symbols = wb[SYMBOLS_SHEET]
        symbol_headers = _instrument_averages_header_map(symbols)
        target_col = symbol_headers[TARGET_RECOMMENDATION_HEADER]
        symbol_col = symbol_headers["Symbol"]
        rec_by_symbol = {
            str(symbols.cell(row, symbol_col).value or "").strip().upper(): symbols.cell(row, target_col).value
            for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1)
        }
        for symbol in ("EURUSD", "USDJPY", "BTCUSDT", "ETHUSDT"):
            _assert_target_recommendation_is_numeric_or_insufficient(rec_by_symbol.get(symbol))
        for symbol in ("BTCUSDT", "ETHUSDT"):
            assert rec_by_symbol.get(symbol) != TARGET_RECOMMENDATION_INSUFFICIENT
        assert any(value == TARGET_RECOMMENDATION_INSUFFICIENT for value in rec_by_symbol.values())
    finally:
        wb.close()


def test_checked_in_trading_journal_target_recommendations_match_backend_calculations():
    path = Path("journal") / "Trading Journal.xlsx"
    if not path.exists():
        pytest.skip("checked-in Trading Journal.xlsx is not available")
    from tools.master_journal_workbook import _target_r_recommendation

    snap = read_master_journal_source(path)
    active_rows = [
        row for row in (snap.get("items") or [])
        if str(row.get("row_type") or "trade").strip().lower() == "trade" and not row.get("is_test_trade")
    ]
    expected_by_group = {
        "overall": _target_r_recommendation(active_rows, scope="overall")["target_recommendation"],
        "fx": _target_r_recommendation([row for row in active_rows if str(row.get("asset_class") or "").lower() == "fx"])["target_recommendation"],
        "crypto": _target_r_recommendation([row for row in active_rows if str(row.get("asset_class") or "").lower() == "crypto"])["target_recommendation"],
    }
    expected_by_symbol = {
        symbol: _target_r_recommendation([row for row in active_rows if str(row.get("symbol") or "").upper() == symbol])["target_recommendation"]
        for symbol in ("BTCUSDT", "ETHUSDT")
    }

    wb = load_workbook(path, data_only=True)
    try:
        stats1 = wb[STATS1_SHEET]
        max_target_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Max target %")
        target_recommendation_row = max_target_row + 2
        assert [stats1.cell(target_recommendation_row, col).value for col in (2, 3, 4)] == [
            expected_by_group["overall"],
            expected_by_group["fx"],
            expected_by_group["crypto"],
        ]

        symbols = wb[SYMBOLS_SHEET]
        headers = _instrument_averages_header_map(symbols)
        rec_by_symbol = {
            str(symbols.cell(row, headers["Symbol"]).value or "").strip().upper(): symbols.cell(row, headers[TARGET_RECOMMENDATION_HEADER]).value
            for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1)
        }
        for symbol, expected in expected_by_symbol.items():
            assert rec_by_symbol.get(symbol) == expected
    finally:
        wb.close()


def _git_workbook_fixture(commit: str, tmp_path: Path) -> Path:
    try:
        blob = subprocess.check_output(["git", "show", f"{commit}:journal/Trading Journal.xlsx"])
    except (subprocess.CalledProcessError, FileNotFoundError) as exc:
        pytest.skip(f"baseline workbook {commit} is not available: {exc}")
    path = tmp_path / f"Trading Journal {commit}.xlsx"
    path.write_bytes(blob)
    return path


def _symbols_a1_at2_border_signatures(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    try:
        symbols = wb[SYMBOLS_SHEET]
        return {
            (row, col): _border_signature(symbols.cell(row, col))
            for row in (1, 2)
            for col in range(1, 47)
        }
    finally:
        wb.close()


def _border_with_test_sides(border: Border, *, left: Side | None = None, right: Side | None = None) -> Border:
    return Border(
        left=copy(left if left is not None else border.left),
        right=copy(right if right is not None else border.right),
        top=copy(border.top),
        bottom=copy(border.bottom),
        diagonal=copy(border.diagonal),
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=copy(border.vertical),
        horizontal=copy(border.horizontal),
        start=copy(border.start),
        end=copy(border.end),
    )


def _symbols_trade_log_formatting_signatures(path: Path) -> dict:
    wb = load_workbook(path, data_only=False)
    try:
        symbols = wb[SYMBOLS_SHEET]
        symbol_headers = _instrument_averages_header_map(symbols)
        symbol_last_col = max(symbol_headers.values())
        symbol_rows = range(1, symbols.max_row + 1)

        trade_log = wb["Trade Log"]
        trade_headers = _trade_log_header_map(trade_log)
        trade_last_col = max(trade_headers.values())
        trade_rows = range(1, trade_log.max_row + 1)

        return {
            "symbols_borders": {
                (row, col): _border_signature(symbols.cell(row, col))
                for row in symbol_rows
                for col in range(1, symbol_last_col + 1)
            },
            "symbols_row1_merges": sorted(
                str(rng) for rng in symbols.merged_cells.ranges if rng.min_row <= 2 and rng.max_row >= 1
            ),
            "symbols_widths": {
                get_column_letter(col): symbols.column_dimensions[get_column_letter(col)].width
                for col in range(1, symbol_last_col + 1)
            },
            "symbols_row_heights": {
                row: symbols.row_dimensions[row].height
                for row in symbol_rows
            },
            "trade_log_borders": {
                (row, col): _border_signature(trade_log.cell(row, col))
                for row in trade_rows
                for col in range(1, trade_last_col + 1)
            },
            "trade_log_fills": {
                (row, col): copy(trade_log.cell(row, col).fill)
                for row in trade_rows
                for col in range(1, trade_last_col + 1)
            },
            "trade_log_header_merges": sorted(
                str(rng) for rng in trade_log.merged_cells.ranges if rng.min_row <= TRADE_LOG_HEADER_ROWS
            ),
            "trade_log_widths": {
                get_column_letter(col): trade_log.column_dimensions[get_column_letter(col)].width
                for col in range(1, trade_last_col + 1)
            },
            "trade_log_row_heights": {
                row: trade_log.row_dimensions[row].height
                for row in trade_rows
            },
        }
    finally:
        wb.close()


def _symbols_cell_value(path: Path, coordinate: str) -> object:
    wb = load_workbook(path, data_only=False)
    try:
        return wb[SYMBOLS_SHEET][coordinate].value
    finally:
        wb.close()


def _formula_error_cells(path: Path) -> list[str]:
    error_tokens = ("#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A")
    errors: set[str] = set()
    for data_only in (False, True):
        wb = load_workbook(path, data_only=data_only, read_only=True)
        try:
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.strip().upper().startswith(error_tokens):
                            errors.add(f"{ws.title}!{cell.coordinate}={value}")
        finally:
            wb.close()
    return sorted(errors)


def _assert_symbols_trade_log_layout_contract(
    path: Path,
    *,
    duration_border_exceptions: set[str] | None = None,
) -> None:
    duration_border_exceptions = duration_border_exceptions or set()
    wb = load_workbook(path, data_only=False)
    try:
        symbols = wb[SYMBOLS_SHEET]
        headers = _instrument_averages_header_map(symbols)
        merges = {str(rng) for rng in symbols.merged_cells.ranges}
        assert "AJ1:AM1" in merges
        assert "AN1:AQ1" in merges
        assert "AR1:AV1" not in merges
        for header in (
            "Shortest duration (DD:HH:MM:SS)",
            "Avg duration (DD:HH:MM:SS)",
            "Longest duration (DD:HH:MM:SS)",
            "Avg winning duration (DD:HH:MM:SS)",
            "Avg losing duration (DD:HH:MM:SS)",
        ):
            col = headers[header]
            letter = get_column_letter(col)
            assert f"{letter}1:{letter}2" in merges
            assert symbols.cell(1, col).value == header
            assert symbols.cell(2, col).value is None
            assert symbols.column_dimensions[letter].width == 43

        longest_col = headers["Longest duration (DD:HH:MM:SS)"]
        winning_col = headers["Avg winning duration (DD:HH:MM:SS)"]
        losing_col = headers["Avg losing duration (DD:HH:MM:SS)"]
        for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1):
            for col, expected in (
                (longest_col, "thin"),
                (winning_col, "thin"),
                (losing_col, "thick"),
            ):
                cell = symbols.cell(row, col)
                if cell.coordinate not in duration_border_exceptions:
                    assert cell.border.right.style == expected

        trade_log = wb["Trade Log"]
        trade_last_col = max(_trade_log_header_map(trade_log).values())
        assert all(
            _cell_fill_rgb(trade_log.cell(row, col)) == ""
            for row in range(1, TRADE_LOG_HEADER_ROWS + 1)
            for col in range(1, trade_last_col + 1)
        )
        for cell in ("X1", "AC1"):
            assert trade_log[cell].border.right.style == "medium"
        for row in range(TRADE_LOG_DATA_START_ROW, trade_log.max_row + 1):
            assert trade_log.row_dimensions[row].height == TRADE_LOG_DATA_ROW_HEIGHT
    finally:
        wb.close()


def _open_save_reopen_cycle(path: Path) -> None:
    wb = load_workbook(path, data_only=False)
    wb.save(path)
    wb.close()
    reopened = load_workbook(path, data_only=False)
    reopened.close()


def test_40f2ef3_symbols_formatting_migration_repairs_manual_borders(tmp_path: Path):
    baseline_path = _git_workbook_fixture("772eedb", tmp_path)
    damaged_path = _git_workbook_fixture("40f2ef3", tmp_path)
    baseline_symbol_header_borders = _symbols_a1_at2_border_signatures(baseline_path)

    snapshot = sample_snapshot()
    result = update_master_journal_workbook_data_only(damaged_path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(damaged_path)
    _open_save_reopen_cycle(damaged_path)

    assert _symbols_a1_at2_border_signatures(damaged_path) == baseline_symbol_header_borders
    _assert_symbols_trade_log_layout_contract(damaged_path)

    wb = load_workbook(damaged_path, data_only=False)
    symbols = wb[SYMBOLS_SHEET]
    manual_side = Side(style="double", color="FFFF0000")
    symbols["AJ1"].border = _border_with_test_sides(symbols["AJ1"].border, left=manual_side)
    wb.save(damaged_path)
    wb.close()
    reopened = load_workbook(damaged_path, data_only=False)
    manual_signature = _border_signature(reopened[SYMBOLS_SHEET]["AJ1"])
    reopened.close()

    result = update_master_journal_workbook_data_only(damaged_path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(damaged_path)
    _open_save_reopen_cycle(damaged_path)
    reopened = load_workbook(damaged_path, data_only=False)
    assert _border_signature(reopened[SYMBOLS_SHEET]["AJ1"]) == manual_signature
    reopened.close()
    assert _formula_error_cells(damaged_path) == []


def test_5c88d4a_symbols_formatting_migration_repairs_manual_borders(tmp_path: Path):
    baseline_path = _git_workbook_fixture("772eedb", tmp_path)
    damaged_path = _git_workbook_fixture("5c88d4a", tmp_path)
    baseline_symbol_header_borders = _symbols_a1_at2_border_signatures(baseline_path)

    result = update_master_journal_workbook_data_only(damaged_path, sample_snapshot())
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(damaged_path)
    _open_save_reopen_cycle(damaged_path)

    assert _symbols_a1_at2_border_signatures(damaged_path) == baseline_symbol_header_borders
    _assert_symbols_trade_log_layout_contract(damaged_path)
    assert _formula_error_cells(damaged_path) == []


def test_checked_in_trading_journal_resync_preserves_symbols_and_trade_log_borders(tmp_path: Path):
    path = Path("journal") / "Trading Journal.xlsx"
    if not path.exists():
        pytest.skip("checked-in Trading Journal.xlsx is not available")
    baseline_path = _git_workbook_fixture("772eedb", tmp_path)
    baseline_symbol_header_borders = _symbols_a1_at2_border_signatures(baseline_path)
    assert _symbols_a1_at2_border_signatures(path) == baseline_symbol_header_borders
    _assert_symbols_trade_log_layout_contract(path)
    assert _formula_error_cells(path) == []

    work = tmp_path / "Trading Journal.xlsx"
    shutil.copy2(path, work)

    before = _symbols_trade_log_formatting_signatures(work)
    snapshot = sample_snapshot()
    for _ in range(2):
        result = update_master_journal_workbook_data_only(work, snapshot)
        assert result["ok"] is True
        Path(result["candidate_path"]).replace(work)
        after = _symbols_trade_log_formatting_signatures(work)
        assert {key: value for key, value in after.items() if key != "trade_log_fills"} == {
            key: value for key, value in before.items() if key != "trade_log_fills"
        }

    _open_save_reopen_cycle(work)
    after = _symbols_trade_log_formatting_signatures(work)
    assert {key: value for key, value in after.items() if key != "trade_log_fills"} == {
        key: value for key, value in before.items() if key != "trade_log_fills"
    }
    assert _symbols_a1_at2_border_signatures(work) == baseline_symbol_header_borders
    _assert_symbols_trade_log_layout_contract(work)
    assert _formula_error_cells(work) == []


def test_source_resync_preserves_manual_symbols_header_and_data_borders(tmp_path: Path):
    checked_in = Path("journal") / "Trading Journal.xlsx"
    if not checked_in.exists():
        pytest.skip("checked-in Trading Journal.xlsx is not available")
    work = tmp_path / "Trading Journal manual borders.xlsx"
    shutil.copy2(checked_in, work)

    before = _symbols_trade_log_formatting_signatures(work)
    wb = load_workbook(work, data_only=False)
    symbols = wb[SYMBOLS_SHEET]
    headers = _instrument_averages_header_map(symbols)
    trades_coordinate = f"{get_column_letter(headers['Trades'])}4"
    expected_trades = symbols[trades_coordinate].value
    symbols[trades_coordinate].value = -987654
    manual_side = Side(style="double", color="FFFF0000")
    symbols["AJ1"].border = _border_with_test_sides(symbols["AJ1"].border, left=manual_side)
    symbols["AT4"].border = _border_with_test_sides(symbols["AT4"].border, right=manual_side)
    wb.save(work)
    wb.close()
    _open_save_reopen_cycle(work)

    wb = load_workbook(work, data_only=False)
    expected_aj1_border = _border_signature(wb[SYMBOLS_SHEET]["AJ1"])
    expected_at4_border = _border_signature(wb[SYMBOLS_SHEET]["AT4"])
    wb.close()

    for attempt in range(2):
        source_snapshot = read_master_journal_source(work)
        result = update_master_journal_workbook_data_only(work, source_snapshot)
        assert result["ok"] is True
        Path(result["candidate_path"]).replace(work)

        wb = load_workbook(work, data_only=False)
        symbols = wb[SYMBOLS_SHEET]
        assert _border_signature(symbols["AJ1"]) == expected_aj1_border
        assert _border_signature(symbols["AT4"]) == expected_at4_border
        wb.close()
        assert _symbols_cell_value(work, trades_coordinate) == expected_trades
        after = _symbols_trade_log_formatting_signatures(work)
        for key in (
            "trade_log_borders",
            "trade_log_fills",
            "trade_log_header_merges",
            "trade_log_widths",
            "trade_log_row_heights",
        ):
            assert after[key] == before[key], f"Trade Log {key} drifted on resync {attempt + 1}"

    _open_save_reopen_cycle(work)
    wb = load_workbook(work, data_only=False)
    symbols = wb[SYMBOLS_SHEET]
    assert _border_signature(symbols["AJ1"]) == expected_aj1_border
    assert _border_signature(symbols["AT4"]) == expected_at4_border
    wb.close()
    assert _symbols_cell_value(work, trades_coordinate) == expected_trades
    _assert_symbols_trade_log_layout_contract(work, duration_border_exceptions={"AT4"})
    assert _formula_error_cells(work) == []


def test_data_only_update_repairs_stale_recommendation_columns_and_stats1_labels(tmp_path: Path):
    out = tmp_path / "stale_recommendations.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)

    wb = load_workbook(out)
    trade_log = wb["Trade Log"]
    current_header_map = _trade_log_header_map(trade_log)
    row_snapshots = []
    for row in range(TRADE_LOG_DATA_START_ROW, trade_log.max_row + 1):
        if not any(trade_log.cell(row, col).value not in (None, "") for col in current_header_map.values()):
            continue
        row_snapshots.append({
            header: trade_log.cell(row, col).value
            for header, col in current_header_map.items()
            if header in TRADE_LOG_HEADERS
        })

    for merged in list(trade_log.merged_cells.ranges):
        if merged.min_row <= 3 and merged.max_row >= 1:
            trade_log.unmerge_cells(str(merged))
    row1, row2, row3 = _trade_log_three_row_header_values_for(RECOMMENDATION_TRADE_LOG_HEADERS)
    for col in range(1, max(trade_log.max_column, len(RECOMMENDATION_TRADE_LOG_HEADERS)) + 1):
        for row in range(1, trade_log.max_row + 1):
            trade_log.cell(row, col).value = None
    for col, header in enumerate(RECOMMENDATION_TRADE_LOG_HEADERS, start=1):
        trade_log.cell(1, col).value = row1[col - 1]
        trade_log.cell(2, col).value = row2[col - 1]
        trade_log.cell(3, col).value = row3[col - 1]
        if header not in MOVE_TO_FIELD_MAP:
            trade_log.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
        else:
            trade_log.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    for group_label, group_headers in (
        ("Move to Break-Even", list(MOVE_TO_FIELD_MAP)[:5]),
        ("Move to Profit", list(MOVE_TO_FIELD_MAP)[5:]),
    ):
        cols = [RECOMMENDATION_TRADE_LOG_HEADERS.index(header) + 1 for header in group_headers]
        trade_log.merge_cells(start_row=1, start_column=min(cols), end_row=1, end_column=max(cols))
        trade_log.cell(1, min(cols)).value = group_label
    for offset, snapshot in enumerate(row_snapshots):
        row = TRADE_LOG_DATA_START_ROW + offset
        for col, header in enumerate(RECOMMENDATION_TRADE_LOG_HEADERS, start=1):
            if header == STOP_RECOMMENDATION_HEADER:
                trade_log.cell(row, col).value = "Reduce stop loss"
            elif header == TARGET_RECOMMENDATION_HEADER:
                trade_log.cell(row, col).value = "Reduce target"
            else:
                trade_log.cell(row, col).value = snapshot.get(header)

    stats1 = wb[STATS1_SHEET]
    max_target_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Max target %")
    stats1.cell(max_target_row, 1).value = None
    stats1.cell(max_target_row + 1, 1).value = None
    stats1.cell(max_target_row + 2, 1).value = None
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)

    checked = load_workbook(out)
    try:
        trade_log = checked["Trade Log"]
        assert _trade_log_header_map(trade_log) == {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
        assert trade_log.max_column == len(TRADE_LOG_HEADERS)
        assert trade_log.cell(1, len(TRADE_LOG_HEADERS)).value == "Row ID"
        assert [
            trade_log.cell(1, col).value
            for col in range(1, trade_log.max_column + 1)
            if str(trade_log.cell(1, col).value or "").strip() == "Recommendation"
        ] == []

        stats1 = checked[STATS1_SHEET]
        first_winners_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Winners")
        max_target_row = next(
            row
            for row in range(1, first_winners_row)
            if stats1.cell(row, 1).value == "Max target %"
        )
        assert stats1.cell(max_target_row + 1, 1).value == "Source"
        assert stats1.cell(max_target_row + 2, 1).value == "Recommendation"
        assert all(stats1.cell(max_target_row + 2, col).value for col in range(2, 5))
    finally:
        checked.close()


def test_trade_log_row_heights_are_normalized_on_data_only_update(tmp_path: Path):
    path = tmp_path / "trade-log-row-heights.xlsx"
    snapshot = sample_snapshot()
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    trade_log = wb["Trade Log"]
    trade_log.row_dimensions[TRADE_LOG_DATA_START_ROW].height = 45
    trade_log.row_dimensions[TRADE_LOG_DATA_START_ROW + 1].height = 30
    trade_log.row_dimensions[23].height = 60
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    checked = load_workbook(path)
    try:
        trade_log = checked["Trade Log"]
        assert trade_log.row_dimensions[TRADE_LOG_DATA_START_ROW].height == 15
        assert trade_log.row_dimensions[TRADE_LOG_DATA_START_ROW + 1].height == 15
        assert trade_log.row_dimensions[23].height == 15
    finally:
        checked.close()


def test_trade_log_new_schema_distances_and_header_aware_update(tmp_path: Path):
    out = tmp_path / "new_schema.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    assert _trade_log_header_map(ws) == {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    stop_distance = ws.cell(_trade_data_row(), _header_col(ws, "Stop Loss Distance"))
    target_distance = ws.cell(_trade_data_row(), _header_col(ws, "Target Distance"))
    assert stop_distance.number_format == "0.00%"
    assert target_distance.number_format == "0.00%"
    assert stop_distance.value == pytest.approx(abs(1.09 - 1.1) / 1.1)
    assert target_distance.value == pytest.approx(abs(1.12 - 1.1) / 1.1)
    assert ws.cell(1, _header_col(ws, "Target Price")).value == "Target Price"
    assert ws.cell(1, _header_col(ws, "Commission")).value == "Commission"
    assert ws.cell(1, _header_col(ws, "Net P/L")).value == "Net P/L"
    assert ws.cell(1, _header_col(ws, "Profit %")).value == "Profit %"
    assert ws.cell(1, _header_col(ws, "Balance After")).value == "Balance After"
    wb.save(out); wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    updated = load_workbook(out)
    ws2 = updated["Trade Log"]
    assert _trade_log_header_map(ws2) == {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    assert ws2.cell(_trade_data_row(), _header_col(ws2, "Commission")).value in (None, "")
    assert ws2.cell(_trade_data_row(), _header_col(ws2, "Net P/L")).value == 120.5
    assert ws2.cell(_trade_data_row(), _header_col(ws2, "Profit %")).value == pytest.approx(0.023)
    assert ws2.cell(_trade_data_row(), _header_col(ws2, "R-Multiple")).value == 1.2
    assert ws2.cell(_trade_data_row(), _header_col(ws2, "Balance After")).value == 1000
    updated.close()


def test_dashboard_market_risk_cells_and_grey_no_metric_cells(tmp_path: Path):
    from openpyxl import Workbook
    out = tmp_path / "dashboard_market.xlsx"
    wb = Workbook()
    dash = wb.active
    dash.title = "Dashboard"
    wb.create_sheet("Trade Log")
    wb.create_sheet("Instrument Averages")
    wb.create_sheet("P&L Calendar")
    dash["B1"] = "Overall"; dash["C1"] = "FX"; dash["D1"] = "Crypto"
    dash["A33"] = "Winners"; dash["A34"] = "Avg result %"; dash["A35"] = "Avg stop %"; dash["A36"] = "Avg target %"; dash["A37"] = "Avg R"
    dash["A38"] = "Losers"; dash["A39"] = "Avg result %"; dash["A40"] = "Avg stop %"; dash["A41"] = "Avg target %"; dash["A42"] = "Avg R"
    dash["A43"] = "Drawdown"; dash["A44"] = "Max drawdown"; dash["A45"] = "Avg drawdown"
    dash["F1"] = "Instrument leaders"; dash["J1"] = "Account Balances"
    dash["J2"] = "Account"; dash["K2"] = "Balance"; dash["L2"] = "Currency"
    dash["D35"].fill = PatternFill("solid", fgColor="EAF2F8")
    _ensure_trade_log_headers(wb)
    wb.save(out); wb.close()

    snap = {
        "items": [],
        "stats": {"totals": {}, "groups": {
            "by_market": {
                "overall": {"max_drawdown_pct": 9.0, "avg_drawdown_pct": 4.5},
                "fx": {"max_drawdown_pct": 10.0, "avg_drawdown_pct": 5.0},
                "crypto": {"max_drawdown_pct": 12.0, "avg_drawdown_pct": 6.0},
            },
            "leaders": {}, "duration": {},
            "risk_expectancy": {
                "avg_result_pct_winners": 1.0,
                "avg_stop_pct_winners": 1.1,
                "avg_target_pct_winners": 2.0,
                "avg_r_multiple_winners": 1.25,
                "avg_result_pct_losers": -3.0,
                "avg_stop_pct_losers": 3.0,
                "avg_target_pct_losers": 4.0,
                "avg_r_multiple_losers": -0.75,
                "max_drawdown_pct": 9.0,
                "avg_drawdown_pct": 4.5,
                "by_market": {
                    "overall": {"avg_result_pct_winners": 1.0, "avg_stop_pct_winners": 1.1, "avg_target_pct_winners": 2.0, "avg_r_multiple_winners": 1.25, "avg_result_pct_losers": -3.0, "avg_stop_pct_losers": 3.0, "avg_target_pct_losers": 4.0, "avg_r_multiple_losers": -0.75, "max_drawdown_pct": 9.0, "avg_drawdown_pct": 4.5},
                    "fx": {"avg_result_pct_winners": 1.5, "avg_stop_pct_winners": 1.5, "avg_target_pct_winners": 2.5, "avg_r_multiple_winners": 1.5, "avg_result_pct_losers": -3.5, "avg_stop_pct_losers": 3.5, "avg_target_pct_losers": 4.5, "avg_r_multiple_losers": -0.5, "max_drawdown_pct": 10.0, "avg_drawdown_pct": 5.0},
                    "crypto": {"avg_result_pct_winners": 5.5, "avg_stop_pct_winners": 5.5, "avg_target_pct_winners": 6.5, "avg_r_multiple_winners": 2.5, "avg_result_pct_losers": -7.5, "avg_stop_pct_losers": 7.5, "avg_target_pct_losers": 8.5, "avg_r_multiple_losers": -1.5, "max_drawdown_pct": 12.0, "avg_drawdown_pct": 6.0},
                },
            },
        }},
        "balances": [],
    }
    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    updated = load_workbook(out)
    d = updated[STATS1_SHEET]
    section_rows = {
        str(d.cell(row, 1).value or ""): row
        for row in range(1, d.max_row + 1)
        if str(d.cell(row, 1).value or "") in {"Winners", "Losers", "Drawdown"}
    }
    def row_in_section(section: str, label: str) -> int:
        start = section_rows[section] + 1
        end = min(
            [row for name, row in section_rows.items() if row > section_rows[section]] or [d.max_row + 1]
        )
        return next(
            row for row in range(start, end)
            if str(d.cell(row, 1).value or "").strip() == label
        )

    winners_pct = row_in_section("Winners", "Percentage expectancy")
    winners_stop = row_in_section("Winners", "Avg stop %")
    winners_r = row_in_section("Winners", "R expectancy")
    losers_pct = row_in_section("Losers", "Percentage expectancy")
    losers_r = row_in_section("Losers", "R expectancy")
    max_drawdown = row_in_section("Drawdown", "Max drawdown")
    avg_drawdown = row_in_section("Drawdown", "Avg drawdown")
    assert d.cell(winners_pct, 3).value == pytest.approx(0.015)
    assert d.cell(winners_pct, 4).value == pytest.approx(0.055)
    assert d.cell(winners_pct, 3).number_format == "0.00%"
    assert d.cell(winners_pct, 4).number_format == "0.00%"
    assert d.cell(winners_stop, 3).value == pytest.approx(0.015)
    assert d.cell(winners_stop, 4).value in (None, "")
    assert d.cell(winners_stop, 3).number_format == "0.00%"
    assert d.cell(winners_stop, 4).number_format != "0.00%"
    assert d.cell(winners_r, 3).value == pytest.approx(1.5)
    assert d.cell(winners_r, 4).value == pytest.approx(2.5)
    assert d.cell(winners_r, 3).number_format == '0.000"R"'
    assert d.cell(winners_r, 4).number_format == '0.000"R"'
    assert d.cell(losers_pct, 3).value == pytest.approx(-0.035)
    assert d.cell(losers_r, 4).number_format == '0.000"R"'
    assert d.cell(max_drawdown, 3).value == pytest.approx(0.10)
    assert d.cell(max_drawdown, 4).value == pytest.approx(0.12)
    assert d.cell(avg_drawdown, 3).value == pytest.approx(0.05)
    assert d.cell(avg_drawdown, 4).value == pytest.approx(0.06)
    assert d.cell(max_drawdown, 3).number_format == "0.00%"
    assert d.cell(avg_drawdown, 4).number_format == "0.00%"
    updated.close()


def _old_trade_log_headers():
    return list(OLD_TRADE_LOG_HEADERS)


def _minimal_old_schema_workbook(path: Path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = "Overall"; ws["D1"] = "FX"; ws["G1"] = "Crypto"; ws["J1"] = "Winners"; ws["J8"] = "Losers"; ws["J14"] = "Drawdown"; ws["M1"] = "Instrument leaders"; ws["P1"] = "Account Balances"
    ws["P2"] = "Account"; ws["Q2"] = "Balance"; ws["R2"] = "Currency"; ws["P3"] = "BINANCE"
    tl = wb.create_sheet("Trade Log")
    for c, h in enumerate(_old_trade_log_headers(), start=1):
        tl.cell(1, c, h)
    tl.append(["2026-01-01", "2026-01-01", "BINANCE", "BTCUSDT", "BUY", 1, 10, 11, "", "", "", "", "", 1, 0.01, 1, 101, "00:00:01:00", "No", "Old setup", "1H", "No", "Old notes", "", "", "USDT", "trade", "old-row"])
    tl.freeze_panes = "A2"
    tl.auto_filter.ref = "A1:AB2"
    inst = wb.create_sheet("Instrument Averages")
    inst.append(["Symbol", "Trades"])
    inst.freeze_panes = "A2"
    inst.auto_filter.ref = "A1:B1"
    wb.create_sheet("P&L Calendar")
    wb.save(path)
    wb.close()


def _validation_for_col(ws, col_letter: str):
    col_idx = coordinate_to_tuple(f"{col_letter}2")[1]
    out = []
    for dv in ws.data_validations.dataValidation:
        if any(rng.min_col <= col_idx <= rng.max_col for rng in dv.cells.ranges):
            out.append(dv)
    return out


def test_trade_log_quality_columns_insert_after_test(tmp_path: Path):
    p = tmp_path / "old_schema.xlsx"
    _minimal_old_schema_workbook(p)
    snap = {"items": [{"id": "old-row", "row_type": "trade", "account": "BINANCE", "symbol": "BTCUSDT", "side": "BUY", "open_time": "2026-01-01", "close_time": "2026-01-01", "net_profit": 1}], "stats": {"totals": {}, "groups": {"by_market": {}, "risk_expectancy": {}, "leaders": {}}}, "balances": [{"account_label": "BINANCE", "balance": 0, "currency": "USDT"}]}
    res = update_master_journal_workbook_data_only(p, snap)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(p)
    wb = load_workbook(p)
    assert wb.sheetnames[:len(SHEET_ORDER)] == SHEET_ORDER
    assert "Dashboard" not in wb.sheetnames
    assert "Instrument Averages" not in wb.sheetnames
    ws = wb["Trade Log"]
    headers = list(_trade_log_header_map(ws))
    assert headers == TRADE_LOG_HEADERS
    wb.close()
    duration_col = _header_col(ws, "Trade Duration (DD:HH:MM:SS)")
    assert headers[duration_col:duration_col + 10] == list(MOVE_TO_FIELD_MAP)
    quality_start = _header_col(ws, "Test") - 1
    assert headers[quality_start:quality_start + 7] == [
        "Test", "Pattern", "EMA", "VWAP", "ATHS/ATLS", "Order", "Round Number",
    ]
    assert "Close" not in headers
    assert _header_col(ws, "Close Stopout") > _header_col(ws, "Pattern")
    row_id_letter = get_column_letter(_header_col(ws, "Row ID"))
    assert ws.cell(1, _header_col(ws, "Row ID")).value == "Row ID"
    assert ws.column_dimensions[row_id_letter].hidden is True
    assert ws.auto_filter.ref == f"A3:{get_column_letter(len(TRADE_LOG_HEADERS))}{ws.max_row}"

def test_trade_log_quality_dropdowns(tmp_path: Path):
    p = tmp_path / "quality_dropdowns.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    ws = load_workbook(p)["Trade Log"]
    def validations(header):
        return _validation_for_col(ws, get_column_letter(_header_col(ws, header)))
    assert any(dv.formula1 == '"Yes,No"' and dv.allow_blank for dv in validations("Test"))
    assert any(dv.formula1 == '"range,channel"' and dv.allow_blank for dv in validations("Pattern"))
    assert any(dv.formula1 == '"All-Time High,All-Time Low"' and dv.allow_blank for dv in validations("ATHS/ATLS"))
    assert any(dv.formula1 == '"Market,Limit"' and dv.allow_blank for dv in validations("Order"))
    for header in ["VWAP", "Round Number", "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close"]:
        assert any(dv.formula1 == '"Yes,No"' and dv.allow_blank for dv in validations(header))
    assert validations("EMA") == []

def test_trade_log_quality_manual_values_survive_resync(tmp_path: Path):
    from datetime import datetime

    p = tmp_path / "quality_resync.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, p)
    wb = load_workbook(p)
    ws = wb["Trade Log"]
    vals = {
        "Trade Number": "C42",
        "Pattern": "legacy-manual",
        "EMA": "20/50",
        "VWAP": "Yes",
        "ATHS/ATLS": "All-Time High",
        "Order": "Limit",
        "Round Number": "Yes",
        "Spiked Out": "No",
        "Close Stopout": "No",
        "Near Perfect Entry": "Yes",
        "Near Win": "No",
        "Early Close": "Yes",
        "Move to Break Even Time": datetime(2026, 1, 1, 10, 0),
        "Move to Break Even Trigger Price": 123.45,
        "Move to Break Even Distance From Entry %": 0.0125,
        "Move to Profit Trigger Price": 125.0,
    }
    for header, value in vals.items():
        ws.cell(_trade_data_row(), _header_col(ws, header), value)
    row_id = ws.cell(_trade_data_row(), _header_col(ws, "Row ID")).value
    wb.save(p); wb.close()
    res = update_master_journal_workbook_data_only(p, snap)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(p)
    ws = load_workbook(p)["Trade Log"]
    rid_col = _header_col(ws, "Row ID")
    row_num = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, rid_col).value == row_id)
    for header, value in vals.items():
        assert ws.cell(row_num, _header_col(ws, header)).value == value
    assert ws.cell(row_num, _header_col(ws, "Trade Number")).number_format == "@"
    assert ws.cell(row_num, _header_col(ws, "Move to Break Even Time")).number_format == "yyyy-mm-dd hh:mm:ss"
    assert ws.cell(row_num, _header_col(ws, "Move to Break Even Distance From Entry %")).number_format == "0.00%"


def test_read_master_journal_manual_overrides_reads_quality_columns(tmp_path: Path):
    p = tmp_path / "manual_quality.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    ws = wb["Trade Log"]
    ws.cell(_trade_data_row(), _header_col(ws, "Pattern"), "Pullback")
    ws.cell(_trade_data_row(), _header_col(ws, "EMA"), "9")
    ws.cell(_trade_data_row(), _header_col(ws, "VWAP"), "No")
    ws.cell(_trade_data_row(), _header_col(ws, "ATHS/ATLS"), "All-Time Low")
    rid = ws.cell(_trade_data_row(), _header_col(ws, "Row ID")).value
    wb.save(p); wb.close()
    ov = read_master_journal_manual_overrides(p)
    assert ov[rid]["pattern"] == "Pullback"
    assert ov[rid]["ema"] == "9"
    assert ov[rid]["vwap"] == "No"
    assert ov[rid]["aths_atls"] == "All-Time Low"


def test_read_master_journal_source_reads_quality_columns(tmp_path: Path):
    p = tmp_path / "source_quality.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    ws = wb["Trade Log"]
    ws.cell(_trade_data_row(), _header_col(ws, "Pattern"), "Range")
    ws.cell(_trade_data_row(), _header_col(ws, "VWAP"), "Yes")
    ws.cell(_trade_data_row(), _header_col(ws, "Order"), "Market")
    ws.cell(_trade_data_row(), _header_col(ws, "Near Win"), "Yes")
    rid = ws.cell(_trade_data_row(), _header_col(ws, "Row ID")).value
    wb.save(p); wb.close()
    item = next(r for r in read_master_journal_source(p)["items"] if r["id"] == rid)
    assert item["pattern"] == "Range"
    assert item["vwap"] == "Yes"
    assert item["order_type"] == "Market"
    assert item["near_win"] == "Yes"


def test_update_data_only_writes_dashboard_horizontal_core_metric_aliases(tmp_path: Path):
    from openpyxl import Workbook
    p = tmp_path / "horizontal.xlsx"
    wb = Workbook(); ws = wb.active; ws.title = "Dashboard"
    wb.create_sheet("Trade Log"); wb.create_sheet("Instrument Averages"); wb.create_sheet("P&L Calendar")
    ws["B1"] = "Overall"; ws["C1"] = "FX"; ws["D1"] = "Crypto"
    ws["A11"] = "Best Win Streak"; ws["A12"] = "Worst Losing Streak"; ws["A15"] = "Avg stop %"; ws["A16"] = "Min stop %"; ws["A17"] = "Max stop %"; ws["A18"] = "Avg target %"; ws["A19"] = "Min target %"; ws["A20"] = "Max target %"
    ws["F1"] = "Winners"; ws["F8"] = "Losers"; ws["F14"] = "Drawdown"; ws["I1"] = "Instrument leaders"; ws["L1"] = "Account Balances"; ws["L2"] = "Account"; ws["M2"] = "Balance"; ws["N2"] = "Currency"; ws["L3"] = "BINANCE"
    _ensure_trade_log_headers(wb)
    for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + 2):
        wb["Trade Log"].row_dimensions[row].height = 15
    wb.save(p); wb.close()
    snap = {"stats": {"totals": {}, "groups": {"by_market": {"overall": {"winning_streak": 4, "losing_streak": 3, "max_stop_pct": 8.5, "max_target_pct": 9.5, STOP_RECOMMENDATION_HEADER: "Reduce stop loss", TARGET_RECOMMENDATION_HEADER: "Increase target — Recommended: 3.0R"}, "fx": {"winning_streak": 2, "losing_streak": 1, "max_stop_pct": 4.0, "max_target_pct": 5.0, STOP_RECOMMENDATION_HEADER: "Keep stop loss", TARGET_RECOMMENDATION_HEADER: "Keep target — Recommended: 2.5R"}, "crypto": {"winning_streak": 6, "losing_streak": 7, "max_stop_pct": 10.75, "max_target_pct": 12.25, STOP_RECOMMENDATION_HEADER: "Increase stop loss", TARGET_RECOMMENDATION_HEADER: "Reduce target — Recommended: 2.0R"}}, "risk_expectancy": {}, "leaders": {}, "duration": {}}}, "balances": [{"account_label": "BINANCE", "balance": 0, "currency": "USDT"}]}
    res = update_master_journal_workbook_data_only(p, snap)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(p)
    ws = load_workbook(p)[STATS1_SHEET]
    assert [ws.cell(11, c).value for c in range(2, 5)] == [4, 2, 6]
    assert [ws.cell(14, c).value for c in range(2, 5)] == [3, 1, 7]
    max_target_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Max target %")
    assert [ws.cell(max_target_row, c).value for c in range(2, 5)] == [pytest.approx(0.095), pytest.approx(0.05), pytest.approx(0.1225)]
    assert [ws.cell(max_target_row, c).number_format for c in range(2, 5)] == ["0.00%", "0.00%", "0.00%"]
    max_stop_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Max stop %")
    stop_recommendation_row = max_stop_row + 2
    target_recommendation_row = max_target_row + 2
    assert ws.cell(max_stop_row + 1, 1).value == "Source"
    assert ws.cell(stop_recommendation_row, 1).value == "Recommendation"
    assert [ws.cell(stop_recommendation_row, c).value for c in range(2, 5)] == ["Need wins & losses", "Need wins & losses", "Need wins & losses"]
    assert ws.cell(max_target_row + 1, 1).value == "Source"
    assert ws.cell(target_recommendation_row, 1).value == "Recommendation"
    assert [ws.cell(target_recommendation_row, c).value for c in range(2, 5)] == [
        "Increase target — Recommended: 3.0R",
        "Increase target \u2014 Recommended: 2.5R",
        "Decrease target \u2014 Recommended: 2.0R",
    ]
    assert all(_cell_fill_rgb(ws.cell(11, col)) == "C6EFCE" for col in range(2, 5))
    assert all(_cell_font_rgb(ws.cell(11, col)) == "006100" for col in range(2, 5))
    assert all(_cell_fill_rgb(ws.cell(14, col)) == "FFC7CE" for col in range(2, 5))
    assert all(_cell_font_rgb(ws.cell(14, col)) == "9C0006" for col in range(2, 5))


def test_stats1_manual_a_column_merges_survive_data_only_update(tmp_path: Path):
    path = tmp_path / "manual-merge.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    stats1.merge_cells("A52:A53")
    stats1["A52"] = "Manual merged note"
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        assert "A52:A53" in {str(rng) for rng in out[STATS1_SHEET].merged_cells.ranges}
        assert out[STATS1_SHEET]["A52"].value == "Manual merged note"
    finally:
        out.close()


def test_stats1_manual_merges_prevent_optional_winners_row_insertion(tmp_path: Path):
    path = tmp_path / "manual-merge-missing-optional.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    winners_row = next(row for row in range(1, stats1.max_row + 1) if stats1.cell(row, 1).value == "Winners")
    stats1.cell(winners_row + 1, 1).value = "Avg stop %"
    stats1.merge_cells("A52:A53")
    stats1["A52"] = "Manual merged note"
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snap)
    assert result["ok"] is True
    assert "Winners: Min stop %" in result["diagnostics"].get("skipped_dashboard_metric_rows_due_to_manual_merges", [])
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        ws = out[STATS1_SHEET]
        assert "A52:A53" in {str(rng) for rng in ws.merged_cells.ranges}
        assert ws["A52"].value == "Manual merged note"
    finally:
        out.close()


def test_generated_stats1_metric_source_merges_are_cleaned_without_touching_manual_notes(tmp_path: Path):
    path = tmp_path / "merged-source-pairs.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    rows = defaultdict(list)
    for row in range(1, stats1.max_row + 1):
        rows[str(stats1.cell(row, 1).value or "")].append(row)
    source_metric_row = rows["Min stop %"][0]
    assert stats1.cell(source_metric_row + 1, 1).value == "Source"
    stats1.merge_cells(start_row=source_metric_row, start_column=1, end_row=source_metric_row + 1, end_column=1)
    stats1.cell(source_metric_row, 1).value = "Min stop %"
    stats1.cell(source_metric_row + 1, 2).value = "EURUSD source detail"
    stats1.merge_cells("A52:A53")
    stats1["A52"] = "Manual merged note"
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        ws = out[STATS1_SHEET]
        merged_ranges = {str(rng) for rng in ws.merged_cells.ranges}
        assert f"A{source_metric_row}:A{source_metric_row + 1}" not in merged_ranges
        assert ws.cell(source_metric_row + 1, 1).value == "Source"
        assert ws.cell(source_metric_row + 1, 2).value not in (None, "")
        assert "A52:A53" in merged_ranges
        assert ws["A52"].value == "Manual merged note"
    finally:
        out.close()


def test_drawdown_detail_rows_are_split_for_stats1_and_reports(tmp_path: Path):
    snap = sample_snapshot()
    snap["items"].append({
        "id": "drawdown-detail-balance-point",
        "row_type": "trade",
        "account": "OANDA DEMO",
        "asset_class": "fx",
        "symbol": "EURUSD",
        "side": "SELL",
        "open_time": "2026-05-02T01:00:00Z",
        "close_time": "2026-05-02T02:00:00Z",
        "net_profit": -20.0,
        "result_pct": -0.2,
        "r_multiple": -0.2,
        "analysis_balance_after_trade": 980.0,
        "trade_duration_seconds": 3600,
    })
    detail = {"start_time": "2026-05-01T01:00:00Z", "end_time": "2026-05-02T02:00:00Z"}
    by_market = snap["stats"]["groups"]["by_market"]
    for market in ("overall", "fx", "crypto"):
        by_market.setdefault(market, {}).update({
            "min_drawdown_pct": 0.75,
            "avg_drawdown_pct": 1.0,
            "max_drawdown_pct": 1.25,
            "min_drawdown_detail": detail,
            "max_drawdown_detail": detail,
        })
    path = tmp_path / "drawdown-detail.xlsx"
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    try:
        for sheet_name, header in ((STATS1_SHEET, "Overall"), (REPORT_YEARLY_SHEET, 2026), ("2026", "May")):
            ws = wb[sheet_name]
            col = 2 if sheet_name == STATS1_SHEET else next(c for c in range(2, ws.max_column + 1) if ws.cell(1, c).value == header)
            rows = defaultdict(list)
            for row in range(1, ws.max_row + 1):
                rows[str(ws.cell(row, 1).value or "")].append(row)
            for label in ("Min drawdown", "Max drawdown"):
                row = rows[label][0]
                assert ws.cell(row, col).number_format == "0.00%"
                assert isinstance(ws.cell(row, col).value, (int, float))
                assert ws.cell(row + 1, 1).value == "Start"
                assert ws.cell(row + 2, 1).value == "End"
                assert str(ws.cell(row + 1, col).value).startswith("2026-05-")
                assert str(ws.cell(row + 2, col).value).startswith("2026-05-")
    finally:
        wb.close()


def test_data_only_update_splits_existing_inline_drawdown_detail(tmp_path: Path):
    path = tmp_path / "inline-drawdown.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    stats1 = wb[STATS1_SHEET]
    rows = {str(stats1.cell(row, 1).value or ""): row for row in range(1, stats1.max_row + 1)}
    min_row = rows["Min drawdown"]
    stats1.cell(min_row, 2).value = "0.00861227% (2020-04-12 16:42:00 to 2020-04-13 19:21:00)"
    stats1.cell(min_row + 1, 2).value = None
    stats1.cell(min_row + 2, 2).value = None
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        ws = out[STATS1_SHEET]
        min_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Min drawdown")
        assert ws.cell(min_row, 2).value == pytest.approx(0.0000861227)
        assert ws.cell(min_row, 2).number_format == "0.00%"
        assert ws.cell(min_row + 1, 1).value == "Start"
        assert ws.cell(min_row + 1, 2).value == "2020-04-12 16:42:00"
        assert ws.cell(min_row + 2, 1).value == "End"
        assert ws.cell(min_row + 2, 2).value == "2020-04-13 19:21:00"
    finally:
        out.close()


def test_stats2_net_pl_percentage_update_resets_stale_borders(tmp_path: Path):
    path = tmp_path / "stats2-net-pl.xlsx"
    snap = sample_snapshot()
    snap["balances"].append({"account_label": "OANDA DEMO", "balance": 980.0, "currency": "AUD"})
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    stats2 = wb[STATS2_SHEET]
    headers = {str(stats2.cell(2, col).value or ""): col for col in range(1, stats2.max_column + 1)}
    net_col = headers["Net P/L Percentage"]
    account_col = headers["Account"]
    stats2.cell(2, headers["As Of"]).value = "Risk Of Ruin"
    row = next(r for r in range(3, stats2.max_row + 1) if stats2.cell(r, account_col).value == "OANDA DEMO")
    custom_border = Border(left=Side(style="thick", color="FF123456"), right=Side(style="double", color="FF654321"))
    stats2.cell(row, net_col).border = custom_border
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        ws = out[STATS2_SHEET]
        final_headers = [str(ws.cell(2, col).value or "") for col in range(1, ws.max_column + 1)]
        assert sum(header.casefold() == "risk of ruin" for header in final_headers) == 1
        assert final_headers[headers["As Of"] - 1] == "As Of"
        assert final_headers[net_col - 1] == "Net P/L Percentage"
        cell = ws.cell(row, net_col)
        assert cell.value == pytest.approx(0.023)
        assert cell.number_format == "0.00%"
        assert _cell_fill_rgb(cell) == ""
        assert _cell_font_rgb(cell) == "000000"
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "medium"
    finally:
        out.close()


def test_report_period_rows_populate_core_counts_duration_and_drawdown(tmp_path: Path):
    snap = sample_snapshot()
    snap["period_reports"] = {}
    snap["items"].append({
        "id": "drawdown-row", "row_type": "trade", "account": "OANDA DEMO", "symbol": "EURUSD",
        "asset_class": "fx", "side": "SELL", "open_time": "2026-05-03T00:00:00Z",
        "close_time": "2026-05-03T00:02:00Z", "result_pct": -0.2,
        "net_profit": -20.0, "trade_duration_seconds": 7215,
        "analysis_balance_after_trade": 980.0,
    })
    snap["items"].append({
        "id": "test-row", "row_type": "trade", "account": "OANDA DEMO", "symbol": "AUDUSD",
        "asset_class": "fx", "side": "BUY", "open_time": "2026-05-04T00:00:00Z",
        "close_time": "2026-05-04T00:01:00Z", "result_pct": 99.0,
        "net_profit": 99.0, "trade_duration_seconds": 60, "is_test_trade": True,
    })
    path = tmp_path / "period-fallback.xlsx"
    build_master_journal_workbook(snap, path)
    wb = load_workbook(path)
    try:
        for sheet_name, header in ((REPORT_YEARLY_SHEET, 2026), ("2026", "May")):
            ws = wb[sheet_name]
            rows = {str(ws.cell(row, 1).value or ""): row for row in range(1, ws.max_row + 1)}
            col = next(c for c in range(2, ws.max_column + 1) if ws.cell(1, c).value == header)
            assert ws.cell(rows["Trades"], col).value == 3
            assert ws.cell(rows["Wins"], col).value == 1
            assert ws.cell(rows["Losses"], col).value == 2
            assert ws.cell(rows["Break-even"], col).value == 0
            assert ws.cell(rows["Test"], col).value == 1
            assert ws.cell(rows["Win rate"], col).value == pytest.approx(1 / 3)
            assert _parse_duration_text(ws.cell(rows["Shortest (DD:HH:MM:SS)"], col).value) is not None
            assert _parse_duration_text(ws.cell(rows["Winners min duration"], col).value) == 3700
            assert _parse_duration_text(ws.cell(rows["Winners avg duration"], col).value) == 3700
            assert _parse_duration_text(ws.cell(rows["Winners max duration"], col).value) == 3700
            assert _parse_duration_text(ws.cell(rows["Losers min duration"], col).value) == 7215
            assert _parse_duration_text(ws.cell(rows["Losers avg duration"], col).value) == 7215
            assert _parse_duration_text(ws.cell(rows["Losers max duration"], col).value) == 7215
            assert ws.cell(rows["Max drawdown"], col).value not in (None, "")
            assert ws.cell(rows["Max drawdown"] + 1, 1).value == "Start"
            assert ws.cell(rows["Max drawdown"] + 1, col).value not in (None, "")
            assert ws.cell(rows["Max drawdown"] + 2, 1).value == "End"
            assert ws.cell(rows["Max drawdown"] + 2, col).value not in (None, "")
    finally:
        wb.close()


def test_update_data_only_repairs_unknown_currency_formats_after_schema_migration(monkeypatch, tmp_path: Path):
    from tools import master_journal_workbook as mjw

    p = tmp_path / "old_schema_unknown_currency.xlsx"
    _minimal_old_schema_workbook(p)
    snap = {
        "items": [{
            "id": "old-row",
            "row_type": "trade",
            "account": "OANDA DEMO",
            "symbol": "EURUSD",
            "side": "BUY",
            "open_time": "2026-01-01",
            "close_time": "2026-01-01",
            "commission": 1.25,
            "net_profit": 10.0,
            "result_pct": 1.0,
        }],
        "stats": {"totals": {}, "groups": {"by_market": {}, "risk_expectancy": {}, "leaders": {}}},
        "balances": [{"account_label": "BINANCE", "balance": 0, "currency": "USDT"}],
    }
    real_build = mjw.build_master_journal_workbook

    def build_with_unknown_formats(snapshot, output_path):
        result = real_build(snapshot, output_path)
        wb = load_workbook(output_path)
        ws = wb["Trade Log"]
        ws.cell(_trade_data_row(), _header_col(ws, "Commission")).number_format = '#,##0.00 "UNKNOWN"'
        ws.cell(_trade_data_row(), _header_col(ws, "Net P/L")).number_format = '#,##0.00 "UNKNOWN"'
        wb.save(output_path)
        wb.close()
        return result

    monkeypatch.setattr(mjw, "build_master_journal_workbook", build_with_unknown_formats)
    res = mjw.update_master_journal_workbook_data_only(p, snap)
    assert res["ok"] is True
    Path(res["candidate_path"]).replace(p)
    ws = load_workbook(p)["Trade Log"]
    assert ws.cell(1, _header_col(ws, "Commission")).value == "Commission"
    assert ws.cell(1, _header_col(ws, "Net P/L")).value == "Net P/L"
    commission_format = str(ws.cell(_trade_data_row(), _header_col(ws, "Commission")).number_format or "")
    net_pl_format = str(ws.cell(_trade_data_row(), _header_col(ws, "Net P/L")).number_format or "")
    assert "UNKNOWN" not in commission_format
    assert "UNKNOWN" not in net_pl_format
    assert "AUD" in commission_format
    assert "AUD" in net_pl_format


def _cf_rule_details(ws):
    details = []
    for key, rules in ws.conditional_formatting._cf_rules.items():
        sqref = str(key.sqref)
        for rule in rules:
            formula = getattr(rule, "formula", None) or []
            dxf = getattr(rule, "dxf", None)
            fill = getattr(getattr(dxf, "fill", None), "fgColor", None)
            details.append((sqref, [str(f) for f in formula], (str(getattr(fill, "rgb", "") or "")[-6:].upper() if fill else "")))
    return details


def test_trade_log_win_loss_row_conditional_formatting_uses_current_schema(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    ws = load_workbook(out)["Trade Log"]
    row_type_letter = get_column_letter(_header_col(ws, "Row Type"))
    net_pl_letter = get_column_letter(_header_col(ws, "Net P/L"))
    expected_range = f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}{ws.max_row}"
    row_rules = [d for d in _cf_rule_details(ws) if d[0] == expected_range]
    formulas = {tuple(d[1]) for d in row_rules}
    assert (f'AND(${row_type_letter}4="trade",${net_pl_letter}4>0)',) in formulas
    assert (f'AND(${row_type_letter}4="trade",${net_pl_letter}4<0)',) in formulas
    assert expected_range == f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}5"
    all_rule_text = " ".join(" ".join(formulas) for _range, formulas, _fill in _cf_rule_details(ws))
    assert "$AA" not in all_rule_text
    assert "A2:AB" not in " ".join(_cf_ranges(ws))


def test_trade_log_stale_conditional_formatting_removed_on_update(tmp_path: Path):
    from openpyxl.formatting.rule import FormulaRule

    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    ws.conditional_formatting.add(
        "A2:AB99",
        FormulaRule(formula=['AND($AA2="trade",$N2<0)'], fill=PatternFill("solid", fgColor="FFC7CE")),
    )
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, sample_snapshot())
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    ws = load_workbook(out)["Trade Log"]
    all_ranges = " ".join(_cf_ranges(ws))
    all_formulas = " ".join(" ".join(formula) for _range, formula, _fill in _cf_rule_details(ws))
    assert "A2:AB" not in all_ranges
    assert "$AA" not in all_formulas
    assert f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}5" in all_ranges


def test_conditional_formatting_uses_worksheet_api_not_unbound_class_method():
    src = Path("tools/master_journal_workbook.py").read_text(encoding="utf-8")
    test_src = Path("tests/test_master_journal_workbook.py").read_text(encoding="utf-8")
    forbidden = "ConditionalFormattingList" + ".add"
    assert forbidden not in src
    assert forbidden not in test_src
    assert "max_priority" not in src


def test_dashboard_trade_log_and_pnl_calendar_loss_profit_fills_match(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snapshot = sample_snapshot()
    snapshot["items"][1]["close_time"] = "2026-06-02T02:00:00Z"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    dash = wb[STATS1_SHEET]
    trade = wb["Trade Log"]
    cal = wb["P&L Calendar"]

    dashboard_fills = {fill for _range, _formula, fill in _cf_rule_details(dash) if fill}
    assert "FFC7CE" in dashboard_fills
    assert "C6EFCE" in dashboard_fills

    expected_trade_range = f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}5"
    trade_row_rules = [d for d in _cf_rule_details(trade) if d[0] == expected_trade_range]
    trade_fills = {fill for _range, _formula, fill in trade_row_rules}
    assert trade_fills == {"FFC7CE", "C6EFCE"}

    assert _cf_rule_details(cal) == []
    assert _cell_fill_rgb(cal["F2"]) == "C6EFCE"
    assert _cell_fill_rgb(cal["G2"]) == "FFC7CE"


def test_pnl_calendar_update_removes_duplicate_generated_profit_loss_rules(tmp_path: Path):
    from openpyxl.formatting.rule import CellIsRule

    p = tmp_path / "calendar_stale_cf.xlsx"
    build_master_journal_workbook(sample_snapshot(), p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    cal.conditional_formatting.add(
        "B2:M2",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    cal.conditional_formatting.add(
        "B2:M2",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="0000FF")),
    )
    wb.save(p)
    wb.close()

    snap = sample_snapshot()
    result = update_master_journal_workbook_data_only(p, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    details = [d for d in _cf_rule_details(cal) if d[0] == "B2:M2"]
    assert details == []
    assert "FFFF00" not in {fill for _range, _formula, fill in _cf_rule_details(cal)}
    assert "0000FF" not in {fill for _range, _formula, fill in _cf_rule_details(cal)}
    assert cal["F2"].value == "1.20%, 2 trades"
    assert _cell_fill_rgb(cal["F2"]) == "C6EFCE"
    assert f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}5" in " ".join(_cf_ranges(wb["Trade Log"]))


def test_generated_pnl_calendar_update_removes_stale_profit_loss_rules(tmp_path: Path):
    from openpyxl.formatting.rule import CellIsRule

    p = tmp_path / "generated_calendar_stale_cf.xlsx"
    snapshot = sample_snapshot()
    snapshot["items"][1]["close_time"] = "2026-06-02T02:00:00Z"
    build_master_journal_workbook(snapshot, p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    cal.conditional_formatting.add(
        "B3:M3",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    cal.conditional_formatting.add(
        "B3:M3",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="0000FF")),
    )
    wb.save(p)
    wb.close()

    result = update_master_journal_workbook_data_only(p, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(p)
    wb = load_workbook(p)
    cal = wb["P&L Calendar"]
    details = [d for d in _cf_rule_details(cal) if d[0] == "B3:M3"]
    assert details == []
    assert _cell_fill_rgb(cal["F2"]) == "C6EFCE"
    assert _cell_fill_rgb(cal["G2"]) == "FFC7CE"
    assert f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}5" in " ".join(_cf_ranges(wb["Trade Log"]))


def _trade_log_row_rule_details(ws):
    row_rules = []
    for cf_range, formulas, fill in _cf_rule_details(ws):
        if cf_range.startswith(f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}") and formulas and formulas[0].startswith('AND($'):
            for rules in ws.conditional_formatting._cf_rules.values():
                for rule in rules:
                    rule_formulas = [str(f) for f in (getattr(rule, "formula", None) or [])]
                    if rule_formulas == formulas:
                        row_rules.append((cf_range, formulas, fill, rule))
    return row_rules

def _trade_log_generated_value_fill_ranges(ws):
    out = []
    for key, rules in ws.conditional_formatting._cf_rules.items():
        sqref = str(key.sqref)
        for part in sqref.split():
            try:
                from openpyxl.utils.cell import range_boundaries
                min_col, min_row, max_col, _max_row = range_boundaries(part)
            except ValueError:
                continue
            if min_row >= 2 and 13 <= min_col <= max_col <= 17:
                for rule in rules:
                    formula = " ".join(str(f) for f in (getattr(rule, "formula", None) or []))
                    if getattr(rule, "type", None) == "cellIs" and formula.strip() == "0":
                        out.append((part, getattr(rule, "operator", None)))
    return out

def test_trade_log_row_rules_stop_if_true_and_no_generated_value_fill_overlap(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    row_rules = _trade_log_row_rule_details(ws)
    assert len(row_rules) == 2
    assert {fill for _range, _formulas, fill, _rule in row_rules} == {"C6EFCE", "FFC7CE"}
    assert all(getattr(rule, "stopIfTrue", None) is True for _range, _formulas, _fill, rule in row_rules)
    assert _trade_log_generated_value_fill_ranges(ws) == []

def test_trade_log_update_removes_stale_generated_value_fill_overlap(tmp_path: Path):
    from openpyxl.formatting.rule import CellIsRule

    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    ws.conditional_formatting.add(
        "M2:M99",
        CellIsRule(operator="notEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    ws.conditional_formatting.add(
        "N2:P99",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="FFFF00")),
    )
    ws.conditional_formatting.add(
        "N2:P99",
        CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="0000FF")),
    )
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, sample_snapshot())
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    ws = load_workbook(out)["Trade Log"]
    assert _trade_log_generated_value_fill_ranges(ws) == []
    row_rules = _trade_log_row_rule_details(ws)
    assert len(row_rules) == 2
    assert all(getattr(rule, "stopIfTrue", None) is True for _range, _formulas, _fill, rule in row_rules)

def test_trade_log_saved_xml_row_rules_stop_if_true_without_overlap(tmp_path: Path):
    import zipfile
    import xml.etree.ElementTree as ET

    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    wb = load_workbook(out)
    trade_sheet_index = wb.sheetnames.index("Trade Log") + 1
    wb.close()
    with zipfile.ZipFile(out) as zf:
        xml = zf.read(f"xl/worksheets/sheet{trade_sheet_index}.xml")
    root = ET.fromstring(xml)
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    overlapping_value_refs = []
    row_rules = []
    for cf in root.findall("x:conditionalFormatting", ns):
        sqref = cf.attrib.get("sqref", "")
        rules = cf.findall("x:cfRule", ns)
        formulas = [rule.findtext("x:formula", default="", namespaces=ns) for rule in rules]
        if sqref.startswith(f"A4:{get_column_letter(len(TRADE_LOG_HEADERS))}") and any('="trade"' in formula for formula in formulas):
            row_rules.extend(rules)
        if sqref in {"M3:M4", "N3:P4"}:
            overlapping_value_refs.append(sqref)
    assert len(row_rules) == 2
    assert all(rule.attrib.get("stopIfTrue") == "1" for rule in row_rules)
    assert overlapping_value_refs == []


def _cell_fill_rgb(cell) -> str:
    rgb = str(getattr(cell.fill.fgColor, "rgb", "") or "")
    return rgb[-6:].upper() if getattr(cell.fill, "fill_type", None) == "solid" and rgb else ""

def _cell_font_rgb(cell) -> str:
    color = getattr(cell.font, "color", None)
    rgb = str(getattr(color, "rgb", "") or "")
    return rgb[-6:].upper() if rgb else ""

def _border_signature(cell):
    def _color_signature(color):
        if color is None:
            return None
        if color.type == "rgb":
            return ("rgb", str(color.rgb))
        if color.type == "indexed":
            return ("indexed", color.indexed)
        if color.type == "theme":
            return ("theme", color.theme, color.tint)
        return (color.type, str(getattr(color, color.type, "")))

    def _side_signature(side):
        if side is None:
            return None
        return (side.style, _color_signature(side.color))

    border = cell.border
    return tuple(_side_signature(side) for side in (border.left, border.right, border.top, border.bottom, border.diagonal))

def _cf_fill_intersects(ws, target_range: str, fill_rgb: str) -> bool:
    target_min_col, target_min_row, target_max_col, target_max_row = range_boundaries(target_range)
    for key, rules in ws.conditional_formatting._cf_rules.items():
        for part in str(key.sqref).split():
            min_col, min_row, max_col, max_row = range_boundaries(part)
            if max_col < target_min_col or target_max_col < min_col or max_row < target_min_row or target_max_row < min_row:
                continue
            for rule in rules:
                color = getattr(getattr(getattr(rule, "dxf", None), "fill", None), "fgColor", None)
                rgb = str(getattr(color, "rgb", "") or "")[-6:].upper()
                if rgb == fill_rgb:
                    return True
    return False


def _trade_log_row_by_id(ws, row_id: str) -> int:
    rid_col = _header_col(ws, "Row ID")
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row, rid_col).value or "") == row_id:
            return row
    raise AssertionError(f"missing row id {row_id!r}")


def test_generated_workbook_repairs_stats1_stats2_and_calendar_structure(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][0]["pattern"] = "Channel"
    snapshot["items"][1]["pattern"] = "Range"
    snapshot["items"].append({
        "id": "structure-drawdown-balance-point",
        "row_type": "trade",
        "account": "OANDA DEMO",
        "asset_class": "fx",
        "symbol": "EURUSD",
        "side": "SELL",
        "open_time": "2026-05-02T01:00:00Z",
        "close_time": "2026-05-02T02:00:00Z",
        "net_profit": -20.0,
        "result_pct": -0.2,
        "r_multiple": -0.2,
        "analysis_balance_after_trade": 980.0,
        "trade_duration_seconds": 3600,
    })
    out = tmp_path / "structure-regressions.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    stats1 = wb[STATS1_SHEET]
    stats2 = wb[STATS2_SHEET]
    cal = wb["P&L Calendar"]

    assert stats1.freeze_panes == "B2"
    rows_by_label = {
        str(stats1.cell(row, 1).value or "").strip(): row
        for row in range(1, stats1.max_row + 1)
    }
    assert stats1.cell(rows_by_label["Min drawdown"], 2).value is not None
    assert stats1.cell(rows_by_label["Avg drawdown"], 2).value is not None
    assert stats1.cell(rows_by_label["Max drawdown"], 2).value is not None
    assert stats1.cell(rows_by_label["Most wins"], 2).value
    assert stats1.cell(rows_by_label["Most losses"], 2).value
    for label in ("Most Profitable", "Least Profitable"):
        row = rows_by_label[label]
        for col in (3, 4):
            assert _cell_fill_rgb(stats1.cell(row, col)) not in {PROFIT_FILL, LOSS_FILL}
        assert not _cf_fill_intersects(stats1, f"C{row}:D{row}", PROFIT_FILL)
        assert not _cf_fill_intersects(stats1, f"C{row}:D{row}", LOSS_FILL)

    assert _cf_fill_intersects(stats2, "F3:F12", PROFIT_FILL)
    assert _cf_fill_intersects(stats2, "F3:F12", LOSS_FILL)
    assert cal.freeze_panes == "B2"
    assert cal["A1"].value == "Month"
    assert cal["B1"].value == 2026
    assert [cal.cell(row, 1).value for row in range(2, 14)] == [
        "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December",
    ]
    assert not any(str(cal.cell(row, 1).value or "").endswith(" Trades") for row in range(2, cal.max_row + 1))
    assert isinstance(cal["B6"].value, str) and re.match(r"^-?\d+\.\d{2}%, \d+ trades?$", cal["B6"].value)
    wb.close()


def test_update_normalizes_every_trade_log_data_row_to_row_5_height(tmp_path: Path):
    snapshot = sample_snapshot()
    template = dict(snapshot["items"][0])
    snapshot["items"] = [
        {
            **template,
            "id": f"height-{idx}",
            "open_time": f"2026-05-{(idx % 28) + 1:02d}T00:00:00Z",
            "close_time": f"2026-05-{(idx % 28) + 1:02d}T01:00:00Z",
            "net_profit": 10.0 if idx % 2 else -5.0,
            "result_pct": 1.0 if idx % 2 else -0.5,
        }
        for idx in range(1, 31)
    ]
    path = tmp_path / "trade-log-heights.xlsx"
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    trade_log = wb["Trade Log"]
    trade_log.row_dimensions[5].height = 21.5
    trade_log.row_dimensions[28].height = 60.0
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(path)
    out = load_workbook(path)
    try:
        trade_log = out["Trade Log"]
        source_height = trade_log.row_dimensions[5].height
        assert source_height == pytest.approx(TRADE_LOG_DATA_ROW_HEIGHT)
        for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + len(snapshot["items"])):
            assert trade_log.row_dimensions[row].height == pytest.approx(source_height)
    finally:
        out.close()


def test_trade_log_direct_row_fills_cover_winning_and_losing_rows(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(sample_snapshot(), out)
    ws = load_workbook(out)["Trade Log"]
    winning_row = _trade_log_row_by_id(ws, "t1")
    losing_row = _trade_log_row_by_id(ws, "t2")
    checked_cols = [1, 3, _header_col(ws, "Net P/L"), len(TRADE_LOG_HEADERS)]
    assert [get_column_letter(col) for col in checked_cols] == [
        "A",
        "C",
        get_column_letter(_header_col(ws, "Net P/L")),
        get_column_letter(len(TRADE_LOG_HEADERS)),
    ]
    assert all(_cell_fill_rgb(ws.cell(winning_row, col)) == "C6EFCE" for col in checked_cols)
    assert all(_cell_fill_rgb(ws.cell(losing_row, col)) == "FFC7CE" for col in checked_cols)


def test_trade_log_direct_row_fills_skip_non_trade_and_zero_rows(tmp_path: Path):
    snap = sample_snapshot()
    snap["items"] = [
        {"id": "win", "row_type": "trade", "account": "A", "symbol": "EURUSD", "open_time": "2026-01-01", "close_time": "2026-01-01", "net_profit": 1.0},
        {"id": "zero", "row_type": "trade", "account": "A", "symbol": "EURUSD", "open_time": "2026-01-02", "close_time": "2026-01-02", "net_profit": 0.0},
        {"id": "cash", "row_type": "cashflow", "account": "A", "symbol": "CASH", "open_time": "2026-01-03", "close_time": "2026-01-03", "net_profit": -50.0, "cashflow_amount": -50.0},
        {"id": "reval", "row_type": "monthly_aud_reval", "account": "A", "period_month": "2026-01-31", "net_profit": 25.0, "result_cash": 25.0},
    ]
    out = tmp_path / "Trading Journal.xlsx"
    build_master_journal_workbook(snap, out)
    ws = load_workbook(out)["Trade Log"]
    assert _cell_fill_rgb(ws.cell(_trade_log_row_by_id(ws, "win"), 1)) == "C6EFCE"
    for row_id in ["zero", "cash", "reval"]:
        row = _trade_log_row_by_id(ws, row_id)
        assert all(_cell_fill_rgb(ws.cell(row, col)) == "" for col in [1, 3, _header_col(ws, "Net P/L"), len(TRADE_LOG_HEADERS)])


def test_update_data_only_restores_trade_log_direct_row_fills(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    ws = wb["Trade Log"]
    winning_row = _trade_log_row_by_id(ws, "t1")
    for col in range(1, len(TRADE_LOG_HEADERS) + 1):
        ws.cell(winning_row, col).fill = PatternFill()
    wb.save(out)
    wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    ws = load_workbook(out)["Trade Log"]
    winning_row = _trade_log_row_by_id(ws, "t1")
    assert all(_cell_fill_rgb(ws.cell(winning_row, col)) == "C6EFCE" for col in [1, 3, _header_col(ws, "Net P/L"), len(TRADE_LOG_HEADERS)])


def test_data_only_update_preserves_manual_trade_log_borders(tmp_path: Path):
    out = tmp_path / "Trading Journal.xlsx"
    snap = sample_snapshot()
    build_master_journal_workbook(snap, out)
    wb = load_workbook(out)
    try:
        ws = wb["Trade Log"]
        headers = _trade_log_header_map(ws)
        target_borders = {
            (TRADE_LOG_DATA_START_ROW, headers["Open Time"]): Border(
                left=Side(style="double", color="FF123456"),
                right=Side(style="thick", color="FF654321"),
                top=Side(style="dashed", color="FFABCDEF"),
                bottom=Side(style="dotted", color="FF111111"),
            ),
            (TRADE_LOG_DATA_START_ROW, headers["Net P/L"]): Border(
                left=Side(style="thick", color="FF222222"),
                right=Side(style="double", color="FF333333"),
                top=Side(style="medium", color="FF444444"),
                bottom=Side(style="dashDot", color="FF555555"),
            ),
            (TRADE_LOG_DATA_START_ROW + 1, headers["Symbol"]): Border(
                left=Side(style="slantDashDot", color="FF666666"),
                right=Side(style="dashed", color="FF777777"),
                top=Side(style="thick", color="FF888888"),
                bottom=Side(style="double", color="FF999999"),
            ),
        }
        before = {}
        for (row, col), border in target_borders.items():
            ws.cell(row, col).border = border
            before[(row, col)] = _border_signature(ws.cell(row, col))
        wb.save(out)
    finally:
        wb.close()

    result = update_master_journal_workbook_data_only(out, snap)
    assert result["ok"] is True
    Path(result["candidate_path"]).replace(out)
    wb = load_workbook(out)
    try:
        ws = wb["Trade Log"]
        for (row, col), signature in before.items():
            assert _border_signature(ws.cell(row, col)) == signature
    finally:
        wb.close()


def _dashboard_core_labels(ws):
    winners_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Winners")
    return [str(ws.cell(row, 1).value or "").strip() for row in range(1, winners_row)]

def _swap_dashboard_test_rows(ws, first: int, second: int) -> None:
    for col in range(1, ws.max_column + 1):
        a = ws.cell(first, col)
        b = ws.cell(second, col)
        a_value, b_value = a.value, b.value
        a_style, b_style = copy(a._style), copy(b._style)
        a_alignment, b_alignment = copy(a.alignment), copy(b.alignment)
        a.value, b.value = b_value, a_value
        a._style, b._style = b_style, a_style
        a.alignment, b.alignment = b_alignment, a_alignment

def test_generated_dashboard_layout_percentages_semantic_fills_and_labels(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["balances"].extend([
        {"account_label": "Bybit Demo", "balance": 25.0, "currency": "USDT"},
        {"account_label": "OANDA DEMO", "balance": 900.0, "currency": "AUD"},
    ])
    out = tmp_path / "generated-dashboard-formatting.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    dash = wb[STATS1_SHEET]
    detail = wb[STATS2_SHEET]

    labels = _dashboard_core_labels(dash)
    assert "Max gain" not in labels
    assert labels[-13:] == [
        "Min duration", "Avg duration", "Max duration",
        "Min Move to Break Even", "Source", "Average Move to Break Even", "Max Move to Break Even", "Source",
        "Min Move to Profit", "Source", "Average Move to Profit", "Max Move to Profit", "Source",
    ]
    assert detail["F2"].value == "Net P/L Percentage"
    assert any(detail.cell(row, 6).value not in (None, "") for row in range(3, 12))
    assert all(detail.cell(row, col).value in (None, "") for row in range(1, 12) for col in range(7, 15))
    assert "BYBIT DEMO" in _dashboard_account_balances(detail)

    assert 0 < dash["C8"].value < 1
    assert -1 <= dash["D8"].value < 0
    assert dash["B9"].value == pytest.approx(0.4)
    assert dash["C9"].value == pytest.approx(1.2)
    assert dash["D9"].value == pytest.approx(-0.8)
    assert dash["C10"].value == pytest.approx(0.023)
    assert dash["C11"].value == pytest.approx(0.0)
    assert dash["D10"].value == pytest.approx(0.0)
    assert dash["D11"].value == pytest.approx(0.011)
    for coordinate in ("C8", "D8", "C10", "D10", "C11", "D11"):
        assert dash[coordinate].number_format == "0.00%"
        assert not any(token in dash[coordinate].number_format.upper() for token in ("AUD", "USDT", "$"))
    for coordinate in ("B9", "C9", "D9"):
        assert dash[coordinate].number_format == '0.000"R"'

    rows_by_label = {str(dash.cell(row, 1).value or "").strip(): row for row in range(1, dash.max_row + 1)}
    green = [
        (rows_by_label[label], col)
        for label in ("Wins", "Gross percent gain", "Gross IR gain")
        for col in (2, 3, 4)
    ]
    red = [
        (rows_by_label[label], col)
        for label in ("Losses", "Gross percent loss", "Gross IR loss")
        for col in (2, 3, 4)
    ]
    assert all(_cell_fill_rgb(dash.cell(row, col)) == "C6EFCE" for row, col in green)
    assert all(_cell_font_rgb(dash.cell(row, col)) == "006100" for row, col in green)
    assert all(_cell_fill_rgb(dash.cell(row, col)) == "FFC7CE" for row, col in red)
    assert all(_cell_font_rgb(dash.cell(row, col)) == "9C0006" for row, col in red)
    assert not _cf_fill_intersects(dash, "B11:D11", "FFC7CE")

    wb.close()

def test_update_repairs_dashboard_order_source_style_max_gain_and_stale_cf(tmp_path: Path):
    from openpyxl.formatting.rule import CellIsRule

    snapshot = sample_snapshot()
    snapshot["balances"].extend([
        {"account_label": "Bybit Demo", "balance": 25.0, "currency": "USDT"},
        {"account_label": "OANDA DEMO", "balance": 900.0, "currency": "AUD"},
    ])
    path = tmp_path / "dashboard-repair.xlsx"
    build_master_journal_workbook(snapshot, path)
    wb = load_workbook(path)
    dash = wb[STATS1_SHEET]
    detail = wb[STATS2_SHEET]

    # Recreate legacy core rows where expectancy sat below streak rows and the
    # displayed Net P/L Percentage was a grey mixed-currency diagnostic.
    _swap_dashboard_test_rows(dash, 14, 16)
    _swap_dashboard_test_rows(dash, 15, 17)
    dash["B8"] = "Unavailable: mixed currencies"
    dash["B8"].fill = PatternFill("solid", fgColor="FFEAF2F8")
    dash["B8"].number_format = "General"

    # Recreate an obsolete generated Max gain pair below the current Dashboard.
    dash["A83"] = "Max gain"
    dash["C83"] = 316.5
    dash["D83"] = 9.75
    dash["C83"].number_format = '#,##0.00 "AUD"'
    dash["D83"].number_format = '#,##0.00 "USDT"'
    dash["A84"] = "Source"
    dash["C84"] = "USDCAD"
    dash["D84"] = "BTCUSDT"
    stale_red = CellIsRule(operator="notEqual", formula=["0"], fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006"))
    dash.conditional_formatting.add("B11:B12", stale_red)
    dash.conditional_formatting.add("C10:D12", copy(stale_red))
    dash.conditional_formatting.add("B48:B74 B81:B82", copy(stale_red))
    dash.conditional_formatting.add(
        "E20",
        CellIsRule(operator="equal", formula=["5"], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    for row in range(3, detail.max_row + 1):
        if detail.cell(row, 1).value == "BYBIT DEMO":
            detail.cell(row, 1).value = "Bybit Demo"
    wb.save(path)
    wb.close()

    result = update_master_journal_workbook_data_only(path, snapshot)
    assert result["ok"] is True
    assert result["diagnostics"]["removed_dashboard_metric_rows"]["Max gain"] == 2
    Path(result["candidate_path"]).replace(path)
    wb = load_workbook(path)
    dash = wb[STATS1_SHEET]
    detail = wb[STATS2_SHEET]
    labels = _dashboard_core_labels(dash)
    assert "Max gain" not in labels
    assert detail["F2"].value == "Net P/L Percentage"
    assert any(detail.cell(row, 6).value not in (None, "") for row in range(3, 12))
    assert all(detail.cell(row, col).value in (None, "") for row in range(1, 12) for col in range(7, 15))
    assert "BYBIT DEMO" in _dashboard_account_balances(detail)
    assert dash["A14"].value == "Percentage expectancy"
    assert dash["A15"].value == "R expectancy"
    assert dash["B8"].value == pytest.approx(0.012)
    assert dash["B8"].number_format == "0.00%"
    rows_by_label = {str(dash.cell(row, 1).value or "").strip(): row for row in range(1, dash.max_row + 1)}
    for label in ("Net P/L Percentage", "Gross percent gain", "Gross percent loss", "Percentage expectancy"):
        assert all(dash.cell(rows_by_label[label], col).number_format == "0.00%" for col in (3, 4))
    for label in ("Net P/L R multiples", "Gross IR gain", "Gross IR loss", "R expectancy"):
        assert all(dash.cell(rows_by_label[label], col).number_format == '0.000"R"' for col in (3, 4))
    for label in ("Gross percent gain", "Gross IR gain"):
        assert all(_cell_fill_rgb(dash.cell(rows_by_label[label], col)) == "C6EFCE" for col in (2, 3, 4))
        assert all(_cell_font_rgb(dash.cell(rows_by_label[label], col)) == "006100" for col in (2, 3, 4))
    for label in ("Gross percent loss", "Gross IR loss"):
        assert all(_cell_fill_rgb(dash.cell(rows_by_label[label], col)) == "FFC7CE" for col in (2, 3, 4))
        assert all(_cell_font_rgb(dash.cell(rows_by_label[label], col)) == "9C0006" for col in (2, 3, 4))
        assert not _cf_fill_intersects(dash, f"B{rows_by_label[label]}:D{rows_by_label[label]}", "FFC7CE")
    for target in ("B49:B54", "B57:B62", "B64:B73", "B75:B90"):
        assert not _cf_fill_intersects(dash, target, "FFC7CE")
    assert _cf_fill_intersects(dash, "E20", "FFF2CC")
    wb.close()


def test_generated_calendar_and_instrument_averages_use_direct_semantic_fills(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][1]["close_time"] = "2026-06-02T02:00:00Z"
    snapshot["stats"]["by_instrument"][0].update({
        "long_wins": 1,
        "long_losses": 0,
        "short_wins": 0,
        "short_losses": 0,
        "net_result_pct": -2.5,
        "avg_result_pct": -1.25,
    })
    out = tmp_path / "semantic-tables.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)

    inst = wb[SYMBOLS_SHEET]
    assert inst.freeze_panes == "B3"
    headers = _instrument_averages_header_map(inst)
    row = INSTRUMENT_AVERAGES_DATA_START_ROW
    for header in ("Wins", "Long wins", "Short wins"):
        assert _cell_fill_rgb(inst.cell(row, headers[header])) == "C6EFCE"
    for header in ("Losses", "Long losses", "Short losses"):
        assert _cell_fill_rgb(inst.cell(row, headers[header])) == "FFC7CE"
    for header in ("Net P/L %", "Avg P/L %"):
        assert _cell_fill_rgb(inst.cell(row, headers[header])) == "C6EFCE"
        assert inst.cell(row, headers[header]).value == pytest.approx(0.023)
    assert _cell_fill_rgb(inst.cell(row, headers["Trades"])) == ""
    assert _cell_fill_rgb(inst.cell(row, headers["Win Rate %"])) == ""
    assert inst.cell(row, headers["Net P/L %"]).number_format == "0.00%"
    assert inst.cell(row, headers["Avg P/L %"]).number_format == "0.00%"
    assert inst.cell(row, headers["Avg duration (DD:HH:MM:SS)"]).number_format == "General"
    assert _parse_duration_text(inst.cell(row, headers["Avg duration (DD:HH:MM:SS)"]).value) is not None

    cal = wb["P&L Calendar"]
    assert cal["A1"].value == "Month"
    assert cal["B1"].value == 2026
    assert cal["A6"].value == "May"
    assert cal["A7"].value == "June"
    assert cal["B6"].value == "2.30%, 1 trade"
    assert cal["B7"].value == "-1.10%, 1 trade"
    assert _cell_fill_rgb(cal["B6"]) == "C6EFCE"
    assert _cell_fill_rgb(cal["B7"]) == "FFC7CE"
    wb.close()


def test_adaptive_trade_formats_keep_tiny_nonzero_values_visible(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["items"][0]["result_pct"] = 0.000001
    snapshot["items"][0]["r_multiple"] = 0.000001
    out = tmp_path / "adaptive.xlsx"
    build_master_journal_workbook(snapshot, out)
    ws = load_workbook(out)["Trade Log"]
    row = _trade_data_row()
    profit = ws.cell(row, _header_col(ws, "Profit %"))
    r_multiple = ws.cell(row, _header_col(ws, "R-Multiple"))
    assert profit.value != 0
    assert r_multiple.value != 0
    assert profit.number_format == adaptive_percent_number_format(profit.value)
    assert r_multiple.number_format == adaptive_number_format(r_multiple.value)
    assert profit.number_format != "0.00%"
    assert r_multiple.number_format != "0.00"


def test_trade_folder_resolver_exact_match_and_ambiguity(tmp_path: Path, monkeypatch):
    forex = tmp_path / "FOREX"
    crypto = tmp_path / "CRYPTO"
    fx_folder = forex / "2025" / "F1024 XAGUSD"
    fx_folder.mkdir(parents=True)
    (forex / "2025" / "F10240 WRONG").mkdir()
    january = crypto / "2025" / "JAN" / "C05 BTC"
    february = crypto / "2025" / "FEB" / "C05 ETH"
    january.mkdir(parents=True)
    february.mkdir(parents=True)

    target, reason = resolve_trade_folder_link("F1024", forex_root=forex, crypto_root=crypto)
    assert reason is None
    assert target == fx_folder.resolve().as_uri()

    target, reason = resolve_trade_folder_link(
        "C05", open_time="2025-02-10", forex_root=forex, crypto_root=crypto
    )
    assert reason is None
    assert target == february.resolve().as_uri()

    target, reason = resolve_trade_folder_link(
        "C05", open_time="2025-03-10", forex_root=forex, crypto_root=crypto
    )
    assert target is None
    assert reason == "ambiguous_trade_folder"

    env_crypto = tmp_path / "env" / "CRYPTO"
    env_folder = env_crypto / "2026" / "JAN" / "C107 BTCUSDT"
    env_folder.mkdir(parents=True)
    zero_padded_folder = env_crypto / "2026" / "FEB" / "C030 ADAUSDT"
    zero_padded_folder.mkdir(parents=True)
    nested_chart_folder = env_crypto / "2026" / "APR" / "C5" / "30"
    nested_chart_folder.mkdir(parents=True)
    (env_crypto / "2026" / "JUL" / "C30").mkdir(parents=True)
    (env_crypto / "2026" / "APR" / "C300").mkdir(parents=True)
    nested_file = env_crypto / "2026" / "MAR" / "screenshots" / "C108 BTCUSDT.png"
    nested_file.parent.mkdir(parents=True)
    nested_file.write_text("capture", encoding="utf-8")
    monkeypatch.setenv("TRADING_JOURNAL_CRYPTO_ROOT", str(env_crypto))
    diagnostics = {}
    target, reason = resolve_trade_folder_link(
        "C107", open_time="2026-01-10", diagnostics=diagnostics
    )
    assert reason is None
    assert target == env_folder.resolve().as_uri()
    assert diagnostics["checked_roots"][0] == str(env_crypto)
    target, reason = resolve_trade_folder_link("C30", open_time="2026-02-10")
    assert reason is None
    assert target == zero_padded_folder.resolve().as_uri()
    target, reason = resolve_trade_folder_link("C30", open_time="2026-04-10")
    assert reason is None
    assert target == nested_chart_folder.resolve().as_uri()
    target, reason = resolve_trade_folder_link("C300", open_time="2026-04-10")
    assert reason is None
    assert target == (env_crypto / "2026" / "APR" / "C300").resolve().as_uri()
    target, reason = resolve_trade_folder_link("C108", open_time="2026-03-10")
    assert reason is None
    assert target == nested_file.resolve().as_uri()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Log"
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        ws.cell(1, col, header)
    _ensure_trade_log_schema(ws)
    headers = _trade_log_header_map(ws)
    ws.cell(TRADE_LOG_DATA_START_ROW, headers[TRADE_NUMBER_HEADER]).value = "C107"
    ws.cell(TRADE_LOG_DATA_START_ROW, headers["Open Time"]).value = "2026-01-10"
    ws.cell(TRADE_LOG_DATA_START_ROW + 1, headers[TRADE_NUMBER_HEADER]).value = "C999"
    existing_target = (tmp_path / "existing").resolve().as_uri()
    ws.cell(TRADE_LOG_DATA_START_ROW + 1, headers[TRADE_NUMBER_HEADER]).hyperlink = existing_target
    link_diagnostics = {}
    _apply_trade_number_hyperlinks(ws, link_diagnostics)
    assert ws.cell(TRADE_LOG_DATA_START_ROW, headers[TRADE_NUMBER_HEADER]).hyperlink.target == env_folder.resolve().as_uri()
    assert ws.cell(TRADE_LOG_DATA_START_ROW + 1, headers[TRADE_NUMBER_HEADER]).hyperlink.target == existing_target
    assert link_diagnostics["trade_number_hyperlinks_added"] == 1
    unresolved = link_diagnostics["trade_number_hyperlink_unresolved"]
    assert unresolved[0]["trade_number"] == "C999"
    assert unresolved[0]["checked_roots"][0] == str(env_crypto)
    assert unresolved[0]["preserved_existing_hyperlink"] is True


def test_result_percentage_totals_segment_demo_balance_resets():
    rows = [
        {
            "row_type": "cashflow", "account": "OANDA DEMO", "asset_class": "fx",
            "symbol": "CASHFLOW", "open_time": "2026-01-01", "close_time": "2026-01-01",
            "cashflow_new_balance": 100.0, "balance_after_trade": 100.0, "currency": "AUD",
        },
        {
            "row_type": "trade", "account": "OANDA DEMO", "asset_class": "fx",
            "symbol": "EURUSD", "open_time": "2026-01-02", "close_time": "2026-01-02",
            "net_profit": -10.0, "balance_after_trade": 90.0, "currency": "AUD",
        },
        {
            "row_type": "trade", "account": "OANDA DEMO", "asset_class": "fx",
            "symbol": "USDJPY", "open_time": "2026-02-01", "close_time": "2026-02-01",
            "net_profit": -1.0, "balance_after_trade": 999.0, "currency": "AUD",
        },
        {
            "row_type": "trade", "account": "OANDA DEMO", "asset_class": "fx",
            "symbol": "GBPUSD", "open_time": "2026-02-02", "close_time": "2026-02-02",
            "net_profit": -9.0, "balance_after_trade": 990.0, "currency": "AUD",
        },
    ]
    balances = [{"account_label": "OANDA DEMO", "balance": 990.0, "currency": "AUD"}]
    stats = _result_percentage_totals_by_market(rows, balances)
    assert stats["fx"]["market_return_pct"] < 0
    account_diag = stats["fx"]["return_diagnostics"][0]
    assert account_diag["reset_count"] == 1
    assert len(account_diag["segments"]) == 2


def test_period_drawdown_uses_balance_segments_after_cashflow_resets():
    rows = [
        {
            "id": "seed", "row_type": "cashflow", "account": "OANDA DEMO",
            "close_time": "2026-01-01", "cashflow_new_balance": 100.0,
        },
        {
            "id": "peak-1", "row_type": "trade", "account": "OANDA DEMO",
            "close_time": "2026-01-02", "analysis_balance_after_trade": 100.0,
            "result_pct": 99.0,
        },
        {
            "id": "trough-1", "row_type": "trade", "account": "OANDA DEMO",
            "close_time": "2026-01-03", "analysis_balance_after_trade": 80.0,
            "result_pct": -99.0,
        },
        {
            "id": "reset", "row_type": "cashflow", "account": "OANDA DEMO",
            "close_time": "2026-01-04", "cashflow_new_balance": 50.0,
        },
        {
            "id": "peak-2", "row_type": "trade", "account": "OANDA DEMO",
            "close_time": "2026-01-05", "analysis_balance_after_trade": 50.0,
            "result_pct": 99.0,
        },
        {
            "id": "trough-2", "row_type": "trade", "account": "OANDA DEMO",
            "close_time": "2026-01-06", "analysis_balance_after_trade": 49.0,
            "result_pct": -99.0,
        },
    ]
    trades = [row for row in rows if row["row_type"] == "trade"]
    metrics = _period_drawdown_metrics(trades, rows)
    assert metrics["drawdown_segments_count"] == 2
    assert metrics["max_drawdown_pct"] == pytest.approx(20.0)
    assert metrics["avg_drawdown_pct"] == pytest.approx(11.0)
    assert metrics["min_drawdown_pct"] == pytest.approx(2.0)
    assert metrics["max_drawdown_detail"]["start_time"] == "2026-01-02"
    assert metrics["max_drawdown_detail"]["end_time"] == "2026-01-03"
    assert metrics["min_drawdown_detail"]["start_time"] == "2026-01-05"
    assert metrics["min_drawdown_detail"]["end_time"] == "2026-01-06"


def test_stats_symbols_and_reports_required_repairs(tmp_path: Path):
    snapshot = sample_snapshot()
    snapshot["balances"].append({
        "account_label": "OANDA DEMO", "balance": 980.0, "currency": "AUD"
    })
    snapshot["items"][0].update({
        "pattern": "Range", "ema": "20", "timeframe": "1H", "commission": 1.25,
    })
    snapshot["items"][1].update({
        "pattern": "Breakout", "ema": "50", "timeframe": "4H", "commission": 0.75,
    })
    snapshot["items"].append({
        "id": "t3", "row_type": "trade", "symbol": "EURUSD", "asset_class": "fx",
        "side": "SELL", "account": "OANDA DEMO", "open_time": "2026-05-03T00:00:00Z",
        "close_time": "2026-05-03T00:03:00Z", "net_profit": -20.0,
        "result_pct": -0.5, "r_multiple": -0.5, "trade_duration_seconds": 180,
        "analysis_balance_after_trade": 980.0, "pattern": "Range", "ema": "20",
        "timeframe": "4H", "commission": 2.0, "currency": "AUD",
    })
    out = tmp_path / "required-repairs.xlsx"
    build_master_journal_workbook(snapshot, out)
    wb = load_workbook(out)
    for sheet_name, coordinate in (
        (STATS1_SHEET, "B8"),
        (STATS2_SHEET, "A1"),
        (SYMBOLS_SHEET, "A2"),
        ("Trade Log", "A1"),
        ("P&L Calendar", "A1"),
        (REPORT_YEARLY_SHEET, "B1"),
    ):
        assert wb[sheet_name][coordinate].alignment.horizontal == "left"

    stats1 = wb[STATS1_SHEET]
    labels = {str(stats1.cell(row, 1).value or ""): row for row in range(1, stats1.max_row + 1)}
    label_rows = defaultdict(list)
    for row in range(1, stats1.max_row + 1):
        label_rows[str(stats1.cell(row, 1).value or "")].append(row)
    assert "Expectancy %" not in labels
    assert "Avg result %" not in labels
    assert "Avg R" not in labels
    assert stats1["A14"].value == "Percentage expectancy"
    assert stats1["A15"].value == "R expectancy"
    assert labels["Percentage expectancy"] < labels["Winners"]
    winners_section = label_rows["Winners"][0]
    primary_avg_duration = next(row for row in label_rows["Avg duration"] if row < winners_section)
    primary_avg_winning_duration = next(row for row in label_rows["Avg winning trade duration"] if row < winners_section)
    primary_avg_losing_duration = next(row for row in label_rows["Avg losing trade duration"] if row < winners_section)
    primary_max_duration = next(row for row in label_rows["Max duration"] if row < winners_section)
    assert primary_avg_duration + 1 == primary_avg_winning_duration
    assert primary_avg_winning_duration + 1 == primary_avg_losing_duration
    assert primary_avg_losing_duration + 1 == primary_max_duration
    for label in ("Avg winning trade duration", "Avg losing trade duration"):
        cell = stats1.cell(next(row for row in label_rows[label] if row < winners_section), 2)
        assert cell.number_format == "General"
        assert isinstance(cell.value, str)
        assert _parse_duration_text(cell.value) is not None
    assert stats1.cell(labels["Net P/L Percentage"], 2).value == pytest.approx(0.007)
    for label in ("Net P/L Percentage", "Gross percent gain", "Gross percent loss", "Gross IR gain", "Gross IR loss"):
        cell = stats1.cell(labels[label], 2)
        assert cell.alignment.horizontal == "left"
        assert cell.font.bold is False
    losers_section_for_style = next(row for row in label_rows["Losers"] if row < labels["Side"])

    categorical_section_rows = [
        row
        for row in range(1, stats1.max_row + 1)
        if str(stats1.cell(row, 1).value or "").strip() in {"Side", "Patterns", "Timeframe", "Commission"}
    ]

    def is_nested_outcome_label(row: int) -> bool:
        label = str(stats1.cell(row, 1).value or "").strip()
        if label not in {"Winners", "Losers"}:
            return False
        previous_sections = [section_row for section_row in categorical_section_rows if section_row < row]
        if not previous_sections:
            return False
        previous = previous_sections[-1]
        next_sections = [section_row for section_row in categorical_section_rows if section_row > previous]
        end = next_sections[0] if next_sections else stats1.max_row + 1
        return previous < row < end

    for row in range(1, stats1.max_row + 1):
        label_cell = stats1.cell(row, 1)
        if label_cell.value in (None, ""):
            continue
        if str(label_cell.value).strip() == "Source" or is_nested_outcome_label(row):
            assert label_cell.font.bold is False
            assert label_cell.font.italic is True
            assert label_cell.alignment.horizontal == "right"
            continue
        assert label_cell.font.bold is True
        assert _cell_font_rgb(label_cell) == "000000"
    for section in ("Side", "Patterns", "Timeframe"):
        section_row = labels[section]
        winner_row = next(
            row for row in range(section_row + 1, stats1.max_row + 1)
            if str(stats1.cell(row, 1).value or "").strip() == "Winners"
        )
        loser_row = next(
            row for row in range(winner_row + 1, stats1.max_row + 1)
            if str(stats1.cell(row, 1).value or "").strip() == "Losers"
        )
        assert _cell_fill_rgb(stats1.cell(winner_row, 2)) == "C6EFCE"
        assert _cell_fill_rgb(stats1.cell(loser_row, 2)) == "FFC7CE"
        for row in (winner_row, loser_row):
            assert stats1.cell(row, 1).font.bold is False
            assert stats1.cell(row, 1).font.italic is True
            assert stats1.cell(row, 1).alignment.horizontal == "right"
    losers_section = next(
        row for row in range(1, stats1.max_row + 1)
        if stats1.cell(row, 1).value == "Losers" and row < labels["Side"]
    )
    loser_row = losers_section + 1
    assert all(stats1.cell(loser_row, col).alignment.horizontal == "left" for col in (2, 3, 4))
    for label in ("Min stop %", "Max stop %", "Min target %", "Max target %"):
        row = label_rows[label][0]
        assert stats1.cell(row + 1, 1).value == "Source"
        assert "·" in str(stats1.cell(row + 1, 2).value)
    for label, expected_fx, expected_crypto in (
        ("Min Commission", 1.25, 0.75),
        ("Avg Commission", 1.625, 0.75),
        ("Max Commission", 2.0, 0.75),
        ("Total Commission", 3.25, 0.75),
    ):
        row = labels[label]
        assert stats1.cell(row, 2).value in (None, "")
        assert stats1.cell(row, 3).value == pytest.approx(expected_fx)
        assert stats1.cell(row, 4).value == pytest.approx(expected_crypto)
        assert "AUD" in stats1.cell(row, 3).number_format
        assert "USDT" in stats1.cell(row, 4).number_format
        assert "%" not in str(stats1.cell(row, 3).number_format) + str(stats1.cell(row, 4).number_format)

    assert stats1.cell(labels["Channel"], 2).value == 0
    assert stats1.cell(labels["Range"], 3).value == 2
    one_hour_row = labels["1H"]
    assert stats1.cell(one_hour_row, 2).value == 1
    assert stats1.cell(one_hour_row + 1, 1).value == "Winners"
    assert stats1.cell(one_hour_row + 1, 2).value == 1
    assert stats1.cell(one_hour_row + 2, 1).value == "Losers"
    assert stats1.cell(one_hour_row + 2, 2).value == 0

    trade_log = wb["Trade Log"]
    trade_duration_cell = trade_log.cell(
        TRADE_LOG_DATA_START_ROW,
        _header_col(trade_log, "Trade Duration (DD:HH:MM:SS)"),
    )
    assert trade_duration_cell.number_format == "General"
    assert isinstance(trade_duration_cell.value, str)
    assert _parse_duration_text(trade_duration_cell.value) is not None
    assert "[>=1000000]" not in DURATION_NUMBER_FORMAT
    for row in range(1, stats1.max_row + 1):
        label = str(stats1.cell(row, 1).value or "").lower()
        if "duration" in label or "move to" in label:
            for col in (2, 3, 4):
                if stats1.cell(row, col).value not in (None, ""):
                    cell = stats1.cell(row, col)
                    assert cell.number_format == "General"
                    assert isinstance(cell.value, str)
                    assert not str(cell.value).startswith(("00 days", "00 hours"))
                    assert _parse_duration_text(cell.value) is not None

    symbols = wb[SYMBOLS_SHEET]
    headers = _instrument_averages_header_map(symbols)
    assert headers["Most Traded Pattern"] == 17
    assert headers["Most Traded EMA"] == 18
    assert headers["Most Profitable Timeframe"] == headers["Most traded timeframe"] + 1
    assert headers["Least Profitable Timeframe"] == headers["Most traded timeframe"] + 2
    timeframe_width = symbols.column_dimensions[
        get_column_letter(headers["Most traded timeframe"])
    ].width
    assert symbols.column_dimensions[
        get_column_letter(headers["Most Profitable Timeframe"])
    ].width >= timeframe_width
    assert symbols.column_dimensions[
        get_column_letter(headers["Least Profitable Timeframe"])
    ].width >= timeframe_width
    assert symbols.cell(3, headers["Most Traded Pattern"]).value == "Range"
    assert symbols.cell(3, headers["Most Traded EMA"]).value == "20"
    assert symbols.cell(3, headers["Most Profitable Timeframe"]).value == "1H"
    assert symbols.cell(3, headers["Least Profitable Timeframe"]).value == "4H"
    for header in (
        "Shortest duration (DD:HH:MM:SS)",
        "Avg duration (DD:HH:MM:SS)",
        "Longest duration (DD:HH:MM:SS)",
        "Avg winning duration (DD:HH:MM:SS)",
        "Avg losing duration (DD:HH:MM:SS)",
    ):
        assert symbols.cell(3, headers[header]).number_format == "General"
    symbol_rows = {
        str(symbols.cell(row, headers["Symbol"]).value or "").strip().upper(): row
        for row in range(INSTRUMENT_AVERAGES_DATA_START_ROW, symbols.max_row + 1)
        if str(symbols.cell(row, headers["Symbol"]).value or "").strip()
    }
    eurusd_row = symbol_rows["EURUSD"]
    for header in (
        "Shortest duration (DD:HH:MM:SS)",
        "Avg duration (DD:HH:MM:SS)",
        "Longest duration (DD:HH:MM:SS)",
        "Avg winning duration (DD:HH:MM:SS)",
        "Avg losing duration (DD:HH:MM:SS)",
    ):
        assert _parse_duration_text(symbols.cell(eurusd_row, headers[header]).value) is not None

    stats2 = wb[STATS2_SHEET]
    stats2_values = [
        str(stats2.cell(row, col).value or "")
        for row in range(1, stats2.max_row + 1)
        for col in range(1, stats2.max_column + 1)
    ]
    assert "Instrument leaders" not in stats2_values
    balance_headers = {
        str(stats2.cell(2, col).value or ""): col for col in range(1, stats2.max_column + 1)
    }
    account_rows = {
        str(stats2.cell(row, balance_headers["Account"]).value or ""): row
        for row in range(3, stats2.max_row + 1)
    }
    risk_cell = stats2.cell(account_rows["OANDA DEMO"], balance_headers["Risk of Ruin"])
    assert 0 <= risk_cell.value <= 1
    assert risk_cell.number_format == "0.00%"
    assert risk_cell.comment and "Balsara" in risk_cell.comment.text

    for sheet_name, period_header in ((REPORT_YEARLY_SHEET, 2026), ("2026", "May")):
        report = wb[sheet_name]
        report_rows = {
            str(report.cell(row, 1).value or ""): row for row in range(1, report.max_row + 1)
        }
        report_label_rows = defaultdict(list)
        for row in range(1, report.max_row + 1):
            report_label_rows[str(report.cell(row, 1).value or "")].append(row)
        assert "Avg result %" not in report_rows
        assert "Avg R" not in report_rows
        for label in ("Percentage expectancy", "R expectancy", "Gross percent gain", "Gross percent loss", "Gross IR gain", "Gross IR loss"):
            assert label in report_rows
        assert report_rows["Min drawdown"] == report_rows["Avg drawdown"] + 1
        period_col = next(
            col for col in range(2, report.max_column + 1)
            if report.cell(1, col).value == period_header
        )
        for label in ("Min stop %", "Max stop %", "Min target %", "Max target %"):
            row = report_label_rows[label][0]
            assert report.cell(row + 1, 1).value == "Source"
            assert "·" in str(report.cell(row + 1, period_col).value)
        total_commission = report.cell(report_rows["Total Commission"], period_col)
        assert "AUD 3.25" in str(total_commission.value)
        assert "USDT 0.75" in str(total_commission.value)
        duration_row = next(
            row for label, row in report_rows.items() if "DD:HH:MM:SS" in label
        )
        duration_cell = report.cell(duration_row, period_col)
        assert duration_cell.number_format == "General"
        assert isinstance(duration_cell.value, str)
        assert not duration_cell.value.startswith(("00 days", "00 hours"))
        assert _parse_duration_text(duration_cell.value) is not None
    wb.close()


def test_report_profit_rows_are_linear_percentages(tmp_path: Path):
    snapshot = sample_snapshot()
    out = tmp_path / "report-percentages.xlsx"
    build_master_journal_workbook(snapshot, out)
    yearly = load_workbook(out)[REPORT_YEARLY_SHEET]
    rows = {str(yearly.cell(row, 1).value): row for row in range(2, yearly.max_row + 1)}
    year_col = next(
        col for col in range(2, yearly.max_column + 1)
        if yearly.cell(1, col).value == 2026
    )
    assert yearly.cell(rows["Net P/L"], year_col).value == pytest.approx(0.012)
    assert yearly.cell(rows["Gross percent gain"], year_col).value == pytest.approx(0.023)
    assert yearly.cell(rows["Gross percent loss"], year_col).value == pytest.approx(0.011)
    assert yearly.cell(rows["Gross IR gain"], year_col).value == pytest.approx(1.2)
    assert yearly.cell(rows["Gross IR loss"], year_col).value == pytest.approx(0.8)
    for label in ("Net P/L", "Gross percent gain", "Gross percent loss"):
        assert yearly.cell(rows[label], year_col).number_format == "0.00%"
    for label in ("Gross IR gain", "Gross IR loss"):
        assert yearly.cell(rows[label], year_col).number_format == '0.000"R"'
