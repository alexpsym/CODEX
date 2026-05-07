from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from TJR.make_trading_journal_replica import (
    compute_journal_stats_replica,
    instrument_display_rows,
    build_output,
    trade_duration_seconds,
    find_journal_dir,
)


def test_gross_loss_is_positive_absolute_value():
    s=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":-5},{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":3}])
    assert s['totals']['gross_loss']==5

def test_money_by_currency_gross_loss_is_positive_absolute_value():
    s=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":-5,"currency":"AUD"}])
    assert s['totals']['money_by_currency']['gross_loss']['AUD']==5

def test_drawdown_average_ignores_zero_peak_points():
    rows=[{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"balance_after":1000,"close_time":datetime(2026,1,1)},{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"balance_after":1100,"close_time":datetime(2026,1,2)},{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":-1,"balance_after":990,"close_time":datetime(2026,1,3)}]
    t=compute_journal_stats_replica(rows)['totals']
    assert t['max_drawdown_pct']==10.0 and t['avg_drawdown_pct']==10.0 and t['min_drawdown_pct']==10.0

def test_duration_group_matches_render_stats_shape():
    rows=[{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"trade_duration_seconds":10},{"symbol":"B","side":"SELL","asset_class":"crypto","net_profit":-1,"trade_duration_seconds":20}]
    d=compute_journal_stats_replica(rows)['groups']['duration']
    keys=["overall_avg_seconds","overall_shortest_seconds","overall_longest_seconds","overall_avg_winner_seconds","overall_avg_loser_seconds","overall_longest_winner_seconds","overall_longest_loser_seconds","fx_avg_seconds","fx_shortest_seconds","fx_longest_seconds","fx_avg_winner_seconds","fx_avg_loser_seconds","fx_shortest_winner_seconds","fx_shortest_loser_seconds","fx_longest_winner_seconds","fx_longest_loser_seconds","crypto_avg_seconds","crypto_shortest_seconds","crypto_longest_seconds","crypto_avg_winner_seconds","crypto_avg_loser_seconds","crypto_shortest_winner_seconds","crypto_shortest_loser_seconds","crypto_longest_winner_seconds","crypto_longest_loser_seconds","metric_sources"]
    for k in keys: assert k in d

def test_risk_expectancy_group_matches_render_stats_shape():
    rows=[{"symbol":"A","side":"BUY","asset_class":"FX","entry":100,"stop_loss":90,"take_profit":110,"result_pct":1,"r_multiple":1,"net_profit":1,"balance_after":1000,"close_time":datetime(2026,1,1)},{"symbol":"A","side":"BUY","asset_class":"FX","entry":100,"stop_loss":90,"take_profit":110,"result_pct":-1,"r_multiple":-1,"net_profit":-1,"balance_after":900,"close_time":datetime(2026,1,2)}]
    r=compute_journal_stats_replica(rows)['groups']['risk_expectancy']
    for k in ["avg_stop_pct","avg_target_pct","avg_result_pct","avg_r_multiple","avg_stop_pct_winners","avg_stop_pct_losers","avg_target_pct_winners","avg_target_pct_losers","avg_result_pct_winners","avg_result_pct_losers","avg_r_multiple_winners","avg_r_multiple_losers","max_drawdown_pct","avg_drawdown_pct","min_drawdown_pct"]: assert k in r

def test_fx_instrument_distances_are_pips_not_raw_price():
    rows=[{"symbol":"EURUSD","side":"BUY","asset_class":"FX","entry":1.1,"stop_loss":1.095,"take_profit":1.11,"net_profit":10,"trade_duration_seconds":1}]
    first=compute_journal_stats_replica(rows)['by_instrument'][0]
    assert round(first['avg_sl_distance_pips'],6)==50.0 and round(first['avg_tp_distance_pips'],6)==100.0
    disp=instrument_display_rows([first])[0]
    assert round(disp['Avg SL W'],6)==50.0 and round(disp['Avg TP W'],6)==100.0

def test_by_instrument_asset_class_is_render_normalized():
    first=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":1}])['by_instrument'][0]
    assert first['asset_class'] in {'fx','crypto'}

def test_market_buckets_have_metric_sources():
    b=compute_journal_stats_replica([{"id":"x1","symbol":"A","side":"BUY","asset_class":"FX","net_profit":-5,"result_pct":-5,"r_multiple":-1,"trade_duration_seconds":2}])['groups']['by_market']['overall']
    assert 'metric_sources' in b and b['metric_sources']['max_loss']['id']=='x1'

def test_unknown_pnl_is_not_break_even():
    t=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","entry":1,"exit":2,"net_profit":None}])['totals']
    assert t['break_even']==0

def test_zero_second_trade_duration_rounds_to_one_second():
    row={"open_time":datetime(2026,1,1),"close_time":datetime(2026,1,1)}
    assert trade_duration_seconds(row)==1.0

def test_dashboard_does_not_collapse_mixed_currency_money(tmp_path: Path):
    j=tmp_path/'journal';j.mkdir();wb=Workbook();ws=wb.active
    ws.append(["opening_time","closing_time","type_buy_sell","symbol","entry_price","closing_price","net_profit","currency"])
    ws.append(["2026-01-01","2026-01-01","Buy","EURUSD",1.1,1.2,10,"AUD"])
    ws.append(["2026-01-02","2026-01-02","Buy","BTCUSDT",100,110,20,"USDT"])
    wb.save(j/'BYBIT DEMO.xlsx')
    out=tmp_path/'o.xlsx';build_output(j,out)
    wb2=load_workbook(out, read_only=True, data_only=True); ws2=wb2['Dashboard']
    vals={(ws2.cell(r,c).value) for r in range(1,200) for c in range(1,5)}
    assert 'AUD' in vals and 'USDT' in vals
    overall_net=None
    for r in range(1,200):
        if ws2.cell(r,1).value=='Net P/L' and any(ws2.cell(rr,1).value=='Overall' for rr in range(max(1,r-25),r)):
            overall_net=ws2.cell(r,2).value; break
    assert isinstance(overall_net,str) and 'AUD' in overall_net and 'USDT' in overall_net and '30' not in overall_net


def test_duration_metric_sources_are_real_refs():
    rows=[{"id":"s","symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"trade_duration_seconds":10},{"id":"l","symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"trade_duration_seconds":20}]
    ms=compute_journal_stats_replica(rows)['groups']['duration']['metric_sources']
    assert ms['overall_shortest_seconds']['id']=='s' and ms['overall_longest_seconds']['id']=='l'

def test_market_bucket_matches_render_shape():
    b=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","entry":100,"stop_loss":90,"take_profit":110,"net_profit":1,"trade_duration_seconds":10}])['groups']['by_market']['overall']
    for k in ['shortest_duration_seconds','longest_duration_seconds','instruments','min_stop_pct','max_stop_pct','min_target_pct','max_target_pct','max_drawdown_pct']:
        assert k in b

def test_totals_contains_render_parity_keys():
    t=compute_journal_stats_replica([{"symbol":"A","side":"BUY","asset_class":"FX","net_profit":1,"trade_duration_seconds":10}])['totals']
    for k in ['min_gain','min_loss','max_winner_duration_seconds','max_loser_duration_seconds','drawdown_balance_points','drawdown_segments_count','unique_instruments','crypto_instruments','fx_instruments']:
        assert k in t

def test_dashboard_duration_values_are_human_readable_or_explicitly_labelled(tmp_path: Path):
    j=tmp_path/'journal';j.mkdir();wb=Workbook();ws=wb.active
    ws.append(["opening_time","closing_time","type_buy_sell","symbol","entry_price","closing_price","net_profit","currency"])
    ws.append(["2026-01-01 00:00","2026-01-01 01:24","Buy","EURUSD",1.1,1.2,10,"AUD"])
    wb.save(j/'BYBIT DEMO.xlsx')
    out=tmp_path/'o2.xlsx';build_output(j,out);wb2=load_workbook(out, read_only=True, data_only=True);d=wb2['Dashboard']
    vals={str(d.cell(r,2).value) for r in range(1,220) if d.cell(r,1).value=='Avg duration'}
    assert any(('h' in v or 'm' in v or 's' in v) for v in vals)


def test_find_journal_dir_uses_repo_arg(tmp_path: Path):
    repo = tmp_path / "repo"
    journal = repo / "journal"
    journal.mkdir(parents=True)
    assert find_journal_dir(str(repo)) == journal


def test_find_journal_dir_uses_codex_repo_env(tmp_path: Path, monkeypatch):
    repo = tmp_path / "env_repo"
    journal = repo / "journal"
    journal.mkdir(parents=True)
    monkeypatch.setenv("CODEX_REPO_DIR", str(repo))
    assert find_journal_dir(None) == journal


def test_find_journal_dir_defaults_updated():
    text = (Path(__file__).resolve().parents[1] / "TJR" / "make_trading_journal_replica.py").read_text(encoding="utf-8")
    assert "/storage/emulated/0/Download/CODEX-master/CODEX-master/journal" in text
    assert "CODEX-master (4)" not in text
