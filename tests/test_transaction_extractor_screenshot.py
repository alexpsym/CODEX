from __future__ import annotations

import csv
import importlib.util
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXTRACTOR_DIR = ROOT / "LEDGER-clone" / "transaction_extractor_local_fixed_v7_no_venv"
EXTRACTOR_PATH = EXTRACTOR_DIR / "extract_transactions_local.py"
RUN_DIR = EXTRACTOR_DIR / "outputs" / "run_20260704_194549"
SIMPLE_LIST_SCREENSHOT = RUN_DIR / "downloads" / "source_01.jpg"
BANK_APP_SCREENSHOT = RUN_DIR / "downloads" / "source_02.jpg"


# side is the extractor's ledger side. Final XLSX/CSV cells must still be positive.
EXPECTED_SIMPLE = [
    ("source_01.jpg", "04/07/2026", "W2695-Woolworths DARRA", "credit", 29.64, False),
    ("source_01.jpg", "27/06/2026", "W2695-Woolworths DARRA", "credit", 37.70, False),
    ("source_01.jpg", "25/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 15.50, False),
    ("source_01.jpg", "21/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 33.30, False),
    ("source_01.jpg", "17/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 5.20, False),
    ("source_01.jpg", "16/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 4.20, False),
    ("source_01.jpg", "14/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 43.50, False),
    ("source_01.jpg", "13/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 6.50, False),
    ("source_01.jpg", "12/06/2026", "W2864-Woolworths BROOKWATER", "credit", 12.80, False),
    ("source_01.jpg", "11/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 23.99, False),
    ("source_01.jpg", "07/06/2026", "W2648-Woolworths FOREST LAKE", "credit", 8.00, False),
]

EXPECTED_BANK = [
    ("source_02.jpg", "04/07/2026", "RENDER.COM +14158304762 US", "credit", 10.17, True),
    ("source_02.jpg", "03/07/2026", "Bridgewater Bake House", "credit", 5.50, False),
    ("source_02.jpg", "03/07/2026", "Bridgewater Bake House", "credit", 8.50, False),
    ("source_02.jpg", "03/07/2026", "ROBINS PIZZA/SHOP 17/2 YE MEADOWBROOK AU", "credit", 19.82, True),
    ("source_02.jpg", "01/07/2026", "Yummy Fried Rice", "credit", 15.90, False),
    ("source_02.jpg", "15/06/2026", "Sp Nutra Nourished", "credit", 55.12, False),
    ("source_02.jpg", "12/06/2026", "Pizza In A Hurryohm Food", "credit", 14.95, False),
    ("source_02.jpg", "09/06/2026", "Foodworks", "credit", 6.50, False),
    ("source_02.jpg", "08/06/2026", "Khao Thai Restaurant", "credit", 29.72, False),
    ("source_02.jpg", "07/06/2026", "Transfer to Alexander Symoniw", "credit", 300.00, False),
    ("source_02.jpg", "07/06/2026", "Transfer from SaveME", "debit", 300.00, False),
    ("source_02.jpg", "05/06/2026", "Robins Pizza/Shop 17/2", "credit", 19.82, False),
    ("source_02.jpg", "05/06/2026", "Render.Com", "credit", 9.83, False),
    ("source_02.jpg", "04/06/2026", "Transfer from SaveME", "debit", 211.00, False),
]

EXPECTED = EXPECTED_SIMPLE + EXPECTED_BANK


@pytest.fixture(scope="module")
def extractor():
    spec = importlib.util.spec_from_file_location("transaction_extractor_local", EXTRACTOR_PATH)
    module = importlib.util.module_from_spec(spec)
    sys.modules["transaction_extractor_local"] = module
    assert spec.loader is not None
    spec.loader.exec_module(module)
    try:
        module.ensure_tesseract()
    except RuntimeError as exc:
        pytest.skip(str(exc))
    return module


@pytest.fixture(scope="module")
def extracted_rows(extractor):
    simple_rows, simple_debug = extractor.source_to_rows(SIMPLE_LIST_SCREENSHOT, RUN_DIR)
    bank_rows, _bank_debug = extractor.source_to_rows(BANK_APP_SCREENSHOT, RUN_DIR)
    assert simple_debug.splitlines()[0].startswith("SIMPLE_LIST_MODE")
    assert len(simple_rows) == 11
    return simple_rows + bank_rows


def _actual_tuple(row):
    if row.debit is not None:
        side = "debit"
        amount = float(row.debit)
    elif row.credit is not None:
        side = "credit"
        amount = float(row.credit)
    else:
        side = ""
        amount = None
    return (
        row.source_name,
        row.date,
        row.description,
        side,
        amount,
        row.pending,
        row.needs_review,
        row.review_reason,
    )


def test_supplied_screenshot_transactions_total_25_rows(extracted_rows):
    expected = [(*row, False, "") for row in EXPECTED]
    assert [_actual_tuple(row) for row in extracted_rows] == expected
    assert len(extracted_rows) == 25

    # Rows 13-15 in the combined output share the "Yesterday, 3 July 2026" heading.
    assert [row.date for row in extracted_rows[12:15]] == ["03/07/2026"] * 3


def test_screenshot_writers_preserve_positive_ledger_amounts(extractor, extracted_rows, tmp_path):
    xlsx_path = tmp_path / "extracted_transactions_test.xlsx"
    csv_path = tmp_path / "review_extraction_test.csv"

    extractor.write_workbook(extracted_rows, xlsx_path)
    extractor.write_review(extracted_rows, csv_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["Transactions"]
    assert [cell.value for cell in ws[1]] == [
        "DATE",
        "ACCOUNT_TYPE",
        "ACCOUNT",
        "DESCRIPTION",
        "DEBIT",
        "CREDIT",
        "NOTES",
        "NOTES",
    ]

    body = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(body) == 25
    for sheet_row, (_source, date, desc, side, amount, pending) in zip(body, EXPECTED):
        assert sheet_row[0] == date
        assert sheet_row[3] == desc
        debit = sheet_row[4]
        credit = sheet_row[5]
        if side == "debit":
            assert float(debit) == amount
            assert credit in (None, "")
        else:
            assert debit in (None, "")
            assert float(credit) == amount
        assert (debit in (None, "") or float(debit) >= 0)
        assert (credit in (None, "") or float(credit) >= 0)
        assert sheet_row[6] == ("pending" if pending else None)

    with csv_path.open(newline="", encoding="utf-8") as f:
        review_rows = list(csv.DictReader(f))
    assert len(review_rows) == 25
    assert {row["needs_review"] for row in review_rows} == {"NO"}
    assert [row["pending"] for row in review_rows].count("YES") == 2
    for row in review_rows:
        for field in ("debit", "credit"):
            if row[field]:
                assert float(row[field]) >= 0


def test_amount_normalization_fails_safe(extractor):
    assert extractor.normalize_amount_candidate("$5.0") is None
    assert extractor.normalize_amount_candidate("319.82") is None
    assert extractor.normalize_amount_candidate("+$21L00")[:2] == (211.0, False)
    assert extractor.normalize_amount_candidate("$650")[:2] == (6.5, False)
