import asyncio
import importlib.util
import json
import os
import shutil
import sys
from pathlib import Path
import types

import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

# dependency-tolerant import shims
bm_pkg = types.ModuleType("bybit_monitor")
bm_mod = types.ModuleType("bybit_monitor.bybit_altcoin_monitor")
bm_mod.__getattr__ = lambda _name: (lambda *a, **k: None)
bm_pkg.bybit_altcoin_monitor = bm_mod
sys.modules.setdefault("bybit_monitor", bm_pkg)
sys.modules.setdefault("bybit_monitor.bybit_altcoin_monitor", bm_mod)
om_pkg = types.ModuleType("oanda_monitor")
om_mod = types.ModuleType("oanda_monitor.oanda_forex_monitor")
om_mod.__getattr__ = lambda _name: (lambda *a, **k: None)
om_pkg.oanda_forex_monitor = om_mod
sys.modules.setdefault("oanda_monitor", om_pkg)
sys.modules.setdefault("oanda_monitor.oanda_forex_monitor", om_mod)
try:
    _httpx_spec = importlib.util.find_spec("httpx")
except ValueError:
    _httpx_spec = None
if _httpx_spec is None:
    httpx_stub = types.ModuleType("httpx")
    httpx_stub.AsyncClient = object
    httpx_stub.Timeout = lambda *a, **k: None
    httpx_stub.TimeoutException = Exception
    httpx_stub.RequestError = Exception
    httpx_stub.HTTPStatusError = Exception
    httpx_stub.Response = object
    httpx_stub.ConnectError = Exception
    sys.modules.setdefault("httpx", httpx_stub)
mp_pkg = types.ModuleType("multipart")
mp_pkg.__version__ = "0.0-test"
mp_sub = types.ModuleType("multipart.multipart")
mp_sub.parse_options_header = lambda *args, **kwargs: ("", {})
sys.modules.setdefault("multipart", mp_pkg)
sys.modules.setdefault("multipart.multipart", mp_sub)
try:
    _requests_spec = importlib.util.find_spec("requests")
except ValueError:
    _requests_spec = None
if _requests_spec is None:
    requests_stub = types.ModuleType("requests")
    requests_adapters = types.ModuleType("requests.adapters")
    requests_adapters.HTTPAdapter = object
    requests_stub.adapters = requests_adapters
    sys.modules.setdefault("requests", requests_stub)
    sys.modules.setdefault("requests.adapters", requests_adapters)
sys.modules.setdefault("dotenv", types.SimpleNamespace(load_dotenv=lambda *args, **kwargs: None))
SPEC = importlib.util.spec_from_file_location(
    "render_master_service_oanda_history_export", ROOT / "render" / "master_service.py"
)
master_service = importlib.util.module_from_spec(SPEC)
assert SPEC and SPEC.loader
sys.modules[SPEC.name] = master_service
SPEC.loader.exec_module(master_service)

OANDA_HISTORY_SPEC = importlib.util.spec_from_file_location(
    "oanda_history_exporter_module", ROOT / "oanda_history-clone" / "oanda_history.py"
)
oanda_history_exporter = importlib.util.module_from_spec(OANDA_HISTORY_SPEC)
assert OANDA_HISTORY_SPEC and OANDA_HISTORY_SPEC.loader
sys.modules[OANDA_HISTORY_SPEC.name] = oanda_history_exporter
OANDA_HISTORY_SPEC.loader.exec_module(oanda_history_exporter)


def test_oanda_history_exporter_writes_opening_loss_quote_home_factor():
    row = oanda_history_exporter._transaction_to_row(
        {
            "id": "100",
            "type": "ORDER_FILL",
            "reason": "MARKET_ORDER",
            "time": "2026-01-01T00:00:00Z",
            "instrument": "EUR_USD",
            "units": "1000",
            "price": "1.1000",
            "homeConversionFactors": {"lossQuoteHome": {"factor": "1.5"}},
            "tradeOpened": {"tradeID": "t1", "units": "1000"},
        }
    )

    assert row["CONVERSION RATE"] == "1.5000"


def test_oanda_history_exporter_supports_deprecated_loss_quote_home_factor():
    row = oanda_history_exporter._transaction_to_row(
        {
            "id": "100",
            "type": "ORDER_FILL",
            "reason": "MARKET_ORDER",
            "time": "2026-01-01T00:00:00Z",
            "instrument": "EUR_USD",
            "units": "1000",
            "price": "1.1000",
            "lossQuoteHomeConversionFactor": "1.4",
            "tradeOpened": {"tradeID": "t1", "units": "1000"},
        }
    )

    assert row["CONVERSION RATE"] == "1.4000"


def test_oanda_history_exporter_does_not_write_closing_conversion_factor():
    row = oanda_history_exporter._transaction_to_row(
        {
            "id": "101",
            "type": "ORDER_FILL",
            "reason": "TAKE_PROFIT_ORDER",
            "time": "2026-01-01T01:00:00Z",
            "instrument": "EUR_USD",
            "units": "-1000",
            "price": "1.1020",
            "homeConversionFactors": {"lossQuoteHome": {"factor": "1.6"}},
            "tradesClosed": [{"tradeID": "t1", "units": "-1000"}],
        }
    )

    assert row["CONVERSION RATE"] == ""


def test_collect_oanda_history_range_uses_base_url_and_splits_windows(monkeypatch: pytest.MonkeyPatch):
    calls = []

    async def fake_fetch(**kwargs):
        calls.append(kwargs)
        return [{"id": f"tx-{len(calls)}"}]

    monkeypatch.setattr(master_service, "_fetch_oanda_transactions_window", fake_fetch)

    start = master_service.datetime(2020, 1, 1, tzinfo=master_service.timezone.utc)
    end = master_service.datetime(2023, 1, 1, tzinfo=master_service.timezone.utc)

    transactions = asyncio.run(
        master_service._collect_oanda_history_range(
            account_id="acc",
            api_key="key",
            base_url="https://api-fxpractice.oanda.com/v3",
            start=start,
            end=end,
        )
    )

    assert len(calls) >= 3
    assert all(call["base_url"].endswith("/v3") for call in calls)
    assert [item["id"] for item in transactions] == ["tx-1", "tx-2", "tx-3", "tx-4"][: len(transactions)]


def test_get_oanda_history_config_live_uses_live_overrides(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("OANDA_API_KEY", "live-key")
    monkeypatch.setenv("OANDA_ACCOUNT_ID", "live-account")
    monkeypatch.setenv("OANDA_API_URL_LIVE", "https://api-fxtrade.oanda.com")
    monkeypatch.setenv("OANDA_BASE_URL_LIVE", "https://ignore-me.example")

    config = master_service._get_oanda_history_config("live")

    assert config["mode"] == "live"
    assert config["base_url"] == "https://api-fxtrade.oanda.com/v3"


def test_get_oanda_history_config_demo_normalizes_v3(monkeypatch: pytest.MonkeyPatch):
    monkeypatch.setenv("OANDA_API_KEY_DEMO", "demo-key")
    monkeypatch.setenv("OANDA_ACCOUNT_ID_DEMO", "demo-account")
    monkeypatch.setenv("OANDA_API_URL_DEMO", "https://api-fxpractice.oanda.com/v3")

    config = master_service._get_oanda_history_config("demo")

    assert config["mode"] == "demo"
    assert config["base_url"] == "https://api-fxpractice.oanda.com/v3"


def test_build_oanda_v3_url_normalizes_base_and_endpoint():
    account_id = "101-001-1234567-001"
    url_plain = master_service._build_oanda_v3_url(
        "https://api-fxpractice.oanda.com",
        "/accounts/{account_id}",
        account_id=account_id,
    )
    url_v3 = master_service._build_oanda_v3_url(
        "https://api-fxpractice.oanda.com/v3",
        "/accounts/{account_id}",
        account_id=account_id,
    )
    url_double_v3_input = master_service._build_oanda_v3_url(
        "https://api-fxpractice.oanda.com/v3/",
        "/v3/accounts/{account_id}",
        account_id=account_id,
    )
    assert url_plain == f"https://api-fxpractice.oanda.com/v3/accounts/{account_id}"
    assert url_v3 == f"https://api-fxpractice.oanda.com/v3/accounts/{account_id}"
    assert url_double_v3_input == f"https://api-fxpractice.oanda.com/v3/accounts/{account_id}"
    assert "/v3/v3/" not in url_plain
    assert "/v3/v3/" not in url_v3
    assert "/v3/v3/" not in url_double_v3_input


def test_fetch_oanda_account_created_time_uses_single_v3_prefix(monkeypatch: pytest.MonkeyPatch):
    seen_urls = []

    class FakeResponse:
        status_code = 200
        content = b'{"account":{"createdTime":"2020-01-01T00:00:00.000000000Z"}}'
        text = '{"account":{"createdTime":"2020-01-01T00:00:00.000000000Z"}}'

        def raise_for_status(self):
            return None

        def json(self):
            return {"account": {"createdTime": "2020-01-01T00:00:00.000000000Z"}}

    class FakeClient:
        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def get(self, url, headers=None):
            seen_urls.append(url)
            return FakeResponse()

    monkeypatch.setattr(master_service.httpx, "AsyncClient", lambda *args, **kwargs: FakeClient())

    created = asyncio.run(
        master_service._fetch_oanda_account_created_time(
            base_url="https://api-fxpractice.oanda.com/v3",
            account_id="101-001-1234567-001",
            api_key="demo-token",
        )
    )
    assert seen_urls == ["https://api-fxpractice.oanda.com/v3/accounts/101-001-1234567-001"]
    assert created == master_service.datetime(2020, 1, 1, tzinfo=master_service.timezone.utc)


def test_oanda_trade_endpoints_use_helper_without_double_v3(monkeypatch: pytest.MonkeyPatch):
    calls = []

    class FakeResponse:
        def __init__(self, payload):
            self.status_code = 200
            self.text = "{}"
            self._payload = payload

        def json(self):
            return self._payload

    class FakeClient:
        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def post(self, url, headers=None, json=None):
            calls.append(("POST", url))
            return FakeResponse({"orderCreateTransaction": {"id": "ord-1"}})

        async def get(self, url, headers=None, params=None):
            calls.append(("GET", url))
            return FakeResponse({"prices": [{"bids": [{"price": "1.0"}], "asks": [{"price": "1.2"}], "instrument": "EUR_USD"}]})

        async def put(self, url, headers=None, json=None):
            calls.append(("PUT", url))
            return FakeResponse({"ok": True})

    monkeypatch.setattr(master_service.httpx, "AsyncClient", lambda *a, **k: FakeClient())
    async def fake_meta(**_k):
        return {"displayPrecision": 5, "tradeUnitsPrecision": 0}

    monkeypatch.setattr(master_service, "_fetch_oanda_instrument_meta", fake_meta)
    monkeypatch.setattr(master_service, "_upsert_trade_context", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda *_a, **_k: None)
    monkeypatch.setattr(master_service, "_parse_limit_cancel_settings", lambda *_a, **_k: (None, None))

    monkeypatch.setattr(master_service, "timeout_s", 10.0, raising=False)
    monkeypatch.setattr(master_service, "connect_s", 2.0, raising=False)
    monkeypatch.setattr(master_service, "read_s", None, raising=False)

    cfg = {"base_url": "https://api-fxpractice.oanda.com/v3", "account_id": "acc-1", "token": "t", "mode": "demo"}
    monkeypatch.setattr(master_service, "_get_oanda_config", lambda *_a, **_k: cfg)

    asyncio.run(master_service._place_oanda_order({"action": "buy", "symbol": "EUR_USD", "quantity": 1, "account": "demo", "order_type": "market"}, request_id="r1"))
    asyncio.run(master_service._fetch_oanda_mid_price(cfg=cfg, instrument="EUR_USD", mode="demo"))
    asyncio.run(master_service._fetch_oanda_mid_prices_batch(cfg=cfg, instruments=["EUR_USD"]))
    asyncio.run(master_service._cancel_oanda_order(cfg=cfg, order_id="ord-2", mode="demo"))
    asyncio.run(master_service._close_oanda_trade(cfg=cfg, trade_id="trd-3", mode="demo"))

    urls = [url for _, url in calls]
    assert "https://api-fxpractice.oanda.com/v3/accounts/acc-1/orders" in urls
    assert "https://api-fxpractice.oanda.com/v3/accounts/acc-1/pricing" in urls
    assert "https://api-fxpractice.oanda.com/v3/accounts/acc-1/orders/ord-2/cancel" in urls
    assert "https://api-fxpractice.oanda.com/v3/accounts/acc-1/trades/trd-3/close" in urls
    assert all("/v3/v3/" not in url for url in urls)


def test_run_oanda_history_export_sanitizes_html_errors(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("OANDA_API_KEY_DEMO", "demo-key")
    monkeypatch.setenv("OANDA_ACCOUNT_ID_DEMO", "demo-account")
    monkeypatch.setenv("OANDA_API_URL_DEMO", "https://api-fxpractice.oanda.com")
    monkeypatch.setattr(master_service, "OANDA_HISTORY_EXPORT_ROOT", tmp_path)

    async def failing_fetch(**kwargs):
        raise RuntimeError("Failed to fetch transactions: 403 <html>Attention Required | Cloudflare</html>")

    monkeypatch.setattr(master_service, "_fetch_oanda_transactions_window", failing_fetch)

    job = master_service.OandaHistoryJob(
        job_id="job1",
        status="queued",
        created_at=0,
        updated_at=0,
        params={"account": "demo", "period": "week", "complete": False},
    )

    asyncio.run(master_service._run_oanda_history_export(job))

    assert job.status == "error"
    assert job.error == (
        "OANDA history export failed with HTTP 403 from upstream. "
        "Check OANDA history base URL and credentials."
    )


def test_oanda_history_export_status_only_returns_download_when_file_exists(tmp_path: Path):
    missing_path = tmp_path / "missing.csv"
    job = master_service.OandaHistoryJob(
        job_id="job-download",
        status="done",
        created_at=0,
        updated_at=0,
        params={},
        output_path=missing_path,
    )
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    try:
        response = asyncio.run(master_service.oanda_history_export_status(job.job_id))
        payload = response.body.decode("utf-8")
        assert "download_url" not in payload
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


def test_oanda_history_export_filename_contains_account_mode(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    monkeypatch.setenv("OANDA_API_KEY_DEMO", "demo-key")
    monkeypatch.setenv("OANDA_ACCOUNT_ID_DEMO", "demo-account")
    monkeypatch.setenv("OANDA_API_URL_DEMO", "https://api-fxpractice.oanda.com")
    monkeypatch.setenv("OANDA_API_KEY", "live-key")
    monkeypatch.setenv("OANDA_ACCOUNT_ID", "live-account")
    monkeypatch.setenv("OANDA_API_URL_LIVE", "https://api-fxtrade.oanda.com")
    monkeypatch.setattr(master_service, "OANDA_HISTORY_EXPORT_ROOT", tmp_path)

    async def fake_fetch(**kwargs):
        return []

    monkeypatch.setattr(master_service, "_fetch_oanda_transactions_window", fake_fetch)
    monkeypatch.setattr(master_service.oanda_history_exporter, "save_to_csv", lambda tx, path: Path(path).write_text("x"))

    demo_job = master_service.OandaHistoryJob(job_id="demo1", status="queued", created_at=0, updated_at=0, params={"account": "demo", "days": 1})
    live_job = master_service.OandaHistoryJob(job_id="live1", status="queued", created_at=0, updated_at=0, params={"account": "live", "days": 1})
    asyncio.run(master_service._run_oanda_history_export(demo_job))
    asyncio.run(master_service._run_oanda_history_export(live_job))
    assert demo_job.output_path is not None and "demo" in demo_job.output_path.name
    assert live_job.output_path is not None and "live" in live_job.output_path.name
    assert demo_job.output_path.with_suffix(".json").exists()
    assert live_job.output_path.with_suffix(".json").exists()


def test_oanda_history_export_backfill_endpoint_preserves_legacy_workbook_path(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    csv_path = tmp_path / "oanda_history_demo_job.csv"
    csv_path.write_text("TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n589,2026-04-08 19:50:46 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.58217,2550,Buy,0,0.57864,0.58888,,,0,1493.64\n594,2026-04-09 19:35:17 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.58308,-2550,Sell,0,,,,,3.2847,1496.92\n")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo","account_label":"OANDA DEMO"}')
    job = master_service.OandaHistoryJob(job_id="jobbf", status="done", created_at=0, updated_at=0, params={"account": "demo"}, output_path=csv_path)
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_append_oanda_export_rows_to_local_workbook", lambda *_a, **_k: 1)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *_a, **_k: {"ok": True, "message": "Done"})
    try:
        res = asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id))
        payload = res.body.decode("utf-8")
        assert "oanda_export_trades_seen" in payload
        assert "oanda_export_target_workbook" in payload
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


def test_oanda_history_export_backfill_endpoint_invalidates_view_cache(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    csv_path = tmp_path / "oanda_history_demo_job2.csv"
    csv_path.write_text("TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}')
    job = master_service.OandaHistoryJob(job_id="jobbf2", status="done", created_at=0, updated_at=0, params={"account": "demo"}, output_path=csv_path)
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    called = {"n": 0}
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_append_oanda_export_rows_to_local_workbook", lambda *_a, **_k: 0)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *_a, **_k: {"ok": True, "message": "Done"})
    monkeypatch.setattr(master_service, "_invalidate_trading_journal_view_snapshot", lambda: called.__setitem__("n", called["n"] + 1))
    try:
        asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id))
        assert called["n"] == 1
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


def test_oanda_history_backfill_endpoint_accepts_legacy_int_append_stats(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    csv_path = tmp_path / "oanda_history_demo_job_legacy.csv"
    csv_path.write_text("TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}')
    job = master_service.OandaHistoryJob(job_id="legacyint", status="done", created_at=0, updated_at=0, params={"account": "demo"}, output_path=csv_path)
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    called = {"n": 0}
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_append_oanda_export_rows_to_local_workbook", lambda *_a, **_k: 0)
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *_a, **_k: {"ok": True, "message": "Done"})
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda *a, **k: {"balances": [{"label": "OANDA DEMO", "balance_source": "authoritative_trade_balance"}]})
    monkeypatch.setattr(master_service, "_invalidate_trading_journal_view_snapshot", lambda: called.__setitem__("n", called["n"] + 1))
    try:
        res = asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id))
        payload = res.body.decode("utf-8")
        assert '"ok":true' in payload
        assert called["n"] == 1
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


def test_oanda_backfill_endpoint_fails_if_snapshot_still_stale(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    csv_path = tmp_path / "oanda_history_demo_job_stale.csv"
    csv_path.write_text("TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}')
    job = master_service.OandaHistoryJob(job_id="stale", status="done", created_at=0, updated_at=0, params={"account": "demo"}, output_path=csv_path)
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_append_oanda_export_rows_to_local_workbook", lambda *_a, **_k: {"changed": 0, "inserted": 0, "updated": 0})
    monkeypatch.setattr(master_service, "_import_trading_journal_from_sources", lambda *_a, **_k: {"ok": True, "message": "Done"})
    monkeypatch.setattr(master_service, "_build_trading_journal_view_snapshot", lambda *a, **k: {"balances": [{"label": "OANDA DEMO", "balance_source": "cashflow_anchor_plus_trades"}]})
    try:
        res = asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id))
        payload = res.body.decode("utf-8")
        assert '"ok":false' in payload
        assert "OANDA_BACKFILL_NOT_VISIBLE_IN_JOURNAL_SNAPSHOT" in payload
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


def test_snapshot_balance_items_supports_list_and_dict_shapes():
    a = master_service._snapshot_balance_items({"balances": [{"label": "OANDA DEMO"}]})
    b = master_service._snapshot_balance_items({"balances": {"items": [{"label": "OANDA DEMO"}]}})
    assert len(a) == 1
    assert len(b) == 1


def test_trading_journal_page_versions_static_js_and_no_store_headers():
    response = asyncio.run(master_service.trading_journal_page())
    body = response.body.decode("utf-8")
    assert "/static/trading_journal.js?v=" in body
    assert '/static/trading_journal.js"></script>' not in body
    assert "no-store" in str(response.headers.get("Cache-Control") or "")


def test_history_page_versions_static_js_and_no_store_headers():
    response = asyncio.run(master_service.merged_history_page())
    body = response.body.decode("utf-8")
    assert "/static/history_page.js?v=" in body
    assert '/static/history_page.js"></script>' not in body
    assert "no-store" in str(response.headers.get("Cache-Control") or "")


def test_oanda_history_export_backfill_endpoint_returns_failure_on_append_error(monkeypatch: pytest.MonkeyPatch, tmp_path: Path):
    csv_path = tmp_path / "oanda_history_demo_job3.csv"
    csv_path.write_text("TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n589,2026-04-08 19:50:46 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.58217,2550,Buy,0,0.57864,0.58888,,,0,1493.64\n594,2026-04-09 19:35:17 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.58308,-2550,Sell,0,,,,,3.2847,1496.92\n")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}')
    job = master_service.OandaHistoryJob(job_id="jobbf3", status="done", created_at=0, updated_at=0, params={"account": "demo"}, output_path=csv_path)
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: False)
    monkeypatch.setattr(master_service, "_append_oanda_export_rows_to_local_workbook", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("MISSING_XLRD_FOR_XLS")))
    try:
        res = asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id))
        payload = res.body.decode("utf-8")
        assert '"ok":false' in payload
        assert "MISSING_XLRD_FOR_XLS" in payload
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)


@pytest.mark.parametrize("account_mode", ["demo", "live"])
def test_oanda_master_journal_backfill_uses_authoritative_import_and_is_idempotent(
    account_mode: str,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    journal_dir = tmp_path / "journal"
    journal_dir.mkdir()
    master_path = journal_dir / "Trading Journal.xlsx"
    csv_path = tmp_path / f"oanda_history_{account_mode}_job.csv"
    csv_path.write_text(
        "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n"
        "589,2026-04-08 19:50:46 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.58217,2550,Buy,0,0.57864,0.58888,,,0,1493.64\n"
        "594,2026-04-09 19:35:17 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.58308,-2550,Sell,0,,,,,3.2847,1496.92\n",
        encoding="utf-8",
    )
    csv_path.with_suffix(".json").write_text(
        json.dumps({"account_mode": account_mode}),
        encoding="utf-8",
    )
    job = master_service.OandaHistoryJob(
        job_id=f"master-{account_mode}",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": account_mode},
        output_path=csv_path,
    )
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    expected_id = f"oanda_export:{account_mode}:589:594"
    import_modes = []

    def _authoritative_import(_name, _payload, account_mode=None):
        import_modes.append(account_mode)
        if master_path.exists():
            wb = load_workbook(master_path)
            ws = wb["Trade Log"]
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Trade Log"
            ws.append(["Row ID"])
            ws.append(["existing:trade"])
            ws.append(["cashflow:existing"])
        existing_ids = {
            str(ws.cell(row, 1).value or "")
            for row in range(2, ws.max_row + 1)
        }
        if expected_id in existing_ids:
            rows_upserted = 0
            duplicate_rows_merged = 1
        else:
            ws.append([expected_id])
            rows_upserted = 1
            duplicate_rows_merged = 0
        wb.save(master_path)
        wb.close()
        return {
            "ok": True,
            "status_code": 200,
            "message": "Import complete.",
            "rows_parsed": 1,
            "rows_upserted": rows_upserted,
            "duplicate_rows_merged": duplicate_rows_merged,
            "verified_row_ids_count": 1,
        }

    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: master_path)
    monkeypatch.setattr(master_service, "_import_uploaded_trading_journal_file", _authoritative_import)
    monkeypatch.setattr(
        master_service,
        "_append_oanda_export_rows_to_local_workbook",
        lambda *_a, **_k: (_ for _ in ()).throw(AssertionError("legacy append path must not run")),
    )
    monkeypatch.setattr(master_service, "_invalidate_trading_journal_view_snapshot", lambda: None)
    monkeypatch.setattr(
        master_service,
        "_build_trading_journal_view_snapshot",
        lambda *a, **k: {
            "balances": [{
                "label": f"OANDA {account_mode.upper()}",
                "balance": 1496.92,
                "balance_source": "oanda_transaction_export_balance",
            }]
        },
    )
    try:
        first = json.loads(
            asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id)).body.decode("utf-8")
        )
        second = json.loads(
            asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id)).body.decode("utf-8")
        )
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)

    assert first["ok"] is True
    assert first["oanda_export_target_workbook"] == "Trading Journal.xlsx"
    assert first["oanda_export_trades_backfilled"] == 1
    assert first["oanda_export_latest_balance"] == pytest.approx(1496.92)
    assert second["ok"] is True
    assert second["oanda_export_trades_backfilled"] == 0
    assert second["oanda_export_trades_updated"] == 0
    assert second["oanda_export_latest_balance"] == pytest.approx(1496.92)
    assert import_modes == [account_mode, account_mode]
    assert not (journal_dir / "OANDA LIVE.xlsx").exists()
    assert not (journal_dir / "OANDA DEMO.xlsx").exists()
    json.dumps(first)
    json.dumps(second)

    wb = load_workbook(master_path, data_only=True, read_only=True)
    ids = [str(row[0] or "") for row in wb["Trade Log"].iter_rows(min_row=2, values_only=True)]
    wb.close()
    assert ids.count(expected_id) == 1
    assert ids.count("existing:trade") == 1
    assert ids.count("cashflow:existing") == 1


def test_oanda_master_backfill_persists_later_nontrade_balance_through_workbook_reread(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    journal_dir = tmp_path / "journal"
    data_dir = tmp_path / "data"
    journal_dir.mkdir()
    data_dir.mkdir()
    master_path = journal_dir / "Trading Journal.xlsx"
    shutil.copy2(ROOT / "journal" / "Trading Journal.xlsx", master_path)
    source_before = master_service.read_master_journal_source(master_path)
    existing_trade_id = next(
        str(item.get("id"))
        for item in source_before["items"]
        if item.get("row_type") == "trade" and item.get("id")
    )
    existing_cashflow_id = next(
        str(item.get("id"))
        for item in source_before["items"]
        if item.get("row_type") == "cashflow" and item.get("id")
    )
    wb = load_workbook(master_path)
    trade_log = master_service._get_trade_log_sheet(wb, allow_legacy=False)
    trade_headers = master_service._trade_log_header_map(trade_log)
    data_start = master_service._trade_log_data_start_row(trade_log)
    row_id_col = trade_headers["Row ID"]
    keep_ids = {existing_trade_id, existing_cashflow_id}
    keep_values = []
    for row_idx in range(data_start, trade_log.max_row + 1):
        if str(trade_log.cell(row_idx, row_id_col).value or "").strip() in keep_ids:
            keep_values.append(
                [
                    trade_log.cell(row_idx, col_idx).value
                    for col_idx in range(1, trade_log.max_column + 1)
                ]
            )
    assert len(keep_values) == 2
    for offset, values in enumerate(keep_values):
        for col_idx, value in enumerate(values, start=1):
            trade_log.cell(data_start + offset, col_idx).value = value
    trailing_rows = trade_log.max_row - (data_start + len(keep_values) - 1)
    if trailing_rows > 0:
        trade_log.delete_rows(data_start + len(keep_values), trailing_rows)
    wb.save(master_path)
    wb.close()

    csv_path = tmp_path / "oanda_history_demo_later_financing.csv"
    csv_path.write_text(
        "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n"
        "990001,2030-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1000,1000,Buy,0,1.0900,1.1200,,,0,1000.00\n"
        "990002,2030-01-01 11:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.1100,-1000,Sell,0,,,,,10.00,1010.00\n"
        "990003,2030-01-01 12:00:00 AEST,DAILY_FINANCING,DAILY_FINANCING,,,,,,,,-1.25,,,1008.75\n",
        encoding="utf-8",
    )
    csv_path.with_suffix(".json").write_text(
        json.dumps({"account_mode": "demo"}),
        encoding="utf-8",
    )
    job = master_service.OandaHistoryJob(
        job_id="master-demo-later-financing",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": "demo"},
        output_path=csv_path,
    )
    expected_id = "oanda_export:demo:990001:990002"
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job

    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR", journal_dir)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_LOCAL_DIR_EXPLICIT", True)
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SOURCE", "master_journal")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_PATH", data_dir / "trading_journal.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_STATE_PATH", data_dir / "trading_journal_state.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SYNC_STATE_PATH", data_dir / "trading_journal_sync_state.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_IMPORT_CACHE_PATH", data_dir / "trading_journal_import_cache.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_VIEW_CACHE_PATH", data_dir / "trading_journal_view_cache.json")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_SQLITE_PATH", data_dir / "trading_journal.sqlite")
    monkeypatch.setattr(master_service, "TRADING_JOURNAL_RESYNC_CACHE_PATH", data_dir / "trading_journal_resync_cache.json")
    monkeypatch.setattr(master_service, "MONTHLY_AUD_REVALUATION_PATH", data_dir / "monthly_aud_revaluation.json")
    monkeypatch.setattr(master_service, "MONTHLY_AUD_REVALUATION_STATE_PATH", data_dir / "monthly_aud_revaluation_state.json")
    monkeypatch.setattr(master_service, "TRADE_CONTEXTS_PATH", data_dir / "trade_contexts.json")
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_authoritative_enabled", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: master_path)
    monkeypatch.setattr(master_service, "_schedule_dropbox_upload_state_backup", lambda *_a, **_k: None)
    monkeypatch.setattr(
        master_service,
        "_sync_journal_excel_files_to_github",
        lambda *_a, **_k: {
            "github_sync_enabled": False,
            "github_sync_ok": True,
            "github_sync_noop": True,
            "github_sync_error": "",
            "github_sync_commit": "",
        },
    )
    monkeypatch.setattr(master_service, "_TRADING_JOURNAL_CACHE", None)
    monkeypatch.setattr(
        master_service,
        "_TRADING_JOURNAL_VIEW_CACHE",
        {"key": None, "payload": None},
    )
    monkeypatch.setattr(master_service, "_PENDING_MANUAL_SYNC_ROWS", [])
    monkeypatch.setattr(master_service, "_PENDING_MANUAL_SYNC_BALANCES", [])

    try:
        first = json.loads(
            asyncio.run(
                master_service.backfill_oanda_history_export_to_journal(job.job_id)
            ).body.decode("utf-8")
        )
        assert first["ok"] is True, first["error"]
        workbook_after_first = master_service.read_master_journal_source(master_path)
        first_ids = [
            str(item.get("id") or "")
            for item in workbook_after_first["items"]
        ]
        imported_trade = next(
            item
            for item in workbook_after_first["items"]
            if str(item.get("id") or "") == expected_id
        )
        first_balance = next(
            item
            for item in workbook_after_first["balances"]
            if str(item.get("account") or "").upper() == "OANDA DEMO"
        )

        master_service._TRADING_JOURNAL_CACHE = None
        master_service._TRADING_JOURNAL_VIEW_CACHE = {"key": None, "payload": None}
        master_service._PENDING_MANUAL_SYNC_ROWS = []
        master_service._PENDING_MANUAL_SYNC_BALANCES = []
        master_service._invalidate_trading_journal_view_snapshot()
        rebuilt = master_service._build_master_journal_verification_snapshot()
        rebuilt_balance = next(
            item
            for item in master_service._snapshot_balance_items(rebuilt)
            if str(item.get("account") or item.get("label") or "").upper()
            == "OANDA DEMO"
        )

        second = json.loads(
            asyncio.run(
                master_service.backfill_oanda_history_export_to_journal(job.job_id)
            ).body.decode("utf-8")
        )
        workbook_after_second = master_service.read_master_journal_source(master_path)
        second_ids = [
            str(item.get("id") or "")
            for item in workbook_after_second["items"]
        ]
        second_balance = next(
            item
            for item in workbook_after_second["balances"]
            if str(item.get("account") or "").upper() == "OANDA DEMO"
        )
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)

    assert first["oanda_export_rows_persisted"] is True
    assert first["oanda_export_balance_applied"] is True
    assert first["snapshot_visible"] is True
    assert first_ids.count(expected_id) == 1
    assert imported_trade["balance_after_trade"] == pytest.approx(1010.0)
    assert first_balance["balance"] == pytest.approx(1008.75)
    assert first_balance["balance_source"] == "oanda_transaction_export_balance"
    assert rebuilt_balance["balance"] == pytest.approx(1008.75)
    assert rebuilt_balance["balance_source"] == "oanda_transaction_export_balance"

    assert second["ok"] is True
    assert second["oanda_export_rows_persisted"] is True
    assert second["oanda_export_balance_applied"] is True
    assert second["snapshot_visible"] is True
    assert second["oanda_export_trades_backfilled"] == 0
    assert second_ids.count(expected_id) == 1
    assert second_balance["balance"] == pytest.approx(1008.75)
    assert second_balance["balance_source"] == "oanda_transaction_export_balance"
    assert second_ids.count(existing_trade_id) == 1
    assert second_ids.count(existing_cashflow_id) == 1
    assert not (journal_dir / "OANDA LIVE.xlsx").exists()
    assert not (journal_dir / "OANDA DEMO.xlsx").exists()


def test_oanda_csv_parser_honors_explicit_account_mode_without_filename_hint(tmp_path: Path):
    csv_path = tmp_path / "completed_export.csv"
    csv_path.write_text(
        "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n"
        "1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1,1000,Buy,0,1.0,1.2,,,0,1000\n"
        "2,2026-01-01 11:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.2,-1000,Sell,0,,,,,10,1010\n",
        encoding="utf-8",
    )
    rows, _balance = master_service._parse_local_trading_journal_workbook(
        csv_path,
        original_name=csv_path.name,
        account_mode="live",
    )
    assert rows[0]["id"] == "oanda_export:live:1:2"
    assert rows[0]["account_label"] == "OANDA LIVE"


@pytest.mark.parametrize(
    ("import_result", "expected_error"),
    [
        (
            {
                "ok": False,
                "status_code": 423,
                "code": "EXCEL_WORKBOOK_OPEN",
                "message": "Close Trading Journal.xlsx in Excel before importing.",
                "errors": ["workbook_locked"],
            },
            "Close Trading Journal.xlsx",
        ),
        (
            {
                "ok": False,
                "status_code": 500,
                "message": "Workbook sync failed: validation failed",
                "errors": ["validation failed"],
            },
            "Workbook sync failed",
        ),
    ],
)
def test_oanda_master_backfill_surfaces_authoritative_import_failures(
    import_result,
    expected_error,
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    csv_path = tmp_path / "oanda_history_demo_failure.csv"
    csv_path.write_text(
        "TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE\n"
        "1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1,1000,Buy,0,1.0,1.2,,,0,1000\n"
        "2,2026-01-01 11:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.2,-1000,Sell,0,,,,,10,1010\n",
        encoding="utf-8",
    )
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}', encoding="utf-8")
    job = master_service.OandaHistoryJob(
        job_id="master-failure",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": "demo"},
        output_path=csv_path,
    )
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(master_service, "_import_uploaded_trading_journal_file", lambda *_a, **_k: import_result)
    try:
        payload = json.loads(
            asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id)).body.decode("utf-8")
        )
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)
    assert payload["ok"] is False
    assert expected_error in payload["error"]
    assert payload["oanda_export_target_workbook"] == "Trading Journal.xlsx"
    assert payload["sync"] == import_result


def test_oanda_master_backfill_parse_and_verification_failures_are_structured(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
):
    csv_path = tmp_path / "oanda_history_demo_structured.csv"
    csv_path.write_text("header\n", encoding="utf-8")
    csv_path.with_suffix(".json").write_text('{"account_mode":"demo"}', encoding="utf-8")
    job = master_service.OandaHistoryJob(
        job_id="master-structured",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": "demo"},
        output_path=csv_path,
    )
    master_service.OANDA_HISTORY_JOBS[job.job_id] = job
    monkeypatch.setattr(master_service, "_master_journal_single_file_mode", lambda: True)
    monkeypatch.setattr(master_service, "_master_journal_path", lambda: tmp_path / "Trading Journal.xlsx")
    monkeypatch.setattr(
        master_service,
        "_parse_local_trading_journal_workbook",
        lambda *_a, **_k: (_ for _ in ()).throw(ValueError("bad OANDA CSV")),
    )
    try:
        parse_payload = json.loads(
            asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id)).body.decode("utf-8")
        )
        assert parse_payload["ok"] is False
        assert "Failed to parse completed OANDA export" in parse_payload["error"]

        monkeypatch.setattr(
            master_service,
            "_parse_local_trading_journal_workbook",
            lambda *_a, **_k: (
                [{"id": "oanda_export:demo:1:2", "source": "oanda_transaction_export"}],
                {"balance": 1010.0, "as_of": "2026-01-01T11:00:00+10:00"},
            ),
        )
        monkeypatch.setattr(
            master_service,
            "_import_uploaded_trading_journal_file",
            lambda *_a, **_k: {
                "ok": True,
                "rows_upserted": 1,
                "duplicate_rows_merged": 0,
                "rows_persisted": True,
                "balance_applied": True,
                "snapshot_visible": True,
            },
        )
        monkeypatch.setattr(
            master_service,
            "_verify_trade_log_row_ids_in_workbook",
            lambda *_a, **_k: {
                "ok": False,
                "missing_row_ids": ["oanda_export:demo:1:2"],
            },
        )
        verify_payload = json.loads(
            asyncio.run(master_service.backfill_oanda_history_export_to_journal(job.job_id)).body.decode("utf-8")
        )
    finally:
        master_service.OANDA_HISTORY_JOBS.pop(job.job_id, None)
    assert verify_payload["ok"] is False
    assert verify_payload["oanda_export_rows_persisted"] is True
    assert verify_payload["oanda_export_balance_applied"] is True
    assert verify_payload["snapshot_visible"] is True
    assert "verification failed" in verify_payload["error"]
    assert verify_payload["sync"]["oanda_backfill_verification"]["missing_row_ids"] == [
        "oanda_export:demo:1:2"
    ]


import pandas as pd

def _sample_oanda_history_rows():
    return [
        {"TICKET":588,"TRANSACTION DATE":"2026-04-08 19:50:45 AEST","TRANSACTION TYPE":"MARKET_ORDER","DETAILS":"CLIENT_ORDER","INSTRUMENT":"NZD_USD","PRICE":"","UNITS":2550,"DIRECTION":"Buy","SPREAD COST":"","STOP LOSS":0.57864,"TAKE PROFIT":0.58888,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1493.64},
        {"TICKET":589,"TRANSACTION DATE":"2026-04-08 19:50:46 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"NZD_USD","PRICE":0.58217,"UNITS":2550,"DIRECTION":"Buy","SPREAD COST":-0.1,"STOP LOSS":0.57864,"TAKE PROFIT":0.58888,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1493.64},
        {"TICKET":592,"TRANSACTION DATE":"2026-04-09 07:00:00 AEST","TRANSACTION TYPE":"DAILY_FINANCING","DETAILS":"DAILY_FINANCING","INSTRUMENT":"","PRICE":"","UNITS":"","DIRECTION":"","SPREAD COST":"","STOP LOSS":"","TAKE PROFIT":"","FINANCING":-0.1329,"COMMISSION":"","PL":"","BALANCE":1493.51},
        {"TICKET":594,"TRANSACTION DATE":"2026-04-09 19:35:17 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"NZD_USD","PRICE":0.58308,"UNITS":-2550,"DIRECTION":"Sell","SPREAD COST":-0.1,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":3.2847,"BALANCE":1496.92},
    ]

def test_oanda_transaction_history_allocates_daily_financing_to_single_open_trade():
    df = pd.DataFrame(_sample_oanda_history_rows())
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(df, account_mode='demo', account_label='OANDA DEMO', source_path='/tmp/oanda_history_demo.csv')
    assert len(parsed['rows']) == 1
    row = parsed['rows'][0]
    assert row['swap'] == pytest.approx(-0.1329)
    assert row['metrics']['oanda_export_pl'] == pytest.approx(3.2847)
    assert row['net_profit'] == pytest.approx(3.1518)
    assert row['balance_after_trade'] == pytest.approx(1496.92)


def test_oanda_transaction_history_allocates_financing_by_closed_interval_despite_unmatched_noise():
    rows = [
        {"TICKET": 100, "TRANSACTION DATE": "2024-01-01 00:00:00 AEST", "TRANSACTION TYPE": "ORDER_FILL", "DETAILS": "MARKET_ORDER", "INSTRUMENT": "EUR_USD", "PRICE": 1.1, "UNITS": 999, "DIRECTION": "Buy", "SPREAD COST": 0.1, "BALANCE": 1000},
        *_sample_oanda_history_rows(),
    ]
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.DataFrame(rows),
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="/tmp/oanda_history_demo.csv",
    )
    row = next(item for item in parsed["rows"] if item["id"] == "oanda_export:demo:589:594")
    assert row["swap"] == pytest.approx(-0.1329)
    assert row["net_profit"] == pytest.approx(3.1518)
    assert "ambiguous_oanda_financing_allocation:592" not in parsed["warnings"]

def test_oanda_transaction_history_csv_parses_closed_trades():
    rows = _sample_oanda_history_rows() + [
        {"TICKET":598,"TRANSACTION DATE":"2026-04-09 19:37:46 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"NZD_USD","PRICE":0.58323,"UNITS":2916,"DIRECTION":"Buy","SPREAD COST":0,"STOP LOSS":0.58117,"TAKE PROFIT":0.58717,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1496.92},
        {"TICKET":601,"TRANSACTION DATE":"2026-04-10 02:13:06 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"NZD_USD","PRICE":0.58718,"UNITS":-2916,"DIRECTION":"Sell","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":16.1655,"BALANCE":1513.09},
        {"TICKET":604,"TRANSACTION DATE":"2026-04-22 20:51:56 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"NZD_USD","PRICE":0.59116,"UNITS":30821,"DIRECTION":"Buy","SPREAD COST":0,"STOP LOSS":0.59097,"TAKE PROFIT":0.59166,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1513.09},
        {"TICKET":608,"TRANSACTION DATE":"2026-04-22 20:52:36 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"NZD_USD","PRICE":0.59112,"UNITS":-30821,"DIRECTION":"Sell","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":-1.7385,"BALANCE":1511.35},
        {"TICKET":612,"TRANSACTION DATE":"2026-04-28 20:52:52 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"USD_JPY","PRICE":159.605,"UNITS":13319,"DIRECTION":"Sell","SPREAD COST":0,"STOP LOSS":159.681,"TAKE PROFIT":159.426,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1511.35},
        {"TICKET":615,"TRANSACTION DATE":"2026-04-28 21:13:44 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"USD_JPY","PRICE":159.681,"UNITS":-13319,"DIRECTION":"Buy","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":-8.9385,"BALANCE":1502.41},
        {"TICKET":618,"TRANSACTION DATE":"2026-04-30 19:45:59 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"EUR_USD","PRICE":1.16929,"UNITS":6546,"DIRECTION":"Buy","SPREAD COST":0,"STOP LOSS":1.16824,"TAKE PROFIT":1.17148,"FINANCING":"","COMMISSION":"","PL":"","BALANCE":1502.41},
        {"TICKET":622,"TRANSACTION DATE":"2026-04-30 19:46:41 AEST","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"EUR_USD","PRICE":1.16910,"UNITS":-6546,"DIRECTION":"Sell","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":-1.7591,"BALANCE":1500.65},
    ]
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(pd.DataFrame(rows), account_mode='demo', account_label='OANDA DEMO', source_path='/tmp/oanda_history_demo.csv')
    assert len(parsed['rows']) == 5
    final = parsed['rows'][-1]
    assert final['raw_refs']['close_ticket'] == '622'
    assert final['balance_after_trade'] == pytest.approx(1500.65)
    assert final['net_profit'] == pytest.approx(-1.7591)
    assert final['account_label'] == 'OANDA DEMO'


def test_oanda_transaction_history_real_csv_blanks_do_not_crash(tmp_path: Path):
    csv_text = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,PL,BALANCE
588,2026-04-08 19:50:45 AEST,MARKET_ORDER,CLIENT_ORDER,NZD_USD,,2550,Buy,,0.57864,0.58888,,,,1493.64
589,2026-04-08 19:50:46 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.58217,2550,Buy,-0.1,,,,,1493.64
592,2026-04-09 07:00:00 AEST,DAILY_FINANCING,DAILY_FINANCING,,,,,,,,-0.1329,,,1493.51
594,2026-04-09 19:35:17 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.58308,-2550,Sell,-0.1,,,,0,3.2847,1496.92
598,2026-04-09 19:37:46 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.58323,2916,Buy,0,0.58117,0.58717,,, ,1496.92
601,2026-04-10 02:13:06 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.58718,-2916,Sell,0,,,,0,16.1655,1513.09
604,2026-04-22 20:51:56 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.59116,30821,Buy,0,0.59097,0.59166,,,,1513.09
608,2026-04-22 20:52:36 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.59112,-30821,Sell,0,,,,0,-1.7385,1511.35
612,2026-04-28 20:52:52 AEST,ORDER_FILL,MARKET_ORDER,USD_JPY,159.605,13319,Sell,0,159.681,159.426,,,,1511.35
615,2026-04-28 21:13:44 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,USD_JPY,159.681,-13319,Buy,0,,,,0,-8.9385,1502.41
618,2026-04-30 19:45:59 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.16929,6546,Buy,0,1.16824,1.17148,,,,1502.41
622,2026-04-30 19:46:41 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,EUR_USD,1.1691,-6546,Sell,0,,,,0,-1.7591,1500.65
"""
    path = tmp_path / 'oanda_history_demo.csv'
    path.write_text(csv_text)
    df = pd.read_csv(path, encoding='utf-8-sig')
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(df, account_mode='demo', account_label='OANDA DEMO', source_path=str(path))
    assert len(parsed['rows']) == 5
    assert parsed['unmatched_open_fills'] == []
    assert parsed['unmatched_close_fills'] == []
    final = parsed['rows'][-1]
    assert final['raw_refs']['close_ticket'] == '622'
    assert final['balance_after_trade'] == pytest.approx(1500.65)


def test_oanda_transaction_history_converts_to_brisbane_local_time():
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.DataFrame(_sample_oanda_history_rows()),
        account_mode='demo',
        account_label='OANDA DEMO',
        source_path='/tmp/oanda_history_demo.csv',
    )
    row = parsed['rows'][0]
    assert row['open_time'] == '2026-04-08T19:50:46+10:00'
    assert row['close_time'] == '2026-04-09T19:35:17+10:00'


def test_oanda_transaction_history_aedt_converts_to_brisbane_wall_clock():
    rows = [
        {"TICKET":1,"TRANSACTION DATE":"2024-10-28 22:59:57 AEDT","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER","INSTRUMENT":"EUR_USD","PRICE":1.1,"UNITS":1000,"DIRECTION":"Buy","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":"","PL":"","BALANCE":1000},
        {"TICKET":2,"TRANSACTION DATE":"2024-10-28 22:59:58 AEDT","TRANSACTION TYPE":"ORDER_FILL","DETAILS":"MARKET_ORDER_TRADE_CLOSE","INSTRUMENT":"EUR_USD","PRICE":1.2,"UNITS":-1000,"DIRECTION":"Sell","SPREAD COST":0,"STOP LOSS":"","TAKE PROFIT":"","FINANCING":"","COMMISSION":0,"PL":1.0,"BALANCE":1001},
    ]
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(pd.DataFrame(rows), account_mode='demo', account_label='OANDA DEMO', source_path='/tmp/oanda_history_demo.csv')
    row = parsed['rows'][0]
    assert row['open_time'] == '2024-10-28T21:59:57+10:00'
    assert row['close_time'] == '2024-10-28T21:59:58+10:00'


def test_oanda_transaction_history_uses_client_order_stop_target():
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(pd.DataFrame(_sample_oanda_history_rows()), account_mode='demo', account_label='OANDA DEMO', source_path='/tmp/oanda_history_demo.csv')
    row = parsed['rows'][0]
    assert row['stop_loss'] == pytest.approx(0.57864)
    assert row['take_profit'] == pytest.approx(0.58888)


def test_bybit_history_export_status_download_url_requires_existing_file(tmp_path: Path):
    missing = tmp_path / "missing-bybit.csv"
    missing_job = master_service.BybitHistoryJob(
        job_id="bybit-missing",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": "demo"},
        output_path=missing,
    )
    existing = tmp_path / "existing-bybit.csv"
    existing.write_text("id\n1\n", encoding="utf-8")
    present_job = master_service.BybitHistoryJob(
        job_id="bybit-present",
        status="done",
        created_at=0,
        updated_at=0,
        params={"account": "demo"},
        output_path=existing,
    )
    master_service.BYBIT_HISTORY_JOBS[missing_job.job_id] = missing_job
    master_service.BYBIT_HISTORY_JOBS[present_job.job_id] = present_job
    try:
        missing_payload = asyncio.run(master_service.bybit_history_export_status(missing_job.job_id)).body.decode("utf-8")
        present_payload = asyncio.run(master_service.bybit_history_export_status(present_job.job_id)).body.decode("utf-8")
        assert "download_url" not in missing_payload
        assert f"/api/bybit-history/export/{present_job.job_id}/download" in present_payload
    finally:
        master_service.BYBIT_HISTORY_JOBS.pop(missing_job.job_id, None)
        master_service.BYBIT_HISTORY_JOBS.pop(present_job.job_id, None)


def test_coinspot_history_export_status_download_url_requires_existing_file(tmp_path: Path):
    missing = tmp_path / "missing-coinspot.zip"
    missing_job = master_service.CoinspotHistoryJob(
        job_id="coinspot-missing",
        status="done",
        created_at=0,
        updated_at=0,
        params={},
        output_path=missing,
    )
    existing = tmp_path / "existing-coinspot.zip"
    existing.write_bytes(b"PK\x03\x04")
    present_job = master_service.CoinspotHistoryJob(
        job_id="coinspot-present",
        status="done",
        created_at=0,
        updated_at=0,
        params={},
        output_path=existing,
    )
    master_service.COINSPOT_HISTORY_JOBS[missing_job.job_id] = missing_job
    master_service.COINSPOT_HISTORY_JOBS[present_job.job_id] = present_job
    try:
        missing_payload = asyncio.run(master_service.coinspot_history_export_status(missing_job.job_id)).body.decode("utf-8")
        present_payload = asyncio.run(master_service.coinspot_history_export_status(present_job.job_id)).body.decode("utf-8")
        assert "download_url" not in missing_payload
        assert f"/api/coinspot-history/export/{present_job.job_id}/download" in present_payload
    finally:
        master_service.COINSPOT_HISTORY_JOBS.pop(missing_job.job_id, None)
        master_service.COINSPOT_HISTORY_JOBS.pop(present_job.job_id, None)


def test_oanda_demo_raw_history_spread_cost_not_commission_and_balance_authoritative():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
625,2026-05-26 16:54:33 AEST,MARKET_ORDER,CLIENT_ORDER,EUR_USD,1.16372,6393.00,Buy,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.65
626,2026-05-26 16:54:34 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.16372,6393.00,Buy,0.4460,,,,0.0000,0.0000,0.0000,0.00000,1500.65
627,2026-05-26 16:54:35 AEST,STOP_LOSS_ORDER,CLIENT_ORDER,EUR_USD,1.16272,-6393.00,Sell,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.65
628,2026-05-26 16:54:35 AEST,TAKE_PROFIT_ORDER,CLIENT_ORDER,EUR_USD,1.16472,-6393.00,Sell,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.65
629,2026-05-26 16:55:30 AEST,MARKET_ORDER,CLIENT_ORDER,EUR_USD,1.16365,-6393.00,Sell,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.65
630,2026-05-26 16:55:31 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,EUR_USD,1.16365,6393.00,Sell,0.4906,,,,0.0000,0.0000,0.0000,-0.45050,1500.20
631,2026-05-26 16:55:32 AEST,ORDER_CANCEL,LINKED_TRADE_CLOSED,EUR_USD,,,,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.20
632,2026-05-26 16:55:33 AEST,ORDER_CANCEL,LINKED_TRADE_CLOSED,EUR_USD,,,,0.0000,,,,0.0000,0.0000,0.0000,0.00000,1500.20
"""
    df = master_service.pd.read_csv(master_service.io.StringIO(csv))
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        df,
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="oanda_demo.csv",
    )

    rows = parsed["rows"]
    assert len(rows) == 1
    row = rows[0]
    assert row["id"] == "oanda_export:demo:626:630"
    assert row["commission"] in (None, 0, 0.0)
    assert row["net_profit"] == pytest.approx(-0.4505)
    assert row["balance_after_trade"] == pytest.approx(1500.20)
    assert parsed["account_balance"]["balance"] == pytest.approx(1500.20)
    assert (row.get("metrics") or {}).get("oanda_open_spread_cost") == pytest.approx(0.4460)
    assert (row.get("metrics") or {}).get("oanda_close_spread_cost") == pytest.approx(0.4906)
    assert (row.get("metrics") or {}).get("oanda_total_spread_cost") == pytest.approx(0.9366)
    assert row["commission"] != pytest.approx(0.9366)


def test_oanda_demo_attached_style_spread_costs_never_become_commission():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
604,2026-04-22 20:51:56 AEST,ORDER_FILL,MARKET_ORDER,NZD_USD,0.59116,30821,Buy,3.0127,0.59097,0.59166,0,0,0,0,0,1513.09
608,2026-04-22 20:52:36 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,NZD_USD,0.59112,30821,Sell,2.7972,,,0,0,0,0,-1.7385,1511.35
618,2026-04-30 19:45:59 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.16929,6546,Buy,0.3667,1.16824,1.17148,0,0,0,0,0,1502.41
622,2026-04-30 19:46:41 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,EUR_USD,1.16910,6546,Sell,0.4125,,,0,0,0,0,-1.7591,1500.65
626,2026-05-26 16:54:34 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.16370,6393,Buy,0.4460,1.16260,1.16593,0,0,0,0,0,1500.65
630,2026-05-26 16:55:31 AEST,ORDER_FILL,MARKET_ORDER_TRADE_CLOSE,EUR_USD,1.16365,6393,Sell,0.4906,,,0,0,0,0,-0.4505,1500.20
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="oanda_history_demo.csv",
    )
    by_id = {row["id"]: row for row in parsed["rows"]}

    row_618 = by_id["oanda_export:demo:618:622"]
    assert row_618["commission"] in (None, 0, 0.0)
    assert row_618["metrics"]["oanda_total_spread_cost"] == pytest.approx(0.7792)
    assert row_618["net_profit"] == pytest.approx(-1.7591)
    assert row_618["balance_after_trade"] == pytest.approx(1500.65)

    row_604 = by_id["oanda_export:demo:604:608"]
    assert row_604["commission"] in (None, 0, 0.0)
    assert row_604["metrics"]["oanda_total_spread_cost"] == pytest.approx(5.8099)

    row_626 = by_id["oanda_export:demo:626:630"]
    assert row_626["commission"] in (None, 0, 0.0)
    assert row_626["metrics"]["oanda_total_spread_cost"] == pytest.approx(0.9366)


def test_oanda_live_market_if_touched_open_and_order_closes_parse_canonical_rows():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
100,2025-01-01 09:59:58 AEST,MARKET_IF_TOUCHED_ORDER,CLIENT_ORDER,EUR_USD,,1200,Sell,,1.1060,1.0980,,0,0,0,,1000
101,2025-01-01 10:00:00 AEST,ORDER_FILL,MARKET_IF_TOUCHED_ORDER,EUR_USD,1.1020,1200,Sell,0.31,,,,0,0,0,0,1000
102,2025-01-01 17:00:00 AEST,DAILY_FINANCING,DAILY_FINANCING,,,,,,,,-0.12,0,0,0,,999.88
103,2025-01-01 18:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.0980,1200,Buy,0.28,,,,0,0,0,7.25,1007.13
201,2025-01-02 10:00:00 AEST,MARKET_IF_TOUCHED_ORDER,CLIENT_ORDER,USD_CHF,,900,Buy,,0.8950,0.9020,,0,0,0,,1007.13
202,2025-01-02 10:01:00 AEST,ORDER_FILL,MARKET_IF_TOUCHED_ORDER,USD_CHF,0.8990,900,Buy,0.11,,,,0,0,0,0,1007.13
203,2025-01-02 11:00:00 AEST,ORDER_FILL,STOP_LOSS_ORDER,USD_CHF,0.8950,900,Sell,0.12,,,,0,0,0,-5.50,1001.63
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="live",
        account_label="OANDA LIVE",
        source_path="oanda_history_live.csv",
    )
    by_id = {row["id"]: row for row in parsed["rows"]}

    take_profit = by_id["oanda_export:live:101:103"]
    assert take_profit["commission"] in (None, 0, 0.0)
    assert take_profit["fees"] in (None, 0, 0.0)
    assert take_profit["swap"] == pytest.approx(-0.12)
    assert take_profit["net_profit"] == pytest.approx(7.13)
    assert take_profit["metrics"]["oanda_total_spread_cost"] == pytest.approx(0.59)
    assert take_profit["stop_loss"] == pytest.approx(1.1060)
    assert take_profit["take_profit"] == pytest.approx(1.0980)

    stop_loss = by_id["oanda_export:live:202:203"]
    assert stop_loss["commission"] in (None, 0, 0.0)
    assert stop_loss["net_profit"] == pytest.approx(-5.50)
    assert stop_loss["metrics"]["oanda_total_spread_cost"] == pytest.approx(0.23)


def test_oanda_live_market_order_position_closeout_parses_as_close():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
420,2025-09-05 03:05:05 AEST,ORDER_FILL,MARKET_ORDER,USD_CHF,0.80699,4130,Sell,0.4999,0.80899,0.80299,0,0,0,0,0,1460.29
423,2025-09-05 07:00:00 AEST,DAILY_FINANCING,DAILY_FINANCING,,,,,,,,-0.9771,0,0,0,,1459.31
425,2025-09-05 18:05:54 AEST,ORDER_FILL,MARKET_ORDER_POSITION_CLOSEOUT,USD_CHF,0.80372,4130,Buy,0.5214,,,,0,0,0,25.5489,1484.86
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="live",
        account_label="OANDA LIVE",
        source_path="oanda_history_live.csv",
    )

    assert len(parsed["rows"]) == 1
    row = parsed["rows"][0]
    assert row["id"] == "oanda_export:live:420:425"
    assert row["raw_refs"]["close_details"] == "MARKET_ORDER_POSITION_CLOSEOUT"
    assert row["commission"] in (None, 0, 0.0)
    assert row["swap"] == pytest.approx(-0.9771)
    assert row["net_profit"] == pytest.approx(24.5718)
    assert row["metrics"]["oanda_total_spread_cost"] == pytest.approx(1.0213)
    assert row["commission"] != pytest.approx(1.0213)


def test_oanda_transaction_history_positive_commission_reduces_net_profit_once():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,CONVERSION RATE,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1000,1000,Buy,0,1.0990,1.1020,1.5,0,0,0,0,0,1000
2,2026-01-01 11:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.1020,1000,Sell,0,,,,0,0.5,0,0,10,1009.5
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="oanda_history_demo.csv",
    )

    row = parsed["rows"][0]
    assert row["commission"] == pytest.approx(0.5)
    assert row["net_profit"] == pytest.approx(9.5)


def test_oanda_transaction_history_negative_export_commission_is_not_double_deducted():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,CONVERSION RATE,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1000,1000,Buy,0,1.0990,1.1020,1.5,0,0,0,0,0,1000
2,2026-01-01 11:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.1020,1000,Sell,0,,,,0,-0.5,0,0,10,1009.5
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="oanda_history_demo.csv",
    )

    row = parsed["rows"][0]
    assert row["commission"] == pytest.approx(0.5)
    assert row["net_profit"] == pytest.approx(9.5)


def test_oanda_transaction_history_guaranteed_execution_fees_reduce_net_profit():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,CONVERSION RATE,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
1,2026-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1000,1000,Buy,0,1.0990,1.1020,1.5,0,0,0,0,0,1000
2,2026-01-01 11:00:00 AEST,ORDER_FILL,STOP_LOSS_ORDER,EUR_USD,1.0990,1000,Sell,0,,,,0,0,0.25,0.10,10,1009.65
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="demo",
        account_label="OANDA DEMO",
        source_path="oanda_history_demo.csv",
    )

    row = parsed["rows"][0]
    assert row["commission"] == pytest.approx(0.35)
    assert row["net_profit"] == pytest.approx(9.65)


def test_oanda_live_ambiguous_daily_financing_is_not_allocated():
    csv = """TICKET,TRANSACTION DATE,TRANSACTION TYPE,DETAILS,INSTRUMENT,PRICE,UNITS,DIRECTION,SPREAD COST,STOP LOSS,TAKE PROFIT,FINANCING,COMMISSION,GSL FEE,GSL PREMIUM,PL,BALANCE
1,2025-01-01 10:00:00 AEST,ORDER_FILL,MARKET_ORDER,EUR_USD,1.1000,1000,Buy,0.10,,,,0,0,0,0,1000
2,2025-01-01 10:05:00 AEST,ORDER_FILL,MARKET_ORDER,USD_CHF,0.9000,2000,Buy,0.20,,,,0,0,0,0,1000
3,2025-01-01 12:00:00 AEST,DAILY_FINANCING,DAILY_FINANCING,,,,,,,,-0.33,0,0,0,,999.67
4,2025-01-01 13:00:00 AEST,ORDER_FILL,TAKE_PROFIT_ORDER,EUR_USD,1.1010,1000,Sell,0.11,,,,0,0,0,1.50,1001.17
5,2025-01-01 13:05:00 AEST,ORDER_FILL,STOP_LOSS_ORDER,USD_CHF,0.8990,2000,Sell,0.21,,,,0,0,0,-2.00,999.17
"""
    parsed = master_service._journal_rows_from_oanda_transaction_history_frame(
        pd.read_csv(master_service.io.StringIO(csv)),
        account_mode="live",
        account_label="OANDA LIVE",
        source_path="oanda_history_live.csv",
    )

    assert "ambiguous_oanda_financing_allocation:3" in parsed["warnings"]
    assert len(parsed["rows"]) == 2
    assert all(row.get("swap") in (None, "", 0, 0.0) for row in parsed["rows"])
    assert [row["commission"] for row in parsed["rows"]] == [None, None]
