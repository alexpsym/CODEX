from __future__ import annotations
from collections import Counter, defaultdict, OrderedDict
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, localcontext
from pathlib import Path
from typing import Any, Collection, Dict, List, Mapping, Optional, Sequence, Set, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.formula import Tokenizer
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Pane, Selection
import hashlib
from openpyxl.styles import PatternFill, Border, Side, Alignment, Color
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
import calendar
from copy import copy, deepcopy
import json
import math
import os
import posixpath
import re
import zipfile
import xml.etree.ElementTree as ET
from xml.sax.saxutils import escape as _xml_escape
from urllib.parse import quote, unquote
from zoneinfo import ZoneInfo
from uuid import uuid4

REPO_ROOT = Path(__file__).resolve().parents[1]
TRADE_LOG_SHEET = "Trade Log"
LEGACY_ALL_TRADES_SHEET = "All Trades"
STATS1_SHEET = "STATS1"
STATS2_SHEET = "STATS2"
SYMBOLS_SHEET = "SYMBOLS"
TARGET_R_METADATA_SHEET = "_Target R Metadata"
LEGACY_DASHBOARD_SHEET = "Dashboard"
LEGACY_INSTRUMENT_AVERAGES_SHEET = "Instrument Averages"
# Backward-compatible aliases (do not remove yet; external imports may still reference these).
ALL_TRADES_SHEET = LEGACY_ALL_TRADES_SHEET
LEGACY_TRADE_LOG_SHEET = LEGACY_ALL_TRADES_SHEET
SHEET_ORDER=[STATS1_SHEET, STATS2_SHEET, SYMBOLS_SHEET, TRADE_LOG_SHEET, "P&L Calendar"]
REPORT_YEARLY_SHEET = "YEARLY REPORT"
REPORT_START_YEAR = 2018
REPORT_MIN_END_YEAR = 2026
TRADE_NUMBER_HEADER = "Trade Number"
STOP_RECOMMENDATION_HEADER = "Stop Loss Recommendation"
TARGET_RECOMMENDATION_HEADER = "Target Recommendation"
RECOMMENDATION_DISPLAY_HEADER = "Recommendation"
TARGET_RECOMMENDATION_NO_WINS = "No eligible winning trades"
TARGET_RECOMMENDATION_NO_LOSSES = "No eligible losing trades"
TARGET_R_RECOMMENDATION_FLOOR = Decimal("1.50")
# Backward-compatible alias for callers that still import the old name.
TARGET_RECOMMENDATION_INSUFFICIENT = ""
RECOMMENDATION_REASON_TEXT = {
    "exact_tie_goal_preference_decrease": "exact tie, so a small decrease is preferred",
}
MOVE_TO_FIELD_MAP = {
    "Move to Break Even Time": "move_to_break_even_time",
    "Move to Break Even Duration": "move_to_break_even_duration",
    "Move to Break Even Trigger Price": "move_to_break_even_trigger_price",
    "Move to Break Even Distance From Entry %": "move_to_break_even_distance_from_entry_pct",
    "Move to Break Even Distance From Exit %": "move_to_break_even_distance_from_exit_pct",
    "Move to Profit Time": "move_to_profit_time",
    "Move to Profit Duration": "move_to_profit_duration",
    "Move to Profit Trigger Price": "move_to_profit_trigger_price",
    "Move to Profit Distance From Entry %": "move_to_profit_distance_from_entry_pct",
    "Move to Profit Distance From Exit %": "move_to_profit_distance_from_exit_pct",
}


class IncrementalWorkbookUpdateNotEligible(RuntimeError):
    """Raised when a workbook cannot be updated safely with the local row path."""

    def __init__(
        self,
        *args: object,
        unsafe_full_rebuild: bool = False,
    ) -> None:
        super().__init__(*args)
        self.unsafe_full_rebuild = bool(unsafe_full_rebuild)


MOVE_TO_TIME_FIELDS = {
    "move_to_break_even_time",
    "move_to_profit_time",
}
QUALITY_ANALYSIS_FIELD_MAP = {
    "Pattern": "pattern",
    "EMA": "ema",
    "VWAP": "vwap",
    "ATHS/ATLS": "aths_atls",
    "Order": "order_type",
    "Round Number": "round_number",
    "Spiked Out": "spiked_out",
    "Close Stopout": "close_stopout",
    "Near Perfect Entry": "near_perfect_entry",
    "Near Win": "near_win",
    "Early Close": "early_close",
}
TRADE_NUMBER_FIELD_MAP = {TRADE_NUMBER_HEADER: "trade_number"}
TRADE_LOG_MANUAL_FIELD_MAP = {**TRADE_NUMBER_FIELD_MAP, **MOVE_TO_FIELD_MAP, **QUALITY_ANALYSIS_FIELD_MAP}
EDITABLE_COLS=["Test",*TRADE_LOG_MANUAL_FIELD_MAP.keys(),"Setup","Timeframe","Breakeven","Notes"]

PRE_VWAP_TRADE_LOG_HEADERS_V1 = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    "Target Price", "Target Distance", "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", *MOVE_TO_FIELD_MAP.keys(),
    "Test", "Pattern", "EMA", "ATHS/ATLS", "Order", "Round Number",
    "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close",
    "Setup", "Timeframe", "Breakeven", "Notes", "Cashflow Amount",
    "Cashflow New Balance", "Currency", "Row Type", "Row ID",
]
TRADE_LOG_HEADERS_V1 = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    "Target Price", "Target Distance", "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", *MOVE_TO_FIELD_MAP.keys(),
    "Test", "Pattern", "EMA", "VWAP", "ATHS/ATLS", "Order", "Round Number",
    "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close",
    "Setup", "Timeframe", "Breakeven", "Notes", "Cashflow Amount",
    "Cashflow New Balance", "Currency", "Row Type", "Row ID",
]
PRE_VWAP_TRADE_LOG_HEADERS = [TRADE_NUMBER_HEADER, *PRE_VWAP_TRADE_LOG_HEADERS_V1]
TRADE_LOG_HEADERS = [TRADE_NUMBER_HEADER, *TRADE_LOG_HEADERS_V1]
PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1 = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    STOP_RECOMMENDATION_HEADER, "Target Price", "Target Distance",
    TARGET_RECOMMENDATION_HEADER, "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", *MOVE_TO_FIELD_MAP.keys(),
    "Test", "Pattern", "EMA", "ATHS/ATLS", "Order", "Round Number",
    "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close",
    "Setup", "Timeframe", "Breakeven", "Notes", "Cashflow Amount",
    "Cashflow New Balance", "Currency", "Row Type", "Row ID",
]
RECOMMENDATION_TRADE_LOG_HEADERS_V1 = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    STOP_RECOMMENDATION_HEADER, "Target Price", "Target Distance",
    TARGET_RECOMMENDATION_HEADER, "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", *MOVE_TO_FIELD_MAP.keys(),
    "Test", "Pattern", "EMA", "VWAP", "ATHS/ATLS", "Order", "Round Number",
    "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close",
    "Setup", "Timeframe", "Breakeven", "Notes", "Cashflow Amount",
    "Cashflow New Balance", "Currency", "Row Type", "Row ID",
]
PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS = [TRADE_NUMBER_HEADER, *PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1]
RECOMMENDATION_TRADE_LOG_HEADERS = [TRADE_NUMBER_HEADER, *RECOMMENDATION_TRADE_LOG_HEADERS_V1]
PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1 = PRE_VWAP_TRADE_LOG_HEADERS_V1
PRE_RECOMMENDATION_TRADE_LOG_HEADERS = PRE_VWAP_TRADE_LOG_HEADERS
PRE_MOVE_TRADE_LOG_HEADERS = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    "Target Price", "Target Distance", "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", "Test", "Pattern", "EMA",
    "ATHS/ATLS", "Order", "Round Number", "Spiked Out", "Close",
    "Stop Out", "Near Perfect Entry", "Near Win", "Early Close",
    "Setup", "Timeframe", "Breakeven", "Notes", "Cashflow Amount",
    "Cashflow New Balance", "Currency", "Row Type", "Row ID",
]
OLD_TRADE_LOG_HEADERS = [
    "Open Time", "Close Time", "Account", "Symbol", "Side", "Qty",
    "Entry Price", "Exit Price", "Stop Loss Price", "Stop Loss Distance",
    "Target Price", "Target Distance", "Commission", "Net P/L",
    "Profit %", "R-Multiple", "Balance After",
    "Trade Duration (DD:HH:MM:SS)", "Test", "Setup", "Timeframe",
    "Breakeven", "Notes", "Cashflow Amount", "Cashflow New Balance",
    "Currency", "Row Type", "Row ID",
]
TRADE_LOG_HEADER_ROWS = 3
TRADE_LOG_FILTER_HEADER_ROW = 3
TRADE_LOG_DATA_START_ROW = 4
TRADE_LOG_DATA_ROW_HEIGHT = 15
AVG_WINNING_DURATION_HEADER = "Avg winning duration (DD:HH:MM:SS)"
AVG_LOSING_DURATION_HEADER = "Avg losing duration (DD:HH:MM:SS)"
MOVE_TO_BREAK_EVEN_HEADERS = list(MOVE_TO_FIELD_MAP.keys())[:5]
MOVE_TO_PROFIT_HEADERS = list(MOVE_TO_FIELD_MAP.keys())[5:]
MOVE_TO_SUBHEADERS = ["Time", "Duration", "Trigger Price", "Distance From Entry %", "Distance From Exit %"]
DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL = "Average Move to Break Even"
DASHBOARD_MOVE_TO_PROFIT_LABEL = "Average Move to Profit"
LEGACY_INSTRUMENT_AVERAGES_HEADERS = [
    "Symbol", "Class", "Trades", "Wins", "Losses", "Break-even", "Longs", "Long wins",
    "Long losses", "Long break-even", "Shorts", "Short wins", "Short losses", "Short break-even",
    "Move to break even", "Move to profit", "Pattern", "EMA", "All-time highs", "All-time lows",
    "Order", "Round number", "Spiked out", "Close stop out", "Near perfect entry", "Near win",
    "Early close", "Most traded timeframe", "R Multiple", "Net P/L %", "Avg P/L %", "Win Rate %",
    "Avg stop % (W)", "Avg stop % (L)", "Avg target % (W)", "Avg target % (L)",
    "Shortest duration (DD:HH:MM:SS)", "Avg duration (DD:HH:MM:SS)", "Longest duration (DD:HH:MM:SS)",
]
INSTRUMENT_AVERAGES_GROUP_HEADER_ROW = 1
INSTRUMENT_AVERAGES_FILTER_HEADER_ROW = 2
INSTRUMENT_AVERAGES_DATA_START_ROW = 3
SYMBOLS_HEADERS = [
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[:16],
    "Most Traded Pattern", "Most Traded EMA",
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[18:20],
    "Market", "Limit",
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[21:28],
    "Most Profitable Timeframe", "Least Profitable Timeframe",
    "Net R Multiple",
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[29:32],
    "Avg stop %",
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[32:34],
    STOP_RECOMMENDATION_HEADER,
    "Avg target %",
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[34:36],
    TARGET_RECOMMENDATION_HEADER,
    *LEGACY_INSTRUMENT_AVERAGES_HEADERS[36:],
    AVG_WINNING_DURATION_HEADER,
    AVG_LOSING_DURATION_HEADER,
]
INSTRUMENT_AVERAGES_HEADERS = SYMBOLS_HEADERS
REPORT_METRIC_LABELS = [
    "Trades",
    "Wins",
    "Losses",
    "Break-even",
    "Test",
    "Win rate",
    "Net P/L",
    "Gross percent gain",
    "Gross percent loss",
    "Gross IR gain",
    "Gross IR loss",
    "Best Win Streak",
    "Worst Losing Streak",
    "Percentage expectancy",
    "R expectancy",
    "Avg stop %",
    "Avg target %",
    "Min stop %",
    "Source",
    "Max stop %",
    "Source",
    "Min target %",
    "Source",
    "Max target %",
    "Source",
    "Avg duration (DD:HH:MM:SS)",
    "Move to Break Even (DD:HH:MM:SS)",
    "Move to Profit (DD:HH:MM:SS)",
    "Max win %",
    "Source",
    "Max loss %",
    "Source",
    "Max R loss",
    "Source",
    "Max R win",
    "Source",
    "Shortest (DD:HH:MM:SS)",
    "Source",
    "Longest (DD:HH:MM:SS)",
    "Source",
    "Winners",
    "Avg stop %",
    "Avg target %",
    "Percentage expectancy",
    "R expectancy",
    "Losers",
    "Avg stop %",
    "Avg target %",
    "Percentage expectancy",
    "R expectancy",
    "Drawdown",
    "Max drawdown",
    "Avg drawdown",
    "Longs",
    "Long wins",
    "Long losses",
    "Long break-even",
    "Shorts",
    "Short wins",
    "Short losses",
    "Short break-even",
]
LIGHT_GREY_FILL_RGB = "FFEAF2F8"
JOURNAL_DISPLAY_TZ = ZoneInfo("Australia/Brisbane")
DURATION_NUMBER_FORMAT = r"00\:00\:00\:00"
LEGACY_DURATION_NUMBER_FORMAT_TOKENS = ("DAYS", "HOURS", "MINUTES", "SECONDS")
DEFAULT_FOREX_ROOT = Path.home() / "Dropbox" / "FOREX"
DEFAULT_CRYPTO_ROOT = Path.home() / "Dropbox" / "CRYPTO"
_TRADE_FOLDER_INDEX_CACHE: Dict[Tuple[str, str], Dict[str, List[Path]]] = {}


def _sheet_by_alias(wb: Workbook, canonical: str, legacy_aliases: List[str] | Tuple[str, ...]):
    if canonical in wb.sheetnames:
        return wb[canonical]
    for legacy in legacy_aliases:
        if legacy not in wb.sheetnames:
            continue
        if canonical not in wb.sheetnames:
            wb[legacy].title = canonical
            return wb[canonical]
        return wb[legacy]
    raise RuntimeError(f"Master Journal is missing required sheet '{canonical}'.")


def _stats1_sheet(wb: Workbook):
    return _sheet_by_alias(wb, STATS1_SHEET, (LEGACY_DASHBOARD_SHEET,))


def _stats2_sheet(wb: Workbook, required: bool = False):
    if STATS2_SHEET in wb.sheetnames:
        return wb[STATS2_SHEET]
    if required:
        raise RuntimeError(f"Master Journal is missing required sheet '{STATS2_SHEET}'.")
    return None


def _symbols_sheet(wb: Workbook):
    return _sheet_by_alias(wb, SYMBOLS_SHEET, (LEGACY_INSTRUMENT_AVERAGES_SHEET,))


def _activate_user_facing_sheet(wb: Workbook) -> None:
    if STATS1_SHEET not in wb.sheetnames:
        return
    wb.active = wb.sheetnames.index(STATS1_SHEET)
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = ws.title == STATS1_SHEET


def _migrate_analysis_sheet_names(wb: Workbook, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    before = list(wb.sheetnames)
    _stats1_sheet(wb)
    _symbols_sheet(wb)
    for canonical, legacy in (
        (STATS1_SHEET, LEGACY_DASHBOARD_SHEET),
        (SYMBOLS_SHEET, LEGACY_INSTRUMENT_AVERAGES_SHEET),
    ):
        if canonical in wb.sheetnames and legacy in wb.sheetnames:
            wb.remove(wb[legacy])
            diagnostics.setdefault("removed_duplicate_legacy_analysis_sheets", []).append(legacy)
    if before != wb.sheetnames:
        diagnostics["migrated_analysis_sheet_names"] = {
            "before": before,
            "after": list(wb.sheetnames),
        }

def _canonical_journal_timeframe(value: Any) -> str:
    text = " ".join(str(value or "").strip().split())
    key = text.lower().replace("-", " ")
    key = " ".join(key.split())
    aliases = {"1m":"1MIN","1 min":"1MIN","1 minute":"1MIN","5m":"5MIN","15m":"15MIN","30m":"30MIN","1h":"1H","4h":"4H","1d":"1D","1w":"1W","1mo":"1MO","1 month":"1MO"}
    return aliases.get(key, text)


TIMEFRAME_ORDER = ("1MIN", "5MIN", "15MIN", "30MIN", "1H", "4H", "DAILY", "WEEKLY", "MONTHLY")


def _canonical_analysis_timeframe(value: Any) -> str:
    canonical = _canonical_journal_timeframe(value).upper()
    return {
        "1D": "DAILY",
        "DAILY": "DAILY",
        "1W": "WEEKLY",
        "WEEKLY": "WEEKLY",
        "1MO": "MONTHLY",
        "MONTHLY": "MONTHLY",
    }.get(canonical, canonical if canonical in TIMEFRAME_ORDER else "")
PROFIT_FILL = "C6EFCE"
PROFIT_FONT = "006100"
LOSS_FILL = "FFC7CE"
LOSS_FONT = "9C0006"

# Generator-owned presentation for statistics metrics.  Row-label fonts,
# alignment, indentation, and the STATS1 border grouping are deliberately not
# part of this policy: those are user-owned presentation in preservation mode.
SEMANTIC_FORMAT_POLICY = {
    "section": "General",
    "count": "0",
    "pct": "0.00%",
    "r": '0.000"R"',
    "duration": "General",
    "text": "General",
    "date": "General",
    "streak_date": "General",
    "drawdown_date": "General",
    "commission": None,
}

def _semantic_fill_rgb(cell) -> str:
    fill = getattr(cell, "fill", None)
    color = getattr(fill, "fgColor", None)
    rgb = str(getattr(color, "rgb", "") or "")
    return rgb[-6:].upper() if getattr(fill, "fill_type", None) == "solid" and rgb else ""

def _apply_full_cell_semantic_fill(cell, semantic: str | None) -> None:
    """Apply a direct profit/loss fill while preserving all unrelated cell styling."""
    if semantic not in {"profit", "loss"}:
        return
    fill_rgb, font_rgb = (PROFIT_FILL, PROFIT_FONT) if semantic == "profit" else (LOSS_FILL, LOSS_FONT)
    font = copy(cell.font)
    font.color = font_rgb
    if cell.font != font:
        cell.font = font
    fill = PatternFill(fill_type="solid", fgColor=fill_rgb)
    if cell.fill != fill:
        cell.fill = fill

def _clear_generated_semantic_fill(cell) -> None:
    if _semantic_fill_rgb(cell) not in {PROFIT_FILL, LOSS_FILL}:
        return
    font = copy(cell.font)
    if getattr(font.color, "type", None) == "rgb" and str(font.color.rgb or "")[-6:].upper() in {PROFIT_FONT, LOSS_FONT}:
        font.color = "000000"
    if cell.font != font:
        cell.font = font
    empty_fill = PatternFill()
    if cell.fill != empty_fill:
        cell.fill = empty_fill

def _clear_cell_fill(cell) -> None:
    empty_fill = PatternFill()
    if cell.fill != empty_fill:
        cell.fill = empty_fill


def _semantic_metric_number_format(
    kind: str,
    value: Any,
    *,
    currency: str | None = None,
) -> str:
    """Return the explicit Excel format for a managed statistics value."""
    if isinstance(value, str) and value:
        return "General"
    if kind == "commission":
        return _currency_number_format(currency) if currency else "General"
    return SEMANTIC_FORMAT_POLICY.get(kind, "General") or "General"


def _reset_managed_metric_fill(cell) -> None:
    """Restore Excel's genuine no-fill state and remove generated font colour."""
    font = copy(cell.font)
    color = getattr(font, "color", None)
    if (
        getattr(color, "type", None) == "rgb"
        and str(getattr(color, "rgb", "") or "")[-6:].upper()
        in {PROFIT_FONT, LOSS_FONT}
    ):
        font.color = "FF000000"
    if cell.font != font:
        cell.font = font
    empty_fill = PatternFill()
    if cell.fill != empty_fill:
        cell.fill = empty_fill


def _apply_semantic_metric_policy(
    cell,
    *,
    kind: str,
    semantic: str | None = None,
    currency: str | None = None,
    populated_only_fill: bool = True,
) -> None:
    """Apply number format and direct fill from semantic meaning, never position."""
    number_format = _semantic_metric_number_format(
        kind,
        cell.value,
        currency=currency,
    )
    if cell.number_format != number_format:
        cell.number_format = number_format
    _reset_managed_metric_fill(cell)

    # Managed statistics values are body text.  This removes the known style
    # shifts that made duration/count cells bold while retaining the existing
    # font family, size, underline, and other authored font properties.
    font = copy(cell.font)
    font.bold = False
    font.italic = False
    if cell.font != font:
        cell.font = font

    populated = cell.value not in (None, "")
    if populated_only_fill and not populated:
        return
    if semantic in {"profit", "loss"}:
        _apply_full_cell_semantic_fill(cell, semantic)
    elif semantic == "auto":
        value = _as_float(cell.value)
        if value is not None and value != 0:
            _apply_full_cell_semantic_fill(
                cell,
                "profit" if value > 0 else "loss",
            )


def _reindex_cell_style_registry(wb: Workbook) -> Dict[str, int]:
    """Repair openpyxl's style lookup after in-place StyleArray mutations.

    Style descriptors mutate a cell's StyleArray, which may also be an object
    already indexed by the loaded workbook.  Its hash lookup then becomes
    stale and openpyxl can append byte-identical cellXfs during save.  Rebuild
    only the lookup dictionary against the existing order: no style is
    deleted, remapped, or renumbered.
    """
    styles = wb._cell_styles
    registry: Dict[Any, int] = {}
    for index, style in enumerate(styles):
        registry.setdefault(style, index)
    styles._dict = registry
    styles.clean = True
    return {
        "declared": len(styles),
        "unique": len(registry),
        "duplicates": len(styles) - len(registry),
    }

def _apply_sign_based_full_cell_fill(cell) -> None:
    value = _as_float(cell.value)
    if value is None or value == 0:
        _clear_generated_semantic_fill(cell)
    else:
        _apply_full_cell_semantic_fill(cell, "profit" if value > 0 else "loss")

LEADER_LABEL_TO_KEY = {
    "overall most wins": "most_wins_instrument",
    "overall most losses": "most_losses_instrument",
    "fx most wins": "fx_most_wins_instrument",
    "fx most losses": "fx_most_losses_instrument",
    "crypto most wins": "crypto_most_wins_instrument",
    "crypto most losses": "crypto_most_losses_instrument",
}

def _get_all_trades_sheet(wb: Workbook, *, allow_legacy: bool = True):
    has_trade_log = TRADE_LOG_SHEET in wb.sheetnames
    has_legacy_all_trades = LEGACY_ALL_TRADES_SHEET in wb.sheetnames
    if has_trade_log and has_legacy_all_trades:
        raise RuntimeError("Master Journal has ambiguous trade sheets: both 'Trade Log' and legacy 'All Trades' exist.")
    if has_trade_log:
        return wb[TRADE_LOG_SHEET]
    if allow_legacy and has_legacy_all_trades:
        return wb[LEGACY_ALL_TRADES_SHEET]
    raise RuntimeError("Master Journal is missing required Trade Log sheet.")

def _get_trade_log_sheet(wb: Workbook, *, allow_legacy: bool = True):
    return _get_all_trades_sheet(wb, allow_legacy=allow_legacy)

def _migrate_legacy_trade_log_sheet_name(wb: Workbook, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    has_trade_log = TRADE_LOG_SHEET in wb.sheetnames
    has_legacy_all_trades = LEGACY_ALL_TRADES_SHEET in wb.sheetnames
    if has_trade_log and has_legacy_all_trades:
        raise RuntimeError("Master Journal has ambiguous trade sheets: both 'Trade Log' and legacy 'All Trades' exist.")
    if has_trade_log:
        return
    if has_legacy_all_trades:
        wb[LEGACY_ALL_TRADES_SHEET].title = TRADE_LOG_SHEET
        diagnostics["migrated_trade_log_sheet"] = True
        return
    raise RuntimeError("Master Journal is missing required Trade Log sheet.")

def _remove_legacy_trade_meta_sheet(wb: Workbook, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if "_Trade Meta" in wb.sheetnames:
        wb.remove(wb["_Trade Meta"])
        diagnostics["removed_legacy_trade_meta"] = True

def _repair_legacy_instrument_averages_freeze_pane(wb: Workbook, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if LEGACY_INSTRUMENT_AVERAGES_SHEET not in wb.sheetnames:
        return
    ws = wb[LEGACY_INSTRUMENT_AVERAGES_SHEET]
    previous = str(ws.freeze_panes or "")
    if previous != "B3":
        ws.freeze_panes = "B3"
        diagnostics["repaired_instrument_averages_freeze_pane"] = True
        diagnostics["previous_instrument_averages_freeze_pane"] = previous


def _ensure_symbols_freeze_panes(wb_or_ws: Workbook | Any, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if hasattr(wb_or_ws, "sheetnames"):
        if SYMBOLS_SHEET not in wb_or_ws.sheetnames:
            return
        ws = wb_or_ws[SYMBOLS_SHEET]
    else:
        ws = wb_or_ws
        if getattr(ws, "title", "") not in {SYMBOLS_SHEET, LEGACY_INSTRUMENT_AVERAGES_SHEET}:
            return
    previous = str(ws.freeze_panes or "")
    view = ws.sheet_view
    pane = getattr(view, "pane", None)
    def _selection_view_state(selection: Any) -> Dict[str, Any]:
        pane_name = getattr(selection, "pane", None)
        active_cell = getattr(selection, "activeCell", None)
        sqref = str(getattr(selection, "sqref", "") or "")
        if pane_name in {"topRight", "bottomLeft"} and active_cell in (None, "A1") and sqref in ("", "A1"):
            active_cell = None
            sqref = ""
        return {"pane": pane_name, "activeCell": active_cell, "sqref": sqref}

    previous_pane = {
        "freeze_panes": previous,
        "xSplit": getattr(pane, "xSplit", None),
        "ySplit": getattr(pane, "ySplit", None),
        "topLeftCell": getattr(pane, "topLeftCell", None),
        "activePane": getattr(pane, "activePane", None),
        "state": getattr(pane, "state", None),
        "selection": [
            _selection_view_state(selection)
            for selection in (getattr(view, "selection", None) or [])
        ],
    }
    canonical_selection = [
        Selection(pane="topRight", activeCell=None, sqref=None),
        Selection(pane="bottomLeft", activeCell=None, sqref=None),
        Selection(pane="bottomRight", activeCell="A1", sqref="A1"),
    ]
    canonical_state = {
        "freeze_panes": "B3",
        "xSplit": 1.0,
        "ySplit": 2.0,
        "topLeftCell": "B3",
        "activePane": "bottomRight",
        "state": "frozen",
        "selection": [
            {"pane": "topRight", "activeCell": None, "sqref": ""},
            {"pane": "bottomLeft", "activeCell": None, "sqref": ""},
            {"pane": "bottomRight", "activeCell": "A1", "sqref": "A1"},
        ],
    }
    view.pane = Pane(xSplit=1, ySplit=2, topLeftCell="B3", activePane="bottomRight", state="frozen")
    view.selection = canonical_selection
    key = "symbols" if getattr(ws, "title", "") == SYMBOLS_SHEET else "instrument_averages"
    if previous_pane != canonical_state:
        diagnostics[f"repaired_{key}_freeze_pane"] = True
        diagnostics[f"previous_{key}_freeze_pane"] = previous
        diagnostics[f"previous_{key}_sheet_view"] = previous_pane


def _instrument_averages_header_row(ws) -> int:
    if str(ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, 1).value or "").strip().lower() == "symbol":
        return INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
    if str(ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, 1).value or "").strip().lower() == "symbol":
        return INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
    return 1


def _instrument_averages_data_start_row(ws) -> int:
    return INSTRUMENT_AVERAGES_DATA_START_ROW if _instrument_averages_header_row(ws) == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW else 2


def _vertical_header_anchor_value(ws, col: int) -> str:
    for merged in ws.merged_cells.ranges:
        if merged.min_col == col and merged.max_col == col and merged.min_row <= 1 and merged.max_row >= 2:
            return str(ws.cell(merged.min_row, col).value or "").strip()
    return ""


def _instrument_averages_header_map(ws) -> Dict[str, int]:
    header_row = _instrument_averages_header_row(ws)
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(header_row, col).value or "").strip()
        if (
            header_row == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
            and value.casefold() == "total"
        ):
            group = _symbols_group_for_column(ws, col).strip().casefold()
            if group == "longs":
                value = "Longs"
            elif group == "shorts":
                value = "Shorts"
        if not value and header_row == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
            value = _vertical_header_anchor_value(ws, col) or str(ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col).value or "").strip()
        if value:
            headers[value] = col
    aliases = {
        "Shortest (DD:HH:MM:SS)": "Shortest duration (DD:HH:MM:SS)",
        "Longest (DD:HH:MM:SS)": "Longest duration (DD:HH:MM:SS)",
    }
    for alias, canonical in aliases.items():
        if alias in headers and canonical not in headers:
            headers[canonical] = headers[alias]
    for col in range(1, ws.max_column + 1):
        canonical = _symbols_canonical_header_for_column(ws, col)
        if canonical and canonical not in headers:
            headers[canonical] = col
    return headers


SYMBOLS_GROUP_HEADER_LABELS = {
    "longs", "shorts", "order", "timeframe", "p/l", "pl", "stops", "targets",
}
SYMBOLS_GROUP_WIDTHS = {
    "longs": 4,
    "shorts": 4,
    "order": 2,
    "timeframe": 12,
    "p/l": 4,
    "pl": 4,
    "stops": 4,
    "targets": 4,
}
DURATION_SYMBOLS_HEADERS = (
    "Shortest duration (DD:HH:MM:SS)",
    "Avg duration (DD:HH:MM:SS)",
    "Longest duration (DD:HH:MM:SS)",
    AVG_WINNING_DURATION_HEADER,
    AVG_LOSING_DURATION_HEADER,
)
DURATION_SYMBOLS_METRIC_KEYS = {
    "shortestdurationddhhmmss": "shortest_duration",
    "shortestddhhmmss": "shortest_duration",
    "avgdurationddhhmmss": "avg_duration",
    "averagedurationddhhmmss": "avg_duration",
    "longestdurationddhhmmss": "longest_duration",
    "longestddhhmmss": "longest_duration",
    "avgwinningdurationddhhmmss": "avg_winning_duration",
    "averagewinningdurationddhhmmss": "avg_winning_duration",
    "avgwinnerdurationddhhmmss": "avg_winning_duration",
    "avglosingdurationddhhmmss": "avg_losing_duration",
    "averagelosingdurationddhhmmss": "avg_losing_duration",
    "avgloserdurationddhhmmss": "avg_losing_duration",
}

# Exact manually drawn SYMBOLS A1:AT2 border sides from the 772eedb workbook.
# Each four-character code is left/right/top/bottom: thin, medium, or none.
SYMBOLS_MANUAL_HEADER_BORDER_CODES = (
    "tntt mmmt mmmt mnmt mmmt nmmt mtmt ttmt ttmt tmmt mtmt ttmt ttmt tmmt "
    "mmmt mmmt mmmt mmmt mmmt mmmt mtmt ntmt mmmt mmmt mmmt mmmt mmmt mmmt "
    "mtmt ttmt tmmt mtmt ttmt tmmt mmmt mtmt nnmt nnmt ntmt mtmt nnmt nnmt "
    "ntmt nttt tttt tttt",
    "tnnt mmnt mmnt mnnt mmnt nmnt mtnt ttnt ttnt tmnt mtnt ttnt ttnt tmnt "
    "mmnt mmnt mmnt mmnt mmnt mmnt mttt tmtt mmtt mmtt mmtt mmtt mmtt mmtt "
    "mttt tttt tmtt mttt tttt tmtt mmtt mttt tttt tttt tmtt mttt tttt tttt "
    "tmtt ntnt ttnt ttnt",
)
KNOWN_FLATTENED_SYMBOLS_HEADER_LAYOUT_FINGERPRINTS = {
    "145ab5e0c1d88e8ff6114f83c2b8d97eb09fde58bfb689b85680f8a905780dc7": "40f2ef3",
    "775d5e9f5059e4b32f71a4d173e15b9a105e7c9cecf29035a29c2867374f698d": "5c88d4a",
}


def _trade_log_visible_header(header: str) -> str:
    if header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER}:
        return RECOMMENDATION_DISPLAY_HEADER
    return header


def _symbols_visible_header(header: str) -> str:
    if header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER}:
        return RECOMMENDATION_DISPLAY_HEADER
    return header


def _norm_header_token(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().casefold())


def _symbols_group_for_column(ws, col: int) -> str:
    for merged in ws.merged_cells.ranges:
        if merged.min_row == 1 and merged.max_row == 1 and merged.min_col <= col <= merged.max_col and merged.max_col > merged.min_col:
            return str(ws.cell(1, merged.min_col).value or "").strip()
    active_group = ""
    active_col = 0
    for cursor in range(1, col + 1):
        text = str(ws.cell(1, cursor).value or "").strip()
        if text and text.casefold() in SYMBOLS_GROUP_HEADER_LABELS:
            active_group = text
            active_col = cursor
    if active_group:
        width = SYMBOLS_GROUP_WIDTHS.get(active_group.casefold())
        if width and active_col <= col <= active_col + width - 1:
            return active_group
    for cursor in range(1, col + 1):
        text = str(ws.cell(1, cursor).value or "").strip()
        if not text:
            continue
        token = text.casefold()
        width = SYMBOLS_GROUP_WIDTHS.get(token)
        if width and cursor <= col <= cursor + width - 1:
            return text
    return ""


def _symbols_metric_key_for_column(ws, col: int) -> str:
    header = str(ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, col).value or "").strip()
    if not header:
        header = _vertical_header_anchor_value(ws, col)
    group = _symbols_group_for_column(ws, col).casefold()
    token = _norm_header_token(header)
    duration_key = DURATION_SYMBOLS_METRIC_KEYS.get(token)
    if duration_key and group in {"", "duration"}:
        return duration_key
    if group == "timeframe":
        return {
            "1m": "timeframe_1m",
            "1min": "timeframe_1m",
            "5m": "timeframe_5m",
            "5min": "timeframe_5m",
            "15m": "timeframe_15m",
            "15min": "timeframe_15m",
            "30m": "timeframe_30m",
            "30min": "timeframe_30m",
            "1h": "timeframe_1h",
            "4h": "timeframe_4h",
            "daily": "timeframe_daily",
            "weekly": "timeframe_weekly",
            "monthly": "timeframe_monthly",
        }.get(token, "")
    if group == "stops":
        return {
            "avgstopall": "avg_stop_pct",
            "avgstoppctall": "avg_stop_pct",
            "minstop": "min_stop_pct",
            "minstoppct": "min_stop_pct",
            "min": "min_stop_pct",
            "avgstop": "avg_stop_pct",
            "avgstoppct": "avg_stop_pct",
            "avg": "avg_stop_pct",
            "avgstopw": "avg_stop_pct_winners",
            "avgstoppctw": "avg_stop_pct_winners",
            "avgw": "avg_stop_pct_winners",
            "avgstopl": "avg_stop_pct_losers",
            "avgstoppctl": "avg_stop_pct_losers",
            "avgl": "avg_stop_pct_losers",
            "maxstop": "max_stop_pct",
            "maxstoppct": "max_stop_pct",
            "max": "max_stop_pct",
            "recommendation": "stop_recommendation",
            "stoplossrecommendation": "stop_recommendation",
        }.get(token, "")
    if group == "targets":
        return {
            "avgtargetall": "avg_target_pct",
            "avgtargetpctall": "avg_target_pct",
            "mintarget": "min_target_pct",
            "mintargetpct": "min_target_pct",
            "min": "min_target_pct",
            "avgtarget": "avg_target_pct",
            "avgtargetpct": "avg_target_pct",
            "avg": "avg_target_pct",
            "avgtargetw": "avg_target_pct_winners",
            "avgtargetpctw": "avg_target_pct_winners",
            "avgw": "avg_target_pct_winners",
            "avgtargetl": "avg_target_pct_losers",
            "avgtargetpctl": "avg_target_pct_losers",
            "avgl": "avg_target_pct_losers",
            "maxtarget": "max_target_pct",
            "maxtargetpct": "max_target_pct",
            "max": "max_target_pct",
            "recommendation": "target_recommendation",
            "targetrecommendation": "target_recommendation",
        }.get(token, "")
    return ""


def _symbols_canonical_header_for_column(ws, col: int) -> str:
    key = _symbols_metric_key_for_column(ws, col)
    return {
        "timeframe_1m": "1M",
        "timeframe_5m": "5M",
        "timeframe_15m": "15M",
        "timeframe_30m": "30M",
        "timeframe_1h": "1H",
        "timeframe_4h": "4H",
        "timeframe_daily": "DAILY",
        "timeframe_weekly": "WEEKLY",
        "timeframe_monthly": "MONTHLY",
        "min_stop_pct": "Min stop %",
        "avg_stop_pct": "Avg stop %",
        "max_stop_pct": "Max stop %",
        "avg_stop_pct_winners": "Avg stop % (W)",
        "avg_stop_pct_losers": "Avg stop % (L)",
        "min_target_pct": "Min target %",
        "avg_target_pct": "Avg target %",
        "max_target_pct": "Max target %",
        "avg_target_pct_winners": "Avg target % (W)",
        "avg_target_pct_losers": "Avg target % (L)",
        "stop_recommendation": STOP_RECOMMENDATION_HEADER,
        "target_recommendation": TARGET_RECOMMENDATION_HEADER,
        "shortest_duration": "Shortest duration (DD:HH:MM:SS)",
        "avg_duration": "Avg duration (DD:HH:MM:SS)",
        "longest_duration": "Longest duration (DD:HH:MM:SS)",
        "avg_winning_duration": AVG_WINNING_DURATION_HEADER,
        "avg_losing_duration": AVG_LOSING_DURATION_HEADER,
    }.get(key, "")


def _apply_symbols_filter_header_layout(ws) -> None:
    header_row = _instrument_averages_header_row(ws)
    if header_row != INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
        return
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 36)
    minimum_widths = {
        "Most Traded Pattern": 18,
        "Most Traded EMA": 16,
        "Most traded timeframe": 18,
        "Most Profitable Timeframe": 20,
        "Least Profitable Timeframe": 20,
        "Shortest duration (DD:HH:MM:SS)": 43,
        "Avg duration (DD:HH:MM:SS)": 43,
        "Longest duration (DD:HH:MM:SS)": 43,
        AVG_WINNING_DURATION_HEADER: 43,
        AVG_LOSING_DURATION_HEADER: 43,
        "Move to break even": 16,
        "Move to profit": 14,
        STOP_RECOMMENDATION_HEADER: 18,
        TARGET_RECOMMENDATION_HEADER: 18,
    }
    headers = _instrument_averages_header_map(ws)
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        if cell.value in (None, ""):
            continue
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "center"
        alignment.horizontal = "left"
        cell.alignment = alignment
    for header, minimum in minimum_widths.items():
        col = headers.get(header)
        if not col:
            continue
        letter = get_column_letter(col)
        current = ws.column_dimensions[letter].width or 8
        if current < minimum:
            ws.column_dimensions[letter].width = minimum


def _set_instrument_averages_auto_filter_to_populated_range(ws) -> None:
    header_row = _instrument_averages_header_row(ws)
    headers = _instrument_averages_header_map(ws)
    last_col = max(headers.values()) if headers else ws.max_column
    last_row = header_row
    for row in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            last_row = row
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{max(header_row, last_row)}"


def _unmerge_instrument_averages_header_rows(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        if (
            merged.min_row <= INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
            and merged.max_row >= INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
        ):
            ws.unmerge_cells(str(merged))
    max_col = max(ws.max_column, len(INSTRUMENT_AVERAGES_HEADERS))
    for row in range(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW + 1):
        for col in range(1, max_col + 1):
            cell = ws._cells.get((row, col))
            if cell is not None and cell.__class__.__name__ == "MergedCell":
                del ws._cells[(row, col)]


def _write_instrument_averages_headers(ws, *, preserve_freeze: bool = False) -> None:
    _unmerge_instrument_averages_header_rows(ws)
    for col in range(1, len(INSTRUMENT_AVERAGES_HEADERS) + 1):
        ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col).value = None
        ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, col).value = _symbols_visible_header(INSTRUMENT_AVERAGES_HEADERS[col - 1])
    group_specs = (
        ("Order", "Market", "Limit"),
        ("Stops", "Avg stop %", STOP_RECOMMENDATION_HEADER),
        ("Targets", "Avg target %", TARGET_RECOMMENDATION_HEADER),
    )
    for label, first_header, last_header in group_specs:
        start = INSTRUMENT_AVERAGES_HEADERS.index(first_header) + 1
        end = INSTRUMENT_AVERAGES_HEADERS.index(last_header) + 1
        ws.merge_cells(
            start_row=INSTRUMENT_AVERAGES_GROUP_HEADER_ROW,
            start_column=start,
            end_row=INSTRUMENT_AVERAGES_GROUP_HEADER_ROW,
            end_column=end,
        )
        ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, start).value = label
    _style_header_row(ws, INSTRUMENT_AVERAGES_GROUP_HEADER_ROW)
    _style_header_row(ws, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW)
    for _label, first_header, _last_header in group_specs:
        start = INSTRUMENT_AVERAGES_HEADERS.index(first_header) + 1
        ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, start).alignment = Alignment(horizontal="center")
    ws.freeze_panes = "B3"
    ws.auto_filter.ref = (
        f"A{INSTRUMENT_AVERAGES_FILTER_HEADER_ROW}:"
        f"{get_column_letter(len(INSTRUMENT_AVERAGES_HEADERS))}{max(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, ws.max_row)}"
    )
    _repair_symbols_header_merges_preserving_layout(ws)
    _ensure_symbols_duration_header_layout(ws, migrate_data_boundaries=True)
    _apply_symbols_filter_header_layout(ws)


def _symbols_exact_vertical_header_merge(ws, col: int) -> bool:
    return any(
        rng.min_col == col
        and rng.max_col == col
        and rng.min_row == INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
        and rng.max_row == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
        for rng in ws.merged_cells.ranges
    )


def _visible_border_side(side: Side | None) -> bool:
    return bool(side and side.style)


def _symbols_internal_duration_right_side(ws, row: int, fallback_col: int) -> Side:
    source = ws.cell(row, fallback_col).border.right if fallback_col else None
    if _visible_border_side(source):
        return copy(source)
    return Side(style="thin", color="FFD1D5DB")


def _symbols_outer_duration_right_side(ws, row: int, current_outer_col: int, previous_outer_col: int) -> Side:
    for col in (current_outer_col, previous_outer_col):
        if not col:
            continue
        side = ws.cell(row, col).border.right
        if _visible_border_side(side) and side.style in {"medium", "thick", "double"}:
            return copy(side)
    return Side(style="thick", color="FFD1D5DB")


def _symbols_manual_header_border(code: str) -> Border:
    def side(token: str) -> Side:
        if token == "t":
            return Side(style="thin", color="FFD1D5DB")
        if token == "m":
            return Side(style="medium", color=Color(indexed=64))
        if token == "n":
            return Side()
        raise ValueError(f"unsupported SYMBOLS border token: {token!r}")

    if len(code) != 4:
        raise ValueError(f"invalid SYMBOLS border code: {code!r}")
    return Border(
        left=side(code[0]),
        right=side(code[1]),
        top=side(code[2]),
        bottom=side(code[3]),
        diagonal=Side(),
        diagonalUp=False,
        diagonalDown=False,
        outline=True,
    )


def _symbols_current_header_schema_matches(ws) -> bool:
    if _instrument_averages_header_row(ws) != INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
        return False
    headers = _instrument_averages_header_map(ws)
    return all(
        headers.get(header) == col
        for col, header in enumerate(INSTRUMENT_AVERAGES_HEADERS, start=1)
    )


def _symbols_border_color_signature(color) -> List[Any] | None:
    if color is None:
        return None
    if color.type == "rgb":
        return ["rgb", str(color.rgb)]
    if color.type == "indexed":
        return ["indexed", int(color.indexed)]
    if color.type == "theme":
        return ["theme", int(color.theme), float(color.tint)]
    if color.type == "auto":
        return ["auto", bool(color.auto)]
    return [str(color.type), str(color.value)]


def _symbols_border_side_signature(side: Side | None) -> List[Any] | None:
    if side is None:
        return None
    return [side.style, _symbols_border_color_signature(side.color)]


def _symbols_border_object_signature(border: Border) -> Dict[str, Any]:
    return {
        "left": _symbols_border_side_signature(border.left),
        "right": _symbols_border_side_signature(border.right),
        "top": _symbols_border_side_signature(border.top),
        "bottom": _symbols_border_side_signature(border.bottom),
        "diagonal": _symbols_border_side_signature(border.diagonal),
        "vertical": _symbols_border_side_signature(border.vertical),
        "horizontal": _symbols_border_side_signature(border.horizontal),
        "start": _symbols_border_side_signature(border.start),
        "end": _symbols_border_side_signature(border.end),
        "diagonal_direction": border.diagonal_direction,
        "diagonalUp": bool(border.diagonalUp),
        "diagonalDown": bool(border.diagonalDown),
        "outline": bool(border.outline),
    }


def _symbols_header_border_merge_fingerprint(ws) -> str:
    last_col = len(INSTRUMENT_AVERAGES_HEADERS)
    if ws.max_column < last_col:
        return ""
    payload = {
        "borders": [
            _symbols_border_object_signature(ws.cell(row, col).border)
            for row in (INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW)
            for col in range(1, last_col + 1)
        ],
        "merges": sorted(
            str(merged)
            for merged in ws.merged_cells.ranges
            if merged.min_row <= INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
            and merged.max_row >= INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
            and merged.min_col <= last_col
            and merged.max_col >= 1
        ),
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _repair_symbols_manual_header_borders(
    ws,
    diagnostics: Dict[str, Any] | None = None,
    *,
    legacy_layout_fingerprint: str | None = None,
) -> None:
    if not _symbols_current_header_schema_matches(ws):
        return
    fingerprint = legacy_layout_fingerprint or _symbols_header_border_merge_fingerprint(ws)
    legacy_layout = KNOWN_FLATTENED_SYMBOLS_HEADER_LAYOUT_FINGERPRINTS.get(fingerprint)
    if not legacy_layout:
        return
    headers = _instrument_averages_header_map(ws)

    repaired: List[str] = []
    for row, encoded_row in enumerate(SYMBOLS_MANUAL_HEADER_BORDER_CODES, start=1):
        codes = encoded_row.split()
        if len(codes) != 46:
            raise ValueError("SYMBOLS manual border oracle must cover A1:AT2")
        for col, code in enumerate(codes, start=1):
            cell = ws.cell(row, col)
            expected = _symbols_manual_header_border(code)
            if cell.border != expected:
                cell.border = expected
                repaired.append(cell.coordinate)

    template_col = headers["Longest duration (DD:HH:MM:SS)"]
    for header in (AVG_WINNING_DURATION_HEADER, AVG_LOSING_DURATION_HEADER):
        target_col = headers[header]
        for row in (INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW):
            source = ws.cell(row, template_col)
            target = ws.cell(row, target_col)
            expected = copy(source.border)
            if target.border != expected:
                target.border = expected
                repaired.append(target.coordinate)

    if diagnostics is not None and repaired:
        diagnostics["symbols_manual_header_borders_repaired"] = repaired
        diagnostics["symbols_manual_header_borders_legacy_layout"] = legacy_layout


def _ensure_symbols_duration_data_boundary(ws) -> None:
    headers = _instrument_averages_header_map(ws)
    longest_col = headers.get("Longest duration (DD:HH:MM:SS)")
    winning_col = headers.get(AVG_WINNING_DURATION_HEADER)
    losing_col = headers.get(AVG_LOSING_DURATION_HEADER)
    if not (longest_col and winning_col and losing_col):
        return
    avg_col = headers.get("Avg duration (DD:HH:MM:SS)") or max(1, longest_col - 1)
    data_start = _instrument_averages_data_start_row(ws)
    for row in range(data_start, max(data_start, ws.max_row) + 1):
        outer_right = _symbols_outer_duration_right_side(ws, row, losing_col, longest_col)
        internal_right = _symbols_internal_duration_right_side(ws, row, avg_col)
        for col in (longest_col, winning_col):
            ws.cell(row, col).border = _border_with_side(
                ws.cell(row, col).border,
                right=copy(internal_right),
            )
        ws.cell(row, losing_col).border = _border_with_side(
            ws.cell(row, losing_col).border,
            right=outer_right,
        )


def _ensure_symbols_duration_header_layout(
    ws,
    diagnostics: Dict[str, Any] | None = None,
    *,
    migrate_data_boundaries: bool = False,
) -> None:
    if _instrument_averages_header_row(ws) != INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
        return
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    headers = _instrument_averages_header_map(ws)
    duration_cols = [
        (header, headers.get(header))
        for header in DURATION_SYMBOLS_HEADERS
        if headers.get(header)
    ]
    if not duration_cols:
        return
    legacy_layout_fingerprint = ""
    if _symbols_current_header_schema_matches(ws):
        candidate = _symbols_header_border_merge_fingerprint(ws)
        if candidate in KNOWN_FLATTENED_SYMBOLS_HEADER_LAYOUT_FINGERPRINTS:
            legacy_layout_fingerprint = candidate
    normalize_duration_widths = bool(
        migrate_data_boundaries or legacy_layout_fingerprint
    )
    col_set = {col for _header, col in duration_cols if col}
    removed_merges: List[str] = []
    for merged in list(ws.merged_cells.ranges):
        intersects_duration = any(merged.min_col <= col <= merged.max_col for col in col_set)
        if not intersects_duration:
            continue
        exact_single = (
            merged.min_col == merged.max_col
            and merged.min_row == INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
            and merged.max_row == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
        )
        if exact_single:
            continue
        if (
            merged.min_row <= INSTRUMENT_AVERAGES_FILTER_HEADER_ROW
            and merged.max_row >= INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
        ):
            removed_merges.append(str(merged))
            ws.unmerge_cells(str(merged))

    created_merges: List[str] = []
    for header, col in duration_cols:
        if not col:
            continue
        label = _symbols_visible_header(header)
        if _symbols_exact_vertical_header_merge(ws, col):
            ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col).value = label
        else:
            row1 = ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, col)
            row2 = ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, col)
            if row1.value in (None, "", "Duration") and row2.value not in (None, ""):
                _copy_cell_style(row2, row1)
            row1.value = label
            row2.value = None
            ws.merge_cells(
                start_row=INSTRUMENT_AVERAGES_GROUP_HEADER_ROW,
                start_column=col,
                end_row=INSTRUMENT_AVERAGES_FILTER_HEADER_ROW,
                end_column=col,
            )
            created_merges.append(f"{get_column_letter(col)}1:{get_column_letter(col)}2")
        letter = get_column_letter(col)
        current_width = ws.column_dimensions[letter].width or 0
        if normalize_duration_widths and current_width < 43:
            ws.column_dimensions[letter].width = 43

    if migrate_data_boundaries or legacy_layout_fingerprint:
        _ensure_symbols_duration_data_boundary(ws)
        diagnostics["symbols_duration_data_boundaries_migrated"] = True
    if legacy_layout_fingerprint:
        _repair_symbols_manual_header_borders(
            ws,
            diagnostics,
            legacy_layout_fingerprint=legacy_layout_fingerprint,
        )
    if removed_merges:
        diagnostics.setdefault("symbols_duration_group_merges_removed", []).extend(removed_merges)
    if created_merges:
        diagnostics.setdefault("symbols_duration_vertical_merges_added", []).extend(created_merges)


def _ensure_symbols_schema(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    header_row = _instrument_averages_header_row(ws)
    headers = _instrument_averages_header_map(ws)
    if not {"Symbol", "Trades"}.issubset(headers):
        return False
    legacy_grouped_metrics = any(
        header in headers
        for header in ("Min stop %", "Max stop %", "Min target %", "Max target %")
    )
    changed = False
    duration_columns_added = False
    rename_map = {
        "Pattern": "Most Traded Pattern",
        "EMA": "Most Traded EMA",
    }
    for old, new in rename_map.items():
        col = headers.get(old)
        if col and new not in headers:
            ws.cell(header_row, col).value = new
            changed = True
    headers = _instrument_averages_header_map(ws)
    timeframe_col = headers.get("Most traded timeframe")
    if timeframe_col and (
        "Most Profitable Timeframe" not in headers
        or "Least Profitable Timeframe" not in headers
    ):
        insert_at = timeframe_col + 1
        ws.insert_cols(insert_at, 2)
        template_col = timeframe_col
        for row in range(1, ws.max_row + 1):
            for offset in range(2):
                _copy_cell_style(ws.cell(row, template_col), ws.cell(row, insert_at + offset))
        template_width = ws.column_dimensions[get_column_letter(template_col)].width
        for offset in range(2):
            ws.column_dimensions[get_column_letter(insert_at + offset)].width = template_width or 23
        ws.cell(header_row, insert_at).value = "Most Profitable Timeframe"
        ws.cell(header_row, insert_at + 1).value = "Least Profitable Timeframe"
        changed = True
    headers = _instrument_averages_header_map(ws)

    def grouped_anchor_col(group_name: str) -> int | None:
        tokens = {"avgstopl", "avgtargetl", "avgl"}
        for col in range(ws.max_column, 0, -1):
            if _symbols_group_for_column(ws, col).casefold() != group_name:
                continue
            token = _norm_header_token(ws.cell(header_row, col).value)
            if token in tokens:
                return col
        return None

    def expand_group_header(anchor_col: int, insert_at: int) -> None:
        group_label = _symbols_group_for_column(ws, anchor_col)
        if not group_label:
            return
        for merged in list(ws.merged_cells.ranges):
            if not (
                merged.min_row == INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
                and merged.max_row == INSTRUMENT_AVERAGES_GROUP_HEADER_ROW
                and merged.min_col <= anchor_col <= merged.max_col
            ):
                continue
            start_col = merged.min_col
            end_col = merged.max_col + 1 if insert_at > merged.max_col else merged.max_col
            ws.unmerge_cells(str(merged))
            ws.merge_cells(
                start_row=INSTRUMENT_AVERAGES_GROUP_HEADER_ROW,
                start_column=start_col,
                end_row=INSTRUMENT_AVERAGES_GROUP_HEADER_ROW,
                end_column=end_col,
            )
            ws.cell(INSTRUMENT_AVERAGES_GROUP_HEADER_ROW, start_col).value = group_label
            return

    def insert_recommendation_after(anchor_header: str, recommendation_header: str, group_name: str) -> None:
        nonlocal changed, headers
        if recommendation_header in headers:
            return
        anchor_col = headers.get(anchor_header) or grouped_anchor_col(group_name)
        if not anchor_col:
            return
        insert_at = anchor_col + 1
        ws.insert_cols(insert_at)
        expand_group_header(anchor_col, insert_at)
        for row in range(1, ws.max_row + 1):
            _copy_cell_style(ws.cell(row, anchor_col), ws.cell(row, insert_at))
        width = ws.column_dimensions[get_column_letter(anchor_col)].width
        ws.column_dimensions[get_column_letter(insert_at)].width = max(width or 0, 18) or 18
        ws.cell(header_row, insert_at).value = _symbols_visible_header(recommendation_header)
        changed = True
        headers = _instrument_averages_header_map(ws)

    def insert_overall_average_before(anchor_header: str, average_header: str) -> None:
        nonlocal changed, headers
        if average_header in headers:
            return
        anchor_col = headers.get(anchor_header)
        if not anchor_col:
            return
        insert_at = anchor_col
        ws.insert_cols(insert_at)
        template_col = insert_at + 1
        for row in range(1, ws.max_row + 1):
            _copy_cell_style(ws.cell(row, template_col), ws.cell(row, insert_at))
        width = ws.column_dimensions[get_column_letter(template_col)].width
        ws.column_dimensions[get_column_letter(insert_at)].width = max(width or 0, 14) or 14
        ws.cell(header_row, insert_at).value = average_header
        changed = True
        headers = _instrument_averages_header_map(ws)

    def insert_duration_header_after(anchor_header: str, wanted_header: str) -> None:
        nonlocal changed, duration_columns_added, headers
        if wanted_header in headers:
            return
        anchor_col = headers.get(anchor_header)
        if not anchor_col:
            return
        insert_at = anchor_col + 1
        ws.insert_cols(insert_at)
        expand_group_header(anchor_col, insert_at)
        _copy_column_style_moving_right_boundary(ws, anchor_col, insert_at)
        width = ws.column_dimensions[get_column_letter(anchor_col)].width
        ws.column_dimensions[get_column_letter(insert_at)].width = max(width or 0, 43) or 43
        ws.cell(header_row, insert_at).value = wanted_header
        changed = True
        duration_columns_added = True
        headers = _instrument_averages_header_map(ws)

    insert_overall_average_before("Avg stop % (W)", "Avg stop %")
    insert_overall_average_before("Avg target % (W)", "Avg target %")
    insert_recommendation_after("Avg stop % (L)", STOP_RECOMMENDATION_HEADER, "stops")
    insert_recommendation_after("Avg target % (L)", TARGET_RECOMMENDATION_HEADER, "targets")
    insert_duration_header_after("Longest duration (DD:HH:MM:SS)", AVG_WINNING_DURATION_HEADER)
    insert_duration_header_after(AVG_WINNING_DURATION_HEADER, AVG_LOSING_DURATION_HEADER)
    if changed:
        if not all(header in _instrument_averages_header_map(ws) for header in INSTRUMENT_AVERAGES_HEADERS) and not legacy_grouped_metrics:
            _write_instrument_averages_headers(ws, preserve_freeze=True)
        headers = _instrument_averages_header_map(ws)
        _ensure_symbols_duration_header_layout(
            ws,
            diagnostics,
            migrate_data_boundaries=duration_columns_added,
        )
        if ws.auto_filter and ws.auto_filter.ref:
            _min_col, _min_row, _max_col, max_row = range_boundaries(ws.auto_filter.ref)
            ws.auto_filter.ref = (
                f"A{header_row}:"
                f"{get_column_letter(max(headers.values()))}{max(max_row, ws.max_row)}"
            )
        _apply_symbols_filter_header_layout(ws)
        diagnostics["migrated_symbols_schema"] = True
    return changed


def _ensure_instrument_averages_schema(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if _instrument_averages_header_row(ws) == INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
        changed = _ensure_symbols_schema(ws, diagnostics)
        headers = _instrument_averages_header_map(ws)
        if all(header in headers for header in INSTRUMENT_AVERAGES_HEADERS):
            _ensure_symbols_duration_header_layout(ws, diagnostics)
            _apply_symbols_filter_header_layout(ws)
            headers = _instrument_averages_header_map(ws)
            _ensure_symbols_freeze_panes(ws, diagnostics)
            if ws.auto_filter and ws.auto_filter.ref:
                _min_col, _min_row, _max_col, max_row = range_boundaries(ws.auto_filter.ref)
                ws.auto_filter.ref = (
                    f"A{INSTRUMENT_AVERAGES_FILTER_HEADER_ROW}:"
                    f"{get_column_letter(max(headers.values()))}{max(max_row, ws.max_row)}"
                )
            return changed
    existing_row_two = [
        str(ws.cell(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, col).value or "").strip()
        for col in range(1, ws.max_column + 1)
    ]
    while existing_row_two and not existing_row_two[-1]:
        existing_row_two.pop()
    if existing_row_two == INSTRUMENT_AVERAGES_HEADERS:
        _write_instrument_averages_headers(ws, preserve_freeze=True)
        _ensure_symbols_freeze_panes(ws, diagnostics)
        return False

    existing = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
    while existing and not existing[-1]:
        existing.pop()
    aliases = {
        "Shortest (DD:HH:MM:SS)": "Shortest duration (DD:HH:MM:SS)",
        "Longest (DD:HH:MM:SS)": "Longest duration (DD:HH:MM:SS)",
    }
    canonical_existing = [aliases.get(header, header) for header in existing]
    if not canonical_existing:
        _write_instrument_averages_headers(ws, preserve_freeze=True)
        diagnostics["migrated_instrument_averages_schema"] = True
        return True
    legacy_order_col = LEGACY_INSTRUMENT_AVERAGES_HEADERS.index("Order") + 1
    if (
        len(canonical_existing) < legacy_order_col
        and canonical_existing == LEGACY_INSTRUMENT_AVERAGES_HEADERS[:len(canonical_existing)]
    ):
        ws.insert_rows(1)
        _write_instrument_averages_headers(ws, preserve_freeze=True)
        diagnostics["migrated_instrument_averages_schema"] = True
        return True
    old_compact_headers = LEGACY_INSTRUMENT_AVERAGES_HEADERS[:14] + LEGACY_INSTRUMENT_AVERAGES_HEADERS[29:]
    if canonical_existing == old_compact_headers:
        inserted_headers = LEGACY_INSTRUMENT_AVERAGES_HEADERS[14:29]
        ws.insert_cols(15, len(inserted_headers))
        template = ws.cell(1, 14)
        for offset, header in enumerate(inserted_headers, start=15):
            cell = ws.cell(1, offset)
            _copy_cell_style(template, cell)
            cell.value = header
            ws.column_dimensions[get_column_letter(offset)].width = 18
        for col, header in enumerate(LEGACY_INSTRUMENT_AVERAGES_HEADERS, start=1):
            ws.cell(1, col).value = header
        canonical_existing = list(LEGACY_INSTRUMENT_AVERAGES_HEADERS)
    if canonical_existing != LEGACY_INSTRUMENT_AVERAGES_HEADERS:
        return False

    old_order_col = LEGACY_INSTRUMENT_AVERAGES_HEADERS.index("Order") + 1
    old_widths = {
        col: ws.column_dimensions[get_column_letter(col)].width
        for col in range(old_order_col + 1, ws.max_column + 1)
    }
    ws.insert_rows(1)
    ws.insert_cols(old_order_col + 1)
    for row in range(INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, ws.max_row + 1):
        _copy_cell_style(ws.cell(row, old_order_col), ws.cell(row, old_order_col + 1))
    for col in range(ws.max_column, old_order_col + 1, -1):
        previous_width = old_widths.get(col - 1)
        if previous_width is not None:
            ws.column_dimensions[get_column_letter(col)].width = previous_width
    order_width = ws.column_dimensions[get_column_letter(old_order_col)].width
    if order_width is not None:
        ws.column_dimensions[get_column_letter(old_order_col + 1)].width = order_width
    _write_instrument_averages_headers(ws, preserve_freeze=True)
    _ensure_symbols_schema(ws, diagnostics)
    _ensure_symbols_freeze_panes(ws, diagnostics)
    diagnostics["migrated_instrument_averages_schema"] = True
    return True

def _pct_points_to_excel_fraction(value: Any) -> float | None:
    num = _as_float(value)
    return None if num is None else num / 100.0

def _excel_fraction_to_pct_points(value: Any) -> float | None:
    num = _as_float(value)
    return None if num is None else num * 100.0


def _distance_fraction_from_prices(entry: Any, level: Any) -> float | None:
    entry_num = _as_float(entry)
    level_num = _as_float(level)
    if entry_num is None or level_num is None:
        return None
    if not (math.isfinite(entry_num) and math.isfinite(level_num)) or entry_num <= 0 or level_num <= 0:
        return None
    return abs(level_num - entry_num) / entry_num


def _validated_distance_fraction(row: Dict[str, Any], level_key: str) -> float | None:
    distance = _distance_fraction_from_prices(row.get("entry_price"), row.get(level_key))
    if distance is None:
        return None
    if _trade_row_market(row) == "fx" and distance > 0.50:
        return None
    return distance


def _distance_pct_points(row: Dict[str, Any], level_key: str) -> float | None:
    fraction = _validated_distance_fraction(row, level_key)
    if fraction is not None:
        return fraction * 100.0
    fallback_key = "stop_loss_distance_pct" if level_key == "stop_loss" else "target_distance_pct"
    fallback = _as_float(row.get(fallback_key))
    if fallback is None or not math.isfinite(fallback):
        return None
    if _trade_row_market(row) == "fx" and abs(fallback) > 50.0:
        return None
    return abs(fallback)


def _trade_outcome_sign(row: Dict[str, Any]) -> int:
    for key in ("result_pct", "net_profit", "result_cash"):
        value = _as_float(row.get(key))
        if value is None or not math.isfinite(value):
            continue
        if value > 0:
            return 1
        if value < 0:
            return -1
        return 0
    return 0


def _average_float(values: List[float]) -> float | None:
    clean = [value for value in values if value is not None and math.isfinite(value)]
    return (sum(clean) / len(clean)) if clean else None


def _format_stop_pct_value(value: Any) -> str:
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        return ""
    return f"{number:.2f}%"


def _format_stop_gap_value(value: Any) -> str:
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        return ""
    return f"{abs(number):.2f} pp"


def _size_recommendation(kind: str, winner_avg: Any, loser_avg: Any) -> str:
    if kind != "stop":
        return TARGET_RECOMMENDATION_INSUFFICIENT
    win_dec = _decimal_from_value(winner_avg)
    loss_dec = _decimal_from_value(loser_avg)
    if win_dec is None or loss_dec is None:
        return ""
    win = float(win_dec)
    loss = float(loss_dec)
    recommended = win_dec
    if win_dec == loss_dec:
        recommended = max(Decimal("0.01"), win_dec - Decimal("0.01"))
        return (
            f"Decrease stop \u2014 Recommended: {_format_stop_pct_value(float(recommended))} "
            f"({_format_stop_gap_value(float(recommended - loss_dec))} below loss average; exact-tie goal preference)"
        )
    gap = abs(win_dec - loss_dec)
    if win_dec < loss_dec:
        return (
            f"Decrease stop \u2014 Recommended: {_format_stop_pct_value(win)} "
            f"({_format_stop_gap_value(float(gap))} below loss average)"
        )
    return (
        f"Increase stop \u2014 Recommended: {_format_stop_pct_value(win)} "
        f"({_format_stop_gap_value(float(gap))} above loss average)"
    )


def _target_r_row_is_test_trade(row: Dict[str, Any]) -> bool:
    marker_keys = {
        "is_test_trade",
        "is_test",
        "test_trade",
        "testtrade",
        "test",
        "paper_trade",
        "papertrade",
        "demo_test",
        "manual_test",
        "backtest_trade",
        "backtesttrade",
    }
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key, value in normalized.items():
            key_text = str(key or "").strip().casefold().replace("-", "_").replace(" ", "_")
            if key_text in marker_keys and _is_test_trade_value(value):
                return True
    return False


def _target_r_meaningful_value(value: Any) -> bool:
    if value in (None, ""):
        return False
    if isinstance(value, bool):
        return value
    num = _as_float(value)
    if num is not None and math.isfinite(num):
        return abs(num) > 1e-12
    return bool(str(value).strip())


def _target_r_nested_dicts(value: Any) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    if isinstance(value, dict):
        out.append(value)
        for child in value.values():
            out.extend(_target_r_nested_dicts(child))
    elif isinstance(value, list):
        for child in value:
            out.extend(_target_r_nested_dicts(child))
    return out


def _target_r_has_movement_evidence(row: Dict[str, Any]) -> bool:
    movement_keys = {
        "move_to_break_even_time",
        "move_to_break_even_duration",
        "move_to_break_even_trigger_price",
        "move_to_profit_time",
        "move_to_profit_duration",
        "move_to_profit_trigger_price",
        "break_even_stop",
        "breakeven_stop",
        "profit_protection_stop",
        "take_profit_adjusted",
        "take_profit_adjustment",
        "submit_take_profit_auto_adjusted",
        "stop_loss_adjusted",
        "stop_loss_modified",
        "take_profit_modified",
        "target_modified",
        "moved_stop_loss",
        "moved_take_profit",
        "adjusted_stop_loss",
        "adjusted_stop_price",
        "adjusted_take_profit",
        "adjusted_target_price",
        "tp_sl_moved",
        "tpsl_moved",
    }
    for container in _target_r_nested_dicts(row):
        for key, value in container.items():
            normalized = str(key or "").strip().casefold()
            normalized = normalized.replace("-", "_").replace(" ", "_")
            if normalized in movement_keys and _target_r_meaningful_value(value):
                return True
    return False


def _target_r_first_float(container: Dict[str, Any], keys: Tuple[str, ...]) -> float | None:
    normalized = {
        str(key).strip().casefold().replace("-", "_").replace(" ", "_"): value
        for key, value in container.items()
    }
    for key in keys:
        value = _as_float(normalized.get(key))
        if value is not None and math.isfinite(value) and value > 0:
            return value
    return None


def _decimal_from_value(value: Any) -> Decimal | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return number if number.is_finite() else None


def _target_r_first_decimal(container: Dict[str, Any], keys: Tuple[str, ...]) -> Decimal | None:
    normalized = {
        str(key).strip().casefold().replace("-", "_").replace(" ", "_"): value
        for key, value in container.items()
    }
    for key in keys:
        value = _decimal_from_value(normalized.get(key))
        if value is not None and value > 0:
            return value
    return None


def _decimal_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def _average_decimal(values: List[Any]) -> Decimal | None:
    clean = [value for value in (_decimal_from_value(v) for v in values) if value is not None]
    if not clean:
        return None
    with localcontext() as ctx:
        ctx.prec = 40
        return sum(clean, Decimal("0")) / Decimal(len(clean))


def _percentile_decimal(sorted_values: List[Any], pct: float) -> Decimal | None:
    clean = sorted(value for value in (_decimal_from_value(v) for v in sorted_values) if value is not None)
    if not clean:
        return None
    if len(clean) == 1:
        return clean[0]
    pct_dec = Decimal(str(pct))
    idx = pct_dec * Decimal(len(clean) - 1)
    lo = int(idx)
    hi = lo if idx == Decimal(lo) else lo + 1
    if lo == hi:
        return clean[lo]
    frac = idx - Decimal(lo)
    with localcontext() as ctx:
        ctx.prec = 40
        return clean[lo] * (Decimal("1") - frac) + clean[hi] * frac


def _target_r_triplet_from_container(container: Dict[str, Any]) -> Dict[str, Any] | None:
    planned_entry_keys = (
        "planned_entry_price",
        "planned_entry",
        "planned_entry_level",
    )
    planned_stop_keys = (
        "planned_stop_price",
        "planned_stop_loss",
        "planned_stop_loss_price",
    )
    planned_target_keys = (
        "planned_target_price",
        "planned_take_profit",
        "planned_take_profit_price",
        "planned_target",
    )
    original_entry_keys = (
        "original_entry_price",
        "original_entry",
        "initial_entry_price",
    )
    original_stop_keys = (
        "original_stop_loss",
        "original_stop_price",
        "original_stop_loss_price",
        "initial_stop_loss",
    )
    original_target_keys = (
        "original_take_profit",
        "original_target_price",
        "original_take_profit_price",
        "original_target",
        "initial_take_profit",
    )
    for source, entry_keys, stop_keys, target_keys in (
        ("planned", planned_entry_keys, planned_stop_keys, planned_target_keys),
        ("original", original_entry_keys, original_stop_keys, original_target_keys),
    ):
        entry = _target_r_first_float(container, entry_keys)
        stop = _target_r_first_float(container, stop_keys)
        target = _target_r_first_float(container, target_keys)
        if entry is not None and stop is not None and target is not None:
            return {
                "entry": entry,
                "stop": stop,
                "target": target,
                "source": source,
                "explicit": True,
            }
    return None


def _target_r_triplet_from_container_decimal(container: Dict[str, Any]) -> Dict[str, Any] | None:
    planned_entry_keys = (
        "planned_entry_price",
        "planned_entry",
        "planned_entry_level",
    )
    planned_stop_keys = (
        "planned_stop_price",
        "planned_stop_loss",
        "planned_stop_loss_price",
    )
    planned_target_keys = (
        "planned_target_price",
        "planned_take_profit",
        "planned_take_profit_price",
        "planned_target",
    )
    original_entry_keys = (
        "original_entry_price",
        "original_entry",
        "initial_entry_price",
    )
    original_stop_keys = (
        "original_stop_loss",
        "original_stop_price",
        "original_stop_loss_price",
        "initial_stop_loss",
    )
    original_target_keys = (
        "original_take_profit",
        "original_target_price",
        "original_take_profit_price",
        "original_target",
        "initial_take_profit",
    )
    for source, entry_keys, stop_keys, target_keys in (
        ("planned", planned_entry_keys, planned_stop_keys, planned_target_keys),
        ("original", original_entry_keys, original_stop_keys, original_target_keys),
    ):
        entry = _target_r_first_decimal(container, entry_keys)
        stop = _target_r_first_decimal(container, stop_keys)
        target = _target_r_first_decimal(container, target_keys)
        if entry is not None and stop is not None and target is not None:
            return {
                "entry": entry,
                "stop": stop,
                "target": target,
                "source": source,
                "explicit": True,
            }
    return None


def _target_r_authoritative_containers(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    containers: List[Dict[str, Any]] = []
    seen: set[int] = set()
    for source in (row, row.get("raw_refs"), row.get("metrics"), row.get("source_data")):
        for container in _target_r_nested_dicts(source):
            marker = id(container)
            if marker in seen:
                continue
            seen.add(marker)
            containers.append(container)
    return containers


def _original_plan_levels(row: Dict[str, Any]) -> Dict[str, Any]:
    for container in _target_r_authoritative_containers(row):
        triplet = _target_r_triplet_from_container(container)
        if triplet:
            triplet["movement_evidence"] = _target_r_has_movement_evidence(row)
            return triplet

    movement_evidence = _target_r_has_movement_evidence(row)
    if movement_evidence:
        return {
            "entry": None,
            "stop": None,
            "target": None,
            "source": "missing_original_after_movement",
            "explicit": False,
            "movement_evidence": True,
        }
    return {
        "entry": _as_float(row.get("entry_price")),
        "stop": _as_float(row.get("stop_loss")),
        "target": _as_float(row.get("take_profit")),
        "source": "fallback_current_levels",
        "explicit": False,
        "movement_evidence": False,
    }


def _original_plan_levels_decimal(row: Dict[str, Any]) -> Dict[str, Any]:
    for container in _target_r_authoritative_containers(row):
        triplet = _target_r_triplet_from_container_decimal(container)
        if triplet:
            triplet["movement_evidence"] = _target_r_has_movement_evidence(row)
            return triplet

    movement_evidence = _target_r_has_movement_evidence(row)
    if movement_evidence:
        return {
            "entry": None,
            "stop": None,
            "target": None,
            "source": "missing_original_after_movement",
            "explicit": False,
            "movement_evidence": True,
        }
    return {
        "entry": _decimal_from_value(row.get("entry_price")),
        "stop": _decimal_from_value(row.get("stop_loss")),
        "target": _decimal_from_value(row.get("take_profit")),
        "source": "fallback_current_levels",
        "explicit": False,
        "movement_evidence": False,
    }


def _target_r_trade_side(row: Dict[str, Any]) -> str:
    side = str(row.get("side") or "").strip().upper()
    if side.startswith("BUY") or side == "LONG":
        return "long"
    if side.startswith("SELL") or side == "SHORT":
        return "short"
    return ""


def _recorded_distance_pct_decimal(row: Dict[str, Any], level_key: str) -> Tuple[Decimal | None, str]:
    keys = (
        ("stop_loss_distance_pct", "stop_loss_distance", "stop_distance_pct", "stop_distance")
        if level_key == "stop_loss"
        else ("target_distance_pct", "target_distance", "take_profit_distance_pct", "take_profit_distance")
    )
    wanted = tuple(key.casefold().replace("-", "_").replace(" ", "_") for key in keys)
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key in wanted:
            if key not in normalized:
                continue
            raw = normalized.get(key)
            if raw in (None, "") or isinstance(raw, bool):
                continue
            text = str(raw).strip()
            if text.endswith("%"):
                text = text[:-1].strip()
            try:
                value = Decimal(text)
            except (InvalidOperation, ValueError):
                continue
            if not value.is_finite() or value == 0:
                continue
            value = abs(value)
            if _trade_row_market(row) == "fx" and value > Decimal("50"):
                return None, f"invalid_fx_original_{'stop' if level_key == 'stop_loss' else 'target'}_distance"
            return value, ""
    return None, f"missing_recorded_{'stop' if level_key == 'stop_loss' else 'target'}_distance"


def _planned_target_r_from_recorded_distances_decimal(row: Dict[str, Any]) -> Decimal | None:
    levels = _original_plan_levels_decimal(row)
    if levels.get("movement_evidence"):
        return None
    stop_pct, _stop_reason = _recorded_distance_pct_decimal(row, "stop_loss")
    target_pct, _target_reason = _recorded_distance_pct_decimal(row, "take_profit")
    if stop_pct is None or target_pct is None or stop_pct <= 0 or target_pct <= 0:
        return None
    with localcontext() as ctx:
        ctx.prec = 40
        return target_pct / stop_pct


def _target_r_recorded_r_multiple_decimal(row: Dict[str, Any]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(
            container,
            (
                "r_multiple",
                "r-multiple",
                "r",
                "captured_r",
                "realized_r",
                "realised_r",
            ),
        )
        if value is not None and value > 0:
            return value
    return None


def _planned_target_r_from_original_plan(row: Dict[str, Any]) -> float | None:
    distance_ratio = _planned_target_r_from_recorded_distances_decimal(row)
    if distance_ratio is not None:
        return float(distance_ratio)
    levels = _original_plan_levels(row)
    entry = _as_float(levels.get("entry"))
    stop = _as_float(levels.get("stop"))
    target = _as_float(levels.get("target"))
    if entry is None or stop is None or target is None:
        return None
    if not (math.isfinite(entry) and math.isfinite(stop) and math.isfinite(target)):
        return None
    if entry <= 0 or stop <= 0 or target <= 0 or abs(entry - stop) <= 1e-12:
        return None
    side = _target_r_trade_side(row)
    if side == "long" and not (stop < entry < target):
        return None
    if side == "short" and not (target < entry < stop):
        return None
    if not side:
        if abs(target - entry) <= 1e-12:
            return None
    return abs(target - entry) / abs(entry - stop)


def _planned_target_r_from_original_plan_decimal(row: Dict[str, Any]) -> Decimal | None:
    distance_ratio = _planned_target_r_from_recorded_distances_decimal(row)
    if distance_ratio is not None:
        return distance_ratio
    levels = _original_plan_levels_decimal(row)
    entry = levels.get("entry")
    stop = levels.get("stop")
    target = levels.get("target")
    if not isinstance(entry, Decimal) or not isinstance(stop, Decimal) or not isinstance(target, Decimal):
        return None
    if entry <= 0 or stop <= 0 or target <= 0 or entry == stop:
        return None
    side = _target_r_trade_side(row)
    if side == "long" and not (stop < entry < target):
        return None
    if side == "short" and not (target < entry < stop):
        return None
    if not side and target == entry:
        return None
    with localcontext() as ctx:
        ctx.prec = 40
        return abs(target - entry) / abs(entry - stop)


def _target_r_exit_price(row: Dict[str, Any]) -> float | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_float(
            container,
            (
                "exit_price",
                "closing_price",
                "close_price",
                "average_exit_price",
                "avg_exit_price",
                "filled_exit_price",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_exit_price_decimal(row: Dict[str, Any]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(
            container,
            (
                "exit_price",
                "closing_price",
                "close_price",
                "average_exit_price",
                "avg_exit_price",
                "filled_exit_price",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_original_plan_trusted(row: Dict[str, Any]) -> Tuple[bool, str, Dict[str, Any]]:
    levels = _original_plan_levels(row)
    if levels.get("movement_evidence") and not levels.get("explicit"):
        return False, "moved_without_original_plan", levels
    entry = _as_float(levels.get("entry"))
    stop = _as_float(levels.get("stop"))
    target = _as_float(levels.get("target"))
    if entry is None:
        return False, "missing_original_entry", levels
    if stop is None:
        return False, "missing_original_stop", levels
    if target is None:
        return False, "missing_original_target", levels
    if not (math.isfinite(entry) and math.isfinite(stop) and math.isfinite(target)):
        return False, "invalid_original_plan", levels
    if entry <= 0 or stop <= 0 or target <= 0:
        return False, "invalid_original_plan", levels
    if abs(entry - stop) <= 1e-12:
        return False, "zero_original_stop_risk", levels
    side = _target_r_trade_side(row)
    if not side:
        return False, "missing_trade_side", levels
    if side == "long" and not stop < entry:
        return False, "invalid_long_original_stop", levels
    if side == "short" and not stop > entry:
        return False, "invalid_short_original_stop", levels
    if side == "long" and not target > entry:
        return False, "invalid_long_original_target", levels
    if side == "short" and not target < entry:
        return False, "invalid_short_original_target", levels
    return True, "", levels


def _target_r_normalized_map(container: Dict[str, Any]) -> Dict[str, Any]:
    return {
        str(key).strip().casefold().replace("-", "_").replace(" ", "_"): value
        for key, value in container.items()
    }


def _target_r_text_from_container(container: Dict[str, Any], keys: Tuple[str, ...]) -> str:
    normalized = _target_r_normalized_map(container)
    for key in keys:
        value = normalized.get(key)
        text = str(value or "").strip().upper()
        if text:
            return text
    return ""


def _target_r_net_profit(row: Dict[str, Any]) -> float | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_float(
            container,
            (
                "net_profit",
                "net_pnl",
                "net_pl",
                "result_cash",
                "realized_pnl",
                "realised_pnl",
                "closed_pnl",
                "realized_pl",
                "realised_pl",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_net_profit_currency(row: Dict[str, Any]) -> str:
    explicit = _currency_code(
        row.get("realized_pnl_currency"),
        row.get("realised_pnl_currency"),
        row.get("result_currency"),
        row.get("currency"),
        row.get("account_currency"),
    )
    if explicit != "UNKNOWN":
        return explicit
    inferred = _infer_trade_log_currency(row, field="net_pnl")
    return inferred or "UNKNOWN"


def _target_r_risk_currency_for_key(container: Dict[str, Any], key: str, row: Dict[str, Any]) -> str:
    normalized = _target_r_normalized_map(container)
    key_text = str(key or "").strip().casefold()
    if key_text.endswith("_aud") or key_text in {"risk_aud", "estimated_total_loss_aud"}:
        return "AUD"
    if key_text.endswith("_usd") or key_text in {"risk_usd"}:
        return "USD"
    if key_text.endswith("_usdt") or key_text in {"risk_usdt"}:
        return "USDT"
    if key_text.endswith("_home") or key_text in {"risk_amount_home"}:
        return _currency_code(
            normalized.get("home_currency"),
            normalized.get("account_currency"),
            row.get("account_currency"),
            row.get("currency"),
            row.get("result_currency"),
        )
    return _currency_code(
        normalized.get(f"{key_text}_currency"),
        normalized.get("risk_currency"),
        normalized.get("original_risk_currency"),
        normalized.get("planned_risk_currency"),
        normalized.get("display_currency"),
    )


def _target_r_currency_compatible(candidate_currency: str, net_currency: str) -> bool:
    candidate = str(candidate_currency or "").strip().upper()
    net = str(net_currency or "").strip().upper()
    return bool(candidate and candidate != "UNKNOWN" and net and net != "UNKNOWN" and candidate == net)


def _target_r_close_enough(left: Any, right: Any) -> bool:
    left_num = _as_float(left)
    right_num = _as_float(right)
    if left_num is None or right_num is None:
        return False
    return math.isclose(left_num, right_num, rel_tol=1e-7, abs_tol=1e-9)


def _target_r_quote_triplet_from_container(container: Dict[str, Any]) -> Dict[str, Any] | None:
    entry = _target_r_first_float(container, ("entry_price", "entry_price_used", "quote_entry_price"))
    stop = _target_r_first_float(container, ("stop_price", "stop_loss", "stop_loss_price", "quote_stop_price"))
    target = _target_r_first_float(container, ("target_price", "take_profit", "take_profit_price", "quote_target_price"))
    if entry is not None and stop is not None and target is not None:
        return {"entry": entry, "stop": stop, "target": target, "source": "quote", "explicit": True}
    return None


def _target_r_container_matches_original_levels(container: Dict[str, Any], levels: Dict[str, Any]) -> bool:
    triplet = _target_r_triplet_from_container(container) or _target_r_quote_triplet_from_container(container)
    if not triplet:
        return False
    return (
        _target_r_close_enough(triplet.get("entry"), levels.get("entry"))
        and _target_r_close_enough(triplet.get("stop"), levels.get("stop"))
        and _target_r_close_enough(triplet.get("target"), levels.get("target"))
    )


def _target_r_stored_original_monetary_risk(row: Dict[str, Any], levels: Dict[str, Any]) -> Tuple[float | None, str]:
    net_currency = _target_r_net_profit_currency(row)
    verified_risk_keys = (
        "original_monetary_risk",
        "planned_monetary_risk",
        "original_risk_amount",
        "planned_risk_amount",
        "initial_risk_amount",
        "risk_amount_home",
        "risk_aud",
        "risk_usd",
        "risk_usdt",
        "estimated_total_loss",
        "estimated_total_loss_aud",
        "estimated_total_loss_usd",
        "estimated_total_loss_usdt",
        "planned_total_loss",
        "original_total_loss",
    )
    saw_currency_mismatch = False
    saw_unknown_currency = False
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        container_matches_levels = _target_r_container_matches_original_levels(container, levels)
        for key in verified_risk_keys:
            value = _as_float(normalized.get(key))
            if value is None or not math.isfinite(value) or value <= 0:
                continue
            if key.startswith("estimated_total_loss") and not container_matches_levels:
                continue
            risk_currency = _target_r_risk_currency_for_key(container, key, row)
            if not risk_currency or risk_currency == "UNKNOWN":
                saw_unknown_currency = True
                continue
            if not _target_r_currency_compatible(risk_currency, net_currency):
                saw_currency_mismatch = True
                continue
            return value, ""
        risk_value = _as_float(normalized.get("risk_value"))
        risk_mode = str(normalized.get("risk_mode") or "").strip().casefold()
        if risk_value is not None and math.isfinite(risk_value) and risk_value > 0:
            fixed_risk_mode = any(token in risk_mode for token in ("fixed", "cash", "amount", "aud", "usd", "usdt"))
            percent_risk_mode = "percent" in risk_mode or risk_mode in {"pct", "%"}
            if fixed_risk_mode and not percent_risk_mode and container_matches_levels:
                risk_currency = "AUD" if "aud" in risk_mode else ("USDT" if "usdt" in risk_mode else ("USD" if "usd" in risk_mode else "UNKNOWN"))
                if not risk_currency or risk_currency == "UNKNOWN":
                    saw_unknown_currency = True
                    continue
                if not _target_r_currency_compatible(risk_currency, net_currency):
                    saw_currency_mismatch = True
                    continue
                return risk_value, ""
    if saw_currency_mismatch:
        return None, "stored_risk_currency_mismatch"
    if saw_unknown_currency:
        return None, "missing_original_risk_currency"
    return None, "missing_stored_original_monetary_risk"


def _target_r_stored_original_monetary_risk_decimal(row: Dict[str, Any], levels: Dict[str, Any]) -> Tuple[Decimal | None, str]:
    net_currency = _target_r_net_profit_currency(row)
    verified_risk_keys = (
        "original_monetary_risk",
        "planned_monetary_risk",
        "original_risk_amount",
        "planned_risk_amount",
        "initial_risk_amount",
        "risk_amount_home",
        "risk_aud",
        "risk_usd",
        "risk_usdt",
        "estimated_total_loss",
        "estimated_total_loss_aud",
        "estimated_total_loss_usd",
        "estimated_total_loss_usdt",
        "planned_total_loss",
        "original_total_loss",
    )
    float_levels = {
        "entry": _decimal_to_float(levels.get("entry") if isinstance(levels.get("entry"), Decimal) else None),
        "stop": _decimal_to_float(levels.get("stop") if isinstance(levels.get("stop"), Decimal) else None),
        "target": _decimal_to_float(levels.get("target") if isinstance(levels.get("target"), Decimal) else None),
    }
    saw_currency_mismatch = False
    saw_unknown_currency = False
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        container_matches_levels = _target_r_container_matches_original_levels(container, float_levels)
        for key in verified_risk_keys:
            value = _target_r_positive_decimal(normalized.get(key))
            if value is None:
                continue
            if key.startswith("estimated_total_loss") and not container_matches_levels:
                continue
            risk_currency = _target_r_risk_currency_for_key(container, key, row)
            if not risk_currency or risk_currency == "UNKNOWN":
                saw_unknown_currency = True
                continue
            if not _target_r_currency_compatible(risk_currency, net_currency):
                saw_currency_mismatch = True
                continue
            return value, ""
        risk_value = _target_r_positive_decimal(normalized.get("risk_value"))
        risk_mode = str(normalized.get("risk_mode") or "").strip().casefold()
        if risk_value is not None:
            fixed_risk_mode = any(token in risk_mode for token in ("fixed", "cash", "amount", "aud", "usd", "usdt"))
            percent_risk_mode = "percent" in risk_mode or risk_mode in {"pct", "%"}
            if fixed_risk_mode and not percent_risk_mode and container_matches_levels:
                risk_currency = "AUD" if "aud" in risk_mode else ("USDT" if "usdt" in risk_mode else ("USD" if "usd" in risk_mode else "UNKNOWN"))
                if not risk_currency or risk_currency == "UNKNOWN":
                    saw_unknown_currency = True
                    continue
                if not _target_r_currency_compatible(risk_currency, net_currency):
                    saw_currency_mismatch = True
                    continue
                return risk_value, ""
    if saw_currency_mismatch:
        return None, "stored_risk_currency_mismatch"
    if saw_unknown_currency:
        return None, "missing_original_risk_currency"
    return None, "missing_stored_original_monetary_risk"


def _target_r_net_profit_decimal(row: Dict[str, Any]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(
            container,
            (
                "net_profit",
                "net_pnl",
                "net_pl",
                "result_cash",
                "realized_pnl",
                "realised_pnl",
                "closed_pnl",
                "realized_pl",
                "realised_pl",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_stored_original_net_r(row: Dict[str, Any], levels: Dict[str, Any]) -> Tuple[Decimal | None, str]:
    provenance_tokens = ("original", "entry", "stop", "net")
    r_keys = ("stored_net_r", "original_net_r", "captured_net_r", "realized_net_r", "realised_net_r", "r_multiple")
    float_levels = {
        "entry": _decimal_to_float(levels.get("entry") if isinstance(levels.get("entry"), Decimal) else None),
        "stop": _decimal_to_float(levels.get("stop") if isinstance(levels.get("stop"), Decimal) else None),
        "target": _decimal_to_float(levels.get("target") if isinstance(levels.get("target"), Decimal) else None),
    }
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        provenance = " ".join(
            str(normalized.get(key) or "")
            for key in (
                "stored_net_r_provenance",
                "net_r_provenance",
                "r_multiple_provenance",
                "r_calculation_provenance",
            )
        ).casefold()
        flags_prove = all(
            _is_test_trade_value(normalized.get(key))
            for key in (
                "net_r_uses_original_entry",
                "net_r_uses_original_stop",
                "net_r_uses_net_result",
            )
        )
        text_proves = all(token in provenance for token in provenance_tokens)
        if not flags_prove and not text_proves:
            continue
        if not _target_r_container_matches_original_levels(container, float_levels):
            continue
        for key in r_keys:
            value = _target_r_positive_decimal(normalized.get(key))
            if value is not None:
                return value, ""
    return None, "missing_verified_stored_original_net_r"


def _target_r_quantity(row: Dict[str, Any]) -> float | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_float(
            container,
            (
                "qty",
                "quantity",
                "qty_raw",
                "size",
                "filled_qty",
                "closed_size",
                "exec_qty",
                "units",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_quantity_decimal(row: Dict[str, Any]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(
            container,
            (
                "qty",
                "quantity",
                "qty_raw",
                "size",
                "filled_qty",
                "closed_size",
                "exec_qty",
                "units",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_explicit_multiplier(row: Dict[str, Any]) -> float | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_float(
            container,
            (
                "contract_multiplier",
                "contract_size",
                "multiplier",
                "price_to_pnl_multiplier",
                "point_value",
                "tick_value",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_explicit_multiplier_decimal(row: Dict[str, Any]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(
            container,
            (
                "contract_multiplier",
                "contract_size",
                "multiplier",
                "price_to_pnl_multiplier",
                "point_value",
                "tick_value",
            ),
        )
        if value is not None:
            return value
    return None


def _target_r_is_fx_row(row: Dict[str, Any]) -> bool:
    account = str(row.get("account_label") or row.get("account") or "").upper()
    symbol = str(row.get("symbol") or "").upper()
    asset = str(row.get("asset_class") or "").strip().lower()
    return (
        asset in {"fx", "forex", "foreign_exchange"}
        or any(token in account for token in ("OANDA", "PEPPERSTONE", "FOREX", " FX"))
        or _is_likely_fx_pair(symbol)
    )


def _target_r_is_crypto_row(row: Dict[str, Any]) -> bool:
    account = str(row.get("account_label") or row.get("account") or "").upper()
    symbol = str(row.get("symbol") or "").upper()
    asset = str(row.get("asset_class") or "").strip().lower()
    return (
        asset in {"crypto", "cryptocurrency", "digital_asset", "digitalasset"}
        or any(token in account for token in ("BYBIT", "BINANCE", "COINSPOT"))
        or bool(_symbol_quote_currency(symbol))
    )


def _target_r_commission_abs(row: Dict[str, Any]) -> float:
    for container in _target_r_authoritative_containers(row):
        value = _as_float(container.get("commission"))
        if value is None:
            value = _as_float(container.get("fees"))
        if value is not None and math.isfinite(value):
            return abs(value)
    return 0.0


def _target_r_fx_units(row: Dict[str, Any], qty: float) -> float | None:
    unit_text = ""
    for container in _target_r_authoritative_containers(row):
        unit_text = _target_r_text_from_container(container, ("quantity_unit", "qty_unit", "size_unit", "unit"))
        if unit_text:
            break
    multiplier = _target_r_explicit_multiplier(row)
    if multiplier is not None:
        return qty * multiplier
    if unit_text in {"UNIT", "UNITS"}:
        return qty
    if unit_text in {"LOT", "LOTS", "STANDARD_LOT", "STANDARD_LOTS"}:
        return qty * 100000.0
    if 0 < qty < 1000:
        return qty * 100000.0
    return qty


def _target_r_fx_units_decimal(row: Dict[str, Any], qty: Decimal) -> Decimal | None:
    unit_text = ""
    for container in _target_r_authoritative_containers(row):
        unit_text = _target_r_text_from_container(container, ("quantity_unit", "qty_unit", "size_unit", "unit"))
        if unit_text:
            break
    multiplier = _target_r_explicit_multiplier_decimal(row)
    if multiplier is not None:
        return qty * multiplier
    if unit_text in {"UNIT", "UNITS"}:
        return qty
    if unit_text in {"LOT", "LOTS", "STANDARD_LOT", "STANDARD_LOTS"}:
        return qty * Decimal("100000")
    if Decimal("0") < qty < Decimal("1000"):
        return qty * Decimal("100000")
    return qty


def _target_r_positive_factor(value: Any) -> float | None:
    factor = _as_float(value)
    if factor is not None and math.isfinite(factor) and factor > 0:
        return factor
    return None


def _target_r_positive_decimal(value: Any) -> Decimal | None:
    factor = _decimal_from_value(value)
    if factor is not None and factor > 0:
        return factor
    return None


def _target_r_loss_quote_home_factor(factors: Any) -> float | None:
    if not isinstance(factors, dict):
        return None
    factor_map = _target_r_normalized_map(factors)
    for nested_key in ("lossquotehome", "loss_quote_home"):
        nested = factor_map.get(nested_key)
        if isinstance(nested, dict):
            value = _target_r_positive_factor(_target_r_normalized_map(nested).get("factor"))
            if value is not None:
                return value
    for flattened_key in ("lossquotehomefactor", "loss_quote_home_factor"):
        value = _target_r_positive_factor(factor_map.get(flattened_key))
        if value is not None:
            return value
    return None


def _target_r_loss_quote_home_factor_decimal(factors: Any) -> Decimal | None:
    if not isinstance(factors, dict):
        return None
    factor_map = _target_r_normalized_map(factors)
    for nested_key in ("lossquotehome", "loss_quote_home"):
        nested = factor_map.get(nested_key)
        if isinstance(nested, dict):
            value = _target_r_positive_decimal(_target_r_normalized_map(nested).get("factor"))
            if value is not None:
                return value
    for flattened_key in ("lossquotehomefactor", "loss_quote_home_factor"):
        value = _target_r_positive_decimal(factor_map.get(flattened_key))
        if value is not None:
            return value
    return None


def _target_r_independent_fx_loss_conversion(row: Dict[str, Any]) -> float | None:
    direct_keys = (
        "original_loss_conversion_factor",
        "original_loss_conversion_rate",
        "opening_loss_conversion_factor",
        "opening_loss_conversion_rate",
        "oanda_open_conversion_rate",
        "oanda_export_open_conversion_rate",
        "oanda_export_conversion_rate",
        "lossquotehomeconversionfactor",
        "loss_quote_home_conversion_factor",
    )
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key in direct_keys:
            value = _target_r_positive_factor(normalized.get(key))
            if value is not None:
                return value
        value = _target_r_loss_quote_home_factor(normalized.get("homeconversionfactors"))
        if value is not None:
            return value
    return None


def _target_r_independent_fx_loss_conversion_decimal(row: Dict[str, Any]) -> Decimal | None:
    direct_keys = (
        "original_loss_conversion_factor",
        "original_loss_conversion_rate",
        "opening_loss_conversion_factor",
        "opening_loss_conversion_rate",
        "oanda_open_conversion_rate",
        "oanda_export_open_conversion_rate",
        "oanda_export_conversion_rate",
        "lossquotehomeconversionfactor",
        "loss_quote_home_conversion_factor",
    )
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key in direct_keys:
            value = _target_r_positive_decimal(normalized.get(key))
            if value is not None:
                return value
        value = _target_r_loss_quote_home_factor_decimal(normalized.get("homeconversionfactors"))
        if value is not None:
            return value
    return None


def _target_r_reconstructed_original_monetary_risk(
    row: Dict[str, Any],
    levels: Dict[str, Any],
    net_profit: float,
) -> Tuple[float | None, str]:
    entry = _as_float(levels.get("entry"))
    stop = _as_float(levels.get("stop"))
    if entry is None or stop is None or not (math.isfinite(entry) and math.isfinite(stop)):
        return None, "missing_original_price_risk"
    price_risk = abs(entry - stop)
    if price_risk <= 1e-12:
        return None, "zero_original_stop_risk"
    qty = _target_r_quantity(row)
    if qty is None or not math.isfinite(qty) or qty <= 0:
        return None, "missing_original_risk_quantity"

    net_currency = _target_r_net_profit_currency(row)
    quote_currency = _symbol_quote_currency(row.get("symbol"))
    multiplier = _target_r_explicit_multiplier(row)
    if _target_r_is_crypto_row(row) and not _target_r_is_fx_row(row):
        multiplier = multiplier if multiplier is not None else 1.0
        risk_currency = quote_currency or net_currency
        if not _target_r_currency_compatible(risk_currency, net_currency):
            return None, "reconstructed_risk_currency_mismatch"
        risk = price_risk * qty * multiplier
        if risk > 0 and math.isfinite(risk):
            return risk, ""

    if _target_r_is_fx_row(row):
        units = _target_r_fx_units(row, qty)
        if units is None or not math.isfinite(units) or units <= 0:
            return None, "missing_independent_original_monetary_risk"
        conversion = _target_r_independent_fx_loss_conversion(row)
        if conversion is None:
            return None, "missing_opening_loss_conversion_factor"
        risk = price_risk * units * conversion
        if risk > 0 and math.isfinite(risk):
            return risk, ""
        return None, "missing_opening_loss_conversion_factor"

    if multiplier is not None:
        risk = price_risk * qty * multiplier
        if risk > 0 and math.isfinite(risk):
            return risk, ""
    return None, "missing_original_monetary_risk"


def _target_r_reconstructed_original_monetary_risk_decimal(
    row: Dict[str, Any],
    levels: Dict[str, Any],
) -> Tuple[Decimal | None, str, str]:
    entry = levels.get("entry")
    stop = levels.get("stop")
    if not isinstance(entry, Decimal) or not isinstance(stop, Decimal):
        return None, "missing_original_price_risk", ""
    price_risk = abs(entry - stop)
    if price_risk <= 0:
        return None, "zero_original_stop_risk", ""
    qty = _target_r_quantity_decimal(row)
    if qty is None or qty <= 0:
        return None, "missing_original_risk_quantity", ""

    net_currency = _target_r_net_profit_currency(row)
    quote_currency = _symbol_quote_currency(row.get("symbol"))
    multiplier = _target_r_explicit_multiplier_decimal(row)
    if _target_r_is_crypto_row(row) and not _target_r_is_fx_row(row):
        multiplier = multiplier if multiplier is not None else Decimal("1")
        risk_currency = quote_currency or net_currency
        if not _target_r_currency_compatible(risk_currency, net_currency):
            return None, "reconstructed_risk_currency_mismatch", ""
        with localcontext() as ctx:
            ctx.prec = 40
            risk = price_risk * qty * multiplier
        if risk > 0:
            return risk, "", "verified_net_profit_over_original_monetary_risk"

    if _target_r_is_fx_row(row):
        units = _target_r_fx_units_decimal(row, qty)
        if units is None or units <= 0:
            return None, "missing_independent_original_monetary_risk", ""
        conversion = _target_r_independent_fx_loss_conversion_decimal(row)
        if conversion is None:
            return None, "missing_opening_conversion", ""
        with localcontext() as ctx:
            ctx.prec = 40
            risk = price_risk * units * conversion
        if risk > 0:
            return risk, "", "reconstructed_opening_conversion_net_r"
        return None, "missing_opening_conversion", ""

    if multiplier is not None:
        with localcontext() as ctx:
            ctx.prec = 40
            risk = price_risk * qty * multiplier
        if risk > 0:
            return risk, "", "verified_net_profit_over_original_monetary_risk"
    return None, "missing_original_monetary_risk", ""


def _target_r_is_oanda_row(row: Dict[str, Any]) -> bool:
    account = str(row.get("account_label") or row.get("account") or "").upper()
    source = str(row.get("source") or "").strip().casefold()
    row_id = str(row.get("id") or "").strip().casefold()
    return "OANDA" in account or source.startswith("oanda_") or row_id.startswith("oanda_")


def _target_r_normalized_value(row: Dict[str, Any], keys: Tuple[str, ...]) -> Any:
    wanted = {key.casefold().replace("-", "_").replace(" ", "_") for key in keys}
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key in wanted:
            if key in normalized:
                return normalized.get(key)
    return None


def _target_r_any_key_present(row: Dict[str, Any], keys: Tuple[str, ...]) -> bool:
    wanted = {key.casefold().replace("-", "_").replace(" ", "_") for key in keys}
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        if any(key in normalized for key in wanted):
            return True
    return False


def _target_r_decimal_value_for_keys(row: Dict[str, Any], keys: Tuple[str, ...]) -> Decimal | None:
    for container in _target_r_authoritative_containers(row):
        value = _target_r_first_decimal(container, keys)
        if value is not None:
            return value
        normalized = _target_r_normalized_map(container)
        for key in keys:
            raw = normalized.get(key.casefold().replace("-", "_").replace(" ", "_"))
            dec = _decimal_from_value(raw)
            if dec is not None:
                return dec
    return None


def _target_r_oanda_net_result_provenance(row: Dict[str, Any]) -> Tuple[bool, str]:
    if not _target_r_is_oanda_row(row):
        return True, ""

    commission_present = _target_r_any_key_present(
        row,
        (
            "oanda_actual_commission_total",
            "oanda_raw_commission",
            "commission",
            "fees",
            "gsl_fee_raw",
            "gsl_premium_raw",
        ),
    )
    commission_values = [
        _target_r_decimal_value_for_keys(row, (key,))
        for key in (
            "oanda_actual_commission_total",
            "oanda_raw_commission",
            "commission",
            "fees",
            "gsl_fee_raw",
            "gsl_premium_raw",
        )
    ]
    commission_values = [value for value in commission_values if value is not None]
    if not commission_present:
        return False, "unverified_commission_cost_status"
    if any(value != 0 for value in commission_values):
        return False, "unverified_commission_cost_status"

    financing_present = _target_r_any_key_present(
        row,
        ("oanda_export_financing_allocated", "financing", "swap"),
    )
    if not financing_present:
        return False, "unverified_financing_swap"

    single_complete = _target_r_normalized_value(
        row,
        (
            "single_complete_exit_verified",
            "oanda_single_complete_exit",
            "single_complete_exit",
        ),
    )
    if _is_test_trade_value(single_complete):
        pass
    else:
        open_units = _target_r_decimal_value_for_keys(row, ("opening_units", "open_units", "qty_raw", "units"))
        close_units = _target_r_decimal_value_for_keys(row, ("closing_units", "closed_units", "close_units"))
        open_ticket = str(_target_r_normalized_value(row, ("open_ticket", "original_open_transaction_id", "opening_transaction_id")) or "").strip()
        close_ticket = str(_target_r_normalized_value(row, ("close_ticket", "close_transaction_id", "closing_transaction_id")) or "").strip()
        if not (open_ticket and close_ticket and open_units is not None and close_units is not None and abs(open_units) == abs(close_units)):
            return False, "partial_or_multiple_exits_not_reconstructed"

    executed_prices = _target_r_normalized_value(
        row,
        (
            "entry_exit_prices_executed",
            "executed_entry_exit_prices",
            "oanda_executed_prices",
        ),
    )
    if not _is_test_trade_value(executed_prices):
        source = str(row.get("source") or "").strip().casefold()
        if source != "oanda_transaction_export":
            return False, "unverified_executed_prices"
    return True, ""


def _target_r_original_monetary_risk(
    row: Dict[str, Any],
    levels: Dict[str, Any],
    net_profit: float,
) -> Tuple[float | None, str]:
    stored, stored_reason = _target_r_stored_original_monetary_risk(row, levels)
    if stored is not None:
        return stored, ""
    reconstructed, reconstructed_reason = _target_r_reconstructed_original_monetary_risk(row, levels, net_profit)
    if reconstructed is not None:
        return reconstructed, ""
    if _target_r_is_fx_row(row):
        return None, reconstructed_reason or "missing_opening_loss_conversion_factor"
    if stored_reason != "missing_stored_original_monetary_risk":
        return None, stored_reason
    return None, reconstructed_reason


def _target_r_original_monetary_risk_detail(
    row: Dict[str, Any],
    levels: Dict[str, Any],
) -> Tuple[Decimal | None, str, str]:
    stored, stored_reason = _target_r_stored_original_monetary_risk_decimal(row, levels)
    if stored is not None:
        return stored, "", "verified_net_profit_over_original_monetary_risk"
    reconstructed, reconstructed_reason, method = _target_r_reconstructed_original_monetary_risk_decimal(row, levels)
    if reconstructed is not None:
        ok, reason = _target_r_oanda_net_result_provenance(row)
        if not ok:
            return None, reason, ""
        return reconstructed, "", method or "verified_net_profit_over_original_monetary_risk"
    if _target_r_is_fx_row(row):
        return None, reconstructed_reason or "missing_opening_conversion", ""
    if stored_reason != "missing_stored_original_monetary_risk":
        return None, stored_reason, ""
    return None, reconstructed_reason, ""


def _target_r_price_capture_from_original_plan(
    row: Dict[str, Any],
    levels: Dict[str, Any],
) -> Tuple[float | None, str]:
    entry = _as_float(levels.get("entry"))
    stop = _as_float(levels.get("stop"))
    exit_price = _target_r_exit_price(row)
    if entry is None or stop is None:
        return None, "missing_original_price_risk"
    if exit_price is None:
        return None, "missing_exit_price"
    if not (math.isfinite(entry) and math.isfinite(stop) and math.isfinite(exit_price)):
        return None, "invalid_exit_price"
    original_risk = abs(entry - stop)
    if original_risk <= 1e-12:
        return None, "zero_original_stop_risk"
    side = _target_r_trade_side(row)
    if side == "long":
        captured = (exit_price - entry) / original_risk
    elif side == "short":
        captured = (entry - exit_price) / original_risk
    else:
        return None, "missing_trade_side"
    if not math.isfinite(captured):
        return None, "invalid_realized_r"
    if captured <= 0:
        return None, "non_positive_realized_r"
    return captured, ""


def _target_r_price_capture_from_original_plan_decimal(
    row: Dict[str, Any],
    levels: Dict[str, Any],
) -> Tuple[Decimal | None, str]:
    entry = levels.get("entry")
    stop = levels.get("stop")
    exit_price = _target_r_exit_price_decimal(row)
    if not isinstance(entry, Decimal) or not isinstance(stop, Decimal):
        return None, "missing_original_price_risk"
    if exit_price is None:
        return None, "missing_actual_exit"
    original_risk = abs(entry - stop)
    if original_risk <= 0:
        return None, "zero_original_stop_risk"
    side = _target_r_trade_side(row)
    with localcontext() as ctx:
        ctx.prec = 40
        if side == "long":
            captured = (exit_price - entry) / original_risk
        elif side == "short":
            captured = (entry - exit_price) / original_risk
        else:
            return None, "missing_trade_side"
    if captured <= 0:
        return None, "non_positive_realized_r"
    return captured, ""


def _target_r_price_capture_from_recorded_stop_distance_decimal(
    row: Dict[str, Any],
    levels: Dict[str, Any],
) -> Tuple[Decimal | None, str]:
    entry = levels.get("entry")
    exit_price = _target_r_exit_price_decimal(row)
    if not isinstance(entry, Decimal):
        return None, "missing_original_entry"
    if exit_price is None:
        return None, "missing_actual_exit"
    stop_pct, reason = _recorded_distance_pct_decimal(row, "stop_loss")
    if stop_pct is None:
        return None, reason or "missing_recorded_stop_distance"
    with localcontext() as ctx:
        ctx.prec = 40
        original_risk = entry * (stop_pct / Decimal("100"))
        if original_risk <= 0:
            return None, "zero_original_stop_risk"
        side = _target_r_trade_side(row)
        if side == "long":
            captured = (exit_price - entry) / original_risk
        elif side == "short":
            captured = (entry - exit_price) / original_risk
        else:
            return None, "missing_trade_side"
    if captured <= 0:
        return None, "non_positive_realized_r"
    return captured, ""


def _target_r_price_capture_net_equivalent(row: Dict[str, Any]) -> Tuple[bool, str]:
    explicit_verified = _target_r_normalized_value(
        row,
        (
            "net_equivalent_price_r_verified",
            "price_r_net_equivalent_verified",
            "verified_net_equivalent_price_r",
        ),
    )
    if _is_test_trade_value(explicit_verified):
        return True, ""

    if _target_r_is_fx_row(row):
        ok, reason = _target_r_oanda_net_result_provenance(row)
        if not ok:
            return False, reason
        return False, "unverified_net_equivalent_price_r"

    cost_tokens = (
        "commission",
        "fee",
        "fees",
        "swap",
        "financing",
        "funding",
        "spread_cost",
        "slippage",
        "rebate",
    )
    partial_tokens = (
        "partial",
        "scale_out",
        "scaleout",
        "split_exit",
        "splitexit",
        "multiple_exit",
        "multipleexit",
        "reduce_only",
        "close_qty",
        "closed_qty",
        "filled_exit_qty",
        "remaining_qty",
    )
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key, value in normalized.items():
            key_text = str(key or "").strip().casefold().replace("-", "_").replace(" ", "_")
            if any(token in key_text for token in cost_tokens):
                amount = _as_float(value)
                if amount is not None and math.isfinite(amount):
                    if abs(amount) > 1e-12:
                        return False, "price_r_not_net_due_to_costs"
                    continue
                if _target_r_meaningful_value(value):
                    return False, "price_r_unverified_costs"
            if any(token in key_text for token in partial_tokens) and _target_r_meaningful_value(value):
                return False, "price_r_unverified_partial_exit"
    return False, "unverified_net_equivalent_price_r"


def _target_r_realized_from_original_plan_detail(row: Dict[str, Any]) -> Tuple[float | None, str, str, Decimal | None]:
    decimal_levels = _original_plan_levels_decimal(row)
    if decimal_levels.get("movement_evidence") and not decimal_levels.get("explicit"):
        return None, "moved_without_original_plan", "", None
    recorded_r = _target_r_recorded_r_multiple_decimal(row)
    if recorded_r is not None:
        return float(recorded_r), "", "recorded_r_multiple", recorded_r
    price_r, price_reason = _target_r_price_capture_from_original_plan_decimal(row, decimal_levels)
    if price_r is not None:
        return float(price_r), "", "price_captured_r_from_original_plan", price_r
    distance_r, distance_reason = _target_r_price_capture_from_recorded_stop_distance_decimal(row, decimal_levels)
    if distance_r is not None:
        return float(distance_r), "", "price_captured_r_from_recorded_stop_distance", distance_r
    return None, price_reason or distance_reason or "invalid_realized_r", "", None


def _target_r_realized_from_original_plan(row: Dict[str, Any]) -> Tuple[float | None, str]:
    realized, reason, _method, _realized_dec = _target_r_realized_from_original_plan_detail(row)
    return realized, reason


def _target_r_metadata_payload(row: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(row, dict) or str(row.get("row_type") or "trade").strip().lower() != "trade":
        return {}
    payload: Dict[str, Any] = {}
    levels = _original_plan_levels(row)
    for source_key, output_key in (
        ("entry", "original_entry_price"),
        ("stop", "original_stop_loss"),
        ("target", "original_take_profit"),
    ):
        value = _target_r_positive_factor(levels.get(source_key))
        if value is not None:
            payload[output_key] = value
    qty = _target_r_quantity(row)
    if qty is not None and math.isfinite(qty) and qty > 0:
        payload["original_quantity"] = qty
        payload.setdefault("qty", qty)
    for container in _target_r_authoritative_containers(row):
        normalized = _target_r_normalized_map(container)
        for key in ("qty_raw", "units"):
            value = _target_r_positive_factor(normalized.get(key))
            if value is not None:
                payload[key] = value
                break
        unit_text = _target_r_text_from_container(container, ("qty_unit", "quantity_unit", "size_unit", "unit"))
        if unit_text:
            payload["qty_unit"] = unit_text.lower()
            break
    net_currency = _target_r_net_profit_currency(row)
    stored, _stored_reason = _target_r_stored_original_monetary_risk(row, levels)
    if stored is not None:
        payload["original_monetary_risk"] = stored
        payload["original_risk_currency"] = net_currency
    else:
        net_profit = _target_r_net_profit(row)
        if net_profit is not None:
            reconstructed, _reason = _target_r_reconstructed_original_monetary_risk(row, levels, net_profit)
            if reconstructed is not None:
                payload["original_monetary_risk"] = reconstructed
                payload["original_risk_currency"] = net_currency
    conversion = _target_r_independent_fx_loss_conversion(row)
    if conversion is not None:
        payload["original_loss_conversion_factor"] = conversion
        payload["opening_loss_conversion_factor"] = conversion
        payload["original_loss_conversion_factor_source"] = str(
            row.get("original_loss_conversion_factor_source")
            or row.get("opening_loss_conversion_factor_source")
            or "opening_loss_quote_home"
        )
        payload["original_loss_conversion_factor_time"] = str(
            row.get("original_loss_conversion_factor_time")
            or row.get("opening_loss_conversion_factor_time")
            or row.get("open_time")
            or ""
        )
    refs = row.get("raw_refs") if isinstance(row.get("raw_refs"), dict) else {}
    open_tx = (
        row.get("original_open_transaction_id")
        or row.get("opening_transaction_id")
        or refs.get("original_open_transaction_id")
        or refs.get("open_transaction_id")
        or refs.get("open_ticket")
        or refs.get("transactionId")
    )
    if open_tx not in (None, ""):
        payload["original_open_transaction_id"] = str(open_tx)
    return {key: value for key, value in payload.items() if value not in (None, "")}


def _read_target_r_metadata_sheet(wb, diagnostics: Dict[str, Any] | None = None) -> Dict[str, Dict[str, Any]]:
    if TARGET_R_METADATA_SHEET not in wb.sheetnames:
        return {}
    ws = wb[TARGET_R_METADATA_SHEET]
    out: Dict[str, Dict[str, Any]] = {}
    seen: set[str] = set()
    duplicates: set[str] = set()
    for row_id, payload_json in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        rid = str(row_id or "").strip()
        if not rid or not payload_json:
            continue
        if rid in seen:
            duplicates.add(rid)
            out.pop(rid, None)
            continue
        if rid in duplicates:
            continue
        try:
            payload = json.loads(str(payload_json))
        except Exception:
            continue
        if isinstance(payload, dict):
            seen.add(rid)
            out[rid] = payload
    if diagnostics is not None and duplicates:
        diagnostics["target_r_metadata_duplicate_row_ids"] = sorted(duplicates)
    return out


def _merge_target_r_metadata(row: Dict[str, Any], metadata: Dict[str, Any]) -> Dict[str, Any]:
    if not metadata:
        return row
    merged = dict(row)
    for key, value in metadata.items():
        if key in {"raw_refs", "metrics"} and isinstance(value, dict):
            existing = merged.get(key) if isinstance(merged.get(key), dict) else {}
            merged[key] = {**value, **existing}
            continue
        if merged.get(key) in (None, ""):
            merged[key] = value
    return merged


def _remove_target_r_metadata_sheet(
    wb,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if TARGET_R_METADATA_SHEET in wb.sheetnames:
        wb.remove(wb[TARGET_R_METADATA_SHEET])
        diagnostics["removed_target_r_metadata_sheet"] = True


def _is_eligible_target_r_win(row: Dict[str, Any]) -> Tuple[bool, str]:
    if str(row.get("row_type") or "trade").strip().lower() != "trade":
        return False, "not_trade"
    if _target_r_row_is_test_trade(row):
        return False, "test_trade"
    if _trade_outcome_sign(row) <= 0:
        return False, "not_winning_trade"
    realized_r, reason = _target_r_realized_from_original_plan(row)
    if realized_r is None:
        return False, reason
    return True, ""


def _target_r_bucket(value: float, increment: float) -> float:
    return round(math.floor(value / increment) * increment, 10)


def _target_r_distribution(values: List[float], increment: float) -> Dict[float, int]:
    counts: Counter[float] = Counter()
    for value in values:
        if value is None or not math.isfinite(value) or value <= 0:
            continue
        counts[_target_r_bucket(value, increment)] += 1
    return OrderedDict((bucket, counts[bucket]) for bucket in sorted(counts))


def _target_r_bucket_label(bucket: float, increment: float) -> str:
    return f"{_format_target_r_value(bucket)}R-{_format_target_r_value(bucket + increment)}R"


def _target_r_bucket_values(values: List[float], increment: float) -> "OrderedDict[float, List[float]]":
    grouped: Dict[float, List[float]] = defaultdict(list)
    for value in values:
        if value is None or not math.isfinite(value) or value <= 0:
            continue
        grouped[_target_r_bucket(value, increment)].append(value)
    return OrderedDict((bucket, grouped[bucket]) for bucket in sorted(grouped))


def _target_r_percentile(sorted_values: List[float], percentile: float) -> float | None:
    if not sorted_values:
        return None
    if len(sorted_values) == 1:
        return sorted_values[0]
    position = (len(sorted_values) - 1) * percentile
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return sorted_values[lower]
    weight = position - lower
    return sorted_values[lower] + ((sorted_values[upper] - sorted_values[lower]) * weight)


def _choose_target_r_increment(values: List[float]) -> float:
    clean = [value for value in values if value is not None and math.isfinite(value) and value > 0]
    if len(clean) < 4:
        return 0.5
    ordered = sorted(clean)
    q1 = _target_r_percentile(ordered, 0.25)
    q3 = _target_r_percentile(ordered, 0.75)
    if q1 is None or q3 is None:
        return 0.5
    iqr = q3 - q1
    if iqr <= 0 or not math.isfinite(iqr):
        return 0.5
    width = 2.0 * iqr / (len(ordered) ** (1.0 / 3.0))
    if width <= 0 or not math.isfinite(width):
        return 0.5
    return min((0.25, 0.5, 1.0), key=lambda increment: (abs(width - increment), increment))


def _format_target_r_value(value: Any) -> str:
    num = _as_float(value)
    if num is None or not math.isfinite(num):
        return ""
    if abs(num - round(num)) < 1e-9:
        return f"{num:.1f}"
    return f"{num:.2f}".rstrip("0").rstrip(".")


def _target_r_display_decimal(value: Any) -> Decimal | None:
    decimal_value = _decimal_from_value(value)
    if decimal_value is None:
        return None
    return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _finalize_target_r_recommendation_decimal(
    recommended_r_decimal: Decimal,
    current_median_decimal: Decimal | None,
    realized_decimal_values: List[Decimal],
    direction: str,
) -> Tuple[Decimal, str, str]:
    recommended = max(recommended_r_decimal, TARGET_R_RECOMMENDATION_FLOOR)
    recommended_display = _target_r_display_decimal(recommended)
    current_display = _target_r_display_decimal(current_median_decimal) if current_median_decimal is not None else None
    if recommended_display is None:
        return recommended, direction, ""
    if current_display is None:
        return recommended, "Increase target", "missing_current_display"

    if recommended_display > current_display:
        return recommended, "Increase target", "display_actionable"
    if recommended_display < current_display:
        return recommended, "Decrease target", "display_actionable"

    clean = sorted(value for value in realized_decimal_values if value is not None and value > 0)
    preferred_direction = direction if direction in {"Increase target", "Decrease target"} else "Increase target"
    if preferred_direction == "Increase target":
        higher = [
            value for value in clean
            if (_target_r_display_decimal(value) or Decimal("0")) > current_display
        ]
        if higher:
            return max(min(higher), TARGET_R_RECOMMENDATION_FLOOR), "Increase target", "nearest_data_supported_display_increase"
        return max(current_display + Decimal("0.01"), TARGET_R_RECOMMENDATION_FLOOR), "Increase target", "final_0_01_display_increase"

    lower = [
        value for value in clean
        if (_target_r_display_decimal(value) or Decimal("0")) < current_display
        and (_target_r_display_decimal(value) or Decimal("0")) >= TARGET_R_RECOMMENDATION_FLOOR
    ]
    if lower:
        return max(max(lower), TARGET_R_RECOMMENDATION_FLOOR), "Decrease target", "nearest_data_supported_display_decrease"
    lower_adjustment = current_display - Decimal("0.01")
    if lower_adjustment >= TARGET_R_RECOMMENDATION_FLOOR:
        return lower_adjustment, "Decrease target", "final_0_01_display_decrease"
    return current_display + Decimal("0.01"), "Increase target", "floor_prevents_display_decrease"


def _target_r_empty_payload(reason: str = "not_enough_eligible_wins") -> Dict[str, Any]:
    return {
        TARGET_RECOMMENDATION_HEADER: TARGET_RECOMMENDATION_INSUFFICIENT,
        "target_recommendation": TARGET_RECOMMENDATION_INSUFFICIENT,
        "target_r_total_winning_trades": 0,
        "target_r_total_fx_wins": 0,
        "target_r_total_crypto_wins": 0,
        "eligible_target_r_wins": 0,
        "target_r_eligible_winning_trades": 0,
        "target_r_eligible_fx_wins": 0,
        "target_r_eligible_crypto_wins": 0,
        "target_r_increment": None,
        "target_r_distribution": {},
        "target_r_peak_bucket": None,
        "target_r_peak_bucket_low": None,
        "target_r_peak_bucket_high": None,
        "target_r_peak_instances": None,
        "target_r_recommended": None,
        "current_median_original_planned_target_r": None,
        "current_avg_original_planned_target_r": None,
        "target_r_excluded_count": 0,
        "target_r_excluded_winning_trades": 0,
        "target_r_excluded_reasons": {reason: 1} if reason else {},
        "target_r_winning_exclusion_reasons": {},
        "target_r_coverage_note": "",
    }


def _legacy_target_r_recommendation_unused(rows: List[Dict[str, Any]], *, scope: str = "standard") -> Dict[str, Any]:
    normalized_scope = str(scope or "standard").strip().lower()
    overall_scope = normalized_scope == "overall"
    realized_values: List[float] = []
    fx_realized_values: List[float] = []
    crypto_realized_values: List[float] = []
    planned_values: List[float] = []
    excluded: Counter[str] = Counter()
    winning_excluded: Counter[str] = Counter()
    total_wins = 0
    total_fx_wins = 0
    total_crypto_wins = 0
    eligible_fx_wins = 0
    eligible_crypto_wins = 0

    for row in rows or []:
        if not isinstance(row, dict):
            excluded["invalid_row"] += 1
            continue
        planned_r = _planned_target_r_from_original_plan(row)
        if planned_r is not None and math.isfinite(planned_r) and planned_r > 0:
            if str(row.get("row_type") or "trade").strip().lower() == "trade" and not _target_r_row_is_test_trade(row):
                planned_values.append(planned_r)
        is_trade = str(row.get("row_type") or "trade").strip().lower() == "trade"
        is_test = _target_r_row_is_test_trade(row)
        is_winner = is_trade and not is_test and _trade_outcome_sign(row) > 0
        is_fx = _target_r_is_fx_row(row)
        is_crypto = _target_r_is_crypto_row(row) and not is_fx
        if is_winner:
            total_wins += 1
            if is_fx:
                total_fx_wins += 1
            if is_crypto:
                total_crypto_wins += 1
        ok, reason = _is_eligible_target_r_win(row)
        if not ok:
            excluded[reason or "excluded"] += 1
            if is_winner:
                winning_excluded[reason or "excluded"] += 1
            continue
        realized_r, reason = _target_r_realized_from_original_plan(row)
        if realized_r is None:
            excluded[reason or "invalid_realized_r"] += 1
            if is_winner:
                winning_excluded[reason or "invalid_realized_r"] += 1
            continue
        realized_values.append(realized_r)
        if is_fx:
            eligible_fx_wins += 1
            fx_realized_values.append(realized_r)
        if is_crypto:
            eligible_crypto_wins += 1
            crypto_realized_values.append(realized_r)

    current_median = _target_r_percentile(sorted(planned_values), 0.5) if planned_values else None
    current_avg = (sum(planned_values) / len(planned_values)) if planned_values else None
    fx_sufficient = bool(eligible_fx_wins)
    crypto_sufficient = bool(eligible_crypto_wins)
    coverage_note = ""
    coverage_label = ""
    recommendation_values = realized_values
    if overall_scope:
        if crypto_sufficient and not fx_sufficient:
            coverage_note = "FX: insufficient eligible wins"
            coverage_label = "Crypto data"
            recommendation_values = crypto_realized_values
        elif fx_sufficient and not crypto_sufficient:
            coverage_note = "Crypto: insufficient eligible wins"
            coverage_label = "FX data"
            recommendation_values = fx_realized_values
        elif not fx_sufficient and not crypto_sufficient:
            recommendation_values = []
    coverage_payload = {
        "target_r_scope": "overall" if overall_scope else normalized_scope,
        "target_r_total_winning_trades": total_wins,
        "target_r_total_fx_wins": total_fx_wins,
        "target_r_total_crypto_wins": total_crypto_wins,
        "target_r_eligible_winning_trades": len(realized_values),
        "target_r_eligible_fx_wins": eligible_fx_wins,
        "target_r_eligible_crypto_wins": eligible_crypto_wins,
        "target_r_excluded_winning_trades": sum(winning_excluded.values()),
        "target_r_winning_exclusion_reasons": dict(winning_excluded),
        "target_r_coverage_note": coverage_note,
    }
    if not recommendation_values:
        payload = _target_r_empty_payload("not_enough_eligible_wins")
        payload["eligible_target_r_wins"] = len(recommendation_values)
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        payload.update(coverage_payload)
        return payload

    increment = _choose_target_r_increment(recommendation_values)
    bucket_values = _target_r_bucket_values(recommendation_values, increment)
    if not bucket_values:
        payload = _target_r_empty_payload("empty_distribution")
        payload["eligible_target_r_wins"] = len(recommendation_values)
        payload["target_r_increment"] = increment
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        payload.update(coverage_payload)
        return payload

    peak_bucket, modal_values = max(
        bucket_values.items(),
        key=lambda item: (len(item[1]), -item[0]),
    )
    modal_values_sorted = sorted(modal_values)
    recommended_r = _target_r_percentile(modal_values_sorted, 0.5)
    if recommended_r is None or not math.isfinite(recommended_r) or recommended_r <= 0:
        payload = _target_r_empty_payload("invalid_modal_recommended_r")
        payload["eligible_target_r_wins"] = len(recommendation_values)
        payload["target_r_increment"] = increment
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        payload.update(coverage_payload)
        return payload
    peak_instances = len(modal_values)
    distribution = OrderedDict(
        (_target_r_bucket_label(bucket, increment), len(values))
        for bucket, values in bucket_values.items()
    )
    peak_bucket_label = _target_r_bucket_label(peak_bucket, increment)
    if current_median is None:
        direction = "Increase target"
    elif recommended_r < current_median:
        direction = "Decrease target"
    elif recommended_r > current_median:
        direction = "Increase target"
    else:
        direction = "Increase target"
    if coverage_label and coverage_note:
        recommendation = f"{coverage_label} — Recommended: {_format_target_r_value(recommended_r)}R — {coverage_note}"
    else:
        recommendation = f"{direction} — Recommended: {_format_target_r_value(recommended_r)}R"
    if current_median is not None and not coverage_label:
        recommendation += f" (current median: {_format_target_r_value(current_median)}R)"

    return {
        TARGET_RECOMMENDATION_HEADER: recommendation,
        "target_recommendation": recommendation,
        **coverage_payload,
        "eligible_target_r_wins": len(recommendation_values),
        "target_r_increment": increment,
        "target_r_distribution": dict(distribution),
        "target_r_peak_bucket": peak_bucket_label,
        "target_r_peak_bucket_low": peak_bucket,
        "target_r_peak_bucket_high": peak_bucket + increment,
        "target_r_peak_instances": peak_instances,
        "target_r_recommended": recommended_r,
        "current_median_original_planned_target_r": current_median,
        "current_avg_original_planned_target_r": current_avg,
        "target_r_excluded_count": sum(excluded.values()),
        "target_r_excluded_reasons": dict(excluded),
    }


def _is_eligible_target_r_loss(row: Dict[str, Any]) -> Tuple[bool, str]:
    if str(row.get("row_type") or "trade").strip().lower() != "trade":
        return False, "not_trade"
    if _target_r_row_is_test_trade(row):
        return False, "test_trade"
    if _trade_outcome_sign(row) >= 0:
        return False, "not_losing_trade"
    stop_pct, reason = _original_stop_distance_pct_points(row)
    if stop_pct is None:
        return False, reason or "invalid_original_stop_distance"
    return True, ""


def _target_r_small_adjustment_step(values: List[float], reference: float, increment: float | None) -> float:
    candidates = [
        abs(value - reference)
        for value in values
        if value is not None and math.isfinite(value) and value != reference
    ]
    if increment is not None and math.isfinite(increment) and increment > 0:
        candidates.append(increment)
    candidates.append(0.01)
    return min(value for value in candidates if value > 0)


def _target_r_exact_tie_adjustment(
    values: List[float],
    current_median: float,
    increment: float | None,
) -> Tuple[float, str, str, bool]:
    mean_value = _average_float(values)
    if mean_value is not None and mean_value > current_median:
        direction = "Increase target"
        reason = "mean_captured_r_above_current_median"
        goal_preference = False
    elif mean_value is not None and mean_value < current_median:
        direction = "Decrease target"
        reason = "mean_captured_r_below_current_median"
        goal_preference = False
    else:
        direction = "Increase target"
        reason = "exact_tie_goal_preference_increase"
        goal_preference = True

    clean = sorted(
        value for value in values
        if value is not None and math.isfinite(value) and value > 0
    )
    if direction == "Increase target":
        higher = [value for value in clean if value > current_median]
        recommended = min(higher) if higher else current_median + _target_r_small_adjustment_step(clean, current_median, increment)
    else:
        lower = [value for value in clean if value < current_median]
        recommended = max(lower) if lower else max(0.01, current_median - _target_r_small_adjustment_step(clean, current_median, increment))
    return recommended, direction, reason, goal_preference


def _target_r_exact_tie_adjustment_decimal(
    values: List[Any],
    current_median: Decimal,
    increment: float | None,
) -> Tuple[Decimal, str, str, bool]:
    clean = sorted(value for value in (_decimal_from_value(v) for v in values) if value is not None and value > 0)
    mean_value = _average_decimal(clean)
    if mean_value is not None and mean_value > current_median:
        direction = "Increase target"
        reason = "mean_captured_r_above_current_median"
        goal_preference = False
    elif mean_value is not None and mean_value < current_median:
        direction = "Decrease target"
        reason = "mean_captured_r_below_current_median"
        goal_preference = False
    else:
        direction = "Increase target"
        reason = "exact_tie_goal_preference_increase"
        goal_preference = True

    step = Decimal(str(_target_r_small_adjustment_step([float(value) for value in clean], float(current_median), increment)))
    if direction == "Increase target":
        higher = [value for value in clean if value > current_median]
        recommended = min(higher) if higher else current_median + step
    else:
        lower = [value for value in clean if value < current_median]
        recommended = max(lower) if lower else max(Decimal("0.01"), current_median - step)
    return recommended, direction, reason, goal_preference


def _target_r_empty_payload(reason: str = TARGET_RECOMMENDATION_NO_WINS) -> Dict[str, Any]:
    return {
        TARGET_RECOMMENDATION_HEADER: "",
        "target_recommendation": "",
        "target_recommendation_missing_reason": reason,
        "target_r_missing_reason": reason,
        "target_r_scope": "standard",
        "target_r_total_winning_trades": 0,
        "target_r_total_losing_trades": 0,
        "target_r_total_fx_wins": 0,
        "target_r_total_crypto_wins": 0,
        "eligible_target_r_wins": 0,
        "eligible_target_r_losses": 0,
        "target_r_eligible_winning_trades": 0,
        "target_r_eligible_losing_trades": 0,
        "target_r_eligible_fx_wins": 0,
        "target_r_eligible_crypto_wins": 0,
        "target_r_eligible_fx_losses": 0,
        "target_r_eligible_crypto_losses": 0,
        "target_r_increment": None,
        "target_r_distribution": {},
        "target_r_peak_bucket": None,
        "target_r_peak_bucket_low": None,
        "target_r_peak_bucket_high": None,
        "target_r_peak_instances": None,
        "target_r_recommended": None,
        "target_r_histogram_recommended_before_tie": None,
        "target_r_recommendation_direction": None,
        "target_r_exact_tie": False,
        "target_r_exact_tie_goal_preference": False,
        "target_r_tie_break_reason": "",
        "current_median_original_planned_target_r": None,
        "current_avg_original_planned_target_r": None,
        "target_r_excluded_count": 0,
        "target_r_excluded_winning_trades": 0,
        "target_r_excluded_losing_trades": 0,
        "target_r_excluded_reasons": {},
        "target_r_winning_exclusion_reasons": {},
        "target_r_losing_exclusion_reasons": {},
        "target_r_coverage_note": "",
    }


def _target_r_recommendation(rows: List[Dict[str, Any]], *, scope: str = "standard") -> Dict[str, Any]:
    normalized_scope = str(scope or "standard").strip().lower()
    realized_values: List[float] = []
    realized_decimal_values: List[Decimal] = []
    planned_values: List[float] = []
    planned_decimal_values: List[Decimal] = []
    excluded: Counter[str] = Counter()
    winning_excluded: Counter[str] = Counter()
    losing_excluded: Counter[str] = Counter()
    method_counts: Counter[str] = Counter()
    method_counts_by_market: Dict[str, Counter[str]] = {
        "overall": Counter(),
        "fx": Counter(),
        "crypto": Counter(),
    }
    winning_excluded_by_market: Dict[str, Counter[str]] = {
        "overall": Counter(),
        "fx": Counter(),
        "crypto": Counter(),
    }
    total_wins = 0
    total_losses = 0
    total_fx_wins = 0
    total_crypto_wins = 0
    eligible_fx_wins = 0
    eligible_crypto_wins = 0
    eligible_losses = 0
    eligible_fx_losses = 0
    eligible_crypto_losses = 0

    for row in rows or []:
        if not isinstance(row, dict):
            excluded["invalid_row"] += 1
            continue
        is_trade = str(row.get("row_type") or "trade").strip().lower() == "trade"
        is_test = _target_r_row_is_test_trade(row)
        is_fx = _target_r_is_fx_row(row)
        is_crypto = _target_r_is_crypto_row(row) and not is_fx
        if not is_trade:
            excluded["not_trade"] += 1
            continue
        if is_test:
            excluded["test_trade"] += 1
            continue
        outcome = _trade_outcome_sign(row)
        if outcome > 0:
            total_wins += 1
            if is_fx:
                total_fx_wins += 1
            if is_crypto:
                total_crypto_wins += 1
            realized_r, reason, method, realized_r_decimal = _target_r_realized_from_original_plan_detail(row)
            if realized_r is None:
                excluded[reason or "invalid_realized_r"] += 1
                winning_excluded[reason or "invalid_realized_r"] += 1
                winning_excluded_by_market["overall"][reason or "invalid_realized_r"] += 1
                if is_fx:
                    winning_excluded_by_market["fx"][reason or "invalid_realized_r"] += 1
                if is_crypto:
                    winning_excluded_by_market["crypto"][reason or "invalid_realized_r"] += 1
                continue
            planned_r = _planned_target_r_from_original_plan(row)
            planned_r_decimal = _planned_target_r_from_original_plan_decimal(row)
            if planned_r is None or not math.isfinite(planned_r) or planned_r <= 0 or planned_r_decimal is None or planned_r_decimal <= 0:
                excluded["invalid_original_planned_target_r"] += 1
                winning_excluded["invalid_original_planned_target_r"] += 1
                winning_excluded_by_market["overall"]["invalid_original_planned_target_r"] += 1
                if is_fx:
                    winning_excluded_by_market["fx"]["invalid_original_planned_target_r"] += 1
                if is_crypto:
                    winning_excluded_by_market["crypto"]["invalid_original_planned_target_r"] += 1
                continue
            if realized_r_decimal is None:
                realized_r_decimal = _decimal_from_value(realized_r)
            if realized_r_decimal is None:
                excluded["invalid_realized_r"] += 1
                winning_excluded["invalid_realized_r"] += 1
                winning_excluded_by_market["overall"]["invalid_realized_r"] += 1
                if is_fx:
                    winning_excluded_by_market["fx"]["invalid_realized_r"] += 1
                if is_crypto:
                    winning_excluded_by_market["crypto"]["invalid_realized_r"] += 1
                continue
            realized_values.append(realized_r)
            realized_decimal_values.append(realized_r_decimal)
            planned_values.append(planned_r)
            planned_decimal_values.append(planned_r_decimal)
            method_key = method or "unknown"
            method_counts[method_key] += 1
            method_counts_by_market["overall"][method_key] += 1
            if is_fx:
                eligible_fx_wins += 1
                method_counts_by_market["fx"][method_key] += 1
            if is_crypto:
                eligible_crypto_wins += 1
                method_counts_by_market["crypto"][method_key] += 1
            continue
        if outcome < 0:
            total_losses += 1
            ok_loss, reason = _is_eligible_target_r_loss(row)
            if ok_loss:
                eligible_losses += 1
                if is_fx:
                    eligible_fx_losses += 1
                if is_crypto:
                    eligible_crypto_losses += 1
            else:
                excluded[reason or "invalid_losing_trade"] += 1
                losing_excluded[reason or "invalid_losing_trade"] += 1
            continue
        excluded["break_even_or_zero"] += 1

    current_median_decimal = _percentile_decimal(planned_decimal_values, 0.5) if planned_decimal_values else None
    current_avg_decimal = _average_decimal(planned_decimal_values) if planned_decimal_values else None
    current_median = _decimal_to_float(current_median_decimal)
    current_avg = _decimal_to_float(current_avg_decimal)
    coverage_payload = {
        "target_r_scope": "overall" if normalized_scope == "overall" else normalized_scope,
        "target_r_total_winning_trades": total_wins,
        "target_r_total_losing_trades": total_losses,
        "target_r_total_fx_wins": total_fx_wins,
        "target_r_total_crypto_wins": total_crypto_wins,
        "target_r_eligible_winning_trades": len(realized_values),
        "target_r_eligible_losing_trades": eligible_losses,
        "target_r_eligible_fx_wins": eligible_fx_wins,
        "target_r_eligible_crypto_wins": eligible_crypto_wins,
        "target_r_eligible_fx_losses": eligible_fx_losses,
        "target_r_eligible_crypto_losses": eligible_crypto_losses,
        "target_r_excluded_winning_trades": sum(winning_excluded.values()),
        "target_r_winning_exclusion_reasons": dict(winning_excluded),
        "target_r_winning_exclusion_reasons_by_market": {
            market: dict(counter) for market, counter in winning_excluded_by_market.items()
        },
        "target_r_excluded_losing_trades": sum(losing_excluded.values()),
        "target_r_losing_exclusion_reasons": dict(losing_excluded),
        "target_r_calculation_method_counts": dict(method_counts),
        "target_r_calculation_method_counts_by_market": {
            market: dict(counter) for market, counter in method_counts_by_market.items()
        },
        "target_r_coverage_note": "",
    }

    if not realized_values:
        payload = _target_r_empty_payload(TARGET_RECOMMENDATION_NO_WINS)
        payload.update(coverage_payload)
        payload["eligible_target_r_losses"] = eligible_losses
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        return payload
    if eligible_losses < 1:
        payload = _target_r_empty_payload(TARGET_RECOMMENDATION_NO_LOSSES)
        payload.update(coverage_payload)
        payload["eligible_target_r_wins"] = len(realized_values)
        payload["eligible_target_r_losses"] = eligible_losses
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        return payload

    increment = _choose_target_r_increment(realized_values)
    bucket_values = _target_r_bucket_values(realized_values, increment)
    if not bucket_values:
        payload = _target_r_empty_payload("empty_distribution")
        payload.update(coverage_payload)
        payload["eligible_target_r_wins"] = len(realized_values)
        payload["eligible_target_r_losses"] = eligible_losses
        payload["target_r_increment"] = increment
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        return payload

    peak_bucket, modal_values = max(bucket_values.items(), key=lambda item: (len(item[1]), -item[0]))
    bucket_decimal_values: Dict[float, List[Decimal]] = defaultdict(list)
    for value, decimal_value in zip(realized_values, realized_decimal_values):
        bucket = math.floor(value / increment) * increment
        bucket_decimal_values[bucket].append(decimal_value)
    modal_values_decimal = sorted(bucket_decimal_values.get(peak_bucket) or [])
    recommended_r_decimal = _percentile_decimal(modal_values_decimal, 0.5)
    recommended_r = _decimal_to_float(recommended_r_decimal)
    if recommended_r_decimal is None or recommended_r is None or not math.isfinite(recommended_r) or recommended_r_decimal <= 0:
        payload = _target_r_empty_payload("invalid_modal_recommended_r")
        payload.update(coverage_payload)
        payload["eligible_target_r_wins"] = len(realized_values)
        payload["eligible_target_r_losses"] = eligible_losses
        payload["target_r_increment"] = increment
        payload["current_median_original_planned_target_r"] = current_median
        payload["current_avg_original_planned_target_r"] = current_avg
        payload["target_r_excluded_count"] = sum(excluded.values())
        payload["target_r_excluded_reasons"] = dict(excluded)
        return payload

    peak_instances = len(modal_values)
    distribution = OrderedDict(
        (_target_r_bucket_label(bucket, increment), len(values))
        for bucket, values in bucket_values.items()
    )
    peak_bucket_label = _target_r_bucket_label(peak_bucket, increment)
    raw_recommended_r = recommended_r
    exact_tie = False
    exact_tie_goal_preference = False
    tie_break_reason = ""
    if current_median_decimal is None:
        direction = "Increase target"
        tie_break_reason = "missing_current_median_goal_preference_increase"
    elif recommended_r_decimal < current_median_decimal:
        direction = "Decrease target"
    elif recommended_r_decimal > current_median_decimal:
        direction = "Increase target"
    else:
        exact_tie = True
        recommended_r_decimal, direction, tie_break_reason, exact_tie_goal_preference = _target_r_exact_tie_adjustment_decimal(
            realized_decimal_values,
            current_median_decimal,
            increment,
        )
        recommended_r = float(recommended_r_decimal)
    recommended_before_actionable_adjustment = _decimal_to_float(recommended_r_decimal)
    recommended_r_decimal, direction, actionable_adjustment_reason = _finalize_target_r_recommendation_decimal(
        recommended_r_decimal,
        current_median_decimal,
        realized_decimal_values,
        direction,
    )
    recommended_r = float(recommended_r_decimal)
    recommendation = f"{direction} \u2014 Recommended: {_format_target_r_value(recommended_r)}R"
    if current_median is not None:
        recommendation += f" (current median: {_format_target_r_value(current_median)}R)"

    return {
        TARGET_RECOMMENDATION_HEADER: recommendation,
        "target_recommendation": recommendation,
        **coverage_payload,
        "target_r_missing_reason": "",
        "eligible_target_r_wins": len(realized_values),
        "eligible_target_r_losses": eligible_losses,
        "target_r_increment": increment,
        "target_r_distribution": dict(distribution),
        "target_r_peak_bucket": peak_bucket_label,
        "target_r_peak_bucket_low": peak_bucket,
        "target_r_peak_bucket_high": peak_bucket + increment,
        "target_r_peak_instances": peak_instances,
        "target_r_recommended": recommended_r,
        "target_r_histogram_recommended_before_tie": raw_recommended_r,
        "target_r_recommended_before_actionable_adjustment": recommended_before_actionable_adjustment,
        "target_r_recommendation_direction": direction,
        "target_r_exact_tie": exact_tie,
        "target_r_exact_tie_goal_preference": exact_tie_goal_preference,
        "target_r_tie_break_reason": tie_break_reason,
        "target_r_actionable_adjustment_reason": actionable_adjustment_reason,
        "current_median_original_planned_target_r": current_median,
        "current_avg_original_planned_target_r": current_avg,
        "target_r_excluded_count": sum(excluded.values()),
        "target_r_excluded_reasons": dict(excluded),
    }


def _original_stop_distance_pct_points(row: Dict[str, Any]) -> Tuple[float | None, str]:
    decimal_pct, reason = _original_stop_distance_pct_points_decimal(row)
    if decimal_pct is not None:
        return float(decimal_pct), ""
    if reason:
        return None, reason
    levels = _original_plan_levels(row)
    if levels.get("movement_evidence") and not levels.get("explicit"):
        return None, "moved_without_original_stop"
    entry = _as_float(levels.get("entry"))
    stop = _as_float(levels.get("stop"))
    if entry is None:
        return None, "missing_original_entry"
    if stop is None:
        return None, "missing_original_stop"
    if not (math.isfinite(entry) and math.isfinite(stop)):
        return None, "invalid_original_stop"
    if entry <= 0 or stop <= 0:
        return None, "invalid_original_stop"
    side = _target_r_trade_side(row)
    if side == "long" and not stop < entry:
        return None, "invalid_long_original_stop"
    if side == "short" and not stop > entry:
        return None, "invalid_short_original_stop"
    if not side:
        return None, "missing_trade_side"
    pct = abs(entry - stop) / entry * 100.0
    if not math.isfinite(pct) or pct <= 0:
        return None, "invalid_original_stop_distance"
    if _trade_row_market(row) == "fx" and pct > 50.0:
        return None, "invalid_fx_original_stop_distance"
    return pct, ""


def _original_stop_distance_pct_points_decimal(row: Dict[str, Any]) -> Tuple[Decimal | None, str]:
    levels = _original_plan_levels_decimal(row)
    if levels.get("movement_evidence") and not levels.get("explicit"):
        return None, "moved_without_original_stop"
    if not levels.get("movement_evidence"):
        recorded_pct, recorded_reason = _recorded_distance_pct_decimal(row, "stop_loss")
        if recorded_pct is not None:
            return recorded_pct, ""
    entry = levels.get("entry")
    stop = levels.get("stop")
    if not isinstance(entry, Decimal):
        return None, "missing_original_entry"
    if not isinstance(stop, Decimal):
        return None, "missing_original_stop"
    if entry <= 0 or stop <= 0:
        return None, "invalid_original_stop"
    side = _target_r_trade_side(row)
    if side == "long" and not stop < entry:
        return None, "invalid_long_original_stop"
    if side == "short" and not stop > entry:
        return None, "invalid_short_original_stop"
    if not side:
        return None, "missing_trade_side"
    with localcontext() as ctx:
        ctx.prec = 40
        pct = abs(entry - stop) / entry * Decimal("100")
    if pct <= 0:
        return None, "invalid_original_stop_distance"
    if _trade_row_market(row) == "fx" and pct > Decimal("50"):
        return None, "invalid_fx_original_stop_distance"
    return pct, ""


def _stop_small_adjustment_step(values: List[float], reference: float) -> float:
    candidates = [
        abs(value - reference)
        for value in values
        if value is not None and math.isfinite(value) and value != reference
    ]
    candidates.append(0.01)
    return min(value for value in candidates if value > 0)


def _stop_exact_tie_adjustment(
    winner_values: List[Any],
    loser_values: List[Any],
    reference: Decimal,
) -> Tuple[Decimal, str, str, bool]:
    winners = sorted(value for value in (_decimal_from_value(v) for v in winner_values) if value is not None and value > 0)
    losers = sorted(value for value in (_decimal_from_value(v) for v in loser_values) if value is not None and value > 0)
    win_median = _percentile_decimal(winners, 0.5)
    loss_median = _percentile_decimal(losers, 0.5)
    if win_median is not None and loss_median is not None and win_median < loss_median:
        direction = "decrease"
        reason = "winner_median_below_loser_median"
        goal_preference = False
    elif win_median is not None and loss_median is not None and win_median > loss_median:
        direction = "increase"
        reason = "winner_median_above_loser_median"
        goal_preference = False
    else:
        direction = "decrease"
        reason = "exact_tie_goal_preference_decrease"
        goal_preference = True

    if direction == "increase":
        higher = [value for value in winners if value > reference]
        step = Decimal(str(_stop_small_adjustment_step([float(value) for value in winners + losers], float(reference))))
        recommended = min(higher) if higher else reference + step
    else:
        lower = [value for value in winners if value < reference]
        step = Decimal(str(_stop_small_adjustment_step([float(value) for value in winners + losers], float(reference))))
        recommended = max(lower) if lower else max(Decimal("0.01"), reference - step)
    return recommended, direction, reason, goal_preference


def _stop_recommendation_text(direction: str, recommended: float, loss_mean: float, *, tie_reason: str = "") -> str:
    gap = recommended - loss_mean
    if direction == "decrease":
        relation = "below"
    else:
        relation = "above"
    visible_reason = RECOMMENDATION_REASON_TEXT.get(str(tie_reason or "").strip(), str(tie_reason or "").strip())
    suffix = f"; {visible_reason}" if visible_reason else ""
    return (
        f"{'Decrease' if direction == 'decrease' else 'Increase'} stop \u2014 "
        f"Recommended: {_format_stop_pct_value(recommended)} "
        f"({_format_stop_gap_value(gap)} {relation} loss average{suffix})"
    )


def _stop_recommendation_payload(
    winner_values: List[float],
    loser_values: List[float],
    excluded: Counter[str] | None = None,
) -> Dict[str, Any]:
    win_mean_dec = _average_decimal(winner_values)
    loss_mean_dec = _average_decimal(loser_values)
    win_mean = _decimal_to_float(win_mean_dec)
    loss_mean = _decimal_to_float(loss_mean_dec)
    recommendation = ""
    gap = None
    direction = None
    recommended = None
    recommended_gap = None
    exact_tie = False
    exact_tie_goal_preference = False
    tie_break_reason = ""
    if win_mean_dec is not None and loss_mean_dec is not None:
        gap_dec = abs(win_mean_dec - loss_mean_dec)
        gap = float(gap_dec)
        if win_mean_dec == loss_mean_dec:
            exact_tie = True
            recommended_dec, direction, tie_break_reason, exact_tie_goal_preference = _stop_exact_tie_adjustment(
                winner_values,
                loser_values,
                loss_mean_dec,
            )
            recommended = float(recommended_dec)
            recommended_gap = float(abs(recommended_dec - loss_mean_dec))
            recommendation = _stop_recommendation_text(
                direction,
                recommended,
                loss_mean,
                tie_reason=tie_break_reason,
            )
        elif win_mean_dec < loss_mean_dec:
            direction = "decrease"
            recommended = win_mean
            recommended_gap = gap
            recommendation = _stop_recommendation_text(direction, recommended, loss_mean)
        else:
            direction = "increase"
            recommended = win_mean
            recommended_gap = gap
            recommendation = _stop_recommendation_text(direction, recommended, loss_mean)
    return {
        STOP_RECOMMENDATION_HEADER: recommendation,
        "stop_recommendation": recommendation,
        "stop_recommendation_missing_reason": (
            ""
            if recommendation
            else "no_eligible_winning_trades"
            if not winner_values
            else "no_eligible_losing_trades"
        ),
        "eligible_stop_loss_wins": len(winner_values),
        "eligible_stop_loss_losses": len(loser_values),
        "stop_loss_winner_mean_pct": win_mean,
        "stop_loss_loser_mean_pct": loss_mean,
        "stop_loss_difference_pp": gap,
        "stop_loss_recommended_pct": recommended,
        "stop_loss_recommended_difference_pp": recommended_gap,
        "stop_loss_recommendation_direction": direction,
        "stop_loss_exact_tie": exact_tie,
        "stop_loss_exact_tie_goal_preference": exact_tie_goal_preference,
        "stop_loss_tie_break_reason": tie_break_reason,
        "stop_loss_excluded_reasons": dict(excluded or {}),
    }


def _distance_recommendation_summary(rows: List[Dict[str, Any]], *, scope: str = "standard") -> Dict[str, Any]:
    buckets = {
        "stop_wins": [],
        "stop_losses": [],
    }
    stop_excluded: Counter[str] = Counter()
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            stop_excluded["not_trade"] += 1
            continue
        if _target_r_row_is_test_trade(row):
            stop_excluded["test_trade"] += 1
            continue
        outcome = _trade_outcome_sign(row)
        if outcome == 0:
            stop_excluded["break_even_or_zero"] += 1
            continue
        stop_pct, reason = _original_stop_distance_pct_points_decimal(row)
        if stop_pct is None:
            stop_excluded[reason or "invalid_original_stop_distance"] += 1
            continue
        suffix = "wins" if outcome > 0 else "losses"
        buckets[f"stop_{suffix}"].append(stop_pct)
    target_payload = _target_r_recommendation(rows, scope=scope)
    return {
        **_stop_recommendation_payload(
            buckets["stop_wins"],
            buckets["stop_losses"],
            stop_excluded,
        ),
        **target_payload,
    }


def _distance_recommendations_by_symbol(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    active: List[Dict[str, Any]] = []
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        if _target_r_row_is_test_trade(row):
            continue
        active.append(row)
        symbol = str(row.get("symbol") or "").strip().upper()
        if symbol:
            grouped[symbol].append(row)
    overall = _distance_recommendation_summary(active, scope="overall")
    out = {"": overall}
    for symbol, symbol_rows in grouped.items():
        out[symbol] = _distance_recommendation_summary(symbol_rows)
    return out


def _recommendation_html_escape(value: Any) -> str:
    return _xml_escape(str(value if value is not None else ""), {'"': "&quot;", "'": "&#39;"})


def _recommendation_trade_identity(row: Mapping[str, Any]) -> Dict[str, str]:
    symbol = str(row.get("symbol") or row.get("instrument") or "-").strip().upper() or "-"
    trade_date = _as_date(row.get("close_time") or row.get("open_time"))
    row_id = str(row.get("id") or row.get("row_id") or stable_row_id(dict(row)) or "-").strip()
    return {
        "symbol": symbol,
        "date": trade_date.isoformat() if trade_date else "-",
        "id": row_id,
        "label": f"{symbol} {trade_date.isoformat() if trade_date else '-'} ({row_id})",
    }


def _recommendation_chart_observations(
    rows: Sequence[Dict[str, Any]],
    metric: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    eligible: List[Dict[str, Any]] = []
    excluded: List[Dict[str, Any]] = []
    for row in rows:
        identity = _recommendation_trade_identity(row)
        if metric == "stop":
            outcome = _trade_outcome_sign(row)
            if outcome == 0:
                excluded.append({**identity, "reason": "break_even_or_zero"})
                continue
            value, reason = _original_stop_distance_pct_points_decimal(row)
            if value is None:
                excluded.append({**identity, "reason": reason or "invalid_original_stop_distance"})
                continue
            eligible.append({
                **identity,
                "value": float(value),
                "outcome": "winner" if outcome > 0 else "loser",
                "context": "original stop distance",
            })
            continue

        outcome = _trade_outcome_sign(row)
        if outcome > 0:
            realized_r, reason, method, _realized_decimal = _target_r_realized_from_original_plan_detail(row)
            planned_r = _planned_target_r_from_original_plan(row)
            planned_decimal = _planned_target_r_from_original_plan_decimal(row)
            if realized_r is None:
                excluded.append({**identity, "reason": reason or "invalid_realized_r"})
                continue
            if planned_r is None or not math.isfinite(planned_r) or planned_r <= 0 or planned_decimal is None or planned_decimal <= 0:
                excluded.append({**identity, "reason": "invalid_original_planned_target_r"})
                continue
            eligible.append({
                **identity,
                "value": float(realized_r),
                "outcome": "winner",
                "context": f"realized from original plan; planned {planned_r:.4f}R; {method or 'unknown method'}",
            })
        elif outcome < 0:
            ok_loss, reason = _is_eligible_target_r_loss(row)
            if ok_loss:
                eligible.append({
                    **identity,
                    "value": None,
                    "outcome": "eligible loss gate",
                    "context": "qualifies the minimum losing-trade coverage gate",
                })
            else:
                excluded.append({**identity, "reason": reason or "invalid_losing_trade"})
        else:
            excluded.append({**identity, "reason": "break_even_or_zero"})
    def identity_key(item: Mapping[str, Any]) -> Tuple[str, str, str, str]:
        return (
            str(item.get("symbol") or "").casefold(),
            str(item.get("date") or ""),
            str(item.get("id") or ""),
            str(item.get("outcome") or item.get("reason") or "").casefold(),
        )

    return sorted(eligible, key=identity_key), sorted(excluded, key=identity_key)


def _recommendation_svg(
    observations: Sequence[Dict[str, Any]],
    recommended: float,
    unit: str,
) -> str:
    points = [item for item in observations if _as_float(item.get("value")) is not None]
    values = [float(item["value"]) for item in points]
    all_values = [*values, float(recommended), 0.0]
    y_min = min(all_values)
    y_max = max(all_values)
    if y_min == y_max:
        y_max = y_min + 1.0
    padding = max((y_max - y_min) * 0.08, 0.01)
    y_min -= padding
    y_max += padding
    # The chart deliberately keeps generous viewBox padding because the SVG is
    # responsive: at narrow browser widths every label scales with the plot.
    # Keep these margins in sync with the focused bounds regression test.
    width, height = 1040, 520
    left, right, top, bottom = 118, 44, 48, 104
    plot_width = width - left - right
    plot_height = height - top - bottom
    point_inset = 10.0

    def x_pos(index: int) -> float:
        usable_width = max(0.0, plot_width - (point_inset * 2.0))
        return left + point_inset + (
            usable_width / 2.0
            if len(points) <= 1
            else index * usable_width / (len(points) - 1)
        )

    def y_pos(value: float) -> float:
        return top + (y_max - value) / (y_max - y_min) * plot_height

    svg: List[str] = [
        f'<svg class="chart" viewBox="0 0 {width} {height}" '
        f'data-plot-left="{left}" data-plot-right="{width-right}" '
        f'data-plot-top="{top}" data-plot-bottom="{height-bottom}" '
        'role="img" aria-label="Eligible trade observations and recommended value">',
        f'<rect data-role="plot-border" x="{left}" y="{top}" width="{plot_width}" height="{plot_height}" fill="#fbfdff" stroke="#cbd5e1"/>',
    ]
    for tick in range(6):
        value = y_min + (y_max - y_min) * tick / 5.0
        y = y_pos(value)
        svg.append(f'<line x1="{left}" y1="{y:.2f}" x2="{width-right}" y2="{y:.2f}" stroke="#e2e8f0"/>')
        svg.append(f'<text data-role="tick-label" x="{left-14}" y="{y+4:.2f}" text-anchor="end" class="axis-label">{value:.3f}{_recommendation_html_escape(unit)}</text>')
    rec_y = y_pos(float(recommended))
    svg.append(f'<line x1="{left}" y1="{rec_y:.2f}" x2="{width-right}" y2="{rec_y:.2f}" stroke="#7c3aed" stroke-width="3" stroke-dasharray="10 6"/>')
    recommendation_label_y = (
        rec_y + 22.0 if rec_y - top < 28.0 else rec_y - 10.0
    )
    recommendation_label_y = min(
        height - bottom - 10.0,
        max(top + 16.0, recommendation_label_y),
    )
    svg.append(f'<text data-role="recommendation-label" x="{width-right-8}" y="{recommendation_label_y:.2f}" text-anchor="end" class="recommendation-label">Recommended {recommended:.8g}{_recommendation_html_escape(unit)}</text>')
    for index, item in enumerate(points):
        value = float(item["value"])
        colour = "#15803d" if item.get("outcome") == "winner" else "#b91c1c"
        title = _recommendation_html_escape(f"{item['label']}: {value:.8g}{unit} — {item.get('outcome')}")
        svg.append(
            f'<circle data-role="observation" cx="{x_pos(index):.2f}" cy="{y_pos(value):.2f}" r="5.5" fill="{colour}" stroke="#ffffff" stroke-width="1.5"><title>{title}</title></circle>'
        )
    svg.extend([
        f'<text data-role="x-axis-title" x="{left + plot_width/2:.2f}" y="{height-28}" text-anchor="middle" class="axis-title">Eligible trade observations (deterministic input order)</text>',
        f'<text data-role="y-axis-title" x="26" y="{top + plot_height/2:.2f}" text-anchor="middle" transform="rotate(-90 26 {top + plot_height/2:.2f})" class="axis-title">Value ({_recommendation_html_escape(unit or 'R')})</text>',
        "</svg>",
    ])
    return "".join(svg)


def _recommendation_chart_html(
    *,
    scope_label: str,
    metric: str,
    payload: Mapping[str, Any],
    observations: Sequence[Dict[str, Any]],
    excluded: Sequence[Dict[str, Any]],
) -> str:
    if metric == "stop":
        recommendation = str(payload.get(STOP_RECOMMENDATION_HEADER) or payload.get("stop_recommendation") or "")
        recommended = _as_float(payload.get("stop_loss_recommended_pct"))
        unit = "%"
        metric_label = "Stop-loss recommendation"
        eligible_count = int(payload.get("eligible_stop_loss_wins") or 0) + int(payload.get("eligible_stop_loss_losses") or 0)
        exclusion_summary = payload.get("stop_loss_excluded_reasons") or {}
        method = "Recommended stop is derived by the existing winner/loser original-stop mean comparison and its deterministic tie rule."
    else:
        recommendation = str(payload.get(TARGET_RECOMMENDATION_HEADER) or payload.get("target_recommendation") or "")
        recommended = _as_float(payload.get("target_r_recommended"))
        unit = "R"
        metric_label = "Target recommendation"
        eligible_count = int(payload.get("target_r_eligible_winning_trades") or payload.get("eligible_target_r_wins") or 0)
        exclusion_summary = payload.get("target_r_excluded_reasons") or {}
        method = "Recommended target is the existing modal realized-R bucket median, including its floor, coverage gate, and deterministic tie adjustment."
    if recommended is None:
        raise ValueError(f"Cannot render {metric} recommendation without a numeric marker")
    numeric_observations = [item for item in observations if _as_float(item.get("value")) is not None]
    gate_observations = [item for item in observations if item.get("value") is None]
    observation_rows: List[str] = []
    for item in observations:
        display_value = "" if item.get("value") is None else f"{float(item['value']):.8g}{unit}"
        observation_rows.append(
            "<tr>"
            f"<td>{_recommendation_html_escape(item.get('symbol'))}</td>"
            f"<td>{_recommendation_html_escape(item.get('date'))}</td>"
            f"<td>{_recommendation_html_escape(item.get('id'))}</td>"
            f"<td>{_recommendation_html_escape(item.get('outcome'))}</td>"
            f"<td>{_recommendation_html_escape(display_value)}</td>"
            f"<td>{_recommendation_html_escape(item.get('context'))}</td>"
            "</tr>"
        )
    rows_html = "".join(observation_rows)
    excluded_rows = "".join(
        "<tr>"
        f"<td>{_recommendation_html_escape(item.get('symbol'))}</td>"
        f"<td>{_recommendation_html_escape(item.get('date'))}</td>"
        f"<td>{_recommendation_html_escape(item.get('id'))}</td>"
        f"<td>{_recommendation_html_escape(item.get('reason'))}</td>"
        "</tr>"
        for item in excluded
    ) or '<tr><td colspan="4">None</td></tr>'
    distribution = payload.get("target_r_distribution") if metric == "target" else None
    distribution_html = ""
    if isinstance(distribution, Mapping) and distribution:
        largest = max(int(value) for value in distribution.values()) or 1
        bars = "".join(
            f'<div class="bar-row"><span>{_recommendation_html_escape(label)}</span><div class="bar"><i style="width:{int(count)/largest*100:.2f}%"></i></div><b>{int(count)}</b></div>'
            for label, count in distribution.items()
        )
        distribution_html = f"<h2>Algorithm distribution</h2><div class=\"bars\">{bars}</div>"
    exclusion_json = _recommendation_html_escape(json.dumps(exclusion_summary, indent=2, sort_keys=True))
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{_recommendation_html_escape(scope_label)} — {_recommendation_html_escape(metric_label)}</title>
<style>
body{{margin:0;background:#f1f5f9;color:#0f172a;font:15px/1.45 Segoe UI,Arial,sans-serif}}main{{max-width:1180px;margin:0 auto;padding:28px}}h1{{margin:.2rem 0}}h2{{margin-top:2rem}}.sub{{color:#475569}}.cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin:18px 0}}.card,section{{background:white;border:1px solid #cbd5e1;border-radius:10px;padding:16px;box-shadow:0 1px 2px #0001}}.card b{{display:block;font-size:1.25rem}}.recommendation{{border-left:5px solid #7c3aed}}.chart{{display:block;width:100%;max-width:100%;height:auto;overflow:visible;background:white;border-radius:8px}}.axis-label{{font-size:11px;fill:#475569}}.axis-title{{font-size:13px;fill:#334155}}.recommendation-label{{font-size:13px;font-weight:700;fill:#6d28d9}}table{{width:100%;border-collapse:collapse;background:white}}th,td{{padding:8px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top}}th{{background:#e2e8f0;position:sticky;top:0}}.table-wrap{{overflow:auto;max-height:520px;border:1px solid #cbd5e1;border-radius:8px}}pre{{white-space:pre-wrap}}.bar-row{{display:grid;grid-template-columns:140px 1fr 48px;gap:8px;align-items:center;margin:7px 0}}.bar{{height:14px;background:#e2e8f0;border-radius:99px;overflow:hidden}}.bar i{{display:block;height:100%;background:#2563eb}}.offline{{font-size:.85rem;color:#475569}}@media(max-width:600px){{main{{padding:12px}}.card,section{{padding:10px}}}}
</style></head><body><main>
<p class="offline">Self-contained offline artifact. No Trading Tools server or network connection is required.</p>
<h1>{_recommendation_html_escape(scope_label)} — {_recommendation_html_escape(metric_label)}</h1>
<p class="sub">{_recommendation_html_escape(method)}</p>
<section class="recommendation" data-recommended-value="{recommended:.12g}"><strong>Workbook recommendation</strong><p>{_recommendation_html_escape(recommendation)}</p></section>
<div class="cards"><div class="card">Eligible algorithm sample<b>{eligible_count}</b></div><div class="card">Plotted numeric observations<b>{len(numeric_observations)}</b></div><div class="card">Eligible loss-gate observations<b>{len(gate_observations)}</b></div><div class="card">Excluded observations<b>{len(excluded)}</b></div></div>
<section>{_recommendation_svg(observations, recommended, unit)}</section>
{distribution_html}
<h2>Eligible observations</h2><div class="table-wrap"><table><thead><tr><th>Symbol</th><th>Date</th><th>Trade ID</th><th>Role</th><th>Value</th><th>Context</th></tr></thead><tbody>{rows_html}</tbody></table></div>
<h2>Excluded observations and filtering</h2><p>Global filters exclude non-trades and test trades before scope analysis. Algorithm-specific exclusion counts:</p><pre>{exclusion_json}</pre><div class="table-wrap"><table><thead><tr><th>Symbol</th><th>Date</th><th>Trade ID</th><th>Reason</th></tr></thead><tbody>{excluded_rows}</tbody></table></div>
</main></body></html>"""


def _recommendation_asset_relative_path(workbook_path: Path, filename: str) -> str:
    relative = posixpath.join(
        f"{workbook_path.stem}.assets",
        "recommendations",
        filename,
    )
    return quote(unquote(relative).replace("\\", "/"), safe="/")


def _prepare_recommendation_chart_bundle(
    rows: Sequence[Dict[str, Any]],
    workbook_path: Path,
) -> Dict[str, Any]:
    active = [
        row for row in rows
        if isinstance(row, dict)
        and str(row.get("row_type") or "trade").strip().lower() == "trade"
        and not _target_r_row_is_test_trade(row)
    ]
    scopes: Dict[str, Tuple[str, List[Dict[str, Any]], str]] = {
        "overall": ("Overall", active, "overall"),
        "fx": ("FX", [row for row in active if _trade_row_market(row) == "fx"], "standard"),
        "crypto": ("Crypto", [row for row in active if _trade_row_market(row) == "crypto"], "standard"),
    }
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in active:
        symbol = str(row.get("symbol") or "").strip().upper()
        if symbol:
            grouped[symbol].append(row)
    for symbol, symbol_rows in sorted(grouped.items()):
        scopes[f"symbol:{symbol}"] = (f"Symbol {symbol}", symbol_rows, "standard")

    files: Dict[str, str] = {}
    links: Dict[Tuple[str, str], str] = {}
    index_items: List[str] = []
    for scope_key, (scope_label, scope_rows, algorithm_scope) in scopes.items():
        payload = _distance_recommendation_summary(scope_rows, scope=algorithm_scope)
        safe_scope = re.sub(r"[^a-z0-9]+", "-", scope_key.casefold()).strip("-") or "overall"
        for metric, header, marker_key in (
            ("stop", STOP_RECOMMENDATION_HEADER, "stop_loss_recommended_pct"),
            ("target", TARGET_RECOMMENDATION_HEADER, "target_r_recommended"),
        ):
            recommendation = _sanitize_recommendation_text(header, payload.get(header))
            if not recommendation or _as_float(payload.get(marker_key)) is None:
                continue
            observations, excluded = _recommendation_chart_observations(scope_rows, metric)
            filename = f"{safe_scope}-{metric}.html"
            relative = _recommendation_asset_relative_path(workbook_path, filename)
            links[(scope_key, metric)] = relative
            files[filename] = _recommendation_chart_html(
                scope_label=scope_label,
                metric=metric,
                payload=payload,
                observations=observations,
                excluded=excluded,
            )
            index_items.append(f'<li><a href="{_recommendation_html_escape(filename)}">{_recommendation_html_escape(scope_label)} — {metric.title()}</a></li>')
    files["index.html"] = (
        "<!doctype html><html lang=\"en\"><head><meta charset=\"utf-8\"><title>Trading Journal recommendation charts</title></head>"
        "<body><h1>Trading Journal recommendation charts</h1><p>Self-contained offline artifacts.</p><ul>"
        + "".join(index_items)
        + "</ul></body></html>"
    )
    return {"files": files, "links": links, "workbook_path": Path(workbook_path)}


def _apply_recommendation_chart_hyperlinks(wb, bundle: Mapping[str, Any]) -> None:
    links = bundle.get("links") if isinstance(bundle, Mapping) else {}
    if not isinstance(links, Mapping):
        return

    def apply_link(cell, target: str | None) -> None:
        current = getattr(cell, "hyperlink", None)
        current_target = str(getattr(current, "target", "") or "")
        canonical_target = (
            quote(
                unquote(str(target)).replace("\\", "/"),
                safe="/",
            )
            if target
            else ""
        )
        if current_target != canonical_target:
            cell.hyperlink = canonical_target or None
        if canonical_target:
            font = copy(cell.font)
            font.color = "0563C1"
            font.underline = "single"
            if cell.font != font:
                cell.font = font

    if STATS1_SHEET in wb.sheetnames:
        ws = wb[STATS1_SHEET]
        market_cols = _stats1_market_columns(ws)
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row, 1).value or "").strip().casefold() != "recommendation":
                continue
            previous = str(ws.cell(row - 1, 1).value or "").strip().casefold()
            metric = "stop" if "stop" in previous else "target" if "target" in previous else ""
            if not metric:
                continue
            for market, col in market_cols.items():
                target = links.get((market, metric)) if ws.cell(row, col).value not in (None, "") else None
                apply_link(ws.cell(row, col), target)

    if SYMBOLS_SHEET in wb.sheetnames:
        ws = wb[SYMBOLS_SHEET]
        headers = _instrument_averages_header_map(ws)
        symbol_col = headers.get("Symbol", 1)
        for row in range(_instrument_averages_data_start_row(ws), ws.max_row + 1):
            symbol = str(ws.cell(row, symbol_col).value or "").strip().upper()
            if not symbol:
                continue
            for metric, header in (("stop", STOP_RECOMMENDATION_HEADER), ("target", TARGET_RECOMMENDATION_HEADER)):
                col = headers.get(header)
                if col:
                    target = links.get((f"symbol:{symbol}", metric)) if ws.cell(row, col).value not in (None, "") else None
                    apply_link(ws.cell(row, col), target)

    if TRADE_LOG_SHEET in wb.sheetnames:
        ws = wb[TRADE_LOG_SHEET]
        headers = _trade_log_header_map(ws)
        symbol_col = headers.get("Symbol")
        if symbol_col:
            for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
                symbol = str(ws.cell(row, symbol_col).value or "").strip().upper()
                for metric, header in (("stop", STOP_RECOMMENDATION_HEADER), ("target", TARGET_RECOMMENDATION_HEADER)):
                    col = headers.get(header)
                    if col:
                        target = links.get((f"symbol:{symbol}", metric)) if ws.cell(row, col).value not in (None, "") else None
                        apply_link(ws.cell(row, col), target)


def _publish_recommendation_chart_bundle(bundle: Mapping[str, Any]) -> Path:
    workbook_path = Path(bundle["workbook_path"])
    files = dict(bundle.get("files") or {})
    asset_dir = workbook_path.parent / f"{workbook_path.stem}.assets" / "recommendations"
    asset_dir.mkdir(parents=True, exist_ok=True)
    expected = set(files)
    for filename, contents in files.items():
        target = asset_dir / filename
        temporary = asset_dir / f".{filename}.{uuid4().hex}.tmp"
        temporary.write_text(str(contents), encoding="utf-8", newline="\n")
        os.replace(temporary, target)
    for stale in asset_dir.glob("*.html"):
        if stale.name not in expected:
            stale.unlink()
    return asset_dir


def _row_distance_recommendation(
    row: Dict[str, Any],
    recommendations_by_symbol: Dict[str, Dict[str, Any]],
    header: str,
) -> str:
    symbol = str(row.get("symbol") or "").strip().upper()
    symbol_text = (recommendations_by_symbol.get(symbol) or {}).get(header)
    return str(_sanitize_recommendation_text(header, symbol_text) or "")


def _sanitize_recommendation_text(header: str, value: Any) -> Any:
    if value in (None, ""):
        return value
    text = str(value).strip()
    for reason_key, visible_reason in RECOMMENDATION_REASON_TEXT.items():
        text = text.replace(reason_key, visible_reason)
    text = re.sub(
        r"\b[a-z][a-z0-9]*(?:_[a-z0-9]+){2,}\b",
        lambda match: match.group(0).replace("_", " "),
        text,
    )
    lowered = text.casefold()
    normalized = re.sub(r"[_\s-]+", " ", lowered).strip()
    insufficient_fragments = (
        "need wins & losses",
        "need more target data",
        "no eligible winning trades",
        "no eligible losing trades",
        "insufficient eligible wins",
        "insufficient eligible losses",
        "recommendation unavailable",
    )
    insufficient_qualifier = re.search(
        r"\b(?:need|requires?|missing|not enough|too few|insufficient|no|unavailable)\b",
        normalized,
    )
    insufficient_subject = re.search(
        r"\b(?:wins?|winning|loss(?:es|ing)?|samples?|data|target|stop|recommendation)\b",
        normalized,
    )
    if (
        any(fragment in normalized for fragment in insufficient_fragments)
        or (insufficient_qualifier and insufficient_subject)
    ):
        return ""
    if "keep" in lowered or "maintain" in lowered:
        return ""
    if header == TARGET_RECOMMENDATION_HEADER:
        if "recommended:" not in lowered:
            return ""
        if lowered.startswith("reduce target"):
            return re.sub(r"^reduce\s+target", "Decrease target", text, count=1, flags=re.IGNORECASE)
        return text
    if header == STOP_RECOMMENDATION_HEADER:
        if "recommended:" not in lowered:
            return ""
        text = re.sub(r"^reduce\s+stop(?:\s+loss)?", "Decrease stop", text, count=1, flags=re.IGNORECASE)
        text = re.sub(r"^increase\s+stop\s+loss", "Increase stop", text, count=1, flags=re.IGNORECASE)
        return text
    return value


def _apply_recommendation_cell_style(cell) -> None:
    alignment = copy(cell.alignment)
    alignment.wrap_text = True
    alignment.horizontal = "left"
    cell.alignment = alignment
    cell.number_format = "General"
    if _semantic_fill_rgb(cell) == "FFF2CC":
        cell.fill = PatternFill()
        font = copy(cell.font)
        if getattr(font.color, "type", None) == "rgb" and str(font.color.rgb or "")[-6:].upper() == "9C6500":
            font.color = "000000"
        cell.font = font


def _linear_profit_percentage_totals(rows: List[Dict[str, Any]]) -> Dict[str, float | None]:
    values = [
        value
        for value in (_as_float(row.get("result_pct")) for row in rows)
        if value is not None and math.isfinite(value)
    ]
    r_values = [
        value
        for value in (_as_float(row.get("r_multiple")) for row in rows)
        if value is not None and math.isfinite(value)
    ]
    return {
        "net_result_pct": sum(values) if values else None,
        "gross_gain_result_pct": sum(value for value in values if value > 0) if values else None,
        "gross_loss_result_pct": abs(sum(value for value in values if value < 0)) if values else None,
        "gross_ir_gain": sum(value for value in r_values if value > 0) if r_values else None,
        "gross_ir_loss": abs(sum(value for value in r_values if value < 0)) if r_values else None,
        "avg_result_pct": (sum(values) / len(values)) if values else None,
        "min_result_pct": min(values) if values else None,
        "max_result_pct": max(values) if values else None,
    }


def _journal_usdt_to_aud_fallback() -> float:
    value = _as_float(os.getenv("JOURNAL_USDT_TO_AUD_FALLBACK", "1.45"))
    return value if value is not None and value > 0 else 1.45


def _currency_to_aud_factor(currency: Any) -> float | None:
    code = str(currency or "").strip().upper()
    if not code or code == "AUD":
        return 1.0
    if code in {"USDT", "USDC", "USD"}:
        return _journal_usdt_to_aud_fallback()
    return None


def _net_result_pct_by_account(rows: List[Dict[str, Any]]) -> Dict[str, float]:
    by_account: Dict[str, float] = defaultdict(float)
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        if _target_r_row_is_test_trade(row):
            continue
        account = _canonical_account_label(row.get("account_label") or row.get("account") or row.get("source") or "")
        if not account:
            continue
        result_pct = _as_float(row.get("result_pct"))
        if result_pct is None or not math.isfinite(result_pct):
            continue
        by_account[account] += result_pct
    return dict(by_account)


def _merge_metric_buckets(*buckets: Dict[str, Any] | None) -> Dict[str, Any]:
    merged: Dict[str, Any] = {}
    metric_sources: Dict[str, Any] = {}
    target_payload_keys = {
        TARGET_RECOMMENDATION_HEADER,
        "target_recommendation",
        "target_r_scope",
        "eligible_target_r_wins",
        "eligible_target_r_losses",
        "target_r_total_losing_trades",
        "target_r_eligible_winning_trades",
        "target_r_eligible_losing_trades",
        "target_r_eligible_fx_wins",
        "target_r_eligible_crypto_wins",
        "target_r_eligible_fx_losses",
        "target_r_eligible_crypto_losses",
        "target_r_increment",
        "target_r_distribution",
        "target_r_peak_bucket",
        "target_r_peak_bucket_low",
        "target_r_peak_bucket_high",
        "target_r_peak_instances",
        "target_r_recommended",
        "target_r_histogram_recommended_before_tie",
        "target_r_recommendation_direction",
        "target_r_exact_tie",
        "target_r_exact_tie_goal_preference",
        "target_r_tie_break_reason",
        "current_median_original_planned_target_r",
        "current_avg_original_planned_target_r",
        "target_r_excluded_count",
        "target_r_excluded_reasons",
        "target_r_excluded_winning_trades",
        "target_r_excluded_losing_trades",
        "target_r_winning_exclusion_reasons",
        "target_r_losing_exclusion_reasons",
        "target_r_coverage_note",
    }
    for bucket in buckets:
        if not isinstance(bucket, dict):
            continue
        incoming_target_text = str(
            bucket.get(TARGET_RECOMMENDATION_HEADER) or bucket.get("target_recommendation") or ""
        ).strip()
        existing_target_text = str(
            merged.get(TARGET_RECOMMENDATION_HEADER) or merged.get("target_recommendation") or ""
        ).strip()
        insufficient_target_texts = {"Need more target data", TARGET_RECOMMENDATION_NO_WINS, TARGET_RECOMMENDATION_NO_LOSSES}
        keep_existing_target_payload = (
            incoming_target_text in insufficient_target_texts
            and existing_target_text
            and existing_target_text not in insufficient_target_texts
        )
        for key, value in bucket.items():
            if key == "metric_sources":
                if isinstance(value, dict):
                    metric_sources.update(value)
            elif keep_existing_target_payload and key in target_payload_keys:
                continue
            else:
                merged[key] = value
    if metric_sources:
        merged["metric_sources"] = metric_sources
    return merged


def adaptive_percent_number_format(value: Any, *, max_decimals: int = 12) -> str:
    number = _as_float(value)
    if number is None or number == 0:
        return "0.00%"
    for decimals in range(2, max_decimals + 1):
        if round(abs(number) * 100.0, decimals) != 0:
            return "0." + ("0" * decimals) + "%"
    return "0." + ("0" * max_decimals) + "%"


def adaptive_number_format(value: Any, *, max_decimals: int = 12) -> str:
    number = _as_float(value)
    if number is None or number == 0:
        return "0.00"
    for decimals in range(2, max_decimals + 1):
        if round(abs(number), decimals) != 0:
            return "0." + ("0" * decimals)
    return "0." + ("0" * max_decimals)


def _trade_number_aliases(trade_number: Any) -> List[str]:
    text = str(trade_number or "").strip().upper()
    match = re.fullmatch(r"([FC])0*(\d+)", text)
    if not match:
        return [text] if text else []
    prefix, digits = match.groups()
    canonical = f"{prefix}{int(digits)}"
    raw = f"{prefix}{digits}"
    zero_padded = f"{prefix}{int(digits):03d}"
    return list(dict.fromkeys([text, canonical, raw, zero_padded]))


def _trade_folder_index(root: Path, prefix: str, *, include_files: bool = False) -> Dict[str, List[Path]]:
    cache_key = (str(root).lower(), prefix.upper(), "files" if include_files else "dirs")
    cached = _TRADE_FOLDER_INDEX_CACHE.get(cache_key)
    if cached is not None:
        return cached
    index: Dict[str, List[Path]] = defaultdict(list)
    if root.is_dir():
        matcher = re.compile(
            rf"(?<![A-Z0-9])({re.escape(prefix.upper())}0*\d+)(?!\d)",
            re.IGNORECASE,
        )
        for current_root, dir_names, file_names in os.walk(root):
            names = list(dir_names)
            if include_files:
                names.extend(file_names)
            for name in names:
                path = Path(current_root) / name
                match = matcher.search(name.strip())
                if match:
                    for alias in _trade_number_aliases(match.group(1)):
                        index[alias].append(path)
                nested_match = re.fullmatch(r"0*(\d+)", name.strip())
                if nested_match and matcher.search(Path(current_root).name.strip()):
                    nested_number = f"{prefix.upper()}{int(nested_match.group(1))}"
                    for alias in _trade_number_aliases(nested_number):
                        index[alias].append(path)
    result = dict(index)
    _TRADE_FOLDER_INDEX_CACHE[cache_key] = result
    return result


def _trade_date_hints(*values: Any) -> Tuple[set[str], set[str]]:
    years: set[str] = set()
    months: set[str] = set()
    for value in values:
        dt = _as_datetime(value)
        if dt is None:
            continue
        years.add(str(dt.year))
        months.update({
            calendar.month_name[dt.month].upper(),
            calendar.month_abbr[dt.month].upper(),
        })
    return years, months


def _trade_folder_roots(prefix: str, explicit_root: Path | None = None) -> List[Path]:
    market = "FOREX" if prefix == "F" else "CRYPTO"
    env_name = f"TRADING_JOURNAL_{market}_ROOT"
    repo_root = REPO_ROOT
    candidates: List[Path] = []
    if explicit_root is not None:
        candidates.append(Path(explicit_root))
    env_root = str(os.getenv(env_name) or "").strip()
    if env_root:
        candidates.append(Path(env_root))
    candidates.extend([
        repo_root / "journal" / ("Forex" if prefix == "F" else "Crypto"),
        Path(r"C:\Users\User\Documents\TRADING") / market,
        Path.home() / "Documents" / "TRADING" / market,
        Path.home() / "Dropbox" / market,
        Path.home() / "Dropbox" / "TRADING" / market,
    ])
    unique: List[Path] = []
    seen: set[str] = set()
    for candidate in candidates:
        key = str(candidate.expanduser()).rstrip("\\/").casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        unique.append(candidate.expanduser())
    return unique


def _trade_file_fallback_root_keys(prefix: str, explicit_root: Path | None = None) -> set[str]:
    market = "FOREX" if prefix == "F" else "CRYPTO"
    env_name = f"TRADING_JOURNAL_{market}_ROOT"
    repo_root = REPO_ROOT
    candidates: List[Path] = [repo_root / "journal" / ("Forex" if prefix == "F" else "Crypto")]
    if explicit_root is not None:
        candidates.append(Path(explicit_root))
    env_root = str(os.getenv(env_name) or "").strip()
    if env_root:
        candidates.append(Path(env_root))
    return {str(path.expanduser()).rstrip("\\/").casefold() for path in candidates}


def _trade_match_specificity(path: Path, prefix: str, trade_number: str) -> int:
    matcher = re.compile(
        rf"(?<![A-Z0-9])({re.escape(prefix.upper())}0*\d+)(?!\d)",
        re.IGNORECASE,
    )
    wanted = set(_trade_number_aliases(trade_number))
    direct = matcher.search(path.name.strip())
    if direct and any(alias in wanted for alias in _trade_number_aliases(direct.group(1))):
        return 2
    nested = re.fullmatch(r"0*(\d+)", path.name.strip())
    if nested and matcher.search(path.parent.name.strip()):
        nested_number = f"{prefix.upper()}{int(nested.group(1))}"
        if any(alias in wanted for alias in _trade_number_aliases(nested_number)):
            return 1
    return 0


def resolve_trade_folder_link(
    trade_number: Any,
    *,
    open_time: Any = None,
    close_time: Any = None,
    forex_root: Path | None = None,
    crypto_root: Path | None = None,
    diagnostics: Dict[str, Any] | None = None,
) -> Tuple[str | None, str | None]:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    number = str(trade_number or "").strip().upper()
    match = re.fullmatch(r"([FC])0*\d+", number)
    if not match:
        diagnostics["checked_roots"] = []
        return None, "invalid_trade_number"
    prefix = match.group(1)
    roots = _trade_folder_roots(prefix, forex_root if prefix == "F" else crypto_root)
    diagnostics["checked_roots"] = [str(root) for root in roots]
    candidates: List[Path] = []
    matched_root: Path | None = None
    file_fallback_root_keys = _trade_file_fallback_root_keys(
        prefix,
        forex_root if prefix == "F" else crypto_root,
    )

    def _matches_for_root(root: Path, *, include_files: bool = False) -> List[Path]:
        index = _trade_folder_index(root, prefix, include_files=include_files)
        matches = []
        seen_matches: set[str] = set()
        for alias in _trade_number_aliases(number):
            for path in index.get(alias, []):
                key = str(path).casefold()
                if key not in seen_matches:
                    seen_matches.add(key)
                    matches.append(path)
        return matches

    for root in roots:
        matches = _matches_for_root(root)
        if matches:
            candidates = matches
            matched_root = root
            break
        root_key = str(root.expanduser()).rstrip("\\/").casefold()
        if root_key in file_fallback_root_keys:
            matches = _matches_for_root(root, include_files=True)
            if matches:
                candidates = matches
                matched_root = root
                break
    if not candidates:
        return None, "missing_trade_folder"
    diagnostics["matched_root"] = str(matched_root) if matched_root else ""
    if len(candidates) > 1:
        years, months = _trade_date_hints(open_time, close_time)
        scored: List[Tuple[int, int, Path]] = []
        for candidate in candidates:
            parts = {part.upper() for part in candidate.parts}
            date_score = (4 if parts & years else 0) + (2 if parts & months else 0)
            specificity = _trade_match_specificity(candidate, prefix, number)
            scored.append((date_score, specificity, candidate))
        best_score = max((date_score, specificity) for date_score, specificity, _candidate in scored)
        best = [
            candidate for date_score, specificity, candidate in scored
            if (date_score, specificity) == best_score
        ]
        if best_score[0] <= 0 or len(best) != 1:
            return None, "ambiguous_trade_folder"
        candidates = best
    return candidates[0].resolve().as_uri(), None


def _normalize_pct_distance_cell(value: Any, number_format: Any = None) -> float | None:
    """Return internal percentage points from a workbook distance cell.

    Excel percentage-formatted cells store fractions (0.01 displays as 1%).
    Plain numeric values are already treated as internal percent points.
    """
    num = _as_float(value)
    if num is None:
        return None
    fmt = str(number_format or "")
    return num * 100.0 if "%" in fmt else num


def _cell_fill_rgb(cell) -> str:
    fg = getattr(getattr(cell, "fill", None), "fgColor", None)
    rgb = str(getattr(fg, "rgb", "") or "").upper()
    return rgb


def _is_light_grey_no_metric_cell(cell) -> bool:
    rgb = _cell_fill_rgb(cell)
    return rgb == LIGHT_GREY_FILL_RGB or rgb.endswith(LIGHT_GREY_FILL_RGB[-6:])


def _is_likely_fx_pair(value: str) -> bool:
    token = str(value or '').upper().replace('/','').replace('-','').replace('_','')
    
    if not (len(token) == 6 and token.isalpha()):
        return False
    known = {"USD","EUR","GBP","JPY","AUD","NZD","CAD","CHF"}
    return token[:3] in known and token[3:] in known
def _is_test_trade_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) == 1.0
    text = str(value or '').strip().lower()
    return text in {'yes','y','true','1'}


def _as_float(v: Any) -> float | None:
    try:
        if v in (None, ""):
            return None
        return float(v)
    except Exception:
        return None


def _as_date(v: Any) -> date | None:
    s = str(v or "").strip()
    if not s:
        return None
    s = s.replace("Z", "")
    for fmt in ["%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"]:
        try:
            return datetime.strptime(s[:19] if "T" in s or " " in s else s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def _duration_seconds_to_ddhhmmss_number(seconds: Any) -> int | None:
    v = _as_float(seconds)
    if v is None:
        return None
    s = max(0, int(v))
    days, rem = divmod(s, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, secs = divmod(rem, 60)
    return days * 1000000 + hours * 10000 + minutes * 100 + secs


def _format_duration_display(seconds: Any) -> str:
    v = _as_float(seconds)
    if v is None:
        return ""
    total_seconds = max(0, int(v))
    days, rem = divmod(total_seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, secs = divmod(rem, 60)
    parts: List[str] = []
    if days:
        parts.append(f"{days}D")
    if hours or parts:
        parts.append(f"{hours}H")
    if minutes or parts:
        parts.append(f"{minutes}M")
    parts.append(f"{secs}S")
    return " ".join(parts)


def _duration_ddhhmmss_cell_to_seconds(value: Any) -> int | None:
    if value in (None, ""):
        return None
    raw = value
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return None
    try:
        n = int(float(raw))
    except Exception:
        return None
    if n < 0:
        return None
    dd = n // 1_000_000
    hh = (n // 10_000) % 100
    mm = (n // 100) % 100
    ss = n % 100
    if hh >= 24 or mm >= 60 or ss >= 60:
        return None
    return dd * 86400 + hh * 3600 + mm * 60 + ss


def _is_ddhhmmss_number_format(number_format: Any) -> bool:
    text = str(number_format or "").upper()
    return r"\:" in text or all(token in text for token in ("DAYS", "HOURS", "MINUTES", "SECONDS"))


def _is_legacy_duration_number_format(number_format: Any) -> bool:
    text_upper = str(number_format or "").upper()
    return all(token in text_upper for token in LEGACY_DURATION_NUMBER_FORMAT_TOKENS)


def _duration_display_cell_value(value: Any, number_format: Any = None) -> str:
    if value in (None, ""):
        return ""
    source_suffix = ""
    duration_value = value
    if isinstance(value, str) and " - " in value:
        duration_value, source_suffix = value.split(" - ", 1)
    seconds = _duration_ddhhmmss_cell_to_seconds(duration_value) if _is_ddhhmmss_number_format(number_format) else None
    if seconds is None:
        seconds = _parse_duration_text(duration_value)
    rendered = _format_duration_display(seconds) if seconds is not None else str(duration_value)
    return f"{rendered} - {source_suffix}" if source_suffix else rendered


def _normalize_trade_log_display_durations(ws) -> None:
    headers = _trade_log_header_map(ws)
    duration_headers = (
        "Trade Duration (DD:HH:MM:SS)",
        "Move to Break Even Duration",
        "Move to Profit Duration",
    )
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        for header in duration_headers:
            col = headers.get(header)
            if not col:
                continue
            cell = ws.cell(row, col)
            if cell.value not in (None, ""):
                cell.value = _duration_display_cell_value(cell.value, cell.number_format)
            cell.number_format = "General"


def _fmt_duration(seconds: Any) -> str:
    n = _duration_seconds_to_ddhhmmss_number(seconds)
    if n is None:
        return "—"
    return f"{n:08d}"





def _fmt_duration_full(seconds: Any) -> int | None:
    return _duration_seconds_to_ddhhmmss_number(seconds)

def _as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    elif isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        raw = raw.replace("Z", "")
        dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(raw, fmt)
                break
            except Exception:
                continue
        if dt is None:
            try:
                dt = datetime.fromisoformat(raw)
            except Exception:
                return None
    else:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(JOURNAL_DISPLAY_TZ).replace(tzinfo=None)
    return dt

def _round_trade_duration_seconds(delta_seconds: Any) -> int | None:
    val = _as_float(delta_seconds)
    if val is None or not math.isfinite(val) or val < 0:
        return None
    if val < 1:
        return 1
    return max(1, int(val + 0.5))

def _infer_trade_duration_seconds(row: Dict[str, Any]) -> int | None:
    if str(row.get("row_type") or "trade").strip().lower() != "trade":
        return None
    for key in ("trade_duration_seconds", "duration_seconds"):
        val = _as_float(row.get(key))
        if val is not None and val >= 0:
            return _round_trade_duration_seconds(val)
    ot = _as_datetime(row.get("open_time"))
    ct = _as_datetime(row.get("close_time"))
    if not ot or not ct:
        return None
    delta = (ct - ot).total_seconds()
    return _round_trade_duration_seconds(delta)

def _resolve_balance_after(row: Dict[str, Any]) -> float | None:
    for key in ("analysis_balance_after_trade", "balance_after_trade", "cashflow_new_balance"):
        val = _as_float(row.get(key))
        if val is not None:
            return val
    return None

def _resolved_all_trade_balances(rows: List[Dict[str, Any]]) -> Dict[str, float]:
    indexed = list(enumerate(rows))
    running: Dict[str, float] = {}
    out: Dict[str, float] = {}
    def _sort_key(item):
        i, row = item
        acct = str(row.get("account_label") or row.get("account") or "")
        ts = str(row.get("close_time") or row.get("open_time") or "")
        return (acct, ts, i)
    for i, row in sorted(indexed, key=_sort_key):
        acct = str(row.get("account_label") or row.get("account") or "")
        resolved = _resolve_balance_after(row)
        if resolved is not None:
            running[acct] = resolved
            out[str(i)] = resolved
            continue
        if acct in running:
            pnl = _as_float(row.get("net_profit"))
            if pnl is not None:
                running[acct] = running[acct] + pnl
                out[str(i)] = running[acct]
    return out

ZERO_HIDE_FORMAT = "0;-0;;@"

_MONTHLY_AUD_REVAL_ROW_ID_RE = re.compile(r"^monthly_aud_reval:bybit_live:(\d{4}-\d{2})$")

def _monthly_aud_reval_row_id_month(row_id: Any) -> str:
    m = _MONTHLY_AUD_REVAL_ROW_ID_RE.match(str(row_id or "").strip())
    return m.group(1) if m else ""

def _is_monthly_aud_reval_semantic_row(row: Dict[str, Any]) -> bool:
    row_type = str(row.get("row_type") or "").strip().lower()
    symbol = str(row.get("symbol") or "").strip().upper()
    account = _canonical_account_label(row.get("account_label") or row.get("account"))
    return row_type == "monthly_aud_reval" and symbol == "MONTHLY AUD P/L" and account == "BYBIT"

def _monthly_aud_reval_record_rank(row: Dict[str, Any]) -> Tuple[int, int, int, int, int, int, str]:
    """Rank canonical monthly rows so duplicate repair keeps the best valid record."""
    row_id_month = _monthly_aud_reval_row_id_month(row.get("id"))
    result_cash = _as_float(
        row.get("result_cash")
        if row.get("result_cash") not in (None, "")
        else row.get("net_profit")
    )
    currency = str(row.get("result_currency") or row.get("currency") or "").strip().upper()
    refs = row.get("raw_refs") if isinstance(row.get("raw_refs"), dict) else {}
    record_month = str(refs.get("period_month") or "").strip()[:7]
    if not record_month:
        record_month = str(row.get("close_time") or row.get("open_time") or "").strip()[:7]
    semantic = _is_monthly_aud_reval_semantic_row(row)
    result_valid = result_cash is not None and math.isfinite(result_cash)
    currency_valid = currency == "AUD"
    month_valid = bool(row_id_month) and (not record_month or record_month == row_id_month)
    valid = semantic and result_valid and currency_valid and month_valid
    completeness_fields = (
        "account", "account_label", "symbol", "open_time", "close_time",
        "result_cash", "result_currency", "currency", "source", "notes", "raw_refs",
    )
    completeness = sum(row.get(field) not in (None, "", [], {}) for field in completeness_fields)
    freshness = str(row.get("updated_at") or row.get("close_time") or "")
    return (
        int(valid),
        int(result_valid),
        int(currency_valid),
        int(month_valid),
        int(semantic),
        completeness,
        freshness,
    )

def _dedupe_monthly_aud_reval_rows(
    rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Deduplicate only canonical monthly P&L IDs, leaving ordinary trades alone."""
    canonical_indexes: Dict[str, int] = {}
    result: List[Dict[str, Any]] = []
    removed_by_id: Counter[str] = Counter()
    for row in rows:
        row_id = str(row.get("id") or "").strip()
        if not _monthly_aud_reval_row_id_month(row_id):
            result.append(row)
            continue
        existing_index = canonical_indexes.get(row_id)
        if existing_index is None:
            canonical_indexes[row_id] = len(result)
            result.append(row)
            continue
        removed_by_id[row_id] += 1
        if _monthly_aud_reval_record_rank(row) > _monthly_aud_reval_record_rank(result[existing_index]):
            result[existing_index] = row
    duplicate_ids = sorted(removed_by_id)
    return result, {
        "duplicate_monthly_aud_reval_groups_removed": len(duplicate_ids),
        "duplicate_monthly_aud_reval_rows_removed": sum(removed_by_id.values()),
        "deduplicated_monthly_aud_reval_row_ids": duplicate_ids,
    }

def _canonical_account_label(label: Any) -> str:
    raw = str(label or "").strip()
    low = raw.lower().replace("_", " ").replace("-", " ")
    parts = {p for p in low.split() if p}
    if "bybit" in parts and "demo" in parts:
        return "BYBIT DEMO"
    if "bybit" in parts and ("live" in parts or len(parts) == 1):
        return "BYBIT"
    if "pepperstone" in parts and "demo" in parts:
        return "PEPPERSTONE DEMO"
    if "pepperstone" in parts and "live" in parts:
        return "PEPPERSTONE LIVE"
    return raw


def _stats2_net_pl_percentage_is_applicable(account_label: Any) -> bool:
    return _canonical_account_label(account_label) != "BYBIT DEMO"

def _repair_or_flag_zero_trade_qty(row: Dict[str, Any]) -> Dict[str, Any]:
    if str(row.get("row_type") or "trade").lower() != "trade":
        return row
    qty=_as_float(row.get("qty"))
    if qty is None or qty!=0:
        return row
    refs=row.get("raw_refs") if isinstance(row.get("raw_refs"),dict) else {}
    for k in ("qty_raw","closedSize","closed_size","execQty","exec_qty","cumExecQty","qty","size","Filled Qty","Size Quantity"):
        cand=_as_float(row.get(k) if k in row else refs.get(k))
        if cand is not None and cand>0:
            row["qty"]=cand; row.setdefault("diagnostics",[]).append("qty_repaired_from_source")
            return row
    sym=str(row.get("symbol") or "").upper()
    acct=str(row.get("account") or row.get("account_label") or "").upper()
    if any(x in acct for x in ("OANDA","PEPPERSTONE")) or ("/" in sym and len(sym)==6):
        row.setdefault("diagnostics",[]).append("zero_qty_unrepaired_fx")
        return row
    ep=_as_float(row.get("entry_price")); xp=_as_float(row.get("exit_price")); np=_as_float(row.get("net_profit")); fee=_as_float(row.get("commission") if row.get("commission") is not None else row.get("fees")) or 0.0
    side=str(row.get("side") or "").upper()
    if (sym.endswith("USDT") or sym.endswith("USDC")) and None not in (ep,xp,np) and side in {"BUY","SELL"}:
        den=(xp-ep) if side=="BUY" else (ep-xp)
        if den and den!=0:
            q=(np+abs(fee))/den
            if q>0 and math.isfinite(q):
                chk=(q*den)-abs(fee)
                if abs(chk-np)<=max(1e-6,abs(np)*1e-5):
                    row["qty"]=q; row.setdefault("diagnostics",[]).append("qty_inferred_from_pnl")
                    return row
    row.setdefault("diagnostics",[]).append("zero_qty_unrepaired")
    return row


def _collect_zero_qty_validation(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    out = {"crypto_zero_qty_unrepaired": [], "fx_zero_qty_unrepaired": []}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        if str(row.get("row_type") or "trade").lower() != "trade":
            continue
        qty = _as_float(row.get("qty"))
        if qty != 0:
            continue
        diag = row.get("diagnostics") if isinstance(row.get("diagnostics"), list) else []
        entry = {
            "id": row.get("id"),
            "account": row.get("account_label") or row.get("account"),
            "symbol": row.get("symbol"),
            "open_time": row.get("open_time"),
            "close_time": row.get("close_time"),
            "source": row.get("source"),
            "diagnostics": list(diag),
        }
        acct = str(row.get("account") or row.get("account_label") or "").upper()
        sym = str(row.get("symbol") or "").upper()
        is_fx = any(x in acct for x in ("OANDA", "PEPPERSTONE")) or ("/" in sym and len(sym) == 7)
        if is_fx:
            out["fx_zero_qty_unrepaired"].append(entry)
        else:
            out["crypto_zero_qty_unrepaired"].append(entry)
    return out


def _canonicalize_and_dedupe_balances(balances: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def _source_rank(value: Any) -> int:
        src = str(value or "").strip().lower()
        if src == "cashflow_anchor_plus_trades":
            return 300
        if (
            src == "oanda_transaction_export_balance"
            or "broker" in src
            or "account_summary" in src
            or "wallet_balance_anchor" in src
        ):
            return 200
        if src in {"authoritative_trade_balance", "trade_timeline", "master_journal"}:
            return 100
        if src == "timeline_missing":
            return 0
        return 50
    def _asof_rank(value: Any) -> float:
        dt = _as_datetime(value)
        return dt.timestamp() if dt else float("-inf")
    def _pick(prev: Dict[str, Any], now: Dict[str, Any]) -> Dict[str, Any]:
        prev_bal = _as_float(prev.get("balance"))
        now_bal = _as_float(now.get("balance"))
        if prev_bal is None and now_bal is not None:
            return now
        if now_bal is None and prev_bal is not None:
            return prev
        prev_score = (_source_rank(prev.get("balance_source") or prev.get("source")), _asof_rank(prev.get("as_of") or prev.get("updated_at")))
        now_score = (_source_rank(now.get("balance_source") or now.get("source")), _asof_rank(now.get("as_of") or now.get("updated_at")))
        return now if now_score >= prev_score else prev

    merged: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for rec in balances or []:
        if not isinstance(rec, dict):
            continue
        label = _canonical_account_label(rec.get("account_label") or rec.get("account") or rec.get("label"))
        if not label:
            continue
        key = label.upper()
        payload = dict(rec)
        payload["account_label"] = label
        payload["account"] = label
        if key not in merged:
            order.append(key)
            merged[key] = payload
            continue
        merged[key] = _pick(merged[key], payload)
    return [merged[k] for k in order]


_ACCOUNT_BALANCE_SOURCE_DEFINED_NAME_PREFIX = "_CODEX_ACCOUNT_BALANCE_SOURCE_"


def _account_balance_source_defined_name(account_label: Any) -> str:
    canonical = _canonical_account_label(account_label).upper()
    digest = hashlib.sha256(canonical.encode("utf-8")).hexdigest()[:20].upper()
    return f"{_ACCOUNT_BALANCE_SOURCE_DEFINED_NAME_PREFIX}{digest}"


def _read_account_balance_source_metadata(
    wb,
    account_label: Any,
) -> Dict[str, str]:
    defined = wb.defined_names.get(_account_balance_source_defined_name(account_label))
    raw = str(getattr(defined, "attr_text", "") or "").strip()
    if len(raw) >= 2 and raw.startswith('"') and raw.endswith('"'):
        raw = raw[1:-1].replace('""', '"')
    raw = raw.strip()
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = None
    if isinstance(parsed, dict):
        return {
            "source": str(parsed.get("source") or "").strip(),
            "timeline_as_of": str(parsed.get("timeline_as_of") or "").strip(),
        }
    return {"source": raw, "timeline_as_of": ""}


def _account_balance_timeline_as_of(value: Any) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None).isoformat()
    text = str(value).strip()
    if not text:
        return ""
    normalized = text[:-1] + "+00:00" if text.endswith("Z") else text
    try:
        parsed = datetime.fromisoformat(normalized)
    except Exception:
        return text
    return parsed.replace(tzinfo=None).isoformat()


def _write_account_balance_source_metadata(
    wb,
    account_label: Any,
    source: Any,
    as_of: Any = None,
) -> None:
    name = _account_balance_source_defined_name(account_label)
    wb.defined_names.pop(name, None)
    clean_source = str(source or "").strip()
    if not clean_source:
        return
    metadata = json.dumps(
        {
            "source": clean_source,
            "timeline_as_of": _account_balance_timeline_as_of(as_of),
        },
        separators=(",", ":"),
        ensure_ascii=True,
    )
    excel_string = f'"{metadata.replace(chr(34), chr(34) * 2)}"'
    wb.defined_names.add(
        DefinedName(
            name,
            attr_text=excel_string,
            hidden=True,
            description="Codex Account Balances provenance",
        )
    )


def _clear_account_balance_source_metadata(wb) -> None:
    for name in list(wb.defined_names.keys()):
        if str(name).startswith(_ACCOUNT_BALANCE_SOURCE_DEFINED_NAME_PREFIX):
            wb.defined_names.pop(name, None)

def _currency_code(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip().upper()
        if text:
            return text
    return "UNKNOWN"

def _symbol_quote_currency(symbol: Any) -> str:
    token = str(symbol or "").upper().replace("/", "").replace("-", "").replace("_", "").strip()
    if not token:
        return ""
    for quote in ("USDT", "USDC", "USD", "BTC", "ETH", "AUD"):
        if token.endswith(quote) and len(token) > len(quote):
            return quote
    return ""

def _infer_trade_log_currency(row: Dict[str, Any], *, field: str) -> str:
    row_type = str(row.get("row_type") or "").strip().lower()
    symbol = str(row.get("symbol") or "").strip().upper()
    if row_type == "monthly_aud_reval" or symbol == "MONTHLY AUD P/L":
        return "AUD"
    explicit_fields = {
        "commission": ("commission_currency", "fee_currency", "currency", "account_currency"),
        "net_pnl": ("realized_pnl_currency", "result_currency", "currency", "account_currency"),
        "balance_after": ("balance_after_trade_currency", "account_currency", "currency", "result_currency"),
    }.get(field, ())
    explicit = _currency_code(*(row.get(k) for k in explicit_fields))
    if explicit != "UNKNOWN":
        return explicit
    account_fingerprint = " ".join(
        str(row.get(k) or "").upper() for k in ("account", "account_label", "source")
    )
    if any(tok in account_fingerprint for tok in ("OANDA", "PEPPERSTONE", "FOREX", " FX")):
        return "AUD"
    normalized_symbol = symbol.replace("/", "").replace("-", "").replace("_", "")
    if _is_likely_fx_pair(normalized_symbol):
        return "AUD"
    is_crypto_account = any(tok in account_fingerprint for tok in ("BYBIT", "BINANCE", "COINSPOT"))
    is_crypto_row = is_crypto_account or str(row.get("asset_class") or "").strip().lower() == "crypto"
    if is_crypto_row:
        quote = _symbol_quote_currency(symbol)
        if quote in {"USDT", "USDC", "USD", "BTC", "ETH"}:
            return quote
    return ""

def _is_crypto_currency(code: str) -> bool:
    c = str(code or "").upper()
    return c in {"USDT", "BTC", "ETH", "SOL", "XRP", "USDC"}

def _currency_number_format(code: str, *, force_decimals: int | None = None) -> str:
    c = _currency_code(code)
    if force_decimals is not None:
        if _is_crypto_currency(c):
            decimals = "#" * min(8, max(0, force_decimals))
            return f'#,##0.{decimals} "{c}"' if decimals else f'#,##0 "{c}"'
        decimals = "0" * max(0, force_decimals)
        return f'#,##0.{decimals} "{c}"'
    if _is_crypto_currency(c):
        return f'#,##0.######## "{c}"'
    return f'#,##0.00 "{c}"'

def _fmt_detail_src(src: Any) -> str:
    if not isinstance(src, dict):
        return "-"
    sym = str(src.get("symbol") or src.get("instrument") or "").strip() or "-"
    d = _as_date(src.get("close_time") or src.get("date") or src.get("open_time"))
    return f"{sym} {d.isoformat() if d else '-'}"


INLINE_SOURCE_METRIC_KEYS = {
    "min_stop_pct", "max_stop_pct", "min_target_pct", "max_target_pct",
    "min_result_pct", "max_result_pct", "min_r_multiple", "max_r_multiple",
    "shortest_duration_seconds", "longest_duration_seconds",
    "min_duration_seconds", "max_duration_seconds",
    "min_duration_seconds_winners", "max_duration_seconds_winners",
    "min_duration_seconds_losers", "max_duration_seconds_losers",
    "min_move_to_break_even_duration_seconds", "max_move_to_break_even_duration_seconds",
    "min_move_to_profit_duration_seconds", "max_move_to_profit_duration_seconds",
    "min_stop_pct_winners", "max_stop_pct_winners",
    "min_target_pct_winners", "max_target_pct_winners",
    "min_stop_pct_losers", "max_stop_pct_losers",
    "min_target_pct_losers", "max_target_pct_losers",
    "min_result_pct_winners", "max_result_pct_winners",
    "min_r_multiple_winners", "max_r_multiple_winners",
    "min_result_pct_losers", "max_result_pct_losers",
    "min_r_multiple_losers", "max_r_multiple_losers",
}


def _inline_metric_source(bucket: Mapping[str, Any], key: str | None) -> Any:
    if not key or key not in INLINE_SOURCE_METRIC_KEYS:
        return None
    sources = bucket.get("metric_sources") if isinstance(bucket, Mapping) else None
    if not isinstance(sources, Mapping):
        return None
    source = sources.get(key)
    if source:
        return source
    aliases = {
        "min_duration_seconds": "shortest_duration_seconds",
        "max_duration_seconds": "longest_duration_seconds",
        "shortest_duration_seconds": "min_duration_seconds",
        "longest_duration_seconds": "max_duration_seconds",
    }
    return sources.get(aliases.get(key, ""))


def _inline_metric_source_has_symbol_and_date(source: Any) -> bool:
    if not isinstance(source, Mapping):
        return False
    symbol = str(
        source.get("symbol") or source.get("instrument") or ""
    ).strip()
    trade_date = _as_date(
        source.get("close_time")
        or source.get("date")
        or source.get("open_time")
    )
    return bool(symbol and trade_date)


def _format_inline_metric_value(value: Any, kind: str, source: Any = None) -> Any:
    if value in (None, ""):
        return ""
    if kind == "pct":
        number = _as_float(value)
        rendered = str(value) if number is None else f"{number:.8f}".rstrip("0").rstrip(".") + "%"
    elif kind == "r":
        number = _as_float(value)
        rendered = str(value) if number is None else f'{number:.3f}R'
    elif kind == "duration":
        rendered = _format_duration_display(value)
    else:
        rendered = str(value)
    source_text = _fmt_detail_src(source) if source else ""
    return f"{rendered} - {source_text}" if source_text and source_text != "-" else rendered


def _stable_trade_identity_key(row: Mapping[str, Any]) -> Tuple[str, str, str]:
    symbol = str(row.get("symbol") or row.get("symbol_raw") or row.get("instrument") or "").strip().upper()
    trade_date = _as_date(row.get("close_time") or row.get("date") or row.get("open_time"))
    row_id = str(row.get("id") or row.get("row_id") or stable_row_id(dict(row)) or "").strip()
    return symbol, trade_date.isoformat() if trade_date else "", row_id


def _pick_deterministic_extreme(
    samples: Sequence[Tuple[float, Mapping[str, Any]]],
    mode: str,
) -> Tuple[float, Mapping[str, Any]] | None:
    clean = [
        (float(value), row)
        for value, row in samples
        if value is not None and math.isfinite(float(value))
    ]
    if not clean:
        return None
    extreme_value = (min if mode == "min" else max)(value for value, _row in clean)
    tied = [(value, row) for value, row in clean if value == extreme_value]
    return min(tied, key=lambda item: _stable_trade_identity_key(item[1]))


def _trade_metric_ref(row: Dict[str, Any], metric_key: str | None = None, metric_value: Any = None) -> Dict[str, Any]:
    return {
        "id": row.get("id"),
        "symbol": row.get("symbol") or row.get("symbol_raw") or row.get("instrument"),
        "asset_class": row.get("asset_class"),
        "side": row.get("side"),
        "open_time": row.get("open_time"),
        "close_time": row.get("close_time"),
        "date": row.get("close_time") or row.get("open_time"),
        "account": row.get("account_label") or row.get("account"),
        "source": row.get("source"),
        "metric_key": metric_key,
        "metric_value": metric_value,
    }


def _fmt_detail_datetime(value: Any) -> str:
    dt = _as_datetime(value)
    if dt is None:
        text = str(value or "").strip()
        text = text.replace("T", " ")
        return text[:19] if len(text) >= 19 else text
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def _fmt_period_detail(detail: Any) -> str:
    if not isinstance(detail, dict):
        return ""
    start = detail.get("start_time")
    end = detail.get("end_time")
    if not (start or end):
        return ""
    return f"{_fmt_detail_datetime(start) or '?'} to {_fmt_detail_datetime(end) or '?'}"


def _fmt_streak_detail(detail: Any) -> str:
    if not isinstance(detail, dict):
        return ""
    start = detail.get("start_time")
    end = detail.get("end_time")
    if not (start or end):
        return ""
    return f"{_fmt_detail_datetime(start) or '?'} to {_fmt_detail_datetime(end) or '?'}"


def _fmt_pct_with_detail(value: Any, detail: Any) -> str:
    number = _as_float(value)
    if number is None:
        return ""
    suffix = _fmt_period_detail(detail)
    text = f"{number:.6g}%"
    return f"{text} ({suffix})" if suffix else text


def _fmt_count_with_detail(value: Any, detail: Any) -> str:
    number = _as_float(value)
    if number is None:
        return ""
    suffix = _fmt_streak_detail(detail)
    text = str(int(number))
    return f"{text} ({suffix})" if suffix else text


_DRAWDOWN_INLINE_RE = re.compile(
    r"^\s*(?P<pct>[-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*%?\s*(?:\((?P<start>.*?)\s+to\s+(?P<end>.*?)\))?\s*$",
    re.IGNORECASE,
)


def _drawdown_detail_start_end(detail: Any) -> Tuple[str, str]:
    if isinstance(detail, dict):
        return _fmt_detail_datetime(detail.get("start_time")), _fmt_detail_datetime(detail.get("end_time"))
    text = str(detail or "").strip()
    if not text:
        return "", ""
    match = _DRAWDOWN_INLINE_RE.match(text)
    if not match:
        match = re.search(r"\((?P<start>.*?)\s+to\s+(?P<end>.*?)\)", text, re.IGNORECASE)
    if not match:
        return "", ""
    return _fmt_detail_datetime(match.group("start")), _fmt_detail_datetime(match.group("end"))


def _drawdown_inline_pct_fraction(value: Any) -> float | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    if "%" not in text and "(" not in text:
        return None
    match = _DRAWDOWN_INLINE_RE.match(text)
    if not match:
        return None
    number = _as_float(match.group("pct"))
    return None if number is None else number / 100.0


def _drawdown_detail_cell_value(detail: Any, endpoint: str) -> str:
    start, end = _drawdown_detail_start_end(detail)
    return start if endpoint == "start" else end


_STREAK_INLINE_RE = re.compile(
    r"^\s*(?P<count>\d+(?:\.0+)?)\s*(?:\((?P<start>.*?)\s+to\s+(?P<end>.*?)\))?\s*$",
    re.IGNORECASE,
)


def _streak_detail_start_end(detail: Any) -> Tuple[str, str]:
    if isinstance(detail, dict):
        return _fmt_detail_datetime(detail.get("start_time")), _fmt_detail_datetime(detail.get("end_time"))
    text = str(detail or "").strip()
    if not text:
        return "", ""
    match = _STREAK_INLINE_RE.match(text)
    if not match:
        match = re.search(r"\((?P<start>.*?)\s+to\s+(?P<end>.*?)\)", text, re.IGNORECASE)
    if not match:
        return "", ""
    return _fmt_detail_datetime(match.group("start")), _fmt_detail_datetime(match.group("end"))


def _streak_inline_count(value: Any) -> int | None:
    number = _as_float(value)
    if number is not None and math.isfinite(number):
        return int(number)
    if not isinstance(value, str):
        return None
    match = _STREAK_INLINE_RE.match(value.strip())
    if not match:
        return None
    count = _as_float(match.group("count"))
    return int(count) if count is not None else None


def _apply_streak_detail_value_font(cell) -> None:
    if cell.value in (None, ""):
        return
    font = copy(cell.font)
    font.color = "FF000000"
    cell.font = font


def _write_streak_detail_rows(
    ws,
    metric_label: str,
    details_by_col: Dict[int, Any],
    counts_by_col: Dict[int, Any] | None = None,
    *,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    wanted = metric_label.casefold()
    metric_rows = [
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip().casefold() == wanted
    ]
    for metric_row in metric_rows:
        detail_rows = _ensure_start_end_rows_after(ws, metric_row, diagnostics)
        if detail_rows is None:
            continue
        start_row, end_row = detail_rows
        ws.cell(start_row, 1).value = "Start"
        ws.cell(end_row, 1).value = "End"
        _apply_dashboard_source_label_style(ws.cell(start_row, 1))
        _apply_dashboard_source_label_style(ws.cell(end_row, 1))
        for col in range(2, ws.max_column + 1):
            inline_value = ws.cell(metric_row, col).value
            count = _streak_inline_count((counts_by_col or {}).get(col))
            if count is None:
                count = _streak_inline_count(inline_value)
            if count is not None:
                ws.cell(metric_row, col).value = count
                ws.cell(metric_row, col).number_format = "0"
            inline_start, inline_end = _streak_detail_start_end(inline_value)
            start, end = _streak_detail_start_end(details_by_col.get(col))
            start = start or inline_start
            end = end or inline_end
            if start:
                ws.cell(start_row, col).value = start
                ws.cell(start_row, col).number_format = "General"
            if end:
                ws.cell(end_row, col).value = end
                ws.cell(end_row, col).number_format = "General"
            _apply_streak_detail_value_font(ws.cell(start_row, col))
            _apply_streak_detail_value_font(ws.cell(end_row, col))


def _copy_row_style_without_values(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target.value = None
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    source_height = ws.row_dimensions[source_row].height
    if source_height is not None:
        ws.row_dimensions[target_row].height = source_height


def _has_dashboard_manual_merge_below(ws, metric_row: int) -> bool:
    if getattr(ws, "title", "") != STATS1_SHEET:
        return False
    return any(
        merged.min_col == 1
        and merged.max_col == 1
        and merged.min_row > metric_row
        for merged in ws.merged_cells.ranges
    )


def _ensure_start_end_rows_after(ws, metric_row: int, diagnostics: Dict[str, Any] | None = None) -> Tuple[int, int] | None:
    def label_at(row: int) -> str:
        return str(ws.cell(row, 1).value or "").strip().casefold()

    if label_at(metric_row + 1) == "start" and label_at(metric_row + 2) == "end":
        return metric_row + 1, metric_row + 2
    if _has_dashboard_manual_merge_below(ws, metric_row):
        if diagnostics is not None:
            diagnostics.setdefault("skipped_detail_rows_due_to_manual_merges", []).append(f"{ws.title}: {metric_row}")
        return None
    if label_at(metric_row + 1) == "start":
        ws.insert_rows(metric_row + 2, 1)
        _copy_row_style_without_values(ws, metric_row + 1, metric_row + 2)
        ws.cell(metric_row + 2, 1).value = "End"
        if diagnostics is not None:
            diagnostics.setdefault("drawdown_detail_rows_inserted", []).append(f"{ws.title}: End")
        return metric_row + 1, metric_row + 2

    ws.insert_rows(metric_row + 1, 2)
    _copy_row_style_without_values(ws, metric_row, metric_row + 1)
    _copy_row_style_without_values(ws, metric_row, metric_row + 2)
    ws.cell(metric_row + 1, 1).value = "Start"
    ws.cell(metric_row + 2, 1).value = "End"
    if diagnostics is not None:
        diagnostics.setdefault("drawdown_detail_rows_inserted", []).append(f"{ws.title}: Start/End")
    return metric_row + 1, metric_row + 2


def _ensure_drawdown_start_end_rows(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    metric_rows = [
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip().casefold() in {"min drawdown", "max drawdown"}
    ]
    for row in sorted(metric_rows, reverse=True):
        _ensure_start_end_rows_after(ws, row, diagnostics)


def _write_drawdown_detail_rows(
    ws,
    metric_label: str,
    details_by_col: Dict[int, Any],
    *,
    diagnostics: Dict[str, Any] | None = None,
    preserve_inline_cells: set | None = None,
) -> None:
    wanted = metric_label.casefold()
    metric_rows = [
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip().casefold() == wanted
    ]
    for metric_row in metric_rows:
        detail_rows = _ensure_start_end_rows_after(ws, metric_row, diagnostics)
        if detail_rows is None:
            continue
        start_row, end_row = detail_rows
        ws.cell(start_row, 1).value = "Start"
        ws.cell(end_row, 1).value = "End"
        _apply_dashboard_source_label_style(ws.cell(start_row, 1))
        _apply_dashboard_source_label_style(ws.cell(end_row, 1))
        for col in range(2, ws.max_column + 1):
            preserving_inline = bool(preserve_inline_cells and (metric_row, col) in preserve_inline_cells)
            detail = None if preserving_inline else details_by_col.get(col)
            inline_value = ws.cell(metric_row, col).value
            inline_start, inline_end = _drawdown_detail_start_end(inline_value)
            start, end = _drawdown_detail_start_end(detail)
            start = inline_start or start
            end = inline_end or end
            inline_pct = _drawdown_inline_pct_fraction(inline_value)
            if inline_pct is not None:
                ws.cell(metric_row, col).value = inline_pct
                ws.cell(metric_row, col).number_format = "0.00%"
            if start:
                ws.cell(start_row, col).value = start
                ws.cell(start_row, col).number_format = "General"
            if end:
                ws.cell(end_row, col).value = end
                ws.cell(end_row, col).number_format = "General"


def _fmt_currency_with_trade_detail(value: Any, currency: str, source: Any) -> str:
    number = _as_float(value)
    if number is None:
        return ""
    detail = str(source or "").strip()
    text = f"{number:,.10f}".rstrip("0").rstrip(".")
    base = f"{currency} {text}" if currency else text
    return f"{base} ({detail})" if detail else base


def _excel_scalar(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, (int, float, bool, str)):
        return value
    if isinstance(value, dict):
        if 'symbol' in value and any(k in value for k in ('wins','losses','total_trades','trades')):
            symbol = value.get('symbol') or 'N/A'
            wins = value.get('wins', '')
            losses = value.get('losses', '')
            trades = value.get('total_trades', value.get('trades', ''))
            return f"{symbol} - Wins {wins} / Losses {losses} / Trades {trades}"
        return ', '.join(f"{k}={value.get(k)}" for k in sorted(value.keys()))
    if isinstance(value, (list, tuple, set)):
        return ', '.join(str(_excel_scalar(v)) for v in value)
    return str(value)


def _instrument_leader_scalar(value: Any, count_key: str) -> Any:
    if not isinstance(value, dict):
        return _excel_scalar(value)
    symbol = str(value.get("symbol") or "N/A").strip() or "N/A"
    if count_key == "wins":
        count = value.get("wins")
        label = "Wins"
    else:
        count = value.get("losses")
        label = "Losses"
    if count is None:
        count = ""
    return f"{symbol} - {label} {count}"

def stable_row_id(row: Dict[str, Any]) -> str:
    rid=str(row.get('id') or row.get('__row_id') or '').strip()
    if rid and not rid.startswith('monthly_aud_reval:'):
        return rid
    if rid and _monthly_aud_reval_row_id_month(rid) and _is_monthly_aud_reval_semantic_row(row):
        return rid
    refs=row.get('raw_refs') if isinstance(row.get('raw_refs'),dict) else {}
    parts=[str(row.get('account_label') or row.get('account') or ''),str(row.get('symbol') or ''),str(row.get('side') or ''),str(row.get('open_time') or ''),str(row.get('close_time') or ''),str(row.get('qty') or row.get('qty_raw') or ''),str(row.get('entry_price') or ''),str(row.get('exit_price') or ''),str(row.get('net_profit') or row.get('result_cash') or ''),str(row.get('source') or ''),str(row.get('source_file') or ''),str(row.get('workbook_name') or ''),str(refs.get('source_file') or ''),str(refs.get('workbook') or ''),str(refs.get('sheet') or ''),str(refs.get('source_row') or ''),str(refs.get('period_month') or '')]
    return 'sig:'+hashlib.sha1('|'.join(parts).encode('utf-8')).hexdigest()[:24]


def _execution_datetime_token(value: Any) -> str:
    parsed = _as_datetime(value)
    if parsed is None:
        return ""
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=JOURNAL_DISPLAY_TZ)
    return str(int(parsed.timestamp()))


def _execution_number_token(value: Any, digits: int = 8) -> str:
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        return ""
    return f"{number:.{digits}f}"


def _trade_execution_fingerprint(row: Dict[str, Any]) -> Optional[str]:
    if not isinstance(row, dict) or str(row.get("row_type") or "trade").strip().lower() != "trade":
        return None
    account = _canonical_account_label(row.get("account_label") or row.get("account")).upper()
    asset_class = str(row.get("asset_class") or _trade_row_market(row) or "").strip().lower()
    symbol = "".join(ch for ch in str(row.get("symbol") or row.get("symbol_raw") or "").upper() if ch.isalnum())
    side_raw = str(row.get("side") or "").strip().upper()
    side = "BUY" if side_raw in {"BUY", "LONG"} or side_raw.startswith("BUY") else (
        "SELL" if side_raw in {"SELL", "SHORT"} or side_raw.startswith("SELL") else side_raw
    )
    open_time = _execution_datetime_token(row.get("open_time"))
    close_time = _execution_datetime_token(row.get("close_time"))
    qty = _execution_number_token(row.get("qty") if row.get("qty") is not None else row.get("qty_raw"))
    entry = _execution_number_token(row.get("entry_price"))
    exit_price = _execution_number_token(row.get("exit_price"))
    if not all((account, asset_class, symbol, side, open_time, close_time, qty, entry, exit_price)):
        return None
    return "|".join([
        account,
        asset_class,
        symbol,
        side,
        open_time,
        close_time,
        qty,
        entry,
        exit_price,
        _execution_number_token(row.get("stop_loss")),
        _execution_number_token(row.get("take_profit")),
    ])


def _trade_execution_row_id(row: Dict[str, Any]) -> str:
    fingerprint = _trade_execution_fingerprint(row)
    if fingerprint:
        return "sig:" + hashlib.sha1(fingerprint.encode("utf-8")).hexdigest()[:24]
    row_without_id = dict(row)
    row_without_id.pop("id", None)
    row_without_id.pop("__row_id", None)
    return stable_row_id(row_without_id)


def _trade_source_execution_fingerprint(row: Dict[str, Any]) -> Optional[str]:
    execution = _trade_execution_fingerprint(row)
    if not execution:
        return None
    parts = execution.split("|")
    return "|".join([*parts[:7], *parts[9:11]])


def _trade_row_source_rank(row: Dict[str, Any]) -> int:
    row_id = str(row.get("id") or row.get("__row_id") or "").strip().lower()
    source = str(row.get("source") or "").strip().lower()
    import_source = str(row.get("import_source") or "").strip().lower()
    combined = " ".join((row_id, source, import_source))
    if row_id.startswith(("oanda_export:", "bybit:", "bybit_")):
        return 500
    if source in {"oanda_transaction_export", "oanda_export", "bybit"} or import_source in {
        "oanda_transaction_export", "oanda_export", "bybit"
    }:
        return 500
    if row_id.startswith("sig:"):
        return 300
    if row_id.startswith("excel:") or source in {"excel", "local_excel", "master_journal"}:
        return 100
    if "oanda_export" in combined or "bybit" in combined:
        return 500
    return 200


def _merge_duplicate_execution_rows(primary: Dict[str, Any], duplicate: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(primary)
    preserve_fields = {
        "trade_number", "manual_overrides", "manual_override_fields", "notes", "pre_trade_comments",
        "entry_comments", "trade_management", "exit_comments", "flags", "setup", "timeframe",
        "breakeven", "_normalization_manual_overrides", *TRADE_LOG_MANUAL_FIELD_MAP.values(),
    }
    for field in preserve_fields:
        if merged.get(field) in (None, "", [], {}) and duplicate.get(field) not in (None, "", [], {}):
            merged[field] = duplicate.get(field)
    return merged


def _dedupe_trade_rows_by_execution(
    rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    groups: Dict[str, List[Tuple[int, Dict[str, Any]]]] = defaultdict(list)
    passthrough: Dict[int, Dict[str, Any]] = {}
    for index, row in enumerate(rows):
        fingerprint = _trade_execution_fingerprint(row)
        if fingerprint:
            groups[fingerprint].append((index, row))
        else:
            passthrough[index] = row

    replacements: Dict[int, Dict[str, Any]] = {}
    removed_indexes: set[int] = set()
    duplicate_groups = 0
    rows_removed = 0
    ambiguous: List[Dict[str, Any]] = []
    for fingerprint, entries in groups.items():
        if len(entries) == 1:
            replacements[entries[0][0]] = entries[0][1]
            continue
        ranked = sorted(
            entries,
            key=lambda item: (-_trade_row_source_rank(item[1]), item[0]),
        )
        best_rank = _trade_row_source_rank(ranked[0][1])
        best = [entry for entry in ranked if _trade_row_source_rank(entry[1]) == best_rank]
        best_ids = {str(entry[1].get("id") or "").strip() for entry in best}
        if len(best) > 1 and len(best_ids) > 1:
            ambiguous.append({
                "fingerprint": fingerprint,
                "row_ids": [str(entry[1].get("id") or "") for entry in entries],
            })
            for index, row in entries:
                replacements[index] = row
            continue
        canonical_index, canonical = ranked[0]
        for index, duplicate in ranked[1:]:
            canonical = _merge_duplicate_execution_rows(canonical, duplicate)
            removed_indexes.add(index)
            rows_removed += 1
        replacements[canonical_index] = canonical
        duplicate_groups += 1

    result: List[Dict[str, Any]] = []
    for index in range(len(rows)):
        if index in removed_indexes:
            continue
        if index in passthrough:
            result.append(passthrough[index])
        elif index in replacements:
            result.append(replacements[index])

    source_execution_groups: Dict[str, List[Tuple[int, Dict[str, Any]]]] = defaultdict(list)
    source_execution_passthrough: Dict[int, Dict[str, Any]] = {}
    for index, row in enumerate(result):
        fingerprint = _trade_source_execution_fingerprint(row)
        if fingerprint:
            source_execution_groups[fingerprint].append((index, row))
        else:
            source_execution_passthrough[index] = row
    source_execution_replacements: Dict[int, Dict[str, Any]] = {}
    source_execution_removed: set[int] = set()
    for fingerprint, entries in source_execution_groups.items():
        if len(entries) == 1:
            source_execution_replacements[entries[0][0]] = entries[0][1]
            continue
        ranked = sorted(entries, key=lambda item: (-_trade_row_source_rank(item[1]), item[0]))
        best_rank = _trade_row_source_rank(ranked[0][1])
        best = [entry for entry in ranked if _trade_row_source_rank(entry[1]) == best_rank]
        best_ids = {str(entry[1].get("id") or "").strip() for entry in best}
        if len(best) > 1 and len(best_ids) > 1:
            ambiguous.append({
                "fingerprint": fingerprint,
                "row_ids": [str(entry[1].get("id") or "") for entry in entries],
                "reason": "duplicate_source_execution",
            })
            for index, row in entries:
                source_execution_replacements[index] = row
            continue
        canonical_index, canonical = ranked[0]
        for index, duplicate in ranked[1:]:
            canonical = _merge_duplicate_execution_rows(canonical, duplicate)
            source_execution_removed.add(index)
            rows_removed += 1
        source_execution_replacements[canonical_index] = canonical
        duplicate_groups += 1
    if source_execution_removed:
        result = [
            source_execution_passthrough.get(index, source_execution_replacements.get(index))
            for index in range(len(result))
            if index not in source_execution_removed
        ]
    return result, {
        "duplicate_execution_groups_removed": duplicate_groups,
        "duplicate_execution_rows_removed": rows_removed,
        "ambiguous_duplicate_execution_groups": ambiguous,
    }


_EXCEL_ROW_ID_RE = re.compile(
    r"^excel:(?P<account>[^:]+):(?P<sheet>[^:]+):(?P<row>\d+):(?P<symbol>[^:]+):(?P<opened>.+)$",
    re.IGNORECASE,
)


def _stale_excel_row_id_reasons(row_id: str, row: Dict[str, Any]) -> List[str]:
    match = _EXCEL_ROW_ID_RE.match(str(row_id or "").strip())
    if not match:
        return []
    reasons: List[str] = []
    encoded_account = _canonical_account_label(match.group("account")).upper()
    visible_account = _canonical_account_label(row.get("account_label") or row.get("account")).upper()
    if encoded_account and visible_account and encoded_account != visible_account:
        reasons.append("account")
    encoded_symbol = "".join(ch for ch in match.group("symbol").upper() if ch.isalnum())
    visible_symbol = "".join(ch for ch in str(row.get("symbol") or "").upper() if ch.isalnum())
    if encoded_symbol and visible_symbol and encoded_symbol != visible_symbol:
        reasons.append("symbol")
    encoded_date = _as_date(match.group("opened"))
    visible_date = _as_date(row.get("open_time") or row.get("close_time"))
    if encoded_date and visible_date and encoded_date != visible_date:
        reasons.append("date")
    return reasons




def _all_trades_row_fingerprint_from_map(values: Dict[str, Any]) -> str:
    parts = [str(values.get(k) or '') for k in ['Account','Symbol','Side','Open Time','Close Time','Qty','Entry Price','Exit Price','Net P/L']]
    return 'sig:' + hashlib.sha1('|'.join(parts).encode('utf-8')).hexdigest()[:24]


def _master_journal_manual_override_keys(row: Mapping[str, Any]) -> List[str]:
    """Return every stable key by which a workbook override may identify a row."""

    row_map = dict(row)
    keys = [
        str(row_map.get("id") or "").strip(),
        str(row_map.get("__row_id") or "").strip(),
        stable_row_id(row_map),
        _all_trades_row_fingerprint_from_map(
            {
                "Account": row_map.get("account_label") or row_map.get("account"),
                "Symbol": row_map.get("symbol"),
                "Side": row_map.get("side"),
                "Open Time": row_map.get("open_time"),
                "Close Time": row_map.get("close_time"),
                "Qty": row_map.get("qty"),
                "Entry Price": row_map.get("entry_price"),
                "Exit Price": row_map.get("exit_price"),
                "Net P/L": row_map.get("net_profit"),
            }
        ),
    ]
    return list(dict.fromkeys(key for key in keys if key))


def _normalize_master_journal_rows(
    rows: Sequence[Mapping[str, Any]],
    manual_overrides: Optional[Mapping[str, Mapping[str, Any]]] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """Canonicalize rows once for workbook values and recommendation artifacts.

    Manual edits are attached before execution de-duplication so an edit belonging
    to a lower-ranked duplicate can still follow the surviving execution.  The
    edit values themselves are applied after de-duplication, which prevents them
    from changing execution identity or weighting recommendation calculations.
    """

    overrides = manual_overrides if isinstance(manual_overrides, Mapping) else {}
    prepared: List[Dict[str, Any]] = []
    ignored_rows = 0
    override_rows_matched = 0
    for raw_row in rows or []:
        if not isinstance(raw_row, Mapping):
            ignored_rows += 1
            continue
        row = dict(raw_row)
        row_type = str(row.get("row_type") or "trade").strip().lower()
        if row_type not in {"trade", "monthly_aud_reval", "cashflow"}:
            ignored_rows += 1
            continue
        row["row_type"] = row_type
        row = _repair_or_flag_zero_trade_qty(row)
        matched_override: Dict[str, Any] = {}
        for key in _master_journal_manual_override_keys(row):
            candidate = overrides.get(key)
            if isinstance(candidate, Mapping):
                matched_override = dict(candidate)
                break
        if matched_override:
            override_rows_matched += 1
            row["_normalization_manual_overrides"] = matched_override
        prepared.append(row)

    normalized, execution_diagnostics = _dedupe_trade_rows_by_execution(prepared)
    normalized, monthly_diagnostics = _dedupe_monthly_aud_reval_rows(normalized)
    overrides_applied = 0
    for row in normalized:
        matched_override = row.pop("_normalization_manual_overrides", None)
        if not isinstance(matched_override, Mapping) or not matched_override:
            continue
        safe_override = dict(matched_override)
        row.update(safe_override)
        existing = row.get("manual_overrides")
        combined = dict(existing) if isinstance(existing, Mapping) else {}
        combined.update(safe_override)
        row["manual_overrides"] = combined
        row["manual_override_fields"] = sorted(combined)
        overrides_applied += 1

    diagnostics = {
        **execution_diagnostics,
        **monthly_diagnostics,
        "ignored_master_journal_rows": ignored_rows,
        "manual_override_rows_matched": override_rows_matched,
        "manual_override_rows_applied": overrides_applied,
        "normalized_master_journal_rows": len(normalized),
    }
    return normalized, diagnostics



def _trade_log_two_row_header_values_for(headers: List[str]) -> Tuple[List[str], List[str]]:
    row1: List[str] = []
    row2: List[str] = []
    for header in headers:
        if header in MOVE_TO_BREAK_EVEN_HEADERS:
            row1.append("Move to Break-Even" if header == MOVE_TO_BREAK_EVEN_HEADERS[0] else "")
            row2.append(MOVE_TO_SUBHEADERS[MOVE_TO_BREAK_EVEN_HEADERS.index(header)])
        elif header in MOVE_TO_PROFIT_HEADERS:
            row1.append("Move to Profit" if header == MOVE_TO_PROFIT_HEADERS[0] else "")
            row2.append(MOVE_TO_SUBHEADERS[MOVE_TO_PROFIT_HEADERS.index(header)])
        else:
            row1.append(_trade_log_visible_header(header))
            row2.append("")
    return row1, row2


def _trade_log_two_row_header_values() -> Tuple[List[str], List[str]]:
    return _trade_log_two_row_header_values_for(TRADE_LOG_HEADERS)


def _trade_log_three_row_header_values_for(headers: List[str]) -> Tuple[List[str], List[str], List[str]]:
    row1, row2 = _trade_log_two_row_header_values_for(headers)
    return row1, row2, [""] * len(headers)


def _trade_log_three_row_header_values() -> Tuple[List[str], List[str], List[str]]:
    return _trade_log_three_row_header_values_for(TRADE_LOG_HEADERS)


def _trade_log_has_three_row_headers_for(ws, headers: List[str]) -> bool:
    row1, row2, row3 = _trade_log_three_row_header_values_for(headers)
    found1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, len(headers) + 1)]
    found2 = [str(ws.cell(2, c).value or "").strip() for c in range(1, len(headers) + 1)]
    found3 = [str(ws.cell(3, c).value or "").strip() for c in range(1, len(headers) + 1)]
    if found1 != row1 or found2 != row2 or found3 != row3:
        return False
    merged_cells = getattr(ws, "merged_cells", None)
    if merged_cells is None:
        return True
    found_merges = {str(merged) for merged in merged_cells.ranges}
    expected_vertical_merges = {
        f"{get_column_letter(col)}1:{get_column_letter(col)}3"
        for col, header in enumerate(headers, start=1)
        if header not in MOVE_TO_FIELD_MAP
    }
    expected_subheader_merges = {
        f"{get_column_letter(col)}2:{get_column_letter(col)}3"
        for col, header in enumerate(headers, start=1)
        if header in MOVE_TO_FIELD_MAP
    }
    return expected_vertical_merges.issubset(found_merges) and expected_subheader_merges.issubset(found_merges)


def _trade_log_has_three_row_headers(ws) -> bool:
    return _trade_log_has_three_row_headers_for(ws, TRADE_LOG_HEADERS)


def _trade_log_has_two_row_headers_for(ws, headers: List[str]) -> bool:
    row1, row2 = _trade_log_two_row_header_values_for(headers)
    found1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, len(headers) + 1)]
    found2 = [str(ws.cell(2, c).value or "").strip() for c in range(1, len(headers) + 1)]
    expected_vertical_merges = {
        f"{get_column_letter(col)}1:{get_column_letter(col)}2"
        for col, header in enumerate(headers, start=1)
        if header not in MOVE_TO_FIELD_MAP
    }
    merged_cells = getattr(ws, "merged_cells", None)
    if merged_cells is None:
        return found1 == row1 and found2 == row2
    found_merges = {str(merged) for merged in merged_cells.ranges}
    return found1 == row1 and found2 == row2 and expected_vertical_merges.issubset(found_merges)


def _trade_log_has_two_row_headers(ws) -> bool:
    return _trade_log_has_two_row_headers_for(ws, TRADE_LOG_HEADERS)


def _trade_log_has_legacy_duplicate_two_row_headers_for(ws, headers: List[str]) -> bool:
    row1, row2 = _trade_log_two_row_header_values_for(headers)
    duplicate_row2 = [
        _trade_log_visible_header(header) if header not in MOVE_TO_FIELD_MAP else row2[col - 1]
        for col, header in enumerate(headers, start=1)
    ]
    found1 = [str(ws.cell(1, c).value or "").strip() for c in range(1, len(headers) + 1)]
    found2 = [str(ws.cell(2, c).value or "").strip() for c in range(1, len(headers) + 1)]
    return found1 == row1 and found2 == duplicate_row2


def _trade_log_has_legacy_duplicate_two_row_headers(ws) -> bool:
    return _trade_log_has_legacy_duplicate_two_row_headers_for(ws, TRADE_LOG_HEADERS)


def _trade_log_has_v1_grouped_two_row_headers(ws) -> bool:
    return _trade_log_has_two_row_headers_for(ws, TRADE_LOG_HEADERS_V1) or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, TRADE_LOG_HEADERS_V1)


def _trade_log_has_recommendation_grouped_three_row_headers(ws) -> bool:
    return _trade_log_has_three_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS)


def _trade_log_has_recommendation_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS)
    )


def _trade_log_has_recommendation_v1_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS_V1)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    )


def _trade_log_has_pre_vwap_recommendation_grouped_three_row_headers(ws) -> bool:
    return _trade_log_has_three_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)


def _trade_log_has_pre_vwap_recommendation_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
    )


def _trade_log_has_pre_vwap_recommendation_v1_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    )


def _trade_log_has_pre_recommendation_grouped_three_row_headers(ws) -> bool:
    return _trade_log_has_three_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS)


def _trade_log_has_pre_recommendation_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
    )


def _trade_log_has_pre_recommendation_v1_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
        or _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    )


def _trade_log_uses_grouped_two_row_headers(ws) -> bool:
    return (
        _trade_log_has_three_row_headers(ws)
        or _trade_log_has_two_row_headers(ws)
        or _trade_log_has_legacy_duplicate_two_row_headers(ws)
        or _trade_log_has_v1_grouped_two_row_headers(ws)
        or _trade_log_has_recommendation_grouped_three_row_headers(ws)
        or _trade_log_has_recommendation_grouped_two_row_headers(ws)
        or _trade_log_has_recommendation_v1_grouped_two_row_headers(ws)
        or _trade_log_has_pre_vwap_recommendation_grouped_three_row_headers(ws)
        or _trade_log_has_pre_vwap_recommendation_grouped_two_row_headers(ws)
        or _trade_log_has_pre_vwap_recommendation_v1_grouped_two_row_headers(ws)
        or _trade_log_has_pre_recommendation_grouped_three_row_headers(ws)
        or _trade_log_has_pre_recommendation_grouped_two_row_headers(ws)
        or _trade_log_has_pre_recommendation_v1_grouped_two_row_headers(ws)
    )


def _trade_log_header_map(ws) -> Dict[str, int]:
    if _trade_log_has_three_row_headers(ws):
        return {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    if _trade_log_has_two_row_headers(ws) or _trade_log_has_legacy_duplicate_two_row_headers(ws):
        return {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    if _trade_log_has_pre_recommendation_grouped_three_row_headers(ws) or _trade_log_has_pre_recommendation_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(PRE_RECOMMENDATION_TRADE_LOG_HEADERS, start=1)}
    if _trade_log_has_recommendation_grouped_three_row_headers(ws) or _trade_log_has_recommendation_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(RECOMMENDATION_TRADE_LOG_HEADERS, start=1)}
    if _trade_log_has_pre_vwap_recommendation_grouped_three_row_headers(ws) or _trade_log_has_pre_vwap_recommendation_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS, start=1)}
    if _trade_log_has_v1_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(TRADE_LOG_HEADERS_V1, start=1)}
    if _trade_log_has_recommendation_v1_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(RECOMMENDATION_TRADE_LOG_HEADERS_V1, start=1)}
    if _trade_log_has_pre_vwap_recommendation_v1_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1, start=1)}
    if _trade_log_has_pre_recommendation_v1_grouped_two_row_headers(ws):
        return {header: col for col, header in enumerate(PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1, start=1)}
    return {
        str(ws.cell(1, c).value or "").strip(): c
        for c in range(1, ws.max_column + 1)
        if str(ws.cell(1, c).value or "").strip()
    }


def _trade_log_data_start_row(ws) -> int:
    if _trade_log_has_three_row_headers(ws):
        return TRADE_LOG_DATA_START_ROW
    if _trade_log_has_recommendation_grouped_three_row_headers(ws):
        return TRADE_LOG_DATA_START_ROW
    if _trade_log_has_pre_vwap_recommendation_grouped_three_row_headers(ws):
        return TRADE_LOG_DATA_START_ROW
    if _trade_log_has_pre_recommendation_grouped_three_row_headers(ws):
        return TRADE_LOG_DATA_START_ROW
    if _trade_log_has_two_row_headers(ws) or _trade_log_has_legacy_duplicate_two_row_headers(ws) or _trade_log_has_v1_grouped_two_row_headers(ws):
        return 3
    if _trade_log_has_recommendation_grouped_two_row_headers(ws) or _trade_log_has_recommendation_v1_grouped_two_row_headers(ws):
        return 3
    if _trade_log_has_pre_vwap_recommendation_grouped_two_row_headers(ws) or _trade_log_has_pre_vwap_recommendation_v1_grouped_two_row_headers(ws):
        return 3
    if _trade_log_has_pre_recommendation_grouped_two_row_headers(ws) or _trade_log_has_pre_recommendation_v1_grouped_two_row_headers(ws):
        return 3
    return 2


def _trade_log_data_row_count(ws) -> int:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    start_row = _trade_log_data_start_row(ws)
    count = 0
    for row in range(start_row, ws.max_row + 1):
        if row_id_col and ws.cell(row, row_id_col).value not in (None, ""):
            count += 1
        elif any(ws.cell(row, col).value not in (None, "") for col in range(1, min(ws.max_column, len(TRADE_LOG_HEADERS)) + 1)):
            count += 1
    return count


def _trade_log_last_populated_row(ws) -> int:
    headers = _trade_log_header_map(ws)
    start_row = _trade_log_data_start_row(ws)
    last_col = max(headers.values(), default=ws.max_column)
    last_row = start_row - 1
    for row in range(start_row, ws.max_row + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in range(1, last_col + 1)):
            last_row = row
    return last_row


def _set_trade_log_auto_filter(ws) -> None:
    last_col = len(TRADE_LOG_HEADERS)
    last_row = max(
        TRADE_LOG_FILTER_HEADER_ROW,
        _trade_log_last_populated_row(ws),
    )
    ws.auto_filter.ref = f"A{TRADE_LOG_FILTER_HEADER_ROW}:{get_column_letter(last_col)}{last_row}"


def _normalize_trade_log_row_heights(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    if getattr(ws, "title", "") not in {TRADE_LOG_SHEET, LEGACY_ALL_TRADES_SHEET}:
        return
    row_keys = [idx for idx in ws.row_dimensions.keys() if isinstance(idx, int)]
    max_row = max([ws.max_row, *row_keys], default=ws.max_row)
    changed = 0
    source_height = TRADE_LOG_DATA_ROW_HEIGHT
    ws.sheet_format.defaultRowHeight = float(source_height)
    for row in range(TRADE_LOG_DATA_START_ROW, max_row + 1):
        if ws.row_dimensions[row].height != source_height:
            changed += 1
        ws.row_dimensions[row].height = source_height
    if diagnostics is not None and changed:
        diagnostics["normalized_trade_log_row_heights"] = changed
        diagnostics["trade_log_row_height_source"] = source_height


def _hide_trade_log_row_id(ws) -> None:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col:
        raise RuntimeError("Trade Log schema repair failed: missing Row ID header.")
    ws.column_dimensions[get_column_letter(row_id_col)].hidden = True


def _clear_trade_log_dropdown_validations(ws) -> None:
    keep = []
    editable_cols = {_trade_log_header_map(ws).get(h) for h in ["Test", "Pattern", "VWAP", "ATHS/ATLS", "Order", "Round Number", "Spiked Out", "Close Stopout", "Near Perfect Entry", "Near Win", "Early Close"]}
    editable_cols.discard(None)
    for dv in list(ws.data_validations.dataValidation):
        touches_editable = False
        for sq in dv.cells.ranges:
            if any(sq.min_col <= c <= sq.max_col for c in editable_cols):
                touches_editable = True
                break
        if not touches_editable:
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def _apply_trade_log_dropdown_validations(ws) -> None:
    headers = _trade_log_header_map(ws)
    _clear_trade_log_dropdown_validations(ws)
    max_row = max(TRADE_LOG_DATA_START_ROW, ws.max_row)
    specs = {
        "Test": '"Yes,No"',
        "ATHS/ATLS": '"All-Time High,All-Time Low"',
        "Order": '"Market,Limit"',
        "VWAP": '"Yes,No"',
        "Round Number": '"Yes,No"',
        "Spiked Out": '"Yes,No"',
        "Pattern": '"range,channel"',
        "Close Stopout": '"Yes,No"',
        "Near Perfect Entry": '"Yes,No"',
        "Near Win": '"Yes,No"',
        "Early Close": '"Yes,No"',
    }
    for header, formula in specs.items():
        col = headers.get(header)
        if not col:
            continue
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(col)
        dv.add(f"{letter}{TRADE_LOG_DATA_START_ROW}:{letter}{max_row}")


def _copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def _border_with_side(border: Border, **sides: Side) -> Border:
    return Border(
        left=sides.get("left", border.left),
        right=sides.get("right", border.right),
        top=sides.get("top", border.top),
        bottom=sides.get("bottom", border.bottom),
        diagonal=border.diagonal,
        diagonal_direction=border.diagonal_direction,
        diagonalUp=border.diagonalUp,
        diagonalDown=border.diagonalDown,
        outline=border.outline,
        vertical=border.vertical,
        horizontal=border.horizontal,
    )


def _copy_column_style_moving_right_boundary(ws, template_col: int, target_col: int) -> None:
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row, template_col)
        target = ws.cell(row, target_col)
        original_right = copy(source.border.right)
        internal_right = copy(source.border.left) if source.border.left and source.border.left.style else Side(style="thin", color="D1D5DB")
        _copy_cell_style(source, target)
        source.border = _border_with_side(source.border, right=internal_right)
        target.border = _border_with_side(target.border, right=original_right)


def _snapshot_cell(cell) -> Dict[str, Any]:
    return {
        "value": cell.value,
        "style": copy(cell._style),
        "comment": copy(cell.comment),
        "hyperlink": copy(cell.hyperlink),
    }


def _restore_cell_snapshot(cell, snapshot: Dict[str, Any], *, value: Any = None, use_snapshot_value: bool = True) -> None:
    cell.value = snapshot["value"] if use_snapshot_value else value
    cell._style = copy(snapshot["style"])
    cell.comment = copy(snapshot["comment"])
    cell.hyperlink = copy(snapshot["hyperlink"])


def _write_trade_log_three_row_headers(ws, header_templates: Dict[str, Dict[str, Any]]) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 3 and merged.max_row >= 1:
            ws.unmerge_cells(str(merged))
    row1, row2, row3 = _trade_log_three_row_header_values()
    for col, logical_header in enumerate(TRADE_LOG_HEADERS, start=1):
        template = header_templates.get(logical_header) or next(iter(header_templates.values()))
        _restore_cell_snapshot(ws.cell(1, col), template, value=row1[col - 1], use_snapshot_value=False)
        _restore_cell_snapshot(ws.cell(2, col), template, value=row2[col - 1], use_snapshot_value=False)
        _restore_cell_snapshot(ws.cell(3, col), template, value=row3[col - 1], use_snapshot_value=False)
    for col, logical_header in enumerate(TRADE_LOG_HEADERS, start=1):
        if logical_header not in MOVE_TO_FIELD_MAP:
            ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)
            anchor = ws.cell(1, col)
            anchor.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
                text_rotation=anchor.alignment.text_rotation,
                shrink_to_fit=anchor.alignment.shrink_to_fit,
                indent=anchor.alignment.indent,
            )
        else:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            anchor = ws.cell(2, col)
            anchor.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
                text_rotation=anchor.alignment.text_rotation,
                shrink_to_fit=anchor.alignment.shrink_to_fit,
                indent=anchor.alignment.indent,
            )
    for label, group_headers in (
        ("Move to Break-Even", MOVE_TO_BREAK_EVEN_HEADERS),
        ("Move to Profit", MOVE_TO_PROFIT_HEADERS),
    ):
        group_cols = [
            col for col, logical_header in enumerate(TRADE_LOG_HEADERS, start=1)
            if logical_header in group_headers
        ]
        if not group_cols:
            continue
        first_col, last_col = min(group_cols), max(group_cols)
        ws.merge_cells(start_row=1, start_column=first_col, end_row=1, end_column=last_col)
        ws.cell(1, first_col).value = label
        for col in group_cols:
            ws.cell(1, col).alignment = copy(ws.cell(2, col).alignment)
            ws.cell(1, col).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
                text_rotation=ws.cell(1, col).alignment.text_rotation,
                shrink_to_fit=ws.cell(1, col).alignment.shrink_to_fit,
                indent=ws.cell(1, col).alignment.indent,
            )
    _normalize_trade_log_header_fills(ws)


def _normalize_trade_log_header_fills(ws) -> None:
    if getattr(ws, "title", "") not in {TRADE_LOG_SHEET, LEGACY_ALL_TRADES_SHEET}:
        return
    max_col = max(ws.max_column, len(TRADE_LOG_HEADERS))
    empty_fill = PatternFill()
    for row in range(1, TRADE_LOG_HEADER_ROWS + 1):
        for col in range(1, max_col + 1):
            if _is_merged_non_anchor(ws, row, col):
                continue
            ws.cell(row, col).fill = empty_fill
    _restore_trade_log_move_group_header_right_borders(ws)


def _restore_trade_log_move_group_header_right_borders(ws) -> None:
    if getattr(ws, "title", "") not in {TRADE_LOG_SHEET, LEGACY_ALL_TRADES_SHEET}:
        return
    headers = _trade_log_header_map(ws)
    for header in (
        "Move to Break Even Distance From Exit %",
        "Move to Profit Distance From Exit %",
    ):
        col = headers.get(header)
        if not col:
            continue
        source_side = ws.cell(2, col).border.right
        if not _visible_border_side(source_side) or source_side.style != "medium":
            source_side = ws.cell(3, col).border.right
        right_side = copy(source_side) if _visible_border_side(source_side) else Side(style="medium")
        ws.cell(1, col).border = _border_with_side(ws.cell(1, col).border, right=right_side)
        for merged in ws.merged_cells.ranges:
            if merged.min_row <= 1 <= merged.max_row and merged.min_col <= col <= merged.max_col:
                anchor = ws.cell(merged.min_row, merged.min_col)
                anchor.border = _border_with_side(anchor.border, right=copy(right_side))
                break


def _write_trade_log_two_row_headers(ws, header_templates: Dict[str, Dict[str, Any]]) -> None:
    """Backward-compatible alias retained for external imports."""
    _write_trade_log_three_row_headers(ws, header_templates)


def _repair_trade_log_move_to_durations(ws, diagnostics: Dict[str, Any] | None = None) -> int:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    headers = _trade_log_header_map(ws)
    open_col = headers.get("Open Time")
    if not open_col:
        return 0
    repaired = 0
    for time_header, duration_header in (
        ("Move to Break Even Time", "Move to Break Even Duration"),
        ("Move to Profit Time", "Move to Profit Duration"),
    ):
        time_col = headers.get(time_header)
        duration_col = headers.get(duration_header)
        if not time_col or not duration_col:
            continue
        for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
            duration_cell = ws.cell(row, duration_col)
            if duration_cell.value not in (None, ""):
                continue
            open_time = _as_datetime(ws.cell(row, open_col).value)
            move_time = _as_datetime(ws.cell(row, time_col).value)
            if open_time is None or move_time is None:
                continue
            duration_cell.value = _fmt_duration_full(max(0.0, (move_time - open_time).total_seconds()))
            duration_cell.number_format = DURATION_NUMBER_FORMAT
            repaired += 1
    if repaired:
        diagnostics["repaired_trade_log_move_to_duration_cells"] = (
            int(diagnostics.get("repaired_trade_log_move_to_duration_cells") or 0) + repaired
        )
    return repaired


def _apply_trade_log_adaptive_formats(ws) -> None:
    headers = _trade_log_header_map(ws)
    profit_col = headers.get("Profit %")
    r_col = headers.get("R-Multiple")
    trade_duration_col = headers.get("Trade Duration (DD:HH:MM:SS)")
    move_duration_cols = [
        headers.get("Move to Break Even Duration"),
        headers.get("Move to Profit Duration"),
    ]
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        if profit_col:
            cell = ws.cell(row, profit_col)
            cell.number_format = adaptive_percent_number_format(cell.value)
        if r_col:
            cell = ws.cell(row, r_col)
            cell.number_format = adaptive_number_format(cell.value)
        if trade_duration_col:
            cell = ws.cell(row, trade_duration_col)
            if cell.value not in (None, ""):
                cell.value = _duration_display_cell_value(cell.value, cell.number_format)
            cell.number_format = "General"
        for col in move_duration_cols:
            if col and ws.cell(row, col).value not in (None, ""):
                ws.cell(row, col).number_format = DURATION_NUMBER_FORMAT


def _repair_legacy_duration_number_formats(wb: Workbook, diagnostics: Dict[str, Any] | None = None) -> int:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    repaired = 0
    registry_repaired = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fmt = str(cell.number_format or "")
                if not _is_legacy_duration_number_format(fmt):
                    continue
                cell.number_format = DURATION_NUMBER_FORMAT
                repaired += 1
    number_formats = getattr(wb, "_number_formats", None)
    if number_formats is not None:
        for index, fmt in enumerate(list(number_formats)):
            if not _is_legacy_duration_number_format(fmt):
                continue
            number_formats[index] = DURATION_NUMBER_FORMAT
            registry_repaired += 1
        if registry_repaired and hasattr(number_formats, "_dict"):
            number_formats._dict = {fmt: idx for idx, fmt in enumerate(number_formats)}
    if repaired:
        diagnostics["repaired_legacy_duration_number_formats"] = repaired
    if registry_repaired:
        diagnostics["repaired_legacy_duration_number_format_registry_entries"] = registry_repaired
    return repaired + registry_repaired


def _apply_trade_number_hyperlinks(
    ws,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    headers = _trade_log_header_map(ws)
    trade_number_col = headers.get(TRADE_NUMBER_HEADER)
    open_time_col = headers.get("Open Time")
    close_time_col = headers.get("Close Time")
    if not trade_number_col:
        return
    linked = 0
    unresolved: List[Dict[str, Any]] = []
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        cell = ws.cell(row, trade_number_col)
        number = str(cell.value or "").strip()
        if not number:
            cell.hyperlink = None
            continue
        link_diagnostics: Dict[str, Any] = {}
        target, reason = resolve_trade_folder_link(
            number,
            open_time=ws.cell(row, open_time_col).value if open_time_col else None,
            close_time=ws.cell(row, close_time_col).value if close_time_col else None,
            diagnostics=link_diagnostics,
        )
        if target:
            cell.hyperlink = target
            cell.number_format = "@"
            linked += 1
        else:
            unresolved.append({
                "trade_number": number,
                "row": row,
                "reason": reason,
                "checked_roots": list(link_diagnostics.get("checked_roots") or []),
                "preserved_existing_hyperlink": bool(cell.hyperlink),
            })
    diagnostics["trade_number_hyperlinks_added"] = linked
    if unresolved:
        diagnostics["trade_number_hyperlink_unresolved"] = unresolved


def _ensure_trade_log_schema(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    before_data_rows = _trade_log_data_row_count(ws)
    already_current = _trade_log_has_three_row_headers(ws)
    legacy_two_row_headers = _trade_log_has_two_row_headers(ws)
    legacy_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers(ws)
    legacy_v1_grouped_headers = _trade_log_has_two_row_headers_for(ws, TRADE_LOG_HEADERS_V1)
    legacy_v1_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, TRADE_LOG_HEADERS_V1)
    legacy_rec_three_row_headers = _trade_log_has_recommendation_grouped_three_row_headers(ws)
    legacy_rec_grouped_headers = _trade_log_has_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_rec_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_rec_v1_grouped_headers = _trade_log_has_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    legacy_rec_v1_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    legacy_pre_vwap_rec_three_row_headers = _trade_log_has_pre_vwap_recommendation_grouped_three_row_headers(ws)
    legacy_pre_vwap_rec_grouped_headers = _trade_log_has_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_pre_vwap_rec_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_pre_vwap_rec_v1_grouped_headers = _trade_log_has_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    legacy_pre_vwap_rec_v1_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    legacy_pre_rec_three_row_headers = _trade_log_has_pre_recommendation_grouped_three_row_headers(ws)
    legacy_pre_rec_grouped_headers = _trade_log_has_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_pre_rec_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
    legacy_pre_rec_v1_grouped_headers = _trade_log_has_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    legacy_pre_rec_v1_duplicate_headers = _trade_log_has_legacy_duplicate_two_row_headers_for(ws, PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
    if already_current:
        if ws.max_column > len(TRADE_LOG_HEADERS):
            extra_count = ws.max_column - len(TRADE_LOG_HEADERS)
            ws.delete_cols(len(TRADE_LOG_HEADERS) + 1, extra_count)
            diagnostics["deleted_extra_trade_log_columns"] = extra_count
        headers = _trade_log_header_map(ws)
        trade_number_col = headers.get(TRADE_NUMBER_HEADER)
        if trade_number_col:
            for row in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1):
                ws.cell(row, trade_number_col).number_format = "@"
        for header in ("Move to Break Even Duration", "Move to Profit Duration"):
            for row in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1):
                ws.cell(row, headers[header]).number_format = DURATION_NUMBER_FORMAT
        for header in (
            "Move to Break Even Distance From Entry %", "Move to Break Even Distance From Exit %",
            "Move to Profit Distance From Entry %", "Move to Profit Distance From Exit %",
        ):
            for row in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1):
                ws.cell(row, headers[header]).number_format = "0.00%"
        for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
            col = headers.get(header)
            if not col:
                continue
            for row in range(TRADE_LOG_DATA_START_ROW, ws.max_row + 1):
                _apply_recommendation_cell_style(ws.cell(row, col))
        ws.freeze_panes = "A4"
        _hide_trade_log_row_id(ws)
        _set_trade_log_auto_filter(ws)
        _apply_trade_log_dropdown_validations(ws)
        _repair_trade_log_move_to_durations(ws, diagnostics)
        _apply_trade_log_adaptive_formats(ws)
        _apply_trade_number_hyperlinks(ws, diagnostics)
        _apply_trade_log_win_loss_row_formatting(ws)
        _apply_trade_log_win_loss_direct_row_fills(ws)
        _normalize_trade_log_row_heights(ws, diagnostics)
        _normalize_trade_log_header_fills(ws)
        if _trade_log_data_row_count(ws) != before_data_rows:
            raise RuntimeError("Trade Log schema validation changed the data row count unexpectedly.")
        return
    else:
        if legacy_two_row_headers or legacy_duplicate_headers:
            source_headers = list(TRADE_LOG_HEADERS)
            source_start_row = 3
        elif legacy_rec_three_row_headers:
            source_headers = list(RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = TRADE_LOG_DATA_START_ROW
        elif legacy_rec_grouped_headers or legacy_rec_duplicate_headers:
            source_headers = list(RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = 3
        elif legacy_pre_vwap_rec_three_row_headers:
            source_headers = list(PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = TRADE_LOG_DATA_START_ROW
        elif legacy_pre_vwap_rec_grouped_headers or legacy_pre_vwap_rec_duplicate_headers:
            source_headers = list(PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = 3
        elif legacy_pre_rec_three_row_headers:
            source_headers = list(PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = TRADE_LOG_DATA_START_ROW
        elif legacy_pre_rec_grouped_headers or legacy_pre_rec_duplicate_headers:
            source_headers = list(PRE_RECOMMENDATION_TRADE_LOG_HEADERS)
            source_start_row = 3
        elif legacy_v1_grouped_headers or legacy_v1_duplicate_headers:
            source_headers = list(TRADE_LOG_HEADERS_V1)
            source_start_row = 3
        elif legacy_rec_v1_grouped_headers or legacy_rec_v1_duplicate_headers:
            source_headers = list(RECOMMENDATION_TRADE_LOG_HEADERS_V1)
            source_start_row = 3
        elif legacy_pre_vwap_rec_v1_grouped_headers or legacy_pre_vwap_rec_v1_duplicate_headers:
            source_headers = list(PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
            source_start_row = 3
        elif legacy_pre_rec_v1_grouped_headers or legacy_pre_rec_v1_duplicate_headers:
            source_headers = list(PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1)
            source_start_row = 3
        else:
            source_headers = [str(ws.cell(1, col).value or "").strip() for col in range(1, ws.max_column + 1)]
            while source_headers and not source_headers[-1]:
                source_headers.pop()
        if source_headers not in (
            PRE_MOVE_TRADE_LOG_HEADERS,
            OLD_TRADE_LOG_HEADERS,
            PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1,
            PRE_RECOMMENDATION_TRADE_LOG_HEADERS,
            PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1,
            PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS,
            RECOMMENDATION_TRADE_LOG_HEADERS_V1,
            RECOMMENDATION_TRADE_LOG_HEADERS,
            PRE_VWAP_TRADE_LOG_HEADERS_V1,
            PRE_VWAP_TRADE_LOG_HEADERS,
            TRADE_LOG_HEADERS_V1,
            TRADE_LOG_HEADERS,
        ):
            raise RuntimeError(
                "Trade Log headers cannot be migrated safely: "
                f"found {source_headers!r}; expected current two-row headers or one of "
                f"{[PRE_MOVE_TRADE_LOG_HEADERS, OLD_TRADE_LOG_HEADERS, PRE_RECOMMENDATION_TRADE_LOG_HEADERS_V1, PRE_RECOMMENDATION_TRADE_LOG_HEADERS, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS_V1, PRE_VWAP_RECOMMENDATION_TRADE_LOG_HEADERS, RECOMMENDATION_TRADE_LOG_HEADERS_V1, RECOMMENDATION_TRADE_LOG_HEADERS, PRE_VWAP_TRADE_LOG_HEADERS_V1, PRE_VWAP_TRADE_LOG_HEADERS, TRADE_LOG_HEADERS_V1, TRADE_LOG_HEADERS]!r}."
            )
        if not (
            legacy_two_row_headers or legacy_duplicate_headers
            or legacy_v1_grouped_headers or legacy_v1_duplicate_headers
            or legacy_rec_three_row_headers
            or legacy_rec_grouped_headers or legacy_rec_duplicate_headers
            or legacy_rec_v1_grouped_headers or legacy_rec_v1_duplicate_headers
            or legacy_pre_vwap_rec_three_row_headers
            or legacy_pre_vwap_rec_grouped_headers or legacy_pre_vwap_rec_duplicate_headers
            or legacy_pre_vwap_rec_v1_grouped_headers or legacy_pre_vwap_rec_v1_duplicate_headers
            or legacy_pre_rec_three_row_headers
            or legacy_pre_rec_grouped_headers or legacy_pre_rec_duplicate_headers
            or legacy_pre_rec_v1_grouped_headers or legacy_pre_rec_v1_duplicate_headers
        ):
            source_start_row = 2

    source_by_header = {header: idx for idx, header in enumerate(source_headers, start=1)}
    if len(source_by_header) != len(source_headers):
        raise RuntimeError(f"Trade Log headers cannot be migrated safely because duplicate logical headers were found: {source_headers!r}.")

    source_data_rows: List[Dict[str, Dict[str, Any]]] = []
    for row in range(source_start_row, ws.max_row + 1):
        if not any(ws.cell(row, col).value not in (None, "") for col in range(1, len(source_headers) + 1)):
            continue
        row_snapshot: Dict[str, Dict[str, Any]] = {}
        for header, col in source_by_header.items():
            row_snapshot[header] = _snapshot_cell(ws.cell(row, col))
        source_data_rows.append(row_snapshot)

    header_templates: Dict[str, Dict[str, Any]] = {}
    def _first_available_template(*headers: str) -> str:
        for candidate in headers:
            if candidate in source_by_header:
                return candidate
        return next(iter(source_by_header))

    source_header_row = 2 if (
        already_current
        or legacy_two_row_headers or legacy_duplicate_headers
        or legacy_v1_grouped_headers or legacy_v1_duplicate_headers
        or legacy_rec_three_row_headers
        or legacy_rec_grouped_headers or legacy_rec_duplicate_headers
        or legacy_rec_v1_grouped_headers or legacy_rec_v1_duplicate_headers
        or legacy_pre_vwap_rec_three_row_headers
        or legacy_pre_vwap_rec_grouped_headers or legacy_pre_vwap_rec_duplicate_headers
        or legacy_pre_vwap_rec_v1_grouped_headers or legacy_pre_vwap_rec_v1_duplicate_headers
        or legacy_pre_rec_three_row_headers
        or legacy_pre_rec_grouped_headers or legacy_pre_rec_duplicate_headers
        or legacy_pre_rec_v1_grouped_headers or legacy_pre_rec_v1_duplicate_headers
    ) else 1
    for header in TRADE_LOG_HEADERS:
        source_header = header
        if header == "Close Stopout" and source_header not in source_by_header:
            source_header = "Stop Out"
        template_header = source_header if source_header in source_by_header else _first_available_template("Open Time", "Test")
        if header == "VWAP" and source_header not in source_by_header:
            template_header = _first_available_template("Round Number", "EMA", "Order", "ATHS/ATLS", "Pattern", "Test", "Setup", "Open Time")
        if header == STOP_RECOMMENDATION_HEADER:
            template_header = "Stop Loss Distance"
        elif header == TARGET_RECOMMENDATION_HEADER:
            template_header = "Target Distance"
        elif "Trigger Price" in header:
            template_header = "Entry Price"
        elif "Distance From" in header:
            template_header = "Stop Loss Distance"
        elif header.endswith("Duration"):
            template_header = "Trade Duration (DD:HH:MM:SS)"
        if template_header not in source_by_header:
            template_header = _first_available_template("Open Time", "Test", "Setup")
        header_templates[header] = _snapshot_cell(ws.cell(source_header_row, source_by_header[template_header]))

    source_dimensions = {
        header: copy(ws.column_dimensions[get_column_letter(col)])
        for header, col in source_by_header.items()
    }
    max_old_row = ws.max_row
    max_old_col = ws.max_column
    for row in range(1, max(max_old_row, TRADE_LOG_DATA_START_ROW + len(source_data_rows) - 1) + 1):
        for col in range(1, max(max_old_col, len(TRADE_LOG_HEADERS)) + 1):
            if _is_merged_non_anchor(ws, row, col):
                continue
            cell = ws.cell(row, col)
            cell.value = None
            cell.comment = None
            cell.hyperlink = None
    _write_trade_log_three_row_headers(ws, header_templates)
    if ws.max_column > len(TRADE_LOG_HEADERS):
        extra_count = ws.max_column - len(TRADE_LOG_HEADERS)
        ws.delete_cols(len(TRADE_LOG_HEADERS) + 1, extra_count)
        diagnostics["deleted_extra_trade_log_columns"] = extra_count
    for target_col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        source_header = header if header in source_by_header else ("Stop Out" if header == "Close Stopout" and "Stop Out" in source_by_header else None)
        template_header = source_header or _first_available_template("Open Time", "Test")
        if header == "VWAP" and source_header is None:
            template_header = _first_available_template("Round Number", "EMA", "Order", "ATHS/ATLS", "Pattern", "Test", "Setup", "Open Time")
        if header == STOP_RECOMMENDATION_HEADER:
            template_header = "Stop Loss Distance"
        elif header == TARGET_RECOMMENDATION_HEADER:
            template_header = "Target Distance"
        elif "Trigger Price" in header:
            template_header = "Entry Price"
        elif "Distance From" in header:
            template_header = "Stop Loss Distance"
        elif header.endswith("Duration"):
            template_header = "Trade Duration (DD:HH:MM:SS)"
        if template_header not in source_by_header:
            template_header = _first_available_template("Open Time", "Test", "Setup")
        dimension = source_dimensions.get(source_header or "") or source_dimensions.get(template_header)
        letter = get_column_letter(target_col)
        ws.column_dimensions[letter].width = dimension.width if dimension and dimension.width else 14
        ws.column_dimensions[letter].hidden = bool(dimension.hidden) if dimension else False
        for offset, row_snapshot in enumerate(source_data_rows):
            target_row = TRADE_LOG_DATA_START_ROW + offset
            snapshot = row_snapshot.get(source_header) if source_header else None
            if snapshot:
                _restore_cell_snapshot(ws.cell(target_row, target_col), snapshot)
            else:
                template_snapshot = row_snapshot.get(template_header)
                if template_snapshot:
                    _restore_cell_snapshot(ws.cell(target_row, target_col), template_snapshot, value=None, use_snapshot_value=False)
        if header in ("Move to Break Even Duration", "Move to Profit Duration"):
            for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + len(source_data_rows)):
                ws.cell(row, target_col).number_format = DURATION_NUMBER_FORMAT
        elif header == TRADE_NUMBER_HEADER:
            for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + len(source_data_rows)):
                ws.cell(row, target_col).number_format = "@"
        elif header in {STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER}:
            for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + len(source_data_rows)):
                _apply_recommendation_cell_style(ws.cell(row, target_col))
        elif "Distance From" in header:
            for row in range(TRADE_LOG_DATA_START_ROW, TRADE_LOG_DATA_START_ROW + len(source_data_rows)):
                ws.cell(row, target_col).number_format = "0.00%"

    _normalize_trade_log_row_heights(ws, diagnostics)
    ws.row_dimensions[1].height = ws.row_dimensions[1].height or 24
    ws.row_dimensions[2].height = ws.row_dimensions[2].height or 24
    ws.row_dimensions[3].height = ws.row_dimensions[3].height or 24
    ws.freeze_panes = "A4"
    _hide_trade_log_row_id(ws)
    _set_trade_log_auto_filter(ws)
    _apply_trade_log_dropdown_validations(ws)
    _repair_trade_log_move_to_durations(ws, diagnostics)
    _apply_trade_log_adaptive_formats(ws)
    _apply_trade_number_hyperlinks(ws, diagnostics)
    _apply_trade_log_win_loss_row_formatting(ws)
    _apply_trade_log_win_loss_direct_row_fills(ws)
    _normalize_trade_log_header_fills(ws)

    after_data_rows = _trade_log_data_row_count(ws)
    if after_data_rows != before_data_rows:
        raise RuntimeError(
            "Trade Log schema migration aborted because data row count changed: "
            f"before={before_data_rows}, after={after_data_rows}."
        )
    if before_data_rows and not after_data_rows:
        raise RuntimeError("Trade Log schema migration aborted because it would blank the Trade Log.")
    if not already_current:
        diagnostics["migrated_trade_log_schema"] = True
        diagnostics["migrated_trade_log_from_headers"] = source_headers

def _conditional_formatting_formula_text(rule) -> str:
    formula = getattr(rule, "formula", None) or []
    if isinstance(formula, (list, tuple)):
        return " ".join(str(part or "") for part in formula)
    return str(formula or "")


def _is_generated_trade_log_win_loss_row_rule(rule, sqref: str) -> bool:
    """Identify only the row-level win/loss rules emitted by this module."""
    formulas = getattr(rule, "formula", None) or []
    if not isinstance(formulas, (list, tuple)) or len(formulas) != 1:
        return False
    formula_text = re.sub(r"\s+", "", str(formulas[0] or ""))
    match = re.fullmatch(
        r'AND\(\$([A-Z]{1,3})(\d+)="trade",\$([A-Z]{1,3})(\d+)([<>])0\)',
        formula_text,
        flags=re.IGNORECASE,
    )
    if not match or match.group(2) != match.group(4):
        return False
    try:
        sqref_parts = str(sqref or "").split()
        if len(sqref_parts) != 1:
            return False
        min_col, min_row, _max_col, _max_row = range_boundaries(sqref_parts[0])
    except ValueError:
        return False
    if min_col != 1 or min_row != int(match.group(2)):
        return False
    fill = getattr(getattr(rule, "dxf", None), "fill", None)
    fill_color = getattr(fill, "fgColor", None)
    fill_rgb = str(getattr(fill_color, "rgb", "") or "")[-6:].upper()
    expected_fill = PROFIT_FILL if match.group(5) == ">" else LOSS_FILL
    return (
        getattr(rule, "type", None) == "expression"
        and getattr(fill, "fill_type", None) == "solid"
        and fill_rgb == expected_fill
    )


def _remove_trade_log_win_loss_row_formatting(ws) -> Dict[str, int]:
    """Remove only generated Trade Log row-level win/loss CF rules."""
    cf = ws.conditional_formatting
    sanitized = OrderedDict()
    generated_priorities: Dict[str, int] = {}
    for key, rules in list(getattr(cf, "_cf_rules", {}).items()):
        sqref = str(getattr(key, "sqref", key))
        retained_rules = []
        for rule in rules:
            if not _is_generated_trade_log_win_loss_row_rule(rule, sqref):
                retained_rules.append(rule)
                continue
            formula_text = re.sub(
                r"\s+",
                "",
                str((getattr(rule, "formula", None) or [""])[0]),
            )
            direction = "profit" if ">0)" in formula_text else "loss"
            priority = int(getattr(rule, "priority", 0) or 0)
            if priority > 0:
                generated_priorities.setdefault(direction, priority)
        if retained_rules:
            sanitized[key] = retained_rules
    cf._cf_rules = sanitized
    return generated_priorities

def _is_generated_trade_log_value_fill_rule(rule) -> bool:
    formula_text = _conditional_formatting_formula_text(rule).strip()
    return (
        getattr(rule, "type", None) == "cellIs"
        and getattr(rule, "operator", None) in {"greaterThan", "lessThan", "notEqual"}
        and formula_text == "0"
    )

def _range_is_trade_log_generated_value_fill_range(range_ref: str) -> bool:
    try:
        min_col, min_row, max_col, _max_row = range_boundaries(range_ref)
    except ValueError:
        return False
    return min_row >= 2 and 13 <= min_col <= max_col <= 17

def _remove_trade_log_generated_value_fill_formatting(ws) -> None:
    cf = ws.conditional_formatting
    refs_to_remove = []
    for key, rules in list(getattr(cf, "_cf_rules", {}).items()):
        sqref = str(getattr(key, "sqref", key))
        sqref_parts = sqref.split()
        if (
            sqref_parts
            and all(_range_is_trade_log_generated_value_fill_range(part) for part in sqref_parts)
            and all(_is_generated_trade_log_value_fill_rule(rule) for rule in rules)
        ):
            refs_to_remove.append(sqref)
    for sqref in refs_to_remove:
        del cf[sqref]

def _cell_fill_rgb(cell) -> str:
    color = getattr(getattr(cell, "fill", None), "fgColor", None)
    rgb = str(getattr(color, "rgb", "") or "")
    return rgb[-6:].upper() if rgb else ""

def _cell_has_generated_trade_log_win_loss_fill(cell) -> bool:
    return getattr(cell.fill, "fill_type", None) == "solid" and _cell_fill_rgb(cell) in {PROFIT_FILL, LOSS_FILL}

def _apply_trade_log_win_loss_direct_row_fills(ws) -> None:
    headers = _trade_log_header_map(ws)
    row_type_col = headers.get("Row Type")
    net_pl_col = headers.get("Net P/L")
    if not row_type_col or not net_pl_col:
        return
    last_col = max((col for header, col in headers.items() if header), default=ws.max_column)
    profit_fill = PatternFill("solid", fgColor=PROFIT_FILL)
    loss_fill = PatternFill("solid", fgColor=LOSS_FILL)
    empty_fill = PatternFill()
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
        net_pl = _as_float(ws.cell(row, net_pl_col).value)
        fill = None
        if row_type == "trade" and net_pl is not None:
            if net_pl > 0:
                fill = profit_fill
            elif net_pl < 0:
                fill = loss_fill
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            if fill is not None:
                cell.fill = fill
            elif _cell_has_generated_trade_log_win_loss_fill(cell):
                cell.fill = empty_fill


def _apply_trade_log_win_loss_direct_row_fills_for_rows(
    ws,
    row_numbers: Collection[int],
) -> None:
    """Apply generated row fills only to explicitly changed Trade Log rows."""
    headers = _trade_log_header_map(ws)
    row_type_col = headers.get("Row Type")
    net_pl_col = headers.get("Net P/L")
    if not row_type_col or not net_pl_col:
        return
    last_col = max((col for header, col in headers.items() if header), default=ws.max_column)
    profit_fill = PatternFill("solid", fgColor=PROFIT_FILL)
    loss_fill = PatternFill("solid", fgColor=LOSS_FILL)
    empty_fill = PatternFill()
    for row in sorted({int(value) for value in row_numbers}):
        row_type = str(ws.cell(row, row_type_col).value or "").strip().lower()
        net_pl = _as_float(ws.cell(row, net_pl_col).value)
        fill = None
        if row_type == "trade" and net_pl is not None:
            if net_pl > 0:
                fill = profit_fill
            elif net_pl < 0:
                fill = loss_fill
        for col in range(1, last_col + 1):
            cell = ws.cell(row, col)
            if fill is not None:
                cell.fill = copy(fill)
            elif _cell_has_generated_trade_log_win_loss_fill(cell):
                cell.fill = copy(empty_fill)

def _apply_trade_log_recommendation_conditional_formatting(ws, start_row: int, last_row: int, headers: Dict[str, int]) -> None:
    yellow_fill = PatternFill("solid", fgColor="FFF2CC")
    yellow_font = Font(color="9C6500")
    keep_fill = PatternFill("solid", fgColor=PROFIT_FILL)
    keep_font = Font(color=PROFIT_FONT)
    for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
        col = headers.get(header)
        if not col:
            continue
        letter = get_column_letter(col)
        cell_ref = f"{letter}{start_row}"
        cell_range = f"{letter}{start_row}:{letter}{last_row}"
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'OR(LEFT(LOWER({cell_ref}),6)="reduce",LEFT(LOWER({cell_ref}),8)="increase")'],
                fill=yellow_fill,
                font=yellow_font,
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'LEFT(LOWER({cell_ref}),4)="keep"'],
                fill=keep_fill,
                font=keep_font,
                stopIfTrue=True,
            ),
        )

def _apply_trade_log_win_loss_row_formatting(ws) -> None:
    headers = _trade_log_header_map(ws)
    row_type_col = headers.get("Row Type")
    net_pl_col = headers.get("Net P/L")
    if not row_type_col or not net_pl_col:
        return
    generated_priorities = _remove_trade_log_win_loss_row_formatting(ws)
    _remove_trade_log_generated_value_fill_formatting(ws)
    last_col = max((col for header, col in headers.items() if header), default=ws.max_column)
    start_row = _trade_log_data_start_row(ws)
    last_row = max(start_row, _trade_log_last_populated_row(ws))
    row_type_letter = get_column_letter(row_type_col)
    net_pl_letter = get_column_letter(net_pl_col)
    cell_range = f"A{start_row}:{get_column_letter(last_col)}{last_row}"
    profit_fill = PatternFill("solid", fgColor=PROFIT_FILL)
    loss_fill = PatternFill("solid", fgColor=LOSS_FILL)
    profit_rule = FormulaRule(
        formula=[
            f'AND(${row_type_letter}{start_row}="trade",'
            f'${net_pl_letter}{start_row}>0)'
        ],
        fill=profit_fill,
        stopIfTrue=True,
    )
    loss_rule = FormulaRule(
        formula=[
            f'AND(${row_type_letter}{start_row}="trade",'
            f'${net_pl_letter}{start_row}<0)'
        ],
        fill=loss_fill,
        stopIfTrue=True,
    )
    if generated_priorities.get("profit"):
        profit_rule.priority = generated_priorities["profit"]
    if generated_priorities.get("loss"):
        loss_rule.priority = generated_priorities["loss"]
    ws.conditional_formatting.add(cell_range, profit_rule)
    ws.conditional_formatting.add(cell_range, loss_rule)

def _is_generated_profit_loss_rule(rule) -> bool:
    formula = getattr(rule, "formula", None) or []
    formula_text = " ".join(str(part or "") for part in formula)
    return (
        getattr(rule, "type", None) == "cellIs"
        and getattr(rule, "operator", None) in {"greaterThan", "lessThan"}
        and formula_text.strip() == "0"
    )

def _pnl_calendar_profit_loss_ranges(ws) -> List[str]:
    year_cols = _detect_calendar_year_columns(ws)
    month_rows = _detect_calendar_month_rows(ws)
    if year_cols and month_rows:
        first_col = min(year_cols.values())
        last_col = max(year_cols.values())
        return [
            f"{get_column_letter(first_col)}{row}:{get_column_letter(last_col)}{row}"
            for _month, row in sorted(month_rows.items())
        ]

    month_cols = _detect_calendar_month_columns(ws)
    if month_cols:
        first_col = min(month_cols.values())
        last_col = max(month_cols.values())
        ranges = []
        if first_col == 2:
            for row in range(2, ws.max_row + 1):
                year_value = _as_float(ws.cell(row, 1).value)
                if year_value is not None:
                    ranges.append(f"{get_column_letter(first_col)}{row}:{get_column_letter(last_col)}{row}")
            return ranges
        for row in range(2, ws.max_row + 1):
            label = str(ws.cell(row, 2).value or "").strip().lower()
            if label == "p/l %":
                ranges.append(f"{get_column_letter(first_col)}{row}:{get_column_letter(last_col)}{row}")
        return ranges

    month_names = {calendar.month_name[i].lower() for i in range(1, 13)}
    pnl_cols = []
    for col in range(2, ws.max_column + 1):
        header = str(ws.cell(1, col).value or "").strip().lower()
        subheader = str(ws.cell(2, col).value or "").strip().lower()
        if header.endswith(" p/l %") or subheader in month_names:
            pnl_cols.append(col)
    if not pnl_cols:
        return []
    first_col = min(pnl_cols)
    last_col = max(pnl_cols)
    ranges = []
    for row in range(3, ws.max_row + 1, 2):
        year_value = _as_float(ws.cell(row, 1).value)
        if year_value is not None:
            ranges.append(f"{get_column_letter(first_col)}{row}:{get_column_letter(last_col)}{row}")
    return ranges

def _remove_pnl_calendar_generated_profit_loss_formatting(ws) -> None:
    refs_to_remove = []
    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        sqref = str(getattr(key, "sqref", key))
        sqref_parts = sqref.split()
        if sqref_parts and all(_is_generated_profit_loss_rule(rule) for rule in rules):
            refs_to_remove.append(sqref)
    for sqref in refs_to_remove:
        del ws.conditional_formatting[sqref]

def _apply_pnl_calendar_profit_loss_formatting(ws) -> None:
    _remove_pnl_calendar_generated_profit_loss_formatting(ws)
    transposed_layout = bool(_detect_calendar_year_columns(ws) and _detect_calendar_month_rows(ws))
    month_cols = _detect_calendar_month_columns(ws)
    one_line_layout = bool(month_cols and min(month_cols.values()) == 2)
    for cell_range in _pnl_calendar_profit_loss_ranges(ws):
        if not (one_line_layout or transposed_layout):
            _profit_loss_rules(ws, cell_range)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row, col)
                value = _pnl_calendar_text_pct_fraction(cell.value)
                if value is None or value == 0:
                    _clear_cell_fill(cell)
                elif value > 0:
                    _apply_full_cell_semantic_fill(cell, "profit")
                else:
                    _apply_full_cell_semantic_fill(cell, "loss")

    # Total Trades rows are deliberately neutral, including when repairing stale direct fills.
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").strip().lower() == "total trades":
            for col in range(3, ws.max_column + 1):
                _clear_cell_fill(ws.cell(row, col))

def _apply_instrument_averages_semantic_fills(ws) -> None:
    headers = {header.lower(): col for header, col in _instrument_averages_header_map(ws).items()}
    profit_headers = {"wins", "long wins", "short wins"}
    loss_headers = {"losses", "long losses", "short losses"}
    signed_headers = {"net r multiple", "net p/l %", "avg p/l %"}
    symbol_col = headers.get("symbol", 1)
    for row in range(_instrument_averages_data_start_row(ws), ws.max_row + 1):
        if ws.cell(row, symbol_col).value in (None, ""):
            continue
        for header in profit_headers:
            col = headers.get(header)
            if col:
                _apply_full_cell_semantic_fill(ws.cell(row, col), "profit")
        for header in loss_headers:
            col = headers.get(header)
            if col:
                _apply_full_cell_semantic_fill(ws.cell(row, col), "loss")
        for header in signed_headers:
            col = headers.get(header)
            if col:
                _apply_sign_based_full_cell_fill(ws.cell(row, col))


def _apply_instrument_averages_requested_style(ws, *, preserve_layout: bool = False) -> None:
    headers = _instrument_averages_header_map(ws)
    header_row = _instrument_averages_header_row(ws)
    data_start_row = _instrument_averages_data_start_row(ws)
    symbol_col = headers.get("Symbol", 1)
    duration_headers = {
        "Shortest duration (DD:HH:MM:SS)", "Avg duration (DD:HH:MM:SS)",
        "Longest duration (DD:HH:MM:SS)", AVG_WINNING_DURATION_HEADER,
        AVG_LOSING_DURATION_HEADER,
    }
    count_headers = {
        "Trades", "Wins", "Losses", "Break-even", "Longs", "Long wins", "Long losses",
        "Long break-even", "Shorts", "Short wins", "Short losses", "Short break-even",
        "All-time highs", "All-time lows", "Market", "Limit", "Round number", "Spiked out",
        "Close stop out", "Near perfect entry", "Near win", "Early close",
        "Move to break even", "Move to profit",
    }
    if preserve_layout:
        for row in range(data_start_row, ws.max_row + 1):
            if ws.cell(row, symbol_col).value in (None, ""):
                continue
            for header in duration_headers:
                if headers.get(header):
                    cell = ws.cell(row, headers[header])
                    if cell.value not in (None, ""):
                        cell.value = _duration_display_cell_value(cell.value, cell.number_format)
                    cell.number_format = "General"
            for header in ("Move to break even", "Move to profit"):
                if headers.get(header):
                    ws.cell(row, headers[header]).number_format = ZERO_HIDE_FORMAT
            for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
                if headers.get(header):
                    _apply_recommendation_cell_style(ws.cell(row, headers[header]))
        _apply_symbols_filter_header_layout(ws)
        _ensure_symbols_duration_header_layout(ws)
        return
    for row in range(data_start_row, ws.max_row + 1):
        if ws.cell(row, symbol_col).value in (None, ""):
            continue
        symbol_value = ws.cell(row, symbol_col).value
        _copy_cell_style(ws.cell(header_row, symbol_col), ws.cell(row, symbol_col))
        ws.cell(row, symbol_col).value = symbol_value
        for col in range(2, ws.max_column + 1):
            alignment = copy(ws.cell(row, col).alignment)
            alignment.horizontal = "center"
            ws.cell(row, col).alignment = alignment
        for header in duration_headers:
            if headers.get(header):
                cell = ws.cell(row, headers[header])
                if cell.value not in (None, ""):
                    cell.value = _duration_display_cell_value(cell.value, cell.number_format)
                cell.number_format = "General"
        for header in count_headers:
            if headers.get(header):
                ws.cell(row, headers[header]).number_format = ZERO_HIDE_FORMAT
        for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
            if headers.get(header):
                _apply_recommendation_cell_style(ws.cell(row, headers[header]))
        if headers.get("Net R Multiple"):
            ws.cell(row, headers["Net R Multiple"]).number_format = '0.000"R"'
        for header in ("Move to break even", "Move to profit"):
            col = headers.get(header)
            if col:
                font = copy(ws.cell(row, col).font)
                font.color = "FF000000"
                ws.cell(row, col).font = font
    move_cols = [headers.get("Move to break even"), headers.get("Move to profit")]
    move_cols = [col for col in move_cols if col]
    if move_cols:
        cut_range = (
            f"{get_column_letter(min(move_cols))}1:"
            f"{get_column_letter(max(move_cols))}{max(data_start_row, ws.max_row)}"
        )
        cf = ws.conditional_formatting
        for key, rules in list(getattr(cf, "_cf_rules", {}).items()):
            sqref = str(getattr(key, "sqref", key))
            replacement_ranges: List[str] = []
            changed = False
            for part in sqref.split():
                remaining = _subtract_range_rectangle(part, cut_range)
                replacement_ranges.extend(remaining)
                changed = changed or remaining != [part]
            if not changed:
                continue
            del cf[sqref]
            for replacement in replacement_ranges:
                for rule in rules:
                    cf.add(replacement, copy(rule))
    if not preserve_layout:
        _write_instrument_averages_headers(ws)
    _apply_symbols_filter_header_layout(ws)


def _apply_instrument_averages_profit_loss_formatting(ws) -> None:
    headers = _instrument_averages_header_map(ws)
    required_headers = {"Net R Multiple", "Net P/L %", "Avg P/L %"}
    if not required_headers.issubset(headers):
        return
    start_row = _instrument_averages_data_start_row(ws)
    last_row = max(start_row, ws.max_row)
    target_ranges = [
        f"{get_column_letter(headers['Net R Multiple'])}{start_row}:"
        f"{get_column_letter(headers['Net R Multiple'])}{last_row}",
        f"{get_column_letter(headers['Net P/L %'])}{start_row}:"
        f"{get_column_letter(headers['Avg P/L %'])}{last_row}",
    ]
    cf = ws.conditional_formatting
    for key, rules in list(getattr(cf, "_cf_rules", {}).items()):
        if not rules or not all(_is_generated_profit_loss_rule(rule) for rule in rules):
            continue
        sqref = str(getattr(key, "sqref", key))
        replacement_ranges: List[str] = []
        changed = False
        for part in sqref.split():
            remaining = [part]
            for target in target_ranges:
                remaining = [
                    piece
                    for current in remaining
                    for piece in _subtract_range_rectangle(current, target)
                ]
            replacement_ranges.extend(remaining)
            changed = changed or remaining != [part]
        if not changed:
            continue
        del cf[sqref]
        for replacement in replacement_ranges:
            for rule in rules:
                cf.add(replacement, copy(rule))
    for target in target_ranges:
        _profit_loss_rules(ws, target)


def _normalize_symbols_performance_formatting(ws) -> None:
    headers = _instrument_averages_header_map(ws)
    data_start = _instrument_averages_data_start_row(ws)
    symbol_col = headers.get("Symbol", 1)
    last_row = max(
        [
            row for row in range(data_start, ws.max_row + 1)
            if ws.cell(row, symbol_col).value not in (None, "")
        ]
        or [data_start]
    )
    net_r_col = headers.get("Net R Multiple")
    signed_pct_cols = [headers.get("Net P/L %"), headers.get("Avg P/L %")]
    signed_pct_cols = [col for col in signed_pct_cols if col]
    percentage_cols = {
        col
        for col in range(1, ws.max_column + 1)
        if (
            (_symbols_metric_key_for_column(ws, col) or "").endswith("_pct")
            or "_pct_" in (_symbols_metric_key_for_column(ws, col) or "")
        )
    }
    win_rate_col = headers.get("Win Rate %")
    if win_rate_col:
        percentage_cols.add(win_rate_col)
    percentage_cols.update(signed_pct_cols)

    for row in range(data_start, last_row + 1):
        if ws.cell(row, symbol_col).value in (None, ""):
            continue
        if net_r_col:
            cell = ws.cell(row, net_r_col)
            cell.number_format = '0.000"R"'
            _apply_sign_based_full_cell_fill(cell)
        for col in percentage_cols:
            ws.cell(row, col).number_format = "0.00%"
        for col in signed_pct_cols:
            _apply_sign_based_full_cell_fill(ws.cell(row, col))

    _apply_instrument_averages_profit_loss_formatting(ws)


def _repair_instrument_timeframe_columns(ws) -> None:
    headers = _instrument_averages_header_map(ws)
    source_col = headers.get("Most traded timeframe")
    target_cols = [
        headers.get("Most Profitable Timeframe"),
        headers.get("Least Profitable Timeframe"),
    ]
    target_cols = [col for col in target_cols if col]
    if not source_col or not target_cols:
        return
    start_row = _instrument_averages_data_start_row(ws)
    last_row = max(start_row, ws.max_row)
    source_range = f"{get_column_letter(source_col)}{start_row}:{get_column_letter(source_col)}{last_row}"
    target_ranges = [
        f"{get_column_letter(col)}{start_row}:{get_column_letter(col)}{last_row}"
        for col in target_cols
    ]
    cf = ws.conditional_formatting
    source_rules: List[Any] = []
    for key, rules in list(getattr(cf, "_cf_rules", {}).items()):
        sqref = str(getattr(key, "sqref", key))
        if source_range in sqref.split():
            source_rules.extend(copy(rule) for rule in rules)
        replacement_ranges: List[str] = []
        changed = False
        for part in sqref.split():
            remaining = [part]
            for target in target_ranges:
                remaining = [
                    piece
                    for current in remaining
                    for piece in _subtract_range_rectangle(current, target)
                ]
            replacement_ranges.extend(remaining)
            changed = changed or remaining != [part]
        if changed:
            del cf[sqref]
            for replacement in replacement_ranges:
                for rule in rules:
                    cf.add(replacement, copy(rule))
    if source_rules:
        for target in target_ranges:
            for rule in source_rules:
                cf.add(target, copy(rule))
    for row in range(start_row, ws.max_row + 1):
        if not any(ws.cell(row, col).value not in (None, "") for col in (source_col, *target_cols)):
            continue
        source_cell = ws.cell(row, source_col)
        for col in target_cols:
            cell = ws.cell(row, col)
            cell.fill = copy(source_cell.fill)
            font = copy(source_cell.font)
            font.color = "FF000000"
            cell.font = font
            cell.number_format = source_cell.number_format


def _repair_symbols_header_merges_preserving_layout(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    header_row = _instrument_averages_header_row(ws)
    if header_row != INSTRUMENT_AVERAGES_FILTER_HEADER_ROW:
        return
    existing = {str(rng) for rng in ws.merged_cells.ranges}
    group_columns = {
        col
        for col in range(1, ws.max_column + 1)
        if _symbols_group_for_column(ws, col)
    }
    created = []
    for col in range(1, ws.max_column + 1):
        if col in group_columns:
            continue
        if any(rng.min_col == col and rng.max_col == col and rng.min_row <= 1 and rng.max_row >= 2 for rng in ws.merged_cells.ranges):
            continue
        label = str(ws.cell(1, col).value or "").strip() or str(ws.cell(2, col).value or "").strip()
        if not label:
            continue
        if ws.cell(1, col).value in (None, ""):
            if ws.cell(2, col).value not in (None, ""):
                _copy_cell_style(ws.cell(2, col), ws.cell(1, col))
            ws.cell(1, col).value = label
        ws.cell(2, col).value = None
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        created.append(f"{get_column_letter(col)}1:{get_column_letter(col)}2")
    if diagnostics is not None and created:
        diagnostics.setdefault("symbols_vertical_header_merges_added", []).extend(
            item for item in created if item not in existing
        )


def _populate_symbols_metrics_preserving_layout(
    ws,
    rows: List[Dict[str, Any]],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    headers = _instrument_averages_header_map(ws)
    symbol_col = headers.get("Symbol", 1)
    analysis = _instrument_analysis_by_symbol(rows)
    populated: List[str] = []
    skipped: List[Dict[str, str]] = []
    data_start = _instrument_averages_data_start_row(ws)
    metric_cols = {
        col: _symbols_metric_key_for_column(ws, col)
        for col in range(1, ws.max_column + 1)
    }
    metric_cols = {col: key for col, key in metric_cols.items() if key}
    if not metric_cols:
        return
    for row in range(data_start, ws.max_row + 1):
        symbol = str(ws.cell(row, symbol_col).value or "").strip().upper()
        if not symbol:
            continue
        payload = analysis.get(symbol)
        if not payload:
            for col, key in metric_cols.items():
                if key.endswith("_recommendation"):
                    cell = ws.cell(row, col)
                    cell.value = ""
                    cell.number_format = "General"
                    _apply_recommendation_cell_style(cell)
                skipped.append({"column": get_column_letter(col), "metric": key, "reason": "symbol_not_in_trade_log"})
            continue
        for col, key in metric_cols.items():
            cell = ws.cell(row, col)
            value = payload.get(key)
            if value is None:
                if key.endswith("_recommendation"):
                    cell.value = ""
                    cell.number_format = "General"
                    _apply_recommendation_cell_style(cell)
                skipped.append({"column": get_column_letter(col), "metric": key, "reason": "metric_not_derivable"})
                continue
            cell.value = value
            if key.startswith("timeframe_"):
                cell.number_format = ZERO_HIDE_FORMAT
            elif key.endswith("_recommendation"):
                cell.number_format = "General"
                _apply_recommendation_cell_style(cell)
            elif key in {"shortest_duration", "avg_duration", "longest_duration", "avg_winning_duration", "avg_losing_duration"}:
                cell.value = _format_duration_display(value)
                cell.number_format = "General"
            else:
                cell.value = value / 100.0
                cell.number_format = "0.00%"
            populated.append(f"{get_column_letter(col)}:{key}")
    if populated:
        diagnostics["symbols_populated_columns"] = sorted(set(populated))
    if skipped:
        diagnostics["symbols_skipped_columns"] = skipped[:200]


def _apply_dashboard_source_label_style(cell) -> None:
    font = copy(cell.font)
    font.bold = False
    font.italic = True
    cell.font = font
    alignment = copy(cell.alignment)
    alignment.horizontal = "right"
    cell.alignment = alignment


def _apply_dashboard_child_label_style(cell) -> None:
    font = copy(cell.font)
    font.bold = False
    font.italic = True
    font.color = "FF000000"
    cell.font = font
    alignment = copy(cell.alignment)
    alignment.horizontal = "right"
    cell.alignment = alignment


def _repair_stats1_child_label_styles(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}

    def has_subordinate_style(cell) -> bool:
        return (
            cell.font.bold is False
            and cell.font.italic is True
            and str(cell.alignment.horizontal or "").lower() == "right"
        )

    def has_core_metric_style(cell) -> bool:
        return (
            cell.font.bold is True
            and cell.font.italic is not True
            and str(cell.alignment.horizontal or "").lower() in {"", "left", "general"}
        )

    def copy_label_style(src, dst) -> None:
        value = dst.value
        border = copy(dst.border)
        _copy_cell_style(src, dst)
        dst.value = value
        dst.border = border

    def first_template(coordinates: Tuple[str, ...], *, source_label: bool):
        for coordinate in coordinates:
            cell = ws[coordinate]
            label = str(cell.value or "").strip().casefold()
            if not label:
                continue
            if source_label and label != "source":
                continue
            if not source_label and label in {"source", "winners", "losers"}:
                continue
            if _semantic_fill_rgb(cell):
                continue
            if not has_subordinate_style(cell):
                continue
            return cell
        return None

    core_templates: Dict[str, Any] = {}
    generic_core_template = None
    first_winners_bounds = _stats1_section_bounds(ws, "Winners")
    core_end = (first_winners_bounds[0] - 1) if first_winners_bounds else ws.max_row
    for row in range(1, core_end + 1):
        cell = ws.cell(row, 1)
        label = _stats1_label_at(ws, row)
        label_key = label.casefold()
        if not label or label_key in {"source", "winners", "losers"}:
            continue
        if _semantic_fill_rgb(cell):
            continue
        if not has_core_metric_style(cell):
            continue
        core_templates.setdefault(label_key, cell)
        if generic_core_template is None:
            generic_core_template = cell

    source_template = first_template(("A61",), source_label=True)
    if source_template is None:
        source_template = next(
            (
                ws.cell(row, 1)
                for row in range(1, ws.max_row + 1)
                if _stats1_label_at(ws, row).casefold() == "source"
                and has_subordinate_style(ws.cell(row, 1))
            ),
            None,
        )
    child_template = first_template(("A97", "A98"), source_label=False)
    if child_template is None:
        for section_name in ("Winners", "Losers"):
            bounds = _stats1_section_bounds(ws, section_name)
            if not bounds:
                continue
            child_template = next(
                (
                    ws.cell(row, 1)
                    for row in range(bounds[0] + 1, bounds[1] + 1)
                    if _stats1_label_at(ws, row)
                    and _stats1_label_at(ws, row).casefold() != "source"
                    and has_subordinate_style(ws.cell(row, 1))
                    and not _semantic_fill_rgb(ws.cell(row, 1))
                ),
                None,
            )
            if child_template is not None:
                break

    repaired_sources = 0
    repaired_metric_labels = 0
    for row in range(1, ws.max_row + 1):
        if _stats1_label_at(ws, row).casefold() == "source":
            if source_template is not None:
                copy_label_style(source_template, ws.cell(row, 1))
            else:
                _apply_dashboard_source_label_style(ws.cell(row, 1))
            repaired_sources += 1
    for section_name in ("Winners", "Losers"):
        bounds = _stats1_section_bounds(ws, section_name)
        if not bounds:
            continue
        for row in range(bounds[0] + 1, bounds[1] + 1):
            label = _stats1_label_at(ws, row)
            if not label or label.casefold() == "source":
                continue
            if _semantic_fill_rgb(ws.cell(row, 1)) == LIGHT_GREY_FILL_RGB[-6:]:
                continue
            template = core_templates.get(label.casefold()) or generic_core_template
            if template is not None:
                copy_label_style(template, ws.cell(row, 1))
            else:
                font = copy(ws.cell(row, 1).font)
                font.bold = True
                font.italic = False
                font.color = "FF000000"
                ws.cell(row, 1).font = font
                alignment = copy(ws.cell(row, 1).alignment)
                alignment.horizontal = "left"
                ws.cell(row, 1).alignment = alignment
            repaired_metric_labels += 1
    categorical_sections = {"side", "patterns", "timeframe", "commission", "drawdown"}
    repaired_nested_outcomes = 0
    for section_name in ("Side", "Patterns", "Timeframe", "Commission"):
        bounds = _stats1_section_bounds(ws, section_name, categorical_sections)
        if not bounds:
            continue
        for row in range(bounds[0] + 1, bounds[1] + 1):
            label = _stats1_label_at(ws, row).casefold()
            if label not in {"winners", "losers"}:
                continue
            if child_template is not None:
                copy_label_style(child_template, ws.cell(row, 1))
            else:
                _apply_dashboard_child_label_style(ws.cell(row, 1))
            repaired_nested_outcomes += 1
    if repaired_sources:
        diagnostics["repaired_stats1_source_label_styles"] = repaired_sources
    if repaired_metric_labels:
        diagnostics["repaired_stats1_metric_label_styles"] = repaired_metric_labels
    if repaired_nested_outcomes:
        diagnostics["repaired_stats1_nested_outcome_label_styles"] = repaired_nested_outcomes

def _is_generated_dashboard_semantic_rule(rule) -> bool:
    if getattr(rule, "type", None) != "cellIs":
        return False
    if getattr(rule, "operator", None) not in {"greaterThan", "lessThan", "notEqual"}:
        return False
    if [str(value) for value in (getattr(rule, "formula", None) or [])] != ["0"]:
        return False
    dxf = getattr(rule, "dxf", None)
    fill = getattr(getattr(dxf, "fill", None), "fgColor", None)
    font = getattr(getattr(dxf, "font", None), "color", None)
    fill_rgb = str(getattr(fill, "rgb", "") or "")[-6:].upper()
    font_rgb = str(getattr(font, "rgb", "") or "")[-6:].upper()
    return (fill_rgb, font_rgb) in {(PROFIT_FILL, PROFIT_FONT), (LOSS_FILL, LOSS_FONT)}

def _subtract_range_rectangle(cell_range: str, cut_range: str) -> List[str]:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cut_min_col, cut_min_row, cut_max_col, cut_max_row = range_boundaries(cut_range)
    inter_min_col = max(min_col, cut_min_col)
    inter_min_row = max(min_row, cut_min_row)
    inter_max_col = min(max_col, cut_max_col)
    inter_max_row = min(max_row, cut_max_row)
    if inter_min_col > inter_max_col or inter_min_row > inter_max_row:
        return [cell_range]
    rectangles = []
    if min_row < inter_min_row:
        rectangles.append((min_col, min_row, max_col, inter_min_row - 1))
    if inter_max_row < max_row:
        rectangles.append((min_col, inter_max_row + 1, max_col, max_row))
    if min_col < inter_min_col:
        rectangles.append((min_col, inter_min_row, inter_min_col - 1, inter_max_row))
    if inter_max_col < max_col:
        rectangles.append((inter_max_col + 1, inter_min_row, max_col, inter_max_row))
    return [
        f"{get_column_letter(left)}{top}:{get_column_letter(right)}{bottom}"
        for left, top, right, bottom in rectangles
    ]

def _sanitize_dashboard_semantic_conditional_formatting(ws) -> None:
    protected_ranges: List[str] = ["B3:D4", "C21:D28"]
    market_cols = _stats1_market_columns(ws)
    managed_rows = _stats1_semantic_row_policies(ws) if market_cols else {}
    if managed_rows:
        protected_ranges.append(
            f"{get_column_letter(min(market_cols.values()))}{min(managed_rows)}:"
            f"{get_column_letter(max(market_cols.values()))}{max(managed_rows)}"
        )
    protected_labels = {
        "wins",
        "losses",
        "side",
        "patterns",
        "timeframe",
        "commission",
        "gross percent gain",
        "gross ir gain",
        "gross percent loss",
        "gross ir loss",
        "max win %",
        "avg win %",
        "min win %",
        "max r win",
        "avg r win",
        "min r win",
        "max loss %",
        "avg loss %",
        "min loss %",
        "max r loss",
        "avg r loss",
        "min r loss",
        "best win streak",
        "winning streak",
        "worst losing streak",
        "losing streak",
        "most profitable",
        "least profitable",
    }
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip().casefold()
        if label in protected_labels:
            protected_ranges.append(f"B{row}:D{row}")
    winners_start = next(
        (row for row in range(1, ws.max_row + 1) if str(ws.cell(row, 1).value or "").strip().casefold() == "winners"),
        None,
    )
    if winners_start:
        next_section = next(
            (
                row
                for row in range(winners_start + 1, ws.max_row + 1)
                if str(ws.cell(row, 1).value or "").strip().casefold() in {"side", "patterns", "timeframe", "commission"}
            ),
            ws.max_row + 1,
        )
        if next_section > winners_start:
            protected_ranges.append(f"B{winners_start}:D{next_section - 1}")
    sanitized = OrderedDict()

    def add_rules(key, rules) -> None:
        if not rules:
            return
        if key in sanitized:
            sanitized[key].extend(rules)
        else:
            sanitized[key] = list(rules)

    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        generated = [rule for rule in rules if _is_generated_dashboard_semantic_rule(rule)]
        manual = [rule for rule in rules if not _is_generated_dashboard_semantic_rule(rule)]
        if manual:
            add_rules(copy(key), manual)
        if not generated:
            continue
        remaining_parts: List[str] = []
        for part in str(key.sqref).split():
            pieces = [part]
            for protected in protected_ranges:
                pieces = [piece for current in pieces for piece in _subtract_range_rectangle(current, protected)]
            remaining_parts.extend(pieces)
        if remaining_parts:
            generated_key = copy(key)
            generated_key.sqref = " ".join(remaining_parts)
            add_rules(generated_key, generated)
    ws.conditional_formatting._cf_rules = sanitized

def _apply_dashboard_requested_semantic_fills(ws) -> None:
    _sanitize_dashboard_semantic_conditional_formatting(ws)
    for coordinate in ("B3", "C3", "D3"):
        _apply_full_cell_semantic_fill(ws[coordinate], "profit")
    for coordinate in ("B4", "C4", "D4"):
        _apply_full_cell_semantic_fill(ws[coordinate], "loss")

    semantic_by_label = {
        "best win streak": "profit",
        "winning streak": "profit",
        "worst losing streak": "loss",
        "losing streak": "loss",
        "gross percent gain": "profit",
        "gross ir gain": "profit",
        "gross percent loss": "loss",
        "gross ir loss": "loss",
        "max win %": "profit",
        "avg win %": "profit",
        "min win %": "profit",
        "max r win": "profit",
        "avg r win": "profit",
        "min r win": "profit",
        "max loss %": "loss",
        "avg loss %": "loss",
        "min loss %": "loss",
        "max r loss": "loss",
        "avg r loss": "loss",
        "min r loss": "loss",
        "max r win": "profit",
    }
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip().lower()
        semantic = semantic_by_label.get(label)
        if semantic:
            for col in (2, 3, 4):
                _apply_full_cell_semantic_fill(ws.cell(row, col), semantic)

    for section in ("Side", "Patterns", "Timeframe"):
        bounds = _stats1_section_bounds(ws, section)
        if not bounds:
            continue
        start, end = bounds
        for row in range(start + 1, end + 1):
            label = str(ws.cell(row, 1).value or "").strip().casefold()
            if label == "winners":
                for col in (2, 3, 4):
                    _apply_full_cell_semantic_fill(ws.cell(row, col), "profit")
            elif label in {"losers", "losses"}:
                for col in (2, 3, 4):
                    _apply_full_cell_semantic_fill(ws.cell(row, col), "loss")

    _remove_orphan_dashboard_semantic_fill_cells(ws)


def _remove_orphan_dashboard_semantic_fill_cells(ws) -> List[str]:
    """Drop style-only cells outside the dashboard's semantic market columns.

    An earlier layout path created a fixed green/red block in otherwise absent
    columns.  Treat a column as absent only when it has no populated cells and
    is not part of a merge, so future nonblank dashboard columns are retained.
    """
    market_cols = _stats1_market_columns(ws)
    if not market_cols:
        return []
    last_market_col = max(market_cols.values())
    populated_columns = {
        col
        for (_row, col), cell in ws._cells.items()
        if cell.value not in (None, "")
    }
    merged_columns = {
        col
        for merged in ws.merged_cells.ranges
        for col in range(merged.min_col, merged.max_col + 1)
    }
    removed: List[str] = []
    for coordinate, cell in list(ws._cells.items()):
        row, col = coordinate
        if (
            col <= last_market_col
            or col in populated_columns
            or col in merged_columns
            or cell.value not in (None, "")
            or _semantic_fill_rgb(cell) not in {PROFIT_FILL, LOSS_FILL}
        ):
            continue
        removed.append(f"{get_column_letter(col)}{row}")
        del ws._cells[coordinate]
    return removed


def _stats1_market_columns(ws) -> Dict[str, int]:
    for row in range(1, min(6, ws.max_row) + 1):
        tokens = {
            str(ws.cell(row, col).value or "").strip().lower(): col
            for col in range(1, min(10, ws.max_column) + 1)
        }
        cols = {
            "overall": tokens.get("overall"),
            "fx": tokens.get("fx") or tokens.get("forex"),
            "crypto": tokens.get("crypto"),
        }
        if all(cols.values()):
            return {key: int(value) for key, value in cols.items() if value}
    return {}


def _stats1_section_bounds(
    ws,
    label: str,
    stop_labels: set[str] | None = None,
) -> Tuple[int, int] | None:
    section_labels = stop_labels or {
        "winners", "losers", "side", "patterns", "timeframe", "commission", "drawdown"
    }
    start = next(
        (row for row in range(1, ws.max_row + 1)
         if str(ws.cell(row, 1).value or "").strip().casefold() == label.casefold()),
        None,
    )
    if start is None:
        return None
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip().casefold() in section_labels:
            end = row - 1
            break
    return start, end


def _stats1_semantic_row_policies(
    ws,
) -> Dict[int, Tuple[str, str | None, str]]:
    """Map STATS1 rows to (kind, fill semantic, semantic section)."""
    categorical_sections = {"side", "patterns", "timeframe", "commission", "drawdown"}
    bounds: Dict[str, Tuple[int, int]] = {}
    for section_name in ("Winners", "Losers"):
        section_bounds = _stats1_section_bounds(ws, section_name)
        if section_bounds:
            bounds[section_name.casefold()] = section_bounds
    for section_name in ("Side", "Patterns", "Timeframe", "Commission", "Drawdown"):
        section_bounds = _stats1_section_bounds(
            ws,
            section_name,
            categorical_sections,
        )
        if section_bounds:
            bounds[section_name.casefold()] = section_bounds

    def section_for_row(row: int) -> str:
        for section_name, (start, end) in bounds.items():
            if start <= row <= end:
                return section_name
        return "core"

    policies: Dict[int, Tuple[str, str | None, str]] = {}
    for row in range(2, ws.max_row + 1):
        label = _stats1_label_at(ws, row)
        if not label:
            continue
        label_key = " ".join(label.strip().casefold().split())
        section = section_for_row(row)
        section_start = bounds.get(section, (0, 0))[0]
        if row == section_start and section != "core":
            policies[row] = ("section", None, section)
            continue

        kind = "text"
        semantic: str | None = None
        if section == "commission":
            kind = "commission"
            semantic = "loss"
        elif section == "drawdown":
            if label_key in {"start", "end"}:
                kind = "drawdown_date"
            elif "drawdown" in label_key:
                kind = "pct"
                semantic = "loss"
        elif section in {"side", "timeframe"}:
            kind = "count"
            if label_key in {"winners", "winner"}:
                semantic = "profit"
            elif label_key in {"losers", "loser", "losses"}:
                semantic = "loss"
        elif section == "patterns":
            if label_key in {
                "most traded", "least traded", "most profitable", "least profitable"
            }:
                kind = "text"
            else:
                kind = "count"
                if label_key in {"winners", "winner"}:
                    semantic = "profit"
                elif label_key in {"losers", "loser", "losses"}:
                    semantic = "loss"
        elif section in {"winners", "losers"}:
            if "duration" in label_key or "move to" in label_key:
                kind = "duration"
            elif label_key in {"most wins", "most losses"}:
                kind = "text"
            elif (
                label_key.startswith(("min r", "avg r", "max r", "r expectancy"))
                or " r " in f" {label_key} "
            ):
                kind = "r"
            elif "%" in label_key or "percent" in label_key or "expectancy" in label_key:
                kind = "pct"
            if section == "winners" and (
                "win %" in label_key
                or "r win" in label_key
                or label_key in {"percentage expectancy", "r expectancy", "min result %", "max result %", "min r", "max r"}
            ):
                semantic = "profit"
            elif section == "losers" and (
                "loss %" in label_key
                or "r loss" in label_key
                or label_key in {"percentage expectancy", "r expectancy", "min result %", "max result %", "min r", "max r"}
            ):
                semantic = "loss"
        else:
            if label_key in {
                "trades", "wins", "losses", "break-even", "break even", "test",
                "best win streak", "winning streak", "worst losing streak", "losing streak",
            }:
                kind = "count"
            elif label_key in {"start", "end"}:
                kind = "streak_date"
            elif "duration" in label_key or "move to" in label_key:
                kind = "duration"
            elif (
                "r multiple" in label_key
                or "gross ir" in label_key
                or label_key in {"r expectancy", "avg r", "average r"}
            ):
                kind = "r"
            elif (
                "%" in label_key
                or "percent" in label_key
                or "win rate" in label_key
                or label_key in {"percentage expectancy", "avg result", "average result"}
            ):
                kind = "pct"

            if label_key == "wins" or label_key in {"best win streak", "winning streak"}:
                semantic = "profit"
            elif label_key == "losses" or label_key in {"worst losing streak", "losing streak"}:
                semantic = "loss"
            elif label_key in {
                "gross percent gain", "gross ir gain", "max win %", "avg win %",
                "min win %", "max r win", "avg r win", "min r win",
            }:
                semantic = "profit"
            elif label_key in {
                "gross percent loss", "gross ir loss", "max loss %", "avg loss %",
                "min loss %", "max r loss", "avg r loss", "min r loss",
            }:
                semantic = "loss"
            elif label_key in {
                "net p/l", "net p/l percentage", "net p/l %", "net p/l r multiples",
                "percentage expectancy", "r expectancy", "avg result %", "avg r",
            }:
                semantic = "auto"

        policies[row] = (kind, semantic, section)
    return policies


def _apply_stats1_semantic_formatting(
    ws,
    extended_metrics: Dict[str, Dict[str, Any]] | None = None,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols = _stats1_market_columns(ws)
    if not market_cols:
        return
    policies = _stats1_semantic_row_policies(ws)
    commission_keys = {
        "min": "min_commission",
        "min commission": "min_commission",
        "avg": "avg_commission",
        "average": "avg_commission",
        "avg commission": "avg_commission",
        "max": "max_commission",
        "max commission": "max_commission",
        "total": "total_commission",
        "total commission": "total_commission",
    }
    repaired = 0
    for row, (kind, semantic, section) in policies.items():
        label_cell = ws.cell(row, 1)
        # STATS1 labels are user-owned except for the known corrupted fills.
        # Setting PatternFill() leaves every font/alignment/indent/border intact.
        label_cell.fill = PatternFill()
        label_key = " ".join(str(label_cell.value or "").strip().casefold().split())

        commission_key = commission_keys.get(label_key) if section == "commission" else None
        if commission_key:
            overall = ws.cell(row, market_cols["overall"])
            overall.value = None
            _apply_semantic_metric_policy(overall, kind="text")
            repaired += 1
            for market, currency in (("fx", "AUD"), ("crypto", "USDT")):
                cell = ws.cell(row, market_cols[market])
                value = ((extended_metrics or {}).get(market) or {}).get(commission_key)
                if value is not None:
                    cell.value = value
                elif cell.value in (None, ""):
                    diagnostics.setdefault("missing_stats1_commission_values", []).append(
                        f"{market} {commission_key}"
                    )
                _apply_semantic_metric_policy(
                    cell,
                    kind="commission",
                    semantic="loss",
                    currency=currency,
                )
                repaired += 1
            continue

        for market, col in market_cols.items():
            cell = ws.cell(row, col)
            currency = None
            if kind == "commission" and market in {"fx", "crypto"}:
                currency = "AUD" if market == "fx" else "USDT"
            cell_semantic = semantic
            if kind == "commission" and market == "overall":
                cell_semantic = None
            _apply_semantic_metric_policy(
                cell,
                kind=kind,
                semantic=cell_semantic,
                currency=currency,
            )
            repaired += 1
    if repaired:
        diagnostics["repaired_stats1_semantic_metric_cells"] = repaired


def _repair_stats1_formatting(
    ws,
    extended_metrics: Dict[str, Dict[str, Any]] | None = None,
    diagnostics: Dict[str, Any] | None = None,
    *,
    preserve_user_presentation: bool = False,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols = _stats1_market_columns(ws)
    if not market_cols:
        return
    if not preserve_user_presentation and ws.freeze_panes != "B2":
        diagnostics["previous_stats1_freeze_panes"] = str(ws.freeze_panes or "")
        ws.freeze_panes = "B2"
        diagnostics["repaired_stats1_freeze_panes"] = "B2"
    _write_streak_detail_rows(ws, "Best Win Streak", {}, diagnostics=diagnostics)
    _write_streak_detail_rows(ws, "Worst Losing Streak", {}, diagnostics=diagnostics)
    if not preserve_user_presentation:
        # A newly generated workbook has no user-authored label typography to
        # retain, so establish its template defaults before the role-specific
        # child-label pass below.  Preservation mode never enters this branch.
        for row in range(2, ws.max_row + 1):
            label_cell = ws.cell(row, 1)
            if label_cell.value in (None, ""):
                continue
            font = copy(label_cell.font)
            font.bold = True
            font.italic = False
            font.color = "FF000000"
            label_cell.font = font
            alignment = copy(label_cell.alignment)
            alignment.horizontal = "left"
            label_cell.alignment = alignment
    for row in range(2, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip().casefold()
        if "duration" in label or label.startswith(
            ("min move to ", "average move to ", "max move to ")
        ):
            for col in market_cols.values():
                cell = ws.cell(row, col)
                if cell.value not in (None, ""):
                    cell.value = _duration_display_cell_value(
                        cell.value,
                        cell.number_format,
                    )
    _apply_stats1_semantic_formatting(ws, extended_metrics, diagnostics)
    if not preserve_user_presentation:
        for row in _stats1_semantic_row_policies(ws):
            for col in market_cols.values():
                alignment = copy(ws.cell(row, col).alignment)
                alignment.horizontal = "left"
                ws.cell(row, col).alignment = alignment
        _apply_stats1_clean_table_borders(ws, diagnostics)
        _repair_stats1_child_label_styles(ws, diagnostics)

def read_master_journal_manual_overrides(path: Path) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    if not path.exists():
        return out
    wb=load_workbook(path, data_only=True)
    try:
        try:
            ws=_get_all_trades_sheet(wb)
        except RuntimeError:
            return out
        header_map = _trade_log_header_map(ws)
        idx = {header: col - 1 for header, col in header_map.items()}
        data_start_row = _trade_log_data_start_row(ws)
        rid_by_row={}
        if '_Trade Meta' in wb.sheetnames:
            meta=wb['_Trade Meta']
            rid_by_row={int(r[0]):str(r[1] or '').strip() for r in meta.iter_rows(min_row=2,values_only=True) if r and r[0] and r[1]}
        for row_num,r in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            comment_rid = ""
            cmt = ws.cell(row_num, 1).comment
            if cmt and isinstance(cmt.text, str) and cmt.text.startswith("row_id:"):
                comment_rid = cmt.text.split("row_id:", 1)[1].strip()
            meta_rid = rid_by_row.get(row_num,'')
            rowid_i = idx.get('Row ID')
            inline_rid = str(r[rowid_i] or '').strip() if rowid_i is not None and rowid_i < len(r) else ''
            rid = comment_rid or inline_rid or meta_rid
            if not rid:
                row_map = {h: (r[i] if i < len(r) else None) for h, i in idx.items()}
                rid = _all_trades_row_fingerprint_from_map(row_map)
            edits={}
            test_i=idx.get('Test')
            if test_i is not None:
                t=str(r[test_i] or '').strip().lower()
                edits['is_test_trade']=t in {'yes','true','1'}
            for col,field in [('Setup','setup'),('Timeframe','timeframe'),('Breakeven','breakeven'),('Notes','notes'), *TRADE_LOG_MANUAL_FIELD_MAP.items()]:
                i=idx.get(col)
                if i is None:
                    continue
                raw_value = r[i]
                if field in {"move_to_break_even_duration", "move_to_profit_duration"} and raw_value not in (None, ""):
                    number_format = str(ws.cell(row_num, i + 1).number_format or "")
                    parsed_duration = _duration_ddhhmmss_cell_to_seconds(raw_value) if _is_ddhhmmss_number_format(number_format) else _parse_duration_text(raw_value)
                    edits[field] = parsed_duration if parsed_duration is not None else raw_value
                elif field in MOVE_TO_TIME_FIELDS:
                    edits[field] = _excel_datetime_to_iso(raw_value)
                elif field in MOVE_TO_FIELD_MAP.values():
                    edits[field] = raw_value
                else:
                    edits[field] = '' if raw_value is None else str(raw_value)
            if 'close_stopout' not in edits and 'Stop Out' in idx:
                i = idx['Stop Out']
                edits['close_stopout'] = '' if r[i] is None else str(r[i])
            out[rid]=edits
    finally:
        wb.close()
    return out


def _trade_row_market(row: Dict[str, Any]) -> str | None:
    asset_class = str(row.get("asset_class") or row.get("class") or "").strip().lower()
    if asset_class in {"fx", "forex"}:
        return "fx"
    if asset_class == "crypto":
        return "crypto"
    account = str(row.get("account_label") or row.get("account") or "").strip().lower()
    if any(token in account for token in ("bybit", "binance", "coinspot", "crypto")):
        return "crypto"
    if any(token in account for token in ("oanda", "pepperstone", "forex", "fx")):
        return "fx"
    if _is_likely_fx_pair(str(row.get("symbol") or "")):
        return "fx"
    return None


def _move_duration_seconds(row: Dict[str, Any], prefix: str) -> float | None:
    duration = _parse_duration_text(row.get(f"{prefix}_duration"))
    if duration is not None and duration >= 0:
        return duration
    move_time = _as_datetime(row.get(f"{prefix}_time"))
    open_time = _as_datetime(row.get("open_time"))
    if move_time is None or open_time is None:
        return None
    seconds = (move_time - open_time).total_seconds()
    return seconds if seconds >= 0 else None


def _trade_move_duration_metrics(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    values: Dict[str, Dict[str, List[Tuple[float, Dict[str, Any]]]]] = {
        market: {"move_to_break_even_duration_seconds": [], "move_to_profit_duration_seconds": []}
        for market in ("overall", "fx", "crypto")
    }
    for row in rows:
        if str(row.get("row_type") or "trade") != "trade" or _target_r_row_is_test_trade(row):
            continue
        markets = ["overall"]
        market = _trade_row_market(row)
        if market:
            markets.append(market)
        for prefix, key in (
            ("move_to_break_even", "move_to_break_even_duration_seconds"),
            ("move_to_profit", "move_to_profit_duration_seconds"),
        ):
            seconds = _move_duration_seconds(row, prefix)
            if seconds is None:
                continue
            for market_name in markets:
                values[market_name][key].append((seconds, row))
    output: Dict[str, Dict[str, Any]] = {}
    for market, metrics in values.items():
        bucket: Dict[str, Any] = {"metric_sources": {}}
        for key, samples in metrics.items():
            values_only = [value for value, _row in samples]
            bucket[key] = (sum(values_only) / len(values_only)) if values_only else None
            if not samples:
                continue
            prefix = key.removesuffix("_duration_seconds")
            min_value, min_row = _pick_deterministic_extreme(samples, "min")
            max_value, max_row = _pick_deterministic_extreme(samples, "max")
            bucket[f"min_{key}"] = min_value
            bucket[f"avg_{key}"] = bucket[key]
            bucket[f"max_{key}"] = max_value
            bucket["metric_sources"][f"min_{key}"] = _trade_metric_ref(min_row, f"min_{prefix}", min_value)
            bucket["metric_sources"][f"max_{key}"] = _trade_metric_ref(max_row, f"max_{prefix}", max_value)
        if not bucket["metric_sources"]:
            bucket.pop("metric_sources", None)
        output[market] = bucket
    return output


def _mode_nonblank(values: List[Any]) -> Any:
    cleaned = [str(value).strip() for value in values if value not in (None, "") and str(value).strip()]
    if not cleaned:
        return ""
    counts = defaultdict(int)
    first_seen: Dict[str, int] = {}
    display: Dict[str, str] = {}
    for index, value in enumerate(cleaned):
        key = value.casefold()
        counts[key] += 1
        first_seen.setdefault(key, index)
        display.setdefault(key, value)
    winner = min(counts, key=lambda key: (-counts[key], first_seen[key]))
    return display[winner]


def _is_positive_manual_value(value: Any) -> bool:
    if value is True or value == 1:
        return True
    return str(value or "").strip().lower() in {"yes", "y", "true", "1"}


def _instrument_analysis_by_symbol(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        if _is_test_trade_value(row.get("is_test_trade")):
            continue
        symbol = str(row.get("symbol") or "").strip().upper()
        if symbol:
            grouped[symbol].append(row)

    result: Dict[str, Dict[str, Any]] = {}
    mode_fields = {
        "pattern": "pattern",
        "ema": "ema",
        "most_traded_timeframe": "timeframe",
    }
    count_fields = {
        "round_number": "round_number",
        "spiked_out": "spiked_out",
        "close_stop_out": "close_stopout",
        "near_perfect_entry": "near_perfect_entry",
        "near_win": "near_win",
        "early_close": "early_close",
    }
    for symbol, symbol_rows in grouped.items():
        payload: Dict[str, Any] = {}
        for output_key, field in mode_fields.items():
            values = [
                _canonical_analysis_timeframe(row.get(field)) if field == "timeframe" else row.get(field)
                for row in symbol_rows
            ]
            payload[output_key] = _mode_nonblank(values)
        all_timeframe_counts: Counter[str] = Counter()
        for row in symbol_rows:
            timeframe = _canonical_analysis_timeframe(row.get("timeframe"))
            if timeframe:
                all_timeframe_counts[timeframe] += 1
        timeframe_totals: Dict[str, float] = defaultdict(float)
        timeframe_counts: Counter[str] = Counter()
        for row in symbol_rows:
            timeframe = _canonical_analysis_timeframe(row.get("timeframe"))
            if not timeframe:
                continue
            pnl = _as_float(row.get("net_profit"))
            if pnl is None:
                pnl = _as_float(row.get("result_cash"))
            if pnl is None:
                pnl = _as_float(row.get("result_pct"))
            if pnl is None or not math.isfinite(pnl):
                continue
            timeframe_totals[timeframe] += pnl
            timeframe_counts[timeframe] += 1
        timeframe_rank = {name: index for index, name in enumerate(TIMEFRAME_ORDER)}
        if timeframe_totals:
            best = min(
                timeframe_totals,
                key=lambda name: (
                    -timeframe_totals[name],
                    -timeframe_counts[name],
                    timeframe_rank.get(name, len(TIMEFRAME_ORDER)),
                    name,
                ),
            )
            worst = min(
                timeframe_totals,
                key=lambda name: (
                    timeframe_totals[name],
                    -timeframe_counts[name],
                    timeframe_rank.get(name, len(TIMEFRAME_ORDER)),
                    name,
                ),
            )
            payload["most_profitable_timeframe"] = best
            payload["least_profitable_timeframe"] = worst
        else:
            payload["most_profitable_timeframe"] = ""
            payload["least_profitable_timeframe"] = ""
        be_count = sum(
            1 for value in (_move_duration_seconds(row, "move_to_break_even") for row in symbol_rows)
            if value is not None and value > 0
        )
        profit_count = sum(
            1 for value in (_move_duration_seconds(row, "move_to_profit") for row in symbol_rows)
            if value is not None and value > 0
        )
        r_values = [
            value for value in (_as_float(row.get("r_multiple")) for row in symbol_rows)
            if value is not None and math.isfinite(value)
        ]
        duration_values = [
            value for value in (_infer_trade_duration_seconds(row) for row in symbol_rows)
            if value is not None and value >= 0
        ]
        winning_duration_values = [
            value for value in (_infer_trade_duration_seconds(row) for row in symbol_rows if _trade_outcome_sign(row) > 0)
            if value is not None and value >= 0
        ]
        losing_duration_values = [
            value for value in (_infer_trade_duration_seconds(row) for row in symbol_rows if _trade_outcome_sign(row) < 0)
            if value is not None and value >= 0
        ]
        aths_values = [str(row.get("aths_atls") or "").strip().lower() for row in symbol_rows]
        order_values = [str(row.get("order_type") or "").strip().lower() for row in symbol_rows]
        payload.update({
            "move_to_break_even": be_count,
            "move_to_profit": profit_count,
            "all_time_highs": sum(value in {"all-time high", "all time high", "aths", "ath"} for value in aths_values),
            "all_time_lows": sum(value in {"all-time low", "all time low", "atls", "atl"} for value in aths_values),
            "market_orders": sum("market" in value for value in order_values),
            "limit_orders": sum("limit" in value for value in order_values),
            "net_r_multiple": sum(r_values) if r_values else None,
            "shortest_duration": min(duration_values) if duration_values else None,
            "avg_duration": (sum(duration_values) / len(duration_values)) if duration_values else None,
            "longest_duration": max(duration_values) if duration_values else None,
            "avg_winning_duration": (sum(winning_duration_values) / len(winning_duration_values)) if winning_duration_values else None,
            "avg_losing_duration": (sum(losing_duration_values) / len(losing_duration_values)) if losing_duration_values else None,
        })
        payload.update({
            "timeframe_1m": all_timeframe_counts["1MIN"],
            "timeframe_5m": all_timeframe_counts["5MIN"],
            "timeframe_15m": all_timeframe_counts["15MIN"],
            "timeframe_30m": all_timeframe_counts["30MIN"],
            "timeframe_1h": all_timeframe_counts["1H"],
            "timeframe_4h": all_timeframe_counts["4H"],
            "timeframe_daily": all_timeframe_counts["DAILY"],
            "timeframe_weekly": all_timeframe_counts["WEEKLY"],
            "timeframe_monthly": all_timeframe_counts["MONTHLY"],
        })
        rec_summary = _distance_recommendation_summary(symbol_rows)
        payload[STOP_RECOMMENDATION_HEADER] = rec_summary.get(STOP_RECOMMENDATION_HEADER)
        payload[TARGET_RECOMMENDATION_HEADER] = rec_summary.get(TARGET_RECOMMENDATION_HEADER)
        payload["stop_recommendation"] = payload[STOP_RECOMMENDATION_HEADER]
        payload["target_recommendation"] = payload[TARGET_RECOMMENDATION_HEADER]
        for source_key, prefix in (("stop_loss", "stop"), ("take_profit", "target")):
            values = [
                pct
                for pct in (_distance_pct_points(row, source_key) for row in symbol_rows)
                if pct is not None and math.isfinite(pct)
            ]
            if values:
                payload[f"min_{prefix}_pct"] = min(values)
                payload[f"avg_{prefix}_pct"] = sum(values) / len(values)
                payload[f"max_{prefix}_pct"] = max(values)
        payload.update(_linear_profit_percentage_totals(symbol_rows))
        for output_key, field in count_fields.items():
            payload[output_key] = sum(_is_positive_manual_value(row.get(field)) for row in symbol_rows)
        result[symbol] = payload
    return result


def _instrument_summary_rows_from_trade_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    display_symbols: Dict[str, str] = {}
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        if _is_test_trade_value(row.get("is_test_trade")):
            continue
        symbol = str(row.get("symbol") or "").strip()
        if not symbol:
            continue
        key = symbol.upper()
        display_symbols.setdefault(key, symbol)
        grouped[key].append(row)

    summaries: List[Dict[str, Any]] = []
    for symbol_key in sorted(grouped):
        symbol_rows = grouped[symbol_key]
        wins = [row for row in symbol_rows if _trade_outcome_sign(row) > 0]
        losses = [row for row in symbol_rows if _trade_outcome_sign(row) < 0]
        breakeven = [row for row in symbol_rows if _trade_outcome_sign(row) == 0]
        longs = [row for row in symbol_rows if str(row.get("side") or "").upper().startswith(("BUY", "LONG"))]
        shorts = [row for row in symbol_rows if str(row.get("side") or "").upper().startswith(("SELL", "SHORT"))]
        durations = [
            value for value in (_infer_trade_duration_seconds(row) for row in symbol_rows)
            if value is not None and value >= 0
        ]
        winning_durations = [
            value for value in (_infer_trade_duration_seconds(row) for row in wins)
            if value is not None and value >= 0
        ]
        losing_durations = [
            value for value in (_infer_trade_duration_seconds(row) for row in losses)
            if value is not None and value >= 0
        ]

        def avg_distance(items: List[Dict[str, Any]], key: str) -> float | None:
            values = [
                value for value in (_distance_pct_points(row, key) for row in items)
                if value is not None and math.isfinite(value)
            ]
            return (sum(values) / len(values)) if values else None

        markets = [_trade_row_market(row) for row in symbol_rows]
        market_counts = Counter(market for market in markets if market)
        market = market_counts.most_common(1)[0][0] if market_counts else None
        net_profit_values = [
            value for value in (_as_float(row.get("net_profit")) for row in symbol_rows)
            if value is not None and math.isfinite(value)
        ]
        summaries.append({
            "symbol": display_symbols.get(symbol_key) or symbol_key,
            "asset_class": market,
            "total_trades": len(symbol_rows),
            "trades": len(symbol_rows),
            "wins": len(wins),
            "losses": len(losses),
            "break_even": len(breakeven),
            "long_trades": len(longs),
            "long_wins": sum(_trade_outcome_sign(row) > 0 for row in longs),
            "long_losses": sum(_trade_outcome_sign(row) < 0 for row in longs),
            "long_break_even": sum(_trade_outcome_sign(row) == 0 for row in longs),
            "short_trades": len(shorts),
            "short_wins": sum(_trade_outcome_sign(row) > 0 for row in shorts),
            "short_losses": sum(_trade_outcome_sign(row) < 0 for row in shorts),
            "short_break_even": sum(_trade_outcome_sign(row) == 0 for row in shorts),
            "win_rate_pct": (len(wins) / len(symbol_rows) * 100.0) if symbol_rows else None,
            "net_profit_total": sum(net_profit_values) if net_profit_values else None,
            "avg_net_profit": (sum(net_profit_values) / len(net_profit_values)) if net_profit_values else None,
            "avg_sl_pct": avg_distance(symbol_rows, "stop_loss"),
            "avg_sl_pct_wins": avg_distance(wins, "stop_loss"),
            "avg_sl_pct_losses": avg_distance(losses, "stop_loss"),
            "avg_tp_pct": avg_distance(symbol_rows, "take_profit"),
            "avg_tp_pct_wins": avg_distance(wins, "take_profit"),
            "avg_tp_pct_losses": avg_distance(losses, "take_profit"),
            "min_trade_duration_seconds": min(durations) if durations else None,
            "avg_trade_duration_seconds": (sum(durations) / len(durations)) if durations else None,
            "max_trade_duration_seconds": max(durations) if durations else None,
            "avg_winning_trade_duration_seconds": (sum(winning_durations) / len(winning_durations)) if winning_durations else None,
            "avg_losing_trade_duration_seconds": (sum(losing_durations) / len(losing_durations)) if losing_durations else None,
        })
    return summaries


def _instrument_leader_payloads_from_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    summaries = _instrument_summary_rows_from_trade_rows(rows)

    def pick(metric: str, market: str | None = None) -> Dict[str, Any] | None:
        candidates = [
            row for row in summaries
            if (market is None or str(row.get("asset_class") or "").casefold() == market)
            and _as_float(row.get(metric)) is not None
        ]
        if not candidates:
            return None
        return max(
            candidates,
            key=lambda row: (
                _as_float(row.get(metric)) or 0.0,
                _as_float(row.get("total_trades")) or 0.0,
                str(row.get("symbol") or ""),
            ),
        )

    return {
        "most_wins_instrument": pick("wins"),
        "most_losses_instrument": pick("losses"),
        "fx_most_wins_instrument": pick("wins", "fx"),
        "fx_most_losses_instrument": pick("losses", "fx"),
        "crypto_most_wins_instrument": pick("wins", "crypto"),
        "crypto_most_losses_instrument": pick("losses", "crypto"),
    }


def _result_percentage_totals_by_market(
    rows: List[Dict[str, Any]], balances: List[Dict[str, Any]]
) -> Dict[str, Dict[str, Any]]:
    """Return linear trade-result percentages plus account-return diagnostics by market.

    ``net_result_pct`` is the sum of non-test Trade Log Profit % values and is
    the displayed Net P/L Percentage. ``market_return_pct`` is a capital-based
    account return diagnostic and must not be relabeled as Net P/L Percentage.
    """
    grouped: Dict[str, List[Dict[str, Any]]] = {
        market: [] for market in ("overall", "fx", "crypto")
    }
    for row in rows:
        if str(row.get("row_type") or "trade") != "trade" or _is_test_trade_value(row.get("is_test_trade")):
            continue
        grouped["overall"].append(row)
        market = _trade_row_market(row)
        if market in {"fx", "crypto"}:
            grouped[market].append(row)

    balance_map: Dict[str, Dict[str, Any]] = {}
    for balance in balances or []:
        if not isinstance(balance, dict):
            continue
        account = _canonical_account_label(balance.get("account_label") or balance.get("account") or balance.get("label"))
        amount = _as_float(balance.get("balance"))
        if account and amount is not None and math.isfinite(amount):
            balance_map[account] = {
                "balance": amount,
                "currency": str(balance.get("currency") or balance.get("account_currency") or "").strip().upper(),
            }

    def _row_account(row: Dict[str, Any]) -> str:
        return _canonical_account_label(row.get("account_label") or row.get("account") or row.get("source") or "")

    def _row_currency(row: Dict[str, Any], market: str) -> str:
        currency = str(
            row.get("account_currency")
            or row.get("currency")
            or row.get("result_currency")
            or row.get("commission_currency")
            or ""
        ).strip().upper()
        return currency or ("AUD" if market == "fx" else "USDT" if market == "crypto" else "")

    def _cashflow_amount(row: Dict[str, Any]) -> float | None:
        amount = _as_float(row.get("cashflow_amount"))
        if amount is not None and math.isfinite(amount):
            return amount
        notes = str(row.get("notes") or "")
        match = re.search(r"\b(Deposit|Withdrawal)\s+(-?[0-9]+(?:\.[0-9]+)?)", notes, re.IGNORECASE)
        if not match:
            return None
        parsed = float(match.group(2))
        return abs(parsed) if match.group(1).lower() == "deposit" else -abs(parsed)

    def _account_return(account: str, trade_rows: List[Dict[str, Any]], market: str) -> Dict[str, Any]:
        account_rows = [
            row for row in rows
            if isinstance(row, dict)
            and _row_account(row) == account
            and str(row.get("row_type") or "trade").strip().lower() in {"trade", "cashflow"}
            and not _target_r_row_is_test_trade(row)
        ]
        account_rows.sort(key=lambda row: _as_datetime(row.get("close_time") or row.get("open_time")) or datetime.min)
        currency = next((_row_currency(row, market) for row in trade_rows if _row_currency(row, market)), "")
        if not currency:
            currency = str((balance_map.get(account) or {}).get("currency") or "")
        authoritative_ending = (balance_map.get(account) or {}).get("balance")
        segments: List[Dict[str, Any]] = []
        current_start: float | None = None
        current_start_type = ""
        current_last_balance: float | None = None
        current_pnl = 0.0
        current_trades = 0
        reset_count = 0
        discontinuities: List[Dict[str, Any]] = []

        def _close_segment(reason: str, ending_override: float | None = None) -> None:
            nonlocal current_start, current_start_type, current_last_balance, current_pnl, current_trades
            if current_start is None or current_trades <= 0:
                return
            ending_balance = ending_override if ending_override is not None else current_last_balance
            if ending_balance is None:
                return
            segments.append({
                "starting_capital": current_start,
                "ending_equity": ending_balance,
                "return_numerator": current_pnl,
                "pnl_total": current_pnl,
                "trades": current_trades,
                "starting_anchor": current_start_type,
                "close_reason": reason,
            })

        def _reset_segment(start: float | None, anchor: str) -> None:
            nonlocal current_start, current_start_type, current_last_balance, current_pnl, current_trades
            current_start = start
            current_start_type = anchor
            current_last_balance = start
            current_pnl = 0.0
            current_trades = 0

        for row in account_rows:
            row_type = str(row.get("row_type") or "trade").strip().lower()
            if row_type == "cashflow":
                amount = _cashflow_amount(row)
                event_balance = _as_float(row.get("cashflow_new_balance"))
                if event_balance is None:
                    event_balance = _as_float(row.get("balance_after_trade"))
                if event_balance is None or not math.isfinite(event_balance):
                    continue
                _close_segment("cashflow_anchor")
                _reset_segment(event_balance, "cashflow_anchor")
                if amount is None:
                    discontinuities.append({
                        "reason": "cashflow_amount_inferred_or_missing",
                        "date": row.get("close_time") or row.get("open_time"),
                    })
                continue
            else:
                event_balance = _as_float(row.get("analysis_balance_after_trade"))
                if event_balance is None:
                    event_balance = _as_float(row.get("balance_after_trade"))
            if event_balance is None or not math.isfinite(event_balance):
                continue
            pnl = _as_float(row.get("net_profit"))
            if pnl is None:
                pnl = _as_float(row.get("result_cash"))
            if pnl is None or not math.isfinite(pnl):
                pnl = 0.0
            if current_start is None:
                _reset_segment(event_balance - pnl, "first_trade_balance")
            elif current_last_balance is not None:
                expected = current_last_balance + pnl
                tolerance = max(0.05, abs(expected) * 0.002)
                diff = event_balance - expected
                if abs(diff) > tolerance:
                    _close_segment("balance_reset")
                    reset_count += 1
                    discontinuities.append({
                        "reason": "balance_reset",
                        "date": row.get("close_time") or row.get("open_time"),
                        "expected_balance": expected,
                        "actual_balance": event_balance,
                        "difference": diff,
                    })
                    _reset_segment(event_balance - pnl, "balance_reset")
            current_pnl += pnl
            current_trades += 1
            current_last_balance = event_balance

        final_ending = authoritative_ending if authoritative_ending is not None else current_last_balance
        _close_segment("final_balance", final_ending)
        valid_segments = [
            segment for segment in segments
            if _as_float(segment.get("starting_capital")) is not None
            and _as_float(segment.get("ending_equity")) is not None
            and float(segment["starting_capital"]) > 0
            and math.isfinite(float(segment["starting_capital"]))
            and math.isfinite(float(segment["ending_equity"]))
        ]
        if not valid_segments:
            pnl_total = sum(
                value for value in (
                    _as_float(row.get("net_profit")) if _as_float(row.get("net_profit")) is not None
                    else _as_float(row.get("result_cash"))
                    for row in trade_rows
                )
                if value is not None and math.isfinite(value)
            )
            if authoritative_ending is not None and math.isfinite(authoritative_ending):
                inferred_start = authoritative_ending - pnl_total
                if inferred_start > 0:
                    valid_segments = [{
                        "starting_capital": inferred_start,
                        "ending_equity": authoritative_ending,
                        "return_numerator": pnl_total,
                        "pnl_total": pnl_total,
                        "trades": len(trade_rows),
                        "starting_anchor": "inferred_from_ending_balance",
                        "close_reason": "ending_balance_fallback",
                    }]
        if not valid_segments:
            return {"available": False, "reason": "missing_balance_anchor", "account": account, "currency": currency}
        first_balance = sum(float(segment["starting_capital"]) for segment in valid_segments)
        numerator = sum(float(segment["return_numerator"]) for segment in valid_segments)
        ending = sum(float(segment["ending_equity"]) for segment in valid_segments)
        pnl_values = []
        for trade_row in trade_rows:
            pnl_value = _as_float(trade_row.get("net_profit"))
            if pnl_value is None:
                pnl_value = _as_float(trade_row.get("result_cash"))
            if pnl_value is not None and math.isfinite(pnl_value):
                pnl_values.append(pnl_value)
        gross_gain = sum(value for value in pnl_values if value > 0)
        gross_loss = abs(sum(value for value in pnl_values if value < 0))
        return {
            "available": True,
            "account": account,
            "currency": currency,
            "starting_capital": first_balance,
            "ending_equity": ending,
            "net_cashflow": None,
            "return_numerator": numerator,
            "gross_gain_numerator": gross_gain,
            "gross_loss_numerator": gross_loss,
            "return_pct": numerator / first_balance * 100.0,
            "gross_gain_return_pct": gross_gain / first_balance * 100.0,
            "gross_loss_return_pct": gross_loss / first_balance * 100.0,
            "starting_anchor": "segmented_account_balance",
            "segments": valid_segments,
            "reset_count": reset_count,
            "discontinuities": discontinuities,
        }

    result: Dict[str, Dict[str, Any]] = {}
    for market in ("overall", "fx", "crypto"):
        totals = _linear_profit_percentage_totals(grouped[market])
        by_account: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for row in grouped[market]:
            account = _row_account(row)
            if account:
                by_account[account].append(row)
        account_returns = [
            _account_return(account, trade_rows, _trade_row_market(trade_rows[0]) or market)
            for account, trade_rows in sorted(by_account.items())
        ]
        available = [item for item in account_returns if item.get("available")]
        unavailable_reason: str | None = None
        market_return: float | None = None
        gross_gain_return: float | None = None
        gross_loss_return: float | None = None
        if not by_account:
            unavailable_reason = "missing_trade_accounts"
        elif len(available) != len(by_account):
            unavailable_reason = "missing_balance_anchor"
        elif available:
            weighted_items: List[Tuple[Dict[str, Any], float]] = []
            unsupported = []
            for item in available:
                factor = _currency_to_aud_factor(item.get("currency"))
                if factor is None:
                    unsupported.append(item)
                else:
                    weighted_items.append((item, factor))
            if unsupported:
                unavailable_reason = "unsupported_currency_for_aud_weighting"
            else:
                denominator = sum(float(item["starting_capital"]) * factor for item, factor in weighted_items)
                numerator = sum(float(item["return_numerator"]) * factor for item, factor in weighted_items)
                gross_gain_numerator = sum(float(item.get("gross_gain_numerator") or 0.0) * factor for item, factor in weighted_items)
                gross_loss_numerator = sum(float(item.get("gross_loss_numerator") or 0.0) * factor for item, factor in weighted_items)
                if denominator > 0:
                    market_return = numerator / denominator * 100.0
                    gross_gain_return = gross_gain_numerator / denominator * 100.0
                    gross_loss_return = gross_loss_numerator / denominator * 100.0
                    proven_full_loss = any(
                        float(segment.get("return_numerator") or 0.0) <= -float(segment.get("starting_capital") or 0.0)
                        for item in available
                        for segment in (item.get("segments") or [])
                        if _as_float(segment.get("starting_capital")) is not None
                    )
                    if market_return < -100.0 and not proven_full_loss:
                        market_return = None
                        unavailable_reason = "return_below_minus_100_unverified"
                else:
                    unavailable_reason = "invalid_combined_starting_capital"
        result[market] = {
            **totals,
            "market_return_pct": market_return,
            "gross_gain_return_pct": gross_gain_return,
            "gross_loss_return_pct": gross_loss_return,
            "return_method": "capital_weighted_account_return_aud",
            "return_unavailable_reason": unavailable_reason,
            "return_diagnostics": account_returns,
        }
    return result


def _risk_of_ruin_by_account(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        if _is_test_trade_value(row.get("is_test_trade")):
            continue
        account = _canonical_account_label(row.get("account_label") or row.get("account") or "")
        if account:
            grouped[account].append(row)
    result: Dict[str, Dict[str, Any]] = {}
    for account, account_rows in grouped.items():
        def _risk_payload(reason: str | None, risk_of_ruin: float | None = None, **extra: Any) -> Dict[str, Any]:
            payload = {
                "risk_of_ruin": risk_of_ruin,
                "reason": reason,
                "model": "fixed_fractional_balsara",
                "trade_count": len(account_rows),
                "win_rate": None,
                "payoff_ratio": None,
                "risk_per_trade_fraction": None,
                "edge": None,
                "capital_units": None,
                "risk_source": None,
            }
            payload.update(extra)
            return payload

        wins: List[float] = []
        losses: List[float] = []
        balance_risks: List[float] = []
        fallback_risks: List[float] = []
        for row in account_rows:
            pnl = _as_float(row.get("net_profit"))
            if pnl is None:
                pnl = _as_float(row.get("result_cash"))
            if pnl is None:
                result_pct = _as_float(row.get("result_pct"))
                if result_pct is not None:
                    pnl = result_pct
            if pnl is not None and math.isfinite(pnl):
                if pnl > 0:
                    wins.append(pnl)
                elif pnl < 0:
                    losses.append(abs(pnl))
                    balance_after = _as_float(row.get("analysis_balance_after_trade"))
                    if balance_after is None:
                        balance_after = _as_float(row.get("balance_after_trade"))
                    if balance_after is not None:
                        balance_before = balance_after + abs(pnl)
                        if balance_before > 0:
                            balance_risks.append(abs(pnl) / balance_before)
            risk_pct = _as_float(row.get("stop_loss_distance_pct"))
            if risk_pct is None:
                distance = _validated_distance_fraction(row, "stop_loss")
                risk_pct = distance * 100.0 if distance is not None else None
            if risk_pct is not None and math.isfinite(risk_pct) and risk_pct > 0:
                fallback_risks.append(risk_pct / 100.0)
        if not wins or not losses:
            result[account] = _risk_payload("requires_wins_and_losses")
            continue
        risk_samples = balance_risks or fallback_risks
        if not risk_samples:
            result[account] = _risk_payload("missing_risk_per_trade")
            continue
        sorted_risks = sorted(risk_samples)
        midpoint = len(sorted_risks) // 2
        risk_fraction = (
            sorted_risks[midpoint]
            if len(sorted_risks) % 2
            else (sorted_risks[midpoint - 1] + sorted_risks[midpoint]) / 2.0
        )
        win_rate = len(wins) / (len(wins) + len(losses))
        loss_rate = 1.0 - win_rate
        avg_win = sum(wins) / len(wins)
        avg_loss = sum(losses) / len(losses)
        payoff_ratio = avg_win / avg_loss if avg_loss > 0 else None
        if payoff_ratio is None or risk_fraction <= 0 or not math.isfinite(risk_fraction):
            result[account] = _risk_payload("invalid_model_inputs")
            continue
        edge = (win_rate * payoff_ratio) - loss_rate
        capital_units = 1.0 / risk_fraction
        # A non-positive edge is mathematically certain ruin in this model.
        # Positive-edge accounts use the Balsara-style curve directly instead
        # of being coerced to 100%.
        if edge <= 0:
            risk_of_ruin = 1.0
        elif edge >= 1:
            risk_of_ruin = 0.0
        else:
            risk_of_ruin = ((1.0 - edge) / (1.0 + edge)) ** capital_units
        risk_of_ruin = max(0.0, min(1.0, risk_of_ruin))
        result[account] = _risk_payload(
            None,
            risk_of_ruin,
            win_rate=win_rate,
            loss_rate=loss_rate,
            payoff_ratio=payoff_ratio,
            risk_per_trade_fraction=risk_fraction,
            risk_source="median_loss_over_balance" if balance_risks else "median_stop_loss_distance",
            edge=edge,
            capital_units=capital_units,
        )
    return result


def _empty_risk_of_ruin_payload(reason: str, trade_count: int | None = None) -> Dict[str, Any]:
    return {
        "risk_of_ruin": None,
        "reason": reason,
        "model": "fixed_fractional_balsara",
        "trade_count": trade_count,
        "win_rate": None,
        "payoff_ratio": None,
        "risk_per_trade_fraction": None,
        "edge": None,
        "capital_units": None,
        "risk_source": None,
    }


def _risk_of_ruin_comment_text(payload: Dict[str, Any]) -> str:
    wanted = {
        key: payload.get(key)
        for key in (
            "win_rate",
            "payoff_ratio",
            "risk_per_trade_fraction",
            "edge",
            "capital_units",
            "risk_source",
            "trade_count",
            "reason",
        )
    }
    return f"Fixed-fractional/Balsara-style risk of ruin. Model inputs: {wanted}"


def _dashboard_extended_metrics(
    rows: List[Dict[str, Any]],
    by_market: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    active = [
        row for row in rows
        if str(row.get("row_type") or "trade") == "trade"
        and not _is_test_trade_value(row.get("is_test_trade"))
    ]

    def _subset(market: str) -> List[Dict[str, Any]]:
        return active if market == "overall" else [
            row for row in active if _trade_row_market(row) == market
        ]

    def _test_subset(market: str) -> List[Dict[str, Any]]:
        return [
            row for row in rows
            if str(row.get("row_type") or "trade") == "trade"
            and _is_test_trade_value(row.get("is_test_trade"))
            and (market == "overall" or _trade_row_market(row) == market)
        ]

    def _outcome(row: Dict[str, Any]) -> int:
        value = _as_float(row.get("result_pct"))
        if value is None:
            value = _as_float(row.get("net_profit"))
        if value is None or value == 0:
            return 0
        return 1 if value > 0 else -1

    def _distance_samples(items: List[Dict[str, Any]], key: str) -> List[Tuple[float, Dict[str, Any]]]:
        samples: List[Tuple[float, Dict[str, Any]]] = []
        for item in items:
            pct = _distance_pct_points(item, key)
            if pct is not None:
                samples.append((pct, item))
        return samples

    def _distance_values(items: List[Dict[str, Any]], key: str) -> List[float]:
        return [value for value, _item in _distance_samples(items, key)]

    def _distance_extreme_ref(items: List[Dict[str, Any]], key: str, mode: str, source_key: str) -> Dict[str, Any] | None:
        samples = _distance_samples(items, key)
        picked = _pick_deterministic_extreme(samples, mode)
        if not picked:
            return None
        picked_value, picked_row = picked
        return _trade_metric_ref(picked_row, source_key, picked_value)

    def _metric_extreme_ref(items: List[Dict[str, Any]], key: str, mode: str, source_key: str) -> Dict[str, Any] | None:
        samples = [
            (value, item)
            for item in items
            for value in [_as_float(item.get(key))]
            if value is not None and math.isfinite(value)
        ]
        picked = _pick_deterministic_extreme(samples, mode)
        if not picked:
            return None
        picked_value, picked_row = picked
        return _trade_metric_ref(picked_row, source_key, picked_value)

    def _derived_extreme_ref(
        items: List[Dict[str, Any]],
        value_getter,
        mode: str,
        source_key: str,
    ) -> Dict[str, Any] | None:
        samples = [
            (float(value), item)
            for item in items
            for value in [value_getter(item)]
            if value is not None and math.isfinite(float(value)) and float(value) >= 0
        ]
        picked = _pick_deterministic_extreme(samples, mode)
        if not picked:
            return None
        picked_value, picked_row = picked
        return _trade_metric_ref(dict(picked_row), source_key, picked_value)

    def _summary(values: List[float], prefix: str) -> Dict[str, Any]:
        return {
            f"min_{prefix}": min(values) if values else None,
            f"avg_{prefix}": (sum(values) / len(values)) if values else None,
            f"max_{prefix}": max(values) if values else None,
        }

    def _pattern_metrics(items: List[Dict[str, Any]]) -> Dict[str, Any]:
        pattern_display: Dict[str, str] = {}
        patterns = []
        for item in items:
            pattern = str(item.get("pattern") or "").strip()
            if not pattern:
                continue
            key = pattern.casefold()
            pattern_display.setdefault(key, pattern)
            patterns.append(key)
        if not patterns:
            return {
                "pattern_channel_total": 0,
                "pattern_channel_wins": 0,
                "pattern_channel_losses": 0,
                "pattern_range_total": 0,
                "pattern_range_wins": 0,
                "pattern_range_losses": 0,
            }
        counts = Counter(patterns)
        pnl_pct: Dict[str, float] = defaultdict(float)
        for item in items:
            pattern = str(item.get("pattern") or "").strip().casefold()
            result_pct = _as_float(item.get("result_pct"))
            if pattern and result_pct is not None:
                pnl_pct[pattern] += result_pct
        count_order = sorted(counts, key=lambda pattern: (counts[pattern], pattern.lower()))
        profit_order = sorted(counts, key=lambda pattern: (pnl_pct.get(pattern, 0.0), pattern.lower()))
        metrics = {
            "most_traded_pattern": pattern_display.get(count_order[-1], count_order[-1]),
            "least_traded_pattern": pattern_display.get(count_order[0], count_order[0]),
            "most_profitable_pattern": pattern_display.get(profit_order[-1], profit_order[-1]),
            "least_profitable_pattern": pattern_display.get(profit_order[0], profit_order[0]),
        }
        for wanted in ("channel", "range"):
            subset = [
                item for item in items
                if str(item.get("pattern") or "").strip().casefold() == wanted
            ]
            metrics[f"pattern_{wanted}_total"] = len(subset)
            metrics[f"pattern_{wanted}_wins"] = sum(_outcome(item) > 0 for item in subset)
            metrics[f"pattern_{wanted}_losses"] = sum(_outcome(item) < 0 for item in subset)
        return metrics

    timeframe_aliases = {
        "1MIN": "1MIN", "5MIN": "5MIN", "15MIN": "15MIN", "30MIN": "30MIN",
        "1H": "1H", "4H": "4H", "1D": "DAILY", "DAILY": "DAILY",
        "1W": "WEEKLY", "WEEKLY": "WEEKLY", "1MO": "MONTHLY", "MONTHLY": "MONTHLY",
    }
    result: Dict[str, Dict[str, Any]] = {}
    for market in ("overall", "fx", "crypto"):
        items = _subset(market)
        winners = [item for item in items if _outcome(item) > 0]
        losers = [item for item in items if _outcome(item) < 0]
        durations = [
            value for value in (_as_float(item.get("trade_duration_seconds")) for item in items)
            if value is not None and value >= 0
        ]
        winner_durations = [
            value for value in (_as_float(item.get("trade_duration_seconds")) for item in winners)
            if value is not None and value >= 0
        ]
        loser_durations = [
            value for value in (_as_float(item.get("trade_duration_seconds")) for item in losers)
            if value is not None and value >= 0
        ]
        break_even_moves = [
            value for value in (_move_duration_seconds(item, "move_to_break_even") for item in items)
            if value is not None
        ]
        profit_moves = [
            value for value in (_move_duration_seconds(item, "move_to_profit") for item in items)
            if value is not None
        ]
        winner_results = [
            value for value in (_as_float(item.get("result_pct")) for item in winners)
            if value is not None
        ]
        loser_results = [
            value for value in (_as_float(item.get("result_pct")) for item in losers)
            if value is not None
        ]
        result_values = [
            value for value in (_as_float(item.get("result_pct")) for item in items)
            if value is not None and math.isfinite(value)
        ]
        winner_r = [
            value for value in (_as_float(item.get("r_multiple")) for item in winners)
            if value is not None and value > 0
        ]
        loser_r = [
            value for value in (_as_float(item.get("r_multiple")) for item in losers)
            if value is not None and value < 0
        ]
        r_values = [
            value for value in (_as_float(item.get("r_multiple")) for item in items)
            if value is not None and math.isfinite(value)
        ]
        commission_rows: List[Tuple[float, Dict[str, Any]]] = []
        for item in items:
            value = _as_float(item.get("commission"))
            if value is not None and value != 0:
                commission_rows.append((abs(value), item))
        commissions = [value for value, _item in commission_rows]
        timeframe_counts: Counter[str] = Counter()
        timeframe_wins: Counter[str] = Counter()
        timeframe_losses: Counter[str] = Counter()
        for item in items:
            timeframe = timeframe_aliases.get(_canonical_journal_timeframe(item.get("timeframe")).upper())
            if timeframe:
                timeframe_counts[timeframe] += 1
                if _outcome(item) > 0:
                    timeframe_wins[timeframe] += 1
                elif _outcome(item) < 0:
                    timeframe_losses[timeframe] += 1
        stats_bucket = by_market.get(market) if isinstance(by_market.get(market), dict) else {}
        min_commission_source = min(commission_rows, key=lambda pair: (pair[0], str(pair[1].get("symbol") or "")))[1] if commission_rows else None
        max_commission_source = max(commission_rows, key=lambda pair: (pair[0], str(pair[1].get("symbol") or "")))[1] if commission_rows else None
        streaks = _period_streak_metrics(items)
        drawdown = _period_drawdown_metrics(items, rows)
        target_scope = "overall" if market == "overall" else "standard"
        result[market] = {
            "trades": len(items),
            "wins": len(winners),
            "losses": len(losers),
            "break_even": sum(_outcome(item) == 0 for item in items),
            "test_trades": len(_test_subset(market)),
            "win_rate_pct": (len(winners) / len(items) * 100.0) if items else None,
            **_summary(durations, "duration_seconds"),
            **_summary(winner_durations, "duration_seconds_winners"),
            **_summary(loser_durations, "duration_seconds_losers"),
            **_summary(break_even_moves, "move_to_break_even_duration_seconds"),
            **_summary(profit_moves, "move_to_profit_duration_seconds"),
            **_summary(result_values, "result_pct"),
            **_summary(_distance_values(items, "stop_loss"), "stop_pct"),
            **_summary(_distance_values(items, "take_profit"), "target_pct"),
            **_summary(_distance_values(winners, "stop_loss"), "stop_pct_winners"),
            **_summary(_distance_values(winners, "take_profit"), "target_pct_winners"),
            **_summary(winner_results, "result_pct_winners"),
            **_summary(winner_r, "r_multiple_winners"),
            **_summary(_distance_values(losers, "stop_loss"), "stop_pct_losers"),
            **_summary(_distance_values(losers, "take_profit"), "target_pct_losers"),
            **_summary(loser_results, "result_pct_losers"),
            **_summary(loser_r, "r_multiple_losers"),
            **_summary(r_values, "r_multiple"),
            "long_trades": sum(str(item.get("side") or "").upper().startswith(("BUY", "LONG")) for item in items),
            "long_wins": sum(str(item.get("side") or "").upper().startswith(("BUY", "LONG")) and _outcome(item) > 0 for item in items),
            "long_losses": sum(str(item.get("side") or "").upper().startswith(("BUY", "LONG")) and _outcome(item) < 0 for item in items),
            "long_break_even": sum(str(item.get("side") or "").upper().startswith(("BUY", "LONG")) and _outcome(item) == 0 for item in items),
            "short_trades": sum(str(item.get("side") or "").upper().startswith(("SELL", "SHORT")) for item in items),
            "short_wins": sum(str(item.get("side") or "").upper().startswith(("SELL", "SHORT")) and _outcome(item) > 0 for item in items),
            "short_losses": sum(str(item.get("side") or "").upper().startswith(("SELL", "SHORT")) and _outcome(item) < 0 for item in items),
            "short_break_even": sum(str(item.get("side") or "").upper().startswith(("SELL", "SHORT")) and _outcome(item) == 0 for item in items),
            **_pattern_metrics(items),
            **{f"timeframe_{label.lower()}": timeframe_counts[label] for label in timeframe_aliases.values()},
            **{f"timeframe_{label.lower()}_wins": timeframe_wins[label] for label in timeframe_aliases.values()},
            **{f"timeframe_{label.lower()}_losses": timeframe_losses[label] for label in timeframe_aliases.values()},
            **_summary(commissions, "commission"),
            **_distance_recommendation_summary(items, scope=target_scope),
            "avg_r_multiple": (sum(r_values) / len(r_values)) if r_values else None,
            "net_r_multiple": sum(r_values) if r_values else None,
            "gross_ir_gain": sum(value for value in r_values if value > 0) if r_values else None,
            "gross_ir_loss": abs(sum(value for value in r_values if value < 0)) if r_values else None,
            "total_commission": sum(commissions) if commissions else None,
            "min_commission_source": _fmt_detail_src(min_commission_source) if min_commission_source else "",
            "max_commission_source": _fmt_detail_src(max_commission_source) if max_commission_source else "",
            "metric_sources": {
                "min_stop_pct": _distance_extreme_ref(items, "stop_loss", "min", "min_stop_pct"),
                "max_stop_pct": _distance_extreme_ref(items, "stop_loss", "max", "max_stop_pct"),
                "min_target_pct": _distance_extreme_ref(items, "take_profit", "min", "min_target_pct"),
                "max_target_pct": _distance_extreme_ref(items, "take_profit", "max", "max_target_pct"),
                "min_stop_pct_winners": _distance_extreme_ref(winners, "stop_loss", "min", "min_stop_pct_winners"),
                "max_stop_pct_winners": _distance_extreme_ref(winners, "stop_loss", "max", "max_stop_pct_winners"),
                "min_target_pct_winners": _distance_extreme_ref(winners, "take_profit", "min", "min_target_pct_winners"),
                "max_target_pct_winners": _distance_extreme_ref(winners, "take_profit", "max", "max_target_pct_winners"),
                "min_stop_pct_losers": _distance_extreme_ref(losers, "stop_loss", "min", "min_stop_pct_losers"),
                "max_stop_pct_losers": _distance_extreme_ref(losers, "stop_loss", "max", "max_stop_pct_losers"),
                "min_target_pct_losers": _distance_extreme_ref(losers, "take_profit", "min", "min_target_pct_losers"),
                "max_target_pct_losers": _distance_extreme_ref(losers, "take_profit", "max", "max_target_pct_losers"),
                "min_result_pct": _metric_extreme_ref(items, "result_pct", "min", "min_result_pct"),
                "max_result_pct": _metric_extreme_ref(items, "result_pct", "max", "max_result_pct"),
                "min_r_multiple": _metric_extreme_ref(items, "r_multiple", "min", "min_r_multiple"),
                "max_r_multiple": _metric_extreme_ref(items, "r_multiple", "max", "max_r_multiple"),
                "min_result_pct_winners": _metric_extreme_ref(winners, "result_pct", "min", "min_result_pct_winners"),
                "max_result_pct_winners": _metric_extreme_ref(winners, "result_pct", "max", "max_result_pct_winners"),
                "min_r_multiple_winners": _metric_extreme_ref(winners, "r_multiple", "min", "min_r_multiple_winners"),
                "max_r_multiple_winners": _metric_extreme_ref(winners, "r_multiple", "max", "max_r_multiple_winners"),
                "min_result_pct_losers": _metric_extreme_ref(losers, "result_pct", "min", "min_result_pct_losers"),
                "max_result_pct_losers": _metric_extreme_ref(losers, "result_pct", "max", "max_result_pct_losers"),
                "min_r_multiple_losers": _metric_extreme_ref(losers, "r_multiple", "min", "min_r_multiple_losers"),
                "max_r_multiple_losers": _metric_extreme_ref(losers, "r_multiple", "max", "max_r_multiple_losers"),
                "min_duration_seconds": _metric_extreme_ref(items, "trade_duration_seconds", "min", "min_duration_seconds"),
                "max_duration_seconds": _metric_extreme_ref(items, "trade_duration_seconds", "max", "max_duration_seconds"),
                "shortest_duration_seconds": _metric_extreme_ref(items, "trade_duration_seconds", "min", "shortest_duration_seconds"),
                "longest_duration_seconds": _metric_extreme_ref(items, "trade_duration_seconds", "max", "longest_duration_seconds"),
                "min_duration_seconds_winners": _metric_extreme_ref(winners, "trade_duration_seconds", "min", "min_duration_seconds_winners"),
                "max_duration_seconds_winners": _metric_extreme_ref(winners, "trade_duration_seconds", "max", "max_duration_seconds_winners"),
                "min_duration_seconds_losers": _metric_extreme_ref(losers, "trade_duration_seconds", "min", "min_duration_seconds_losers"),
                "max_duration_seconds_losers": _metric_extreme_ref(losers, "trade_duration_seconds", "max", "max_duration_seconds_losers"),
                "min_move_to_break_even_duration_seconds": _derived_extreme_ref(items, lambda row: _move_duration_seconds(row, "move_to_break_even"), "min", "min_move_to_break_even_duration_seconds"),
                "max_move_to_break_even_duration_seconds": _derived_extreme_ref(items, lambda row: _move_duration_seconds(row, "move_to_break_even"), "max", "max_move_to_break_even_duration_seconds"),
                "min_move_to_profit_duration_seconds": _derived_extreme_ref(items, lambda row: _move_duration_seconds(row, "move_to_profit"), "min", "min_move_to_profit_duration_seconds"),
                "max_move_to_profit_duration_seconds": _derived_extreme_ref(items, lambda row: _move_duration_seconds(row, "move_to_profit"), "max", "max_move_to_profit_duration_seconds"),
            },
            "min_drawdown_pct": stats_bucket.get("min_drawdown_pct") if stats_bucket.get("min_drawdown_pct") is not None else drawdown.get("min_drawdown_pct"),
            "avg_drawdown_pct": stats_bucket.get("avg_drawdown_pct") if stats_bucket.get("avg_drawdown_pct") is not None else drawdown.get("avg_drawdown_pct"),
            "max_drawdown_pct": stats_bucket.get("max_drawdown_pct") if stats_bucket.get("max_drawdown_pct") is not None else drawdown.get("max_drawdown_pct"),
            "min_drawdown_detail": stats_bucket.get("min_drawdown_detail") or drawdown.get("min_drawdown_detail"),
            "max_drawdown_detail": stats_bucket.get("max_drawdown_detail") or drawdown.get("max_drawdown_detail"),
            "winning_streak": stats_bucket.get("winning_streak") or streaks.get("winning_streak"),
            "losing_streak": stats_bucket.get("losing_streak") or streaks.get("losing_streak"),
            "longest_winning_streak": stats_bucket.get("longest_winning_streak") or streaks.get("longest_winning_streak"),
            "longest_losing_streak": stats_bucket.get("longest_losing_streak") or streaks.get("longest_losing_streak"),
        }
    return result


def _trade_log_datetime_value(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        parsed = value
    else:
        parsed = datetime.fromisoformat(value.replace("Z", ""))
    if isinstance(parsed, datetime) and parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)
    return parsed


def _trade_log_row_values(
    row: Mapping[str, Any],
    *,
    resolved_balance: Any = None,
    recommendations_by_symbol: Optional[Mapping[str, Any]] = None,
) -> Dict[str, Any]:
    """Serialize a journal row to the canonical Trade Log logical columns."""
    row_map = dict(row)
    pct = _as_float(row_map.get("result_pct"))
    is_monthly = str(row_map.get("row_type") or "") == "monthly_aud_reval"
    symbol = row_map.get("symbol") or ("MONTHLY AUD P/L" if is_monthly else "")
    account = row_map.get("account_label") or row_map.get("account") or ("BYBIT" if is_monthly else "")
    notes = row_map.get("notes") or (
        "Monthly BYBIT AUD P/L bookkeeping note (excluded from metrics)."
        if is_monthly
        else ""
    )
    net_pnl = (
        row_map.get("net_profit")
        if row_map.get("net_profit") is not None
        else row_map.get("result_cash")
    )
    open_time = row_map.get("open_time") or row_map.get("period_month")
    close_time = row_map.get("close_time") or row_map.get("period_month")
    commission = _as_float(row_map.get("commission"))
    commission_value = "" if commission in (None, 0.0) else commission
    cashflow_new_balance = row_map.get("cashflow_new_balance")
    row_type = str(row_map.get("row_type") or "trade").strip().lower()
    if row_type == "cashflow" and cashflow_new_balance in (None, ""):
        cashflow_new_balance = resolved_balance
    side = str(row_map.get("side") or "").upper()
    setup_value = "" if row_type in {"monthly_aud_reval", "cashflow"} else (row_map.get("setup") or "")
    stop_loss_distance: Any = ""
    target_distance: Any = ""
    stop_loss_price = row_map.get("stop_loss")
    target_price = row_map.get("take_profit")
    if row_type == "trade":
        stop_loss_distance = _validated_distance_fraction(row_map, "stop_loss")
        target_distance = _validated_distance_fraction(row_map, "take_profit")
        if (
            _distance_fraction_from_prices(row_map.get("entry_price"), stop_loss_price) is not None
            and stop_loss_distance is None
        ):
            stop_loss_price = ""
        if (
            _distance_fraction_from_prices(row_map.get("entry_price"), target_price) is not None
            and target_distance is None
        ):
            target_price = ""
        stop_loss_distance = "" if stop_loss_distance is None else stop_loss_distance
        target_distance = "" if target_distance is None else target_distance
    close_stopout = row_map.get("close_stopout")
    if close_stopout in (None, ""):
        close_stopout = row_map.get("close_stop_out")
    if close_stopout in (None, ""):
        close_stopout = row_map.get("stop_out")
    recommendations = dict(recommendations_by_symbol or {})
    values: Dict[str, Any] = {
        TRADE_NUMBER_HEADER: str(row_map.get("trade_number") or ""),
        "Open Time": _trade_log_datetime_value(open_time),
        "Close Time": _trade_log_datetime_value(close_time),
        "Account": account,
        "Symbol": symbol,
        "Side": side,
        "Qty": row_map.get("qty"),
        "Entry Price": row_map.get("entry_price"),
        "Exit Price": row_map.get("exit_price"),
        "Stop Loss Price": stop_loss_price,
        "Stop Loss Distance": stop_loss_distance,
        STOP_RECOMMENDATION_HEADER: (
            _row_distance_recommendation(row_map, recommendations, STOP_RECOMMENDATION_HEADER)
            if row_type == "trade" and recommendations
            else ""
        ),
        "Target Price": target_price,
        "Target Distance": target_distance,
        TARGET_RECOMMENDATION_HEADER: (
            _row_distance_recommendation(row_map, recommendations, TARGET_RECOMMENDATION_HEADER)
            if row_type == "trade" and recommendations
            else ""
        ),
        "Commission": commission_value,
        "Net P/L": net_pnl,
        "Profit %": (pct / 100.0 if pct is not None else ""),
        "R-Multiple": row_map.get("r_multiple"),
        "Balance After": resolved_balance,
        "Trade Duration (DD:HH:MM:SS)": _format_duration_display(
            _infer_trade_duration_seconds(row_map)
        ),
        "Test": "Yes" if _is_test_trade_value(row_map.get("is_test_trade")) else "No",
        "Pattern": row_map.get("pattern") or "",
        "EMA": row_map.get("ema") or "",
        "VWAP": row_map.get("vwap") or "",
        "ATHS/ATLS": row_map.get("aths_atls") or "",
        "Order": row_map.get("order_type") or "",
        "Round Number": row_map.get("round_number") or "",
        "Spiked Out": row_map.get("spiked_out") or "",
        "Close Stopout": close_stopout or "",
        "Near Perfect Entry": row_map.get("near_perfect_entry") or "",
        "Near Win": row_map.get("near_win") or "",
        "Early Close": row_map.get("early_close") or "",
        "Setup": setup_value,
        "Timeframe": _canonical_journal_timeframe(row_map.get("timeframe") or ""),
        "Breakeven": row_map.get("breakeven") or "",
        "Notes": notes,
        "Cashflow Amount": row_map.get("cashflow_amount"),
        "Cashflow New Balance": cashflow_new_balance,
        "Currency": (
            row_map.get("currency")
            or row_map.get("account_currency")
            or row_map.get("result_currency")
            or ""
        ),
        "Row Type": row_map.get("row_type") or "trade",
        "Row ID": stable_row_id(row_map),
    }
    for header, field in MOVE_TO_FIELD_MAP.items():
        raw_value = row_map.get(field)
        if field in {"move_to_break_even_duration", "move_to_profit_duration"} and raw_value not in (None, ""):
            values[header] = _format_duration_display(raw_value)
        elif field in MOVE_TO_TIME_FIELDS and raw_value not in (None, ""):
            values[header] = _trade_log_datetime_value(raw_value)
        else:
            values[header] = raw_value or ""
    return values


def _apply_trade_log_row_number_formats(
    ws,
    row_idx: int,
    row: Mapping[str, Any],
) -> Dict[str, str]:
    """Apply full-builder number formats to one Trade Log row."""
    headers = _trade_log_header_map(ws)
    applied: Dict[str, str] = {}

    def set_format(header: str, number_format: str) -> None:
        col = headers.get(header)
        if not col:
            return
        ws.cell(row_idx, col).number_format = number_format
        applied[header] = number_format

    commission_currency = _infer_trade_log_currency(dict(row), field="commission")
    pnl_currency = _infer_trade_log_currency(dict(row), field="net_pnl")
    balance_currency = _infer_trade_log_currency(dict(row), field="balance_after")
    set_format(TRADE_NUMBER_HEADER, "@")
    set_format("Qty", "#,##0.##########")
    set_format("Open Time", "yyyy-mm-dd hh:mm:ss")
    set_format("Close Time", "yyyy-mm-dd hh:mm:ss")
    set_format("Stop Loss Distance", "0.00%")
    set_format("Target Distance", "0.00%")
    if commission_currency:
        set_format("Commission", _currency_number_format(commission_currency))
    if pnl_currency:
        set_format("Net P/L", _currency_number_format(pnl_currency))
    profit_col = headers.get("Profit %")
    if profit_col:
        set_format("Profit %", adaptive_percent_number_format(ws.cell(row_idx, profit_col).value))
    r_col = headers.get("R-Multiple")
    if r_col:
        set_format("R-Multiple", adaptive_number_format(ws.cell(row_idx, r_col).value))
    if balance_currency:
        set_format(
            "Balance After",
            "#,##0.0000000000" if _is_crypto_currency(balance_currency) else "#,##0.00",
        )
    set_format("Trade Duration (DD:HH:MM:SS)", "General")
    set_format("Move to Break Even Time", "yyyy-mm-dd hh:mm:ss")
    set_format("Move to Break Even Duration", DURATION_NUMBER_FORMAT)
    set_format("Move to Profit Time", "yyyy-mm-dd hh:mm:ss")
    set_format("Move to Profit Duration", DURATION_NUMBER_FORMAT)
    for header in (
        "Move to Break Even Distance From Entry %",
        "Move to Break Even Distance From Exit %",
        "Move to Profit Distance From Entry %",
        "Move to Profit Distance From Exit %",
    ):
        set_format(header, "0.00%")
    for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
        col = headers.get(header)
        if col:
            _apply_recommendation_cell_style(ws.cell(row_idx, col))
    return applied


def build_master_journal_workbook(
    snapshot: Dict[str, Any],
    output_path: Path,
    *,
    recommendation_workbook_path: Optional[Path] = None,
    publish_recommendation_assets: bool = True,
    manual_overrides: Optional[Mapping[str, Mapping[str, Any]]] = None,
) -> Dict[str, Any]:
    wb=Workbook(); wb.remove(wb.active)
    for s in SHEET_ORDER: wb.create_sheet(s)
    hyperlink_workbook_path = Path(recommendation_workbook_path or output_path)
    resolved_manual_overrides = manual_overrides
    if resolved_manual_overrides is None:
        resolved_manual_overrides = (
            read_master_journal_manual_overrides(hyperlink_workbook_path)
            if hyperlink_workbook_path.exists()
            else {}
        )
    rows, row_normalization_diagnostics = _normalize_master_journal_rows(
        snapshot.get('items') or [],
        resolved_manual_overrides,
    )
    metric_rows=[r for r in rows if str(r.get('row_type') or 'trade')=='trade']
    non_test=[r for r in metric_rows if not _is_test_trade_value(r.get('is_test_trade'))]
    stats = snapshot.get('stats') or {}
    totals = stats.get('totals') or {}
    groups = stats.get('groups') or {}

    dash = _stats1_sheet(wb)
    for col, width in (("A", 32), ("B", 18), ("C", 18), ("D", 18), ("E", 3), ("F", 24), ("G", 20), ("H", 12), ("I", 12), ("J", 12)):
        dash.column_dimensions[col].width = width

    by_market = groups.get("by_market") or {}
    risk = groups.get("risk_expectancy") or {}
    duration = groups.get("duration") or {}
    leaders = {
        **_instrument_leader_payloads_from_rows(metric_rows),
        **dict(groups.get("leaders") or {}),
    }
    move_duration_metrics = _trade_move_duration_metrics(metric_rows)
    percentage_totals = _result_percentage_totals_by_market(rows, snapshot.get("balances") or stats.get("balances") or [])
    extended_metrics = _dashboard_extended_metrics(rows, by_market)
    buckets = {
        market: _merge_metric_buckets(
            dict(bucket or {}),
            percentage_totals[market],
            {key: value for key, value in extended_metrics[market].items() if value is not None},
            {key: value for key, value in move_duration_metrics[market].items() if value is not None},
        )
        for market, bucket in {
            "overall": by_market.get("overall") or totals,
            "fx": by_market.get("fx") or {},
            "crypto": by_market.get("crypto") or {},
        }.items()
    }
    buckets["overall"]["most_wins_instrument"] = leaders.get("most_wins_instrument")
    buckets["overall"]["most_losses_instrument"] = leaders.get("most_losses_instrument")
    buckets["fx"]["most_wins_instrument"] = leaders.get("fx_most_wins_instrument")
    buckets["fx"]["most_losses_instrument"] = leaders.get("fx_most_losses_instrument")
    buckets["crypto"]["most_wins_instrument"] = leaders.get("crypto_most_wins_instrument")
    buckets["crypto"]["most_losses_instrument"] = leaders.get("crypto_most_losses_instrument")
    market_cols = {"overall": 2, "fx": 3, "crypto": 4}
    for market, col in market_cols.items():
        dash.cell(1, col, {"overall": "Overall", "fx": "FX", "crypto": "Crypto"}[market])

    core_rows = [
        ("Trades", "trades", "count", None),
        ("Wins", "wins", "count", "profit"),
        ("Losses", "losses", "count", "loss"),
        ("Break-even", "break_even", "count", None),
        ("Test", "test_trades", "count", None),
        ("Win rate", "win_rate_pct", "pct", None),
        ("Net P/L Percentage", "net_result_pct", "pct", "auto"),
        ("Net P/L R multiples", "net_r_multiple", "r", "auto"),
        ("Gross percent gain", "gross_gain_result_pct", "pct", "profit"),
        ("Gross percent loss", "gross_loss_result_pct", "pct", "loss"),
        ("Gross IR gain", "gross_ir_gain", "r", "profit"),
        ("Gross IR loss", "gross_ir_loss", "r", "loss"),
        ("Percentage expectancy", "avg_result_pct", "pct", "auto"),
        ("R expectancy", "avg_r_multiple", "r", "auto"),
        ("Best Win Streak", "winning_streak", "count", "profit"),
        ("Start", "start:longest_winning_streak", "streak_date", None),
        ("End", "end:longest_winning_streak", "streak_date", None),
        ("Worst Losing Streak", "losing_streak", "count", "loss"),
        ("Start", "start:longest_losing_streak", "streak_date", None),
        ("End", "end:longest_losing_streak", "streak_date", None),
        ("Avg stop %", "avg_stop_pct", "pct", None),
        ("Min stop %", "min_stop_pct", "pct", None),
        ("Max stop %", "max_stop_pct", "pct", None),
        ("Recommendation", STOP_RECOMMENDATION_HEADER, "text", None),
        ("Avg target %", "avg_target_pct", "pct", None),
        ("Min target %", "min_target_pct", "pct", None),
        ("Max target %", "max_target_pct", "pct", None),
        ("Recommendation", TARGET_RECOMMENDATION_HEADER, "text", None),
    ]

    row = 2
    for label, key, kind, semantic in core_rows:
        dash.cell(row, 1, label).font = Font(bold=True)
        for market, col in market_cols.items():
            bucket = buckets[market]
            value = bucket.get(key)
            cell = dash.cell(row, col)
            inline_source = _inline_metric_source(bucket, key)
            if inline_source and kind in {"pct", "r", "duration"}:
                cell.value = _format_inline_metric_value(value, kind, inline_source)
                cell.number_format = "General"
            elif kind == "pct":
                number = _as_float(value)
                cell.value = "" if number is None else number / 100.0
                cell.number_format = adaptive_percent_number_format(cell.value)
            elif kind == "r":
                cell.value = "" if value is None else value
                cell.number_format = '0.000"R"'
            elif kind == "source":
                source_key = str(key or "").split(":", 1)[1] if ":" in str(key or "") else str(key or "")
                source = (bucket.get("metric_sources") or {}).get(source_key)
                cell.value = _fmt_detail_src(source) if source else ""
                cell.number_format = "General"
            elif kind == "count":
                cell.value = "" if value is None else value
                cell.number_format = "0"
            elif kind == "duration":
                cell.value = _format_duration_display(value) if value is not None else ""
                cell.number_format = "General"
            elif kind == "streak_date":
                endpoint, detail_key = str(key or "").split(":", 1)
                start, end = _streak_detail_start_end(bucket.get(detail_key))
                cell.value = start if endpoint == "start" else end
                cell.number_format = "General"
                _apply_streak_detail_value_font(cell)
            elif kind == "number":
                cell.value = "" if value is None else value
                cell.number_format = "#,##0.##########"
            elif kind == "text":
                if key == "most_wins_instrument":
                    cell.value = "" if value is None else _instrument_leader_scalar(value, "wins")
                elif key == "most_losses_instrument":
                    cell.value = "" if value is None else _instrument_leader_scalar(value, "losses")
                else:
                    cell.value = "" if value is None else _excel_scalar(value)
            elif kind == "money":
                money_map = (bucket.get("money_by_currency") or {}).get(key) or {}
                if isinstance(money_map, dict) and len(money_map) == 1:
                    currency, amount = next(iter(money_map.items()))
                    cell.value = amount
                    cell.number_format = _currency_number_format(currency)
                else:
                    cell.value = "" if value is None else _excel_scalar(value)
            if semantic == "profit":
                _apply_full_cell_semantic_fill(cell, "profit")
            elif semantic == "loss":
                _apply_full_cell_semantic_fill(cell, "loss")
            elif semantic == "auto":
                _apply_sign_based_full_cell_fill(cell)
        row += 1

    section_specs = [
        ("Duration", [
            ("Min duration", "min_duration_seconds", "duration", None),
            ("Avg duration", "avg_duration_seconds", "duration", None),
            ("Avg winning trade duration", "avg_duration_seconds_winners", "duration", None),
            ("Avg losing trade duration", "avg_duration_seconds_losers", "duration", None),
            ("Max duration", "max_duration_seconds", "duration", None),
            ("Min Move to Break Even", "min_move_to_break_even_duration_seconds", "duration", None),
            ("Average Move to Break Even", "avg_move_to_break_even_duration_seconds", "duration", None),
            ("Max Move to Break Even", "max_move_to_break_even_duration_seconds", "duration", None),
            ("Min Move to Profit", "min_move_to_profit_duration_seconds", "duration", None),
            ("Average Move to Profit", "avg_move_to_profit_duration_seconds", "duration", None),
            ("Max Move to Profit", "max_move_to_profit_duration_seconds", "duration", None),
        ], False),
        ("Winners", [
            ("Min stop %", "min_stop_pct_winners", "pct", None),
            ("Avg stop %", "avg_stop_pct_winners", "pct", None),
            ("Max stop %", "max_stop_pct_winners", "pct", None),
            ("Min target %", "min_target_pct_winners", "pct", None),
            ("Avg target %", "avg_target_pct_winners", "pct", None),
            ("Max target %", "max_target_pct_winners", "pct", None),
            ("Min duration", "min_duration_seconds_winners", "duration", None),
            ("Avg duration", "avg_duration_seconds_winners", "duration", None),
            ("Max duration", "max_duration_seconds_winners", "duration", None),
            ("Min win %", "min_result_pct_winners", "pct", "profit"),
            ("Avg win %", "avg_result_pct_winners", "pct", "profit"),
            ("Max win %", "max_result_pct_winners", "pct", "profit"),
            ("Min R win", "min_r_multiple_winners", "r", "profit"),
            ("Avg R win", "avg_r_multiple_winners", "r", "profit"),
            ("Max R win", "max_r_multiple_winners", "r", "profit"),
            ("Most wins", "most_wins_instrument", "text", "profit"),
        ], True),
        ("Losers", [
            ("Min stop %", "min_stop_pct_losers", "pct", None),
            ("Avg stop %", "avg_stop_pct_losers", "pct", None),
            ("Max stop %", "max_stop_pct_losers", "pct", None),
            ("Min target %", "min_target_pct_losers", "pct", None),
            ("Avg target %", "avg_target_pct_losers", "pct", None),
            ("Max target %", "max_target_pct_losers", "pct", None),
            ("Min duration", "min_duration_seconds_losers", "duration", None),
            ("Avg duration", "avg_duration_seconds_losers", "duration", None),
            ("Max duration", "max_duration_seconds_losers", "duration", None),
            ("Max loss %", "min_result_pct_losers", "pct", "loss"),
            ("Avg loss %", "avg_result_pct_losers", "pct", "loss"),
            ("Min loss %", "max_result_pct_losers", "pct", "loss"),
            ("Max R loss", "min_r_multiple_losers", "r", "loss"),
            ("Avg R loss", "avg_r_multiple_losers", "r", "loss"),
            ("Min R loss", "max_r_multiple_losers", "r", "loss"),
            ("Most losses", "most_losses_instrument", "text", "loss"),
        ], True),
        ("Side", [
            ("Long", "long_trades", "count", None),
            ("Winners", "long_wins", "count", "profit"),
            ("Losers", "long_losses", "count", "loss"),
            ("Short", "short_trades", "count", None),
            ("Winners", "short_wins", "count", "profit"),
            ("Losers", "short_losses", "count", "loss"),
        ], True),
        ("Patterns", [
            ("Channel", "pattern_channel_total", "count", None),
            ("Winners", "pattern_channel_wins", "count", "profit"),
            ("Losers", "pattern_channel_losses", "count", "loss"),
            ("Range", "pattern_range_total", "count", None),
            ("Winners", "pattern_range_wins", "count", "profit"),
            ("Losers", "pattern_range_losses", "count", "loss"),
            ("Most Traded", "most_traded_pattern", "text", None),
            ("Least Traded", "least_traded_pattern", "text", None),
            ("Most Profitable", "most_profitable_pattern", "text", None),
            ("Least Profitable", "least_profitable_pattern", "text", None),
        ], True),
        ("Timeframe", [
            ("1MIN", "timeframe_1min", "count", None),
            ("Winners", "timeframe_1min_wins", "count", "profit"),
            ("Losers", "timeframe_1min_losses", "count", "loss"),
            ("5MIN", "timeframe_5min", "count", None),
            ("Winners", "timeframe_5min_wins", "count", "profit"),
            ("Losers", "timeframe_5min_losses", "count", "loss"),
            ("15MIN", "timeframe_15min", "count", None),
            ("Winners", "timeframe_15min_wins", "count", "profit"),
            ("Losers", "timeframe_15min_losses", "count", "loss"),
            ("30MIN", "timeframe_30min", "count", None),
            ("Winners", "timeframe_30min_wins", "count", "profit"),
            ("Losers", "timeframe_30min_losses", "count", "loss"),
            ("1H", "timeframe_1h", "count", None),
            ("Winners", "timeframe_1h_wins", "count", "profit"),
            ("Losers", "timeframe_1h_losses", "count", "loss"),
            ("4H", "timeframe_4h", "count", None),
            ("Winners", "timeframe_4h_wins", "count", "profit"),
            ("Losers", "timeframe_4h_losses", "count", "loss"),
            ("DAILY", "timeframe_daily", "count", None),
            ("Winners", "timeframe_daily_wins", "count", "profit"),
            ("Losers", "timeframe_daily_losses", "count", "loss"),
            ("WEEKLY", "timeframe_weekly", "count", None),
            ("Winners", "timeframe_weekly_wins", "count", "profit"),
            ("Losers", "timeframe_weekly_losses", "count", "loss"),
            ("MONTHLY", "timeframe_monthly", "count", None),
            ("Winners", "timeframe_monthly_wins", "count", "profit"),
            ("Losers", "timeframe_monthly_losses", "count", "loss"),
        ], True),
        ("Commission", [
            ("Min Commission", "min_commission", "number", None),
            ("Avg Commission", "avg_commission", "number", None),
            ("Max Commission", "max_commission", "number", None),
            ("Total Commission", "total_commission", "number", None),
        ], True),
        ("Drawdown", [
            ("Min drawdown", "min_drawdown_pct", "pct", "loss"),
            ("Start", "start:min_drawdown_detail", "drawdown_date", None),
            ("End", "end:min_drawdown_detail", "drawdown_date", None),
            ("Avg drawdown", "avg_drawdown_pct", "pct", "loss"),
            ("Max drawdown", "max_drawdown_pct", "pct", "loss"),
            ("Start", "start:max_drawdown_detail", "drawdown_date", None),
            ("End", "end:max_drawdown_detail", "drawdown_date", None),
        ], True),
    ]
    for title, metrics, show_header in section_specs:
        if show_header:
            dash.cell(row, 1, title).font = Font(bold=True)
            for col in market_cols.values():
                dash.cell(row, col).fill = PatternFill("solid", fgColor="EAF2F8")
            row += 1
        for label, key, kind, semantic in metrics:
            dash.cell(row, 1, label).font = Font(bold=True)
            for market, col in market_cols.items():
                value = None if title == "Commission" and market == "overall" else buckets[market].get(key)
                cell = dash.cell(row, col)
                inline_source = _inline_metric_source(buckets[market], key)
                if inline_source and kind in {"pct", "r", "duration"}:
                    cell.value = _format_inline_metric_value(value, kind, inline_source)
                    cell.number_format = "General"
                elif kind == "pct":
                    number = _as_float(value)
                    cell.value = "" if number is None else number / 100.0
                    cell.number_format = adaptive_percent_number_format(cell.value)
                elif kind == "r":
                    cell.value = "" if value is None else value
                    cell.number_format = '0.000"R"'
                elif kind == "source":
                    source_key = str(key or "").split(":", 1)[1] if ":" in str(key or "") else str(key or "")
                    source = (buckets[market].get("metric_sources") or {}).get(source_key)
                    cell.value = _fmt_detail_src(source) if source else ""
                    cell.number_format = "General"
                elif kind == "duration":
                    cell.value = _format_duration_display(value) if value is not None else ""
                    cell.number_format = "General"
                elif kind == "drawdown_date":
                    endpoint, detail_key = str(key or "").split(":", 1)
                    cell.value = _drawdown_detail_cell_value(buckets[market].get(detail_key), endpoint)
                    cell.number_format = "General"
                elif kind == "count":
                    cell.value = "" if value is None else int(value)
                    cell.number_format = "0"
                elif kind == "number":
                    cell.value = "" if value is None else value
                    if title == "Commission" and market in {"fx", "crypto"}:
                        currency = "AUD" if market == "fx" else "USDT"
                        cell.number_format = _currency_number_format(currency)
                        _clear_generated_semantic_fill(cell)
                    else:
                        cell.number_format = "#,##0.##########"
                else:
                    cell.value = "" if value is None else _excel_scalar(value)
                if semantic in {"profit", "loss"} and cell.value not in (None, ""):
                    _apply_full_cell_semantic_fill(cell, semantic)
            row += 1

    _style_header_row(dash, 1)
    _table_border(dash, 1, 1, row - 1, 4)
    labels_by_name = {str(dash.cell(rr, 1).value or "").strip(): rr for rr in range(1, row)}
    for label in ("Percentage expectancy", "R expectancy"):
        metric_row = labels_by_name.get(label)
        if metric_row:
            _profit_loss_rules(dash, f"B{metric_row}:D{metric_row}")

    detail = _stats2_sheet(wb, required=True)
    balances = _canonicalize_and_dedupe_balances(snapshot.get("balances") or stats.get("balances") or [])
    _clear_account_balance_source_metadata(wb)
    detail.cell(1, 1, "Account Balances").font = Font(bold=True)
    stats2_headers = ("Account", "Balance", "Currency", "Risk of Ruin", "Net P/L Percentage")
    account_net_pct = _net_result_pct_by_account(rows)
    for col, header in enumerate(stats2_headers, start=1):
        detail.cell(2, col, header).font = Font(bold=True)
    risk_of_ruin = _risk_of_ruin_by_account(rows)
    for target_row, rec in enumerate(balances, start=3):
        if not isinstance(rec, dict):
            continue
        currency = _currency_code(rec.get("currency"), rec.get("account_currency"))
        account_label = _canonical_account_label(rec.get("account_label") or rec.get("account") or rec.get("source") or "")
        detail.cell(target_row, 1, account_label)
        detail.cell(target_row, 2, _as_float(rec.get("balance")))
        _write_account_balance_source_metadata(
            wb,
            account_label,
            rec.get("balance_source") or rec.get("source"),
            rec.get("as_of"),
        )
        detail.cell(target_row, 2).number_format = "#,##0.0000000000" if _is_crypto_currency(currency) else "#,##0.00"
        detail.cell(target_row, 3, currency)
        risk_payload = risk_of_ruin.get(account_label) or _empty_risk_of_ruin_payload("no_usable_trade_history")
        detail.cell(target_row, 4, risk_payload.get("risk_of_ruin"))
        detail.cell(target_row, 4).number_format = "0.00%"
        detail.cell(target_row, 4).comment = Comment(_risk_of_ruin_comment_text(risk_payload), "Codex")
        net_pct = account_net_pct.get(account_label)
        if net_pct is not None and _stats2_net_pl_percentage_is_applicable(account_label):
            detail.cell(target_row, 5, net_pct / 100.0)
            detail.cell(target_row, 5).number_format = adaptive_percent_number_format(detail.cell(target_row, 5).value)

    _apply_dashboard_requested_semantic_fills(dash)

    resolved_balances = _resolved_all_trade_balances(rows)
    recommendations_by_symbol = _distance_recommendations_by_symbol(rows)
    ws=_get_all_trades_sheet(wb); headers=TRADE_LOG_HEADERS; ws.append(headers)
    for i, row in enumerate(rows):
        resolved_balance = resolved_balances.get(str(i))
        values = _trade_log_row_values(
            row,
            resolved_balance=resolved_balance,
            recommendations_by_symbol=recommendations_by_symbol,
        )
        ws.append([values.get(header, '') for header in TRADE_LOG_HEADERS])
    _style_table_sheet(ws,1,'A2',True)
    for rr in range(2, ws.max_row + 1):
        row_ctx = rows[rr - 2] if rr - 2 < len(rows) else {}
        _apply_trade_log_row_number_formats(ws, rr, row_ctx)
    _ensure_trade_log_schema(ws)
    trade_cols = _trade_log_header_map(ws)
    last_trade_row = max(TRADE_LOG_DATA_START_ROW, ws.max_row)
    commission_col = trade_cols.get("Commission")
    net_col = trade_cols.get("Net P/L")
    r_col = trade_cols.get("R-Multiple")
    if commission_col:
        letter = get_column_letter(commission_col)
        _negative_impact_rule(ws, f"{letter}{TRADE_LOG_DATA_START_ROW}:{letter}{last_trade_row}")
    if net_col and r_col:
        _profit_loss_rules(ws, f"{get_column_letter(net_col)}{TRADE_LOG_DATA_START_ROW}:{get_column_letter(r_col)}{last_trade_row}")
    _apply_trade_log_win_loss_row_formatting(ws)
    _apply_trade_log_win_loss_direct_row_fills(ws)

    inst=_symbols_sheet(wb)
    _write_instrument_averages_headers(inst)
    instrument_analysis = _instrument_analysis_by_symbol(rows)
    instrument_rows = stats.get('by_instrument') or _instrument_summary_rows_from_trade_rows(rows)
    next_instrument_row = INSTRUMENT_AVERAGES_DATA_START_ROW
    for rec in instrument_rows:
        cls=str(rec.get("asset_class") or rec.get("class") or "").lower()
        analysis = instrument_analysis.get(str(rec.get("symbol") or "").strip().upper(), {})
        row_idx = next_instrument_row
        netp = _as_float(analysis.get("net_result_pct"))
        avgp = _as_float(analysis.get("avg_result_pct"))
        if netp is None:
            netp = _as_float(rec.get("net_result_pct"))
        if avgp is None:
            avgp = _as_float(rec.get("avg_result_pct"))
        instrument_values = [
            rec.get("symbol"), cls.upper() if cls else None, rec.get("total_trades", rec.get("trades")),
            rec.get("wins"), rec.get("losses"), rec.get("break_even"),
            rec.get("long_trades", rec.get("longs")), rec.get("long_wins"), rec.get("long_losses"), rec.get("long_break_even"),
            rec.get("short_trades", rec.get("shorts")), rec.get("short_wins"), rec.get("short_losses"), rec.get("short_break_even"),
            analysis.get("move_to_break_even"),
            analysis.get("move_to_profit"),
            analysis.get("pattern"), analysis.get("ema"), analysis.get("all_time_highs"), analysis.get("all_time_lows"),
            analysis.get("market_orders"), analysis.get("limit_orders"),
            analysis.get("round_number"), analysis.get("spiked_out"), analysis.get("close_stop_out"),
            analysis.get("near_perfect_entry"), analysis.get("near_win"), analysis.get("early_close"),
            analysis.get("most_traded_timeframe"),
            analysis.get("most_profitable_timeframe"), analysis.get("least_profitable_timeframe"),
            analysis.get("net_r_multiple"),
            (netp/100.0 if netp is not None else ''), (avgp/100.0 if avgp is not None else ''),
            rec.get("win_rate_pct"),
            analysis.get("avg_stop_pct") if analysis.get("avg_stop_pct") is not None else rec.get("avg_sl_pct"),
            rec.get('avg_sl_pct_wins'), rec.get('avg_sl_pct_losses'),
            _sanitize_recommendation_text(STOP_RECOMMENDATION_HEADER, analysis.get(STOP_RECOMMENDATION_HEADER) or ""),
            analysis.get("avg_target_pct") if analysis.get("avg_target_pct") is not None else rec.get("avg_tp_pct"),
            rec.get('avg_tp_pct_wins'), rec.get('avg_tp_pct_losses'),
            _sanitize_recommendation_text(TARGET_RECOMMENDATION_HEADER, analysis.get(TARGET_RECOMMENDATION_HEADER) or ""),
            _format_duration_display(rec.get("min_trade_duration_seconds", rec.get("shortest_duration_seconds"))),
            _format_duration_display(rec.get("avg_trade_duration_seconds", rec.get("avg_duration_seconds"))),
            _format_duration_display(rec.get("max_trade_duration_seconds", rec.get("longest_duration_seconds"))),
            _format_duration_display(rec.get("avg_winning_trade_duration_seconds", rec.get("avg_winning_duration"))),
            _format_duration_display(rec.get("avg_losing_trade_duration_seconds", rec.get("avg_losing_duration"))),
        ]
        for col_idx, value in enumerate(instrument_values, start=1):
            inst.cell(row_idx, col_idx).value = value
        next_instrument_row += 1
        header_cols = {header: index + 1 for index, header in enumerate(INSTRUMENT_AVERAGES_HEADERS)}
        for header in ("Win Rate %", "Avg stop %", "Avg stop % (W)", "Avg stop % (L)", "Avg target %", "Avg target % (W)", "Avg target % (L)"):
            cc = header_cols[header]
            cell = inst.cell(row_idx, cc)
            val = _as_float(cell.value)
            if val is not None:
                cell.value = val / 100.0
                cell.number_format = "0.00%"
        for header in (STOP_RECOMMENDATION_HEADER, TARGET_RECOMMENDATION_HEADER):
            _apply_recommendation_cell_style(inst.cell(row_idx, header_cols[header]))
        for zc in [
            4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14,
            header_cols["All-time highs"], header_cols["All-time lows"],
            header_cols["Market"], header_cols["Limit"], header_cols["Round number"],
            header_cols["Spiked out"], header_cols["Close stop out"], header_cols["Near perfect entry"],
            header_cols["Near win"], header_cols["Early close"],
        ]:
            inst.cell(row_idx, zc).number_format = ZERO_HIDE_FORMAT
        inst.cell(row_idx, header_cols["Net R Multiple"]).number_format = '0.000"R"'
        inst.cell(row_idx, header_cols["Net P/L %"]).number_format = "0.00%"
        inst.cell(row_idx, header_cols["Avg P/L %"]).number_format = "0.00%"
        for col in (
            header_cols["Shortest duration (DD:HH:MM:SS)"],
            header_cols["Avg duration (DD:HH:MM:SS)"],
            header_cols["Longest duration (DD:HH:MM:SS)"],
            header_cols[AVG_WINNING_DURATION_HEADER],
            header_cols[AVG_LOSING_DURATION_HEADER],
        ):
            inst.cell(row_idx, col).number_format = "General"
        for col in (header_cols["Move to break even"], header_cols["Move to profit"]):
            inst.cell(row_idx, col).number_format = ZERO_HIDE_FORMAT
    _style_table_sheet(inst, INSTRUMENT_AVERAGES_FILTER_HEADER_ROW, 'B3', True)
    _style_header_row(inst, INSTRUMENT_AVERAGES_GROUP_HEADER_ROW)
    _apply_instrument_averages_requested_style(inst)
    _apply_instrument_averages_profit_loss_formatting(inst)
    _apply_instrument_averages_semantic_fills(inst)
    _repair_instrument_timeframe_columns(inst)
    _populate_symbols_metrics_preserving_layout(inst, rows)
    _normalize_symbols_performance_formatting(inst)

    cal = wb["P&L Calendar"]
    _write_pnl_calendar_one_line_layout(cal, {"items": non_test})
    _ensure_pnl_calendar_freeze_panes(cal)
    _ensure_report_sheets(wb, snapshot)
    _remove_target_r_metadata_sheet(wb)
    _repair_stats1_formatting(dash, extended_metrics)
    _repair_legacy_duration_number_formats(wb)

    if STATS2_SHEET in wb.sheetnames:
        _repair_stats2_account_balance_formatting(wb[STATS2_SHEET])
    _apply_workbook_left_alignment(wb)
    _repair_stats1_child_label_styles(dash)
    _repair_stats2_as_of_datetime_style(wb)
    _normalize_generated_statistics_row_heights(wb)
    _normalize_generated_duration_text(wb)
    _normalize_trade_log_display_durations(_get_trade_log_sheet(wb, allow_legacy=False))
    chart_bundle = _prepare_recommendation_chart_bundle(
        rows,
        hyperlink_workbook_path,
    )
    _apply_recommendation_chart_hyperlinks(wb, chart_bundle)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    if publish_recommendation_assets:
        _publish_recommendation_chart_bundle(chart_bundle)
    return {
        'ok': True,
        'path': str(output_path),
        'diagnostics': {
            **row_normalization_diagnostics,
            'recommendation_chart_assets_deferred': not publish_recommendation_assets,
        },
    }

def _table_border(ws, top_row, left_col, bottom_row, right_col):
    thin=Side(style='thin', color='D1D5DB')
    thick=Side(style='thick', color='D1D5DB')
    for r in range(top_row,bottom_row+1):
        for c in range(left_col,right_col+1):
            ws.cell(r,c).border=Border(
                left=thick if c==left_col else thin,
                right=thick if c==right_col else thin,
                top=thick if r==top_row else thin,
                bottom=thick if r==bottom_row else thin,
            )


def _clear_borders_in_range(ws, top_row: int, left_col: int, bottom_row: int, right_col: int) -> None:
    empty = Border()
    for row in range(top_row, bottom_row + 1):
        for col in range(left_col, right_col + 1):
            ws.cell(row, col).border = empty


def _apply_group_box_border(
    ws,
    top_row: int,
    left_col: int,
    bottom_row: int,
    right_col: int,
    *,
    outer_style: str = "medium",
) -> None:
    if top_row > bottom_row or left_col > right_col:
        return
    inner = Side(style="thin", color="D1D5DB")
    outer = Side(style=outer_style, color="808080")
    for row in range(top_row, bottom_row + 1):
        for col in range(left_col, right_col + 1):
            ws.cell(row, col).border = Border(
                left=outer if col == left_col else inner,
                right=outer if col == right_col else inner,
                top=outer if row == top_row else inner,
                bottom=outer if row == bottom_row else inner,
            )


def _last_nonblank_row_in_columns(ws, columns: List[int]) -> int:
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row, col).value not in (None, "") for col in columns):
            return row
    return 1


def _stats1_label_at(ws, row: int) -> str:
    return str(ws.cell(row, 1).value or "").strip()


def _find_stats1_first_label(ws, label: str, *, after: int = 0, before: int | None = None) -> int | None:
    wanted = label.casefold()
    last = before if before is not None else ws.max_row
    return next(
        (
            row for row in range(max(1, after + 1), min(ws.max_row, last) + 1)
            if _stats1_label_at(ws, row).casefold() == wanted
        ),
        None,
    )


def _apply_stats1_clean_table_borders(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols = _stats1_market_columns(ws)
    right_col = max(market_cols.values(), default=4)
    right_col = max(4, min(right_col, ws.max_column))
    last_row = _last_nonblank_row_in_columns(ws, list(range(1, right_col + 1)))
    if last_row < 1:
        return

    generated_merge_labels = {
        *(label.casefold() for label in REPORT_METRIC_LABELS),
        "trades",
        "wins",
        "losses",
        "break-even",
        "test",
        "win rate",
        "net p/l percentage",
        "net p/l r multiples",
        "gross percent gain",
        "gross percent loss",
        "gross ir gain",
        "gross ir loss",
        "percentage expectancy",
        "r expectancy",
        "best win streak",
        "worst losing streak",
        "start",
        "end",
        "stops",
        "targets",
        "duration",
        "side",
        "patterns",
        "timeframe",
        "commission",
        "drawdown",
        "source",
        "recommendation",
        "avg stop %",
        "min stop %",
        "max stop %",
        "avg target %",
        "min target %",
        "max target %",
        "min duration",
        "avg duration",
        "max duration",
        "average",
        "move to break even",
        "move to profit",
        "average move to break even",
        "average move to profit",
        "min win %",
        "avg win %",
        "max win %",
        "min r win",
        "avg r win",
        "max r win",
        "most wins",
        "max loss %",
        "avg loss %",
        "min loss %",
        "max r loss",
        "avg r loss",
        "min r loss",
        "most losses",
        "long",
        "short",
        "channel",
        "range",
        "1min",
        "5min",
        "15min",
        "30min",
        "1h",
        "4h",
        "daily",
        "weekly",
        "monthly",
        "min commission",
        "avg commission",
        "max commission",
        "total commission",
    }
    unmerged = []
    for merged in list(ws.merged_cells.ranges):
        anchor_label = _stats1_label_at(ws, merged.min_row).casefold()
        if (
            merged.min_row <= last_row
            and merged.max_row >= 1
            and merged.min_col <= right_col
            and merged.max_col >= 1
            and anchor_label in generated_merge_labels
        ):
            ws.unmerge_cells(str(merged))
            unmerged.append(str(merged))
    if unmerged:
        diagnostics["unmerged_stats1_layout_blocks"] = unmerged

    source_parent_labels = {
        "min stop %",
        "max stop %",
        "min target %",
        "max target %",
        "min move to break even",
        "max move to break even",
        "min move to profit",
        "max move to profit",
    }
    restored_source_labels = 0
    for row in range(1, last_row):
        if _stats1_label_at(ws, row).casefold() not in source_parent_labels:
            continue
        if _stats1_label_at(ws, row + 1):
            continue
        if not any(ws.cell(row + 1, col).value not in (None, "") for col in market_cols.values()):
            continue
        ws.cell(row + 1, 1).value = "Source"
        _apply_dashboard_source_label_style(ws.cell(row + 1, 1))
        restored_source_labels += 1
    if restored_source_labels:
        diagnostics["restored_stats1_source_labels_after_unmerge"] = restored_source_labels

    manual_move_label_borders: Dict[int, Border] = {}
    for row in range(1, last_row + 1):
        label = _stats1_label_at(ws, row).casefold()
        if label not in {
            DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL.casefold(),
            DASHBOARD_MOVE_TO_PROFIT_LABEL.casefold(),
            "move to break even (dd:hh:mm:ss)",
            "move to profit (dd:hh:mm:ss)",
        }:
            continue
        border = ws.cell(row, 1).border
        if any(getattr(side, "style", None) for side in (border.left, border.right, border.top, border.bottom)):
            manual_move_label_borders[row] = copy(border)

    _clear_borders_in_range(ws, 1, 1, last_row, right_col)

    static_section_labels = {
        "stops",
        "targets",
        "duration",
        "drawdown",
        "side",
        "patterns",
        "timeframe",
        "commission",
    }
    section_rows = [
        row for row in range(1, last_row + 1)
        if _stats1_label_at(ws, row).casefold() in static_section_labels
    ]

    ranges: List[Tuple[int, int]] = []
    first_section = min(section_rows) if section_rows else last_row + 1
    r_expectancy = _find_stats1_first_label(ws, "R expectancy", before=first_section - 1)
    if r_expectancy:
        ranges.append((1, r_expectancy))
    cursor = r_expectancy or 1
    best = _find_stats1_first_label(ws, "Best Win Streak", after=cursor - 1, before=first_section - 1)
    worst = _find_stats1_first_label(ws, "Worst Losing Streak", after=cursor - 1, before=first_section - 1)
    if best:
        best_end = best + 2 if best + 2 <= last_row and _stats1_label_at(ws, best + 1).casefold() == "start" else best
        ranges.append((best, min(best_end, first_section - 1)))
        cursor = max(cursor, min(best_end, first_section - 1))
    if worst:
        worst_end = worst + 2 if worst + 2 <= last_row and _stats1_label_at(ws, worst + 1).casefold() == "start" else worst
        ranges.append((worst, min(worst_end, first_section - 1)))
        cursor = max(cursor, min(worst_end, first_section - 1))

    stop_start = _find_stats1_first_label(ws, "Avg stop %", after=cursor, before=first_section - 1)
    target_start = _find_stats1_first_label(ws, "Avg target %", after=stop_start or cursor, before=first_section - 1)
    if stop_start:
        ranges.append((stop_start, (target_start - 1) if target_start else first_section - 1))
    if target_start:
        ranges.append((target_start, first_section - 1))

    side_row = _find_stats1_first_label(ws, "Side") or last_row + 1
    duration_row = _find_stats1_first_label(ws, "Duration") or 0
    winners_section = _find_stats1_first_label(ws, "Winners", after=duration_row, before=side_row - 1)
    losers_section = _find_stats1_first_label(ws, "Losers", after=winners_section or duration_row, before=side_row - 1)
    for row in (winners_section, losers_section):
        if row and row not in section_rows:
            section_rows.append(row)
    section_rows = sorted(set(section_rows))

    for index, start in enumerate(section_rows):
        end = (section_rows[index + 1] - 1) if index + 1 < len(section_rows) else last_row
        ranges.append((start, end))
        for col in range(1, right_col + 1):
            cell = ws.cell(start, col)
            font = copy(cell.font)
            font.bold = True
            font.color = "FF000000"
            cell.font = font
            # Section rows are intentionally neutral.  A genuine no-fill is
            # required so the generator cannot reintroduce shifted grey fills.
            cell.fill = PatternFill()

    normalized_ranges: List[Tuple[int, int]] = []
    for start, end in ranges:
        start = max(1, start)
        end = min(last_row, end)
        if start > end:
            continue
        if normalized_ranges and start <= normalized_ranges[-1][1]:
            normalized_ranges[-1] = (normalized_ranges[-1][0], max(normalized_ranges[-1][1], end))
        else:
            normalized_ranges.append((start, end))

    for start, end in normalized_ranges:
        _apply_group_box_border(ws, start, 1, end, right_col, outer_style="medium")
    for row, border in manual_move_label_borders.items():
        ws.cell(row, 1).border = copy(border)
    diagnostics["repaired_stats1_table_borders"] = [f"A{start}:{get_column_letter(right_col)}{end}" for start, end in normalized_ranges]


def _apply_stats2_clean_table_borders(
    ws,
    section: Dict[str, int],
    header_row: int,
    col_map: Dict[str, int],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    account_col = col_map.get("account")
    if not account_col:
        return
    left_col = section.get("start_col", 1)
    right_col = max(col_map.values())
    last_row = header_row
    for row in range(header_row + 1, min(ws.max_row, section.get("end_row", ws.max_row)) + 1):
        if str(ws.cell(row, account_col).value or "").strip():
            last_row = row
    top_row = max(1, section.get("anchor_row", 1))
    _clear_borders_in_range(ws, top_row, left_col, max(last_row, header_row), right_col)
    for row in range(top_row, max(last_row, header_row) + 1):
        for col in range(left_col, right_col + 1):
            cell = ws.cell(row, col)
            cell.fill = PatternFill()
            font = copy(cell.font)
            font.color = "FF000000"
            font.bold = row in {top_row, header_row}
            cell.font = font
    _apply_group_box_border(ws, top_row, left_col, max(last_row, header_row), right_col, outer_style="medium")
    diagnostics["repaired_stats2_table_borders"] = (
        f"{get_column_letter(left_col)}{top_row}:{get_column_letter(right_col)}{max(last_row, header_row)}"
    )


def _write_stat_section(ws, start_row, start_col, title, rows, use_detail_col=False, apply_semantic_cf=False):
    right_col = start_col + 1
    ws.merge_cells(start_row=start_row,start_column=start_col,end_row=start_row,end_column=right_col)
    h=ws.cell(start_row,start_col,title); h.font=Font(bold=True,color='00000000'); h.fill=PatternFill('solid',fgColor='00EAF2F8')
    r=start_row+1
    for row in rows:
        label,val,sem,kind,money_key,detail_text,money_map = (list(row)+[None]*7)[:7]
        ws.cell(r,start_col,_excel_scalar(label)).font=Font(color='00000000', bold=True)
        vcell=ws.cell(r,start_col+1)
        if kind=='pct':
            x = _as_float(val); vcell.value = '' if x is None else x/100.0; vcell.number_format="0.00%"
        elif kind=='r':
            x = _as_float(val); vcell.value = '' if x is None else x; vcell.number_format='0.000"R"'
        elif kind=='count':
            vcell.value = _as_float(val) if val is not None else '—'; vcell.number_format='0'
        elif kind=='money':
            mm=(money_map or {}).get(money_key or '') if isinstance(money_map,dict) else {}
            if isinstance(mm, dict) and len(mm)==1:
                ccy = list(mm.keys())[0]; vcell.value=list(mm.values())[0]; vcell.number_format=f'"{ccy}" #,##0.00;[Red]-"{ccy}" #,##0.00'
            elif isinstance(mm, dict) and len(mm)>1:
                vcell.value=' / '.join(f"{k} {float(v):.2f}" for k,v in sorted(mm.items()))
            else:
                vcell.value = _as_float(val) if _as_float(val) is not None else '—'
        elif kind=='duration':
            vcell.value = _format_duration_display(val)
            vcell.number_format = "General"
        else:
            vcell.value = '—' if val is None else val
        if apply_semantic_cf and isinstance(vcell.value, (int, float)):
            if kind == "count":
                pass
            elif sem == "auto":
                _profit_loss_rules(ws, f"{vcell.coordinate}:{vcell.coordinate}")
            elif sem in {"loss", "drawdown"}:
                _negative_impact_rule(ws, f"{vcell.coordinate}:{vcell.coordinate}")
            elif sem == "profit" and kind in {"pct", "r", "money"}:
                _profit_loss_rules(ws, f"{vcell.coordinate}:{vcell.coordinate}")
        if detail_text not in (None, "", "—"):
            r += 1
            ws.cell(r, start_col, "Source").font = Font(bold=True)
            ws.cell(r, start_col+1, detail_text)
        r+=1
    _table_border(ws,start_row,start_col,r-1,right_col)
    return r - start_row


def _style_header_row(ws, row=1):
    fill=PatternFill('solid', fgColor='E5E7EB')
    thin=Side(style='thin', color='D1D5DB')
    for c in ws[row]:
        c.font=Font(bold=True)
        c.fill=fill
        c.border=Border(left=thin,right=thin,top=thin,bottom=thin)

def _profit_loss_rules(ws, cell_range: str):
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', fgColor=PROFIT_FILL), font=Font(color=PROFIT_FONT)))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', fgColor=LOSS_FILL), font=Font(color=LOSS_FONT)))


def _negative_impact_rule(ws, cell_range: str):
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='notEqual', formula=['0'], fill=PatternFill('solid', fgColor=LOSS_FILL), font=Font(color=LOSS_FONT)))


def _style_table_sheet(ws, header_row=1, freeze='A2', autofilter=True):
    _style_header_row(ws, header_row)
    ws.freeze_panes=freeze
    if autofilter:
        ws.auto_filter.ref=f"A{header_row}:{get_column_letter(ws.max_column)}{max(header_row+1,ws.max_row)}"
    _table_border(ws, header_row, 1, ws.max_row, ws.max_column)
    for cell in ws[header_row]:
        cell.font = Font(name='Calibri', size=11, bold=True, color='00000000')
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 15


_REPORT_SECTION_ROWS = {"Winners", "Losers", "Drawdown", "Longs", "Shorts", "Patterns", "Timeframe", "Commission"}
_REPORT_SPECS = [
    ("Trades", "trades", "count", None),
    ("Wins", "wins", "count", "profit"),
    ("Losses", "losses", "count", "loss"),
    ("Break-even", "break_even", "count", None),
    ("Test", "test_trades", "count", None),
    ("Win rate", "win_rate_pct", "pct", None),
    ("Net P/L", "net_result_pct", "pct", "auto"),
    ("Gross percent gain", "gross_gain_result_pct", "pct", "profit"),
    ("Gross percent loss", "gross_loss_result_pct", "pct", "loss"),
    ("Gross IR gain", "gross_ir_gain", "r", "profit"),
    ("Gross IR loss", "gross_ir_loss", "r", "loss"),
    ("Best Win Streak", "winning_streak", "count", "profit"),
    ("Start", "start:longest_winning_streak", "streak_date", None),
    ("End", "end:longest_winning_streak", "streak_date", None),
    ("Worst Losing Streak", "losing_streak", "count", "loss"),
    ("Start", "start:longest_losing_streak", "streak_date", None),
    ("End", "end:longest_losing_streak", "streak_date", None),
    ("Percentage expectancy", "avg_result_pct", "pct", "auto"),
    ("R expectancy", "avg_r_multiple", "r", "auto"),
    ("Avg stop %", "avg_stop_pct", "pct", None),
    ("Avg target %", "avg_target_pct", "pct", None),
    ("Min stop %", "min_stop_pct", "pct", None),
    ("Max stop %", "max_stop_pct", "pct", None),
    ("Min target %", "min_target_pct", "pct", None),
    ("Max target %", "max_target_pct", "pct", None),
    ("Avg duration (DD:HH:MM:SS)", "avg_duration_seconds", "duration", None),
    ("Move to Break Even (DD:HH:MM:SS)", "move_to_break_even_duration_seconds", "duration", None),
    ("Move to Profit (DD:HH:MM:SS)", "move_to_profit_duration_seconds", "duration", None),
    ("Max win %", "max_result_pct", "pct", "profit"),
    ("Max loss %", "min_result_pct", "pct", "loss"),
    ("Max R loss", "min_r_multiple", "r", "loss"),
    ("Max R win", "max_r_multiple", "r", "profit"),
    ("Shortest (DD:HH:MM:SS)", "shortest_duration_seconds", "duration", None),
    ("Longest (DD:HH:MM:SS)", "longest_duration_seconds", "duration", None),
    ("Winners", None, "section", None),
    ("Avg stop %", "avg_stop_pct_winners", "pct", None),
    ("Avg target %", "avg_target_pct_winners", "pct", None),
    ("Percentage expectancy", "avg_result_pct_winners", "pct", "profit"),
    ("R expectancy", "avg_r_multiple_winners", "r", "profit"),
    ("Winners min duration", "min_duration_seconds_winners", "duration", None),
    ("Winners avg duration", "avg_duration_seconds_winners", "duration", None),
    ("Winners max duration", "max_duration_seconds_winners", "duration", None),
    ("Losers", None, "section", None),
    ("Avg stop %", "avg_stop_pct_losers", "pct", None),
    ("Avg target %", "avg_target_pct_losers", "pct", None),
    ("Percentage expectancy", "avg_result_pct_losers", "pct", "loss"),
    ("R expectancy", "avg_r_multiple_losers", "r", "loss"),
    ("Losers min duration", "min_duration_seconds_losers", "duration", None),
    ("Losers avg duration", "avg_duration_seconds_losers", "duration", None),
    ("Losers max duration", "max_duration_seconds_losers", "duration", None),
    ("Drawdown", None, "section", None),
    ("Max drawdown", "max_drawdown_pct", "pct", "loss"),
    ("Start", "start:max_drawdown_detail", "drawdown_date", None),
    ("End", "end:max_drawdown_detail", "drawdown_date", None),
    ("Avg drawdown", "avg_drawdown_pct", "pct", "loss"),
    ("Min drawdown", "min_drawdown_pct", "pct", "loss"),
    ("Start", "start:min_drawdown_detail", "drawdown_date", None),
    ("End", "end:min_drawdown_detail", "drawdown_date", None),
    ("Longs", "long_trades", "count", None),
    ("Long wins", "long_wins", "count", "profit"),
    ("Long losses", "long_losses", "count", "loss"),
    ("Long break-even", "long_break_even", "count", None),
    ("Shorts", "short_trades", "count", None),
    ("Short wins", "short_wins", "count", "profit"),
    ("Short losses", "short_losses", "count", "loss"),
    ("Short break-even", "short_break_even", "count", None),
    ("Min duration", "min_duration_seconds", "duration", None),
    ("Max duration", "max_duration_seconds", "duration", None),
    ("Min Move to Break Even", "min_move_to_break_even_duration_seconds", "duration", None),
    ("Max Move to Break Even", "max_move_to_break_even_duration_seconds", "duration", None),
    ("Min Move to Profit", "min_move_to_profit_duration_seconds", "duration", None),
    ("Max Move to Profit", "max_move_to_profit_duration_seconds", "duration", None),
    ("Winners min stop %", "min_stop_pct_winners", "pct", None),
    ("Winners max stop %", "max_stop_pct_winners", "pct", None),
    ("Winners min target %", "min_target_pct_winners", "pct", None),
    ("Winners max target %", "max_target_pct_winners", "pct", None),
    ("Winners min result %", "min_result_pct_winners", "pct", "profit"),
    ("Winners max result %", "max_result_pct_winners", "pct", "profit"),
    ("Winners min R", "min_r_multiple_winners", "r", "profit"),
    ("Winners max R", "max_r_multiple_winners", "r", "profit"),
    ("Losers min stop %", "min_stop_pct_losers", "pct", None),
    ("Losers max stop %", "max_stop_pct_losers", "pct", None),
    ("Losers min target %", "min_target_pct_losers", "pct", None),
    ("Losers max target %", "max_target_pct_losers", "pct", None),
    ("Losers min result %", "min_result_pct_losers", "pct", "loss"),
    ("Losers max result %", "max_result_pct_losers", "pct", "loss"),
    ("Losers min R", "min_r_multiple_losers", "r", "loss"),
    ("Losers max R", "max_r_multiple_losers", "r", "loss"),
    ("Patterns", None, "section", None),
    ("Most Traded", "most_traded_pattern", "text", None),
    ("Least Traded", "least_traded_pattern", "text", None),
    ("Most Profitable", "most_profitable_pattern", "text", "profit"),
    ("Least Profitable", "least_profitable_pattern", "text", "loss"),
    ("Timeframe", None, "section", None),
    ("1MIN", "timeframe_1min", "count", None),
    ("5MIN", "timeframe_5min", "count", None),
    ("15MIN", "timeframe_15min", "count", None),
    ("30MIN", "timeframe_30min", "count", None),
    ("1H", "timeframe_1h", "count", None),
    ("4H", "timeframe_4h", "count", None),
    ("DAILY", "timeframe_daily", "count", None),
    ("WEEKLY", "timeframe_weekly", "count", None),
    ("MONTHLY", "timeframe_monthly", "count", None),
    ("Commission", None, "section", None),
    ("Min Commission", "min_commission_by_currency", "commission", None),
    ("Avg Commission", "avg_commission_by_currency", "commission", None),
    ("Max Commission", "max_commission_by_currency", "commission", None),
    ("Total Commission", "total_commission_by_currency", "commission", None),
]
REPORT_METRIC_LABELS = [label for label, _key, _kind, _semantic in _REPORT_SPECS]

# Semantic groups parallel the authoritative STATS1 grouping without assuming
# identical labels or row numbers on the report sheets.
_REPORT_GROUP_SPEC_RANGES = (
    (0, 18),    # core counts / percentages / R metrics
    (19, 24),   # stops and targets
    (25, 27),   # duration and move averages
    (28, 31),   # result / R extrema
    (32, 33),   # duration extrema
    (34, 41),   # winner section
    (42, 49),   # loser section
    (50, 57),   # drawdown section
    (58, 65),   # side counts
    (66, 71),   # duration extrema
    (72, 79),   # winner extrema
    (80, 87),   # loser extrema
    (88, 92),   # patterns
    (93, 102),  # timeframe
    (103, 107), # commission
)


def _effective_semantic_fill(kind: str, semantic: str | None) -> str | None:
    if kind == "commission":
        return "loss"
    if kind in {
        "section", "duration", "text", "date", "streak_date", "drawdown_date"
    }:
        return None
    return semantic


def _currency_from_number_format(number_format: str) -> str:
    text = str(number_format or "").upper()
    for currency in ("USDT", "USDC", "AUD", "USD", "BTC", "ETH", "SOL", "XRP"):
        if currency in text:
            return currency
    return ""


def _report_trailing_blank_rows(ws) -> List[int]:
    spec_rows = _report_spec_rows(ws)
    if not spec_rows:
        return []
    last_managed_row = max(row for row, _spec in spec_rows)
    return [
        row
        for row in range(last_managed_row + 1, ws.max_row + 1)
        if all(
            ws.cell(row, col).value in (None, "")
            for col in range(1, ws.max_column + 1)
        )
    ]


def _apply_report_semantic_formatting(
    ws,
    *,
    header_cols: Sequence[int] | None = None,
    buckets: Sequence[Dict[str, Any]] | None = None,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    bucket_by_col = {
        col: bucket
        for col, bucket in zip(header_cols or (), buckets or ())
    }
    repaired = 0
    for row, (_label, key, kind, semantic) in _report_spec_rows(ws):
        ws.cell(row, 1).fill = PatternFill()
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row, col)
            currency = ""
            if kind == "commission" and key:
                values = (bucket_by_col.get(col) or {}).get(key)
                if isinstance(values, dict) and len(values) == 1:
                    currency = _currency_code(next(iter(values)))
                if not currency:
                    currency = _currency_from_number_format(cell.number_format)
            _apply_semantic_metric_policy(
                cell,
                kind=kind,
                semantic=_effective_semantic_fill(kind, semantic),
                currency=currency or None,
            )
            repaired += 1
    if repaired:
        diagnostics.setdefault("repaired_report_semantic_metric_cells", {})[
            ws.title
        ] = repaired
    trailing_cells = 0
    for row in _report_trailing_blank_rows(ws):
        for col in range(1, ws.max_column + 1):
            # Trailing blank rows are outside the managed report metrics.
            # Normalize their fill/format without touching user-owned font
            # flags, which may be restored from an older workbook layout.
            cell = ws.cell(row, col)
            if cell.number_format != "General":
                cell.number_format = "General"
            _clear_cell_fill(cell)
            trailing_cells += 1
    if trailing_cells:
        diagnostics.setdefault("neutralized_report_trailing_blank_cells", {})[
            ws.title
        ] = trailing_cells


def _apply_report_semantic_group_borders(
    ws,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    spec_rows = _report_spec_rows(ws)
    rows_by_index = {
        index: row
        for index, (row, _spec) in enumerate(spec_rows)
    }
    if not rows_by_index:
        return
    last_col = max(1, ws.max_column)
    last_managed_row = max(rows_by_index.values())
    _clear_borders_in_range(ws, 1, 1, last_managed_row, last_col)
    for row in _report_trailing_blank_rows(ws):
        for col in range(1, last_col + 1):
            ws.cell(row, col).border = Border()
    applied: List[str] = []
    for start_index, end_index in _REPORT_GROUP_SPEC_RANGES:
        rows = [
            rows_by_index[index]
            for index in range(start_index, end_index + 1)
            if index in rows_by_index
        ]
        if not rows:
            continue
        start_row = 1 if start_index == 0 else min(rows)
        end_row = max(rows)
        _apply_group_box_border(
            ws,
            start_row,
            1,
            end_row,
            last_col,
            outer_style="medium",
        )
        applied.append(
            f"A{start_row}:{get_column_letter(last_col)}{end_row}"
        )
    diagnostics.setdefault("repaired_report_semantic_border_groups", {})[
        ws.title
    ] = applied


def _report_stats1_style_context(spec_index: int) -> str:
    if 34 <= spec_index <= 41 or 72 <= spec_index <= 79:
        return "winners"
    if 42 <= spec_index <= 49 or 80 <= spec_index <= 87:
        return "losers"
    if 50 <= spec_index <= 57:
        return "drawdown"
    if 58 <= spec_index <= 65:
        return "side"
    if 88 <= spec_index <= 92:
        return "patterns"
    if 93 <= spec_index <= 102:
        return "timeframe"
    if 103 <= spec_index <= 107:
        return "commission"
    return "core"


def _stats1_label_template_for_report(
    stats1,
    *,
    spec_index: int,
    label: str,
):
    context = _report_stats1_style_context(spec_index)
    label_key = " ".join(str(label or "").strip().casefold().split())
    aliases = {
        "net p/l": "net p/l percentage",
        "avg duration (dd:hh:mm:ss)": "avg duration",
        "move to break even (dd:hh:mm:ss)": "average move to break even",
        "move to profit (dd:hh:mm:ss)": "average move to profit",
        "shortest (dd:hh:mm:ss)": "min duration",
        "longest (dd:hh:mm:ss)": "max duration",
        "longs": "long",
        "shorts": "short",
        "min commission": "min",
        "avg commission": "avg",
        "max commission": "max",
        "total commission": "total",
    }
    candidate = aliases.get(label_key, label_key)
    if label_key.startswith("winners "):
        candidate = label_key[len("winners "):]
    elif label_key.startswith("losers "):
        candidate = label_key[len("losers "):]
    if context == "side":
        if "wins" in label_key:
            candidate = "winners"
        elif "loss" in label_key:
            candidate = "losers"
        elif "long" in label_key:
            candidate = "long"
        elif "short" in label_key:
            candidate = "short"

    categorical_sections = {"side", "patterns", "timeframe", "commission", "drawdown"}
    if context == "core":
        rows = range(1, (_stats1_section_bounds(stats1, "Winners") or (stats1.max_row + 1, 0))[0])
    else:
        section_bounds = _stats1_section_bounds(
            stats1,
            context.title(),
            categorical_sections if context in categorical_sections else None,
        )
        rows = range(section_bounds[0], section_bounds[1] + 1) if section_bounds else ()
    for row in rows:
        if " ".join(_stats1_label_at(stats1, row).strip().casefold().split()) == candidate:
            return stats1.cell(row, 1)

    # Fallbacks retain the authoritative typography roles even where a report
    # label has no exact STATS1 counterpart.
    fallback_candidates = []
    if context in {"winners", "losers"}:
        fallback_candidates = ["avg stop %", context]
    elif context == "side":
        fallback_candidates = ["winners" if "win" in label_key else "losers" if "loss" in label_key else "long"]
    elif context == "commission":
        fallback_candidates = ["min", "commission"]
    elif context in categorical_sections:
        fallback_candidates = [context]
    else:
        fallback_candidates = ["trades"]
    for fallback in fallback_candidates:
        for row in range(1, stats1.max_row + 1):
            if " ".join(_stats1_label_at(stats1, row).strip().casefold().split()) == fallback:
                return stats1.cell(row, 1)
    return None


def _apply_report_label_fonts_from_stats1(stats1, report_ws) -> None:
    for spec_index, (row, spec) in enumerate(_report_spec_rows(report_ws)):
        template = _stats1_label_template_for_report(
            stats1,
            spec_index=spec_index,
            label=spec[0],
        )
        if template is None:
            continue
        target = report_ws.cell(row, 1)
        target.font = copy(template.font)
        target.alignment = copy(template.alignment)
        target.fill = PatternFill()


def _apply_report_workbook_semantic_presentation(
    wb,
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    if STATS1_SHEET not in wb.sheetnames:
        return
    stats1 = wb[STATS1_SHEET]
    for sheet_name in (
        REPORT_YEARLY_SHEET,
        *[name for name in wb.sheetnames if name.isdigit() and int(name) >= REPORT_START_YEAR],
    ):
        if sheet_name not in wb.sheetnames:
            continue
        report_ws = wb[sheet_name]
        _apply_report_semantic_formatting(report_ws, diagnostics=diagnostics)
        _apply_report_label_fonts_from_stats1(stats1, report_ws)
        _apply_report_semantic_group_borders(report_ws, diagnostics)


def _report_trade_years_from_snapshot(snapshot: Dict[str, Any]) -> List[int]:
    years = {REPORT_START_YEAR, REPORT_MIN_END_YEAR, datetime.now(JOURNAL_DISPLAY_TZ).year}
    for row in snapshot.get("items") or []:
        if not isinstance(row, dict) or str(row.get("row_type") or "trade") != "trade":
            continue
        dt = _as_date(row.get("close_time") or row.get("open_time"))
        if dt:
            years.add(int(dt.year))
    reports = ((snapshot.get("stats") or {}).get("period_reports") or {})
    for key in (reports.get("years") or {}).keys():
        try:
            years.add(int(key))
        except Exception:
            pass
    end_year = max(max(years), REPORT_MIN_END_YEAR)
    return list(range(REPORT_START_YEAR, end_year + 1))


def expected_report_sheet_names(snapshot: Dict[str, Any] | None = None) -> List[str]:
    snapshot = snapshot if isinstance(snapshot, dict) else {}
    return [REPORT_YEARLY_SHEET, *[str(year) for year in _report_trade_years_from_snapshot(snapshot)]]


def _period_report_lookup(snapshot: Dict[str, Any], *, year: int, month: int | None = None) -> Dict[str, Any]:
    reports = ((snapshot.get("stats") or {}).get("period_reports") or {})
    if month is None:
        years = reports.get("years") if isinstance(reports.get("years"), dict) else {}
        return years.get(year) or years.get(str(year)) or {}
    months = reports.get("months") if isinstance(reports.get("months"), dict) else {}
    year_months = months.get(year) or months.get(str(year)) or {}
    if not isinstance(year_months, dict):
        return {}
    return year_months.get(month) or year_months.get(str(month)) or {}


def _report_bucket_from_stats(stats: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(stats, dict):
        return {}
    groups = stats.get("groups") if isinstance(stats.get("groups"), dict) else {}
    totals = stats.get("totals") if isinstance(stats.get("totals"), dict) else {}
    by_market = groups.get("by_market") if isinstance(groups.get("by_market"), dict) else {}
    overall = by_market.get("overall") if isinstance(by_market.get("overall"), dict) else {}
    risk = groups.get("risk_expectancy") if isinstance(groups.get("risk_expectancy"), dict) else {}
    risk_by_market = risk.get("by_market") if isinstance(risk.get("by_market"), dict) else {}
    duration = groups.get("duration") if isinstance(groups.get("duration"), dict) else {}
    bucket: Dict[str, Any] = {}
    bucket.update(totals)
    bucket.update(overall)
    bucket.update(risk)
    if isinstance(risk_by_market.get("overall"), dict):
        bucket.update(risk_by_market["overall"])
    bucket.setdefault("shortest_duration_seconds", duration.get("overall_shortest_seconds"))
    bucket.setdefault("longest_duration_seconds", duration.get("overall_longest_seconds"))
    bucket.setdefault("avg_duration_seconds", duration.get("overall_avg_seconds"))
    bucket.setdefault("min_duration_seconds", duration.get("overall_shortest_seconds"))
    bucket.setdefault("max_duration_seconds", duration.get("overall_longest_seconds"))
    bucket.setdefault("net_result_pct", overall.get("net_result_pct", totals.get("net_result_pct")))
    bucket.setdefault("gross_gain_result_pct", overall.get("gross_gain_result_pct", totals.get("gross_gain_result_pct")))
    bucket.setdefault("gross_loss_result_pct", overall.get("gross_loss_result_pct", totals.get("gross_loss_result_pct")))
    metric_sources: Dict[str, Any] = {}
    for source_map in (
        overall.get("metric_sources") if isinstance(overall, dict) else None,
        duration.get("metric_sources") if isinstance(duration, dict) else None,
    ):
        if isinstance(source_map, dict):
            metric_sources.update(source_map)
    if metric_sources:
        bucket["metric_sources"] = metric_sources
    return bucket


def _report_cell_value(bucket: Dict[str, Any], key: str | None, kind: str) -> Any:
    if kind == "section" or key is None:
        return ""
    if kind == "source":
        source_key = key.split(":", 1)[1] if ":" in key else key
        source = (bucket.get("metric_sources") or {}).get(source_key)
        return _fmt_detail_src(source) if source else ""
    value = bucket.get(key)
    inline_source = _inline_metric_source(bucket, key)
    if key in INLINE_SOURCE_METRIC_KEYS and kind in {"pct", "r", "duration"}:
        if _inline_metric_source_has_symbol_and_date(inline_source):
            return _format_inline_metric_value(value, kind, inline_source)
        # A new workbook may still expose a valid numeric extrema when the
        # upstream source metadata is incomplete.  Preservation mode applies
        # the stronger existing-value guard in
        # _update_report_sheet_preserving_layout.
    if kind == "pct":
        number = _as_float(value)
        return "" if number is None else number / 100.0
    if kind == "drawdown_date":
        endpoint, detail_key = str(key or "").split(":", 1)
        return _drawdown_detail_cell_value(bucket.get(detail_key), endpoint)
    if kind == "r":
        number = _as_float(value)
        return "" if number is None else number
    if kind == "count":
        number = _as_float(value)
        return "" if number is None else int(number)
    if kind == "streak_date":
        endpoint, detail_key = str(key or "").split(":", 1)
        start, end = _streak_detail_start_end(bucket.get(detail_key))
        return start if endpoint == "start" else end
    if kind == "duration":
        return _format_duration_display(value) if value not in (None, "") else ""
    if kind == "number":
        number = _as_float(value)
        return "" if number is None else number
    if kind == "commission":
        if not isinstance(value, dict) or not value:
            return ""
        if len(value) == 1:
            return next(iter(value.values()))
        return " / ".join(f"{currency} {amount:,.8f}".rstrip("0").rstrip(".") for currency, amount in sorted(value.items()))
    return "" if value is None else value


def _format_report_sheet(ws, last_col: int) -> None:
    current_freeze = getattr(ws.freeze_panes, "coordinate", ws.freeze_panes)
    if current_freeze != "B2":
        ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    for col in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14
    _style_header_row(ws, 1)
    for row in range(2, len(REPORT_METRIC_LABELS) + 2):
        ws.row_dimensions[row].height = 14.5
    _apply_report_semantic_formatting(ws)
    _apply_report_semantic_group_borders(ws)


def _normalize_generated_statistics_row_heights(wb) -> None:
    managed = [
        ws.title
        for ws in wb.worksheets
        if ws.title in {STATS1_SHEET, REPORT_YEARLY_SHEET}
        or (ws.title.isdigit() and int(ws.title) >= REPORT_START_YEAR)
    ]
    for sheet_name in managed:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            if any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1)):
                ws.row_dimensions[row].height = 14.5


def _normalize_generated_duration_text(wb) -> None:
    managed_names = {
        STATS1_SHEET,
        SYMBOLS_SHEET,
        REPORT_YEARLY_SHEET,
        *[
            ws.title
            for ws in wb.worksheets
            if ws.title.isdigit() and int(ws.title) >= REPORT_START_YEAR
        ],
    }
    verbose_pattern = re.compile(r"\b\d+(?:\.\d+)?\s*(?:days?|hours?|minutes?|seconds?)\b", re.IGNORECASE)
    for sheet_name in managed_names:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not verbose_pattern.search(cell.value):
                    continue
                value_text = cell.value
                duration_text, separator, source_text = value_text.partition(" - ")
                seconds = _parse_duration_text(duration_text)
                if seconds is None:
                    continue
                cell.value = _format_duration_display(seconds) + (f" - {source_text}" if separator else "")
                cell.number_format = "General"


def _clear_report_sheet(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)
    ws.conditional_formatting._cf_rules.clear()


def _write_report_sheet(ws, headers: List[Any], buckets: List[Dict[str, Any]]) -> None:
    _clear_report_sheet(ws)
    ws.cell(1, 1, "")
    for col, header in enumerate(headers, start=2):
        ws.cell(1, col, header)
    for row_idx, (label, key, kind, _semantic) in enumerate(_REPORT_SPECS, start=2):
        ws.cell(row_idx, 1, label)
        for col, bucket in enumerate(buckets, start=2):
            ws.cell(row_idx, col, _report_cell_value(bucket, key, kind))
    _format_report_sheet(ws, max(1, len(headers) + 1))
    for row_idx, (_label, key, kind, _semantic) in enumerate(_REPORT_SPECS, start=2):
        if kind != "commission" or key is None:
            continue
        for col, bucket in enumerate(buckets, start=2):
            values = bucket.get(key)
            if isinstance(values, dict) and len(values) == 1:
                ws.cell(row_idx, col).number_format = _currency_number_format(next(iter(values)))
            elif isinstance(values, dict) and len(values) > 1:
                ws.cell(row_idx, col).number_format = "General"
    _apply_report_semantic_formatting(
        ws,
        header_cols=list(range(2, len(headers) + 2)),
        buckets=buckets,
    )
    _apply_report_semantic_group_borders(ws)


def _copy_report_row(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        target.value = source.value
        target.number_format = source.number_format
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.comment = copy(source.comment) if source.comment else None
        target.hyperlink = copy(source.hyperlink) if source.hyperlink else None
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _report_label_rows(ws, label: str) -> List[int]:
    wanted = str(label or "").strip().casefold()
    return [
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip().casefold() == wanted
    ]


def _repair_report_layout(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    source_rows = _report_label_rows(ws, "Source")
    for source_row in reversed(source_rows):
        ws.delete_rows(source_row, 1)
    if source_rows:
        diagnostics.setdefault("removed_report_source_rows", {})[ws.title] = len(source_rows)
    label_replacements = {
        "avg result %": "Percentage expectancy",
        "average result percent": "Percentage expectancy",
        "average result %": "Percentage expectancy",
        "avg r": "R expectancy",
        "average r": "R expectancy",
        "gross gain": "Gross percent gain",
        "gross loss": "Gross percent loss",
    }
    for row in range(1, ws.max_row + 1):
        current = str(ws.cell(row, 1).value or "").strip()
        replacement = label_replacements.get(current.casefold())
        if replacement and current != replacement:
            ws.cell(row, 1).value = replacement
            diagnostics.setdefault("renamed_report_metric_labels", []).append(f"{ws.title}: {replacement}")

    expectancy_rows = _report_label_rows(ws, "Expectancy %")
    for row in reversed(expectancy_rows):
        ws.delete_rows(row, 1)
    if expectancy_rows:
        diagnostics.setdefault("report_expectancy_rows_removed", []).append(ws.title)

    drawdown_rows = _report_label_rows(ws, "Drawdown")
    min_rows = _report_label_rows(ws, "Min drawdown")
    if drawdown_rows and min_rows:
        drawdown_row = drawdown_rows[0]
        avg_row = next(
            (row for row in range(drawdown_row + 1, min(ws.max_row, drawdown_row + 8) + 1)
             if str(ws.cell(row, 1).value or "").strip().casefold() == "avg drawdown"),
            None,
        )
        source_row = min_rows[0]
        target_row = (avg_row + 1) if avg_row else None
        if target_row and source_row != target_row:
            ws.insert_rows(target_row, 1)
            if source_row >= target_row:
                source_row += 1
            _copy_report_row(ws, source_row, target_row)
            ws.delete_rows(source_row, 1)
            diagnostics.setdefault("report_min_drawdown_relocated", []).append(ws.title)

    _ensure_drawdown_start_end_rows(ws, diagnostics)
    _write_drawdown_detail_rows(ws, "Min drawdown", {}, diagnostics=diagnostics)
    _write_drawdown_detail_rows(ws, "Max drawdown", {}, diagnostics=diagnostics)
    _write_streak_detail_rows(ws, "Best Win Streak", {}, diagnostics=diagnostics)
    _write_streak_detail_rows(ws, "Worst Losing Streak", {}, diagnostics=diagnostics)

    commission_rows = _report_label_rows(ws, "Commission")
    if commission_rows and not _report_label_rows(ws, "Total Commission"):
        commission_row = commission_rows[0]
        max_row = next(
            (row for row in range(commission_row + 1, min(ws.max_row, commission_row + 8) + 1)
             if str(ws.cell(row, 1).value or "").strip().casefold() == "max commission"),
            None,
        )
        if max_row:
            target_row = max_row + 1
            ws.insert_rows(target_row, 1)
            _copy_report_row(ws, max_row, target_row)
            ws.cell(target_row, 1).value = "Total Commission"
            for col in range(2, ws.max_column + 1):
                ws.cell(target_row, col).value = None
            diagnostics.setdefault("report_total_commission_rows_added", []).append(ws.title)


def _report_spec_rows(ws) -> List[Tuple[int, Tuple[str, str | None, str, str | None]]]:
    output: List[Tuple[int, Tuple[str, str | None, str, str | None]]] = []
    cursor = 2
    for spec in _REPORT_SPECS:
        label = spec[0].casefold()
        found = next(
            (row for row in range(cursor, ws.max_row + 1)
             if str(ws.cell(row, 1).value or "").strip().casefold() == label),
            None,
        )
        if found is None:
            continue
        output.append((found, spec))
        cursor = found + 1
    return output


def _update_report_sheet_preserving_layout(
    ws,
    headers: List[Any],
    buckets: List[Dict[str, Any]],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    _repair_report_layout(ws, diagnostics)
    header_cols: List[int] = []
    for header in headers:
        found = next(
            (col for col in range(2, ws.max_column + 1)
             if str(ws.cell(1, col).value or "").strip().casefold() == str(header).strip().casefold()),
            None,
        )
        if found is not None:
            header_cols.append(found)
        else:
            header_cols.append(len(header_cols) + 2)
            ws.cell(1, header_cols[-1]).value = header
    spec_rows = _report_spec_rows(ws)
    found_labels = [spec[0] for _row, spec in spec_rows]
    missing = [spec[0] for spec in _REPORT_SPECS if spec[0] not in found_labels]
    if missing:
        diagnostics.setdefault("missing_report_metric_labels", {})[ws.title] = missing
        _write_report_sheet(ws, headers, buckets)
        diagnostics.setdefault("rewrote_report_sheets_for_requested_metrics", []).append(ws.title)
        return
    for row, (_label, key, kind, _semantic) in spec_rows:
        for col, bucket in zip(header_cols, buckets):
            cell = ws.cell(row, col)
            inline_source = _inline_metric_source(bucket, key) if key else None
            bucket_trades = _as_float(bucket.get("trades"))
            bucket_test_trades = _as_float(bucket.get("test_trades"))
            test_only_bucket = (
                (bucket_trades is None or bucket_trades == 0)
                and bucket_test_trades is not None
                and bucket_test_trades > 0
            )
            preserve_existing_inline = (
                key in INLINE_SOURCE_METRIC_KEYS
                and kind in {"pct", "r", "duration"}
                and cell.value not in (None, "")
                and not test_only_bucket
                and not _inline_metric_source_has_symbol_and_date(
                    inline_source
                )
            )
            next_value = (
                cell.value
                if preserve_existing_inline
                else _report_cell_value(bucket, key, kind)
            )
            if preserve_existing_inline:
                diagnostics.setdefault(
                    "preserved_report_values_without_safe_source", {}
                ).setdefault(ws.title, []).append(cell.coordinate)
            cell.value = next_value
            if isinstance(cell.value, str) and cell.value:
                cell.number_format = "General"
            elif kind == "pct":
                cell.number_format = "0.00%"
            elif kind == "r":
                cell.number_format = '0.000"R"'
            elif kind == "count":
                cell.number_format = "0"
            elif kind == "streak_date":
                cell.number_format = "General"
                _apply_streak_detail_value_font(cell)
            elif kind == "duration":
                cell.number_format = "General"
            elif kind == "drawdown_date":
                cell.number_format = "General"
            elif kind == "commission":
                values = bucket.get(key) if key else None
                if isinstance(values, dict) and len(values) == 1:
                    cell.number_format = _currency_number_format(next(iter(values)))
                elif isinstance(values, dict) and len(values) > 1:
                    cell.number_format = "General"
                else:
                    cell.number_format = "General"
    _apply_report_semantic_formatting(
        ws,
        header_cols=header_cols,
        buckets=buckets,
        diagnostics=diagnostics,
    )
    _apply_report_semantic_group_borders(ws, diagnostics)
    diagnostics.setdefault("updated_report_sheets", []).append(ws.title)


def _report_rows_for_period(
    snapshot: Dict[str, Any],
    *,
    year: int,
    month: int | None = None,
) -> List[Dict[str, Any]]:
    cache = snapshot.get("_report_rows_by_period") if isinstance(snapshot, dict) else None
    if not isinstance(cache, dict):
        by_year: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
        by_month: Dict[Tuple[int, int], List[Dict[str, Any]]] = defaultdict(list)
        for row in snapshot.get("items") or []:
            if not isinstance(row, dict):
                continue
            if str(row.get("row_type") or "trade").strip().lower() not in {"trade", "cashflow", "monthly_aud_reval"}:
                continue
            timestamp = _as_datetime(row.get("close_time") or row.get("open_time"))
            if timestamp is None:
                continue
            by_year[timestamp.year].append(row)
            by_month[(timestamp.year, timestamp.month)].append(row)
        cache = {"year": by_year, "month": by_month}
        if isinstance(snapshot, dict):
            snapshot["_report_rows_by_period"] = cache
    if month is None:
        return list((cache.get("year") or {}).get(year, []))
    return list((cache.get("month") or {}).get((year, month), []))


def _report_outcome(row: Dict[str, Any]) -> int:
    value = _as_float(row.get("result_pct"))
    if value is None:
        value = _as_float(row.get("net_profit"))
    if value is None:
        value = _as_float(row.get("result_cash"))
    if value is None or not math.isfinite(value) or value == 0:
        return 0
    return 1 if value > 0 else -1


def _symbol_win_rate_snapshot_fingerprint(
    rows: Sequence[Mapping[str, Any]],
) -> str:
    """Fingerprint only the authoritative inputs that can change win rates."""
    samples: List[Tuple[str, str, bool, int]] = []
    for source in rows:
        row = dict(source)
        if str(row.get("row_type") or "trade").strip().casefold() != "trade":
            continue
        identity = str(row.get("id") or "").strip()
        if not identity:
            identity = _trade_execution_fingerprint(row) or ""
        samples.append(
            (
                identity,
                str(row.get("symbol") or "").strip().upper(),
                _is_test_trade_value(row.get("is_test_trade")),
                _trade_outcome_sign(row),
            )
        )
    payload = json.dumps(
        sorted(samples),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _period_streak_metrics(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    ordered = sorted(
        rows,
        key=lambda row: _as_datetime(row.get("close_time") or row.get("open_time")) or datetime.min,
    )
    best: Dict[int, Tuple[int, Dict[str, Any] | None]] = {1: (0, None), -1: (0, None)}
    current_sign = 0
    current_rows: List[Dict[str, Any]] = []
    for row in ordered:
        sign = _report_outcome(row)
        if sign == 0:
            current_sign = 0
            current_rows = []
            continue
        if sign != current_sign:
            current_sign = sign
            current_rows = [row]
        else:
            current_rows.append(row)
        if len(current_rows) > best[sign][0]:
            best[sign] = (
                len(current_rows),
                {
                    "start_time": current_rows[0].get("close_time") or current_rows[0].get("open_time"),
                    "end_time": current_rows[-1].get("close_time") or current_rows[-1].get("open_time"),
                },
            )
    result: Dict[str, Any] = {
        "winning_streak": best[1][0] or None,
        "losing_streak": best[-1][0] or None,
    }
    if best[1][1]:
        result["longest_winning_streak"] = best[1][1]
    if best[-1][1]:
        result["longest_losing_streak"] = best[-1][1]
    return result


def _drawdown_account_key(row: Dict[str, Any]) -> str:
    account = _canonical_account_label(row.get("account_label") or row.get("account") or "")
    return re.sub(r"[^a-z0-9]+", "", account.casefold())


def _drawdown_row_timestamp(row: Dict[str, Any]) -> Tuple[float, Any] | None:
    raw = row.get("close_time") or row.get("open_time")
    dt = _as_datetime(raw)
    if dt is None:
        return None
    return dt.timestamp(), raw


def balance_drawdown_metrics(
    rows: List[Dict[str, Any]],
    all_rows: List[Dict[str, Any]] | None = None,
) -> Dict[str, Any]:
    anchor_source = all_rows if all_rows is not None else rows
    events_by_account: Dict[str, List[Tuple[float, str]]] = defaultdict(list)
    for row in anchor_source:
        row_type = str(row.get("row_type") or "trade").strip().lower()
        if row_type not in {"cashflow", "monthly_aud_reval"}:
            continue
        account_key = _drawdown_account_key(row)
        stamped = _drawdown_row_timestamp(row)
        if not account_key or stamped is None:
            continue
        ts, raw_time = stamped
        events_by_account[account_key].append((ts, str(row.get("id") or raw_time)))
    for account_key in list(events_by_account):
        events_by_account[account_key] = sorted(events_by_account[account_key], key=lambda item: item[0])

    segments: Dict[Tuple[str, str], List[Tuple[float, float, Any, str]]] = defaultdict(list)
    for row in rows:
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        account_key = _drawdown_account_key(row)
        stamped = _drawdown_row_timestamp(row)
        if not account_key or stamped is None:
            continue
        ts, raw_time = stamped
        balance = _as_float(row.get("analysis_balance_after_trade"))
        if balance is None:
            balance = _as_float(row.get("balance_after_trade"))
        if balance is None or not math.isfinite(balance) or balance <= 0:
            continue
        anchor_id = "__no_anchor__"
        for event_ts, event_id in events_by_account.get(account_key, []):
            if event_ts <= ts:
                anchor_id = event_id
            else:
                break
        account_label = str(row.get("account_label") or row.get("account") or account_key).strip()
        segments[(account_key, anchor_id)].append((ts, balance, raw_time, account_label))

    drawdown_values: List[float] = []
    drawdown_details: List[Tuple[float, Dict[str, Any]]] = []
    segment_count = 0
    balance_points = 0
    for points in segments.values():
        ordered = sorted(points, key=lambda item: item[0])
        if len(ordered) < 2:
            continue
        segment_count += 1
        balance_points += len(ordered)
        peak: float | None = None
        peak_time: Any = None
        peak_account = ""
        for _ts, balance, raw_time, account in ordered:
            if peak is None or balance > peak:
                peak = balance
                peak_time = raw_time
                peak_account = account
            if peak is None or peak <= 0:
                continue
            drawdown = (peak - balance) / peak * 100.0
            if drawdown > 0 and math.isfinite(drawdown):
                drawdown_values.append(drawdown)
                drawdown_details.append((
                    drawdown,
                    {
                        "account": account or peak_account,
                        "start_time": peak_time,
                        "end_time": raw_time,
                        "peak": peak,
                        "trough": balance,
                    },
                ))

    if drawdown_values:
        min_item = min(drawdown_details, key=lambda item: item[0])
        max_item = max(drawdown_details, key=lambda item: item[0])
        return {
            "min_drawdown_pct": min(drawdown_values),
            "avg_drawdown_pct": sum(drawdown_values) / len(drawdown_values),
            "max_drawdown_pct": max(drawdown_values),
            "min_drawdown_detail": min_item[1],
            "max_drawdown_detail": max_item[1],
            "drawdown_balance_points": balance_points,
            "drawdown_segments_count": segment_count,
        }
    if segment_count:
        return {
            "min_drawdown_pct": 0.0,
            "avg_drawdown_pct": 0.0,
            "max_drawdown_pct": 0.0,
            "drawdown_balance_points": balance_points,
            "drawdown_segments_count": segment_count,
        }
    return {}


def _period_drawdown_metrics(
    rows: List[Dict[str, Any]],
    all_rows: List[Dict[str, Any]] | None = None,
) -> Dict[str, Any]:
    return balance_drawdown_metrics(rows, all_rows)


def _report_bucket_from_period_rows(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    active = [
        row for row in rows
        if str(row.get("row_type") or "trade").strip().lower() == "trade"
        and not _is_test_trade_value(row.get("is_test_trade"))
    ]
    outcomes = [_report_outcome(row) for row in active]
    trades = len(active)
    duration_samples = [
        (value, row)
        for row in active
        for value in [_as_float(row.get("trade_duration_seconds"))]
        if value is not None and value >= 0 and math.isfinite(value)
    ]
    bucket: Dict[str, Any] = {
        "trades": trades,
        "wins": sum(outcome > 0 for outcome in outcomes),
        "losses": sum(outcome < 0 for outcome in outcomes),
        "break_even": sum(outcome == 0 for outcome in outcomes),
        "test_trades": sum(
            str(row.get("row_type") or "trade").strip().lower() == "trade"
            and _is_test_trade_value(row.get("is_test_trade"))
            for row in rows
        ),
    }
    bucket["win_rate_pct"] = (bucket["wins"] / trades * 100.0) if trades else None
    if duration_samples:
        shortest_value, shortest_row = min(duration_samples, key=lambda item: item[0])
        longest_value, longest_row = max(duration_samples, key=lambda item: item[0])
        bucket.update({
            "shortest_duration_seconds": shortest_value,
            "longest_duration_seconds": longest_value,
            "avg_duration_seconds": sum(value for value, _row in duration_samples) / len(duration_samples),
            "min_duration_seconds": shortest_value,
            "max_duration_seconds": longest_value,
            "metric_sources": {
                "shortest_duration_seconds": _trade_metric_ref(shortest_row, "shortest_duration_seconds", shortest_value),
                "longest_duration_seconds": _trade_metric_ref(longest_row, "longest_duration_seconds", longest_value),
            },
        })
    bucket.update(_period_streak_metrics(active))
    bucket.update(_period_drawdown_metrics(active, rows))
    return bucket


def _report_bucket_for_period(
    snapshot: Dict[str, Any],
    *,
    year: int,
    month: int | None = None,
) -> Dict[str, Any]:
    bucket = _report_bucket_from_stats(_period_report_lookup(snapshot, year=year, month=month))
    rows = _report_rows_for_period(snapshot, year=year, month=month)
    if rows:
        active_rows = [
            row for row in rows
            if str(row.get("row_type") or "trade").strip().lower() == "trade"
            and not _is_test_trade_value(row.get("is_test_trade"))
        ]
        fallback = _report_bucket_from_period_rows(rows)
        extended = _dashboard_extended_metrics(rows, {"overall": {**bucket, **fallback}})["overall"]
        move_durations = _trade_move_duration_metrics(active_rows)["overall"]
        bucket = _merge_metric_buckets(
            bucket,
            fallback,
            _linear_profit_percentage_totals(active_rows),
            {key: value for key, value in extended.items() if value is not None},
            {key: value for key, value in move_durations.items() if value is not None},
        )
        commission_by_currency: Dict[str, List[float]] = defaultdict(list)
        for row in active_rows:
            commission = _as_float(row.get("commission"))
            if commission is None or commission == 0:
                continue
            market = _trade_row_market(row)
            currency = str(
                row.get("commission_currency")
                or row.get("currency")
                or row.get("account_currency")
                or ("AUD" if market == "fx" else "USDT" if market == "crypto" else "")
            ).strip().upper()
            if currency:
                commission_by_currency[currency].append(abs(commission))
        bucket["min_commission_by_currency"] = {
            currency: min(values) for currency, values in commission_by_currency.items() if values
        }
        bucket["avg_commission_by_currency"] = {
            currency: sum(values) / len(values) for currency, values in commission_by_currency.items() if values
        }
        bucket["max_commission_by_currency"] = {
            currency: max(values) for currency, values in commission_by_currency.items() if values
        }
        bucket["total_commission_by_currency"] = {
            currency: sum(values) for currency, values in commission_by_currency.items() if values
        }
    return bucket


def _ensure_report_sheets(wb, snapshot: Dict[str, Any], diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    current_order = list(wb.sheetnames)
    current_active = wb.active.title if wb.worksheets else ""
    current_selected = [
        ws.title for ws in wb.worksheets if bool(ws.sheet_view.tabSelected)
    ]
    structure = (
        snapshot.get("_workbook_structure")
        if isinstance(snapshot, dict)
        and isinstance(snapshot.get("_workbook_structure"), dict)
        else {}
    )
    requested_order = structure.get("sheet_order")
    preserved_order = (
        [str(name) for name in requested_order if str(name) in wb.sheetnames]
        if isinstance(requested_order, list)
        else current_order
    )
    preserved_order = list(dict.fromkeys(preserved_order))
    preserved_active = str(structure.get("active_sheet") or current_active)
    requested_selected = structure.get("selected_sheets")
    preserved_selected = (
        [str(name) for name in requested_selected if str(name) in wb.sheetnames]
        if isinstance(requested_selected, list)
        else current_selected
    )
    years = _report_trade_years_from_snapshot(snapshot if isinstance(snapshot, dict) else {})
    expected_names = [REPORT_YEARLY_SHEET, *[str(year) for year in years]]
    created: set[str] = set()
    for name in expected_names:
        if name not in wb.sheetnames:
            wb.create_sheet(name)
            diagnostics.setdefault("created_report_sheets", []).append(name)
            created.add(name)
    yearly_buckets = [_report_bucket_for_period(snapshot, year=year) for year in years]
    yearly_ws = wb[REPORT_YEARLY_SHEET]
    if REPORT_YEARLY_SHEET in created or not any(yearly_ws.cell(1, col).value not in (None, "") for col in range(2, yearly_ws.max_column + 1)):
        _write_report_sheet(yearly_ws, years, yearly_buckets)
    else:
        _update_report_sheet_preserving_layout(yearly_ws, years, yearly_buckets, diagnostics)
    for year in years:
        start_month = 5 if year == 2018 else 1
        month_headers = [calendar.month_name[month] for month in range(start_month, 13)]
        month_buckets = [
            _report_bucket_for_period(snapshot, year=year, month=month)
            for month in range(start_month, 13)
        ]
        year_ws = wb[str(year)]
        if str(year) in created or not any(year_ws.cell(1, col).value not in (None, "") for col in range(2, year_ws.max_column + 1)):
            _write_report_sheet(year_ws, month_headers, month_buckets)
        else:
            _update_report_sheet_preserving_layout(year_ws, month_headers, month_buckets, diagnostics)
    remaining = [
        sheet.title for sheet in wb._sheets if sheet.title not in preserved_order
    ]
    final_order = [*preserved_order, *remaining]
    wb._sheets = [wb[name] for name in final_order]
    if preserved_active in wb.sheetnames:
        wb.active = wb.sheetnames.index(preserved_active)
    selected_names = {
        name for name in preserved_selected if name in wb.sheetnames
    }
    if not selected_names and preserved_active in wb.sheetnames:
        selected_names.add(preserved_active)
    for ws in wb.worksheets:
        ws.sheet_view.tabSelected = ws.title in selected_names
    _apply_report_workbook_semantic_presentation(wb, diagnostics)

def _write_instrument_leaders_section(ws, start_row, start_col, leaders):
    ws.merge_cells(start_row=start_row,start_column=start_col,end_row=start_row,end_column=start_col+4)
    ws.cell(start_row,start_col,"Instrument leaders").font=Font(bold=True)
    headers=["Metric","Symbol","Wins","Losses","Trades"]
    for i,h in enumerate(headers):
        ws.cell(start_row+1,start_col+i,h).font=Font(bold=True)
    rows=[("Overall most wins","most_wins_instrument"),("Overall most losses","most_losses_instrument"),("FX most wins","fx_most_wins_instrument"),("FX most losses","fx_most_losses_instrument"),("Crypto most wins","crypto_most_wins_instrument"),("Crypto most losses","crypto_most_losses_instrument")]
    rr=start_row+2
    for label,key in rows:
        v=leaders.get(key) or {}
        ws.cell(rr,start_col,label).font=Font(bold=True)
        ws.cell(rr,start_col+1,v.get("symbol") or "—")
        ws.cell(rr,start_col+2,v.get("wins"))
        ws.cell(rr,start_col+3,v.get("losses"))
        ws.cell(rr,start_col+4,v.get("total_trades"))
        rr += 1
    _table_border(ws,start_row,start_col,rr-1,start_col+4)
    return rr - 1




def _parse_duration_text(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        num = float(value)
        if num >= 1000000:
            n = int(num)
            days = n // 1000000
            n %= 1000000
            hours = n // 10000
            n %= 10000
            minutes = n // 100
            seconds = n % 100
            if 0 <= hours < 24 and 0 <= minutes < 60 and 0 <= seconds < 60:
                return float(days * 86400 + hours * 3600 + minutes * 60 + seconds)
        return float(value)
    text = str(value).strip().lower()
    if not text:
        return None
    total = 0.0
    matched = False
    for n,u in re.findall(
        r'([0-9]+(?:\.[0-9]+)?)\s*(days?|hours?|minutes?|seconds?|d|h|m|s)\b',
        text,
    ):
        matched = True
        num = float(n)
        if u.startswith('day') or u == 'd': total += num*86400
        elif u.startswith('hour') or u == 'h': total += num*3600
        elif u.startswith('minute') or u == 'm': total += num*60
        else: total += num
    if matched:
        return total
    try:
        return float(text)
    except ValueError:
        return None

def _excel_datetime_to_iso(v: Any) -> str:
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).isoformat()
    if isinstance(v, (int,float)):
        try:
            base=datetime(1899,12,30)
            return (base+timedelta(days=float(v))).isoformat()
        except Exception:
            return str(v)
    return str(v or '')

def _alias_index(idx: Dict[str, int], *names: str) -> int | None:
    for n in names:
        if n in idx:
            return idx[n]
    return None


def _is_merged_non_anchor(ws, row: int, col: int) -> bool:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return not (row == merged.min_row and col == merged.min_col)
    return False


def _merged_bounds(ws, row: int, col: int) -> Tuple[int, int, int, int] | None:
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return merged.min_row, merged.min_col, merged.max_row, merged.max_col
    return None


def _merged_row_end(ws, row: int, col: int) -> int | None:
    bounds = _merged_bounds(ws, row, col)
    if not bounds:
        return None
    min_row, _min_col, max_row, _max_col = bounds
    return max_row if max_row > min_row else None


def _set_cell_horizontal_alignment(cell, horizontal: str) -> None:
    if cell.alignment.horizontal == horizontal:
        return
    alignment = copy(cell.alignment)
    alignment.horizontal = horizontal
    cell.alignment = alignment


def _apply_workbook_left_alignment(wb) -> None:
    left_alignment_id_by_source: Dict[int, int] = {}

    def left_alignment_id(source_id: int) -> int:
        if source_id in left_alignment_id_by_source:
            return left_alignment_id_by_source[source_id]
        source = wb._alignments[source_id]
        if source.horizontal == "left":
            left_alignment_id_by_source[source_id] = source_id
            return source_id
        target = copy(source)
        target.horizontal = "left"
        target_id = wb._alignments.add(target)
        left_alignment_id_by_source[source_id] = target_id
        return target_id

    default_left_alignment = Alignment(horizontal="left")
    for ws in wb.worksheets:
        non_anchor_cells = {
            (row, col)
            for merged in ws.merged_cells.ranges
            for row in range(merged.min_row, merged.max_row + 1)
            for col in range(merged.min_col, merged.max_col + 1)
            if not (row == merged.min_row and col == merged.min_col)
        }
        style_cache: Dict[Tuple[int, ...], Any] = {}
        for (row, col), cell in list(ws._cells.items()):
            if (row, col) in non_anchor_cells:
                continue
            if cell.value in (None, ""):
                continue
            source_style = cell._style
            if source_style is None:
                cell.alignment = default_left_alignment
                continue
            source_alignment_id = source_style.alignmentId
            target_alignment_id = left_alignment_id(source_alignment_id)
            if target_alignment_id == source_alignment_id:
                continue
            style_key = tuple(source_style)
            target_style = style_cache.get(style_key)
            if target_style is None:
                target_style = copy(source_style)
                target_style.alignmentId = target_alignment_id
                style_cache[style_key] = target_style
            cell._style = target_style


def _header_map(ws, header_row: int = 1) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        key = str(ws.cell(header_row, c).value or "").strip()
        if key and key not in out:
            out[key] = c
    return out




def _style_signature(cell) -> Dict[str, Any]:
    return {
        "fill": copy(cell.fill),
        "font": copy(cell.font),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
    }


def _header_style_snapshot(ws) -> List[Dict[str, Any]]:
    return [
        {
            "value": ws.cell(1, c).value,
            "style": _style_signature(ws.cell(1, c)),
        }
        for c in range(1, ws.max_column + 1)
    ]


def _auto_filter_layout_signature(ws) -> Any:
    ref = ws.auto_filter.ref if ws.auto_filter else None
    if not ref:
        return None
    try:
        min_col, min_row, max_col, _max_row = range_boundaries(ref)
        return (min_col, min_row, max_col)
    except Exception:
        return ref


def _trade_log_auto_filter_signature_is_acceptable(signature: Any) -> bool:
    """Return whether a captured Trade Log filter covers the canonical header."""
    if not isinstance(signature, (tuple, list)) or len(signature) != 3:
        return False
    min_col, min_row, max_col = signature
    return (
        min_col == 1
        and min_row == TRADE_LOG_FILTER_HEADER_ROW
        and isinstance(max_col, int)
        and max_col >= len(TRADE_LOG_HEADERS)
    )


def _trade_log_auto_filter_signature_is_canonical(signature: Any) -> bool:
    return signature == (
        1,
        TRADE_LOG_FILTER_HEADER_ROW,
        len(TRADE_LOG_HEADERS),
    )


def _worksheet_layout_snapshot(ws) -> Dict[str, Any]:
    return {
        "merged": sorted(str(r) for r in ws.merged_cells.ranges),
        "row_heights": {k: v.height for k, v in ws.row_dimensions.items() if v.height is not None},
        "col_widths": {k: v.width for k, v in ws.column_dimensions.items()},
        "hidden_cols": {k: bool(v.hidden) for k, v in ws.column_dimensions.items() if v.hidden},
        "freeze": ws.freeze_panes,
        "auto_filter": _auto_filter_layout_signature(ws),
    }


def _contract_xml(value: Any) -> str:
    """Return a stable XML representation for an openpyxl serialisable value."""
    if value is None:
        return ""
    to_tree = getattr(value, "to_tree", None)
    if not callable(to_tree):
        return str(value)
    return ET.tostring(to_tree(), encoding="unicode")


def _contract_cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return {"type": type(value).__name__, "value": value.isoformat()}
    if isinstance(value, Decimal):
        return {"type": "Decimal", "value": str(value)}
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if hasattr(value, "text") or hasattr(value, "ref"):
        return {
            "type": type(value).__name__,
            "text": str(getattr(value, "text", "") or ""),
            "ref": str(getattr(value, "ref", "") or ""),
        }
    return {"type": type(value).__name__, "value": str(value)}


def _dimension_contract(dimension: Any, *, column: bool) -> Dict[str, Any]:
    attributes = (
        (
            "width", "hidden", "bestFit", "outlineLevel", "collapsed",
            "style_id", "min", "max",
        )
        if column
        else (
            "height", "hidden", "outlineLevel", "collapsed", "style_id",
            "thickTop", "thickBot",
        )
    )
    return {
        attribute: getattr(dimension, attribute, None)
        for attribute in attributes
    }


def _row_dimensions_preservation_contract(ws) -> Dict[str, Any]:
    if ws.title != TRADE_LOG_SHEET:
        return {
            str(key): _dimension_contract(dimension, column=False)
            for key, dimension in sorted(ws.row_dimensions.items())
        }

    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    data_start = _trade_log_data_start_row(ws)
    coordinate_dimensions: Dict[str, Any] = {}
    dimensions_by_row_id: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for key, dimension in sorted(ws.row_dimensions.items()):
        signature = _dimension_contract(dimension, column=False)
        if not isinstance(key, int) or key < data_start:
            coordinate_dimensions[str(key)] = signature
            continue

        is_generated_data_height = (
            signature.get("height") in (None, TRADE_LOG_DATA_ROW_HEIGHT)
            and not bool(signature.get("hidden"))
            and int(signature.get("outlineLevel") or 0) == 0
            and not bool(signature.get("collapsed"))
            and int(signature.get("style_id") or 0) == 0
            and not bool(signature.get("thickTop"))
            and not bool(signature.get("thickBot"))
        )
        if is_generated_data_height:
            continue

        row_id = (
            str(ws.cell(key, row_id_col).value or "").strip()
            if row_id_col
            else ""
        )
        if row_id:
            dimensions_by_row_id[row_id].append(signature)
        else:
            coordinate_dimensions[str(key)] = signature

    return {
        "coordinates": coordinate_dimensions,
        "data_rows_by_id": {
            row_id: sorted(
                signatures,
                key=lambda item: json.dumps(item, sort_keys=True, default=str),
            )
            for row_id, signatures in sorted(dimensions_by_row_id.items())
        },
    }


def _conditional_formatting_contract(ws) -> List[Dict[str, Any]]:
    def generator_owned(key, rule) -> bool:
        sqref = str(getattr(key, "sqref", key))
        if ws.title in {STATS1_SHEET, STATS2_SHEET}:
            return _is_generated_dashboard_semantic_rule(rule)
        if ws.title == SYMBOLS_SHEET:
            return _is_generated_profit_loss_rule(rule)
        if ws.title in {TRADE_LOG_SHEET, LEGACY_ALL_TRADES_SHEET}:
            return (
                _is_generated_trade_log_win_loss_row_rule(rule, sqref)
                or _is_generated_trade_log_value_fill_rule(rule)
            )
        if ws.title == "P&L Calendar":
            return _is_generated_profit_loss_rule(rule)
        return False

    def rule_xml(rule) -> str:
        tree = rule.to_tree()
        # dxfId is a package-local style-table index and can be reassigned by
        # openpyxl while the differential formatting itself remains identical.
        tree.attrib.pop("dxfId", None)
        return ET.tostring(tree, encoding="unicode")

    entries: List[Dict[str, Any]] = []
    for key, rules in getattr(
        ws.conditional_formatting, "_cf_rules", {}
    ).items():
        kept_rules = [
            rule
            for rule in rules
            if not generator_owned(key, rule)
        ]
        if not kept_rules:
            continue
        entries.append(
            {
                "sqref": str(key.sqref),
                "rules": [
                    {
                        "rule": rule_xml(rule),
                        "dxf": _contract_xml(getattr(rule, "dxf", None)),
                    }
                    for rule in kept_rules
                ],
            }
        )
    return entries


def _is_managed_derived_sheet(sheet_name: str) -> bool:
    return sheet_name in {
        STATS1_SHEET,
        STATS2_SHEET,
        SYMBOLS_SHEET,
        TRADE_LOG_SHEET,
        LEGACY_ALL_TRADES_SHEET,
        "P&L Calendar",
        REPORT_YEARLY_SHEET,
    } or (sheet_name.isdigit() and int(sheet_name) >= REPORT_START_YEAR)


def _is_generator_owned_hyperlink_cell(
    ws,
    row: int,
    col: int,
) -> bool:
    """Return whether a managed value cell owns its hyperlink presentation."""
    if ws.title == STATS1_SHEET:
        # All STATS1 market metrics are regenerated.  Their hyperlinks may be
        # added (recommendation rows) or cleared (ordinary metric rows) as the
        # corresponding generated value changes.  Label-column links remain
        # user-owned and are still covered by the preservation contract.
        return row >= 2 and col in set(_stats1_market_columns(ws).values())

    if ws.title == SYMBOLS_SHEET:
        headers = _instrument_averages_header_map(ws)
        recommendation_columns = {
            value
            for value in (
                headers.get(STOP_RECOMMENDATION_HEADER),
                headers.get(TARGET_RECOMMENDATION_HEADER),
            )
            if value is not None
        }
        return (
            row >= _instrument_averages_data_start_row(ws)
            and col in recommendation_columns
        )

    if ws.title == TRADE_LOG_SHEET:
        headers = _trade_log_header_map(ws)
        recommendation_columns = {
            value
            for value in (
                headers.get(STOP_RECOMMENDATION_HEADER),
                headers.get(TARGET_RECOMMENDATION_HEADER),
            )
            if value is not None
        }
        return (
            row >= _trade_log_data_start_row(ws)
            and col in recommendation_columns
        )

    return False


def _is_generator_owned_comment_cell(ws, row: int, col: int) -> bool:
    if ws.title != STATS2_SHEET:
        return False
    for header_row in range(max(1, row - 8), row):
        if (
            str(ws.cell(header_row, col).value or "").strip().casefold()
            != "risk of ruin"
        ):
            continue
        row_headers = {
            str(ws.cell(header_row, candidate_col).value or "")
            .strip()
            .casefold()
            for candidate_col in range(1, ws.max_column + 1)
        }
        if {"account", "balance", "currency"}.issubset(row_headers):
            return True
    return False


def _comment_preservation_value(comment) -> Dict[str, Any]:
    return {
        "text": comment.text,
        "author": comment.author,
        "height": comment.height,
        "width": comment.width,
    }


def _hyperlink_preservation_value(hyperlink) -> Dict[str, Any]:
    return {
        # OOXML consumers treat percent-encoded and literal spaces in an
        # internal relationship target equivalently.
        "target": unquote(
            str(getattr(hyperlink, "target", "") or "")
        ).replace("\\", "/"),
        "location": str(getattr(hyperlink, "location", "") or ""),
        "display": str(getattr(hyperlink, "display", "") or ""),
        "tooltip": str(getattr(hyperlink, "tooltip", "") or ""),
    }


def _trade_log_annotation_contract(ws) -> Dict[str, Any]:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col:
        return {"row_ids": [], "records_by_row_id": {}}
    header_by_col = {col: header for header, col in headers.items() if header}
    records_by_row_id: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        row_id = str(ws.cell(row, row_id_col).value or "").strip()
        if not row_id:
            continue
        record: Dict[str, Dict[str, Any]] = {
            "comments": {},
            "formulas": {},
            "hyperlinks": {},
        }
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            header = header_by_col.get(col, f"column:{col}")
            if cell.comment is not None:
                record["comments"][header] = _comment_preservation_value(
                    cell.comment
                )
            if cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                record["formulas"][header] = _contract_cell_value(cell.value)
            hyperlink = getattr(cell, "hyperlink", None)
            if (
                hyperlink is not None
                and not _is_generator_owned_hyperlink_cell(ws, row, col)
            ):
                record["hyperlinks"][header] = (
                    _hyperlink_preservation_value(hyperlink)
                )
        records_by_row_id[row_id].append(record)
    return {
        "row_ids": sorted(records_by_row_id),
        "records_by_row_id": {
            row_id: sorted(
                records,
                key=lambda item: json.dumps(item, sort_keys=True, default=str),
            )
            for row_id, records in sorted(records_by_row_id.items())
        },
    }


def _stats1_schema_is_canonical(ws) -> bool:
    labels = [
        str(ws.cell(row, 1).value or "").strip().casefold()
        for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip()
    ]
    required = {
        "avg winning trade duration",
        "avg losing trade duration",
        "average move to break even",
        "average move to profit",
        "commission",
        "drawdown",
    }
    return (
        "source" not in labels
        and required.issubset(labels)
        and labels.count("recommendation") == 2
    )


def _workbook_preservation_contract(wb) -> Dict[str, Any]:
    """Snapshot user-owned state that preservation mode must not mutate.

    Statistics value/fill formats and report border groups are intentionally
    absent: those are generator-owned by the semantic repair.  STATS1 label
    typography and every STATS1 border edge are explicitly retained.
    """
    active = wb.active
    workbook_views = [
        _contract_xml(view)
        for view in (getattr(wb, "views", None) or [])
    ]
    contract: Dict[str, Any] = {
        "sheet_order": list(wb.sheetnames),
        "active_sheet": active.title if active is not None else None,
        "selected_sheets": [
            ws.title
            for ws in wb.worksheets
            if bool(getattr(ws.sheet_view, "tabSelected", False))
        ],
        "workbook_views": workbook_views,
        "sheets": {},
        "formulas": {},
        "unmanaged_values": {},
        "comments": {},
        "hyperlinks": {},
        "trade_log_annotations": {
            "row_ids": [],
            "records_by_row_id": {},
        },
        "stats1_schema_canonical": False,
    }
    for ws in wb.worksheets:
        trade_log_annotation_rows: Set[int] = set()
        if ws.title == TRADE_LOG_SHEET:
            contract["trade_log_annotations"] = (
                _trade_log_annotation_contract(ws)
            )
            row_id_col = _trade_log_header_map(ws).get("Row ID")
            if row_id_col:
                trade_log_annotation_rows = {
                    row
                    for row in range(
                        _trade_log_data_start_row(ws),
                        ws.max_row + 1,
                    )
                    if str(ws.cell(row, row_id_col).value or "").strip()
                }
        contract["sheets"][ws.title] = {
            "state": ws.sheet_state,
            "merged": sorted(str(item) for item in ws.merged_cells.ranges),
            "freeze": str(ws.freeze_panes or ""),
            "column_dimensions": {
                str(key): _dimension_contract(dimension, column=True)
                for key, dimension in sorted(ws.column_dimensions.items())
            },
            "row_dimensions": _row_dimensions_preservation_contract(ws),
            # The terminal row of a Trade Log filter may legitimately grow or
            # shrink with the managed data.  Its header/column footprint may not.
            "auto_filter": _auto_filter_layout_signature(ws),
            "data_validations": _contract_xml(ws.data_validations),
            "manual_conditional_formatting": _conditional_formatting_contract(ws),
            "views": _contract_xml(ws.views),
            "sheet_properties": _contract_xml(ws.sheet_properties),
            "sheet_format": _contract_xml(ws.sheet_format),
            "page_margins": _contract_xml(ws.page_margins),
            "page_setup": _contract_xml(ws.page_setup),
            "print_options": _contract_xml(ws.print_options),
            "print_area": str(ws.print_area or ""),
            "print_title_rows": str(ws.print_title_rows or ""),
            "print_title_cols": str(ws.print_title_cols or ""),
        }
        formulas: Dict[str, Any] = {}
        unmanaged_values: Dict[str, Any] = {}
        comments: Dict[str, Any] = {}
        hyperlinks: Dict[str, Any] = {}
        for (row, col), cell in sorted(ws._cells.items()):
            coordinate = f"{get_column_letter(col)}{row}"
            if (
                (
                    cell.data_type == "f"
                    or (
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    )
                )
                and row not in trade_log_annotation_rows
            ):
                formulas[coordinate] = _contract_cell_value(cell.value)
            if (
                not _is_managed_derived_sheet(ws.title)
                and cell.value not in (None, "")
            ):
                unmanaged_values[coordinate] = _contract_cell_value(cell.value)
            if (
                cell.comment is not None
                and not _is_generator_owned_comment_cell(ws, row, col)
                and row not in trade_log_annotation_rows
            ):
                comments[coordinate] = _comment_preservation_value(cell.comment)
            hyperlink = getattr(cell, "hyperlink", None)
            if (
                hyperlink is not None
                and not _is_generator_owned_hyperlink_cell(
                    ws,
                    row,
                    col,
                )
                and row not in trade_log_annotation_rows
            ):
                hyperlinks[coordinate] = _hyperlink_preservation_value(
                    hyperlink
                )
        if formulas:
            contract["formulas"][ws.title] = formulas
        if unmanaged_values:
            contract["unmanaged_values"][ws.title] = unmanaged_values
        if comments:
            contract["comments"][ws.title] = comments
        if hyperlinks:
            contract["hyperlinks"][ws.title] = hyperlinks

    if STATS1_SHEET in wb.sheetnames:
        stats1 = wb[STATS1_SHEET]
        contract["stats1_schema_canonical"] = (
            _stats1_schema_is_canonical(stats1)
        )
        contract["stats1_row_labels"] = {
            str(row): {
                "value": _contract_cell_value(stats1.cell(row, 1).value),
                "font": _contract_xml(stats1.cell(row, 1).font),
                "alignment": _contract_xml(stats1.cell(row, 1).alignment),
            }
            for row in range(1, stats1.max_row + 1)
            if stats1.cell(row, 1).value not in (None, "")
        }
        contract["stats1_borders"] = {
            f"{get_column_letter(col)}{row}": _contract_xml(
                stats1.cell(row, col).border
            )
            for row in range(1, stats1.max_row + 1)
            for col in range(1, min(4, stats1.max_column) + 1)
        }
    return contract


def _contract_sha256(contract: Mapping[str, Any]) -> str:
    payload = json.dumps(
        contract,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _assert_trade_log_annotations_preserved(
    before: Mapping[str, Any],
    after: Mapping[str, Any],
) -> None:
    before_contract = before.get("trade_log_annotations") or {}
    after_contract = after.get("trade_log_annotations") or {}
    after_row_ids = set(after_contract.get("row_ids") or [])
    before_records = before_contract.get("records_by_row_id") or {}
    after_records = after_contract.get("records_by_row_id") or {}
    changes = []

    def contains_annotations(candidate: Mapping[str, Any], source: Mapping[str, Any]) -> bool:
        for kind in ("comments", "formulas", "hyperlinks"):
            candidate_items = candidate.get(kind) or {}
            for header, value in (source.get(kind) or {}).items():
                if candidate_items.get(header) != value:
                    return False
        return True

    for row_id, source_records in before_records.items():
        if row_id not in after_row_ids:
            continue
        candidate_records = list(after_records.get(row_id) or [])
        used: Set[int] = set()
        for source_record in source_records:
            if not any(
                source_record.get(kind)
                for kind in ("comments", "formulas", "hyperlinks")
            ):
                continue
            matching_index = next(
                (
                    index
                    for index, candidate in enumerate(candidate_records)
                    if index not in used
                    and contains_annotations(candidate, source_record)
                ),
                None,
            )
            if matching_index is None:
                changes.append(
                    (row_id, source_record, candidate_records)
                )
            else:
                used.add(matching_index)

    if changes:
        preview = changes[:20]
        remainder = len(changes) - len(preview)
        suffix = f" (+{remainder} more)" if remainder else ""
        raise RuntimeError(
            "Workbook preservation contract changed Trade Log annotations/formulas "
            f"for surviving rows: {preview!r}{suffix}."
        )


def _assert_trade_log_row_dimensions_preserved(
    before_sheet: Mapping[str, Any],
    after_sheet: Mapping[str, Any],
    after_row_ids: Set[str],
) -> None:
    before_contract = before_sheet.get("row_dimensions") or {}
    after_contract = after_sheet.get("row_dimensions") or {}
    before_coordinates = before_contract.get("coordinates") or {}
    after_coordinates = after_contract.get("coordinates") or {}
    coordinate_changes = [
        (
            coordinate,
            signature,
            after_coordinates.get(coordinate),
        )
        for coordinate, signature in before_coordinates.items()
        if after_coordinates.get(coordinate) != signature
    ]
    coordinate_changes.extend(
        [
            (
                coordinate,
                None,
                after_coordinates.get(coordinate),
            )
            for coordinate in set(after_coordinates) - set(before_coordinates)
            if (
                not str(coordinate).isdigit()
                or int(coordinate) < TRADE_LOG_DATA_START_ROW
            )
        ]
    )
    if coordinate_changes:
        coordinate_changes.sort(
            key=lambda item: (
                not str(item[0]).isdigit(),
                int(item[0]) if str(item[0]).isdigit() else str(item[0]),
            )
        )
        preview = coordinate_changes[:20]
        remainder = len(coordinate_changes) - len(preview)
        suffix = f" (+{remainder} more)" if remainder else ""
        raise RuntimeError(
            "Workbook preservation contract changed coordinate-owned Trade "
            f"Log row dimensions: {preview!r}{suffix}."
        )

    before_by_id = before_contract.get("data_rows_by_id") or {}
    after_by_id = after_contract.get("data_rows_by_id") or {}
    changes = []
    for row_id, source_signatures in before_by_id.items():
        if row_id not in after_row_ids:
            continue
        candidate_signatures = list(after_by_id.get(row_id) or [])
        used: Set[int] = set()
        for signature in source_signatures:
            matching_index = next(
                (
                    index
                    for index, candidate in enumerate(candidate_signatures)
                    if index not in used and candidate == signature
                ),
                None,
            )
            if matching_index is None:
                changes.append((row_id, signature, candidate_signatures))
            else:
                used.add(matching_index)
    if changes:
        preview = changes[:20]
        remainder = len(changes) - len(preview)
        suffix = f" (+{remainder} more)" if remainder else ""
        raise RuntimeError(
            "Workbook preservation contract changed custom Trade Log row "
            f"dimensions for surviving rows: {preview!r}{suffix}."
        )


def _assert_workbook_preservation_contract(
    before: Mapping[str, Any],
    after: Mapping[str, Any],
) -> None:
    before_order = list(before.get("sheet_order") or [])
    after_order = list(after.get("sheet_order") or [])
    retained_before_order = [
        name for name in before_order if name != "_Trade Meta"
    ]
    existing_after = [
        name for name in after_order if name in retained_before_order
    ]
    if existing_after != retained_before_order:
        raise RuntimeError(
            "Workbook preservation contract changed existing sheet order: "
            f"before={before_order!r}; after={after_order!r}."
        )
    unexpected_new = [
        name
        for name in after_order
        if name not in before_order and not _is_managed_derived_sheet(name)
    ]
    if unexpected_new:
        raise RuntimeError(
            "Workbook preservation contract added unexpected sheets: "
            f"{unexpected_new!r}."
        )

    scalar_keys = [
        "active_sheet", "selected_sheets", "workbook_views", "formulas",
        "unmanaged_values", "comments", "hyperlinks",
    ]
    # Canonical STATS1 is the user's authoritative presentation and must be
    # identical.  Historical Source-row schemas are intentionally compacted by
    # the existing migration path, so their row coordinates and border edges
    # cannot be compared before versus after that structural migration.
    if bool(before.get("stats1_schema_canonical")):
        scalar_keys.extend(("stats1_row_labels", "stats1_borders"))
    for key in scalar_keys:
        if before.get(key) != after.get(key):
            detail = ""
            if key in {"comments", "hyperlinks"}:
                before_links = before.get(key) or {}
                after_links = after.get(key) or {}
                changes = []
                for sheet_name in sorted(set(before_links) | set(after_links)):
                    sheet_before = before_links.get(sheet_name) or {}
                    sheet_after = after_links.get(sheet_name) or {}
                    for coordinate in sorted(
                        set(sheet_before) | set(sheet_after)
                    ):
                        if sheet_before.get(coordinate) != sheet_after.get(coordinate):
                            changes.append(
                                (
                                    sheet_name,
                                    coordinate,
                                    sheet_before.get(coordinate),
                                    sheet_after.get(coordinate),
                                )
                            )
                preview = changes[:20]
                remainder = len(changes) - len(preview)
                suffix = f" (+{remainder} more)" if remainder else ""
                detail = f" at {preview!r}{suffix}"
            elif key == "stats1_row_labels":
                before_labels = before.get(key) or {}
                after_labels = after.get(key) or {}
                changes = [
                    (
                        row,
                        before_labels.get(row),
                        after_labels.get(row),
                    )
                    for row in sorted(
                        set(before_labels) | set(after_labels),
                        key=lambda value: (
                            not str(value).isdigit(),
                            int(value) if str(value).isdigit() else str(value),
                        ),
                    )
                    if before_labels.get(row) != after_labels.get(row)
                ]
                preview = changes[:20]
                remainder = len(changes) - len(preview)
                suffix = f" (+{remainder} more)" if remainder else ""
                detail = f" at {preview!r}{suffix}"
            raise RuntimeError(
                f"Workbook preservation contract changed {key}{detail}."
            )
    _assert_trade_log_annotations_preserved(before, after)
    before_sheets = before.get("sheets") or {}
    after_sheets = after.get("sheets") or {}
    for sheet_name, sheet_before in before_sheets.items():
        sheet_after = after_sheets.get(sheet_name)
        if sheet_name == "_Trade Meta" and sheet_after is None:
            continue
        comparable_before = dict(sheet_before)
        comparable_after = dict(sheet_after or {})
        if sheet_name == TRADE_LOG_SHEET:
            _assert_trade_log_row_dimensions_preserved(
                sheet_before,
                sheet_after or {},
                set(
                    (
                        after.get("trade_log_annotations") or {}
                    ).get("row_ids")
                    or []
                ),
            )
            comparable_before.pop("row_dimensions", None)
            comparable_after.pop("row_dimensions", None)
            before_filter = comparable_before.get("auto_filter")
            after_filter = comparable_after.get("auto_filter")
            if (
                before_filter != after_filter
                and not _trade_log_auto_filter_signature_is_acceptable(
                    before_filter
                )
                and _trade_log_auto_filter_signature_is_canonical(
                    after_filter
                )
            ):
                comparable_before["auto_filter"] = after_filter
        if comparable_after != comparable_before:
            changed = sorted(
                key
                for key in set(comparable_before) | set(comparable_after)
                if comparable_before.get(key) != comparable_after.get(key)
            )
            raise RuntimeError(
                "Workbook preservation contract changed user-owned layout on "
                f"{sheet_name}: {changed!r}."
            )


def _managed_statistics_presentation_contract(wb) -> Dict[str, Any]:
    """Fingerprint generator-owned metric formats/fills and report borders."""
    contract: Dict[str, Any] = {}
    if STATS1_SHEET in wb.sheetnames:
        ws = wb[STATS1_SHEET]
        market_cols = _stats1_market_columns(ws)
        policies = _stats1_semantic_row_policies(ws)
        contract[STATS1_SHEET] = {
            f"{get_column_letter(col)}{row}": {
                "kind": policies[row][0],
                "semantic": policies[row][1],
                "number_format": ws.cell(row, col).number_format,
                "fill": _contract_xml(ws.cell(row, col).fill),
                "font": _contract_xml(ws.cell(row, col).font),
            }
            for row in sorted(policies)
            for col in sorted(market_cols.values())
        }
    for sheet_name in wb.sheetnames:
        if not (
            sheet_name == REPORT_YEARLY_SHEET
            or (sheet_name.isdigit() and int(sheet_name) >= REPORT_START_YEAR)
        ):
            continue
        ws = wb[sheet_name]
        rows = _report_spec_rows(ws)
        report_contract: Dict[str, Any] = {}
        for row, (_label, _key, kind, semantic) in rows:
            for col in range(2, ws.max_column + 1):
                cell = ws.cell(row, col)
                report_contract[f"{get_column_letter(col)}{row}"] = {
                    "kind": kind,
                    "semantic": _effective_semantic_fill(kind, semantic),
                    "number_format": cell.number_format,
                    "fill": _contract_xml(cell.fill),
                    "font": _contract_xml(cell.font),
                }
        trailing_rows = _report_trailing_blank_rows(ws)
        for row in trailing_rows:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                report_contract[f"{get_column_letter(col)}{row}"] = {
                    "kind": "trailing_blank",
                    "semantic": None,
                    "number_format": cell.number_format,
                    "fill": _contract_xml(cell.fill),
                    "font": _contract_xml(cell.font),
                }
        last_contract_row = max(
            [r for r, _spec in rows] + trailing_rows,
            default=0,
        )
        report_contract["__borders__"] = {
            f"{get_column_letter(col)}{row}": _contract_xml(
                ws.cell(row, col).border
            )
            for row in range(1, last_contract_row + 1)
            for col in range(1, ws.max_column + 1)
        }
        contract[sheet_name] = report_contract
    return contract


def _snapshot_trade_log_data_presentations_by_row_id(
    ws,
    cell_presentations: Dict[Tuple[int, int], Dict[str, Any]],
) -> Dict[str, List[Dict[str, Any]]]:
    """Capture Trade Log data presentation by stable row identity, not coordinate."""
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col:
        return {}

    presentations_by_row_id: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    header_columns = [(header, col) for header, col in headers.items() if header]
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        row_id = str(ws.cell(row, row_id_col).value or "").strip()
        if not row_id:
            continue
        values: Dict[str, Any] = {}
        cells: Dict[str, Dict[str, Any]] = {}
        for header, col in header_columns:
            coordinate = (row, col)
            presentation = cell_presentations.get(coordinate)
            if presentation is None:
                cell = ws.cell(row, col)
                presentation = {
                    "value": cell.value,
                    "style": copy(cell._style),
                    "hyperlink": copy(getattr(cell, "hyperlink", None)),
                    "comment": copy(getattr(cell, "comment", None)),
                }
            presentation = dict(presentation)
            value = presentation.get("value")
            presentation["authored_formula"] = (
                value
                if isinstance(value, str) and value.startswith("=")
                else None
            )
            values[header] = value
            cells[header] = presentation
        presentations_by_row_id[row_id].append(
            {
                "source_row": row,
                "row_height": ws.row_dimensions[row].height,
                "values": values,
                "cells": cells,
            }
        )
    return dict(presentations_by_row_id)


def _trade_log_presentation_values_equal(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, (int, float, Decimal)) and not isinstance(left, bool):
        if isinstance(right, (int, float, Decimal)) and not isinstance(right, bool):
            try:
                return abs(Decimal(str(left)) - Decimal(str(right))) <= Decimal("1e-12")
            except (InvalidOperation, ValueError):
                pass
    if left == right:
        return True
    return str(left).strip() == str(right).strip()


def _trade_log_presentation_match_score(
    candidate: Dict[str, Any],
    ws,
    row: int,
    headers: Dict[str, int],
) -> Tuple[int, int, int, int, int, int]:
    identity_headers = {
        "Open Time", "Close Time", "Account", "Symbol", "Side",
        TRADE_NUMBER_HEADER, "Net P/L", "Row Type",
    }
    values = candidate.get("values") if isinstance(candidate.get("values"), dict) else {}
    identity_matches = 0
    identity_mismatches = 0
    matches = 0
    mismatches = 0
    completeness = 0
    for header, col in headers.items():
        if not header or header == "Row ID":
            continue
        previous = values.get(header)
        current = ws.cell(row, col).value
        if previous not in (None, ""):
            completeness += 1
        if previous in (None, "") and current in (None, ""):
            continue
        equal = _trade_log_presentation_values_equal(previous, current)
        if equal:
            matches += 1
            if header in identity_headers:
                identity_matches += 1
        else:
            mismatches += 1
            if header in identity_headers:
                identity_mismatches += 1
    source_row = int(candidate.get("source_row") or row)
    return (
        identity_matches,
        -identity_mismatches,
        matches,
        -mismatches,
        completeness,
        -abs(source_row - row),
    )


def _restore_trade_log_data_presentations_by_row_id(
    ws,
    presentations_by_row_id: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, Any]:
    """Move authored presentation with surviving Trade Log rows after a rebuild."""
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col or not presentations_by_row_id:
        return {
            "trade_log_row_presentations_restored_by_row_id": 0,
            "trade_log_cell_presentations_restored_by_row_id": 0,
            "trade_log_formula_cells_restored_by_row_id": 0,
            "trade_log_formula_rows_restored_by_row_id": 0,
        }

    used_candidate_indexes: Dict[str, Set[int]] = defaultdict(set)
    unmatched_row_ids: Set[str] = set()
    restored_rows = 0
    restored_cells = 0
    restored_formula_cells = 0
    restored_formula_row_ids: Set[str] = set()
    last_row = _trade_log_last_populated_row(ws)
    for row in range(_trade_log_data_start_row(ws), last_row + 1):
        row_id = str(ws.cell(row, row_id_col).value or "").strip()
        if not row_id:
            continue
        candidates = presentations_by_row_id.get(row_id) or []
        available = [
            (index, candidate)
            for index, candidate in enumerate(candidates)
            if index not in used_candidate_indexes[row_id]
        ]
        if not available:
            unmatched_row_ids.add(row_id)
            continue
        candidate_index, candidate = max(
            available,
            key=lambda item: _trade_log_presentation_match_score(
                item[1], ws, row, headers
            ),
        )
        used_candidate_indexes[row_id].add(candidate_index)
        cells = candidate.get("cells") if isinstance(candidate.get("cells"), dict) else {}
        for header, col in headers.items():
            if not header:
                continue
            presentation = cells.get(header)
            if not isinstance(presentation, dict):
                continue
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell._style = copy(presentation["style"])
            authored_formula = presentation.get("authored_formula")
            if (
                isinstance(authored_formula, str)
                and authored_formula.startswith("=")
            ):
                cell.value = authored_formula
                restored_formula_cells += 1
                restored_formula_row_ids.add(row_id)
            if hasattr(cell, "hyperlink"):
                cell.hyperlink = copy(presentation.get("hyperlink"))
            if hasattr(cell, "comment"):
                cell.comment = copy(presentation.get("comment"))
            restored_cells += 1
        ws.row_dimensions[row].height = candidate.get("row_height")
        restored_rows += 1

    diagnostics: Dict[str, Any] = {
        "trade_log_row_presentations_restored_by_row_id": restored_rows,
        "trade_log_cell_presentations_restored_by_row_id": restored_cells,
        "trade_log_formula_cells_restored_by_row_id": restored_formula_cells,
        "trade_log_formula_rows_restored_by_row_id": len(
            restored_formula_row_ids
        ),
    }
    if unmatched_row_ids:
        diagnostics["trade_log_row_presentation_unmatched_row_ids"] = sorted(
            unmatched_row_ids
        )
    return diagnostics


def _snapshot_invariants(wb) -> Dict[str, Any]:
    out: Dict[str, Any] = {"sheetnames": list(wb.sheetnames)}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out[f"{sheet_name}_layout"] = _worksheet_layout_snapshot(ws)
    dash = wb[STATS1_SHEET] if STATS1_SHEET in wb.sheetnames else (wb[LEGACY_DASHBOARD_SHEET] if LEGACY_DASHBOARD_SHEET in wb.sheetnames else None)
    if dash is not None:
        out["dash_cf"] = [str(k.sqref) for k in dash.conditional_formatting._cf_rules.keys()]
        out["dash_styles"] = {
            (r, c): _style_signature(dash.cell(r, c))
            for r in range(1, dash.max_row + 1)
            for c in range(1, dash.max_column + 1)
        }
    for name, prefix in ((ALL_TRADES_SHEET, "all_trades"), (TRADE_LOG_SHEET, "trade_log"), (SYMBOLS_SHEET, "instrument")):
        try:
            ws = _get_all_trades_sheet(wb) if name in {ALL_TRADES_SHEET, TRADE_LOG_SHEET} and prefix in {"all_trades", "trade_log"} else (wb[name] if name in wb.sheetnames else None)
        except Exception:
            ws = None
        if ws is not None:
            out[f"{prefix}_headers"] = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            out[f"{prefix}_header_styles"] = _header_style_snapshot(ws)
        ref = ws.auto_filter.ref if ws and ws.auto_filter else None
        out[f"{prefix}_filter_present"] = bool(ref)
        if ref:
            min_col, min_row, _, _ = range_boundaries(ref)
            out[f"{prefix}_filter_min_col"] = min_col
            out[f"{prefix}_filter_min_row"] = min_row
        else:
            out[f"{prefix}_filter_min_col"] = None
            out[f"{prefix}_filter_min_row"] = None
    if "P&L Calendar" in wb.sheetnames:
        ws = wb["P&L Calendar"]
        out["pnl_calendar_layout"] = _worksheet_layout_snapshot(ws)
        out["pnl_calendar_dimensions"] = (ws.max_row, ws.max_column)
    return out


def _workbook_content_snapshot(wb) -> Dict[str, int]:
    trade_log = _get_trade_log_sheet(wb)
    instrument = wb[SYMBOLS_SHEET] if SYMBOLS_SHEET in wb.sheetnames else (wb[LEGACY_INSTRUMENT_AVERAGES_SHEET] if LEGACY_INSTRUMENT_AVERAGES_SHEET in wb.sheetnames else None)
    calendar_ws = wb["P&L Calendar"] if "P&L Calendar" in wb.sheetnames else None
    instrument_rows = 0
    if instrument is not None:
        instrument_data_start = _instrument_averages_data_start_row(instrument)
        instrument_rows = sum(
            1 for row in range(instrument_data_start, instrument.max_row + 1)
            if any(instrument.cell(row, col).value not in (None, "") for col in range(1, instrument.max_column + 1))
        )
    calendar_cells = 0
    if calendar_ws is not None:
        calendar_cells = sum(
            1 for row in calendar_ws.iter_rows()
            for cell in row if cell.value not in (None, "")
        )
    return {
        "trade_log_data_rows": _trade_log_data_row_count(trade_log),
        "instrument_average_data_rows": instrument_rows,
        "pnl_calendar_populated_cells": calendar_cells,
    }


def _assert_workbook_content_not_wiped(
    before: Dict[str, int],
    after: Dict[str, int],
    *,
    migration_performed: bool,
    allow_empty_test_only_derived_sections: bool = False,
) -> None:
    labels = {
        "trade_log_data_rows": "Trade Log data rows",
        "instrument_average_data_rows": "Instrument Averages data rows",
        "pnl_calendar_populated_cells": "P&L Calendar populated cells",
    }
    for key, label in labels.items():
        before_value = int(before.get(key) or 0)
        after_value = int(after.get(key) or 0)
        if before_value > 0 and after_value == 0:
            if (
                allow_empty_test_only_derived_sections
                and key
                in {
                    "instrument_average_data_rows",
                    "pnl_calendar_populated_cells",
                }
            ):
                continue
            raise RuntimeError(f"Workbook update aborted because {label} would be wiped (before={before_value}, after=0).")
        if migration_performed and after_value < before_value:
            raise RuntimeError(
                f"Workbook schema migration aborted because {label} dropped: "
                f"before={before_value}, after={after_value}."
            )


def _assert_invariants_unchanged(before: Dict[str, Any], after: Dict[str, Any]) -> None:
    skipped = {
        "pnl_calendar_layout",
        "pnl_calendar_dimensions",
        "P&L Calendar_layout",
        "dash_cf",
        "dash_styles",
        "STATS1_layout",
        "SYMBOLS_layout",
        "instrument_filter_min_row",
        "all_trades_filter_min_row",
        "trade_log_filter_min_row",
    }
    for key in before.keys() | after.keys():
        if key in skipped:
            continue
        if key == f"{TRADE_LOG_SHEET}_layout":
            before_layout = dict(before.get(key) or {})
            after_layout = dict(after.get(key) or {})
            before_layout.pop("row_heights", None)
            after_layout.pop("row_heights", None)
            if before_layout == after_layout:
                continue
        if before.get(key) != after.get(key):
            raise RuntimeError(
                f"Workbook structural invariant changed: {key}; "
                f"before={before.get(key)!r}; after={after.get(key)!r}"
            )


def _assert_filter_covers_data(ws, *, sheet_name: str, header_row: int = 1, required_headers: List[str] | None = None, header_map: Dict[str, int] | None = None) -> None:
    ref = ws.auto_filter.ref if ws.auto_filter else None
    if not ref:
        raise RuntimeError(f"{sheet_name} filter missing.")
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if min_row != header_row or min_col != 1:
        raise RuntimeError(f"{sheet_name} filter starts at invalid range {ref}.")
    headers = header_map or _header_map(ws, header_row=header_row)
    required_headers = required_headers or []
    for h in required_headers:
        col = headers.get(h)
        if col and col > max_col:
            raise RuntimeError(f"{sheet_name} filter does not include required column '{h}'.")
    last_row = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1)):
            last_row = r
    if max_row < last_row:
        raise RuntimeError(f"{sheet_name} filter excludes populated rows.")


def _shift_dashboard_range_rows(cell_range: str, start_row: int, amount: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if max_row < start_row:
        return cell_range
    if min_row >= start_row:
        min_row += amount
    max_row += amount
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _delete_dashboard_rows_preserving_layout(ws, row_idx: int, amount: int) -> None:
    delete_end = row_idx + amount - 1

    def shifted_range(cell_range: str) -> str | None:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        if max_row < row_idx:
            return cell_range
        if min_row > delete_end:
            min_row -= amount
            max_row -= amount
        elif min_row >= row_idx and max_row <= delete_end:
            return None
        else:
            removed = max(0, min(max_row, delete_end) - max(min_row, row_idx) + 1)
            if min_row >= row_idx:
                min_row = row_idx
            max_row -= removed
            if max_row < min_row:
                return None
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    merged_ranges = [str(merged) for merged in ws.merged_cells.ranges]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))

    shifted_cf = OrderedDict()
    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        parts = [shifted_range(part) for part in str(key.sqref).split()]
        parts = [part for part in parts if part]
        if not parts:
            continue
        shifted_key = copy(key)
        shifted_key.sqref = " ".join(parts)
        shifted_cf[shifted_key] = rules

    row_heights = {idx: dim.height for idx, dim in ws.row_dimensions.items()}
    ws.delete_rows(row_idx, amount)
    for idx in list(ws.row_dimensions.keys()):
        del ws.row_dimensions[idx]
    for idx, height in row_heights.items():
        if row_idx <= idx <= delete_end:
            continue
        target = idx - amount if idx > delete_end else idx
        ws.row_dimensions[target].height = height

    for cell_range in merged_ranges:
        shifted = shifted_range(cell_range)
        if shifted:
            ws.merge_cells(shifted)
    ws.conditional_formatting._cf_rules = shifted_cf

def _remove_dashboard_metric_pair(
    ws, label: str, diagnostics: Dict[str, Any] | None = None
) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    wanted = str(label or "").strip().lower()
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value or "").strip().lower() != wanted:
            continue
        amount = 1
        if row + 1 <= ws.max_row and str(ws.cell(row + 1, 1).value or "").strip().lower() == "source":
            amount = 2
        _delete_dashboard_rows_preserving_layout(ws, row, amount)
        diagnostics.setdefault("removed_dashboard_metric_rows", {})[label] = amount
        return True
    return False

def _dashboard_row_snapshot(ws, row: int) -> Dict[str, Any]:
    cells = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        cells.append({
            "value": cell.value,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
            "protection": copy(cell.protection),
            "comment": copy(cell.comment),
            "hyperlink": copy(cell.hyperlink),
        })
    return {"cells": cells, "height": ws.row_dimensions[row].height}

def _restore_dashboard_row_snapshot(ws, row: int, snapshot: Dict[str, Any]) -> None:
    ws.row_dimensions[row].height = snapshot.get("height")
    for col, state in enumerate(snapshot["cells"], start=1):
        cell = ws.cell(row, col)
        cell.value = state["value"]
        cell.font = copy(state["font"])
        cell.fill = copy(state["fill"])
        cell.border = copy(state["border"])
        cell.alignment = copy(state["alignment"])
        cell.number_format = state["number_format"]
        cell.protection = copy(state["protection"])
        cell.comment = copy(state["comment"])
        cell.hyperlink = copy(state["hyperlink"])

def _move_dashboard_row_preserving_layout(ws, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    snapshot = _dashboard_row_snapshot(ws, source_row)
    insert_at = target_row if source_row > target_row else target_row + 1
    style_row = source_row + 1 if insert_at <= source_row else source_row
    insert_at = _insert_dashboard_rows_preserving_layout(ws, insert_at, 1, style_row)
    _restore_dashboard_row_snapshot(ws, insert_at, snapshot)
    delete_at = source_row + 1 if insert_at <= source_row else source_row
    _delete_dashboard_rows_preserving_layout(ws, delete_at, 1)

def _repair_dashboard_extreme_metric_order(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    positions = {}
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip().lower()
        if label in {"max win %", "max loss %"}:
            positions[label] = row
    win_row = positions.get("max win %")
    loss_row = positions.get("max loss %")
    if not win_row or not loss_row or win_row < loss_row:
        return False
    if win_row != loss_row + 2:
        raise RuntimeError("Dashboard Max win % / Max loss % rows are not a movable adjacent pair.")
    if any(str(ws.cell(row + 1, 1).value or "").strip().lower() != "source" for row in (loss_row, win_row)):
        raise RuntimeError("Dashboard Max win % / Max loss % source rows are missing.")
    loss_pair = [_dashboard_row_snapshot(ws, loss_row), _dashboard_row_snapshot(ws, loss_row + 1)]
    win_pair = [_dashboard_row_snapshot(ws, win_row), _dashboard_row_snapshot(ws, win_row + 1)]
    for offset, snapshot in enumerate(win_pair + loss_pair):
        _restore_dashboard_row_snapshot(ws, loss_row + offset, snapshot)
    diagnostics["reordered_dashboard_extreme_metric_rows"] = True
    return True

def _repair_dashboard_core_layout(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    _remove_dashboard_metric_pair(ws, "Max gain", diagnostics)
    _repair_dashboard_extreme_metric_order(ws, diagnostics)
    _apply_dashboard_requested_semantic_fills(ws)

def _insert_dashboard_rows_preserving_layout(ws, row_idx: int, amount: int, style_row: int) -> int:
    while True:
        blocking = next(
            (
                merged for merged in ws.merged_cells.ranges
                if merged.min_row < row_idx <= merged.max_row
            ),
            None,
        )
        if blocking is None:
            break
        row_idx = blocking.max_row + 1
    merged_ranges = [str(merged) for merged in ws.merged_cells.ranges]
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    shifted_merges = [_shift_dashboard_range_rows(cell_range, row_idx, amount) for cell_range in merged_ranges]

    shifted_cf = OrderedDict()
    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        shifted_key = copy(key)
        shifted_key.sqref = " ".join(
            _shift_dashboard_range_rows(part, row_idx, amount) for part in str(key.sqref).split()
        )
        shifted_cf[shifted_key] = rules

    ws.insert_rows(row_idx, amount)
    for cell_range in shifted_merges:
        ws.merge_cells(cell_range)
    ws.conditional_formatting._cf_rules = shifted_cf

    source_height = ws.row_dimensions[style_row].height
    for target_row in range(row_idx, row_idx + amount):
        if source_height is not None:
            ws.row_dimensions[target_row].height = source_height
        for col in range(1, ws.max_column + 1):
            if _is_merged_non_anchor(ws, target_row, col):
                continue
            source = ws.cell(style_row, col)
            target = ws.cell(target_row, col)
            _copy_cell_style(source, target)
            target.value = None
            target.comment = None
            target.hyperlink = None
    return row_idx


def _ensure_dashboard_move_duration_rows(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols: Dict[str, int] = {}
    for row in range(1, min(5, ws.max_row) + 1):
        tokens = {str(ws.cell(row, col).value or "").strip().lower(): col for col in range(1, min(8, ws.max_column) + 1)}
        candidate = {
            "overall": tokens.get("overall"),
            "fx": tokens.get("fx") or tokens.get("forex"),
            "crypto": tokens.get("crypto"),
        }
        if all(candidate.values()) and candidate["overall"] + 1 == candidate["fx"] and candidate["fx"] + 1 == candidate["crypto"]:
            market_cols = {key: int(value) for key, value in candidate.items() if value}
            break
    if not market_cols or market_cols["overall"] <= 1:
        return False

    label_col = market_cols["overall"] - 1
    aliases = {
        DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL.lower(): DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        "move to break even (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        "move to break-even (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        "average move to break-even (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        "average move to break even (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        DASHBOARD_MOVE_TO_PROFIT_LABEL.lower(): DASHBOARD_MOVE_TO_PROFIT_LABEL,
        "move to profit (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_PROFIT_LABEL,
        "average move to profit (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_PROFIT_LABEL,
    }

    def label_rows() -> Dict[str, int]:
        found: Dict[str, int] = {}
        for row in range(1, ws.max_row + 1):
            raw = str(ws.cell(row, label_col).value or "").strip().lower()
            canonical = aliases.get(raw)
            if canonical and canonical not in found:
                found[canonical] = row
        return found

    changed = False
    existing_rows = label_rows()
    for canonical in (DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL, DASHBOARD_MOVE_TO_PROFIT_LABEL):
        row = existing_rows.get(canonical)
        if row and ws.cell(row, label_col).value != canonical:
            ws.cell(row, label_col).value = canonical
            diagnostics.setdefault("renamed_dashboard_metric_labels", []).append(canonical)
            changed = True

    duration_row = next((
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, label_col).value or "").strip().lower() in {"avg duration", "avg duration (dd:hh:mm:ss)"}
    ), None)
    if duration_row is None:
        diagnostics.setdefault("missing_dashboard_metric_labels", []).append("Avg duration (DD:HH:MM:SS)")
        return changed

    rows = label_rows()
    if DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL not in rows:
        insert_at = rows.get(DASHBOARD_MOVE_TO_PROFIT_LABEL, duration_row + 1)
        insert_at = _insert_dashboard_rows_preserving_layout(ws, insert_at, 1, duration_row)
        ws.cell(insert_at, label_col).value = DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL
        diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL)
        changed = True

    rows = label_rows()
    if DASHBOARD_MOVE_TO_PROFIT_LABEL not in rows:
        break_even_row = rows[DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL]
        insert_at = break_even_row + 1
        insert_at = _insert_dashboard_rows_preserving_layout(ws, insert_at, 1, duration_row)
        ws.cell(insert_at, label_col).value = DASHBOARD_MOVE_TO_PROFIT_LABEL
        diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(DASHBOARD_MOVE_TO_PROFIT_LABEL)
        changed = True

    rows = label_rows()
    for canonical in (DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL, DASHBOARD_MOVE_TO_PROFIT_LABEL):
        row = rows.get(canonical)
        if row and ws.cell(row, label_col).value != canonical:
            ws.cell(row, label_col).value = canonical
            diagnostics.setdefault("renamed_dashboard_metric_labels", []).append(canonical)
            changed = True

    return changed


def _remove_stats1_generated_source_rows(
    ws,
    diagnostics: Dict[str, Any] | None = None,
) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    source_rows = [
        row for row in range(1, ws.max_row + 1)
        if any(
            str(ws.cell(row, col).value or "").strip().casefold() == "source"
            for col in range(1, ws.max_column + 1)
        )
    ]
    for source_row in reversed(source_rows):
        _delete_dashboard_rows_preserving_layout(ws, source_row, 1)
    if source_rows:
        diagnostics.setdefault("removed_dashboard_source_rows", []).extend(source_rows)
    return bool(source_rows)


def _ensure_dashboard_requested_metric_rows(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols = _stats1_market_columns(ws)
    if not market_cols or market_cols["overall"] <= 1:
        return False
    label_col = market_cols["overall"] - 1
    changed = False

    def label_at(row: int) -> str:
        return str(ws.cell(row, label_col).value or "").strip()

    def market_values_at(row: int) -> List[Any]:
        if row < 1 or row > ws.max_row:
            return []
        return [ws.cell(row, col).value for col in market_cols.values()]

    def row_has_market_values(row: int) -> bool:
        return any(value not in (None, "") for value in market_values_at(row))

    def row_values_look_like_recommendation(row: int) -> bool:
        for value in market_values_at(row):
            text = str(value or "").strip().casefold()
            if text.startswith(("reduce ", "increase ", "keep ", "need wins")):
                return True
        return False

    def row_values_look_like_source(row: int) -> bool:
        values = [value for value in market_values_at(row) if value not in (None, "")]
        if not values:
            return False
        if row_values_look_like_recommendation(row):
            return False
        return any(isinstance(value, str) for value in values)

    def set_blank_label(row: int, label: str, reason: str) -> bool:
        nonlocal changed
        if row < 1 or row > ws.max_row or label_at(row) or _is_merged_non_anchor(ws, row, label_col):
            return False
        ws.cell(row, label_col).value = label
        diagnostics.setdefault("repaired_dashboard_blank_metric_labels", []).append(reason)
        changed = True
        return True

    def find_row(label: str, *, before: int | None = None) -> int | None:
        wanted = label.casefold()
        last_row = min(ws.max_row, before - 1) if before else ws.max_row
        return next(
            (row for row in range(1, last_row + 1) if label_at(row).casefold() == wanted),
            None,
        )

    def core_boundary_row() -> int:
        return min(
            [
                row for row in range(1, ws.max_row + 1)
                if label_at(row).casefold() in {"min duration", "min duration (dd:hh:mm:ss)", "winners"}
            ] or [ws.max_row + 1]
        )

    def find_core_row(label: str) -> int | None:
        return find_row(label, before=core_boundary_row())

    replacements = {
        "avg result %": "Percentage expectancy",
        "average result percent": "Percentage expectancy",
        "average result %": "Percentage expectancy",
        "avg r": "R expectancy",
        "average r": "R expectancy",
    }
    for row in range(1, ws.max_row + 1):
        replacement = replacements.get(label_at(row).casefold())
        if replacement and label_at(row) != replacement:
            ws.cell(row, label_col).value = replacement
            diagnostics.setdefault("renamed_dashboard_metric_labels", []).append(replacement)
            changed = True

    gross_replacements = {
        "gross gain": "Gross percent gain",
        "gross loss": "Gross percent loss",
    }
    for row in range(1, min(ws.max_row, 40) + 1):
        replacement = gross_replacements.get(label_at(row).casefold())
        if replacement:
            ws.cell(row, label_col).value = replacement
            diagnostics.setdefault("renamed_dashboard_metric_labels", []).append(replacement)
            changed = True

    def insert_after(anchor_label: str, wanted_label: str) -> int | None:
        nonlocal changed
        existing = find_row(wanted_label)
        if existing:
            return existing
        anchor = find_row(anchor_label)
        if not anchor:
            return None
        insert_at = anchor + 1
        insert_at = _insert_dashboard_rows_preserving_layout(ws, insert_at, 1, anchor)
        ws.cell(insert_at, label_col).value = wanted_label
        diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(wanted_label)
        changed = True
        return insert_at

    gross_ir_gain_row = insert_after("Gross percent loss", "Gross IR gain")
    if gross_ir_gain_row and not find_row("Gross IR loss"):
        inserted = _insert_dashboard_rows_preserving_layout(ws, gross_ir_gain_row + 1, 1, gross_ir_gain_row)
        ws.cell(inserted, label_col).value = "Gross IR loss"
        diagnostics.setdefault("inserted_dashboard_metric_rows", []).append("Gross IR loss")
        changed = True

    canonical_core_order = [
        "Net P/L Percentage",
        "Net P/L R multiples",
        "Gross percent gain",
        "Gross percent loss",
        "Gross IR gain",
        "Gross IR loss",
        "Percentage expectancy",
        "R expectancy",
    ]
    target_row = find_row(canonical_core_order[0])
    if target_row:
        for metric_label in canonical_core_order:
            current_row = find_row(metric_label)
            if not current_row:
                continue
            if current_row != target_row:
                _move_dashboard_row_preserving_layout(ws, current_row, target_row)
                diagnostics.setdefault("reordered_dashboard_core_metric_rows", []).append(metric_label)
                changed = True
            target_row += 1

    best_streak_row = find_core_row("Best Win Streak") or find_core_row("Winning Streak")
    if best_streak_row and best_streak_row + 5 < core_boundary_row():
        labels = [label_at(best_streak_row + offset).casefold() for offset in range(6)]
        if labels in (
            ["best win streak", "worst losing streak", "start", "end", "start", "end"],
            ["winning streak", "losing streak", "start", "end", "start", "end"],
        ):
            snapshots = [
                _dashboard_row_snapshot(ws, best_streak_row),
                _dashboard_row_snapshot(ws, best_streak_row + 2),
                _dashboard_row_snapshot(ws, best_streak_row + 3),
                _dashboard_row_snapshot(ws, best_streak_row + 1),
                _dashboard_row_snapshot(ws, best_streak_row + 4),
                _dashboard_row_snapshot(ws, best_streak_row + 5),
            ]
            for offset, snapshot in enumerate(snapshots):
                _restore_dashboard_row_snapshot(ws, best_streak_row + offset, snapshot)
            diagnostics.setdefault("reordered_dashboard_streak_detail_rows", []).append(
                f"{best_streak_row}:{best_streak_row + 5}"
            )
            changed = True

    legacy_expectancy_row = find_core_row("Expectancy")
    if legacy_expectancy_row and find_core_row("Percentage expectancy") and find_core_row("R expectancy"):
        amount = 1
        if (
            legacy_expectancy_row + 2 <= ws.max_row
            and label_at(legacy_expectancy_row + 1).casefold() == "%"
            and label_at(legacy_expectancy_row + 2).casefold() == "r"
        ):
            amount = 3
        _delete_dashboard_rows_preserving_layout(ws, legacy_expectancy_row, amount)
        diagnostics.setdefault("removed_dashboard_metric_rows", {})["Expectancy"] = (
            diagnostics.setdefault("removed_dashboard_metric_rows", {}).get("Expectancy", 0) + amount
        )
        changed = True

    cursor_anchor = "Avg stop %"
    for metric_label in ("Min stop %", "Max stop %", "Avg target %", "Min target %", "Max target %"):
        metric_row = find_core_row(metric_label)
        if not metric_row:
            anchor = find_core_row(cursor_anchor)
            if not anchor:
                continue
            candidate = anchor + 1
            if candidate <= ws.max_row and label_at(candidate).casefold() == "source":
                candidate += 1
            elif candidate <= ws.max_row and not label_at(candidate) and row_values_look_like_source(candidate):
                set_blank_label(candidate, "Source", f"{cursor_anchor}: Source")
                candidate += 1
            if candidate <= ws.max_row and not label_at(candidate) and row_has_market_values(candidate):
                if set_blank_label(candidate, metric_label, metric_label):
                    metric_row = candidate
                else:
                    metric_row = None
            else:
                metric_row = None
            if not metric_row:
                metric_row = anchor + 1
                metric_row = _insert_dashboard_rows_preserving_layout(ws, metric_row, 1, anchor)
                ws.cell(metric_row, label_col).value = metric_label
                diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(metric_label)
                changed = True
        cursor_anchor = metric_label

    if _remove_stats1_generated_source_rows(ws, diagnostics):
        changed = True

    def ensure_recommendation_after(anchor_label: str) -> int | None:
        nonlocal changed
        anchor = find_core_row(anchor_label)
        if not anchor:
            return None
        expected = anchor + 1
        if expected <= ws.max_row and label_at(expected).casefold() == "recommendation":
            return expected
        if expected <= ws.max_row and not label_at(expected) and row_values_look_like_recommendation(expected):
            if set_blank_label(expected, "Recommendation", f"{anchor_label}: Recommendation"):
                return expected
        inserted = _insert_dashboard_rows_preserving_layout(ws, expected, 1, anchor)
        ws.cell(inserted, label_col).value = "Recommendation"
        diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(f"{anchor_label}: Recommendation")
        changed = True
        return inserted

    ensure_recommendation_after("Max stop %")
    ensure_recommendation_after("Max target %")

    def compact_recommendation_after(anchor_label: str) -> None:
        nonlocal changed
        anchor = find_core_row(anchor_label)
        if not anchor:
            return
        boundary_labels = {"targets", "duration", "winners"}
        boundary = next(
            (
                row
                for row in range(anchor + 1, ws.max_row + 1)
                if label_at(row).casefold() in boundary_labels
            ),
            ws.max_row + 1,
        )
        rec_row = anchor + 1
        if rec_row >= boundary:
            return
        for merged in list(ws.merged_cells.ranges):
            if merged.min_col <= label_col <= merged.max_col and merged.min_row < rec_row <= merged.max_row:
                ws.unmerge_cells(str(merged))
                diagnostics.setdefault("unmerged_dashboard_recommendation_label_blocks", []).append(str(merged))
                changed = True
        current_label = label_at(rec_row).casefold()
        later_recommendation_row = next(
            (
                row
                for row in range(rec_row + 1, boundary)
                if row_values_look_like_recommendation(row)
            ),
            None,
        )
        if current_label != "recommendation" and (not current_label or later_recommendation_row):
            ws.cell(rec_row, label_col).value = "Recommendation"
            diagnostics.setdefault("repaired_dashboard_blank_metric_labels", []).append(f"{anchor_label}: Recommendation")
            changed = True
        if not row_values_look_like_recommendation(rec_row) and later_recommendation_row:
            for col in market_cols.values():
                source = ws.cell(later_recommendation_row, col)
                target = ws.cell(rec_row, col)
                target.value = source.value
                target.number_format = "General"
            delete_start = rec_row + 1
            delete_amount = later_recommendation_row - rec_row
            _delete_dashboard_rows_preserving_layout(ws, delete_start, delete_amount)
            diagnostics.setdefault("removed_dashboard_stale_recommendation_rows", []).append(
                f"{anchor_label}: {delete_start}-{later_recommendation_row}"
            )
            changed = True

    compact_recommendation_after("Max stop %")
    compact_recommendation_after("Max target %")

    first_winners_row = next(
        (row for row in range(1, ws.max_row + 1) if label_at(row).casefold() == "winners"),
        ws.max_row + 1,
    )

    def find_primary_row(label: str) -> int | None:
        return find_row(label, before=first_winners_row)

    def find_primary_row_any(labels: Tuple[str, ...]) -> int | None:
        aliases = {label.casefold() for label in labels}
        return next(
            (
                row for row in range(1, first_winners_row)
                if label_at(row).casefold() in aliases
            ),
            None,
        )

    avg_duration_row = find_primary_row_any(("Avg duration", "Avg duration (DD:HH:MM:SS)"))
    if avg_duration_row:
        primary_labels = [label_at(row).casefold() for row in range(1, first_winners_row) if label_at(row)]
        primary_label_counts = Counter(primary_labels)
        preserved_primary_heights = {
            label_at(row).casefold(): ws.row_dimensions[row].height
            for row in range(1, first_winners_row)
            if (
                label_at(row)
                and label_at(row).casefold() != "source"
                and primary_label_counts[label_at(row).casefold()] == 1
                and ws.row_dimensions[row].height is not None
            )
        }
        insert_after_row = avg_duration_row
        for wanted_label, aliases in (
            ("Avg winning trade duration", ("Avg winning trade duration", "Avg winning trade duration (DD:HH:MM:SS)")),
            ("Avg losing trade duration", ("Avg losing trade duration", "Avg losing trade duration (DD:HH:MM:SS)")),
        ):
            existing = find_primary_row_any(aliases)
            if existing:
                if label_at(existing) != wanted_label:
                    ws.cell(existing, label_col).value = wanted_label
                    diagnostics.setdefault("renamed_dashboard_metric_labels", []).append(wanted_label)
                    changed = True
                insert_after_row = existing
                continue
            inserted = _insert_dashboard_rows_preserving_layout(ws, insert_after_row + 1, 1, insert_after_row)
            ws.cell(inserted, label_col).value = wanted_label
            diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(wanted_label)
            insert_after_row = inserted
            changed = True
        if preserved_primary_heights:
            for row in range(1, ws.max_row + 1):
                height = preserved_primary_heights.get(label_at(row).casefold())
                if height is not None:
                    ws.row_dimensions[row].height = height

    return changed


def _ensure_dashboard_expectancy_row(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    market_cols = _stats1_market_columns(ws)
    if not market_cols or market_cols["overall"] <= 1:
        return False
    label_col = market_cols["overall"] - 1
    removed = False
    for row in reversed(range(1, ws.max_row + 1)):
        if str(ws.cell(row, label_col).value or "").strip().casefold() != "expectancy %":
            continue
        _delete_dashboard_rows_preserving_layout(ws, row, 1)
        diagnostics.setdefault("removed_dashboard_metric_rows", {})["Expectancy %"] = (
            diagnostics.setdefault("removed_dashboard_metric_rows", {}).get("Expectancy %", 0) + 1
        )
        removed = True
    return removed


def _ensure_dashboard_extended_layout(ws, diagnostics: Dict[str, Any] | None = None) -> bool:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    labels = {
        str(ws.cell(row, 1).value or "").strip().lower()
        for row in range(1, ws.max_row + 1)
    }
    if not {"side", "patterns", "timeframe", "commission"}.issubset(labels):
        return False

    canonical_duration_labels = {
        "min duration (dd:hh:mm:ss)": "Min duration",
        "avg duration (dd:hh:mm:ss)": "Avg duration",
        "avg winning trade duration (dd:hh:mm:ss)": "Avg winning trade duration",
        "average winning trade duration (dd:hh:mm:ss)": "Avg winning trade duration",
        "avg losing trade duration (dd:hh:mm:ss)": "Avg losing trade duration",
        "average losing trade duration (dd:hh:mm:ss)": "Avg losing trade duration",
        "max duration (dd:hh:mm:ss)": "Max duration",
        "min move to break even (dd:hh:mm:ss)": "Min Move to Break Even",
        "average move to break even (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
        "max move to break even (dd:hh:mm:ss)": "Max Move to Break Even",
        "min move to profit (dd:hh:mm:ss)": "Min Move to Profit",
        "average move to profit (dd:hh:mm:ss)": DASHBOARD_MOVE_TO_PROFIT_LABEL,
        "max move to profit (dd:hh:mm:ss)": "Max Move to Profit",
    }
    changed = False
    for row in range(1, ws.max_row + 1):
        raw = str(ws.cell(row, 1).value or "").strip().lower()
        replacement = canonical_duration_labels.get(raw)
        if replacement:
            ws.cell(row, 1).value = replacement
            changed = True

    desired_by_section = {
        "Winners": [
            ("Min stop %", {"min stop %"}),
            ("Avg stop %", {"avg stop %"}),
            ("Max stop %", {"max stop %"}),
            ("Min target %", {"min target %"}),
            ("Avg target %", {"avg target %"}),
            ("Max target %", {"max target %"}),
            ("Min duration", {"min duration", "min duration (dd:hh:mm:ss)"}),
            ("Avg duration", {"avg duration", "avg duration (dd:hh:mm:ss)"}),
            ("Max duration", {"max duration", "max duration (dd:hh:mm:ss)"}),
            ("Min result %", {"min result %", "min win %"}),
            ("Percentage expectancy", {"percentage expectancy", "avg result %", "avg win %"}),
            ("Max result %", {"max result %", "max win %"}),
            ("Min R", {"min r", "min r win"}),
            ("R expectancy", {"r expectancy", "avg r", "avg r win"}),
            ("Max R", {"max r", "max r win"}),
        ],
        "Losers": [
            ("Min stop %", {"min stop %"}),
            ("Avg stop %", {"avg stop %"}),
            ("Max stop %", {"max stop %"}),
            ("Min target %", {"min target %"}),
            ("Avg target %", {"avg target %"}),
            ("Max target %", {"max target %"}),
            ("Min duration", {"min duration", "min duration (dd:hh:mm:ss)"}),
            ("Avg duration", {"avg duration", "avg duration (dd:hh:mm:ss)"}),
            ("Max duration", {"max duration", "max duration (dd:hh:mm:ss)"}),
            ("Min result %", {"min result %", "max loss %"}),
            ("Percentage expectancy", {"percentage expectancy", "avg result %", "avg loss %"}),
            ("Max result %", {"max result %", "min loss %"}),
            ("Min R", {"min r", "max r loss"}),
            ("R expectancy", {"r expectancy", "avg r", "avg r loss"}),
            ("Max R", {"max r", "min r loss"}),
        ],
    }
    for section_name, next_sections in (
        ("Winners", {"losers"}),
        ("Losers", {"side"}),
    ):
        section_row = next((
            row for row in range(1, ws.max_row + 1)
            if str(ws.cell(row, 1).value or "").strip().lower() == section_name.lower()
        ), None)
        if not section_row:
            continue
        section_end = next((
            row for row in range(section_row + 1, ws.max_row + 1)
            if str(ws.cell(row, 1).value or "").strip().lower() in next_sections
        ), ws.max_row + 1)
        protect_manual_merges = any(
            merged.min_col == 1
            and merged.max_col == 1
            and section_row < merged.min_row < section_end
            for merged in ws.merged_cells.ranges
        )
        cursor = section_row + 1
        for wanted, aliases in desired_by_section[section_name]:
            while cursor <= ws.max_row and str(ws.cell(cursor, 1).value or "").strip().casefold() in {"", "source"}:
                cursor += 1
            merged_end = _merged_row_end(ws, cursor, 1)
            if merged_end and merged_end >= cursor:
                cursor = merged_end + 1
                while cursor <= ws.max_row and str(ws.cell(cursor, 1).value or "").strip().casefold() in {"", "source"}:
                    cursor += 1
            current = str(ws.cell(cursor, 1).value or "").strip()
            if current.lower() in aliases:
                cursor += 1
                continue
            next_anchor = next((
                row for row in range(cursor, ws.max_row + 1)
                if str(ws.cell(row, 1).value or "").strip().lower() in next_sections
            ), ws.max_row + 1)
            found = next((
                row for row in range(cursor + 1, next_anchor)
                if str(ws.cell(row, 1).value or "").strip().lower() in aliases
            ), None)
            if found == cursor:
                cursor += 1
                continue
            if protect_manual_merges:
                diagnostics.setdefault("skipped_dashboard_metric_rows_due_to_manual_merges", []).append(f"{section_name}: {wanted}")
                continue
            template_row = min(max(section_row + 1, cursor), max(section_row + 1, next_anchor - 1))
            cursor = _insert_dashboard_rows_preserving_layout(ws, cursor, 1, template_row)
            ws.cell(cursor, 1).value = wanted
            diagnostics.setdefault("inserted_dashboard_metric_rows", []).append(f"{section_name}: {wanted}")
            changed = True
            cursor += 1
    return changed


def _find_anchor_sections(ws, anchors: List[str], optional: List[str] | None = None) -> Dict[str, Dict[str, int]]:
    optional = optional or []
    all_anchors = list(dict.fromkeys([*anchors, *optional]))
    found: Dict[str, Dict[str, int]] = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            text = str(ws.cell(r, c).value or "").strip().lower()
            for a in all_anchors:
                if text == a.lower() and a not in found:
                    found[a] = {"anchor_row": r, "anchor_col": c}
    missing = [a for a in anchors if a not in found]
    if missing:
        raise RuntimeError(f"Dashboard section anchors missing: {', '.join(missing)}")

    for name, meta in list(found.items()):
        ar, ac = meta["anchor_row"], meta["anchor_col"]
        same_row_right = [m["anchor_col"] for n,m in found.items() if m["anchor_row"] == ar and m["anchor_col"] > ac]
        end_col = (min(same_row_right)-1) if same_row_right else ws.max_column
        same_band_below = []
        for n,m in found.items():
            if m["anchor_row"] <= ar:
                continue
            if ac <= m["anchor_col"] <= end_col:
                same_band_below.append(m["anchor_row"])
        end_row = (min(same_band_below)-1) if same_band_below else ws.max_row
        found[name].update({"start_row": ar+1, "end_row": end_row, "start_col": ac, "end_col": end_col})
    return found


def _find_label_in_section(ws, label: str, section: Dict[str, int]) -> Tuple[int, int] | None:
    wanted = str(label or "").strip().lower()
    max_row = ws.max_row
    max_col = ws.max_column
    start_row = max(1, section.get("start_row", 1))
    end_row = min(max_row, section.get("end_row", max_row))
    start_col = max(1, section.get("start_col", 1))
    end_col = min(max_col, section.get("end_col", max_col) - 1)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            if str(ws.cell(r,c).value or "").strip().lower() == wanted:
                return r,c
    return None
def _find_label_cell(ws, label: str, search_cols: List[int] | None = None) -> tuple[int, int] | None:
    wanted = str(label or "").strip().lower()
    if not wanted:
        return None
    cols = search_cols or list(range(1, ws.max_column + 1))
    for r in range(1, ws.max_row + 1):
        for c in cols:
            if str(ws.cell(r, c).value or "").strip().lower() == wanted:
                return (r, c)
    return None

def _find_instrument_leaders_table(ws) -> tuple[int | None, Dict[str, int], Dict[str, int], int]:
    anchors: List[tuple[int, int]] = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip().lower() == "instrument leaders":
                anchors.append((r, c))
    if not anchors:
        return None, {}, {}, 1
    candidates: List[tuple[int, int, int, Dict[str, int], Dict[str, int]]] = []
    for ar, ac in anchors:
        header_row = None
        header_map: Dict[str, int] = {}
        for r in range(ar + 1, min(ws.max_row, ar + 12) + 1):
            wanted = ["metric", "symbol", "wins", "losses", "trades"]
            for c in range(ac, min(ws.max_column, ac + 8) - len(wanted) + 3):
                tokens = [str(ws.cell(r, c + offset).value or "").strip().lower() for offset in range(len(wanted))]
                if tokens != wanted:
                    continue
                header_row = r
                header_map = {token: c + offset for offset, token in enumerate(wanted)}
                break
            if header_row:
                break
        if not header_row:
            continue
        metric_rows: Dict[str, int] = {}
        for r in range(header_row + 1, min(ws.max_row, header_row + 24) + 1):
            label = str(ws.cell(r, header_map["metric"]).value or "").strip().lower()
            if label:
                metric_rows[label] = r
        candidates.append((ac, ar, header_row, header_map, metric_rows))
    if candidates:
        ac, ar, header_row, header_map, metric_rows = sorted(candidates, key=lambda t: (t[0], t[1]))[0]
        return header_row, header_map, metric_rows, ac
    first_anchor = sorted(anchors, key=lambda t: (t[1], t[0]))[0]
    return None, {}, {}, first_anchor[1]



def _copy_leader_row_cell_style(ws, source_row: int, target_row: int, header_map: Dict[str, int]) -> None:
    for col in header_map.values():
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def _repair_missing_market_leader_counterpart_row(ws, metric_rows: Dict[str, int], header_map: Dict[str, int], label: str) -> int | None:
    wanted = str(label or "").strip().lower()
    if not wanted or wanted in metric_rows or not header_map:
        return metric_rows.get(wanted)
    parts = wanted.split()
    if len(parts) < 3 or parts[-2:] not in (["most", "wins"], ["most", "losses"]):
        return None
    market = " ".join(parts[:-2])
    counterpart_suffix = "losses" if parts[-1] == "wins" else "wins"
    counterpart = f"{market} most {counterpart_suffix}"
    counterpart_row = metric_rows.get(counterpart)
    if not counterpart_row:
        return None

    metric_col = header_map["metric"]
    table_cols = sorted(header_map.values())
    first_data_row = min(metric_rows.values()) if metric_rows else counterpart_row
    last_scan_row = max(max(metric_rows.values(), default=counterpart_row) + 8, counterpart_row + 1)
    for row in range(counterpart_row + 1, min(ws.max_row + 24, last_scan_row) + 1):
        if any(ws.cell(row, col).value not in (None, "") for col in table_cols):
            continue
        source_row = counterpart_row
        existing = [r for r in metric_rows.values() if first_data_row <= r < row]
        if existing:
            source_row = max(existing)
        _copy_leader_row_cell_style(ws, source_row, row, header_map)
        ws.cell(row, metric_col).value = label
        metric_rows[wanted] = row
        return row
    raise RuntimeError(f"Instrument leaders is missing required row '{label}' and no safe blank row is available to restore it.")

def _write_value_preserving_cell(ws, row: int, col: int, value: Any) -> bool:
    if _is_merged_non_anchor(ws, row, col):
        return False
    ws.cell(row, col).value = value
    return True

def _detect_calendar_month_columns(ws) -> Dict[int, int]:
    month_cols: Dict[int, int] = {}
    names = {calendar.month_name[i].lower(): i for i in range(1, 13)}
    for c in range(1, ws.max_column + 1):
        token = str(ws.cell(1, c).value or "").strip().lower()
        if token in names:
            month_cols[names[token]] = c
    return month_cols


def _detect_calendar_month_rows(ws) -> Dict[int, int]:
    month_rows: Dict[int, int] = {}
    names = {calendar.month_name[i].lower(): i for i in range(1, 13)}
    for row in range(2, ws.max_row + 1):
        token = str(ws.cell(row, 1).value or "").strip().lower()
        if token in names:
            month_rows[names[token]] = row
    return month_rows


def _detect_calendar_year_columns(ws) -> Dict[int, int]:
    year_cols: Dict[int, int] = {}
    for col in range(2, ws.max_column + 1):
        year_value = _as_float(ws.cell(1, col).value)
        if year_value is None:
            continue
        year = int(year_value)
        if 1900 <= year <= 3000:
            year_cols[year] = col
    return year_cols


def _existing_pnl_calendar_years(ws) -> Set[int]:
    years = set(_detect_calendar_year_columns(ws).keys())
    for row in range(2, ws.max_row + 1):
        year_value = _as_float(ws.cell(row, 1).value)
        if year_value is None:
            continue
        year = int(year_value)
        if 1900 <= year <= 3000:
            years.add(year)
    return years


def _pnl_calendar_monthly_values(snapshot_or_rows: Any) -> Dict[Tuple[int, int], Dict[str, Any]]:
    rows = snapshot_or_rows.get("items") if isinstance(snapshot_or_rows, dict) else snapshot_or_rows
    monthly: Dict[Tuple[int, int], Dict[str, Any]] = defaultdict(lambda: {"pct": 0.0, "count": 0})
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        if _target_r_row_is_test_trade(row):
            continue
        if str(row.get("row_type") or "trade").strip().lower() != "trade":
            continue
        timestamp = _as_date(row.get("close_time") or row.get("open_time"))
        pct = _as_float(row.get("result_pct"))
        if not timestamp or pct is None or not math.isfinite(pct):
            continue
        key = (timestamp.year, timestamp.month)
        monthly[key]["pct"] += pct
        monthly[key]["count"] += 1
    return monthly


def _format_pnl_calendar_month_cell(pct_points: float, count: int) -> str:
    label = "trade" if int(count) == 1 else "trades"
    return f"{pct_points:.2f}%, {int(count)} {label}"


def _pnl_calendar_text_pct_fraction(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value).strip()
    if not text:
        return None
    match = re.match(r"^\s*([+-]?(?:\d+(?:\.\d*)?|\.\d+))\s*%", text)
    if not match:
        return None
    try:
        return float(match.group(1)) / 100.0
    except ValueError:
        return None


def _copy_calendar_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy(src._style)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def _write_pnl_calendar_one_line_layout(ws, snapshot: Dict[str, Any], diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    monthly = _pnl_calendar_monthly_values(snapshot)
    years = sorted(_existing_pnl_calendar_years(ws) | {year for year, _month in monthly.keys()})
    target_rows = 13
    target_cols = max(1, len(years) + 1)

    row_heights = {idx: dim.height for idx, dim in ws.row_dimensions.items() if dim.height is not None}
    month_header_template = copy(ws.cell(1, 1))
    year_header_template = copy(ws.cell(1, min(max(ws.max_column, 1), 2)))
    month_label_template = copy(ws.cell(min(max(ws.max_row, 1), 2), 1))
    value_template = copy(ws.cell(min(max(ws.max_row, 1), 2), min(max(ws.max_column, 1), 2)))

    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    if ws.max_row > target_rows:
        ws.delete_rows(target_rows + 1, ws.max_row - target_rows)
    if ws.max_column > target_cols:
        ws.delete_cols(target_cols + 1, ws.max_column - target_cols)

    for row in range(1, target_rows + 1):
        for col in range(1, target_cols + 1):
            cell = ws.cell(row, col)
            template = (
                month_header_template
                if row == 1 and col == 1
                else year_header_template
                if row == 1
                else month_label_template
                if col == 1
                else value_template
            )
            _copy_calendar_cell_style(template, cell)
            cell.value = None
            cell.number_format = "General"
            base_alignment = copy(cell.alignment)
            cell.alignment = Alignment(
                horizontal="center" if col > 1 else (base_alignment.horizontal or "left"),
                vertical=base_alignment.vertical or "center",
                wrap_text=False,
            )
            if row > 1 and col > 1:
                _clear_cell_fill(cell)

    ws.cell(1, 1).value = "Month"
    for col_idx, year in enumerate(years, start=2):
        ws.cell(1, col_idx).value = year

    for month in range(1, 13):
        row_idx = month + 1
        ws.cell(row_idx, 1).value = calendar.month_name[month]
        for col_idx, year in enumerate(years, start=2):
            values = monthly.get((year, month))
            cell = ws.cell(row_idx, col_idx)
            if not values:
                cell.value = None
                _clear_cell_fill(cell)
                continue
            pct = float(values["pct"])
            count = int(values["count"])
            cell.value = _format_pnl_calendar_month_cell(pct, count)
            if pct > 0:
                _apply_full_cell_semantic_fill(cell, "profit")
            elif pct < 0:
                _apply_full_cell_semantic_fill(cell, "loss")
            else:
                _clear_cell_fill(cell)

    for idx, height in row_heights.items():
        if idx <= target_rows:
            ws.row_dimensions[idx].height = height
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 12)
    for col in range(2, target_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0,
            18,
        )
    ws.auto_filter.ref = f"A1:{get_column_letter(target_cols)}{target_rows}"
    _style_header_row(ws, 1)
    _table_border(ws, 1, 1, target_rows, target_cols)
    ws.freeze_panes = "B2"
    _apply_pnl_calendar_profit_loss_formatting(ws)
    diagnostics["pnl_calendar_transposed_years_written"] = len(years)


def _ensure_pnl_calendar_freeze_panes(ws) -> None:
    if _detect_calendar_year_columns(ws) and _detect_calendar_month_rows(ws):
        ws.freeze_panes = "B2"
        return
    month_cols = _detect_calendar_month_columns(ws)
    if month_cols and min(month_cols.values()) == 2:
        ws.freeze_panes = "B2"
        return
    if month_cols and min(month_cols.values()) == 3:
        ws.freeze_panes = "C2"
        return
    names = {calendar.month_name[i].lower() for i in range(1, 13)}
    row_two_months = {
        c for c in range(1, ws.max_column + 1)
        if str(ws.cell(2, c).value or "").strip().lower() in names
    }
    row_one_has_month_headers = any(
        str(ws.cell(1, c).value or "").strip().lower().endswith(" p/l %")
        for c in range(1, ws.max_column + 1)
    )
    if row_one_has_month_headers and row_two_months and min(row_two_months) == 2:
        ws.freeze_panes = "B3"

def _update_pnl_calendar_preserving_layout(dst_ws, snapshot: Dict[str, Any], diagnostics: Dict[str, Any] | None = None) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    _write_pnl_calendar_one_line_layout(dst_ws, snapshot, diagnostics)

def _normalize_stats2_account_balance_headers(
    ws,
    section: Dict[str, int],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    start_row = max(1, section.get("start_row", 1))
    end_row = min(section.get("end_row", ws.max_row), start_row + 7)
    for row in range(start_row, end_row + 1):
        headers = {
            str(ws.cell(row, col).value or "").strip().casefold(): col
            for col in range(section["start_col"], section["end_col"] + 1)
            if str(ws.cell(row, col).value or "").strip()
        }
        if not {"account", "balance", "currency"}.issubset(headers):
            continue
        risk_cols = [
            col for col in range(section["start_col"], section["end_col"] + 1)
            if str(ws.cell(row, col).value or "").strip().casefold() == "risk of ruin"
        ]
        as_of_cols = [
            col for col in range(section["start_col"], section["end_col"] + 1)
            if str(ws.cell(row, col).value or "").strip().casefold() in {"as of", "as_of"}
        ]
        net_cols = [
            col for col in range(section["start_col"], section["end_col"] + 1)
            if str(ws.cell(row, col).value or "").strip().casefold() in {"net p/l percentage", "net p/l %", "net p/l percent"}
        ]
        if as_of_cols:
            target_col = as_of_cols[0]
            source_col = net_cols[0] if net_cols else None
            if source_col and source_col != target_col:
                for data_row in range(row + 1, section["end_row"] + 1):
                    source_cell = ws.cell(data_row, source_col)
                    target_cell = ws.cell(data_row, target_col)
                    _copy_cell_style(source_cell, target_cell)
                    target_cell.value = source_cell.value
                    target_cell.comment = copy(source_cell.comment) if source_cell.comment else None
                    source_cell.value = None
                    source_cell.comment = None
                ws.cell(row, source_col).value = None
            else:
                for data_row in range(row + 1, section["end_row"] + 1):
                    ws.cell(data_row, target_col).value = None
                    ws.cell(data_row, target_col).comment = None
            ws.cell(row, target_col).value = "Net P/L Percentage"
            net_cols = [target_col]
            as_of_cols = []
            diagnostics.setdefault("normalized_stats2_duplicate_headers", []).append("removed As Of")
        if len(risk_cols) > 1:
            for duplicate_col in risk_cols[1:]:
                ws.cell(row, duplicate_col).value = None
                for data_row in range(row + 1, section["end_row"] + 1):
                    ws.cell(data_row, duplicate_col).value = None
                    ws.cell(data_row, duplicate_col).comment = None
                diagnostics.setdefault("normalized_stats2_duplicate_headers", []).append("removed duplicate Risk Of Ruin")
        if not net_cols and ws.title == STATS2_SHEET:
            last_known = max(
                [
                    col for col in [
                        headers.get("account"),
                        headers.get("balance"),
                        headers.get("currency"),
                        risk_cols[0] if risk_cols else None,
                    ]
                    if col
                ],
                default=section["start_col"],
            )
            target_col = last_known + 1
            if target_col <= section["end_col"]:
                target_has_content = any(
                    str(ws.cell(data_row, target_col).value or "").strip()
                    for data_row in range(row, section["end_row"] + 1)
                )
                if target_has_content:
                    return
                ws.cell(row, target_col).value = "Net P/L Percentage"
                diagnostics.setdefault("normalized_stats2_duplicate_headers", []).append("added Net P/L Percentage")
        return


def _find_dashboard_table_headers(ws, section: Dict[str, int], *, scan_rows: int = 8) -> tuple[int | None, Dict[str, int]]:
    required = {"account", "balance", "currency"}
    header_row = None
    col_map: Dict[str, int] = {}
    start_row = max(1, section.get("start_row", 1))
    end_row = min(section.get("end_row", ws.max_row), start_row + max(1, scan_rows) - 1)
    for r in range(start_row, end_row + 1):
        row_map: Dict[str, int] = {}
        for c in range(section["start_col"], section["end_col"] + 1):
            h = str(ws.cell(r, c).value or "").strip().lower()
            if h == "account":
                row_map["account"] = c
            elif h == "balance":
                row_map["balance"] = c
            elif h == "currency":
                row_map["currency"] = c
            elif h == "risk of ruin" and "risk_of_ruin" not in row_map:
                row_map["risk_of_ruin"] = c
            elif h in {"as of", "as_of"} and "as_of" not in row_map:
                row_map["as_of"] = c
            elif h in {"net p/l percentage", "net p/l %", "net p/l percent"} and "net_pl_percentage" not in row_map:
                row_map["net_pl_percentage"] = c
        if required.issubset(row_map.keys()):
            header_row = r
            col_map = row_map
            break
    return header_row, col_map


def _repair_stats2_account_balance_formatting(ws, diagnostics: Dict[str, Any] | None = None) -> None:
    try:
        sections = _find_anchor_sections(ws, ["Account Balances"])
    except Exception:
        return
    section = sections.get("Account Balances")
    if not section:
        return
    _normalize_stats2_account_balance_headers(ws, section, diagnostics)
    header_row, col_map = _find_dashboard_table_headers(ws, section)
    if not header_row or "balance" not in col_map:
        return
    balance_col = col_map["balance"]
    template_row = 4 if header_row < 4 <= section["end_row"] else header_row + 1
    template = ws.cell(template_row, balance_col)
    account_col = col_map.get("account")
    currency_col = col_map.get("currency")
    for row in range(header_row + 1, section["end_row"] + 1):
        if account_col and not str(ws.cell(row, account_col).value or "").strip():
            continue
        cell = ws.cell(row, balance_col)
        value = cell.value
        _copy_cell_style(template, cell)
        cell.value = value
        currency = str(ws.cell(row, currency_col).value or "").strip().upper() if currency_col else ""
        if currency:
            cell.number_format = _currency_number_format(currency, force_decimals=8 if _is_crypto_currency(currency) else 2)
        _set_cell_horizontal_alignment(cell, "right")
    _apply_stats2_clean_table_borders(ws, section, header_row, col_map, diagnostics)
    if "net_pl_percentage" in col_map:
        _apply_stats2_net_pl_percentage_conditional_formatting(
            ws, header_row, section, col_map["net_pl_percentage"]
        )
    if diagnostics is not None:
        diagnostics["repaired_stats2_account_balance_formatting"] = True


def _repair_stats2_as_of_datetime_style(wb, diagnostics: Dict[str, Any] | None = None) -> None:
    if STATS2_SHEET not in wb.sheetnames:
        return
    stats2 = wb[STATS2_SHEET]
    try:
        section = _find_anchor_sections(stats2, ["Account Balances"])["Account Balances"]
    except Exception:
        return
    header_row, col_map = _find_dashboard_table_headers(stats2, section)
    as_of_col = col_map.get("as_of") if header_row else None
    if not as_of_col:
        if diagnostics is not None:
            diagnostics["repaired_stats2_as_of_datetime_cells"] = 0
        return
    repaired = 0
    for row in range(header_row + 1, section["end_row"] + 1):
        cell = stats2.cell(row, as_of_col)
        if cell.value in (None, ""):
            continue
        value = cell.value
        if isinstance(value, str):
            value = value.replace("T", " ")
            if len(value) >= 19:
                value = value[:19]
        cell.value = value
        cell.fill = PatternFill()
        font = copy(cell.font)
        font.color = "FF000000"
        cell.font = font
        cell.number_format = "yyyy-mm-dd hh:mm:ss" if isinstance(value, datetime) else "General"
        repaired += 1
    if diagnostics is not None:
        diagnostics["repaired_stats2_as_of_datetime_cells"] = repaired


def _read_stats2_account_balances(wb) -> List[Dict[str, Any]]:
    if STATS2_SHEET not in wb.sheetnames:
        return []
    ws = wb[STATS2_SHEET]
    try:
        sections = _find_anchor_sections(ws, ["Account Balances"])
    except Exception:
        return []
    section = sections.get("Account Balances")
    if not section:
        return []
    header_row, col_map = _find_dashboard_table_headers(ws, section)
    if not header_row or not {"account", "balance", "currency"}.issubset(col_map.keys()):
        return []
    balances: List[Dict[str, Any]] = []
    account_col = col_map["account"]
    for row in range(header_row + 1, section["end_row"] + 1):
        account = str(ws.cell(row, account_col).value or "").strip()
        if not account:
            continue
        balance = _as_float(ws.cell(row, col_map["balance"]).value)
        currency = str(ws.cell(row, col_map["currency"]).value or "").strip().upper()
        balance_metadata = _read_account_balance_source_metadata(wb, account)
        balance_source = balance_metadata.get("source") or "stats2_account_balances"
        payload: Dict[str, Any] = {
            "account": account,
            "account_label": account,
            "label": account,
            "balance": balance,
            "currency": currency,
            "source": balance_source,
            "balance_source": balance_source,
        }
        if "as_of" in col_map:
            payload["as_of"] = _excel_datetime_to_iso(ws.cell(row, col_map["as_of"]).value)
        if balance_metadata.get("timeline_as_of"):
            payload["timeline_as_of"] = balance_metadata["timeline_as_of"]
        balances.append(payload)
    return balances

def _ensure_account_balance_row(ws, section: Dict[str, int], header_row: int, col_map: Dict[str, int], account_label: str) -> int:
    account_col = col_map["account"]
    wanted = _canonical_account_label(account_label)
    for r in range(header_row + 1, section["end_row"] + 1):
        lbl = _canonical_account_label(ws.cell(r, account_col).value)
        if lbl and lbl == wanted:
            return r
    for r in range(header_row + 1, section["end_row"] + 1):
        lbl = str(ws.cell(r, account_col).value or "").strip()
        if lbl:
            continue
        bal_blank = ws.cell(r, col_map["balance"]).value in (None, "")
        cur_blank = ws.cell(r, col_map["currency"]).value in (None, "")
        asof_blank = ("as_of" not in col_map) or (ws.cell(r, col_map["as_of"]).value in (None, ""))
        if bal_blank and cur_blank and asof_blank:
            return r

    if section["end_row"] < ws.max_row:
        raise RuntimeError(f"Account Balances section has no writable row for '{account_label}' without shifting dashboard layout.")

    row = section["end_row"] + 1
    template_row = section["end_row"] if section["end_row"] > header_row else header_row + 1
    for c in range(section["start_col"], section["end_col"] + 1):
        src = ws.cell(template_row, c)
        dst = ws.cell(row, c)
        dst.number_format = src.number_format
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    section["end_row"] = row
    return row


def _clear_account_balance_row(ws, row: int, col_map: Dict[str, int]) -> None:
    ws.cell(row, col_map["account"]).value = None
    ws.cell(row, col_map["balance"]).value = None
    ws.cell(row, col_map["currency"]).value = None
    if "risk_of_ruin" in col_map:
        ws.cell(row, col_map["risk_of_ruin"]).value = None
        ws.cell(row, col_map["risk_of_ruin"]).comment = None
    if "as_of" in col_map:
        ws.cell(row, col_map["as_of"]).value = None
    if "net_pl_percentage" in col_map:
        ws.cell(row, col_map["net_pl_percentage"]).value = None


def _apply_stats2_net_pl_percentage_conditional_formatting(
    ws,
    header_row: int,
    section: Dict[str, int],
    net_col: int,
) -> None:
    start_row = header_row + 1
    end_row = max(start_row, section.get("end_row", ws.max_row))
    col_letter = get_column_letter(net_col)
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    target_bounds = range_boundaries(cell_range)
    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        try:
            existing_bounds = range_boundaries(str(getattr(key, "sqref", key)))
        except ValueError:
            existing_bounds = None
        if (
            existing_bounds == target_bounds
            and len(rules) == 2
            and all(_is_generated_dashboard_semantic_rule(rule) for rule in rules)
        ):
            for row in range(start_row, end_row + 1):
                _clear_generated_semantic_fill(ws.cell(row, net_col))
            return
    sanitized = OrderedDict()
    for key, rules in list(getattr(ws.conditional_formatting, "_cf_rules", {}).items()):
        manual = [rule for rule in rules if not _is_generated_dashboard_semantic_rule(rule)]
        if manual:
            new_key = copy(key)
            sanitized[new_key] = manual
    ws.conditional_formatting._cf_rules = sanitized
    for row in range(start_row, end_row + 1):
        _clear_generated_semantic_fill(ws.cell(row, net_col))
    _profit_loss_rules(ws, cell_range)


def _write_stats2_net_pl_percentages(
    ws,
    section: Dict[str, int],
    header_row: int,
    col_map: Dict[str, int],
    rows: List[Dict[str, Any]],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    net_col = col_map.get("net_pl_percentage")
    account_col = col_map.get("account")
    if not net_col or not account_col:
        return
    totals = _net_result_pct_by_account(rows)
    written = 0
    for row in range(header_row + 1, section["end_row"] + 1):
        account = _canonical_account_label(ws.cell(row, account_col).value)
        if not account:
            continue
        value = totals.get(account)
        cell = ws.cell(row, net_col)
        if value is None or not _stats2_net_pl_percentage_is_applicable(account):
            cell.value = None
            _clear_generated_semantic_fill(cell)
            continue
        cell.value = value / 100.0
        cell.number_format = adaptive_percent_number_format(cell.value)
        written += 1
    _apply_stats2_net_pl_percentage_conditional_formatting(ws, header_row, section, net_col)
    diagnostics["stats2_net_pl_percentage_cells_written"] = written


def _write_stats2_risk_of_ruin(
    ws,
    section: Dict[str, int],
    header_row: int,
    col_map: Dict[str, int],
    rows: List[Dict[str, Any]],
    diagnostics: Dict[str, Any] | None = None,
) -> None:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    risk_col = col_map.get("risk_of_ruin")
    if not risk_col:
        diagnostics["risk_of_ruin_unavailable"] = "Risk of Ruin header missing"
        return
    risk_by_account = _risk_of_ruin_by_account(rows)
    account_col = col_map["account"]
    written = 0
    unavailable: List[Dict[str, Any]] = []
    for row in range(header_row + 1, section["end_row"] + 1):
        account = _canonical_account_label(ws.cell(row, account_col).value)
        if not account:
            continue
        payload = risk_by_account.get(account) or _empty_risk_of_ruin_payload("no_usable_trade_history")
        cell = ws.cell(row, risk_col)
        cell.value = payload.get("risk_of_ruin")
        cell.number_format = "0.00%"
        cell.comment = Comment(_risk_of_ruin_comment_text(payload), "Codex")
        if payload.get("risk_of_ruin") is None:
            unavailable.append({"account": account, "reason": payload.get("reason")})
        else:
            written += 1
    diagnostics["risk_of_ruin_cells_written"] = written
    if unavailable:
        diagnostics["risk_of_ruin_unavailable_accounts"] = unavailable



def _repair_trade_log_row_ids_from_rows(ws, rows, diagnostics):
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    headers = _trade_log_header_map(ws)
    rid_col = headers.get('Row ID')
    if not rid_col:
        return
    start_row = _trade_log_data_start_row(ws)
    repaired=0
    for rr in range(start_row, start_row + len(rows)):
        row_ctx = rows[rr-start_row]
        expected = str(row_ctx.get('id') or stable_row_id(row_ctx)).strip() if isinstance(row_ctx, dict) else ''
        if not expected:
            continue
        cell=ws.cell(rr, rid_col)
        if str(cell.value or '').strip()!=expected:
            cell.value=expected
            repaired+=1
    if repaired:
        diagnostics['repaired_trade_log_row_ids']=repaired

def read_master_journal_source(path: Path) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Master Journal workbook not found: {path}")
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        diagnostics={'repaired_corrupted_row_ids': []}
        balances = _read_stats2_account_balances(wb)
        ws = _get_all_trades_sheet(wb)
        header_map = _trade_log_header_map(ws)
        headers = list(header_map.keys())
        idx = {header: col - 1 for header, col in header_map.items()}
        data_start_row = _trade_log_data_start_row(ws)
        required = {'Open Time','Close Time','Account','Symbol','Side'}
        if not required.issubset(set(idx.keys())):
            raise RuntimeError('Master Journal Trade Log headers are invalid.')
        items=[]; cashflow_ledger=defaultdict(list)
        def _num(v):
            try:
                if v in (None, ""): return None
                return float(v)
            except Exception: return None
        i_stop = _alias_index(idx, 'Stop Loss', 'Stop Loss Price')
        i_tp = _alias_index(idx, 'Take Profit', 'Target Price', 'Target')
        i_stop_dist = _alias_index(idx, 'Stop Loss Distance', 'Stop Loss Distance %')
        i_target_dist = _alias_index(idx, 'Target Distance', 'Target Distance %')
        i_pnl = _alias_index(idx, 'Net P/L', 'Net Profit', 'Realized PnL')
        i_result_pct = _alias_index(idx, 'Result %', 'Profit %', 'P/L %', 'Result Percent')
        i_dur = _alias_index(idx, 'Trade Duration (DD:HH:MM:SS)', 'Trade Duration', 'Trade Duration Seconds', 'Duration')
        max_col = max(1, len(headers))
        max_row = ws.max_row or 1
        try:
            _min_col, _min_row, dim_max_col, dim_max_row = range_boundaries(ws.calculate_dimension())
            max_col = min(max_col, max(1, dim_max_col))
            max_row = max(1, dim_max_row)
        except Exception:
            pass
        for row_cells in ws.iter_rows(min_row=data_start_row, max_row=max_row, max_col=max_col):
            r = [cell.value for cell in row_cells]
            if not any(v not in (None,'') for v in r):
                continue
            symbol = str(r[idx.get('Symbol',3)] or '').strip()
            side = str(r[idx.get('Side',4)] or '').strip()
            account = str(r[idx.get('Account',2)] or '').strip()
            row_id = str(r[idx.get('Row ID',len(r)-1)] or '').strip() if 'Row ID' in idx else ''
            row_type_raw = str(r[idx.get('Row Type')]).strip().lower() if 'Row Type' in idx and idx.get('Row Type') is not None else ''
            row_type = row_type_raw if row_type_raw in {'cashflow','monthly_aud_reval','trade'} else ('cashflow' if symbol.upper()=='CASHFLOW' else ('monthly_aud_reval' if symbol.upper()=='MONTHLY AUD P/L' else 'trade'))
            open_time = _excel_datetime_to_iso(r[idx.get('Open Time',0)])
            close_time = _excel_datetime_to_iso(r[idx.get('Close Time',1)])
            duration = _duration_ddhhmmss_cell_to_seconds(r[i_dur]) if i_dur is not None else None
            if duration is None and i_dur is not None:
                duration = _parse_duration_text(r[i_dur])
            if duration is None and row_type == "trade":
                ot = _as_datetime(open_time)
                ct = _as_datetime(close_time)
                if ot and ct:
                    sec = int((ct - ot).total_seconds())
                    duration = max(1, sec) if sec >= 0 else None
            account_u = account.upper()
            symbol_u = symbol.upper().replace('_','/').replace('-','/')
            if any(t in account_u for t in ('OANDA','PEPPERSTONE','FOREX',' FX')):
                asset_class = 'fx'
            elif any(t in account_u for t in ('BYBIT','BINANCE','COINSPOT')):
                asset_class = 'crypto'
            elif _is_likely_fx_pair(symbol_u):
                asset_class = 'fx'
            elif any(t in symbol_u for t in ('USDT','USDC','BTC','ETH','PERP')):
                asset_class = 'crypto'
            else:
                asset_class = ''
            balance_after = _num(r[idx.get('Balance After')]) if 'Balance After' in idx else None
            cashflow_amount = _num(r[idx.get('Cashflow Amount')]) if 'Cashflow Amount' in idx else (_num(r[i_pnl]) if i_pnl is not None else None)
            cashflow_new_balance = _num(r[idx.get('Cashflow New Balance')]) if 'Cashflow New Balance' in idx else None
            if row_type == 'cashflow' and cashflow_new_balance is None:
                cashflow_new_balance = balance_after
            execution_row = {
                'row_type': row_type,
                'account': account,
                'asset_class': asset_class,
                'symbol': symbol,
                'side': side,
                'open_time': open_time,
                'close_time': close_time,
                'qty': _num(r[idx.get('Qty')]) if 'Qty' in idx else None,
                'entry_price': _num(r[idx.get('Entry Price')]) if 'Entry Price' in idx else None,
                'exit_price': _num(r[idx.get('Exit Price')]) if 'Exit Price' in idx else None,
                'stop_loss': _num(r[i_stop]) if i_stop is not None else None,
                'take_profit': _num(r[i_tp]) if i_tp is not None else None,
            }
            computed_id = _trade_execution_row_id(execution_row)
            monthly_like_id = row_id.startswith('monthly_aud_reval:')
            monthly_semantic = _is_monthly_aud_reval_semantic_row({'row_type': row_type, 'symbol': symbol, 'account': account})
            if row_id and (('PEPPERSTONE' in row_id.upper() or 'OANDA' in row_id.upper()) and ('BYBIT' in account_u or ('USDT' in symbol_u))):
                diagnostics['repaired_corrupted_row_ids'].append({'old_row_id': row_id, 'new_row_id': computed_id, 'reason': 'broker_account_mismatch'})
                row_id = computed_id
            elif row_id and _stale_excel_row_id_reasons(row_id, execution_row):
                diagnostics['repaired_corrupted_row_ids'].append({
                    'old_row_id': row_id,
                    'new_row_id': computed_id,
                    'reason': 'stale_excel_row_id',
                    'mismatched_fields': _stale_excel_row_id_reasons(row_id, execution_row),
                })
                row_id = computed_id
            elif row_id and monthly_like_id and (row_type != 'monthly_aud_reval' or not monthly_semantic or not _monthly_aud_reval_row_id_month(row_id)):
                diagnostics['repaired_corrupted_row_ids'].append({'old_row_id': row_id, 'new_row_id': computed_id, 'reason': 'invalid_monthly_aud_reval_row_id', 'row_type': row_type, 'symbol': symbol, 'account': account})
                row_id = computed_id
            currency = str(r[idx.get('Currency')] or '').strip() if 'Currency' in idx else ''
            item={'id': row_id or computed_id, 'row_type':row_type,'account':account,'symbol':symbol,'side':side,'open_time':open_time,'close_time':close_time,'qty':_num(r[idx.get('Qty')]) if 'Qty' in idx else None,'entry_price':_num(r[idx.get('Entry Price')]) if 'Entry Price' in idx else None,'exit_price':_num(r[idx.get('Exit Price')]) if 'Exit Price' in idx else None,'stop_loss':_num(r[i_stop]) if i_stop is not None else None,'take_profit':_num(r[i_tp]) if i_tp is not None else None,'stop_loss_distance_pct':_normalize_pct_distance_cell(r[i_stop_dist], row_cells[i_stop_dist].number_format) if i_stop_dist is not None and i_stop_dist < len(row_cells) else None,'target_distance_pct':_normalize_pct_distance_cell(r[i_target_dist], row_cells[i_target_dist].number_format) if i_target_dist is not None and i_target_dist < len(row_cells) else None,'commission':_num(r[idx.get('Commission')]) if 'Commission' in idx else None,'net_profit':_num(r[i_pnl]) if i_pnl is not None else None,'result_pct':_excel_fraction_to_pct_points(r[i_result_pct]) if i_result_pct is not None else None,'r_multiple':_num(r[idx.get('R-Multiple')]) if 'R-Multiple' in idx else None,'balance_after_trade':balance_after,'balance_after_trade_source':'master_journal','trade_duration_seconds':duration,'is_test_trade':str(r[idx.get('Test')]).strip().lower() in {'yes','y','true','1'} if 'Test' in idx else False,'setup':r[idx.get('Setup',17)] if 'Setup' in idx else '','timeframe':r[idx.get('Timeframe',18)] if 'Timeframe' in idx else '','breakeven':r[idx.get('Breakeven',19)] if 'Breakeven' in idx else '','notes':r[idx.get('Notes',20)] if 'Notes' in idx else '','cashflow_amount':cashflow_amount,'cashflow_new_balance':cashflow_new_balance,'currency':currency, 'asset_class': asset_class, 'source':'master_journal'}
            if TRADE_NUMBER_HEADER in idx:
                trade_number_value = r[idx[TRADE_NUMBER_HEADER]] if idx[TRADE_NUMBER_HEADER] < len(r) else ""
                item["trade_number"] = "" if trade_number_value is None else str(trade_number_value).strip()
            for header, field in TRADE_LOG_MANUAL_FIELD_MAP.items():
                if header in idx:
                    field_index = idx[header]
                    raw_value = r[field_index] if field_index < len(r) else ''
                    if field in {"move_to_break_even_duration", "move_to_profit_duration"} and raw_value not in (None, ""):
                        number_format = str(row_cells[field_index].number_format or "")
                        parsed_duration = _duration_ddhhmmss_cell_to_seconds(raw_value) if _is_ddhhmmss_number_format(number_format) else _parse_duration_text(raw_value)
                        item[field] = parsed_duration if parsed_duration is not None else raw_value
                    elif field in MOVE_TO_TIME_FIELDS:
                        item[field] = _excel_datetime_to_iso(raw_value)
                    else:
                        item[field] = raw_value
            if "close_stopout" not in item and "Stop Out" in idx:
                item["close_stopout"] = r[idx["Stop Out"]] if idx["Stop Out"] < len(r) else ''
            if row_type == "monthly_aud_reval":
                monthly_currency = (currency or str(item.get("result_currency") or "").strip() or "AUD").upper()
                item["result_cash"] = _num(r[i_pnl]) if i_pnl is not None else None
                item["result_currency"] = monthly_currency
                item["account_label"] = account
                refs: Dict[str, Any] = {}
                month_source = close_time or open_time
                month_dt = _as_datetime(month_source)
                if month_dt is not None:
                    refs["period_month"] = month_dt.strftime("%Y-%m")
                item["raw_refs"] = refs
                item["source"] = str(item.get("source") or "").strip() or "bybit_monthly_aud_reval"
                item.pop("net_profit", None)
                item.pop("realized_pnl", None)
            items.append(item)
            if row_type=='cashflow':
                cashflow_ledger[account].append({'account':account,'date':item['close_time'] or item['open_time'],'amount':cashflow_amount,'new_balance':cashflow_new_balance,'currency':str(r[idx.get('Currency')] or '').strip() if 'Currency' in idx else '','reason':item.get('notes') or '', 'side':side})
        trade_row_id_counts = Counter(
            str(item.get("id") or "").strip()
            for item in items
            if str(item.get("row_type") or "trade").strip().lower() == "trade" and str(item.get("id") or "").strip()
        )
        duplicate_trade_row_ids = sorted(rid for rid, count in trade_row_id_counts.items() if count > 1)
        if duplicate_trade_row_ids:
            diagnostics["target_r_duplicate_trade_row_ids"] = duplicate_trade_row_ids
        items, monthly_dedupe_diagnostics = _dedupe_monthly_aud_reval_rows(items)
        diagnostics.update(monthly_dedupe_diagnostics)
        items, dedupe_diagnostics = _dedupe_trade_rows_by_execution(items)
        diagnostics.update(dedupe_diagnostics)
        return {
            'items': items,
            'balances': balances,
            'cashflow_ledger': dict(cashflow_ledger),
            'diagnostics': diagnostics,
            '_workbook_structure': {
                'sheet_order': list(wb.sheetnames),
                'active_sheet': (
                    wb.sheetnames[
                        min(
                            max(int(getattr(wb, "_active_sheet_index", 0)), 0),
                            len(wb.sheetnames) - 1,
                        )
                    ]
                    if wb.sheetnames
                    else ''
                ),
                'selected_sheets': (
                    [
                        wb.sheetnames[
                            min(
                                max(
                                    int(getattr(wb, "_active_sheet_index", 0)),
                                    0,
                                ),
                                len(wb.sheetnames) - 1,
                            )
                        ]
                    ]
                    if wb.sheetnames
                    else []
                ),
            },
        }
    finally:
        wb.close()


_SPREADSHEETML_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_DRAWINGML_CHART_NAMESPACE = (
    "http://schemas.openxmlformats.org/drawingml/2006/chart"
)
_OFFICE_DOCUMENT_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_PACKAGE_RELATIONSHIP_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_EXTERNAL_WORKBOOK_FILE_REFERENCE = re.compile(
    r"\[[^\]\r\n]+\.(?:xls|xlsx|xlsm|xlsb|xlt|xltx|xltm|xlam)\]",
    re.IGNORECASE,
)
_EXTERNAL_WORKBOOK_INDEX_REFERENCE = re.compile(
    r"\[\d+\][^!\r\n]{0,255}!"
)

_TRADE_LOG_FORMULA_CACHE_SAFE_FUNCTIONS = frozenset({
    "ABS", "AND", "AVERAGE", "CEILING", "CONCATENATE", "COUNT", "COUNTA",
    "COUNTBLANK", "EXACT", "IF", "IFERROR", "INT", "LEFT", "LEN", "LOWER",
    "MAX", "MID", "MIN", "MOD", "NOT", "OR", "PRODUCT", "RIGHT", "ROUND",
    "ROUNDDOWN", "ROUNDUP", "SIGN", "SUM", "TRIM", "TRUNC", "UPPER",
})
_TRADE_LOG_FORMULA_CACHE_A1_RANGE = re.compile(
    r"^\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?$",
    re.IGNORECASE,
)
_TRADE_LOG_FORMULA_CACHE_MAX_DEPENDENCY_CELLS = 4096
_TRADE_LOG_FORMULA_CACHE_MAX_DEPENDENCY_DEPTH = 32


def _worksheet_ooxml_part_name(package: zipfile.ZipFile, sheet_name: str) -> str:
    workbook_root = ET.fromstring(package.read("xl/workbook.xml"))
    relationship_id = ""
    for sheet in workbook_root.findall(
        f".//{{{_SPREADSHEETML_NAMESPACE}}}sheet"
    ):
        if str(sheet.attrib.get("name") or "") == str(sheet_name):
            relationship_id = str(
                sheet.attrib.get(
                    f"{{{_OFFICE_DOCUMENT_RELATIONSHIP_NAMESPACE}}}id"
                )
                or ""
            )
            break
    if not relationship_id:
        raise RuntimeError(f"Worksheet {sheet_name!r} is missing from workbook.xml.")
    relationships_root = ET.fromstring(
        package.read("xl/_rels/workbook.xml.rels")
    )
    target = ""
    for relationship in relationships_root.findall(
        f"{{{_PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship"
    ):
        if str(relationship.attrib.get("Id") or "") == relationship_id:
            target = str(relationship.attrib.get("Target") or "")
            break
    if not target:
        raise RuntimeError(
            f"Worksheet relationship {relationship_id!r} has no package target."
        )
    if target.startswith("/"):
        part_name = target.lstrip("/")
    else:
        part_name = posixpath.normpath(posixpath.join("xl", target))
    if part_name not in package.namelist():
        raise RuntimeError(
            f"Worksheet package part {part_name!r} is missing for {sheet_name!r}."
        )
    return part_name


def _workbook_calculation_signature(path: Path) -> Tuple[Tuple[str, str], ...]:
    with zipfile.ZipFile(path) as package:
        workbook_root = ET.fromstring(package.read("xl/workbook.xml"))
    calc_properties = workbook_root.find(
        f"{{{_SPREADSHEETML_NAMESPACE}}}calcPr"
    )
    if calc_properties is None:
        return ()
    return tuple(sorted(
        (str(key), str(value))
        for key, value in calc_properties.attrib.items()
    ))


def _apply_workbook_calculation_signature(
    wb: Workbook,
    signature: Sequence[Tuple[str, str]],
) -> None:
    """Keep calcPr byte-semantics stable across an openpyxl save.

    openpyxl treats an omitted ``fullCalcOnLoad`` attribute as ``True`` and
    writes the explicit default back on save.  That is semantically equivalent
    to Excel, but it violates this updater's preservation contract.  Rehydrate
    only the attributes that were actually present in the source package.
    """
    values = {str(key): str(value) for key, value in signature}
    calculation = wb.calculation
    boolean_fields = {
        "fullCalcOnLoad",
        "iterate",
        "fullPrecision",
        "calcCompleted",
        "calcOnSave",
        "concurrentCalc",
        "forceFullCalc",
    }
    integer_fields = {"calcId", "iterateCount", "concurrentManualCount"}
    float_fields = {"iterateDelta"}
    for field in calculation.__attrs__:
        raw = values.get(field)
        if raw is None:
            setattr(calculation, field, None)
        elif field in boolean_fields:
            setattr(calculation, field, raw.strip().lower() in {"1", "true"})
        elif field in integer_fields:
            setattr(calculation, field, int(raw))
        elif field in float_fields:
            setattr(calculation, field, float(raw))
        else:
            setattr(calculation, field, raw)


def _chart_relationship_signature_from_package(
    package: zipfile.ZipFile,
) -> Tuple[Tuple[str, str, str, str, str], ...]:
    def canonical_target(relationship_part: str, target: str) -> str:
        raw = str(target or "").replace("\\", "/")
        if raw.startswith("/"):
            return posixpath.normpath(raw.lstrip("/"))
        rel_dir, rel_name = posixpath.split(relationship_part)
        if posixpath.basename(rel_dir) == "_rels" and rel_name.endswith(".rels"):
            owner_dir = posixpath.dirname(rel_dir)
            owner_name = rel_name[:-5]
            owner_part = posixpath.join(owner_dir, owner_name)
            return posixpath.normpath(posixpath.join(posixpath.dirname(owner_part), raw))
        return posixpath.normpath(raw)

    relationships: List[Tuple[str, str, str, str, str]] = []
    for part_name in sorted(
        name for name in package.namelist() if name.endswith(".rels")
    ):
        root = ET.fromstring(package.read(part_name))
        for relationship in root.findall(
            f"{{{_PACKAGE_RELATIONSHIP_NAMESPACE}}}Relationship"
        ):
            relationship_type = str(relationship.attrib.get("Type") or "")
            if not relationship_type.rstrip("/").endswith("/chart"):
                continue
            relationships.append((
                part_name,
                str(relationship.attrib.get("Id") or ""),
                relationship_type,
                canonical_target(
                    part_name,
                    str(relationship.attrib.get("Target") or ""),
                ),
                str(relationship.attrib.get("TargetMode") or ""),
            ))
    return tuple(sorted(relationships))


def _snapshot_chart_ooxml(path: Path) -> Dict[str, Any]:
    with zipfile.ZipFile(path) as package:
        parts = {
            part_name: package.read(part_name)
            for part_name in sorted(package.namelist())
            if part_name.startswith("xl/charts/")
        }
        relationships = _chart_relationship_signature_from_package(package)
    return {
        "parts": parts,
        "relationships": relationships,
    }


def _verify_chart_ooxml_snapshot(
    path: Path,
    snapshot: Mapping[str, Any],
) -> None:
    expected_parts = snapshot.get("parts") or {}
    expected_relationships = tuple(snapshot.get("relationships") or ())
    with zipfile.ZipFile(path) as package:
        actual_part_names = {
            part_name
            for part_name in package.namelist()
            if part_name.startswith("xl/charts/")
        }
        if actual_part_names != set(expected_parts):
            raise RuntimeError(
                "Workbook chart OOXML part set changed during save."
            )
        for part_name, expected_payload in expected_parts.items():
            if package.read(part_name) != expected_payload:
                raise RuntimeError(
                    f"Workbook chart OOXML part changed: {part_name}."
                )
        if (
            _chart_relationship_signature_from_package(package)
            != expected_relationships
        ):
            raise RuntimeError(
                "Workbook chart relationship set changed during save."
            )


def _trade_log_formula_reference_bounds(
    token_value: str,
    sheet_name: str,
) -> Tuple[int, int, int, int]:
    reference = str(token_value or "").strip()
    if "!" in reference:
        sheet_prefix, reference = reference.rsplit("!", 1)
        sheet_prefix = sheet_prefix.strip()
        if (
            len(sheet_prefix) >= 2
            and sheet_prefix.startswith("'")
            and sheet_prefix.endswith("'")
        ):
            sheet_prefix = sheet_prefix[1:-1].replace("''", "'")
        if sheet_prefix.casefold() != str(sheet_name or "").casefold():
            raise ValueError("cross-sheet reference")
    reference = reference.replace("$", "")
    if not _TRADE_LOG_FORMULA_CACHE_A1_RANGE.fullmatch(reference):
        raise ValueError("named, structured, whole-row, or whole-column reference")
    bounds = range_boundaries(reference)
    if bounds[2] > 16384 or bounds[3] > 1048576:
        raise ValueError("reference is outside worksheet bounds")
    return bounds


def _trade_log_formula_dependency_signature(
    ws,
    coordinate: str,
    formula: str,
) -> Dict[str, Any]:
    """Return a conservative same-sheet dependency closure for one formula."""
    dependencies: Dict[str, Tuple[Any, ...]] = {}
    active: Set[str] = set()
    checked_cells = 0

    def fail(reason: str) -> Dict[str, Any]:
        return {
            "safe": False,
            "reason": reason,
            "coordinates": tuple(sorted(dependencies)),
            "signature": tuple(sorted(dependencies.items())),
            "checked_cells": checked_cells,
        }

    def visit_formula(
        formula_coordinate: str,
        formula_text: str,
        depth: int,
    ) -> Optional[str]:
        nonlocal checked_cells
        if depth > _TRADE_LOG_FORMULA_CACHE_MAX_DEPENDENCY_DEPTH:
            return "formula dependency depth exceeds the conservative limit"
        if formula_coordinate in active:
            return "circular formula dependency"
        active.add(formula_coordinate)
        try:
            try:
                tokens = Tokenizer(formula_text).items
            except Exception:
                return "formula could not be tokenized"
            reference_bounds: List[Tuple[int, int, int, int]] = []
            for token in tokens:
                if token.type == "FUNC" and token.subtype == "OPEN":
                    function_name = str(token.value or "").rstrip("(").upper()
                    if function_name not in _TRADE_LOG_FORMULA_CACHE_SAFE_FUNCTIONS:
                        return f"function {function_name or '?'} is not proven deterministic"
                if token.type == "OPERAND" and token.subtype == "RANGE":
                    try:
                        reference_bounds.append(
                            _trade_log_formula_reference_bounds(token.value, ws.title)
                        )
                    except ValueError as exc:
                        return str(exc)
            for min_col, min_row, max_col, max_row in reference_bounds:
                cell_count = (max_col - min_col + 1) * (max_row - min_row + 1)
                if (
                    checked_cells + cell_count
                    > _TRADE_LOG_FORMULA_CACHE_MAX_DEPENDENCY_CELLS
                ):
                    return "formula dependency range exceeds the conservative limit"
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        checked_cells += 1
                        dependency_coordinate = f"{get_column_letter(col)}{row}"
                        dependency_cell = ws._cells.get((row, col))
                        dependency_value = (
                            dependency_cell.value
                            if dependency_cell is not None
                            else None
                        )
                        dependency_is_formula = (
                            dependency_cell is not None
                            and str(dependency_cell.data_type or "") == "f"
                        )
                        if dependency_is_formula and not (
                            isinstance(dependency_value, str)
                            and dependency_value.startswith("=")
                        ):
                            return (
                                "dependency uses a non-plain formula "
                                f"representation at {dependency_coordinate}"
                            )
                        if (
                            isinstance(dependency_value, str)
                            and dependency_value.startswith("=")
                        ):
                            dependencies[dependency_coordinate] = (
                                "formula",
                                dependency_value,
                                dependency_cell.data_type,
                            )
                            nested_error = visit_formula(
                                dependency_coordinate,
                                dependency_value,
                                depth + 1,
                            )
                            if nested_error:
                                return nested_error
                        else:
                            if isinstance(dependency_value, (datetime, date)):
                                comparable_value: Any = dependency_value.isoformat()
                            elif isinstance(dependency_value, Decimal):
                                comparable_value = str(dependency_value)
                            else:
                                comparable_value = dependency_value
                            dependencies[dependency_coordinate] = (
                                "value",
                                (
                                    dependency_cell.data_type
                                    if dependency_cell is not None
                                    else "n"
                                ),
                                comparable_value,
                            )
            return None
        finally:
            active.discard(formula_coordinate)

    error = visit_formula(str(coordinate), str(formula), 0)
    if error:
        return fail(error)
    return {
        "safe": True,
        "reason": "",
        "coordinates": tuple(sorted(dependencies)),
        "signature": tuple(sorted(dependencies.items())),
        "checked_cells": checked_cells,
    }


def _trade_log_authored_formula_entries(ws) -> List[Dict[str, Any]]:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col:
        return []
    entries: List[Dict[str, Any]] = []
    for row in range(_trade_log_data_start_row(ws), _trade_log_last_populated_row(ws) + 1):
        row_id = str(ws.cell(row, row_id_col).value or "").strip()
        if not row_id:
            continue
        for header, col in headers.items():
            value = ws.cell(row, col).value
            if not (isinstance(value, str) and value.startswith("=")):
                continue
            entries.append({
                "row_id": row_id,
                "header": header,
                "row": row,
                "column": col,
                "coordinate": ws.cell(row, col).coordinate,
                "formula": value,
                "dependency": _trade_log_formula_dependency_signature(
                    ws, ws.cell(row, col).coordinate, value
                ),
            })
    return entries


def _incremental_assert_trade_log_formula_dependencies_safe(
    ws,
    affected_rows: Collection[int],
) -> Dict[str, int]:
    affected = {int(row) for row in affected_rows}
    entries = _trade_log_authored_formula_entries(ws)
    checked_dependency_cells = 0
    for entry in entries:
        if int(entry["row"]) in affected:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Trade Log row {entry['row']} contains a formula in managed "
                f"column {entry['header']!r}.",
                unsafe_full_rebuild=True,
            )
        dependency = entry["dependency"]
        checked_dependency_cells += int(dependency.get("checked_cells") or 0)
        if not dependency.get("safe"):
            raise IncrementalWorkbookUpdateNotEligible(
                "Incremental workbook update cannot prove cached formula safety for "
                f"Trade Log {entry['coordinate']}: {dependency.get('reason') or 'unknown dependency'}.",
                unsafe_full_rebuild=True,
            )
        intersecting = sorted(
            coordinate
            for coordinate in dependency.get("coordinates") or ()
            if range_boundaries(coordinate)[1] in affected
        )
        if intersecting:
            raise IncrementalWorkbookUpdateNotEligible(
                "Incremental workbook update would invalidate the cached result for "
                f"Trade Log {entry['coordinate']}; changed dependencies: "
                + ", ".join(intersecting[:20]),
                unsafe_full_rebuild=True,
            )
    return {
        "trade_log_authored_formula_cells_checked": len(entries),
        "trade_log_formula_dependency_cells_checked": checked_dependency_cells,
    }


def _formula_ooxml_signature(formula_element: ET.Element) -> Tuple[Any, ...]:
    return (
        str(formula_element.text or ""),
        tuple(sorted(
            (str(key), str(value))
            for key, value in formula_element.attrib.items()
        )),
    )


def _trade_log_ooxml_formula_records(
    path: Path,
    sheet_name: str,
) -> Tuple[Dict[str, Any], ...]:
    """Inspect every formula stored in the managed Trade Log data rectangle."""
    with zipfile.ZipFile(path) as package:
        part_name = _worksheet_ooxml_part_name(package, sheet_name)
        root = ET.fromstring(package.read(part_name))
    records: List[Dict[str, Any]] = []
    for xml_cell in root.findall(f".//{{{_SPREADSHEETML_NAMESPACE}}}c"):
        formula_element = xml_cell.find(
            f"{{{_SPREADSHEETML_NAMESPACE}}}f"
        )
        if formula_element is None:
            continue
        coordinate = str(xml_cell.attrib.get("r") or "")
        try:
            min_col, min_row, max_col, max_row = range_boundaries(coordinate)
        except (TypeError, ValueError):
            continue
        if (
            min_row != max_row
            or min_col != max_col
            or min_row < TRADE_LOG_DATA_START_ROW
            or min_col < 1
            or min_col > len(TRADE_LOG_HEADERS)
        ):
            continue
        attributes = {
            str(key): str(value)
            for key, value in formula_element.attrib.items()
        }
        formula_type = str(attributes.get("t") or "normal").casefold()
        non_plain_attributes = {
            key: value
            for key, value in attributes.items()
            if key != "t"
        }
        records.append({
            "coordinate": coordinate,
            "formula_ooxml_signature": _formula_ooxml_signature(
                formula_element
            ),
            "formula_type": formula_type,
            "non_plain": (
                formula_type not in {"", "normal"}
                or bool(non_plain_attributes)
            ),
        })
    return tuple(records)


def _trade_log_formula_representation_issue(
    ws,
    records: Sequence[Mapping[str, Any]],
) -> str:
    non_plain = [record for record in records if record.get("non_plain")]
    if non_plain:
        first = non_plain[0]
        return (
            f"Trade Log {first.get('coordinate') or '?'} uses non-plain "
            f"formula representation {first.get('formula_type') or 'unknown'!r}"
        )
    mapped_coordinates = {
        str(entry["coordinate"])
        for entry in _trade_log_authored_formula_entries(ws)
    }
    unmapped = [
        str(record.get("coordinate") or "?")
        for record in records
        if str(record.get("coordinate") or "") not in mapped_coordinates
    ]
    if unmapped:
        return (
            "Trade Log formula cells have no unambiguous stable Row ID/header "
            "mapping: " + ", ".join(unmapped[:20])
        )
    return ""


def _cell_contains_formula(cell) -> bool:
    value = cell.value
    return (
        str(getattr(cell, "data_type", "") or "") == "f"
        or (isinstance(value, str) and value.startswith("="))
    )


def _managed_stats2_account_balance_formula_cells(wb) -> Tuple[str, ...]:
    if STATS2_SHEET not in wb.sheetnames:
        return ()
    ws = wb[STATS2_SHEET]
    try:
        section = _find_anchor_sections(ws, ["Account Balances"])[
            "Account Balances"
        ]
    except Exception:
        return ()
    header_row, col_map = _find_dashboard_table_headers(ws, section)
    if not header_row or not {"account", "balance", "currency"}.issubset(
        col_map
    ):
        return ()
    managed_columns = [
        int(col_map[name])
        for name in ("account", "currency", "balance", "as_of")
        if name in col_map
    ]
    return tuple(
        ws.cell(row, col).coordinate
        for row in range(header_row + 1, section["end_row"] + 1)
        for col in managed_columns
        if _cell_contains_formula(ws.cell(row, col))
    )


def _snapshot_trade_log_formula_caches(
    path: Path,
    ws,
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    entries = {
        str(entry["coordinate"]): entry
        for entry in _trade_log_authored_formula_entries(ws)
    }
    if not entries:
        return {}
    with zipfile.ZipFile(path) as package:
        part_name = _worksheet_ooxml_part_name(package, ws.title)
        root = ET.fromstring(package.read(part_name))
    xml_cells = {
        str(cell.attrib.get("r") or ""): cell
        for cell in root.findall(f".//{{{_SPREADSHEETML_NAMESPACE}}}c")
    }
    records: Dict[Tuple[str, str], Dict[str, Any]] = {}
    ambiguous: Set[Tuple[str, str]] = set()
    for coordinate, entry in entries.items():
        xml_cell = xml_cells.get(coordinate)
        if xml_cell is None:
            continue
        formula_element = xml_cell.find(
            f"{{{_SPREADSHEETML_NAMESPACE}}}f"
        )
        cached_value = xml_cell.find(
            f"{{{_SPREADSHEETML_NAMESPACE}}}v"
        )
        if (
            formula_element is None
            or cached_value is None
            or cached_value.text is None
        ):
            continue
        key = (str(entry["row_id"]), str(entry["header"]))
        if key in records:
            ambiguous.add(key)
            continue
        records[key] = {
            **entry,
            "formula_ooxml_signature": _formula_ooxml_signature(formula_element),
            "cached_value": str(cached_value.text),
            "cached_type": xml_cell.attrib.get("t"),
        }
    for key in ambiguous:
        records.pop(key, None)
    return records


def _plan_trade_log_formula_cache_transplant(
    ws,
    source_records: Mapping[Tuple[str, str], Mapping[str, Any]],
    *,
    strict: bool,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    rows_by_id: Dict[str, List[int]] = defaultdict(list)
    if row_id_col:
        for row in range(
            _trade_log_data_start_row(ws),
            _trade_log_last_populated_row(ws) + 1,
        ):
            row_id = str(ws.cell(row, row_id_col).value or "").strip()
            if row_id:
                rows_by_id[row_id].append(row)
    plan: List[Dict[str, Any]] = []
    skipped: List[Dict[str, str]] = []
    for (row_id, header), source_record in source_records.items():
        target_rows = rows_by_id.get(str(row_id)) or []
        target_col = headers.get(str(header))
        reason = ""
        if len(target_rows) != 1 or not target_col:
            reason = "stable row/header target is missing or ambiguous"
        else:
            target_cell = ws.cell(target_rows[0], target_col)
            if target_cell.value != source_record.get("formula"):
                reason = "formula text changed"
            else:
                target_dependency = _trade_log_formula_dependency_signature(
                    ws, target_cell.coordinate, str(target_cell.value)
                )
                source_dependency = source_record.get("dependency") or {}
                if not source_dependency.get("safe"):
                    reason = str(
                        source_dependency.get("reason")
                        or "source dependencies are not proven safe"
                    )
                elif not target_dependency.get("safe"):
                    reason = str(
                        target_dependency.get("reason")
                        or "target dependencies are not proven safe"
                    )
                elif (
                    source_dependency.get("signature")
                    != target_dependency.get("signature")
                ):
                    reason = "formula dependency values changed"
                else:
                    plan.append({
                        **dict(source_record),
                        "target_coordinate": target_cell.coordinate,
                    })
        if reason:
            skipped.append({
                "row_id": str(row_id),
                "header": str(header),
                "reason": reason,
            })
    if strict and skipped:
        first = skipped[0]
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update cannot preserve a cached authored formula "
            f"for {first['row_id']!r}/{first['header']!r}: {first['reason']}.",
            unsafe_full_rebuild=True,
        )
    return plan, {
        "trade_log_formula_cache_source_cells": len(source_records),
        "trade_log_formula_cache_planned_cells": len(plan),
        "trade_log_formula_cache_skipped_cells": len(skipped),
        "trade_log_formula_cache_skipped": skipped,
    }


def _verify_trade_log_formula_cache_transplant(
    path: Path,
    sheet_name: str,
    plan: Sequence[Mapping[str, Any]],
) -> None:
    with zipfile.ZipFile(path) as package:
        part_name = _worksheet_ooxml_part_name(package, sheet_name)
        root = ET.fromstring(package.read(part_name))
    xml_cells = {
        str(cell.attrib.get("r") or ""): cell
        for cell in root.findall(f".//{{{_SPREADSHEETML_NAMESPACE}}}c")
    }
    for record in plan:
        coordinate = str(record.get("target_coordinate") or "")
        xml_cell = xml_cells.get(coordinate)
        if xml_cell is None:
            raise RuntimeError(
                f"Formula-cache verification lost Trade Log cell {coordinate}."
            )
        formula_element = xml_cell.find(
            f"{{{_SPREADSHEETML_NAMESPACE}}}f"
        )
        cached_value = xml_cell.find(
            f"{{{_SPREADSHEETML_NAMESPACE}}}v"
        )
        if (
            formula_element is None
            or _formula_ooxml_signature(formula_element)
            != record.get("formula_ooxml_signature")
            or cached_value is None
            or str(cached_value.text) != str(record.get("cached_value"))
            or xml_cell.attrib.get("t") != record.get("cached_type")
        ):
            raise RuntimeError(
                f"Formula-cache verification failed for Trade Log {coordinate}."
            )


def _patch_formula_caches_in_worksheet_payload(
    payload: bytes,
    plan: Sequence[Mapping[str, Any]],
) -> bytes:
    text = payload.decode("utf-8")
    for record in plan:
        coordinate = str(record.get("target_coordinate") or "")
        coordinate_pattern = re.escape(coordinate)
        cell_pattern = re.compile(
            r'(<(?:[A-Za-z_][\w.-]*:)?c\b(?=[^>]*\br="'
            + coordinate_pattern
            + r'")[^>]*>)(.*?)(</(?:[A-Za-z_][\w.-]*:)?c>)',
            re.DOTALL,
        )
        matches = list(cell_pattern.finditer(text))
        if len(matches) != 1:
            raise RuntimeError(
                f"Formula-cache patch could not uniquely locate Trade Log {coordinate}."
            )
        match = matches[0]
        start_tag, body, end_tag = match.groups()
        start_tag = re.sub(r'\s+t="[^"]*"', "", start_tag)
        cached_type = record.get("cached_type")
        if cached_type is not None:
            escaped_type = _xml_escape(str(cached_type), {'"': "&quot;"})
            start_tag = start_tag[:-1] + f' t="{escaped_type}">'
        cached_text = _xml_escape(str(record.get("cached_value") or ""))
        value_element = f"<v>{cached_text}</v>"
        value_pattern = re.compile(
            r"<(?:[A-Za-z_][\w.-]*:)?v(?:\s[^>]*)?>.*?"
            r"</(?:[A-Za-z_][\w.-]*:)?v\s*>",
            re.DOTALL,
        )
        if value_pattern.search(body):
            body = value_pattern.sub(value_element, body, count=1)
        else:
            self_closing_value = re.compile(
                r"<(?:[A-Za-z_][\w.-]*:)?v(?:\s[^>]*)?/\s*>"
            )
            if self_closing_value.search(body):
                body = self_closing_value.sub(value_element, body, count=1)
            else:
                formula_close = re.search(
                    r"</(?:[A-Za-z_][\w.-]*:)?f\s*>", body
                )
                if formula_close is None:
                    raise RuntimeError(
                        f"Formula-cache patch found no formula at Trade Log {coordinate}."
                    )
                body = (
                    body[:formula_close.end()]
                    + value_element
                    + body[formula_close.end():]
                )
        replacement = start_tag + body + end_tag
        text = text[:match.start()] + replacement + text[match.end():]
    return text.encode("utf-8")


def _preserve_candidate_ooxml(
    candidate_path: Path,
    sheet_name: str,
    formula_cache_plan: Sequence[Mapping[str, Any]],
    chart_snapshot: Mapping[str, Any],
) -> Dict[str, Any]:
    expected_chart_parts = dict(chart_snapshot.get("parts") or {})
    expected_chart_relationships = tuple(
        chart_snapshot.get("relationships") or ()
    )
    replacements: Dict[str, bytes] = dict(expected_chart_parts)
    with zipfile.ZipFile(candidate_path) as package:
        candidate_chart_parts = {
            name for name in package.namelist() if name.startswith("xl/charts/")
        }
        if candidate_chart_parts != set(expected_chart_parts):
            raise RuntimeError(
                "Workbook chart OOXML part set changed during candidate save."
            )
        if (
            _chart_relationship_signature_from_package(package)
            != expected_chart_relationships
        ):
            raise RuntimeError(
                "Workbook chart relationship set changed during candidate save."
            )
        if formula_cache_plan:
            worksheet_part = _worksheet_ooxml_part_name(package, sheet_name)
            worksheet_root = ET.fromstring(package.read(worksheet_part))
            xml_cells = {
                str(cell.attrib.get("r") or ""): cell
                for cell in worksheet_root.findall(
                    f".//{{{_SPREADSHEETML_NAMESPACE}}}c"
                )
            }
            for record in formula_cache_plan:
                coordinate = str(record.get("target_coordinate") or "")
                xml_cell = xml_cells.get(coordinate)
                formula_element = (
                    xml_cell.find(f"{{{_SPREADSHEETML_NAMESPACE}}}f")
                    if xml_cell is not None
                    else None
                )
                if (
                    formula_element is None
                    or _formula_ooxml_signature(formula_element)
                    != record.get("formula_ooxml_signature")
                ):
                    raise RuntimeError(
                        "Formula-cache transplant refused changed formula at "
                        f"Trade Log {coordinate}."
                    )
            replacements[worksheet_part] = (
                _patch_formula_caches_in_worksheet_payload(
                    package.read(worksheet_part), formula_cache_plan
                )
            )
        changed_replacements = {
            part_name: payload
            for part_name, payload in replacements.items()
            if package.read(part_name) != payload
        }
    if changed_replacements:
        rewrite_path = candidate_path.with_name(
            f".{candidate_path.name}.ooxml-preserve-{uuid4().hex}.tmp"
        )
        replaced = False
        try:
            with zipfile.ZipFile(candidate_path, "r") as source_package:
                with zipfile.ZipFile(rewrite_path, "w") as target_package:
                    target_package.comment = source_package.comment
                    for info in source_package.infolist():
                        if info.filename in changed_replacements:
                            payload = changed_replacements[info.filename]
                        else:
                            payload = source_package.read(info.filename)
                        target_package.writestr(info, payload)
            with rewrite_path.open("r+b") as handle:
                handle.flush()
                os.fsync(handle.fileno())
            if formula_cache_plan:
                _verify_trade_log_formula_cache_transplant(
                    rewrite_path, sheet_name, formula_cache_plan
                )
            _verify_chart_ooxml_snapshot(rewrite_path, chart_snapshot)
            os.replace(rewrite_path, candidate_path)
            replaced = True
        finally:
            if not replaced and rewrite_path.exists():
                rewrite_path.unlink(missing_ok=True)
    else:
        if formula_cache_plan:
            _verify_trade_log_formula_cache_transplant(
                candidate_path, sheet_name, formula_cache_plan
            )
        _verify_chart_ooxml_snapshot(candidate_path, chart_snapshot)
    changed_chart_parts = sum(
        1
        for part_name in expected_chart_parts
        if part_name in changed_replacements
    )
    return {
        "trade_log_formula_caches_transplanted": len(formula_cache_plan),
        "trade_log_formula_cache_candidate_verified": True,
        "chart_ooxml_parts_verified": len(expected_chart_parts),
        "chart_ooxml_parts_restored": changed_chart_parts,
        "chart_relationships_verified": len(expected_chart_relationships),
        "candidate_ooxml_rewritten_once": bool(changed_replacements),
    }


def _external_workbook_formula_references(path: Path) -> List[str]:
    """Return formula/name/chart references that depend on another workbook.

    Master Journal workbooks are intentionally self-contained.  This guard is
    run before loading with ``keep_links=False`` so a genuine dependency can
    never be removed silently.
    """

    references: List[str] = []
    with zipfile.ZipFile(path) as package:
        members = set(package.namelist())
        workbook_xml = ET.fromstring(package.read("xl/workbook.xml"))
        main_namespace = {"x": _SPREADSHEETML_NAMESPACE}
        has_external_reference_table = bool(
            workbook_xml.findall(
                "x:externalReferences/x:externalReference",
                main_namespace,
            )
        )

        def is_external_formula(value: Any) -> bool:
            text = str(value or "").strip()
            if not text:
                return False
            if _EXTERNAL_WORKBOOK_FILE_REFERENCE.search(text):
                return True
            return bool(
                has_external_reference_table
                and _EXTERNAL_WORKBOOK_INDEX_REFERENCE.search(text)
            )

        for part_name in sorted(
            name
            for name in members
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        ):
            root = ET.fromstring(package.read(part_name))
            for cell in root.findall(
                f".//{{{_SPREADSHEETML_NAMESPACE}}}c"
            ):
                formula = cell.find(
                    f"{{{_SPREADSHEETML_NAMESPACE}}}f"
                )
                if formula is not None and is_external_formula(formula.text):
                    references.append(
                        f"{part_name}!{cell.attrib.get('r', '?')}={formula.text}"
                    )

        for defined_name in workbook_xml.findall(
            "x:definedNames/x:definedName",
            main_namespace,
        ):
            if is_external_formula(defined_name.text):
                references.append(
                    "definedName:"
                    f"{defined_name.attrib.get('name', '?')}={defined_name.text}"
                )

        for part_name in sorted(
            name
            for name in members
            if name.startswith("xl/charts/") and name.endswith(".xml")
        ):
            root = ET.fromstring(package.read(part_name))
            for index, formula in enumerate(
                root.findall(
                    f".//{{{_DRAWINGML_CHART_NAMESPACE}}}f"
                ),
                start=1,
            ):
                if is_external_formula(formula.text):
                    references.append(
                        f"{part_name}!series[{index}]={formula.text}"
                    )
    return references


def update_master_journal_workbook_data_only(
    path: Path,
    snapshot: Dict[str, Any],
    expected_survivor_row_ids: Optional[List[str]] = None,
    *,
    preserve_existing_layout: bool = False,
    publish_recommendation_assets: bool = True,
) -> Dict[str, Any]:
    existing_symbol_win_rate_fingerprint: str | None = None
    symbol_win_rate_fingerprint_error = ""
    if preserve_existing_layout:
        try:
            existing_symbol_win_rate_fingerprint = (
                _symbol_win_rate_snapshot_fingerprint(
                    read_master_journal_source(path).get("items") or []
                )
            )
        except Exception as exc:
            # A failed comparison must never freeze a derived metric.  The
            # normal regeneration path remains safe and the reason is exposed
            # in diagnostics.
            symbol_win_rate_fingerprint_error = str(exc)
    source_calculation_signature = (
        _workbook_calculation_signature(path)
        if preserve_existing_layout
        else ()
    )
    source_chart_snapshot = (
        _snapshot_chart_ooxml(path)
        if preserve_existing_layout
        else {"parts": {}, "relationships": ()}
    )
    external_formula_references = _external_workbook_formula_references(path)
    if external_formula_references:
        raise RuntimeError(
            "Master Journal contains genuine external workbook formula/name/chart "
            "references; refusing to discard external link metadata: "
            + "; ".join(external_formula_references[:10])
        )
    wb = load_workbook(path, keep_links=False)
    diagnostics: Dict[str, Any] = {"missing_accounts": [], "updated_cells": 0}
    if symbol_win_rate_fingerprint_error:
        diagnostics["symbol_win_rate_fingerprint_error"] = (
            symbol_win_rate_fingerprint_error
        )
    source_preservation_contract = (
        _workbook_preservation_contract(wb)
        if preserve_existing_layout
        else {}
    )
    if source_preservation_contract:
        diagnostics["preservation_contract_source_sha256"] = (
            _contract_sha256(source_preservation_contract)
        )
    preserved_layout = {
        ws.title: {
            "column_widths": {
                key: dimension.width
                for key, dimension in ws.column_dimensions.items()
            },
            "row_heights": {
                key: dimension.height
                for key, dimension in ws.row_dimensions.items()
            },
            "freeze_panes": ws.freeze_panes,
            "views": deepcopy(ws.views),
            "auto_filter": deepcopy(ws.auto_filter),
            "merged_cells": [str(item) for item in ws.merged_cells.ranges],
            "conditional_formatting": deepcopy(
                ws.conditional_formatting
            ),
            "data_validations": deepcopy(ws.data_validations),
            "sheet_format": deepcopy(ws.sheet_format),
            "cell_presentations": {
                coordinate: {
                    "value": cell.value,
                    "style": copy(cell._style),
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                    "protection": copy(cell.protection),
                    "hyperlink": copy(getattr(cell, "hyperlink", None)),
                    "comment": copy(getattr(cell, "comment", None)),
                }
                for coordinate, cell in ws._cells.items()
            },
        }
        for ws in wb.worksheets
    } if preserve_existing_layout else {}
    preserved_trade_log_presentations: Dict[str, List[Dict[str, Any]]] = {}
    preserved_trade_log_formula_caches: Dict[
        Tuple[str, str], Dict[str, Any]
    ] = {}
    if preserve_existing_layout:
        try:
            source_trade_log = _get_all_trades_sheet(wb)
            source_trade_log_formula_records = (
                _trade_log_ooxml_formula_records(
                    path, source_trade_log.title
                )
            )
            representation_issue = _trade_log_formula_representation_issue(
                source_trade_log,
                source_trade_log_formula_records,
            )
            if representation_issue:
                raise RuntimeError(
                    "Preservation-mode workbook update cannot safely preserve "
                    f"{representation_issue}."
                )
            stats2_formula_cells = (
                _managed_stats2_account_balance_formula_cells(wb)
            )
            if stats2_formula_cells:
                raise RuntimeError(
                    "Preservation-mode workbook update cannot safely preserve "
                    "a formula in a managed cell of STATS2 Account Balances: "
                    + ", ".join(stats2_formula_cells[:20])
                    + "."
                )
            diagnostics["trade_log_ooxml_formula_cells_checked"] = len(
                source_trade_log_formula_records
            )
            source_layout = preserved_layout.get(source_trade_log.title) or {}
            preserved_trade_log_presentations = (
                _snapshot_trade_log_data_presentations_by_row_id(
                    source_trade_log,
                    source_layout.get("cell_presentations") or {},
                )
            )
            preserved_trade_log_formula_caches = (
                _snapshot_trade_log_formula_caches(path, source_trade_log)
            )
        except Exception:
            wb.close()
            raise
    try:
        _repair_legacy_instrument_averages_freeze_pane(wb, diagnostics)
        _migrate_analysis_sheet_names(wb, diagnostics)
        content_before = _workbook_content_snapshot(wb)
        _migrate_legacy_trade_log_sheet_name(wb, diagnostics)
        _remove_legacy_trade_meta_sheet(wb, diagnostics)
        instrument_ws = _symbols_sheet(wb)
        _ensure_instrument_averages_schema(instrument_ws, diagnostics)
        _ensure_symbols_freeze_panes(instrument_ws, diagnostics)
        if not preserve_existing_layout:
            _apply_instrument_averages_requested_style(
                instrument_ws, preserve_layout=True
            )
        def _repair_trade_log_unknown_currency_formats(ws, rows: List[Dict[str, Any]], diagnostics: Dict[str, Any] | None = None) -> None:
            diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
            repaired = 0
            headers = _trade_log_header_map(ws)
            repair_cols = (
                (headers.get("Commission"), "commission"),
                (headers.get("Net P/L"), "net_pnl"),
            )
            start_row = _trade_log_data_start_row(ws)
            for rr in range(start_row, ws.max_row + 1):
                row_ctx = rows[rr - start_row] if rr - start_row < len(rows) else {}
                for col, field in repair_cols:
                    if not col:
                        continue
                    cell = ws.cell(rr, col)
                    fmt = str(cell.number_format or "")
                    if "UNKNOWN" not in fmt:
                        continue
                    ccy = _infer_trade_log_currency(row_ctx, field=field)
                    if not ccy:
                        continue
                    cell.number_format = _currency_number_format(ccy)
                    repaired += 1
            if repaired:
                diagnostics["repaired_trade_log_unknown_currency_formats"] = True
                diagnostics["repaired_trade_log_unknown_currency_format_cells"] = repaired
        trade_log_ws = _get_trade_log_sheet(wb, allow_legacy=False)
        _ensure_trade_log_schema(trade_log_ws, diagnostics)
        dash = _stats1_sheet(wb)
        detail_dash = _stats2_sheet(wb) or dash
        if not preserve_existing_layout:
            _repair_dashboard_core_layout(dash, diagnostics)
            _ensure_dashboard_expectancy_row(dash, diagnostics)
            _ensure_dashboard_move_duration_rows(dash, diagnostics)
            _ensure_dashboard_requested_metric_rows(dash, diagnostics)
            _ensure_dashboard_extended_layout(dash, diagnostics)
        _remove_stats1_generated_source_rows(dash, diagnostics)

        stats = snapshot.get("stats") or {}
        existing_manual_overrides = read_master_journal_manual_overrides(path) if path.exists() else {}
        rows, row_normalization_diagnostics = _normalize_master_journal_rows(
            snapshot.get("items") or [],
            existing_manual_overrides,
        )
        diagnostics.update(row_normalization_diagnostics)
        snapshot = dict(snapshot)
        snapshot["items"] = rows
        incoming_symbol_win_rate_fingerprint = (
            _symbol_win_rate_snapshot_fingerprint(rows)
        )
        preserve_symbol_win_rates = bool(
            preserve_existing_layout
            and existing_symbol_win_rate_fingerprint
            and existing_symbol_win_rate_fingerprint
            == incoming_symbol_win_rate_fingerprint
        )
        diagnostics["symbol_win_rate_input_fingerprint"] = (
            incoming_symbol_win_rate_fingerprint
        )
        diagnostics["symbol_win_rate_inputs_unchanged"] = (
            preserve_symbol_win_rates
        )
        if STATS2_SHEET not in wb.sheetnames:
            wb.create_sheet(STATS2_SHEET, 1)
            diagnostics["created_stats2_from_legacy_layout"] = True
        _ensure_report_sheets(wb, snapshot, diagnostics)
        _remove_target_r_metadata_sheet(wb, diagnostics)
        before = _snapshot_invariants(wb)
        groups = stats.get("groups") or {}
        by_market = groups.get("by_market") or {}
        risk = groups.get("risk_expectancy") or {}
        metric_rows = [
            row for row in rows
            if isinstance(row, dict) and str(row.get("row_type") or "trade").strip().lower() == "trade"
        ]
        leaders = {
            **_instrument_leader_payloads_from_rows(metric_rows),
            **dict(groups.get("leaders") or {}),
        }
        totals = stats.get("totals") or {}
        move_duration_metrics = _trade_move_duration_metrics(rows)
        extended_metrics = _dashboard_extended_metrics(rows, by_market)

        anchors = _find_anchor_sections(
            dash,
            ["Overall", "Winners", "Losers", "Drawdown"],
            optional=["Duration", "Side", "Patterns", "Timeframe", "Commission"],
        )
        detail_anchors = _find_anchor_sections(
            detail_dash,
            ["Account Balances"],
            optional=["Instrument leaders", "FX", "Crypto"],
        )
        section_sheets = {
            **{name: dash for name in anchors},
            **{name: detail_dash for name in detail_anchors},
        }
        section_anchors = {**anchors, **detail_anchors}

        def _format_metric_value(value: Any, metric_type: str = "raw"):
            if value is None:
                return None
            if metric_type == "pct":
                pct = _as_float(value)
                if pct is None:
                    return None
                return _pct_points_to_excel_fraction(pct)
            if metric_type == "r":
                r_value = _as_float(value)
                return r_value if r_value is not None else value
            if metric_type == "duration":
                return _format_duration_display(value)
            if metric_type == "count":
                f = _as_float(value)
                return int(f) if f is not None else value
            if metric_type == "source":
                return _fmt_detail_src(value)
            return value

        def _set_dashboard_metric_number_format(cell, metric_type: str) -> None:
            if metric_type == "pct":
                cell.number_format = adaptive_percent_number_format(cell.value)
            elif metric_type == "r":
                cell.number_format = '0.000"R"'
            elif metric_type == "count":
                cell.number_format = "0"
            elif metric_type == "duration":
                cell.number_format = "General"

        def _apply_dashboard_metric_semantic_style(cell, semantic: str | None) -> None:
            value = _as_float(cell.value)
            if value is None or value == 0 or not semantic:
                return
            if semantic == "profit_loss":
                _apply_full_cell_semantic_fill(cell, "profit" if value > 0 else "loss")
            elif semantic in {"loss", "drawdown"}:
                _apply_full_cell_semantic_fill(cell, "loss")
            elif semantic == "profit":
                _apply_full_cell_semantic_fill(cell, "profit")

        def _write_dashboard_metric_cell(
            target_ws,
            row: int,
            col: int,
            value: Any,
            metric_type: str = "raw",
            semantic: str | None = None,
            *,
            allow_grey: bool = False,
        ) -> bool:
            existing_cell = target_ws.cell(row, col)
            existing_text = str(existing_cell.value or "").strip().casefold()
            grey_locked = _is_light_grey_no_metric_cell(existing_cell) and not allow_grey
            if grey_locked and existing_text.startswith("unavailable:"):
                grey_locked = False
            if value is None or grey_locked:
                return False
            if _write_value_preserving_cell(target_ws, row, col, value):
                diagnostics["updated_cells"] += 1
                cell = target_ws.cell(row, col)
                _set_dashboard_metric_number_format(cell, metric_type)
                _apply_dashboard_metric_semantic_style(cell, semantic)
                return True
            return False

        def write_metric(
            section: str,
            label: str,
            value: Any,
            metric_type: str = "raw",
            source: Any = None,
        ):
            out = (
                _format_inline_metric_value(value, metric_type, source)
                if source and metric_type in {"pct", "r", "duration"}
                else _format_metric_value(value, metric_type)
            )
            if out is None:
                return
            target_ws = section_sheets[section]
            pos = _find_label_in_section(target_ws, label, section_anchors[section])
            if not pos:
                return
            _write_dashboard_metric_cell(
                target_ws,
                pos[0],
                pos[1] + 1,
                out,
                "raw" if source else metric_type,
            )

        def _main_dashboard_market_columns() -> Dict[str, int]:
            cols: Dict[str, int] = {}
            for r in range(1, min(5, dash.max_row) + 1):
                for c in range(1, min(8, dash.max_column) + 1):
                    token = str(dash.cell(r, c).value or "").strip().lower()
                    if token == "overall":
                        cols["overall"] = c
                    elif token in {"fx", "forex"}:
                        cols["fx"] = c
                    elif token == "crypto":
                        cols["crypto"] = c
                if {"overall", "fx", "crypto"}.issubset(cols):
                    break
            return cols

        manual_drawdown_metric_cells = set()

        def write_market_metric(
            section: str,
            label: str | List[str],
            values_by_market: Dict[str, Any],
            metric_type: str = "raw",
            semantic: str | None = None,
            sources_by_market: Dict[str, Any] | None = None,
        ):
            labels = label if isinstance(label, list) else [label]
            pos = None
            for candidate_label in labels:
                pos = _find_label_in_section(dash, candidate_label, anchors[section])
                if pos:
                    break
            if not pos:
                return
            market_cols = _main_dashboard_market_columns()
            missing_markets = []
            for market, col in market_cols.items():
                raw_value = values_by_market.get(market)
                source = (sources_by_market or {}).get(market)
                out = (
                    _format_inline_metric_value(raw_value, metric_type, source)
                    if source and metric_type in {"pct", "r", "duration"}
                    else _format_metric_value(raw_value, metric_type)
                )
                if out is None:
                    if section == "Drawdown" and market in {"fx", "crypto"}:
                        missing_markets.append(market)
                    continue
                if section == "Drawdown" and (pos[0], col) in manual_drawdown_metric_cells:
                    continue
                wrote = _write_dashboard_metric_cell(
                    dash,
                    pos[0],
                    col,
                    out,
                    "raw" if source else metric_type,
                    semantic,
                )
                if wrote and source and semantic and _as_float(raw_value) not in (None, 0):
                    if semantic == "profit_loss":
                        _apply_full_cell_semantic_fill(dash.cell(pos[0], col), "profit" if float(raw_value) > 0 else "loss")
                    elif semantic in {"loss", "drawdown"}:
                        _apply_full_cell_semantic_fill(dash.cell(pos[0], col), "loss")
                    elif semantic == "profit":
                        _apply_full_cell_semantic_fill(dash.cell(pos[0], col), "profit")
            if missing_markets:
                diagnostics.setdefault("missing_market_drawdown_values", []).extend(
                    f"{market} {label}" for market in missing_markets
                )

        def _dashboard_label_rows_by_col(label_col: int = 1) -> Dict[str, List[int]]:
            rows_by_label: Dict[str, List[int]] = defaultdict(list)
            for r in range(1, dash.max_row + 1):
                label = str(dash.cell(r, label_col).value or "").strip().lower()
                if label:
                    rows_by_label[label].append(r)
            return rows_by_label

        def write_horizontal_core_market_metrics() -> None:
            market_cols: Dict[str, int] = {}
            for r in range(1, min(5, dash.max_row) + 1):
                row_tokens = {str(dash.cell(r, c).value or "").strip().lower(): c for c in range(1, min(8, dash.max_column) + 1)}
                candidate = {
                    "overall": row_tokens.get("overall"),
                    "fx": row_tokens.get("fx") or row_tokens.get("forex"),
                    "crypto": row_tokens.get("crypto"),
                }
                if all(candidate.values()) and candidate["overall"] + 1 == candidate["fx"] and candidate["fx"] + 1 == candidate["crypto"] and candidate["overall"] > 1:
                    market_cols = {k: int(v) for k, v in candidate.items() if v}
                    break
            if not {"overall", "fx", "crypto"}.issubset(market_cols):
                return
            label_rows = _dashboard_label_rows_by_col(1)
            for legacy_label in ("net p/l", "net p/l %"):
                for row_num in label_rows.get(legacy_label, []):
                    if not _is_merged_non_anchor(dash, row_num, 1):
                        dash.cell(row_num, 1).value = "Net P/L Percentage"
            label_rows = _dashboard_label_rows_by_col(1)
            metric_specs = [
                (["Trades"], "trades", "count", None),
                (["Wins"], "wins", "count", None),
                (["Losses"], "losses", "count", None),
                (["Break-even"], "break_even", "count", None),
                (["Test"], "test_trades", "count", None),
                (["Win rate"], "win_rate_pct", "pct", None),
                (["Net P/L Percentage", "Net P/L %", "Net P/L"], "net_result_pct", "pct", "profit_loss"),
                (["Net P/L R multiples", "Net P/L R multiple", "Net R Multiple"], "net_r_multiple", "r", "profit_loss"),
                (["Gross percent gain", "Gross gain"], "gross_gain_result_pct", "pct", "profit_loss"),
                (["Gross percent loss", "Gross loss"], "gross_loss_result_pct", "pct", "loss"),
                (["Gross IR gain"], "gross_ir_gain", "r", "profit"),
                (["Gross IR loss"], "gross_ir_loss", "r", "loss"),
                (["Best Win Streak", "Winning Streak"], "winning_streak", "count", None),
                (["Worst Losing Streak", "Losing Streak"], "losing_streak", "count", None),
                (["Percentage expectancy", "Avg result %", "Average result percent"], "avg_result_pct", "pct", "profit_loss"),
                (["R expectancy", "Avg R", "Average R"], "avg_r_multiple", "r", "profit_loss"),
                (["Avg stop %"], "avg_stop_pct", "pct", None),
                (["Avg target %"], "avg_target_pct", "pct", None),
                (["Min stop %"], "min_stop_pct", "pct", None),
                (["Max stop %"], "max_stop_pct", "pct", None),
                (["Min target %"], "min_target_pct", "pct", None),
                (["Max target %"], "max_target_pct", "pct", None),
                (["Min duration", "Min duration (DD:HH:MM:SS)"], "min_duration_seconds", "duration", None),
                (["Avg duration", "Avg duration (DD:HH:MM:SS)"], "avg_duration_seconds", "duration", None),
                (["Avg winning trade duration", "Avg winning trade duration (DD:HH:MM:SS)"], "avg_duration_seconds_winners", "duration", None),
                (["Avg losing trade duration", "Avg losing trade duration (DD:HH:MM:SS)"], "avg_duration_seconds_losers", "duration", None),
                (["Max duration", "Max duration (DD:HH:MM:SS)"], "max_duration_seconds", "duration", None),
                (["Min Move to Break Even", "Min Move to Break Even (DD:HH:MM:SS)"], "min_move_to_break_even_duration_seconds", "duration", None),
                ([
                    DASHBOARD_MOVE_TO_BREAK_EVEN_LABEL,
                    "Move to Break Even (DD:HH:MM:SS)",
                    "Move to Break-Even (DD:HH:MM:SS)",
                    "Average Move to Break Even (DD:HH:MM:SS)",
                ], "move_to_break_even_duration_seconds", "duration", None),
                (["Max Move to Break Even", "Max Move to Break Even (DD:HH:MM:SS)"], "max_move_to_break_even_duration_seconds", "duration", None),
                (["Min Move to Profit", "Min Move to Profit (DD:HH:MM:SS)"], "min_move_to_profit_duration_seconds", "duration", None),
                ([
                    DASHBOARD_MOVE_TO_PROFIT_LABEL,
                    "Move to Profit (DD:HH:MM:SS)",
                    "Average Move to Profit (DD:HH:MM:SS)",
                ], "move_to_profit_duration_seconds", "duration", None),
                (["Max Move to Profit", "Max Move to Profit (DD:HH:MM:SS)"], "max_move_to_profit_duration_seconds", "duration", None),
                (["Max loss %"], "min_result_pct", "pct", "loss"),
                (["Max win %"], "max_result_pct", "pct", None),
                (["Max R loss"], "min_r_multiple", "r", "loss"),
                (["Max R win"], "max_r_multiple", "r", None),
            ]
            percentage_totals = _result_percentage_totals_by_market(rows, snapshot.get("balances") or stats.get("balances") or [])
            buckets = {
                market: _merge_metric_buckets(
                    dict(bucket or {}),
                    percentage_totals[market],
                    {key: value for key, value in extended_metrics[market].items() if value is not None},
                    {key: value for key, value in move_duration_metrics[market].items() if value is not None},
                )
                for market, bucket in {
                    "overall": by_market.get("overall") or totals,
                    "fx": by_market.get("fx") or {},
                    "crypto": by_market.get("crypto") or {},
                }.items()
            }
            for labels, key, metric_type, semantic in metric_specs:
                rows_for_metric: List[int] = []
                for label in labels:
                    rows_for_metric.extend(label_rows.get(label.lower(), []))
                if not rows_for_metric:
                    diagnostics.setdefault("missing_dashboard_metric_labels", []).append(" / ".join(labels))
                    continue
                for row_num in sorted(set(rows_for_metric)):
                    for market, col in market_cols.items():
                        bucket = buckets.get(market) or {}
                        raw_value = bucket.get(key)
                        inline_source = _inline_metric_source(bucket, key)
                        value = (
                            _format_inline_metric_value(raw_value, metric_type, inline_source)
                            if inline_source and metric_type in {"pct", "r", "duration"}
                            else _format_metric_value(raw_value, metric_type)
                        )
                        if value is None:
                            diagnostics.setdefault("missing_dashboard_metric_values", []).append(f"{market} {labels[0]}")
                            continue
                        wrote = _write_dashboard_metric_cell(
                            dash,
                            row_num,
                            col,
                            value,
                            "raw" if inline_source else metric_type,
                            semantic,
                            allow_grey=False,
                        )
                        if wrote and inline_source and semantic and _as_float(raw_value) not in (None, 0):
                            if semantic == "profit_loss":
                                _apply_full_cell_semantic_fill(dash.cell(row_num, col), "profit" if float(raw_value) > 0 else "loss")
                            elif semantic in {"loss", "drawdown"}:
                                _apply_full_cell_semantic_fill(dash.cell(row_num, col), "loss")
                            elif semantic == "profit":
                                _apply_full_cell_semantic_fill(dash.cell(row_num, col), "profit")

            first_winners_row = next(
                (row for row in range(1, dash.max_row + 1)
                 if str(dash.cell(row, 1).value or "").strip().casefold() == "winners"),
                dash.max_row + 1,
            )
            def write_core_recommendation_after(anchor_label: str, key: str) -> None:
                for row_num in label_rows.get(anchor_label.lower(), []):
                    if row_num >= first_winners_row:
                        continue
                    rec_row = row_num + 1
                    if rec_row > dash.max_row or str(dash.cell(rec_row, 1).value or "").strip().casefold() != "recommendation":
                        continue
                    for market, col in market_cols.items():
                        value = _sanitize_recommendation_text(key, (buckets.get(market) or {}).get(key))
                        if _write_value_preserving_cell(dash, rec_row, col, _excel_scalar(value)):
                            dash.cell(rec_row, col).number_format = "General"
                            _apply_recommendation_cell_style(dash.cell(rec_row, col))
                            diagnostics["updated_cells"] += 1
                    break

            write_core_recommendation_after("Max stop %", STOP_RECOMMENDATION_HEADER)
            write_core_recommendation_after("Max target %", TARGET_RECOMMENDATION_HEADER)

            for labels, detail_key in (
                (["Worst Losing Streak", "Losing Streak"], "longest_losing_streak"),
                (["Best Win Streak", "Winning Streak"], "longest_winning_streak"),
            ):
                rows_for_metric = []
                for label in labels:
                    rows_for_metric.extend(label_rows.get(label.lower(), []))
                for row_num in sorted(set(rows_for_metric)):
                    metric_label = str(dash.cell(row_num, 1).value or "").strip()
                    details = {
                        col: (buckets.get(market) or {}).get(detail_key)
                        for market, col in market_cols.items()
                    }
                    counts = {
                        col: (buckets.get(market) or {}).get("winning_streak" if detail_key == "longest_winning_streak" else "losing_streak")
                        for market, col in market_cols.items()
                    }
                    _write_streak_detail_rows(dash, metric_label, details, counts, diagnostics=diagnostics)

        write_horizontal_core_market_metrics()

        section_maps = {
            "Overall": by_market.get("overall") or totals,
            "FX": by_market.get("fx") or {},
            "Crypto": by_market.get("crypto") or {},
        }
        for section, bucket in section_maps.items():
            if section not in section_sheets:
                continue
            metric_sources = bucket.get("metric_sources") or {}
            write_metric(section, "Trades", bucket.get("trades"), "count")
            write_metric(section, "Wins", bucket.get("wins"), "count")
            write_metric(section, "Losses", bucket.get("losses"), "count")
            write_metric(section, "Best Win Streak", bucket.get("winning_streak"), "count")
            write_metric(section, "Winning Streak", bucket.get("winning_streak"), "count")
            write_metric(section, "Worst Losing Streak", bucket.get("losing_streak"), "count")
            write_metric(section, "Losing Streak", bucket.get("losing_streak"), "count")
            write_metric(section, "Break-even", bucket.get("break_even"), "count")
            write_metric(section, "Test", bucket.get("test_trades"), "count")
            write_metric(section, "Win rate", bucket.get("win_rate_pct"), "pct")
            write_metric(section, "Net P/L", bucket.get("net_profit_total"))
            for label in ("Percentage expectancy", "Avg result %"):
                write_metric(section, label, bucket.get("avg_result_pct"), "pct")
            for label in ("R expectancy", "Avg R"):
                write_metric(section, label, bucket.get("avg_r_multiple"), "r")
            write_metric(section, "Gross gain", bucket.get("gross_gain"))
            write_metric(section, "Gross loss", bucket.get("gross_loss"))
            write_metric(section, "Max loss %", bucket.get("min_result_pct"), "pct", metric_sources.get("min_result_pct"))
            write_metric(section, "Max win %", bucket.get("max_result_pct"), "pct", metric_sources.get("max_result_pct"))
            write_metric(section, "Max R loss", bucket.get("min_r_multiple"), "r", metric_sources.get("min_r_multiple"))
            write_metric(section, "Max R win", bucket.get("max_r_multiple"), "r", metric_sources.get("max_r_multiple"))
            write_metric(section, "Avg stop %", bucket.get("avg_stop_pct"), "pct")
            write_metric(section, "Avg target %", bucket.get("avg_target_pct"), "pct")
            write_metric(section, "Max target %", bucket.get("max_target_pct"), "pct", metric_sources.get("max_target_pct"))
            write_metric(section, "Avg duration", bucket.get("avg_duration_seconds"), "duration")


        risk_by_market = risk.get("by_market") if isinstance(risk.get("by_market"), dict) else {}
        def _risk_market_values(key: str) -> Dict[str, Any]:
            return {
                "overall": risk.get(key),
                "fx": (risk_by_market.get("fx") or {}).get(key),
                "crypto": (risk_by_market.get("crypto") or {}).get(key),
            }

        def _extended_market_values(key: str) -> Dict[str, Any]:
            return {
                market: (extended_metrics.get(market) or {}).get(key)
                for market in ("overall", "fx", "crypto")
            }

        def _extended_market_sources(key: str) -> Dict[str, Any]:
            return {
                market: ((extended_metrics.get(market) or {}).get("metric_sources") or {}).get(key)
                for market in ("overall", "fx", "crypto")
            }

        def _winner_loser_market_values(key: str) -> Dict[str, Any]:
            """Prefer values derived from the matching outcome subset.

            Some upstream risk payloads historically exposed unfiltered market
            averages under winner/loser keys.  The normalized trade rows are
            authoritative whenever they can derive the requested metric; the
            key-specific risk value remains a fallback for snapshots that do
            not include rows.
            """
            derived = _extended_market_values(key)
            fallback = _risk_market_values(key)
            return {
                market: (
                    derived.get(market)
                    if derived.get(market) is not None
                    else fallback.get(market)
                )
                for market in ("overall", "fx", "crypto")
            }

        write_market_metric("Winners", ["Percentage expectancy", "Avg result %", "Avg win %"], _winner_loser_market_values("avg_result_pct_winners"), "pct", "profit_loss")
        write_market_metric("Winners", ["R expectancy", "Avg R", "Avg R win"], _winner_loser_market_values("avg_r_multiple_winners"), "r", "profit_loss")
        write_market_metric("Winners", "Avg stop %", _winner_loser_market_values("avg_stop_pct_winners"), "pct")
        write_market_metric("Winners", "Avg target %", _winner_loser_market_values("avg_target_pct_winners"), "pct")
        write_market_metric("Losers", ["Percentage expectancy", "Avg result %", "Avg loss %"], _winner_loser_market_values("avg_result_pct_losers"), "pct", "loss")
        write_market_metric("Losers", ["R expectancy", "Avg R", "Avg R loss"], _winner_loser_market_values("avg_r_multiple_losers"), "r", "loss")
        write_market_metric("Losers", "Avg stop %", _winner_loser_market_values("avg_stop_pct_losers"), "pct")
        write_market_metric("Losers", "Avg target %", _winner_loser_market_values("avg_target_pct_losers"), "pct")
        for section, suffix in (("Winners", "winners"), ("Losers", "losers")):
            for label, key in (
                ("Min duration", f"min_duration_seconds_{suffix}"),
                ("Avg duration", f"avg_duration_seconds_{suffix}"),
                ("Max duration", f"max_duration_seconds_{suffix}"),
            ):
                write_market_metric(
                    section,
                    label,
                    _extended_market_values(key),
                    "duration",
                    sources_by_market=_extended_market_sources(key),
                )
        for section, suffix, semantic in (("Winners", "winners", "profit"), ("Losers", "losers", "loss")):
            label_aliases = {
                ("Winners", "Min result %"): ["Min result %", "Min win %"],
                ("Winners", "Max result %"): ["Max result %", "Max win %"],
                ("Winners", "Min R"): ["Min R", "Min R win"],
                ("Winners", "Max R"): ["Max R", "Max R win"],
                ("Losers", "Min result %"): ["Min result %", "Max loss %"],
                ("Losers", "Max result %"): ["Max result %", "Min loss %"],
                ("Losers", "Min R"): ["Min R", "Max R loss"],
                ("Losers", "Max R"): ["Max R", "Min R loss"],
            }
            for label, key, metric_type in (
                ("Min stop %", f"min_stop_pct_{suffix}", "pct"),
                ("Max stop %", f"max_stop_pct_{suffix}", "pct"),
                ("Min target %", f"min_target_pct_{suffix}", "pct"),
                ("Max target %", f"max_target_pct_{suffix}", "pct"),
                ("Min result %", f"min_result_pct_{suffix}", "pct"),
                ("Max result %", f"max_result_pct_{suffix}", "pct"),
                ("Min R", f"min_r_multiple_{suffix}", "r"),
                ("Max R", f"max_r_multiple_{suffix}", "r"),
            ):
                write_market_metric(
                    section,
                    label_aliases.get((section, label), label),
                    _extended_market_values(key),
                    metric_type,
                    semantic if "result" in key or "r_multiple" in key else None,
                    _extended_market_sources(key),
                )

        if "Side" in anchors:
            side_rows = [
                ("Long", "long_trades", None),
                ("Winners", "long_wins", "profit"),
                ("Losers", "long_losses", "loss"),
                ("Short", "short_trades", None),
                ("Winners", "short_wins", "profit"),
                ("Losers", "short_losses", "loss"),
            ]
            market_cols = _main_dashboard_market_columns()
            for offset, (_label, key, semantic) in enumerate(side_rows):
                row_num = anchors["Side"]["start_row"] + offset
                for market, col in market_cols.items():
                    value = (extended_metrics.get(market) or {}).get(key)
                    if value is not None:
                        _write_dashboard_metric_cell(dash, row_num, col, int(value), "count", semantic)

        if "Patterns" in anchors:
            for label, key in (
                ("Most Traded", "most_traded_pattern"),
                ("Least Traded", "least_traded_pattern"),
                ("Most Profitable", "most_profitable_pattern"),
                ("Least Profitable", "least_profitable_pattern"),
            ):
                write_market_metric("Patterns", label, _extended_market_values(key))
            market_cols = _main_dashboard_market_columns()
            current_pattern = ""
            for row_num in range(anchors["Patterns"]["start_row"], anchors["Patterns"]["end_row"] + 1):
                label = str(dash.cell(row_num, 1).value or "").strip().casefold()
                key_suffix = None
                semantic = None
                if label in {"channel", "range"}:
                    current_pattern = label
                    key_suffix = "total"
                elif label in {"winners", "winner"} and current_pattern:
                    key_suffix = "wins"
                    semantic = "profit"
                elif label in {"losers", "loser", "losses"} and current_pattern:
                    key_suffix = "losses"
                    semantic = "loss"
                else:
                    current_pattern = "" if label in {"most traded", "least traded", "most profitable", "least profitable"} else current_pattern
                if not current_pattern or not key_suffix:
                    continue
                metric_key = f"pattern_{current_pattern}_{key_suffix}"
                for market, col in market_cols.items():
                    value = (extended_metrics.get(market) or {}).get(metric_key)
                    if value is not None:
                        _write_dashboard_metric_cell(dash, row_num, col, int(value), "count", semantic)

        if "Timeframe" in anchors:
            market_cols = _main_dashboard_market_columns()
            current_timeframe = ""
            for label in ("1MIN", "5MIN", "15MIN", "30MIN", "1H", "4H", "DAILY", "WEEKLY", "MONTHLY"):
                write_market_metric(
                    "Timeframe",
                    label,
                    _extended_market_values(f"timeframe_{label.lower()}"),
                    "count",
                )
            timeframe_labels = {"1MIN", "5MIN", "15MIN", "30MIN", "1H", "4H", "DAILY", "WEEKLY", "MONTHLY"}
            for row_num in range(anchors["Timeframe"]["start_row"], anchors["Timeframe"]["end_row"] + 1):
                label_raw = str(dash.cell(row_num, 1).value or "").strip().upper()
                key_suffix = None
                semantic = None
                if label_raw in timeframe_labels:
                    current_timeframe = label_raw
                    continue
                if label_raw in {"WINNERS", "WINNER"} and current_timeframe:
                    key_suffix = "wins"
                    semantic = "profit"
                elif label_raw in {"LOSERS", "LOSER", "LOSSES"} and current_timeframe:
                    key_suffix = "losses"
                    semantic = "loss"
                if not current_timeframe or not key_suffix:
                    continue
                metric_key = f"timeframe_{current_timeframe.lower()}_{key_suffix}"
                for market, col in market_cols.items():
                    value = (extended_metrics.get(market) or {}).get(metric_key)
                    if value is not None:
                        _write_dashboard_metric_cell(dash, row_num, col, int(value), "count", semantic)

        if "Commission" in anchors:
            for label, key in (
                ("Min Commission", "min_commission"),
                ("Avg Commission", "avg_commission"),
                ("Max Commission", "max_commission"),
                ("Total Commission", "total_commission"),
            ):
                write_market_metric("Commission", label, _extended_market_values(key))

        if "Drawdown" in anchors:
            market_cols = _main_dashboard_market_columns()
            for drawdown_label in ("Min drawdown", "Max drawdown"):
                drawdown_pos = _find_label_in_section(dash, drawdown_label, anchors["Drawdown"])
                if not drawdown_pos:
                    continue
                drawdown_row = drawdown_pos[0]
                for col in market_cols.values():
                    if _drawdown_inline_pct_fraction(dash.cell(drawdown_row, col).value) is not None:
                        manual_drawdown_metric_cells.add((drawdown_row, col))
            _write_drawdown_detail_rows(
                dash,
                "Min drawdown",
                {
                    col: (extended_metrics.get(market) or {}).get("min_drawdown_detail")
                    for market, col in market_cols.items()
                },
                diagnostics=diagnostics,
                preserve_inline_cells=manual_drawdown_metric_cells,
            )
            _write_drawdown_detail_rows(
                dash,
                "Max drawdown",
                {
                    col: (extended_metrics.get(market) or {}).get("max_drawdown_detail")
                    for market, col in market_cols.items()
                    },
                    diagnostics=diagnostics,
                    preserve_inline_cells=manual_drawdown_metric_cells,
                )
            anchors = _find_anchor_sections(
                dash,
                ["Overall", "Winners", "Losers", "Drawdown"],
                optional=["Duration", "Side", "Patterns", "Timeframe", "Commission"],
            )
            section_sheets = {
                **{name: dash for name in anchors},
                **{name: detail_dash for name in detail_anchors},
            }
            section_anchors = {**anchors, **detail_anchors}

        write_market_metric(
            "Drawdown",
            "Min drawdown",
            _extended_market_values("min_drawdown_pct"),
            "pct",
            "drawdown",
        )
        write_market_metric("Drawdown", "Max drawdown", _extended_market_values("max_drawdown_pct"), "pct", "drawdown")
        write_market_metric("Drawdown", "Avg drawdown", _extended_market_values("avg_drawdown_pct"), "pct", "drawdown")
        if "Drawdown" in anchors:
            market_cols = _main_dashboard_market_columns()
            _write_drawdown_detail_rows(
                dash,
                "Min drawdown",
                {
                    col: (extended_metrics.get(market) or {}).get("min_drawdown_detail")
                    for market, col in market_cols.items()
                },
                diagnostics=diagnostics,
                preserve_inline_cells=manual_drawdown_metric_cells,
            )
            _write_drawdown_detail_rows(
                dash,
                "Max drawdown",
                {
                    col: (extended_metrics.get(market) or {}).get("max_drawdown_detail")
                    for market, col in market_cols.items()
                },
                diagnostics=diagnostics,
                preserve_inline_cells=manual_drawdown_metric_cells,
            )
        duration = groups.get("duration") or {}
        dsrc = duration.get("metric_sources") or {}
        if "FX" in section_sheets:
            write_metric("FX", "FX shortest", duration.get("fx_shortest_seconds"), "duration", dsrc.get("fx_shortest_seconds"))
            write_metric("FX", "FX longest", duration.get("fx_longest_seconds"), "duration", dsrc.get("fx_longest_seconds"))
        if "Crypto" in section_sheets:
            write_metric("Crypto", "Crypto shortest", duration.get("crypto_shortest_seconds"), "duration", dsrc.get("crypto_shortest_seconds"))
            write_metric("Crypto", "Crypto longest", duration.get("crypto_longest_seconds"), "duration", dsrc.get("crypto_longest_seconds"))
        if "Duration" in anchors:
            write_metric("Duration", "Overall avg", duration.get("overall_avg_seconds"), "duration")
            write_metric("Duration", "Overall shortest", duration.get("overall_shortest_seconds"), "duration", dsrc.get("overall_shortest_seconds"))
            write_metric("Duration", "Overall longest", duration.get("overall_longest_seconds"), "duration", dsrc.get("overall_longest_seconds"))
            write_metric("Duration", "FX shortest", duration.get("fx_shortest_seconds"), "duration", dsrc.get("fx_shortest_seconds"))
            write_metric("Duration", "FX longest", duration.get("fx_longest_seconds"), "duration", dsrc.get("fx_longest_seconds"))
            write_metric("Duration", "Crypto shortest", duration.get("crypto_shortest_seconds"), "duration", dsrc.get("crypto_shortest_seconds"))
            write_metric("Duration", "Crypto longest", duration.get("crypto_longest_seconds"), "duration", dsrc.get("crypto_longest_seconds"))

        diagnostics.setdefault("leader_payload_keys", [])
        market_cols = _stats1_market_columns(dash)
        leader_row_specs = [
            ("Winners", "Most wins", {
                "overall": "most_wins_instrument",
                "fx": "fx_most_wins_instrument",
                "crypto": "crypto_most_wins_instrument",
            }),
            ("Losers", "Most losses", {
                "overall": "most_losses_instrument",
                "fx": "fx_most_losses_instrument",
                "crypto": "crypto_most_losses_instrument",
            }),
        ]
        for section_name, label, key_by_market in leader_row_specs:
            bounds = _stats1_section_bounds(dash, section_name)
            if not bounds:
                continue
            row_idx = next(
                (row for row in range(bounds[0] + 1, bounds[1] + 1)
                 if str(dash.cell(row, 1).value or "").strip().casefold() == label.casefold()),
                None,
            )
            if not row_idx:
                diagnostics.setdefault("missing_dashboard_metric_labels", []).append(label)
                continue
            for market, col in market_cols.items():
                payload = leaders.get(key_by_market[market]) or {}
                if not payload:
                    continue
                diagnostics["leader_payload_keys"].append(key_by_market[market])
                normalized_payload = dict(payload)
                if normalized_payload.get("trades") is None and normalized_payload.get("total_trades") is not None:
                    normalized_payload["trades"] = normalized_payload.get("total_trades")
                if normalized_payload.get("total_trades") is None and normalized_payload.get("trades") is not None:
                    normalized_payload["total_trades"] = normalized_payload.get("trades")
                count_key = "wins" if label.casefold() == "Most wins".casefold() else "losses"
                if _write_value_preserving_cell(dash, row_idx, col, _instrument_leader_scalar(normalized_payload, count_key)):
                    dash.cell(row_idx, col).number_format = "General"
                    diagnostics["updated_cells"] += 1

        leader_header_row, leader_headers, metric_rows, leader_start_col = _find_instrument_leaders_table(detail_dash)
        if leader_header_row and leader_headers:
            clear_cols = range(leader_start_col, max(leader_headers.values()) + 1)
            clear_start_row = max(1, leader_header_row - 1)
            clear_end_row = max(metric_rows.values(), default=leader_header_row)
            for row_idx in range(clear_start_row, clear_end_row + 1):
                for col in clear_cols:
                    cell = detail_dash.cell(row_idx, col)
                    cell.value = None
                    cell.comment = None
                    cell.hyperlink = None
            diagnostics["cleared_stats2_instrument_leaders_table"] = True

        balances = _canonicalize_and_dedupe_balances(snapshot.get("balances") or [])
        _clear_account_balance_source_metadata(wb)
        diagnostics.setdefault("non_numeric_balance_accounts", [])
        section = detail_anchors["Account Balances"]
        _normalize_stats2_account_balance_headers(detail_dash, section, diagnostics)
        header_row, col_map = _find_dashboard_table_headers(detail_dash, section)
        if not header_row or "account" not in col_map or "balance" not in col_map or "currency" not in col_map:
            raise RuntimeError("Account Balances headers missing in section.")
        diagnostics.setdefault("stale_account_balance_rows_cleared", 0)
        account_col = col_map["account"]
        existing_rows_by_canonical: Dict[str, List[int]] = {}
        existing_rows_by_raw: Dict[str, List[int]] = {}
        for rr in range(header_row + 1, section["end_row"] + 1):
            raw_label = str(detail_dash.cell(rr, account_col).value or "").strip()
            if not raw_label:
                continue
            canonical_label = _canonical_account_label(raw_label)
            existing_rows_by_raw.setdefault(raw_label, []).append(rr)
            existing_rows_by_canonical.setdefault(canonical_label, []).append(rr)
        account_balance_targets: Dict[str, Dict[str, Any]] = {}
        for b in balances:
            label = _canonical_account_label(b.get("account_label") or b.get("account"))
            if not label:
                continue
            bal_num = _as_float(b.get("balance"))
            if bal_num is None:
                diagnostics["non_numeric_balance_accounts"].append(label)
                continue
            if label == "BYBIT":
                bybit_rows = existing_rows_by_canonical.get("BYBIT", [])
                bybit_live_rows = [rr for rr in range(header_row + 1, section["end_row"] + 1) if _canonical_account_label(detail_dash.cell(rr, account_col).value) == "BYBIT" and str(detail_dash.cell(rr, account_col).value or "").strip() != "BYBIT"]
                if not bybit_rows and bybit_live_rows:
                    target = bybit_live_rows[0]
                    if _write_value_preserving_cell(detail_dash, target, col_map["account"], "BYBIT"):
                        diagnostics["updated_cells"] += 1
                    existing_rows_by_canonical.setdefault("BYBIT", []).append(target)
                    bybit_rows = existing_rows_by_canonical["BYBIT"]
            try:
                row = _ensure_account_balance_row(detail_dash, section, header_row, col_map, label)
            except Exception as exc:
                diagnostics["missing_accounts"].append(label)
                diagnostics.setdefault("account_balance_write_errors", []).append(str(exc))
                continue
            if _write_value_preserving_cell(detail_dash, row, col_map["account"], label):
                diagnostics["updated_cells"] += 1
            if _write_value_preserving_cell(detail_dash, row, col_map["balance"], bal_num):
                diagnostics["updated_cells"] += 1
            _write_account_balance_source_metadata(
                wb,
                label,
                b.get("balance_source") or b.get("source"),
                b.get("as_of"),
            )
            curr = str(b.get("currency") or "").strip()
            existing_fmt = str(detail_dash.cell(row, col_map["balance"]).number_format or "")
            if curr:
                if not existing_fmt or existing_fmt == "General":
                    detail_dash.cell(row, col_map["balance"]).number_format = _currency_number_format(curr)
                elif _is_crypto_currency(curr) and "#" not in existing_fmt:
                    detail_dash.cell(row, col_map["balance"]).number_format = _currency_number_format(curr, force_decimals=10)
                elif (not _is_crypto_currency(curr)) and "#" not in existing_fmt:
                    detail_dash.cell(row, col_map["balance"]).number_format = _currency_number_format(curr, force_decimals=2)
            if curr:
                if _write_value_preserving_cell(detail_dash, row, col_map["currency"], curr):
                    diagnostics["updated_cells"] += 1
            account_balance_targets[label] = {"row": row, "balance": bal_num, "currency": curr}
            if "as_of" in col_map:
                as_of = str(b.get("as_of") or "").strip()
                if as_of:
                    if _write_value_preserving_cell(detail_dash, row, col_map["as_of"], as_of):
                        diagnostics["updated_cells"] += 1
            if label == "BYBIT":
                stale_rows = []
                for rr in range(header_row + 1, section["end_row"] + 1):
                    raw_here = str(detail_dash.cell(rr, account_col).value or "").strip()
                    if _canonical_account_label(raw_here) == "BYBIT" and raw_here != "BYBIT" and rr != row:
                        stale_rows.append(rr)
                for stale_row in stale_rows:
                    _clear_account_balance_row(detail_dash, stale_row, col_map)
                    diagnostics["stale_account_balance_rows_cleared"] += 1

        diagnostics.setdefault("account_balance_verified", [])
        diagnostics.setdefault("account_balance_mismatches", [])
        for label, target in account_balance_targets.items():
            row = int(target["row"])
            expected = _as_float(target.get("balance"))
            actual = _as_float(detail_dash.cell(row, col_map["balance"]).value)
            if expected is not None and actual is not None and abs(actual - expected) <= max(1e-9, abs(expected) * 1e-12):
                diagnostics["account_balance_verified"].append(label)
            else:
                diagnostics["account_balance_mismatches"].append({"account": label, "expected": expected, "actual": actual, "row": row})
        if diagnostics["account_balance_mismatches"]:
            return {"ok": False, "error": "dashboard_account_balance_verification_failed", "diagnostics": diagnostics}
        _write_stats2_risk_of_ruin(detail_dash, section, header_row, col_map, rows, diagnostics)
        _write_stats2_net_pl_percentages(detail_dash, section, header_row, col_map, rows, diagnostics)

        zero_qty = _collect_zero_qty_validation(rows)
        diagnostics.update(zero_qty)
        if zero_qty["crypto_zero_qty_unrepaired"]:
            sample = ", ".join(str(x.get("id") or x.get("symbol") or "?") for x in zero_qty["crypto_zero_qty_unrepaired"][:5])
            return {"ok": False, "error": f"Unrepaired crypto zero-quantity trade rows detected: {sample}", "diagnostics": diagnostics}

        _apply_dashboard_requested_semantic_fills(dash)
        _repair_stats1_formatting(
            dash,
            extended_metrics,
            diagnostics,
            preserve_user_presentation=preserve_existing_layout,
        )
        _repair_legacy_duration_number_formats(wb, diagnostics)

        tmp = path.with_suffix(".update.tmp.xlsx")
        build_master_journal_workbook(
            snapshot,
            tmp,
            publish_recommendation_assets=False,
            manual_overrides={},
        )
        gen = load_workbook(tmp, data_only=False)
        try:
            if diagnostics.get("created_stats2_from_legacy_layout"):
                src_detail = _stats2_sheet(gen, required=True)
                dst_detail = _stats2_sheet(wb, required=True)
                for row in src_detail.iter_rows():
                    for source in row:
                        target = dst_detail.cell(source.row, source.column)
                        target.value = source.value
                        target.number_format = source.number_format
                        target.font = copy(source.font)
                        target.fill = copy(source.fill)
                        target.border = copy(source.border)
                        target.alignment = copy(source.alignment)
                        target.protection = copy(source.protection)
                        target.comment = copy(source.comment) if source.comment else None
            def _copy_data_rows(src_ws, dst_ws, start_row: int, *, force_all_columns: bool = False):
                if force_all_columns:
                    src_map = _trade_log_header_map(src_ws)
                    dst_map = _trade_log_header_map(dst_ws)
                    missing_src = [header for header in TRADE_LOG_HEADERS if header not in src_map]
                    missing_dst = [header for header in TRADE_LOG_HEADERS if header not in dst_map]
                    if missing_src or missing_dst:
                        raise RuntimeError(
                            "Trade Log logical headers do not match expected template: "
                            f"missing_source={missing_src!r}, missing_destination={missing_dst!r}."
                        )
                    header_pairs = [(src_map[header], dst_map[header], header) for header in TRADE_LOG_HEADERS]
                    max_col = len(TRADE_LOG_HEADERS)
                else:
                    src_headers = [str(src_ws.cell(1, c).value or "").strip() for c in range(1, src_ws.max_column + 1)]
                    dst_headers = [str(dst_ws.cell(1, c).value or "").strip() for c in range(1, dst_ws.max_column + 1)]
                    src_map = {header: idx + 1 for idx, header in enumerate(src_headers) if header}
                    dst_map = {header: idx + 1 for idx, header in enumerate(dst_headers) if header}
                    max_col = min(src_ws.max_column, dst_ws.max_column)
                    header_pairs = [
                        (src_map[header], dst_col, header)
                        for header, dst_col in dst_map.items()
                        if header in src_map and dst_col <= max_col
                    ]
                clear_max_col = max(dst_ws.max_column, max_col) if force_all_columns else max_col
                for row in range(start_row, dst_ws.max_row + 1):
                    for col in range(1, clear_max_col + 1):
                        if _is_merged_non_anchor(dst_ws, row, col):
                            continue
                        cell = dst_ws.cell(row, col)
                        cell.value = None
                        cell.comment = None
                        cell.hyperlink = None
                src_start_row = _trade_log_data_start_row(src_ws) if force_all_columns else start_row
                dst_row = start_row
                for src_row in range(src_start_row, src_ws.max_row + 1):
                    if force_all_columns and not any(src_ws.cell(src_row, col).value not in (None, "") for col in range(1, max_col + 1)):
                        continue
                    for src_col, dst_col, _header in header_pairs:
                        if _is_merged_non_anchor(dst_ws, dst_row, dst_col):
                            continue
                        src_cell = src_ws.cell(src_row, src_col)
                        dst_cell = dst_ws.cell(dst_row, dst_col)
                        dst_cell.value = src_cell.value
                        dst_cell.number_format = src_cell.number_format
                        dst_cell.comment = copy(src_cell.comment) if src_cell.comment else None
                        dst_cell.hyperlink = copy(src_cell.hyperlink) if src_cell.hyperlink else None
                    dst_row += 1
                last_row = max(start_row - 1, dst_row - 1)
                if force_all_columns:
                    _set_trade_log_auto_filter(dst_ws)
                    _hide_trade_log_row_id(dst_ws)
                    _apply_trade_log_dropdown_validations(dst_ws)
                elif dst_ws.auto_filter and dst_ws.auto_filter.ref:
                    last_col_letter = get_column_letter(max_col)
                    dst_ws.auto_filter.ref = f"A1:{last_col_letter}{max(1,last_row)}"


            gen_trade_log = _get_all_trades_sheet(gen, allow_legacy=False)
            live_trade_log = _get_all_trades_sheet(wb, allow_legacy=False)
            _copy_data_rows(gen_trade_log, live_trade_log, TRADE_LOG_DATA_START_ROW, force_all_columns=True)
            _normalize_trade_log_row_heights(live_trade_log, diagnostics)
            _repair_trade_log_row_ids_from_rows(live_trade_log, rows, diagnostics)
            if expected_survivor_row_ids:
                header_map = _trade_log_header_map(live_trade_log)
                ridx = header_map.get("Row ID")
                if not ridx:
                    return {
                        "ok": False,
                        "error": "workbook_row_survivor_verification_failed",
                        "missing_row_ids": sorted([rid for rid in expected_survivor_row_ids if rid]),
                        "reason": "missing_row_id_header",
                        "diagnostics": diagnostics,
                    }
                present = {str(live_trade_log.cell(rr, ridx).value or "").strip() for rr in range(_trade_log_data_start_row(live_trade_log), live_trade_log.max_row + 1)}
                missing = sorted([rid for rid in expected_survivor_row_ids if rid and rid not in present])
                if missing:
                    return {"ok": False, "error": "workbook_row_survivor_verification_failed", "missing_row_ids": missing, "diagnostics": diagnostics}
            _repair_trade_log_unknown_currency_formats(live_trade_log, rows, diagnostics)
            _repair_trade_log_move_to_durations(live_trade_log, diagnostics)
            _apply_trade_log_adaptive_formats(live_trade_log)
            _apply_trade_number_hyperlinks(live_trade_log, diagnostics)
            _apply_trade_log_win_loss_row_formatting(live_trade_log)
            _apply_trade_log_win_loss_direct_row_fills(live_trade_log)

            def _copy_instrument_rows_header_aware(src_ws, dst_ws):
                aliases = {
                    'trades': ['trades','total trades','total_trades'],
                    'wins':['wins'],'losses':['losses'],'break-even':['break-even','break even'],
                    'longs':['longs','long trades'],'shorts':['shorts','short trades'],
                    'long wins':['long wins'],'long losses':['long losses'],'long break-even':['long break-even'],
                    'short wins':['short wins'],'short losses':['short losses'],'short break-even':['short break-even'],
                    'net p/l %':['net p/l %'],'avg p/l %':['avg p/l %'],'win rate %':['win rate %'],
                    'avg stop % (w)':['avg stop % (w)'],'avg stop % (l)':['avg stop % (l)'],'avg target % (w)':['avg target % (w)'],'avg target % (l)':['avg target % (l)'],
                    'stop recommendation':['stop loss recommendation'],
                    'target recommendation':['target recommendation'],
                    'shortest':['shortest duration (dd:hh:mm:ss)','shortest (dd:hh:mm:ss)'],
                    'avgdur':['avg duration (dd:hh:mm:ss)'],
                    'longest':['longest duration (dd:hh:mm:ss)','longest (dd:hh:mm:ss)'],
                    'movebe':['move to break even'],'moveprofit':['move to profit'],
                    'pattern':['most traded pattern','pattern'],'ema':['most traded ema','ema'],
                    'ath':['all-time highs'],'atl':['all-time lows'],
                    'market':['market'],'limit':['limit'],'round':['round number'],'spiked':['spiked out'],
                    'closestop':['close stop out'],'nearentry':['near perfect entry'],
                    'nearwin':['near win'],'earlyclose':['early close'],
                    'timeframe':['most traded timeframe'],
                    'mostprofitabletimeframe':['most profitable timeframe'],
                    'leastprofitabletimeframe':['least profitable timeframe'],
                    'rmultiple':['net r multiple','r multiple'],
                    'symbol':['symbol'],'class':['class']
                }
                dst_header_row = _instrument_averages_header_row(dst_ws)
                src_start_row = _instrument_averages_data_start_row(src_ws)
                dst_start_row = _instrument_averages_data_start_row(dst_ws)
                src_headers = {
                    str(header or "").strip().casefold(): col
                    for header, col in _instrument_averages_header_map(src_ws).items()
                }
                dst_headers = {
                    str(header or "").strip().casefold(): col
                    for header, col in _instrument_averages_header_map(dst_ws).items()
                }
                def find_col(headers, keys):
                    for k in keys:
                        if k in headers:
                            return headers[k]
                    return None
                pair_by_destination = {}
                for header, dc in dst_headers.items():
                    sc = src_headers.get(header)
                    if sc:
                        pair_by_destination[dc] = (sc, dc)
                for _,keys in aliases.items():
                    sc=find_col(src_headers, keys); dc=find_col(dst_headers, keys)
                    if sc and dc:
                        pair_by_destination[dc] = (sc, dc)
                pairs = [
                    pair_by_destination[dc]
                    for dc in sorted(pair_by_destination)
                ]
                if not pairs:
                    return
                src_symbol_col = find_col(src_headers, aliases["symbol"])
                dst_symbol_col = find_col(dst_headers, aliases["symbol"])
                populated_rows = []
                if preserve_existing_layout and src_symbol_col and dst_symbol_col:
                    source_records = []
                    for src_row in range(src_start_row, src_ws.max_row + 1):
                        symbol = str(src_ws.cell(src_row, src_symbol_col).value or "").strip().upper()
                        if symbol:
                            source_records.append((symbol, src_row))

                    destination_rows_by_symbol = defaultdict(list)
                    existing_symbol_rows = []
                    for dst_row in range(dst_start_row, dst_ws.max_row + 1):
                        symbol = str(dst_ws.cell(dst_row, dst_symbol_col).value or "").strip().upper()
                        if not symbol:
                            continue
                        destination_rows_by_symbol[symbol].append(dst_row)
                        existing_symbol_rows.append(dst_row)

                    matched_records = []
                    new_records = []
                    for symbol, src_row in source_records:
                        candidates = destination_rows_by_symbol.get(symbol) or []
                        if candidates:
                            matched_records.append((src_row, candidates.pop(0)))
                        else:
                            new_records.append((symbol, src_row))

                    used_rows = {dst_row for _src_row, dst_row in matched_records}
                    removed_rows = [
                        dst_row for dst_row in existing_symbol_rows
                        if dst_row not in used_rows
                    ]
                    for dst_row in removed_rows:
                        for _src_col, dst_col in pairs:
                            cell = dst_ws.cell(dst_row, dst_col)
                            cell.value = None
                            cell.comment = None
                            cell.hyperlink = None

                    next_append_row = max(existing_symbol_rows, default=dst_start_row - 1) + 1
                    for _symbol, src_row in sorted(new_records, key=lambda item: (item[0], item[1])):
                        while any(
                            dst_ws.cell(next_append_row, dst_col).value not in (None, "")
                            for _src_col, dst_col in pairs
                        ):
                            next_append_row += 1
                        matched_records.append((src_row, next_append_row))
                        next_append_row += 1

                    for src_row, dst_row in matched_records:
                        for src_col, dst_col in pairs:
                            source_cell = src_ws.cell(src_row, src_col)
                            destination_cell = dst_ws.cell(dst_row, dst_col)
                            destination_cell.value = source_cell.value
                            destination_cell.number_format = source_cell.number_format
                        populated_rows.append(dst_row)
                    diagnostics["symbols_rows_preserved_by_identity"] = len(used_rows)
                    diagnostics["symbols_rows_appended"] = len(new_records)
                    diagnostics["symbols_rows_cleared"] = len(removed_rows)
                else:
                    max_dst_row = max(dst_ws.max_row, src_ws.max_row)
                    for r in range(dst_start_row, max_dst_row + 1):
                        for _,dc in pairs:
                            dst_ws.cell(r,dc).value=None
                    dst_row = dst_start_row
                    for r in range(src_start_row, src_ws.max_row+1):
                        for sc,dc in pairs:
                            s=src_ws.cell(r,sc); d=dst_ws.cell(dst_row,dc)
                            d.value=s.value; d.number_format=s.number_format
                        if any(dst_ws.cell(dst_row, dc).value not in (None, "") for _sc, dc in pairs):
                            populated_rows.append(dst_row)
                        dst_row += 1
                last_col = max(_instrument_averages_header_map(dst_ws).values())
                dst_ws.auto_filter.ref=(
                    f"A{dst_header_row}:"
                    f"{get_column_letter(last_col)}{max([dst_header_row, *populated_rows])}"
                )
            if SYMBOLS_SHEET in wb.sheetnames and SYMBOLS_SHEET in gen.sheetnames:
                instrument_ws = _symbols_sheet(wb)
                preserved_win_rates: Dict[str, List[Any]] = defaultdict(list)
                if preserve_symbol_win_rates:
                    existing_headers = _instrument_averages_header_map(
                        instrument_ws
                    )
                    existing_symbol_col = existing_headers.get("Symbol")
                    existing_win_rate_col = existing_headers.get("Win Rate %")
                    if existing_symbol_col and existing_win_rate_col:
                        for existing_row in range(
                            _instrument_averages_data_start_row(instrument_ws),
                            instrument_ws.max_row + 1,
                        ):
                            symbol = str(
                                instrument_ws.cell(
                                    existing_row, existing_symbol_col
                                ).value
                                or ""
                            ).strip().upper()
                            if symbol:
                                preserved_win_rates[symbol].append(
                                    instrument_ws.cell(
                                        existing_row, existing_win_rate_col
                                    ).value
                                )
                _copy_instrument_rows_header_aware(_symbols_sheet(gen), instrument_ws)
                if not preserve_existing_layout:
                    _apply_instrument_averages_requested_style(
                        instrument_ws, preserve_layout=True
                    )
                _apply_instrument_averages_profit_loss_formatting(instrument_ws)
                _apply_instrument_averages_semantic_fills(instrument_ws)
                _repair_instrument_timeframe_columns(instrument_ws)
                _populate_symbols_metrics_preserving_layout(instrument_ws, rows, diagnostics)
                if preserved_win_rates:
                    current_headers = _instrument_averages_header_map(
                        instrument_ws
                    )
                    current_symbol_col = current_headers.get("Symbol")
                    current_win_rate_col = current_headers.get("Win Rate %")
                    restored_win_rates = 0
                    if current_symbol_col and current_win_rate_col:
                        remaining = {
                            symbol: list(values)
                            for symbol, values in preserved_win_rates.items()
                        }
                        for current_row in range(
                            _instrument_averages_data_start_row(instrument_ws),
                            instrument_ws.max_row + 1,
                        ):
                            symbol = str(
                                instrument_ws.cell(
                                    current_row, current_symbol_col
                                ).value
                                or ""
                            ).strip().upper()
                            values = remaining.get(symbol) or []
                            if not values:
                                continue
                            prior_value = values.pop(0)
                            instrument_ws.cell(
                                current_row, current_win_rate_col
                            ).value = prior_value
                            restored_win_rates += 1
                    diagnostics[
                        "symbols_win_rates_preserved_for_unchanged_inputs"
                    ] = restored_win_rates
                if not preserve_existing_layout:
                    _ensure_symbols_duration_header_layout(
                        instrument_ws, diagnostics
                    )
                _set_instrument_averages_auto_filter_to_populated_range(instrument_ws)
                _ensure_symbols_freeze_panes(instrument_ws, diagnostics)
            if "P&L Calendar" in wb.sheetnames and "P&L Calendar" in gen.sheetnames:
                cal_ws = wb["P&L Calendar"]
                _write_pnl_calendar_one_line_layout(cal_ws, snapshot, diagnostics)
        finally:
            gen.close()
            tmp.unlink(missing_ok=True)

        trade_log = _get_all_trades_sheet(wb, allow_legacy=False)
        content_after = _workbook_content_snapshot(wb)
        normalized_trade_rows = [
            row
            for row in rows
            if isinstance(row, dict)
            and str(row.get("row_type") or "trade").strip().lower() == "trade"
        ]
        test_only_derived_sections = bool(normalized_trade_rows) and not any(
            not _is_test_trade_value(row.get("is_test_trade"))
            for row in normalized_trade_rows
        )
        _assert_workbook_content_not_wiped(
            content_before,
            content_after,
            migration_performed=bool(diagnostics.get("migrated_trade_log_schema")),
            allow_empty_test_only_derived_sections=test_only_derived_sections,
        )
        _assert_filter_covers_data(trade_log, sheet_name="Trade Log", header_row=TRADE_LOG_FILTER_HEADER_ROW, required_headers=["Open Time", "Close Time", "Row ID"], header_map=_trade_log_header_map(trade_log))
        symbols_ws = _symbols_sheet(wb)
        _ensure_symbols_freeze_panes(symbols_ws, diagnostics)
        _assert_filter_covers_data(
            symbols_ws,
            sheet_name=SYMBOLS_SHEET,
            header_row=_instrument_averages_header_row(symbols_ws),
            required_headers=["Symbol", "Trades"],
            header_map=_instrument_averages_header_map(symbols_ws),
        )

        if preserve_existing_layout:
            for sheet_name, layout in preserved_layout.items():
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                for key in list(ws.column_dimensions):
                    if key not in layout["column_widths"]:
                        del ws.column_dimensions[key]
                for key, width in layout["column_widths"].items():
                    ws.column_dimensions[key].width = width
                for key in list(ws.row_dimensions):
                    if key not in layout["row_heights"]:
                        del ws.row_dimensions[key]
                for key, height in layout["row_heights"].items():
                    ws.row_dimensions[key].height = height
                ws.freeze_panes = layout["freeze_panes"]
                ws.views = deepcopy(layout["views"])
                restore_auto_filter = True
                if sheet_name == TRADE_LOG_SHEET:
                    preserved_filter = layout["auto_filter"]
                    preserved_ref = (
                        preserved_filter.ref
                        if preserved_filter is not None
                        else None
                    )
                    try:
                        min_col, min_row, max_col, max_row = (
                            range_boundaries(preserved_ref)
                        )
                        required_columns = [
                            _trade_log_header_map(ws).get(header)
                            for header in ("Open Time", "Close Time", "Row ID")
                        ]
                        restore_auto_filter = (
                            min_col == 1
                            and min_row == TRADE_LOG_FILTER_HEADER_ROW
                            and max_row >= _trade_log_last_populated_row(ws)
                            and all(
                                column is not None and column <= max_col
                                for column in required_columns
                            )
                        )
                    except Exception:
                        restore_auto_filter = False
                    if not restore_auto_filter:
                        diagnostics[
                            "repaired_trade_log_auto_filter_layout"
                        ] = True
                if restore_auto_filter:
                    ws.auto_filter = deepcopy(layout["auto_filter"])
                ws.conditional_formatting = deepcopy(
                    layout["conditional_formatting"]
                )
                ws.data_validations = deepcopy(
                    layout["data_validations"]
                )
                ws.sheet_format = deepcopy(layout["sheet_format"])
                wanted_merges = set(layout["merged_cells"])
                current_merges = {
                    str(item) for item in ws.merged_cells.ranges
                }
                for extra in sorted(current_merges - wanted_merges):
                    ws.unmerge_cells(extra)
                for missing in sorted(wanted_merges - current_merges):
                    ws.merge_cells(missing)
                stats1_managed_rows: set[int] = set()
                stats1_last_managed_col = 0
                report_managed_rows: set[int] = set()
                if sheet_name == STATS1_SHEET:
                    stats1_managed_rows = set(
                        _stats1_semantic_row_policies(ws)
                    )
                    stats1_last_managed_col = max(
                        _stats1_market_columns(ws).values(),
                        default=4,
                    )
                elif sheet_name == REPORT_YEARLY_SHEET or (
                    sheet_name.isdigit() and int(sheet_name) >= REPORT_START_YEAR
                ):
                    report_managed_rows = {
                        row for row, _spec in _report_spec_rows(ws)
                    }
                    report_managed_rows.update(
                        _report_trailing_blank_rows(ws)
                    )
                for coordinate, presentation in layout[
                    "cell_presentations"
                ].items():
                    cell = ws._cells.get(coordinate)
                    if cell is None:
                        cell = ws.cell(
                            row=coordinate[0],
                            column=coordinate[1],
                        )
                    managed_statistics_cell = False
                    if sheet_name == STATS1_SHEET and coordinate[0] >= 2:
                        managed_statistics_cell = (
                            coordinate[0] in stats1_managed_rows
                            and coordinate[1] <= stats1_last_managed_col
                        )
                    elif (
                        sheet_name == REPORT_YEARLY_SHEET
                        or (sheet_name.isdigit() and int(sheet_name) >= REPORT_START_YEAR)
                    ) and coordinate[0] >= 2:
                        managed_statistics_cell = coordinate[0] in report_managed_rows

                    if managed_statistics_cell:
                        # User-owned presentation survives, while number formats
                        # and semantic fills remain generator-owned.  Restore
                        # component ids directly so duplicate-but-equivalent
                        # legacy fonts/borders cannot be remapped to a different
                        # id on the next openpyxl save.
                        source_style = presentation["style"]
                        managed_style = copy(cell._style or source_style)
                        managed_style.fontId = source_style.fontId
                        managed_style.borderId = source_style.borderId
                        managed_style.alignmentId = source_style.alignmentId
                        managed_style.protectionId = source_style.protectionId
                        cell._style = managed_style
                    else:
                        cell._style = copy(presentation["style"])
                    if (
                        not isinstance(cell, MergedCell)
                        and cell.value == presentation["value"]
                    ):
                        if hasattr(cell, "_hyperlink"):
                            cell._hyperlink = copy(
                                presentation["hyperlink"]
                            )
                        if hasattr(cell, "comment"):
                            cell.comment = copy(
                                presentation["comment"]
                            )
                    elif not isinstance(cell, MergedCell):
                        old_link = presentation["hyperlink"]
                        current_link = getattr(cell, "hyperlink", None)
                        if (
                            old_link is not None
                            and current_link is not None
                            and (
                                getattr(current_link, "target", None),
                                getattr(current_link, "location", None),
                            )
                            == (
                                getattr(old_link, "target", None),
                                getattr(old_link, "location", None),
                            )
                        ):
                            cell._hyperlink = None
                        old_comment = presentation["comment"]
                        current_comment = getattr(cell, "comment", None)
                        if (
                            old_comment is not None
                            and current_comment is not None
                            and (
                                current_comment.text,
                                current_comment.author,
                            )
                            == (
                                old_comment.text,
                                old_comment.author,
                            )
                        ):
                            cell.comment = None
            _sanitize_dashboard_semantic_conditional_formatting(dash)
            _apply_stats1_semantic_formatting(
                dash,
                extended_metrics,
                diagnostics,
            )
            removed_orphan_fills = (
                _remove_orphan_dashboard_semantic_fill_cells(dash)
            )
            if removed_orphan_fills:
                diagnostics[
                    "removed_orphan_stats1_semantic_fill_cells"
                ] = removed_orphan_fills
            _apply_report_workbook_semantic_presentation(wb, diagnostics)
        if preserve_existing_layout:
            trade_log = _get_all_trades_sheet(wb, allow_legacy=False)
            diagnostics.update(
                _restore_trade_log_data_presentations_by_row_id(
                    trade_log,
                    preserved_trade_log_presentations,
                )
            )
            _apply_trade_log_win_loss_row_formatting(trade_log)
            _apply_trade_log_win_loss_direct_row_fills(trade_log)
            row_id_col = _trade_log_header_map(trade_log).get("Row ID")
            current_rows_by_id = {
                str(row.get("id") or stable_row_id(row)).strip(): row
                for row in rows
                if isinstance(row, dict)
            }
            formats_reapplied = 0
            if row_id_col:
                for row_number in range(_trade_log_data_start_row(trade_log), trade_log.max_row + 1):
                    row_id = str(trade_log.cell(row_number, row_id_col).value or "").strip()
                    row_context = current_rows_by_id.get(row_id)
                    if row_context is None:
                        continue
                    formats_reapplied += len(
                        _apply_trade_log_row_number_formats(trade_log, row_number, row_context)
                    )
            diagnostics["reapplied_trade_log_generator_number_formats"] = formats_reapplied
            diagnostics["reapplied_generated_trade_log_formatting_after_layout_restore"] = True
        _assert_filter_covers_data(
            _get_all_trades_sheet(wb, allow_legacy=False),
            sheet_name="Trade Log",
            header_row=TRADE_LOG_FILTER_HEADER_ROW,
            required_headers=["Open Time", "Close Time", "Row ID"],
            header_map=_trade_log_header_map(
                _get_all_trades_sheet(wb, allow_legacy=False)
            ),
        )
        _normalize_trade_log_display_durations(_get_trade_log_sheet(wb, allow_legacy=False))
        _normalize_symbols_performance_formatting(_symbols_sheet(wb))
        chart_bundle = _prepare_recommendation_chart_bundle(rows, path)
        _apply_recommendation_chart_hyperlinks(wb, chart_bundle)
        after = _snapshot_invariants(wb)
        _assert_invariants_unchanged(before, after)
        candidate = path.with_suffix(".update-candidate.tmp.xlsx")
        formula_cache_plan: List[Dict[str, Any]] = []
        if preserve_existing_layout:
            formula_cache_plan, formula_cache_diagnostics = (
                _plan_trade_log_formula_cache_transplant(
                    _get_all_trades_sheet(wb, allow_legacy=False),
                    preserved_trade_log_formula_caches,
                    strict=False,
                )
            )
            diagnostics.update(formula_cache_diagnostics)
        if STATS2_SHEET in wb.sheetnames:
            _repair_stats2_account_balance_formatting(wb[STATS2_SHEET], diagnostics)
        if not preserve_existing_layout:
            _apply_workbook_left_alignment(wb)
            _repair_stats1_child_label_styles(dash, diagnostics)
            _repair_stats2_as_of_datetime_style(wb, diagnostics)
        if not preserve_existing_layout:
            _normalize_generated_statistics_row_heights(wb)
        _normalize_generated_duration_text(wb)
        if preserve_existing_layout:
            _apply_workbook_calculation_signature(
                wb,
                source_calculation_signature,
            )
        diagnostics["cell_style_registry_before_save"] = (
            _reindex_cell_style_registry(wb)
        )
        wb.save(candidate)
        if preserve_existing_layout:
            diagnostics.update(
                _preserve_candidate_ooxml(
                    candidate,
                    TRADE_LOG_SHEET,
                    formula_cache_plan,
                    source_chart_snapshot,
                )
            )
            if (
                _workbook_calculation_signature(candidate)
                != source_calculation_signature
            ):
                raise RuntimeError(
                    "Master Journal calculation settings changed during preservation update."
                )
            candidate_wb = load_workbook(
                candidate,
                data_only=False,
                keep_links=False,
            )
            try:
                candidate_preservation_contract = (
                    _workbook_preservation_contract(candidate_wb)
                )
                _assert_workbook_preservation_contract(
                    source_preservation_contract,
                    candidate_preservation_contract,
                )
                managed_presentation_contract = (
                    _managed_statistics_presentation_contract(candidate_wb)
                )
                diagnostics["preservation_contract_candidate_sha256"] = (
                    _contract_sha256(candidate_preservation_contract)
                )
                diagnostics["managed_statistics_presentation_sha256"] = (
                    _contract_sha256(managed_presentation_contract)
                )
                diagnostics["preservation_contract_verified"] = True
            finally:
                candidate_wb.close()
        if publish_recommendation_assets:
            diagnostics["recommendation_chart_asset_directory"] = str(
                _publish_recommendation_chart_bundle(chart_bundle)
            )
        else:
            diagnostics["recommendation_chart_assets_deferred"] = True
        return {"ok": True, "path": str(path), "candidate_path": str(candidate), "diagnostics": diagnostics}
    finally:
        wb.close()

def _incremental_require_canonical_trade_log(wb) -> Any:
    try:
        ws = _get_trade_log_sheet(wb, allow_legacy=False)
    except Exception as exc:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update requires the canonical Trade Log sheet."
        ) from exc
    expected = {header: col for col, header in enumerate(TRADE_LOG_HEADERS, start=1)}
    if (
        not _trade_log_has_three_row_headers(ws)
        or _trade_log_header_map(ws) != expected
        or ws.max_column != len(TRADE_LOG_HEADERS)
    ):
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update requires the current canonical Trade Log schema."
        )
    return ws


def _incremental_trade_log_row_index(ws) -> Tuple[Dict[str, int], int]:
    headers = _trade_log_header_map(ws)
    row_id_col = headers.get("Row ID")
    if not row_id_col:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update requires a Row ID column."
        )
    row_by_id: Dict[str, int] = {}
    last_populated = _trade_log_data_start_row(ws) - 1
    last_col = len(TRADE_LOG_HEADERS)
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        populated = any(
            ws.cell(row, col).value not in (None, "")
            for col in range(1, last_col + 1)
        )
        if not populated:
            continue
        last_populated = row
        row_id = str(ws.cell(row, row_id_col).value or "").strip()
        if not row_id:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Trade Log row {row} is populated but has no stable Row ID."
            )
        if row_id in row_by_id:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Trade Log contains duplicate Row ID {row_id!r}."
            )
        row_by_id[row_id] = row
    return row_by_id, last_populated


def _incremental_changed_rows_by_id(
    changed_rows: Sequence[Mapping[str, Any]],
    *,
    allow_empty: bool = False,
) -> Tuple[List[str], Dict[str, Dict[str, Any]]]:
    order: List[str] = []
    rows_by_id: Dict[str, Dict[str, Any]] = {}
    for index, raw_row in enumerate(changed_rows):
        if not isinstance(raw_row, Mapping):
            raise IncrementalWorkbookUpdateNotEligible(
                f"Changed workbook row {index} is not a mapping."
            )
        row = dict(raw_row)
        row_id = str(row.get("id") or row.get("__row_id") or "").strip()
        if not row_id:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Changed workbook row {index} has no stable Row ID."
            )
        if row_id in rows_by_id:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Changed rows contain duplicate Row ID {row_id!r}."
            )
        if stable_row_id(row) != row_id:
            raise IncrementalWorkbookUpdateNotEligible(
                f"Changed workbook row {row_id!r} does not have a canonical stable Row ID."
            )
        row["id"] = row_id
        order.append(row_id)
        rows_by_id[row_id] = row
    if not order and not allow_empty:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update requires at least one changed row."
        )
    return order, rows_by_id


def _incremental_assert_no_managed_formulas(ws, row: int) -> None:
    for col, header in enumerate(TRADE_LOG_HEADERS, start=1):
        if _cell_contains_formula(ws.cell(row, col)):
            raise IncrementalWorkbookUpdateNotEligible(
                f"Trade Log row {row} contains a formula in managed column {header!r}.",
                unsafe_full_rebuild=True,
            )


def _incremental_clone_trade_log_row_presentation(ws, source_row: int, target_row: int) -> None:
    if source_row < _trade_log_data_start_row(ws):
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental append requires an adjacent canonical Trade Log data row."
        )
    if any(
        item.min_row <= target_row <= item.max_row
        for item in ws.merged_cells.ranges
    ):
        raise IncrementalWorkbookUpdateNotEligible(
            f"First unused Trade Log row {target_row} contains an authored merge.",
            unsafe_full_rebuild=True,
        )
    for col in range(1, len(TRADE_LOG_HEADERS) + 1):
        target = ws.cell(target_row, col)
        if (
            target.value not in (None, "")
            or target.comment is not None
            or target.hyperlink is not None
        ):
            raise IncrementalWorkbookUpdateNotEligible(
                f"First unused Trade Log row {target_row} contains authored content.",
                unsafe_full_rebuild=True,
            )
        source = ws.cell(source_row, col)
        # A blank row may be intentionally preformatted for the next trade.
        # Retain any target presentation already present; only fill genuinely
        # unformatted cells from the adjacent canonical data-row baseline.
        if target._style is None:
            target._style = copy(source._style)
            target.protection = copy(source.protection)
            target.number_format = source.number_format
        target.comment = None
        target.hyperlink = None
    target_dimension = ws.row_dimensions[target_row]
    if target_dimension.height is None:
        target_dimension.height = ws.row_dimensions[source_row].height


_INCREMENTAL_GENERATED_DV_FORMULAS = {
    '"Yes,No"': {
        "Test", "VWAP", "Round Number", "Spiked Out", "Close Stopout",
        "Near Perfect Entry", "Near Win", "Early Close",
    },
    '"All-Time High,All-Time Low"': {"ATHS/ATLS"},
    '"Market,Limit"': {"Order"},
    '"range,channel"': {"Pattern"},
}


def _incremental_generated_validation_ranges(ws, validation) -> Optional[List[Any]]:
    if str(getattr(validation, "type", "") or "") != "list":
        return None
    formula = str(getattr(validation, "formula1", "") or "")
    expected_headers = _INCREMENTAL_GENERATED_DV_FORMULAS.get(formula)
    if not expected_headers or not bool(getattr(validation, "allow_blank", False)):
        return None
    headers = _trade_log_header_map(ws)
    expected_cols = {headers.get(header) for header in expected_headers}
    expected_cols.discard(None)
    ranges = list(getattr(getattr(validation, "sqref", None), "ranges", []) or [])
    if not ranges:
        return None
    start_row = _trade_log_data_start_row(ws)
    covered_cols: Set[int] = set()
    for item in ranges:
        if item.min_row != start_row:
            return None
        item_cols = set(range(item.min_col, item.max_col + 1))
        if not item_cols or not item_cols.issubset(expected_cols):
            return None
        covered_cols.update(item_cols)
    return ranges if covered_cols else None


def _incremental_extend_generated_validations(ws, last_row: int) -> int:
    extended = 0
    for validation in list(ws.data_validations.dataValidation):
        ranges = _incremental_generated_validation_ranges(ws, validation)
        if not ranges:
            continue
        replacements: List[str] = []
        changed = False
        for item in ranges:
            target_last = max(last_row, int(item.max_row))
            replacements.append(
                f"{get_column_letter(item.min_col)}{item.min_row}:"
                f"{get_column_letter(item.max_col)}{target_last}"
            )
            changed = changed or target_last != int(item.max_row)
        if changed:
            validation.sqref = " ".join(replacements)
            extended += 1
    return extended


def _incremental_formula_value_signature(value: Any) -> Tuple[Any, ...]:
    if isinstance(value, str):
        return ("plain", value)
    attributes = getattr(value, "__dict__", {})
    return (
        "special",
        f"{type(value).__module__}.{type(value).__qualname__}",
        tuple(sorted(
            (str(key), repr(attribute_value))
            for key, attribute_value in attributes.items()
        )),
    )


def _incremental_formula_signature(
    wb,
) -> Tuple[Tuple[str, str, Tuple[Any, ...]], ...]:
    return tuple(sorted(
        (
            ws.title,
            cell.coordinate,
            _incremental_formula_value_signature(cell.value),
        )
        for ws in wb.worksheets
        for cell in ws._cells.values()
        if _cell_contains_formula(cell)
    ))


def _incremental_defined_name_signature(
    wb,
    *,
    excluded_names: Collection[str] = (),
) -> Tuple[Tuple[Any, ...], ...]:
    excluded = {str(name) for name in excluded_names}
    fields = (
        "attr_text", "localSheetId", "hidden", "function", "vbProcedure",
        "xlm", "functionGroupId", "shortcutKey", "publishToServer",
        "workbookParameter", "description",
    )
    return tuple(sorted(
        (
            str(name),
            *(getattr(defined, field, None) for field in fields),
        )
        for name, defined in wb.defined_names.items()
        if str(name) not in excluded
    ))


def _incremental_structural_signature(wb) -> Dict[str, Any]:
    sheet_layouts: Dict[str, Any] = {}
    for ws in wb.worksheets:
        layout = _worksheet_layout_snapshot(ws)
        if ws.title == TRADE_LOG_SHEET:
            layout = dict(layout)
            layout.pop("row_heights", None)
            layout.pop("auto_filter", None)
        sheet_layouts[ws.title] = layout
    return {
        "sheetnames": tuple(wb.sheetnames),
        "active": wb.active.title if wb.worksheets else "",
        "selected": tuple(
            ws.title for ws in wb.worksheets if bool(ws.sheet_view.tabSelected)
        ),
        "sheet_states": tuple((ws.title, ws.sheet_state) for ws in wb.worksheets),
        "layouts": sheet_layouts,
        "chart_counts": tuple((ws.title, len(ws._charts)) for ws in wb.worksheets),
    }


def _incremental_serialized_tree(value: Any) -> str:
    if value is None or not hasattr(value, "to_tree"):
        return ""
    return ET.tostring(value.to_tree(), encoding="unicode")


def _incremental_trade_log_preservation_signature(ws) -> Dict[str, Any]:
    """Capture Trade Log state that a balance-only save must not rewrite."""
    cells: List[Tuple[Any, ...]] = []
    for (row, col), cell in sorted(ws._cells.items()):
        hyperlink = cell.hyperlink
        comment = cell.comment
        cells.append(
            (
                row,
                col,
                cell.value,
                cell.data_type,
                tuple(cell._style) if cell._style is not None else None,
                cell.number_format,
                None
                if hyperlink is None
                else (
                    hyperlink.target,
                    hyperlink.location,
                    hyperlink.display,
                    hyperlink.tooltip,
                ),
                None
                if comment is None
                else (comment.text, comment.author, comment.width, comment.height),
            )
        )
    conditional_formatting = tuple(
        (
            str(getattr(key, "sqref", key)),
            tuple(_incremental_serialized_tree(rule) for rule in rules),
        )
        for key, rules in ws.conditional_formatting._cf_rules.items()
    )
    validations = tuple(
        _incremental_serialized_tree(validation)
        for validation in ws.data_validations.dataValidation
    )
    tables = tuple(
        (str(table.name), _incremental_serialized_tree(table))
        for table in sorted(
            ws.tables.values(), key=lambda value: str(value.name)
        )
    )
    return {
        "cells": tuple(cells),
        "merged": tuple(sorted(str(item) for item in ws.merged_cells.ranges)),
        "row_dimensions": tuple(
            (str(key), _incremental_serialized_tree(dimension))
            for key, dimension in sorted(ws.row_dimensions.items())
        ),
        "column_dimensions": tuple(
            (str(key), _incremental_serialized_tree(dimension))
            for key, dimension in sorted(ws.column_dimensions.items())
        ),
        "auto_filter": _incremental_serialized_tree(ws.auto_filter),
        "conditional_formatting": conditional_formatting,
        "data_validations": validations,
        "tables": tables,
        "views": _incremental_serialized_tree(ws.views),
        "sheet_properties": _incremental_serialized_tree(ws.sheet_properties),
        "sheet_format": _incremental_serialized_tree(ws.sheet_format),
        "protection": _incremental_serialized_tree(ws.protection),
        "print_options": _incremental_serialized_tree(ws.print_options),
        "page_margins": _incremental_serialized_tree(ws.page_margins),
        "page_setup": _incremental_serialized_tree(ws.page_setup),
        "chart_count": len(ws._charts),
        "image_count": len(ws._images),
    }


def _incremental_chart_formula_signature(path: Path) -> Tuple[Tuple[str, str], ...]:
    with zipfile.ZipFile(path) as package:
        return tuple(sorted(
            (part_name, str(formula.text or ""))
            for part_name in package.namelist()
            if part_name.startswith("xl/charts/") and part_name.endswith(".xml")
            for formula in ET.fromstring(package.read(part_name)).findall(
                f".//{{{_DRAWINGML_CHART_NAMESPACE}}}f"
            )
        ))


def _incremental_value_equal(left: Any, right: Any) -> bool:
    if left in (None, "") and right in (None, ""):
        return True
    if isinstance(left, (int, float, Decimal)) and not isinstance(left, bool):
        if isinstance(right, (int, float, Decimal)) and not isinstance(right, bool):
            try:
                return abs(Decimal(str(left)) - Decimal(str(right))) <= Decimal("1e-9")
            except (InvalidOperation, ValueError):
                return False
    if isinstance(left, (datetime, date)) and isinstance(right, (datetime, date)):
        return left == right
    return left == right


def _incremental_account_balance_label(account_balance: Optional[Mapping[str, Any]]) -> str:
    if not isinstance(account_balance, Mapping):
        return ""
    return _canonical_account_label(
        account_balance.get("account_label")
        or account_balance.get("account")
        or account_balance.get("label")
    )


def _incremental_update_account_balance(
    wb,
    account_balance: Mapping[str, Any],
) -> Dict[str, Any]:
    label = _incremental_account_balance_label(account_balance)
    balance = _as_float(account_balance.get("balance"))
    if not label or balance is None:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental account balance update requires an account label and numeric balance."
        )
    if STATS2_SHEET not in wb.sheetnames:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental account balance update requires STATS2."
        )
    ws = wb[STATS2_SHEET]
    try:
        section = _find_anchor_sections(ws, ["Account Balances"])["Account Balances"]
    except Exception as exc:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental account balance update could not locate Account Balances."
        ) from exc
    header_row, col_map = _find_dashboard_table_headers(ws, section)
    if not header_row or not {"account", "balance", "currency"}.issubset(col_map):
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental account balance update requires canonical STATS2 balance headers."
        )
    matching_rows = [
        row
        for row in range(header_row + 1, section["end_row"] + 1)
        if _canonical_account_label(ws.cell(row, col_map["account"]).value) == label
    ]
    if len(matching_rows) != 1:
        raise IncrementalWorkbookUpdateNotEligible(
            f"Incremental account balance update requires exactly one STATS2 row for {label!r}."
        )
    row = matching_rows[0]
    managed_cols = [col_map["balance"], col_map["currency"]]
    if "as_of" in col_map:
        managed_cols.append(col_map["as_of"])
    for col in managed_cols:
        if _cell_contains_formula(ws.cell(row, col)):
            raise IncrementalWorkbookUpdateNotEligible(
                f"STATS2 account balance row {row} contains a formula in a managed cell.",
                unsafe_full_rebuild=True,
            )
    ws.cell(row, col_map["balance"]).value = balance
    currency = str(account_balance.get("currency") or "").strip().upper()
    if currency:
        ws.cell(row, col_map["currency"]).value = currency
        ws.cell(row, col_map["balance"]).number_format = _currency_number_format(
            currency,
            force_decimals=10 if _is_crypto_currency(currency) else 2,
        )
    as_of = account_balance.get("as_of")
    if as_of not in (None, "") and "as_of" in col_map:
        ws.cell(row, col_map["as_of"]).value = str(as_of)
    source = account_balance.get("balance_source") or account_balance.get("source")
    _write_account_balance_source_metadata(wb, label, source, as_of)
    return {
        "label": label,
        "row": row,
        "balance": balance,
        "currency": currency,
        "as_of": "" if as_of in (None, "") else str(as_of),
        "source": str(source or "").strip(),
    }


def _incremental_validate_account_balance(
    wb,
    expected: Mapping[str, Any],
) -> Dict[str, Any]:
    balances = {
        _canonical_account_label(item.get("account_label") or item.get("account")): item
        for item in _read_stats2_account_balances(wb)
        if isinstance(item, dict)
    }
    actual = balances.get(str(expected.get("label") or ""))
    if not actual or not _incremental_value_equal(actual.get("balance"), expected.get("balance")):
        raise RuntimeError("Incremental workbook account balance verification failed.")
    expected_currency = str(expected.get("currency") or "")
    if expected_currency and str(actual.get("currency") or "") != expected_currency:
        raise RuntimeError("Incremental workbook account currency verification failed.")
    metadata = _read_account_balance_source_metadata(wb, expected.get("label"))
    expected_as_of = str(expected.get("as_of") or "")
    metadata_as_of = str(metadata.get("timeline_as_of") or "")
    if expected_as_of and _account_balance_timeline_as_of(metadata_as_of) != _account_balance_timeline_as_of(expected_as_of):
        raise RuntimeError("Incremental workbook account balance timestamp verification failed.")
    if str(metadata.get("source") or "") != str(expected.get("source") or ""):
        raise RuntimeError("Incremental workbook account balance source verification failed.")
    return {
        "account_label": str(actual.get("account_label") or actual.get("account") or ""),
        "row": int(expected.get("row") or 0),
        "balance": actual.get("balance"),
        "currency": str(actual.get("currency") or ""),
        "as_of": expected_as_of,
        "source": str(metadata.get("source") or ""),
        "provenance_defined_name": (
            _account_balance_source_defined_name(expected.get("label"))
            if metadata
            else ""
        ),
    }


def _incremental_validate_trade_log_ranges(ws, last_row: int) -> None:
    expected_filter = (
        f"A{TRADE_LOG_FILTER_HEADER_ROW}:"
        f"{get_column_letter(len(TRADE_LOG_HEADERS))}{last_row}"
    )
    if not ws.auto_filter or str(ws.auto_filter.ref or "") != expected_filter:
        raise RuntimeError("Incremental workbook Trade Log filter verification failed.")
    expected_cf = (
        f"A{TRADE_LOG_DATA_START_ROW}:"
        f"{get_column_letter(len(TRADE_LOG_HEADERS))}{last_row}"
    )
    generated_ranges: List[str] = []
    for key, rules in ws.conditional_formatting._cf_rules.items():
        sqref = str(getattr(key, "sqref", key))
        for rule in rules:
            if _is_generated_trade_log_win_loss_row_rule(rule, sqref):
                generated_ranges.append(sqref)
    if generated_ranges.count(expected_cf) != 2 or any(
        item != expected_cf for item in generated_ranges
    ):
        raise RuntimeError("Incremental workbook Trade Log conditional formatting verification failed.")
    for validation in ws.data_validations.dataValidation:
        ranges = _incremental_generated_validation_ranges(ws, validation)
        if ranges and any(int(item.max_row) < last_row for item in ranges):
            raise RuntimeError("Incremental workbook Trade Log validation range verification failed.")


def _incremental_validate_changed_rows(
    wb,
    changed_order: Sequence[str],
    rows_by_id: Mapping[str, Mapping[str, Any]],
    expected_values: Mapping[str, Mapping[str, Any]],
    expected_formats: Mapping[str, Mapping[str, str]],
    expected_survivor_row_ids: Optional[Collection[str]],
) -> Tuple[Dict[str, int], int]:
    ws = _incremental_require_canonical_trade_log(wb)
    row_index, last_row = _incremental_trade_log_row_index(ws)
    for row_id in changed_order:
        row_number = row_index.get(row_id)
        if row_number is None:
            raise RuntimeError(f"Incremental workbook verification lost Row ID {row_id!r}.")
        projection = expected_values[row_id]
        for header, col in _trade_log_header_map(ws).items():
            if not _incremental_value_equal(ws.cell(row_number, col).value, projection.get(header, "")):
                raise RuntimeError(
                    f"Incremental workbook value verification failed for {row_id!r} column {header!r}."
                )
        for header, number_format in expected_formats[row_id].items():
            col = _trade_log_header_map(ws).get(header)
            if col and str(ws.cell(row_number, col).number_format or "") != str(number_format):
                raise RuntimeError(
                    f"Incremental workbook number-format verification failed for {row_id!r} column {header!r}."
                )
    if expected_survivor_row_ids is not None:
        expected = {str(value or "").strip() for value in expected_survivor_row_ids}
        expected.discard("")
        missing = sorted(expected - set(row_index))
        if missing:
            raise IncrementalWorkbookUpdateNotEligible(
                "Incremental workbook survivor verification failed: " + ", ".join(missing[:20])
            )
    _incremental_validate_trade_log_ranges(ws, last_row)
    return row_index, last_row


def update_master_journal_workbook_incremental(
    path: Path,
    changed_rows: Sequence[Mapping[str, Any]],
    *,
    account_balance: Optional[Mapping[str, Any]] = None,
    expected_survivor_row_ids: Optional[Collection[str]] = None,
) -> Dict[str, Any]:
    """Atomically update only selected canonical Trade Log rows and one balance."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Master Journal workbook not found: {path}")
    source_calculation_signature = _workbook_calculation_signature(path)
    source_chart_snapshot = _snapshot_chart_ooxml(path)
    external_references = _external_workbook_formula_references(path)
    if external_references:
        raise IncrementalWorkbookUpdateNotEligible(
            "Incremental workbook update cannot safely preserve external workbook references.",
            unsafe_full_rebuild=True,
        )
    changed_order, rows_by_id = _incremental_changed_rows_by_id(
        changed_rows,
        allow_empty=isinstance(account_balance, Mapping),
    )
    candidate = path.with_name(
        f".{path.stem}.incremental-{uuid4().hex}.tmp{path.suffix}"
    )
    wb = load_workbook(path, data_only=False, keep_links=False)
    balance_name = ""
    if isinstance(account_balance, Mapping):
        balance_label = _incremental_account_balance_label(account_balance)
        if balance_label:
            balance_name = _account_balance_source_defined_name(balance_label)
    excluded_names = {balance_name} if balance_name else set()
    before_formula_signature = _incremental_formula_signature(wb)
    before_name_signature = _incremental_defined_name_signature(
        wb, excluded_names=excluded_names
    )
    before_structural_signature = _incremental_structural_signature(wb)
    before_chart_signature = _incremental_chart_formula_signature(path)
    inserted_ids: List[str] = []
    updated_ids: List[str] = []
    affected_rows: List[int] = []
    expected_values: Dict[str, Dict[str, Any]] = {}
    expected_formats: Dict[str, Dict[str, str]] = {}
    balance_expected: Optional[Dict[str, Any]] = None
    verified_account_balance: Optional[Dict[str, Any]] = None
    trade_log_preservation_signature: Optional[Dict[str, Any]] = None
    source_formula_caches: Dict[Tuple[str, str], Dict[str, Any]] = {}
    formula_cache_plan: List[Dict[str, Any]] = []
    formula_cache_diagnostics: Dict[str, Any] = {}
    formula_dependency_diagnostics: Dict[str, int] = {}
    validation_extended = 0
    replaced = False
    try:
        ws = _incremental_require_canonical_trade_log(wb)
        source_formula_records = _trade_log_ooxml_formula_records(
            path, ws.title
        )
        representation_issue = _trade_log_formula_representation_issue(
            ws, source_formula_records
        )
        if representation_issue:
            raise IncrementalWorkbookUpdateNotEligible(
                "Incremental workbook update cannot safely preserve "
                f"{representation_issue}.",
                unsafe_full_rebuild=True,
            )
        stats2_formula_cells = _managed_stats2_account_balance_formula_cells(wb)
        if stats2_formula_cells:
            raise IncrementalWorkbookUpdateNotEligible(
                "Incremental workbook update cannot safely preserve a formula in "
                "a managed cell of STATS2 Account Balances: "
                + ", ".join(stats2_formula_cells[:20])
                + ".",
                unsafe_full_rebuild=True,
            )
        row_index, last_populated = _incremental_trade_log_row_index(ws)
        source_formula_caches = _snapshot_trade_log_formula_caches(path, ws)
        if not changed_order:
            trade_log_preservation_signature = (
                _incremental_trade_log_preservation_signature(ws)
            )
        next_append_row = max(_trade_log_data_start_row(ws), last_populated + 1)
        # Reserve and validate every append row before mutating the adjacent
        # populated row. This keeps the canonical presentation baseline stable
        # when one import appends several rows or also updates the former last row.
        for row_id in changed_order:
            row_number = row_index.get(row_id)
            if row_number is None:
                _incremental_clone_trade_log_row_presentation(
                    ws, last_populated, next_append_row
                )
                row_number = next_append_row
                next_append_row += 1
                row_index[row_id] = row_number
                inserted_ids.append(row_id)
        planned_affected_rows = [row_index[row_id] for row_id in changed_order]
        formula_dependency_diagnostics = (
            _incremental_assert_trade_log_formula_dependencies_safe(
                ws, planned_affected_rows
            )
        )
        formula_dependency_diagnostics[
            "trade_log_ooxml_formula_cells_checked"
        ] = len(source_formula_records)
        for row_id in changed_order:
            row = rows_by_id[row_id]
            row_number = row_index[row_id]
            if row_id not in inserted_ids:
                updated_ids.append(row_id)
            _incremental_assert_no_managed_formulas(ws, row_number)
            explicit_balance = _as_float(row.get("balance_after_trade"))
            resolved_balance = (
                explicit_balance
                if explicit_balance is not None
                else _resolve_balance_after(dict(row))
            )
            projection = _trade_log_row_values(
                row,
                resolved_balance=resolved_balance,
                recommendations_by_symbol=None,
            )
            if str(projection.get("Row ID") or "").strip() != row_id:
                raise IncrementalWorkbookUpdateNotEligible(
                    f"Incremental serializer changed Row ID {row_id!r}."
                )
            for header, col in _trade_log_header_map(ws).items():
                ws.cell(row_number, col).value = projection.get(header, "")
            expected_values[row_id] = dict(projection)
            expected_formats[row_id] = _apply_trade_log_row_number_formats(
                ws, row_number, row
            )
            affected_rows.append(row_number)
        if changed_order:
            _apply_trade_log_win_loss_direct_row_fills_for_rows(ws, affected_rows)
            _apply_trade_log_win_loss_row_formatting(ws)
            last_populated = _trade_log_last_populated_row(ws)
            ws.auto_filter.ref = (
                f"A{TRADE_LOG_FILTER_HEADER_ROW}:"
                f"{get_column_letter(len(TRADE_LOG_HEADERS))}{last_populated}"
            )
            validation_extended = _incremental_extend_generated_validations(
                ws, last_populated
            )
        if expected_survivor_row_ids is not None:
            expected_survivors = {
                str(value or "").strip() for value in expected_survivor_row_ids
            }
            expected_survivors.discard("")
            missing = sorted(expected_survivors - set(row_index))
            if missing:
                raise IncrementalWorkbookUpdateNotEligible(
                    "Incremental workbook survivor verification failed: "
                    + ", ".join(missing[:20])
                )
        if isinstance(account_balance, Mapping):
            balance_expected = _incremental_update_account_balance(wb, account_balance)
        formula_cache_plan, formula_cache_diagnostics = (
            _plan_trade_log_formula_cache_transplant(
                ws,
                source_formula_caches,
                strict=True,
            )
        )
        _apply_workbook_calculation_signature(
            wb,
            source_calculation_signature,
        )
        wb.save(candidate)
        wb.close()
        wb = None
        formula_cache_diagnostics.update(
            _preserve_candidate_ooxml(
                candidate,
                TRADE_LOG_SHEET,
                formula_cache_plan,
                source_chart_snapshot,
            )
        )

        verification_wb = load_workbook(candidate, data_only=False, keep_links=False)
        try:
            if changed_order:
                verified_index, verified_last_row = _incremental_validate_changed_rows(
                    verification_wb,
                    changed_order,
                    rows_by_id,
                    expected_values,
                    expected_formats,
                    expected_survivor_row_ids,
                )
            else:
                verified_ws = _incremental_require_canonical_trade_log(
                    verification_wb
                )
                verified_index, verified_last_row = _incremental_trade_log_row_index(
                    verified_ws
                )
                if expected_survivor_row_ids is not None:
                    expected_survivors = {
                        str(value or "").strip()
                        for value in expected_survivor_row_ids
                    }
                    expected_survivors.discard("")
                    missing = sorted(expected_survivors - set(verified_index))
                    if missing:
                        raise IncrementalWorkbookUpdateNotEligible(
                            "Incremental workbook survivor verification failed: "
                            + ", ".join(missing[:20])
                        )
                if (
                    _incremental_trade_log_preservation_signature(verified_ws)
                    != trade_log_preservation_signature
                ):
                    raise RuntimeError(
                        "Incremental balance-only Trade Log preservation verification failed."
                    )
            if balance_expected is not None:
                verified_account_balance = _incremental_validate_account_balance(
                    verification_wb, balance_expected
                )
            if _incremental_formula_signature(verification_wb) != before_formula_signature:
                raise RuntimeError("Incremental workbook formula preservation verification failed.")
            if _incremental_defined_name_signature(
                verification_wb, excluded_names=excluded_names
            ) != before_name_signature:
                raise RuntimeError("Incremental workbook defined-name preservation verification failed.")
            if _incremental_structural_signature(verification_wb) != before_structural_signature:
                raise RuntimeError("Incremental workbook structural preservation verification failed.")
        finally:
            verification_wb.close()
        if _incremental_chart_formula_signature(candidate) != before_chart_signature:
            raise RuntimeError("Incremental workbook chart preservation verification failed.")
        if (
            _workbook_calculation_signature(candidate)
            != source_calculation_signature
        ):
            raise RuntimeError(
                "Incremental workbook calculation settings changed."
            )
        with candidate.open("r+b") as handle:
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(candidate, path)
        replaced = True
        return {
            "ok": True,
            "path": str(path),
            "incremental": True,
            "inserted_row_ids": inserted_ids,
            "updated_row_ids": updated_ids,
            "affected_row_ids": list(changed_order),
            "affected_row_numbers": affected_rows,
            "account_balance_updated": balance_expected is not None,
            "diagnostics": {
                "trade_log_last_populated_row": verified_last_row,
                "trade_log_row_ids_verified": len(verified_index),
                "generated_data_validations_extended": validation_extended,
                "candidate_verified_before_replace": True,
                "verified_account_balance": verified_account_balance,
                **formula_dependency_diagnostics,
                **formula_cache_diagnostics,
            },
        }
    finally:
        if wb is not None:
            wb.close()
        if not replaced and candidate.exists():
            candidate.unlink(missing_ok=True)


def refresh_master_journal_derived_sheets(path: Path, snapshot: Dict[str, Any]) -> Dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Master Journal workbook not found: {path}")
    result = update_master_journal_workbook_data_only(
        path,
        snapshot,
        preserve_existing_layout=True,
    )
    if not result.get("ok"):
        return result
    candidate = Path(str(result.get("candidate_path") or ""))
    if not candidate.exists():
        return {"ok": False, "error": "derived_sheet_candidate_missing", "path": str(path)}
    candidate.replace(path)
    return {"ok": True, "path": str(path), "diagnostics": result.get("diagnostics") or {}}
