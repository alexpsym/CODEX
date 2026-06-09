from __future__ import annotations

import argparse
import json
import math
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.master_journal_workbook import (
    MOVE_TO_FIELD_MAP,
    TRADE_NUMBER_HEADER,
    _as_datetime,
    _as_float,
    _canonical_journal_timeframe,
    _duration_ddhhmmss_cell_to_seconds,
    _ensure_trade_log_schema,
    _fmt_duration_full,
    _parse_duration_text,
    _repair_trade_log_move_to_durations,
    _trade_log_data_start_row,
    _trade_log_header_map,
)


LEGACY_SHEET = "TRADE LOG"


@dataclass
class LegacyTrade:
    row: int
    trade_number: str
    symbol: str
    side: str
    open_time: Optional[datetime]
    close_time: Optional[datetime]
    entry: Optional[float]
    exit: Optional[float]
    stop: Optional[float]
    target: Optional[float]
    result_pct: Optional[float]
    r_multiple: Optional[float]
    move_be: Dict[str, Any]
    move_profit: Dict[str, Any]
    manual: Dict[str, Any]
    uncertain_breakeven: Dict[str, Any]


@dataclass
class RepoTrade:
    row: int
    symbol: str
    side: str
    open_time: Optional[datetime]
    close_time: Optional[datetime]
    entry: Optional[float]
    exit: Optional[float]
    stop: Optional[float]
    target: Optional[float]
    result_pct: Optional[float]
    r_multiple: Optional[float]


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def _cell(ws, row: int, col: str, ignored: List[Dict[str, Any]]) -> Any:
    cell = ws.cell(row, _col(col))
    value = cell.value
    if getattr(cell, "data_type", None) == "e" or (isinstance(value, str) and value.startswith("#")):
        ignored.append({"row": row, "column": col, "value": str(value)})
        return None
    return value


def _safe_float(value: Any) -> Optional[float]:
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        return None
    return float(number)


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)):
        try:
            return from_excel(float(value)).replace(tzinfo=None)
        except Exception:
            return None
    return _as_datetime(value)


def _parse_duration(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, timedelta):
        return max(0.0, value.total_seconds())
    if isinstance(value, time):
        return float(value.hour * 3600 + value.minute * 60 + value.second)
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 < number < 1:
            return number * 86400.0
        parsed = _duration_ddhhmmss_cell_to_seconds(number)
        return float(parsed) if parsed is not None else (number if number >= 0 else None)
    parsed = _parse_duration_text(value)
    return float(parsed) if parsed is not None else None


def _canon_symbol(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _canon_side(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in {"LONG", "BUY"} or text.startswith("BUY"):
        return "BUY"
    if text in {"SHORT", "SELL"} or text.startswith("SELL"):
        return "SELL"
    return text


def _yes_no(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"yes", "y", "true", "1", "x"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return ""


def _order(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "market" in text:
        return "Market"
    if "limit" in text:
        return "Limit"
    return ""


def _aths_atls(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ").replace("-", " ")
    if text in {"aths", "ath", "all time high", "all time highs"}:
        return "All-Time High"
    if text in {"atls", "atl", "all time low", "all time lows"}:
        return "All-Time Low"
    return ""


def _pattern(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "channel" in text:
        return "channel"
    if "range" in text:
        return "range"
    return ""


def _distance_fraction(value: Any) -> Optional[float]:
    if value in (None, ""):
        return None
    text = str(value).strip()
    number = _safe_float(text[:-1] if text.endswith("%") else value)
    if number is None:
        return None
    if text.endswith("%") or abs(number) > 1:
        return number / 100.0
    return number


def parse_legacy_trades(path: Path) -> Tuple[List[LegacyTrade], List[Dict[str, Any]], int]:
    wb = load_workbook(path, data_only=True)
    ignored: List[Dict[str, Any]] = []
    fallback_used = 0
    try:
        ws = wb[LEGACY_SHEET]
        trades: List[LegacyTrade] = []
        for row in range(1, ws.max_row + 1):
            raw_number = _cell(ws, row, "A", ignored)
            if not isinstance(raw_number, (int, float)) or int(raw_number) != raw_number:
                continue
            symbol = _canon_symbol(_cell(ws, row, "B", ignored))
            side = _canon_side(_cell(ws, row, "T", ignored))
            if not symbol or side not in {"BUY", "SELL"}:
                continue
            primary_profit = {
                "move_to_profit_time": _parse_datetime(_cell(ws, row, "BO", ignored)),
                "move_to_profit_duration": _parse_duration(_cell(ws, row, "BP", ignored)),
                "move_to_profit_trigger_price": _safe_float(_cell(ws, row, "BQ", ignored)),
                "move_to_profit_distance_from_entry_pct": _distance_fraction(_cell(ws, row, "BR", ignored)),
                "move_to_profit_distance_from_exit_pct": _distance_fraction(_cell(ws, row, "BS", ignored)),
            }
            fallback_profit = {
                "move_to_profit_trigger_price": _safe_float(_cell(ws, row, "BT", ignored)),
                "move_to_profit_distance_from_entry_pct": _distance_fraction(_cell(ws, row, "BU", ignored)),
                "move_to_profit_distance_from_exit_pct": _distance_fraction(_cell(ws, row, "BV", ignored)),
            }
            if not any(value not in (None, "") for value in primary_profit.values()) and any(
                value not in (None, "") for value in fallback_profit.values()
            ):
                fallback_used += 1
            for key, value in fallback_profit.items():
                if primary_profit.get(key) in (None, "") and value not in (None, ""):
                    primary_profit[key] = value
            manual = {
                "Pattern": _pattern(_cell(ws, row, "BB", ignored)),
                "EMA": _yes_no(_cell(ws, row, "AY", ignored)) or str(_cell(ws, row, "AY", ignored) or "").strip(),
                "ATHS/ATLS": _aths_atls(_cell(ws, row, "AW", ignored)),
                "Order": _order(_cell(ws, row, "AV", ignored)),
                "Round Number": _yes_no(_cell(ws, row, "AR", ignored)),
                "Spiked Out": _yes_no(_cell(ws, row, "AQ", ignored)),
                "Close Stopout": _yes_no(_cell(ws, row, "AP", ignored)),
                "Near Perfect Entry": _yes_no(_cell(ws, row, "AO", ignored)),
                "Near Win": _yes_no(_cell(ws, row, "AN", ignored)),
                "Early Close": _yes_no(_cell(ws, row, "AM", ignored)),
                "Timeframe": _canonical_journal_timeframe(_cell(ws, row, "AU", ignored)),
            }
            manual = {key: value for key, value in manual.items() if value not in (None, "")}
            uncertain = {
                col: _cell(ws, row, col, ignored)
                for col in ("AX", "BA", "BC", "BD")
                if _cell(ws, row, col, ignored) not in (None, "")
            }
            trades.append(LegacyTrade(
                row=row,
                trade_number=f"F{int(raw_number)}",
                symbol=symbol,
                side=side,
                open_time=_parse_datetime(_cell(ws, row, "C", ignored)),
                close_time=_parse_datetime(_cell(ws, row, "D", ignored)),
                entry=_safe_float(_cell(ws, row, "R", ignored)),
                exit=_safe_float(_cell(ws, row, "S", ignored)),
                stop=_safe_float(_cell(ws, row, "X", ignored)),
                target=_safe_float(_cell(ws, row, "AB", ignored)),
                result_pct=_safe_float(_cell(ws, row, "H", ignored)),
                r_multiple=_safe_float(_cell(ws, row, "J", ignored)),
                move_be={key: value for key, value in {
                    "move_to_break_even_time": _parse_datetime(_cell(ws, row, "BJ", ignored)),
                    "move_to_break_even_duration": _parse_duration(_cell(ws, row, "BK", ignored)),
                    "move_to_break_even_trigger_price": _safe_float(_cell(ws, row, "BL", ignored)),
                    "move_to_break_even_distance_from_entry_pct": _distance_fraction(_cell(ws, row, "BM", ignored)),
                    "move_to_break_even_distance_from_exit_pct": _distance_fraction(_cell(ws, row, "BN", ignored)),
                }.items() if value not in (None, "")},
                move_profit={key: value for key, value in primary_profit.items() if value not in (None, "")},
                manual=manual,
                uncertain_breakeven=uncertain,
            ))
        return trades, ignored, fallback_used
    finally:
        wb.close()


def parse_repo_trades(ws) -> Tuple[List[RepoTrade], Dict[str, int]]:
    _ensure_trade_log_schema(ws)
    headers = _trade_log_header_map(ws)
    trades: List[RepoTrade] = []
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        row_type = str(ws.cell(row, headers.get("Row Type", 0)).value or "trade").strip().lower()
        if row_type != "trade":
            continue
        account = str(ws.cell(row, headers.get("Account", 0)).value or "").upper()
        symbol = _canon_symbol(ws.cell(row, headers.get("Symbol", 0)).value)
        if not any(token in account for token in ("OANDA", "PEPPERSTONE", "FOREX", " FX")) and len(symbol) != 6:
            continue
        trades.append(RepoTrade(
            row=row,
            symbol=symbol,
            side=_canon_side(ws.cell(row, headers.get("Side", 0)).value),
            open_time=_as_datetime(ws.cell(row, headers.get("Open Time", 0)).value),
            close_time=_as_datetime(ws.cell(row, headers.get("Close Time", 0)).value),
            entry=_safe_float(ws.cell(row, headers.get("Entry Price", 0)).value),
            exit=_safe_float(ws.cell(row, headers.get("Exit Price", 0)).value),
            stop=_safe_float(ws.cell(row, headers.get("Stop Loss Price", 0)).value),
            target=_safe_float(ws.cell(row, headers.get("Target Price", 0)).value),
            result_pct=(
                _safe_float(ws.cell(row, headers.get("Profit %", 0)).value) * 100.0
                if _safe_float(ws.cell(row, headers.get("Profit %", 0)).value) is not None else None
            ),
            r_multiple=_safe_float(ws.cell(row, headers.get("R-Multiple", 0)).value),
        ))
    return trades, headers


def _time_delta(legacy: LegacyTrade, repo: RepoTrade) -> Optional[float]:
    if not all((legacy.open_time, legacy.close_time, repo.open_time, repo.close_time)):
        return None
    return abs((repo.open_time - legacy.open_time).total_seconds()) + abs((repo.close_time - legacy.close_time).total_seconds())


def _price_score(legacy: LegacyTrade, repo: RepoTrade) -> Tuple[int, float, float]:
    matches = 0
    difference = 0.0
    for left, right in ((legacy.entry, repo.entry), (legacy.exit, repo.exit), (legacy.stop, repo.stop), (legacy.target, repo.target)):
        if left is None or right is None:
            continue
        tolerance = max(1e-8, abs(left) * 0.0002)
        delta = abs(left - right)
        difference += delta / max(abs(left), 1e-8)
        if delta <= tolerance:
            matches += 1
    result_diff = abs((legacy.result_pct or 0.0) - (repo.result_pct or 0.0)) if legacy.result_pct is not None and repo.result_pct is not None else 999999.0
    return matches, difference, result_diff


def _load_overrides(path: Optional[Path]) -> Dict[str, int]:
    if not path:
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Override mapping must be a JSON object of trade number or legacy row to repo row.")
    return {str(key): int(value) for key, value in payload.items()}


def match_trades(
    legacy_trades: List[LegacyTrade],
    repo_trades: List[RepoTrade],
    overrides: Optional[Dict[str, int]] = None,
) -> Tuple[Dict[int, Tuple[LegacyTrade, RepoTrade]], List[LegacyTrade], List[Dict[str, Any]]]:
    overrides = overrides or {}
    repo_by_row = {trade.row: trade for trade in repo_trades}
    matches: Dict[int, Tuple[LegacyTrade, RepoTrade]] = {}
    unmatched: List[LegacyTrade] = []
    ambiguous: List[Dict[str, Any]] = []
    for legacy in legacy_trades:
        override_row = overrides.get(legacy.trade_number, overrides.get(str(legacy.row)))
        if override_row is not None:
            repo = repo_by_row.get(override_row)
            if repo is None:
                unmatched.append(legacy)
            else:
                matches[legacy.row] = (legacy, repo)
            continue
        candidates: List[Tuple[float, Tuple[int, float, float], RepoTrade]] = []
        for repo in repo_trades:
            if legacy.symbol != repo.symbol or legacy.side != repo.side:
                continue
            delta = _time_delta(legacy, repo)
            if delta is not None and delta <= 600:
                candidates.append((delta, _price_score(legacy, repo), repo))
        if not candidates:
            unmatched.append(legacy)
            continue
        candidates.sort(key=lambda item: (item[0], -item[1][0], item[1][1], item[1][2], item[2].row))
        best_delta = candidates[0][0]
        same_execution = [candidate for candidate in candidates if candidate[0] == best_delta]
        if len(same_execution) > 1:
            ambiguous.append({
                "legacy_row": legacy.row,
                "trade_number": legacy.trade_number,
                "repo_rows": [candidate[2].row for candidate in same_execution],
            })
            continue
        matches[legacy.row] = (legacy, candidates[0][2])
    return matches, unmatched, ambiguous


def _write_fields(ws, headers: Dict[str, int], row: int, values: Dict[str, Any], *, overwrite: bool) -> Tuple[int, int]:
    field_to_header = {field: header for header, field in MOVE_TO_FIELD_MAP.items()}
    written = 0
    skipped = 0
    for field, value in values.items():
        header = field_to_header.get(field)
        col = headers.get(header or "")
        if not col or value in (None, ""):
            continue
        cell = ws.cell(row, col)
        if cell.value not in (None, "") and not overwrite:
            skipped += 1
            continue
        if field.endswith("_duration"):
            cell.value = _fmt_duration_full(value)
            cell.number_format = r'00\:00\:00\:00'
        elif field.endswith("_pct"):
            cell.value = value
            cell.number_format = "0.00%"
        else:
            cell.value = value
            if field.endswith("_time"):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        written += 1
    return written, skipped


def run_backfill(
    journal_path: Path,
    legacy_path: Path,
    *,
    charts_zip: Optional[Path] = None,
    apply_changes: bool = False,
    override_map: Optional[Path] = None,
    overwrite: bool = False,
) -> Dict[str, Any]:
    if charts_zip and not charts_zip.exists():
        raise FileNotFoundError(charts_zip)
    legacy_trades, ignored, fallback_used = parse_legacy_trades(legacy_path)
    overrides = _load_overrides(override_map)
    wb = load_workbook(journal_path)
    try:
        ws = wb["Trade Log"]
        repo_trades, headers = parse_repo_trades(ws)
        matches, unmatched, ambiguous = match_trades(legacy_trades, repo_trades, overrides)
        target_counts = Counter(repo.row for _legacy, repo in matches.values())
        duplicate_targets = sorted(row for row, count in target_counts.items() if count > 1)
        summary: Dict[str, Any] = {
            "legacy_trades_parsed": len(legacy_trades),
            "matches": len(matches),
            "unmatched": len(unmatched),
            "ambiguous": len(ambiguous),
            "duplicated_repo_targets": len(duplicate_targets),
            "ignored_invalid_cells": len(ignored),
            "move_profit_fallback_used": fallback_used,
            "manual_override_count": len(overrides),
            "uncertain_breakeven_candidates": sum(bool(trade.uncertain_breakeven) for trade in legacy_trades),
            "sample_mappings": [
                {"trade_number": legacy.trade_number, "legacy_row": legacy.row, "repo_row": repo.row, "symbol": legacy.symbol}
                for legacy, repo in list(matches.values())[:20]
            ],
            "unmatched_legacy_rows": [trade.row for trade in unmatched[:20]],
            "ambiguous_matches": ambiguous[:20],
            "duplicated_repo_target_rows": duplicate_targets[:20],
            "applied": False,
        }
        if apply_changes and (unmatched or ambiguous or duplicate_targets):
            summary["error"] = "unsafe_match_result"
            return summary
        if not apply_changes:
            return summary

        written = 0
        skipped_nonblank = 0
        trade_col = headers[TRADE_NUMBER_HEADER]
        for legacy, repo in matches.values():
            trade_cell = ws.cell(repo.row, trade_col)
            if trade_cell.value in (None, "") or overwrite:
                trade_cell.value = legacy.trade_number
                trade_cell.number_format = "@"
                written += 1
            else:
                skipped_nonblank += 1
            for values in (legacy.move_be, legacy.move_profit):
                count, skipped = _write_fields(ws, headers, repo.row, values, overwrite=overwrite)
                written += count
                skipped_nonblank += skipped
            for header, value in legacy.manual.items():
                col = headers.get(header)
                if not col or value in (None, ""):
                    continue
                cell = ws.cell(repo.row, col)
                if cell.value not in (None, "") and not overwrite:
                    skipped_nonblank += 1
                    continue
                cell.value = value
                written += 1
        _repair_trade_log_move_to_durations(ws, summary)
        wb.save(journal_path)
        summary.update({"applied": True, "cells_written": written, "skipped_nonblank_cells": skipped_nonblank})
        return summary
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Backfill legacy Forex trade metadata into Trading Journal.xlsx.")
    parser.add_argument("--journal", required=True, type=Path)
    parser.add_argument("--legacy", required=True, type=Path)
    parser.add_argument("--charts-zip", type=Path)
    parser.add_argument("--override-map", type=Path)
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--apply", action="store_true", help="Write changes. Omit for dry-run.")
    args = parser.parse_args()
    summary = run_backfill(
        args.journal,
        args.legacy,
        charts_zip=args.charts_zip,
        apply_changes=args.apply,
        override_map=args.override_map,
        overwrite=args.overwrite,
    )
    print(json.dumps(summary, indent=2, sort_keys=True, default=str))
    return 2 if summary.get("error") else 0


if __name__ == "__main__":
    raise SystemExit(main())
