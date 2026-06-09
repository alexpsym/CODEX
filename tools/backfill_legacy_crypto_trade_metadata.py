from __future__ import annotations

import argparse
import json
import math
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.datetime import from_excel

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.master_journal_workbook import (
    MOVE_TO_FIELD_MAP,
    TRADE_LOG_DATA_START_ROW,
    TRADE_NUMBER_HEADER,
    _as_datetime,
    _as_float,
    _canonical_journal_timeframe,
    _duration_ddhhmmss_cell_to_seconds,
    _ensure_trade_log_schema,
    _fmt_duration_full,
    _parse_duration_text,
    _repair_trade_log_move_to_durations,
    _trade_log_data_start_row,
    _trade_log_header_map,
)


LEGACY_SHEET = "TRADE LOG"
CRYPTO_ACCOUNTS = ("BINANCE", "BYBIT", "COINSPOT")
CRYPTO_SYMBOL_TOKENS = ("USDT", "USDC", "BTC", "ETH", "PERP")


@dataclass
class LegacyTrade:
    row: int
    trade_number: str
    symbol: str
    side: str
    local_open: Optional[datetime]
    local_close: Optional[datetime]
    utc_open: Optional[datetime]
    utc_close: Optional[datetime]
    entry: Optional[float]
    exit: Optional[float]
    stop: Optional[float]
    target: Optional[float]
    r_multiple: Optional[float]
    result_pct: Optional[float]
    move_be: Dict[str, Any]
    move_profit: Dict[str, Any]
    manual: Dict[str, Any]
    uncertain_breakeven: Dict[str, Any]


@dataclass
class RepoTrade:
    row: int
    symbol: str
    side: str
    open_time: Optional[datetime]
    close_time: Optional[datetime]
    entry: Optional[float]
    exit: Optional[float]
    stop: Optional[float]
    target: Optional[float]
    r_multiple: Optional[float]
    result_pct: Optional[float]


def _col(letter: str) -> int:
    return column_index_from_string(letter)


def _is_excel_error(value: Any) -> bool:
    return isinstance(value, str) and value.strip().startswith("#")


def _cell_value(ws, row: int, col: str, ignored: List[Dict[str, Any]]) -> Any:
    cell = ws.cell(row, _col(col))
    value = cell.value
    if getattr(cell, "data_type", None) == "e" or _is_excel_error(value):
        ignored.append({"row": row, "column": col, "value": str(value or "")})
        return None
    return value


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value in (None, "") or _is_excel_error(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    if isinstance(value, (int, float)):
        try:
            return from_excel(float(value)).replace(tzinfo=None)
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("Z", "")
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(text[:19] if fmt.startswith("%Y") and len(text) > 19 else text, fmt)
        except Exception:
            continue
    try:
        parsed = datetime.fromisoformat(text)
        return parsed.replace(tzinfo=None)
    except Exception:
        return None


def _safe_float(value: Any) -> Optional[float]:
    number = _as_float(value)
    if number is None or not math.isfinite(number):
        return None
    return float(number)


def _parse_duration_seconds(value: Any) -> Optional[float]:
    if value in (None, "") or _is_excel_error(value):
        return None
    if isinstance(value, timedelta):
        return max(0.0, value.total_seconds())
    if isinstance(value, time):
        return float(value.hour * 3600 + value.minute * 60 + value.second)
    if isinstance(value, (int, float)):
        number = float(value)
        if 0 < number < 1:
            return number * 86400.0
        parsed = _duration_ddhhmmss_cell_to_seconds(number)
        if parsed is not None:
            return float(parsed)
        return number if number >= 0 else None
    parsed = _parse_duration_text(value)
    return float(parsed) if parsed is not None else None


def _parse_pct_fraction(value: Any) -> Optional[float]:
    if value in (None, "") or _is_excel_error(value):
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("%"):
            number = _safe_float(text[:-1])
            return None if number is None else number / 100.0
    number = _safe_float(value)
    if number is None:
        return None
    return number / 100.0 if abs(number) > 1 else number


def _distance_fraction(value: Any, base_price: Any, comparison_price: Any) -> Optional[float]:
    if value in (None, "") or _is_excel_error(value):
        return None
    if isinstance(value, str) and value.strip().endswith("%"):
        return _parse_pct_fraction(value)
    base = _safe_float(base_price)
    comparison = _safe_float(comparison_price)
    if base not in (None, 0.0) and comparison is not None:
        return abs(float(comparison) - float(base)) / abs(float(base))
    return _parse_pct_fraction(value)


def _canon_symbol(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _canon_side(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in {"LONG", "BUY"} or text.startswith("BUY"):
        return "BUY"
    if text in {"SHORT", "SELL"} or text.startswith("SELL"):
        return "SELL"
    return text


def _yes_no(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"yes", "y", "true", "1", "x"}:
        return "Yes"
    if text in {"no", "n", "false", "0"}:
        return "No"
    return ""


def _order(value: Any) -> str:
    text = str(value or "").strip().lower()
    if "market" in text:
        return "Market"
    if "limit" in text:
        return "Limit"
    return ""


def _aths_atls(value: Any) -> str:
    text = str(value or "").strip().lower().replace("-", " ").replace("_", " ")
    if text in {"aths", "ath", "all time high", "all time highs"}:
        return "All-Time High"
    if text in {"atls", "atl", "all time low", "all time lows"}:
        return "All-Time Low"
    return ""


def _is_crypto_repo_row(account: str, symbol: str) -> bool:
    account_u = str(account or "").upper()
    symbol_u = str(symbol or "").upper()
    return any(token in account_u for token in CRYPTO_ACCOUNTS) or any(token in symbol_u for token in CRYPTO_SYMBOL_TOKENS)


def _price_matches(legacy: LegacyTrade, repo: RepoTrade) -> int:
    pairs = (
        (legacy.entry, repo.entry),
        (legacy.exit, repo.exit),
        (legacy.stop, repo.stop),
        (legacy.target, repo.target),
    )
    matches = 0
    for left, right in pairs:
        if left is None or right is None:
            continue
        tolerance = max(1e-8, abs(float(left)) * 0.0002)
        if abs(float(left) - float(right)) <= tolerance:
            matches += 1
    return matches


def _time_delta_seconds(legacy: LegacyTrade, repo: RepoTrade) -> Optional[float]:
    candidates: List[float] = []
    for open_time, close_time in (
        (legacy.local_open, legacy.local_close),
        (
            legacy.utc_open + timedelta(hours=10) if legacy.utc_open else None,
            legacy.utc_close + timedelta(hours=10) if legacy.utc_close else None,
        ),
    ):
        if open_time and close_time and repo.open_time and repo.close_time:
            candidates.append(
                abs((repo.open_time - open_time).total_seconds())
                + abs((repo.close_time - close_time).total_seconds())
            )
    return min(candidates) if candidates else None


def parse_legacy_trades(path: Path) -> Tuple[List[LegacyTrade], List[Dict[str, Any]], int]:
    wb = load_workbook(path, data_only=True)
    ignored: List[Dict[str, Any]] = []
    fallback_used = 0
    try:
        if LEGACY_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Legacy workbook missing sheet {LEGACY_SHEET!r}.")
        ws = wb[LEGACY_SHEET]
        trades: List[LegacyTrade] = []
        for row in range(2, ws.max_row + 1):
            number = _safe_float(_cell_value(ws, row, "A", ignored))
            if number is None or int(number) != number:
                continue
            trade_number = f"C{int(number)}"
            symbol = str(_cell_value(ws, row, "B", ignored) or "").strip()
            side = _canon_side(_cell_value(ws, row, "AJ", ignored))
            if not symbol or side not in {"BUY", "SELL"}:
                continue
            entry = _safe_float(_cell_value(ws, row, "Z", ignored))
            exit_price = _safe_float(_cell_value(ws, row, "AA", ignored))
            stop = _safe_float(_cell_value(ws, row, "AC", ignored))
            target = _safe_float(_cell_value(ws, row, "AE", ignored))
            break_even_trigger = _safe_float(_cell_value(ws, row, "BN", ignored))
            move_be = {
                "move_to_break_even_time": _parse_datetime(_cell_value(ws, row, "BL", ignored)),
                "move_to_break_even_duration": _parse_duration_seconds(_cell_value(ws, row, "BM", ignored)),
                "move_to_break_even_trigger_price": break_even_trigger,
                "move_to_break_even_distance_from_entry_pct": _distance_fraction(
                    _cell_value(ws, row, "BO", ignored), entry, break_even_trigger
                ),
                "move_to_break_even_distance_from_exit_pct": _distance_fraction(
                    _cell_value(ws, row, "BP", ignored), target, break_even_trigger
                ),
            }
            profit_primary_trigger = _safe_float(_cell_value(ws, row, "BV", ignored))
            profit_primary = {
                "move_to_profit_trigger_price": profit_primary_trigger,
                "move_to_profit_distance_from_entry_pct": _distance_fraction(
                    _cell_value(ws, row, "BW", ignored), entry, profit_primary_trigger
                ),
                "move_to_profit_distance_from_exit_pct": _distance_fraction(
                    _cell_value(ws, row, "BX", ignored), target, profit_primary_trigger
                ),
            }
            profit_fallback_trigger = _safe_float(_cell_value(ws, row, "BS", ignored))
            profit_fallback = {
                "move_to_profit_trigger_price": profit_fallback_trigger,
                "move_to_profit_distance_from_entry_pct": _distance_fraction(
                    _cell_value(ws, row, "BT", ignored), entry, profit_fallback_trigger
                ),
                "move_to_profit_distance_from_exit_pct": _distance_fraction(
                    _cell_value(ws, row, "BU", ignored), target, profit_fallback_trigger
                ),
            }
            use_fallback = any(v is not None for v in profit_fallback.values()) and not any(v is not None for v in profit_primary.values())
            if use_fallback:
                fallback_used += 1
            move_profit = {
                "move_to_profit_time": _parse_datetime(_cell_value(ws, row, "BQ", ignored)),
                "move_to_profit_duration": _parse_duration_seconds(_cell_value(ws, row, "BR", ignored)),
                **(profit_fallback if use_fallback else {k: profit_primary.get(k) if profit_primary.get(k) is not None else profit_fallback.get(k) for k in profit_primary}),
            }
            channel_value = _cell_value(ws, row, "AP", ignored)
            channel = "channel" if _yes_no(channel_value) == "Yes" or "channel" in str(channel_value or "").lower() else ""
            manual = {
                "Pattern": channel,
                "ATHS/ATLS": _aths_atls(_cell_value(ws, row, "BF", ignored)),
                "Order": _order(_cell_value(ws, row, "BH", ignored)),
                "Round Number": _yes_no(_cell_value(ws, row, "BB", ignored)),
                "Spiked Out": _yes_no(_cell_value(ws, row, "BD", ignored)),
                "Close Stopout": _yes_no(_cell_value(ws, row, "BC", ignored)),
                "Near Perfect Entry": _yes_no(_cell_value(ws, row, "AQ", ignored)),
                "Near Win": _yes_no(_cell_value(ws, row, "AR", ignored)),
                "Early Close": _yes_no(_cell_value(ws, row, "AF", ignored)),
                "Timeframe": _canonical_journal_timeframe(_cell_value(ws, row, "AI", ignored)),
            }
            uncertain_breakeven = {
                col: _cell_value(ws, row, col, ignored)
                for col in ("AN", "AO")
                if _cell_value(ws, row, col, ignored) not in (None, "")
            }
            trades.append(
                LegacyTrade(
                    row=row,
                    trade_number=trade_number,
                    symbol=_canon_symbol(symbol),
                    side=side,
                    local_open=_parse_datetime(_cell_value(ws, row, "W", ignored)),
                    local_close=_parse_datetime(_cell_value(ws, row, "X", ignored)),
                    utc_open=_parse_datetime(_cell_value(ws, row, "U", ignored)),
                    utc_close=_parse_datetime(_cell_value(ws, row, "V", ignored)),
                    entry=entry,
                    exit=exit_price,
                    stop=stop,
                    target=target,
                    r_multiple=_safe_float(_cell_value(ws, row, "N", ignored)),
                    result_pct=_safe_float(_cell_value(ws, row, "M", ignored)),
                    move_be={k: v for k, v in move_be.items() if v not in (None, "")},
                    move_profit={k: v for k, v in move_profit.items() if v not in (None, "")},
                    manual={k: v for k, v in manual.items() if v not in (None, "")},
                    uncertain_breakeven=uncertain_breakeven,
                )
            )
        return trades, ignored, fallback_used
    finally:
        wb.close()


def parse_repo_trades(ws) -> Tuple[List[RepoTrade], Dict[str, int]]:
    _ensure_trade_log_schema(ws)
    headers = _trade_log_header_map(ws)
    start_row = _trade_log_data_start_row(ws)
    trades: List[RepoTrade] = []
    for row in range(start_row, ws.max_row + 1):
        row_type = str(ws.cell(row, headers.get("Row Type", 0)).value or "trade").strip().lower() if headers.get("Row Type") else "trade"
        if row_type != "trade":
            continue
        account = str(ws.cell(row, headers.get("Account", 0)).value or "") if headers.get("Account") else ""
        symbol = str(ws.cell(row, headers.get("Symbol", 0)).value or "") if headers.get("Symbol") else ""
        if not _is_crypto_repo_row(account, symbol):
            continue
        trades.append(
            RepoTrade(
                row=row,
                symbol=_canon_symbol(symbol),
                side=_canon_side(ws.cell(row, headers.get("Side", 0)).value if headers.get("Side") else ""),
                open_time=_as_datetime(ws.cell(row, headers.get("Open Time", 0)).value) if headers.get("Open Time") else None,
                close_time=_as_datetime(ws.cell(row, headers.get("Close Time", 0)).value) if headers.get("Close Time") else None,
                entry=_safe_float(ws.cell(row, headers.get("Entry Price", 0)).value) if headers.get("Entry Price") else None,
                exit=_safe_float(ws.cell(row, headers.get("Exit Price", 0)).value) if headers.get("Exit Price") else None,
                stop=_safe_float(ws.cell(row, headers.get("Stop Loss Price", 0)).value) if headers.get("Stop Loss Price") else None,
                target=_safe_float(ws.cell(row, headers.get("Target Price", 0)).value) if headers.get("Target Price") else None,
                r_multiple=_safe_float(ws.cell(row, headers.get("R-Multiple", 0)).value) if headers.get("R-Multiple") else None,
                result_pct=(_safe_float(ws.cell(row, headers.get("Profit %", 0)).value) * 100.0 if headers.get("Profit %") and _safe_float(ws.cell(row, headers.get("Profit %", 0)).value) is not None else None),
            )
        )
    return trades, headers


def match_trades(legacy_trades: List[LegacyTrade], repo_trades: List[RepoTrade]) -> Tuple[Dict[int, Tuple[LegacyTrade, RepoTrade]], List[LegacyTrade], List[Dict[str, Any]]]:
    matches: Dict[int, Tuple[LegacyTrade, RepoTrade]] = {}
    unmatched: List[LegacyTrade] = []
    ambiguous: List[Dict[str, Any]] = []
    for legacy in legacy_trades:
        qualified: List[Tuple[Tuple[float, int, float, float], RepoTrade]] = []
        for repo in repo_trades:
            if legacy.symbol != repo.symbol or legacy.side != repo.side:
                continue
            delta = _time_delta_seconds(legacy, repo)
            if delta is None:
                continue
            price_count = _price_matches(legacy, repo)
            if delta <= 300 or (delta <= 3600 and price_count >= 3):
                r_diff = abs((legacy.r_multiple or 0.0) - (repo.r_multiple or 0.0)) if legacy.r_multiple is not None and repo.r_multiple is not None else 999999.0
                pct_diff = abs((legacy.result_pct or 0.0) - (repo.result_pct or 0.0)) if legacy.result_pct is not None and repo.result_pct is not None else 999999.0
                qualified.append(((float(delta), -price_count, r_diff, pct_diff), repo))
        qualified.sort(key=lambda item: item[0])
        if not qualified:
            unmatched.append(legacy)
            continue
        if len(qualified) > 1 and qualified[0][0] == qualified[1][0]:
            ambiguous.append({
                "legacy_row": legacy.row,
                "trade_number": legacy.trade_number,
                "repo_rows": [qualified[0][1].row, qualified[1][1].row],
            })
            continue
        matches[legacy.row] = (legacy, qualified[0][1])
    return matches, unmatched, ambiguous


def _has_any_move_data(values: Dict[str, Any]) -> bool:
    return any(value not in (None, "") for value in values.values())


def _write_move_fields(ws, headers: Dict[str, int], row: int, values: Dict[str, Any], *, overwrite: bool = False) -> int:
    field_to_header = {field: header for header, field in MOVE_TO_FIELD_MAP.items()}
    written = 0
    for field, value in values.items():
        header = field_to_header.get(field)
        col = headers.get(header or "")
        if not col or value in (None, ""):
            continue
        cell = ws.cell(row, col)
        if cell.value not in (None, "") and not overwrite:
            continue
        if field.endswith("_duration"):
            cell.value = _fmt_duration_full(value)
            cell.number_format = r'00\:00\:00\:00'
        elif field.endswith("_pct"):
            cell.value = value
            cell.number_format = "0.00%"
        else:
            cell.value = value
            if field.endswith("_time"):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
        written += 1
    return written


def run_backfill(journal_path: Path, legacy_path: Path, *, apply_changes: bool = False, overwrite: bool = False) -> Dict[str, Any]:
    legacy_trades, ignored, fallback_used = parse_legacy_trades(legacy_path)
    wb = load_workbook(journal_path)
    try:
        if "Trade Log" not in wb.sheetnames:
            raise RuntimeError("Journal workbook missing Trade Log sheet.")
        ws = wb["Trade Log"]
        repo_trades, headers = parse_repo_trades(ws)
        matches, unmatched, ambiguous = match_trades(legacy_trades, repo_trades)
        target_counts = Counter(repo.row for _legacy, repo in matches.values())
        duplicate_targets = sorted(row for row, count in target_counts.items() if count > 1)
        move_be_rows = sum(1 for legacy, _repo in matches.values() if _has_any_move_data(legacy.move_be))
        move_profit_rows = sum(1 for legacy, _repo in matches.values() if _has_any_move_data(legacy.move_profit))
        summary: Dict[str, Any] = {
            "legacy_trades_parsed": len(legacy_trades),
            "matches": len(matches),
            "unmatched": len(unmatched),
            "ambiguous": len(ambiguous),
            "duplicated_repo_targets": len(duplicate_targets),
            "trade_numbers_to_write": len(matches),
            "move_be_rows_to_write": move_be_rows,
            "move_profit_rows_to_write": move_profit_rows,
            "ignored_invalid_cells": len(ignored),
            "fallback_used": fallback_used,
            "uncertain_breakeven_candidates": sum(bool(trade.uncertain_breakeven) for trade in legacy_trades),
            "applied": False,
            "unmatched_legacy_rows": [trade.row for trade in unmatched[:20]],
            "ambiguous_matches": ambiguous[:20],
            "duplicated_repo_target_rows": duplicate_targets[:20],
        }
        if apply_changes and (unmatched or ambiguous or duplicate_targets):
            summary["error"] = "unsafe_match_result"
            return summary
        if apply_changes:
            trade_col = headers.get(TRADE_NUMBER_HEADER)
            if not trade_col:
                raise RuntimeError("Journal Trade Log missing Trade Number column after schema migration.")
            for legacy, repo in matches.values():
                cell = ws.cell(repo.row, trade_col)
                if cell.value in (None, "") or overwrite:
                    cell.value = legacy.trade_number
                    cell.number_format = "@"
                _write_move_fields(ws, headers, repo.row, legacy.move_be, overwrite=overwrite)
                _write_move_fields(ws, headers, repo.row, legacy.move_profit, overwrite=overwrite)
                for header, value in legacy.manual.items():
                    col = headers.get(header)
                    if not col or value in (None, ""):
                        continue
                    target = ws.cell(repo.row, col)
                    if target.value not in (None, "") and not overwrite:
                        continue
                    target.value = value
            _repair_trade_log_move_to_durations(ws, summary)
            wb.save(journal_path)
            summary["applied"] = True
        return summary
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Backfill legacy crypto Trade Number and move-to metadata into Trading Journal.xlsx.")
    parser.add_argument("--journal", required=True, type=Path)
    parser.add_argument("--legacy", required=True, type=Path)
    parser.add_argument("--apply", action="store_true", help="Write changes. Omit for dry-run.")
    parser.add_argument("--overwrite", action="store_true", help="Replace existing manual values.")
    args = parser.parse_args()
    summary = run_backfill(args.journal, args.legacy, apply_changes=bool(args.apply), overwrite=bool(args.overwrite))
    print(json.dumps(summary, indent=2, sort_keys=True, default=str))
    return 2 if summary.get("error") else 0


if __name__ == "__main__":
    raise SystemExit(main())
