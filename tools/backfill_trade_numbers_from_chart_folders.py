from __future__ import annotations

import argparse
import io
import json
import math
import re
import sys
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.master_journal_workbook import (
    TRADE_NUMBER_HEADER,
    _as_datetime,
    _as_float,
    _ensure_trade_log_schema,
    _trade_log_data_start_row,
    _trade_log_header_map,
)


FOREX_FOLDER_RE = re.compile(
    r"^FOREX/(?P<year>\d{4})/(?P<trade>F\d+)(?:\s+(?P<token>[^/]+))?/",
    re.IGNORECASE,
)
CRYPTO_FOLDER_RE = re.compile(
    r"^CRYPTO/(?P<year>\d{4})/(?:(?P<month>[^/]+)/)?"
    r"(?P<trade>C\d+)(?:\s+(?P<token>[^/]+))?/",
    re.IGNORECASE,
)
DATE_RE = re.compile(r"(20\d{2})[-_.](\d{2})[-_.](\d{2})")
COMPACT_DATE_RE = re.compile(r"(20\d{2})(\d{2})(\d{2})")
MONTHS = {
    name: number
    for number, name in enumerate(
        ("", "JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")
    )
    if name
}


@dataclass
class ChartFolder:
    market: str
    year: int
    month: Optional[int]
    trade_number: str
    label_trade_number: str
    token: str
    symbol: str
    folder_path: str
    entries: List[str] = field(default_factory=list)
    dates: List[date] = field(default_factory=list)
    direction: str = ""
    entry: Optional[float] = None
    stop: Optional[float] = None
    target: Optional[float] = None


@dataclass
class RepoTrade:
    row: int
    trade_number: str
    market: str
    year: Optional[int]
    month: Optional[int]
    symbol: str
    side: str
    open_time: Optional[datetime]
    close_time: Optional[datetime]
    entry: Optional[float]
    stop: Optional[float]
    target: Optional[float]


def _canon_symbol(value: Any) -> str:
    return "".join(ch for ch in str(value or "").upper() if ch.isalnum())


def _expected_symbol(token: str, market: str = "crypto") -> str:
    symbol = _canon_symbol(token)
    if not symbol:
        return ""
    if market == "forex":
        return symbol
    return symbol if symbol.endswith(("USDT", "USDC")) else f"{symbol}USDT"


def _canon_side(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in {"LONG", "BUY"} or text.startswith("BUY"):
        return "BUY"
    if text in {"SHORT", "SELL"} or text.startswith("SELL"):
        return "SELL"
    return text


def _safe_float(value: Any) -> Optional[float]:
    number = _as_float(value)
    return float(number) if number is not None and math.isfinite(number) else None


def _extract_dates(text: str) -> List[date]:
    values: List[date] = []
    for pattern in (DATE_RE, COMPACT_DATE_RE):
        for match in pattern.finditer(text):
            try:
                values.append(date(*(int(part) for part in match.groups())))
            except ValueError:
                continue
    return values


def _summary_value(text: str, *labels: str) -> str:
    for label in labels:
        match = re.search(rf"(?im)^\s*{re.escape(label)}\s*:\s*([^\r\n]+)", text)
        if match:
            return str(match.group(1)).strip()
    return ""


def _summary_number(text: str, *labels: str) -> Optional[float]:
    value = _summary_value(text, *labels)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", value)
    return _safe_float(match.group(0)) if match else None


def _xlsx_member_text(data: bytes) -> str:
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return ""
    values: List[str] = []
    try:
        for ws in wb.worksheets[:3]:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value not in (None, ""):
                        values.append(str(cell.value))
                        if len(values) >= 500:
                            return "\n".join(values)
    finally:
        wb.close()
    return "\n".join(values)


def _metadata_text(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> str:
    name = info.filename.lower()
    if info.file_size > 5_000_000:
        return ""
    try:
        data = archive.read(info)
    except Exception:
        return ""
    if name.endswith((".txt", ".csv", ".json")):
        return data.decode("utf-8-sig", errors="replace")
    if name.endswith(".xlsx") and re.search(r"trade|position|order|history|summary|scan", name):
        return _xlsx_member_text(data)
    return ""


def parse_chart_folders(path: Path, diagnostics: Optional[Dict[str, Any]] = None) -> List[ChartFolder]:
    diagnostics = diagnostics if isinstance(diagnostics, dict) else {}
    folders: Dict[str, ChartFolder] = {}
    metadata: Dict[str, List[str]] = {}
    skipped_2026_plus: set[str] = set()
    with zipfile.ZipFile(path) as archive:
        for info in archive.infolist():
            normalized = info.filename.replace("\\", "/")
            match = FOREX_FOLDER_RE.match(normalized)
            market = "forex"
            if not match:
                match = CRYPTO_FOLDER_RE.match(normalized)
                market = "crypto"
            if not match:
                continue
            year = int(match.group("year"))
            root_path = normalized[:match.end()].rstrip("/")
            if year >= 2026:
                skipped_2026_plus.add(root_path)
                continue
            label_trade = match.group("trade").upper()
            trade_number = label_trade
            folder_path = root_path
            remainder = normalized[match.end():]
            if market == "crypto":
                nested = re.match(r"(?P<number>\d+)/", remainder)
                if nested:
                    trade_number = f"C{int(nested.group('number'))}"
                    folder_path = f"{root_path}/{nested.group('number')}"
                elif info.is_dir() and not remainder:
                    continue
            elif info.is_dir() and not remainder:
                continue
            token = str(match.groupdict().get("token") or "").strip()
            month_token = str(match.groupdict().get("month") or "").strip().upper()
            folder = folders.setdefault(
                folder_path,
                ChartFolder(
                    market=market,
                    year=year,
                    month=MONTHS.get(month_token[:3]),
                    trade_number=trade_number,
                    label_trade_number=label_trade,
                    token=token,
                    symbol=_expected_symbol(token, market),
                    folder_path=folder_path,
                ),
            )
            folder.entries.append(normalized)
            folder.dates.extend(_extract_dates(normalized))
            text = _metadata_text(archive, info)
            if text:
                metadata.setdefault(folder_path, []).append(text)
                folder.dates.extend(_extract_dates(text))

    symbol_pattern = re.compile(r"\b[A-Z0-9]{2,15}(?:USDT|USDC|USD|JPY|CHF|AUD|NZD|CAD|GBP|EUR)\b")
    for folder_path, texts in metadata.items():
        folder = folders[folder_path]
        text = "\n".join(texts)
        summary_symbol = _summary_value(text, "Symbol", "Pair", "Instrument")
        if summary_symbol and not folder.symbol:
            folder.symbol = _canon_symbol(summary_symbol)
        elif not folder.symbol:
            symbols = {_canon_symbol(value) for value in symbol_pattern.findall(text.upper())}
            if len(symbols) == 1:
                folder.symbol = symbols.pop()
        folder.direction = _canon_side(_summary_value(text, "Direction", "Side"))
        folder.entry = _summary_number(text, "Entry Price", "Entry")
        folder.stop = _summary_number(text, "Stop Price", "Stop Loss", "Stop Loss Price")
        folder.target = _summary_number(text, "Target Price", "Take Profit", "Take Profit Price")
        folder.dates = sorted(set(folder.dates))

    diagnostics["folders_skipped_2026_plus"] = len(skipped_2026_plus)
    return sorted(
        folders.values(),
        key=lambda folder: (folder.market, folder.year, folder.month or 0, int(folder.trade_number[1:]), folder.folder_path),
    )


def _repo_market(account: str, symbol: str) -> str:
    account_u = account.upper()
    symbol_u = symbol.upper()
    if any(token in account_u for token in ("BYBIT", "BINANCE", "COINSPOT")) or symbol_u.endswith(
        ("USDT", "USDC", "PERP")
    ):
        return "crypto"
    return "forex"


def parse_repo_trades(ws, account: str = "") -> Tuple[List[RepoTrade], Dict[str, int]]:
    _ensure_trade_log_schema(ws)
    headers = _trade_log_header_map(ws)
    account_token = account.strip().upper()
    trades: List[RepoTrade] = []
    for row in range(_trade_log_data_start_row(ws), ws.max_row + 1):
        row_type = str(ws.cell(row, headers.get("Row Type", 0)).value or "trade").strip().lower()
        account_value = str(ws.cell(row, headers.get("Account", 0)).value or "").strip().upper()
        if row_type != "trade" or (account_token and account_token not in account_value):
            continue
        symbol = _canon_symbol(ws.cell(row, headers.get("Symbol", 0)).value)
        opened = _as_datetime(ws.cell(row, headers.get("Open Time", 0)).value)
        closed = _as_datetime(ws.cell(row, headers.get("Close Time", 0)).value)
        trade_date = opened or closed
        trades.append(RepoTrade(
            row=row,
            trade_number=str(ws.cell(row, headers[TRADE_NUMBER_HEADER]).value or "").strip(),
            market=_repo_market(account_value, symbol),
            year=trade_date.year if trade_date else None,
            month=trade_date.month if trade_date else None,
            symbol=symbol,
            side=_canon_side(ws.cell(row, headers.get("Side", 0)).value),
            open_time=opened,
            close_time=closed,
            entry=_safe_float(ws.cell(row, headers.get("Entry Price", 0)).value),
            stop=_safe_float(ws.cell(row, headers.get("Stop Loss Price", 0)).value),
            target=_safe_float(ws.cell(row, headers.get("Target Price", 0)).value),
        ))
    return trades, headers


def _date_distance(folder: ChartFolder, repo: RepoTrade) -> int:
    if not folder.dates:
        return 999999
    repo_dates = [value.date() for value in (repo.open_time, repo.close_time) if value is not None]
    if not repo_dates:
        return 999999
    return min(abs((repo_date - folder_date).days) for repo_date in repo_dates for folder_date in folder.dates)


def _price_distance(folder: ChartFolder, repo: RepoTrade) -> float:
    distances: List[float] = []
    for expected, actual in ((folder.entry, repo.entry), (folder.stop, repo.stop), (folder.target, repo.target)):
        if expected is None or actual is None:
            continue
        distances.append(abs(expected - actual) / max(abs(expected), 1e-9))
    return sum(distances) / len(distances) if distances else 999999.0


def _legacy_row_trade_numbers(
    ws,
    market: str,
    legacy_path: Optional[Path],
    override_map: Optional[Path],
) -> Tuple[Dict[int, str], Dict[str, Any]]:
    if not legacy_path:
        return {}, {"legacy_journal_used": False, "legacy_journal_missing": True}
    if not legacy_path.exists():
        return {}, {"legacy_journal_used": False, "legacy_journal_missing": True, "legacy_journal_path": str(legacy_path)}
    if market == "forex":
        from tools.backfill_legacy_forex_trade_metadata import (
            _load_overrides,
            match_trades,
            parse_legacy_trades,
            parse_repo_trades as parse_legacy_repo_trades,
        )
        legacy, _ignored, _fallback = parse_legacy_trades(legacy_path)
        repo, _headers = parse_legacy_repo_trades(ws)
        matches, unmatched, ambiguous = match_trades(
            legacy,
            repo,
            _load_overrides(override_map) if override_map else {},
        )
    else:
        from tools.backfill_legacy_crypto_trade_metadata import (
            match_trades,
            parse_legacy_trades,
            parse_repo_trades as parse_legacy_repo_trades,
        )
        legacy, _ignored, _fallback = parse_legacy_trades(legacy_path)
        repo, _headers = parse_legacy_repo_trades(ws)
        matches, unmatched, ambiguous = match_trades(legacy, repo)
    mapping = {repo_trade.row: legacy_trade.trade_number for legacy_trade, repo_trade in matches.values()}
    return mapping, {
        "legacy_journal_used": True,
        "legacy_journal_missing": False,
        "legacy_journal_path": str(legacy_path),
        "legacy_matches": len(mapping),
        "legacy_unmatched": len(unmatched),
        "legacy_ambiguous": len(ambiguous),
    }


def match_chart_folders(
    folders: List[ChartFolder],
    repo_trades: List[RepoTrade],
    legacy_trade_numbers: Optional[Dict[int, str]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    matches: List[Dict[str, Any]] = []
    existing: List[Dict[str, Any]] = []
    unmatched: List[Dict[str, Any]] = []
    ambiguous: List[Dict[str, Any]] = []
    assigned_rows: set[int] = set()
    legacy_trade_numbers = legacy_trade_numbers or {}
    for folder in folders:
        scoped = [
            trade for trade in repo_trades
            if trade.market == folder.market
            and trade.year == folder.year
            and trade.row not in assigned_rows
            and (not trade.trade_number or trade.trade_number == folder.trade_number)
            and (not folder.month or trade.month == folder.month)
            and (not folder.symbol or trade.symbol == folder.symbol)
            and (not folder.direction or trade.side == folder.direction)
        ]
        existing_number_candidates = [
            trade for trade in scoped
            if trade.trade_number == folder.trade_number
        ]
        legacy_candidates = [
            trade for trade in scoped
            if legacy_trade_numbers.get(trade.row) == folder.trade_number
        ]
        if len(existing_number_candidates) == 1:
            existing_trade = existing_number_candidates[0]
            scored = [(_date_distance(folder, existing_trade), _price_distance(folder, existing_trade), existing_trade.row, existing_trade)]
            evidence = "existing_trade_number"
        elif len(legacy_candidates) == 1:
            scored = [(_date_distance(folder, legacy_candidates[0]), _price_distance(folder, legacy_candidates[0]), legacy_candidates[0].row, legacy_candidates[0])]
            evidence = "legacy_execution_match"
        else:
            scored = sorted(
                [(_date_distance(folder, trade), _price_distance(folder, trade), trade.row, trade) for trade in scoped],
                key=lambda item: item[:3],
            )
            evidence = ""
        if not scored:
            unmatched.append({
                "trade_number": folder.trade_number,
                "folder": folder.folder_path,
                "symbol": folder.symbol,
                "reason": "no_repo_candidate",
            })
            continue
        best = scored[0]
        exact_symbol = bool(folder.symbol and best[3].symbol == folder.symbol)
        date_confident = best[0] <= 3
        price_confident = best[1] <= 0.08
        unique_candidate = len(scored) == 1
        separated = (
            unique_candidate
            or best[0] < scored[1][0]
            or (best[0] == scored[1][0] and best[1] + 0.005 < scored[1][1])
        )
        folder_token_unique = bool(folder.token and exact_symbol and unique_candidate)
        high_confidence = bool(
            evidence
            or (exact_symbol and (date_confident or price_confident) and separated)
            or folder_token_unique
        )
        if not high_confidence:
            payload = {
                "trade_number": folder.trade_number,
                "folder": folder.folder_path,
                "symbol": folder.symbol,
                "repo_rows": [item[3].row for item in scored[:5]],
                "reason": "insufficient_discriminator" if not folder.symbol and not evidence else "ambiguous_repo_candidates",
            }
            (unmatched if payload["reason"] == "insufficient_discriminator" else ambiguous).append(payload)
            continue
        repo = best[3]
        if repo.trade_number and repo.trade_number != folder.trade_number:
            ambiguous.append({
                "trade_number": folder.trade_number,
                "folder": folder.folder_path,
                "symbol": folder.symbol,
                "repo_rows": [repo.row],
                "reason": "conflicting_existing_trade_number",
                "existing_trade_number": repo.trade_number,
            })
            continue
        assigned_rows.add(repo.row)
        payload = {
            "trade_number": folder.trade_number,
            "folder": folder.folder_path,
            "symbol": folder.symbol or repo.symbol,
            "repo_row": repo.row,
            "date_distance_days": None if best[0] >= 999999 else best[0],
            "price_distance": None if best[1] >= 999999 else best[1],
            "confidence": "high",
            "evidence": evidence or ("symbol_date_or_price" if date_confident or price_confident else "unique_symbol_folder"),
        }
        if repo.trade_number == folder.trade_number:
            existing.append(payload)
        else:
            matches.append(payload)
    return matches, existing, unmatched, ambiguous


def _blank_pre_2026_trade_numbers(repo_trades: List[RepoTrade]) -> int:
    return sum(trade.year is not None and trade.year < 2026 and not trade.trade_number for trade in repo_trades)


def run_backfill(
    journal_path: Path,
    charts_zip: Path,
    *,
    account: str = "",
    apply_changes: bool = False,
    legacy_path: Optional[Path] = None,
    override_map: Optional[Path] = None,
) -> Dict[str, Any]:
    parse_diagnostics: Dict[str, Any] = {}
    folders = parse_chart_folders(charts_zip, parse_diagnostics)
    markets = {folder.market for folder in folders}
    market = next(iter(markets)) if len(markets) == 1 else ""
    wb = load_workbook(journal_path)
    try:
        ws = wb["Trade Log"]
        legacy_mapping, legacy_summary = _legacy_row_trade_numbers(ws, market, legacy_path, override_map)
        repo_trades, headers = parse_repo_trades(ws, account)
        matches, existing, unmatched, ambiguous = match_chart_folders(folders, repo_trades, legacy_mapping)
        blanks_before = _blank_pre_2026_trade_numbers(repo_trades)
        summary: Dict[str, Any] = {
            "archive": str(charts_zip),
            "market": market,
            "folders_parsed": len(folders),
            "folders_skipped_2026_plus": int(parse_diagnostics.get("folders_skipped_2026_plus") or 0),
            "high_confidence_matches": len(matches),
            "skipped_existing": len(existing),
            "ambiguous": len(ambiguous),
            "unmatched": len(unmatched),
            "blank_pre_2026_trade_numbers_before": blanks_before,
            "blank_pre_2026_trade_numbers_remaining": max(0, blanks_before - len(matches)),
            "matches": matches,
            "existing_folders": existing,
            "ambiguous_folders": ambiguous,
            "unmatched_folders": unmatched,
            "applied": False,
            **legacy_summary,
        }
        if not apply_changes:
            return summary
        trade_col = headers[TRADE_NUMBER_HEADER]
        written = 0
        skipped_nonblank = 0
        for match in matches:
            cell = ws.cell(match["repo_row"], trade_col)
            if cell.value not in (None, ""):
                skipped_nonblank += 1
                continue
            cell.value = match["trade_number"]
            cell.number_format = "@"
            written += 1
        wb.save(journal_path)
        summary.update({
            "applied": True,
            "trade_numbers_written": written,
            "skipped_nonblank": skipped_nonblank,
            "blank_pre_2026_trade_numbers_remaining": max(0, blanks_before - written),
        })
        return summary
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Backfill Trade Number from FOREX or CRYPTO chart-folder labels.")
    parser.add_argument("--journal", required=True, type=Path)
    parser.add_argument("--charts-zip", required=True, type=Path)
    parser.add_argument("--account", default="")
    parser.add_argument("--legacy", type=Path, help="Optional matching legacy journal workbook.")
    parser.add_argument("--override-map", type=Path, help="Optional Forex legacy override mapping.")
    parser.add_argument("--apply", action="store_true", help="Write high-confidence matches. Omit for dry-run.")
    args = parser.parse_args()
    summary = run_backfill(
        args.journal,
        args.charts_zip,
        account=args.account,
        apply_changes=args.apply,
        legacy_path=args.legacy,
        override_map=args.override_map,
    )
    print(json.dumps(summary, indent=2, sort_keys=True, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
