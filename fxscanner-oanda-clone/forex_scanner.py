import os
import subprocess
import platform
import webbrowser
import logging
import re
import requests
import pandas as pd
from datetime import datetime, timezone
from typing import Optional
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
import time

# Desktop notification functionality has been removed
# to avoid pop-up messages on Windows or other platforms.
notification = None

# Environment variables for authentication
API_KEY = os.getenv('OANDA_API_KEY')
ACCOUNT_ID = os.getenv('OANDA_ACCOUNT_ID')
# Default to the live trading API unless overridden
BASE_URL = os.getenv('OANDA_BASE_URL', 'https://api-fxtrade.oanda.com/v3')

HEADERS = {'Authorization': f'Bearer {API_KEY}'}

# Folder where Excel files are saved
OUTPUT_DIR = 'output'



def get_available_instruments():
    """Return a set of instruments enabled for the account."""
    url = f"{BASE_URL}/accounts/{ACCOUNT_ID}/instruments"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        r.raise_for_status()
        data = r.json()
        return {inst['name'] for inst in data.get('instruments', [])}
    except requests.Timeout:
        print("Timed out retrieving instruments")
        return set()
    except requests.RequestException:
        # Fallback to predefined list if request fails
        return set()


# Instrument lists
MAJOR_PAIRS = [
    'EUR_USD', 'USD_JPY', 'GBP_USD', 'USD_CHF',
    'USD_CAD', 'AUD_USD', 'NZD_USD'
]
MINOR_PAIRS = [
    'EUR_GBP', 'EUR_JPY', 'GBP_JPY', 'CHF_JPY', 'EUR_CHF',
    'EUR_CAD', 'EUR_AUD', 'EUR_NZD', 'GBP_CHF', 'GBP_CAD',
    'GBP_AUD', 'GBP_NZD', 'AUD_JPY', 'AUD_CHF', 'AUD_CAD',
    'AUD_NZD', 'NZD_JPY', 'NZD_CHF', 'NZD_CAD', 'CAD_JPY', 'CAD_CHF'
]
CROSS_PAIRS = [
    'EUR_GBP', 'EUR_JPY', 'EUR_CHF', 'EUR_AUD', 'EUR_NZD', 'EUR_CAD',
    'GBP_JPY', 'GBP_CHF', 'GBP_AUD', 'GBP_NZD', 'GBP_CAD',
    'AUD_JPY', 'AUD_CHF', 'AUD_NZD', 'AUD_CAD',
    'NZD_JPY', 'NZD_CHF', 'NZD_CAD', 'CAD_JPY', 'CAD_CHF', 'CHF_JPY'
]
INDICES = [
    'US30_USD', 'SPX500_USD', 'NAS100_USD', 'UK100_GBP', 'DE40_EUR',
    'FRA40_EUR', 'EU50_EUR', 'ES35_EUR', 'AU200_AUD', 'JP225_USD',
    'HK33_HKD', 'CN50_USD', 'SG30_SGD'
]
METALS = ['XAU_USD', 'XAG_USD']
COMMODITIES = ['NATGAS_USD', 'BCO_USD', 'WTICO_USD']
CRYPTO = ['BTC_USD']

INSTRUMENTS = sorted(set(MAJOR_PAIRS + MINOR_PAIRS + CROSS_PAIRS + INDICES + METALS + COMMODITIES + CRYPTO))

# Granularities used when fetching data
TIMEFRAMES = {
    '5M': 'M5',
    '15M': 'M15',
    '30M': 'M30',
    '1H': 'H1',
    '4H': 'H4',
}


def build_correlations(lookback: int = 50):
    """Calculate return correlations for each timeframe."""
    correlations = {}
    for label, gran in TIMEFRAMES.items():
        returns = {}
        min_len = None
        for inst in INSTRUMENTS:
            candles = fetch_candles(inst, gran, count=lookback + 1, complete_only=True)
            closes = [float(c['mid']['c']) for c in candles]
            if len(closes) < 2:
                continue
            series = pd.Series(closes).pct_change().dropna()
            returns[inst] = series
            if min_len is None or len(series) < min_len:
                min_len = len(series)
        if returns:
            aligned = {k: v.iloc[:min_len].reset_index(drop=True) for k, v in returns.items()}
            df = pd.DataFrame(aligned)
            correlations[label] = df.corr()
    return correlations


def fetch_candles(instrument: str, granularity: str, count: int = 2, *, complete_only: bool = True):
    """Fetch recent candles and optionally filter to completed ones."""
    url = f"{BASE_URL}/instruments/{instrument}/candles"
    params = {
        'granularity': granularity,
        'count': max(count, 3),  # request a few extra in case last is incomplete
        'price': 'M',
    }
    try:
        r = requests.get(url, headers=HEADERS, params=params, timeout=10)
        r.raise_for_status()
    except requests.Timeout:
        print(f"Timed out fetching candles for {instrument}")
        return []
    except requests.RequestException:
        return []
    candles = r.json().get('candles', [])
    if complete_only:
        candles = [c for c in candles if c.get('complete')]
    return candles[-count:]


def calc_price_metrics(instrument: str):
    price_change = {}
    price_range = {}
    for label, gran in TIMEFRAMES.items():
        candles = fetch_candles(instrument, gran, count=2, complete_only=True)
        if len(candles) < 2:
            continue
        open_price = float(candles[0]['mid']['o'])
        close_price = float(candles[-1]['mid']['c'])
        high = max(float(c['mid']['h']) for c in candles)
        low = min(float(c['mid']['l']) for c in candles)
        # Return decimal fractions so Excel percentage formatting displays
        # the correct values (e.g. 0.004 -> 0.4%).
        change = (close_price - open_price) / open_price
        rng = (high - low) / open_price
        price_change[label] = change
        price_range[label] = rng
    return price_change, price_range




def fetch_spreads(instruments):
    """Return the current spread for each instrument as a percentage."""
    url = f"{BASE_URL}/accounts/{ACCOUNT_ID}/pricing"
    spreads = {}
    params = {'instruments': ','.join(instruments)}
    try:
        r = requests.get(url, headers=HEADERS, params=params, timeout=10)
        r.raise_for_status()
    except requests.Timeout:
        print("Timed out retrieving pricing data")
        return spreads
    except requests.RequestException:
        return spreads

    for price in r.json().get('prices', []):
        bid = float(price.get('closeoutBid', 0))
        ask = float(price.get('closeoutAsk', 0))
        if bid and ask:
            # Store decimal fraction so percentage formatting in Excel shows
            # the correct value
            spreads[price['instrument']] = (ask - bid) / ask
    return spreads


def build_data():
    """Run each type of scan sequentially for all instruments."""

    price_change_data = {}
    price_range_data = {}
    for inst in INSTRUMENTS:
        pc, pr = calc_price_metrics(inst)
        price_change_data[inst] = pc
    print("Price change scan complete.")

    price_range_data = {}
    print("Scanning price range for all symbols...")
    for inst in INSTRUMENTS:
        _, pr = calc_price_metrics(inst)
        price_range_data[inst] = pr
    print("Price range scan complete.")

    print("Calculating correlations...")
    correlations = build_correlations()
    print("Correlation matrices complete.")

    # Get the current spread percentage for each instrument
    spreads = fetch_spreads(INSTRUMENTS)

    return price_change_data, price_range_data, spreads, correlations

def export_to_excel(price_change, price_range, spreads, correlations, filename='scan_output.xlsx'):
    price_change_df = pd.DataFrame.from_dict(price_change, orient='index')
    price_change_df.insert(0, 'Symbol', price_change_df.index)
    price_range_df = pd.DataFrame.from_dict(price_range, orient='index')
    price_range_df.insert(0, 'Symbol', price_range_df.index)

    spread_df = pd.DataFrame(list(spreads.items()), columns=['Symbol', 'Spread'])
    spread_df = spread_df.sort_values('Spread', ascending=True).reset_index(drop=True)

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # Price Range sheet
    price_range_df.to_excel(writer, sheet_name='Price Range', index=False, startrow=1)
    ws = writer.sheets['Price Range']
    ws['A1'] = '% Price Range'
    ws.insert_rows(3)

    # Price Change sheet
    price_change_df.to_excel(writer, sheet_name='Price Change', index=False, startrow=1)
    ws = writer.sheets['Price Change']
    ws['A1'] = '% Price Change'
    ws.insert_rows(3)
    start_col = 2
    end_col = ws.max_column
    start_row = 4
    end_row = ws.max_row
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    cell_range = f"{start_letter}{start_row}:{end_letter}{end_row}"
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill))

    # Spread sheet
    spread_df.to_excel(writer, sheet_name='Spread', index=False, startrow=1)
    ws = writer.sheets['Spread']
    ws['A1'] = 'Current Spread %'
    ws.insert_rows(3)

    for label, corr_df in correlations.items():
        sheet_name = f'Corr {label}'
        corr_df.to_excel(writer, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        start_row = 2
        end_row = ws.max_row
        start_col = 2
        end_col = ws.max_column
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        cell_range = f"{start_letter}{start_row}:{end_letter}{end_row}"
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='lessThan', formula=['0'], fill=red_fill),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill),
        )
        # Keep the headers visible when scrolling by freezing the first
        # row and column. This sets the freeze point at cell B2 so row 1
        # and column A stay locked in place.
        ws.freeze_panes = ws["B2"]

    # apply percentage format while keeping numeric values
    for name, sheet in writer.sheets.items():
        if name == 'Spread':
            decimals = 6
            min_row = 4
        elif name.startswith('Corr'):
            decimals = 2
            min_row = 2
        else:
            decimals = 4
            min_row = 4

        fmt = f"0.{''.join('0' for _ in range(decimals))}%"

        for row in sheet.iter_rows(min_row=min_row):
            for cell in row[1:]:
                if isinstance(cell.value, float):
                    cell.number_format = fmt

    writer.close()


_OPENED_PATHS: set[str] = set()


def open_in_edge(file_path: str) -> None:
    """Open ``file_path`` in Microsoft Edge only once per run."""
    if file_path in _OPENED_PATHS:
        print(f"Edge already open for {file_path}")
        return

    _OPENED_PATHS.add(file_path)

    if platform.system() == "Windows":
        try:
            subprocess.Popen(["cmd", "/c", "start", "msedge", file_path])
        except OSError as exc:
            print(f"Failed to open Edge: {exc}")
    else:
        webbrowser.open(file_path)


def export_to_html(
    df: pd.DataFrame,
    filename: str,
    header: str,
    *,
    include_sort_buttons: bool = False,
    refresh_seconds: int = 60,
    sort_by_symbol: bool = True,
) -> None:
    """Write ``df`` to ``filename`` with a dark theme and auto-refresh.

    Parameters
    ----------
    df : DataFrame
        Table of data to export.
    filename : str
        Name of the HTML file.
    header : str
        Title to display at the top of the page.
    include_sort_buttons : bool, optional
        Whether to add buttons for sorting by timeframe columns.
    refresh_seconds : int, optional
        How often the page should refresh itself.
    sort_by_symbol : bool, optional
        If ``True``, sort rows alphabetically by the ``Symbol`` column before
        exporting. Set to ``False`` to preserve the current order.
    """

    if sort_by_symbol and "Symbol" in df.columns and not df.empty:
        df = df.sort_values("Symbol")

    os.makedirs("html", exist_ok=True)
    path = os.path.join("html", filename)
    print(f"Exporting data to HTML: {path}")

    def style_cell(val: float) -> str:
        try:
            num = float(val)
        except (TypeError, ValueError):
            return ""
        return (
            "background-color:#C6EFCE;color:#006100"
            if num >= 0
            else "background-color:#FFC7CE;color:#9C0006"
        )

    numeric_cols = df.select_dtypes("number").columns
    highlight_cols = [c for c in numeric_cols if c != "Spread"]
    styled = df.style.map(style_cell, subset=highlight_cols)

    format_dict: dict[str, callable] = {}
    for col in numeric_cols:
        if col == "Spread":
            decimals = 4
        else:
            decimals = 2
        # Values are stored as decimal fractions (e.g. 0.004 -> 0.4%).
        # Multiply by 100 before formatting so the HTML output matches
        # the Excel export.
        format_dict[col] = lambda x, dec=decimals: f"{x * 100:.{dec}f}%"

    styled = styled.format(format_dict)
    html_table = styled.to_html(index=False, table_uuid="data-table")
    html_table = html_table.replace('id="T_data-table"', 'id="data-table"')
    html_table = re.sub(r'<th[^>]*row_heading[^>]*>.*?</th>', "", html_table, flags=re.DOTALL)
    html_table = re.sub(r'<th[^>]*class="blank level0"[^>]*>.*?</th>', "", html_table, flags=re.DOTALL)

    nav = ""
    if include_sort_buttons:
        sort_buttons = "".join(
            f"<button onclick=\"sortBy('{tf}')\">{tf}</button>" for tf in df.columns if tf in TIMEFRAMES
        )
        nav = (
            "<div style='display:flex;justify-content:flex-start;align-items:center;gap:4px;margin-bottom:8px'>"
            f"{sort_buttons}</div>"
        )

    with open(path, "w", encoding="utf-8") as f:
        f.write(
            "<html><head><meta charset='utf-8'>"
            f"<meta http-equiv='refresh' content='{refresh_seconds}'>"
            "<style>"
            "body{background:#121212;color:#fff;font-family:Arial,Helvetica,sans-serif;}"
            "table{background:#1e1e1e;color:#fff;border-collapse:collapse;width:100%;}"
            "th,td{border:1px solid #333;padding:4px;text-align:right;}"
            "th{background:#333;}"
            "td:first-child,th:first-child{text-align:left;}"
            "button{background:#333;color:#fff;border:1px solid #555;padding:4px 8px;margin-right:4px;cursor:pointer;}"
            "button:hover{background:#444;}"
            "</style>"
            f"<title>{header}</title></head><body>"
        )
        if nav:
            f.write(nav)
        f.write(f"<h1 style='margin-top:8px'>{header}</h1>")
        f.write(html_table)
        if include_sort_buttons:
            f.write(
                "<script>"
                "const sortDirections = {};"
                "function sortBy(col){"
                "  const table=document.getElementById('data-table');"
                "  const headers=Array.from(table.rows[0].cells).map(c=>c.textContent.trim());"
                "  const idx=headers.indexOf(col);"
                "  if(idx===-1)return;"
                "  const dir=sortDirections[col]||'desc';"
                "  const rows=Array.from(table.tBodies[0].rows);"
                "  rows.sort((a,b)=>{"
                "    const aVal=parseFloat(a.cells[idx].textContent.replace(/[%,$]/g,'').replace(/,/g,''))||0;"
                "    const bVal=parseFloat(b.cells[idx].textContent.replace(/[%,$]/g,'').replace(/,/g,''))||0;"
                "    return dir==='desc'?bVal-aVal:aVal-bVal;"
                "  });"
                "  rows.forEach(r=>table.tBodies[0].appendChild(r));"
                "  sortDirections[col]=dir==='desc'?'asc':'desc';"
                "}"
                "</script>"
            )
        f.write("</body></html>")

    open_in_edge(os.path.abspath(path))


def export_spread_html(spreads: dict) -> None:
    """Write the current spreads to ``spread.html``."""

    spread_df = pd.DataFrame(list(spreads.items()), columns=["Symbol", "Spread"])
    spread_df = spread_df.sort_values("Spread", ascending=True).reset_index(drop=True)
    export_to_html(
        spread_df,
        "spread.html",
        header="Current Spread %",
        include_sort_buttons=False,
        refresh_seconds=60,
        sort_by_symbol=False,
    )


def export_price_html(price_change: dict, price_range: dict) -> None:
    """Write price change and range tables to HTML."""

    price_change_df = pd.DataFrame.from_dict(price_change, orient="index")
    price_change_df.insert(0, "Symbol", price_change_df.index)

    price_range_df = pd.DataFrame.from_dict(price_range, orient="index")
    price_range_df.insert(0, "Symbol", price_range_df.index)

    export_to_html(
        price_range_df,
        "price_range.html",
        header="% Price Range",
        include_sort_buttons=True,
        refresh_seconds=15 * 60,
    )
    export_to_html(
        price_change_df,
        "price_change.html",
        header="% Price Change",
        include_sort_buttons=True,
        refresh_seconds=15 * 60,
    )


def export_correlation_html(correlations: dict) -> None:
    """Write correlation matrices to a single HTML file with buttons."""

    if not correlations:
        return

    os.makedirs("html", exist_ok=True)
    path = os.path.join("html", "correlation.html")
    print(f"Exporting data to HTML: {path}")

    first_label = next(iter(correlations))
    buttons = "".join(
        f"<button onclick=\"showTable('{label}')\">{label}</button>" for label in correlations
    )
    nav = (
        "<div style='display:flex;justify-content:flex-start;align-items:center;"
        "gap:4px;margin-bottom:8px'>"
        f"{buttons}</div>"
    )

    tables: list[str] = []
    for label, df in correlations.items():
        df = df.copy()
        df.insert(0, "Symbol", df.index)
        def style_cell(val: float) -> str:
            try:
                num = float(val)
            except (TypeError, ValueError):
                return ""
            return "background-color:#C6EFCE;color:#006100" if num >= 0 else "background-color:#FFC7CE;color:#9C0006"

        numeric_cols = df.select_dtypes("number").columns
        styled = df.style.map(style_cell, subset=numeric_cols)
        styled = styled.format({c: (lambda x: f"{x:.2f}%") for c in numeric_cols})
        html = styled.to_html(index=False, table_uuid=f"data-table-{label}")
        html = html.replace(
            f'id="T_data-table-{label}"', f'id="data-table-{label}"'
        )
        html = re.sub(r'<th[^>]*row_heading[^>]*>.*?</th>', "", html, flags=re.DOTALL)
        html = re.sub(
            r'<th[^>]*class="blank level0"[^>]*>.*?</th>', "", html, flags=re.DOTALL
        )
        display = "block" if label == first_label else "none"
        tables.append(
            f"<div id='{label}' style='display:{display}'>"
            f"<h1 style='margin-top:8px'>Correlation Matrix {label}</h1>"
            f"{html}</div>"
        )

    with open(path, "w", encoding="utf-8") as f:
        f.write(
            "<html><head><meta charset='utf-8'>"
            "<meta http-equiv='refresh' content='1800'>"
            "<style>"
            "body{background:#121212;color:#fff;font-family:Arial,Helvetica,sans-serif;}"
            "table{background:#1e1e1e;color:#fff;border-collapse:collapse;width:100%;}"
            "th,td{border:1px solid #333;padding:4px;text-align:right;}"
            "th{background:#333;}"
            "td:first-child,th:first-child{text-align:left;}"
            "button{background:#333;color:#fff;border:1px solid #555;padding:4px 8px;margin-right:4px;cursor:pointer;}"
            "button:hover{background:#444;}"
            "</style></head><body>"
        )
        f.write(nav)
        for table in tables:
            f.write(table)
        f.write(
            "<script>"
            f"let current='{first_label}';"
            "function showTable(tf){"
            "  if(current){document.getElementById(current).style.display='none';}"
            "  document.getElementById(tf).style.display='block';"
            "  current=tf;"
            "}"
            "</script></body></html>"
        )

    open_in_edge(os.path.abspath(path))


def export_all_html(price_change, price_range, spreads, correlations) -> None:
    """Export all results to individual HTML files."""

    export_price_html(price_change, price_range)
    export_spread_html(spreads)
    export_correlation_html(correlations)



def run_scan():
    """Run all scans and save results to a timestamped Excel file."""
    pc_data, pr_data, spreads, correlations = build_data()
    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = os.path.join(OUTPUT_DIR, f'Scan_{timestamp}.xlsx')
    export_to_excel(pc_data, pr_data, spreads, correlations, filename)
    print('Saved', filename)
    export_all_html(pc_data, pr_data, spreads, correlations)


def run_periodic_scans() -> None:
    """Continuously refresh each metric at its own interval."""

    price_change: dict = {}
    price_range: dict = {}
    spreads: dict = {}
    correlations: dict = {}

    intervals = {
        "spread": 60,
        "price": 15 * 60,
        "corr": 30 * 60,
    }
    next_run = {k: 0 for k in intervals}

    while True:
        now = time.time()

        if now >= next_run["spread"]:
            spreads = fetch_spreads(INSTRUMENTS)
            export_spread_html(spreads)
            next_run["spread"] = now + intervals["spread"]

        if now >= next_run["price"]:
            price_change.clear()
            price_range.clear()
            for inst in INSTRUMENTS:
                pc, pr = calc_price_metrics(inst)
                price_change[inst] = pc
                price_range[inst] = pr
            export_price_html(price_change, price_range)
            next_run["price"] = now + intervals["price"]

        if now >= next_run["corr"]:
            correlations = build_correlations()
            export_correlation_html(correlations)
            next_run["corr"] = now + intervals["corr"]

        time.sleep(60)


if __name__ == '__main__':
    if not API_KEY or not ACCOUNT_ID:
        raise SystemExit('Please set OANDA_API_KEY and OANDA_ACCOUNT_ID environment variables')

    available = get_available_instruments()
    if available:
        INSTRUMENTS[:] = [i for i in INSTRUMENTS if i in available]

    run_periodic_scans()
